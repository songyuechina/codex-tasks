# -*- coding: utf-8 -*-
# 文件位置: D:/codex-tasks/cad/system/common_logger.py
# 版本: V3.1
"""
common_logger.py（收束版）

目标：
- 为整个 cad 系统提供统一日志系统（console + rotating file）
- 提供统一调试开关：set_debug_mode(mode / who / wait / log_level)
- 提供可选 Excel 测试记录：record_test_result / checkpoint / CriticalSection
- 保持库模块导入尽量“无副作用”

规则：
- 业务脚本统一使用：
    from system.common_logger import sys_logger
- 入口/测试/agent 启动脚本才允许调用：
    set_debug_mode(...)
- 普通业务模块禁止 print()；应使用 sys_logger.*
- CriticalSection 不吞异常
"""

from __future__ import annotations

import datetime
import inspect
import logging
import os
import sys
import time
from pathlib import Path
from logging.handlers import RotatingFileHandler
from typing import Any, Dict, Optional

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
except Exception:
    openpyxl = None
    Workbook = None
    load_workbook = None


# ==========================================
# 1. 全局路径配置
# ==========================================
CURRENT_FILE = Path(__file__).resolve()
CAD_ROOT = CURRENT_FILE.parent.parent
TESTS_DIR = CAD_ROOT / "tests"
TEST_EXCEL_PATH = TESTS_DIR / "testfunc.xlsx"

try:
    TESTS_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    pass


# ==========================================
# 2. 全局调试配置
# ==========================================
DEBUG_CONFIG: Dict[str, Any] = {
    "MODE": 0,             # 0=安静 1=调试
    "WHO": "AI",           # AI / HUMAN
    "WAIT": 0,             # HUMAN 暂停秒数
    "LOG_LEVEL": "WARNING" # 默认安静
}

_LOGGER_NAME = "CAD_System"


def _parse_log_level(level: Any) -> int:
    """把 'INFO' / logging.INFO / None 转成 logging level int。"""
    if level is None:
        return logging.INFO
    if isinstance(level, int):
        return level
    s = str(level).strip().upper()
    return getattr(logging, s, logging.INFO)


# ==========================================
# 3. Logger 构造
# ==========================================
def setup_logger(log_file: str = "system_run.log") -> logging.Logger:
    """
    创建或复用全局 logger。
    - logger 名字固定为 CAD_System
    - 文件日志：cad/system/system_run.log
    - 控制台日志：stdout
    """
    logger = logging.getLogger(_LOGGER_NAME)
    initial_level = _parse_log_level(DEBUG_CONFIG.get("LOG_LEVEL", "INFO"))
    logger.setLevel(initial_level)

    formatter = logging.Formatter(
        "%(asctime)s - [%(levelname)s] - %(filename)s:%(lineno)d - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # 避免重复添加 handler
    if logger.handlers:
        for h in logger.handlers:
            try:
                h.setLevel(initial_level)
                h.setFormatter(formatter)
            except Exception:
                pass
        return logger

    # File handler
    try:
        log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), log_file)
        fh = RotatingFileHandler(
            log_path,
            maxBytes=5 * 1024 * 1024,
            backupCount=3,
            encoding="utf-8",
        )
        fh.setFormatter(formatter)
        fh.setLevel(initial_level)
        logger.addHandler(fh)
    except Exception:
        pass

    # Console handler
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(formatter)
    ch.setLevel(initial_level)
    logger.addHandler(ch)

    return logger


sys_logger = setup_logger()


def set_log_level(
    level: Any = "INFO",
    *,
    apply_to_console: bool = True,
    apply_to_file: bool = True,
) -> None:
    """
    动态设置日志级别：同时设置 logger 与各 handler。
    """
    lv = _parse_log_level(level)
    logger = logging.getLogger(_LOGGER_NAME)
    logger.setLevel(lv)

    for h in logger.handlers:
        if apply_to_file and isinstance(h, RotatingFileHandler):
            h.setLevel(lv)
        if apply_to_console and isinstance(h, logging.StreamHandler) and not isinstance(
            h, RotatingFileHandler
        ):
            h.setLevel(lv)


def set_debug_mode(
    mode: int = 1,
    who: str = "AI",
    wait_time: int = 30,
    log_level: Optional[str] = None,
) -> None:
    """
    调试模式总开关。
    - MODE / WHO / WAIT：控制 checkpoint / CriticalSection 的人工暂停逻辑
    - LOG_LEVEL：控制 console + file 的日志输出强度
    """
    DEBUG_CONFIG["MODE"] = int(mode)
    DEBUG_CONFIG["WHO"] = str(who).strip().upper()
    DEBUG_CONFIG["WAIT"] = int(wait_time)

    if log_level is None:
        desired = "WARNING" if DEBUG_CONFIG["MODE"] == 0 else "INFO"
    else:
        desired = str(log_level).strip().upper()

    DEBUG_CONFIG["LOG_LEVEL"] = desired

    try:
        set_log_level(desired, apply_to_console=True, apply_to_file=True)
    except Exception:
        pass

    try:
        sys_logger.info(
            f"🔧 调试配置更新: Mode={DEBUG_CONFIG['MODE']}, "
            f"Who={DEBUG_CONFIG['WHO']}, "
            f"Wait={DEBUG_CONFIG['WAIT']}s, "
            f"LogLevel={desired}"
        )
    except Exception:
        pass


# ==========================================
# 4. Excel 记录
# ==========================================
def record_test_result(
    script_name: Optional[str] = None,
    func_name: Optional[str] = None,
    is_pass: bool = False,
    **variables: Any,
) -> None:
    """
    写一条测试/断点记录到 Excel。
    - 若未传 script_name/func_name，则自动推断调用者
    - 若 openpyxl 不可用，则直接跳过
    """
    if not openpyxl:
        return

    if script_name is None or func_name is None:
        try:
            frame = inspect.currentframe().f_back
            if script_name is None:
                script_name = os.path.basename(frame.f_code.co_filename)
            if func_name is None:
                func_name = frame.f_code.co_name
                if func_name == "<module>":
                    func_name = "[Main]"
        except Exception:
            script_name = script_name or "Unknown"
            func_name = func_name or "Unknown"

    try:
        row_data = [
            f"Time:{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Script:{script_name}",
            f"Func:{func_name}",
        ]

        for k, v in variables.items():
            val = str(v)
            if len(val) > 500:
                val = val[:500] + "..."
            row_data.append(f"{k}:{val}")

        row_data.append(f"Result:{'PASS' if is_pass else 'FAIL'}")

        if os.path.exists(TEST_EXCEL_PATH):
            wb = load_workbook(TEST_EXCEL_PATH)  # type: ignore[misc]
            ws = wb.active
        else:
            wb = Workbook()  # type: ignore[misc]
            ws = wb.active

        ws.append(row_data)
        wb.save(TEST_EXCEL_PATH)
    except Exception as e:
        sys_logger.error(f"❌ Excel Error: {e}")


# ==========================================
# 5. 核心工具
# ==========================================
def checkpoint(desc: str, is_pass: bool = True, **variables: Any) -> None:
    """
    轻量检查点：
    - 记录 Excel
    - HUMAN + WAIT>0 时进行人工暂停
    """
    try:
        frame = inspect.currentframe().f_back
        s_name = os.path.basename(frame.f_code.co_filename)
        f_name = frame.f_code.co_name
    except Exception:
        s_name, f_name = "Unknown", "Unknown"

    record_test_result(s_name, f_name, is_pass, Checkpoint=desc, **variables)

    if DEBUG_CONFIG["MODE"] == 1 and DEBUG_CONFIG["WHO"] == "HUMAN":
        wait = int(DEBUG_CONFIG.get("WAIT", 0))
        if wait > 0:
            print(f"\n✋ [检查点] {desc} | Data: {variables}")
            print(f"⏳ 暂停 {wait}s (Ctrl+C 跳过)...")
            try:
                time.sleep(wait)
            except KeyboardInterrupt:
                print("⏩ 跳过")


class CriticalSection:
    """
    关键区段上下文管理器：
    - 进入时记录 info
    - 失败时记录 error
    - 写 Excel
    - 不吞异常（__exit__ 返回 False）
    """

    def __init__(
        self,
        description: str = "关键操作",
        script_name: Optional[str] = None,
        func_name: Optional[str] = None,
    ):
        self.desc = description

        if script_name is None or func_name is None:
            try:
                frame = inspect.currentframe().f_back
                if script_name is None:
                    script_name = os.path.basename(frame.f_code.co_filename)
                if func_name is None:
                    func_name = frame.f_code.co_name
            except Exception:
                script_name = script_name or "Unknown"
                func_name = func_name or "Unknown"

        self.s_name = script_name
        self.f_name = func_name

        self.mode = int(DEBUG_CONFIG.get("MODE", 0))
        self.who = str(DEBUG_CONFIG.get("WHO", "AI")).upper()
        self.wait_time = int(DEBUG_CONFIG.get("WAIT", 0))

        self.is_active = self.mode == 1
        self.is_human = self.is_active and self.who == "HUMAN"

        self.metrics: Dict[str, Any] = {}
        self.is_pass = False

    def record(self, **kwargs: Any) -> None:
        self.metrics.update(kwargs)

    def __enter__(self):
        sys_logger.info(f"🚩 进入区域: {self.desc} [Mode:{self.mode}|{self.who}]")
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            self.is_pass = False
            self.metrics["error"] = str(exc_val)
            sys_logger.error(f"❌ 区域崩溃: {exc_val}")
        else:
            self.is_pass = "error" not in self.metrics

        try:
            record_test_result(
                self.s_name,
                self.f_name,
                self.is_pass,
                Section=self.desc,
                **self.metrics,
            )
        except Exception:
            pass

        if self.is_human and self.wait_time > 0:
            print(f"\n✋ [人工干预] {self.desc} | Vars: {self.metrics}")
            print(f"⏳ 暂停 {self.wait_time}s...")
            try:
                time.sleep(self.wait_time)
            except KeyboardInterrupt:
                print("⏩ 跳过")

        return False


# ==========================================
# 6. 兼容辅助
# ==========================================
def node(msg: str, *args: Any, **kwargs: Any) -> None:
    """
    兼容旧接口。
    等价于 sys_logger.info(msg)。
    """
    if args or kwargs:
        try:
            msg = msg.format(*args, **kwargs)
        except Exception:
            pass
    sys_logger.info(msg)


__all__ = [
    "sys_logger",
    "set_debug_mode",
    "set_log_level",
    "record_test_result",
    "checkpoint",
    "CriticalSection",
    "node",
]
