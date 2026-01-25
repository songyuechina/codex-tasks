# -*- coding: utf-8 -*-
# 文件位置: D:/claude-tasks/cad/system/common_logger.py
# 版本: V1.4 (增加 AI/Human 模式区分 & 智能等待)

import os
import sys
import logging
import time
import datetime
from pathlib import Path
from logging.handlers import RotatingFileHandler

import openpyxl
from openpyxl import Workbook, load_workbook




# ... (引入 openpyxl 的代码保持不变) ...
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
except ImportError:
    openpyxl = None

# ==========================================
# 1. 全局调试总控参数 (可在其他脚本修改)
# 0 = 关闭调试 (代码不执行)
# 1 = 开启调试 (代码执行)
# ==========================================
DEBUG_MODE = 0 

def set_debug_mode(mode):
    """用于在其他脚本中动态修改调试状态"""
    global DEBUG_MODE
    DEBUG_MODE = mode
    sys_logger.info(f"🔧 调试模式已切换为: {'开启' if mode == 1 else '关闭'}")

# ==========================================
# 1. 全局调试配置字典
# ==========================================
DEBUG_CONFIG = {
    "MODE": 0,          # 0=关闭, 1=开启
    "WHO": "HUMAN",     # "AI"=自动化测试, "HUMAN"=人类观察
    "WAIT": 30          # 人类模式下的等待时间(秒)
}

def set_debug_mode(mode=1, who="HUMAN", wait_time=30):
    """
    设置调试系统的运行模式
    :param mode: 0=关闭调试, 1=开启调试
    :param who: 'AI' (记录数据不等待) 或 'HUMAN' (等待用户观察)
    :param wait_time: 当 who='HUMAN' 时的暂停时间
    """
    global DEBUG_CONFIG
    DEBUG_CONFIG["MODE"] = mode
    # 统一转大写，防止 'ai', 'Ai' 写法不一致
    DEBUG_CONFIG["WHO"] = str(who).upper() 
    DEBUG_CONFIG["WAIT"] = wait_time
    
    sys_logger.info(f"🔧 调试配置更新: 模式={mode}, 对象={DEBUG_CONFIG['WHO']}, 等待={DEBUG_CONFIG['WAIT']}s")

# ==========================================
# 2. 路径系统配置 (动态获取)
# ==========================================
# common_logger.py 位于 .../cad/system/
CURRENT_FILE = Path(__file__).resolve()
SYSTEM_DIR = CURRENT_FILE.parent
CAD_DIR = SYSTEM_DIR.parent
TESTS_DIR = CAD_DIR / "tests"
TEST_EXCEL_PATH = TESTS_DIR / "testfunc.xlsx"

# 确保 tests 目录存在
if not TESTS_DIR.exists():
    try:
        TESTS_DIR.mkdir(parents=True, exist_ok=True)
    except:
        pass

# ==========================================
# 3. Excel 测试记录功能
# ==========================================

def record_test_result(script_name, func_name, is_pass, **variables):
    """
    将测试运行时的变量快照写入 Excel。
    
    参数:
    - script_name: 脚本名
    - func_name: 函数名
    - is_pass: 是否通过 (True/False)
    - **variables: 任意多个关键变量 (使用关键字参数传递，如 x=1, msg="test")
    
    格式: [时间:值, 脚本:值, 函数:值, 变量1:值, 变量2:值, ..., 结果:PASS/FAIL]
    """
    
    # 全局配置检查 (假设 TEST_EXCEL_PATH 已定义，否则使用默认路径)
    excel_path = str(globals().get("TEST_EXCEL_PATH", "D:/claude-tasks/cad/tests/test_log.xlsx"))

    try:
        # 1. 准备行数据 (Row Data)
        row_data = []

        # A. 时间
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row_data.append(f"Time:{current_time}")

        # B. 脚本名
        row_data.append(f"Script:{script_name}")

        # C. 函数名
        row_data.append(f"Func:{func_name}")

        # D. 任意多个变量 (核心逻辑：name:value)
        # 遍历传入的关键字参数，每个变量占一个格子
        for name, value in variables.items():
            # 将值转为字符串，防止 None 或对象导致写入报错
            row_data.append(f"{name}:{str(value)}")

        # E. 结果 (始终放在最后)
        pass_status = "PASS" if is_pass else "FAIL"
        row_data.append(f"Result:{pass_status}")

        # 2. Excel 操作
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            # 注意：根据要求，此处不再写入表头 (Header)

        # 3. 写入 (append 会自动寻找第一个空行)
        ws.append(row_data)
        
        # 4. 保存
        wb.save(excel_path)
        # 如果有日志系统，可以打印一下 (可选)
        # print(f"📝 记录已保存: {row_data}")

    except Exception as e:
        # 简单的错误捕获，防止记录日志本身导致程序崩溃
        print(f"❌ 写入 Excel 失败: {e}")


# ==========================================
# 4. 日志系统 (保持原有逻辑)
# ==========================================

def setup_logger(log_file="system_run.log"):
    """
    配置全局唯一的日志记录器
    """
    logger = logging.getLogger("CAD_System")
    logger.setLevel(logging.INFO)

    # 防止重复添加 Handler (避免日志重复打印)
    if logger.handlers:
        return logger

    # 1. 定义格式: 时间 - 级别 - 文件名:行号 - 消息
    formatter = logging.Formatter(
        '%(asctime)s - [%(levelname)s] - %(filename)s:%(lineno)d - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    # 2. 文件输出 (自动保存在本脚本同级目录下)
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # [优化] 确保目录存在（虽然获取的是当前目录，但为了健壮性）
    if not os.path.exists(current_dir):
        os.makedirs(current_dir, exist_ok=True)
        
    log_path = os.path.join(current_dir, log_file)
    
    try:
        # 5MB 一个文件，最多保留 3 个备份
        file_handler = RotatingFileHandler(
            log_path, maxBytes=5*1024*1024, backupCount=3, encoding='utf-8'
        )
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
    except Exception as e:
        print(f"❌ 无法创建日志文件 handler: {e}")

    # 3. 控制台输出
    # [优化] 解决 Windows 控制台中文输出可能崩溃的问题
    try:
        # 尝试强制设置 stdout 为 utf-8 (Python 3.7+)
        if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

    console_handler = logging.StreamHandler(sys.stdout) 
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    return logger

# 全局单例
sys_logger = setup_logger()

# ==========================================
# 5. 智能调试上下文类 (核心修改)
# ==========================================
class DebugContext:
    def __init__(self, name="调试片段"):
        self.name = name
        self.is_active = (DEBUG_CONFIG["MODE"] == 1)
        self.who = DEBUG_CONFIG["WHO"]
        self.wait_time = DEBUG_CONFIG["WAIT"]

    @property
    def is_ai(self):
        """是否为 AI 自动化模式"""
        return self.is_active and (self.who == "AI")

    @property
    def is_human(self):
        """是否为 人类 观察模式"""
        return self.is_active and (self.who == "HUMAN")

    def __enter__(self):
        if self.is_active:
            sys_logger.info(f"🔰 [DEBUG START]: {self.name} ({self.who} Mode)")
        # 返回 self，以便在 with 内部访问属性
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if not self.is_active:
            return False

        # 如果是人类模式，执行倒计时等待
        if self.is_human and self.wait_time > 0:
            print(f"\n👀 [人类观察模式] 请检查 CAD 窗口...")
            print(f"⏳ 程序将在 {self.wait_time} 秒后继续 (按 Ctrl+C 可强制跳过)...")
            try:
                # 简单的倒计时显示
                for i in range(self.wait_time, 0, -1):
                    # end='' 实现不换行刷新
                    print(f"\r⏳ 剩余时间: {i} 秒   ", end='', flush=True) 
                    time.sleep(1)
                print("\r🚀 继续执行...                   \n")
            except KeyboardInterrupt:
                print("\n⏩ 用户跳过等待")
        
        # AI 模式下通常不等待，或者只等待极短时间确保 I/O 完成
        elif self.is_ai:
            print(f"🤖 [AI 模式] {self.name} 执行完毕，无等待。")

        sys_logger.info(f"⏹️ [DEBUG END]: {self.name}")
        return False

# =========================================================
# 🆕 新增：节点日志辅助函数 (Node Logger)
# =========================================================
def node(msg, *args, **kwargs):
    """
    【简易日志包装器】
    功能：
    1. 支持类似 C# 的 {} 占位符格式化，例如: node("处理: {}", item)
    2. 统一发送到 sys_logger.info
    """
    if args or kwargs:
        try:
            # 尝试 .format() 格式化
            formatted_msg = msg.format(*args, **kwargs)
        except:
            # 如果格式化失败（比如参数不匹配），则原样输出，避免报错
            formatted_msg = f"{msg} [参数: {args} {kwargs}]"
    else:
        formatted_msg = msg
        
    # 发送到系统日志
    sys_logger.info(formatted_msg)
    
if __name__ == "__main__":
    # --- 自测代码 ---
    
    print("--- 测试 1: 默认模式 (DEBUG_MODE=0) ---")
    with DebugContext("测试应该跳过") as on:
        if on:
            print("❌ 这行不应该被打印！")
            time.sleep(5)

    print("--- 测试 2: 开启模式 (DEBUG_MODE=1) ---")
    set_debug_mode(1) # 开启总控
    
    with DebugContext("测试生效") as on:
        if on:
            print("✅ 这行应该被打印！")
            print("正在模拟休眠 2秒...")
            time.sleep(2)
            
    print("测试结束。")
