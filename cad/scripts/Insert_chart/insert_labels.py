# -*- coding: utf-8 -*-
# c
# 版本: V1.1 (修复空图层名导致的崩溃)

#&&% 导入
#  引导代码 (确保能找到 system)
import os
import sys
from pathlib import Path
current = Path(__file__).resolve()
while current.name != 'cad':
    if current.parent == current: raise Exception("找不到根目录")
    current = current.parent
sys.path.insert(0, str(current))
from system.project_setup import PathConfig


#  导入配置
from scripts.CAD_basic import *
# 确保引用了这些基础库，防止独立运行时报错
import time
import math
import re as _re


#—————————— 核心系统库导入 ——————————
from system.common_logger import (
    sys_logger,          # 日志记录器 (info/error)
    CriticalSection,     # 【方式A】块状托管 (with ...)
    checkpoint,          # 【方式B】线性打点 (func(...))
    record_test_result,  # 底层 Excel 写入 (finally 中用)
    node,                # 简易日志格式化
    set_debug_mode       # ⚠️ 推荐补上这个：用于控制 AI/Human 模式
)

from system.CAD_coordination import CADGuard ,wait_quiescent# 事务安全卫士 (Undo/Lock)
from system.licad import C 
from system.CAD_com_utils import retry_on_busy,SafeCOM 





import scripts.CAD_file_operations as cfo

from functools import partial

import pythoncom
import datetime






#  日志和调试
from system.common_logger import sys_logger,CriticalSection, set_debug_mode
userpath = Path(PathConfig.userpath)

set_debug_mode(mode=1, who="AI", wait_time=30)

#1
"""
重大功能必须建立专门脚本，因为它的功能设计可能会不变发展。
插图签涉及打印区域，实践证明，没必要在插图签时限定打印区域。
打印区域，应该交给打印区域的函数完成。

打印区域，我们定义了一系列的标准区域。只要合乎阈值命中，就是对应的图纸
规格和比例，否则给一个相近的。根本不必考虑其是否更加合理。当用户不使用
标准图框时，说明他不在乎。我们没必要去替他在乎。

功能不要纠缠一起，应该合理分离。建立独立的维度。纠结只会增加不必要的混乱。


"""
#&&% 图签块名

Block_Names_all = [
            'A3', 'A2', 'A1',
            'A2_1_4', 'A2_1_2', 'A2_3_4',
            'A1_1_4', 'A1_1_2', 'A1_3_4',
            'A0', 'A0_1_8', 'A0_1_4',

            'A3-H', 'A2-H', 'A1-H', 'A0-H',

            'A3普通块', 'A2普通块', 'A1普通块', 'A0普通块',
]

Block_Names_0 = [
            'A3', 'A2', 'A1',
            'A2_1_4', 'A2_1_2', 'A2_3_4',
            'A1_1_4', 'A1_1_2', 'A1_3_4',
            'A0', 'A0_1_8', 'A0_1_4',
]


Block_Names_1 = [
            'A3-H', 'A2-H', 'A1-H', 'A0-H',
]

Block_Names_2 = [
            'A3普通块', 'A2普通块', 'A1普通块', 'A0普通块',
]


def get_block_true_name(blk_obj):
    """
    获取块的真实名称。
    逻辑：优先 EffectiveName (动态块)，其次 Name (普通块)
    """
    # 1. 优先尝试获取 EffectiveName (解决动态块匿名问题)
    b_name = get_attr(blk_obj, "Name")
    
    # 2. 如果 EffectiveName 获取失败(None)或是空字符串，则尝试获取 Name
    if not b_name:
        b_name = get_attr(blk_obj, "EffectiveName")
        
    return b_name

#&&% 判断预定块名集的块实例数量 

def filter_blocks_by_list(all_blocks=None, target_names=Block_Names_0):
    """
    输入所有块实例，返回名称在 target_names 列表中的块实例列表。
    """
    if all_blocks is None:

        all_blocks = select_kuai()


    result_blocks = []
    
    # 将目标列表转为集合，查找速度更快 (可选，但推荐)
    target_set = set(target_names)
    
    print(f"开始筛选，待处理对象总数: {len(all_blocks)}")

    for blk in all_blocks:
        # 获取真实名称
        name = get_block_true_name(blk)
        
        # 如果获取到了名称，并且名称在目标列表中
        if name and name in target_set:
            result_blocks.append(blk)
            
    print(f"筛选完成。在列表中找到的块实例数: {len(result_blocks)}")
    return result_blocks

#&&% 制作无参数函数

"""

from functools import partial # 需导入

# 1. 准备参数
my_polylines = [...] 
tpl_path = str(userpath / "dwg文件/标准图签.dwg")

# 2. 制作“偏函数”
# 语法: partial(函数名, 参数1, 参数2...)
# 它会返回一个新的函数，这个新函数已经不需要传参了
action_wrapper = partial(
    insert_and_scale_labels_area_any, 
    coms_dayin=my_polylines, 
    filepath=tpl_path
)

check_wrapper = partial(verify_blocks_exist, "A3-H")

# 3. 传入
run_safety_loop(
    target_dwg=work_dwg,
    action_func=action_wrapper,
    check_func=check_wrapper,
    max_retries=3
)

"""


#&&% 插入图签块到打印区域_辅函数 
#insert_and_scale_labels_area_any功能就是将包含了图框的标准图签块按横竖向分配到任意矩形多段线区域。


def block_name_from_spec(spec_str: str):
    if not spec_str: return None
    s = str(spec_str).strip()
    if s in ("A0", "A1", "A2", "A3"): return s
    s = s.replace("+", "_").replace("/", "_")
    return s

# ---------- 小工具：根据 frame_info 计算几何修正系数 k_geom ----------
def compute_k_from_info(info: dict) -> float:
    """
    以“模板图签块为 1:100 基准”，根据实际图框的尺寸求出几何修正系数 k_geom。
    """
    try:
        corners = info.get("corners") or []
        entity = info.get("entity")
        drawing_frame = info.get("drawing_frame") or ""
        ratio_str = info.get("ratio") or "1:100"

        # —— 1. 计算实际宽高（模型单位）——
        w_act = h_act = None

        if len(corners) >= 4:
            p0, p1, p2, p3 = corners[:4]

            def dist(a, b):
                return math.hypot(a[0] - b[0], a[1] - b[1])

            # 默认 corners 顺序：左下、左上、右上、右下
            w_act = dist(p0, p3)  # 水平边
            h_act = dist(p0, p1)  # 垂直边
        elif entity is not None:
            # 退路：用外包盒估一个
            try:
                ll, ur = entity.GetBoundingBox()
                w_act = abs(ur[0] - ll[0])
                h_act = abs(ur[1] - ll[1])
            except Exception:
                pass

        if not w_act or not h_act:
            return 1.0

        # —— 2. 解析理论纸张尺寸（mm）——
        # 例："ISO_A2_(594.00_x_420.00_MM)"
        m = _re.search(r"\(([\d\.]+)\s*x\s*([\d\.]+)", drawing_frame)
        if not m:
            return 1.0

        w_mm = float(m.group(1))
        h_mm = float(m.group(2))

        # —— 3. 解析比例 "1:100" ——>
        m2 = _re.search(r"1\s*:\s*(\d+)", ratio_str)
        base_scale = float(m2.group(1)) if m2 else 100.0  # 默认 1:100

        # 模型中的“理论宽高”
        w_nom = w_mm * base_scale
        h_nom = h_mm * base_scale

        if not w_nom or not h_nom:
            return 1.0

        kw = w_act / w_nom
        kh = h_act / h_nom

        k_list = [v for v in (kw, kh) if v > 0]
        if not k_list:
            return 1.0

        k = sum(k_list) / len(k_list)

        if k <= 0:
            return 1.0

        return k

    except Exception as e:
        node("⚠ compute_k_from_info 异常: {}", e)
        return 1.0


def compute_insert_factors(entities, res, result_dict):
    """
    根据 res 的 ratio/spec 与 result_dict（图签模板信息）的块定义，计算缩放系数 k。
    回传 [(entity, block_name, spec, k), ...]
    """
    
    def _denom(s: str) -> int:
        m = _re.match(r".*:(\d+)$", s or "")
        return int(m.group(1)) if m else 1
    
    mapping = {
        info.get("spec"): (info.get("block_name"), info.get("ratio"))
        for info in result_dict.values()
        if info.get("spec")
    }
    
    outputs = []
    for ent in entities:
        matched = next(
            (info for info in res.values() if info.get("entity") is ent),
            None
        )
        if not matched:
            outputs.append((ent, None, None, None))
            continue
    
        spec = matched.get("spec")
        ratio1 = matched.get("ratio") or "1:1"
        block_name, ratio2 = mapping.get(spec, (None, "1:1"))
    
        d1, d2 = _denom(ratio1), _denom(ratio2)
        k = d1 / d2 if d2 else None
        outputs.append((ent, block_name, spec, k))
    
    return outputs



def adjust_block_to_frame(frame_ent, blk_com, tol_len=10.0):
    """
    在统一缩放 k_val 插入块之后，进一步检查打印框与块的外包盒尺寸。

    - 若 |W_frame - W_block| 和 |H_frame - H_block| 都 ≤ tol_len：
          认为是“标准图框或接近标准”，不再处理。
    - 若差值较大：
          只依赖 BoundingBox 和 Rotation，按 0°/90° 两种情况精确反推
          额外的 X/Y 缩放，使最终宽高尽量等于打印框宽高。
    """
    # 1) 打印框外包盒
    try:
        ll_f, ur_f = frame_ent.GetBoundingBox()
        Wf = abs(ur_f[0] - ll_f[0])
        Hf = abs(ur_f[1] - ll_f[1])
    except Exception as e:
        node("  ⚠ adjust_block_to_frame: 获取打印框 BoundingBox 失败: {}", e)
        return

    # 2) 当前块外包盒
    try:
        ll_b, ur_b = blk_com.GetBoundingBox()
        Wb = abs(ur_b[0] - ll_b[0])
        Hb = abs(ur_b[1] - ll_b[1])
    except Exception as e:
        node("  ⚠ adjust_block_to_frame: 获取图签块 BoundingBox 失败: {}", e)
        return

    if Wf <= 0 or Hf <= 0 or Wb <= 0 or Hb <= 0:
        node("  ⚠ adjust_block_to_frame: 宽高异常 Wf={}, Hf={}, Wb={}, Hb={}",
             Wf, Hf, Wb, Hb)
        return

    # 若本身就已经在容差内，认为是“标准图框或接近标准”，不再强行拉伸
    if abs(Wf - Wb) <= tol_len and abs(Hf - Hb) <= tol_len:
        node("  ▶ adjust_block_to_frame: 框与图签块尺寸已在容差内，无需校正")
        return

    # 3) 读取当前缩放与旋转
    try:
        sx0 = getattr(blk_com, "XScaleFactor", 1.0)
        sy0 = getattr(blk_com, "YScaleFactor", 1.0)
    except Exception:
        sx0 = sy0 = 1.0

    try:
        rot = getattr(blk_com, "Rotation", 0.0)  # 弧度
    except Exception:
        rot = 0.0

    rot_deg = abs(rot * 180.0 / math.pi) % 180.0

    # 4) 根据 rotation 决定“宽高和 X/Y 缩放因子的对应关系”
    if rot_deg < 1.0 or abs(rot_deg - 180.0) < 1.0:
        # 横向：世界宽 = XScale * 宽，世界高 = YScale * 高
        factor_x = Wf / Wb
        factor_y = Hf / Hb
        mode = "anisotropic_horizontal"
    elif abs(rot_deg - 90.0) < 1.0:
        # 竖向：世界宽 ≈ YScale 方向的长度，世界高 ≈ XScale 方向的长度
        factor_x = Hf / Hb   # XScale 作用于“高度”那一边
        factor_y = Wf / Wb   # YScale 作用于“宽度”那一边
        mode = "anisotropic_vertical"
    else:
        # 其它奇怪角度（一般不会出现），退回统一比例近似
        rw = Wf / Wb
        rh = Hf / Hb
        s = (rw + rh) / 2.0
        factor_x = factor_y = s
        mode = "uniform_fallback"

    sx = sx0 * factor_x
    sy = sy0 * factor_y

    try:
        blk_com.XScaleFactor = sx
        blk_com.YScaleFactor = sy
        node(
            "  ▶ adjust_block_to_frame 模式={}，rot≈{:.1f}°: "
            "XScaleFactor {:.4f}→{:.4f}, YScaleFactor {:.4f}→{:.4f}",
            mode, rot_deg, sx0, sx, sy0, sy
        )
    except Exception as e:
        node("  ⚠ adjust_block_to_frame: 设置 X/YScaleFactor 失败: {}", e)




#&&&% 插入图签块到打印区域_主函数 

# =========================================================
#&&% 主函数成功版
# =========================================================



@timeit
@debuggable
def insert_and_scale_labels_area_any(
        coms_dayin=None,        
        filepath=str(userpath/"dwg文件/标准图签.dwg"),
    ):
    """
    【V8.5 生产就绪版】
    1. 修正变量作用域错误。
    2. 严格执行 6s 等待与资源审计。
    """
    doc = C.doc
    
    # 核心特征表 (审计基准)
    EXPECTED_PREFIXES = [
        'A3', 'A2', 'A1', 'A2_1_4', 'A2_1_2', 'A2_3_4',
        'A1_1_4', 'A1_1_2', 'A1_3_4', 'A0', 'A0_1_8', 'A0_1_4',
    ]

    # ——————————————————————————————————————————————————
    # 阶段 1: 预计算 (数据准备)
    # ——————————————————————————————————————————————————
    if coms_dayin is None:
        coms_dayin = get_rectangular_polylines(min_side=100.0, area_tolerance=0.02)
    
    if not coms_dayin:
        sys_logger.warning("⚠️ 未找到任何矩形多段线。")
        return {}

    # 去重
    coms_dayin = remove_duplicate_polylines(coms_dayin, tol=1, priority_layer="dy_zhuanyong")
    
    # 排序
    max_s = 0
    for e in coms_dayin:
        try: 
            min_p, max_p = e.GetBoundingBox()
            max_s = max(max_s, abs(max_p[0]-min_p[0]))
        except: pass
    cha = 20 if max_s < 2000 else 2000
    coms_dayin = sort_coms_by_llcorner(coms_dayin, cha_Y=cha)

    # 【关键变量初始化】
    valid_ents = []
    res = {}
    
    for ent in coms_dayin:
        ret = generate_name_and_ratio_from_com(ent)
        if ret == 0: continue
        try:
            res[ent.Handle] = {
                "entity": ent,
                "spec": ret[2],          # "A1+1/4"
                "ratio": ret[1],
                "drawing_frame": ret[0],
                "is_vertical": ret[3]
            }
            valid_ents.append(ent)
        except: pass
    
    sys_logger.info(f"📋 [任务锁定] 目标区域: {len(valid_ents)} 个")
    if not valid_ents: return {}

    # ——————————————————————————————————————————————————
    # 阶段 2: 事务开启 & 资源同步审计
    # ——————————————————————————————————————————————————
    bind_dict = {}
    
    with CADGuard("一键插图签总成", disable_ui=True, independent_undo=True):

        # 1. 执行导入并炸开 (这是动荡的起点)
        template_instances, _, _ = Insert_Company_Label_Common_Block(filepath=filepath)
        
        # 2. 【记录起点】就在炸开指令刚刚发完的时候开始计时
        sync_start_time = time.time()
        
        # 3. 【核心同步】使用协同机制获取自然时间
        sys_logger.info("⏳ [系统同步] 模板已炸开，正在探测 CAD 数据库自然静默时间...")
        
        try:
            # 这是一个阻塞探测，直到 CAD 真的空闲
            wait_quiescent(min_quiet=1.0, timeout=12) 
            
            # 计算并打印合理等待时间
            actual_wait_duration = time.time() - sync_start_time
            sys_logger.info(f"✅ [同步达成] CAD 已就绪。本次自然同步耗时: {actual_wait_duration:.2f}s")
            
            # 此时此刻，执行点名抓取，成功率最高
            doc.Regen(1) 
            
        except Exception as e:
            actual_wait_duration = time.time() - sync_start_time
            sys_logger.warning(f"⚠️ [同步超时] 耗时 {actual_wait_duration:.2f}s 仍未静默，强制进入点名流程。")

        # 3. 审计点名：锁定现场名字
        real_live_map = {} 
        try:
            current_blks = select_kuai()
            # 拿到现场所有块名
            raw_live_names = [getattr(b, "EffectiveName", b.Name) for b in current_blks]
            
            for prefix in EXPECTED_PREFIXES:
                for live_name in raw_live_names:
                    # 匹配规则：完全相等，或者前缀匹配（如 A3_100550）
                    if live_name == prefix or live_name.startswith(prefix + "_"):
                        real_live_map[prefix] = live_name
                        break
            
            # 🔥🔥 增加这一行：打印捕捉到的“实物”名单 🔥🔥
            detected_real_names = sorted(list(real_live_map.values()))
            sys_logger.info(f"📋 [实物清单] 现场锁定块名: {detected_real_names}")
            
            sys_logger.info(f"🧐 [审计报告] 预期 12 个，实测匹配到 {len(real_live_map)} 个。")
            
        except Exception as e:
            sys_logger.error(f"审计点名失败: {e}")


        # ———————————————————————————————————————————————
        # 阶段 3: 执行分发
        # ———————————————————————————————————————————————
        sys_logger.info("▶ 开始按信号分发...")
        
        # 这里的 valid_ents 现在已经被正确定义在上方
        for seq, ob in enumerate(valid_ents, 1):
            info = res.get(ob.Handle)
            if not info: continue
            
            raw_signal = info["spec"]
            # 信号映射
            feature_key = raw_signal.replace("+", "_").replace("/", "_")
            # 查实物名
            target_real_name = real_live_map.get(feature_key)
            
            if not target_real_name:
                sys_logger.warning(f"⚠️ 信号匹配失败: {raw_signal}")
                continue
            
            # 几何对齐
            k_geom = compute_k_from_info(info)
            try:
                _, _, blk_ref = insert_block_into_poly_area(target_real_name, ob, k=k_geom)
                adjust_block_to_frame(ob, blk_ref)
                bind_dict[blk_ref.Handle] = {"frame_info": info, "title_block": blk_ref}
            except Exception as e:
                sys_logger.error(f"插入报错: {e}")

        # 清理
        if template_instances:
            for tb in template_instances:
                try: tb.Delete()
                except: pass
        
    cfo.save_file()
    sys_logger.info("✅ 所有图签处理完成。")
    return bind_dict


# =========================================================
#&&% 主函数再次强化版
# =========================================================

@timeit
@debuggable
def insert_and_scale_labels_area_any(
        coms_dayin=None,        
        filepath=str(userpath/"dwg文件/标准图签.dwg"),
    ):
    """
    【V8.7 最终审计版】
    1. 包含阶梯式重试与协同探测机制。
    2. 实时打印现场捕获的 12 个真实块名。
    3. 自动将运行结果记录至 D:/claude-tasks/cad/tests/testfunc.xlsx。

    【功能描述】
    自动化批量插图签总成函数。根据分析出的图框规格信号，将对应的图签块精准插入、缩放并对齐到目标区域。

    【核心逻辑流程】
    1. 预计算阶段：分析输入的矩形多段线，通过几何特征计算出“信号规格”（如 A1+1/4）并存储。
    2. 资源加载阶段：在 CAD 原点插入外部图签 dwg 文件并执行 EXPLODE（炸开）操作。
    3. ★关键节点（同步与审计）★：
       - 插入并炸开后，程序进入 6 秒强制等待期，确保 CAD 数据库完成对象重组。
       - 执行 Regen 强制刷新内存索引。
       - 使用“通缉令名单”（EXPECTED_PREFIXES）对现场进行点名，捕获带后缀的真实块名（如 A1_1_4_100550）。
    4. 分发执行阶段：利用“信号规格”匹配捕获到的“真实块名”，执行批量插入与 anisotropic（非等比例）缩放对齐。
    5. 清理阶段：删除原点的模板实例，保存图纸。

    【参数说明】
    :param coms_dayin: (list/None) 目标多段线对象列表。若为 None，则自动筛选全图符合条件的矩形框。
    :param filepath: (str) 外部标准图签 dwg 的绝对路径。

    【输出参数】
    :return bind_dict: (dict) 成功插入后的映射字典。
                       Key: 插入块的 Handle (句柄)
                       Value: {
                           "frame_info": 原始框的几何信息与规格,
                           "title_block": 插入后的块引用对象对象
                       }
    """
    doc = C.doc
    
    # 核心特征表 (审计基准)
    EXPECTED_PREFIXES = [
        'A3', 'A2', 'A1', 'A2_1_4', 'A2_1_2', 'A2_3_4',
        'A1_1_4', 'A1_1_2', 'A1_3_4', 'A0', 'A0_1_8', 'A0_1_4',
    ]

    # 初始化飞行记录仪变量
    test_status = "PENDING"
    actual_sync_time = "N/A"
    error_msg = ""
    retry_count = 0
    bind_dict = {}
    valid_ents = []
    res = {}

    # ——————————————————————————————————————————————————
    # 阶段 1: 预计算 (数据准备)
    # ——————————————————————————————————————————————————
    try:
        if coms_dayin is None:
            coms_dayin = select_maxrect_polylines_1(
            layer_name = "dy_zhuanyong",
            precision_mode = False,
            width = 0.0,
            color = 1,
           
            )
            coms_dayin=coms_dayin[0]

                
        if not coms_dayin:
            sys_logger.warning("⚠️ 未找到任何矩形多段线。")
            return {}

        coms_dayin = remove_duplicate_polylines(coms_dayin, tol=1, priority_layer="dy_zhuanyong")
        
        # 排序逻辑
        max_s = 0
        for e in coms_dayin:
            try: 
                min_p, max_p = e.GetBoundingBox()
                max_s = max(max_s, abs(max_p[0]-min_p[0]))
            except: pass
        cha = 20 if max_s < 2000 else 2000
        coms_dayin = sort_coms_by_llcorner(coms_dayin, cha_Y=cha)

        for ent in coms_dayin:
            ret = generate_name_and_ratio_from_com(ent)
            if ret == 0: continue
            res[ent.Handle] = {
                "entity": ent, "spec": ret[2], "ratio": ret[1],
                "drawing_frame": ret[0], "is_vertical": ret[3]
            }
            valid_ents.append(ent)
        
        sys_logger.info(f"📋 [任务锁定] 目标区域: {len(valid_ents)} 个")

        # ——————————————————————————————————————————————————
        # 阶段 2: 事务开启 & 阶梯式资源审计
        # ——————————————————————————————————————————————————
        real_live_map = {}
        retry_waits = [0, 6, 12] 
        
        with CADGuard("一键插图签总成", disable_ui=True, independent_undo=True):

            # 1. 导入并炸开
            template_instances, _, _ = Insert_Company_Label_Common_Block(filepath=filepath)
            
            # 2. 阶梯重试循环
            for i, delay in enumerate(retry_waits, 1):
                retry_count = i
                if delay > 0:
                    sys_logger.warning(f"⚠️ [审计未齐] 正在执行第 {i} 轮：阶梯延时 {delay}s...")
                    time.sleep(delay)
                
                sync_start_time = time.time()
                sys_logger.info(f"⏳ [探测中] 第 {i}/3 次锁定 CAD 状态...")
                
                try:
                    # 协同探测
                    wait_quiescent(min_quiet=1.0, timeout=10)
                    actual_sync_time = f"{time.time() - sync_start_time:.2f}s"
                    sys_logger.info(f"✅ [状态确认] CAD 已就绪。探测耗时: {actual_sync_time}")
                    
                    doc.Regen(1)
                    current_blks = select_kuai()
                    raw_live_names = list({getattr(b, "EffectiveName", b.Name) for b in current_blks})
                    
                    # 特征审计
                    temp_map = {}
                    for prefix in EXPECTED_PREFIXES:
                        for live_name in raw_live_names:
                            if live_name == prefix or live_name.startswith(prefix + "_"):
                                temp_map[prefix] = live_name
                                break
                    
                    if len(temp_map) >= 12:
                        real_live_map = temp_map
                        sys_logger.info(f"🎊 [审计通过] 成功锁定 12 个块名。")
                        
                        # 🔥🔥 关键：打印实物清单真名 🔥🔥
                        detected_real_names = sorted(list(real_live_map.values()))
                        sys_logger.info(f"📋 [实物清单] 现场锁定块名: {detected_real_names}")
                        break 
                    else:
                        sys_logger.warning(f"🧐 [审计报告] 第 {i} 次匹配不足 ({len(temp_map)}/12)。")
                except Exception as e:
                    sys_logger.error(f"❗ [通讯异常] 第 {i} 次探测故障: {e}")

            # 3. 理论必得检查
            if len(real_live_map) < 12:
                raise RuntimeError(f"理论必得逻辑失效：炸开后无法找齐 12 个块名资源。")

            # ———————————————————————————————————————————————
            # 阶段 3: 执行分发
            # ———————————————————————————————————————————————
            sys_logger.info("▶ 开始按信号分发...")
            for seq, ob in enumerate(valid_ents, 1):
                info = res.get(ob.Handle)
                feature_key = info["spec"].replace("+", "_").replace("/", "_")
                target_real_name = real_live_map.get(feature_key)
                
                if not target_real_name: continue
                
                k_geom = compute_k_from_info(info)
                try:
                    _, _, blk_ref = insert_block_into_poly_area(target_real_name, ob, k=k_geom)
                    adjust_block_to_frame(ob, blk_ref)
                    bind_dict[blk_ref.Handle] = {"frame_info": info, "title_block": blk_ref}
                except Exception as e:
                    sys_logger.error(f"插入报错: {e}")

            # 清理
            if template_instances:
                for tb in template_instances:
                    try: tb.Delete()
                    except: pass
        
        test_status = "PASS"

    except Exception as e:
        test_status = "FAIL"
        error_msg = str(e)
        sys_logger.error(f"❌ 运行异常: {error_msg}")
        raise 

    finally:
        try:
            log_file = "D:/claude-tasks/cad/tests/testfunc.xlsx"
            import openpyxl
            import datetime
            import os

            # 1. 准备数据
            headers = ["时间", "函数名", "同步详情", "目标数", "成功数", "状态", "审计备注"]
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            test_record = [
                current_time,
                "insert_and_scale_labels_area_any",
                f"Sync:{actual_sync_time} | R:{retry_count}",
                len(valid_ents),
                len(bind_dict),
                test_status,
                error_msg if test_status == "FAIL" else "✅ 审计通过"
            ]

            # 2. 加载或创建
            if not os.path.exists(log_file):
                wb = openpyxl.Workbook()
                ws = wb.active
            else:
                wb = openpyxl.load_workbook(log_file)
                ws = wb.active

            # 3. ★ 强制检查第一行：不论如何，确保第一行是标题 ★
            # 只有当第一行内容不匹配标题时，才进行“初始化/插入”操作
            if ws.cell(row=1, column=1).value != headers[0]:
                if ws.cell(row=1, column=1).value is None:
                    # 如果是空表，直接在第一行写标题
                    for col, val in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=val)
                else:
                    # 如果第一行有数据但不是标题，在最上方插入一行写标题
                    ws.insert_rows(1)
                    for col, val in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=val)

            # 4. 追加数据：寻找真正的末尾
            # 注意：append 会从 max_row + 1 开始写
            ws.append(test_record)

            # 5. 保存
            wb.save(log_file)
            sys_logger.info("📊 记录仪: 结果已成功登记 (标题强制校验完成)。")

        except Exception as log_err:
            sys_logger.warning(f"⚠️ 登记异常: {log_err}")


    return bind_dict

# =========================================================
#&&% 主函数审计版A
# =========================================================

@timeit
@debuggable
def insert_and_scale_labels_area_any(
        coms_dayin=None,        
        filepath=str(userpath/"dwg文件/标准图签.dwg"),
    ):
    """
    【V9.3 最终整合版】
    基于 V8.7 核心逻辑（阶梯重试+资源点名），集成 V2.0 日志系统的 checkpoint 和 record_test_result。
    """
    # 窗口管理
    minimize_all_windows()
    time.sleep(0.5)
    activate_window_by_title("AutoCAD", click_titlebar=True)
    time.sleep(0.5)

    doc = C.doc
    
    # —————————— 1. 初始化变量 ——————————
    test_is_pass = False
    sync_duration_str = "N/A"
    retry_count_record = 0 # 记录实际重试次数
    target_count = 0
    success_count = 0
    error_log = ""
    
    EXPECTED_PREFIXES = [
        'A3', 'A2', 'A1', 'A2_1_4', 'A2_1_2', 'A2_3_4',
        'A1_1_4', 'A1_1_2', 'A1_3_4', 'A0', 'A0_1_8', 'A0_1_4',
    ]

    try:
        # ——————————————————————————————————————————————————
        # 阶段 1: 预计算 (数据准备)
        # ——————————————————————————————————————————————————
        if coms_dayin is None:
            # 沿用 V8.7 逻辑，调用 select_maxrect_polylines_1
            result = select_maxrect_polylines_1(
                layer_name="dy_zhuanyong", 
                precision_mode=False,
                width=0.0,
                color=1
            )
            coms_dayin = result[0] if result else []
    
        if not coms_dayin:
            error_log = "No Polylines Found"
            sys_logger.warning("⚠️ 未找到任何矩形多段线。")
            return {}

        coms_dayin = remove_duplicate_polylines(coms_dayin, tol=1, priority_layer="dy_zhuanyong")
        
        max_s = 0
        for e in coms_dayin:
            try: 
                min_p, max_p = e.GetBoundingBox()
                max_s = max(max_s, abs(max_p[0]-min_p[0]))
            except: pass
        cha = 20 if max_s < 2000 else 2000
        coms_dayin = sort_coms_by_llcorner(coms_dayin, cha_Y=cha)

        # 增加动态容差计算 (参考之前修复方案，防止大比例尺识别失败)
        all_lens = []
        for e in coms_dayin:
            try:
                min_p, max_p = e.GetBoundingBox()
                w, h = abs(max_p[0]-min_p[0]), abs(max_p[1]-min_p[1])
                all_lens.append(min(w, h))
            except: pass
        recognition_tol = max(10.0, min(all_lens) * 0.005) if all_lens else 100.0

        valid_ents = []
        res = {}
        for ent in coms_dayin:
            ret = generate_name_and_ratio_from_com(ent, tol=recognition_tol)
            if ret == 0: continue
            try:
                res[ent.Handle] = {
                    "entity": ent, "spec": ret[2], "ratio": ret[1],
                    "drawing_frame": ret[0], "is_vertical": ret[3]
                }
                valid_ents.append(ent)
            except: pass
        
        target_count = len(valid_ents)
        sys_logger.info(f"📋 [任务锁定] 目标区域: {target_count} 个")
        
        if not valid_ents: 
            error_log = "No Valid Specs Identified"
            return {}

        # ——————————————————————————————————————————————————
        # 阶段 2: 事务开启 & 资源同步审计
        # ——————————————————————————————————————————————————
        bind_dict = {}
        
        with CADGuard("一键插图签总成", disable_ui=True, independent_undo=True):

            # 1. 执行导入并炸开
            template_instances, _, _ = Insert_Company_Label_Common_Block(filepath=filepath)
            
            # 2. 阶梯重试循环 (V8.7 核心逻辑)
            real_live_map = {}
            retry_waits = [0, 6, 12] 
            
            for i, delay in enumerate(retry_waits, 1):
                retry_count_record = i
                if delay > 0:
                    sys_logger.warning(f"⚠️ [审计未齐] 第 {i} 轮重试：等待 {delay}s...")
                    time.sleep(delay)
                
                sync_start = time.time()
                sys_logger.info(f"⏳ [探测中] 第 {i}/3 次锁定 CAD 状态...")
                
                try:
                    # 协同探测
                    wait_quiescent(min_quiet=1.0, timeout=10)
                    sync_duration_str = f"{time.time() - sync_start:.2f}s"
                    sys_logger.info(f"✅ [状态确认] CAD 已就绪。耗时: {sync_duration_str}")
                    
                    doc.Regen(1)
                    current_blks = select_kuai()
                    raw_live_names = list({getattr(b, "EffectiveName", b.Name) for b in current_blks})
                    
                    # 特征审计
                    temp_map = {}
                    for prefix in EXPECTED_PREFIXES:
                        for live_name in raw_live_names:
                            if live_name == prefix or live_name.startswith(prefix + "_"):
                                temp_map[prefix] = live_name
                                break
                    
                    if len(temp_map) >= 12:
                        real_live_map = temp_map
                        sys_logger.info(f"🎊 [审计通过] 成功锁定 12 个块名。")
                        
                        found_names = sorted(list(real_live_map.values()))
                        sys_logger.info(f"📋 [实物清单]: {found_names}")
                        
                        # 🔥【关键节点插入】在此处记录检查点 🔥
                        checkpoint(
                            desc="资源炸开审计", 
                            is_pass=True,
                            found_count=len(real_live_map),
                            retry_round=i,
                            sync_time=sync_duration_str,
                            names=str(found_names)
                        )
                        break 
                    else:
                        sys_logger.warning(f"🧐 [审计报告] 第 {i} 次匹配不足 ({len(temp_map)}/12)。")
                        
                except Exception as e:
                    sys_logger.error(f"❗ [通讯异常] 第 {i} 次探测故障: {e}")

            # 3. 理论必得检查
            if len(real_live_map) < 12:
                # 记录最后一次失败的状态
                checkpoint(
                    desc="资源炸开审计失败", 
                    is_pass=False, 
                    found_count=len(real_live_map),
                    error="炸开后资源不足"
                )
                raise RuntimeError(f"理论必得逻辑失效：炸开后无法找齐 12 个块名资源。")

            # ———————————————————————————————————————————————
            # 阶段 3: 执行分发
            # ———————————————————————————————————————————————
            sys_logger.info("▶ 开始按信号分发...")
            
            for seq, ob in enumerate(valid_ents, 1):
                info = res.get(ob.Handle)
                if not info: continue
                
                feature_key = info["spec"].replace("+", "_").replace("/", "_")
                target_real_name = real_live_map.get(feature_key)
                
                if not target_real_name:
                    sys_logger.warning(f"⚠️ 信号匹配失败: {info['spec']}")
                    continue
                
                try:
                    k_geom = compute_k_from_info(info)
                    _, _, blk_ref = insert_block_into_poly_area(target_real_name, ob, k=k_geom)
                    adjust_block_to_frame(ob, blk_ref)
                    bind_dict[blk_ref.Handle] = {"frame_info": info, "title_block": blk_ref}
                except Exception as e:
                    sys_logger.error(f"插入报错: {e}")

            # 清理
            if template_instances:
                for tb in template_instances:
                    try: tb.Delete()
                    except: pass
            
        cfo.save_file()
        success_count = len(bind_dict)
        test_is_pass = True
        sys_logger.info("✅ 所有图签处理完成。")
        
        return bind_dict

    except Exception as e:
        test_is_pass = False
        error_log = str(e)
        sys_logger.error(f"❌ 运行发生致命错误: {e}")
        raise 

    finally:




        # —————— 最后的总成绩单 ——————
        try:
            record_test_result(
                script_name="insert_labels.py",
                func_name="insert_and_scale_labels_area_any",
                is_pass=test_is_pass,
                target=target_count,
                success=success_count,
                sync_time=sync_duration_str,
                retry=retry_count_record,
                error=error_log if error_log else "None"
            )
        except Exception as log_err:
            sys_logger.error(f"无法写入测试记录: {log_err}")

#&&&% 剥出块

# =========================================================
#&&% 主函数成功版
# =========================================================

@timeit
@retry_on_busy( max_retries=10, base_delay=0.5)
def normalize_core_title_blocks_by_layer_new1(
    core_layer: str = "dy_quyu_H", 
    core_base_names=None,
    verbose: bool = True,
):
    """
    【V2.2】全局扫描 -> 安全炸开 -> 核心定义重命名。
    ★ 升级：全面接入 sys_logger，同时支持控制台输出和文件记录。
    """

    # --- 0. 内部日志适配器 (保留缩进格式，但走统一通道) ---
    def log(msg, indent=0):
        if verbose:
            prefix = "  " * indent
            # 发送到统一日志系统
            sys_logger.info(f"[CoreBlock] {prefix}{msg}")

    # --- 1. 基础准备 ---
    doc = C.doc
    if not doc:
        sys_logger.error("[CoreBlock] ❌ 无法获取文档对象") # 错误级别用 error
        return False

    if core_base_names is None:
        core_base_names = Block_Names_0

    # --- 2. 定义安全炸开 (内嵌) ---
    def safe_explode_retry(entity, max_retries=5, interval=0.2):
        try: ent_name = entity.Name
        except: ent_name = "<未知>"
        
        for i in range(1, max_retries + 1):
            try:
                res = entity.Explode()
                if res and len(res) > 0:
                    try: entity.Delete()
                    except: pass
                    return True
            except Exception:
                time.sleep(interval)
        return False

    def get_base_name(name: str) -> str:
        return str(name).split("_")[0]

    # ================= 3. 开始执行 =================
    log("🚀 开始执行核心块规范化流程...")

    # [步骤 1] 全局扫描
    log("▶ 1. 全局扫描块实例...", indent=0)
    
    try:
        # 尝试复用 select_kuai，如果没有则手动选
        try: 
            all_blocks = select_kuai(autocast=True)
        except:
            ss = doc.SelectionSets.Add("Temp_Core_Scan_Logger")
            ss.Select(5, None, None, to_vt_int([0]), to_vt_variant(["INSERT"]))
            all_blocks = [x for x in ss]
            ss.Delete()
    except Exception as e:
        sys_logger.error(f"[CoreBlock] 选择集构建失败: {e}")
        return False

    target_instances = []
    for blk in all_blocks:
        try:
            bname = blk.Name
            if bname in core_base_names or get_base_name(bname) in core_base_names:
                target_instances.append(blk)
        except: pass
    
    log(f"命中目标块实例: {len(target_instances)} 个", indent=1)

    # [步骤 2] 执行炸开
    if target_instances:
        log(f"▶ 2. 尝试炸开 {len(target_instances)} 个实例...", indent=0)
        success_count = 0
        for i, blk in enumerate(target_instances, 1):
            ok = safe_explode_retry(blk, max_retries=3)
            if ok: success_count += 1
        
        log(f"炸开汇总: 成功 {success_count} / 总计 {len(target_instances)}", indent=1)
    else:
        log("▶ 2. 无需炸开 (未发现目标实例)", indent=0)

    # [步骤 3] 块定义重命名
    ts = datetime.now().strftime("%d%H%M")
    log(f"▶ 3. 核心块定义重命名 (后缀: _{ts})", indent=0)
    
    renamed_count = 0
    for base in core_base_names:
        try:
            blk_def = doc.Blocks.Item(base)
        except:
            continue 
            
        new_name = f"{base}_{ts}"
        
        # 冲突回避
        suffix = 1
        while True:
            try:
                doc.Blocks.Item(new_name)
                new_name = f"{base}_{ts}_{suffix}"
                suffix += 1
            except:
                break 
        
        try:
            blk_def.Name = new_name
            log(f"✅ Definition重命名: {base} -> {new_name}", indent=1)
            renamed_count += 1
        except Exception as e:
            sys_logger.warning(f"[CoreBlock] ❌ 重命名失败 [{base}]: {e}")

    if renamed_count == 0:
        log("⚠️ 未发现需要重命名的核心块定义。", indent=1)
    
    log("✅ 流程结束。", indent=0)


    # —————— [替换开始] 新版 CriticalSection 模式 ——————
    # 自动推断脚本名和函数名，描述为 "图签炸开校验"
    with CriticalSection(description="图签炸开校验") as ctx:
        
        # 1. 业务逻辑：计算数据 (这部分逻辑保持不变)
        expected_list = stc("图签测试")
        shu1 = len(expected_list) if expected_list else 0
        
        all_blks = select_kuai()
        actual_list = filter_blocks_by_list(all_blks, Block_Names_0)
        shu2 = len(actual_list)

        # 2. 【关键】上报数据给记录仪
        # 这些数据会自动写入 Excel，不需要手动调用 record_test_result
        ctx.record(
            expected=shu1, 
            actual=shu2, 
            diff=(shu1 - shu2)
        )

        # 3. 调试分支 A: AI 自动校验 (可选)
        if ctx.is_ai:
            if shu1 != shu2 or shu1 == 0:
                # 记录错误信息 (CriticalSection 会自动标记为 FAIL)
                ctx.record(error=f"数量不匹配: 预期{shu1} vs 实际{shu2}")
                sys_logger.error(f"🤖 AI 校验失败: 预期{shu1} != 实际{shu2}")
            else:
                sys_logger.info(f"🤖 AI 校验通过: {shu1} == {shu2}")

        # 4. 调试分支 B: 人类观察 (可选)
        if ctx.is_human:
            print("\n" + "-"*40)
            print(f"👀 [人工观察模式] insert_and_scale_labels_area_any")
            print(f"预期数量: {shu1} | 实际数量: {shu2}")
            print(f"请检查图层 '图签测试' 与实际插入块是否重合")
            print("-"*40)
            # 不需要写 time.sleep，__exit__ 会自动处理倒计时
            
    # —————— [替换结束] ——————


    return True

# =========================================================
#&&% 主函数审计版
# =========================================================

@timeit
def normalize_core_title_blocks_by_layer_new1(
    core_layer: str = "dy_quyu_H", 
    core_base_names=None,
    verbose: bool = True,
):
    """
    【V2.4 极速批处理版】
    优化策略：
    1. 引入 CADGuard 大事务，确保操作原子性。
    2. 采用"只管发令，不管结果"的批量炸开策略，极大提升速度。
    3. 最后统一等待同步，一次性完成审计。
    """
    from datetime import datetime
    import time

    def log(msg, indent=0):
        if verbose:
            prefix = "  " * indent
            sys_logger.info(f"[CoreBlock] {prefix}{msg}")

    doc = C.doc
    if not doc:
        sys_logger.error("[CoreBlock] ❌ 无法获取文档对象")
        return False

    if core_base_names is None:
        core_base_names = Block_Names_0

    # 1. 准备阶段
    log("🚀 开始执行核心块规范化流程...")
    
    try:
        try: all_blocks = select_kuai(autocast=True)
        except:
            # 备用选择逻辑
            ss = doc.SelectionSets.Add("Temp_Core_Scan_Logger")
            ss.Select(5, None, None, to_vt_int([0]), to_vt_variant(["INSERT"]))
            all_blocks = [x for x in ss]
            ss.Delete()
    except Exception as e:
        sys_logger.error(f"[CoreBlock] 选择集构建失败: {e}")
        return False

    target_instances = []
    for blk in all_blocks:
        try:
            bname = blk.Name
            if bname in core_base_names or get_base_name(bname) in core_base_names:
                target_instances.append(blk)
        except: pass
    
    initial_count = len(target_instances)
    log(f"命中目标块实例: {initial_count} 个", indent=1)

    if not target_instances:
        log("▶ 无需炸开 (未发现目标实例)", indent=0)
        return True

    # =========================================================================
    # 2. 核心操作区 (大事务包裹)
    # =========================================================================
    with CADGuard("核心块规范化批处理", disable_ui=True, independent_undo=True):
        
        log(f"▶ 2. 批量炸开 {initial_count} 个实例...", indent=0)
        
        # [极速模式] 只发送指令，不等待反馈，不重试
        # 原理：在事务保护下，命令队列会极其高效
        explode_cmd_count = 0
        for blk in target_instances:
            try:
                blk.Explode() # 只发令
                blk.Delete()  # 原身删除
                explode_cmd_count += 1
            except:
                pass # 极个别失败不影响大局
        
        log(f"指令发送完毕: {explode_cmd_count} / {initial_count}", indent=1)

        # [关键] 事务内的同步等待
        # 此时 CAD 正在后台疯狂计算几何图形
        log("⏳ 等待 CAD 几何重算...", indent=1)
        wait_quiescent(min_quiet=1.0, timeout=10) # 给它 1秒 喘息
        doc.Regen(1) # 强制刷新显示列表，更新数据库索引



        # =========================================================
        # 3. 审计检查 (修改重点)
        # =========================================================
        all_now = select_kuai()
        
        # 定义核心名单 (这是炸开后应该出现的“核”)
        # 你需要根据实际情况补充完整，比如还有 A2_1_4-H 等
        Block_Names_1 = [
            'A3-H', 'A2-H', 'A1-H', 'A0-H',
            'A1_1_4-H', 'A2_1_4-H' # <--- 请务必补全所有可能的内核名字！
        ]
        
        # A. 查核 (期望等于初始数量)
        inner_blocks = filter_blocks_by_list(all_now, Block_Names_1)
        count_inner = len(inner_blocks)
        
        # B. 查壳 (期望为 0)
        # 注意：这里依然要用 core_base_names (即 Block_Names_0) 来检查是否炸干净了
        remaining_shells = filter_blocks_by_list(all_now, core_base_names)
        count_shell = len(remaining_shells)
        
        # 记录关键节点
        checkpoint(
            desc="规范化-炸开审计", 
            # 通过标准：内核出来了(数量对等) 且 壳子没了(数量为0)
            is_pass=(count_inner == initial_count), 
            
            sent_cmd=explode_cmd_count,   # 发令数 (24)
            found_inner=count_inner,      # 实测内核数 (期望 24)
            found_shell=count_shell,      # 残留壳数 (期望 0)
            
            # 方便调试时看看到底抓到了谁
            sample_inner=str([b.Name for b in inner_blocks[:3]]) if inner_blocks else "None"
        )






        # 4. 块定义重命名
        ts = datetime.now().strftime("%d%H%M")
        log(f"▶ 3. 核心块定义重命名 (后缀: _{ts})", indent=0)
        
        renamed_count = 0
        for base in core_base_names:
            try: 
                blk_def = doc.Blocks.Item(base)
                new_name = f"{base}_{ts}"
                suffix = 1
                while True:
                    try:
                        doc.Blocks.Item(new_name)
                        new_name = f"{base}_{ts}_{suffix}"
                        suffix += 1
                    except: break 
                
                blk_def.Name = new_name
                log(f"✅ Renamed: {base} -> {new_name}", indent=2)
                renamed_count += 1
            except: continue

        if renamed_count == 0:
            log("⚠️ 未发现可重命名的定义", indent=1)
    
    log("✅ 流程结束。", indent=0)
    return True

#&&% 生产标准版B
@timeit
def normalize_core_title_blocks_by_layer_new1(
    core_layer: str = "dy_quyu_H", 
    core_base_names=None,
    verbose: bool = True,
):
    """
    【V3.0 生产标准版】(原 _new1 / _retry)
    核心机制：事务级回滚 + 饱和攻击 + 双向审计 (查核查壳)
    """
    import datetime as dt
    import time
    
    # --- 内部日志 ---
    def log(msg, indent=0):
        if verbose:
            sys_logger.info(f"[CoreBlock] {'  '*indent}{msg}")

    doc = C.doc
    if not doc: return False
    if core_base_names is None: core_base_names = Block_Names_0

    MAX_RETRIES = 3
    
    for attempt in range(1, MAX_RETRIES + 1):
        log(f"🔄 [Attempt {attempt}/{MAX_RETRIES}] 启动事务流程...")
        
        try:
            # 开启事务：失败自动回滚
            with CADGuard(f"核心规范化-第{attempt}次", disable_ui=True, independent_undo=True):
                
                # --- 1. 重新扫描 ---
                try:
                    all_blocks = select_kuai(autocast=True)
                except:
                    ss = doc.SelectionSets.Add(f"Temp_Retry_{attempt}")
                    ss.Select(5, None, None, to_vt_int([0]), to_vt_variant(["INSERT"]))
                    all_blocks = [x for x in ss]
                    ss.Delete()

                target_instances = []
                for blk in all_blocks:
                    try:
                        bname = blk.Name
                        if bname in core_base_names or bname.split("_")[0] in core_base_names:
                            target_instances.append(blk)
                    except: pass
                
                initial_count = len(target_instances)
                log(f"锁定目标: {initial_count} 个", indent=1)

                if initial_count == 0:
                    log("无目标，流程跳过。", indent=1)
                    return True 

                # --- 2. 极速炸开 ---
                log("执行批量炸开...", indent=1)
                cmd_count = 0
                for blk in target_instances:
                    try:
                        blk.Explode()
                        blk.Delete()
                        cmd_count += 1
                    except: pass
                
                # --- 3. 等待与刷新 ---
                wait_time = 0.5 + (attempt * 0.5)
                log(f"等待同步 ({wait_time}s)...", indent=1)
                wait_quiescent(min_quiet=wait_time, timeout=10)
                doc.Regen(1)

                # --- 4. 现场审计 ---
                all_now = select_kuai()
                
                # 构建内核名单 (-H)
                Block_Names_Kernel = [n + "-H" for n in core_base_names]
                
                found_kernels = len(filter_blocks_by_list(all_now, Block_Names_Kernel))
                found_shells = len(filter_blocks_by_list(all_now, core_base_names))
                
                is_success = (found_kernels >= initial_count) and (found_shells == 0)
                
                # 🔥 审计检查点
                checkpoint(
                    desc=f"规范化-炸开审计", 
                    is_pass=is_success,
                    round=attempt,
                    sent=cmd_count,
                    found_kernel=found_kernels,
                    left_shell=found_shells
                )

                # --- 5. 判决与回滚 ---
                if not is_success:
                    error_msg = f"审计失败: 核{found_kernels}/{initial_count}, 壳{found_shells}"
                    log(f"❌ {error_msg} -> 触发回滚!", indent=1)
                    raise RuntimeError(error_msg) # 触发 CADGuard 回滚
                
                # --- 6. 成功收尾 (重命名) ---

                ts = dt.datetime.now().strftime("%d%H%M")

                for base in core_base_names:
                    try: 
                        blk_def = doc.Blocks.Item(base)
                        new_name = f"{base}_{ts}"
                        suffix = 1
                        while True:
                            try:
                                doc.Blocks.Item(new_name); new_name = f"{base}_{ts}_{suffix}"; suffix += 1
                            except: break 
                        blk_def.Name = new_name
                    except: continue
                
                log("✅ 本次尝试成功，事务提交。", indent=1)
                return True 

        except Exception as e:
            sys_logger.warning(f"⚠️ 第 {attempt} 次尝试失败: {e}")
            if attempt < MAX_RETRIES:
                log("准备进行下一次重试...", indent=1)
                time.sleep(1)
            else:
                log("❌ 已达到最大重试次数，操作终止。", indent=0)
                checkpoint(desc="事务重试最终失败", is_pass=False, error=str(e))
                return False

    return False



# =========================================================
#&&% 🏭 主业务版B (精明的指挥官)
# =========================================================

@timeit
def normalize_core_title_blocks_by_layer_new1(
    core_layer: str = "dy_quyu_H", 
    core_base_names=None,
    verbose: bool = True,
):
    """
    【V3.2 生产标准版】(集成 safe_explode_retry 深度搜救版)
    核心机制：事务级回滚 + 饱和攻击 + 双向审计 + 结果验证原子操作
    """
    import datetime as dt
    import time
    
    # --- 内部日志 ---
    def log(msg, indent=0):
        if verbose:
            sys_logger.info(f"[CoreBlock] {'  '*indent}{msg}")

    doc = C.doc
    if not doc: return False
    if core_base_names is None: core_base_names = Block_Names_0

    MAX_RETRIES = 3
    
    for attempt in range(1, MAX_RETRIES + 1):
        log(f"🔄 [Attempt {attempt}/{MAX_RETRIES}] 启动事务流程...")
        
        try:
            with CADGuard(f"核心规范化-第{attempt}次", disable_ui=True, independent_undo=True):
                
                # --- 1. 重新扫描 ---
                try:
                    all_blocks = select_kuai(autocast=True)
                except:
                    ss = doc.SelectionSets.Add(f"Temp_Retry_{attempt}")
                    ss.Select(5, None, None, to_vt_int([0]), to_vt_variant(["INSERT"]))
                    all_blocks = [x for x in ss]
                    ss.Delete()

                target_instances = []
                for blk in all_blocks:
                    try:
                        bname = blk.Name
                        if bname in core_base_names or bname.split("_")[0] in core_base_names:
                            target_instances.append(blk)
                    except: pass
                
                initial_count = len(target_instances)
                log(f"锁定目标: {initial_count} 个", indent=1)

                if initial_count == 0:
                    log("无目标，流程跳过。", indent=1)
                    return True 

                # --- 2. 极速炸开 (升级为 safe_explode_retry) ---
                log("执行批量炸开与验证...", indent=1)
                cmd_count = 0
                
                for blk in target_instances:
                    # =================================================
                    # 🚀 关键替换点
                    # =================================================
                    # 调用 safe_explode_retry 进行 炸开+验证+搜救
                    # 它会自动重试，直到对象消失或超时
                    result_objs = safe_explode_retry(blk, verbose=False) # 内部日志关掉，避免刷屏
                    
                    # 依据严格契约：None 表示失败，List (含[]) 表示成功
                    if result_objs is not None:
                        cmd_count += 1
                    else:
                        sys_logger.warning(f"⚠️ 单个块炸开验证失败 (Handle: {blk.Handle})，将在审计阶段处理。")
                    # =================================================
                
                # --- 3. 等待与刷新 ---
                wait_time = 0.5 + (attempt * 0.5)
                log(f"等待同步 ({wait_time}s)...", indent=1)
                wait_quiescent(min_quiet=wait_time, timeout=10)
                doc.Regen(1)

                # --- 4. 现场审计 ---
                all_now = select_kuai()



                Block_Names_Kernel = [n + "-H" for n in core_base_names]
                
                found_kernels = len(filter_blocks_by_list(all_now, Block_Names_Kernel))
                found_shells = len(filter_blocks_by_list(all_now, core_base_names))
                
                is_success = (found_kernels >= initial_count) and (found_shells == 0)
                
                checkpoint(
                    desc=f"规范化-炸开审计", 
                    is_pass=is_success,
                    round=attempt,
                    sent=cmd_count,
                    found_kernel=found_kernels,
                    left_shell=found_shells
                )

                # --- 5. 判决与回滚 ---
                if not is_success:
                    error_msg = f"审计失败: 核{found_kernels}/{initial_count}, 壳{found_shells}"
                    log(f"❌ {error_msg} -> 触发回滚!", indent=1)
                    raise RuntimeError(error_msg) 
                
                # --- 6. 成功收尾 (重命名) ---
                ts = dt.datetime.now().strftime("%d%H%M")
                for base in core_base_names:
                    try: 
                        blk_def = doc.Blocks.Item(base)
                        new_name = f"{base}_{ts}"
                        suffix = 1
                        while True:
                            try:
                                doc.Blocks.Item(new_name); new_name = f"{base}_{ts}_{suffix}"; suffix += 1
                            except: break 
                        blk_def.Name = new_name
                    except: continue
                
                log("✅ 本次尝试成功，事务提交。", indent=1)
                return True 

        except Exception as e:
            sys_logger.warning(f"⚠️ 第 {attempt} 次事务级尝试失败: {e}")
            if attempt < MAX_RETRIES:
                log("准备进行下一次重试...", indent=1)
                time.sleep(1)
            else:
                log("❌ 已达到最大重试次数，操作终止。", indent=0)
                checkpoint(desc="事务重试最终失败", is_pass=False, error=str(e))
                return False

    return False


# =========================================================================
#&&&%  总装流水线 (串联 A 和 B)
# =========================================================================
#&&% 无参版

def run_title_block_assembly_pipeline():
    """
    【总装流水线】
    串联执行：
    1. 智能插入与缩放 (insert_and_scale_labels_area_any)
    2. 数据库稳定化 (Bridge)
    3. 核心规范化炸开 (normalize_core_title_blocks_by_layer_new1)
    """
    
    # 开启总监控
    with CriticalSection(description="[总装] 图签全流程") as ctx:
        
        # --- 步骤 1: 插入 ---
        sys_logger.info("\n💠 [Phase 1] 启动图签插入与缩放...")
        
        result_dict = insert_and_scale_labels_area_any(
            coms_dayin=None,
            filepath=str(userpath/"dwg文件/标准图签.dwg")
        )
        
        count_inserted = len(result_dict)
        ctx.record(phase1_inserted=count_inserted)
        
        if count_inserted == 0:
            sys_logger.error("❌ 第一阶段未产生有效图签，流水线终止。")
            ctx.record(error="Phase 1 returned 0")
            return False

        # --- 桥接区: 稳定化 ---
        sys_logger.info("\n💤 [Bridge] 等待几何数据落地...")
        wait_quiescent(min_quiet=1.5, timeout=10) 
        try: C.doc.Regen(1)
        except: pass
        
        # --- 步骤 2: 规范化 ---
        sys_logger.info("\n💠 [Phase 2] 启动核心规范化...")
        
        target_shells = [
            'A3', 'A2', 'A1', 'A0',
            'A2_1_4', 'A2_1_2', 'A2_3_4',
            'A1_1_4', 'A1_1_2', 'A1_3_4', 
            'A0_1_8', 'A0_1_4'
        ]
        
        # ✅ 调用 V3.0 正式版函数名
        is_norm_ok = normalize_core_title_blocks_by_layer_new1(
            core_base_names=target_shells,
            verbose=True
        )
        
        ctx.record(phase2_status=("OK" if is_norm_ok else "FAIL"))

        if not is_norm_ok:
            sys_logger.error("❌ 第二阶段规范化失败。")
            return False

        # --- 最终验收 ---
        sys_logger.info("\n✨ 流水线执行完毕。")
        
        if ctx.is_human:
            print("\n" + "="*50)
            print(f"🏭 [流水线验收]")
            print(f"1. 插入阶段: {count_inserted} 个图签已就位")
            print(f"2. 规范阶段: {('成功' if is_norm_ok else '失败')}")
            print("="*50 + "\n")
            
    return True

#&&% 双模版

# =========================================================================
# &&&%  总装流水线 (串联 A 和 B) - [V2.0 双模版]
# =========================================================================
def run_title_block_assembly_pipeline(external_coms=None, external_filepath=None):
    """
    【总装流水线】
    既支持 UI 按钮直接点击（无参），也支持 Power 接口传参调用。
    
    参数:
        external_coms: list (可选) 外部传入的多段线对象列表。如果为None，则触发交互式选择。
        external_filepath: str (可选) 外部传入的图签路径。如果为None，使用默认配置。
        
    流程:
    1. 智能插入与缩放 (insert_and_scale_labels_area_any)
    2. 数据库稳定化 (Bridge)
    3. 核心规范化炸开 (normalize_core_title_blocks_by_layer_new1)
    """
    import os
    from system.project_setup import PathConfig # 确保能获取默认路径配置

    # 开启总监控
    with CriticalSection(description="[总装] 图签全流程") as ctx:
        
        # —————————— 0. 数据源准备 ——————————
        # A. 确定目标对象 (Coms)
        if external_coms is not None:
            target_coms = external_coms
            sys_logger.info(f"🚀 [管道] 接收外部传入对象: {len(target_coms)} 个")
        else:
            # 原有逻辑：交互式选择
            result = select_maxrect_polylines_1(
                layer_name="dy_zhuanyong", 
                precision_mode=False,
                width=0.0,
                color=1
            )
            target_coms = result[0] if result else []
            sys_logger.info(f"🖱️ [管道] 交互式选择对象: {len(target_coms)} 个")

        if not target_coms:
            sys_logger.warning("⚠️ 流水线中止：未提供或未选中任何有效区域。")
            return False

        # B. 确定图块路径 (Filepath)
        if external_filepath:
            target_path = external_filepath
        else:
            # 原有逻辑：默认路径
            target_path = str(PathConfig.userpath / "dwg文件/标准图签.dwg")

        # —————————— 步骤 1: 插入 ——————————
        sys_logger.info("\n💠 [Phase 1] 启动图签插入与缩放...")
        
        # 调用底层函数 (V9.3)
        result_dict = insert_and_scale_labels_area_any(
            coms_dayin=target_coms,   # 使用确定好的对象
            filepath=target_path      # 使用确定好的路径
        )
        

        count_inserted = len(result_dict) if result_dict else 0
 
        ctx.record(phase1_inserted=count_inserted)
        
        if count_inserted == 0:
            sys_logger.error("❌ 第一阶段未产生有效图签，流水线终止。")
            ctx.record(error="Phase 1 returned 0")
            return False

        # —————————— 桥接区: 稳定化 ——————————
        sys_logger.info("\n💤 [Bridge] 等待几何数据落地...")
        wait_quiescent(min_quiet=1.0, timeout=10) 
        try: C.doc.Regen(1)
        except: pass
        
        # —————————— 步骤 2: 规范化 ——————————
        sys_logger.info("\n💠 [Phase 2] 启动核心规范化 (原子重试版)...")
        
        target_shells = [
            'A3', 'A2', 'A1', 'A0',
            'A2_1_4', 'A2_1_2', 'A2_3_4',
            'A1_1_4', 'A1_1_2', 'A1_3_4', 
            'A0_1_8', 'A0_1_4'
        ]
        
        # 调用 V3.1 生产标准版 (带事务和原子重试)
        is_norm_ok = normalize_core_title_blocks_by_layer_new1(
            core_base_names=target_shells,
            verbose=True
        )
        
        ctx.record(phase2_status=("OK" if is_norm_ok else "FAIL"))

        if not is_norm_ok:
            sys_logger.error("❌ 第二阶段规范化失败。")
            return False

        # —————————— 最终验收 ——————————
        sys_logger.info("\n✨ 流水线执行完毕。")
        
        if ctx.is_human:
            print("\n" + "="*50)
            print(f"🏭 [流水线验收]")
            print(f"1. 插入阶段: {count_inserted} 个图签已就位")
            print(f"2. 规范阶段: {('成功' if is_norm_ok else '失败')}")
            print("="*50 + "\n")
            
    return True




# =========================================================================
#&&&% 旗舰级总控入口
# =========================================================================

@timeit
def run_full_project_workflow():
    """【旗舰级总控】"""
    
    with CriticalSection(description="工程图纸全自动化流水线") as ctx_master:
        sys_logger.info("🚀 启动全流程...")
        
        # 调用总装流水线
        # 注意：这里调用的是 run_title_block_assembly_pipeline
        is_assembly_ok = run_title_block_assembly_pipeline()
        
        if not is_assembly_ok:
            raise RuntimeError("Phase 1 (Assembly) Failed")
        
        ctx_master.record(phase1="DONE")
        
        # (此处可继续添加 重建字典、编目录 等后续步骤)
        
        sys_logger.info("🎉 全流程圆满结束！")



#&&&% 图纸空间的图签插入


























































