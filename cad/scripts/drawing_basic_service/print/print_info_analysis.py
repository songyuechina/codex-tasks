# -*- coding: utf-8 -*-
"""print_info_analysis.py

打印信息分析：
1. 针对指定打印区域，分析是否存在内框线。
2. 再判断内框线右下角是否存在合乎要求的图签块。
3. 图签块若为属性块，读取全部标签和值。
4. 图签块若为普通块，分析其外包盒窗口内的 CAD/T天正文字。
5. 将图纸名称、图号、项目名称等结果整理成按打印区域排序编号索引的字典。
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional
import argparse
import json
import re
import sys
import time


current = Path(__file__).resolve()
MODULE_DIR = current.parent
while current.name != "cad":
    if current.parent == current:
        raise Exception("找不到根目录 cad")
    current = current.parent
if str(current) not in sys.path:
    sys.path.insert(0, str(current))
if str(MODULE_DIR) not in sys.path:
    sys.path.insert(0, str(MODULE_DIR))

from system.licad import C
from system.CAD_core import launch_cad_guardians, li, litz, open_file
from system.CAD_com_utils import retry_on_busy
from system.CAD_coordination import wait_quiescent
from system.runtime_guard_bridge import RuntimeGuardTriggered, assert_runtime_guard_ok, render_guard_error
from system.CAD_selection import get_attr, select_entities_in_window, ss_select
from system.common_logger import sys_logger
from library.cad_annotation import get_text_content

from print_area_content_analysis import analyze_jobs_content
from print_policy import (
    PRINT_MODE_ADAPTIVE,
    PRINT_MODE_BASIC,
    PRINT_MODE_PURIFIED_ADAPTIVE,
    PrintJob,
    build_print_plan,
    normalize_print_mode,
    plan_to_dict,
    reindex_jobs_by_space,
)
from print_area_analysis import (
    _bbox_wh_from_bbox,
    _eps_from_short_side,
    get_rect_polylines_by_space,
    select_all_polylines,
)


TEXT_OBJECTS = {
    "AcDbText",
    "AcDbMText",
    "Text",
    "MText",
    "TDbText",
    "TDbMText",
}
TITLE_TAG_HINTS = ("图名", "图纸名称", "TITLE", "DWG_NAME", "SHEET_NAME")
NUMBER_TAG_HINTS = ("图纸编号", "图号", "DWG_NO", "DRAWING_NO", "SHEET_NO", "NO.")
PROJECT_TAG_HINTS = ("项目名称", "工程名称", "PROJECT", "工程名")
SUBPROJECT_TAG_HINTS = ("子项目名称", "子项名称", "单项工程", "SUBPROJECT")
DRAWING_NO_TOKEN_RE = re.compile(r"[A-Za-z0-9０-９]")
STRICT_DRAWING_NO_RE = re.compile(r"^\s*[A-Za-z]*\s*[-~_]*\s*[\d０-９]+\s*$")


EXCEL_MAIN_COLUMNS = [
    ("sequence_no", "序号"),
    ("sequence_key", "序号键"),
    ("page_key", "页面键"),
    ("layout_name", "布局"),
    ("space_kind", "空间"),
    ("print_handle", "打印区域句柄"),
    ("drawing_title", "图纸名称"),
    ("drawing_no", "图纸编号"),
    ("project_name", "项目名称"),
    ("subproject_name", "子项目名称"),
    ("drawing_title_record_count", "图纸名称对象数"),
    ("drawing_no_record_count", "图纸编号对象数"),
    ("drawing_title_handles", "图纸名称句柄"),
    ("drawing_no_handles", "图纸编号句柄"),
    ("ratio", "比例"),
    ("paper_code", "纸张代号"),
    ("media", "纸张名称"),
    ("rotation", "方向"),
    ("plot_scale", "打印缩放"),
    ("standard_flag", "标准图幅标记"),
    ("inner_frame_exists", "有无内框线"),
    ("inner_frame_handle", "内框线句柄"),
    ("inner_frame_bbox", "内框线包围盒"),
    ("graphic_info_area_bbox", "图签信息区域"),
    ("graphic_info_area_source", "图签信息区域来源"),
    ("graphic_orientation", "图签分析方向"),
    ("right_bottom_title_block_exists", "有无右下图签块"),
    ("title_block_kind", "图签块类型"),
    ("title_block_handle", "图签块句柄"),
    ("title_block_name", "图签块名称"),
    ("title_block_bbox", "图签块包围盒"),
    ("title_no_resolve_method", "编号标题判定方式"),
    ("drawing_title_candidates", "图纸名称候选"),
    ("drawing_no_candidates", "图号候选"),
    ("title_block_texts", "图签块文字"),
    ("title_block_attr_fields", "属性字段"),
    ("analysis_stop_reason", "分析终止原因"),
]


@dataclass
class BlockSnapshot:
    handle: str
    block_name: str
    bbox: tuple[float, float, float, float]
    owner_btr: str
    is_attribute_block: int
    attr_fields: dict[str, str]


@dataclass
class TextRecord:
    handle: str
    text: str
    obj_name: str
    bbox: tuple[float, float, float, float]
    layer: str
    color: int
    height: float
    rotation: float
    insertion_point: tuple[float, float, float] | None
    source_handles: tuple[str, ...] = ()


def _safe_handle(obj: Any) -> str:
    try:
        return str(get_attr(obj, "Handle", getattr(obj, "Handle", "")))
    except Exception:
        return str(id(obj))


def _bbox_xy(ent: Any) -> tuple[float, float, float, float] | None:
    try:
        p1, p2 = ent.GetBoundingBox()
        min_x = min(float(p1[0]), float(p2[0]))
        min_y = min(float(p1[1]), float(p2[1]))
        max_x = max(float(p1[0]), float(p2[0]))
        max_y = max(float(p1[1]), float(p2[1]))
        return min_x, min_y, max_x, max_y
    except Exception:
        return None


def _bbox_width(bbox: tuple[float, float, float, float]) -> float:
    return max(float(bbox[2]) - float(bbox[0]), 0.0)


def _bbox_height(bbox: tuple[float, float, float, float]) -> float:
    return max(float(bbox[3]) - float(bbox[1]), 0.0)


def _bbox_union(bboxes: Iterable[tuple[float, float, float, float]]) -> tuple[float, float, float, float] | None:
    items = [bbox for bbox in bboxes if bbox is not None]
    if not items:
        return None
    return (
        min(item[0] for item in items),
        min(item[1] for item in items),
        max(item[2] for item in items),
        max(item[3] for item in items),
    )


def _record_sort_key_reading(record: TextRecord) -> tuple[float, float]:
    return (-float(record.bbox[1]), float(record.bbox[0]))


def _record_sort_key_x(record: TextRecord) -> tuple[float, float]:
    return (float(record.bbox[0]), float(record.bbox[1]))


def _record_sort_key_bottom(record: TextRecord) -> tuple[float, float]:
    return (float(record.bbox[1]), float(record.bbox[0]))


def _text_record_to_dict(record: TextRecord) -> dict[str, Any]:
    return {
        "handle": record.handle,
        "text": record.text,
        "obj_name": record.obj_name,
        "bbox": record.bbox,
        "layer": record.layer,
        "color": record.color,
        "height": record.height,
        "rotation": record.rotation,
        "insertion_point": list(record.insertion_point) if record.insertion_point else None,
        "source_handles": list(record.source_handles),
    }


def _record_handle_list(records: list[TextRecord]) -> list[str]:
    handles: list[str] = []
    for record in records:
        for handle in (record.source_handles or (record.handle,)):
            value = str(handle or "").strip()
            if value and value not in handles:
                handles.append(value)
    return handles


def _classify_text_record_dicts(
    text_records: list[TextRecord],
    drawing_title_records: list[TextRecord],
    drawing_no_records: list[TextRecord],
) -> list[dict[str, Any]]:
    title_handles = set(_record_handle_list(drawing_title_records))
    number_handles = set(_record_handle_list(drawing_no_records))
    classified: list[dict[str, Any]] = []
    for record in text_records:
        item = _text_record_to_dict(record)
        handle = str(record.handle or "")
        if handle in title_handles:
            item["resolved_role"] = "drawing_title"
        elif handle in number_handles:
            item["resolved_role"] = "drawing_no"
        else:
            item["resolved_role"] = "other"
        classified.append(item)
    return classified


def _normalize_path(path: str | Path) -> str:
    return str(Path(path).resolve()).lower()


def _is_name_only_target(target_path: str | Path) -> bool:
    raw = str(target_path)
    return Path(raw).name.lower() == raw.lower()


def _find_document_by_path(target_path: str | Path):
    wanted = _normalize_path(target_path)
    target_name = Path(target_path).name.lower()
    allow_name_fallback = _is_name_only_target(target_path)
    try:
        for doc in C.acad.Documents:
            try:
                doc_name = str(getattr(doc, "Name", "") or "").lower()
                if _normalize_path(doc.FullName) == wanted or (allow_name_fallback and doc_name == target_name):
                    return doc
            except Exception:
                continue
    except Exception:
        return None
    return None


def _activate_document_by_path(target_path: str | Path, retries: int = 8, delay: float = 0.5) -> bool:
    wanted = _normalize_path(target_path)
    target_name = Path(target_path).name.lower()
    allow_name_fallback = _is_name_only_target(target_path)
    for _ in range(retries):
        doc = _find_document_by_path(target_path)
        if doc is None:
            time.sleep(delay)
            continue
        try:
            doc.Activate()
        except Exception:
            time.sleep(delay)
            continue
        time.sleep(delay)
        try:
            active = C.acad.ActiveDocument
            active_name = str(getattr(active, "Name", "") or "").lower()
            if active and (
                _normalize_path(active.FullName) == wanted
                or (allow_name_fallback and active_name == target_name)
            ):
                return True
        except Exception:
            pass
    return False


def _close_document_by_path(target_path: str | Path, save_changes: bool = False) -> bool:
    doc = _find_document_by_path(target_path)
    if doc is None:
        return True
    try:
        doc.Close(bool(save_changes))
        return True
    except Exception:
        return False


def _find_layout_by_owner_btr(owner_btr_name: str):
    if owner_btr_name == "*MODEL_SPACE":
        try:
            return C.doc.Layouts.Item("Model")
        except Exception:
            return None
    for layout in C.doc.Layouts:
        try:
            if str(layout.Block.Name) == owner_btr_name:
                return layout
        except Exception:
            continue
    return None


def _activate_space_for_owner(owner_btr_name: str) -> bool:
    layout = _find_layout_by_owner_btr(owner_btr_name)
    if layout is None:
        return False
    try:
        C.doc.ActiveLayout = layout
    except Exception:
        return False
    try:
        C.doc.MSpace = owner_btr_name == "*MODEL_SPACE"
    except Exception:
        pass
    time.sleep(0.2)
    return True


def _contains_bbox(
    outer: tuple[float, float, float, float],
    inner: tuple[float, float, float, float],
    tol: float,
) -> bool:
    return (
        outer[0] <= inner[0] + tol
        and outer[1] <= inner[1] + tol
        and outer[2] >= inner[2] - tol
        and outer[3] >= inner[3] - tol
    )


def _area_of_bbox(bbox: tuple[float, float, float, float]) -> float:
    return max((bbox[2] - bbox[0]) * (bbox[3] - bbox[1]), 0.0)


def _dynamic_tol_from_bbox(bbox: tuple[float, float, float, float]) -> float:
    short_side = min(bbox[2] - bbox[0], bbox[3] - bbox[1])
    return max(_eps_from_short_side(short_side), 1.0)


def _contains_chinese(text: str) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", text))


def _looks_like_drawing_no_text(text: str) -> bool:
    text = _clean_text(text)
    if not text or _contains_chinese(text):
        return False
    return bool(STRICT_DRAWING_NO_RE.match(text) or DRAWING_NO_TOKEN_RE.search(text))


def _clean_text(text: str) -> str:
    if not isinstance(text, str):
        return ""
    text = text.replace("\\P", " ").replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _clean_attr_value(value: str) -> str:
    text = _clean_text(value)
    if text.startswith("\\") and ";" in text:
        _, right = text.split(";", 1)
        text = right.strip()
    return text


def _layer_contains(layer_name: str, keyword: str) -> bool:
    return keyword in str(layer_name or "").strip()


def _get_color_index(ent: Any) -> int:
    try:
        return int(get_attr(ent, "Color", getattr(ent, "Color", 0)) or 0)
    except Exception:
        return 0


def _is_portrait_job(job: dict[str, Any], inner_bbox: tuple[float, float, float, float] | None = None) -> bool:
    bbox = inner_bbox or tuple(job["lower_left"] + job["upper_right"])
    return _bbox_height(bbox) > _bbox_width(bbox)


def _select_entities_in_bbox(
    owner_btr_name: str,
    bbox: tuple[float, float, float, float],
) -> list[Any]:
    if not _activate_space_for_owner(owner_btr_name):
        return []
    try:
        entities = select_entities_in_window(
            bbox[0],
            bbox[1],
            bbox[2],
            bbox[3],
            ty=1.0,
            select_mode="_W",
        ) or []
    except Exception:
        entities = []

    out: list[Any] = []
    seen: set[str] = set()
    for ent in entities:
        handle = _safe_handle(ent)
        if handle in seen:
            continue
        seen.add(handle)
        out.append(ent)
    return out


def _make_text_record(ent: Any) -> TextRecord | None:
    obj_name = str(get_attr(ent, "ObjectName", getattr(ent, "ObjectName", "")))
    if obj_name not in TEXT_OBJECTS:
        return None
    ent_bbox = _bbox_xy(ent)
    if ent_bbox is None:
        return None
    text = _clean_text(get_text_content(ent))
    if not text:
        return None
    layer = str(get_attr(ent, "Layer", getattr(ent, "Layer", "")))
    insertion = get_attr(ent, "InsertionPoint", None)
    if insertion is not None:
        try:
            insertion = tuple(float(v) for v in insertion[:3])
        except Exception:
            insertion = None
    return TextRecord(
        handle=_safe_handle(ent),
        text=text,
        obj_name=obj_name,
        bbox=ent_bbox,
        layer=layer,
        color=_get_color_index(ent),
        height=_bbox_height(ent_bbox),
        rotation=float(get_attr(ent, "Rotation", 0.0) or 0.0),
        insertion_point=insertion,
        source_handles=(_safe_handle(ent),),
    )


def _combine_text_values(records: list[TextRecord], *, sort_mode: str = "reading", separator: str = "") -> str:
    if not records:
        return ""
    if sort_mode == "x":
        ordered = sorted(records, key=_record_sort_key_x)
    elif sort_mode == "bottom":
        ordered = sorted(records, key=_record_sort_key_bottom)
    else:
        ordered = sorted(records, key=_record_sort_key_reading)
    return separator.join(record.text for record in ordered if record.text).strip()


def _merge_text_records(
    records: list[TextRecord],
    *,
    sort_mode: str = "reading",
    separator: str = "",
) -> TextRecord | None:
    if not records:
        return None
    if len(records) == 1:
        return records[0]
    if sort_mode == "x":
        ordered = sorted(records, key=_record_sort_key_x)
    elif sort_mode == "bottom":
        ordered = sorted(records, key=_record_sort_key_bottom)
    else:
        ordered = sorted(records, key=_record_sort_key_reading)
    merged_bbox = _bbox_union([record.bbox for record in ordered]) or ordered[0].bbox
    source_handles: list[str] = []
    for record in ordered:
        source_handles.extend(record.source_handles or (record.handle,))
    return TextRecord(
        handle=ordered[0].handle,
        text=_combine_text_values(ordered, sort_mode=sort_mode, separator=separator),
        obj_name=ordered[0].obj_name,
        bbox=merged_bbox,
        layer=ordered[0].layer,
        color=ordered[0].color,
        height=max(record.height for record in ordered),
        rotation=ordered[0].rotation,
        insertion_point=ordered[0].insertion_point,
        source_handles=tuple(dict.fromkeys(source_handles)),
    )


def _vertical_gap_between_records(a: TextRecord, b: TextRecord) -> float:
    a_min_y, a_max_y = a.bbox[1], a.bbox[3]
    b_min_y, b_max_y = b.bbox[1], b.bbox[3]
    if a_max_y < b_min_y:
        return b_min_y - a_max_y
    if b_max_y < a_min_y:
        return a_min_y - b_max_y
    return 0.0


def _extract_attribute_fields(block_ref: Any) -> dict[str, str]:
    out: dict[str, str] = {}
    try:
        if not bool(get_attr(block_ref, "HasAttributes", False)):
            return out
        for attr in block_ref.GetAttributes():
            tag = str(get_attr(attr, "TagString", "")).strip()
            value = _clean_attr_value(str(get_attr(attr, "TextString", "")).strip())
            if tag and value:
                out[tag] = value
    except Exception:
        return out
    return out


def TDbMText_content(comobj: Any, separator: str = "\n") -> str:
    fragments: list[Any] = []
    copy_ent = None
    try:
        li()
    except Exception:
        pass

    try:
        try:
            copy_ent = comobj.Copy()
        except Exception:
            copy_ent = None
        if copy_ent is None:
            return ""

        try:
            exploded = copy_ent.Explode()
            fragments = list(exploded) if exploded else []
        except Exception:
            fragments = []

        if not fragments:
            return ""

        valid = []
        for frag in fragments:
            obj_name = str(get_attr(frag, "ObjectName", ""))
            if obj_name in {"AcDbText", "AcDbMText", "Text", "MText"}:
                valid.append(frag)
        if not valid:
            return ""

        def _sort_key(ent: Any) -> tuple[float, float]:
            bbox = _bbox_xy(ent) or (0.0, 0.0, 0.0, 0.0)
            return (-bbox[1], bbox[0])

        ordered = sorted(valid, key=_sort_key)
        parts: list[str] = []
        last_y: float | None = None
        for frag in ordered:
            text = _clean_text(str(get_attr(frag, "TextString", "")))
            if not text:
                continue
            bbox = _bbox_xy(frag) or (0.0, 0.0, 0.0, 0.0)
            if last_y is not None and abs(bbox[1] - last_y) > 0.3:
                parts.append(separator)
            parts.append(text)
            last_y = bbox[1]
        return "".join(parts).strip()
    except Exception as exc:
        sys_logger.info(f"TDbMText 内容提取失败: {exc}")
        return ""
    finally:
        for frag in fragments:
            try:
                frag.Delete()
            except Exception:
                pass
        if copy_ent is not None:
            try:
                copy_ent.Delete()
            except Exception:
                pass


def _extract_text_string(ent: Any) -> str:
    obj_name = str(get_attr(ent, "ObjectName", getattr(ent, "ObjectName", "")))
    try:
        if obj_name in {"AcDbText", "Text", "AcDbMText", "MText"}:
            return _clean_text(str(get_attr(ent, "TextString", "")))
        if obj_name == "TDbText":
            return _clean_text(
                str(
                    get_attr(
                        ent,
                        "Text",
                        get_attr(ent, "TXT", get_attr(ent, "TextString", "")),
                    )
                )
            )
        if obj_name == "TDbMText":
            return _clean_text(TDbMText_content(ent))
    except Exception:
        return ""
    return ""


def _match_field_from_attrs(fields: dict[str, str], hints: tuple[str, ...]) -> str:
    best_score = 0
    best_value = ""
    for tag, value in fields.items():
        if not value:
            continue
        tag_upper = tag.upper()
        score = 0
        for hint in hints:
            hint_upper = hint.upper()
            if tag_upper == hint_upper:
                score = max(score, 100)
            elif hint_upper in tag_upper:
                score = max(score, 60)
        if score > best_score:
            best_score = score
            best_value = value
    return best_value


def _resolve_owner_btr_name(job: dict[str, Any]) -> str:
    owner_btr = str(job.get("owner_btr", ""))
    if owner_btr.startswith("layout_viewport:"):
        layout_name = owner_btr.split(":", 1)[1]
        try:
            layout = C.doc.Layouts.Item(layout_name)
            return str(layout.Block.Name)
        except Exception:
            return owner_btr
    return owner_btr


def _collect_rectangles_by_owner() -> dict[str, list[Any]]:
    grouped = get_rect_polylines_by_space(select_all_polylines(autocast=True))
    out: dict[str, list[Any]] = {"*MODEL_SPACE": grouped.get("model", [])}
    for owner_btr_name, rects in grouped.get("papers", {}).items():
        out[str(owner_btr_name)] = rects
    return out


def collect_space_block_snapshots(owner_btr_name: str) -> list[BlockSnapshot]:
    if not _activate_space_for_owner(owner_btr_name):
        sys_logger.warning(f"无法激活分析空间: owner={owner_btr_name}")
        return []

    try:
        block_entities = ss_select("all", filter_types=[0], filter_data=["INSERT"], autocast=True) or []
    except Exception as exc:
        sys_logger.warning(f"块选择集采样失败: owner={owner_btr_name} err={exc}")
        block_entities = []

    snapshots: list[BlockSnapshot] = []
    for ent in block_entities:
        obj_name = str(get_attr(ent, "ObjectName", getattr(ent, "ObjectName", "")))
        if obj_name != "AcDbBlockReference":
            continue
        bbox = _bbox_xy(ent)
        if bbox is None:
            continue
        try:
            block_name = str(get_attr(ent, "EffectiveName", get_attr(ent, "Name", getattr(ent, "Name", ""))))
        except Exception:
            block_name = str(get_attr(ent, "Name", getattr(ent, "Name", "")))
        attrs = _extract_attribute_fields(ent)
        snapshots.append(
            BlockSnapshot(
                handle=_safe_handle(ent),
                block_name=block_name,
                bbox=bbox,
                owner_btr=owner_btr_name,
                is_attribute_block=int(bool(attrs)),
                attr_fields=attrs,
            )
        )
    return snapshots


def select_texts_in_bbox(owner_btr_name: str, bbox: tuple[float, float, float, float]) -> list[TextRecord]:
    entities = _select_entities_in_bbox(owner_btr_name, bbox)
    out: list[TextRecord] = []
    for ent in entities:
        record = _make_text_record(ent)
        if record is not None:
            out.append(record)
    out.sort(key=_record_sort_key_reading)
    return out


def _choose_inner_frame(print_job: dict[str, Any], owner_rects: list[Any]) -> tuple[str, tuple[float, float, float, float] | None]:
    print_bbox = tuple(print_job["lower_left"] + print_job["upper_right"])
    print_area = _area_of_bbox(print_bbox)
    tol = _dynamic_tol_from_bbox(print_bbox)
    print_handle = str(print_job.get("handle", ""))
    best_handle = ""
    best_bbox: tuple[float, float, float, float] | None = None
    best_area = 0.0

    for poly in owner_rects:
        handle = _safe_handle(poly)
        if handle == print_handle:
            continue
        bbox = _bbox_xy(poly)
        if bbox is None:
            continue
        if not _contains_bbox(print_bbox, bbox, tol):
            continue
        area = _area_of_bbox(bbox)
        if area >= print_area * 0.999:
            continue
        if area < print_area * 0.85:
            continue
        if area > best_area:
            best_handle = handle
            best_bbox = bbox
            best_area = area

    return best_handle, best_bbox


def _choose_corner_block(
    inner_bbox: tuple[float, float, float, float],
    block_snapshots: list[BlockSnapshot],
    *,
    portrait: bool = False,
) -> BlockSnapshot | None:
    if portrait:
        target_x, target_y = inner_bbox[0], inner_bbox[1]
    else:
        target_x, target_y = inner_bbox[2], inner_bbox[1]
    tol = _dynamic_tol_from_bbox(inner_bbox)
    best: BlockSnapshot | None = None
    best_distance: float | None = None

    for block in block_snapshots:
        block_x = block.bbox[0] if portrait else block.bbox[2]
        block_y = block.bbox[1]
        dx = abs(block_x - target_x)
        dy = abs(block_y - target_y)
        if dx > tol or dy > tol:
            continue
        distance = (dx ** 2 + dy ** 2) ** 0.5
        if best is None or distance < best_distance:
            best = block
            best_distance = distance
    return best


def _explode_block_reference_copy(block_handle: str) -> list[Any]:
    fragments: list[Any] = []
    try:
        block_ref = C.raw_doc.HandleToObject(block_handle)
    except Exception:
        return fragments
    try:
        copy_ent = block_ref.Copy()
    except Exception:
        return fragments
    try:
        exploded = copy_ent.Explode()
        fragments = list(exploded) if exploded else []
    except Exception:
        fragments = []
    finally:
        try:
            copy_ent.Delete()
        except Exception:
            pass
    return fragments


def _is_rectangular_polyline(ent: Any, tol: float = 1e-3) -> bool:
    obj_name = str(get_attr(ent, "ObjectName", getattr(ent, "ObjectName", "")))
    if obj_name not in {"AcDbPolyline", "AcDb2dPolyline", "Polyline", "LWPOLYLINE"}:
        return False
    coords = get_attr(ent, "Coordinates", None)
    if not coords:
        return False
    values = list(coords)
    if len(values) < 8:
        return False
    pairs = []
    for idx in range(0, len(values) - 1, 2):
        pairs.append((round(float(values[idx]) / tol) * tol, round(float(values[idx + 1]) / tol) * tol))
    unique_pairs = list(dict.fromkeys(pairs))
    if len(unique_pairs) == 5 and unique_pairs[0] == unique_pairs[-1]:
        unique_pairs = unique_pairs[:-1]
    xs = {round(item[0] / tol) for item in unique_pairs}
    ys = {round(item[1] / tol) for item in unique_pairs}
    return len(xs) == 2 and len(ys) == 2 and len(unique_pairs) in {4, 5}


def _extract_guide_rectangles(
    selected_block: BlockSnapshot | None,
) -> dict[str, tuple[float, float, float, float]]:
    if selected_block is None:
        return {}
    fragments = _explode_block_reference_copy(selected_block.handle)
    if not fragments:
        return {}
    try:
        guides: dict[str, tuple[float, float, float, float]] = {}
        for frag in fragments:
            layer = str(get_attr(frag, "Layer", getattr(frag, "Layer", "")))
            if layer != "Defpoints":
                continue
            if not _is_rectangular_polyline(frag):
                continue
            bbox = _bbox_xy(frag)
            if bbox is None:
                continue
            color = _get_color_index(frag)
            if color == 1:
                current = guides.get("red")
                if current is None or _area_of_bbox(bbox) > _area_of_bbox(current):
                    guides["red"] = bbox
            elif color == 3:
                current = guides.get("green")
                if current is None or _area_of_bbox(bbox) > _area_of_bbox(current):
                    guides["green"] = bbox
        return guides
    finally:
        for frag in fragments:
            try:
                frag.Delete()
            except Exception:
                pass


def _resolve_by_named_layers(
    text_records: list[TextRecord],
) -> tuple[list[TextRecord], list[TextRecord], str] | None:
    number_records = [record for record in text_records if _layer_contains(record.layer, "图纸编号")]
    title_records = [record for record in text_records if _layer_contains(record.layer, "图纸名称")]
    if not number_records or not title_records:
        return None
    merged_number = _merge_text_records(number_records, sort_mode="x", separator="")
    return title_records, [merged_number] if merged_number else [], "layer_named"


def _resolve_by_guide_rectangles(
    owner_btr_name: str,
    guide_rectangles: dict[str, tuple[float, float, float, float]],
) -> tuple[list[TextRecord], list[TextRecord], str] | None:
    red_bbox = guide_rectangles.get("red")
    green_bbox = guide_rectangles.get("green")
    if red_bbox is None or green_bbox is None:
        return None

    title_records = select_texts_in_bbox(owner_btr_name, red_bbox)
    number_records = select_texts_in_bbox(owner_btr_name, green_bbox)
    if len(number_records) == 2:
        ordered = sorted(number_records, key=_record_sort_key_x)
        y_gap = _vertical_gap_between_records(ordered[0], ordered[1])
        threshold = max(min(ordered[0].height, ordered[1].height), 1.0)
        if y_gap < threshold:
            merged = _merge_text_records(ordered, sort_mode="x", separator="")
            number_records = [merged] if merged else ordered
    return title_records, number_records, "guide_rectangles"


def _resolve_by_fallback_rules(
    text_records: list[TextRecord],
    graphic_info_area_bbox: tuple[float, float, float, float],
) -> tuple[list[TextRecord], list[TextRecord], str]:
    number_candidates = [
        record for record in text_records
        if STRICT_DRAWING_NO_RE.match(record.text or "")
    ]
    number_candidates.sort(key=_record_sort_key_bottom)

    number_records: list[TextRecord] = []
    if number_candidates:
        first = number_candidates[0]
        if len(number_candidates) == 1:
            number_records = [first]
        else:
            second = number_candidates[1]
            local_y_1 = first.bbox[1] - graphic_info_area_bbox[1]
            local_y_2 = second.bbox[1] - graphic_info_area_bbox[1]
            if ((local_y_1 + first.height) * 2) < local_y_2:
                number_records = [first]
            else:
                merged = _merge_text_records([first, second], sort_mode="x", separator="")
                number_records = [merged] if merged else [first, second]

    used_handles = {
        source_handle
        for record in number_records
        for source_handle in (record.source_handles or (record.handle,))
    }
    title_records = [record for record in text_records if record.handle not in used_handles]
    return title_records, number_records, "fallback_regex"


def _make_page_key(job: dict[str, Any]) -> str:
    return f"{job['layout_name']}-{int(job['sequence_no']):02d}"


def _make_analysis_error_row(job: dict[str, Any], error: Exception) -> dict[str, Any]:
    page_key = _make_page_key(job)
    return {
        "page_key": page_key,
        "sequence_key": f"{int(job['sequence_no']):02d}",
        "sequence_no": int(job["sequence_no"]),
        "layout_name": str(job["layout_name"]),
        "space_kind": str(job["space_kind"]),
        "print_handle": str(job["handle"]),
        "drawing_title": "",
        "drawing_no": "",
        "project_name": "",
        "subproject_name": "",
        "drawing_title_record_count": 0,
        "drawing_no_record_count": 0,
        "drawing_title_handles": [],
        "drawing_no_handles": [],
        "ratio": str(job.get("ratio", "")),
        "paper_code": str(job.get("paper_code", "")),
        "media": str(job.get("media", "")),
        "rotation": int(job.get("rotation", 0)),
        "plot_scale": float(job.get("plot_scale", 0.0)),
        "standard_flag": int(job.get("standard_flag", 0)),
        "inner_frame_exists": 0,
        "inner_frame_found": 0,
        "inner_frame_handle": "",
        "inner_frame_bbox": None,
        "graphic_info_area_bbox": None,
        "graphic_info_area_source": "",
        "graphic_orientation": "",
        "right_bottom_title_block_exists": 0,
        "right_bottom_block_exists": 0,
        "title_block_found": 0,
        "title_block_kind": "",
        "title_block_handle": "",
        "title_block_name": "",
        "title_block_is_attribute": 0,
        "title_block_bbox": None,
        "title_block_attr_fields": {},
        "attr_fields": {},
        "title_no_resolve_method": "",
        "title_block_texts": [],
        "title_block_text_records": [],
        "graphic_info_text_records": [],
        "classified_text_records": [],
        "drawing_title_records": [],
        "drawing_no_records": [],
        "drawing_title_candidates": [],
        "drawing_no_candidates": [],
        "selected_texts": [],
        "analysis_stop_reason": f"analysis_error:{type(error).__name__}",
        "analysis_error": str(error),
    }


def analyze_print_job_info(
    job: dict[str, Any],
    *,
    owner_rects: list[Any],
    block_snapshots: list[BlockSnapshot],
) -> dict[str, Any]:
    page_key = _make_page_key(job)
    inner_handle, inner_bbox = _choose_inner_frame(job, owner_rects)
    selected_block = None
    attr_fields: dict[str, str] = {}
    graphic_info_text_records: list[TextRecord] = []
    drawing_title_records: list[TextRecord] = []
    drawing_no_records: list[TextRecord] = []
    drawing_title_candidates: list[str] = []
    drawing_no_candidates: list[str] = []
    drawing_title = ""
    drawing_no = ""
    project_name = ""
    subproject_name = ""
    stop_reason = "no_inner_frame"
    graphic_info_area_bbox: tuple[float, float, float, float] | None = None
    graphic_info_area_source = ""
    title_no_resolve_method = ""
    portrait = _is_portrait_job(job, inner_bbox)
    graphic_orientation = "portrait" if portrait else "landscape"

    if inner_bbox is not None:
        selected_block = _choose_corner_block(inner_bbox, block_snapshots, portrait=portrait)
        if selected_block is None:
            stop_reason = "no_corner_block"
        else:
            graphic_info_area_bbox = selected_block.bbox
            graphic_info_area_source = "corner_block_bbox"
            attr_fields = dict(selected_block.attr_fields)
            project_name = _match_field_from_attrs(attr_fields, PROJECT_TAG_HINTS)
            subproject_name = _match_field_from_attrs(attr_fields, SUBPROJECT_TAG_HINTS)
            graphic_info_text_records = select_texts_in_bbox(selected_block.owner_btr, selected_block.bbox)

            resolved = _resolve_by_named_layers(graphic_info_text_records)
            if resolved is not None:
                drawing_title_records, drawing_no_records, title_no_resolve_method = resolved
                stop_reason = "graphic_info_area_layer_named"
            else:
                guide_rectangles = _extract_guide_rectangles(selected_block)
                guide_resolved = _resolve_by_guide_rectangles(selected_block.owner_btr, guide_rectangles)
                if guide_resolved is not None:
                    drawing_title_records, drawing_no_records, title_no_resolve_method = guide_resolved
                    stop_reason = "graphic_info_area_guides"
                else:
                    drawing_title_records, drawing_no_records, title_no_resolve_method = _resolve_by_fallback_rules(
                        graphic_info_text_records,
                        graphic_info_area_bbox,
                    )
                    stop_reason = "graphic_info_area_fallback"

            drawing_title_candidates = [record.text for record in drawing_title_records if record.text]
            drawing_no_candidates = [record.text for record in drawing_no_records if record.text]
            drawing_title = _combine_text_values(drawing_title_records, sort_mode="reading", separator="")
            drawing_no = _combine_text_values(drawing_no_records, sort_mode="x", separator="")

            if not drawing_title and attr_fields:
                drawing_title = _match_field_from_attrs(attr_fields, TITLE_TAG_HINTS)
            if not drawing_no and attr_fields:
                drawing_no = _match_field_from_attrs(attr_fields, NUMBER_TAG_HINTS)

    title_block_kind = ""
    if selected_block is not None:
        title_block_kind = "attribute_block" if selected_block.is_attribute_block == 1 else "normal_block"

    drawing_title_handles = _record_handle_list(drawing_title_records)
    drawing_no_handles = _record_handle_list(drawing_no_records)
    classified_text_records = _classify_text_record_dicts(
        graphic_info_text_records,
        drawing_title_records,
        drawing_no_records,
    )

    return {
        "page_key": page_key,
        "sequence_key": f"{int(job['sequence_no']):02d}",
        "sequence_no": int(job["sequence_no"]),
        "layout_name": str(job["layout_name"]),
        "space_kind": str(job["space_kind"]),
        "print_handle": str(job["handle"]),
        "drawing_title": drawing_title,
        "drawing_no": drawing_no,
        "project_name": project_name,
        "subproject_name": subproject_name,
        "drawing_title_record_count": len(drawing_title_records),
        "drawing_no_record_count": len(drawing_no_records),
        "drawing_title_handles": drawing_title_handles,
        "drawing_no_handles": drawing_no_handles,
        "ratio": str(job.get("ratio", "")),
        "paper_code": str(job.get("paper_code", "")),
        "media": str(job.get("media", "")),
        "rotation": int(job.get("rotation", 0)),
        "plot_scale": float(job.get("plot_scale", 0.0)),
        "standard_flag": int(job.get("standard_flag", 0)),
        "inner_frame_exists": int(inner_bbox is not None),
        "inner_frame_found": int(inner_bbox is not None),
        "inner_frame_handle": inner_handle,
        "inner_frame_bbox": inner_bbox,
        "graphic_info_area_bbox": graphic_info_area_bbox,
        "graphic_info_area_source": graphic_info_area_source,
        "graphic_orientation": graphic_orientation,
        "right_bottom_title_block_exists": int(selected_block is not None),
        "right_bottom_block_exists": int(selected_block is not None),
        "title_block_found": int(selected_block is not None),
        "title_block_kind": title_block_kind,
        "title_block_handle": selected_block.handle if selected_block else "",
        "title_block_name": selected_block.block_name if selected_block else "",
        "title_block_is_attribute": int(selected_block.is_attribute_block) if selected_block else 0,
        "title_block_bbox": selected_block.bbox if selected_block else None,
        "title_block_attr_fields": attr_fields,
        "attr_fields": attr_fields,
        "title_no_resolve_method": title_no_resolve_method,
        "title_block_texts": [item.text for item in graphic_info_text_records],
        "title_block_text_records": [_text_record_to_dict(item) for item in graphic_info_text_records],
        "graphic_info_text_records": [_text_record_to_dict(item) for item in graphic_info_text_records],
        "classified_text_records": classified_text_records,
        "drawing_title_records": [_text_record_to_dict(item) for item in drawing_title_records],
        "drawing_no_records": [_text_record_to_dict(item) for item in drawing_no_records],
        "drawing_title_candidates": drawing_title_candidates,
        "drawing_no_candidates": drawing_no_candidates,
        "selected_texts": [_text_record_to_dict(item) for item in graphic_info_text_records],
        "analysis_stop_reason": stop_reason,
    }


@retry_on_busy(max_retries=4, base_delay=0.5)
def _analyze_print_job_info_with_retry(
    job: dict[str, Any],
    *,
    owner_rects: list[Any],
    block_snapshots: list[BlockSnapshot],
) -> dict[str, Any]:
    try:
        wait_quiescent(min_quiet=0.2, timeout=10.0)
    except Exception:
        pass
    return analyze_print_job_info(
        job,
        owner_rects=owner_rects,
        block_snapshots=block_snapshots,
    )


def analyze_print_info_jobs(
    jobs_by_space: dict[str, list[dict[str, Any]]],
    *,
    excluded_handles: set[str] | None = None,
) -> dict[str, Any]:
    excluded_handles = {str(item) for item in (excluded_handles or set())}
    rects_by_owner = _collect_rectangles_by_owner()
    blocks_cache: dict[str, list[BlockSnapshot]] = {}
    rows_by_space: dict[str, list[dict[str, Any]]] = {}
    print_info_dict_by_space: dict[str, dict[str, dict[str, Any]]] = {}
    page_info_dict: dict[str, dict[str, Any]] = {}

    for layout_name, jobs in jobs_by_space.items():
        rows: list[dict[str, Any]] = []
        info_rows: dict[str, dict[str, Any]] = {}
        for index, job in enumerate(jobs, start=1):
            handle = str(job.get("handle", ""))
            if handle in excluded_handles:
                continue
            owner_btr_name = _resolve_owner_btr_name(job)
            if owner_btr_name not in blocks_cache:
                blocks_cache[owner_btr_name] = collect_space_block_snapshots(owner_btr_name)
            try:
                if index == 1 or index % 6 == 0:
                    try:
                        C.li()
                    except Exception:
                        pass
                row = _analyze_print_job_info_with_retry(
                    job,
                    owner_rects=rects_by_owner.get(owner_btr_name, []),
                    block_snapshots=blocks_cache.get(owner_btr_name, []),
                )
            except Exception as exc:
                sys_logger.warning(
                    f"print_info 单页分析失败: layout={layout_name} handle={handle} seq={job.get('sequence_no')} err={exc}"
                )
                row = _make_analysis_error_row(job, exc)
            rows.append(row)
            info_rows[row["sequence_key"]] = row
            page_info_dict[row["page_key"]] = row
        rows_by_space[layout_name] = rows
        if info_rows:
            print_info_dict_by_space[layout_name] = info_rows

    all_rows = [row for rows in rows_by_space.values() for row in rows]
    if len(print_info_dict_by_space) == 1:
        print_info_dict = next(iter(print_info_dict_by_space.values()))
    else:
        print_info_dict = {
            row["page_key"]: row
            for row in all_rows
        }
    return {
        "total_jobs": len(all_rows),
        "inner_frame_found_count": sum(1 for row in all_rows if row["inner_frame_found"] == 1),
        "title_block_found_count": sum(1 for row in all_rows if row["title_block_found"] == 1),
        "attribute_block_count": sum(1 for row in all_rows if row["title_block_is_attribute"] == 1),
        "with_title_count": sum(1 for row in all_rows if row["drawing_title"]),
        "with_drawing_no_count": sum(1 for row in all_rows if row["drawing_no"]),
        "with_number_count": sum(1 for row in all_rows if row["drawing_no"]),
        "with_project_count": sum(1 for row in all_rows if row["project_name"]),
        "with_subproject_count": sum(1 for row in all_rows if row.get("subproject_name")),
        "print_info_dict": print_info_dict,
        "print_info_dict_by_space": print_info_dict_by_space,
        "page_info_dict": page_info_dict,
        "jobs_by_space": rows_by_space,
    }


def _load_jobs_from_plan_json(plan_json_path: Path) -> dict[str, list[dict[str, Any]]]:
    raw = json.loads(plan_json_path.read_text(encoding="utf-8"))
    return raw["jobs_by_space"]


def _load_excluded_handles(content_json_path: Path | None) -> set[str]:
    if content_json_path is None or not content_json_path.exists():
        return set()
    raw = json.loads(content_json_path.read_text(encoding="utf-8"))
    return {str(item["handle"]) for item in raw.get("pseudo_candidates", [])}


def _stringify_excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float)):
        return value
    if isinstance(value, list):
        if all(isinstance(item, str) for item in value):
            return "\n".join(item for item in value if item)
        return json.dumps(value, ensure_ascii=False, indent=2)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, indent=2)
    return str(value)


def write_print_info_excel(result: dict[str, Any], output_path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "summary"
    summary_rows = [
        ("dwg_path", result.get("dwg_path", "")),
        ("print_mode", result.get("print_mode", "")),
        ("total_jobs", result.get("total_jobs", 0)),
        ("inner_frame_found_count", result.get("inner_frame_found_count", 0)),
        ("title_block_found_count", result.get("title_block_found_count", 0)),
        ("with_title_count", result.get("with_title_count", 0)),
        ("with_drawing_no_count", result.get("with_drawing_no_count", 0)),
        ("with_project_count", result.get("with_project_count", 0)),
        ("with_subproject_count", result.get("with_subproject_count", 0)),
        ("excluded_handle_count", result.get("excluded_handle_count", 0)),
    ]
    for row_index, (key, value) in enumerate(summary_rows, start=1):
        summary_ws.cell(row=row_index, column=1, value=key).font = Font(bold=True)
        summary_ws.cell(row=row_index, column=2, value=_stringify_excel_value(value))

    detail_ws = workbook.create_sheet("print_info")
    for col_index, (_, title) in enumerate(EXCEL_MAIN_COLUMNS, start=1):
        cell = detail_ws.cell(row=1, column=col_index, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    print_info_dict = result.get("print_info_dict", {}) or {}
    ordered_rows = sorted(
        print_info_dict.values(),
        key=lambda item: (int(item.get("sequence_no", 0)), str(item.get("page_key", ""))),
    )
    for row_index, row in enumerate(ordered_rows, start=2):
        for col_index, (key, _) in enumerate(EXCEL_MAIN_COLUMNS, start=1):
            detail_ws.cell(row=row_index, column=col_index, value=_stringify_excel_value(row.get(key)))

    text_ws = workbook.create_sheet("text_records")
    text_headers = [
        "序号",
        "页面键",
        "打印区域句柄",
        "文字序号",
        "归类角色",
        "文字句柄",
        "文字内容",
        "对象类型",
        "图层",
        "包围盒",
    ]
    for col_index, title in enumerate(text_headers, start=1):
        cell = text_ws.cell(row=1, column=col_index, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    title_ws = workbook.create_sheet("drawing_title_records")
    number_ws = workbook.create_sheet("drawing_no_records")
    classified_headers = ["序号", "页面键", "打印区域句柄", "对象序号", "文字句柄", "文字内容", "对象类型", "图层", "包围盒"]
    for ws in (title_ws, number_ws):
        for col_index, title in enumerate(classified_headers, start=1):
            cell = ws.cell(row=1, column=col_index, value=title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    text_row_index = 2
    title_row_index = 2
    number_row_index = 2
    for row in ordered_rows:
        text_records = row.get("classified_text_records", []) or row.get("graphic_info_text_records", []) or []
        for text_index, text_row in enumerate(text_records, start=1):
            values = [
                row.get("sequence_no", 0),
                row.get("page_key", ""),
                row.get("print_handle", ""),
                text_index,
                text_row.get("resolved_role", ""),
                text_row.get("handle", ""),
                text_row.get("text", ""),
                text_row.get("obj_name", ""),
                text_row.get("layer", ""),
                _stringify_excel_value(text_row.get("bbox")),
            ]
            for col_index, value in enumerate(values, start=1):
                text_ws.cell(row=text_row_index, column=col_index, value=value)
            text_row_index += 1
        for target_ws, target_records, current_row_index in (
            (title_ws, row.get("drawing_title_records", []) or [], title_row_index),
            (number_ws, row.get("drawing_no_records", []) or [], number_row_index),
        ):
            local_row_index = current_row_index
            for item_index, item in enumerate(target_records, start=1):
                item_values = [
                    row.get("sequence_no", 0),
                    row.get("page_key", ""),
                    row.get("print_handle", ""),
                    item_index,
                    item.get("handle", ""),
                    item.get("text", ""),
                    item.get("obj_name", ""),
                    item.get("layer", ""),
                    _stringify_excel_value(item.get("bbox")),
                ]
                for col_index, value in enumerate(item_values, start=1):
                    target_ws.cell(row=local_row_index, column=col_index, value=value)
                local_row_index += 1
            if target_ws is title_ws:
                title_row_index = local_row_index
            else:
                number_row_index = local_row_index

    for ws in (summary_ws, detail_ws, text_ws, title_ws, number_ws):
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            max_len = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                value_len = max((len(part) for part in value.splitlines()), default=0)
                if value_len > max_len:
                    max_len = value_len
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _filter_jobs_by_requested_handles(
    jobs_by_space: dict[str, list[dict[str, Any]]],
    requested_handles: set[str] | None,
) -> dict[str, list[dict[str, Any]]]:
    requested_handles = {str(item) for item in (requested_handles or set())}
    if not requested_handles:
        return jobs_by_space

    filtered: dict[str, list[dict[str, Any]]] = {}
    for layout_name, jobs in jobs_by_space.items():
        selected = [job for job in jobs if str(job.get("handle", "")) in requested_handles]
        if selected:
            filtered[layout_name] = selected
    return _reindex_job_dicts_by_space(filtered)


def _reindex_job_dicts_by_space(
    jobs_by_space: dict[str, list[dict[str, Any]]],
) -> dict[str, list[dict[str, Any]]]:
    normalized: dict[str, list[dict[str, Any]]] = {}
    dataclass_jobs: dict[str, list[PrintJob]] = {}
    for layout_name, jobs in jobs_by_space.items():
        items: list[PrintJob] = []
        for job in jobs:
            items.append(PrintJob(**job))
        if items:
            dataclass_jobs[layout_name] = items

    for layout_name, jobs in reindex_jobs_by_space(dataclass_jobs).items():
        normalized[layout_name] = [
            {
                **job.__dict__,
                "lower_left": list(job.lower_left),
                "upper_right": list(job.upper_right),
            }
            for job in jobs
        ]
    return normalized


def _exclude_handles_and_reindex(
    jobs_by_space: dict[str, list[dict[str, Any]]],
    excluded_handles: set[str] | None,
) -> dict[str, list[dict[str, Any]]]:
    excluded_handles = {str(item) for item in (excluded_handles or set())}
    if not excluded_handles:
        return _reindex_job_dicts_by_space(jobs_by_space)

    filtered: dict[str, list[dict[str, Any]]] = {}
    for layout_name, jobs in jobs_by_space.items():
        kept = [job for job in jobs if str(job.get("handle", "")) not in excluded_handles]
        if kept:
            filtered[layout_name] = kept
    return _reindex_job_dicts_by_space(filtered)


def run_print_info_case(
    *,
    dwg_path: Path,
    output_path: Path,
    source_dwg_path: Path | None = None,
    plan_json_path: Path | None = None,
    content_json_path: Path | None = None,
    mode: str = PRINT_MODE_BASIC,
    include_model: bool = True,
    include_layouts: bool = True,
    only_layouts: list[str] | None = None,
    requested_handles: set[str] | None = None,
    keep_open: bool = False,
) -> dict[str, Any]:
    if not dwg_path.exists():
        raise FileNotFoundError(dwg_path)
    report_dwg_path = Path(source_dwg_path) if source_dwg_path else dwg_path

    launch_cad_guardians()
    if not litz():
        raise RuntimeError("print_info_analysis 启动阶段未能恢复到可信天正环境")
    assert_runtime_guard_ok("print_info_analysis:after_litz")

    need_open = _find_document_by_path(dwg_path) is None
    if need_open:
        if not open_file(str(dwg_path)):
            raise RuntimeError(f"打开 DWG 失败: {dwg_path}")
    if not _activate_document_by_path(dwg_path):
        raise RuntimeError(f"未能激活 DWG: {dwg_path}")
    wait_quiescent(min_quiet=0.5, timeout=20.0)
    assert_runtime_guard_ok("print_info_analysis:after_open_dwg")

    content_json_source = content_json_path

    if plan_json_path and plan_json_path.exists():
        jobs_by_space = _load_jobs_from_plan_json(plan_json_path)
    else:
        plan = build_print_plan(
            str(dwg_path),
            output_path.parent / "_tmp_pdf",
            mode=normalize_print_mode(mode),
            include_model=include_model,
            include_layouts=include_layouts,
            only_layouts=only_layouts,
        )
        jobs_by_space = plan_to_dict(plan)["jobs_by_space"]
        if mode == PRINT_MODE_PURIFIED_ADAPTIVE and content_json_source is None:
            content_payload = analyze_jobs_content(jobs_by_space)
            pseudo_handles = {str(item["handle"]) for item in content_payload.get("pseudo_candidates", [])}
            if pseudo_handles:
                content_json_source = output_path.parent / "content_analysis.json"
                content_json_source.write_text(
                    json.dumps(
                        {
                            "dwg_path": str(report_dwg_path),
                            "analysis_dwg_path": str(dwg_path),
                            "mode": mode,
                            **content_payload,
                        },
                        ensure_ascii=False,
                        indent=2,
                    ),
                    encoding="utf-8",
                )
    assert_runtime_guard_ok("print_info_analysis:before_analyze_jobs")

    excluded_handles = _load_excluded_handles(content_json_source)
    jobs_by_space = _filter_jobs_by_requested_handles(jobs_by_space, requested_handles)
    jobs_by_space = _exclude_handles_and_reindex(jobs_by_space, excluded_handles)
    analysis = analyze_print_info_jobs(jobs_by_space)
    result = {
        "dwg_path": str(report_dwg_path),
        "analysis_dwg_path": str(dwg_path),
        "plan_json_path": str(plan_json_path) if plan_json_path else "",
        "content_json_path": str(content_json_source) if content_json_source else "",
        "print_mode": normalize_print_mode(mode),
        "requested_handles": sorted(str(item) for item in (requested_handles or set())),
        "excluded_handle_count": len(excluded_handles),
        **analysis,
    }
    excel_path = output_path.with_suffix(".xlsx")
    result["excel_path"] = str(excel_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    write_print_info_excel(result, excel_path)

    if not keep_open and need_open:
        _close_document_by_path(dwg_path, save_changes=False)
    return result


def main() -> None:
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8")
            except Exception:
                pass

    parser = argparse.ArgumentParser()
    parser.add_argument("--dwg", required=True, help="source dwg path")
    parser.add_argument("--plan-json", default="", help="existing print plan json")
    parser.add_argument("--content-json", default="", help="content analysis json for excluding pseudo areas")
    parser.add_argument("--output", default="", help="output json path")
    parser.add_argument("--mode", default=PRINT_MODE_BASIC, help="basic/adaptive/purified_adaptive")
    parser.add_argument("--layout", action="append", default=None, help="only analyze specified layout, repeatable")
    parser.add_argument("--handle", action="append", default=None, help="only analyze specified print-area handle, repeatable")
    parser.add_argument("--no-model", action="store_true", help="skip model space when building plan")
    parser.add_argument("--no-layouts", action="store_true", help="skip layout spaces when building plan")
    parser.add_argument("--keep-open", action="store_true", help="keep dwg open after analysis")
    args = parser.parse_args()

    dwg_path = Path(args.dwg)
    plan_json_path = Path(args.plan_json) if args.plan_json else None
    content_json_path = Path(args.content_json) if args.content_json else None
    if args.output:
        output_path = Path(args.output)
    elif plan_json_path:
        output_path = plan_json_path.parent / "print_info_analysis.json"
    else:
        output_path = MODULE_DIR / "cases" / "output" / dwg_path.stem / "print_info_analysis.json"

    try:
        result = run_print_info_case(
            dwg_path=dwg_path,
            output_path=output_path,
            source_dwg_path=dwg_path,
            plan_json_path=plan_json_path,
            content_json_path=content_json_path,
            mode=args.mode,
            include_model=not args.no_model,
            include_layouts=not args.no_layouts,
            only_layouts=args.layout,
            requested_handles=set(args.handle or []),
            keep_open=args.keep_open,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except RuntimeGuardTriggered as exc:
        print(json.dumps(render_guard_error(exc), ensure_ascii=False, indent=2))
        raise SystemExit(2)


if __name__ == "__main__":
    main()
