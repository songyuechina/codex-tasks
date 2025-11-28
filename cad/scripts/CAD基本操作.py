# -*- coding: utf-8 -*-

"""
3.11.5                                                        序


https://blog.csdn.net/Hulunbuir/article/details/83715279?spm=1001.2014.3001.5506 csdn上的参照学习文件

 D:/Myprogramsystem/cad/打印服务/printcad_to_pdf.py  二十一世纪不是人工写小说或论文的适宜时代，但的确是适合调用各种手段例如人工智能来实现某

个重大目标任务的时代。程序系统是一个巨大的工程，需要我们综合很多学科技能。其中最重要的事情之一，就是认识理解这个程序系统本身。我建立了自己

的Myprogramsystem 体系，也许在专业人士看来有点可笑。但实际上，正是从这些可能幼稚、错误的做法上，我们可以逐渐正确理解“合理的概念和本质的规

律”。

    《CAD自动化》这个文件，是整个CAD系统的开发文件，是动态的文件。我们会在某个节点复制副本并标注相应的名称保留它们，以便将来的参照使用。就

目前来说，我需要将打印、编目录这些最底层的工作完成，它们最终将封装到《CAD基本操作》这个文件中。

    接下来的任务是方案图的深化、施工图的绘制、工程图纸的识别等研究。在这些研究中，必然要产生、处理天文数字的数据，不可能用word、记事本等

格式保存和引用，这就是我们应该重视数据库操作的原因。我当然知道，过去、现在、将来都会有专业的公司、专业的团队来开发研究这些问题。要知道

这些成果实际上的确是有商业价值的。但是最重要的是这将训练我们的能力。就是扎根在土木工程并充分运用计算机知识的能力。我认为应系统学习

一下计算机课程。实际上，数学仍然是最重要的，我们需要学习大学本科数学教育的课程，具备做研究数学的能力却并不去做数学研究。

    当然，有人会认为这些要求是不是太高了。那我们要人工智能干什么呢。人们不是认为人工智能无所不能吗？我们应该注意到，动用人工智能写一篇百万字

的小说，实际上是对读者的羞辱。当很多公文、工程文件、科学论文的内容是用人工智能完成的时候，只是表明它是一些必不可少、如果需要我们就认真去阅读

并引用的内容，但肯定不是需要读者耗费大量时间精力阅读学习的内容。目前的人工智能只是对天文数字量的数据进行了科学合理处理的结果，它没有意识也没

有灵魂思想，它就是水中、镜中我们人就智力方面而言的影子镜像而已。

    真正有价值的问题不是人工智能有没有意识，会不会统治控制人类。真正有价值的问题是人工智能将迫使我们理解清楚人的本质、社会的本质、政治的本质、

经济的本质。私人拥有过量财富是不是极大的罪恶？拥有天量资本的集团是不是国家民族生死存亡的巨大威胁？社保医保制度是不是恰恰引发一个国家社会人口

日益锐减最后造成整个体系崩溃的根源？如果有人自认为是哲学家或者爱好哲学，他就不得不为所有人类的生存担忧，不得不为所有人活着的尊严、价值、意义

担忧。


完美解决各个脚本公用变量mp,doc,acad等

每个脚本都定义了L1()

在主脚本定义Lx，里面同时运行 L1() cj.L1() sj.L1()所有函数都会激活

eu不需要和脚本绑定

20240820
                                                                                                                                   20250330






"""

















#&&&&%%   CAD基本操作 

#&&&&%%  第一部分  导入、转换、连接等前置程序 
#_____________________________________________________________________________________________________________________________________________


# 导入模块


# —— 脚本最顶部，就写这两行 —— 
import warnings
warnings.filterwarnings(
    "ignore",
    r".*Revert to STA COM threading mode.*",
    category=UserWarning,
    module=r"pywinauto\..*"
)
from pywinauto.application import Application


#A______________________________________________________________

#B______________________________________________________________

import builtins



#C______________________________________________________________

from collections import defaultdict,deque

from concurrent.futures import ThreadPoolExecutor, TimeoutError

import cv2

import ctypes

from contextlib import redirect_stdout, redirect_stderr,suppress


#D______________________________________________________________


import datetime

#E______________________________________________________________

#F______________________________________________________________
import functools

from fractions import Fraction

from functools import cmp_to_key

from fitz import Rect

import fitz

from functools import wraps

import functools

from functools import wraps

import fire









#G______________________________________________________________

#H______________________________________________________________

#I______________________________________________________________

from itertools import chain

import importlib

import imageio

import inspect

import itertools

import io


#J______________________________________________________________


import  json


#K______________________________________________________________

#L______________________________________________________________

#M______________________________________________________________

import math

import multiprocessing

import mysql.connector

import matplotlib.pyplot as plt


from multiprocessing import Process


#N______________________________________________________________

import networkx as nx

import numpy as np 


#O______________________________________________________________

import os


# ———— 1. 抑制 pygame 欢迎提示 ————
# 必须在任何 import pygame 之前设置
os.environ.setdefault("PYGAME_HIDE_SUPPORT_PROMPT", "1")


from openpyxl import Workbook, load_workbook

from openpyxl.utils import column_index_from_string as col2idx

#P______________________________________________________________

import PyPDF2.errors

import pywintypes

from  pyautocad  import Autocad,APoint,aDouble

import pyautocad.types

import pytesseract

from PIL import ImageGrab

from pythoncom import VT_ARRAY, VT_R8

import pyautogui



from pywinauto import findwindows

import psutil

import pythoncom

import pygame

import pygetwindow as gw

import pyttsx3

from PIL  import Image



import pprint

import pdb

from pathlib import Path

import psutil

from pygetwindow import getWindowsWithTitle

from pypdf import PdfReader, PdfWriter


#Q______________________________________________________________

#R______________________________________________________________

import runpy

import regex as xre

import re

import random

#S______________________________________________________________

from shapely.geometry import Point, Polygon,LineString

import subprocess

import shutil

from sympy import *

import sys

from scipy.spatial import ConvexHull


from subprocess import DETACHED_PROCESS




#T______________________________________________________________

import time

import tempfile

from typing import Tuple, Literal

from typing import Any, Dict, List ,Tuple, Optional

from typing import Sequence, Literal, Optional,Union

from typing import Dict, Set,List

import tkinter as tk

import traceback

import tkinter

import threading


#U______________________________________________________________
import uuid
#V______________________________________________________________

#W______________________________________________________________



# ===== pywin32 / COM 基础 =====
import win32com                      # 确保顶层包存在

import win32com.client

import win32com.client as win32      # 常用入口，统一别名
from win32com.client import CastTo, makepy, constants, VARIANT,Dispatch

# 动态调度（少数场景用）
import win32com.client.dynamic as dyn

# Windows API / GUI / 进程
import win32gui
import win32api
import win32con
import win32process

import pythoncom                     # CoInitialize, COM 错误
import pywintypes                    # com_error


"""
主力用：win32.gencache.EnsureDispatch("AutoCAD.Application")（更稳、配合 MakePy）。

偶尔用：win32.Dispatch(...) 或 dyn.Dispatch(...)（需要强制晚绑定/动态属性时）。

避免把变量命名为 constants，以免遮蔽 win32com.client.constants。

如果你做成脚手架模块（scaffold），可加：

__all__ = [
    "win32com", "win32", "CastTo", "makepy", "constants", "VARIANT",
    "dyn", "win32gui", "win32api", "win32con", "win32process",
    "pythoncom", "pywintypes",
]


"""

#X______________________________________________________________

#Y______________________________________________________________

#Z______________________________________________________________

import psutil
import subprocess
import sys
from subprocess import CREATE_NO_WINDOW, DETACHED_PROCESS


from typing import List, Dict
from pathlib import Path
import datetime
from typing import List, Set, Dict




import datetime
import unicodedata


import psutil
import os
import time
import signal
import subprocess
from subprocess import CREATE_NO_WINDOW, DETACHED_PROCESS


from typing import Tuple
import io
import fitz  # PyMuPDF
from PIL import Image



#自建库


sys.path.append('D:/Myprogramsystem')






#&&% 函数别名
def alias(*names):
    """
    @alias("别名1","别名2",…)
    def foo(...): …
    """
    def decorator(func):
        mod = sys.modules[func.__module__]
        for nm in names:
            setattr(mod, nm, func)
        return func
    return decorator

# —— 保存原始 print ——
_orig_print = builtins.print

# ---------------- 全局劫持 ----------------
_orig_print = builtins.print            # 先保存原 print
def suppress_all_prints():
    builtins.print = lambda *a, **k: None
def restore_all_prints():
    builtins.print = _orig_print

# ---------------- 局部调试控制 ----------------
DEBUG             = False              # 总开关
_DEBUG_CODE_STACK = []                 # ← 新增：调试函数调用栈

def node(msg: str, *args, **kwargs):
    """
    只有 DEBUG = True 且当前帧属于【栈底】函数时才打印。
    """
    if not DEBUG or not _DEBUG_CODE_STACK:
        return
    frame = inspect.currentframe().f_back
    try:
        # 只允许“最外层”调试函数输出
        if frame.f_code is _DEBUG_CODE_STACK[0]:
            _orig_print(msg.format(*args), **kwargs)
    finally:
        del frame

@alias("e")
def enable_debug():
    """开启调试：普通 print → 静默，仅 node() 生效。"""
    global DEBUG
    DEBUG = True
    suppress_all_prints()

@alias("d")
def disable_debug():
    """关闭调试：恢复普通 print，node() 失效。"""
    global DEBUG
    DEBUG = False
    restore_all_prints()

def debuggable(func):
    """
    装饰器：进入时把 func.__code__ 推入栈；离开时弹栈。
    node() 始终只看栈底元素，所以只有第一次进入的函数会真正输出。
    """
    @functools.wraps(func)
    def wrapper(*args, **kw):
        _DEBUG_CODE_STACK.append(func.__code__)     # —— 入栈
        try:
            return func(*args, **kw)
        finally:
            _DEBUG_CODE_STACK.pop()                 # —— 出栈
    return wrapper


#&&% 控制函数运行时间
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --


# 日志配置
LOG_PATH = Path(r"D:/Myprogramsystem/XT/系统运行中断错误记录.xlsx")
LOG_SHEET = "错误记录"

# 初始化日志文件和工作表
if not LOG_PATH.exists():
    wb = Workbook()
    ws = wb.active
    ws.title = LOG_SHEET
    ws.append(["时间", "函数名", "参数repr", "备注"])
    wb.save(LOG_PATH)


def _log(func_name, args, kwargs, note):
    """
    将错误或超时信息写入 Excel 日志。如果工作表不存在则创建。
    """
    from time import strftime
    wb = load_workbook(LOG_PATH)
    # 确保日志表存在
    if LOG_SHEET in wb.sheetnames:
        ws = wb[LOG_SHEET]
    else:
        ws = wb.create_sheet(LOG_SHEET)
        ws.append(["时间", "函数名", "参数repr", "备注"])
    ws.append([strftime("%Y-%m-%d %H:%M:%S"),
               func_name,
               f"args={args!r}, kwargs={kwargs!r}",
               note])
    wb.save(LOG_PATH)


def _kill_acad():
    """
    强制结束所有以 acad 开头的 CAD 进程。
    """
    for proc in psutil.process_iter(("pid", "name")):
        name = proc.info.get("name")
        if name and name.lower().startswith("acad"):
            try:
                proc.kill()
            except Exception:
                pass


def timeout_and_log2(timeout_sec: float):
    """
    装饰器：对关键函数添加双重超时保护。
    - 守护线程监控主函数执行，超过 timeout_sec 秒将
      调用 _kill_acad() 并记录日志。
    - 函数内部可根据 timeout_sec+1 触发二次超时操作。

    用法：
        @timeout_and_log2(20)
        def Fx(..., timeout_sec=20):
            # 可选二次超时监控
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            finished = threading.Event()

            def watchdog():
                # 等待 timeout_sec 秒后，如函数尚未结束，杀CAD并记录
                if not finished.wait(timeout_sec):
                    _kill_acad()
                    _log(func.__name__, args, kwargs,
                         f"超时 {timeout_sec}s，强制终止 CAD")
            threading.Thread(target=watchdog, daemon=True).start()

            try:
                result = func(*args, **kwargs)
            except Exception:
                _log(func.__name__, args, kwargs,
                     f"函数内部异常:\n{traceback.format_exc()}")
                raise
            finally:
                finished.set()

            return result
        return wrapper
    return decorator

# ---------- 示例：双重超时测试函数 ----------
@timeout_and_log2(20)
def test_draw_circle_and_wait(center=(0,0,0), radius=10000, timeout_sec=20):
    """
    测试：在 CAD 中画圆并触发等待命令，
    后台线程在 timeout_sec+1 秒后发送无效命令以触发中断。
    """
    import threading
    from time import sleep, time

    # 获取当前 DWG 文件名
    try:
        current_file = current_dwg_basename()
    except Exception:
        current_file = '<unknown>'

    try:


        # 第1步：画圆
        circ = draw_circle(center, radius)
        if circ is None:
            print("❌ 圆绘制失败，测试中断。")
            return


        # 第2步：触发 CAD 等待
        doc.SendCommand("h\n")

        # 第3步：二次超时线程：timeout_sec+1 秒后发无效命令中断CAD
        def timeout_action():
            start = time()
            while True:
                if time() - start > timeout_sec + 1:
                    try:
                        doc.SendCommand("_.POINT 0,0,0")
                    except Exception as e:
                        print(f"⚠ 第二重超时触发，中断 CAD: {e}")
                    break
                sleep(0.1)

        threading.Thread(target=timeout_action, daemon=True).start()
        return circ

    except Exception as e:
        print(f"❌ test_draw_circle_and_wait 捕获到异常，退出: {e}")
        # 记录日志：函数名、输入参数、当前 DWG、异常信息
        try:
            current_file = getattr(doc, 'Name', '<unknown>')
            _log('test_draw_circle_and_wait', (center, radius), {'timeout_sec': timeout_sec},
                 f"内部异常，文件={current_file}, 异常={e}")
        except Exception:
            pass
        return

#函数计时

def timeit(func):
    """
    运行前后打印耗时，始终使用 _orig_print，
    不会被 enable_debug() 的 print 劫持影响。
    """
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        # 运行前
        _orig_print(f"⏱ 开始 `{func.__name__}` …")
        t0 = time.time()

        result = func(*args, **kwargs)

        elapsed = time.time() - t0
        # 运行后
        _orig_print(f"⏱ 完成 `{func.__name__}`，耗时：{elapsed:.3f} 秒")
        return result

    return wrapper



#&&% 函数打印消息控制
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --



# ---------- 示例：打印消息屏蔽测试函数 ----------

def ceshi_xiaoxi_dayin():

    stc("测试1")
    print("只打印主函数消息")



#&&% JSON数据读取存放设置
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --

# ———————— 全局配置 ————————
SAVE_DIR = r"D:/Myprogramsystem/XT/文件字典信息"
os.makedirs(SAVE_DIR, exist_ok=True)

# 全局缓存所有已经加载的打印字典
PRINT_DICTS: dict[str, dict] = {}

# ———————— 加载指定 DWG 的打印字典 ————————
def com_to_handle(obj):
    """
    如果是 COM 对象，则返回它的 Handle（字符串）；否则原样返回 obj。
    """
    try:
        # 大多数 Dispatch 对象都有 .Handle 属性
        if hasattr(obj, "Handle"):
            return obj.Handle
    except pythoncom.com_error:
        pass
    return obj

def serialize(obj):
    """
    递归地将 dict/list/tuple/COMObject 转为只含基础类型(包括 handle 字符串)的结构。
    """
    if isinstance(obj, dict):
        return {k: serialize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [serialize(v) for v in obj]
    if isinstance(obj, tuple):
        return tuple(serialize(v) for v in obj)
    # 不是容器的，就尝试转换成 handle
    return com_to_handle(obj)

def save_print_dict_generic(dwg_name: str, bind_dict: dict):
    """
    将 bind_dict 序列化并写入 JSON 文件，键为 dwg_name。
    序列化时自动把所有 COM 对象转换为它们的 Handle。
    """
    ser = serialize(bind_dict)
    path = os.path.join(SAVE_DIR, f"{dwg_name}_打印字典.json")
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(ser, f, ensure_ascii=False, indent=2)
    PRINT_DICTS[dwg_name] = ser
    print(f"✅ 已保存打印字典到 {path}")

def load(dwg_name: str) -> dict:
    """
    从 JSON 文件加载之前保存的打印字典（只含基本类型 & handle 字符串）。
    如果内存中已有，则优先返回内存中的。
    """
    if dwg_name in PRINT_DICTS:
        return PRINT_DICTS[dwg_name]

    path = os.path.join(SAVE_DIR, f"{dwg_name}_打印字典.json")
    if os.path.isfile(path):
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        PRINT_DICTS[dwg_name] = data
        print(f"✅ 已加载打印字典自 {path}")
        return data

    print(f"⚠ 未找到 {path}，返回空字典")
    return {}

# ———————— 辅助：获取当前 DWG 的纯文件名 ————————
def current_dwg_basename() -> str:
    """
    acad.ActiveDocument.Name 可能带路径或带 .dwg，
    取不带扩展的文件名。
    """
    name = acad.ActiveDocument.Name
    base = os.path.splitext(os.path.basename(name))[0]
    return base

def current_dwg_folder():
    """
    获取当前打开的 DWG 文件所在的文件夹路径。
    
    """
    try:
        
        full_path = doc.FullName  # e.g. "D:\\Projects\\MyDrawing.dwg"
        folder_path = os.path.dirname(full_path)
        return folder_path
    except Exception as e:
        print(f"❌ 无法获取当前 DWG 文件夹：{e}")
        return None


#&&% 基础函数


def vtpnt(x, y, z):
    """坐标点转化为浮点数"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (x, y, z))

def vtobj(obj):
    """转化为对象数组"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj)

def vtFloat(lst):
    """列表转化为浮点数"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, lst)
    
def vtInt(lst):
    """列表转化为整数"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, lst)

def vtVariant(lst):
    """列表转化为变体"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, lst)

def ConvertArrays2Variant(inputdata, vartype):#例如vartype="Variant"
    import pythoncom
    if vartype == "ArrayofObjects":  # 对象数组
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, inputdata)
    if vartype == "Double":  # 双精度
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, inputdata)
    if vartype == "ShortInteger":  # 短整型
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, inputdata)
    if vartype == "LongInteger":  # 长整型
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I4, inputdata)
    if vartype == "Variant":  # 变体
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, inputdata)
    return outputdata

def vtlist(obj_list):#将com对象列表转换为可用win32数据
    """

    将对象列表转为 VARIANT 类型以供 COM 接口使用

    依赖语句 from win32com.client import VARIANT

    """
    return VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj_list)


warnings.filterwarnings("ignore", category=UserWarning)



#&&% 打开天正start_applicationV9

def start_applicationV9(
    PTH: str = r"C:\Tangent\TArchT20V9",
    max_retries: int = 3,
    retry_delay: float = 2.0
) -> subprocess.Popen | None:
    """
    启动天正 TGStart.exe，失败重试。
    返回 Popen 对象（成功）或 None（失败）。
    """
    exe = os.path.join(PTH, "TGStart.exe")
    for attempt in range(1, max_retries + 1):
        try:
            proc = subprocess.Popen(
                [exe],
                creationflags=DETACHED_PROCESS,
                cwd=PTH
            )
            print("🚀 启动天正CAD 成功")
            return proc
        except Exception as e:
            print(f"第 {attempt} 次启动失败: {e}")
            if attempt < max_retries:
                print(f"等待 {retry_delay:.1f} 秒后重试…")
                time.sleep(retry_delay)
    print(f"已达最大重试次数 ({max_retries})，启动失败。")
    return None




def get_acad_process_id(ming):#获取进程的ID
    for process in psutil.process_iter(attrs=['pid', 'name']):
        if str(ming) in process.info['name'].lower():
            return process.info['pid']
    return None
    
def jingchengshu_wenjian():#查看cad进程数

    CAD_PROCESS_NAME = "acad.exe"

    found_process_i = 0
    for process in psutil.process_iter(['pid', 'name']):
        if process.info['name'] == CAD_PROCESS_NAME:
            found_process_i = found_process_i+1            

    return  found_process_i 
    
def close_all_cad_processes():#关闭所有进程
    max_retries = 3
    for attempt in range(max_retries):
        success = True
        for process in psutil.process_iter(['pid', 'name']):
            if process.info['name'] == "acad.exe":
                try:
                    process.kill()
                    time.sleep(2)
                except Exception as e:
                    print(f"尝试关闭 CAD 进程失败: {e}")
                    success = False
                    break
        if success:
            return
        else:
            print(f"关闭 CAD 进程失败，正在重试... 尝试次数: {attempt + 1}")
            time.sleep(2)
    print("多次尝试关闭 CAD 进程失败，请检查系统。")
    
def close_oldest_cad_process(process_name="acad.exe"):#关闭上一个进程
    cad_processes = [p for p in psutil.process_iter(['pid', 'name', 'create_time']) if p.info['name'] == process_name]
        
    # 检查是否有多个CAD进程
    if len(cad_processes) > 1:
        # 按启动时间排序，最早的进程在前
        oldest_process = sorted(cad_processes, key=lambda p: p.info['create_time'])[0]
            
        try:
            # 关闭最早的进程
            psutil.Process(oldest_process.pid).terminate()
            print(f"已关闭最早的CAD进程，PID: {oldest_process.pid}")
        except Exception as e:
            print(f"关闭进程时出错: {e}")
            pass
    else:
        print("没有多个CAD进程运行。")
        pass


#&&%  cad连接、关闭

def li():#这里的li()不需要写成多次尝试
    global acad, doc, mp, sp
    acad, doc = get_acad_doc()
    mp = doc.ModelSpace
    sp = doc.PaperSpace
    print("当前桌面文件：", doc.Name)
    print("win32已经连接正常—CAD基本操作")
    return True





_RPC_BUSY = (-2147417846, -2147418111)  # 应用程序忙/Call rejected
_RPC_DOWN = (-2147023174,)              # RPC 服务器不可用

def com_retry(fn, retries=30, delay=0.05):
    for _ in range(retries):
        try:
            return fn()
        except pywintypes.com_error as e:
            code = e.args[0] if e.args else None
            if code in _RPC_BUSY + _RPC_DOWN:
                time.sleep(delay); continue
            raise
    return fn()

def _coinit_once():
    try: pythoncom.CoInitialize()
    except pythoncom.error: pass

def ensure_typelib_from_running():
    """从正在运行的 AutoCAD 直接生成 makepy（不依赖注册表）"""
    _coinit_once()
    app = win32.gencache.EnsureDispatch("AutoCAD.Application")
    tlb, _ = app._oleobj_.GetTypeInfo().GetContainingTypeLib()
    makepy.GenerateFromTypeLibSpec(tlb)

def get_acad_doc(max_wait=8.0):
    _coinit_once()
    t0 = time.time()
    app = None
    while True:
        try:
            app = win32.gencache.EnsureDispatch("AutoCAD.Application")
            doc = app.ActiveDocument
            _ = doc.Name
            return app, doc
        except pywintypes.com_error as e:
            code = e.args[0] if e.args else None
            if (code in _RPC_BUSY + _RPC_DOWN) and (time.time()-t0 < max_wait):
                time.sleep(0.05); continue
            # 自愈：若无文档则新建
            try:
                if app is None:
                    app = win32.gencache.EnsureDispatch("AutoCAD.Application")
                doc = app.Documents.Add()
                _ = doc.Name
                return app, doc
            except Exception:
                pass
            raise

class _ComLiveProxy:
    """把 getter 包成“活体”对象：每次访问都拿最新/可用的 COM 实例"""
    def __init__(self, getter): self._getter = getter
    def __getattr__(self, name):
        obj = self._getter()
        return getattr(obj, name)
    # 可选：允许直接作为可迭代或 bool 使用时更自然
    def __dir__(self): return dir(self._getter())

# ===== 你新的 li()（装载代理），兼容旧写法 =====


def li_new():#20250924新版本
    """
    连接/修复 CAD 会话，并把 acad/doc/mp/sp 设置为“活体代理”。
    原有大量函数里继续写 acad./doc./mp./sp. 即可，但它们会自动取到当前有效对象。
    """
    global acad, doc, mp, sp
    # 先确保 makepy 存在（CastTo/强类型更稳）
    try:
        ensure_typelib_from_running()
    except Exception:
        pass

    # 安装代理（getter 内部包含重试与自愈）
    acad = _ComLiveProxy(lambda: get_acad_doc()[0])
    doc  = _ComLiveProxy(lambda: get_acad_doc()[1])
    mp   = _ComLiveProxy(lambda: get_acad_doc()[1].ModelSpace)
    sp   = _ComLiveProxy(lambda: get_acad_doc()[1].PaperSpace)

    # 轻量校验 & 提示
    name = com_retry(lambda: doc.Name)
    print("当前桌面文件：", name)
    print("win32已经连接正常—CAD基本操作")
    return True
    



"""
@require_doc，它最强大的地方在于自动跟随cad文件的切换，跟踪激活文件？更快的_.doc=get_cad_doc()模式则是确定没有文件切换时更快




"""

#装饰器动态模式

def require_doc(fn):
    def wrapper(*args, doc=None, **kw):
        if doc is None:
            _, doc = get_acad_doc()
        return fn(*args, doc=doc, **kw)
    return wrapper

def rpc_safe(fn):
    def wrapper(*a, **k):
        return com_retry(lambda: fn(*a, **k))
    return wrapper

#只读类函数模板

@require_doc
@rpc_safe
def read_active_doc_info(doc=None):
    """示例：读取当前图名、空间实体数量"""
    name = com_retry(lambda: doc.Name)
    msp  = doc.ModelSpace
    count = com_retry(lambda: msp.Count)
    return {"name": name, "model_count": count}

#修改类函数模板（带撤销、失败也能收尾
@require_doc
def move_selected_by(dx, dy, dz=0.0, doc=None):
    """示例：交互拾取若干对象并平移"""
    ents = pmxz(prompt="\n选择要平移的对象，回车结束：")
    if not ents:
        print("未选择对象"); return 0

    com_retry(lambda: doc.StartUndoMark())
    moved = 0
    try:
        for e in ents:
            com_retry(lambda e=e: e.Move((0,0,0), (dx,dy,dz)))
            moved += 1
    finally:
        com_retry(lambda: doc.EndUndoMark())
    print(f"✅ 已平移 {moved} 个对象")
    return moved

#命令行交互模板（SendCommand + 等待）
"""
主要是为了稳定性和可控性，区别在于：

直接用 doc.SendCommand()

最快，直接把字符串送进 CAD 命令行。

但要自己加 \n (chr(13))，否则命令不会生效。

没有等待/确认逻辑，CAD 还没执行完，你的 Python 可能就继续往下跑了。

封装过的模板（SendCommand + 等待）

会在 SendCommand 后自动 time.sleep(x) 或循环检测 CAD 是否空闲（如 While acad.GetAcadState().IsQuiescent == False）。

可以加日志：执行了什么命令、耗时多久。

遇到 CAD 忙碌 / 拒绝调用时，可以自动重试。

一般还会做“换行清理”，避免上一条命令残留影响下一条。

返回结构：尽量返回明确的字典/元组/数量，便于上层组合使用。

异常：普通 COM 波动交给 com_retry；真正的逻辑错误（类型不符、数据不全）直接抛 RuntimeError 或返回 None 并打印友好信息。





"""

def send_cmd(cmd: str, wait_s: float = 0.3):
    """安全发送命令并小等待，避免命令堆积"""
    _, doc = get_acad_doc()
    doc.SendCommand(cmd if cmd.endswith("\n") else (cmd + "\n"))
    time.sleep(wait_s)

@require_doc
def zoom_window(x1, y1, x2, y2, pad_ratio=0.1, doc=None):
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)
    pad = pad_ratio * ((x_hi-x_lo + y_hi-y_lo)/2.0)
    send_cmd(f"_.ZOOM\n_W\n{x_lo-pad},{y_lo-pad}\n{x_hi+pad},{y_hi+pad}\n", wait_s=0.6)



#选择 + Cast + 读取属性模板

@require_doc
def pick_line_and_report(doc=None):
    """交互选直线并输出起终点、长度"""
    ents = ss_select("onscreen", filter_types=[0], filter_data=["LINE"], autocast=True,
                     prompt="\n请选择一根直线：")
    if not ents:
        print("未选到直线"); return None
    ln = ents[0]  # IAcadLine（autocast 已转换）
    sp = com_retry(lambda: ln.StartPoint)
    ep = com_retry(lambda: ln.EndPoint)
    L  = com_retry(lambda: ln.Length)
    info = {"start": tuple(sp), "end": tuple(ep), "length": L}
    print(info)
    return info


#批处理模板（不交互、按过滤批量处理）

"""
🔹 为什么我单独提出“批处理模板”？

因为光有 for 循环 还不够，实际在 CAD 中会遇到一些坑：

撤销保护
如果你循环 1000 个对象，用户后悔时，Ctrl+Z 可能要点 1000 次。
→ 用 doc.StartUndoMark() / doc.EndUndoMark() 包起来，只需撤销一次。

RPC 错误 / CAD 忙
在循环里有时会报 RPC 服务器不可用，所以每个对象的属性访问最好包一层 com_retry。

性能
如果每次循环都重新 get_acad_doc()，会很慢。
→ 批处理模板里只获取一次 doc，然后在循环中直接用。

安全清理
SelectionSet 要记得 Delete()，否则下一次可能报错。

🔹 一个最简单的批处理例子
@require_doc
def batch_move_all_lines_to_layer0(doc=None):
    
    # 1) 选择
    ents = ss_select("all", filter_types=[0], filter_data=["LINE"], autocast=True)
    if not ents:
        print("❌ 没有找到直线"); return 0

    # 2) 撤销块
    com_retry(lambda: doc.StartUndoMark())
    moved = 0
    try:
        # 3) 循环处理
        for e in ents:
            com_retry(lambda: setattr(e, "Layer", "0"))
            moved += 1
    finally:
        # 4) 结束撤销块
        com_retry(lambda: doc.EndUndoMark())

    print(f"✅ 批处理完成，共移动 {moved} 条直线到图层 0")
    return moved


"""



@require_doc
def flatten_all_circles_to_z0(doc=None):
    """把所有圆的 Z 置 0（示例）"""
    circles = ss_select("all", filter_types=[0], filter_data=["CIRCLE"], autocast=True)
    if not circles:
        print("未找到圆"); return 0
    com_retry(lambda: doc.StartUndoMark())
    n = 0
    try:
        for c in circles:
            # 移到 Z=0（Circle 没有 Elevation，就 Transform）
            sp = com_retry(lambda: c.Center)
            com_retry(lambda: c.Move(sp, (sp[0], sp[1], 0.0)))
            n += 1
    finally:
        com_retry(lambda: doc.EndUndoMark())
    print(f"✅ 处理圆 {n} 个")
    return n


#&&&% RGB色彩

def aci_to_rgb(ci: int):
    # 只给出常用 ACI 的近似色；其余返回 None
    table = {
        1:(255,0,0), 2:(255,255,0), 3:(0,255,0), 4:(0,255,255),
        5:(0,0,255), 6:(255,0,255), 7:(255,255,255)
    }
    return table.get(int(ci)) if ci is not None else None
def get_entity_rgb(ent):
    """返回 (rgb_tuple_or_None, source_str)"""
    # 1) 先看 TrueColor 的方式
    tc = ent.TrueColor                 # IAcadAcCmColor
    method = getattr(tc, "ColorMethod", None)
    # acColorMethodByRGB 大多为 3（不同版本枚举值可能不同，但逻辑相同）
    if method == 3 or (hasattr(tc, "Red") and (tc.Red or tc.Green or tc.Blue)):
        return (tc.Red, tc.Green, tc.Blue), "ByRGB(TrueColor)"

    # 2) 看 ColorIndex / Color（ACI / ByLayer / ByBlock）
    ci = getattr(ent, "ColorIndex", None)
    if ci is None:
        ci = getattr(ent, "Color", None)

    # ByLayer
    if ci == 256:
        layer_name = ent.Layer
        layer = ent.Document.Layers.Item(layer_name)
        ltc = layer.TrueColor
        return (ltc.Red, ltc.Green, ltc.Blue), f"ByLayer({layer_name})"

    # ByBlock
    if ci == 0:
        # 真正的显示颜色取决于引用它的块参照/上级容器
        return None, "ByBlock(需根据上级块参照解析)"

    # 3) 纯 ACI（1..255）
    if isinstance(ci, int) and 1 <= ci <= 255:
        rgb = aci_to_rgb(ci)
        return (rgb if rgb else None), f"ACI({ci})"

    # 兜底
    return None, "Unknown"


"""
lj=pmxz()
pl=lj[0]
rgb, src = get_entity_rgb(pl)
print("RGB:", rgb, "| 来源:", src)
RGB: (255, 0, 0) | 来源: ByRGB(TrueColor)
"""

    
def  guanbi_cad_doc():#关闭所有cad文件

        
    shuliang_cad = acad.Documents.Count

    print("当前桌面打开的cad文件数量:",shuliang_cad)

    for i in range(0,shuliang_cad):

        CourentDoc = acad.ActiveDocument

        CourentDoc.Close()

        li()
    
def guanbi_cad_one():#关闭所有cad文件但保留一个并连接
    shuliang_cad = acad.Documents.Count
    print("当前桌面打开的cad文件数量:", shuliang_cad)

    if shuliang_cad > 1:
        for i in range(shuliang_cad - 1):  # 保留最后一个文件
            courent_doc = acad.ActiveDocument
            courent_doc.Close()

        # 重新连接到最后一个文件
        if not li():
            print("重新连接到最后一个文件失败，请检查系统。")
    else:
        print("只有一个CAD文件打开，无需关闭其他文件。")

def st():#启动或连上CAD开始工作
    """
    归零重新归于标准态的处理

    """
    cad_process_count = jingchengshu_wenjian()

    if cad_process_count > 1:
        close_all_cad_processes()
        start_applicationV9()
        li()#生成全局变量acad,mp,doc
    elif cad_process_count == 0:
        start_applicationV9()
        li()#生成全局变量acad,mp,doc
    elif cad_process_count == 1:
        if not li():
            close_all_cad_processes()
            start_applicationV9()



def huifu_xitong():#启动或连上CAD开始工作
    """
    归零重新归于标准态的处理

    """
    close_all_cad_processes()

    start_applicationV9()




#&&&%  * 重复多次调用函数

def chongfu_caozuo(
    Fx,
    dwg_instance=None,
    args: tuple=(),
    kwargs: dict=None,
    max_retries: int=3,
    failure_value=None
):
    """
    对指定函数/方法进行重复调用，直到成功或耗尽重试次数。

    :param Fx:             待调用的函数对象（或方法），不要在这里写 Fx()
    :param dwg_instance:   如果要调用类的方法，就把实例传进来；否则为 None
    :param args:           位置参数，按顺序传给 Fx
    :param kwargs:         关键字参数 dict，传给 Fx
    :param max_retries:    最多尝试次数（>=1）
    :param failure_value:  全部尝试失败时的返回值

    :return: (result, attempts, error)
        - result:   Fx 的返回值，或 failure_value
        - attempts: 成功时为那次尝试编号（1-based），失败时等于 max_retries
        - error:    最后一次捕获到的 Exception 实例；成功时为 None
    """
    if kwargs is None:
        kwargs = {}

    last_exception = None

    for attempt in range(1, max_retries + 1):
        try:
            # 决定如何调用 Fx
            if dwg_instance is not None and not getattr(Fx, "__self__", None):
                # Fx 很可能是类中定义的函数，但还未绑定到实例
                result = Fx(dwg_instance, *args, **kwargs)
            else:
                # Fx 已经是普通函数或已绑定的方法
                result = Fx(*args, **kwargs)

            # 成功：返回 (结果, 尝试次数, None)
            return result, attempt, None

        except Exception as e:
            last_exception = e
            if attempt < max_retries:
                print(f"⚠ 操作出错，正在第 {attempt} 次重试…")
                # （可选）在重试前做些环境刷新、移动窗口等
                try:    li()  # 示例：重新排列 CAD/IDLE 窗口
                except: pass
                time.sleep(2)
            else:
                print("❌ 多次尝试均失败，请检查环境或参数是否正确。")
                # 返回 (failure_value, 用尽的尝试次数, 最后一个异常)
                return failure_value, attempt, last_exception


#  自动计时装饰器

"""
用法
@simple_timer
def heavy_func():
    ...

heavy_func()

"""
def simple_timer(func):
    def wrapper(*args, **kwargs):
        t0 = time.time()
        result = func(*args, **kwargs)
        t1 = time.time()
        print(f"⏱️ {func.__name__} 耗时：{t1 - t0:.4f} 秒")
        return result
    return wrapper            









#&&&&%% 第二部分 列表处理


"""
该模块研究列表的基本问题，特别是com对象列表 


有一个针对全部程序系统的基本文件处理文件Basic_oper.py，例如字符串，列表，集合的操作等等，便于我们展开别的专项工作时引用而不必重复生成。

在每个专项工作文件中，仍然有一套自己的基本处理文件，放在脚本的最前面，这两者并不重复，而且很合理。我们将在专项工作中集中这些基本处理，

将来再移植到Basic_oper.py。虽然函数重复出现在不同的脚本，但并不会混乱，也不会降低效率。这反映了内在的本质规律：Basic_oper.py

不是专门开发研究的成果，而是来自各个专项工作的总结。


"""


# 按com实体对象中提取的坐标排序


def sort_tuples(lst,cha_Y =2000):#对列表按插入点xy坐标排序
    
    """
    这是很有用的一个双值排序函数，对于COM对象，可以先将其转换为元组，即可使用这个函数

    它的价值在于，很容易拓展到n值排序
    """
    
    # 先按照m[1]降序排序
    lst.sort(key=lambda x: -x[2][1])

    i = 0
    while i < len(lst) - 1:
        j = i + 1
        # 查找所有m[1]差距在chaY以内的元素
        while j < len(lst) and abs(lst[i][2][1] - lst[j][2][1]) < cha_Y:
            j += 1
        
        # 如果找到了m[1]值相近的元素，根据m[0]值进行排序
        if j - i > 1:
            lst[i:j] = sorted(lst[i:j], key=lambda x: x[2][0])
        
        i = j

    return lst

def multi_dim_tolerance_sort(lst, key_index=2, tolerances=[10000, 1000, 0]):#高维排序
    """
    对 lst 列表中的元组按多维坐标字段排序，考虑每个维度的容差进行逐层排序。

    参数：
        lst: [(id, name, (x, y, z)), ...]
        key_index: 坐标在元组中的索引（默认是第3项，即元组[2]）
        tolerances: 每个维度允许的容差，例如 [Z差, Y差, X差]

    返回：
        排好序的新列表
    """
    # 最外层排序：按最高维降序排（Z从上往下）
    dim = len(tolerances)
    lst.sort(key=lambda x: -x[key_index][0])  # Z 倒序

    def recursive_sort(sublist, level):
        if level >= dim - 1:
            # 最后一级直接按该维升序
            return sorted(sublist, key=lambda x: x[key_index][level + 1])
        
        i = 0
        while i < len(sublist) - 1:
            j = i + 1
            while j < len(sublist) and abs(sublist[i][key_index][level] - sublist[j][key_index][level]) < tolerances[level]:
                j += 1
            if j - i > 1:
                sublist[i:j] = recursive_sort(sublist[i:j], level + 1)
            i = j
        return sublist

    return recursive_sort(lst, 0)


def get_ll_pt(ent):#提取函数，对象左下角点
    minpt, _ = ent.GetBoundingBox()
    return minpt[0], minpt[1],0

def get_center(ent):#提取函数，中心点
    minpt, maxpt = ent.GetBoundingBox()
    return ((minpt[0]+maxpt[0])/2, (minpt[1]+maxpt[1])/2)

def sort_entities_by_position( entity_list, extract_func, cha_Y=2000):#对com对象按提取坐标分别沿y,x方向排序
        """
        对实体列表根据其坐标（通过 extract_func 获取）进行排序：
        - 先按 Y 值降序（从上到下）
        - Y 值接近（差值 < cha_Y）者再按 X 值升序（从左到右）

        参数：
            entity_list: COM 实体对象列表
            extract_func: 提取坐标函数，返回 (x, y) 元组的函数即可
            cha_Y: 同一行判定的 Y 方向容差

        返回：按坐标顺序排列的新实体对象列表

        调用示例

        sorted_objs = sort_entities_by_position(LB, extract_func=get_ll_pt)
        
        """
        triples = [(ent, *extract_func(ent)) for ent in entity_list]

        # 按 Y 值降序排列
        triples.sort(key=lambda t: -t[2])

        i = 0
        while i < len(triples) - 1:
            j = i + 1
            while j < len(triples) and abs(triples[i][2] - triples[j][2]) < cha_Y:
                j += 1
            triples[i:j] = sorted(triples[i:j], key=lambda t: t[1])  # 按 X 升序
            i = j

        return [t[0] for t in triples]

def get_line_start(ent):
    """
    提取一条直线的起点 (x, y)
    ent: AcDbLine 或类似对象，具有 .StartPoint 属性
    """
    pt = ent.StartPoint   # 假设是一个 (x, y, z) 或 [x, y, z]
    return pt[0], pt[1]

#&&% * 对列表实体进行从上到下、从左到右的排序

def sort_coms_by_llcorner(com_list, cha_Y=2000):
    """
    按 BoundingBox 左下角坐标排序：
      · 先按 y 降序（越大越靠前，即自上而下）
      · 同一行(Δy<cha_Y)内按 x 升序（自左向右）
    """
    wrapped = []
    for ent in com_list:
        try:
            p1, _ = ent.GetBoundingBox()      # p1 已是左下
            x_ll, y_ll = p1[0], p1[1]
        except Exception:
            x_ll = y_ll = float('-inf')       # 取不到的一律放最后
        wrapped.append((ent, x_ll, y_ll))     # (实体, x, y)

    # 先按 y 降序
    wrapped.sort(key=lambda t: -t[2])

    i = 0
    while i < len(wrapped) - 1:
        j = i + 1
        while j < len(wrapped) and abs(wrapped[i][2] - wrapped[j][2]) < cha_Y:
            j += 1
        # 行内再按 x 升序
        if j - i > 1:
            wrapped[i:j] = sorted(wrapped[i:j], key=lambda t: t[1])
        i = j

    return [ent for ent, _, _ in wrapped]




def sort_coms_by_rbcorner(com_list, *, cha_X=100):
    """
    竖向图框（或已整体旋转 -90° 的图纸）使用 ——  
    · 先按列：x₁ 降序（右 → 左）  
    · 列内：  y₂ 降序（上 → 下）

    参数
    ----
    com_list : list[COM object]
        文字实体列表
    cha_X    : float
        判定“同一列”的容差，单位与图纸坐标一致
    """
    wrapped = []
    for ent in com_list:
        try:
            (x1, _, _), (_, y2, _) = ent.GetBoundingBox()  # x1 = 左, y2 = 上
        except Exception:
            x1 = y2 = float("-inf")                        # 失败的排最后
        wrapped.append((ent, x1, y2))

    # — ① 按 x₁ 降序：最右列在前 —
    wrapped.sort(key=lambda t: -t[1])

    # — ② 同列内再按 y₂ 降序：上 → 下 —
    i = 0
    while i < len(wrapped) - 1:
        j = i + 1
        while j < len(wrapped) and abs(wrapped[i][1] - wrapped[j][1]) < cha_X:
            j += 1
        wrapped[i:j] = sorted(wrapped[i:j], key=lambda t: -t[2])
        i = j

    return [t[0] for t in wrapped]


def sort_coms_by_llcorner_custom(objs, tol_x=100):
    """
    按左下角 x, y 坐标对 COM 对象列表 objs 排序：
      1. 先按 x 升序分组：相邻 x 差 ≤ tol_x 视为同一组
      2. 每组内按 y 降序排列（y 大的排前）
      3. 最终把各组按它们的 x 升序依次拼接

    参数
    ----
    objs : List[COMObject]
        要排序的 COM 对象列表，每个对象必须支持 GetBoundingBox()
    tol_x : float
        x 方向上的容差，小于等于此值的 x 差视为同组

    返回
    ----
    List[COMObject]
        排序后的对象列表
    """
    # 提取 (obj, llx, lly)
    items = []
    for e in objs:
        ll, ur = e.GetBoundingBox()
        llx, lly, _ = ll
        items.append((e, llx, lly))

    # 按 x 升序初排
    items.sort(key=lambda t: t[1])

    # 分组：相邻 x 差 ≤ tol_x 的归一组
    clusters = []
    rep_x, current = items[0][1], [items[0]]
    for item in items[1:]:
        _, x, _ = item
        if abs(x - rep_x) <= tol_x:
            current.append(item)
        else:
            clusters.append((rep_x, current))
            rep_x, current = x, [item]
    clusters.append((rep_x, current))

    # 按组的 x 升序排好后，组内按 y 降序，再平铺
    result = []
    for _, group in sorted(clusters, key=lambda c: c[0]):
        group_sorted = sorted(group, key=lambda t: t[2], reverse=True)
        result.extend([t[0] for t in group_sorted])

    return result


def sort_coms_by_center(objs, tol_x=100):
    """
    按外包盒中心坐标对 COM 对象列表 objs 排序：
      1. 先计算每个对象的 bbox 中心 (cx, cy)
      2. 按 cx 升序初排序
      3. 将相邻 cx 差 ≤ tol_x 的对象归同一组
      4. 各组按 cy 降序排列（cy 大的排前）
      5. 最后按组的 cx 升序依次拼接各组内对象

    参数
    ----
    objs : List[COMObject]
        要排序的 COM 对象列表，每个对象必须支持 GetBoundingBox()
    tol_x : float
        x 方向上的容差，小于等于此值的 cx 差视为同组

    返回
    ----
    List[COMObject]
        排序后的对象列表
    """
    # 1. 提取 (obj, cx, cy)
    items = []
    for e in objs:
        ll, ur = e.GetBoundingBox()
        cx = (ll[0] + ur[0]) / 2.0
        cy = (ll[1] + ur[1]) / 2.0
        items.append((e, cx, cy))

    if not items:
        return []

    # 2. 按 cx 升序初排
    items.sort(key=lambda t: t[1])

    # 3. 分组：相邻 cx 差 ≤ tol_x 的归一组
    clusters = []
    rep_x, current = items[0][1], [items[0]]
    for item in items[1:]:
        _, x, _ = item
        if abs(x - rep_x) <= tol_x:
            current.append(item)
        else:
            clusters.append((rep_x, current))
            rep_x, current = x, [item]
    clusters.append((rep_x, current))

    # 4. 各组内部按 cy 降序排列，再平铺
    result = []
    for rep_x, group in sorted(clusters, key=lambda c: c[0]):
        group_sorted = sorted(group, key=lambda t: t[2], reverse=True)
        result.extend([t[0] for t in group_sorted])

    return result


##对列表实体进行正序或逆序编号

def number_entities_by_order(entity_list, prefix="", start=1, k=0):
    """
    对排序好的 COM 实体对象列表进行编号。

    参数：
        entity_list: 实体对象列表
        prefix: 编号前缀，默认为空字符串
        start: 编号起始数值，默认为 1
        k: 排序方向控制变量：
            - k = 0 正序编号
            - k = 1 逆序编号

    返回：
        编号字符串列表（如 ["1", "2", "3"] 或 ["图1", "图2", "图3"]）
    """
    n = len(entity_list)
    index_list = range(n) if k == 0 else reversed(range(n))
    
    result = []
    for i, idx in enumerate(index_list):
        label = f"{prefix}{start + i}"
        result.append(label)
    
    return result


# 重复操作列表对象

def apply_to_each1(obj_list,  action_func):#重复操作列表对象
    """
    对 obj_list 中的每个对象，
    传入 action_func 中处理。

    参数：
        obj_list: 对象列表
        
        action_func: 用于处理提取结果的函数（如 srhd）
    """
    for obj in obj_list:
        
        action_func(obj)
        
def apply_to_each2(obj_list, extract_func, action_func):#双层嵌套重复操作列表对象
    """
    对 obj_list 中的每个对象，先通过 extract_func 提取值，
    再将该值传入 action_func 中处理。

    参数：
        obj_list: 对象列表
        extract_func: 用于提取 (x, y) 或其他值的函数
        action_func: 用于处理提取结果的函数（如 srhd）
    """
    for obj in obj_list:
        value = extract_func(obj)
        action_func(value)

#&&&&%% 第三部分 文件夹文件处理


"""
该模块是有关一般性的文件夹、文件上的操作 

"""
#&&% 确保文件被删除
def ensure_file_absent(save_path: str, ty: float = 1.0) -> None:
    """
    确保指定路径的文件不存在。如果存在，则删除并等待 ty 秒。

    :param save_path: 带路径的文件名，例如 "D:/output/result.txt"
    :param ty: 删除后等待的秒数，默认为 1 秒
    """
    try:
        if os.path.isfile(save_path):
            os.remove(save_path)
            # 等待 ty 秒，以确保文件系统完成删除操作
            time.sleep(ty)
            print(f"✅ 已删除文件：{save_path}，并等待了 {ty} 秒")
        else:
            print(f"ℹ 文件不存在，无需删除：{save_path}")
    except Exception as e:
        print(f"❌ 删除文件时出错：{e}")


def traverse_with_os_walk(root_dir: str):
    """
    遍历 root_dir 及其所有子目录，打印每个目录和文件的完整路径。
    """
    for dirpath, dirnames, filenames in os.walk(root_dir):
        # dirpath: 当前遍历到的目录路径
        print(f"Directory: {dirpath}")
        for dirname in dirnames:
            print(f"  Sub-dir: {os.path.join(dirpath, dirname)}")
        for filename in filenames:
            print(f"  File   : {os.path.join(dirpath, filename)}")



def find_files_with_extensions(directory, extensions):
    #找到以[".dwg"]结尾的文件directory为文件夹及其路径，extensions为后缀或中间位置字符列表
    matching_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if any(file.lower().endswith(ext) for ext in extensions):
                matching_files.append(os.path.join(root, file))
    return matching_files

def get_filename_without_extension(FileandPath):
    # 从完整路径中获取文件名（包含扩展名）
    filename_with_extension = os.path.basename(FileandPath)
    
    # 分离文件名和扩展名
    filename_without_extension, _ = os.path.splitext(filename_with_extension)

    return filename_without_extension


def delete_files_with_patterns(folder_path, patterns):
    """
    删除文件夹中符合指定模式的文件。

    Args:
        folder_path (str): 文件夹的路径。
        patterns (list): 包含要匹配的模式的列表，例如["_t7", "_t3", "_t8"]。

    Returns:
        None
    """
    # 遍历文件夹中的所有文件
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        
        # 检查文件名是否包含任何一个指定的模式
        if any(pattern in filename for pattern in patterns):
            # 如果符合条件，就删除该文件
            os.remove(file_path)
            print(f"Deleted file: {filename}")
            pass

    print("File deletion completed.")
    pass


#确保文件夹中名字含有特殊字符的文件被清空

def clear_files_with_prefix(folder: str, filename_prefix: str = "区域导出", delay: float = 0.5):
    """
    清除指定文件夹中所有文件名包含给定前缀的文件。若文件正在被占用或删除失败，会尝试重试一次。

    :param folder:           要清理的文件夹路径
    :param filename_prefix:  需要匹配的文件名前缀（只要文件名中包含该字符串，就会被删除）
    :param delay:            删除失败时的等待时间（秒），默认 0.5 秒
    """
    if not os.path.isdir(folder):
        print(f"❌ 目标路径不是有效文件夹：{folder}")
        return

    # 列出文件夹中所有条目
    entries = os.listdir(folder)
    # 过滤出文件名中包含指定前缀的文件
    to_delete = [fname for fname in entries if filename_prefix in fname and os.path.isfile(os.path.join(folder, fname))]

    if not to_delete:
        print(f"ℹ️ 文件夹中未发现文件名包含 “{filename_prefix}” 的文件。")
        return

    for fname in to_delete:
        full_path = os.path.join(folder, fname)
        try:
            os.remove(full_path)
            print(f"✅ 已删除：{fname}")
        except Exception as e:
            print(f"⚠ 删除失败（第一次尝试）：{fname}，错误：{e}，稍后重试……")
            time.sleep(delay)
            # 再试一次
            try:
                os.remove(full_path)
                print(f"✅ 重试成功删除：{fname}")
            except Exception as e2:
                print(f"❌ 再次删除仍失败：{fname}，错误：{e2}")



def find_files_with_string(directory, search_string):
    #找到文件夹中含有指定字符串的文件
    matched_files = []
    for file in os.listdir(directory):
        if search_string in file:
            matched_files.append(file)
    return matched_files

#使路径名和文件名合并后合乎预期
def join_paths(p1, p2):
    # 使用 os.path.join 合并路径
    result = os.path.join(p1, p2)
    # 替换反斜杠为正斜杠
    return result.replace("\\", "/")



#&&&&%% 第四部分 选择方法 


# ===================== 前置工具（统一处理） =====================

# —— 通用重试：吞“忙/拒绝/RPC down”
_RPC_BUSY = (-2147417846, -2147418111)
_RPC_DOWN = (-2147023174,)

def com_retry(fn, retries=30, delay=0.05):
    for _ in range(retries):
        try:
            return fn()
        except pywintypes.com_error as e:
            code = e.args[0] if e.args else None
            if code in _RPC_BUSY + _RPC_DOWN:
                time.sleep(delay); continue
            raise
    return fn()

# —— 现取现用：获取当前激活文档（无文档则新建）
def get_acad_doc(max_wait=8.0):
    try: pythoncom.CoInitialize()
    except pythoncom.error: pass
    t0 = time.time()
    app = None
    while True:
        try:
            app = win32.gencache.EnsureDispatch("AutoCAD.Application")
            doc = app.ActiveDocument
            _ = doc.Name
            return app, doc
        except pywintypes.com_error as e:
            code = e.args[0] if e.args else None
            if (code in _RPC_BUSY + _RPC_DOWN) and (time.time()-t0 < max_wait):
                time.sleep(0.05); continue
            # 无打开图 → 新建
            try:
                if app is None:
                    app = win32.gencache.EnsureDispatch("AutoCAD.Application")
                doc = app.Documents.Add(); _ = doc.Name
                return app, doc
            except Exception:
                pass
            raise

# —— 过滤数组（避免与你库已有 vtInt/vtVariant 冲突）
def to_vt_int(seq):
    return VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, list(seq))

def to_vt_variant(seq):
    return VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, list(seq))

# —— 常见对象自动 Cast：拿到专接口（StartPoint/Coordinates/Contents…）
_CAST_MAP = {
    # 基础几何
    "AcDbLine":"IAcadLine", "AcDbCircle":"IAcadCircle", "AcDbArc":"IAcadArc","AcDbPoint":"IAcadPoint",
    "AcDbEllipse":"IAcadEllipse", "AcDbSpline":"IAcadSpline",
    # 多段线
    "AcDbPolyline":"IAcadLWPolyline", "AcDb2dPolyline":"IAcadPolyline", "AcDb3dPolyline":"IAcad3DPolyline",
    # 文字
    "AcDbText":"IAcadText", "AcDbMText":"IAcadMText",
    # 块/属性
    "AcDbBlockReference":"IAcadBlockReference",
    "AcDbAttribute":"IAcadAttributeReference", "AcDbAttributeDefinition":"IAcadAttribute",
    # 引线/标注（常用）
    "AcDbLeader":"IAcadLeader", "AcDbMLeader":"IAcadMLeader",
    "AcDbDimension":"IAcadDimension", "AcDbAlignedDimension":"IAcadDimAligned",
    "AcDbRotatedDimension":"IAcadDimRotated", "AcDbRadialDimension":"IAcadDimRadial",
    "AcDbDiametricDimension":"IAcadDimDiametric", "AcDbArcDimension":"IAcadDimArc",
    "AcDb3PointAngularDimension":"IAcadDim3PointAngular", "AcDb2LineAngularDimension":"IAcadDim2LineAngular",
    "AcDbOrdinateDimension":"IAcadDimOrdinate",
    # 其它
    "AcDbHatch":"IAcadHatch", "AcDbTable":"IAcadTable",
}

def _maybe_cast(ent):
    try:
        name = com_retry(lambda: ent.ObjectName)
        iface = _CAST_MAP.get(name)
        if iface:
            try: return CastTo(ent, iface)
            except Exception: return ent
        return ent
    except Exception:
        return ent

# —— 统一选择器：all/window/crossing/onscreen
# mode: "all" | "window" | "crossing" | "onscreen"
def ss_select(mode="all", p1=None, p2=None, filter_types=None, filter_data=None, autocast=True, prompt=None):
    _, doc = get_acad_doc()
    ss_name = f"SS_{uuid.uuid4().hex[:8]}"
    # 清理旧名 + 新建
    try: doc.SelectionSets.Item(ss_name).Delete()
    except Exception: pass
    ss = doc.SelectionSets.Add(ss_name)

    try:
        if mode == "onscreen":
            if prompt:
                try: doc.Utility.Prompt(prompt)
                except Exception: pass
            ss.SelectOnScreen()
        elif mode in ("all","window","crossing"):
            selmode = 5 if mode=="all" else (3 if mode=="window" else 4)
            ss.Select(
                selmode,
                p1 if mode!="all" else None,
                p2 if mode!="all" else None,
                to_vt_int(filter_types) if filter_types else None,
                to_vt_variant(filter_data) if filter_data else None
            )
        else:
            raise ValueError("未知选择模式")

        n = com_retry(lambda: ss.Count)
        items = [com_retry(lambda i=i: ss.Item(i)) for i in range(n)]
        if autocast:
            items = [_maybe_cast(e) for e in items]
        return items
    finally:
        try: ss.Delete()
        except Exception: pass


# ===================== 选择函数（重写版） =====================

# 1) 选择“包围框覆盖指定点”的对象
def select_entities_through_point(p, tol=0.1):
    """
    查找所有其 BoundingBox 覆盖点 p 的图元。
    不依赖全局 mp；自动重试；失败的对象跳过。
    """
    _, doc = get_acad_doc()
    msp = doc.ModelSpace
    px, py, _ = p
    selected = []
    # 遍历模型空间；大图慎用（必要时可改为小窗口交叉选择）
    for ent in msp:
        try:
            p1, p2 = com_retry(lambda: ent.GetBoundingBox())
            x_min, x_max = sorted([p1[0], p2[0]])
            y_min, y_max = sorted([p1[1], p2[1]])
            x_min -= tol; x_max += tol; y_min -= tol; y_max += tol
            if x_min <= px <= x_max and y_min <= py <= y_max:
                selected.append(_maybe_cast(ent))
        except Exception:
            continue
    print(f"✅ 共找到 {len(selected)} 个对象经过点 {p}")
    return selected


# 2) 按图层选择（支持单个或列表）
def select_tuceng(layer_names, max_retries=5, delay=0.5, autocast=True):
    if isinstance(layer_names, str):
        layers = [layer_names]
    else:
        layers = list(layer_names)
    last = None
    for k in range(1, max_retries+1):
        try:
            ents = ss_select(
                mode="all",
                filter_types=[8],                # 8 = Layer
                filter_data=[layers if len(layers)>1 else layers[0]],
                autocast=autocast
            )
            print(f"✅ 第 {k} 次尝试：选到图层 {layers} 上 {len(ents)} 个对象")
            return ents
        except Exception as e:
            last = e
            print(f"⚠ 第 {k} 次失败：{e!r}")
            try:
                _, doc = get_acad_doc()
                doc.SendCommand("RE\nZ\nE\n")
            except Exception:
                pass
            time.sleep(delay)
    print(f"❌ 重试 {max_retries} 次后仍失败：{last!r}")
    return []

def stc(layer_names, **kwargs):
    return select_tuceng(layer_names, **kwargs)


# 3) 选择所有块（INSERT）
def select_kuai(max_retries: int = 5, autocast=True):
    last = None; t0 = time.time()
    for k in range(1, max_retries+1):
        try:
            ents = ss_select(
                mode="all",
                filter_types=[0],               # 0 = 实体类型
                filter_data=["INSERT"],
                autocast=autocast
            )
            print(f"✅ select_kuai 成功（第 {k} 次），耗时 {time.time()-t0:.3f}s，共 {len(ents)} 个块")
            return ents
        except Exception as e:
            last = e
            print(f"⚠ select_kuai 第 {k} 次失败：{e!r}")
            try:
                _, doc = get_acad_doc(); doc.SendCommand("RE\nZ\nE\n")
            except Exception: pass
            time.sleep(0.5)
    print(f"❌ select_kuai 在 {max_retries} 次尝试后仍失败：{last!r}")
    return []


# 4) 选择所有 TEXT
def select_text(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["TEXT"], autocast=autocast)
    print("耗时：", time.time()-t0)
    return ents

# 5) 选择所有 MTEXT
def select_mtext(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["MTEXT"], autocast=autocast)
    print("耗时：", time.time()-t0)
    return ents

# 6) PUB_TEXT 图层上的“天正文字”分类（按 ObjectName）
def select_pub_text_entities():
    LAYER_NAME = "PUB_TEXT"
    ents = select_tuceng(LAYER_NAME, autocast=False)  # 天正对象通常是代理/专有类，先别 Cast
    tdb_texts, tdb_mtexts = [], []
    for ent in ents:
        name = getattr(ent, "ObjectName", None) or getattr(ent, "EntityName", "")
        if name == "TDbText":
            tdb_texts.append(ent)
        elif name == "TDbMText":
            tdb_mtexts.append(ent)
    return tdb_texts, tdb_mtexts

def collect_all_texts():
    """
    同时收集天正与原生 CAD 文本，并统一图层为 'PUB_TEXT'。
    返回：tianzheng_texts / tianzheng_mtexts / cad_texts / cad_mtexts
    """
    LAYER_NAME = "PUB_TEXT"
    tz_texts, tz_mtexts = select_pub_text_entities()
    cad_texts  = select_text(autocast=True)
    cad_mtexts = select_mtext(autocast=True)

    for ent in (tz_texts + tz_mtexts + cad_texts + cad_mtexts):
        try:
            ent.Layer = LAYER_NAME
        except Exception:
            pass
    return tz_texts, tz_mtexts, cad_texts, cad_mtexts


# 7) 选择 LINE / CIRCLE / ELLIPSE / SPLINE
def select_line(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["LINE"], autocast=autocast)
    print("耗时：", time.time()-t0)
    return ents

def select_circle(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["CIRCLE"], autocast=autocast)
    print("耗时：", time.time()-t0)
    return ents

def select_ellipse(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["ELLIPSE"], autocast=autocast)
    print("耗时：", time.time()-t0)
    return ents

def select_spline(autocast=True):
    t0 = time.time()
    ents = ss_select(mode="all", filter_types=[0], filter_data=["SPLINE"], autocast=autocast)
    print("耗时：", time.time()-t0)
    return ents


# 8) 传统多段线（POLYLINE）与轻量多段线（LWPOLYLINE）
def select_polyline_chuantong(max_retries: int = 5, autocast=True):
    last = None; t0 = time.time()
    for k in range(1, max_retries+1):
        try:
            ents = ss_select(mode="all", filter_types=[0], filter_data=["POLYLINE"], autocast=autocast)
            print(f"✅ select_polyline_chuantong 成功（第 {k} 次），耗时 {time.time()-t0:.3f}s，共 {len(ents)} 条")
            return ents
        except Exception as e:
            last = e
            print(f"⚠ select_polyline_chuantong 第 {k} 次失败：{e!r}")
            try:
                _, doc = get_acad_doc(); doc.SendCommand("RE\nZ\nE\n")
            except Exception: pass
            time.sleep(0.5)
    print(f"❌ select_polyline_chuantong 在 {max_retries} 次后仍失败：{last!r}")
    return []

def select_polyline(max_retries: int = 5, autocast=True):
    last = None; t0 = time.time()
    for k in range(1, max_retries+1):
        try:
            ents = ss_select(mode="all", filter_types=[0], filter_data=["LWPOLYLINE"], autocast=autocast)
            print(f"✅ select_polyline 成功（第 {k} 次），耗时 {time.time()-t0:.3f}s，共 {len(ents)} 条")
            return ents
        except Exception as e:
            last = e
            print(f"⚠ select_polyline 第 {k} 次失败：{e!r}")
            try:
                _, doc = get_acad_doc(); doc.SendCommand("RE\nZ\nE\n")
            except Exception: pass
            time.sleep(0.5)
    print(f"❌ select_polyline 在 {max_retries} 次后仍失败：{last!r}")
    return []


#&&% ##  6  屏幕选择





# --- 新版选择：不依赖全局 doc；不 for-in；自动清理；可选自动 CastTo ---
def pmxz(prompt="\n请在屏幕拾取图元，以Enter键结束：", autocast=True):
    """
    交互拾取对象（显式 on-screen 选择），返回实体列表。
    内部直接复用统一内核 ss_select，保持与其它选择函数一致的行为。
    """
    return ss_select(
        mode="onscreen",
        p1=None, p2=None,
        filter_types=None,     # 如需限制类型，可传 [0] 并配合 filter_data=["LINE"] 等
        filter_data=None,
        autocast=autocast,
        prompt=prompt
    )

"""
如果你想“只允许选某类对象”（比如直线），就传过滤参数即可：

def pmxz_line(prompt="选一根直线："):
    return ss_select("onscreen", filter_types=[0], filter_data=["LINE"], autocast=True, prompt=prompt)


"""


# 8) 隐性 → 显性选择（安全版）
def yin_to_xian_xuanze(LB, wait_s=0.6):
    """
    将 COM 选中的对象列表（LB）转换为命令窗口中的蓝色“高亮选中”状态。
    实现：StartUndoMark → 批量 Delete → U 撤销 → SELECT P。
    """
    _, doc = get_acad_doc()
    deleted = 0
    com_retry(lambda: doc.StartUndoMark())
    try:
        for x in LB:
            try:
                com_retry(lambda x=x: x.Delete())
                deleted += 1
            except Exception:
                pass
        print(f"🧹 尝试删除 {len(LB)} 个对象，成功删除 {deleted} 个。")

        # 撤销删除（恢复并建立 _P）
        doc.SendCommand("_U\n\n")
        time.sleep(wait_s)

        # 使用 P（previous selection）建立显性选择
        doc.SendCommand("_.SELECT\nP\n\n")
        time.sleep(0.5)
        print("✅ 已将隐性选择转换为显性（蓝色高亮）")
    finally:
        try: com_retry(lambda: doc.EndUndoMark())
        except Exception: pass


# 9) 显性 → 隐性选择（保留你原法，但收敛成函数）
def xian_to_yin_xuanze():
    _, doc = get_acad_doc()
    # 基于你的流程：CopyBase → 全选 → Erase → PasteClip → 取 SelectionSetAll
    doc.SendCommand("_COPYBASE\n0,0,0\n\n"); time.sleep(3)
    doc.SendCommand("_AI_SELALL\n\n");       time.sleep(1.5)
    doc.SendCommand("_ERASE\n\n");           time.sleep(0.8)
    doc.SendCommand("_PASTECLIP\n0,0,0\n\n");time.sleep(2.0)

    # 取隐性选择（全部）
    try:
        doc.SelectionSets.Item("MySelectionSet").Delete()
    except Exception:
        pass
    ss = doc.SelectionSets.Add("MySelectionSet")
    ss.Select(5)  # 全选
    lb = [ss.Item(i) for i in range(ss.Count)]
    try: ss.Delete()
    except Exception: pass
    print(f"✅ 共获取对象 {len(lb)} 个，转换为可操作选择列表")
    return lb


# 10) 隐性窗口选择（优先 SelectionSet 窗口；失败兜底遍历包围盒）
def select_objects_in_window_area(x1, y1, x2, y2, max_retry=5):
    _, doc = get_acad_doc()
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)
    doc.SetVariable("TILEMODE", 1)

    # 临时选择集
    try: doc.SelectionSets.Item("MyWindowSelection").Delete()
    except Exception: pass
    ss = doc.SelectionSets.Add("MyWindowSelection")

    entities = []
    # ① 重试窗口选择
    p1, p2 = pt3(x_lo,y_lo), pt3(x_hi,y_hi)
    for _ in range(max_retry):
        try:
            ss.Clear()
            ss.Select(1, p1, p2)  # 1 = Window
            if ss.Count > 0: break
        except Exception:
            pass
        time.sleep(0.25)

    if ss.Count > 0:
        for i in range(ss.Count):
            try: entities.append(_maybe_cast(ss.Item(i)))
            except Exception: continue
        try: ss.Delete()
        except Exception: pass
        print(f"✅ 窗口选择成功，共 {len(entities)} 个对象。")
        return entities

    # ② 兜底：遍历 ModelSpace 的包围盒角点
    try:
        msp = doc.ModelSpace
        def in_win(pt): return x_lo <= pt[0] <= x_hi and y_lo <= pt[1] <= y_hi
        for ent in msp:
            try:
                a,b = com_retry(lambda: ent.GetBoundingBox())
                if in_win(a) or in_win(b):
                    entities.append(_maybe_cast(ent))
            except Exception:
                continue
        print(f"[FALLBACK] 遍历模型空间得到 {len(entities)} 个对象。")
    except Exception as e:
        print(f"❌ 遍历模型空间失败: {e}")

    try: ss.Delete()
    except Exception: pass
    return entities


# 11) 隐显结合的区域选择（高亮选择并返回 PickfirstSelectionSet）
def select_entities_in_window(x1, y1, x2, y2, ty: float = 1.0, select_mode: str = "_W"):
    _, doc = get_acad_doc()
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)

    # 清空 Pickfirst
    try: doc.Pickenabled = False
    except Exception: pass
    try: doc.PickfirstSelectionSet.Clear()
    except Exception: pass

    # Zoom（加 20% 缓冲）
    buf = 0.20 * ((x_hi-x_lo) + (y_hi-y_lo)) / 2.0
    doc.SendCommand(f"_.ZOOM\n_W\n{x_lo-buf},{y_lo-buf}\n{x_hi+buf},{y_hi+buf}\n")
    time.sleep(ty)

    # Select（显性，蓝色高亮）
    doc.SendCommand(f"_.SELECT\n{select_mode}\n{x_lo},{y_lo}\n{x_hi},{y_hi}\n\n")
    time.sleep(ty/2)

    selset = doc.PickfirstSelectionSet
    com_list = [ent for ent in selset]
    try: selset.Clear()
    except Exception: pass
    return com_list


# 12) 显性区域选择（蓝色高亮）
def highlight_entities_in_window(x1, y1, x2, y2):
    _, doc = get_acad_doc()
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)
    h = 0.1 * ((abs(x_hi-x_lo)+abs(y_hi-y_lo))/2.0)

    # ZOOM
    doc.SendCommand(f"_.ZOOM\n_W\n{x_lo-h},{y_lo-h}\n{x_hi+h},{y_hi+h}\n"); time.sleep(1)

    # 显性 SELECT
    doc.SendCommand(f"_.SELECT\n_W\n{x_lo},{y_lo}\n{x_hi},{y_hi}\n\n"); time.sleep(0.5)
    print(f"✅ 已高亮选择区域 ({x_lo},{y_lo}) ~ ({x_hi},{y_hi}) 的对象")


# ====== 辅助：按偏移量扩张包围框（p1,p2 都是 (x,y,?) 序列）======
def expand_rectangle(p1, p2, offset):
    x1, y1 = float(p1[0]), float(p1[1])
    x2, y2 = float(p2[0]), float(p2[1])
    x_lo, x_hi = sorted((x1, x2))
    y_lo, y_hi = sorted((y1, y2))
    return (x_lo - offset, y_lo - offset), (x_hi + offset, y_hi + offset)
# 1) 隔离对象（模型空间区域）
def isolate_modelspace_area(x1, y1, x2, y2):
    """
    在模型空间，将窗口区域内对象隔离显示：
      - 隐性选区（COM SelectionSet/包围盒兜底）→
      - 转显性（蓝色高亮，供 SendCommand 识别）→
      - 发送 _IsolateObjects
    """
    _, doc = get_acad_doc()
    # 确保模型空间
    try: doc.SetVariable("TILEMODE", 1)
    except Exception: pass

    LB = select_objects_in_window_area(x1, y1, x2, y2)
    if not LB:
        print("❌ 没有选择到对象，终止操作")
        return

    print(f"✅ 选中对象 {len(LB)} 个，准备隔离")
    yin_to_xian_xuanze(LB, wait_s=0.6)  # 已实现：隐性→显性
    time.sleep(0.4)
    doc.SendCommand("_.IsolateObjects\n")
    print("🎯 已发送隔离命令")
# 2) 图纸空间的区域选择（隐性列表返回）
def select_paperspace_objects_in_window(x1, y1, x2, y2):
    """
    在图纸空间内，选择窗口 (x1,y1)-(x2,y2) 覆盖的对象；优先用 SelectionSet 窗口，
    若为空则遍历 PaperSpace 的包围盒兜底。返回 COM 对象列表（已自动 Cast）。
    """
    _, doc = get_acad_doc()
    try: doc.SetVariable("TILEMODE", 0)  # 切到图纸空间
    except Exception: pass

    (x_lo, y_lo), (x_hi, y_hi) = normalize_rect(x1, y1, x2, y2)
    # 优先：窗口选择
    try:
        items = ss_select(
            mode="window",
            p1=pt3(x_lo, y_lo),
            p2=pt3(x_hi, y_hi),
            filter_types=None,
            filter_data=None,
            autocast=True
        )
        if items:
            print(f"✅ 图纸空间窗口选择 {len(items)} 个对象。")
            return items
    except Exception:
        pass

    # 兜底：遍历 PaperSpace 包围盒
    selected = []
    ps = doc.PaperSpace
    def intersects(a1, a2, b1, b2):
        return (a1 <= b2) and (a2 >= b1)
    for i in range(ps.Count):
        try:
            ent = com_retry(lambda i=i: ps.Item(i))
            p1, p2 = com_retry(lambda: ent.GetBoundingBox())
            ex1, ey1 = p1[0], p1[1]
            ex2, ey2 = p2[0], p2[1]
            if intersects(ex1, ex2, x_lo, x_hi) and intersects(ey1, ey2, y_lo, y_hi):
                selected.append(_maybe_cast(ent))
        except Exception:
            continue
    print(f"[FALLBACK] 图纸空间包围盒遍历得到 {len(selected)} 个对象。")
    return selected
# 3) 将“单个对象”显性高亮（蓝色夹点）
def highlight_entity_by_bbox(entity):
    """
    对指定 COM 对象 entity，通过扩大其 bounding box 来进行显性高亮选中。
    会自动 Zoom 到对象附近，然后用窗口选择把它高亮出来。
    """
    try:
        # 用对象自身的 App/Doc，避免多文档切换干扰
        doc = com_retry(lambda: entity.Application.ActiveDocument)
        p1, p2 = com_retry(lambda: entity.GetBoundingBox())

        x1, y1 = float(p1[0]), float(p1[1])
        x2, y2 = float(p2[0]), float(p2[1])

        # Zoom 到对象 + 10% 缓冲
        h = 0.1 * ((abs(x1 - x2) + abs(y1 - y2)) / 2.0)
        doc.SendCommand(f"_.ZOOM\n_W\n{x1-h},{y1-h}\n{x2+h},{y2+h}\n")
        time.sleep(1.0)

        # 计算略扩张的选择窗口（墙体特殊偏移）
        x_len = abs(x2 - x1); y_len = abs(y2 - y1)
        max_len = max(x_len, y_len)
        offset = 130 if (getattr(entity, "ObjectName", "") == "TDbWall") else (max_len * 0.1)
        (X1,Y1),(X2,Y2) = expand_rectangle(p1, p2, offset)

        # 显性窗口选择
        highlight_entities_in_window(X1, Y1, X2, Y2)
        print("✅ 已高亮目标对象")
    except Exception as e:
        print("❌ 无法高亮该对象:", e)
# 4) 利用 Visible 进行“显选→隐选”转换的便捷封装（保持你的流程）
def select_visible(x1, y1, x2, y2):
    """
    先显性高亮窗口对象，再调用 xian_to_yin_xuanze() 转换为可操作的隐性列表。
    返回隐性列表（COM 对象集合）。
    """
    highlight_entities_in_window(x1, y1, x2, y2)
    time.sleep(1.0)
    return xian_to_yin_xuanze()
# 5) 显示指定空间中所有隐藏对象（可过滤类型，可选择高亮）
def unhide_all(space=None, filter_names=None, highlight=False):
    """
    显示 space 中所有 Visible=False 的对象。
    参数：
      - space: 可传 doc.ModelSpace / doc.PaperSpace / None / "model" / "paper"
      - filter_names: 仅恢复这些 ObjectName 的对象（例如 ["AcDbPolyline", "AcDbBlockReference"]）
      - highlight: True 则对恢复的对象调用 .Highlight(True)
    返回：revealed 列表（已自动 Cast）
    """
    _, doc = get_acad_doc()

    # 解析空间参数
    if space is None:
        target = doc.ModelSpace
    elif isinstance(space, str):
        s = space.strip().lower()
        if s in ("model", "m", "modelspace"):
            target = doc.ModelSpace
        elif s in ("paper", "p", "paperspace", "layout"):
            target = doc.PaperSpace
        else:
            target = doc.ModelSpace
    else:
        # 传入的是 COM 集合（有 Count/Item）
        target = space

    revealed = []
    cnt = com_retry(lambda: target.Count)
    for i in range(cnt):
        try:
            obj = com_retry(lambda i=i: target.Item(i))
            # 有些对象没有 Visible，跳过
            if not hasattr(obj, "Visible"):
                continue
            if obj.Visible:
                continue
            name = getattr(obj, "ObjectName", "")
            if (filter_names is None) or (name in filter_names):
                obj.Visible = True  # 设为可见
                if highlight and hasattr(obj, "Highlight"):
                    try: obj.Highlight(True)
                    except Exception: pass
                revealed.append(_maybe_cast(obj))
                print(f"✅ 显示对象：{name} | Handle={getattr(obj, 'Handle', '?')}")
        except Exception as e:
            print(f"⚠️ 跳过索引 {i}：{e}")

    print(f"\n📊 共显示 {len(revealed)} 个隐藏对象。")
    return revealed


#&&% #  组
"""

创建组
group = doc.Groups.Add("mygroup")
LB=pmxz()
请在屏幕拾取图元，以Enter键结束

group.AppendItems(vtobj([LB[0], LB[1], LB[2]]))


从组名获取组合组中对象

group = doc.Groups.Item("G001")

entities = [group.Item(i) for i in range(group.Count)]  # 遍历组内对象

entities[0].Handle
'2C3'

entities[0].Move(vtpnt(0,0,0),vtpnt(0,10000,0))


解除组
group.Delete()


group1 = doc.Groups.Add("mygroup")
group1.AppendItems(vtobj([LB[0], LB[1], LB[2]]))
group2 = doc.Groups.Add("mygroupA")
group2.AppendItems(vtobj([LB[3], LB[4], LB[5],LB[6]]))
group3 = doc.Groups.Add("mygroupB")
group3.AppendItems(vtobj([group1,group2]))非法操作
group3.AppendItems(vtobj([LB[0],LB[1],LB[2],LB[3], LB[4], LB[5],LB[6]]))


# 提取 group1 和 group2 的所有成员
group1_entities = [group1.Item(i) for i in range(group1.Count)]
group2_entities = [group2.Item(i) for i in range(group2.Count)]

# 合并成新的列表
all_entities = group1_entities + group2_entities

# 创建 group3
group3 = doc.Groups.Add("mygroupB")
group3.AppendItems(vtobj(all_entities))

 get_boundingbox_from_objects(objs)
"""

# 建立全部列表com对象的最小边界框

def get_boundingbox_from_objects(objs):#从列表com对象建立最小边界框
    """
    从一组图形对象（如 LB）中获取整体包围盒
    返回值：(min_x, min_y, min_z), (max_x, max_y, max_z)
    """
    min_point, max_point = None, None

    for obj in objs:
        try:
            min_pt, max_pt = obj.GetBoundingBox()
            if min_point is None:
                min_point, max_point = list(min_pt), list(max_pt)
            else:
                min_point = [min(min_point[i], min_pt[i]) for i in range(3)]
                max_point = [max(max_point[i], max_pt[i]) for i in range(3)]
        except Exception as e:
            print(f"跳过无法获取边界的对象: {obj.ObjectName}")
            continue

    return tuple(min_point), tuple(max_point)
        

# 建立组的最小边界框

def chuangjian_zu(group_name):

    group = doc.Groups.Add(group_name)

    return group

def nametogroup(group_name):#从组名获取实体com组对象
    group_obj = doc.Groups.Item(group_name)

    return group_obj

##获取所有组

def get_all_group_names():
    """
    获取当前 DWG 文档中所有组的名称列表。
    
    返回:
      List[str] — 包含所有组名称的列表
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    groups = doc.Groups
    return [groups.Item(i).Name for i in range(groups.Count)]

def get_all_groups():
    """
    获取当前 DWG 文档中所有组的 COM 对象列表及其名称。
    
    返回:
      List[Tuple[str, COMObject]] — 每项为 (组名称, 组对象)
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    groups = doc.Groups
    result = []
    for i in range(groups.Count):
        grp = groups.Item(i)
        result.append((grp.Name, grp))
    return result



#将多个com对象对象加入名为group_name的组
def add_objects_to_group(group_name, obj_list):

    """
    将 obj_list 中的所有图形对象加入名为 group_name 的组中
    如果组已存在，使用原组；否则新建
    返回：Group 对象
    """
    groups = doc.Groups
    try:
        group = groups.Item(group_name)
    except:
        group = groups.Add(group_name)

    group.AppendItems(vtlist(obj_list))
    return group


#将单独com对象对象加入名为group_name的组中

def add_object_to_group(group_name, obj):
    """
    将单个图形对象 obj 加入名为 group_name 的组中。
    如果组已存在，则使用该组；否则新建一个新组。
    
    参数：
      group_name (str)：组名称
      obj：要加入组的 COM 对象（如多段线、线段、块参照等）
    
    返回：
      COM Group 对象
    """
    groups = doc.Groups
    try:
        # 尝试获取已存在的组
        group = groups.Item(group_name)
    except Exception:
        # 不存在则新建
        group = groups.Add(group_name)
    
    # vtlist 工具将 Python list 转成 VBA 可接受的 SAFEARRAY
    group.AppendItems(vtlist([obj]))
    return group

#将单独com对象对象移出名为group_name的组
def remove_object_from_group(group_name, obj):
    """
    将单个 COM 对象 obj 从名为 group_name 的组中移出。
    如果组不存在或对象不在组中，则会打印错误信息但不抛异常。
    
    参数:
      group_name: 组名（字符串）
      obj:         要移出的 COM 对象
    
    返回:
      如果组存在，返回该 Group 对象；否则返回 None。
    """
    try:
        group = doc.Groups.Item(group_name)
    except Exception:
        print(f"❌ 组 '{group_name}' 不存在")
        return None

    # 把单个对象包装成长度为1的 COM SafeArray
    variant = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, [obj])
    try:
        group.RemoveItems(variant)
        print(f"✅ 已从组 '{group_name}' 中移除对象 {obj.Handle}")
    except Exception as e:
        print(f"❌ 从组 '{group_name}' 移除对象失败：{e}")

    return group

#将多个com对象对象移出名为group_name的组
def remove_objects_from_group(group_name, obj_list):
    """
    将 obj_list 中的所有图形对象从名为 group_name 的组中移出。
    如果组不存在，会打印提示并返回 None；否则返回该组对象。
    
    :param group_name: 组名称
    :param obj_list: 要移除的 COM 对象列表
    :return: Group 对象 或 None
    """
    groups = doc.Groups
    try:
        group = groups.Item(group_name)
    except Exception:
        print(f"组 '{group_name}' 不存在，无法移除对象。")
        return None

    # 把 Python 列表包装成 VARIANT SafeArray，VT_DISPATCH 表示对象类型
    arr = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj_list)
    try:
        group.RemoveItems(arr)
        print(f"已从组 '{group_name}' 中移除 {len(obj_list)} 个对象。")
    except Exception as e:
        print(f"移除对象时发生错误：{e}")
    return group



#&&% 从名为group_name的组获取内部包含的实体对象


def get_com_from_groupname(group_name):
    """
    根据组名获取对应实体列表。
    - 若组不存在、或组中无实体，均返回空列表，不抛异常。
    """

    try:
        group = nametogroup(group_name)
    except Exception:
        # nametogroup 本身失败（组不存在等）
        return []

    if not group:
        # group 为 None 或空，也直接返回空列表
        return []

    entities = [group.Item(i) for i in range(group.Count)] #从com组中获取全部对象
    
    return entities

#从名为group_name的组返回按类型分类的字典
def get_com_from_groupname_by_type(group_name):
    """
    根据组名获取对应实体，并按类型名分类返回。

    :param group_name: 组名称
    :return: dict，键为实体类型名（ObjectName），值为该类型的实体列表
    """
    # nametogroup 是你已有的“组名→Group 对象”函数
    group = nametogroup(group_name)
    if group is None:
        print(f"组 '{group_name}' 不存在")
        return {}

    by_type = {}
    # Group.Count 是实体数量，Item(i) 取出第 i 个实体
    for i in range(group.Count):
        ent = group.Item(i)
        # AutoCAD COM 对象一般有 ObjectName 属性
        typ = getattr(ent, "ObjectName", None) or getattr(ent, "EntityName", "Unknown")
        by_type.setdefault(typ, []).append(ent)

    # 打印一下各类型数量，方便调试
    for typ, lst in by_type.items():
        print(f"  类型 {typ} ：{len(lst)} 个实体")

    return by_type

#从名为group_name的组返回按类型分类的字典，且类型按各自位置提取函数排好序
def get_group_entities_sorted(group_name, type_extractors, cha_Y=0.5):
    """
    从组中按类型获取实体，并对指定类型按坐标排序。

    :param group_name: str，组名
    :param type_extractors: dict，{ type_name: extract_func }，
           extract_func(ent) 返回 (x, y) 坐标，用于排序。
    :param cha_Y: float，同一行 Y 方向的容差
    :return: dict，{ type_name: [ent, ...] }，已排序或原序

    对天正单行多行文字可以使用GetBoundingBox获取的左下角点
    
    """
    def get_cadtext_pos(ent):
        # CAD单行、多行文字：Position 属性返回 (x, y, z)
        return float(ent.InsertionPoint[0]), float(ent.InsertionPoint[1])


    # 先按类型获取全部实体
    by_type = get_com_from_groupname_by_type(group_name)
    sorted_by_type = {}

    for typ, ents in by_type.items():
        if typ in type_extractors:
            extract_func = type_extractors[typ]
            # 排序
            sorted_list = sort_entities_by_position(ents, extract_func, cha_Y=cha_Y)
            sorted_by_type[typ] = sorted_list
            print(f"Type '{typ}' sorted with {len(sorted_list)} entities")
        else:
            # 保持原序
            sorted_by_type[typ] = list(ents)
            print(f"Type '{typ}' left unsorted ({len(ents)} entities)")

    return sorted_by_type


#从名为group_name的组返回按类型分类的字典，各类型统一按boundingbox中心排好序


def get_group_entities_sorted_by_type_and_bbox(group_name, cha_Y=0.5):
    """
    将组 group_name 中的实体按类型分类，并对每种类型内部按包围盒中心排序：
      1) 先按 center_y 降序（从上到下）
      2) 同一“行”内（|ΔY|<cha_Y）再按 center_x 升序（从左到右）

    参数：
      group_name: 要操作的组名
      cha_Y: 同一“行”Y 方向容差

    返回：
      一个 dict，key=类型名(ObjectName)，value=排序后的实体列表
    """
    # 1. 取组
    group = nametogroup(group_name)
    ents = [group.Item(i) for i in range(group.Count)]

    # 2. 按类型分组
    by_type = {}
    for ent in ents:
        typ = getattr(ent, "ObjectName", None) or ent.EntityName
        by_type.setdefault(typ, []).append(ent)

    # 3. 辅助：计算包围盒中心点

    # 4. 对每个类型内部排序
    for typ, lst in by_type.items():
        triples = [(e, *bbox_center_2(e)) for e in lst]
        # Y 降序
        triples.sort(key=lambda t: -t[2])
        # 同行内按 X 升序
        i = 0
        while i < len(triples) - 1:
            j = i + 1
            while j < len(triples) and abs(triples[i][2] - triples[j][2]) < cha_Y:
                j += 1
            triples[i:j] = sorted(triples[i:j], key=lambda t: t[1])
            i = j
        # 覆盖原列表
        by_type[typ] = [t[0] for t in triples]

    return by_type


# 获取两个组中共有的实体，按类型分类并按包围盒中心排序
def common_group_entities_sorted(group_name1, group_name2, cha_Y=0.5):
    """
    获取两个组中共有的实体，按类型分类并按包围盒中心排序。

    参数：
      group_name1, group_name2: 要比较的两个组名
      cha_Y: 同一行判定的 Y 方向容差

    返回：
      dict => key: ObjectName 类型名, value: 排序后的实体列表
    """
    # 1. 取组
    g1 = nametogroup(group_name1)
    g2 = nametogroup(group_name2)

    ents1 = [g1.Item(i) for i in range(g1.Count)]
    ents2 = [g2.Item(i) for i in range(g2.Count)]

    # 2. 建立 handle->entity 映射
    map1 = {e.Handle: e for e in ents1}
    map2 = {e.Handle: e for e in ents2}

    # 3. 找共有的 handles
    common_handles = set(map1.keys()) & set(map2.keys())

    # 4. 收集共有实体（这里取自 map1）
    common_ents = [map1[h] for h in common_handles]

    # 5. 按类型分组
    by_type = {}
    for ent in common_ents:
        typ = getattr(ent, "ObjectName", None) or ent.EntityName
        by_type.setdefault(typ, []).append(ent)

    # 6. 包围盒中心点计算
    def bbox_center(e):
        min_pt, max_pt = e.GetBoundingBox()
        x1, y1, _ = tuple(min_pt)
        x2, y2, _ = tuple(max_pt)
        return ((x1 + x2) / 2, (y1 + y2) / 2)

    # 7. 排序：Y 降序，同一行内(|ΔY|<cha_Y)按 X 升序
    for typ, lst in by_type.items():
        triples = [(e, *bbox_center(e)) for e in lst]
        triples.sort(key=lambda t: -t[2])
        i = 0
        while i < len(triples) - 1:
            j = i + 1
            while j < len(triples) and abs(triples[i][2] - triples[j][2]) < cha_Y:
                j += 1
            triples[i:j] = sorted(triples[i:j], key=lambda t: t[1])
            i = j
        by_type[typ] = [t[0] for t in triples]

    return by_type

def get_boundingbox_from_group(group):#从com对象group建立最小边界框

    """
    并非组的实际BoundingBOX数据

    """

    entities = [group.Item(i) for i in range(group.Count)] #从组中获取全部对象

    p1,p2 = get_boundingbox_from_objects(entities)

    return p1,p2


def select_group_entities(group_obj): #选择组中对象，转换为高亮显性选择
    """
    让组中的所有对象进入蓝色高亮状态（通过反选法）
    """
    try:
        handles = [entity.Handle for entity in group_obj]
        objs = []

        for obj in mp:
            if obj.Handle in handles:
                objs.append(obj)

        yin_to_xian_xuanze(objs)  # 你已有的高亮函数
        print("✅ 已显性选中组内图元")
        return True
    except Exception as e:
        print("❌ 显性选择失败:", e)
        return False

#&&% #  Handle和Label


"""
Handle身份信息和天正标签Label

与列表和组不同，身份信息的标识使我们可以从字典存储的名称信息精准控制每个对象，而不是总依赖对全部对象的遍历


可以在对象创建时就打上标签，使对象归为某个类

可以在一批对象完成后调整标签

可以临时标记一组标签

可以通过Label给天正对象加标签
q1= doc.ModelSpace.Item(doc.ModelSpace.Count - 1)#自动获取图纸空间之前最后一个绘制的对象


从Handle回溯com对象

doc.HandleToObject('2BD')
q1= doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
q1.Label="A1"
q1= doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
q1.Label="B1"
LB=pmxz()
请在屏幕拾取图元，以Enter键结束
LB[0].Label
'B1'
LB[1].Label
'A1'


"""

# 将列表对象按分类将其handle身份标识存入字典

@alias("h")
def HandleToObject(ZF):#从Handle身份信息值回溯com对象

    """
    对连接在墙上的门窗测试无效

    """

    obj = doc.HandleToObject(ZF)

    return obj

def print_coms_handle(LB):

    LC=[]

    for x in LB:

        LC.append(x.Handle)

    print(f"com对象列表对应的Handle句柄列表：{LC} ")




@alias("H")
def handles_to_coms(LB_handles):

    """
    对连接在墙上的门窗测试无效

    """
    LC=[]

    for xx in LB_handles:

        obj = doc.HandleToObject(xx)
        LC.append(obj)

    return LC


def get_all_handles():#获取所有Handle
    """
    获取当前图纸中所有对象（通常在 ModelSpace）的 Handle 值列表。

    返回：
        handle_list: 所有图元的 Handle 字符串列表
    """
    handle_list = []

    for obj in mp:  # 使用你已经定义好的全局 ModelSpace mp
        try:
            handle_list.append(obj.Handle)
        except:
            continue  # 跳过无 Handle 或异常对象

    print(f"✅ 已获取 {len(handle_list)} 个对象的 Handle")
    return handle_list

def find_entity_by_handle(handle_str):#从Handle获取实体（适合包括天正的文件）
    """
    遍历当前图纸所有对象，手动比对 Handle 值，找到指定的实体对象。

    参数：
        handle_str: 目标 Handle（字符串）

    返回：
        对象（若找到），否则 None
    """
    for obj in mp:  # 可扩展：遍历 sp 也行
        try:
            if obj.Handle == handle_str:
                return obj
        except:
            continue

    return None



def group_objects_by_type_and_handle(LB):#将列表对象的Handle身份信息分类存入字典返回
    """
    将com对象列表 LB 中的对象按 ObjectName 分类，并存储其 Handle。
    每类按 LB 中出现顺序编号。

    参数：
        LB - AutoCAD 实体对象列表（如 select_objects_in_window_area() 返回）

    返回：
        ZD - dict 格式 {ObjectName: [Handle1, Handle2, ...]}
    """
    ZD = {}  # 初始化字典

    for obj in LB:
        try:
            obj_type = obj.ObjectName
            handle = obj.Handle

            if obj_type not in ZD:
                ZD[obj_type] = []

            ZD[obj_type].append(handle)

        except Exception as e:
            print(f"⚠️ 跳过对象，原因: {e}")
            continue

    # 输出提示信息
    for obj_type, handles in ZD.items():
        print(f"✅ {obj_type}: 共 {len(handles)} 个对象")

    return ZD

# 通过名称存储对象信息反回溯对象

def record_handle_with_type(LB, typename, prefix="OBJ"):#将一批对象的 Handle 存储到结构化的字典中，并记录类型名和编号
    """
    替代 XData 方法：记录对象 Handle、类型名、编号，返回结构化字典。
    """
    ZD = {typename: {}}
    for i, obj in enumerate(LB, start=1):
        try:
            h = obj.Handle
            tag = f"{prefix}_{i:03d}"
            ZD[typename][h] = tag
        except:
            continue
    print(f"✅ 已记录 {len(ZD[typename])} 个“{typename}”对象（Handle+编号）")
    return ZD

def convert_named_dict(ZD, typename):# 构建编号 → COM 对象 的映射字典
    """
    将 ZD["门"] 的结构由 Handle: 编号 转换为 编号: COM对象
    返回：新的字典 {编号: COM实体}
    """
    doc = win32com.client.Dispatch("AutoCAD.Application").ActiveDocument
    named_dict = {}

    handle_map = ZD.get(typename, {})
    for handle, name in handle_map.items():
        try:
            obj = doc.HandleToObject(handle)
            named_dict[name] = obj
        except:
            print(f"⚠️ 无法找到对象（handle={handle}）")
            continue

    return named_dict

def get_named_object(tag, ZD, typename="门"):#从标签获取对象
    named = convert_named_dict(ZD, typename)
    return named.get(tag)



def draw_tags_on_objects_fixed(named_dict, height=250, offset=(1000, 1000, 0)):#直接将编号文字写在每个图元上，通常居中或偏移一点点
    """
    在每个对象的中心点附近绘制标注文字。
    
    参数:
        named_dict - 如 {"Men_001": <COMObject>, ...}
        height     - 文字高度
        offset     - 偏移量（用于防止文字盖住对象）
    """
    import win32com.client
    import pythoncom

    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument
    ms = doc.ModelSpace

    for name, obj in named_dict.items():
        try:
            # 使用 GetBoundingBox 获取中心点
            min_pt, max_pt = obj.GetBoundingBox()
            center_pt = (
                (min_pt[0] + max_pt[0]) / 2,
                (min_pt[1] + max_pt[1]) / 2,
                (min_pt[2] + max_pt[2]) / 2
            )

            # 加上偏移量
            label_pt = (
                center_pt[0] + offset[0],
                center_pt[1] + offset[1],
                center_pt[2] + offset[2]
            )

            # 确保插入点是三维点
            pt = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, label_pt)
            
            # 添加文字
            ms.AddText(name, pt, height)
            print(f"✅ 已标注对象: {name}")

        except Exception as e:
            print(f"⚠️ 标注失败: {name}, 错误: {e}")

# 给天正对象打上标签存入字典，用于以名称反向回溯操作

def label_tarch_doors(LB1, typename="门", prefix="men"):#给选中的天正门打上标签并存入字典返回，非天正图元没有Label属性
    """
    从对象列表 LB1 中筛选出天正门 (ObjectName == 'TDbOpening')，
    并为其按顺序打上 .Label 标签（如 men001, men002 ...）。

    返回：
        ZD = {typename: {编号: 对象}}
    """
    ZD = {typename: {}}
    LB2 = []

    for obj in LB1:
        try:
            if hasattr(obj, "ObjectName") and obj.ObjectName == "TDbOpening":
                LB2.append(obj)
        except:
            continue

    for i, obj in enumerate(LB2, start=1):
        try:
            tag = f"{prefix}{i:03d}"
            obj.Label = tag
            ZD[typename][tag] = obj
            print(f"✅ 已标记: {tag}")
        except Exception as e:
            print(f"⚠️ 设置标签失败：{e}")

    print(f"\n📦 共找到并标注 {len(LB2)} 个天正门")
    return ZD


# 获取模型空间上绘制的最后一个图元

def last_obj():

    obj = doc.ModelSpace.Item(doc.ModelSpace.Count - 1)

    return obj

# 获取模型空间上的Handle


"""
target_handles = ['5F', '60', '61']
map = get_handle_object_map(doc.ModelSpace)
objs = [map[h] for h in target_handles if h in map]
这比每次都遍历 ModelSpace 快得多，尤其是大图纸中上千个图元时。

"""

def get_handle_object_map(ms):
    """返回 {handle: object} 映射"""
    return {ent.Handle: ent for ent in ms}



#&&% XData

"""

在 RegAppTable 中注册,在第一次给图元附加 XData 时，AutoCAD 内部会检查 RegAppTable（注册应用程序表）中是否已经存在 “TestApp” 这个名称。

如果不存在，AutoCAD 会自动往 RegAppTable 里插入一条记录，把 “TestApp” 注册进去。

如果你希望手动控制，也可以先调用 doc.Application.RegistryModes.Add("TestApp")（或使用 AutoLISP：(regapp "TestApp")）

app_name    = "TestApp"            # 自定义的应用程序名
data_types  = [1000, 1040, 1070]
data_values = ["示例文字", 3.14159, 12345]
set_xdata(lineObj, app_name, data_types, data_values)
types_out, data_out = get_xdata(lineObj, app_name)
types_out
[1001, 1000, 1040, 1070]
data_out
['TestApp', '示例文字', 3.14159, 12345]


"""
def set_xdata(
    com_obj,
    app_name: str,
    data_types: list[int],
    data_values: list,
):
    """
    向任意 AutoCAD COM 对象（如 Line、Circle、BlockReference 等）附加 XData。

    参数：
      com_obj        -- 任意支持 SetXData() 方法的 COM 对象
      app_name       -- 注册过的应用程序名（字符串）；第一个 DataType 一定是 1001，对应的第一个 Data 存放此 app_name
      data_types     -- 后续的 DataType 列表（不含第一项 1001）；例如 [1000, 1040, 1070] 等
      data_values    -- 与 data_types 对应的数据值列表；长度与 data_types 一一对应。例如 ["文字串", 3.14, 42]

    说明：
      AutoCAD 规定 XData 的第一对元素必须是 (1001, 应用程序名)。后面才是按顺序出现的其他 (DataType, Data)。
      因此实际发送给 SetXData 的 DataType 数组第一个元素要放 1001，Data 数组第一个元素要放 app_name。
    """
    def vtint(val):
        """
        将 Python 列表转换为 VARIANT 类型的整数数组，
        以便传给 COM 对象作为 XData 的 DataType。
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, val)

    def vtvariant(var):
        """
        将 Python 列表转换为 VARIANT 类型的 VARIANT 数组，
        以便传给 COM 对象作为 XData 的 Data。
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, var)


    # 1. 拼接完整的 DataType 列表：第一个为 1001
    full_types = [1001] + data_types

    # 2. 拼接完整的 Data 列表：第一个为 app_name
    full_data = [app_name] + data_values

    # 3. 将 Python 列表转换为 VARIANT 数组
    vt_types = vtint(full_types)
    vt_data  = vtvariant(full_data)

    # 4. 调用 COM 的 SetXData 方法
    com_obj.SetXData(vt_types, vt_data)

def get_xdata(
    com_obj,
    app_name: str,
):
    """
    从任意 AutoCAD COM 对象（如 Line、Circle、BlockReference 等）读取 XData。

    参数：
      com_obj   -- 任意支持 GetXData() 方法的 COM 对象
      app_name  -- 申请读取的应用程序名（必须与 set_xdata 时使用的相同）

    返回：
      (type_codes, data_values) 二元组，其中
        type_codes: Python 列表，对应每个 XData 项目的 DataType（包括第一项 1001）
        data_values: Python 列表，对应每个 XData 项目的 Data（包括第一项 app_name）
    
    如果该对象没有附加此 app_name 下的 XData，则 GetXData 会抛出错误；建议调用前先用 Error Handling 包裹或
    通过 com_obj.GetXData(app_name) 进行捕获并返回 None。
    """
    def vtint(val):
        """
        将 Python 列表转换为 VARIANT 类型的整数数组，
        以便传给 COM 对象作为 XData 的 DataType。
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, val)

    def vtvariant(var):
        """
        将 Python 列表转换为 VARIANT 类型的 VARIANT 数组，
        以便传给 COM 对象作为 XData 的 Data。
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, var)


    try:
        type_codes, data_values = com_obj.GetXData(app_name)
        # 注意：返回的 type_codes 和 data_values 都是 tuple，转换为 list 更易处理
        return list(type_codes), list(data_values)
    except pythoncom.com_error:
        # 对象上不存在该 app_name 的 XData，或读取失败
        return None, None


#&&% Xdata标记

def set_xdata_tab(entitycom):

    app_name    = "PrintApp"
    data_types  = [1000]
    data_values = ["增补目录模板"]
    set_xdata(entitycom, app_name, data_types, data_values)

    return

def is_printApp_xdata_com(entitycom):

    try:

        get_xdata( entitycom, "PrintApp")

        return True

    except:

        return  False


#&&&&%%  第五部分  线面分析 


#_____________________________________________________________________________________________________________________________________________

#  模块使用说明

"""
该模块研究dwg图纸中的线段、圆曲线、平面等基本几何问题 

"""
# 线段分析

def compute_line_angle(line):#按绘制顺序度量线段角度
    """
    计算直线的方向角（单位：度），基于 StartPoint / EndPoint。
    非直线对象将抛出异常。
    0-360，从起点处画横线，旋转到终点的角度
    """

    try:
        x1, y1, _ = line.StartPoint
        x2, y2, _ = line.EndPoint
        dx = x2 - x1
        dy = y2 - y1
        angle_rad = math.atan2(dy, dx)
        angle_deg = math.degrees(angle_rad)
        if angle_deg < 0:
            angle_deg += 360
        return angle_deg
    except AttributeError:
        print("❌ 该对象不具备 StartPoint / EndPoint")
        return None


def draw_point(pt):
    """
    在模型空间绘制一个 AutoCAD 点实体。

    参数：
        pt: (x, y, z) 三维坐标元组

    返回：
        新创建的 Point 对象；失败时返回 None
    """
    try:
        # AutoCAD 的“点”由 AddPoint 创建，需传 VARIANT
        obj = mp.AddPoint(vtpnt(*pt))
        return obj
    except Exception as e:
        print(f"❌ 无法绘制点: {e}")
        return None

def draw_line(p1, p2):#从两点坐标返回直线段
    """
    在模型空间中绘制从 p1 到 p2 的直线段。

    参数：
        p1, p2: 三维坐标元组 (x, y, z)

    返回：
        新创建的直线对象（COM 对象）
    """
    try:
        line_obj = mp.AddLine(vtpnt(*p1), vtpnt(*p2))
        return line_obj
    except Exception as e:
        print(f"❌ 无法绘制直线: {e}")
        return None


def draw_circle(center, radius):
    """
    以 center 为圆心、radius 为半径绘制圆。

    参数：
        center: (x, y, z)
        radius: 浮点半径

    返回：
        新创建的 Circle 对象；失败时返回 None
    """
    try:
        obj = mp.AddCircle(vtpnt(*center), radius)
        return obj
    except Exception as e:
        print(f"❌ 无法绘制圆: {e}")
        return None


def draw_regular_polygon(center, radius, sides):
    """
    绘制正多边形（LWPolyline，已闭合）
    :param center: 圆心 (x,y,z)
    :param radius: 外接圆半径
    :param sides : 边数 ≥3
    """
    if sides < 3:
        print("❌ 边数必须 ≥ 3"); return None
    cx, cy, cz = (center + (0.0,))[:3]

    pts_flat = []
    for k in range(sides):
        ang = 2 * math.pi * k / sides
        pts_flat.extend([cx + radius*math.cos(ang),
                         cy + radius*math.sin(ang)])

    try:
        v_pts = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8,
            pts_flat
        )
        poly = mp.AddLightWeightPolyline(v_pts)
        poly.Closed = True
        return poly
    except Exception as e:
        print(f"❌ 无法绘制正多边形: {e}")
        return None



def prioritize_horizontal(lines, tol=0.5):
    """
    将列表中所有“水平”直线段（起点和终点的 y 差小于 tol）放在最前面，
    其它直线保留原有相对顺序。

    :param lines: 直线对象列表，每个对象具有 .StartPoint 和 .EndPoint 属性，
                  这两个属性应返回 (x, y, z) 或类似可下标的三元组。
    :param tol:   判定为水平的 y 方向容差（默认 0.5）
    :return:      新列表，水平直线段在前，非水平在后
    """
    horizontals = []
    non_horizontals = []
    for ln in lines:
        y1 = ln.StartPoint[1]
        y2 = ln.EndPoint[1]
        if abs(y1 - y2) < tol:
            horizontals.append(ln)
        else:
            non_horizontals.append(ln)
    return horizontals,non_horizontals

    
def get_spline_length_by_conversion(spline_entity):#返回样条曲线的长度（按默认10分断拟合）
    
    """
    将样条曲线对象复制、高亮并通过 _SPLINEDIT 转为多段线，
    然后读取长度，并删除该多段线。

    返回：
        样条曲线转换后的长度值（float）
    """
    try:
        # Step 1：复制 spline 对象
        new_spline = spline_entity.Copy()

        # Step 2：显性高亮
        highlight_entity_by_bbox(new_spline)

        # Step 3：模拟命令 SPLINEDIT → P → Enter → Enter
        doc.SendCommand("_SPLINEDIT\nP\n\n\n")
        time.sleep(1.2)  # 等待 CAD 完成处理（可根据机器速度调整）

        # Step 4：获取新生成的对象（最后一个）
        last_index = doc.ModelSpace.Count - 1
        poly = doc.ModelSpace.Item(last_index)

        # Step 5：检查 Length 属性
        if hasattr(poly, "Length"):
            length = poly.Length
            poly.Delete()  # 删除临时 polyline
            return length
        else:
            print("❌ 转换后对象没有 Length 属性")
            return None

    except Exception as e:
        print(f"❌ 获取样条曲线长度失败：{e}")
        return None

def estimate_ellipse_length(ellipse):#返回椭圆长度
    """
    估算椭圆对象的长度（周长），使用 Ramanujan 公式。
    """
    try:
        a = ellipse.MajorRadius
        b = ellipse.MinorRadius

        pi = math.pi
        h = 3 * (a + b) - math.sqrt((3 * a + b) * (a + 3 * b))
        length = pi * h
        return length
    except Exception as e:
        print(f"❌ 无法估算椭圆长度: {e}")
        return None



def get_entity_geometry_info(obj):#返回图形关键几何信息
    """
    根据图元类型返回其关键几何信息：
    - 点：坐标
    - 直线：起点、终点、长度
    - 圆：圆心、半径、长度、面积
    - 椭圆：中心、主轴、次轴、长度、面积
    - 多段线：起点、终点、长度、面积（若闭合）
    - 样条曲线：起点、终点、长度（需转换），面积（若闭合）
    """
    try:
        name = obj.ObjectName.lower()

        # 点
        if "point" in name:
            return {"type": "Point", "position": obj.Coordinates}

        # 直线
        elif "line" in name and "xline" not in name:
            p1 = obj.StartPoint
            p2 = obj.EndPoint
            length = math.dist(p1, p2)
            return {"type": "Line", "start": p1, "end": p2, "length": length}

        # 圆
        elif "circle" in name:
            center = obj.Center
            radius = obj.Radius
            length = 2 * math.pi * radius
            area = math.pi * radius ** 2
            return {"type": "Circle", "center": center, "radius": radius, "length": length, "area": area}

        # 椭圆
        elif "ellipse" in name:
            center = obj.Center
            a = obj.MajorRadius
            b = obj.MinorRadius
            area = math.pi * a * b
            h = 3 * (a + b) - math.sqrt((3 * a + b) * (a + 3 * b))
            length = math.pi * h  # Ramanujan 公式
            return {
                "type": "Ellipse",
                "center": center,
                "major_radius": a,
                "minor_radius": b,
                "length": length,
                "area": area
            }

        # 多段线
        elif "polyline" in name:
            coords = obj.Coordinates
            start = (coords[0], coords[1], 0)
            end = (coords[-2], coords[-1], 0)
            length = getattr(obj, "Length", 0)
            area = obj.Area if obj.Closed else 0
            return {
                "type": "Polyline",
                "start": start,
                "end": end,
                "length": length,
                "area": area
            }

        # 样条曲线（需转换测量）
        elif "spline" in name:
            p1 = obj.GetFitPoint(0)
            p2 = obj.GetFitPoint(obj.NumberOfFitPoints - 1)
            length = get_spline_length_by_conversion(obj)
            area = obj.Area if obj.Closed else 0
            return {
                "type": "Spline",
                "start": p1,
                "end": p2,
                "length": length,
                "area": area
            }

        else:
            return {"type": "Unknown", "ObjectName": obj.ObjectName}

    except Exception as e:
        return {"type": "Error", "message": str(e)}



##在两点确定的方向上，返回与对象点指定距离的点

def points_on_line_at_distance_3d(
    p1: Tuple[float, float, float],
    p2: Tuple[float, float, float],
    px: Tuple[float, float, float],
    distance: float
) -> List[Tuple[float, float, float]]:
    """
    已知 px 在由 p1->p2 确定的直线上，返回在该直线上与 px 距离为 distance 的两个点。

    :param p1: 起点 (x1, y1, z1)
    :param p2: 终点 (x2, y2, z2)
    :param px: 参考点 (x, y, z)，位于直线上
    :param distance: 与 px 的目标距离
    :return: [(ax, ay, az), (bx, by, bz)]，分别为正向和反向移动 distance 后的点
    """
    x1, y1, z1 = p1
    x2, y2, z2 = p2
    xx, yy, zz = px

    # 1) 计算方向向量并归一化
    dx, dy, dz = x2 - x1, y2 - y1, z2 - z1
    length = math.sqrt(dx*dx + dy*dy + dz*dz)
    if length == 0:
        raise ValueError("p1 和 p2 重合，无法确定方向向量")
    ux, uy, uz = dx / length, dy / length, dz / length

    # 2) 沿正向和反向各移动 distance
    ax = xx + ux * distance
    ay = yy + uy * distance
    az = zz + uz * distance

    bx = xx - ux * distance
    by = yy - uy * distance
    bz = zz - uz * distance

    return [(ax, ay, az), (bx, by, bz)]


# 找出一组直线段内的伪相交区域

def find_fake_intersection_regions(lines, tol=10, real_tol=0.01):
    """
    查找伪相交区域：对于任意线段 A 的端点 P，若：
    - 存在其他线段 B 满足 P 到 B 距离 < tol，且
    - 对所有 B，P 到 B 的距离 >= real_tol
    则判定为伪相交点。
    在模型空间中绘制圆（半径 1000）表示这些点。
    """
    ensure_layer("测试辅助")
    ms = doc.ModelSpace
    added = []

    def point_to_line_distance(p, a1, a2):
        x0, y0 = p[:2]
        x1, y1 = a1[:2]
        x2, y2 = a2[:2]
        dx, dy = x2 - x1, y2 - y1
        if dx == dy == 0:
            return math.hypot(x0 - x1, y0 - y1)
        t = max(0, min(1, ((x0 - x1) * dx + (y0 - y1) * dy) / (dx*dx + dy*dy)))
        proj_x = x1 + t * dx
        proj_y = y1 + t * dy
        return math.hypot(x0 - proj_x, y0 - proj_y)

    for A in lines:
        try:
            p1 = A.StartPoint
            p2 = A.EndPoint
        except Exception:
            continue

        for pt in [p1, p2]:
            pt_key = tuple(round(c, 3) for c in pt)
            if pt_key in added:
                continue

            min_dist = 1e10
            has_near = False
            has_real_near = False

            for B in lines:
                if B == A:
                    continue
                try:
                    b1, b2 = B.StartPoint, B.EndPoint
                    dist = point_to_line_distance(pt, b1, b2)
                    if dist < tol:
                        has_near = True
                    if dist < real_tol:
                        has_real_near = True
                        break
                except:
                    continue

            if has_near and not has_real_near:
                ms.AddCircle(vtpnt(*pt), 1000).Layer = "测试辅助"
                added.append(pt_key)
                print(f"✅ 伪相交区域点: {pt}")

    print("✅ 伪相交区域绘制完成")



# 把区域内的直线段交点打断

def lines_daduan(start_point,end_point):#全部脚本统一采用三维坐标点模式

    """
    这个命令对于避免天正墙体没有出现不相交的覆盖是非常重要的，直接应用天正的tlinebk

    还要先处理假相交点区域待优化20250409

    """

    # 使用 f-string 语法将三维坐标变量插入命令字符串中
    start_point_str = f"{start_point[0]},{start_point[1]},{start_point[2]}"

    end_point_str = f"{end_point[0]},{end_point[1]},{end_point[2]}"

    command = f"tlinebk{chr(13)}{start_point_str}{chr(13)}{end_point_str}{chr(13)}{chr(13)}{chr(13)}"

    acad.ActiveDocument.SendCommand(command)






#找出一组直线段中的所有直线段中所有重复的线段并删除

def delete_duplicate_lines(lines, tol=0.01):
    """
    删除重复的直线段，仅保留每组中一条。

    参数：
        lines: 模型空间中所有线段对象列表（ObjectName 为 'AcDbLine'）
        tol: 距离容差，小于此值认为两点重合
    """
    def is_same_point(p1, p2):
        return all(abs(a - b) < tol for a, b in zip(p1, p2))

    def is_duplicate(line1, line2):
        try:
            a1, a2 = line1.StartPoint, line1.EndPoint
            b1, b2 = line2.StartPoint, line2.EndPoint
            return (
                (is_same_point(a1, b1) and is_same_point(a2, b2)) or
                (is_same_point(a1, b2) and is_same_point(a2, b1))
            )
        except:
            return False

    keep = []
    to_delete = []

    for i, line in enumerate(lines):
        is_dup = False
        for j in range(i):
            if is_duplicate(line, lines[j]):
                is_dup = True
                break
        if is_dup:
            to_delete.append(line)
        else:
            keep.append(line)

    count = 0
    for dup in to_delete:
        try:
            dup.Delete()
            count += 1
        except:
            continue

    print(f"✅ 删除了 {count} 条重复直线段，保留 {len(keep)} 条。")
    return keep



#删除完全或局部重复线段

def delete_redundant_lines(lines, tol=0.01):
    """
    删除重复线段和局部重复线段，只保留每组中的一条。
    """
    def is_same_point(p1, p2):
        return abs(p1[0] - p2[0]) < tol and abs(p1[1] - p2[1]) < tol

    def point_on_segment(p, a, b):
        ax, ay = a[:2]
        bx, by = b[:2]
        px, py = p[:2]
        cross = abs((bx - ax) * (py - ay) - (by - ay) * (px - ax))
        if cross > tol:
            return False
        dot = (px - ax) * (bx - ax) + (py - ay) * (by - ay)
        if dot < 0:
            return False
        sq_len = (bx - ax)**2 + (by - ay)**2
        if dot > sq_len:
            return False
        return True

    def is_completely_duplicate(l1, l2):
        try:
            p1, p2 = l1.StartPoint, l1.EndPoint
            q1, q2 = l2.StartPoint, l2.EndPoint
            return (
                (is_same_point(p1, q1) and is_same_point(p2, q2)) or
                (is_same_point(p1, q2) and is_same_point(p2, q1))
            )
        except:
            return False

    def is_locally_duplicate(short_line, long_line):
        try:
            p1, p2 = short_line.StartPoint, short_line.EndPoint
            q1, q2 = long_line.StartPoint, long_line.EndPoint
            return point_on_segment(p1, q1, q2) and point_on_segment(p2, q1, q2)
        except:
            return False

    to_delete_handles = set()
    total = len(lines)

    for i in range(total):
        l1 = lines[i]
        h1 = l1.Handle
        if h1 in to_delete_handles:
            continue
        for j in range(i + 1, total):
            l2 = lines[j]
            h2 = l2.Handle
            if h2 in to_delete_handles:
                continue

            if is_completely_duplicate(l1, l2):
                to_delete_handles.add(h2)
            elif is_locally_duplicate(l2, l1):
                to_delete_handles.add(h2)
            elif is_locally_duplicate(l1, l2):
                to_delete_handles.add(h1)
                break

    deleted = 0
    for ent in lines:
        if ent.Handle in to_delete_handles:
            try:
                ent.Delete()
                deleted += 1
            except:
                continue

    print(f"✅ 删除重复/局部重复线段 {deleted} 条，保留 {total - deleted} 条。")

#找出一组直线段中的孤立线段产生的交点


def find_isolated_intersections(LB, tol=0.5):
    """
    找出线段列表 LB 中的孤立线段，并计算它们与其它线段的所有交点。

    用于人工标记的门窗位置

    参数：
      LB:   线段列表，每个元素是 [(x1,y1,z1), (x2,y2,z2)]
      tol:  端点重合判断容差

    返回：
      intersections: 交点列表，每个元素是 (x, y, z)
    """
    def segment_intersection(seg1, seg2, tol):
        """
        计算线段 seg1=(A,B) 与 seg2=(C,D) 的交点（二维），
        若相交且唯一，返回 (x, y, z)，否则返回 None。
        z 取 seg1 第一端点的 z。
        """
        (x1, y1, z1), (x2, y2, _) = seg1
        (x3, y3, _),  (x4, y4, _) = seg2

        # 方向向量
        r = (x2-x1, y2-y1)
        s = (x4-x3, y4-y3)
        # 叉积 r × s
        rxs = r[0]*s[1] - r[1]*s[0]
        if abs(rxs) < tol:
            return None  # 平行或共线，不处理
        # 解 t, u
        qp = (x3-x1, y3-y1)
        t = (qp[0]*s[1] - qp[1]*s[0]) / rxs
        u = (qp[0]*r[1] - qp[1]*r[0]) / rxs
        # 只考虑严格交于段内
        if -tol <= t <= 1+tol and -tol <= u <= 1+tol:
            xi = x1 + t*r[0]
            yi = y1 + t*r[1]
            zi = z1
            return (xi, yi, zi)
        return None

    # 1. 找出孤立线段
    isolated = []
    for i, seg in enumerate(LB):
        p1, p2 = seg
        shared = False
        for j, other in enumerate(LB):
            if i == j:
                continue
            q1, q2 = other
            if same_point(p1, q1, tol) or same_point(p1, q2, tol) \
            or same_point(p2, q1, tol) or same_point(p2, q2, tol):
                shared = True
                break
        if not shared:
            isolated.append(seg)

    # 2. 对每根孤立线段，与其余线段求交点
    intersections = []
    for seg in isolated:
        for other in LB:
            if other is seg:
                continue
            ip = segment_intersection(seg, other, tol)
            if ip is not None:
                intersections.append(ip)

    #删除孤立线段20250420
    for seg in isolated:

        seg.Delete()

    return intersections    


##doc.sendcommand("TSpOutline"+chr(13)+"41849.69465957, 12250.50102376, 0"+chr(13)+chr(13)+chr(13))

##doc.sendcommand("TRoflna"+chr(13)+"0"+chr(13))


def get_inner_point_of_polygon(polygon: Polygon):
    """
    获取给定 polygon 的一个保证在其内部的点。

    参数：
        polygon (shapely.geometry.Polygon): 目标多边形

    返回：
        (x, y): 内部点坐标元组
    """
    if not isinstance(polygon, Polygon):
        raise ValueError("❌ 输入必须是 shapely.geometry.Polygon")

    inner_point = polygon.representative_point()
    return inner_point.x, inner_point.y


#&&%____________________________________________________       获取一组直线段所有的封闭多边形和外轮廓线       ________________
#…………………………………………………………………………………………………………………………………………………………………
# 线面分析 - 获取一组直线段所有的封闭多边形和外轮廓线

#__________________________________________________________________________________________________________________________
#…………………………………………………………………………………………………………………………………………………………………
#  使用说明

"""
获取一组直线段所有的封闭多边形和外轮廓线

主要函数有：
(1)寻找直线段组中最右下角封闭多边性
(2)从com边或顶点坐标列表用PL复线绘制多边形
(3)获取一组直线段的外轮廓线
(4)获取最右下角的封闭多边形不影响其他封闭多边形的连续边的顶点列表
(5)获取全部封闭多边形，但不完全
(6)获取全部封闭多边形

"""
def get_room_outline_from_point(x, y, z=0):## 获取输入点所在房间的轮廓
    """
    自动发送 TSpOutline 命令，从指定点获取房间轮廓。

    参数:
        x, y, z: 点的坐标，z 默认为 0。

    也许会有用处，能从多边形内点迅速得到多边形
        
    """
    try:
        point_str = f"{x},{y},{z}"
        cmd = (
            "TSpOutline" + chr(13) +  # 启动命令
            point_str + chr(13) +     # 输入点
            chr(13) +                 # 确认默认设置
            chr(13)                   # 确认生成
        )
        doc.SendCommand(cmd)
        print(f"✅ 已请求获取点 ({x},{y},{z}) 的房间轮廓。")

    except Exception as e:
        print(f"❌ 获取房间轮廓失败：{e}")


def connect_lines_to_polyline_if_closed(lines, tol=0.5):##判断一组闭合直线段是否构成封闭多段线，是就返回PL复线
    """
    判断线段是否首尾连接成闭合多边形，如果是，则绘制PL多段线。
    
    参数:
        lines: AutoCAD中选中的 AcDbLine 对象列表。
        tol: 容许的端点闭合误差。
    
    返回:
        多段线对象，或 None。
    """

    try:
        # 提取二维端点集合（忽略z）
        segments = []
        for ln in lines:
            try:
                p1 = (ln.StartPoint[0], ln.StartPoint[1])
                p2 = (ln.EndPoint[0], ln.EndPoint[1])
                segments.append((p1, p2))
            except Exception:
                continue

        if not segments:
            print("⚠️ 无有效线段")
            return None

        # 构建连接链条
        used = set()
        sequence = [segments[0][0]]
        current = segments[0][1]
        used.add(0)

        while True:
            found = False
            for idx, (a, b) in enumerate(segments):
                if idx in used:
                    continue
                if Point(current).distance(Point(a)) < tol:
                    sequence.append(current)
                    current = b
                    used.add(idx)
                    found = True
                    break
                elif Point(current).distance(Point(b)) < tol:
                    sequence.append(current)
                    current = a
                    used.add(idx)
                    found = True
                    break
            if not found:
                break

        # 检查是否闭合
        if Point(current).distance(Point(sequence[0])) > tol:
            print("❌ 线段未构成闭合区域")
            return None
        sequence.append(sequence[0])  # 闭合环

        # 构造二维点数组
        pts = []
        for pt in sequence:
            pts.extend([pt[0], pt[1]])

        # 绘制PL
        poly = doc.ModelSpace.AddLightWeightPolyline(vtFloat(pts))
        poly.Closed = True
        print("✅ 成功绘制封闭PL线")
        return poly

    except Exception as e:
        print(f"❌ Polyline 创建失败: {e}")
        return None


def is_closed_polygon_from_lines(lines, tol=0.5):##判断一组闭合直线段是否构成封闭多段线，不返回PL复线
    """
    判断一组 AutoCAD 直线段是否首尾连接形成闭合多边形。
    
    参数:
        lines: AcDbLine 类型的 COM 对象列表
        tol: 闭合判断容差，单位与CAD一致（如mm）

    返回:
        True 表示首尾闭合形成多边形，False 否则
    """
    try:
        # 提取二维端点 (x, y)
        segments = []
        for ln in lines:
            try:
                p1 = (ln.StartPoint[0], ln.StartPoint[1])
                p2 = (ln.EndPoint[0], ln.EndPoint[1])
                segments.append((p1, p2))
            except Exception:
                continue

        if not segments:
            return False

        # 构造首尾连接链
        used = set()
        sequence = [segments[0][0]]
        current = segments[0][1]
        used.add(0)

        while True:
            found = False
            for idx, (a, b) in enumerate(segments):
                if idx in used:
                    continue
                if Point(current).distance(Point(a)) < tol:
                    sequence.append(current)
                    current = b
                    used.add(idx)
                    found = True
                    break
                elif Point(current).distance(Point(b)) < tol:
                    sequence.append(current)
                    current = a
                    used.add(idx)
                    found = True
                    break
            if not found:
                break

        # 判断是否回到起点（闭合）
        if Point(current).distance(Point(sequence[0])) < tol and len(used) == len(segments):
            return True
        else:
            return False

    except Exception as e:
        print(f"❌ 判断失败: {e}")
        return False

def same_point(p1, p2, tol=0.5):
    """判断两个点是否在容差范围内相同（只比较 X、Y 坐标）"""
    return abs(p1[0] - p2[0]) <= tol and abs(p1[1] - p2[1]) <= tol


def same_line(ln1, ln2, tol=0.5):
    """
    判断两条线段 ln1 和 ln2 是否“相同”
    线段被认为相同的条件是：
      ln1 的 StartPoint 与 ln2 的 StartPoint 近似相同且 ln1 的 EndPoint 与 ln2 的 EndPoint 近似相同，
      或者 ln1 的 StartPoint 与 ln2 的 EndPoint 近似相同且 ln1 的 EndPoint 与 ln2 的 StartPoint 近似相同。
    """
    p1 = tuple(ln1.StartPoint)
    p2 = tuple(ln1.EndPoint)
    q1 = tuple(ln2.StartPoint)
    q2 = tuple(ln2.EndPoint)

    return (same_point(p1, q1, tol) and same_point(p2, q2, tol)) or \
           (same_point(p1, q2, tol) and same_point(p2, q1, tol))


def calculate_absolute_angle(line, P, tol=0.5):
    """
    计算线段（line）从点 P 出发的绝对角度（0-360°）
    如果 line 的起点等于 P，则返回从 P 到终点的角度，否则返回从 P 到起点的角度。
    """
    sp = tuple(line.StartPoint)
    ep = tuple(line.EndPoint)
    if same_point(sp, P, tol):
        dx = ep[0] - P[0]
        dy = ep[1] - P[1]
    else:
        dx = sp[0] - P[0]
        dy = sp[1] - P[1]
    return math.degrees(math.atan2(dy, dx)) % 360

def calculate_relative_angle(line, P, current_line, tol=0.5):
    """
    计算当前参考线（current_line）与候选线段（line）之间的相对角度，
    角度是以共点 P 为中心，当前线从 P 到其非 P 的端点的绝对角度为基准，
    计算 candidate line 与该基准角度之间顺时针或逆时针的角度差（逆时针方向）。
    结果为 0 到 360 之间的数值。
    """
    sp_current = tuple(current_line.StartPoint)
    ep_current = tuple(current_line.EndPoint)
    sp = tuple(line.StartPoint)
    ep = tuple(line.EndPoint)

    # 选择当前线段中不等于 P 的端点作为参考
    if same_point(sp_current, P, tol):
        current_point = ep_current
    else:
        current_point = sp_current

    def angle(p1, p2):
        dx = p2[0] - p1[0]
        dy = p2[1] - p1[1]
        return math.degrees(math.atan2(dy, dx)) % 360

    angle_current = angle(P, current_point)
    # 对候选线段，选择不等于 P 的端点
    if same_point(sp, P, tol):
        target_point = ep
    else:
        target_point = sp
    angle_target = angle(P, target_point)
    angle_diff = (angle_target - angle_current) % 360
    return angle_diff

#####################
# 函数：查找给定点P经过的线段，按照绝对角度排序

def find_lines_angle(lines, P, tol=0.5):
    """
    查找与指定点 P 共端点的所有线段，并按从 P 出发离开的绝对几何角度排序。

    参数:
        lines: 直线段对象列表，每个对象要求具备 StartPoint, EndPoint, Handle 属性。
        P: 三元组 (x, y, z)，目标共点。
        tol: 判断共点的容差（仅比较 x 和 y 坐标）。

    返回:
        按绝对角度从小到大排序的共点线段列表。
    """
    shared_lines = []
    P = tuple(P)
    print(f"调试：目标共点 P = {P}")
    
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
            dx_sp = abs(sp[0] - P[0])
            dy_sp = abs(sp[1] - P[1])
            dx_ep = abs(ep[0] - P[0])
            dy_ep = abs(ep[1] - P[1])
            print(f"线段 {ln.Handle}: sp={sp} 差值=({dx_sp:.4f},{dy_sp:.4f}), ep={ep} 差值=({dx_ep:.4f},{dy_ep:.4f})")
            if (dx_sp <= tol and dy_sp <= tol) or (dx_ep <= tol and dy_ep <= tol):
                shared_lines.append(ln)
        except Exception as e:
            print(f"⚠️ 跳过无效线段 {getattr(ln, 'Handle', '未知')} : {e}")
            continue

    shared_lines.sort(key=lambda ln: calculate_absolute_angle(ln, P, tol))
    print("调试：按绝对角度排序后的共点线段：")
    for ln in shared_lines:
        print(f"  线段 {ln.Handle}：角度 = {calculate_absolute_angle(ln, P, tol):.2f}°")
    return shared_lines

#####################
# 函数：查找与P共点的线段，按照与当前线段的相对角度排序

def find_lines_sharing_point(lines, P, current_line, tol=0.5):
    """
    查找与指定点 P 共端点的所有线段，并按从 current_line 逆时针旋转到其他线段的相对角度排序。
    其中 current_line 的角度定义为 0°。
    
    参数:
        lines: 直线段对象列表，每个对象要求具备 StartPoint, EndPoint, Handle 属性。
        P: 共点，三元组 (x, y, z)。
        current_line: 当前参考线段，对应角度定义为 0°。
        tol: 判断共点的容差（只比较 x 和 y 坐标）。
        
    返回:
        按相对角度从小到大排序的经过 P 的线段列表（current_line 亦包括其中）。
    """
    shared_lines = []
    P = tuple(P)
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
            if (abs(sp[0]-P[0])<=tol and abs(sp[1]-P[1])<=tol) or (abs(ep[0]-P[0])<=tol and abs(ep[1]-P[1])<=tol):
                shared_lines.append(ln)
        except Exception as e:
            print(f"⚠️ 跳过无效线段 {getattr(ln, 'Handle', '未知')} : {e}")
            continue
    # 根据从当前线段旋转（逆时针）的相对角度排序
    shared_lines.sort(key=lambda ln: calculate_relative_angle(ln, P, current_line, tol))
    return shared_lines

#####################
# 函数：根据当前线段和共点 P，选择下一条后继线段（选择相对角度最大的那条），返回 (后继线段, 新共点)

def find_successor_line_max(current_line, lines, P, tol=0.5):
    """
    在给定共点 P 处，排除当前线段（current_line）后，
    选择在该点处与当前线段相对旋转角度最大的那条线段作为后继线段，
    返回 (后继线段, 新共点)。

    参数：
      current_line: 当前线段对象。
      lines: 所有直线段对象列表，每个对象必须具备 StartPoint, EndPoint, Handle 属性。
      P: 共点（三元组）。
      tol: 共点判断的容差（默认 0.5）。

    返回：
      (后继线段, 新共点)，若找不到合适的后继线段，则返回 (None, P)。
    """
    # 调用 find_lines_sharing_point 获取所有经过 P 的线段，并使用 current_line 的相对角度排序
    candidates = find_lines_sharing_point(lines, P, current_line, tol)
    # 排除当前线段
    candidates = [ln for ln in candidates if ln.Handle != current_line.Handle]

    if not candidates:
        print(f"❌ 在点 {P} 处找不到除当前线外的候选后继线段")
        return None, P

    best_line = None
    max_angle = -1
    new_point = P
    for ln in candidates:
        relative_angle = calculate_relative_angle(ln, P, current_line, tol)
        if relative_angle > max_angle:
            max_angle = relative_angle
            if same_point(tuple(ln.StartPoint), P, tol):
                new_point = tuple(ln.EndPoint)
            else:
                new_point = tuple(ln.StartPoint)
            best_line = ln

    if best_line is None:
        print(f"❌ 没有找到有效的后继线段")
        return None, P

    print(f"选中后继线段 {best_line.Handle}，新共点 {new_point}")
    return best_line, new_point

#&&%#####################
# 辅助函数：从所有线段中找出最右下角的点
def find_rightbottom_point(lines, tol=0.5):
    """
    从所有线段端点中，找出 y 值最小的点；若有多个，则选 x 最大的点作为最右下角点。
    """
    all_points = []
    for line in lines:
        if hasattr(line, "StartPoint") and hasattr(line, "EndPoint"):
            all_points.append(tuple(line.StartPoint))
            all_points.append(tuple(line.EndPoint))
    if not all_points:
        return None
    min_y = min(p[1] for p in all_points)
    candidates = [p for p in all_points if abs(p[1]-min_y) <= tol]
    rb = max(candidates, key=lambda p: p[0])
    print(f"✅ 最右下角点为：{rb}")
    return rb



#  主函数
#  (1)
# 寻找直线段组中最右下角封闭多边性

#  该函数系列包括如下一些函数


"""
辅助函数

same_point(p1, p2, tol) 判断两点是否在容差范围内相同

calculate_absolute_angle(line, P, tol) 计算从点 P 出发到某条线段（取与 P 不相同的端点）的绝对角度

calculate_relative_angle(line, P, current_line, tol) 计算当前线（以 P 为起点，选取与 P 不相同的端点）

和候选线（以 P 为起点）的角度差（逆时针方向，结果在 0～360 之间）

函数 find_lines_angle
用于初始阶段：给定一个共点 P，找到所有经过该点的线段，并按它们的绝对角度排序（从小到大）。

函数 find_lines_sharing_point
给定共点 P 与一个“当前线段”（作为参考），返回所有经过 P 的线段，并按照从当前线段出发旋转（逆时针）的相对角度排序。

函数 find_successor_line_max
根据当前线段和共点 P，从通过 P 的候选线中选取“转角”最大的作为后继线段，并返回后继线段及该线段另一端的新共点。

函数 find_rightbottom_closed_polygon
利用上述函数构造封闭多边形。初始时先根据所有线段端点确定“最右下角”点（函数 find_rightbottom_point），然后在该点处调用

find_lines_angle 得到所有经过该点的线段，选取绝对角度最小的那一根作为第一条边（初始 current_line），再依次调用

find_successor_line_max 推进多边形直到闭合或达到最大步数。


"""

def find_rightbottom_closed_polygon(lines, tol=0.5, max_steps=50):
    """
    利用所有线段构造封闭多边形：
      1. 先查找所有线段端点中的最右下角点（RB）。
      2. 在 RB 处获取所有经过该点的线段（按照绝对角度排序），选择绝对角度最小的那根作为初始边。
      3. 以该初始边为第一条边，之后依次利用 find_successor_line_max 选择后继边，
         直到新共点回到初始点（闭合）或超过最大步数。
         
    返回：
      构成封闭多边形的点列表（依次为每条边的终点），若无法构成则返回 None。
    """
    # 定位最右下角点 RB
    rb = find_rightbottom_point(lines, tol)
    if rb is None:
        print("❌ 无右下角点")
        return None

    # 初始共点
    current_point = rb
    # 从 RB 处按绝对角度排序取候选线段
    candidates = find_lines_angle(lines, rb, tol)
    if not candidates:
        print(f"❌ 在右下角点 {rb} 处没有找到经过的线段")
        return None
    # 选择绝对角度最小的线段作为初始边
    current_line = candidates[0]
    print(f"调试：选中初始线段 {current_line.Handle}，绝对角度 = {calculate_absolute_angle(current_line, rb, tol):.2f}°")

    # 得到初始边的另一端
    sp = tuple(current_line.StartPoint)
    ep = tuple(current_line.EndPoint)
    if same_point(sp, rb, tol):
        next_point = ep
    else:
        next_point = sp

    polygon_points = [rb, next_point]
    visited_handles = {current_line.Handle}
    current_point = next_point
    steps = 1

    while steps < max_steps:
        # 调用 find_successor_line_max 得到下一条线段及其另一端的点
        successor, new_point = find_successor_line_max(current_line, lines, current_point, tol)
        if successor is None:
            print("❌ 无后继线段，构造失败")
            return None
        # 检查是否闭合（新共点与起始点接近）
        if same_point(new_point, rb, tol):
            polygon_points.append(rb)
            print(f"✅ 成功构建封闭多边形，步数 = {steps}")
            return polygon_points

        if successor.Handle in visited_handles:
            print("🔁 检测到重复线段，构造失败")
            return None

        polygon_points.append(new_point)
        visited_handles.add(successor.Handle)
        current_line = successor
        current_point = new_point
        steps += 1

    print("⚠️ 达到最大步数，未能构造出闭合多边形")
    return None


# 从com边或顶点坐标列表用PL复线绘制多边形


def draw_polygon_as_polyline(polygon, layer_name="测试辅助", tol=0.5):
    """
    将构造的多边形（polygon）转换为顶点序列，并在当前 AutoCAD 文档 doc 的 ModelSpace 中添加
    一个多段线（PLINE）。如果顶点序列的首尾两点重合，则绘制闭合多段线，否则绘制开放多段线。
    
    参数:
      polygon:
        1. 如果是由线段组成的列表，每个元素要求具备 StartPoint 和 EndPoint 属性，
           则按照这些线段的端点顺序构造顶点序列。
        2. 如果是顶点列表，例如 [(x, y, z), (x, y, z), ...]，则直接使用。
      layer_name: 绘制多段线所在的图层名称（默认为 "测试辅助"）。
      tol: 判断点是否重合的容差值（仅比较 x 和 y 坐标）。
      
    返回:
      成功时返回创建的多段线对象（PLINE），否则返回 None。
    """

    # 内部函数：判断两个点是否近似相等（仅比较 x 和 y 坐标）
##    def same_point(p1, p2, tol=tol):
##        return abs(p1[0] - p2[0]) <= tol and abs(p1[1] - p2[1]) <= tol

    if not polygon:
        print("❌ 未提供有效的 polygon 数据")
        return None

    vertices = []
    is_closed = False  # 是否绘制闭合多段线

    # 判断 polygon 是顶点列表还是线段列表
    if isinstance(polygon[0], (tuple, list)):
        # 判断传入是否为顶点列表：检查第一个元素为 tuple/list 且长度>=3
        if len(polygon[0]) < 3:
            print("❌ 顶点数据格式错误")
            return None
        # 直接使用顶点列表（转换为 tuple 形式）
        vertices = [tuple(pt) for pt in polygon]
        # 如果首尾已经重合，则视为闭合；否则保持开放（不自动补充首点）
        if same_point(vertices[0], vertices[-1], tol):
            is_closed = True
        else:
            is_closed = False
    else:
        # 假定 polygon 是线段列表，每个元素具有 StartPoint 和 EndPoint 属性
        first_line = polygon[0]
        start_pt = tuple(first_line.StartPoint)
        last_line = polygon[-1]
        # 判断哪一端与起始点相连，作为初始点
        if same_point(start_pt, tuple(last_line.StartPoint), tol) or same_point(start_pt, tuple(last_line.EndPoint), tol):
            # start_pt 可以作为起点
            pass
        else:
            # 否则选用第一条线段的 EndPoint作为起始点
            start_pt = tuple(first_line.EndPoint)
        vertices.append(start_pt)
        current_pt = start_pt
        # 遍历每条线段构造顶点序列
        for line in polygon:
            sp = tuple(line.StartPoint)
            ep = tuple(line.EndPoint)
            if same_point(current_pt, sp, tol):
                next_pt = ep
            elif same_point(current_pt, ep, tol):
                next_pt = sp
            else:
                print(f"❌ 线段 {line.Handle} 与当前点 {current_pt} 未连接，构造多边形失败")
                return None
            vertices.append(next_pt)
            current_pt = next_pt
        # 检查是否闭合（首尾重合）
        if same_point(vertices[0], vertices[-1], tol):
            is_closed = True
        else:
            is_closed = False

    # 输出调试信息：打印顶点序列
    print("调试：多段线顶点序列：")
    for i, pt in enumerate(vertices):
        print(f"  顶点 {i}: {pt}")

    # 将顶点序列转换为一维坐标数组：[x1, y1, z1, x2, y2, z2, …]
    coords = []
    for pt in vertices:
        coords.extend([pt[0], pt[1], pt[2]])
    coords_tuple = tuple(coords)
    coords_variant = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, coords_tuple)

    # 确保图层存在，否则创建新图层
    try:
        _ = doc.Layers.Item(layer_name)
    except Exception:
        doc.Layers.Add(layer_name)

    # 在 ModelSpace 中添加多段线并设置相应属性
    try:
        ms = doc.ModelSpace
        polyline = ms.AddPolyline(coords_variant)
        polyline.Closed = is_closed
        polyline.Layer = layer_name
        # 可选设置颜色和宽度，按需要调整
        polyline.Color = 1
        polyline.ConstantWidth = 20
        polyline.Update()
        doc.Regen(0)
        print(f"✅ 成功在图层 '{layer_name}' 绘制多段线, Closed={is_closed}")
        return polyline
    except Exception as e:
        print("❌ 绘制多段线失败：", e)
        return None



#&&%###外轮廓线

# -----------------------------------------------------
# 辅助函数：判断两个点是否近似相等（仅比较 x 和 y 坐标）
def is_nearly_equal(p1, p2, tol):
    return abs(p1[0] - p2[0]) <= tol and abs(p1[1] - p2[1]) <= tol



# -----------------------------------------------------
# 寻找后继线段：在共点 P 处，从当前边之外的候选边中选择相对角度最小的边
def find_successor_line_min(current_line, lines, P, tol=0.5):
    # 先获取共点 P 的所有线段，使用 find_lines_sharing_point 逻辑，但这里直接采用 find_lines_angle 排序后筛选
    candidates = []
    P = tuple(P)
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
            if (abs(sp[0]-P[0]) <= tol and abs(sp[1]-P[1]) <= tol) or (abs(ep[0]-P[0]) <= tol and abs(ep[1]-P[1]) <= tol):
                # 排除当前线段自身
                if ln.Handle == current_line.Handle:
                    continue
                candidates.append(ln)
        except Exception as e:
            continue

    if not candidates:
        print(f"❌ 在点 {P} 处找不到候选后继线段")
        return None, P

    best_line = None
    min_angle = 360
    new_point = P
    for candidate in candidates:
        angle_diff = calculate_relative_angle(candidate, P, current_line, tol)
        # 选择相对角度最小的候选
        if angle_diff < min_angle:
            min_angle = angle_diff
            best_line = candidate
            # 更新新共点：选择 candidate 中不等于 P 的端点
            sp_cand = tuple(candidate.StartPoint)
            ep_cand = tuple(candidate.EndPoint)
            if is_nearly_equal(sp_cand, P, tol):
                new_point = ep_cand
            else:
                new_point = sp_cand

    if best_line is None:
        print(f"❌ 没有找到合适的后继线段")
        return None, P

    print(f"选中后继线段 {best_line.Handle}，新共点 {new_point}，相对转角 = {min_angle:.2f}°")
    return best_line, new_point



#  主函数
#  (3)
# 获取一组直线段的外轮廓线

def get_outer_contour(lines, tol=0.5, max_steps=50):
    """
    获取一组直线段的外轮廓线
    规则：
      1. 计算所有线段端点的最右下角点 P（绝对方法）。
      2. 在 P 处，按绝对角度排序，选择绝对角度最小的边作为第一条边；
      3. 以当前边的另一端点作为新的共点，从当前边出发，选择相对于当前边逆时针转角最小的候选边，
         直到新共点回到初始 P 点（闭合）或达到最大步数为止。
    返回：
      外轮廓线构成的线段列表，如果无法构成封闭轮廓则返回空列表。
    """
    # 内部函数：获取所有线段的最右下角点
    rb = find_rightbottom_point(lines, tol)
    if rb is None:
        print("❌ 无最右下角点")
        return []

    # 设置初始共点为 rb
    P = rb
    print(f"调试：最右下角点为 {P}")

    # 在 P 处，按照绝对角度排序获取所有共点的线段
    candidate_lines = find_lines_angle(lines, P, tol)
    if not candidate_lines:
        print(f"❌ 在点 {P} 处未找到共点线段")
        return []

    # 根据题目要求，第一条边选择绝对角度最小的线段（即排序后第 1 根线段）
    initial_line = candidate_lines[0]
    # 确定初始线段的另一端点
    sp_init = tuple(initial_line.StartPoint)
    ep_init = tuple(initial_line.EndPoint)
    if is_nearly_equal(sp_init, P, tol):
        next_point = ep_init
    else:
        next_point = sp_init
    print(f"调试：选中初始线段 {initial_line.Handle}，起点 {P} -> 终点 {next_point}")
    
    contour_lines = [initial_line]
    visited_handles = {initial_line.Handle}
    current_line = initial_line
    current_point = next_point
    steps = 0

    # 迭代构造外轮廓
    while steps < max_steps:
        print(f"调试：目标共点 P = {current_point}")
        # 在当前共点处查找候选的后继线段，使用最小相对角度策略
        successor, new_point = find_successor_line_min(current_line, lines, current_point, tol)
        if successor is None:
            print("❌ 无后继线段，构造失败")
            return []
        if successor.Handle in visited_handles:
            print(f"🔁 检测到重复线段 {successor.Handle}，构造失败")
            return []
        contour_lines.append(successor)
        visited_handles.add(successor.Handle)
        print(f"步 {steps+1}: 选中线段 {successor.Handle}，新共点 {new_point}")
        # 检查是否闭合：如果新共点与最右下角点近似相等，则认为闭合
        if is_nearly_equal(new_point, rb, tol):
            print("✅ 成功构造封闭轮廓线")
            return contour_lines
        # 更新当前线段及共点
        current_line = successor
        current_point = new_point
        steps += 1

    print("⚠️ 达到最大步数，未能构造封闭轮廓线")
    return []


#&&%##获取所有封闭多边形

#删除多边形的一些边

"""
有lines中多于两条直线段经过，就称该点为多分枝点

编制函数，PL是lines中的封闭多边形，用顶点列表表示，p1是lines的最右下角点，

且p1是PL的一个顶点，用坐标点表示。将PL的顶点排序，从p1按逆时针方向推进，

将遇到的非多分枝顶点放入列表LN，直到遇到一个多分枝点D1结束；再从p1按顺时

针方向推进，将遇到的顶点放入列表LN，直到遇到一个多分枝点D2结束。顺时针方

向推进时，把p1考虑在内，即p1如果为多分枝点则立即结束，结束时遇到的多分枝

点D2就是p1。逆时针方向推进时，不把p1考虑在内，即p1为多分枝点仍然推进，遇

到一个多分枝点D1结束。函数返回多分枝点D1，D2，以及它们之间的非多分枝顶点

列表LN。


"""
def deduplicate_vertices(vertices, tol=0.5):
    """
    去掉顶点列表中相邻重复的顶点：
    如果两个相邻顶点（顺序出现）之间的二维距离小于 tol，则认为它们重复，只保留前一个，
    但如果这个重复点是列表中的最后一个且与第一个顶点相同（表示闭合多边形），则保留此重复点。
    
    参数:
      vertices: 顶点列表，每个顶点格式为 (x, y, z) 的元组。
      tol: 距离阈值。
      
    返回:
      处理后的顶点列表，只有中间连续重复的点被去除，保留闭合多边形的首尾重复。
    """
    if not vertices:
        return []
    
    deduped = [vertices[0]]
    n = len(vertices)
    
    for i in range(1, n):
        pt = vertices[i]
        prev = deduped[-1]
        dx = pt[0] - prev[0]
        dy = pt[1] - prev[1]
        dist = math.sqrt(dx*dx + dy*dy)
        
        # 如果距离大于等于 tol，则认为是不同的顶点，保留它
        if dist >= tol:
            deduped.append(pt)
        else:
            # 如果当前顶点是最后一个，并且与第一个顶点相同，则保留它（表示闭合）
            if i == n - 1 and same_point(pt, vertices[0], tol):
                deduped.append(pt)
            # 否则跳过该点
    return deduped


#  主函数
#  (4)
# 获取最右下角的封闭多边形不影响其他封闭多边形的连续边的顶点列表 

def analyze_polygon_branches(PL, lines, p1, tol=0.5):
    """
    分析封闭多边形 PL 的分枝情况。PL 为封闭多边形的顶点列表（按逆时针顺序排列），
    p1 为最右下角点，必在 PL 中。
    
    规则：
      1. 从 p1 出发：
         - 沿逆时针方向（不将 p1 视为候选）推进，累计遇到的非多分枝顶点，直至遇到第一个多分枝点 D1；
         - 沿顺时针方向（将 p1 也考虑在内）推进，累计遇到的非多分枝顶点，直至遇到第一个多分枝点 D2。
      2. 将这两个方向累计得到的非多分枝顶点 LN（其中顺时针方向得到的顶点需要反转后与逆时针方向的顶点相接）打印出来。
      3. 由于 PL 是闭合多边形，从 D1 到 D2有两条连续顶点序列，从中选择包含 LN 的那条作为最终返回结果。

    返回：
      返回从 D1 到 D2（包含 D1 和 D2）的连续顶点序列，此序列包含了 LN 的顶点。
      如果任一方向未能找到多分枝点，则打印提示，并返回 None。
    """


    # 内部辅助：判断两个点是否近似相等（仅比较 x,y 坐标）
##    def same_point(a, b, tol=tol):
##        return abs(a[0] - b[0]) <= tol and abs(a[1] - b[1]) <= tol
##


    # 内部辅助：判断某顶点是否为多分枝点
    def is_multi_branch(vertex):
        cnt = 0
        for ln in lines:
            try:
                sp = tuple(ln.StartPoint)
                ep = tuple(ln.EndPoint)
            except Exception:
                continue
            if same_point(vertex, sp, tol) or same_point(vertex, ep, tol):
                cnt += 1
        return cnt > 2

    # 找出 p1 在 PL 中的索引
    try:
        idx = PL.index(p1)
    except Exception as e:
        print("错误：p1 不在 PL 中")
        return None

    n = len(PL)

    # 沿逆时针方向推进（不包括 p1），累计非多分枝顶点 LN_ccw，直到遇到第一个多分枝点 D1
    LN_ccw = []
    D1 = None
    i = (idx + 1) % n
    while i != idx:
        v = PL[i]
        if is_multi_branch(v):
            D1 = v
            break
        else:
            LN_ccw.append(v)
        i = (i + 1) % n
    if D1 is None:
        print("调试：逆时针方向未遇到多分枝点")
        return None

    # 沿顺时针方向推进（包含 p1），累计非多分枝顶点 LN_cw，直到遇到第一个多分枝点 D2
    LN_cw = []
    if is_multi_branch(PL[idx]):
        D2 = PL[idx]
    else:
        D2 = None
        i = (idx - 1) % n
        while i != idx:
            v = PL[i]
            if is_multi_branch(v):
                D2 = v
                break
            else:
                LN_cw.append(v)
            i = (i - 1) % n
        if D2 is None:
            print("调试：顺时针方向未遇到多分枝点")
            return None

    # 合并顺时针方向的顶点（需要反转）和逆时针方向的顶点
    LN = list(reversed(LN_cw)) + LN_ccw

    # 调试打印
    print("调试信息：")
    print("目标共点 p1 =", p1)
    print("逆时针方向收集的非多分枝顶点 (LN_ccw):")
    for pt in LN_ccw:
        print("  ", pt)
    print("顺时针方向收集的非多分枝顶点 (LN_cw, 原顺序):")
    for pt in LN_cw:
        print("  ", pt)
    print("合并后的非多分枝顶点 LN (从 D2 到 D1):")
    for pt in LN:
        print("  ", pt)
    print("逆时针方向遇到的多分枝点 D1:", D1)
    print("顺时针方向遇到的多分枝点 D2:", D2)

    # 计算在 PL 中 D1 和 D2 的索引
    try:
        idx_D1 = PL.index(D1)
        idx_D2 = PL.index(D2)
    except Exception as e:
        print("错误：无法在 PL 中查找 D1 或 D2:", e)
        return None

    # 由于 PL 是闭合的，我们有两条连接 D1 和 D2：
    if idx_D1 <= idx_D2:
        branch_A = PL[idx_D1: idx_D2 + 1]
        branch_B = PL[idx_D2:] + PL[:idx_D1 + 1]
    else:
        branch_A = PL[idx_D1:] + PL[:idx_D2 + 1]
        branch_B = PL[idx_D2: idx_D1 + 1]

    # 选择包含 LN 中顶点的那个分枝
    selected_branch = None
    if LN:
        found_in_A = any(same_point(v, LN[0], tol) for v in branch_A)
        found_in_B = any(same_point(v, LN[0], tol) for v in branch_B)
        if found_in_A and not found_in_B:
            selected_branch = branch_A
        elif found_in_B and not found_in_A:
            selected_branch = branch_B
        elif found_in_A and found_in_B:
            # 如果两条分枝都包含，则选择较短的那条，通常这才是“局部”的分枝
            selected_branch = branch_A if len(branch_A) <= len(branch_B) else branch_B
        else:
            print("调试：LN 中的顶点不在任一分枝上，默认选择 branch_A")
            selected_branch = branch_A
    else:
        print("调试：LN 为空，无法确定包含 LN 的分枝，默认选择 branch_A")
        selected_branch = branch_A

    print("最终返回的从 D1 到 D2（包含 LN 的分枝）顶点序列：")
    for pt in selected_branch:
        print("  ", pt)


    #去重
    selected_branch=deduplicate_vertices(selected_branch, tol=tol)

    return selected_branch

##根据输入的顶点列表判断，将lines的其顶点在该顶点列表的线段移出列表


def remove_lines_in_LBv(lines, LB_v, tol=0.1):
    """
    从 COM 线段列表 lines 中移除那些其两个顶点都在 LB_v 中的线段。
    
    参数：
      lines: COM 线段对象列表，每个对象要求具有 StartPoint 与 EndPoint 属性，
             其值为形如 (x, y, z) 的可迭代对象。
      LB_v: 顶点列表，每个顶点为 (x, y, z) 的元组。
      tol: 判断两顶点近似相等的容差值（默认 0.1）。
    
    返回：
      返回移除满足条件（即两端点都在 LB_v 中）的线段后剩余的线段列表。
    """

    def same_point(pt1, pt2, tol=tol):
        # 只比较 x, y, z 坐标的差值，判断两点是否近似相等
        return abs(pt1[0] - pt2[0]) <= tol and abs(pt1[1] - pt2[1]) <= tol and abs(pt1[2] - pt2[2]) <= tol

    remaining_lines = []
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
        except Exception as e:
            print(f"⚠️ 跳过无效线段，原因: {e}")
            continue

        # 判断该线段的起点和终点是否均存在于 LB_v 中
        sp_in = any(same_point(sp, lb, tol) for lb in LB_v)
        ep_in = any(same_point(ep, lb, tol) for lb in LB_v)
        if not (sp_in and ep_in):
            remaining_lines.append(ln)
        else:
            print(f"删除线段 {ln.Handle}，其两个端点均在 LB_v 中：sp={sp}, ep={ep}")
    return remaining_lines

#  主函数
#  (5)
#&&% 获取全部封闭多边形，但不完全

def process_polygons(lines, tol=0.5, max_steps=50, layer_name="测试辅助"):
    """
    递归提取并绘制直线段集 lines 中所有封闭多边形。
    
    流程：
      1. 从 lines 中提取最右下角封闭多边形（调用 find_rightbottom_closed_polygon）。
      2. 对该多边形：
            - 获取最右下角点 p1（调用 find_rightbottom_point）；
            - 将多边形（polygon）加入列表 LB；
            - 调用 analyze_polygon_branches 分析多边形分枝，得到用来判断移除线段的顶点列表 Lv；
            - 调用 remove_lines_in_LBv，将所有两端点均在 Lv 中的直线段从 lines 中移除；
            - 调用 draw_polygon_as_polyline 绘制该多边形，将生成的多段线 COM 对象加入 LBcom；
            - 然后调用 draw_polygon_as_polyline 绘制顶点列表 Lv（蓝色绘制），将生成的多段线 COM 对象加入 LB_yc；
      3. 重复上述过程直至无法提取出封闭多边形；
    
    返回：
      (LB, LBcom, LB_yc)，其中：
         LB: 封闭多边形顶点列表集合（每个为顶点序列）。
         LBcom: 绘制出的多段线 COM 对象列表。
         LB_yc: 绘制蓝色辅助多段线（用于检查移除的线段顶点）的 COM 对象列表。
    """
    LB = []
    LBcom = []
    LB_yc = []  # 存储蓝色辅助多段线 COM 对象，用于检查移除的顶点
    Ly = []  # 用来记录每次移除的直线段
    iteration = 0

    while True:
        iteration += 1
        print(f"\n【迭代 {iteration}】剩余直线段数量 = {len(lines)}")
        p1 = find_rightbottom_point(lines, tol)
        if p1 is None:
            print("未找到最右下角点，结束迭代")
            break
        print(f"当前最右下角点 p1 = {p1}")
        
        polygon = find_rightbottom_closed_polygon(lines, tol=tol, max_steps=max_steps)
        if polygon is None:
            print("无法提取封闭多边形，结束迭代")
            break
        LB.append(polygon)
        print("提取的封闭多边形顶点：")
        for pt in polygon:
            print("  ", pt)
        
        Lv = analyze_polygon_branches(polygon, lines, p1, tol=tol)
        if Lv is None:
            print("分析多分枝失败，结束本次迭代")
        else:
            print("用于移除线段的顶点列表 Lv：")
            for pt in Lv:
                print("  ", pt)
            # 移除 lines 中两端点均在 Lv 中的直线段（调用 remove_lines_in_LBv）
            new_lines = remove_lines_in_LBv(lines, Lv, tol=0.1)
            removed_count = len(lines) - len(new_lines)
            print(f"本次移除直线段数: {removed_count}")
            # 记录被移除的直线段
            for ln in lines:
                if ln not in new_lines:
                    Ly.append(ln)
            lines = new_lines
        
        # 绘制提取出的封闭多边形
        polyline = draw_polygon_as_polyline(polygon, layer_name=layer_name, tol=tol)
        if polyline:
            LBcom.append(polyline)
        
        # 调用 draw_polygon_as_polyline 绘制 Lv（辅助多段线，注意Lv是顶点列表）
        if Lv is not None and len(Lv) > 0:
            poly_blue = draw_polygon_as_polyline(Lv, layer_name=layer_name, tol=tol)
            if poly_blue:
                LB_yc.append(poly_blue)
        
        if len(lines) < 3:
            print("剩余直线段不足以构成封闭多边形，退出")
            break

    return LB, LBcom, LB_yc



def extract_polygon_from_lines(lines, tol=0.5):
    """
    将表示封闭多边形边缘的线段（COM对象列表）转换为顶点列表（按顺序排列），
    消除中间相邻的重复顶点。若能成功构成闭合多边形，则返回顶点列表（闭环不重复），否则返回 None。
    
    参数：
      lines: 直线段对象列表，每个对象要求具有 StartPoint 和 EndPoint 属性
      tol: 判断顶点相等的容差（仅比较 x 和 y 坐标）。
    
    返回：
      顶点列表，如 [p1, p2, ..., pn]，其中 p1 表示多边形的起点（不重复列出闭合顶点）。
    """
    if not lines:
        return None
    
    # 定义同点判断函数
##    def same_point(a, b, tol=tol):
##        return abs(a[0] - b[0]) <= tol and abs(a[1] - b[1]) <= tol
    
    # 拷贝线段列表，方便修改（注意：此处不会复制COM对象，仅复制引用）
    remaining_lines = list(lines)
    
    try:
        start_pt = tuple(remaining_lines[0].StartPoint)
    except Exception as e:
        print("提取第一条线段起点失败：", e)
        return None
    vertices = [start_pt]
    current_pt = start_pt

    # 由于多边形是闭合的，最多尝试 2*len(remaining_lines) 次
    for _ in range(len(remaining_lines) * 2):
        found_edge = False
        for ln in remaining_lines:
            try:
                sp = tuple(ln.StartPoint)
                ep = tuple(ln.EndPoint)
            except Exception:
                continue

            nxt = None
            if same_point(current_pt, sp) and (not same_point(current_pt, ep)):
                nxt = ep
            elif same_point(current_pt, ep) and (not same_point(current_pt, sp)):
                nxt = sp

            if nxt is None:
                continue

            # 一旦找到与当前点相连的边，移除此边
            remaining_lines.remove(ln)
            # 如果得到的 nxt 与起点相同，则说明闭合结束
            if same_point(nxt, start_pt):
                vertices.append(start_pt)
                return deduplicate_vertices(vertices, tol)
            else:
                vertices.append(nxt)
                current_pt = nxt
                found_edge = True
                break
        if not found_edge:
            # 未能找到与当前点相连的边，构成闭合失败
            break
    # 如果循环结束后闭合未完成，则返回 None
    return None



#将多段线列表炸开为线段，返回线段列表



def explode_polylines(LB):
    """
    对多段线列表 LB 中的每一个多段线，调用 .Explode() 方法，
    并将炸开得到的所有线段合并为一个新的线段列表返回。

    参数:
      LB: 多段线 COM 对象列表，每个对象应支持 .Explode() 方法，
          如 pl = polyline_explode_object.Explode() 返回该多段线炸开后的线段集合。
          
    返回:
      一个包含所有炸开后直线段 COM 对象的列表。
    """
    exploded_lines = []
    for pl in LB:
        try:
            # 调用 Explode() 方法，返回一个集合（例如 COM Collection）
            exploded = pl.Explode()
            # 遍历返回的集合，将每根线段添加到列表中
            # 注意：遍历 COM Collection 的方法可能因环境不同而不同，此处假设可以直接迭代
            for ln in exploded:
                exploded_lines.append(ln)
        except Exception as e:
            handle = getattr(pl, 'Handle', '未知')
            print(f"⚠️ 处理多段线 {handle} 时出错: {e}")
    return exploded_lines


#lines1 中那些不在 lines2 中的线段

def subtract_line_sets(lines1, lines2, tol=0.5):
    """
    比较两个线段集合 lines1 和 lines2，返回 lines1 中那些不在 lines2 中的线段。

    参数：
      lines1: 第一组线段（候选集合），每个对象须具有 StartPoint 和 EndPoint 属性。
      lines2: 第二组线段，参照集合。
      tol: 判断点是否相同的容差，默认值为 0.5。
      
    返回：
      lines1 中所有不与 lines2 中任意线段相同的线段构成的列表。
    """
    result = []
    for ln in lines1:
        found = False
        for ln2 in lines2:
            if same_line(ln, ln2, tol):
                found = True
                break
        if not found:
            result.append(ln)
    return result

#  主函数
#  (6)
#&&% 获取全部封闭多边形


def process_final(lines, tol=0.5, max_steps=50, layer_name="测试辅助"):

    print("len(lines):",len(lines))

    L1, L2, L3 = process_polygons(lines, tol=tol, max_steps=max_steps, layer_name=layer_name)

    print(f"process_polygons 完成：\n  L1 数量 = {len(L1)}\n  L2 数量 = {len(L2)}\n  L3 数量 = {len(L3)}")
    
    print(">>> 对 L3 中的多段线执行 Explode() 操作...")

    exploded_lines=explode_polylines(L3)



    shengyu = subtract_line_sets(lines, exploded_lines, tol=tol)

    for x in shengyu:

        print(x.StartPoint,x.EndPoint)

    L_shengyu=extract_polygon_from_lines(shengyu, tol=tol)
   

    ld = draw_polygon_as_polyline(L_shengyu, layer_name=layer_name, tol=tol)

    L2.append(ld)

    L1.append(L_shengyu)
    for x in exploded_lines:
        x.Delete()

    return L1, L2, L3



#&&&% PL复线处理


"""
研究PL复线处理的问题

(1)多段线的基本操作 get_unique_vertices_from_pl_com

(2)PL打印线 generate_name_and_ratio_from_polyline(polyline,A3dy=0)

(3)将正交六边形多段线分成两个矩形区域 split_hexagon_combined(polygon, tol=0.1, simplify_tol=0.5)

(4)获取多段线的上下左右边界的直线段，返回线段端点列表 get_bbox_edge_segments(pl, tol=0.5)

(5)获取多段线的内部的文字 get_texts_in_polyline(com_pl, tol=0.5)

(6)多段线上的均分插入 distribute_points_on_entity(entity, n, block, scale_factor, ys)

(7)返回 pl1 中与 pl2 “共线且有重叠”的区段列表 common_segments_between_polylines(pl1, pl2, tol=0.5)

(8)找到全部“两根多段线耦合成一个矩形”的多段线 two_plines_making_rectangle(pl1, pl2, tol=0.5)

"""
#  主函数
#  (1)
# 多段线的基本操作

#  该函数系列包括如下一些函数
"""

标准顶点坐标列表,是像[(0, 0, 0), (0, 100, 0), (0, 900, 0)]这样的列表

它表示3个连续顶点2根连续线段，公共点坐标不重复

"""
## 0 轻量多段线和一般传统多段线各有用处，后者才能在三维中使用，更广泛




@alias("画轻量多段线")
def draw_lwpolyline(
    coords3d: list[tuple[float, float, float]],
    layer_name: str = "0",
    width: float = 0.0,
    color: int = 256,
    closed: bool = False
):
    """
    根据一组 (x, y, z) 坐标点绘制轻量级多段线（LWPOLYLINE）。

    :param coords3d: 形如 [(x1, y1, z1), (x2, y2, z2), …] 的点列表，
                     仅使用 x,y 坐标，忽略 z。
    :param layer_name: 目标图层名称，不存在则自动创建。
    :param width:      多段线恒宽 (ConstantWidth)。
    :param color:      颜色索引 (AutoCAD Color Index)，256=BYLAYER。
    :param closed:     是否闭合多段线（首尾相连）。

    :return:           新建的轻量级多段线对象 (COM AddLightWeightPolyline)。

    pts = [
        (0.0, 0.0, 0.0),
        (100.0, 0.0, 0.0),
        (100.0, 50.0, 0.0),
        (0.0, 50.0, 0.0),
    ]
    poly = draw_lwpolyline(
        coords3d=pts,
        layer_name="dy_quyu",
        width=0.0,
        color=1,      # 红色
        closed=True
    )
    poly.Coordinates
    (0.0, 0.0, 100.0, 0.0, 100.0, 50.0, 0.0, 50.0)

    len((0.0, 0.0, 100.0, 0.0, 100.0, 50.0, 0.0, 50.0))
    8

    """
    # 1️⃣ 连接 AutoCAD

    # 2️⃣ 确保图层存在
    layers = doc.Layers
    try:
        lyr = layers.Item(layer_name)
    except Exception:
        lyr = layers.Add(layer_name)
    # Optional: 开启图层
    lyr.LayerOn = True

    # 3️⃣ 准备坐标数组：扁平化 x,y
    raw = []
    for x, y, _ in coords3d:
        raw.extend((x, y))
    # 转 COM VARIANT 数组
    arr = win32com.client.VARIANT(
        pythoncom.VT_ARRAY | pythoncom.VT_R8,
        raw
    )

    # 4️⃣ 绘制轻量级多段线

    try:

        pline = mp.AddLightWeightPolyline(arr)
        pline.Layer         = layer_name
        pline.ConstantWidth = width
        pline.color         = color
        pline.Closed        = bool(closed)

        print(f"✅ 已在图层『{layer_name}』绘制多段线，Closed = {closed}")
        return pline
    except Exception as e:
        print("❌ 绘制多段线失败:", e) 

   # 5️⃣ 返回新对象

    return pline



# 1 从com复线获取标准顶点坐标列表

def get_unique_vertices_from_pl_com(pl_com):
    """
    提取多段线的顶点列表，不重复连续线段的公共顶点，返回顶点列表。
    
    参数:
        pl_com: AutoCAD 中的 Polyline COM 对象（AcDbPolyline）
        
    返回:
        顶点列表，每两个点构成一条线段，格式：[ (x1, y1, z1), (x2, y2, z2), ... ]
        
    """
    # 获取二维坐标数据
    coords = pl_com.Coordinates
    vertices = []

    # 以每两个坐标为一组，构成线段
    for i in range(0, len(coords) - 1, 2):
        x1, y1 = coords[i], coords[i + 1]
        z1 = 0  # 假设z坐标为0，如果需要，可以通过某种方式获取真实z坐标
        if not vertices:
            vertices.append((x1, y1, z1))
        else:
            # 如果当前点与上一个点不重复，添加到顶点列表
            if (x1, y1, z1) != vertices[-1]:
                vertices.append((x1, y1, z1))
    
    # 添加最后一个点，避免遗漏
    last_x, last_y = coords[-2], coords[-1]
    z_last = 0  # 同样，假设最后的z坐标为0
    if (last_x, last_y, z_last) != vertices[-1]:
        vertices.append((last_x, last_y, z_last))
    
    return vertices

# 2 将com线段转成顶点坐标列表，一根线段一个列表

def convert_lines_to_points(segments):
    """
    将com线段列表转换为顶点列表，每条线段的两个端点作为一个独立的列表。

    参数:
      segments: 线段对象列表，每个元素具有 StartPoint 和 EndPoint 属性。

    返回:
      包含多个线段顶点的列表，每个线段是一个包含两个端点坐标的列表。
    """
    points_list = []

    for segment in segments:
        # 提取线段的两个端点
        start_point = tuple(segment.StartPoint)
        end_point = tuple(segment.EndPoint)

        # 将线段的两个端点存入一个列表，添加到结果列表
        points_list.append([start_point, end_point])

    return points_list

# 3 合并顶点列表表示的连续线段，允许多根断开的连续线段

def merge_segments_new(LB, tol=0.5):
    """
    使用convert_lines_to_points 将线段实体转成顶点列表表达式后就可以使用此命令

    不断合并连接，能连接的都会连接   

    """
    def same(p, q):
        return abs(p[0]-q[0]) <= tol and abs(p[1]-q[1]) <= tol

    # 1) 为端点做哈希 — 用 round() 把 tol 纳入 key，避免浮点字典键难比较
    def key(pt):
        return (round(pt[0]/tol), round(pt[1]/tol))   # 只 hash XY

    buckets = defaultdict(list)   # key(pt)  ->  [(seg_index, dir), ...]
    for idx, seg in enumerate(LB):
        a, b = seg[0], seg[-1]
        buckets[key(a)].append((idx, +1))   #  +1 表示 seg[0] 方向
        buckets[key(b)].append((idx, -1))   #  -1 表示 seg[-1] 方向

    used = [False]*len(LB)
    sequences = []

    while True:
        # 找到尚未使用的第一条线段
        try:
            seed_idx = next(i for i,u in enumerate(used) if not u)
        except StopIteration:
            break                           # 全部用完
        used[seed_idx] = True
        seq = deque(LB[seed_idx])           # 双端队列便于首尾增长

        # 函数: 把可接的线段拼到 deque 的一头
        def grow(at_tail: bool):
            while True:
                end_pt = seq[-1] if at_tail else seq[0]
                bucket = buckets[key(end_pt)]
                # 移除已用完的骨牌
                bucket[:] = [pair for pair in bucket if not used[pair[0]]]
                if not bucket:              # 再也接不上
                    break
                idx, direction = bucket.pop()
                used[idx] = True
                seg = LB[idx]
                # 根据 direction 决定正向还是反向加入
                if direction == +1:         # bucket 点是 seg[0]
                    add = seg[1:]           # 去掉公共点再拼
                else:                       # bucket 点是 seg[-1]
                    add = seg[-2::-1]       # 反向、去掉公共点
                if at_tail:
                    seq.extend(add)
                else:
                    seq.extendleft(add[::-1])   # extendleft 要反转

        # 先往尾巴拼，再往头拼（顺序无所谓，都会做到极限）
        grow(True)
        grow(False)
        sequences.append(list(seq))

    return sequences

# 4 绘制连续PL多段线，断开的PL多段线要分开绘制

def draw_polyline(vertices,
                  layer_name="测试辅助",
                  tol=0.5,
                  width=20,
                  color=1):
    """
    复线和多边形都应该按标准顶点坐标列表表达
    根据顶点序列 vertices 在当前 AutoCAD 文档 (全局变量 doc) 的 ModelSpace
    绘制多段线 (PLine)。

    参数
    ----
    vertices : list[tuple]
        仅支持顶点列表形式：[(x, y, z), (x, y, z), ...]。
    layer_name : str
        目标图层名；不存在则自动创建。
    tol : float
        判断首尾是否闭合的容差(只比较 X / Y)。
    width : float
        ConstantWidth 设置。
    color : int
        AutoCAD 颜色号。

    返回
    ----
    acad_polyline : PLine COM 对象 | None
    """

    # ----------------- 内部工具 -----------------
##    def same_point(p1, p2, _tol=tol):
##        """只比较 x、y 坐标的近似相等"""
##        return abs(p1[0] - p2[0]) <= _tol and abs(p1[1] - p2[1]) <= _tol
    # --------------------------------------------

    # ---------- 输入校验 ----------
    if not vertices or not isinstance(vertices, (list, tuple)):
        print("❌ 请输入有效的顶点列表")
        return None
    if not isinstance(vertices[0], (list, tuple)) or len(vertices[0]) < 3:
        print("❌ 顶点格式错误，应为 (x, y, z)")
        return None

    # --------- 处理闭合性 ---------
    is_closed = same_point(vertices[0], vertices[-1])

    # --------- 打印调试信息 --------
    print("调试：绘制多段线的顶点序列")
    for idx, pt in enumerate(vertices):
        print(f"  {idx}: {pt}")

    # --------- 生成 SAFEARRAY ------
    flat = []
    for x, y, z in vertices:
        flat.extend([x, y, z])
    coords_var = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, tuple(flat))

    # ---------- 确保图层存在 -------
    try:
        _ = doc.Layers.Item(layer_name)
    except Exception:
        doc.Layers.Add(layer_name)

    # ---------- 绘制多段线 ----------
    try:
        pl = doc.ModelSpace.AddPolyline(coords_var)
        pl.Closed = is_closed
        pl.Layer = layer_name
        pl.Color = color
        pl.ConstantWidth = width
        pl.Update()
        doc.Regen(0)
        print(f"✅ 已在图层『{layer_name}』绘制多段线，Closed = {is_closed}")
        return pl
    except Exception as e:
        print("❌ 绘制多段线失败:", e)
        return None



# 5 获取多段线后的线段列表，原来的多段线仍然存在


"""
pl=  Pl_obj.Explode()

"""



# 6 将多条直线段（允许不连续）连接成PL复线

def lines_to_polylines(Lc,
                       tol=0.5,
                       layer_name="某某图层",
                       width=20,
                       color=1):
    """
    将若干直线段 (COM Line 对象) 连续合并成多段线：
      1. 直线段 → 顶点对列表  (convert_lines_to_points)
      2. 连续段合并         (merge_segments_new)
      3. 生成多段线         (draw_polyline)
      4. 删除原直线段

    参数
    ----
    Lc : list[Line COM]
        要合并的 AcadLine 对象列表。
    tol : float
        顶点判同容差。
    layer_name / width / color
        传递给 draw_polyline 的控制参数。

    返回
    ----
    PLs : list[PLine COM]
        新生成的多段线对象列表。
    """

    # ---------- 0. 边界检查 ----------
    if not Lc:
        print("❌ 输入线段列表为空")
        return []

    # ---------- 1. 线段 → 顶点对 --------
    # convert_lines_to_points 应返回形如 [[p1,p2],[p3,p4] ...]
    LB = convert_lines_to_points(Lc)

    # ---------- 2. 合并连续顶点 --------
    # merge_segments_new 会把 LB 合并成若干连续顶点序列
    LK = merge_segments_new(LB, tol=tol)

    # ---------- 3. 生成多段线 ----------
    PLs = []
    for verts in LK:
        pl = draw_polyline(verts,
                           layer_name=layer_name,
                           tol=tol,
                           width=width,
                           color=color)
        if pl:
            PLs.append(pl)

    # ---------- 4. 删除原直线 ----------
    for ln in Lc:
        try:
            ln.Delete()
        except Exception:
            pass          # 若已被删除或无效则忽略

    print(f"✅ 已生成 {len(PLs)} 条多段线，并删除 {len(Lc)} 条原直线")
    return PLs





# 7 找到多段线的最左下角的点

def find_min_point(obj):
    """
    获取任意对象的左下角坐标（通过其外包盒）。

    :param obj: 支持 GetBoundingBox() 方法的 COM 对象（如多段线、块参照等）。
    :return:    (min_x, min_y) 元组，表示对象外包盒的左下角坐标
    """
    try:
        ll_point, _ = obj.GetBoundingBox()
        min_x, min_y, _ = ll_point
        return min_x, min_y
    except Exception as e:
        print(f"❌ 获取外包盒失败: {e}")
        return None, None

# 8 找到多段线的最右上角的点

def find_max_point(obj):
    """
    获取任意对象的右上角坐标（通过其外包盒）。

    :param obj: 支持 GetBoundingBox() 方法的 COM 对象（如多段线、块参照等）。
    :return:    (max_x, max_y) 元组，表示对象外包盒的右上角坐标
    """
    try:
        _, ur_point = obj.GetBoundingBox()
        max_x, max_y, _ = ur_point
        return max_x, max_y
    except Exception as e:
        print(f"❌ 获取外包盒失败: {e}")
        return None, None


def distance(point1, point2):
    """计算两点之间的距离"""
    return ((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)**0.5

# 9 删除多段线列表中重复的多段线

def remove_duplicate_polylines(LB_1):
    """处理多段线列表，删除重复的多段线，并将处理后的多段线添加到新列表中。"""
    LB_2 = []

    for XX in LB_1:

        biaoji=0
        try:
            XX.color = 6  # 尝试设置颜色，如果出错则跳过
        except :
            biaoji=1

        if biaoji == 0:
           
            # 检查是否有重复的多段线
            for YY in LB_1:

                try:
                    
                    if XX != YY:
                        
                        min_distance = distance(find_min_point(XX), find_min_point(YY))
                        
                        max_distance = distance(find_max_point(XX), find_max_point(YY))

                        if min_distance < 10 and max_distance < 10:
                            
                            YY.Delete()  # 尝试删除重复的多段线
                           
                except:

                    pass
           
            LB_2.append(XX)

    return LB_2

# 10 定义矩形

def define_rectangle_by_diagonal(p1, p2):
    """
    使用两个对角顶点定义矩形，矩形的边与坐标轴平行或垂直。
    p1, p2: 分别为矩形的一对对角顶点，格式为 (x, y)。
    返回矩形的四个顶点、长和宽。
    """
    x1, y1 = p1
    x2, y2 = p2

    # 计算两个边长
    length_x = abs(x2 - x1)
    length_y = abs(y2 - y1)

    # 确定长和宽
    length = max(length_x, length_y)
    width = min(length_x, length_y)

    # 确定矩形的四个顶点
    rectangle_points = [(x1, y1), (x1, y2), (x2, y2), (x2, y1)]

    return rectangle_points, length, width

def define_rectangle_by_diagonal_x(p1, p2):
    """
    使用两个对角顶点定义矩形，矩形的边与坐标轴平行或垂直。
    p1, p2: 分别为矩形的一对对角顶点，格式为 (x, y)。
    返回矩形的四个顶点、长和宽。
    """
    x1, y1 = p1
    x2, y2 = p2

    # 计算两个边长
    length_x = abs(x2 - x1)
    length_y = abs(y2 - y1)

    # 确定长和宽
    length = max(length_x, length_y)
    width = min(length_x, length_y)

    # 确定矩形的四个顶点
    rectangle_points = [x1, y1, x2, y1, x2, y2, x1, y2]

    return rectangle_points




# 11 矩形框的扩张

def expand_rectangle(p1, p2, offset=130):
    """
    给定矩形框的两个对角点（p1 和 p2），
    返回在四个方向扩展 offset 后的新对角点 P1 和 P2。
    """

    x1, y1, z1 = p1
    x2, y2, z2 = p2

    # 确保 p1 是左下角，p2 是右上角（即使输入反了）
    x_min, x_max = sorted([x1, x2])
    y_min, y_max = sorted([y1, y2])
    z = z1  # z 坐标保持一致

    # 四向扩展 offset，构造新的矩形框
    P1 = (x_min - offset, y_min - offset, z)
    P2 = (x_max + offset, y_max + offset, z)

    return P1, P2


# 12 矩形标准化

def parse_rectangle_points(*args):
    """
    接收多种坐标格式输入，统一解析为矩形四角点：
    返回：
        (左下, 右上, 左上, 右下)，每个为三元组 (x, y, z)
    
    合法输入形式：
        - (x1, y1, z1), (x2, y2, z2)
        - [(x1, y1, z1), (x2, y2, z2)]
        - (x1, y1, x2, y2)
        - (x1, y1, 0, x2, y2, 0)
    """
    try:
        # 解包列表输入
        if len(args) == 1 and isinstance(args[0], list):
            args = args[0]

        # 标准化为两个三维点
        if len(args) == 2 and all(isinstance(pt, (tuple, list)) and len(pt) == 3 for pt in args):
            p1, p2 = args
        elif len(args) == 4:
            x1, y1, x2, y2 = args
            p1, p2 = (x1, y1, 0), (x2, y2, 0)
        elif len(args) == 6:
            x1, y1, z1, x2, y2, z2 = args
            p1, p2 = (x1, y1, z1), (x2, y2, z2)
        else:
            raise ValueError("输入格式不合法")

        # 解析 min/max 点坐标
        x_min = min(p1[0], p2[0])
        x_max = max(p1[0], p2[0])
        y_min = min(p1[1], p2[1])
        y_max = max(p1[1], p2[1])
        z = p1[2] if len(p1) > 2 else 0

        # 四角坐标
        left_bottom = (x_min, y_min, z)
        right_top = (x_max, y_max, z)
        left_top = (x_min, y_max, z)
        right_bottom = (x_max, y_min, z)

        return left_bottom, right_top, left_top, right_bottom

    except Exception as e:
        print(f"❌ 解析矩形点失败: {e}")
        return None





#  主函数
#  (2)
#  PL打印线

#  该函数系列包括如下一些函数


#&&%  1 判断多段线是否合乎打印要求的多段线返回其图号和比例或0

def generate_name_and_ratio_from_polyline(comobj,A3dy=0,Fandy=("ISO_A3_(420.00_x_297.00_MM)","0:0","A3"),tol=10):



    """
    函数已经修改为针对任意对象，按其外包盒来分析

    当一般选择为空时自动作比例选择。不再重新编制函数和修改该函数，直接取值A3dy=1即可


    1)在使用该函数之前，应使用多段线快速选择并逐步去掉多余的PL线

    2)这一段表达这样的思想：对于那些不是很接近但误差也不是很大的情况，我们会使用已有图框中最接近的一个去作为结论，但这个可能也是错误的
    elif diff1 < 200 and diff2 < 200:
                    total_diff = diff1 + diff2  # 计算总差值
                    if total_diff < closest_diff:
                        closest_diff = total_diff
                        index_pl = i


    index_pl和closest_diff都在不断地更新直到得到最接近的.当48基元有靠近对象时，会记录下靠近痕迹，但不会影响取值。仅当其未匹配偏差到更准确的基元时菜有必要打印这个参考消息。

    另外，包含着5点合法打印线的可能。因此逻辑仍待优化，但从根本上没问题了。20250507

    3)define_rectangle_by_diagonal所定义的length就是长边，不是x方向的距离

    4)Fandy=("ISO_A3_(420.00_x_297.00_MM)","0:0","A3"),("ISO_A2_(594.00_x_420.00_MM)","0:0","A2"),("ISO_A1_(841.00_x_594.00_MM)","0:0","A1"),("ISO_A0_(1189.00_x_841.00_MM)","0:0","A0")
        
    """
    #函数返回打印多段线的图幅，比例，图号信息。对于不合法的PL线返回0值。

    # 1） 定义基本列表
    
    LB_dayingkuang = [
            (118900,84100,100),(178350,126150,150),(59450,42050,50),(29725,21025,25),
            (133763,84100,100),(200644.5,126150,150),(66881.5,42050,50),(33440.75,21025,25),
            (148625,84100,100),(222937.5,126150,150),(74312.5,42050,50),(37156.25,21025,25),
            (84100,59400,100),(126150,89100,150),(42050,29700,50),(21025,14850,25),
            (105125,59400,100),(157687.5,89100,150),(52562.5,29700,50),(26281.25,14850,25),            
            (126150,59400,100),(189225,89100,150),(63075,29700,50),(31537.5,14850,25),            
            (147175,59400,100),(220762.5,89100,150),(73587.5,29700,50),(36793.75,14850,25),
            (59400,42000,100),(89100,63000,150),(29700,21000,50),(14850,10500,25),            
            (74250,42000,100),(111375,63000,150),(37125,21000,50),(18562.5,10500,25),            
            (89100,42000,100),(133650,63000,150),(44550,21000,50),(22275,10500,25),            
            (103950,42000,100),(155925,63000,150),(51975,21000,50),(25987.5,10500,25),            
            (42000,29700,100),(63000,44550,150),(21000,14850,50),(10500,7425,25)            
        ]
    
    drawing_map_ml = [("A0","1:100"),("A0","1:150"),("A0","1:50"),("A0","1:25"),
                      ("A0+1/8","1:100"),("A0+1/8","1:150"),("A0+1/8","1:50"),("A0+1/8","1:25"),
                      ("A0+1/4","1:100"),("A0+1/4","1:150"),("A0+1/4","1:50"),("A0+1/4","1:25"),
                      ("A1","1:100"),("A1","1:150"),("A1","1:50"),("A1","1:25"),
                      ("A1+1/4","1:100"),("A1+1/4","1:150"),("A1+1/4","1:50"),("A1+1/4","1:25"),
                      ("A1+1/2","1:100"),("A1+1/4","1:150"),("A1+1/2","1:50"),("A1+1/2","1:25"),
                      ("A1+3/4","1:100"),("A1+3/4","1:150"),("A1+3/4","1:50"),("A1+3/4","1:25"),
                      ("A2","1:100"),("A2","1:150"),("A2","1:50"),("A2","1:25"),
                      ("A2+1/4","1:100"),("A2+1/4","1:150"),("A2+1/4","1:50"),("A2+1/4","1:25"),
                      ("A2+1/2","1:100"),("A2+1/2","1:150"),("A2+1/2","1:50"),("A2+1/2","1:25"),
                      ("A2+3/4","1:100"),("A2+3/4","1:150"),("A2+3/4","1:50"),("A2+3/4","1:25"),
                      ("A3","1:100"),("A3","1:150"),("A3","1:50"),("A3","1:25")
        ]                      
                      
    drawing_map = ["ISO_A0_(1189.00_x_841.00_MM)","ISO_A0_(1189.00_x_841.00_MM)","ISO_A0_(1189.00_x_841.00_MM)","ISO_A0_(1189.00_x_841.00_MM)",
                   "UserDefinedMetric (1337.63 x 841.00毫米)", "UserDefinedMetric (1337.63 x 841.00毫米)","UserDefinedMetric (1337.63 x 841.00毫米)","UserDefinedMetric (1337.63 x 841.00毫米)",
                   "UserDefinedMetric (1486.25 x 841.00毫米)","UserDefinedMetric (1486.25 x 841.00毫米)","UserDefinedMetric (1486.25 x 841.00毫米)","UserDefinedMetric (1486.25 x 841.00毫米)",
                   "ISO_A1_(841.00_x_594.00_MM)","ISO_A1_(841.00_x_594.00_MM)","ISO_A1_(841.00_x_594.00_MM)","ISO_A1_(841.00_x_594.00_MM)",
                   "UserDefinedMetric (1051.25 x 594.00毫米)","UserDefinedMetric (1051.25 x 594.00毫米)","UserDefinedMetric (1051.25 x 594.00毫米)","UserDefinedMetric (1051.25 x 594.00毫米)",
                   "UserDefinedMetric (1261.50 x 594.00毫米)","UserDefinedMetric (1261.50 x 594.00毫米)","UserDefinedMetric (1261.50 x 594.00毫米)","UserDefinedMetric (1261.50 x 594.00毫米)",
                   "UserDefinedMetric (1471.75 x 594.00毫米)","UserDefinedMetric (1471.75 x 594.00毫米)","UserDefinedMetric (1471.75 x 594.00毫米)","UserDefinedMetric (1471.75 x 594.00毫米)",
                   "ISO_A2_(594.00_x_420.00_MM)","ISO_A2_(594.00_x_420.00_MM)","ISO_A2_(594.00_x_420.00_MM)","ISO_A2_(594.00_x_420.00_MM)",
                   "UserDefinedMetric (742.50 x 420.00毫米)","UserDefinedMetric (742.50 x 420.00毫米)","UserDefinedMetric (742.50 x 420.00毫米)","UserDefinedMetric (742.50 x 420.00毫米)",                   
                   "UserDefinedMetric (891.00 x 420.00毫米)","UserDefinedMetric (891.00 x 420.00毫米)","UserDefinedMetric (891.00 x 420.00毫米)","UserDefinedMetric (891.00 x 420.00毫米)",
                   "UserDefinedMetric (1039.50 x 420.00毫米)","UserDefinedMetric (1039.50 x 420.00毫米)","UserDefinedMetric (1039.50 x 420.00毫米)","UserDefinedMetric (1039.50 x 420.00毫米)",
                   "ISO_A3_(420.00_x_297.00_MM)","ISO_A3_(420.00_x_297.00_MM)","ISO_A3_(420.00_x_297.00_MM)","ISO_A3_(420.00_x_297.00_MM)"]
                      
    # 2） 确定多段线的长与宽（长度值大的为length）
    PL_min = find_min_point(comobj)
    
    PL_max = find_max_point(comobj)
    
    _, length, width = define_rectangle_by_diagonal(PL_min, PL_max)


    # 3） 确定序号

    index_pl = ""
    closest_diff = float('inf')  # 初始化最接近的差值为无穷大
    if A3dy == 0:

        print("length,width",length,width)
        for i in range(len(LB_dayingkuang)):
            obj = LB_dayingkuang[i]
            obj_length, obj_width, _ = obj
            diff1 = abs(length - obj_length)
            diff2 = abs(width - obj_width)

            if diff1 < tol and diff2 < tol:
                index_pl = i
                break  # 如果找到了非常接近的，就不再继续查找


            

            elif diff1 < 200 and diff2 < 200:

                print("偏差值在10--200之间,请注意该尺寸")
                total_diff = diff1 + diff2  # 计算总差值
                if total_diff < closest_diff:
                    closest_diff = total_diff
                    index_pl = i

        # 退出循环后统一判断
        if index_pl != "":

            print("位置：",comobj.GetBoundingBox()[0],comobj.GetBoundingBox()[1])

            print("数据：",drawing_map[index_pl], drawing_map_ml[index_pl][1], drawing_map_ml[index_pl][0])
            
            return drawing_map[index_pl], drawing_map_ml[index_pl][1], drawing_map_ml[index_pl][0]
        else:


            print("位置：",comobj.GetBoundingBox()[0],comobj.GetBoundingBox()[1])

            #对原0值对象加广义判断

            fanhui = get_print_template_info(comobj, tol=tol)

            return fanhui 

    # 4） 返回值

    elif A3dy == 1:

        print("length,width",length,width)
        
        if abs(width/length - 0.707) <= 0.01:

            print("位置：",comobj.GetBoundingBox()[0],comobj.GetBoundingBox()[1])

            return Fandy[0],Fandy[1],Fandy[2]

        else:            

            print("位置：",comobj.GetBoundingBox()[0],comobj.GetBoundingBox()[1])

            return 0

    else:

        print("参数输入错误")

        return 0
##新的四元信息分析检测函数

def generate_name_and_ratio_from_com(
    comobj,
    A3dy=0,
    Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3",0),
    tol=10
):
    """
    函数已经修改为针对任意对象，按其外包盒来分析。
    返回值由原来的三元组 (图幅, 比例, 图号) 扩展为四元组 (图幅, 比例, 图号, 竖向标志)：
      - 竖向标志 = 1 表示外包盒竖向 (height > width)
      - 否则返回 0（横向或正方形）

    参数：
      comobj -- 任意支持 GetBoundingBox() 的多段线 COM 对象
      A3dy   -- 当 A3dy=1 时，直接按照 Fandy 返回。否则使用 LB_dayingkuang 进行匹配。
      Fandy  -- 默认 ("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3",0)，规模 A3 情况下返回的信息
      tol    -- 与标准尺寸比较的容差，默认为 10

    返回：
      (图幅, 比例, 图号, 竖向标志) 或者 0
    """
    # 1）定义基本尺寸库
    LB_dayingkuang = [
        (118900, 84100, 100),   (178350, 126150, 150),  (59450, 42050, 50),   (29725, 21025, 25),
        (133763, 84100, 100),   (200644.5, 126150, 150), (66881.5, 42050, 50), (33440.75, 21025, 25),
        (148625, 84100, 100),   (222937.5, 126150, 150),(74312.5, 42050, 50), (37156.25, 21025, 25),
        (84100, 59400, 100),    (126150, 89100, 150),   (42050, 29700, 50),   (21025, 14850, 25),
        (105125, 59400, 100),   (157687.5, 89100, 150), (52562.5, 29700, 50), (26281.25, 14850, 25),
        (126150, 59400, 100),   (189225, 89100, 150),   (63075, 29700, 50),   (31537.5, 14850, 25),
        (147175, 59400, 100),   (220762.5, 89100, 150), (73587.5, 29700, 50), (36793.75, 14850, 25),
        (59400, 42000, 100),    (89100, 63000, 150),    (29700, 21000, 50),   (14850, 10500, 25),
        (74250, 42000, 100),    (111375, 63000, 150),   (37125, 21000, 50),   (18562.5, 10500, 25),
        (89100, 42000, 100),    (133650, 63000, 150),   (44550, 21000, 50),   (22275, 10500, 25),
        (103950, 42000, 100),   (155925, 63000, 150),   (51975, 21000, 50),   (25987.5, 10500, 25),
        (42000, 29700, 100),    (63000, 44550, 150),    (21000, 14850, 50),   (10500, 7425, 25)
    ]

    drawing_map_ml = [
        ("A0", "1:100"), ("A0", "1:150"), ("A0", "1:50"),  ("A0", "1:25"),
        ("A0+1/8", "1:100"), ("A0+1/8", "1:150"), ("A0+1/8", "1:50"),  ("A0+1/8", "1:25"),
        ("A0+1/4", "1:100"), ("A0+1/4", "1:150"), ("A0+1/4", "1:50"),  ("A0+1/4", "1:25"),
        ("A1", "1:100"), ("A1", "1:150"), ("A1", "1:50"), ("A1", "1:25"),
        ("A1+1/4", "1:100"), ("A1+1/4", "1:150"), ("A1+1/4", "1:50"),  ("A1+1/4", "1:25"),
        ("A1+1/2", "1:100"), ("A1+1/2", "1:150"), ("A1+1/2", "1:50"),  ("A1+1/2", "1:25"),
        ("A1+3/4", "1:100"), ("A1+3/4", "1:150"), ("A1+3/4", "1:50"),  ("A1+3/4", "1:25"),
        ("A2", "1:100"), ("A2", "1:150"), ("A2", "1:50"),  ("A2", "1:25"),
        ("A2+1/4", "1:100"), ("A2+1/4", "1:150"), ("A2+1/4", "1:50"),  ("A2+1/4", "1:25"),
        ("A2+1/2", "1:100"), ("A2+1/2", "1:150"), ("A2+1/2", "1:50"),  ("A2+1/2", "1:25"),
        ("A2+3/4", "1:100"), ("A2+3/4", "1:150"), ("A2+3/4", "1:50"),  ("A2+3/4", "1:25"),
        ("A3", "1:100"), ("A3", "1:150"), ("A3", "1:50"),  ("A3", "1:25")
    ]

    drawing_map = [
        "ISO_A0_(1189.00_x_841.00_MM)", "ISO_A0_(1189.00_x_841.00_MM)",
        "ISO_A0_(1189.00_x_841.00_MM)", "ISO_A0_(1189.00_x_841.00_MM)",
        "UserDefinedMetric (1337.63 x 841.00毫米)", "UserDefinedMetric (1337.63 x 841.00毫米)",
        "UserDefinedMetric (1337.63 x 841.00毫米)", "UserDefinedMetric (1337.63 x 841.00毫米)",
        "UserDefinedMetric (1486.25 x 841.00毫米)", "UserDefinedMetric (1486.25 x 841.00毫米)",
        "UserDefinedMetric (1486.25 x 841.00毫米)", "UserDefinedMetric (1486.25 x 841.00毫米)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "UserDefinedMetric (1051.25 x 594.00毫米)", "UserDefinedMetric (1051.25 x 594.00毫米)",
        "UserDefinedMetric (1051.25 x 594.00毫米)", "UserDefinedMetric (1051.25 x 594.00毫米)",
        "UserDefinedMetric (1261.50 x 594.00毫米)", "UserDefinedMetric (1261.50 x 594.00毫米)",
        "UserDefinedMetric (1261.50 x 594.00毫米)", "UserDefinedMetric (1261.50 x 594.00毫米)",
        "UserDefinedMetric (1471.75 x 594.00毫米)", "UserDefinedMetric (1471.75 x 594.00毫米)",
        "UserDefinedMetric (1471.75 x 594.00毫米)", "UserDefinedMetric (1471.75 x 594.00毫米)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "UserDefinedMetric (742.50 x 420.00毫米)", "UserDefinedMetric (742.50 x 420.00毫米)",
        "UserDefinedMetric (742.50 x 420.00毫米)", "UserDefinedMetric (742.50 x 420.00毫米)",
        "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)",
        "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)",
        "UserDefinedMetric (1039.50 x 420.00毫米)", "UserDefinedMetric (1039.50 x 420.00毫米)",
        "UserDefinedMetric (1039.50 x 420.00毫米)", "UserDefinedMetric (1039.50 x 420.00毫米)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)"
    ]

    # 2）计算外包盒的长、宽，以及横竖信息
    PL_min = find_min_point(comobj)
    PL_max = find_max_point(comobj)
    _, length, width = define_rectangle_by_diagonal(PL_min, PL_max)

    # 计算 dx, dy 以判断竖向或横向
    dx = abs(PL_max[0] - PL_min[0])
    dy = abs(PL_max[1] - PL_min[1])
    orientation_flag = 1 if dy > dx else 0

    # 3）确定匹配序号 index_pl
    index_pl = ""
    closest_diff = float('inf')

    if A3dy == 0:
        print("length, width =", length, width)
        for i, (obj_length, obj_width, _) in enumerate(LB_dayingkuang):
            diff1 = abs(length - obj_length)
            diff2 = abs(width - obj_width)

            # 完全匹配
            if diff1 < tol and diff2 < tol:
                index_pl = i
                break

            # 近似匹配的候选
            elif diff1 < 200 and diff2 < 200:
                total_diff = diff1 + diff2
                if total_diff < closest_diff:
                    closest_diff = total_diff
                    index_pl = i

        # 判断并返回
        if index_pl != "":
            print("位置：", comobj.GetBoundingBox()[0], comobj.GetBoundingBox()[1])
            return (
                drawing_map[index_pl],
                drawing_map_ml[index_pl][1],
                drawing_map_ml[index_pl][0],
                orientation_flag
            )
        else:
            print("位置：", comobj.GetBoundingBox()[0], comobj.GetBoundingBox()[1])
            fanhui = get_print_template_info(comobj, tol=tol)
            # 如果 fanhui 是三元组，则拼接 orientation_flag；否则返回 0
            if isinstance(fanhui, tuple) and len(fanhui) == 3:
                return (*fanhui, orientation_flag)
            return 0

    # 4）当 A3dy == 1 时，直接根据长宽比返回 Fandy 并附加 orientation_flag
    elif A3dy == 1:
        print("length, width =", length, width)
        if abs(width / length - 0.707) <= 0.01:
            print("位置：", comobj.GetBoundingBox()[0], comobj.GetBoundingBox()[1])
            return (*Fandy, orientation_flag)
        else:
            print("位置：", comobj.GetBoundingBox()[0], comobj.GetBoundingBox()[1])
            return 0

    else:
        print("参数输入错误")
        return 0


##20250711修改
def generate_name_and_ratio_from_com(
    comobj,
    A3dy=0,
    Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3", 0),
    tol=10
):
    """
    返回 (图幅, 比例, 图号, 竖向标志)。未命中时返回 0。

    ① 精确命中 → 原样返回  
    ② 命中 “×1.2”  → 返回原标准数据，并把对象改成蓝色  
    ③ 近似命中     → 返回原标准数据，并把对象改成红色 + 宽度 (200 / 2)

    对拟合标准框线，加了红色警告和加粗提示，对放大1.2倍的打印框线加了蓝色提示

    """

    # ————————————— 1. 基础数据区 —————————————
    LB_dayingkuang = [
        (118900, 84100, 100),   (178350, 126150, 150),  (59450, 42050, 50),   (29725, 21025, 25),
        (133763, 84100, 100),   (200644.5, 126150, 150), (66881.5, 42050, 50), (33440.75, 21025, 25),
        (148625, 84100, 100),   (222937.5, 126150, 150),(74312.5, 42050, 50), (37156.25, 21025, 25),
        (84100, 59400, 100),    (126150, 89100, 150),   (42050, 29700, 50),   (21025, 14850, 25),
        (105125, 59400, 100),   (157687.5, 89100, 150), (52562.5, 29700, 50), (26281.25, 14850, 25),
        (126150, 59400, 100),   (189225, 89100, 150),   (63075, 29700, 50),   (31537.5, 14850, 25),
        (147175, 59400, 100),   (220762.5, 89100, 150), (73587.5, 29700, 50), (36793.75, 14850, 25),
        (59400, 42000, 100),    (89100, 63000, 150),    (29700, 21000, 50),   (14850, 10500, 25),
        (74250, 42000, 100),    (111375, 63000, 150),   (37125, 21000, 50),   (18562.5, 10500, 25),
        (89100, 42000, 100),    (133650, 63000, 150),   (44550, 21000, 50),   (22275, 10500, 25),
        (103950, 42000, 100),   (155925, 63000, 150),   (51975, 21000, 50),   (25987.5, 10500, 25),
        (42000, 29700, 100),    (63000, 44550, 150),    (21000, 14850, 50),   (10500, 7425, 25)
    ]

    drawing_map_ml = [
        ("A0", "1:100"), ("A0", "1:150"), ("A0", "1:50"),  ("A0", "1:25"),
        ("A0+1/8", "1:100"), ("A0+1/8", "1:150"), ("A0+1/8", "1:50"),  ("A0+1/8", "1:25"),
        ("A0+1/4", "1:100"), ("A0+1/4", "1:150"), ("A0+1/4", "1:50"),  ("A0+1/4", "1:25"),
        ("A1", "1:100"), ("A1", "1:150"), ("A1", "1:50"), ("A1", "1:25"),
        ("A1+1/4", "1:100"), ("A1+1/4", "1:150"), ("A1+1/4", "1:50"),  ("A1+1/4", "1:25"),
        ("A1+1/2", "1:100"), ("A1+1/2", "1:150"), ("A1+1/2", "1:50"),  ("A1+1/2", "1:25"),
        ("A1+3/4", "1:100"), ("A1+3/4", "1:150"), ("A1+3/4", "1:50"),  ("A1+3/4", "1:25"),
        ("A2", "1:100"), ("A2", "1:150"), ("A2", "1:50"),  ("A2", "1:25"),
        ("A2+1/4", "1:100"), ("A2+1/4", "1:150"), ("A2+1/4", "1:50"),  ("A2+1/4", "1:25"),
        ("A2+1/2", "1:100"), ("A2+1/2", "1:150"), ("A2+1/2", "1:50"),  ("A2+1/2", "1:25"),
        ("A2+3/4", "1:100"), ("A2+3/4", "1:150"), ("A2+3/4", "1:50"),  ("A2+3/4", "1:25"),
        ("A3", "1:100"), ("A3", "1:150"), ("A3", "1:50"),  ("A3", "1:25")
    ]

    drawing_map = [
        "ISO_A0_(1189.00_x_841.00_MM)", "ISO_A0_(1189.00_x_841.00_MM)",
        "ISO_A0_(1189.00_x_841.00_MM)", "ISO_A0_(1189.00_x_841.00_MM)",
        "UserDefinedMetric (1337.63 x 841.00毫米)", "UserDefinedMetric (1337.63 x 841.00毫米)",
        "UserDefinedMetric (1337.63 x 841.00毫米)", "UserDefinedMetric (1337.63 x 841.00毫米)",
        "UserDefinedMetric (1486.25 x 841.00毫米)", "UserDefinedMetric (1486.25 x 841.00毫米)",
        "UserDefinedMetric (1486.25 x 841.00毫米)", "UserDefinedMetric (1486.25 x 841.00毫米)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "UserDefinedMetric (1051.25 x 594.00毫米)", "UserDefinedMetric (1051.25 x 594.00毫米)",
        "UserDefinedMetric (1051.25 x 594.00毫米)", "UserDefinedMetric (1051.25 x 594.00毫米)",
        "UserDefinedMetric (1261.50 x 594.00毫米)", "UserDefinedMetric (1261.50 x 594.00毫米)",
        "UserDefinedMetric (1261.50 x 594.00毫米)", "UserDefinedMetric (1261.50 x 594.00毫米)",
        "UserDefinedMetric (1471.75 x 594.00毫米)", "UserDefinedMetric (1471.75 x 594.00毫米)",
        "UserDefinedMetric (1471.75 x 594.00毫米)", "UserDefinedMetric (1471.75 x 594.00毫米)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "UserDefinedMetric (742.50 x 420.00毫米)", "UserDefinedMetric (742.50 x 420.00毫米)",
        "UserDefinedMetric (742.50 x 420.00毫米)", "UserDefinedMetric (742.50 x 420.00毫米)",
        "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)",
        "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)",
        "UserDefinedMetric (1039.50 x 420.00毫米)", "UserDefinedMetric (1039.50 x 420.00毫米)",
        "UserDefinedMetric (1039.50 x 420.00毫米)", "UserDefinedMetric (1039.50 x 420.00毫米)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)"
    ]

    # —————————— 2. 外包盒与朝向 ——————————
    PL_min = find_min_point(comobj)
    PL_max = find_max_point(comobj)
    _, length, width = define_rectangle_by_diagonal(PL_min, PL_max)
    dx, dy = abs(PL_max[0] - PL_min[0]), abs(PL_max[1] - PL_min[1])
    orientation_flag = 1 if dy > dx else 0

    # —————————— 3. 匹配类型判定 ——————————
    best_i, best_type = None, None      # 'exact' | 'scale12' | 'approx'
    best_score = float('inf')           # 仅用于 approx

    if A3dy == 0:
        for i, (std_len, std_wid, _) in enumerate(LB_dayingkuang):
            # 精确
            if abs(length - std_len) < tol and abs(width - std_wid) < tol:
                best_i, best_type = i, 'exact'
                break

            # ×1.2
            if abs(length - 1.2*std_len) < tol and abs(width - 1.2*std_wid) < tol:
                best_i, best_type = i, 'scale12'
                break

            # 近似
            diff = max(abs(length - std_len), abs(width - std_wid))
            if diff < 200 and diff < best_score:
                best_i, best_type = i, 'approx'
                best_score = diff

        # —————————— 4. 命中后的处理 ——————————
        if best_i is not None:
            zhi = (
                drawing_map[best_i],
                drawing_map_ml[best_i][1],
                drawing_map_ml[best_i][0],
                orientation_flag
            )

            if best_type == 'scale12':          # 蓝
                try:
                    comobj.Color = 5
                except Exception:
                    pass

            elif best_type == 'approx':         # 红 + 加粗
                try:
                    comobj.Color = 1
                except Exception:
                    pass
                try:
                    total_len = float(getattr(comobj, "Length", 0))
                except Exception:
                    total_len = 0
                new_wid = 2 if total_len and total_len < 10000 else 200
                try:
                    comobj.ConstantWidth = new_wid
                except Exception:
                    try:
                        comobj.SetWidth(new_wid, new_wid)
                    except Exception:
                        pass

            return zhi

        # —— 未命中：调用外部模板判定 —— 
        fanhui = get_print_template_info(comobj, tol=tol)
        return (*fanhui, orientation_flag) if isinstance(fanhui, tuple) else 0

    # —————————— 5. A3dy == 1 直接用长宽比判定 ——————————
    elif A3dy == 1:
        if abs(width / length - 0.707) <= 0.01:
            return (*Fandy, orientation_flag)
        return 0

    # —————————— 6. 参数错误 ——————————
    else:
        print("参数输入错误")
        return 0





#  2 com复线从上到下从左到右排序


def polyline_sort(polyline_list):
    """对com多段线按照特定规则进行排序"""

    # 存储多段线及其最左下角点
    polylines_with_min_points = [(pl, find_min_point(pl)) for pl in polyline_list]

    # 先按照y值降序排序
    polylines_with_min_points.sort(key=lambda item: -item[1][1])

    # 再对y值差距在1000以内的多段线按照x值升序排序
    i = 0
    while i < len(polylines_with_min_points) - 1:
        j = i + 1
        # 查找所有y值差距在1000以内的多段线
        while j < len(polylines_with_min_points) and abs(polylines_with_min_points[i][1][1] - polylines_with_min_points[j][1][1]) < 1000:
            j += 1
        
        # 如果找到了y值相近的多段线，根据x值进行排序
        if j - i > 1:
            polylines_with_min_points[i:j] = sorted(polylines_with_min_points[i:j], key=lambda item: item[1][0])
        
        i = j

    # 只返回多段线对象
    return [item[0] for item in polylines_with_min_points]




#&&&%  *** 3 将PLcom线列表的坐标信息存储


def plcom_to_coor(plines):
    """
    接受多根轻量级多段线或常规多段线的 COM 对象列表，返回它们的坐标列表及闭合状态。

    :param plines: 可迭代的一组 LWPOLYLINE 或 POLYLINE COM 对象
    :return: 列表，每个元素为 (pts, closed_flag)：
             - pts: 顶点列表 [(x0, y0), (x1, y1), …]
             - closed_flag: 1 表示闭合，0 表示未闭合

    兼容两种情况：
      1. LWPOLYLINE.Coordinates → 偶数长度，例如 [x0,y0, x1,y1, x2,y2, …]
      2. POLYLINE.Coordinates  → 3 的倍数长度，例如 [x0,y0,z0, x1,y1,z1, …]

    如果既不是偶数也不是 3 的倍数，将跳过该条多段线并打印 WARN 提示。
    """
    plines = ensure_list(plines)
    all_info = []

    for pl in plines:
        raw = list(pl.Coordinates)  # 可能是偶数长度 (LWPOLYLINE) 或 3 的倍数长度 (POLYLINE)

        pts = []
        # —— 情况 A：如果长度能被 3 整除，认为是常规 POLYLINE → 每 3 个数一组 (x,y,z)
        if len(raw) % 3 == 0 and len(raw) > 0:
            for i in range(0, len(raw), 3):
                x = raw[i]
                y = raw[i + 1]
                pts.append((x, y))

        # —— 情况 B：否则如果长度能被 2 整除，认为是轻量级 LWPOLYLINE → 每 2 个数一组 (x,y)
        elif len(raw) % 2 == 0 and len(raw) > 0:
            for i in range(0, len(raw), 2):
                x = raw[i]
                y = raw[i + 1]
                pts.append((x, y))

        else:
            # 既不是 2 的倍数也不是 3 的倍数：坐标数据异常，跳过这一条，打印 WARN
            handle = getattr(pl, "Handle", "<unknown>")
            print(f"[WARN] plcom_to_coor：跳过 Handle={handle} 的多段线，"
                  f"Coordinates 长度={len(raw)} 既非 2 的倍数也非 3 的倍数。")
            continue

        # 读取 Closed 属性，True 表示闭合
        closed_flag = 1 if getattr(pl, "Closed", False) else 0

        all_info.append((pts, closed_flag))

    return all_info


# 4 从坐标信息列表返回PLcom线列表

    
def plcoor_to_com(coord_info, layer_name="测试辅助", width=0, color=256):
    """
    在当前 DWG 中根据坐标和封闭标志绘制多条轻量级多段线。

    :param coord_info: 列表，每个元素为 (pts, closed_flag)，
                       pts 为 [(x0,y0),…] 顶点列表，
                       closed_flag 为 1（闭合）或 0（不闭合）。
    :param layer_name: 目标图层名称（不存在则创建），默认 "测试辅助"
    :param width:      多段线宽度，默认 0
    :param color:      颜色索引，默认 256（BYLAYER）
    :return:           绘制的多段线对象列表
    """
    # 1) 连接 AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    ms   = doc.ModelSpace

    # 2) 确保图层存在
    layers = doc.Layers
    try:
        lyr = layers.Item(layer_name)
    except Exception:
        lyr = layers.Add(layer_name)
    lyr.LayerOn = True

    created = []
    for pts, closed_flag in coord_info:
        # 将 pts 展平为 [x0,y0,x1,y1,…]
        raw = []
        for x, y in pts:
            raw.extend((x, y))
        # 转为 COM 数组
        arr = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8,
            raw
        )
        # 添加轻量级多段线
        lw = ms.AddLightWeightPolyline(arr)
        lw.Layer         = layer_name
        lw.ConstantWidth = width
        lw.Color         = color
        lw.Closed        = bool(closed_flag)
        created.append(lw)

    # 可选：缩放到可见范围
    acad.ZoomExtents()
    print(f"✅ 已绘制 {len(created)} 条轻量级多段线到图层 “{layer_name}”")
    return created






# 5 确定多段线打印框是否竖向

def panduan_shuxiangkuang(polyline):

    PL_min = find_min_point(polyline)
    
    PL_max = find_max_point(polyline)

    cha_x  = abs(PL_max[0] - PL_min[0])   

    cha_y  = abs(PL_max[1] - PL_min[1])

    if  cha_y > cha_x:

        return True

    else :

        return False


# 6 统一为A2的图幅打印

def tongyi_tufu(LB,TFname):


    """
    将打印线列表的每根线对应的图纸尺寸统一为一个TFname
        
    """

    LB_xin = []

    for ob in LB:

        LB_xin.append(TFname)#"ISO_A2_(594.00_x_420.00_MM)","ISO_A3_(420.00_x_297.00_MM)"

    return LB_xin



#  主函数
#  (3)
#  将正交六边形多段线分成两个矩形区域

#  该函数系列包括如下一些函数

"""
该函数将正交六边形多段线分成两个矩形区域，只针对标准六点六边的PL多边形
在处理六点六边的正交PL形多边形之前，使用simplify_polygon(polygon, tol=0.5)将伪边点清除
            
"""

# 消除多边形的伪边点

def simplify_polygon(poly, tol=0.5):
    """
    简化多边形顶点列表：如果某顶点 P 与其前后两点共线（在容差 tol 范围内），则将其移除。
    首尾相连处理：第一个点的“前点”是最后一个点，最后一个点的“次点”是第一个点。
    
    参数:
        poly: [(x, y, z), …]  原始顶点列表（可能首尾重复或有多余顶点）
        tol:  共线判断的容差（对应叉积绝对值）
    
    返回:
        简化后的顶点列表（同样保留首尾是否闭合的形式，不会自动去重首尾）
    """
    # 先做一次“去掉首尾重复”以免无限循环
    if len(poly) > 1 and poly[0] == poly[-1]:
        poly = poly[:-1]

    def is_colinear(p_prev, p, p_next):
        # 只比较 x,y，计算 (p - p_prev) × (p_next - p) 的“叉积”
        x1, y1 = p[0] - p_prev[0], p[1] - p_prev[1]
        x2, y2 = p_next[0] - p[0],   p_next[1] - p[1]
        cross = x1 * y2 - y1 * x2
        return abs(cross) <= tol

    # 重复扫一遍能删就删，直到一轮下来没有删除
    changed = True
    while changed and len(poly) >= 3:
        changed = False
        n = len(poly)
        for i in range(n):
            prev_idx = (i - 1) % n
            next_idx = (i + 1) % n
            if is_colinear(poly[prev_idx], poly[i], poly[next_idx]):
                # 删除第 i 个点，重置一轮扫描
                del poly[i]
                changed = True
                break
    return poly



# 1. 标准化多边形顶点列表，去掉相邻和首尾重复点
def normalize_polygon(polygon):
    """
    标准化多边形顶点列表：  
      - 去掉任意相邻重复的点  
      - 如果首尾相同，则去掉末尾那个  
      
    参数:
        polygon: 原始顶点列表，每个点为 (x, y, z)
    返回:
        去重后的顶点列表
    """
    if not polygon:
        return []
    normalized = [polygon[0]]
    for pt in polygon[1:]:
        if pt != normalized[-1]:
            normalized.append(pt)
    # 首尾相同则删尾
    if len(normalized) > 1 and normalized[0] == normalized[-1]:
        normalized.pop()
    return normalized


# 2. 找到某顶点的前驱/后继（按循环多边形）
def get_adjacent_points(polygon, p):
    """
    在多边形 polygon 中返回顶点 p 的前后相邻点（支持循环）。
    会先调用 normalize_polygon 清理重复点。
    """
    poly = normalize_polygon(polygon)
    if not poly:
        raise ValueError("多边形为空")
    try:
        idx = poly.index(p)
    except ValueError:
        raise ValueError(f"点 {p} 不在多边形顶点列表中")
    prev_pt = poly[idx - 1] if idx > 0 else poly[-1]
    next_pt = poly[idx + 1] if idx < len(poly) - 1 else poly[0]
    return prev_pt, next_pt


# 3. 点是否在多边形内部（射线法，仅在 XY 平面判断）
def point_in_polygon(pt, polygon):
    """
    判断三维点 pt=(x,y,z) 在多边形 polygon 的 XY 投影内否。
    polygon 中点格式为 (x,y,z)，首尾可重复或不重复均可。
    """
    x, y, _ = pt
    poly2d = [(q[0], q[1]) for q in normalize_polygon(polygon)]
    inside = False
    n = len(poly2d)
    for i in range(n):
        x1, y1 = poly2d[i]
        x2, y2 = poly2d[(i + 1) % n]
        if (y1 > y) != (y2 > y):
            xint = x1 + (y - y1) * (x2 - x1) / (y2 - y1)
            if xint > x:
                inside = not inside
    return inside


# 4. 无穷直线 vs 线段在 XY 平面相交
def line_segment_intersection_2d(p, d, a, b, tol=1e-8):
    """
    计算射线 L(t)=p + t·d 与线段 AB 在 XY 平面上的交点。
    p,d,a,b 皆为 (x,y,z)，但只取 x,y 分量参与计算。
    返回 (xi, yi, t) 或 None。
    """
    px, py, _ = p
    dx, dy, _ = d
    ax, ay, _ = a
    bx, by, _ = b
    ux, uy = bx - ax, by - ay

    det = dx * (-uy) - dy * (-ux)
    if abs(det) < tol:
        return None

    rhsx, rhsy = ax - px, ay - py
    t_param = (rhsx * (-uy) - rhsy * (-ux)) / det
    u_param = (dx * rhsy - dy * rhsx) / det

    if -tol <= u_param <= 1 + tol:
        xi = px + t_param * dx
        yi = py + t_param * dy
        return xi, yi, t_param
    return None


# 5. 计算 p 和其相邻点中点 c，如果 c 内部则返回 c，否则沿 p->c 的射线找到第一个交点
def get_auxiliary_point(p, p_prev, p_next, polygon, tol=1e-8):
    """
    对于多边形顶点 p 及其前后相邻点 p_prev, p_next，
    返回一个位于多边形内部的辅助点 q：
      1. 先取 c = (p_prev + p_next)/2；若 c 在内部，则返回 c
      2. 否则沿射线 p->c 与多边形其它边求最靠近 p 的交点
    返回点格式为 (x,y,z)。
    """
    # 1) 中点 c
    cx = (p_prev[0] + p_next[0]) / 2
    cy = (p_prev[1] + p_next[1]) / 2
    cz = (p_prev[2] + p_next[2]) / 2
    c = (cx, cy, cz)

    if point_in_polygon(c, polygon):
        return c

    # 2) 构造方向 d = c - p 并归一化
    dx, dy = cx - p[0], cy - p[1]
    mag = math.hypot(dx, dy)
    if mag < tol:
        raise RuntimeError("辅助点方向向量过小")
    d = (dx / mag, dy / mag, 0.0)

    # 在每条不含 p 的边上求交
    poly = normalize_polygon(polygon)
    intersects = []
    for i in range(len(poly)):
        a = poly[i]
        b = poly[(i + 1) % len(poly)]
        if a == p or b == p:
            continue
        res = line_segment_intersection_2d(p, d, a, b, tol)
        if res:
            xi, yi, t_param = res
            if abs(t_param) > tol:
                # 插值出对应的 z
                zi = p[2] + t_param * d[2]
                intersects.append((xi, yi, zi, t_param))

    if not intersects:
        raise RuntimeError("未找到有效交点")
    # 取最小 |t| 的那一个
    intersects.sort(key=lambda it: abs(it[3]))
    xi, yi, zi, _ = intersects[0]
    return (xi, yi, zi)


# 6. 计算 p 点的“凹凸度量角”
def concavity_measure(p, p_prev, p_next, q):
    """
    给定 p, p_prev, p_next, q（均为 (x,y,z)），
    计算度量角：
      angle = 360 - larger_angle + smaller_angle  
    其中 smaller/larger 是 pq→p_prev 与 pq→p_next 的逆时针夹角。
    凸点约 ~90°，凹点约 ~270°。
    """
    def angle_of(vx, vy):
        a = math.degrees(math.atan2(vy, vx))
        return a if a >= 0 else a + 360

    # 构造 2D 向量
    vq = (q[0] - p[0], q[1] - p[1])
    v1 = (p_prev[0] - p[0], p_prev[1] - p[1])
    v2 = (p_next[0] - p[0], p_next[1] - p[1])

    a_q = angle_of(*vq)
    a1  = angle_of(*v1)
    a2  = angle_of(*v2)

    d1 = (a1 - a_q) % 360
    d2 = (a2 - a_q) % 360

    small, large = (d1, d2) if d1 < d2 else (d2, d1)
    return 360 - large + small


# 7. 直接给出 p 在多边形上的度量角
def concavity_angle(p, polygon):
    """
    直接计算多边形 polygon 上顶点 p 的凹凸度量角。
    """
    p_prev, p_next = get_adjacent_points(polygon, p)
    q = get_auxiliary_point(p, p_prev, p_next, polygon)
    return concavity_measure(p, p_prev, p_next, q)





# 8.合理分割PL正交六边形

def split_orthogonal_hexagon(polygon, tol=0.1):#水平分割
    """
    将正交六边形 polygon 按凹顶点所在水平线切成两个矩形。
    polygon: 6 个 (x,y,z) 顶点的列表，允许首尾重合或相邻重复，会自动规范化。
    """
    # 1. 规范化，去掉相邻重复和首尾同点
    poly = normalize_polygon(polygon)
    if len(poly) != 6:
        raise ValueError("必须传入6点正交多边形")
    # 2. 找出唯一的凹点 p
    concaves = [pt for pt in poly
                if abs(concavity_angle(pt, poly) - 270) < tol]
    if len(concaves) != 1:
        raise RuntimeError(f"没能唯一定位凹点，找到 {len(concaves)} 个")
    p = concaves[0]
    y0 = p[1]

    # 3. 只对真正“跨越” y=y0 的边求交
    intersections = []
    n = len(poly)
    for i in range(n):
        a = poly[i]
        b = poly[(i+1) % n]
        y1, y2 = a[1], b[1]
        # 仅当严格跨越才算：一端在上 (y>y0)，一端在下 (y<y0)
        if (y1 - y0) * (y2 - y0) < -tol**2:
            t = (y0 - y1) / (y2 - y1)
            xi = a[0] + t * (b[0] - a[0])
            zi = a[2] + t * (b[2] - a[2])
            intersections.append((xi, y0, zi, i))

    # 4. 应该只剩一个真的 crossing 交点
    if len(intersections) != 1:
        raise RuntimeError(f"没能唯一定位 q，找到 {len(intersections)} 个候选点")
    xi, yi, zi, edge_idx = intersections[0]
    q = (xi, yi, zi)

    # 5. 把 q 插回那条边之后
    newpoly = []
    for i in range(n):
        newpoly.append(poly[i])
        if i == edge_idx:
            newpoly.append(q)
    # now len(newpoly)==7

    # 6. 找 p, q 的索引
    i_p = newpoly.index(p)
    i_q = newpoly.index(q)

    # 7. 分割成两段多边形
    if i_q < i_p:
        rect1 = newpoly[i_q:i_p+1]
        rect2 = newpoly[i_p:] + newpoly[:i_q+1]
    else:
        rect1 = newpoly[i_p:i_q+1]
        rect2 = newpoly[i_q:] + newpoly[:i_p+1]

    # 8. 计算2D面积
    def area2d(pts):
        s = 0
        m = len(pts)
        for j in range(m):
            x1,y1,_ = pts[j]
            x2,y2,_ = pts[(j+1)%m]
            s += x1*y2 - x2*y1
        return abs(s)/2

    A1, A2 = area2d(rect1), area2d(rect2)
    # 面积小的放前面
    return (rect1, rect2) if A1 <= A2 else (rect2, rect1)


def split_orthogonal_hexagon_vertical(polygon, tol=0.1):#竖向分割
    """
    将正交六边形 polygon 按凹顶点所在竖线切成两个矩形。
    polygon: 6 个 (x,y,z) 顶点的列表，允许首尾重合或相邻重复，会自动规范化。
    tol: 用于识别凹点和跨越判断的容差。
    返回: (rect1, rect2)，面积小的放前面。
    """
    # 规范化：去掉相邻重复和首尾同点
    poly = normalize_polygon(polygon)
    if len(poly) != 6:
        raise ValueError("必须传入6点正交多边形")

    # 找唯一凹点
    concaves = [pt for pt in poly
                if abs(concavity_angle(pt, poly) - 270) < tol]
    if len(concaves) != 1:
        raise RuntimeError(f"没能唯一定位凹点，找到 {len(concaves)} 个")
    p = concaves[0]
    x0 = p[0]

    # 求竖线 x=x0 与真正跨越边的交点
    intersections = []
    n = len(poly)
    for i in range(n):
        a = poly[i]
        b = poly[(i+1)%n]
        x1, x2 = a[0], b[0]
        # 仅对严格跨越的边求交
        if (x1 - x0)*(x2 - x0) < -tol**2:
            # 插值比例
            t = (x0 - x1)/(x2 - x1)
            yi = a[1] + t*(b[1] - a[1])
            zi = a[2] + t*(b[2] - a[2])
            intersections.append((x0, yi, zi, i))

    if len(intersections) != 1:
        raise RuntimeError(f"没能唯一定位 q，找到 {len(intersections)} 个候选点")
    xi, yi, zi, edge_idx = intersections[0]
    q = (xi, yi, zi)

    # 把 q 插回那条边
    newpoly = []
    for i in range(n):
        newpoly.append(poly[i])
        if i == edge_idx:
            newpoly.append(q)
    # newpoly 长度应为7

    # 定位 p、q 索引
    i_p = newpoly.index(p)
    i_q = newpoly.index(q)

    # 分割两段
    if i_q < i_p:
        rect1 = newpoly[i_q:i_p+1]
        rect2 = newpoly[i_p:] + newpoly[:i_q+1]
    else:
        rect1 = newpoly[i_p:i_q+1]
        rect2 = newpoly[i_q:] + newpoly[:i_p+1]

    # 面积计算（2D）
    def area2d(pts):
        s = 0
        m = len(pts)
        for j in range(m):
            x1,y1,_ = pts[j]
            x2,y2,_ = pts[(j+1)%m]
            s += x1*y2 - x2*y1
        return abs(s)*0.5

    A1, A2 = area2d(rect1), area2d(rect2)
    return (rect1, rect2) if A1 <= A2 else (rect2, rect1)


# 合理分割PL正交六边形

def area_of(verts):
    """多边形面积计算（顶点首尾闭合或不闭合均可）"""
    s = 0
    n = len(verts)
    for i in range(n):
        x1, y1, *_ = verts[i]
        x2, y2, *_ = verts[(i+1) % n]
        s += x1 * y2 - x2 * y1
    return abs(s) * 0.5



def split_hexagon_combined(polygon, tol=0.1, simplify_tol=0.5):# 合理分割PL正交六边形
    """
    合理分割一个正交（近似）六边形 PL：
      1) 如果传入的是 COM PL 对象，则先提取唯一顶点列表；
      2) 对顶点列表做简化（去除伪顶点/伪边）；
      3) 先做横向分割，再做竖向分割，比较两种分割最小矩形面积，
         将最小的那对矩形排在前面，返回四个矩形顶点列表：
         [min1, partner1, min2, partner2]
    参数:
      polygon: 要分割的多段线，既可以是 [(x,y,z),...] 顶点列表，也可以是 COM PL 对象
      tol: 分割时判断凹角的容差
      simplify_tol: 简化多边形时去除“伪顶点/伪边”的容差
    返回:
      四个矩形的顶点列表：最小矩形、其同组矩形、次小矩形、其同组矩形
    """
    # —— 1. 如果是 COM PL 对象，先提取顶点列表 —— 
    # （假定已有 get_unique_vertices_from_pl_com(com_pl) -> [(x,y,z),...]）
    if not isinstance(polygon, list):
        polygon = get_unique_vertices_from_pl_com(polygon)

    # —— 2. 简化顶点列表 —— 
    # （假定已有 simplify_polygon(verts, tol) -> 清理后的顶点列表）
    polygon = simplify_polygon(polygon, simplify_tol)

    # —— 3. 横向与竖向分割 —— 
    rects_h = split_orthogonal_hexagon(polygon, tol)
    rects_v = split_orthogonal_hexagon_vertical(polygon, tol)

    A_h1, A_h2 = rects_h
    A_v1, A_v2 = rects_v

    # 计算每对的“更小”矩形和“配对”矩形
    if area_of(A_h1) <= area_of(A_h2):
        min_h, partner_h = A_h1, A_h2
    else:
        min_h, partner_h = A_h2, A_h1

    if area_of(A_v1) <= area_of(A_v2):
        min_v, partner_v = A_v1, A_v2
    else:
        min_v, partner_v = A_v2, A_v1

    # —— 4. 根据最小面积值决定输出顺序 —— 
    if area_of(min_h) <= area_of(min_v):
        return [min_h, partner_h, min_v, partner_v]
    else:
        return [min_v, partner_v, min_h, partner_h]



#  主函数
#  (4)
#  获取多段线的上下左右边界的直线段，返回线段端点列表

#  该函数系列包括如下一些函数

def get_bbox_edge_segments(pl, tol=0.5):
    """
    获取对象 pl 的包围盒四条边，分别作为独立的列表返回：
      top    = [(xmin, ymax, z), (xmax, ymax, z)]
      bottom = [(xmin, ymin, z), (xmax, ymin, z)]
      left   = [(xmin, ymin, z), (xmin, ymax, z)]
      right  = [(xmax, ymin, z), (xmax, ymax, z)]
    并打印调试信息。
    """
    # ----- 1. 调用 GetBoundingBox -----
    try:
        min_pt, max_pt = pl.GetBoundingBox()
    except Exception:
        mins = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (0.0, 0.0, 0.0))
        maxs = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (0.0, 0.0, 0.0))
        pl.GetBoundingBox(mins, maxs)
        min_pt = tuple(mins.value)
        max_pt = tuple(maxs.value)

    xmin, ymin, zmin = min_pt
    xmax, ymax, _    = max_pt

    print(f"▶ BoundingBox Min 点: {min_pt}")
    print(f"▶ BoundingBox Max 点: {max_pt}")

    # ----- 2. 构造四个顶点（顺时针） -----
    p1 = (xmin, ymin, zmin)
    p2 = (xmax, ymin, zmin)
    p3 = (xmax, ymax, zmin)
    p4 = (xmin, ymax, zmin)

    print("▶ 矩形四个顶点 (顺时针):")
    for i, pt in enumerate((p1, p2, p3, p4), 1):
        print(f"   {i}: {pt}")

    # ----- 3. 四条边，各自用列表表达 -----
    top    = [p4, p3]    # y = ymax
    bottom = [p1, p2]    # y = ymin
    left   = [p1, p4]    # x = xmin
    right  = [p2, p3]    # x = xmax

    return top, bottom, left, right



#  主函数
#  (5)
#&&%  获取多段线的内部的文字

#  该函数系列包括如下一些函数



def get_texts_in_polyline(com_pl, tol=0.5):
    """
    在多段线 com_pl 内部筛选文字，并返回文字对象列表和对应的文字内容列表。

    参数:
      com_pl:  COM 多段线对象
      tol:     点-in-多边形时的容差（目前未用到，可留作将来扩展）

    返回:
      inside:   落在 com_pl 内部的文字 COM 对象列表
      contents: 对应 inside 中每个对象的文字内容列表
    """
    # 1) 多段线转顶点列表（标准化后的三维点）
    poly = get_unique_vertices_from_pl_com(com_pl)

    # 2) 收集所有文字实体（天正＋原生 CAD）
    tzh_text, tzh_mtext, cad_text, cad_mtext = collect_all_texts()
    all_texts = tzh_text + tzh_mtext + cad_text + cad_mtext

    inside = []
    contents = []

    for txt in all_texts:
        # 用左下角点判断是否在多段线内
        min_pt, _ = txt.GetBoundingBox()
        if point_in_polygon(min_pt, poly):
            inside.append(txt)

            # 根据对象类型取出内容
            name = getattr(txt, "ObjectName", "") or getattr(txt, "EntityName", "")
            if name in ("AcDbText", "AcDbMText"):
                # AutoCAD 原生单/多行
                contents.append(txt.TextString)
            elif name == "TDbText":
                # 天正单行
                contents.append(txt.Text)
            elif name == "TDbMText":
                # 天正多行，需要炸开取内容
                contents.append( TDbMText_content(txt))
            else:
                # 其它（万一有），留空或用 repr
                contents.append("")

    print(f"总共找到 {len(inside)} 条落在多段线内部的文字。")
    return inside, contents




# 获取单独一行的天正多行文字内容

def TDbMText_content(comobj):

    """
    对复杂多行文本内容，需要专门分析

    本函数仅针对单个文字，获取其文字内容

    不改变其本身属性和其它对象的关系

    """

    highlight_entity_by_bbox( comobj)
    cmd = "x\n"
    doc.SendCommand(cmd)

    ob=doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
    neirong=ob.TextString

    cmd = "u\n"

    doc.SendCommand(cmd)

    return neirong







#  主函数
#  (6)
#  多段线上的均分插入


"""
该函数用于在dwg文件沿着PL线快速均衡放置树木等图块

            
"""
def distribute_points_on_entity(entity, n, block, scale_factor, ys):

    def distance(p1, p2):
        return math.sqrt((p1[0] - p2[0])**2 + (p1[1] - p2[1])**2)

    model_space = doc.ModelSpace
    block_name = block.Name  # 获取块的名称
    
    # 如果实体是直线
    if entity.ObjectName == "AcDbLine":
        start_point = entity.StartPoint
        end_point = entity.EndPoint
        for i in range(n):
            x = start_point[0] + i * (end_point[0] - start_point[0]) / (n - 1)
            y = start_point[1] + i * (end_point[1] - start_point[1]) / (n - 1)
            inserted_block = model_space.InsertBlock(vtpnt(x, y, 0), block_name, scale_factor, scale_factor, scale_factor, 0)
            inserted_block.Color = ys  # 设置颜色为红色

    # 如果实体是圆弧
    elif entity.ObjectName == "AcDbArc":
        start_angle = entity.StartAngle
        end_angle = entity.EndAngle
        center = entity.Center
        radius = entity.Radius
        for i in range(n):
            angle = start_angle + i * (end_angle - start_angle) / (n - 1)
            x = center[0] + radius * math.cos(angle)
            y = center[1] + radius * math.sin(angle)
            inserted_block = model_space.InsertBlock(vtpnt(x, y, 0), block_name, scale_factor, scale_factor, scale_factor, 0)
            inserted_block.Color = ys  # 设置颜色为红色

    # 如果实体是多段线
    elif entity.ObjectName == "AcDbPolyline":
        coords = entity.Coordinates
        points = [(coords[i], coords[i+1]) for i in range(0, len(coords), 2)]
        total_length = sum(distance(points[i], points[i+1]) for i in range(len(points)-1))
        segment_length = total_length / n
        current_length = 0

        for i in range(n):
            accumulated_length = 0
            for j in range(len(points) - 1):
                segment = distance(points[j], points[j+1])
                if accumulated_length + segment > current_length:
                    ratio = (current_length - accumulated_length) / segment
                    x = points[j][0] + ratio * (points[j+1][0] - points[j][0])
                    y = points[j][1] + ratio * (points[j+1][1] - points[j][1])
                    inserted_block = model_space.InsertBlock(vtpnt(x, y, 0), block_name, scale_factor, scale_factor, scale_factor, 0)
                    inserted_block.Color = ys  # 设置颜色为红色
                    break
                accumulated_length += segment
            current_length += segment_length
            




#  主函数
#  (7)
# 返回 pl1 中与 pl2 “共线且有重叠”的区段列表

#  该函数系列包括如下一些函数

# 1 判断一条直线是否完全在另一条直线上

def is_segment_contained(seg_a, seg_b, tol=1e-4):
    """
    判断 seg_a 是否完全位于 seg_b 上（包含端点）。
    
    参数
    ----
    seg_a, seg_b :  ( (x1,y1,z1), (x2,y2,z2) )   或   AcDbLine
        两条待比较的线段。先判断 seg_a 是否被 seg_b 覆盖；
        若想双向判断，调用两次并对调顺序即可。
    tol : float
        距离与投影误差容差，默认 1e‑4 (CAD 单位)。

    返回
    ----
    bool
        True  —— seg_a 整段落在 seg_b 上  
        False —— 否则
    """
    # -------- 把输入统一转成端点元组 ----------
    def get_endpoints(entity):
        if hasattr(entity, "StartPoint"):            # COM 线段
            return (tuple(entity.StartPoint), tuple(entity.EndPoint))
        else:                                        # 纯坐标二元组
            return (tuple(entity[0]), tuple(entity[1]))

    a1, a2 = get_endpoints(seg_a)
    b1, b2 = get_endpoints(seg_b)

    # -------- 基本几何工具 ----------
    def dist(p, q):
        return math.hypot(p[0]-q[0], p[1]-q[1])

    def dot(u, v):
        return u[0]*v[0] + u[1]*v[1]

    def colinear(p, q, r, tol):
        """三点是否近似共线（面积≈0）"""
        return abs( (q[0]-p[0])*(r[1]-p[1]) - (q[1]-p[1])*(r[0]-p[0]) ) <= tol

    # -------- 先检测共线性 ----------
    if not (colinear(b1, b2, a1, tol) and colinear(b1, b2, a2, tol)):
        return False          # seg_a 端点不与 seg_b 共线 → 不可能包含

    # -------- 再检测投影是否在 b1‑b2 区间内 ----------
    # 令 b1 为原点，b_dir 为方向向量
    b_dir = (b2[0]-b1[0], b2[1]-b1[1])
    b_len2 = dot(b_dir, b_dir)
    if b_len2 == 0:           # seg_b 长度为 0 → 无法包含他段
        return False

    def proj_param(p):
        # 标准化投影参数 t，若 0<=t<=1 则投影落在 seg_b 上
        return dot( (p[0]-b1[0], p[1]-b1[1]), b_dir ) / b_len2

    t_a1 = proj_param(a1)
    t_a2 = proj_param(a2)

    # 允许介于 0±tol 到 1±tol 之间
    inside_1 = -tol <= t_a1 <= 1+tol
    inside_2 = -tol <= t_a2 <= 1+tol

    return inside_1 and inside_2

#&&% 2 返回 PL线pl1 中与 pl2 “共线且有重叠”的区段列表

def common_segments_between_polylines(pl1, pl2, tol=0.5):
    """
    返回 pl1 中与 pl2 “共线且有重叠”的区段列表，每个区段用
      [(x1,y1,0.0),(x2,y2,0.0)] 表示。

    参数
    ----
    pl1, pl2 : AutoCAD AcDbPolyline COM 对象 (或伪造对象，只需有 .Coordinates / .Closed)
    tol      : 共线 & 距离容差 (同 CAD 单位)

    返回
    ----
    overlaps : list[ list[(x,y,z),(x,y,z)] ]
    """

    # ────────── 内部小工具 ──────────
    def coords_to_xy_pairs(coords):
        """把 (x1,y1,x2,y2,…) 转成 [(x,y),…]；自动丢弃尾部残值"""
        pairs = []
        for i in range(0, len(coords) - 1, 2):
            pairs.append((coords[i], coords[i + 1]))
        return pairs

    def build_segments(verts, closed=False):
        """由顶点顺序生成线段 [P,Q] 列表"""
        segs = []
        for i in range(len(verts) - 1):
            segs.append([verts[i], verts[i + 1]])
        if closed and len(verts) > 2:
            segs.append([verts[-1], verts[0]])
        return segs

    def dist(p, q):
        return math.hypot(p[0] - q[0], p[1] - q[1])

    def colinear(p, q, r):
        """三点共线判定 |叉积| < tol · max(边长)"""
        return abs((q[0] - p[0]) * (r[1] - p[1]) -
                   (q[1] - p[1]) * (r[0] - p[0])) <= tol * max(dist(p, q), dist(p, r), dist(q, r), 1)

    def project(p, axis):
        """根据 axis==0 用 x，否则 y"""
        return p[axis]

    def segment_overlap(seg_a, seg_b):
        """
        若共线且区间有重叠，返回实际重叠区间端点 (pa, pb)（二维点）。
        否则返回 None
        """
        (p1, p2), (q1, q2) = seg_a, seg_b
        # 共线检测
        if not (colinear(p1, p2, q1) and colinear(p1, p2, q2)):
            return None

        # 选投影轴
        axis = 0 if abs(p2[0] - p1[0]) >= abs(p2[1] - p1[1]) else 1
        a1, a2 = project(p1, axis), project(p2, axis)
        b1, b2 = project(q1, axis), project(q2, axis)
        # 使 a1 <= a2, b1 <= b2
        if a1 > a2:
            p1, p2 = p2, p1
            a1, a2 = a2, a1
        if b1 > b2:
            q1, q2 = q2, q1
            b1, b2 = b2, b1

        # 计算 1‑D 重叠区间
        left, right = max(a1, b1), min(a2, b2)
        if right - left <= tol:          # “长度”≈0 视为无重叠
            return None

        # 把投影点还原到 2D 端点 —— 在线段 p1‑p2 上按比例取值
        def interp(t):
            """t 为 0~1"""
            return (p1[0] + t * (p2[0] - p1[0]),
                    p1[1] + t * (p2[1] - p1[1]))

        len_p = a2 - a1 if a2 != a1 else 1e-9
        pa = interp((left - a1) / len_p)
        pb = interp((right - a1) / len_p)
        return pa, pb

    # ────────── 主流程 ──────────
    v1 = coords_to_xy_pairs(pl1.Coordinates)
    v2 = coords_to_xy_pairs(pl2.Coordinates)

    # 若为闭合 polyline，补一个首尾顶点
    if getattr(pl1, "Closed", False) and v1 and v1[0] != v1[-1]:
        v1.append(v1[0])
    if getattr(pl2, "Closed", False) and v2 and v2[0] != v2[-1]:
        v2.append(v2[0])

    segs1 = build_segments(v1, closed=False)
    segs2 = build_segments(v2, closed=False)

    overlaps = []

    for s1 in segs1:
        for s2 in segs2:
            ov = segment_overlap(s1, s2)
            if ov:
                pa, pb = ov
                overlaps.append([(pa[0], pa[1], 0.0), (pb[0], pb[1], 0.0)])
                break        # 一条 s1 找到重叠就够了，可跳出

    # --- 打印摘要 ---
    print("★ 与 pl2 重叠 (或被包含) 的 pl1 线段数：", len(overlaps))
    for idx, seg in enumerate(overlaps, 1):
        print(f"  {idx}. {seg[0]}  →  {seg[1]}")

    return overlaps




#  主函数
#  (8)
# 找到全部“两根多段线耦合成一个矩形”的多段线

#  该函数系列包括如下一些函数

"""
函数是用来分析主房间带卫生间这种情况的，因此对输入变量是有较严格假定的，并非针对任意情况

"""
# 1 判断矩形是否包含另一个矩形

def is_rect_inside_rect(rect_outer, rect_inner, tol=1e-6):
    """
    判定 axis‑aligned 的矩形 rect_inner 是否被（含边界）完全包在 rect_outer 内。

    参数
    ----
    rect_outer : ((xmin, ymin), (xmax, ymax))
    rect_inner : ((xmin, ymin), (xmax, ymax))
        两个元组分别给出矩形左下、右上坐标（假定 Z 全 0）。
    tol        : float
        容差（允许轻微数值误差；AutoCAD double 转 python float 时推荐 1e‑6～1e‑4）。

    返回
    ----
    bool   ——  rect_inner ⊆ rect_outer ？
    """
    (ox0, oy0), (ox1, oy1) = rect_outer
    (ix0, iy0), (ix1, iy1) = rect_inner

    return (
        ix0 >= ox0 - tol and
        iy0 >= oy0 - tol and
        ix1 <= ox1 + tol and
        iy1 <= oy1 + tol
    )




# 2 判断两条正交多段线拼在一起后是否正好是一个矩形


def two_plines_making_rectangle(pl1, pl2, tol=0.5):#
    """
    判断两条正交多段线拼在一起后是否正好是一个矩形。
    假设：两 PLine 没有面积重叠，只可能共用完整边或边的一部分。
    """

    import math

    def same_point(a, b, tol):
        return abs(a[0] - b[0]) <= tol and abs(a[1] - b[1]) <= tol

    def pline_vertices(pl):
        # 从 pl.Coordinates 提取 (x,y) 顶点列表，自动闭合
        c = pl.Coordinates
        verts = []
        for i in range(0, len(c), 2):
            verts.append((c[i], c[i+1]))
        if not same_point(verts[0], verts[-1], tol):
            verts.append(verts[0])
        return verts

    def poly_area(verts):
        # 计算首尾闭合的多边形面积
        s = 0
        for i in range(len(verts)-1):
            x1,y1 = verts[i]
            x2,y2 = verts[i+1]
            s += x1*y2 - x2*y1
        return abs(s)*0.5

    def collect_segments(verts):
        # 从顶点列表生成线段列表 [((x1,y1),(x2,y2)), ...]
        segs = []
        for i in range(len(verts)-1):
            segs.append((verts[i], verts[i+1]))
        return segs

    def covers_edge(edge, segs):
        """
        判断给定的 bbox 边 edge=((x0,y0),(x1,y1)) 能否被 segs 中若干共线线段
        连续覆盖（无缝隙）。
        """
        (x0,y0),(x1,y1) = edge
        # 算出方向向量和长度
        dx, dy = x1-x0, y1-y0
        L = math.hypot(dx, dy)
        if L < tol:
            return False
        ux, uy = dx/L, dy/L  # 单位方向向量

        # 投影每条 seg 到 [0,L] 参数区间
        intervals = []
        for (ax,ay),(bx,by) in segs:
            # 判断端点是否在同一直线上
            # cross == 0
            cross = (ax-x0)*dy - (ay-y0)*dx
            if abs(cross) > tol*L:  
                continue
            # 计算两端投影参数
            t1 = (  (ax-x0)*ux + (ay-y0)*uy )
            t2 = (  (bx-x0)*ux + (by-y0)*uy )
            a, b = min(t1,t2), max(t1,t2)
            # 只保留与 [0,L] 有交集的部分
            if b < -tol or a > L+tol:
                continue
            intervals.append((max(0.0, a), min(L, b)))

        if not intervals:
            return False

        # 合并所有区间
        intervals.sort(key=lambda iv: iv[0])
        cur_start, cur_end = intervals[0]
        for a,b in intervals[1:]:
            if a > cur_end + tol:
                # 出现间隙
                return False
            cur_end = max(cur_end, b)
        # 最后检测是否覆盖了整个 [0, L]
        return (cur_start <= tol) and (cur_end >= L - tol)

    # 1) 提取顶点及面积
    v1 = pline_vertices(pl1)
    v2 = pline_vertices(pl2)
    A1 = poly_area(v1)
    A2 = poly_area(v2)

    # 2) 计算公共外包矩形
    xs = [p[0] for p in v1[:-1]] + [p[0] for p in v2[:-1]]
    ys = [p[1] for p in v1[:-1]] + [p[1] for p in v2[:-1]]
    xmin, xmax = min(xs), max(xs)
    ymin, ymax = min(ys), max(ys)
    A_bbox = (xmax-xmin)*(ymax-ymin)

    # 3) 面积校验
    if abs((A1 + A2) - A_bbox) > tol:
        return False

    # 4) 检查四条边被覆盖
    bbox_edges = [
        ((xmin,ymin),(xmax,ymin)),
        ((xmax,ymin),(xmax,ymax)),
        ((xmax,ymax),(xmin,ymax)),
        ((xmin,ymax),(xmin,ymin))
    ]
    segs = collect_segments(v1) + collect_segments(v2)

    hits = 0
    for edge in bbox_edges:
        if covers_edge(edge, segs):
            hits += 1
    return hits == 4



#  主函数
#  (9)
#  判断多段线PL2是否在多段线PL1多边形中

#  该函数系列包括如下一些函数

def are_all_vertices_inside(pl1, pl2):
    """
    判断多段线 pl2 的所有顶点是否都在多段线 pl1 构成的多边形内部。

    参数：
      pl1, pl2: COM 多段线对象（LWPOLYLINE 或 POLYLINE），视为封闭正多边形。
    返回：
      (all_inside, outside_pts)
      - all_inside: 如果 pl2 的每个顶点都在 pl1 内部，返回 True，否则 False。
      - outside_pts: 列表，包含所有落在 pl1 外部的 pl2 顶点 (x, y, z)。
    """
    # 先把 COM 多段线转成顶点列表 [(x,y,z), ...]
    verts1 = get_unique_vertices_from_pl_com(pl1)
    verts2 = get_unique_vertices_from_pl_com(pl2)

    outside_pts = []
    for pt in verts2:
        if not point_in_polygon(pt, verts1):
            outside_pts.append(pt)

    all_inside = len(outside_pts) == 0
    if all_inside:
        print("✅ pl2 的所有顶点都在 pl1 的内部。")
    else:
        print(f"❌ pl2 有 {len(outside_pts)} 个顶点不在 pl1 内部：")
        for p in outside_pts:
            print("   ", p)

    return all_inside, outside_pts




#&&&&%% 第六部分 综合控制管理

#  模块使用说明

"""



估计马工那样的文件打开时会遇到大量的窗口，需要按照插入图签函数那样采用双线程进行7次取消处理 20250720





没有那么复杂，根本上讲，使用代码将一个作业过程完成，再将其流程标准化梳理，这个过程的自动化就完成了。不需要太多耗费心神研究这个过程本身。





通过输入输出字典，不仅收集了信息，统一了信息处理平台，将来在构建大函数节点时，也可以通过查看字典的信息输出，判断每一个环节的重要参数

控制系统运转

抓住主要核心部分，忽略次要的要素，免得陷入无意义的劳作

先把整体大框架骨架建立好

关键概念，思想，节点，消息要打印出来分析

加强数据结构和算法的训练，这是要点

我们不仅仅是解决了某个问题，更是研究清楚了其包含的内在规律。这些骨架，结构，相关的思想方法概念就是内在规律的表达。

要深刻确定每个功能函数的设计思想和骨架结构，不能含糊依赖人工智能，否则就会飘忽不确定，没力度。有了本质的明确的认知，需要的时候，无惧重来


系统尽可能完全自动化   尽可能少干预系统自动化 

关于整个系统处理的优化分析

文件处理之前，应该对所有对象选择后分类放入各自的预设图层，这样将来可以快速找到它们
同样文件操作中，将相应对象放入它们的图层，也便于找到它们
图层的选择很快，也可以使用Handle标签等
如果对象有很强的逻辑关联，可以建立组用来关联它们

20250423

对自动化操作有了新的提升。我们有pywin32的API控制，还有SendComman发送窗口命令控制，还有pyautogui等拟人操作控制，
还有autolisp控制。同时还要注意整个电脑系统 的控制设置。它们是综合解决问题的。

把CAD窗口设置为1/4，把IDLE设置到右侧，这是使用拟人操作的基础。

select_objects_in_window_area 实际选择不可靠且慢，可行方法是使用高亮显示，再通过pick方法转隐性选择 pick方法转隐性选择
也会改变文件之间的复制粘贴

基于在一次操作运算中获得的信息应该进行充分的整合分析，减少重复选择计算和在CAD复杂环境操作的风险


应该总结经验教训，针对一个任务的深入分析，分解成函数，数据结构，各个结构的咬合，精准到位。比如编目录，它的要点，每个任务的实现，组合实现。转场，参照起点，雷边，删除，选择

插图纸目录模板时跳出shx字体对话框 

activate_window_by_title("缺少 SHX 文件")
✅ 已激活窗口：“缺少 SHX 文件” 位置(1060,510) 大小424×318
(1060, 510, 424, 318)

已获取坐标：(1232, 733)  点击忽略
(1232, 733)




activate_window_by_title("AutoCAD 错误中断", click_titlebar= True)
✅ 已激活窗口：“AutoCAD 错误中断” 位置(485,227) 大小364×211
(485, 227, 364, 211)

ceshubiao_weizhi()
请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置… 确定
已获取坐标：(775, 404)
(775, 404)


activate_window_by_title("AutoCAD 警告", click_titlebar= True)
✅ 已激活窗口：“AutoCAD 警告” 位置(392,247) 大小497×227
(392, 247, 497, 227)

ceshubiao_weizhi()
请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置…  否
已获取坐标：(811, 434)
(811, 434)


✅ 已激活窗口：“ AutoCAD 错误报告” 位置(1055,425) 大小450×549
(1055, 425, 450, 549)

ceshubiao_weizhi()
请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置…  x（关闭）
已获取坐标：(1483, 440)
(1483, 440)

activate_window_by_title("错误报告 - 已取消", click_titlebar= True)
✅ 已激活窗口：“错误报告 - 已取消” 位置(1127,610) 大小307×179
(1127, 610, 307, 179)

ceshubiao_weizhi()
请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置… 确定
已获取坐标：(1364, 753)
(1364, 753)


出问题以后的CAD重新打开

activate_window_by_title("图形修复", click_titlebar = True)
✅ 已激活窗口：“图形修复” 位置(1089,562) 大小366×213
(1089, 562, 366, 213)

ceshubiao_weizhi()
请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置…关闭(汉字按钮)
已获取坐标：(1397, 724)
(1397, 724)



"""
#&&&% 如果获取不到cad文字的处理办法


from win32com.client import CastTo

# —— 可选：简单清洗 MText 控制码（保留数字用）
#_MTEXT_CTRL_RE = re.compile(r"(\\[A-Za-z])|({|})|(;)|(\\\S[^;]*;)")
def _clean_mtext(s: str) -> str:
    # 把 \P 换成空格，去掉常见控制码；如需保留换行，可改成 '\n'
    s = str(s).replace("\\P", " ")
    s = _MTEXT_CTRL_RE.sub("", s)
    return s.strip()

def get_text_and_ip(ent):
    """
    通吃 AcDbText / AcDbMText / Attribute(Reference)：
    返回 (text, (x,y,z)) 或 (None, None)
    """
    try:
        name = ent.ObjectName
    except Exception:
        name = ""

    # 1) 先尝试“就地访问”，兼容以前版本（有些环境直接能用）
    for attr in ("TextString", "Contents"):  # TEXT / MTEXT
        try:
            txt = getattr(ent, attr)
            ip  = ent.InsertionPoint  # 这一步有时也需要 CastTo，下面会兜底
            if txt:
                return str(txt), (float(ip[0]), float(ip[1]), float(ip[2]) if len(ip) > 2 else 0.0)
        except Exception:
            pass

    # 2) 按类型 CastTo 到具体接口再取
    try:
        if name == "AcDbText":
            obj = CastTo(ent, "IAcadText")
            return str(obj.TextString), tuple(map(float, obj.InsertionPoint))
        elif name == "AcDbMText":
            obj = CastTo(ent, "IAcadMText")
            return _clean_mtext(obj.Contents), tuple(map(float, obj.InsertionPoint))
        elif name in ("AcDbAttribute", "AcDbAttributeReference"):
            # 属性文字（块里的标签）
            # 两种接口都试一下
            try:
                obj = CastTo(ent, "IAcadAttributeReference")
            except Exception:
                obj = CastTo(ent, "IAcadAttribute")
            return str(obj.TextString), tuple(map(float, obj.InsertionPoint))
    except Exception:
        pass

    return None, None


# —— 你原来“按插入点 y 从上到下配对 J1..Jn”的封装（适配 get_text_and_ip）
#_num_re = re.compile(r"-?\d+(?:\.\d+)?")

def _first_number(s: str) -> float:
    m = _num_re.search(str(s))
    if not m:
        raise ValueError(f"无法解析数字: {s!r}")
    return float(m.group(0))

def build_J_points_from_selected_texts(LB, n_points=61, prefix_x="30", prefix_y="37"):
    """
    输入：LB=select_text() 返回的选择集（含 TEXT/MTEXT/属性文字等）
    规则：
      - 文本以 '30' 开头的是 X；以 '37' 开头的是 Y
      - 按“插入点 y”从大到小（即从上到下）排序；同一行按 x 升序稳定
      - 依序配给 J1..Jn
    返回：
      pts_dict: {'J1': (X, Y), ...}
      pts_list: [('J1', X, Y), ...]  # 已按编号排序
    """
    items = []
    for ent in LB:
        txt, ip = get_text_and_ip(ent)
        if not txt or not ip:
            continue
        x0, y0 = float(ip[0]), float(ip[1])
        t = str(txt).strip().replace(" ", "")

        if t.startswith(prefix_x):
            items.append(("x", _first_number(t), y0, x0))
        elif t.startswith(prefix_y):
            items.append(("y", _first_number(t), y0, x0))

    # 上->下：y 降序；同行 x 升序
    items.sort(key=lambda it: (-it[2], it[3]))

    xs = [v for typ, v, _, _ in items if typ == "x"]
    ys = [v for typ, v, _, _ in items if typ == "y"]

    if len(xs) != n_points or len(ys) != n_points:
        raise RuntimeError(f"X={len(xs)} / Y={len(ys)} 与期望 {n_points} 不符，请检查选择集或前缀。")

    pts_dict = {f"J{i+1}": (xs[i], ys[i]) for i in range(n_points)}
    pts_list = [(f"J{i+1}", xs[i], ys[i]) for i in range(n_points)]
    return pts_dict, pts_list

"""

把你项目里所有直接用：

obj.TextString

obj.Contents

obj.InsertionPoint

的地方，统一替换为：

txt, ip = get_text_and_ip(obj)
# txt 为字符串（已做了基本清洗），ip 为 (x, y, z)


"""
import math
import traceback
from win32com.client import CastTo, gencache, Dispatch

def _flatten_coords(obj):
    """将各种返回格式（tuple/list of tuples OR flat list）统一为平面坐标 [x,y,x,y,...]"""
    if obj is None:
        return None
    # 如果是扁平数值列表
    try:
        if isinstance(obj, (list, tuple)) and obj and all(isinstance(x, (int,float)) for x in obj):
            return list(obj)
    except Exception:
        pass
    # 如果是 list/tuple of points [(x,y,z),...]
    try:
        coords = []
        for p in obj:
            if p is None:
                continue
            # p 可以是 (x,y,z)
            coords.extend([float(p[0]), float(p[1])])
        if coords:
            return coords
    except Exception:
        pass
    return None

def poly_length_from_coords_flat(coords):
    if not coords or len(coords) < 4:
        return 0.0
    total = 0.0
    for i in range(0, len(coords)-2, 2):
        dx = coords[i+2] - coords[i]
        dy = coords[i+3] - coords[i+1]
        total += math.hypot(dx, dy)
    return total

def try_read_coords_by_props(ent):
    """尝试各种常见属性/方法读取多段线顶点或坐标列表"""
    cand_props = ("Coordinates","coordinates","GetCoordinates","Points","PointsXY","Vertexes","Vertices")
    for p in cand_props:
        try:
            val = getattr(ent, p)
            coords = val() if callable(val) else val
            flat = _flatten_coords(coords)
            if flat:
                return flat
        except Exception:
            continue
    return None

def try_cast_and_read(ent):
    """尝试 CastTo 到常见 Polyline 接口并读取 .Coordinates"""
    try:
        # IAca dLWPolyline 或 IAcadPolyline（在 makepy 中常见）
        for iface in ("IAcadLWPolyline","IAcadPolyline"):
            try:
                pev = CastTo(ent, iface)
                # 常见属性
                for p in ("Coordinates","coordinates","GetCoordinates"):
                    try:
                        val = getattr(pev, p)
                        coords = val() if callable(val) else val
                        flat = _flatten_coords(coords)
                        if flat:
                            return flat
                    except Exception:
                        pass
            except Exception:
                continue
    except Exception:
        pass
    return None

def get_length_via_explode(ent, doc):
    """通过复制+Explode的方式收集爆炸后分段的长度（对线段/弧段处理），并清理临时对象
       返回 (length, had_temp_entities_flag)
    """
    created = []
    total = 0.0
    try:
        # 1) 复制实体（Copy 会返回新的 COM 对象）
        try:
            ent_copy = ent.Copy()
            created.append(ent_copy)
        except Exception:
            # 有些 wrapper Copy 方法需要文档操作，尝试 alternative：Add copy via clone? 若失败直接 explode 原对象（危险）
            ent_copy = ent  # 退而求其次（但我们不想直接 explode 原件）
        # 2) Explode（可能返回一个集合或产生新的实体在 ModelSpace）
        try:
            res = ent_copy.Explode()
        except Exception as e:
            # 如果 Explode 失败，尝试直接读取 Length 属性（最后的兜底）
            try:
                return float(ent.Length), False
            except Exception:
                raise

        # res 可能是一个 sequence（生成的实体集合）或 None（有些接口直接在 ModelSpace 插入）
        segs = []
        try:
            # 如果 res 本身是一个 Python iterable
            for item in (res or []):
                segs.append(item)
        except Exception:
            # 当 res 为 None 时，尝试扫描最近创建的实体（不能可靠）
            segs = []

        # 如果没有通过返回的 res 获取到分段，尝试在 ModelSpace 中查找由 Copy 产生的后续对象
        # 为简单稳妥，我们也把 ent_copy 加入 created（若 ent_copy != ent）
        for s in segs:
            # 对每个分段，优先读取 Length 属性（圆弧也可能有 Length）
            try:
                L = getattr(s, "Length")
                if L is not None:
                    total += float(L)
                    created.append(s)
                    continue
            except Exception:
                pass
            # 对直线段，使用 StartPoint/EndPoint
            try:
                sp = getattr(s, "StartPoint", None)
                ep = getattr(s, "EndPoint", None)
                if sp and ep:
                    dx = float(ep[0]) - float(sp[0])
                    dy = float(ep[1]) - float(sp[1])
                    total += math.hypot(dx, dy)
                    created.append(s)
                    continue
            except Exception:
                pass
            # 对弧段，尝试用 Radius 和 StartAngle/EndAngle 计算
            try:
                r = getattr(s, "Radius", None)
                sa = getattr(s, "StartAngle", None)
                ea = getattr(s, "EndAngle", None)
                if r is not None and sa is not None and ea is not None:
                    # 角度可能以弧度给出
                    ang = abs(float(ea) - float(sa))
                    total += abs(float(r)) * ang
                    created.append(s)
                    continue
            except Exception:
                pass
            # 若无法识别，忽略该段
        # 3) 清理：擦除临时创建的实体（注意：如果 ent_copy is original ent, 不要擦除原实体）
        for obj in created:
            try:
                # 不擦除原对象
                if obj is ent:
                    continue
                obj.Erase()
            except Exception:
                pass
        # 如果 ent_copy was the original ent (we didn't want that), we didn't erase it above
        return total, (len(created) > 0)
    except Exception as e:
        # 尝试兜底读 Length
        try:
            return float(ent.Length), False
        except Exception:
            # 最终失败
            # 打印 traceback 以便调试
            traceback.print_exc()
            return None, False

def get_polyline_coords_and_length(ent, doc=None):
    """
    主要对外函数：先尝试直接取顶点坐标（返回 flat coords list），并计算长度；
    如果无法得到顶点，则用 explode 回退计算长度。
    返回 (coords_flat_or_None, length_or_None, used_explode_flag)
    """
    # 1) 直接按常见属性名尝试
    coords = try_read_coords_by_props(ent)
    if coords:
        length = poly_length_from_coords_flat(coords)
        return coords, length, False

    # 2) 尝试 CastTo 常见接口
    coords = try_cast_and_read(ent)
    if coords:
        length = poly_length_from_coords_flat(coords)
        return coords, length, False

    # 3) Explode 回退（需要提供 doc 用于清理）
    if doc is None:
        # 尝试获取 doc（实体有 Document 属性）
        try:
            doc = ent.Document
        except Exception:
            doc = None
    length, used_temp = get_length_via_explode(ent, doc)
    if length is not None:
        return None, length, used_temp

    # 4) 最后兜底：尝试读取 Length 属性直接返回
    try:
        L = getattr(ent, "Length")
        return None, float(L), False
    except Exception:
        return None, None, False



import win32com.client as win32

def ensure_acad():
    # 早绑定：生成并加载类型库（比 Dispatch 稳定，属性/方法更齐全）
    acad = win32.gencache.EnsureDispatch("AutoCAD.Application")
    return acad

def cast_polyline(ent):
    """把选到的多段线实体转成更具体的接口，方便取专有属性。"""
    name = ent.ObjectName  # 如 'AcDbPolyline' / 'AcDb2dPolyline' / 'AcDb3dPolyline'
    if name == "AcDbPolyline":     # 轻量 2D 多段线
        try:
            return win32.CastTo(ent, "IAcadLWPolyline")
        except Exception:
            return ent  # 退化为动态调度
    elif name == "AcDb2dPolyline":
        try:
            return win32.CastTo(ent, "IAcadPolyline")  # 旧 2D 多段线
        except Exception:
            return ent
    elif name == "AcDb3dPolyline":
        try:
            return win32.CastTo(ent, "IAcad3DPolyline")
        except Exception:
            return ent
    else:
        return ent

def get_ent_truecolor_rgb(ent):
    """
    返回实体实际显示颜色(尽力)：(r,g,b)
    如果是 ByLayer / ByBlock，会回退去查图层或给出 None。
    """
    try:
        tc = ent.TrueColor  # AcadAcCmColor
        # ColorMethod：0=ByBlock,1=ByLayer,2=ByColor,3=ByACI,4=ByRGB,5=ByPen
        cm = getattr(tc, "ColorMethod", None)
        if cm in (4,):  # ByRGB
            return (tc.Red, tc.Green, tc.Blue)
        if cm in (2, 3):  # 直接颜色/ACI
            # 直接给出 ACI 对应的 RGB 可能需要映射；简单返回 None
            return None
        if cm == 1:  # ByLayer
            try:
                layer = ent.Document.Layers.Item(ent.Layer)
                ltc = layer.TrueColor
                if getattr(ltc, "ColorMethod", None) == 4:
                    return (ltc.Red, ltc.Green, ltc.Blue)
            except Exception:
                pass
        # 其它情况回退
        return None
    except Exception:
        return None

def polyline_vertices(pl):
    """
    返回多段线顶点列表：
    - LWPolyline: 用 Coordinates（扁平数组 x,y[,x,y...]）
    - Polyline/3DPolyline: 用 Coordinate(i)
    """
    name = pl.ObjectName
    pts = []
    try:
        if name == "AcDbPolyline":  # LW
            coords = list(pl.Coordinates)  # [x1,y1, x2,y2, ...]
            # LWPolyline 固定在平面，Z 一般等于 Elevation
            z = getattr(pl, "Elevation", 0.0)
            for i in range(0, len(coords), 2):
                pts.append((coords[i], coords[i+1], z))
        elif name in ("AcDb2dPolyline", "AcDb3dPolyline"):
            n = pl.NumberOfVertices
            for i in range(n):
                p = pl.Coordinate(i)  # (x, y, z)
                pts.append((p[0], p[1], p[2]))
        else:
            # 兜底尝试 Coordinates
            coords = list(pl.Coordinates)
            z = getattr(pl, "Elevation", 0.0)
            for i in range(0, len(coords), 2):
                pts.append((coords[i], coords[i+1], z))
    except Exception:
        pass
    return pts

def inspect_polyline(ent):
    """打印多段线的关键信息：类型、图层、颜色、是否闭合、长度、起点、顶点列表等。"""
    pl = cast_polyline(ent)
    info = {}
    info["ObjectName"] = ent.ObjectName
    info["Layer"] = ent.Layer
    # Color: 256 表示 ByLayer；更可靠的是 TrueColor
    try:
        info["ColorIndex_or_Int"] = ent.Color
    except Exception:
        info["ColorIndex_or_Int"] = None
    info["TrueColorRGB"] = get_ent_truecolor_rgb(ent)

    # 是否闭合
    closed = None
    for key in ("Closed", "Closed2d", "Closed3d"):
        if hasattr(pl, key):
            try:
                closed = bool(getattr(pl, key))
                break
            except Exception:
                pass
    info["Closed"] = closed

    # 长度/面积
    for key in ("Length", "Area"):
        try:
            info[key] = float(getattr(pl, key))
        except Exception:
            info[key] = None

    # 顶点与起点
    verts = polyline_vertices(pl)
    info["VertexCount"] = len(verts)
    info["StartPoint"] = verts[0] if verts else None
    info["Vertices"] = verts

    return info
















#&&&% 重要基础知识（python、算法、数据结构）

#&&% ***🧠  抛出异常的处理逻辑

"""
1 try+except+finaly

成功后执行finaly后语句，失败后也会执行finaly后语句，必定预期整个代码段要做的就是finaly提供的功能


2 try+except+else

成功后执行else后语句，失败后不会会执行else后语句，必定预期成功情况下 要做的就是else提供的功能

3
try:
    risky_operation()
except SomeError as e:
    handle_error(e)
    # 这段只在 try 失败时运行
    try:
        do_special_on_failure()
    except Exception as e2:
        log("failure-only block failed:", e2)
else:
    # 成功时走这里
    do_if_success()
finally:
    cleanup()

do_special_on_failure() 就是专门为代码段失败时预期必定要执行的语句。它不同于直接放在except中，因为except可能的失败操作不能保证它被预期必然执行。

我们需要在某个操作失败时，必须执行某个操作，就使用这个逻辑



"""




#&&&% 脚本注释标准模块
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --

#📌 节点 1：形成正式“打印框线”并重排图形
#─────────────────────────────────────

"""🔍 输入:"""

"""🔧 关键函数："""
"""   ⮞ F1: `find_standard_printframes()` – 返回标准图框列表"""

"""🧠 处理逻辑:"""

"""📤 输出:"""


 

"""
🄐  🄑  🄒  🄓  🄔  🄕  🄖  🄗  🄘  🄙  🄚  🄛  🄜  🄝  🄞  🄟  🄠  🄡  🄢  🄣  🄤  🅅  🅆  🅇   🅈   🅉







⓿ ① ② ③ ④ ⑤ ⑥ ⑦ ⑧ ⑨ ⑩ ⑪ ⑫ ️⃣  

# ❖❖❖—— 重要函数：foo() ——❖❖❖

# ✺✺✺—— 关键逻辑：process_data()——✺✺✺

# ⚡⚡⚡—— 高优先：init_config() ——⚡⚡⚡

# ★★★—— 核心入口：main() ——★★★
常用符号推荐：

❖ （四角菱形）

✺ （放射状花瓣）

⚡ ★ （闪电）

★ （实心星）

✦ （中实菱形）

❗ （粗感叹号）

 
+--------------------------+
| 方法1：从打印图签块或打印线获取DX |
+--------------------------+
      ↓ 成功          ↓ 失败
+------------------+   +--------------------------+
| 从图签块重建框线 |   | 方法2：从"公司图签"图块提取DX |
+------------------+   +--------------------------+
           ↓                     ↓ 成功    ↓ 失败


+--------------------------------+
| 初始状态：未开始打印           |
+--------------------------------+
               ↓


        +---------+
        | F3函数  | ← 识别拟合框线对应图形
        +---------+
             ↓
 
📌 节点 1：确定“准打印框线”
─────────────────────────────────────
🔁 初始化阶段：
    ⮞ 清空 "准打印框线" 图层
    ⮞ 若已存在旧对象，则全部删除

🔍 提取打印框候选集合 DX 的三种方法：

① 方法一（推荐方式）
    ⮞ 查找 “图签块” + “打印框线” 块
    ⮞ 若存在图签块：
        ⮞ 提取包围盒 → 重绘为矩形 → 存入“准打印框线”
        ⮞ 清除原“打印框线”对象
        ⮞ ✅ 完成本节点目标


📦 输出：DX 对象集合 + “准打印框线”图层绘制

------------------------------------------------------

📌 节点 2：形成正式“打印框线”并重排图形
─────────────────────────────────────
🔍 输入：DX 集合（来自上节点）

🧠 图形分析与处理：
    ⮞ 识别标准打印框（F1）
    ⮞ 识别拟合打印框（F2）

🧱 图框替换与排版：
    ⮞ 拟合框 → 替换为最近 A1 标准比例图框
    ⮞ 重新排列打印图框，避免重叠

📦 输出：
    ⮞ 正式“打印框线”写入目标图层
    ⮞ 准打印框图层清空
    ⮞ Handle → 信息存储用于后续匹配

🔧 函数：
    ⮞ F1: 标准框识别
    ⮞ F2: 拟合框识别
    ⮞ F3: 图形识别归属（局部）
    ⮞ F4: 图形重排整体逻辑

┌────────────┐⮞ F1: 标准框识别 
│ 节点 1：   │
│ 准打印框线 │
└────┬───────┘
     │ 清空图层
     ▼
┌────────────┐
│ 方法一：图签块或框线 │◄─────┐
└────┬──────────────┘      │
     │ 找不到             │
     ▼                    │

┌────────────┐
│ 排序 + 移动图形 F3/F4 │
└────┬────────┘
     ▼
┌────────────────────┐
│ 写入“打印框线”图层 │
└────────────────────┘

📌 节点 X：<简要标题>
─────────────────────────────────────
🔍 输入：
    ⮞ 输入数据或图元说明（如 DX 集合）
    ⮞ 来源说明（如来自上游节点）
    
🧠 处理逻辑：
    ⮞ 步骤说明 1（如：识别标准打印框）
    ⮞ 步骤说明 2（如：判断拟合框）

🔧 关键函数：
    ⮞ F1: 函数名 – 功能说明
    ⮞ F2: 函数名 – 功能说明

📤 输出：
    ⮞ 输出数据目标（如写入正式图层）
    ⮞ 副作用（如清除中间层、更新 handle 记录）

⚠️ 异常处理：
    ⮞ 未匹配图框 → 输出警告
    ⮞ 多图重叠 → 排序并修正重排

🗂 示例调用：
    process_printframes(DX)


📌	节点起始	📌 节点 2：形成正式“打印框线”并重排图形
────────────	分割线	──────── 节点分隔 ────────
🧭	逻辑导向	🧭 下一步识别拟合图框
🔀	路径分支	🔀 分流处理：标准框 vs 拟合框
⤵ / ⤴	子流程跳转	⤵ 进入子流程 F3
🔂	循环处理	🔂 针对所有 F2 框执行替换逻辑

🔧 功能或模块说明
符号	用途	示例
🔧	函数定义	🔧 F1: 标准框识别函数
🧠	处理逻辑	🧠 图形分析与处理
🧪	测试/调试块	🧪 临时打印 bbox 信息
📦	输出结果	📦 输出图层对象
📥 / 📤	输入/输出	📤 输出写入 LAYER_PRINT 正式图框
🧱	数据结构/实体对象	🧱 实体对象替换操作

🔍 分析与识别相关
符号	用途	示例
🔍	搜索/识别	🔍 查找符合 A1 比例的图框
🧬	特征提取	🧬 提取图形外包盒坐标与宽高比
🔎	精细识别	🔎 判断是否为完整封闭矩形

✅ 状态反馈
符号	用途	示例
✅	成功	✅ 图框替换完成
⚠️	警告	⚠️ 找不到合适图框匹配项
❌	错误	❌ 无法识别多段线边界
⏳	等待	⏳ 等待 CAD 对象响应

🔁 控制结构（可嵌入注释中）
符号	用途	示例
⬅️ ➡️ ⬆️ ⬇️	方位、移动方向	⬆️ 上移图框 1000 单位避免重叠
🔄	重复执行	🔄 对所有图框执行一次旋转检测
↪️	返回	↪️ 返回原图层处理状态

📊 数据操作与计算
符号	用途	示例
📐	几何计算	📐 计算图框宽高比
📏	测量	📏 测量距离与偏移
🧮	数值处理	🧮 总计图框数量 = len(F1) + len(F2)



"""


#&&%  *** 常用功能代码块

## 1 如果某个操作流程出错，最多间隔1秒尝试3次即无视之，进入下一流程


##try:
##    feng_lines = stc("bianmulu_lp")
##    print(f"🗑 将 {len(split_pairs)} 条分裂线标记为红色 (尝试 {attempt})")
##    for A, B in split_pairs:
##        _, yA, _ = bbox_center_3(A)
##        _, yB, _ = bbox_center_3(B)
##        y_mid = (yA + yB) / 2.0
##
##        for ent in feng_lines:
##            try:
##                obj = dyn.Dispatch(ent._oleobj_)
##            except:
##                continue
##            if obj.ObjectName not in ("AcDbLine", "AcDbPolyline", "AcDb2dPolyline"):
##                continue
##            _, cy, _ = bbox_center_3(obj)
##            if abs(cy - y_mid) <= y_tol:
##                print(f"  🔴 标记 Handle={obj.Handle} centerY={cy:.1f}")
##                try:
##                    obj.Color = 1
##                    obj.Layer = "测试辅助"
##                except:
##                    try:
##                        obj.TrueColor = 0xFF0000
##                    except:
##                        print(f"    ⚠️ 上色失败 Handle={obj.Handle}")
##    # 如果执行到这里不出异常，就退出重试循环
##    break
##except Exception as e:
##    print(f"⚠ 分裂线标记第 {attempt} 次失败: {e}")
##    if attempt < 3:
##        time.sleep(1)
##    else:
##        print("ℹ 分裂线标记跳过，进入下一流程")


def kill_wps(verbose: bool = False):
    """
    结束所有 WPS/金山办公相关进程，特别是 wpspdf.exe。
    verbose=True 时打印被终止的映像名。
    """
    # 需要关掉的映像清单（全部小写）
    targets = {
        "wps.exe",          # WPS 主进程
        "wpspdf.exe",       # ★ PDF 预览器
        "wpp.exe",          # 演示
        "et.exe",           # 表格
        "ksolaunch.exe",    # 启动器
        "wpscloudsvr.exe",  # 云同步
        "wpsupdate.exe",    # 更新器
        "wpscenter.exe",    # 消息中心
    }

    killed = set()
    for name in targets:
        # /T 连子进程一起；/F 强制；/IM 按映像名
        rc = subprocess.call(
            f'taskkill /T /F /IM {name}',
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        if rc == 0:      # 返回 0 表示确有此进程并已结束
            killed.add(name)

    if verbose:
        if killed:
            print("✓ 已结束进程:", ", ".join(sorted(killed)))
        else:
            print("ℹ️ 未检测到 WPS 相关进程")


#&&% 人机结合的函数


"""

通过input输入等待，在复杂环节引入人工操作，而其余部分仍交给计算机自动完成
因为最终的目的是节省时间
从普通图签块制作标准属性块就是一个典范
"""

#&&% 函数运行节点控制


"""
通过打印对象的Handle我们能很方便检测函数运行过程中的结果，而不需要从函数返回值或麻烦的调试中检查错误
print(f"从块{blk.Name}中对象按比例和面积特征过滤得到的com对象Handle句柄")

print_coms_handle(filtered)

"""


#&&% 函数状态控制

# 函数运行状态的标准化重构(重装软件使重启状态标准化)

"""
如果重启天正CAD系统5次都未能运行，则系统将自己重新安装天正和CAD

天正画图错误被强制中断CAD进程消息窗口
activate_window_by_title("图形修复", click_titlebar= True)
✅ 已激活窗口：“图形修复” 位置(1089,562) 大小366×213
(1089, 562, 366, 213)
同一测试的其窗口“关闭”按钮的点击位置 (1394, 724)，点击后即正常运行



"""

# 函数时间控制

"""
如果一个系统函数运转受阻，将会被时间监控中断进程，天正CAD系统会被中断重启


"""
def cad_zhongduan_ceshi():

    i=0

    while True:

        LP=[(0, i, 0.0), (90509.17232155753, 38080.72757883748, 0.0), (72659.40193916814, 48550.64173869454, 0.0), (61087.391734572855, 35617.55453017057, 0.0), (62182.914098626905, 23340.76471079199, 0.0), (43443.713317261485, 21669.276832727846, 0.0), (32018.9079127094, 29207.203355399004, 0.0), (22430.211826160492, 19081.580148860856, 0.0)]

        draw_lwpolyline(
    LP,
    layer_name= "0",
    width = 0.0,
    color= 256,
    closed = False
)

    i=i+10000




# 函数运行消息打印和结构标准控制

"""
在系统的更高级层面，函数不需要显示底层函数的运转信息，从提速的角度讲应该去掉print语句，但这个可以在稳定运转一段时间之后处理

"""


# 函数逻辑结构

"""
应该设置函数的明确逻辑构成，在各个环节打印消息用于分析运转是否错误。通过定义内部函数或小函数，建立正确的概念和模块，使得人工智能系统更容易被驾驭

"""


# 函数案例生成

"""
生成足够多的测试文件，节省程序测试时间，并通过完整案例使函数逻辑更合理

"""

# 打印案例生成器

"""
generate_test_cases(
    r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/自动化(动态)/目录图签/生成测试文件/目录图签.dwg",
    num = 3,
    jianju_x = 10000,
    juli_tukuang = 6000,
    juli_y = 400000,
)

"""

# 文字候选
TEXT_OPTIONS = [
    "一层平面图", "二层平面图", "1-18轴立面图", "2-2剖面图",
    "1-1剖面图", "楼梯大样图", "门窗详图图"
]

# 辅助：创建文字实体
def create_text(txt: str, position: Tuple[float, float, float], height: float, layer: str):
    vpt = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, list(position))
    ent = mp.AddText(txt, vpt, height)
    try:
        ent.Layer = layer
    except Exception:
        pass
    return ent

# 随机绘制基本图形，置于layer_tu
def draw_random_shape(xmin: float, xmax: float, ymin: float, ymax: float, layer_tu: str):
    """
    在给定范围内随机绘制一点、线段、圆或正多边形，
    并放置到指定图层。
    """
    typ = random.choice(["point", "line", "circle", "polygon"])
    shape = None
    if typ == "point":
        pt = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        shape = draw_point(pt)
    elif typ == "line":
        p1 = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        p2 = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        shape = draw_line(p1, p2)
    elif typ == "circle":
        center = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        r = random.uniform(1000, min(xmax - xmin, ymax - ymin) / 4)
        shape = draw_circle(center, r)
    else:
        center = (random.uniform(xmin, xmax), random.uniform(ymin, ymax), 0)
        r = random.uniform(1000, min(xmax - xmin, ymax - ymin) / 4)
        sides = random.randint(3, 8)
        shape = draw_regular_polygon(center, r, sides)
    # 设置图层
    try:
        if shape is not None:
            shape.Layer = layer_tu
    except Exception:
        pass
def generate_test_cases(
    output_path: str,
    num: int = 3,
    jianju_x: float = 10000,
    juli_tukuang: float = 6000,
    juli_y: float = 400000,
):
    """
    生成测试 DWG：多行图框+图形+边缘文字，文字区域由 get_frame_edge 计算。
    """
    sizes = {"A3": (42000, 29700), "A2": (59400, 42000), "A1": (84000, 59400)}
    seq_fixed = [
        ("A3", False), ("A2", False), ("A1", False),
        ("A3", True),  ("A2", False), ("A1", True),
        ("A2", False), ("A2", False),
    ]

    layer_kuang = "测试绘制_图框"
    layer_tu    = "测试绘制_图形"
    layer_wz    = "测试绘制_文字"
    #不覆盖从当前时间给带路径文件名命名
    output_path = rename_time(output_path)
    
    create_new_dwg_file_no(output_path)
    time.sleep(1)

    base_x = random.uniform(0, 1000000)
    base_y = random.uniform(0, 1000000)

    for row in range(num):
        y0 = base_y + row * juli_y
        x = base_x
        seq = seq_fixed if row == 0 else [
            (random.choice(list(sizes.keys())), random.choice([False, True]))
            for _ in seq_fixed
        ]

        for name, vertical in seq:
            w, h = sizes[name]
            if vertical:
                pts = [(x, y0, 0), (x, y0 + w, 0), (x + h, y0 + w, 0),
                       (x + h, y0, 0), (x, y0, 0)]
            else:
                pts = [(x, y0, 0), (x + w, y0, 0), (x + w, y0 + h, 0),
                       (x, y0 + h, 0), (x, y0, 0)]
            # 绘制并获取多段线实体
            poly_ent = draw_polyline(pts, layer_name=layer_kuang, tol=0.5, width=20, color=1)
            # 计算边域四角点
            edge_pts = get_frame_edge(poly_ent, juli_tukuang)
            if not edge_pts:
                continue
            ex0, ey0, _ = edge_pts[0]
            _, ey1, _ = edge_pts[1]
            ex1, _, _ = edge_pts[2]

            # 图框内部随机图形，保留内边缘
            # 计算内部区域
            if vertical:
                inner_xmin, inner_xmax = x, x + h
                inner_ymin, inner_ymax = y0 + juli_tukuang, y0 + w
                x_shift = h
            else:
                inner_xmin, inner_xmax = x, x + w
                inner_ymin, inner_ymax = y0, y0 + h - juli_tukuang
                x_shift = w
            # 随机图形
            for _ in range(random.randint(5, 10)):
                ensure_layer_current(layer_name="测试绘制_图形")                
                draw_random_shape(
                    inner_xmin, inner_xmax,
                    inner_ymin, inner_ymax,
                    layer_tu
                )

            # 边域文字：1-6 文本，分 1-3 排，限定在 edge 区域
            n_texts = random.randint(1, 6)
            rows_t = random.randint(1, min(3, n_texts))
            counts = [n_texts // rows_t] * rows_t
            for i in range(n_texts % rows_t):
                counts[i] += 1
            text_height = 300
            line_spacing = text_height + 100

            for r, cnt in enumerate(counts):
                choices = random.sample(TEXT_OPTIONS, cnt)
                line_str = '、'.join(choices)
                if not vertical:
                    tx = random.uniform(ex0, ex1 - text_height)
                    ty = ey1 - (r + 1) * line_spacing
                    ty = min(max(ty, ey0), ey1 - text_height)
                else:
                    ty = random.uniform(ey0 + text_height, ey1)
                    tx = ex0 + r * line_spacing
                    tx = min(max(tx, ex0), ex1 - text_height)
                print(f"[DEBUG] 写入文字 '{line_str}' 在 ({tx:.2f},{ty:.2f}) 图层 {layer_wz}")
                ensure_layer_current("测试绘制_文字")                
                create_text(line_str, (tx, ty, 0), text_height, layer_wz)

            x += x_shift + jianju_x

    time.sleep(1)
    doc.SendCommand("Z\nE\n")
    savefile()
    guanbifile()

def get_frame_edge(poly_ent, juli_tukuang: float = 6000.0):
    """
    给定一个闭合多段线实体，判断其横向/竖向，并返回边域四角坐标。

    :param poly_ent:    闭合的 polyline 实体（AcDbPolyline）
    :param juli_tukuang:边域宽度
    :return: [(x1,y1,0), (x1,y2,0), (x2,y2,0), (x2,y1,0)]
    """
    try:
        (pmin, pmax) = poly_ent.GetBoundingBox()
    except Exception as e:
        print(f"[ERROR] 获取包围盒失败: {e}")
        return None
    xmin, ymin, _ = pmin
    xmax, ymax, _ = pmax
    w = xmax - xmin
    h = ymax - ymin
    if w >= h:
        orientation = 'horizontal'
        # 横向：边域在右侧
        x1 = xmax - juli_tukuang
        x2 = xmax
        y1 = ymin
        y2 = ymax
    else:
        orientation = 'vertical'
        # 竖向：边域在底部
        x1 = xmin
        x2 = xmax
        y1 = ymin
        y2 = ymin + juli_tukuang
    print(f"[DEBUG] 框为 {orientation}，边域坐标: ({x1:.2f},{y1:.2f}), ({x1:.2f},{y2:.2f}), ({x2:.2f},{y2:.2f}), ({x2:.2f},{y1:.2f})")
    return [(x1, y1, 0), (x1, y2, 0), (x2, y2, 0), (x2, y1, 0)]

#&&% 打印设置和pyautogui基本操作

"""
run_py("测鼠标位置.py")#测位置

pyautogui.moveTo(1415, 796)#定位

pyautogui.write("EXPLODE", interval=0.1)#写文字

pyautogui.press("enter")#回车

pyautogui.doubleClick()#双击

pyautogui.hotkey('ctrl', 'a')#全选

pyautogui.press('delete')#删除

# 然后从该位置向上滚动 200
pyautogui.scroll(200)

# 假设已经 moveTo(x,y)
pyautogui.mouseDown()
pyautogui.dragRel(200, 100, duration=0.3)  # 向右200，下100的框选
pyautogui.mouseUp()

如果你的“预先有的数据”不是全在一个文本框里，而是一个列表/表格中的多行，你可以
在单击一次之后，用 pyautogui.dragTo() 或 pyautogui.dragRel() 做框选
然后按 delete

# 方法一：直接右键
pyautogui.click(button='right')
# 或者使用快捷函数
# pyautogui.rightClick()

# 如果需要在弹出的上下文菜单里选择某项，比如向下移动两次再回车：
pyautogui.press('down', presses=2, interval=0.1)
pyautogui.press('enter')


# 发送 Ctrl+1
pyautogui.hotkey('ctrl', '1')
# 发送 Ctrl+9
pyautogui.hotkey('ctrl', '9')
# 发送 Shift+PrintScreen
# 注意 PyAutoGUI 里 PrintScreen 键通常叫 'printscreen' 或 'prtsc'
pyautogui.hotkey('shift', 'printscreen')

# Ctrl+1 手动版
pyautogui.keyDown('ctrl')
pyautogui.press('1')
pyautogui.keyUp('ctrl')
pyautogui.press("esc")

currentLayout = acad.ActiveDocument.ActiveLayout
currentLayout.CanonicalMediaName = "ISO_A1_(841.00_x_594.00_MM)"
currentLayout.CanonicalMediaName#查看当前图纸设置


标准A0,A1,A2,A3的可打印区域是可单独修改的，其0-0-0-0效果已经打印出来 


"""

dy_yonghu=[

                                           #A0 1189.00_x_841.00

    ("1337.63","841.00","17","17","5","5"),#A0+1/8


    ("1486.25","841.00","17","17","5","5"),#A0+1/4

                                           #A1 841.00_x_594.00

    ("1051.25","594.00","17","17","5","5"),#A1+1/4


    ("1261.50","594.00","17","17","5","5"),#A1+1/2


    ("1471.75","594.00","17","17","5","5"),#A1+3/4

                                           #A2 594.00_x_420.00


    ("742.50","420.00","17","17","5","5"), #A2+1/4

                                         

    ("891.00","420.00","17","17","5","5"), #A2+1/2


    ("1039.50","420.00","17","17","5","5") #A2+3/4


                                           #A3 420.00_x_297.00
]



def 批量设置用户打印尺寸(dy_yonghu):
    """
    先连接，再人工打开ctr+P选择DWG TO PDF打印后再运行此函数

    """

    #按用户给定数据批量生成用户打印尺寸
    for i in range(0,len(dy_yonghu)):
        dyshu = dy_yonghu[i]
 
        pyautogui.moveTo(1331, 569)#点击打印窗口的 特性 按钮
        pyautogui.click(1331, 569)        
        time.sleep(2)


        pyautogui.moveTo(1189, 627)#点击打印窗口的 自定义图纸尺寸 按钮
        pyautogui.click(1189, 627)        
        time.sleep(2)

        pyautogui.moveTo(1415, 796)#点击打印窗口的 编辑 按钮
        pyautogui.click(1415, 796)        
        time.sleep(2)

        pyautogui.moveTo(1161, 690)#点击宽度窗口位置
        pyautogui.click(1161, 690)        
        time.sleep(2)

        pyautogui.doubleClick()#选预设数据
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[0], interval=0.1)#写宽度
        time.sleep(2)

        pyautogui.moveTo(1171, 745)#点击高度窗口位置
        pyautogui.click(1171, 745)        
        time.sleep(2)

        pyautogui.doubleClick()#选预设数据
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[1], interval=0.1)#写高度
        time.sleep(2)
        
        pyautogui.moveTo(1452, 915)#点击下一页
        pyautogui.click(1452, 915)        
        time.sleep(2)

        pyautogui.moveTo(1144, 684)#点击上边距窗口位置
        pyautogui.click(1144, 684)        
        time.sleep(2)

        pyautogui.doubleClick()#选预设数据
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[2], interval=0.1)#写上边距
        time.sleep(2)

        pyautogui.moveTo(1144, 729)#点击下边距窗口位置
        pyautogui.click(1144, 729)        
        time.sleep(2)


        pyautogui.doubleClick()#选预设数据
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[3], interval=0.1)#写下边距
        time.sleep(2)

        pyautogui.moveTo(1143, 770)#点击左边距窗口位置
        pyautogui.click(1143, 770)        
        time.sleep(2)

        pyautogui.doubleClick()#选预设数据
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[4], interval=0.1)#写左边距
        time.sleep(2)

        pyautogui.moveTo(1147, 809)#点击右边距窗口位置
        pyautogui.click(1147, 809)        
        time.sleep(2)

        pyautogui.doubleClick()#选预设数据
        pyautogui.press('delete')
        time.sleep(1)
        pyautogui.write(dyshu[5], interval=0.1)#写右边距
        time.sleep(2)

        pyautogui.moveTo(1443, 919)#点击下一页
        pyautogui.click(1443, 919)        
        time.sleep(2)

        pyautogui.moveTo(1443, 922)#点击下一页
        pyautogui.click(1443, 922)        
        time.sleep(2)

        pyautogui.moveTo(1460, 921)#点击完成
        pyautogui.click(1460, 921)        
        time.sleep(2)


        pyautogui.moveTo(1239, 920)#点击确定
        pyautogui.click(1239, 920)        
        time.sleep(2)

##确保删除对象

def safe_delete(ob, retries: int = 5, delay: float = 1.0) -> bool:
    """
    尝试删除 CAD 对象 ob，最多重试 retries 次，每次间隔 delay 秒。
    只捕获 COM 错误，成功返回 True，否则返回 False。
    """


    for attempt in range(1, retries + 1):
        try:
            ob.Delete()
            return True
        except pywintypes.com_error:
            time.sleep(delay)
    return False





#&&&%  * 高亮选择转隐性移动区域内全部对象


"""
高亮选择窗口操作更可靠从而更快，但要考虑窗口不能挡

"""

def move_entities_in_region(coms, target=(0,0,0), ty=1, max_iter=3):
    """
    将 `coms` 对象的包围盒内所有实体，沿向量 (target - 左下角) 移动，
    每轮等待 `ty` 秒，最多循环 `max_iter` 次。

    参数：
      coms      -- 支持 GetBoundingBox() 的 COM 对象（如多段线、块参照等）
      target    -- 目标基点坐标，默认为 (0,0,0)
      ty        -- 每轮移动后等待秒数
      max_iter  -- 最多尝试的轮数
    """
    # 1. 读取包围盒，计算左下角 (x1,y1) 和右上角 (x2,y2)
    p1, p2 = coms.GetBoundingBox()
    x1, y1 = min(p1[0], p2[0]), min(p1[1], p2[1])
    x2, y2 = max(p1[0], p2[0]), max(p1[1], p2[1])

    # 2. 计算位移向量 Δ = target - (x1,y1)
    dx = target[0] - x1
    dy = target[1] - y1
    dz = target[2] - 0.0

    for i in range(1, max_iter + 1):
        # 清空显性选择集
        try:
            doc.PickfirstSelectionSet.Clear()
        except Exception:
            pass

        # 3. 在 CAD 里用窗口选择高亮区域内对象
        highlight_entities_in_window(x1, y1, x2, y2)

        # 4. 拿到显性选择集里的实体
        pick = doc.PickfirstSelectionSet
        count = pick.Count if hasattr(pick, 'Count') else len(list(pick))
        if count == 0:
            print(f"✅ 第 {i} 轮：区域已清空，停止。")
            break

        print(f"♻️ 第 {i} 轮：检测到 {count} 个对象，正在移动…")
        # 5. 对每个实体按计算好的向量执行 Move
        for ent in pick:
            try:
                # 从 (x1,y1,0) 移动到 (x1+dx, y1+dy, dz)
                ent.Move(vtpnt(x1, y1, 0.0),
                         vtpnt(x1 + dx, y1 + dy, dz))
            except Exception as e:
                print(f"  ⚠ 对象 {ent.Handle} 移动失败：{e}")

        # 6. 等待 CAD 完成命令再进入下一轮
        time.sleep(ty)
    else:
        print("⚠ 达到最大迭代次数，可能仍有残留对象。")





###矩形标准化及整体优化
"""
使用parse_rectangle_points函数接受用户对矩形对角点两种方式的自由理解,使用它

改造已有函数，从而修复这种因为对角点不同理解的混乱

jd()应该在连接时就设置，控制绘图精度

有些函数处理还可以优化，显性选择不要在窗口上遮挡，不要动窗口

稳定性和强壮性随使用反馈变化

复制一个参照版本便于查阅函数

以数据库为基础不断扩展数据，扩张整个绘图系统的想法是对的


应该有一个文件初始化函数，将处理文件转入标准设置状态，从精度，字体，墙线显示加粗，打印空间安全，打印配置，视图辅助环境等等。


允许输入一条多段线或多根多段线，改善函数接口，这种错误是合乎人类思维的


在两个字典之间定义函数，通过字典查找确定信息








"""


#&&% 字典的生成、使用、扩展

"""
##{
##  # 顶层键：每个图签框的 BlockReference.Handle（字符串）
##  "2BC3": {
##      # ① 这条记录对应的打印框信息，统一放到一个子字典里
##      "frame_info": {
##          # 4 个角点（3D）    
##          "corners": [
##              (x1, y1, z1),
##              (x1, y2, z1),
##              (x2, y2, z1),
##              (x2, y1, z1),
##          ],
##          # ISO 编号或名称
##          "format": "ISO_A2_(594.00_x_420.00_MM)",
##          # 表示比例
##          "scale":  "1:100",
##          # 纸张规格（A2/A3/A1…）
##          "paper_size": "A2",
##      },
##      # ② 插入成功后的标题块引用  
##      "title_block_handle": "2C10",         # Handle 字符串
##      "title_block_entity": <COMObject ...> # （可选，方便立刻操作）
##  },
##
##  "2BC4": {
##    # ……
##  }
##}
##
##在上面示例里：
##
##顶层 key 用 Handle（字符串形式）来索引每一条记录；
##
##frame_info：一个“打印框”相关的子字典，里面再细分
##
##corners：四个 3D 坐标点
##
##format：图幅名称（ISO_…）
##
##scale：打印比例（比如 "1:100"）
##
##paper_size：规格（比如 "A2"、"A3"）
##
##title_block_handle：新插入的图签块引用的 Handle，用于快速查回或二次处理；
##
##title_block_entity（可选）：保留一份 COM 对象，以免你后面还想马上调用它的属性／方法。
##
##为什么这样命名更易维护？
##语义明确
##
##format、scale、paper_size 一看就知道分别代表图幅、比例、规格。
##
##层次分明
##
##所有跟“打印框”相关的内容都放 frame_info，以后要扩展（比如加 “偏移量 offset”）也可以放进去。
##
##与之平行的是标题块信息 (title_block_…)。
##
##Handle 做主键
##
##外层直接用 BlockReference.Handle 字符串，既不易冲突，也方便序列化、存盘。
##
##保留实体或句柄
##
##存下 title_block_entity 可直接调用；存下 title_block_handle 可在重载文档后重新查找实例。
##
##你可以据此把已有的 zd_tuqian 做一次重命名／重构，例如：
##
##python
##复制
##编辑
##def normalize_tuqian_dict(raw: dict) -> dict:
##    
##    将原始的 zd_tuqian 转换为更清晰的结构：
##    {
##      handle: {
##        "frame_info": {
##          "corners":    …,
##          "format":     …,
##          "scale":      …,
##          "paper_size": …,
##        },
##        "title_block_handle": …,
##        "title_block_entity": …
##      }, …
##    }
##    
##    new = {}
##    for handle, info in raw.items():
##        fi = info["frame_info"]
##        new[handle] = {
##            "frame_info": {
##                "corners":    fi["corners"],
##                "format":     fi["drawing_frame"],
##                "scale":      fi["ratio"],
##                "paper_size": fi["spec"],
##            },
##            "title_block_handle": getattr(info["title_block"], "Handle", None),
##            "title_block_entity": info["title_block"],
##        }
##    return new
##这样，后续所有基于 zd_tuqian 读取比例、规格等，都能写成：
##
##entry = zd_tuqian[handle]
##scale     = entry["frame_info"]["scale"]
##paper_sz  = entry["frame_info"]["paper_size"]
##blk_handle = entry["title_block_handle"]
##希望这个命名规范能帮助你把字典结构理得更清晰，也更容易维护和扩展。
##
##
##entities = [info["entity"] for info in res.values()]
##specs     = [info["spec"]   for info in res.values()]
##ratios    = [info["ratio"]  for info in res.values()]
##
##info =bindy.get("2D07")
##info['title_block']
##字典的产生，保存，调用非常有价值，它将为我们存取大量信息，也改变了整个程序结构
##
### 拿到某个 title_block_handle 的 info
##info = bind_dict[handle]
##
### 移动图签块
##tb = info['title_block']
##tb.Move(vtpnt(0,0,0), vtpnt(0,10000,0))
##
### 移动对应的打印框 polyline
##pl = info['frame_info']['entity']
##pl.Move(vtpnt(0,0,0), vtpnt(0,10000,0))


"""

def 圆点(tz=1):


    """
    控制点的显示

    """

    zhi=0

    if tz ==1:

        zhi=35

    acad.ActiveDocument.SetVariable("PDMODE", zhi)#就是点最好的圆加十字形显示



def 图纸背景(zhi = 16777215 ):

    acad.ActiveDocument.Application.preferences.Display.GraphicsWinModelBackgrndColor = zhi#0即变成黑色





#&&&%  *** 按区域调整视图

def  shitu_region(x1,y1,x2,y2):

    """
    按按对象外包盒调整视图
    使用时可shitu_region(*p),p=(x1,y1,x2,y2)

    """

    h=0.3*(abs(x1-x2)+abs(y1-y2))/2

    # 1️⃣ 缩放视图到合适窗口
    zoom_cmd = (
        f"_.ZOOM\n"      # 调用 Zoom
        f"_W\n"          # 窗口选项
        f"{x1-h},{y1-h}\n"   # 第一个角点
        f"{x2+h},{y2+h}\n"   # 第二个角点
    )
    doc.SendCommand(zoom_cmd)

#按对象外包盒调整视图

def  shitu_entity(obj):

    """
    按按对象外包盒调整视图

    """
    p1,p2=obj.GetBoundingBox()

    x1,y1=p1[0],p1[1]

    x2,y2=p2[0],p2[1]

    h=0.3*(abs(x1-x2)+abs(y1-y2))/2

    # 1️⃣ 缩放视图到合适窗口
    zoom_cmd = (
        f"_.ZOOM\n"      # 调用 Zoom
        f"_W\n"          # 窗口选项
        f"{x1-h},{y1-h}\n"   # 第一个角点
        f"{x2+h},{y2+h}\n"   # 第二个角点
    )
    doc.SendCommand(zoom_cmd)


def ensure_list(obj, element_type=None):
    """
    如果 obj 已经是列表，则原样返回；
    否则，如果没有指定 element_type，或 obj 是 element_type 的实例，
      则将其包装成单元素列表返回；
    否则抛出 TypeError。

    :param obj: 单个元素或元素列表
    :param element_type: 列表元素应有的类型（可选）
    :return: 元素列表
    """
    # 1) 如果已经是列表，直接返回
    if isinstance(obj, list):
        return obj

    # 2) 要么不限制类型，要么 obj 是指定类型
    if element_type is None or isinstance(obj, element_type):
        return [obj]

    # 3) 其它情况报错
    raise TypeError(
        f"期望类型 {element_type.__name__} 或 List[{element_type.__name__}]，"
        f"但收到 {type(obj).__name__}"
    )


#  主函数
#  (1)
# 某某功能作用的函数

#  该函数系列包括如下一些函数

"""
文件处理之前执行
collect_all_texts()
将所有文字放入天正"PUB_TEXT"图层，后续无需再去重复

执行大文件变文件夹+每张图一个文件，减少不必要的选择运算


"""


#&&% 窗口键盘控制
"""
# 1. 截取全屏，返回一个 PIL.Image 对象
img = pyautogui.screenshot()

# 2. 如果想直接保存到文件：
pyautogui.screenshot('full_screen.png')

# 3. 只截取屏幕的一部分（x, y, width, height）
#    例如：从左上角 (100,100) 开始，截取 300×200 的区域
region_img = pyautogui.screenshot(region=(100, 100, 300, 200))
region_img.save('partial.png')


"""

#GIF录屏

def record_screen_gif(
    output_path: str,
    duration: float = 5.0,
    fps: int = 10,
    region: tuple[int,int,int,int] | None = None
):
    """
    录制屏幕并保存为 GIF。

    :param output_path: 输出 GIF 文件路径，比如 "demo.gif"
    :param duration:    录制时长（秒）
    :param fps:         帧率（每秒截多少帧）
    :param region:      可选，(left, top, width, height) 只录制这一区域
    """
    import imageio 

    frames = []
    interval = 1.0 / fps
    end_time = time.time() + duration

    print(f"▶ 开始录制：时长={duration}s，帧率={fps}fps，区域={region or '全屏'}")
    while time.time() < end_time:
        img = pyautogui.screenshot(region=region)  # PIL.Image 对象
        frames.append(img)
        time.sleep(interval)

    print(f"🛑 录制结束，共捕获 {len(frames)} 帧，正在合成 GIF……")
    # imageio 能直接接受 PIL.Image
    imageio.mimsave(output_path, frames, fps=fps)
    print(f"✅ 已保存为 {output_path}")


#动画录屏










def minimize_all_windows():
    """
    模拟按下 Win+M，将所有窗口最小化。

    """

    import ctypes

    user32 = ctypes.windll.user32
    VK_LWIN = 0x5B   # 左 Win 键
    VK_M    = 0x4D   # M 键
    KEYEVENTF_KEYUP = 0x2

    # 按下 Win
    user32.keybd_event(VK_LWIN, 0, 0, 0)
    # 按下 M
    user32.keybd_event(VK_M,    0, 0, 0)
    # 松开 M
    user32.keybd_event(VK_M,    0, KEYEVENTF_KEYUP, 0)
    # 松开 Win
    user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)

    # 给系统一点反应时间
    time.sleep(0.1)



#&&%#控制CAD屏幕窗口在左上角


def set_autocad_window_to_top_left(resize_half: bool = True):
    """
    将 AutoCAD 窗口还原并移动到屏幕左上角，可选将其调整为半屏大小。
    """
    # 1️⃣ 找到可见或最小化的 AutoCAD 窗口
    windows = [w for w in gw.getWindowsWithTitle('AutoCAD')]
    if not windows:
        print("❌ 未找到 AutoCAD 窗口！")
        return
    win = windows[0]

    # 2️⃣ 如果窗口最小化，先还原
    if win.isMinimized:
        win.restore()
        time.sleep(0.3)

    # 3️⃣ 激活窗口，确保我们操作的是真正的前台窗口
    try:
        win.activate()
    except Exception:
        # 有时 activate 也会失败，短暂等待再试
        time.sleep(0.2)
        win.activate()
    time.sleep(0.2)

    # 4️⃣ 移动到左上角 (0,0)
    win.moveTo(0, 0)
    time.sleep(0.2)

    # 5️⃣ 可选：将窗口调整为屏幕一半大小
    if resize_half:
        screen_w, screen_h = pyautogui.size()
        win.resizeTo(screen_w // 2, screen_h // 2)
        time.sleep(0.2)
        print(f"✅ AutoCAD 窗口已恢复并移动到左上角，尺寸设为 {screen_w//2} x {screen_h//2}")
    else:
        print("✅ AutoCAD 窗口已恢复并移动到左上角")



def l():

    set_autocad_window_to_top_left()



#&&% 更合理控制窗口函数
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --

def minimize_all_windows_d():
    """
    模拟 Win + D，将所有窗口最小化（切换）。
    """
    # VK_LWIN = 0x5B, VK_D = 0x44, KEYEVENTF_KEYUP = 0x2
    ctypes.windll.user32.keybd_event(0x5B, 0, 0, 0)       # 按下 Win
    ctypes.windll.user32.keybd_event(0x44, 0, 0, 0)       # 按下 D
    ctypes.windll.user32.keybd_event(0x44, 0, 0x2, 0)     # 松开 D
    ctypes.windll.user32.keybd_event(0x5B, 0, 0x2, 0)     # 松开 Win
    time.sleep(0.3)


def minimize_all_windows_m():
    """
    模拟按下 Win+M，将所有窗口最小化。
    """
    user32 = ctypes.windll.user32
    VK_LWIN = 0x5B   # 左 Win 键
    VK_M    = 0x4D   # M 键
    KEYEVENTF_KEYUP = 0x2

    # 按下 Win
    user32.keybd_event(VK_LWIN, 0, 0, 0)
    # 按下 M
    user32.keybd_event(VK_M,    0, 0, 0)
    # 松开 M
    user32.keybd_event(VK_M,    0, KEYEVENTF_KEYUP, 0)
    # 松开 Win
    user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)

    # 给系统一点反应时间
    time.sleep(0.1)


def restore_and_position(
    name: str = "AutoCAD",
    width_ratio: float = 0.5,
    height_ratio: float = 0.5,
    x: int = 0,
    y: int = 0
) -> bool:
    """
    将第一个标题中包含 name 的窗口恢复、激活，并调整到指定位置和大小。

    :param name:          窗口标题关键字，默认 "AutoCAD"
    :param width_ratio:   窗口宽度占屏幕宽度的比例 (0 < ratio ≤ 1)
    :param height_ratio:  窗口高度占屏幕高度的比例 (0 < ratio ≤ 1)
    :param x:             窗口左上角 X 坐标，默认 0
    :param y:             窗口左上角 Y 坐标，默认 0
    :return:              找到并操作成功返回 True，否则 False
    """
    # ① 查找窗口
    candidates = [w for w in gw.getWindowsWithTitle(name) if w.title]
    if not candidates:
        print(f"❌ 未找到标题包含 “{name}” 的窗口。")
        return False
    win = candidates[0]

    # ② 如果最小化，则还原
    if win.isMinimized:
        try:
            win.restore()
            time.sleep(0.2)
        except Exception as e:
            print(f"⚠ 无法还原窗口：{e}")

    # ③ 激活窗口（置于最前）
    try:
        win.activate()
    except Exception:
        time.sleep(0.1)
        try:
            win.activate()
        except Exception as e:
            print(f"⚠ 激活窗口失败：{e}")

    time.sleep(0.2)

    # ④ 移动到指定位置
    try:
        win.moveTo(x, y)
    except Exception as e:
        print(f"⚠ 移动窗口失败：{e}")

    time.sleep(0.2)

    # ⑤ 调整窗口大小
    sw, sh = pyautogui.size()
    # 限制比例范围
    wr = max(0.01, min(1.0, width_ratio))
    hr = max(0.01, min(1.0, height_ratio))
    new_w = int(sw * wr)
    new_h = int(sh * hr)
    try:
        win.resizeTo(new_w, new_h)
        time.sleep(0.2)
    except Exception as e:
        print(f"⚠ 调整窗口大小失败：{e}")

    print(f"✅ 已将窗口“{win.title}”移动到 ({x},{y})，并调整为 {new_w}×{new_h}（占屏幕 {wr*100:.0f}% × {hr*100:.0f}%）")

    return True


"""
restore_and_position_cad(
    "微信",
    width_ratio = 0.5,
    height_ratio = 0.5,
    x= 0,
    y= 0
)

"""

def list_open_window_titles() -> list[str]:
    """
    获取当前所有可见窗口的标题列表包括子窗口。

    :return: 一个字符串列表，每一项都是一个非空窗口标题。
    """
    titles = []
    for w in gw.getAllWindows():
        title = w.title.strip()
        if title:
            titles.append(title)
    return titles
#&&% * 测鼠标位置
def ceshubiao_weizhi():
    """
    提示用户 5 秒内将鼠标移动到 AutoCAD 命令栏输入位置，
    然后采集当前鼠标坐标并返回 (x, y)。
    """
    print("请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置…")
    time.sleep(5)
    x, y = pyautogui.position()
    print(f"已获取坐标：({x}, {y})")
    return x, y

def run_idle_background(script_path: str):
    """
    用后台模式启动 IDLE 去运行某个脚本，返回 Popen 实例。
    """
    # Windows 上隐藏窗口的 flag
    CREATE_NO_WINDOW = 0x08000000
    proc = subprocess.Popen([
            sys.executable, "-m", "idlelib", "-r", script_path
        ],
        creationflags=CREATE_NO_WINDOW,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL
    )
    return proc



"""
idle_proc = run_idle_background(r"D:/Myprogramsystem/cad/CAD基本操作.py")
# 结束刚才启动的 IDLE
idle_proc.terminate()
idle_proc.wait(timeout=5)
"""

#按下鼠标左键再拖动

def click_and_drag(x: int, y: int, juli: int):
    """
    在屏幕坐标 (x, y) 按下左键，然后向纵向拖动距离 juli。
    若 juli 为正值，则向上拖动；若为负值，则向下拖动。
    """
    # 1. 移动到 (x, y)
    pyautogui.moveTo(x, y)

    time.sleep(2) 

    # 2. 按住左键
    pyautogui.mouseDown(button='left')

    # 3. 拖动：拖动到 (x, y + juli)
    #    如果仅需相对拖动，也可用 dragRel(0, juli)。
    dest_x = x
    dest_y = y - juli  # 注意：屏幕坐标里，向上是 y 减小，向下是 y 增大
    pyautogui.moveTo(dest_x, dest_y, duration=0.2)

    # 4. 松开左键
    pyautogui.mouseUp(button='left')

#寻找指定图片形状
def click_and_find_image_shape(x: int, y: int, tupian_path: str, timeout: float = 10.0):
    """
    在 (x, y) 单击一次，然后不断在整个屏幕上查找与 tupian_path 对应的图片形状，
    一旦找到，就把鼠标移到它的中心并返回中心坐标 (Python int)；超时仍未找到则返回 None。

    :param x, y:        单击的初始坐标（可触发界面更新）
    :param tupian_path: 要识别的目标图片路径（如微信笑脸）
    :param timeout:     最长等待时间（秒）
    :return:            找到时返回 (x, y)；否则返回 None
    """
    if not os.path.isfile(tupian_path):
        raise FileNotFoundError(f"找不到图片文件：{tupian_path}")

    # 1. 右键点击 (x, y)
    pyautogui.moveTo(x, y)
    pyautogui.click()
    time.sleep(0.3)  # 等待界面刷新

    start = time.time()
    while True:
        # 2. 搜索图片并获取中心坐标
        loc = pyautogui.locateCenterOnScreen(tupian_path, confidence=0.9)
        if loc:
            cx, cy = loc
            # 转为 Python int 再移动
            cx, cy = int(cx), int(cy)
            pyautogui.moveTo(cx, cy)
            return cx, cy

        # 3. 检查超时
        if time.time() - start > timeout:
            print("❌ 超时，未能识别到指定图片形状。")
            return None

        time.sleep(0.2)

#右键菜单
def right_click_and_move(x: int, y: int, x_xiangdui: int, y_xiangdui: int):
    """
    在屏幕坐标 (x, y) 处执行右键点击，然后将鼠标相对于 (x, y) 
    移动到 (x + x_xiangdui, y + y_xiangdui) 并停留。

    :param x:            初始水平坐标
    :param y:            初始垂直坐标
    :param x_xiangdui:   相对水平偏移（正值向右，负值向左）
    :param y_xiangdui:   相对垂直偏移（正值向下，负值向上）
    """
    # 移动到目标位置并右键点击
    pyautogui.moveTo(x, y)
    pyautogui.click(button='right')
    time.sleep(2)
    # 计算相对目标位置
    dest_x = x + x_xiangdui
    dest_y = y + y_xiangdui
    # 平滑移动到新位置
    pyautogui.moveTo(dest_x, dest_y, duration=0.2)

def kill_all_idle():
    """
    终止所有名为 'idle' 或 'idle.exe' 的进程（不再需要任务管理器）。
    """
    for p in psutil.process_iter(("pid", "name")):
        name = (p.info["name"] or "").lower()
        if name.startswith("idle"):
            try:
                p.terminate()
            except Exception:
                pass






##控制IDLE屏幕窗口在右上角
def set_idle_window_to_top_right():

    # 获取IDLE窗口句柄
    windows = [w for w in gw.getWindowsWithTitle('IDLE') if w.visible]
    
    if not windows:
        print("❌ 未找到IDLE窗口！")
        return
    
    # 获取IDLE主窗口
    win = windows[0]
    
    # 获取屏幕分辨率
    screen_width, screen_height = pyautogui.size()

    # 计算窗口大小：横竖各占1/2
    half_width = screen_width // 2
    half_height = screen_height // 2

    # 将窗口移动到屏幕的右上角 (屏幕宽度的一半，顶部为0)
    win.moveTo(half_width, 0)

    # 调整窗口大小，设置为屏幕的1/2宽度和1/2高度
    win.resizeTo(half_width, half_height)
    
    print(f"✅ IDLE窗口已调整到屏幕右上角，尺寸为 {half_width} x {half_height}")

def r():

    set_idle_window_to_top_right()

##控制OBS窗口在右下角
def place_obs_bottom_right():
    """
    将 OBS Studio 主窗口移动到屏幕右下角，并缩放为屏幕宽高的一半。
    - 若找不到 OBS 窗口，会打印错误信息。
    - 若有多个 OBS 窗口，仅操作第一个可见窗口。
    """
    # 1️⃣ 获取 OBS 主窗口
    obs_windows = [w for w in gw.getWindowsWithTitle('OBS') if w.visible]
    if not obs_windows:
        print('❌ 未找到可见的 OBS Studio 窗口')
        return

    obs = obs_windows[0]            # 取第一个
    print(f'🔍 找到窗口: {obs.title}')

    # 2️⃣ 计算目标尺寸与位置 —— 右下角 1/4 区域
    screen_w, screen_h = pyautogui.size()
    half_w, half_h = screen_w // 2, screen_h // 2
    target_left = screen_w - half_w      # 右半屏起点
    target_top  = screen_h - half_h      # 下半屏起点

    # 3️⃣ 移动并缩放
    obs.moveTo(target_left, target_top)
    time.sleep(0.1)                      # 给系统一点缓冲
    obs.resizeTo(half_w, half_h)

    print(f'✅ 已将 OBS 窗口调整到右下角 {half_w}×{half_h}')

def r2():

    place_obs_bottom_right()




##最小、最大化窗口
def minimize_window(window_keyword: str = 'OBS') -> bool:
    """
    通用：最小化第一个标题包含 window_keyword 的可见窗口。

    :param window_keyword: 要匹配的窗口标题关键字（子串匹配），默认 'OBS'
    :return: 如果成功最小化返回 True，否则返回 False
    """
    # 1) 找到所有匹配的可见窗口
    windows = [w for w in gw.getWindowsWithTitle(window_keyword) if w.visible]
    if not windows:
        print(f'❌ 未找到标题包含 "{window_keyword}" 的可见窗口')
        return False

    # 2) 取第一个并最小化
    win = windows[0]
    print(f'🔍 找到窗口: "{win.title}"，执行最小化')
    win.minimize()
    return True

def maximize_autocad_window(window_keyword: str = 'AutoCAD') -> bool:
    """
    强制最大化第一个标题包含 window_keyword 的可见窗口。
    优先尝试使用 win32gui，如不可用则退回 ctypes 调用 user32。
    """
    # 1) 找到目标窗口
    wins = [w for w in gw.getWindowsWithTitle(window_keyword) if w.visible]
    if not wins:
        print(f"❌ 未找到标题包含 “{window_keyword}” 的可见窗口")
        return False

    win = wins[0]
    hWnd = win._hWnd

    # 2) 先恢复（避免最小化状态），再最大化
    #    尝试使用 win32gui
    try:
        import win32gui, win32con
        win32gui.ShowWindow(hWnd, win32con.SW_RESTORE)
        time.sleep(0.1)
        win32gui.ShowWindow(hWnd, win32con.SW_MAXIMIZE)
    except ImportError:
        # 如果没有 pywin32，就退回 ctypes
        SW_RESTORE  = 9
        SW_MAXIMIZE = 3
        ctypes.windll.user32.ShowWindow(hWnd, SW_RESTORE)
        time.sleep(0.1)
        ctypes.windll.user32.ShowWindow(hWnd, SW_MAXIMIZE)

    time.sleep(0.2)  # 确保窗口完成最大化
    print(f"✅ 已将窗口 “{win.title}” 最大化")
    return True

##控制屏幕位置开始录制或关闭OBS
def start_obs_recording_by_click(x: int = 1768, y: int = 872,
                                 button: str = 'left',
                                 clicks: int = 1,
                                 move_duration: float = 0.2):
    """
    通过鼠标点击屏幕上 (x,y) 坐标来控制 OBS 开始/停止录制。
    
    :param x: 目标点击点 X 坐标
    :param y: 目标点击点 Y 坐标
    :param button: 鼠标按钮，'left'、'right' 或 'middle'
    :param clicks: 点击次数，默认为单击
    :param move_duration: 从当前位置移动到目标的耗时（秒）
    """
    # 安全设置（可选）
    pyautogui.FAILSAFE = True
    pyautogui.PAUSE    = 0.1

    # 1) 平滑移动到目标
    pyautogui.moveTo(x, y, duration=move_duration)
    # 2) 点击
    pyautogui.click(x=x, y=y, clicks=clicks, button=button)
    print(f"✅ 已点击 ({x}, {y})，请检查 OBS 是否已开始/停止录制。")


#&&&% *** 录屏
def fs(x1,y1):
    """
    微信调到0.5窗口

    """
    
    pyautogui.moveTo(x1+595,y1+190)
    pyautogui.click(x1+595,y1+190)
    pyautogui.press("enter")

def xuanqun(x1,y1,neirong):
    
    copy_to_clipboard(neirong)

    pyautogui.moveTo(x1+128,y1+33) 
    pyautogui.click(x1+128,y1+33)   
    time.sleep(2)
    activate_window_by_title("微信")

    time.sleep(2)

    pyautogui.moveTo(x1+128,y1+33)

    activate_window_by_title("微信")

    time.sleep(2)

    pyautogui.click(x1+128,y1+33)
  

    pyautogui.hotkey('ctrl', 'v') 


    pyautogui.moveTo(x1+158,y1+49) 
    pyautogui.click(x1+158,y1+49)   
 

    pyautogui.press("enter")


def copy_to_clipboard(text: str):
    """
    将传入的 text 文本写入系统剪贴板，供后续右键→粘贴使用。
    
    """
    
    # 创建一个隐藏的 tk 根窗口
    r = tkinter.Tk()
    r.withdraw()  # 隐藏主窗口

    # 清空剪贴板并写入新的文本
    r.clipboard_clear()
    r.clipboard_append(text)
    # 必须 update() 一下，确保数据真正存到剪贴板
    r.update()

    # 销毁隐藏窗口
    r.destroy()

def xieweixin(x1,y1,neirong):
        
        copy_to_clipboard(neirong)

        time.sleep(2)
   
        pyautogui.moveTo(x1+532,y1+151)

        time.sleep(2)
        activate_window_by_title("微信")

        
        pyautogui.moveTo(x1+532,y1+151)

        time.sleep(2)

        pyautogui.click(x1+532,y1+151)
      

        pyautogui.hotkey('ctrl', 'v')        

        pyautogui.click(x1+532,y1+151)
    
def 主操作函数():
    restore_and_position(
        name = "微信",
        width_ratio = 0.5,
        height_ratio = 0.5,
        x = 0,
        y = 0
    )  
 
    time.sleep(1)


    x1,y1,_,_ = activate_window_by_title("微信")

    xuanqun(x1,y1,"华新工作群")

    time.sleep(3)

    x2,y2=click_and_find_image_shape(358, 594, r"D:/Myprogramsystem/XT/weixin_xiaolian.png", timeout = 10.0)

    time.sleep(5)

    pyautogui.click(x2,y2) 

    time.sleep(5)

    x3,y3=click_and_find_image_shape(358, 594, r"D:/Myprogramsystem/XT/weixin_daixao_biaoqingbao.png", timeout = 10.0)

    time.sleep(2)
    pyautogui.click(x3,y3) 
    time.sleep(2)


 
    xieweixin(x1,y1,"Hello!我是公司管理员小化身，从今天起将参与公司管理，与各位一起奋进！")

    fs(x1,y1)

    
    time.sleep(1)


def  main_func(folder_path=r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/工业园整理/三期/测试"):

    打印输出PDF() 




def luping_1(main_func):
    """
    1) 开始 OBS 录制
    2) 最小化所有窗口；恢复并定位 OBS
    3) 调用传入的 main_func() 执行主操作
    4) 恢复 OBS，停止录制

    main_func: 零参数函数，负责执行插入或其他主操作
    """
    # 最小化所有窗口
    minimize_all_windows_d()

    # 恢复并定位 OBS 窗口
    restore_and_position(
        name="OBS",
        width_ratio=0.5,
        height_ratio=0.5,
        x=0,
        y=0
    )
    time.sleep(1)

    # 点击 OBS “开始录制”按钮（相对坐标）
    x0, y0, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x0 + 824, y0 + 328)
    time.sleep(1)
    pyautogui.click(x0 + 824, y0 + 328)

    # 最小化所有窗口，切换到 CAD
    minimize_all_windows_d()

    # 3️⃣ 执行主操作

    time.sleep(1)
    restore_and_position(
        name="AutoCAD",
        width_ratio=1,
        height_ratio=1,
        x=0,
        y=0
    )
 
    
    main_func()

    print("主操作函数执行完毕")

    # 4️⃣ 恢复并激活 OBS，停止录制
    time.sleep(0.5)
    minimize_all_windows_d()
    restore_and_position(
        name="OBS",
        width_ratio=0.5,
        height_ratio=0.5,
        x=0,
        y=0
    )
    time.sleep(0.2)

    x4, y4, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x4 + 824, y4 + 328)
    time.sleep(1)
    pyautogui.click(x4 + 824, y4 + 328)

    print("录制已停止")


def luping(main_func, *args, **kwargs):
    """
    1) 开始 OBS 录制  
    2) 最小化所有窗口；恢复并定位 OBS  
    3) 切换到 CAD 窗口，调用 main_func(*args, **kwargs）  
    4) 切换回 OBS，停止录制  

    main_func : 任意可调用对象  
    *args, **kwargs : 会原样传给 main_func  
    """
    # —— 最小化所有窗口 ——  
    minimize_all_windows_d()

    # —— 定位并激活 OBS ——  
    restore_and_position(name="OBS", width_ratio=0.5, height_ratio=0.5, x=0, y=0)
    time.sleep(1)

    # 点击 OBS “开始录制”  
    x0, y0, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x0 + 824, y0 + 328)
    time.sleep(0.2)
    pyautogui.click(x0 + 824, y0 + 328)

    # —— 切换到 CAD ——  
    minimize_all_windows_d()
    time.sleep(0.5)
    restore_and_position(name="AutoCAD", width_ratio=1.0, height_ratio=1.0, x=0, y=0)
    time.sleep(1)

    # —— 3️⃣ 执行主操作 ——  
    main_func(*args, **kwargs)
    print("主操作函数执行完毕")

    # —— 停止录制 ——  
    time.sleep(0.5)
    minimize_all_windows_d()
    restore_and_position(name="OBS", width_ratio=0.5, height_ratio=0.5, x=0, y=0)
    time.sleep(0.2)

    x4, y4, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x4 + 824, y4 + 328)
    time.sleep(0.2)
    pyautogui.click(x4 + 824, y4 + 328)

    print("录制已停止")













def mccs():
    
    dw=DWG()

    MC_com =[]

    shitu_entity(kuang)

    for x in LD:

        Ent=dw.insert_door(x,width=600,btn_x=351, btn_y=129, cmd_x=471, cmd_y=455)    

        MC_com.append(Ent)
        
    time.sleep(10)

    for yuan in LByuan:
        
        for x in MC_com:

            transfer_props_by_matchprop(yuan, x, delay=0.5)

            shitu_entity(x)

            time.sleep(2)

        name = (yuan.Layer)[3:]

        print("************************************************************")

        print(f"**************       替换门成{name} 构件       *************")

        print("************************************************************")


        shitu_entity(kuang)        

        time.sleep(10)


def 魔方():

    import 魔方分析

    魔方分析.魔方控制台(r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/自动化(动态)/魔方/分析")



#&&%#运行指定程序名的程序

def run_py(pyname):
    try:
        # 运行指定的 Python 程序，隐藏命令行窗口
        result = subprocess.run(
            [sys.executable, pyname],
            check=True,
            text=True,
            capture_output=True,
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        print(f"✅ 程序 {pyname} 执行成功。输出:\n{result.stdout}")
    except subprocess.CalledProcessError as e:
        print(f"❌ 运行 {pyname} 时发生错误: {e}")
        print(f"错误信息: {e.stderr}")
    except FileNotFoundError:
        print(f"❌ 未找到程序 {pyname}。请检查文件名和路径。")




##把鼠标移动到命令行窗口
def focus_cmdline(cmd_x, cmd_y, delay=0.2):
    """
    把鼠标移到命令行并单击，确保焦点回到 AutoCAD 命令栏。

    
    """
    pyautogui.moveTo(cmd_x, cmd_y, duration=delay)
    pyautogui.click()



#&&% 激活窗口和子窗口

def activate_window_by_title(title: str, click_titlebar: bool = True) -> bool:
    """
    激活一个指定标题的窗口。
    
    1. 用 pygetwindow 找到窗口对象；
    2. 如果窗口被最小化，先还原；
    3. 将窗口置为前台（优先使用 win.activate()，失败时用 ctypes 强制）；
    4. （可选）在标题栏中点击一次，确保焦点。
    
    :param title: 要激活的窗口标题（部分匹配）。
    :param click_titlebar: 是否模拟一次点击标题栏，以确保窗口获得焦点。
    :return: True=激活成功，False=未找到窗口。
    """
    USER32 = ctypes.windll.user32
    SW_RESTORE = 9


    # ① 查找标题包含关键字的窗口
    wins = [w for w in gw.getWindowsWithTitle(title) if title in w.title]
    if not wins:
        print(f"❌ 未找到标题包含 “{title}” 的窗口")
        return False
    win = wins[0]

    # ② 如果最小化，先还原
    if win.isMinimized:
        win.restore()
        time.sleep(0.2)

    
    # ③ 置为前台：先尝试 pygetwindow.activate()
    try:
        win.activate()
        time.sleep(0.2)
    except Exception:
        # ctypes 强制还原并置前
        hwnd = win._hWnd
        USER32.ShowWindow(hwnd, SW_RESTORE)          # 还原
        USER32.SetForegroundWindow(hwnd)             # 置前
        time.sleep(0.2)

    # ④ 可选：在标题栏中点一下
    if click_titlebar:
        x = win.left + win.width // 2
        y = win.top + max(1, min(30, win.height // 10))
        pyautogui.click(x, y)
        time.sleep(0.1)

    print(f"✅ 已激活窗口：“{win.title}” 位置({win.left},{win.top}) 大小{win.width}×{win.height}")

    return  win.left,win.top,win.width,win.height


def click_in_window(title: str, offset_x: float, offset_y: float, click_titlebar: bool = False) -> bool:
    """
    在指定窗口的某个相对像素位置点击一次（相对于窗口左上角的偏移量）。

    :param title:         窗口标题关键字（部分匹配）
    :param offset_x:      从窗口左上角算起的水平像素偏移量（0 表示左边缘，width 表示右边缘）
    :param offset_y:      从窗口左上角算起的垂直像素偏移量（0 表示顶边缘，height 表示底边缘）
    :param click_titlebar: 是否先在标题栏点击一次以确保窗口获取焦点
    :return:              True=点击成功，False=未找到窗口
    """
    # 1) 查找窗口
    wins = [w for w in gw.getWindowsWithTitle(title) if title in w.title]
    if not wins:
        print(f"❌ 未找到标题包含 “{title}” 的窗口")
        return False
    win = wins[0]

    # 2) 如果最小化，就先还原
    if win.isMinimized:
        win.restore()
        time.sleep(0.2)

    # 3) 尝试激活
    try:
        win.activate()
        time.sleep(0.2)
    except Exception:
        # ctypes 强制还原并置前
        hwnd = win._hWnd
        SW_RESTORE = 9
        ctypes.windll.user32.ShowWindow(hwnd, SW_RESTORE)
        ctypes.windll.user32.SetForegroundWindow(hwnd)
        time.sleep(0.2)

    # 4) 可选：点击标题栏让焦点真正到窗口
    if click_titlebar:
        tx = win.left + win.width // 2
        ty = win.top + 10  # 标题栏一般在窗口顶部 10px 左右
        pyautogui.click(tx, ty)
        time.sleep(0.1)

    # 5) 计算目标点的绝对屏幕坐标
    abs_x = int(win.left + offset_x)
    abs_y = int(win.top  + offset_y)

    # 6) 点击
    pyautogui.moveTo(abs_x, abs_y, duration=0.2)
    pyautogui.click(abs_x, abs_y)
    time.sleep(0.1)
    print(f"🔘 已在窗口 “{win.title}” 内部点击 ({offset_x}, {offset_y}) → 绝对 ({abs_x}, {abs_y})")
    return True

"""
click_in_window("图形导出", offset_x=600-10, offset_y=600-10, click_titlebar=True)

"""

#&&% 链接和关闭艾可云

def activate_and_click_aikeyun():


    try:
    
        left, top, width, height = activate_window_by_title("艾可云", click_titlebar=True)
        
        # Calculate relative offsets from provided data
        rel_x1 = 887 - 682  # 205
        rel_y1 = 542 - 93   # 449
        
        # Move to first click position and left click
        pyautogui.moveTo(left + rel_x1, top + rel_y1)
        pyautogui.click(button='left')
        
        # Wait for 2 seconds
        time.sleep(2)
        
        # Calculate close position: using width - 22 for x, as it aligns with top-right close button
        rel_x2 = width - 22  # Equivalent to 400 - 22 = 378 in data
        rel_y2 = 18          # 111 - 93
        
        # Move to close position and left click
        pyautogui.moveTo(left + rel_x2, top + rel_y2)
        pyautogui.click(button='left')
    
    
    
    except:
    
        pass
    
    









def drag_in_window_simple(
    title: str,
    start: tuple[float,float],
    offset: tuple[float,float],
    duration: float = 0.3,
    button: str = 'left',
    absolute_start: bool = False
):
    """
    拖拽函数，支持相对或绝对起点：

    :param title: 窗口标题关键字
    :param start: 起点坐标，(x, y)；
                  如果 absolute_start==False，则当成“窗口内部”坐标
                  如果 absolute_start==True，则当成“屏幕”坐标
    :param offset: 拖拽向量 (dx, dy)
    :param duration: 拖动时长
    :param button: 'left' 或 'right'
    :param absolute_start: True 则 start 视为屏幕绝对坐标
    """
    left, top, w, h = activate_window_by_title(title)

    if absolute_start:
        x1, y1 = start
    else:
        x1 = left + start[0]
        y1 = top  + start[1]

    x2 = x1 + offset[0]
    y2 = y1 + offset[1]

    pyautogui.moveTo(x1, y1)
    pyautogui.mouseDown(button=button)
    pyautogui.moveTo(x2, y2, duration=duration)
    pyautogui.mouseUp(button=button)
    time.sleep(0.1)
    print(f"已在“{title}”窗口拖拽：起点{(x1,y1)} → 终点{(x2,y2)}")


"""
drag_in_window_simple(
    "图形导出",
    start=(10,10),
    offset=(100,50),
    absolute_start=False
)

"""

#&&% 屏幕截图快照

"""
img = pyautogui.screenshot()

img.save('D:/Myprogramsystem/BaiduSyncdisk/宋岳/工业园整理/三期/CAD打印/01建筑/screenshot.png')

局部

# region=(left, top, width, height)
img = pyautogui.screenshot(region=(0, 0, 800, 600))

left=0 表示从屏幕最左边开始
top=0 表示从屏幕最上边开始
截取从左上角往右 800 像素、往下 600 像素范围内的区域。

应该是对角点 测算多次得CAD屏幕
img = pyautogui.screenshot(region=(175,101,2350,1250))
img.save('D:/Myprogramsystem/BaiduSyncdisk/宋岳/工业园整理/三期/CAD打印/01建筑/screenshot.png')


"""


##自动炸开区域内对象
    
def run_auto_explode_area(x1, y1, x2, y2, cmd_x, cmd_y, delay=2):

    """
    这是一个未使用pywin32API控制天正CAD的典型函数
    它炸开区域x1, y1, x2, y2的所有对象，不适合反复操作

    合理的处理方式还是配合视窗调整（发送z\na或e\n）先slect显性选择对象，再发送命令

    """
    
    script = Path(__file__).with_name('auto_EXPLODE.py')
    cmd = [
        sys.executable, str(script),
        f'--x1={x1}', f'--y1={y1}',
        f'--x2={x2}', f'--y2={y2}',
        f'--cmd_x={cmd_x}', f'--cmd_y={cmd_y}',
        f'--delay={delay}'
    ]
    # 隐藏子进程窗口
    flags = subprocess.CREATE_NO_WINDOW
    try:
        subprocess.run(cmd, check=True, creationflags=flags)
        print('▶ 子程序完成')
    except subprocess.CalledProcessError as e:
        print(f'❌ 子程序退出码 {e.returncode}')





##确保CAD窗口输入法为英文
"""
本函数研究了ctr+空格的转换，如需要可进一步分析ctr+shift转换的问题        

它真正的意义是提供了ocr截图分析方法，这让自动化处理窗口问题更为强大

"""
def ensure_english_input_method(cmd_x: int, cmd_y: int, delay: float = 0.2):
    """
    聚焦 AutoCAD 命令行 -> 输入 a -> OCR 判断是否出现 A(ARC)/AA(AREA)。
    若未出现则执行 Win+Space 强制切换输入法一次。

    参数
    ----
    cmd_x, cmd_y : 命令行输入框的屏幕坐标
    delay        : 鼠标移动耗时（秒）
    """

    # ---------- 辅助 ---------- #
    def click_cmdline():
        pyautogui.moveTo(cmd_x, cmd_y, duration=delay)
        pyautogui.click()

    def set_to_english():
        pyautogui.hotkey("win", "space")       # 系统输入法切换
        time.sleep(2.0)                        # 给系统充足时间

    def is_english() -> bool:
        """输入 a 截图 → OCR → 宽松匹配"""
        pyautogui.write("a")                   # 测试字符
        time.sleep(1.2)                        # 等列表弹出
        img = ImageGrab.grab(bbox=(310, 381, 510, 581))   # 按需调整
        raw = pytesseract.image_to_string(img)
        print(f"OCR 捕获: {raw!r}")

        # 去空白→大写，统一匹配
        cleaned = re.sub(r"\s+", "", raw).upper()
        return ("A(ARC)" in cleaned) or ("AA(AREA)" in cleaned)

    # ---------- 主流程 ---------- #
    click_cmdline()                # 聚焦命令栏

    if is_english():
        print("✅ 已是英文输入法，直接结束。")
        return

    print("❌ 检测到非英文，执行切换 …")
    set_to_english()
    click_cmdline()                # 切换后重新聚焦

    # 切换一次后不再循环；如仍失败可手动检查
    if is_english():
        print("✅ 切换成功。")
    else:
        print("⚠️ 切换后仍未检测到英文，可能 OCR 失效或坐标需调整。")



##列出所有当前窗口

def list_all_windows():
    # 获取所有可见的窗口
    windows = gw.getWindowsWithTitle('')
    
    if not windows:
        print("❌ 没有找到任何可见窗口。")
    else:
        print("当前桌面上的所有窗口：")
        for win in windows:
            print(f"窗口标题: {win.title}, 窗口大小: {win.width}x{win.height}, 位置: ({win.left}, {win.top})")


#关闭垃圾干扰窗口
def close_360_popup_window():
    # 获取所有窗口标题
    windows = gw.getWindowsWithTitle('360')  # '360' 为弹出窗口的部分标题，具体根据实际情况调整

    # 遍历所有窗口，找到匹配的窗口
    for window in windows:
        print(f"找到窗口: {window.title}")

        if '360' in window.title:  # 确认窗口属于360软件（你可以根据实际窗口标题调整）
            # 激活窗口
            window.activate()
            time.sleep(1)

            # 通过模拟按键关闭窗口（如果是弹出的提示框，常见的关闭按钮是 'ESC' 或 'Alt+F4'）
            pyautogui.hotkey('alt', 'f4')  # 模拟 Alt + F4 来关闭窗口
            print(f"✅ 关闭了窗口: {window.title}")
            return

    print("❌ 没有找到匹配的360窗口")







#  主函数
#  (1)
# 天正墙中轴线显示与隐藏

#  该函数系列包括如下一些函数

"""
墙是否加粗边线，墙中线显示，墙中线隐藏是文件属性，因此我们用D:/Myprogramsystem/cad/xitongjicuwenjian/墙基线打开.dwg，

D:/Myprogramsystem/cad/xitongjicuwenjian/墙基线关闭.dwg，D:/Myprogramsystem/cad/xitongjicuwenjian/墙线加粗.dwg三个空文件来

控制文件和程序的运行。因为获取天正墙信息的函数依赖墙中线的显示，所以对不熟悉的文件进行处理时，就可以通过这几个基本文件实现确定

的控制，即我们预期要文件的墙中线显示出来，或者墙边线要加粗，不加粗，依赖实际的需要。




"""
## 测试示例

##__________







def toggle_wall_centerline(conf=0.8, wait=2):
    """
    若 CAD 底栏中 “墙中线显示” 图标为灰色则点击启用，
    否则直接报告已开启。
    """
    import pyautogui as pg

    pg.FAILSAFE = True
    pg.PAUSE    = 0.1           # 全局轻微停顿

    # 模板文件
    BASE = Path(__file__).parent
    BAR_IMG  = BASE / 'bar_block.png'     # 底栏块
    OFF_IMG  = BASE / 'icon_off.png'      # 灰  ≡
    ON_IMG   = BASE / 'icon_on.png'       # 蓝  ≡


    # 1️⃣ 找到底栏大块坐标
    region = pg.locateOnScreen(str(BAR_IMG), confidence=conf)
    if not region:
        print('❌ 未找到底部状态栏块，请确认模板 bar_block.png')
        return
    bar_region = (region.left, region.top, region.width, region.height)

    # 2️⃣ 先查“蓝色已开”图标
    if pg.locateOnScreen(str(ON_IMG), region=bar_region, confidence=conf):
        print('✅ 天正墙中线已经显示（蓝色）。')
        return

    # 3️⃣ 查灰色图标
    off = pg.locateOnScreen(str(OFF_IMG), region=bar_region, confidence=conf)
    if not off:
        print('❌ 未找到灰色图标，可能界面皮肤不同或模板需重截。')
        return

    # 4️⃣ 点击灰色图标
    pg.click(pg.center(off))
    time.sleep(wait)

    # 5️⃣ 再次验证
    if pg.locateOnScreen(str(ON_IMG), region=bar_region, confidence=conf):
        print('✅ 已点击，天正墙中线现在已显示（蓝色）。')
    else:
        print('⚠️ 点击后仍未检测到蓝色图标，请手动检查。')





##使用辅助大矩形控制视图范围

def draw_shitu_rectangle_lw(length=500000.0, width=500000.0, center=(0.0, 0.0), layer_name="shitu"):
    """
    在 AutoCAD 中绘制一个以 center 为中心、长 length、宽 width 的闭合轻量级多段线矩形，
    并放到 layer_name 图层（不存在则新建）。

    参数
    ----
    length, width : float
        矩形的长和宽
    center : tuple(float, float)
        矩形中心点 (x, y)
    layer_name : str
        指定图层名称
    """

    # 2️⃣ 确保目标图层存在
    layers = doc.Layers
    try:
        lyr = layers.Item(layer_name)
    except Exception:
        lyr = layers.Add(layer_name)
        lyr.Color = 3  # 设置图层颜色（可选）

    # 3️⃣ 计算矩形顶点（轻量级多段线只要二维坐标）
    cx, cy = center
    hl = length / 2.0
    hw = width  / 2.0

    pts = [
        cx - hl, cy - hw,
        cx + hl, cy - hw,
        cx + hl, cy + hw,
        cx - hl, cy + hw,
        cx - hl, cy - hw,  # 回到起点闭合
    ]

    # 4️⃣ 把 Python 列表转换为 COM-safe 的双精度数组
    variant_pts = win32com.client.VARIANT(
        pythoncom.VT_ARRAY | pythoncom.VT_R8,
        pts
    )

    # 5️⃣ 添加轻量级多段线并设置属性
    lwpoly = mp.AddLightWeightPolyline(variant_pts)
    lwpoly.Closed = True
    lwpoly.Layer  = layer_name

    # 6️⃣ 可选：缩放到图形 extents
    acad.ZoomExtents()

    print(f"✅ 已在图层 “{layer_name}” 上绘制 {length}×{width} 矩形（轻量级多段线）")














#&&%设置长度角度单位精度

def set_dwg_units_precision():
    """
    设置当前 DWG 文件的单位及精度：
    - 长度单位：单位类型不变，仅设置精度为 0.00000000
    - 角度单位：单位类型不变，精度为 0.00000000
    """

    try:
        vars = doc.GetVariable

        # 设置长度精度（LUPREC = 8 表示 8 位小数）
        doc.SetVariable("LUPREC", 8)

        # 设置角度精度（AUPREC = 8 表示 8 位小数）
        doc.SetVariable("AUPREC", 8)

        print("✅ 已将长度和角度单位精度设置为 8 位小数 (0.00000000)")
    except Exception as e:
        print(f"❌ 设置失败: {e}")

def jd():
    set_dwg_units_precision()


##标注样式

def list_dim_styles():
    """
    列出当前 DWG 文件中所有标注样式名称。
    """
    try:
        styles = doc.DimStyles
        names = [styles.Item(i).Name for i in range(styles.Count)]
        print("📐 当前标注样式列表：")
        for name in names:
            print(" -", name)
        return names
    except Exception as e:
        print(f"❌ 获取标注样式失败：{e}")
        return []



def set_current_dimstyle_via_command(style_name="_TCH_ARCH"):
    """
    使用命令行方式设置当前标注样式，兼容天正。

    参数:
        style_name (str): 要设为当前标注样式的名称（如 "_TCH_ARCH"）
    """
    try:
        doc.SendCommand(f"-DIMSTYLE\nR\n{style_name}\n")
        print(f"✅ 已尝试通过命令行设置标注样式为：{style_name}")
    except Exception as e:
        print(f"❌ 命令行设置标注样式失败：{e}")



#&&% 文字左对齐到垂直线

def align_texts_to_x_by_llcorner(texts, x_target):
    """
    将 texts 列表中每个文字对象的外包盒左下角 X 对齐到 x_target，Y、Z 不变。

    texts: List[COMObject]    单行或多行文字对象
    x_target: float           目标 X 坐标
    """
    base_pt = vtpnt(0, 0, 0)
    for txt in texts:
        try:
            # 取左下角点
            ll_pt, _ = txt.GetBoundingBox()
            x0, y0, z0 = ll_pt
            # 计算平移向量
            dx = x_target - x0
            move_vec = vtpnt(dx, 0, 0)
            # 平移
            txt.Move(base_pt, move_vec)
        except Exception as e:
            h = getattr(txt, "Handle", "?")
            print(f"⚠ 对象 {h} 对齐失败：{e}")




#&&% 获取文字内容

def extract_text_content(ent):
    """
    根据实体类型获取文本内容，并仅移除文字首尾的空格：
      - AcDbText 或 AcDbMText: ent.TextString
      - TDbText: ent.Text
      - TDbMText: TDbMText_content(ent)
    只调用 strip() 删除首尾空白，不影响中间空格。
    """
    obj = ent.ObjectName
    if obj in ("AcDbText", "AcDbMText"):
        raw = ent.TextString
    elif obj == "TDbText":
        raw = ent.Text
    elif obj == "TDbMText":
        raw = TDbMText_content(ent)
    else:
        return ""
    # 仅删除首尾空白
    return raw.strip()





#&&% 设置当前使用的字体样式

        
def set_current_text_style(style_name="Standard"):
    """
    设置当前文字样式（通过 COM 接口方式）。

    参数:
        style_name (str): 要设置为当前的文字样式名称
    """
    try:
        text_styles = doc.TextStyles
        style = text_styles.Item(style_name)  # 获取指定样式
        doc.ActiveTextStyle = style           # 设置为当前样式
        print(f"✅ 当前文字样式已设置为：{style_name}")
    except Exception as e:
        print(f"❌ 设置文字样式失败：{e}")

#&&% 获取所有当前字体样式

def huoqu_ziti_style():

    styles = {ts.Name for ts in doc.TextStyles}
    
    return styles




#&&% 设置字体样式

"""
1，可多次对某种样式设置再更新

2，对汉字字体设置，不存在要设大字体。可用字体在C:/Windows/Fonts查找，直接用汉字名，使用create_text_style命令

3， 可以单独设置shx字体，使用set_text_style_onlyshx(style_name="style01", font_file="gbenor.shx", big_font_file=None)

所造字体缺失大字体，对汉字会产生问号

4，如不是单独设置汉字字体，完整的字体设置应包括英文部分的设置和大字体设置，用于中文显示而且是shx字体

 

"""
def create_text_style(sty_name="style01", ziti="宋体"):
    """
    在当前 DWG 中创建（或更新）一个中文文字样式。
    

    :param sty_name: 样式名称，默认 "style01"
    :param ziti:     字体名称，AutoCAD 会在系统字体中查找，例如 "宋体"
    # acad.ActiveDocument.ActiveTextStyle.SetFont(Typeface, Bold, Italic, charSet, PitchandFamily)
    # Typeface 字体名称；
    # Bold 加粗，布尔值，False为不加粗字体；
    # Italic 倾斜，布尔值，False为倾斜字体；
    # CharSet 字体字符集，1为默认字符集；
    # PitchAndFamily 字节及笔画形式。

    ts = doc.TextStyles.Item("HIT_TxtStyle")对shx

    ts.fontFile = "bigfont.shx"



    """

    # 1️⃣ 确保样式存在
    styles = doc.TextStyles
    try:
        ts = styles.Item(sty_name)
        print(f"⚠ 样式 '{sty_name}' 已存在，正在更新其属性。")
    except Exception:
        ts = styles.Add(sty_name)
        print(f"✅ 已创建文字样式 '{sty_name}'。")

    # 2️⃣ 设置字体
    try:
        acad.ActiveDocument.TextStyles.Item(sty_name).SetFont(ziti, False, False, 1, 0 or 0)
    except Exception:
        try:
            acad.ActiveDocument.TextStyles.Item(sty_name).SetFont(ziti, False, False, 1, 0 or 0)
        except Exception as e:
            print(f"⚠ 无法设置字体为 '{ziti}'：{e}")

    # 3️⃣ 置为当前

    acad.ActiveDocument.ActiveTextStyle = acad.ActiveDocument.TextStyles.Item(sty_name)
    
    # 5️⃣ 通知用户
    print(f"✅ 样式 '{sty_name}' 属性已更新")




def set_text_style_onlyshx(style_name="style01", font_file="gbenor.shx", big_font_file=None):
    """
    C:/Program Files/Autodesk/AutoCAD 2021/Fonts查找可用shx字体    
    设置 AutoCAD 的文字样式：
    - font_file：英文字体（.shx 或 .ttf）
    - big_font_file：可选的大字体（如 gbcbig.txt），默认为 None 表示不设置，仅设置英文字体
    set_text_style_onlyshx(style_name="TEST_STYLE", font_file="gbxxx.shx", big_font_file=None)
    不成功消息可用于判定shx字体在不在左侧英文字体


    """
    try:
        import win32com.client

        acad = win32com.client.Dispatch("AutoCAD.Application")
        doc = acad.ActiveDocument
        styles = doc.TextStyles

        # 查找或创建字体样式
        if style_name in [s.Name for s in styles]:
            text_style = styles.Item(style_name)
        else:
            text_style = styles.Add(style_name)

        text_style.FontFile = font_file

        # 只有传入合法的大字体路径才设置 BigFontFile，否则跳过设置
        if big_font_file and isinstance(big_font_file, str):
            text_style.BigFontFile = big_font_file

        print(f"字体样式 '{style_name}' 设置成功：英文字体 = {font_file}，大字体 = {big_font_file or '未设置'}")
        return True

    except Exception as e:
        print(f"设置字体样式失败：{e}")
        return False



def set_text_style(style_name="style01", font_file="gbenor.shx", big_font_file="gbcbig.shx"):
    """
    设置CAD中文字样式：英文shx文件 + 中文大字体文件
    """
    try:

        styles = doc.TextStyles

        # 判断是否已有该样式
        if style_name in [s.Name for s in styles]:
            text_style = styles.Item(style_name)
        else:
            text_style = styles.Add(style_name)

        # 设置字体和大字体
        text_style.FontFile = font_file
        text_style.BigFontFile = big_font_file

        print(f"字体样式 '{style_name}' 设置成功，英文字体: {font_file}, 中文大字体: {big_font_file}")
        return True

    except Exception as e:
        print(f"设置字体样式失败：{e}")
        return False

#&&% 列出可用shx非大字表 shx大字表 

"""

到CAD字体设置下拉菜单找

"""

##&&% 问号字体替换
"""
网友的ftst方案已经彻底解决此问题，我们需要的是备份好文件和进一步编制点击窗口的函数
我建议使用Standard样式的gbenor.shx和大字gbcbig.shx替换更安全
"""



#&&% ## 处理字体样式同名问题

def rename_conflicting_text_styles(file1_path: str,
                                   file2_path: str,
                                   suffix: str = "_1",
                                   retry_delay: float = 0.2,
                                   max_retries: int = 10):
    """
    在两个 DWG 中找出同名（用户）文字样式，
    并在第一个文件中将它们重命名（原名 + suffix）：
      1) 通过 -RENAME 命令重命名样式
      2) 确认样式表里旧名已消失、新名已出现
      3) REGEN 强制刷新
      4) 遍历 ModelSpace，将引用旧样式的实体指向新样式
      5) 保存并关闭
    系统默认样式会被自动跳过。
    """
    SYSTEM_STYLES = {"Standard", "ASHADE", "Annotative", "BigFont"}

    acad = win32com.client.Dispatch("AutoCAD.Application")
    acad.Visible = True

    doc1 = acad.Documents.Open(os.path.abspath(file1_path))
    doc2 = acad.Documents.Open(os.path.abspath(file2_path))

    try:
        # 1) 收集两文件的样式
        styles1 = {ts.Name for ts in doc1.TextStyles}
        styles2 = {ts.Name for ts in doc2.TextStyles}
        conflicts = (styles1 & styles2) - SYSTEM_STYLES
        if not conflicts:
            print("✅ 未发现需要重命名的用户样式。")
            return

        print(f"⚠ 发现同名用户样式：{conflicts}，将在 “{os.path.basename(file1_path)}” 中重命名：")
        ms = doc1.ModelSpace

        for old_name in conflicts:
            new_name = old_name + suffix
            # 如果新名也冲突，就多加后缀，直到独一
            while new_name in styles1:
                new_name += suffix

            # 2) 发送 RENAME 命令
            cmd = f"-RENAME\nStyle\n{old_name}\n{new_name}\n\n"
            doc1.SendCommand(cmd)
            # 等待命令被处理、样式表更新
            for attempt in range(max_retries):
                time.sleep(retry_delay)
                # 强制刷新图形状态
                doc1.SendCommand("REGEN\n")
                try:
                    # 如果旧名已不存在并且新名存在，就跳出重试
                    doc1.TextStyles.Item(new_name)
                    try:
                        doc1.TextStyles.Item(old_name)
                        # 旧名仍然存在，继续等
                        continue
                    except Exception:
                        # 旧名被正确移除
                        break
                except Exception:
                    # 新名还没出现，继续等
                    continue
            else:
                print(f"  ❗ 重命名 “{old_name}” → “{new_name}” 可能未生效（超时）。")

            # 3) 再把所有实体中引用旧样式的改成新样式
            for ent in ms:
                try:
                    ename = getattr(ent, "EntityName", "").upper()
                    oname = getattr(ent, "ObjectName", "")
                    # 原生 TEXT/MTEXT
                    if ename in ("TEXT", "MTEXT"):
                        if ent.TextStyle == old_name:
                            ent.TextStyle = new_name
                    # 天正文字
                    elif oname in ("TDbText", "TDbMText"):
                        if ent.TextStyle == old_name:
                            ent.TextStyle = new_name
                except Exception:
                    # 有些实体可能不允许改样式，忽略它们
                    pass

            # 更新本地样式集合
            styles1.discard(old_name)
            styles1.add(new_name)
            print(f"  · 样式 “{old_name}” → “{new_name}”")

        # 4) 最后保存并反馈
        doc1.Save()
        print(f"✅ 已保存改动到 “{os.path.basename(file1_path)}”。")

    finally:
        # 关闭文档，不保存对第二个文件的任何改动
        doc1.Close(False)
        doc2.Close(False)




















##将一个对象属性传给多个对象

def transfer_props_by_matchprop(entity, Ob, max_try=3, delay=0.4):

    CR = chr(13)

    def wait_idle(acad, dt=0.2, n=50):
        """轮询 IsQuiescent，最多 n 次，每次 dt 秒"""
        for _ in range(n):
            if acad.GetAcadState().IsQuiescent:
                return
            time.sleep(dt)

    def expand_rectangle(p1, p2, offset):
        return (p1[0]-offset, p1[1]-offset), (p2[0]+offset, p2[1]+offset)


    """
    把 entity 的属性批量复制到 Ob。若 Layer 未变化则重试，最多 3 次。
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument

    src_layer   = entity.Layer
    orig_layer  = Ob.Layer        # 复制前目标的图层

    # 目标包围盒窗口
    p1, p2 = Ob.GetBoundingBox()
    x1, y1, x2, y2 = p1[0], p1[1], p2[0], p2[1]
    h = 0.1 * (abs(x1 - x2) + abs(y1 - y2)) / 2
    (wx1, wy1), (wx2, wy2) = expand_rectangle(p1, p2, h)

    match_cmd = (
        "_MATCHPROP" + CR +
        "P"          + CR +            # Previous 作为源
        "_W"         + CR +
        f"{wx1},{wy1}" + CR +
        f"{wx2},{wy2}" + CR + CR
    )

    for attempt in range(1, max_try + 1):
        try:
            # ——— 1. 设定源对象为 Previous ———
            try:
                highlight_entity_by_bbox(entity)
            except Exception:
                entity.Select(True)        # 退而求其次
            time.sleep(delay)

            # ——— 2. 发送 MATCHPROP ———
            doc.SendCommand(match_cmd)
            wait_idle(acad)

            # ——— 3. 判断是否成功 ———
            if Ob.Layer == src_layer:
                print(f"[OK] 第 {attempt} 次匹配成功，Layer 改为 {src_layer}")
                return True

            print(f"[WARN] 第 {attempt} 次后 Layer 未变，重试…")
            time.sleep(delay)

        except Exception as e:
            print(f"[ERR] 第 {attempt} 次匹配异常：{e}")

    print(f"[FAIL] 连续 {max_try} 次仍未把属性复制给目标")
    return False


#视图合理化控制
"""
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_swiso"+chr(13))#西南轴测
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_seiso"+chr(13))#东南轴测
acad.ActiveDocument.SendCommand(
    "_-view" + chr(13) +
    "_nwiso" + chr(13)
)#西北轴测
acad.ActiveDocument.SendCommand(
    "_-view" + chr(13) +
    "_neiso" + chr(13)
)#东北轴测
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_top"+chr(13))#俯视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_bottom"+chr(13))#仰视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_front"+chr(13))#前视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_Right"+chr(13))#右视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_Back"+chr(13))#后视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_Left"+chr(13))#左视图
acad.ActiveDocument.SendCommand("_vscurrent"+chr(13)+"_R"+chr(13))#真实视觉样式

##合理显示对象

acad.ActiveDocument.SendCommand(
    "_z" + chr(13) +
    "_e" + chr(13)
)

acad.ActiveDocument.SendCommand(
    "_zoom" + chr(13) +
    "s"     + chr(13) +
    "0.8xp" + chr(13)
)




"""
#&&% 双线程生成器

def run_dual_threads_1(f1,                     # 线程1 函数

                     f2,                     # 线程2 函数

                     f1_args=(), f1_kwargs=None,

                     f2_args=(), f2_kwargs=None,

                     timeout_sec=180):



    """
    通用“双线程-GUI”调度器

    - f1 必须负责触发一个阻塞性窗口（如 SendCommand、ShowModalDlg 等）。
    - f2 负责侦测并自动化处理该窗口，当处理完毕后通知 f1 继续或退出。
    - f1、f2 的函数签名都应以 (timeout_event, done_event, …) 开头：
    
      def f1(timeout_event, done_event, …):
          pythoncom.CoInitialize()
          try:
              # ……主体代码：例如触发打印对话框
              timeout_event.wait()    # 等待线程2通知或超时
          except Exception:
              # 打印日志或其他处理
          finally:
              pythoncom.CoUninitialize()
              done_event.set()        # 通知“双线程”总控：我结束了

      def f2(timeout_event, done_event, …):
          pythoncom.CoInitialize()
          try:
              # ……主体代码：例如等待“创建打印文件”对话框出现并点击“回车”键
          except Exception:
              # 打印日志或其他处理
          finally:
              pythoncom.CoUninitialize()
              timeout_event.set()    # 通知 f1 线程：窗口已处理完，可以结束等待
              done_event.set()       # 通知“双线程”总控：我结束了

    - 参数说明：
        f1, f2：传入上面那种签名的函数
        f1_args、f1_kwargs：给 f1 传递的位置参数和关键字参数
        f2_args、f2_kwargs：给 f2 传递的位置参数和关键字参数
        timeout_sec：最多等待多少秒，如果超过则报告超时并返回 False

    - 返回值：
        True  ：两个子线程在 time_limit 内都完成了
        False ：超时，至少一个线程没及时调用 done_event.set()
    """
    if f1_kwargs is None:
        f1_kwargs = {}
    if f2_kwargs is None:
        f2_kwargs = {}

    # ① 为线程 1、2 准备同一个事件对
    timeout_event = threading.Event()
    done_event    = threading.Event()

    # ② 启动“线程1”：负责弹出、阻塞窗口
    t1 = threading.Thread(
        target=f1,
        args=(timeout_event, done_event, *f1_args),
        kwargs=f1_kwargs,
        daemon=True
    )
    # ③ 启动“线程2”：负责侦测窗口并点击、关闭
    t2 = threading.Thread(
        target=f2,
        args=(timeout_event, done_event, *f2_args),
        kwargs=f2_kwargs,
        daemon=True
    )

    start_time = time.time()
    t1.start()
    t2.start()

    # ④ 等待线程1 在 timeout_sec 秒内结束
    t1.join(timeout=timeout_sec)
    
    t2.join(timeout=timeout_sec)
    
        
        

    # ⑥ 检查 done_event 是否已被 set()
    if not done_event.is_set():
        # 超时：由调度器负责给线程1 发送退出信号
        timeout_event.set()
        print(f"⚠ 双线程执行超过 {timeout_sec}s —— 已触发 timeout_event")
        return False
    else:
        print("✅ 双线程任务在时限内完成")
        return True



#&&% CAD取消选择操作



def cancel_cad_selection(attempts: int = 3, delay: float = 0.5) -> bool:

    for i in range(1, attempts + 1):
        try:
            highlight_entities_in_window(0, 0, 0, 0)
            print(f"✅ 第{i}次尝试：cancel_cad_selection 成功")
            return True
        except Exception as e:
            print(f"⚠ 第{i}次尝试失败：{e}")
            if i < attempts:
                time.sleep(delay)
    print("❌ 已重试多次，仍未能执行 cancel_cad_selection")
    return False






#&&&&%% 第七部分  测试辅助




#   CAD基本操作-测试辅助
#####&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&




#最小化窗口
def min_w():
    import ctypes
    
    VK_MENU = 0x12  # Alt键
    VK_TAB = 0x09   # Tab键
    VK_LWIN = 0x5B  # 左Win键
    VK_M = 0x4D     # M键
    KEYEVENTF_KEYUP = 0x2

    # 模拟Alt + Tab
    ctypes.windll.user32.keybd_event(VK_MENU, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_TAB, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_TAB, 0, KEYEVENTF_KEYUP, 0)
    ctypes.windll.user32.keybd_event(VK_MENU, 0, KEYEVENTF_KEYUP, 0)

    # 模拟Win + M
    ctypes.windll.user32.keybd_event(VK_LWIN, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_M, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_M, 0, KEYEVENTF_KEYUP, 0)
    ctypes.windll.user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)




def ql():#清除测试辅助图层上的对象

    ensure_layer("测试辅助")

    




    
def srhd(*args):#根据输入坐标在模型空间画点
    """
    在模型空间绘制点并标注序号，支持以下调用形式：
    - srhd((x1,y1,z1), (x2,y2,z2))   # 多个点元组
    - srhd([(x1,y1,z1), (x2,y2,z2)]) # 列表形式
    - srhd((x1,y1,z1))               # 单点
    """
    doc = acad.ActiveDocument
    ms = doc.ModelSpace
    layer_name = "测试辅助"

    # 创建图层（如果不存在）
    try:
        doc.Layers.Item(layer_name)
    except:
        doc.Layers.Add(layer_name)
        print(f"✅ 已创建图层：{layer_name}")

    # 统一格式处理：允许单点、多个点、列表传入
    if len(args) == 1:
        if isinstance(args[0], (list, tuple)):
            if len(args[0]) == 3 and all(isinstance(i, (int, float)) for i in args[0]):
                points = [args[0]]
            else:
                points = args[0]
        else:
            print("❌ 输入格式不正确")
            return
    else:
        points = args

    # 绘制点与编号
    for idx, P in enumerate(points, 1):
        try:
            pt = vtpnt(*P)
            point = ms.AddPoint(pt)
            point.Layer = layer_name

            text = ms.AddText(str(idx), pt, 10)
            text.Layer = layer_name
        except Exception as e:
            print(f"❌ 添加点失败: {e}")

    return "✅ 点与编号已绘制"



def srhd_p(*args):#根据输入坐标在图纸空间画点
    """
    在图纸空间绘制点和编号，支持：
    - 多个坐标元组：srhd_p((x1,y1,z1), (x2,y2,z2))
    - 单个列表：srhd_p([(x1,y1,z1), (x2,y2,z2)])
    - 单个点：srhd_p((x1,y1,z1))
    """
    ps = doc.PaperSpace
    layer_name = "测试辅助"

    # 创建图层（如果不存在）
    try:
        doc.Layers.Item(layer_name)
    except:
        doc.Layers.Add(layer_name)
        print(f"✅ 已创建图层：{layer_name}")

    # 判断参数结构
    if len(args) == 1:
        if isinstance(args[0], (list, tuple)):
            if len(args[0]) == 3 and all(isinstance(i, (int, float)) for i in args[0]):
                # 单个点元组
                points = [args[0]]
            else:
                # 列表形式
                points = args[0]
        else:
            print("❌ 输入格式不正确")
            return
    else:
        # 多个点元组
        points = args

    for idx, P in enumerate(points, 1):
        try:
            pt = vtpnt(*P)
            point = ps.AddPoint(pt)
            point.Layer = layer_name

            text = ps.AddText(str(idx), pt, 10)
            text.Layer = layer_name
        except Exception as e:
            print(f"❌ 添加点失败: {e}")

    return "✅ 图纸空间中的点与编号已绘制"


def comtomath(LBcom):#将com点列表转为数学点列表

    LB_point=[]

    for i in range(0,len(LBcom)):

        point = LBcom[i].Coordinates

        LB_point.append(point)

    return LB_point        

        

#&&% 隔远查看

def fuzhi_chakan(LBcom,K=1):#K为放大倍数

    LK=[]

    for xx in LBcom:

        try:

            copy_obj = xx.Copy()
    
            point1 = vtpnt(0,0,0)
    
            point2 = vtpnt(0,K*1000000,0)
    
            copy_obj.Move(point1,point2)


            LK.append(copy_obj)

        except Exception as e:

            print(f"❌ 移动对象{xx.Handle}失败: {e}")               

    return LK



#测量已有文字长度

def celiang_wenzichangdu(TEXTCOM):

    text_copy = TEXTCOM.Copy()

    text_copy.Alignment = 2

    text_copy.TextAlignmentPoint =vtpnt(0,0,0)

    chang = abs(text_copy.InsertionPoint[0])

    text_copy.Delete()

    return chang

#测量新写文字长度

def celiang_wenzichangdu_write(ZF,style="图签",height=270,scalefactor=0.8):

    #根据字符串按样式字高宽度因子写入cad后的测量长度

    text_obj = acad.ActiveDocument.ModelSpace.AddText(ZF, vtpnt(0,0,0), height)

    text_obj.StyleName = style

    text_obj.ScaleFactor =scalefactor #宽度因子

    chang = celiang_wenzichangdu(text_obj)

    text_obj.Delete()

    return chang




##清空文件夹
def qingkong_wenjianjia(FolderPath):

     #清空文件夹B
    folder_path_1 = FolderPath 

    # 遍历文件夹中的所有文件
    for filename in os.listdir(folder_path_1):
        
        file_pathx = os.path.join(folder_path_1, filename)
        
        # 确保它是一个文件而不是文件夹
        if os.path.isfile(file_pathx):
            
            os.remove(file_pathx)  # 删除文件

        print(f"{FolderPath}文件夹已清空")


#&&% 返回对象外包盒的长，宽，横竖向，角点信息

def get_bbox_info(com_obj):
    """
    获取传入 AutoCAD COM 对象（如 Line、Circle、BlockReference 等）的外包盒信息，
    并计算其长（length）、宽（width）以及横向或竖向（orientation）状态。

    参数：
      com_obj -- 任意支持 GetBoundingBox() 方法的 AutoCAD COM 对象

    返回：
      (length, width, orientation)
        length      -- 外包盒较长一边的长度（在 X/Y 平面上）
        width       -- 外包盒较短一边的长度（在 X/Y 平面上）
        orientation -- 字符串：
                         "horizontal" 表示 X 方向跨度 ≥ Y 方向跨度，
                         "vertical"   表示 Y 方向跨度 >  X 方向跨度
    如果对象不支持 GetBoundingBox，会抛出异常；也可根据需要自行捕获处理。
    """
 


   # 调用 GetBoundingBox 方法，返回两个点：minPt、maxPt
    # minPt、maxPt 都是 3 元素的 tuple 或 list，形如 (x, y, z)
    




    try:

        minPt, maxPt = com_obj.GetBoundingBox()

    except Exception as e:

        print(f"获取外包盒失败: {e}")               

        return None


    # 计算 X、Y 方向跨度
    dx = maxPt[0] - minPt[0]
    dy = maxPt[1] - minPt[1]

    # 将较大值定义为 length，较小值定义为 width
    length = max(dx, dy)
    width  = min(dx, dy)

    # 判断横向（X 跨度 ≥ Y 跨度）还是竖向（Y 跨度 > X 跨度）
    if dx >= dy:
        orientation = "horizontal"
    else:
        orientation = "vertical"

    return minPt, maxPt,length, width, orientation

#&&% 判断对象外包盒的横竖

def bbox_orientation_flag(com_obj):
    """
    判断任意 COM 对象的外包盒是竖向、横向还是正方形：
      - 如果 Y 方向跨度 > X 方向跨度，返回 1（竖向）
      - 否则（包括 X 方向跨度 >= Y 方向跨度），返回 0
        —— 即当外包盒为正方形（X 跨度 == Y 跨度）时，也返回 0

    参数：
      com_obj -- 支持 GetBoundingBox() 的 AutoCAD COM 对象

    返回：
      int -- 竖向返回 1，横向或正方形返回 0
    """
    # 获取外包盒的最小点和最大点
    min_pt, max_pt = com_obj.GetBoundingBox()
    # 计算 X、Y 方向跨度
    dx = abs(max_pt[0] - min_pt[0])
    dy = abs(max_pt[1] - min_pt[1])
    # 如果是竖向（dy > dx），返回 1；否则（横向或正方形）返回 0
    return 1 if dy > dx else 0

#&&% 获取多个对象的外包盒数据
def group_bbox_corners(com_objs):
    """
    计算一组 COM 对象的整体外包盒，并按顺序返回四个角点坐标：
      1. 左下角 (minX, minY)
      2. 右上角 (maxX, maxY)
      3. 左上角 (minX, maxY)
      4. 右下角 (maxX, minY)

    参数：
      com_objs -- 可迭代的一组支持 GetBoundingBox() 方法的 COM 对象列表

    返回：
      四元组：(
        (minX, minY, z),
        (maxX, maxY, z),
        (minX, maxY, z),
        (maxX, minY, z)
      )
      其中 z 取自各对象外包盒的 z 值范围，统一使用最小 z（如需不同，可按需调整）
    """
    # 初始化为极端值
    global_min_x = float('inf')
    global_min_y = float('inf')
    global_max_x = float('-inf')
    global_max_y = float('-inf')
    global_min_z = float('inf')  # 如果需要统一 z 值，可使用最小 z
    # 如果不关心 z，只返回 0 即可。这里以最小 z 作为统一 z
    for obj in com_objs:
        try:
            min_pt, max_pt = obj.GetBoundingBox()
        except Exception:
            # 如果对象不支持 GetBoundingBox，跳过
            continue

        x1, y1, z1 = min_pt
        x2, y2, z2 = max_pt

        # 更新 X/Y extremes
        if x1 < global_min_x:
            global_min_x = x1
        if y1 < global_min_y:
            global_min_y = y1
        if x2 > global_max_x:
            global_max_x = x2
        if y2 > global_max_y:
            global_max_y = y2

        # 更新 Z extremes（如果需要统一使用最小 z）
        if z1 < global_min_z:
            global_min_z = z1

    # 如果所有对象都被跳过（列表为空或都不支持 GetBoundingBox），直接返回 None
    if global_min_x == float('inf'):
        return None

    # 采用 global_min_z 作为所有角点的 z 分量
    z = global_min_z

    # 左下、右上、左上、右下
    bottom_left  = (global_min_x, global_min_y, z)
    top_right    = (global_max_x, global_max_y, z)
    top_left     = (global_min_x, global_max_y, z)
    bottom_right = (global_max_x, global_min_y, z)

    return bottom_left, top_right, top_left, bottom_right

#&&% zip的用法
"""
seq_x0  = (x0, y0, z0)
P_start = (dx, dy, dz)
要得到新坐标 (x0+dx, y0+dy, z0+dz)，就可以写：
seq_x0 = tuple(a + b for a, b in zip(seq_x0, P_start))
zip(seq_x0, P_start) 会产出 (x0, dx), (y0, dy), (z0, dz)
a + b for a, b in ... 就对每一对做相加
最后用 tuple(...) 把结果收成三元组
对两个列表求和
xs = [10, 20, 30]
ys = [1, 2, 3]
sums = [x + y for x, y in zip(xs, ys)]
# sums == [11, 22, 33]
并行遍历三组数据
names = ["A", "B", "C"]
ages  = [30, 25, 40]
scores= [85, 92, 78]
for n, a, s in zip(names, ages, scores):
    print(f"{n} 年龄{a} 分数{s}")
解压
如果有一个列表 pairs = [(1,4),(2,5),(3,6)]，要拆成两个列表：
a, b = zip(*pairs)
# a == (1,2,3), b == (4,5,6)

"""

#&&% 从两点绘制矩形

def draw_rectangle_by_corners(p1: tuple[float, float, float],
                              p2: tuple[float, float, float],
                              layer_name: str = "测试辅助",
                              width: float = 0.0,
                              color: int = 256) -> object:
    """
    基于两点绘制一个闭合矩形：
      - p1: (x_min, y_min, z)
      - p2: (x_max, y_max, z)
    调用 draw_lwpolyline 绘制四边闭合多段线，返回新创建的多段线对象。
    """
    # 左下、右下、右上、左上
    x1, y1, z1 = p1
    x2, y2, z2 = p2
    coords = [
        (x1, y1, z1),
        (x2, y1, z1),
        (x2, y2, z1),
        (x1, y2, z1)
    ]
    # 绘制闭合矩形
    rect = draw_lwpolyline(
        coords3d=coords,
        layer_name=layer_name,
        width=width,
        color=color,
        closed=True
    )
    return rect

#&&% 获取外包盒中心
def bbox_center_2(e):
    # GetBoundingBox 返回两个 Point (minPt, maxPt)
    min_pt, max_pt = e.GetBoundingBox()
    x1, y1, _ = tuple(min_pt)
    x2, y2, _ = tuple(max_pt)
    return ((x1 + x2) / 2, (y1 + y2) / 2)

def bbox_center_3(ent):
    """返回实体外包盒中心 (cx, cy, cz)"""
    mn, mx = ent.GetBoundingBox()
    return ((mn[0] + mx[0]) / 2.0,
            (mn[1] + mx[1]) / 2.0,
            (mn[2] + mx[2]) / 2.0)


#&&&&%% 第八部分   CAD文件操作


#_________________________________________________________________________________________________________________________

#  模块使用说明

"""
该模块解决CAD文件的转换、打开、关闭、文件之间的复制、分解组合等问题 

"""

#  主函数
#  (1)
# 文件转成t7、t3格式

#  该函数系列包括如下一些函数
"""
zhuancheng_t7()
zhuancheng_t3()
在li()连接激活文件后，直接执行该命令即可转换
一般理解所得文件和当前文件同文件夹，但这次测试结果却在文件夹
C:/Users/Administrator/Documents/WeChat Files/wxid_mhrzdlppqacc22/FileStorage/File/2025-04/
"""    
#&&% 对获得的两行大地坐标数据建立字典

"""
大量数据的整理读取，AI不能绝对保证正确

"""

def build_J_points_from_selected_texts(LB, n_points=61, prefix_x="30", prefix_y="37"):
    num_re = re.compile(r"-?\d+(?:\.\d+)?")

    def first_number(s):
        m = num_re.search(str(s))
        if not m:
            raise ValueError(f"无法解析数字: {s!r}")
        return float(m.group(0))

    items = []
    for ent in LB:
        txt, ip = get_text_and_ip(ent)
        if not txt or not ip:
            continue
        txt = str(txt).strip().replace(" ", "")
        x0, y0, *_ = ip
        if txt.startswith(prefix_x):
            items.append(("x", first_number(txt), float(y0), float(x0)))
        elif txt.startswith(prefix_y):
            items.append(("y", first_number(txt), float(y0), float(x0)))

    # 从上到下（y降序），同行按x升序稳定
    items.sort(key=lambda t: (-t[2], t[3]))
    xs = [v for typ, v, _, _ in items if typ == "x"]
    ys = [v for typ, v, _, _ in items if typ == "y"]

    if len(xs) != n_points or len(ys) != n_points:
        raise RuntimeError(f"X={len(xs)} / Y={len(ys)} 与期望 {n_points} 不符。")

    pts_dict = {f"J{i+1}": (xs[i], ys[i]) for i in range(n_points)}
    pts_list = [(f"J{i+1}", xs[i], ys[i]) for i in range(n_points)]
    return pts_dict, pts_list


#从界点大地坐标计算经纬度
def convert_pts_dict_to_latlon(pts_dict, central_lon=111):
    """
    输入: pts_dict = {'J1': (N, E), 'J2': (N, E), ...}
          N=北坐标, E=东坐标 (CGCS2000 高斯-克吕格投影)
    输出: geo_dict = {'J1': (lon, lat), 'J2': (lon, lat), ...}  (经度, 纬度, 单位:度)
    """
    a = 6378137.0
    f = 1 / 298.257222101
    FE = 500000
    FN = 0
    e2 = 2*f - f*f
    e_prime_2 = e2 / (1.0 - e2)

    def one_point(N, E):
        m = N - FN
        mu = m / (a * (1 - e2/4 - 3*e2**2/64 - 5*e2**3/256))
        e1 = (1 - math.sqrt(1 - e2)) / (1 + math.sqrt(1 - e2))
        Bf = (mu
              + (3*e1/2 - 27*e1**3/32) * math.sin(2*mu)
              + (21*e1**2/16 - 55*e1**4/32) * math.sin(4*mu)
              + (151*e1**3/96) * math.sin(6*mu)
              + (1097*e1**4/512) * math.sin(8*mu))
        x_prime = E - FE
        cosBf = math.cos(Bf)
        tanBf = math.tan(Bf)
        eta2 = e_prime_2 * (cosBf**2)
        Nf = a / math.sqrt(1 - e2 * (math.sin(Bf)**2))
        Mf = a * (1 - e2) / (1 - e2 * (math.sin(Bf)**2))**1.5
        lat = (Bf
               - tanBf/(2* Mf* Nf) * (x_prime**2)
               + tanBf/(24* Mf* Nf**3) * (5 + 3*tanBf**2 + eta2 - 9*tanBf**2 * eta2) * (x_prime**4)
               - tanBf/(720* Mf* Nf**5) * (61 + 90*tanBf**2 + 45*tanBf**4) * (x_prime**6))
        lon = (math.radians(central_lon)
               + x_prime/(Nf* cosBf)
               - (1 + 2*tanBf**2 + eta2) * (x_prime**3)/(6* Nf**3* cosBf)
               + (5 + 28*tanBf**2 + 24*tanBf**4 + 6* eta2 + 8*tanBf**2 * eta2) * (x_prime**5)/(120* Nf**5* cosBf))
        return math.degrees(lon), math.degrees(lat)

    geo_dict = {}
    for k, (N, E) in pts_dict.items():
        geo_dict[k] = one_point(N, E)

    return geo_dict

# 文件基本操作



"""
使用共同的文件全局变量acad,mp,doc,sp是我们编制脚本控制不同文件的基础            


"""


    

def rename_time(output_path):#不覆盖从当前时间给带路径文件名命名

    from datetime import datetime
    now = datetime.now()
    time_str = now.strftime("%m-%d-%H-%M")

    dir_name, file_name = os.path.split(output_path)
    name, ext = os.path.splitext(file_name)
    new_file_name = f"{name}_{time_str}{ext}"
    new_path = os.path.join(dir_name, new_file_name)
    return new_path

#供参考优化
def open_dwg(path: str, visible: bool = True):
    # 初始化 COM（尤其在多线程或非脚本交互环境下推荐调用）
    pythoncom.CoInitialize()

    # 启动并连接到 AutoCAD 应用
    # Dispatch 会在已有实例上复用，DispatchEx 会强制新开一个实例
    acad = Dispatch("AutoCAD.Application")  # 或者 DispatchEx("AutoCAD.Application")

    # 可见性（True 时会弹出界面）
    acad.Visible = visible

    # 打开 DWG 文档
    doc = acad.Documents.Open(path)

    print(f"已打开：{doc.Name}")

    return acad, doc









##打开文件
def Open_By_Omission_wenjian(file_path):

    """
    不能指望通过这函数控制所有打开文件时遇到的窗口跳出，报错等，它们应该在别的地方解决，例如字体问题，报错窗口，代理错误等等

    """
    
    t1 = time.time()
    
    max_retries = 3

    for attempt in range(max_retries):
        
        try:
            
            # 尝试打开文件
            new_doc = acad.Documents.Open(file_path)
            
            print(f"Opened file: {file_path}")
            
            t2 = time.time()
            
            print("文件打开耗时:", t2 - t1, "秒")

            li() #所有的函数都使用同样的全局变量acad,mp,doc等，这样在不同的文件上函数仍然通用
            
            return new_doc
        
        except Exception as e:
            
            print(f"Attempt {attempt + 1} failed: {e}")
            pass
            
            time.sleep(2)  # 等待2秒后再次尝试

    # 如果所有尝试都失败
    t2 = time.time()
    
    print("文件打开耗时:", t2 - t1, "秒")
    
    return  None



##打开文件时跳过缺字体窗口


def f1_openfile_getwindaow(timeout_event, done_event,
                     file_path):
    pythoncom.CoInitialize()
    try:

        

        print("线程1启动")

        li()

        time.sleep(2)

        Open_By_Omission_wenjian(file_path)

        time.sleep(10)

        timeout_event.wait()          # 等待线程2完工 / 或调度器超时

    except Exception as e:
        print("f1_openfile_getwindaow:", e, traceback.format_exc())
    finally:
        pythoncom.CoUninitialize()
        done_event.set()
        


def f2_delwindaow(timeout_event, done_event):
    pythoncom.CoInitialize()
    try:
        print("线程2启动")
        # 闪动窗口（示例：刷新一下窗口列表）
       
        time.sleep(1)

        BT = list_open_window_titles()
        print("当前窗口标题：", BT)

        if '缺少 SHX 文件' in BT:
            # 点击对话框中的“忽略”按钮，假设它在对话框中的大致坐标 (148, 220)
            click_in_window("缺少 SHX 文件", 148, 220, click_titlebar=True)
            print("🖱 已点击“忽略”按钮")
        else:
            print("ℹ️ 未检测到“缺少 SHX 文件”对话框")

    except Exception as e:
        print("f2_delwindaow 异常:", e, traceback.format_exc())
    finally:
        # 先发信号给线程1 退出阻塞，再通知总控已经完成
        timeout_event.set()
        pythoncom.CoUninitialize()
        done_event.set()




def Open_file_nic(file_path ):

    """
    实测似乎人工打开文件时会跳出SHX字体缺失窗口，而Open_By_Omission_wenjian命令不会，也许这个命令是多余的，但仍然保留这个机制，因为我们不能断定
    
    """

    ok = run_dual_threads(

        f1=f1_openfile_getwindaow,

        f2=f2_delwindaow,

        f1_args=(file_path,),

        f2_args=(),

        timeout_sec=600
    )
    if ok:
        print(f"🎉 成功打开文件 → {file_path}")
    else:
        print("🚨 打开文件超时 / 失败")





##保存文件#另存就是doc.SaveAs()

def savefile():

    doc.Save()


##关闭文件(别乱删)

def guanbifile():

    doc.Close()



##确保关闭当前文件


def close_current_drawing_safely():
    """
    安全关闭当前 DWG 文件，确保确实关闭并重新连接。
    最多尝试 3 次。
    """
    
    try:
        Name1 = doc.Name
    except:
        print("⚠️ 当前 doc 无法获取名称，可能未连接。")

        li()

        Name1 = doc.Name
    
    for attempt in range(1, 4):  # 最多尝试3次
        print(f"🔄 第 {attempt} 次尝试关闭 '{Name1}'")
        close_dwg_by_name(Name1)

        li()  # 重新连接 acad、doc、mp、sp 等
        try:
            Name2 = doc.Name
        except:
            Name2 = None

        if Name2 != Name1:
            print(f"🟢 已确认文件 '{Name1}' 关闭，当前打开文件为 '{Name2}'")
            li()  # 再执行一次，确保变量正确
            return
        else:
            print("⚠️ 文件仍未关闭，继续尝试...")

    print(f"❌ 多次尝试仍未成功关闭 '{Name1}'，请手动检查。")



#&&% 2 两个文件A,B的操作
"""
文件作为块插入不稳定，甚至在交互模式可以同样的代码转为函数模式却不行，反倒是使用全选复制粘贴比较靠谱

除了使用这个命令合并文件，还有转换为纯数据命令

        
    """

#新建一个空白文件(不打开)
def create_new_dwg_file(name_with_path):
    """
    创建一个新的 DWG 文件，并保存为指定完整路径 name_with_path（应以 .dwg 结尾）。

    参数：
        name_with_path - 完整的文件保存路径（例如 D:\CADXT\Export\新图01.dwg）
    """
   
    # 检查并创建目录（如果不存在）
    folder = os.path.dirname(name_with_path)
    if not os.path.exists(folder):
        os.makedirs(folder)
        print(f"📁 已创建文件夹：{folder}")

    # 新建 DWG 文件
    acad.Documents.Add()
    time.sleep(1)

    # 连接当前文档
    new_doc = acad.ActiveDocument

    # 保存到指定路径
    new_doc.SaveAs(name_with_path)

    new_doc.Close()
    li()
    print(f"✅ 新建并保存 DWG 文件：{name_with_path}")

#新建一个空白文件
def create_new_dwg_file_no(name_with_path):
    """
    创建一个新的 DWG 文件，并保存为指定完整路径 name_with_path（应以 .dwg 结尾）。

    参数：
        name_with_path - 完整的文件保存路径（例如 D:\CADXT\Export\新图01.dwg）
    """
   
    # 检查并创建目录（如果不存在）
    folder = os.path.dirname(name_with_path)
    if not os.path.exists(folder):
        os.makedirs(folder)
        print(f"📁 已创建文件夹：{folder}")

    # 新建 DWG 文件
    acad.Documents.Add()
    time.sleep(1)

    # 连接当前文档
    new_doc = acad.ActiveDocument

    # 保存到指定路径
    new_doc.SaveAs(name_with_path)
  
    li()
    print(f"✅ 新建并保存 DWG 文件：{name_with_path}")



#  主函数
#  (1)
# 关闭文件

#  该函数系列包括如下一些函数

def close_all_except_active_safe():#关闭除当前激活文档外的所有 DWG 文件
    """
    更稳定地关闭除当前激活文档外的所有 DWG 文件，避免 COM 对象断链。
    """
    try:
        active_name = acad.ActiveDocument.Name
        all_names = [acad.Documents.Item(i).Name for i in range(acad.Documents.Count)]
        closed = 0

        for name in all_names:
            if name != active_name:
                try:
                    doc = acad.Documents.Item(name)
                    doc.Close(False)  # 不保存直接关闭
                    print(f"🗂️ 已关闭：{name}")
                    closed += 1
                except Exception as e:
                    print(f"⚠️ 无法关闭 {name}：{e}")

        print(f"✅ 成功关闭 {closed} 个文档，仅保留 {active_name}")

    except Exception as e:
        print(f"❌ 安全关闭失败：{e}")


## 测试示例
##close_all_except_active_safe()
##🗂️ 已关闭：Drawing3.dwg
##🗂️ 已关闭：Drawing4.dwg
##🗂️ 已关闭：空白.dwg
##🗂️ 已关闭：测试1.dwg
##✅ 成功关闭 4 个文档，仅保留 cs.dwg
##__________
    

# 双文件操作
"""
刚开始启动天正，默认打开了一个drawing1的文件，使用create_new_dwg_file新建一个文件后，新建的文件会关闭，仍然是drawing1在，仍需li()连接一下

此时使用Open_By_Omission_wenjian打开新文件，就会同时自动关闭drawing1。同样需要li()连接一下
因此我们面临两种可能，一个是天正的drawing1作为基本连接状态，此时文件数为1，但要注意create_new_dwg_file新建 不改变这种状态，而 Open_By_Omission_wenjian打开桌面就会变成只有1个文件

另一种情况时我们用自己系统的“空白.dwg”作为当前桌面的唯一基础文件。
事实上，在这个桌面上只有一个文件的情况下，我们需要有两个文件的状态。继续打开一个新文件即可。但此时，就需要在打开新文件之前，运行doc1=acad.ActiveDocument，用来标记当前文件。再打开新文件
li()新文件，运行doc2=acad.ActiveDocument，记录当前文件。此时，连接的是doc2,但我们可以执行doc1.Close()关闭前面的文件，就恢复到了单一文件状态。
当然，通过dir(doc1)可以查阅文件的属性和方法，不仅可以查看文件的名字，还可以获取这个文件的其它属性。例如doc1.Groups，PlotConfigurations,我们是否可以在两个文件都在的情况下拷贝一个到另一个来。


#当前激活桌面文件的激活布局
acad.ActiveDocument.SetVariable("TILEMODE",0)

layouts = doc.Layouts

existing_names = [layout.Name for layout in layouts]

doc.ActiveLayout = layouts.Item("布局1")
    
layout = source_doc.ActiveLayout

#获取当前激活文件的打印配置

zg_pdf_config = source_doc.PlotConfigurations.Item("ZG_PDF")

# 使用 CopyFrom 方法将 "ZG_PDF" 的设置复制到指定布局

layout.CopyFrom(zg_pdf_config)

也就是说，两个文件同时打开的操作是必不可少的。虽然具体运行时我们尽量保持一个文件。但两个文件需要交换信息。

"""

def copy_doc_to_current(file_path):#将目标文件原位复制粘贴到当前文件

    """
    若操作不当，竟至不能复制粘贴
    将目标文件粘贴进来，当然是最自然的需求
    不要用“空白.dwg”作测试，也不要用天正开启时默认生成的文件，因为它们都会破坏整个流程
    很遗憾，块插入命令不能生效

    """   

    li()
    doc.Save()
    file_path1=file_path
    file_path2=doc.FullName#获取当前文件带路径的文件名
    
    
    name2=doc.Name
    doc2=get_doc_by_name(name2)   
    
    #打开辅助文件
    
    Open_By_Omission_wenjian("D:/Myprogramsystem/cad/xitongjicuwenjian/空白.dwg")

    li()

    name3=doc.Name

    doc3=get_doc_by_name(name3)     

    doc2.Close()

    insert_dwg_2(file_path2,file_path1, insert_point=(0, 0, 0))

    time.sleep(2)

    Open_By_Omission_wenjian(file_path2)

    li()

    doc.SendCommand("z\ne\n")

    doc3.Close()


def insert_dwg_2(new_file_path_A,new_file_path_B, insert_point=(0, 0, 0)):#文件B合成到文件A中

    """
    如果刚打开天正 ，会默认产生一个dwg1文件，也会激活，用来进行后续操作。可当我们操作到后面关闭新建的文件A时，注意新打开A新建A时这个dwg1也关闭了。?

    """

    

    li()

    Open_By_Omission_wenjian(new_file_path_B) #打开文件B       

    li()

    acad.ActiveDocument.SendCommand('_ai_selall'+chr(13))

    time.sleep(4)

    acad.ActiveDocument.SendCommand('_copybase'+chr(13)+'0,0,0'+chr(13)+chr(13))#全选复制

    time.sleep(4)

    guanbifile()
    
    li()

    Open_By_Omission_wenjian(new_file_path_A)#打开文件A

    li()        
    
    acad.ActiveDocument.SendCommand('TPasteClip'+chr(13)+'0,0,0'+chr(13)+chr(13))#粘贴

    time.sleep(4)

    doc.save()

    time.sleep(4)
   
    doc.close()
    
    li()



def copy_group_S1_from_doc1_to_doc2(doc1, doc2, group_name="S1"):#将名为“S1”的组复制到当前桌面上另一个文档
    """
    将 doc1 中名为 group_name 的组复制到 doc2 中，并重新组装组。
    粘贴点为 0,0,0。粘贴后通过 handle 差集识别新对象。

    mp要不断更新

    def refresh_modelspace(doc):
        return doc.ModelSpace
    逻辑更清晰

    直接发送命令复制粘贴或许更好
    """
    try:
        # 1. 激活源文档
        set_active_doc(doc1)
        li()

        group = doc.Groups.Item(group_name)
        handles = [ent.Handle for ent in group]
        objs = [ent for ent in mp if ent.Handle in handles]

        yin_to_xian_xuanze(objs)

        doc.SendCommand("_copybase\n0,0,0\n")
        time.sleep(0.5)
        doc.SendCommand("_copyclip\n\n")
        time.sleep(1)

        # 2. 激活目标文档
        set_active_doc(doc2)
        li()
        ms2 = doc.ModelSpace

        # 3. 粘贴前记录已有对象
        pre_map = get_handle_object_map(ms2)

        # 4. 粘贴
        doc.SendCommand("_pasteclip"+chr(13)+"0,0,0"+chr(13)+chr(13))#
        time.sleep(1.5)

        # 5. 粘贴后重新记录对象
        ms2 = doc.ModelSpace
        
        post_map = get_handle_object_map(ms2)

        # 6. 取出新对象（通过 handle 差集）
        new_handles = set(post_map) - set(pre_map)
        new_objs = [post_map[h] for h in new_handles]

        print(f"✅ 粘贴完成，识别出 {len(new_objs)} 个新图元")

        # 7. 添加这些对象到组中（使用你提供的方法）
        add_objects_to_group(group_name, new_objs)

        print(f"✅ 成功将粘贴对象加入组 '{group_name}'")

    except Exception as e:
        print(f"❌ 复制组失败: {e}")

        



#&&% 合并两个不重叠的文件    
def insert_dwg_pyautocad(new_file_path_A,new_file_path_B, insert_point=(0, 0, 0)):#文件B合成到新建文件A中

    """
    如果刚打开天正 ，会默认产生一个dwg1文件，也会激活，用来进行后续操作。可当我们操作到后面关闭新建的文件A时，注意新打开A新建A时这个dwg1也关闭了。?

    """

    create_new_dwg_file(new_file_path_A)#新建文件命令已经默认了关闭新建的这个文件的操作

    li()

    Open_By_Omission_wenjian(new_file_path_B) #打开文件B       

    li()

    acad.ActiveDocument.SendCommand('_ai_selall'+chr(13))

    time.sleep(4)

    acad.ActiveDocument.SendCommand('_copybase'+chr(13)+'0,0,0'+chr(13)+chr(13))#全选复制

    time.sleep(4)

    guanbifile()
    
    li()

    Open_By_Omission_wenjian(new_file_path_A)#打开文件A

    li()        
    
    acad.ActiveDocument.SendCommand('TPasteClip'+chr(13)+'0,0,0'+chr(13)+chr(13))#粘贴

    time.sleep(4)

    doc.save()

    time.sleep(4)
   
    doc.close()
    
    li()

#批量合成文件
def insert_multiple_dwgs_to_new_file(new_file_path_A, source_files_list):
    """
    创建一个新 DWG 文件 A，并将多个已有 DWG 文件（B1, B2, ...）的内容
    原位复制粘贴到 A 中，最终保存并关闭 A，返回初始 dwg 状态。

    参数：
        new_file_path_A      - 要创建的新 DWG 文件路径（完整路径）
        source_files_list    - 要复制进来的多个 DWG 文件路径列表
    """
    # Step 1: 创建新文件 A（新建后自动关闭）
    create_new_dwg_file(new_file_path_A)
    time.sleep(1)

    # Step 2: 逐个处理每个源文件 B
    for idx, fileB in enumerate(source_files_list):
        print(f"📥 处理第 {idx+1} 个源文件：{fileB}")
        Open_By_Omission_wenjian(fileB)
        li()
        time.sleep(1)

        acad.ActiveDocument.SendCommand('_ai_selall' + chr(13))
        time.sleep(2)

        acad.ActiveDocument.SendCommand('_copybase' + chr(13) + '0,0,0' + chr(13) + chr(13))
        time.sleep(3)

        guanbifile()
        time.sleep(1)

        Open_By_Omission_wenjian(new_file_path_A)
        li()
        time.sleep(1)

        acad.ActiveDocument.SendCommand('TPasteClip' + chr(13) + '0,0,0' + chr(13) + chr(13))
        time.sleep(3)

        acad.ActiveDocument.Save()
        time.sleep(1)

        acad.ActiveDocument.Close(False)
        time.sleep(1)

    # Step 3: 操作结束后回到 dwg1 初始状态
    li()
    print(f"✅ 已成功将 {len(source_files_list)} 个 DWG 合并至文件：{new_file_path_A}")





    

def xianshi_yincangtuxing():#  显示文件中可能隐藏的对象
        
    acad.ActiveDocument.SendCommand("HFKJ"+chr(13)+chr(13))#在V4状态下可能要改成"HFKJ"+chr(13)+"Y"+chr(13)

    doc.Save()



# 创建一个事件对象

timeout_event = threading.Event()

event = threading.Event()

def run_cad_program(timeout_event, event):
    pythoncom.CoInitialize()
    
    try:
        acad = win32com.client.Dispatch("AutoCAD.Application")
        print(acad.ActiveDocument.Name)
        
        acad.ActiveDocument.SendCommand("TSaveAs"+chr(13))
        
        print("CAD命令已发送，等待窗口操作完成...")
        
        # 等待timeout_event信号，如果收到信号，则退出
        timeout_event.wait()

    except Exception as e:
        print(f"run_cad_program 出现错误: {e}")
        pass
    finally:
        pythoncom.CoUninitialize()
        event.set()  # 通知主线程完成

        
def automate_window_with_pywinauto_t7(timeout_event, event):
    pythoncom.CoInitialize()

    try:
        jincheng_acad = get_acad_process_id('acad.exe')
        handle = findwindows.find_windows(process=jincheng_acad)[0]

        app = Application().connect(handle=handle)
        time.sleep(2)

        try:
            window = app.window(title="图形导出", class_name="#32770")
        except Exception as e:
            print("尝试重新获取窗口:", e)
            time.sleep(2)
            window = app.window(title="图形导出", class_name="#32770")

        if window.exists():
            print("窗口存在")
            child_windows = window.children()

            for child_window in child_windows:
                try:
                    child_window.select("天正7文件 (*.dwg) ")
                    print("正在导出t7")
                    pass
                except Exception as e:
                    pass
        else:
            print("窗口不存在")
            pass

        save_button = window.child_window(title="保存(&S)", class_name="Button")

        try:
            save_button.set_focus()
            
            save_button.click()


        except Exception as e:
            print("pywinauto处理窗口出现问题:", e)
            time.sleep(2)
            save_button.set_focus()
            save_button.click()

        print("已使用pywinauto自动化窗口")
        pass

    except Exception as e:
        print(f"automate_window_with_pywinauto_t7 出现错误: {e}")
        pass
    finally:
        pythoncom.CoUninitialize()
        timeout_event.set()
        event.set()  # 通知主线程完成

#&&% # 天正转t7    
def zhuancheng_t7():
    
    """
    在li()连接激活文件后，直接执行该命令即可转换
    一般理解所得文件和当前文件同文件夹，但这次测试结果却在文件夹
    C:/Users/Administrator/Documents/WeChat Files/wxid_mhrzdlppqacc22/FileStorage/File/2025-04/
    """    
    t1 = time.time()

    chaoshibiaoji = 1

    # 创建一个Event，用于通知子线程终止
    timeout_event = threading.Event()
    
    # 创建一个Event，用于通知主线程完成
    event = threading.Event()

    # 创建两个线程，分别执行run_cad_program和automate_window_with_pywinauto_t7
    thread1 = threading.Thread(target=run_cad_program, args=(timeout_event, event))
    thread2 = threading.Thread(target=automate_window_with_pywinauto_t7, args=(timeout_event, event))

    thread1.start()
    thread2.start()

    thread1.join(timeout=180)
    thread2.join(timeout=180)

    if not event.is_set():
        print("操作超时，正在中断...")

        pass

        
        # 设置timeout_event，通知run_cad_program()终止
        timeout_event.set()

    print("文件转T7格式操作结束")

    t2 = time.time()
    print("文件转T7格式操作总共用时", t2 - t1, "秒")


# 创建一个事件对象

timeout_event = threading.Event()

event = threading.Event()


def automate_window_with_pywinauto_t3(timeout_event, event):
    pythoncom.CoInitialize()

    try:
        jincheng_acad = get_acad_process_id('acad.exe')
        handle = findwindows.find_windows(process=jincheng_acad)[0]

        app = Application().connect(handle=handle)
        time.sleep(2)

        try:
            window = app.window(title="图形导出", class_name="#32770")
        except Exception as e:
            print("尝试重新获取窗口:", e)
            time.sleep(2)
            window = app.window(title="图形导出", class_name="#32770")

        if window.exists():
            print("窗口存在")
            pass
            child_windows = window.children()

            for child_window in child_windows:
                try:
                    child_window.select("天正3文件 (*.dwg) ")
                    print("正在导出t3")
                    pass
                except Exception as e:
                    pass
        else:
            print("窗口不存在")
            pass

        save_button = window.child_window(title="保存(&S)", class_name="Button")

        try:
            save_button.set_focus()
            save_button.click()
        except Exception as e:
            print("pywinauto处理窗口出现问题:", e)
            pass
            time.sleep(2)
            save_button.set_focus()
            save_button.click()

        print("已使用pywinauto自动化窗口")
        pass

    except Exception as e:
        print(f"automate_window_with_pywinauto_t3 出现错误: {e}")
        pass
    finally:
        pythoncom.CoUninitialize()
        timeout_event.set()
        event.set()  # 通知主线程完成
#&&% # 天正转t3
        
def zhuancheng_t3():

    """
    在li()连接激活文件后，直接执行该命令即可转换
    一般理解所得文件和当前文件同文件夹，但这次测试结果却在文件夹
    C:/Users/Administrator/Documents/WeChat Files/wxid_mhrzdlppqacc22/FileStorage/File/2025-04/
    """    

    t1 = time.time()

    chaoshibiaoji = 1

    # 创建一个Event，用于通知子线程终止
    timeout_event = threading.Event()
    
    # 创建一个Event，用于通知主线程完成
    event = threading.Event()

    # 创建两个线程，分别执行run_cad_program和automate_window_with_pywinauto_t3
    thread1 = threading.Thread(target=run_cad_program, args=(timeout_event, event))
    thread2 = threading.Thread(target=automate_window_with_pywinauto_t3, args=(timeout_event, event))

    thread1.start()
    thread2.start()

    thread1.join(timeout=180)
    thread2.join(timeout=180)

    if not event.is_set():
        print("操作超时，正在中断...")

        pass

        
        # 设置timeout_event，通知run_cad_program()终止
        timeout_event.set()

    print("文件转T3格式操作结束")

    t2 = time.time()
    print("文件转T3格式操作总共用时", t2 - t1, "秒")


#  主函数
#  (2)
# 确定一个文件中所有的对象类型

#  该函数系列包括如下一些函数


def get_all_object_types():
    """
    获取当前文件中模型空间中所有图形对象的类型名称（如 AcDbText, AcDbLine 等）

    返回：
        类型名列表（去重，已排序）
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument
    modelspace = doc.ModelSpace

    types_set = set()

    for obj in modelspace:
        try:
            types_set.add(obj.ObjectName)
        except:
            continue

    types_list = sorted(types_set)
    print(f"✅ 共发现 {len(types_list)} 种对象类型：")
    for t in types_list:
        print(f" - {t}")

    return types_list



#  主函数
#  (3)
# 关闭文件名为Name的文件

#  该函数系列包括如下一些函数

def close_dwg_by_name(Name):
    """
    关闭当前桌面中名为 Name 的 DWG 文件。
    如果文件已打开，则关闭该文件。
    
    参数：
        Name: 要关闭的 DWG 文件的名称（如 "example.dwg"）不含路径的名
    """
    try:
        acad = win32com.client.Dispatch("AutoCAD.Application")  # 获取 AutoCAD 应用
        doc = acad.Documents.Item(Name)  # 获取指定名称的文档
        
        if doc:
            doc.Close(False)  # 关闭文件，不提示保存
            print(f"✅ 文件 '{Name}' 已关闭")
        else:
            print(f"❌ 未找到名为 '{Name}' 的文件")
    except Exception as e:
        print(f"❌ 关闭文件 '{Name}' 失败: {e}")




def set_active_doc(doc):# 设置文件doc为激活对象

    """
    doc为com对象
    将指定文档 doc 设置为当前激活窗口。
    将其中一个打开时，要使用li()，然后执行doc1=acad.ActiveDocument
    再打开另一个，使用li()连接，执行doc2=acad.ActiveDocument就实际上获取了两个文件com实体对象
    然后就可以使用这个函数随时激活其中一个文件，激活后仍然要使用li()连接才能正确运行
    """
    try:
        doc.Activate()
        time.sleep(0.3)  # 稍作延时，确保激活生效
        print("✅ 当前激活文档：", doc.Name)
        return True
    except Exception as e:
        print("❌ 激活文档失败：", e)
        return False


def get_doc_by_name(name): #从文件名获取文件对象
    """
    通过文件名获取 AutoCAD 文档对象，例如 '空白.dwg'
    如果未找到，返回 None
    """
    for doc in acad.Documents:
        if doc.Name.lower() == name.lower():
            return doc
    return None

def get_open_document_names():#返回所有打开的文件名
    return [doc.Name for doc in acad.Documents]



#&&% 当前CAD文件数

def dwgs_count():#当前桌面的dwg文件数量

    shu = acad.Documents.Count

    return shu



#  主函数
#  (3)
# 跨文件复制粘贴

#&&%  文件按打印区域分解成多个文件


def export_region_to_new_file(x1, y1, x2, y2, filename,ty=1):##将区域(x1,y1,x2,y2)内的对象原位剪切粘贴到一个新文件中。
    """
    将当前文件中指定窗口区域内的图形对象，剪切粘贴到一个新 DWG 文件中，并保存为 filename.dwg。
    最后关闭新建文件、重新连接当前文件。
    """


    # 聚焦区域
    shitu_region(x1,y1,x2,y2)


    # 步骤 1：选择区域对象并转为蓝色夹点选中


    
    li()
    LB = select_objects_in_window_area(x1, y1, x2, y2)
    if not LB:
        print("❌ 区域内未找到任何对象，操作中止")
        return

    yin_to_xian_xuanze(LB,ty=ty)  # 关键：转换为命令行状态选中

    # 步骤 2：复制选中对象
    
    
    original_doc = doc  # 记录原始文档句柄

    doc.SendCommand("_copybase" + chr(13) + "0,0,0" + chr(13) + chr(13))
    time.sleep(3)

    # 步骤 3：新建 DWG 文件并连接
    acad.Documents.Add()
    time.sleep(1)
    li()
    new_doc = acad.ActiveDocument

    # 步骤 4：粘贴对象
    new_doc.SendCommand("TPasteClip" + chr(13) + "0,0,0" + chr(13) + chr(13))
    time.sleep(3)

    #视点聚焦
    shitu_region(x1,y1,x2,y2)

    # 步骤 5：保存新文件
    original_path = original_doc.FullName
    folder = os.path.dirname(original_path)
    save_path = os.path.join(folder, filename + ".dwg")


    new_doc.SaveAs(save_path)
    print(f"✅ 对象已导出到：{save_path}")

    # ✅ 新增部分 ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓

    # 步骤 6：关闭新建文件
    new_doc.Close(False)  # 不提示保存

    # 步骤 7：重新连接原始文件
    li()
    time.sleep(1)
    active_doc = acad.ActiveDocument

    



def batch_export_regions_to_files(region_list, filename_prefix="区域导出",ty=1):##文件分解成多个小文件存储
    """
    批量导出多个区域为 DWG 文件。

    参数：
        region_list: 列表，每项为一个 (x1, y1, x2, y2) 元组
        filename_prefix: 保存文件名前缀
        ty---通过删除再恢复对象以将隐性选择转变为显性选择，留一个延迟时间控制参数
    """
    #确保文件夹中名含"区域导出"的文件被删除

    folder_1 = current_dwg_folder()
    clear_files_with_prefix(folder_1, filename_prefix="区域导出", delay = ty)
    time.sleep(ty)



    name1 = current_dwg_basename()

    for idx, region in enumerate(region_list):
        if len(region) != 4:
            print(f"❌ 区域第 {idx+1} 个格式错误，跳过。")
            continue

        x1, y1, x2, y2 = region
        filename = f"{filename_prefix}_{idx+1:02d}"
        filename = name1 + filename

        chongfu_caozuo(
            export_region_to_new_file,
            dwg_instance=None,
            args=(x1, y1, x2, y2, filename),
            kwargs={'ty':ty},
            max_retries=3,
            failure_value=None
        )




#&&% #双文件操作

def copy_region_from_doc1_to_doc2_absolute(doc1, doc2, x1, y1, x2, y2):#从doc1的一个区域原位粘贴到doc2
    """
    将 doc1 中指定区域 (x1, y1)-(x2, y2) 内的对象复制，
    粘贴到 doc2 的相同位置 (0,0,0)。（原位复制）

    当未能完全控制激活文件时，这些函数运行并不稳定
    
    要求：
    - li() 可更新当前全局连接
    - select_objects_in_window_area() 返回 COM 对象列表
    - yin_to_xian_xuanze() 将对象转为命令行选中状态
    """
    try:
        # Step 1: 激活并连接源文档
        set_active_doc(doc1)
        li()

        # Step 2: 选择区域内对象
        LB = select_objects_in_window_area(x1, y1, x2, y2)
        if not LB:
            print("❌ 区域内未找到对象，复制中止")
            return

        # Step 3: 将对象蓝色选中状态
        yin_to_xian_xuanze(LB)

        # Step 4: 执行 _copybase 和 _copyclip
        doc1.SendCommand("_copybase" + chr(13) + "0,0,0" + chr(13) + chr(13))
        time.sleep(2)
        
        """        
        如果出现粘贴回源文档的混乱，可以考虑将源文件在已经执行了set_active_doc(doc2)后关闭doc1

        """        

        # Step 5: 激活并连接目标文档
        set_active_doc(doc2)
        li()

        # Step 6: 粘贴到相同位置
        doc2.SendCommand("TPasteClip" + chr(13) + "0,0,0" + chr(13) + chr(13))#_pastclip效果一样
        time.sleep(1.5)

        print(f"✅ 区域对象已从 {doc1.Name} 粘贴到 {doc2.Name}")

    except Exception as e:
        print(f"❌ 复制区域对象失败: {e}")

def copy_region_from_doc1_to_doc2_relative(doc1, doc2, x1, y1, x2, y2):#从doc1区域零点粘贴到doc2整个空间零点
    """
    将 doc1 中指定区域 (x1, y1)-(x2, y2) 内的对象复制，
    粘贴到 doc2 的相同位置 (0,0,0)。（原位复制）
    
    要求：
    - li() 可更新当前全局连接
    - select_objects_in_window_area() 返回 COM 对象列表
    - yin_to_xian_xuanze() 将对象转为命令行选中状态
    """
    try:
        # Step 1: 激活并连接源文档
        set_active_doc(doc1)
        li()

        # Step 2: 选择区域内对象
        LB = select_objects_in_window_area(x1, y1, x2, y2)
        if not LB:
            print("❌ 区域内未找到对象，复制中止")
            return

        # Step 3: 将对象蓝色选中状态
        yin_to_xian_xuanze(LB)

        # Step 4: 执行 _copybase 和 _copyclip
        doc1.SendCommand("_copybase\n0,0,0\n\n")
        time.sleep(0.5)
        doc1.SendCommand("_copyclip\n\n")
        time.sleep(2)

        
        """        
        如果出现粘贴回源文档的混乱，可以考虑将源文件在已经执行了set_active_doc(doc2)后关闭doc1

        """        


        # Step 5: 激活并连接目标文档
        set_active_doc(doc2)
        li()

        # Step 6: 粘贴到相同位置
        doc2.SendCommand("_pasteclip\n0,0,0\n\n")
        time.sleep(1.5)

        print(f"✅ 区域对象已从 {doc1.Name} 粘贴到 {doc2.Name}")

    except Exception as e:
        print(f"❌ 复制区域对象失败: {e}")


def copy_region_from_doc1_to_doc2_at_point(doc1, doc2, x1, y1, x2, y2, x0, y0, z0=0):#将doc1指定区域的最低点，粘贴到doc2的目标点
    """
    将 doc1 中指定区域 (x1, y1)-(x2, y2) 内的对象复制，
    粘贴到 doc2 的指定位置 (x0, y0, z0)。
    """
    try:
        # Step 1: 激活并连接源文档
        set_active_doc(doc1)
        li()

        # Step 2: 选择区域内对象
        LB = select_objects_in_window_area(x1, y1, x2, y2)
        if not LB:
            print("❌ 区域内未找到对象，复制中止")
            return

        # Step 3: 蓝色高亮
        yin_to_xian_xuanze(LB)

        # Step 4: 拷贝
        doc1.SendCommand("_copybase\n0,0,0\n\n")
        time.sleep(0.5)
        doc1.SendCommand("_copyclip\n\n")
        time.sleep(2)

        # Step 5: 激活目标文档
        set_active_doc(doc2)

        # 尝试多次连接目标文档
        for _ in range(5):
            try:
                li()
                break
            except:
                time.sleep(0.5)
        else:
            print("❌ 激活目标文档失败，无法连接")
            return

        # Step 6: 粘贴到指定点
        paste_cmd = f"_pasteclip\n{x0},{y0},{z0}\n\n"
        doc2.SendCommand(paste_cmd)
        time.sleep(1.5)

        print(f"✅ 区域对象已从 {doc1.Name} 粘贴到 {doc2.Name} 的指定点 ({x0}, {y0}, {z0})")

    except Exception as e:
        print(f"❌ 复制区域对象失败：{e}")






def insert_as_block(p,block_path = r"D:/Myprogramsystem/XT/MC_yuan.dwg"):#以块的插入将MC_yuan.dwg插入当前激活文件
    """
    将 MC_yuan.dwg 文件以块形式插入当前文档，并立即分解。

    MC_yuan.dwg 文件分别将9种门窗基元放入各自的图层

    jz-danmen jz-shuangmen jz-tuilamen jz-juanlianmen jz-zimumen jz-pingchuang jz-tuchuang jz-baiyechuang jz-gaochuang

    墙都在jizhunwall
    p   (x,y,z)三维坐标
    
    """
    
    if not os.path.exists(block_path):
        print("❌ 文件不存在：", block_path)
        return

    ensure_layer(layer_name="jizhunwall")

    x=p[0]

    y=p[1]

    z=p[2]

    # 插入块（插入点0,0,0，比例1）
    cmd = f"-insert\n{block_path}\n{x},{y},{z}\n1\n1\n0\n"
    doc.SendCommand(cmd)

    # 稳定等待图层对象出现（最多等 3 秒）
    LB = []
    for _ in range(30):
        time.sleep(0.1)
        LB = select_tuceng("jizhunwall")
        if LB:
            break
    else:
        print("❌ 图层中未能及时检测到对象")
        return

    print("选到的jizhutuceng对象数量", len(LB))

    block = LB[0]

    for attempt in range(3):
        try:
            block.Explode()
            print(f"✅ 第 {attempt+1} 次炸块成功")
            break
        except Exception as e:
            print(f"⚠️ 第 {attempt + 1} 次炸块失败：{e}")
            time.sleep(0.2)
    else:
        print("❌ 多次尝试炸块失败")
        return

    try:
        block.Delete()
        print("🗑️ 原块对象已删除")
    except:
        print("⚠️ 删除原块失败")



#&&&&%%  第九部分 CAD图块 

##给属性块赋予新值，可以局部赋值


def huoqukuai_shuxing_zhi(XX):#XX为属性块实体


    attributes = XX.GetAttributes()

    tags=[]
    values=[]

    for attr in attributes:
               
        tag = attr.TagString

            
        value = attr.TextString

        tags.append(tag)

        values.append(value)

    return tags,values


def set_attributes_values(block, tags_order, new_values):
    """
    为属性块的标签设置新的值。

    参数:
    - block: 要修改的属性块（COM BlockReference 对象）。
    - tags_order: 一个列表，包含你希望按照哪种顺序为属性块的标签设置新的值。
    - new_values: 一个列表，包含按标签顺序排列的新值。

    返回:
    None
    """
    # 先尝试获取属性列表
    try:
        attributes = block.GetAttributes()
    except Exception as e:
        print(f"⚠ 实体 {block.ObjectName}({getattr(block, 'Handle', '?')}) 无法获取属性，跳过: {e}")
        return

    index = 0
    for tag in tags_order:
        # 找到对应标签的属性
        found = False
        for attr in attributes:
            if attr.TagString == tag:
                found = True
                try:
                    attr.TextString = new_values[index]
                    print(f"标签: {tag}  新值: {new_values[index]}")
                except Exception as e:
                    print(f"⚠ 设置标签 '{tag}' 时出错: {e}")
                index += 1
                break
        if not found:
            print(f"⚠ 未找到属性标签 '{tag}'，已跳过")

    # 更新块
    try:
        block.Update()
    except Exception as e:
        print(f"⚠ 更新块时出错: {e}")              

##>>> tags_order=["1.0","施工图","2023.10","1:100","1.0","专业名称"]
##>>> new_values=["1.2版","初步设计","2021.07","1:25","JS-09","建施"]



def resize_block_attribute(block_ref, tag: str, *, height: float = 200.0, width: float = 4500.0):
    """
    将块参照 block_ref 中指定 Tag 的属性文字改成给定字高并设置边界宽度。

    适用对象
    --------
    block_ref : AcadBlockReference
        必须是包含属性 (HasAttributes=True) 的块参照
    tag       : str
        目标属性 TagString（不区分大小写）
    height    : float
        目标字高（Drawing Units）
    width     : float
        多行属性 (MText attribute) 的边界宽度；若属性不是多行，
        尝试设置 WidthFactor 以近似效果。

    返回
    ----
    bool
        True  : 至少找到并修改了一个属性
        False : 没有找到指定 tag 或调整失败
    """
    if (getattr(block_ref, "ObjectName", "") != "AcDbBlockReference"
            or not getattr(block_ref, "HasAttributes", False)):
        print("⚠ 传入对象不是带属性的块参照")
        return False

    modified = False
    target_tag = tag.strip().upper()

    try:
        for attr in block_ref.GetAttributes():
            if attr.TagString.strip().upper() != target_tag:
                continue

            # ——— 字高 ———
            try:
                attr.Height = height
            except Exception as e:
                print(f"⚠ 设置 Height 失败: {e}")

            # ——— 边界宽度 / 宽度因子 ———
            # 多行属性是 AcDbAttributeReference，但内核中仍带 MText，
            # COM 暴露 'Width'；若没有就退而求其次改 WidthFactor
            if hasattr(attr, "Width"):
                try:
                    attr.Width = width
                except Exception as e:
                    print(f"⚠ 设置 Width 失败: {e}")
            else:
                try:
                    # 估算一个宽度因子使单行文本占据近似宽度
                    # 【经验】WidthFactor * 字符数 * 字高 ≈ 宽度
                    char_count = max(len(attr.TextString.replace("\\P", "")), 1)
                    wf = width / (char_count * height)
                    attr.WidthFactor = wf
                except Exception as e:
                    print(f"⚠ 设置 WidthFactor 失败: {e}")

            modified = True
    except Exception as e:
        print(f"⚠ GetAttributes() 失败: {e}")

    return modified










## 3 获取属性块里多段线矩形的坐标值

def huoqu_kuai_pl(blocka):#输入实体块，得到实体块中多段线矩形的坐标，其坐标以插入点的定义点为原点
    # 连接到AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    kuaiming=blocka.Name

    # 获取块定义
    block_def = doc.Blocks.Item(str(kuaiming))

    # 获取块定义中的所有对象
    block_objects = list(block_def)

    # 查找三角形并删除
    for obj in block_objects:

##        print(obj.ObjectName)
        
        if obj.ObjectName == "AcDbPolyline":

            print(obj.Coordinates)

##>>> huoqu_kuai_pl(kuai[0])
##AcDbPolyline
##(-6000.0, 0.0, 0.0, 0.0, 0.0, 57400.0, -6000.0, 57400.0)
## 



#定义基点的块

def create_block_with_basepoint():
    # 连接到AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # 定义块的基点位置
    base_point = vtpnt(10, 10, 0)

    # 创建一个新的块
    block = doc.Blocks.Add(base_point, "MyBlock")

    # 在块中添加一个圆形实体
    block.AddCircle(base_point, 5)

#块的添加


def create_block_with_triangle_and_text():
    # 连接到AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # 创建新块
    grip = vtpnt(0, 0)
    blockObj = doc.Blocks.Add(grip, "MyBlock")

    # 在块中添加三角形
    pt1 = vtpnt(0, 0, 0)
    pt2 = vtpnt(10, 0, 0)
    pt3 = vtpnt(5, 10, 0)
    blockObj.AddLine(pt1, pt2)
    blockObj.AddLine(pt2, pt3)
    blockObj.AddLine(pt3, pt1)

    # 在块中添加文字对象
    text_point = vtpnt(2, 2, 0)
    blockObj.AddText("太美了", text_point, 2)

    print("块 'MyBlock' 创建成功")



def huoqu_kuai_pl(blocka):
    # 连接到AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    kuaiming=blocka.Name

    # 获取块定义
    block_def = doc.Blocks.Item(str(kuaiming))

    # 获取块定义中的所有对象
    block_objects = list(block_def)

    # 查找三角形并删除
    for obj in block_objects:

        print(obj.ObjectName)
        if obj.ObjectName == "AcDbPolyline":

            print(obj.Coordinates)

    

# 块的边界

def get_bounding_box_of_block(block_name):
    # 连接到AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # 获取块定义
    block_def = doc.Blocks.Item(block_name)

    min_x, min_y, min_z = float('inf'), float('inf'), float('inf')
    max_x, max_y, max_z = float('-inf'), float('-inf'), float('-inf')

    # 遍历块定义中的所有对象
    for obj in block_def:
        try:
            # 尝试获取对象的边界框
            lower_left, upper_right = obj.GetBoundingBox()
            
            min_x = min(min_x, lower_left[0])
            min_y = min(min_y, lower_left[1])
            min_z = min(min_z, lower_left[2])
            
            max_x = max(max_x, upper_right[0])
            max_y = max(max_y, upper_right[1])
            max_z = max(max_z, upper_right[2])
        except:
            pass

    return ((min_x, min_y, min_z), (max_x, max_y, max_z))


def create_new_block_with_insert_and_line():
    # 连接到AutoCAD
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # 检查块名称"块1"是否已经存在
    if "块3" in [blk.Name for blk in doc.Blocks]:
        print("块名称'块3'已经存在。请选择一个新的名称或删除现有的块。")
        return

    # 创建新块的插入点
    grip = vtpnt(0, 0, 0)
    blockObj1 = doc.Blocks.Add(grip, "块3")

    # 在块1中插入MyBlock块
    insertion_point_for_myblock = vtpnt(10, 10, 0)
    blockObj1.InsertBlock(insertion_point_for_myblock, "MyBlock", 1, 1, 1, 0)

    # 在块1中添加一根直线段
    start_point = vtpnt(0, 0, 0)
    end_point = vtpnt(50, 50, 0)
    blockObj1.AddLine(start_point, end_point)

    print("块1已创建并添加了MyBlock和直线段")

##块的搜索
    
def copy_and_move_blocks_from_layer(layer_name, block_prefix):
    
        
    # 使用select_tuceng函数选择指定图层上的所有对象
    all_objects = select_tuceng(layer_name)
    
    # 过滤出块对象，且块名的前两个字母与指定的前缀匹配
    blocks = [obj for obj in all_objects if obj.ObjectName == "AcDbBlockReference" and obj.Name[:2] == block_prefix]
    
    # 定义移动的起始点和结束点
    vtpnt_from = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0, 0, 0])
    vtpnt_to = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0, 2000000, 0])
    
    # 对每个块进行复制和移动操作
    for block in blocks:
        # 复制块
        copied_block = block.Copy()
        
        # 移动复制的块
        copied_block.Move(vtpnt_from, vtpnt_to)

    print(f"Copied and moved {len(blocks)} blocks from layer {layer_name} with prefix {block_prefix}.")

#块名的清除

def delete_block_name(block_name):#先删除块名为block_name的实体再执行该命令删除块名，免得将来发生替换警告

    t1=time.time()

    # 获取块定义的集合
    blocks = acad.ActiveDocument.Database.Blocks

    print("len(blocks)",len(blocks))

    # 遍历所有块定义
    for block in blocks:
        # 检查块名称是否匹配
        if block.Name == block_name:
            try:
                # 尝试删除块
                block.Delete()
                print(f"块 '{block_name}' 已被删除。")
            except Exception as e:
                print(f"删除块 '{block_name}' 时发生错误: {str(e)}")
            break
    else:
        # 如果未找到块
        print(f"未找到名为 '{block_name}' 的块。")

    t2 = time.time()

    print("删除块名耗时：",t2-t1)

#更改实体块名

def rename_block_entity(ent, new_name):
    """
    将给定块参照实体 ent 的块名改为 new_name。
    如果 new_name 在 Block 表中已存在，则该实体将指向已有定义；
    否则将重命名它当前所引用的块定义。
    
    参数：
      ent       -- 一个 COM 块参照对象（如 BlockReference）
      new_name  -- 目标块名（字符串）
    """
    # 获取当前文档和块表

    blocks = doc.Blocks
    old_name = ent.Name

    try:
        # 尝试查找 new_name 是否已存在
        blocks.Item(new_name)
        # 存在：直接让该实体引用此定义
        ent.Name = new_name
    except Exception:
        # 不存在：重命名它当前所引用的定义
        blk_def = blocks.Item(old_name)
        blk_def.Name = new_name

#&&% 由块名选择实例

def get_block_instances(block_name: str, max_retries: int = 5):
    """
    根据给定的块定义名，检索当前图形中所有对应的块参照实例（BlockReference），
    返回这些 COM 对象的列表。

    参数：
      block_name     – 块定义的名称（字符串），如 "MyBlock"
      max_retries    – 调用 select_kuai 时的最大重试次数

    返回：
      instances     – 包含所有匹配块参照的列表（如果未找到或者出错，返回空列表）
    """
    # ① 先调用 select_kuai() 拿到所有块实例 COM 对象（扁平列表）
    try:
        all_blocks = select_kuai(max_retries)
    except Exception as e:
        print(f"❌ 调用 select_kuai 失败：{e}")
        return []

    instances = []
    # ② select_kuai 已经是所有块实例的列表，直接遍历
    for ent in all_blocks:
        try:
            # 对于块参照，EntityName 通常是 "AcDbBlockReference"
            # 且我们要筛选 Name 恰好等于 block_name 的那些
            if getattr(ent, "EntityName", "") == "AcDbBlockReference" and getattr(ent, "Name", "") == block_name:
                instances.append(ent)
        except Exception:
            # 如果某个 COM 对象没有 EntityName/Name 属性，就跳过
            continue

    print(f"选择到名为{block_name}的实例块{len(instances)}个")

    return instances



#&&% 从块实体对象获取其内部com对象

def get_entities_from_block_reference(block_ref):
    """
    获取块引用对象中的所有子实体（COM对象形式）。

    参数：
        block_ref: 块引用对象（AcDbBlockReference）
        doc: 当前 AutoCAD 文档对象（Document）

    返回：
        entities: 子实体列表
    """
    try:
        block_name = block_ref.EffectiveName
        block_def = doc.Blocks.Item(block_name)
        entities = [ent for ent in block_def]
        print(f"✅ 获取到 {len(entities)} 个子对象")
        return entities
    except Exception as e:
        print(f"❌ 获取失败：{e}")
        return []

    




#以块的方式插入文件

def insert_block_into_autocad(block_file_path, insertion_point=(0, 0, 0), scale=(1, 1, 1), rotation=0):
    """
    以块的方式插入 DWG 文件到 AutoCAD 中

    :param block_file_path: 块文件的路径，通常为 DWG 文件路径
    :param insertion_point: 插入位置的三元组 (x, y, z)
    :param scale: 块的缩放比例三元组 (sx, sy, sz)
    :param rotation: 块的旋转角度（弧度）
    """
    try:

        # 定义块插入点
        insertion_point = (insertion_point[0], insertion_point[1], insertion_point[2])

        # 插入块，使用 InsertBlock 方法
        block = ms.InsertBlock(insertion_point, block_file_path, scale[0], scale[1], scale[2], rotation)

        print(f"✅ 块已插入，文件：{block_file_path}，插入点：{insertion_point}，缩放：{scale}，旋转角度：{rotation}")

    except Exception as e:
        print(f"❌ 插入块时出错：{e}")



#不炸开

def insert_standard_block(block_dwg,
                          insertion_point=(0, 0, 0),
                          scale=(1, 1, 1),
                          rotation=0,
                          wait=0.3):
    """
    不炸开，
    全程无交互对话框。

    block_dwg: 标准块 DWG 路径
    insertion_point: (x,y,z)
    scale: (sx,sy,sz)
    rotation: 旋转角度（度）
    """
    before = select_kuai()
    before_handles = {b.Handle for b in before}


    if not os.path.isfile(block_dwg):
        raise FileNotFoundError(block_dwg)

    # 准备参数
    x, y, z    = insertion_point
    sx, sy, sz = scale
    path       = os.path.abspath(block_dwg).replace("\\", "/")

    # 1) 插入块
    insert_cmd = (
        "-INSERT\n"
        f"\"{path}\"\n"    # 文件路径要加双引号
        f"{x},{y},{z}\n"
        f"{sx}\n"
        f"{sy}\n"
        f"{sz}\n"
        f"{rotation}\n"
    )
    doc.SendCommand(insert_cmd)
    time.sleep(wait)

    doc.SendCommand("RE\n")
    doc.SendCommand("Z\nE\n")
    time.sleep(wait)

    after = select_kuai()
    new_refs = [b for b in after if b.Handle not in before_handles]
    if not new_refs:
        print("⚠ 未检测到任何新插入的块引用")
        return []

    results = []
    for blk in new_refs:
        # 5. 先将它旋转归零
        try:
            blk.Rotation = 0
        except Exception:
            pass

        # 7. 取它的包围盒四角
        p1, p2 = blk.GetBoundingBox()
        minx, miny, minz = p1
        maxx, maxy, maxz = p2
        corners = [
            (minx, miny, minz),  # 左下
            (minx, maxy, minz),  # 左上
            (maxx, maxy, minz),  # 右上
            (maxx, miny, minz),  # 右下
        ]

        results.append((blk, corners))

    return results


#炸开
def insert_and_explode_dwg(block_dwg,
                           insertion_point=(0, 0, 0),
                           scale=(1, 1, 1),
                           rotation=0,
                           wait=0.3):
    """
    将一个 WBLOCK 导出的标准块 DWG 插入到当前图，
    并立即 EXPLODE 成普通图元（不保留块引用）。

    参数:
        block_dwg: 标准块 DWG 路径
        insertion_point: 插入点 (x,y,z)
        scale: (sx,sy,sz)
        rotation: 旋转角度（度）
        wait: 每步命令后等待秒数
    """

    before = select_kuai()
    before_handles = {b.Handle for b in before}


    if not os.path.isfile(block_dwg):
        raise FileNotFoundError(block_dwg)

    # 准备参数
    x, y, z    = insertion_point
    sx, sy, sz = scale
    path       = os.path.abspath(block_dwg).replace("\\", "/")

    # 1) 插入块
    insert_cmd = (
        "-INSERT\n"
        f"\"{path}\"\n"    # 文件路径要加双引号
        f"{x},{y},{z}\n"
        f"{sx}\n"
        f"{sy}\n"
        f"{sz}\n"
        f"{rotation}\n"
    )
    doc.SendCommand(insert_cmd)
    time.sleep(wait)

    doc.SendCommand("RE\n")
    doc.SendCommand("Z\nE\n")
    time.sleep(wait)

    # 2) EXPLODE “Last”   （炸开最新插入的块引用）
    explode_cmd = (
        "EXPLODE\n"
        "L\n"    # Last
        "\n"     # 完成选择
    )
    doc.SendCommand(explode_cmd)
    time.sleep(wait)

    print(f"✅ 已插入并炸开：{os.path.basename(path)} @ ({x},{y},{z})")


    after = select_kuai()
    new_refs = [b for b in after if b.Handle not in before_handles]
    if not new_refs:
        print("⚠ 未检测到任何新插入的块引用")
        return []

    results = []
    for blk in new_refs:
        # 5. 先将它旋转归零
        try:
            blk.Rotation = 0
        except Exception:
            pass

        # 7. 取它的包围盒四角
        p1, p2 = blk.GetBoundingBox()
        minx, miny, minz = p1
        maxx, maxy, maxz = p2
        corners = [
            (minx, miny, minz),  # 左下
            (minx, maxy, minz),  # 左上
            (maxx, maxy, minz),  # 右上
            (maxx, miny, minz),  # 右下
        ]

        results.append((blk, corners))

    return results,blk


#&&% 双线程插入块


##确保炸开块并获取块内对象
def safe_explode_and_delete(bk, ci=3, delay=1.0):
    """
    对块对象 bk 执行安全的 Explode 与 Delete 操作：
      1. 最多尝试 ci 次调用 bk.Explode()，
         每次调用后等待 delay 秒，再检查返回的实体列表长度是否 > 1。
      2. 如果 ci 次都没有成功（len(LP) <= 1），抛出 RuntimeError。
      3. 如果 Explode 成功，最多尝试 ci 次调用 bk.Delete()，遇错则继续重试。
      4. 返回第一次成功 Explode 时得到的实体列表 LP。

    :param bk:    要炸开的块引用 COM 对象
    :param ci:    最大尝试次数，默认 3
    :param delay: 每次 Explode 后的等待时间（秒），默认 1.0
    :return:      成功 Explode 后返回的新实体列表 LP
    :raises:      RuntimeError 如果 Explode 在 ci 次内都失败
    """
    LP = []
    # 1️⃣ 重试 Explode
    for attempt in range(1, ci + 1):
        try:
            LP = bk.Explode()
        except Exception as e:
            LP = []  # 如果调用失败，视为没有炸开
        time.sleep(delay)

        # 尝试获取长度
        try:
            count = len(LP)
        except Exception:
            # 如果 LP 不是标准序列，就把它转换一下
            LP = list(LP)
            count = len(LP)

        if count > 1:
            # 炸开成功
            break
        else:
            # 第 attempt 次 Explode 未成功，继续重试
            continue
    else:
        raise RuntimeError(f"Explode 在 {ci} 次尝试后仍未成功 (len(LP)={len(LP)})")

    # 2️⃣ Explode 成功后，重试 Delete
    for attempt in range(1, ci + 1):
        try:
            bk.Delete()
            break
        except Exception:
            # Delete 出错，继续下次重试
            continue
    # 如果 ci 次都失败，也不抛错误，只是放弃删除
    return LP





#&&% 获取面积足够大的全部非同名块实例

def get_large_block_instances(
    area_threshold: float,
    tol: float = 100.0,
    max_retries: int = 5
) -> list:
    """
    获取模型空间中所有块实例，筛选出“包围盒面积大于 area_threshold” 的块，
    并按面积去重：如果两个块的面积差值小于 tol，则只保留其中一个。
    返回一个按出现顺序去重后的 COM 对象列表。

    :param area_threshold: 面积阈值（单位与 CAD 坐标相同，例如以平方图纸单位计）。
    :param tol:            面积去重的容差值（默认 100），若两个块面积差小于 tol，视为同一块。
    :param max_retries:    调用 select_kuai 时的最大重试次数，默认 5 次。
    :return:               包含所有“面积 > area_threshold”且去重后的块引用 COM 对象列表，
                          如果 select_kuai 调用失败，会返回空列表。
    """
    # ① 调用 select_kuai，每隔 max_retries 次重连一次
    try:
        all_blocks = select_kuai(max_retries)
    except Exception as e:
        print(f"❌ 调用 select_kuai 失败：{e}")
        return []

    # 临时列表，用于收集已加入结果的块面积（用于去重）
    seen_areas = []
    large_blocks = []

    for blk in all_blocks:
        try:
            # GetBoundingBox 返回 (ll_point, ur_point)
            ll_point, ur_point = blk.GetBoundingBox()
            x1, y1, _ = ll_point
            x2, y2, _ = ur_point
            width = abs(x2 - x1)
            height = abs(y2 - y1)
            area = width * height
        except Exception:
            # 如果某些块无法获取包围盒，则跳过
            continue

        # 面积判断
        if area > area_threshold:
            # 检查当前 area 是否已近似出现在 seen_areas 中
            duplicate = False
            for existing_area in seen_areas:
                if abs(area - existing_area) < tol:
                    duplicate = True
                    break

            if not duplicate:
                seen_areas.append(area)
                large_blocks.append(blk)

    return large_blocks

#从com对象中，根据其外包盒的矩形的长边与短边的比值和面积在160000000到1000000000两个条件筛算

#&&% 确定合乎标准打印要求的自建多段线区域

def get_large_block_instances_with_tolerance(max_retries: int = 5, area_threshold: float = 70 ):
    """
    获取当前 DWG 中所有大尺寸块实例
    A3 1:100的正常值> 1240000000.000000
   
    """
    # ① 尝试获取所有块实例
    try:
        all_blocks = select_kuai(max_retries)
    except Exception as e:
        print(f"❌ 调用 select_kuai 失败：{e}")
        return []

    # ② 先筛选面积 >= area_threshold 的块
    LB = []
    for blk in all_blocks:
        try:
            ll_point, ur_point = blk.GetBoundingBox()
            x1, y1, _ = ll_point
            x2, y2, _ = ur_point
            width = abs(x2 - x1)
            height = abs(y2 - y1)
            area = width * height
        except Exception:
            # 读取包围盒失败，跳过
            continue

        if area >= area_threshold:
            # 附带面积信息，便于后续比较
            LB.append((blk, area))

    return LC


#&&% 确定合乎广义标准打印要求的自建多段线区域




#&&% 块内坐标转换成世界坐标（适合平面上的一般块）

def transform_point_by_block(block_ref, local_pt):
    """
    将块内部坐标 local_pt = (lx, ly, lz) 转换为世界坐标：
      1. 按 block_ref 的比例缩放 (XScaleFactor, YScaleFactor, ZScaleFactor)
      2. 按 block_ref.Rotation 旋转（绕 Z 轴，弧度制）
      3. 平移到 block_ref.InsertionPoint = (ix, iy, iz)

    返回 (wx, wy, wz)：
      wx = ix + (lx * sx * cosθ - ly * sy * sinθ)
      wy = iy + (lx * sx * sinθ + ly * sy * cosθ)
      wz = iz + (lz * sz)
    """
    ix, iy, iz = block_ref.InsertionPoint
    sx = block_ref.XScaleFactor
    sy = block_ref.YScaleFactor
    sz = block_ref.ZScaleFactor
    theta = block_ref.Rotation  # 单位：弧度

    lx, ly, lz = local_pt
    # 缩放后再旋转
    x_scaled = lx * sx
    y_scaled = ly * sy

    cos_t = math.cos(theta)
    sin_t = math.sin(theta)
    xr = x_scaled * cos_t - y_scaled * sin_t
    yr = x_scaled * sin_t + y_scaled * cos_t
    zr = lz * sz

    wx = ix + xr
    wy = iy + yr
    wz = iz + zr
    return (wx, wy, wz)


def select_block_by_name(block_name: str, max_retries: int = 5):
    """
    从 *模型空间* 快速选出指定块名的所有实例，返回实体列表。
    不遍历 ModelSpace，速度与 SelectionSet 相同量级。
    """
    t0, last_exc = time.time(), None

    for attempt in range(1, max_retries + 1):
        try:
            # 1) 删旧选择集
            with suppress(Exception):
                doc.SelectionSets.Item("SS_block_by_name").Delete()

            # 2) 新建选择集
            ss = doc.SelectionSets.Add("SS_block_by_name")

            # 3) 过滤器：0=INSERT, 2=块名, 410=布局名"Model"
            filterType  = vtInt([0, 2, 410])
            filterData  = vtVariant(["INSERT", block_name, "Model"])

            # acSelectionSetAll = 5
            ss.Select(5, 0, 0, filterType, filterData)

            lb = list(ss)
            print(f"✅ 选到 {len(lb)} 个 “{block_name}”（{time.time() - t0:.3f}s，第{attempt}次）")
            return lb

        except Exception as e:
            last_exc = e
            print(f"⚠ 第 {attempt} 次失败：{e!r}")
            time.sleep(0.3)

    print(f"❌ {max_retries} 次仍失败：{last_exc!r}")
    return []

def get_all_block_definitions(doc=None):
    """
    返回当前 DWG 文档中所有块定义（BlockTableRecord）列表。

    :param doc: AutoCAD Document 对象，若为 None 则从当前激活文档获取
    :return: list of BlockTableRecord COM 对象
    blk.Name获取块名

    """
    # 如果外部没传入 doc，就从当前激活文档拿
    if doc is None:
        acad = win32com.client.Dispatch("AutoCAD.Application")
        doc  = acad.ActiveDocument

    blocks = []
    count = doc.Blocks.Count  # 块定义总数
    # Blocks 集合在 COM 中是 0…Count−1 编号
    for i in range(count):
        try:
            blk = doc.Blocks.Item(i)
            blocks.append(blk)
        except Exception:
            # 跳过任何访问不成功的索引
            continue

    return blocks

def purge_block(block_name: str, quiet: bool = False):
    """
    删除指定块的所有实例，并彻底清除该块定义。
    
    步骤：
      1. 在模型空间删除所有同名 INSERT 实例
      2. 在每个布局的 Block (PaperSpace) 删除所有同名 INSERT 实例
      3. 调用 PurgeAll() 清理未用定义
      4. 再次尝试删除块定义
    
    :param block_name: 要清理的块名称（区分大小写）
    :param quiet: True 则不打印过程信息
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument

    # --- 1) 模型空间 ---
    removed = 0
    for ent in list(doc.ModelSpace):
        if ent.ObjectName in ("AcDbBlockReference", "AcDbMInsertBlock") and getattr(ent, "Name", "") == block_name:
            with suppress(Exception):
                ent.Delete()
                removed += 1

    # --- 2) 各布局块空间（PaperSpace） ---
    for layout in doc.Layouts:
        with suppress(Exception):
            block_space = layout.Block
            for ent in list(block_space):
                if ent.ObjectName in ("AcDbBlockReference", "AcDbMInsertBlock") and getattr(ent, "Name", "") == block_name:
                    with suppress(Exception):
                        ent.Delete()
                        removed += 1

    time.sleep(0.2)
    if not quiet:
        print(f"ℹ 共删除 {removed} 个 “{block_name}” 实例")

    # --- 3) PurgeAll 清理所有未被引用的定义 ---
    with suppress(Exception):
        doc.PurgeAll()
        if not quiet:
            print("✅ PurgeAll 清理未用定义")

    # --- 4) 删除块定义 ---
    try:
        blk = doc.Blocks.Item(block_name)
        blk.Delete()
        if not quiet:
            print(f"✅ 已删除块定义：{block_name}")
    except Exception as e:
        if not quiet:
            print(f"⚠ 删除块定义失败（仍有隐性引用？）：{e}")

    if not quiet:
        print(f"ℹ 完成对 '{block_name}' 的清理。")

def purge_unused_blocks(quiet: bool = False):
    """
    一次性清除所有未被任何 INSERT 实例引用的块定义。
    速度快、可靠性高（不用逐个 SelectionSet 检测）。
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument

    # 1) 记录清理前的块名列表
    before = []
    for i in range(doc.Blocks.Count):
        with suppress(Exception):
            name = doc.Blocks.Item(i).Name
            before.append(name)

    # 2) 调用 PurgeAll 一次性清理
    t0 = time.time()
    with suppress(Exception):
        doc.PurgeAll()
    t1 = time.time()

    # 3) 记录清理后的块名列表
    after = []
    for i in range(doc.Blocks.Count):
        with suppress(Exception):
            name = doc.Blocks.Item(i).Name
            after.append(name)

    # 4) 计算差集
    removed = [name for name in before if name not in after]

    if not quiet:
        print(f"✅ PurgeAll 清理完成，耗时 {t1 - t0:.3f}s")
        print(f"ℹ 共移除 {len(removed)} 个未使用块：")
        for nm in removed:
            print("   ·", nm)

    return removed


def get_selected_blockreference_names():
    """
    使用 pmxz() 选择实体，并返回所有块引用（AcDbBlockReference）的块名列表。

    :return: list of str，所选块引用的 Name 属性列表
    """
    try:
        # 调用已有函数获取当前选择集（返回 COM 对象列表）
        entities = pmxz()
    except Exception as e:
        print(f"❌ 调用 pmxz() 失败：{e}")
        return []

    block_names = []
    for ent in entities:
        try:
            # 只处理块引用
            if getattr(ent, "ObjectName", "") == "AcDbBlockReference":
                # Name 属性就是块名
                name = getattr(ent, "Name", None)
                if name:
                    block_names.append(name)
        except Exception as e:
            # 某些 COM 对象可能不支持上述属性，忽略它
            continue

    return block_names



def delete_layer(layername: str):
    """
    删除当前 DWG 文件中名为 layername 的图层。
    - 如果图层不存在，直接返回。
    - 如果图层是当前层，则切换到 0 层后再删除。
    - 删除前会尝试解锁、去掉冻结/打印锁。
    """
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument
    layers = doc.Layers

    try:
        layer = layers.Item(layername)
    except Exception:
        print(f"⚠ 图层 {layername} 不存在。")
        return

    # 如果该图层是当前层，切换到 "0"
    if doc.ActiveLayer.Name == layername:
        doc.ActiveLayer = layers.Item("0")

    # 解锁、解冻、去掉打印锁定，避免删除失败
    try:
        layer.Lock = False
        layer.Freeze = False
        layer.Plottable = True
    except Exception:
        pass

    try:
        layer.Delete()
        print(f"✅ 图层 {layername} 已删除。")
    except Exception as e:
        print(f"❌ 删除失败：{e}")

# 使用方法
# delete_layer("MyLayer")



#&&&&%% 第十部分  非图形对象处理


"""
该模块研究非图形对象处理问题 

"""


#  主函数
#  (1)
# 图层操作


#&&% *** 将屏幕所选对象赋予到指定图层

@alias("s1")
def sc_objs_to_layer(layer_name,cl=256):

    def pmxz_new():
        """
        人工选择对象，返回所选实体对象列表。
        自动清理已有的 "SS1" 选择集。
        """
        try:
            # 如果 "SS1" 已存在，先删除
            try:
                ss = doc.SelectionSets.Item("SS1")
                ss.Delete()
            except:
                pass  # 如果不存在就忽略

            selection = doc.SelectionSets.Add("SS1")
            selection.SelectOnScreen()
            objs = [selection.Item(i) for i in range(selection.Count)]
            selection.Delete()
            return objs
        except Exception as e:
            print("❌ 选择失败:", e)
            return []



    layers = doc.Layers

    try:
        layer = layers.Item(layer_name)
    except:
        layer = layers.Add(layer_name)
        print(f"🟢 已新建图层：{layer_name}")

    LB=pmxz_new()

    for x in LB:

        x.Layer = layer_name

        x.Color = cl


    return LB



#  该函数系列包括如下一些函数

def create_layers_from_list(layer_names):
    """
    创建列表中指定的图层，如果图层已存在则跳过。

    参数：
        layer_names: 包含图层名称的字符串列表
    """
    try:
        layers = doc.Layers
        created = 0
        skipped = 0

        for name in layer_names:
            try:
                _ = layers.Item(name)  # 检查是否已存在
                print(f"⏩ 图层已存在：{name}")
                skipped += 1
            except:
                layers.Add(name)
                print(f"✅ 新建图层：{name}")
                created += 1

        print(f"\n📊 总计：新建 {created} 个图层，跳过 {skipped} 个已有图层")

    except Exception as e:
        print("❌ 创建图层时出错：", e)


def ensure_layer(layer_name="jizhunwall"):
    """
    确保图层存在并设为当前图层，同时删除该图层上所有对象（最多重试 3 次）。
    """
    try:
        layers = doc.Layers
        # 1) 获取或新建图层
        try:
            layer = layers.Item(layer_name)
        except Exception:
            layer = layers.Add(layer_name)
            print(f"🟢 已新建图层：{layer_name}")
        # 2) 切换到图层
        doc.ActiveLayer = layer
        print(f"✅ 当前图层已设置为：{layer_name}")

        # 3) 删除图层中全部对象，重试 up to 5
        for attempt in range(1, 6):
            ents = select_tuceng(layer_name)
            if not ents:
                # 已经没有对象，提前退出
                print(f"🧹 图层 '{layer_name}' 已清空（共尝试 {attempt - 1} 次）")
                break

            deleted = 0
            for ent in ents:
                try:
                    ent.Delete()
                    deleted += 1
                except:
                    continue
            print(f"  第 {attempt} 次删除：共删除 {deleted} 个对象")

            time.sleep(0.1)  # 短暂等待，确保对象被移除


            #刷新

            doc.SendCommand("RE\n")
            doc.SendCommand("Z\nE\n")

        else:
            # 五次都还有残留
            remaining = len(select_tuceng(layer_name))
            print(f"⚠ 重试 3 次后，图层 '{layer_name}' 仍有 {remaining} 个对象未能删除")

    except Exception as e:
        print("❌ 创建/切换图层或清理失败：", e)


@alias("s2")

def ensure_layer_current(layer_name="jizhunwall", max_retries=3):
    """
    确保图层存在并设为当前图层，失败时最多重试 max_retries 次。
    """
    layers = doc.Layers
    for attempt in range(1, max_retries + 1):
        try:
            # 获取或新建图层
            try:
                layer = layers.Item(layer_name)
            except Exception:
                layer = layers.Add(layer_name)
                print(f"🟢 已新建图层：{layer_name}")
            # 切换到图层
            doc.ActiveLayer = layer
            print(f"✅ 当前图层已设置为：{layer_name} (尝试 {attempt})")
            return True
        except Exception as e:
            print(f"❌ 尝试 {attempt} 创建/设置图层失败：{e}")
    print(f"❌ 达到最大重试次数 ({max_retries})，无法创建或切换到图层：{layer_name}")
    return False


# 设置指定图层的颜色、线型、开关状态和冻结状态

@alias("s3")

def set_layer_properties(layer_name, color_index=9, linetype="Continuous", on=True, frozen=False):
    """
    设置指定图层的颜色、线型、开关状态和冻结状态。

    参数：
        layer_name (str): 图层名称
        color_index (int): 图层颜色索引（默认 9）
        linetype (str): 图层线型名称（默认 'Continuous'）
        on (bool): 图层是否打开（默认 True）
        frozen (bool): 图层是否冻结（默认 False）
    """
    try:
        layers = doc.Layers
        try:
            layer = layers.Item(layer_name)
        except:
            layer = layers.Add(layer_name)
            print(f"✅ 已新建图层：{layer_name}")

        # 设置颜色
        layer.color = color_index

        # 设置线型
        try:
            ltype = doc.Linetypes.Item(linetype)
        except:
            doc.Linetypes.Load(linetype, linetype)
            ltype = doc.Linetypes.Item(linetype)
        layer.Linetype = linetype

        # 设置开关状态
        layer.LayerOn = on

        # 设置冻结状态
        layer.Freeze = frozen

        print(f"🔧 图层属性已更新：{layer_name} | 颜色={color_index} | 线型={linetype} | 开关={'开' if on else '关'} | 冻结={'是' if frozen else '否'}")

    except Exception as e:
        print(f"❌ 设置图层属性失败：{e}")




#&&% 将列表中的对象图层设为目标图层

def set_layer_with_retry(LB, layername, ci=3):
    """
    将给定 COM 对象列表 LB 中的每个对象的 Layer 属性设为 layername。
    每次失败后等待 1 秒，最多尝试 ci 次。

    参数：
      LB         -- 可迭代的 COM 对象序列，每个对象支持设置 .Layer
      layername  -- 要设置的目标图层名称（字符串）
      ci         -- 每个对象最大重试次数（默认 3）

    返回：
      成功设置图层的对象列表，以及失败未能设置的对象列表：
        (success_list, failed_list)
    """
    success = []
    failed = []

    for obj in LB:
        for attempt in range(1, ci + 1):
            try:
                obj.Layer = layername
                success.append(obj)
                break
            except Exception as e:
                if attempt == ci:
                    # 最后一次仍失败
                    failed.append(obj)
                    print(f"⚠️ 对象 {getattr(obj, 'Handle', '<unknown>')} 设置图层“{layername}”失败：{e}")
                else:
                    time.sleep(1)
    return success, failed



#&&&&%% 第十一部分  天正DWG文件操作

### start


class DWG:
    def __init__(self, filepath=None, create_new=0):
        """
        参数:
          filepath: 指定的文件路径（可选）
          create_new: 标志，若为1则产生（新建或打开）DWG文件；若为0则不产生新文件，而使用当前活动文档。
                    默认值为0。
        """
        self.acad = win32com.client.Dispatch("AutoCAD.Application")
        self.acad.Visible = True
        if create_new:
            if filepath:
                self.doc = self.acad.Documents.Open(filepath)
            else:
                self.doc = self.acad.Documents.Add()
        else:
            # 不产生新文件，直接使用当前活动文档
            self.doc = self.acad.ActiveDocument

    def save(self, filepath=None):
        if filepath:
            self.doc.SaveAs(filepath)
        else:
            self.doc.Save()




    #&&%____________________________________________________    天正墙门窗      __________________________________________________
    #…………………………………………………………………………………………………………………………………………………………………
    # DWG类 - 天正墙门窗

    #__________________________________________________________________________________________________________________________
    #…………………………………………………………………………………………………………………………………………………………………
    #  使用说明

    """
    研究天正墙的基本操作等问题



    #dir(LB[0])查看属性

    ##天正门的有用信息清单

    ##有用的属性/方法清单（可读写或用于标注）：
    ##属性名	含义（推测）	类型
    ##Handle	对象唯一编号（✔ 可用）	str
    ##Layer	图层名（✔ 可读写）	str
    ##Color	颜色（✔ 可设置）	int
    ##Width	门宽（✔ 可用）	float
    ##Height	门高（✔ 可用）	float
    ##DoorSill	门槛高（✔）	float
    ##EntityName	实体名称（可能是 "门"）	str
    ##OutlineArea	投影面积（✔）	float
    ##Label	标签名或门类型名	str
    ##ProStyle	门型样式名	str
    ##Usage	用途（可能是开向、类型）	str
    ##SlotWidth	门缝宽度	float
    ##WinSill	门窗下口高度（如是组合构件）	float



    ##天正窗信息清单

    ##性名	含义说明	类型	可用性
    ##Handle	对象唯一编号（AutoCAD级）	str	✅ 必用
    ##Label	用户可写的编号标签（✔你可以赋值）	str	✅ 推荐
    ##Width	窗宽，单位为毫米	float	✅ 核心属性
    ##Height	窗高，单位为毫米	float	✅ 核心属性
    ##WinSill	窗台高度（距地）	float	✅ 设计用
    ##Usage	用途，如“飘窗”、“凸窗”	str	✅ 分类
    ##ProStyle	窗型样式，如“单扇窗”、“推拉窗”	str	✅ 分类
    ##Structure	构造信息（墙体厚度/嵌入方式）	str/int	⚠️ 不一定都有
    ##SlotWidth	槽宽（窗框槽）	float	可选
    ##SlabThickness	窗下板厚度	float	可选
    ##Layer	所在图层	str	✅ 管理用
    ##color	当前颜色	int	✅ 可控制
    ##Visible	是否可见（布尔值）	bool	✅ 筛选




    #天正墙信息清单

    ##✅ 高价值字段速查表（天正墙对象）
    ##属性名	含义（推测）	类型	可用性
    ##Handle	AutoCAD唯一标识符	str	✅ 必备
    ##Label	自定义编号标签（可读写）	str	✅ 推荐用于编号
    ##Layer	所在图层	str	✅ 管理用
    ##Height	墙高	float	✅ 用于统计
    ##StruHeight	结构墙高（可能不同于装饰墙）	float	✅ 分析结构层高
    ##TotalWidth	墙总厚度	float	✅ 常用于分类
    ##InsuThick	墙体保温层厚度	float	✅ 用于节能统计
    ##Material	墙体材料	str	✅ 分类使用
    ##Usage	用途，如“承重”、“隔墙”	str	✅ 类型分析
    ##IsArc / StruIsArc	是否为弧形墙	bool	✅ 图形过滤
    ##HoleArea	孔洞总面积（门窗洞口）	float	✅ 净面积核算
    ##Volume	墙体体积	float	✅ 工程量
    ##HoleVolume	洞口体积	float	✅ 净体积计算
    ##LSurfaceArea / RSurfaceArea	左/右立面面积	float	✅ 节能分析
    ##Visible	是否可见	bool	✅ 图形控制
    ##color	当前颜色	int	✅ 可调整



    LB[0].Handle
    '331'
    LB[1].Handle
    '40E'
    LB[2].Handle
    '41F'
    LB[3].Handle
    '438'
    LB[4].Handle
    '425'
    插入点都可以从BongdingBox的对角点数据换算出来
    LB=pmxz()
    请在屏幕拾取图元，以Enter键结束
    LB[0].Height
    2100.0
    LB[0].Width
    300.0
    LB[0].Label
    'FM甲0321'
    LB[0].DoorSill
    0.0
    LB[0].Label=""

    PL分析应该加上误差信息反馈，yintoxian对于非天正对象偏移值应该按5%计算，

    在整个空间在块中按色彩选择线将它们按pl加粗， 嵌套块处理 ，线型处理，

    天正标注，轴线，标高，引线，楼梯，其它一些菜单工具


    通过区域选择和对象类型过滤，即可精准选择到D:/Myprogramsystem/XT/MC_yuan.dwg文件中的门窗基元。

    门窗翻转只是一个镜像操作问题。属性参数可以从dir(LB[0])查询到。这样就可以自己编函数实现门窗编号，且对其进行根本控制统计。


    LB=select_objects_in_window_area(1014085.4304097779, 1029776.5683047362, 1013594.3737339601, 1030360.0671773665)
    ✅ 成功选择 2 个对象。
    LB[0].ObjectName
    'TDbOpening'
    LB[1].ObjectName
    'TDbWall'



    """


    #  主函数
    #  (1)
    # 绘制墙体

    #  该函数系列包括如下一些函数

    def draw_wall(self,start_point,end_point,width=240):## 绘制墙体

        """
        可以通过属性.TotalWidth,.LeftWidth,RightWidth确定墙宽度的信息。

        如果要修改某堵墙的宽度，可以先选择一个区域的墙实体，再通过端点坐标信息筛选出来使用命令行的TWallThick命令，对应极少见的左右墙厚不一致的情况交给人工即可

        """
        # 两点的中点 临时位置命令

        print("画墙开始")

        def midpoint(start_point,end_point):

            return ((start_point[0]+end_point[0])/2,(start_point[1]+end_point[1])/2,(start_point[2]+end_point[2])/2)

        Midstart_point = midpoint(start_point,end_point)

        Midend_point = (Midstart_point[0]-100,Midstart_point[1]+100,Midstart_point[2])


        # 使用 f-string 语法将三维坐标变量插入命令字符串中
        start_point_str = f"{start_point[0]},{start_point[1]},{start_point[2]}"

        end_point_str = f"{end_point[0]},{end_point[1]},{end_point[2]}"

        width_str = f"{width}"

        Midstart_point_str = f"{Midstart_point[0]},{Midstart_point[1]},{Midstart_point[2]}"        

        Midend_point_str = f"{Midend_point[0]},{Midend_point[1]},{Midend_point[2]}"        

        command = f"tgwall{chr(13)}{start_point_str}{chr(13)}{end_point_str}{chr(13)}{chr(13)}{chr(13)}"


        # 改墙宽
        
        acad.ActiveDocument.SendCommand(command)

        command = f"TWallThick{chr(13)}{Midstart_point_str}{chr(13)}{Midend_point_str}{chr(13)}{chr(13)}{width_str}{chr(13)}"

        acad.ActiveDocument.SendCommand(command)

    def convert_lines_to_walls_and_delete(self,x1,y1,x2,y2, width=240,tol=0.5):#将区域内一组直线转为天正墙
        """
        将一组直线段转为墙体，并删除原线段。

        直线段的处理前期多依赖点人工，后期可进一步强化细节处理

        参数：
        - lines: 包含 AcDbLine 类型的对象列表；
        - width: 墙体宽度，默认 240。
        """

        
        LK = select_objects_in_window_area(x1, y1, x2, y2)

        LX =[]
        
        for obj in LK:

            if obj.ObjectName =="AcDbLine" :

                LX.append(obj)

        success = 0
        failed = 0
        LQ =[]
        for i in range(0,len(LX)):

            line = LX[i]
            try:
                p1 = line.StartPoint
                print("端点1：",line.StartPoint) 
                p2 = line.EndPoint
                print("端点2：",line.EndPoint)
                self.draw_wall(p1, p2, width)

                Lq = select_objects_in_window_area(x1, y1, x2, y2)
               
                for obj in Lq:

                    if obj.ObjectName =="TDbWall" :

                        LQ.append(obj)

                        obj.Move(vtpnt(0,0,0),vtpnt(0,500000,0))
                
                line.Delete()
                success += 1
            except Exception as e:
                print(f"⚠️ 绘制失败或删除失败: {e}")
                failed += 1

        

        time.sleep(1)
        for xx in LQ:

            xx.Move(vtpnt(0,500000,0),vtpnt(0,0,0))

        print(f"✅ 转换完成：绘制墙体 {success} 根，失败 {failed} 根，原线段已删除。")

        

    def is_shear_wall(self,wall, tol=0.01):# 判断是否斜墙
        """
        判断天正墙是否为斜墙（非水平、非垂直墙），如果是斜墙则高亮显示该墙。
        
        参数:
            wall: 天正墙对象（包含 TotalWidth 和 GetBoundingBox 方法）
            tol: 长度误差控制容差，默认 0.01
            
        返回:
            bool: 如果为斜墙，返回 True；否则返回 False
        """
        try:
            # 获取墙的外包盒对角点
            min_pt, max_pt = wall.GetBoundingBox()

            # 计算外包盒的长和宽
            length = abs(max_pt[0] - min_pt[0])  # x方向的长度
            width = abs(max_pt[1] - min_pt[1])   # y方向的宽度

            # 获取墙的总宽度
            total_width = wall.TotalWidth

            # 判断长或宽是否与墙的总宽度相等
            if abs(length - total_width) > tol and abs(width - total_width) > tol:
                # 斜墙（长宽都与总宽度不相等）
                print("✅ 斜墙，已高亮显示")
                wall.Highlight(True)  # 高亮显示斜墙
                return True
            else:
                # 正交墙（水平或垂直墙）
                print("✅ 正交墙")
                return False
        except Exception as e:
            print(f"❌ 错误：{e}")
            return False


    def get_wall_centerline_points(self,wall,tol=0.5):

        """
        获取天正斜墙体的中心线坐标

        需要预先绘制500000*500000PL矩形配合视图范围对象可见性控制
        
        """

        def do_wall_shear(wall):##针对斜墙
            """
            获取天正斜墙体的中心线坐标（颜色为 46 的直线）并返回其包围区域。
            返回：p1, p2, (x1, y1, x2, y2) —— 中心线两端点 & bounding box

            注意及时删除函数返回(x1, y1, x2, y2)区域内的垃圾对象

            要确保天正墙轴线显示
            
            """
            try:

                # 1. 复制墙体并上移
                wall_copy = wall.Copy()
                wall.Visible = False
                move_vec = vtpnt(0, 500000, 0)
                wall_copy.Move(vtpnt(0, 0, 0), move_vec)

                # 2. 获取并扩展 bounding box（用于窗口选择乱线）
                p1_box, p2_box = wall_copy.GetBoundingBox()
                x1 = min(p1_box[0], p2_box[0]) - 300
                y1 = min(p1_box[1], p2_box[1]) - 300
                x2 = max(p1_box[0], p2_box[0]) + 300
                y2 = max(p1_box[1], p2_box[1]) + 300

                # 3. 缩放视图 + 高亮
                highlight_entity_by_bbox(wall_copy)
                time.sleep(2)

                # 4. 炸开
                doc.SendCommand("x\n")
                time.sleep(2)

                # 5. 选择该区域内颜色为 46 的直线
                objs = list(select_objects_in_window_area(x1, y1, x2, y2))
                center_line = None
                for obj in objs:
                    try:
                        if obj.ObjectName == "AcDbLine" and obj.Color == 46:
                            center_line = obj
                            break
                    except:
                        continue

                if not center_line:
                    raise Exception("未找到颜色为 46 的中心线")

                # 6. 提取中心线坐标，修正Y偏移
                sp = center_line.StartPoint
                ep = center_line.EndPoint
                p1 = (sp[0], sp[1] - 500000, sp[2])
                p2 = (ep[0], ep[1] - 500000, ep[2])

                wall.Visible = True

                return p1, p2, (x1, y1, x2, y2)

            except Exception as e:
                print(f"❌ get_wall_centerline_points 失败：{e}")
                return None, None, None

        def do_wall_Orthogonal(wall,tol): # # 获取墙体中轴线端点_仅针对正交墙
              
                """
                获取天正墙体的中心轴线的两个端点坐标（使用 TotalWidth 替代 Width）
                
                参数:
                    wall: 天正墙对象
                
                返回:
                    po1, po2: 中心线的两个端点坐标
                """
                try:
                    min_pt, max_pt = wall.GetBoundingBox()
                    x1, y1, z1 = min_pt
                    x2, y2, z2 = max_pt
                    z = z1

                    # 四角点
                    p1 = (x1, y1, z)
                    p2 = (x2, y1, z)
                    p3 = (x2, y2, z)
                    p4 = (x1, y2, z)

                    # 左右中线
                    mid_lr_1 = ((p1[0] + p4[0]) / 2, (p1[1] + p4[1]) / 2, z)
                    mid_lr_2 = ((p2[0] + p3[0]) / 2, (p2[1] + p3[1]) / 2, z)

                    # 上下中线
                    mid_tb_1 = ((p1[0] + p2[0]) / 2, (p1[1] + p2[1]) / 2, z)
                    mid_tb_2 = ((p3[0] + p4[0]) / 2, (p3[1] + p4[1]) / 2, z)

                    len_lr = math.dist(mid_lr_1, mid_lr_2)
                    len_tb = math.dist(mid_tb_1, mid_tb_2)

                    width = wall.TotalWidth
                    diff_lr = abs(len_lr - width)
                    diff_tb = abs(len_tb - width)

                    if diff_lr < tol:
                        return mid_tb_1, mid_tb_2
                    elif diff_tb < tol:
                        return mid_lr_1, mid_lr_2
                    else:
                        return (mid_lr_1, mid_lr_2) if len_lr > len_tb else (mid_tb_1, mid_tb_2)

                except Exception as e:
                    print("错误：", e)
                    return None, None

        if self.is_shear_wall(wall):

            p1, p2, bbox = do_wall_shear(wall)

            obj_del = select_objects_in_window_area(bbox[0], bbox[1], bbox[2], bbox[3])#必须隐性选择

            #清除垃圾辅助对象

            for xx in obj_del:

                xx.Delete()

                print("清除辅助对象")
                

            return p1,p2,bbox 

        else:
            
            p1, p2 = do_wall_Orthogonal(wall,tol)
            
            return p1,p2,None     


    def set_wall_width(self,wall,width):#改墙厚
        """
        使用 TWallThick 命令设置天正墙厚度。

        参数：
            
            width: 要设置的墙厚，单位通常为 mm（例如 200）
        """
        #要把隐性的选择对象转换为蓝色亮显对象才能在CAD命令行操作
        
        #获取BoundingBox

        P1,P2 = wall.GetBoundingBox()

        p1,p2 = expand_rectangle(P1, P2, offset=130)

        #窗口选择

        highlight_entities_in_window(p1[0],p1[1],p2[0],p2[1])
        
        try:
            cmd = f"TWallThick{chr(13)}{width}{chr(13)}"
            doc.SendCommand(cmd)
            print(f"已发送 TWallThick 命令，设置墙厚为 {width}")
        except Exception as e:
            print("发送命令失败:", e)






    #获取门窗基点

    def get_door_window_base_points(self,primiOb,wall_width=240):

        """
        以D:/Myprogramsystem/XT/MC_yuan.dwg文件的基元为准

        primiOb 基元对象

        返回点坐标

        """
        if primiOb.Layer in ["jz-danmen","jz-shuangmen","jz-zimumen"]:

            p1,p2=primiOb.GetBoundingBox()

            x = (p1[0]+p2[0])/2

            y =  p1[1]

            z = (p1[2]+p2[2])/2 

            return (x,y,z)

        elif primiOb.Layer in ["jz-tuilamen","jz-juanlianmen","jz-pingchuang","jz-baiyechuang","jz-gaochuang","jz-dong"]:

            p1,p2=primiOb.GetBoundingBox()

            x = (p1[0]+p2[0])/2

            y = (p1[1]+p2[1])/2

            z = (p1[2]+p2[2])/2 

            return (x,y,z)        

        elif primiOb.Layer in ["jz-tuchuang"]:

            p1,p2=primiOb.GetBoundingBox()

            x = (p1[0]+p2[0])/2

            y = p1[1]+wall_width/2

            z = (p1[2]+p2[2])/2 

            return (x,y,z)        


        elif primiOb.Layer in ["jz-huchuang"]:

            p1,p2=primiOb.GetBoundingBox()

            x = (p1[0]+p2[0])/2

            y = (p1[1]+p2[1])/2

            z = (p1[2]+p2[2])/2 

            ##240墙

            if wall_width == 240:

                x = x+2.01054791

                y = y+1.66388336

            elif wall_width == 200:

                x = x+2.52864298

                y = y+1.79479496

            else :

                x = x+2.52864298

                y = y+1.79479496

                print("非240圆弧墙非200圆弧墙，所得基点可能产生不允许误差")

            return (x,y,z)        




        elif primiOb.Layer in ["jz-menlianchuang"]:

            p1,p2=primiOb.GetBoundingBox()

            x = (p1[0]+p2[0])/2

            y = p1[1]+wall_width/2

            z = (p1[2]+p2[2])/2 

            return (x,y,z)

        else:

            return None






    def insert_mc_on_wall(self,template_entity, wall,entity_width=900, target_point=(0,0,0), tol=10):##在指定有墙的位置插入指定门窗
        """
        基元以D:/Myprogramsystem/XT/MC_yuan.dwg为准
        参数：
            template_entity: 原始门对象（作为复制来源）
            target_point: 门插入的中心位置 (x, y, z)
            tol: 与墙体对齐的误差容限
        """
        
        try:
            
            #获取基元基点

            template_entity.Label = ""
            
            base_point = self.get_door_window_base_points(template_entity,wall_width=240)

            new_entity = template_entity.Copy()

            #移动到目标点

            move_vec = (
                target_point[0] - base_point[0],
                target_point[1] - base_point[1],
                target_point[2] - base_point[2]
            )
            new_entity.Move(vtpnt(0, 0, 0), vtpnt(*move_vec))

            #获取墙方向
                      
            wp1, wp2,_ = self.get_wall_centerline_points(wall)

            #获取墙方向上target_point距离100的点

            ob_jing = points_on_line_at_distance_3d(wp1, wp2, target_point, 100)

            ob_jing = ob_jing[0]

            #激活天正智能门窗系统
            
            new_entity.Move(vtpnt(target_point[0],target_point[1],target_point[2]),vtpnt(ob_jing[0],ob_jing[1],ob_jing[2]))

            new_entity.Move(vtpnt(ob_jing[0],ob_jing[1],ob_jing[2]),vtpnt(target_point[0],target_point[1],target_point[2]))

            new_entity.Move(vtpnt(target_point[0],target_point[1],target_point[2]),vtpnt(ob_jing[0],ob_jing[1],ob_jing[2]))

            new_entity.Move(vtpnt(ob_jing[0],ob_jing[1],ob_jing[2]),vtpnt(target_point[0],target_point[1],target_point[2]))

            new_entity.Move(vtpnt(target_point[0],target_point[1],target_point[2]),vtpnt(ob_jing[0],ob_jing[1],ob_jing[2]))

            new_entity.Move(vtpnt(ob_jing[0],ob_jing[1],ob_jing[2]),vtpnt(target_point[0],target_point[1],target_point[2]))


            #设置图层和宽度

            new_entity.Layer = "WINDOW"

            new_entity.Width = entity_width
           
            print(f"✅ 已插入，容差: {tol}")
            return new_entity

        except Exception as e:
            print(f"❌ 插入失败：{e}")
            return None


    # 获取门的几何信息

    def get_door_geometry(self,door_entity):
        """
        获取指定门对象的几何信息，例如宽度、高度、位置、门框厚度等。
        
        参数:
            door_entity: COM对象，表示门实体

        返回:
            dict: 包含几何属性的字典
        """
        men1 = door_entity.Copy()
        door_entity.Visible=False
        
        highlight_entity_by_bbox( men1)
        cmd = "x\n"
        doc.SendCommand(cmd)

        ob=doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
        LS = get_entities_from_block_reference(ob)

        ob.Delete()

        door_entity.Visible=True

        return LS


    ##基于insert_INSERTMC_with_coord方法重新编制的插入门窗凸窗等函数 

    def insert_door(self,coord,width=900,btn_x=351, btn_y=129, cmd_x=471, cmd_y=455):#插单门

        """

        coord 三维坐标位置，基准数据依电脑改变

        我们不需要插双门的函数，只需要通过单门和匹配函数transfer_props_by_matchprop(ob1, ob2, delay=0.5)将ob1属性传给ob2
        以及将目标文件原位粘贴到当前文件的函数copy_doc_to_current("D:/Myprogramsystem/XT/MC_yuan.dwg")
        就可以解决双门等问题

        可以通过房间组，隐藏，图层，隔离等多种手段解决局部多个对象干扰GetBoundingBox的选择问题

        """

        insert_INSERTMC_with_coord(coord, btn_x=btn_x, btn_y=btn_y, cmd_x=cmd_x, cmd_y=cmd_y)

        time.sleep(1)

        entity=last_obj()

        entity.Width = width

        return entity




    #__________  天正墙门窗   _________________________  20250403  ________________________________________________________________
    #…………………………………………………………………………………………………………………………………………………………………







#&&% #____________________________________________________      天正房间      __________________________________________________
#…………………………………………………………………………………………………………………………………………………………………
# DWG类 - 天正房间

#__________________________________________________________________________________________________________________________
#…………………………………………………………………………………………………………………………………………………………………
#  使用说明

"""
研究天正房间的基本操作等问题 

如下的属性是非常有用的
'TDbSpace'
ShowArea
LB[0].Name
'房间'
LB[0].Name="客厅"
ShowName
Update
LB[0].ShowArea
'是'
LB[0].OccuArea
'23.86'
LB[0].WallSurfArea
'55.75'
LB[0].UseArea
'21.57'

#轮廓线的获取
obj.ObjectName
'AcDbPolyline'
obj.Length
97033.99119585901
obj2 = doc.ModelSpace.Item(doc.ModelSpace.Count - 2)
obj2.ObjectName
'TDbText'
obj2.Text
'341.87m^U2^U'
obj3 = doc.ModelSpace.Item(doc.ModelSpace.Count - 3)
obj3.Text
'建筑面积'

lt[0].ShowOutLine="是"


"""


#  主函数
#  (1)
# 获取天正房间、房间PL线、房间名称

#  该函数系列包括如下一些函数

# 1 建立天正房间


def run_auto_TUPDSPACE_with_coord(coord):
    """
    调用 auto_TUPDSPACE.py 子程序，传入坐标，并等待其执行完成（无黑框弹出）。

    参数:
        coord: 三元组坐标 (x, y, z)
 

    请注意，该命令的运行前提是先让天正墙所表达的房间，处于高亮蓝色选择状态。命令是有效的，这是基础





   """
    try:
        # 校验 coord 格式
        if not isinstance(coord, (tuple, list)) or len(coord) != 3:
            raise ValueError("坐标必须是一个包含三个数值的元组或列表")

        x1, y1, z1 = coord
        coord_str = f"{x1},{y1},{z1}"

        # 构造命令行
        cmd = [
            sys.executable,               # 保证调用当前 Python 解释器
            "auto_TUPDSPACE.py",
            "--coord", coord_str
        ]

        # 在 Windows 下隐藏子进程 cmd 窗口
        subprocess.run(
            cmd,
            check=True,
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        print(f"✅ 已完成子程序调用：{coord_str}")

    except Exception as e:
        print(f"❌ 调用失败：{e}")


def insert_INSERTMC_with_coord(coord, btn_x=374, btn_y=127, cmd_x=471, cmd_y=455):
    """
    调用 auto_click_opening.py 子程序，传入坐标，并等待其执行完成（无黑框弹出）。

    参数:
        coord: 三元组坐标 (x, y, z)
        btn_x, btn_y: 门窗对话框中“画窗”按钮的屏幕坐标
        cmd_x, cmd_y: CAD命令行输入的屏幕坐标

        单线变墙也可以类似处理，但目前不是重点问题
    """
    try:
        # 校验 coord 格式
        if not isinstance(coord, (tuple, list)) or len(coord) != 3:
            raise ValueError("坐标必须是一个包含三个数值的元组或列表")

        x1, y1, z1 = coord
        coord_str = f"{x1},{y1},{z1}"

        # 构造命令行，传入子程序
        cmd = [
            sys.executable,               # 保证调用当前 Python 解释器
            "auto_click_opening.py",       # 你子程序的路径
            "--coord", coord_str,         # 传入坐标
            "--btn_x", str(btn_x),        # 传入按钮位置
            "--btn_y", str(btn_y),        # 传入按钮位置
            "--cmd_x", str(cmd_x),        # 传入CAD命令行位置
            "--cmd_y", str(cmd_y)         # 传入CAD命令行位置
        ]

        # 在 Windows 下隐藏子进程 cmd 窗口
        subprocess.run(
            cmd,
            check=True,
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        print(f"✅ 已完成子程序调用：{coord_str}")

    except Exception as e:
        print(f"❌ 调用失败：{e}")



# 获取天正房间、房间中轴PL线及其文字内容

def get_centerline_polyline_and_texts_from_rooms(x1, y1, x2, y2):

    """
    
    
    批量提取多个 TDbSpace 房间对象的中轴线 Polyline 和名称文字：
      - 隐藏 WALL 图层，隐藏房间轮廓
      - 逐个让房间可见、复制、炸开、筛选出 Polyline 和文字
      - 恢复所有可见性和属性

    :x1,y1,x2,y2: 天正房间对象所在区域，不能稳定控制天正对象本身，采用区域
    :return: [(room, LB_coor, text_cn, text_other), ...]
        room: 房间对象
        LB_coor: 房间中轴线坐标
        text_cn: 中文房间名
        text_other: 其他房间名

    """

    # 1️⃣ 获取房间对象列表
    LSP = select_objects_in_window_area(x1, y1, x2, y2)
    room_list = [obj for obj in LSP if obj.ObjectName == "TDbSpace"and obj.Name != "建筑面积" ]
    print("房间数量：", len(room_list))

    room_data = []  # 存储每个房间的结果

    # 2️⃣ 遍历每个房间，提取信息
    for i, room in enumerate(room_list):

        print("序号：",i)



        # 随机编号房间名称
        name_room = room.Name
        room.Name = name_room + str(i)

        # 复制房间实体，然后移动到外部区域
        copy_ent = room.Copy()

        copy_ent.Move(vtpnt(0,0,0),vtpnt(0,500000+i*50000,0))

        P1,P2=copy_ent.GetBoundingBox()
        

        X1,Y1=P1[0],P1[1]

        X2,Y2=P2[0],P2[1]

        h=0.1*(abs(X1-X2)+abs(Y1-Y2))/2

        X1,Y1=P1[0]-h,P1[1]-h

        X2,Y2=P2[0]+h,P2[1]+h


        # 高亮并炸开
        highlight_entity_by_bbox(copy_ent)
        time.sleep(0.3)
        doc.SendCommand("x\n")
        time.sleep(2)

        # 从 ModelSpace 当前区域取出最新对象

        pline = None
        text_cn = None
        text_other = None
        time.sleep(2)

        # 抓取当前区域对象
        new_objs = select_objects_in_window_area(X1, Y1, X2, Y2)

        # 3️⃣ 处理炸开后的对象
        num_poly=0
        for obj in new_objs:

            obj.Layer = "测试辅助"
            try:
                if obj.ObjectName == "AcDbPolyline":
                    print(f"确认选择到每个房间的多段线：{obj.ObjectName}")
                    print(obj.Coordinates)
                    num_poly=num_poly+1

                    # 收集多段线坐标，从坐标生成新PL线，斩断与天正房间对象的关联
                    pline = obj

                    pline.Move(vtpnt(0,500000+i*50000,0),vtpnt(0,0,0))

                    obj.Layer = "房间中轴线"
                  

                elif hasattr(obj, "Text"):
                    txt = obj.Text
                    # 含中文优先
                    if re.search(r'[\u4e00-\u9fff]', txt):
                        text_cn = txt
                    else:
                        text_other = txt
                    
                else:
                    pass
            except Exception as e:
                print(f"❌ 错误处理对象 {obj}: {e}")

        print(f"确认选择到该房间的多段线根数：{num_poly}") 
        # 4️⃣ 添加房间数据到结果列表
        room_data.append((room, pline, text_cn, text_other or ""))

           
    # 5️⃣ 恢复所有房间的可见性和轮廓状态

    time.sleep(2)

    ql()

    print(f"✅ 成功提取 {len(room_data)} 个房间的中轴线和名称")

    return room_data

            

# 获取操作区域

def get_operation_area(tol=0.5):
    
    LP = select_tuceng("方案深化")
    

    p1 = p2 = LB_coor1 = LB_coor2 = None
    for ob in LP:
        if ob.ConstantWidth <tol:
            p1 = find_min_point(ob)#左下角
            p2 = find_max_point(ob)#右上角
            LB_coor1 = plcom_to_coor(ob)
            ob.Delete()
        else:
            LB_coor2 = plcom_to_coor(ob)  # 存储多段线的坐标，允许输入多段线或多段线列表
            ob.Delete()

    # 检查是否成功获取 p1/p2
    if p1 is None or p2 is None:
        raise ValueError("❌ 未在“方案深化”图层中找到 ConstantWidth==0 的对象")

    x1, y1 = p1
    x2, y2 = p2
    return (x1, y1, x2, y2), LB_coor1,LB_coor2
            
#  ——————————————
#  核心函数
#  (1)
#  获取天正房间、房间PL线、房间名称
#  ____________________________
#  ——————————————

def get_tianzhengroom_roompolyline_name(mianji_shuzi_weizhi=3000):
    """
    获取天正房间对象、房间中轴线多段线以及房间名称文本。
    1. 从“方案深化”图层获取操作区域和原始外框坐标 LB_coor
    2. 计算 TUPDSPACE 插入点 coord
    3. 高亮选择区域内对象，运行 TUPDSPACE
    4. 筛选 TDbSpace 对象（天正房间）
    5. 提取每个房间的中轴线 Polyline 和文字
    6. 清除区域内垃圾直线
    7. 恢复原始外框（用 plcoor_to_com 重绘）
    返回 (tianzheng_fangjian_list, room_pl_list, room_name_list)


    列表room_data[i] 每个房间的天正房间、房间中轴线、房间名称、房间面积
    """
    # 1️⃣ 获取操作区域
    p, LB_coor1,LB_coor2 = get_operation_area(tol=0.5)

    x1,y1=p[0], p[1]

    x2,y2=p[2], p[3]

    # 左下、右上点（三元组）

    # 计算插入坐标
    coord_x = (p[0] + p[2]) / 2
    coord_y = p[1] - mianji_shuzi_weizhi
    coord   = (coord_x, coord_y, 0)


    # 2️⃣ 缩放视图窗口


    h=0.1*(abs(x1-x2)+abs(y1-y2))/2

    zoom_cmd = (
        f"_.ZOOM\n"      # 调用 Zoom
        f"_W\n"          # 窗口选项
        f"{x1-h},{y1-h}\n"   # 第一个角点
        f"{x2+h},{y2+h}\n"   # 第二个角点
    )
    doc.SendCommand(zoom_cmd)
    time.sleep(1)  # 等待缩放完成



    # 2 线变墙


    dw=DWG()

    dw.convert_lines_to_walls_and_delete(x1,y1,x2,y2, width=240,tol=0.5)

    time.sleep(2)

    # 3️⃣ 高亮选择区域内所有对象
    highlight_entities_in_window(x1, y1, x2, y2)
    time.sleep(3)

    # 4️⃣ 调用 TUPDSPACE 子程序生成天正房间
    run_auto_TUPDSPACE_with_coord(coord)#调用时应隐藏cmd

    time.sleep(3)

    # 5️⃣ 获取房间数据
    
    room_data = get_centerline_polyline_and_texts_from_rooms(x1, y1, x2, y2)

    # 6️⃣ 恢复原始外框：用 LB_coor 重绘外框多段线

    if LB_coor2 == None:

        pass

    else:
    
        plcoor_to_com(LB_coor2, layer_name="方案深化", width=100, color=6)


    if LB_coor1 == None:

        pass

    else:


        plcoor_to_com(LB_coor1, layer_name="方案深化", width=0, color=6)

    return room_data


##锁住图层以保护对象不被删除
def lock_layer_objects(layername: str) -> bool:
    """
    锁定指定图层，使该图层上的所有对象不可编辑。

    :param layername: 要锁定的图层名称
    :return: 锁定成功返回 True，找不到图层返回 False
   """

    layers = doc.Layers

    # 尝试获取图层
    try:
        lyr = layers.Item(layername)
    except Exception:
        print(f"❌ 未找到图层 “{layername}”")
        return False

    # 确保图层开启，否则锁定无效
    lyr.LayerOn = True

    # 锁定图层
    lyr.Lock = True
    print(f"✅ 图层 “{layername}” 已锁定，图层上的对象不可编辑")
    return True


##解锁图层
def unlock_layer(layername: str) -> bool:
    """
    解锁指定图层，使该图层上的对象可编辑。

    :param layername: 要解锁的图层名称
    :return: 解锁成功返回 True，找不到图层返回 False
    """
    # 1) 连接到正在运行的 AutoCAD
    layers = doc.Layers

    # 2) 尝试获取图层
    try:
        lyr = layers.Item(layername)
    except Exception:
        print(f"❌ 未找到图层 “{layername}”")
        return False

    # 3) 如果图层处于关闭状态，可以先打开
    lyr.LayerOn = True

    # 4) 解锁图层
    lyr.Lock = False
    print(f"✅ 图层 “{layername}” 已解锁，图层上的对象可编辑")
    return True






#&&% #____________________________________________________       天正标注等辅助处理       ______________________________________
#…………………………………………………………………………………………………………………………………………………………………
# DWG类 - 天正标注等辅助处理

#__________________________________________________________________________________________________________________________
#…………………………………………………………………………………………………………………………………………………………………
#  使用说明

"""
研究天正标注等基本操作等问题

由房间边界或者轴线边界提供标注两侧端点的依据，我们的基础标注命令从一个点到另一个点，尺寸线点，扩展点

自动化分析则是基于对房间，轴网，门窗的分析产生的控制点集

"""


#  主函数
#  (1)
# 逐点标注

#  该函数系列包括如下一些函数


def dim_by_points(*args):
    """
    使用 TDimMP 命令进行标注。
    
    参数格式支持：
      - dim_by_points(p1, p2, p3)
      - dim_by_points(p1, p2, p3, q1, q2, ...)
      - dim_by_points([p1, p2, p3, q1, q2, ...])

    说明：
      p1: 第一个标注起点
      p2: 第一个标注终点
      p3: 尺寸线的定位点
      q1, q2, ...：扩展标注点，可选（也可以没有），位置自由
    """
    try:
        # 解包嵌套列表输入
        if len(args) == 1 and isinstance(args[0], list):
            pts = args[0]
        else:
            pts = list(args)

        if len(pts) < 3:
            print("❌ 至少需要三个点：起点、终点、尺寸线定位点")
            return

        p1, p2, p3, *extension_pts = pts

        def pt_str(pt):
            return f"{pt[0]},{pt[1]},{pt[2]}"

        command = (
            "TDimMP" + chr(13) +
            pt_str(p1) + chr(13) +
            pt_str(p2) + chr(13) +
            pt_str(p3) + chr(13)
        )

        # 扩展点部分
        for pt in extension_pts:
            if not isinstance(pt, (list, tuple)) or len(pt) != 3:
                print(f"⚠️ 非法扩展点：{pt}，已跳过")
                continue
            command += pt_str(pt) + chr(13)

        command += chr(13)  # 结束命令

        acad.ActiveDocument.SendCommand(command)
        print(f"✅ 已标注主段：{pt_str(p1)} → {pt_str(p2)}，尺寸线位置：{pt_str(p3)}")
        if extension_pts:
            print(f"➕ 扩展标注点数量：{len(extension_pts)}")

    except Exception as e:
        print(f"❌ 执行标注失败：{e}")



#_____________________________________________________________________________________________________________________________________________


#&&&&%% 第十二部分  文件调度体系


#&&% ⮞ F5: 文件调度


"""

使用百度同步，可能还会产生其它一些问题，这个要等将来在调整分析。不追求3台服务电脑的完全一致。每台电脑会有不同的一些处理。

dispatch_inputs()函数

修改dispatch_inputs()，原来的函数是直接在百度端的输入文件夹选择时间最早的文件夹或文件，传入中央端的对应输入文件夹。并且是读取D:/Myprogramsystem/BaiduSyncdisk/基础服务/processed_inputs.txt中是文件夹或文件名，排除这些读取到的文件夹名或文件名，选择其时间最早的。现在要这样修改：必须从D:/Myprogramsystem/BaiduSyncdisk/基础服务/processed_inputs.txt读取文件夹或文件名，这个文件夹名或文件名，是由其它函数或脚本写入的，它带有文件路径。我们恰恰要根据processed_inputs.txt中存在的带路径项目名去百度端的输入文件夹核实相应的文件，然后将它传入中央端的对应文件夹。
processed_inputs.txt遵循这样的格式和规范：一个带路径的完整文件夹名文件名，占一行。紧接下一行无空格，是0或1，1表示已经开始执行。再紧接下一行是0或1，1表示已经完成任务。
下一个文件夹或文件的信息和上一个之间用空格断开，每个文件夹或文件的信息占三行。保留原来的读取文件延迟时间120秒的机制。新机制是依据processed_inputs.txt的信息来传输百度端的输入文件夹的文件夹或文件到中央端。

将processed_inputs.txt文件作为其它函数或脚本完成的任务。它基于一个合理的分配，在一个循环中，对所有输入的文件夹或文件进行合理的任务分配。分配之后写入各台电脑的processed_inputs.txt，因为要通过百度同步传递文件
所以每台电脑上的该文件名实际上要加上A，B，C不同的标识。



系统架构图：
![Architecture](D:/Myprogramsystem/CentralControlProgram/流程图/文件调度系统整体流程图.png)

编辑一下这个功能还是很有意思的


- 核心处理函数:ttt

- 后台进程: [log](run_daemon)




分配脚本

分配脚本必须在A电脑-服务器上运行，进行分配。分配的结果写入不同的processed_inputs.txt文件，电脑不同，命名应该不同。


分配要考虑服务器电脑是最后的保障，另外两台是加速。如果网络中断，仍然要完成任务。

30分钟一次循环，找到输入的文件夹或文件进行分配。5分钟内还没正式执行，取消任务，也不再读取回传结果。30分钟内还没有完成，重新分配。

第一个循环找到的全部文件夹或文件处理完毕后，才会进行下一轮分配。



模拟CAD处理脚本 


##模拟用户输入.py 

模拟CAD处理脚本.py


发现今天的问题是开机启动时CAD出错，连别的无限循环程序也不运行了。必须知道列表中的py脚本是否在运行


D:/Myprogramsystem/CentralControlProgram
下的allocation_script.py,cad_dialog_killer.py,central_dispatch.py,sync_from_baidu.py


"""

# 1 查询服务数据流


def analyze_log_for_user(log_input, username, output_file=None):
    """
    Analyzes the log file to track the trajectory of the most recent input item for a user.
    
    Args:
        log_input (str): Path to the log file or the log text itself.
        username (str): The username to analyze (e.g., "曾德权").
        output_file (str, optional): Path to save the output with proper newlines. If None, print to console.
    
    Returns:
        str: A formatted string with the trajectory of the most recent input item, or an error message.


    analyze_log_for_user(r"D:/Myprogramsystem/XT/Basic_service_processing/central_loop.log", "曾德权", output_file=r"D:/Myprogramsystem/XT/Basic_service_processing/服务数据流查询.txt")



    """
    # Handle both file path and direct log text
    if os.path.isfile(log_input):
        try:
            with open(log_input, 'r', encoding='utf-8') as f:
                log_text = f.read()
        except Exception as e:
            return f"Error reading log file {log_input}: {e}"
    else:
        log_text = log_input  # Assume log_input is the log text

    lines = log_text.split('\n')
    input_item = None
    input_from = None
    backup_source = None
    backup_baidu = None
    central_copy = None
    deleted_baidu = None
    output_time = None
    return_from = None
    return_to_baidu = None
    return_to_source = None
    source = 'Central processing (external)'

    # Search for the latest input sync for the user
    for i in range(len(lines) - 1, -1, -1):
        line = lines[i]
        if input_item is None and '【SYNC】复制输入项：' in line and username in line:
            parts = line.split(' → ')
            if len(parts) == 2:
                from_path = parts[0].split('：')[1].strip()
                to_path = parts[1].strip()
                item_name = from_path.split('\\')[-1]
                input_item = item_name
                input_from = from_path
                # Find source backup
                for j in range(i + 1, len(lines)):
                    if '【BACKUP】移动源输入项：' in lines[j] and item_name in lines[j]:
                        backup_parts = lines[j].split(' → ')
                        backup_source = backup_parts[1].strip()
                        break
                # Find dispatch and baidu backup, central copy, deleted
                for j in range(i + 1, len(lines)):
                    if '[PICK]' in lines[j] and username in lines[j] and item_name in lines[j]:
                        for k in range(j + 1, len(lines)):
                            if '备份：' in lines[k] and item_name in lines[k]:
                                backup_baidu_parts = lines[k].split(' → ')
                                backup_baidu = backup_baidu_parts[1].strip()
                            if '复制到中央：' in lines[k] and item_name in lines[k]:
                                central_parts = lines[k].split(' → ')
                                central_copy = central_parts[1].strip()
                            if '【DISPATCH】已删除百度端输入项：' in lines[k] and item_name in lines[k]:
                                deleted_baidu = lines[k].split('：')[1].strip()
                                break
                        break
                # Find return and sync
                for j in range(i + 1, len(lines)):
                    if '回传: ' in lines[j] and item_name in lines[j]:
                        return_parts = lines[j].split(' -> ')
                        return_from = return_parts[0].split('回传: ')[1].strip() + ' (from Central output)'
                        return_to_baidu = return_parts[1].strip()
                        # Get time from cycle start
                        for m in range(j, -1, -1):
                            if '--- 循环 #' in lines[m]:
                                output_time = lines[m].split(' ---')[0].strip()
                                break
                        # Find sync to source
                        for k in range(j + 1, len(lines)):
                            if '【RETURN】同步输出项: ' in lines[k] and item_name in lines[k]:
                                sync_parts = lines[k].split(' → ')
                                return_to_source = sync_parts[1].strip()
                                break
                        break
                break

    if input_item is None:
        return f"No recent input found for user {username}. Check if username is correct or if log contains relevant [SYNC] entries."

    # Build the result with proper newlines
    result = (
        f"用户 {username} 最近一次输入文件或文件夹: {input_item}\n"
        f"从哪里传入: {input_from}\n"
        f"到哪里备份 (Microcloud): {backup_source or 'Not found'}\n"
        f"到哪里备份 (Baidu Sync): {backup_baidu or 'Not found'}\n"
        f"在哪里停止被删除: {deleted_baidu or 'Not found'}\n"
        f"对应的输出文件夹传回时间: {output_time or 'Not found'}\n"
        f"从哪里传来: {return_from or 'Not found'}\n"
        f"源头在哪里: {source}\n"
        f"传回 Baidu 输出: {return_to_baidu or 'Not found'}\n"
        f"传回 Microcloud 输出: {return_to_source or 'Not found'}\n"
    )

    # Output to file if specified
    if output_file:
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(result)
            print(f"Output written to {output_file}")
        except Exception as e:
            print(f"Error writing to output file {output_file}: {e}")
    
    return result


# 2 监测系统架构文件夹内的一级文件夹和文件名

"""
run_daemon("D:/Myprogramsystem/CentralControlProgram/directory_monitor.py")直接随开机运行无限循环
D:/Myprogramsystem/XT/Basic_service_processing/directory_monitor.log


"""


# 3 清理用户输入输出（07插图签辅助输入）
def clear_specific_folders():
    """
    清空每个用户下指定的输入文件夹、07插图签辅助文件夹和输出文件夹内部的所有文件和子文件夹。
    注意：此函数会永久删除文件，请在运行前备份重要数据！
    """

    # 定义用户列表
    USERS = ['吴后建', '宋一水', '宋宇杰', '宋岳', '宋杰', '彭程', '曾德权', '杨淞杰', '蓝竟铭', '谭长孟']
    
    # 定义基础路径
    BASE_PATH = r"D:/MyPro/基础服务"
    SOURCE_FILE = r"D:/MyPro/基础服务/公用文件/测试文件/属性图签插图签样板文件1.dwg"
    
    # 定义相对路径
    REL_INPUTS = [
        r"01插图签/01只插图签输入",
        r"01插图签/02插图签编目录输入",
        r"01插图签/03插图签编目录打印输入",
        r"01插图签/04修改标签值输入",
        r"02编目录/01只编目录输入",
        r"02编目录/02编目录打印输入",
        r"03打印/01施工图模式打印/01输入",
        r"03打印/02A3文本模式打印/01输入",
    ]
    REL_OUTPUTS = [
        r"01插图签/05输出",
        r"02编目录/03输出",
        r"03打印/01施工图模式打印/02输出",
        r"03打印/02A3文本模式打印/02输出",
    ]
    OTHER_REL_DIRS = [
        r"04Excel控制文件/01不控制图纸名称",
        r"04Excel控制文件/02控制图纸名称",
        r"04Excel控制文件/03修改标签值",
        r"04Excel控制文件/04实际反馈",
        r"05Fonts文件夹",
        r"06打印样式",
        r"07备份",
        r"01插图签/06基本图签",
        r"01插图签/07插图签辅助输入"
    ]
    
    # 定义要清空的相对路径列表
    TARGET_REL_DIRS = REL_INPUTS + [r"01插图签/07插图签辅助输入"] + REL_OUTPUTS
    
    for user in USERS:
        user_path = os.path.join(BASE_PATH, user)
        if not os.path.exists(user_path):
            print(f"用户文件夹 {user_path} 不存在，跳过。")
            continue
        
        for rel_dir in TARGET_REL_DIRS:
            target_path = os.path.join(user_path, rel_dir)
            if not os.path.exists(target_path):
                print(f"文件夹 {target_path} 不存在，跳过。")
                continue
            
            # 遍历文件夹内部的所有内容
            for item in os.listdir(target_path):
                item_path = os.path.join(target_path, item)
                if os.path.isfile(item_path):
                    # 删除文件
                    os.remove(item_path)
                    print(f"删除文件: {item_path}")
                elif os.path.isdir(item_path):
                    # 删除子文件夹及其内容
                    shutil.rmtree(item_path)
                    print(f"删除文件夹: {item_path}")
    
    print("清空操作完成。")



#清空体系中全部文件夹文件

def clear_directories(filepath):
    """
    递归遍历指定根路径 filepath，找到所有相对路径以 REL_GROUPS 和 OTHER_REL_DIRS 中列出的路径结尾的子目录，
    并清空这些子目录内的所有文件和子文件夹。
    
    参数：
        filepath (str): 根目录路径
    """
    REL_GROUPS = [
        {
            "inputs": [
                r"01插图签/01只插图签输入",
                r"01插图签/02插图签编目录输入",
                r"01插图签/03插图签编目录打印输入",
                r"01插图签/04修改标签值输入",
                r"01插图签/07插图签辅助输入",  # 添加并改名
            ],
            "output": r"01插图签/05输出"
        },
        {
            "inputs": [
                r"02编目录/01只编目录输入",
                r"02编目录/02编目录打印输入",
            ],
            "output": r"02编目录/03输出"
        },
        {
            "inputs": [
                r"03打印/01施工图模式打印/01输入",
            ],
            "output": r"03打印/01施工图模式打印/02输出"
        },
        {
            "inputs": [
                r"03打印/02A3文本模式打印/01输入",
            ],
            "output": r"03打印/02A3文本模式打印/02输出"
        },
    ]
    
    OTHER_REL_DIRS = [
        r"04Excel控制文件/01不控制图纸名称",
        r"04Excel控制文件/02控制图纸名称",
        r"04Excel控制文件/03修改标签值",
        r"04Excel控制文件/04实际反馈",
        r"05Fonts文件夹",
        r"06打印样式",
        r"07备份",
        r"01插图签/06基本图签",
        # r"01插图签/07插图签辅助" # 移除，原目录改为07插图签辅助输入
    ]
    
    # 收集所有相关的相对目录路径（作为目标匹配的路径模式）
    target_dirs = []
    
    # 从 REL_GROUPS 中提取输入和输出路径
    for group in REL_GROUPS:
        target_dirs.append(group["output"])
        target_dirs.extend(group["inputs"])
    
    # 添加 OTHER_REL_DIRS 中的路径
    target_dirs.extend(OTHER_REL_DIRS)
    
    # 将 filepath 转换为 Path 对象
    root_path = Path(filepath)
    
    # 递归遍历根目录下的所有子目录
    for dir_root, dir_names, file_names in os.walk(root_path):
        current_path = Path(dir_root)
        
        # 对于每个遍历到的目录，检查是否匹配目标相对路径
        for target in target_dirs:
            # 构建目标路径的字符串（统一使用 / 分隔符）
            target_str = target.replace('\\', '/')
            
            try:
                relative = current_path.relative_to(root_path)
                relative_str = str(relative).replace('\\', '/')
                
                # 检查相对路径是否以目标路径结尾
                if relative_str.endswith(target_str):
                    print(f"找到匹配目录：{current_path}")
                    try:
                        # 遍历匹配目录中的所有项目
                        for item in current_path.iterdir():
                            try:
                                if item.is_file():
                                    # 删除文件
                                    item.unlink()
                                    print(f"已删除文件：{item}")
                                elif item.is_dir():
                                    # 删除子文件夹及其内容
                                    shutil.rmtree(item)
                                    print(f"已删除文件夹：{item}")
                            except Exception as e:
                                print(f"删除 {item} 时出错：{e}")
                    except Exception as e:
                        print(f"访问目录 {current_path} 时出错：{e}")
            except ValueError:
                # 如果 current_path 不是 root_path 的子路径，跳过（理论上不会发生）
                pass

# 4 检查当前输入输出口的操作对象(文件夹或文件)

def scan_directories() -> Dict:
    """
    遍历三端目录，扫描 REL_INPUTS、REL_OUTPUTS 和 OTHER_REL_DIRS 中的文件夹，
    返回子文件夹名和文件列表，并将结果保存到 directory_scan_YYYYMMDD_HHMMSS.txt。

    Returns:
        Dict: 包含扫描时间和三端目录内容的字典，结构如下：
        {
            'scan_time': 'YYYY-MM-DD HH:MM:SS JST',
            'Microcloud': { project_name: { rel_path: [item1, item2, ...], ... }, ... },
            'BaiduSync': { project_name: { rel_path: [item1, item2, ...], ... }, ... },
            'Central': { rel_path: [item1, item2, ...], ... }
        }
    """
    # 目录根定义
    SOURCE_ROOT = Path(r"D:/Mypro/基础服务")
    PROJECT_ROOT = Path(r"D:/Myprogramsystem/BaiduSyncdisk/基础服务")
    CENTRAL_ROOT = Path(r"D:/Myprogramsystem/XT/Basic_service_processing")

    # 配置
    EXCLUDE_PROJECT_NAMES: Set[str] = {"公用文件"}
    OCCUPANCY_IGNORE_NAMES: Set[str] = {"desktop.ini"}

    # 相对路径定义
    REL_INPUTS = [
        r"01插图签/01只插图签输入",
        r"01插图签/02插图签编目录输入",
        r"01插图签/03插图签编目录打印输入",
        r"01插图签/04修改标签值输入",
        r"02编目录/01只编目录输入",
        r"02编目录/02编目录打印输入",
        r"03打印/01施工图模式打印/01输入",
        r"03打印/02A3文本模式打印/01输入",
    ]
    REL_OUTPUTS = [
        r"01插图签/05输出",
        r"02编目录/03输出",
        r"03打印/01施工图模式打印/02输出",
        r"03打印/02A3文本模式打印/02输出",
    ]
    OTHER_REL_DIRS = [
        r"04Excel控制文件/01不控制图纸名称",
        r"04Excel控制文件/02控制图纸名称",
        r"04Excel控制文件/03修改标签值",
        r"04Excel控制文件/04实际反馈",
        r"05Fonts文件夹",
        r"06打印样式",
        r"07备份",
        r"01插图签/06基本图签",
        r"01插图签/07插图签辅助"
    ]

    # 获取当前时间（JST 时区）
    current_time = datetime.datetime.now()
    time_str = current_time.strftime("%Y-%m-%d %H:%M:%S JST")
    timestamp = current_time.strftime("%Y%m%d_%H%M%S")
    OUTPUT_FILE = CENTRAL_ROOT / f"directory_scan_{timestamp}.txt"

    # 加载项目名称
    project_names = [
        d.name for d in PROJECT_ROOT.iterdir()
        if d.is_dir() and d.name not in EXCLUDE_PROJECT_NAMES
    ]
    project_names.sort()

    # 初始化结果字典
    result = {
        'scan_time': time_str,
        'Microcloud': {},
        'BaiduSync': {},
        'Central': {}
    }

    # 扫描 Microcloud 端
    all_rel_dirs = REL_INPUTS + REL_OUTPUTS + OTHER_REL_DIRS
    for project in project_names:
        result['Microcloud'][project] = {}
        for rel in all_rel_dirs:
            dir_path = SOURCE_ROOT / project / rel
            if dir_path.exists():
                contents = [
                    str(p.relative_to(dir_path))
                    for p in dir_path.rglob('*')
                    if p.name not in OCCUPANCY_IGNORE_NAMES
                ]
                if contents:
                    result['Microcloud'][project][rel] = sorted(contents)

    # 扫描 Baidu Sync 端
    for project in project_names:
        result['BaiduSync'][project] = {}
        for rel in all_rel_dirs:
            dir_path = PROJECT_ROOT / project / rel
            if dir_path.exists():
                contents = [
                    str(p.relative_to(dir_path))
                    for p in dir_path.rglob('*')
                    if p.name not in OCCUPANCY_IGNORE_NAMES
                ]
                if contents:
                    result['BaiduSync'][project][rel] = sorted(contents)

    # 扫描 Central 端（无项目层级）
    for rel in all_rel_dirs:
        dir_path = CENTRAL_ROOT / rel
        if dir_path.exists():
            contents = [
                str(p.relative_to(dir_path))
                for p in dir_path.rglob('*')
                if p.name not in OCCUPANCY_IGNORE_NAMES
            ]
            if contents:
                result['Central'][rel] = sorted(contents)

    # 生成文本输出
    text_content = [
        f"目录扫描报告",
        f"扫描时间: {time_str}",
        ""
    ]

    # Microcloud 端
    text_content.append("Microcloud 端 (D:/Mypro/基础服务)")
    text_content.append("=" * 50)
    for project in project_names:
        if project in result['Microcloud'] and result['Microcloud'][project]:
            text_content.append(f"项目: {project}")
            for rel, items in result['Microcloud'][project].items():
                text_content.append(f"  {rel}:")
                for item in items:
                    text_content.append(f"    - {item}")
                text_content.append("")
    
    # Baidu Sync 端
    text_content.append("Baidu Sync 端 (D:/Myprogramsystem/BaiduSyncdisk/基础服务)")
    text_content.append("=" * 50)
    for project in project_names:
        if project in result['BaiduSync'] and result['BaiduSync'][project]:
            text_content.append(f"项目: {project}")
            for rel, items in result['BaiduSync'][project].items():
                text_content.append(f"  {rel}:")
                for item in items:
                    text_content.append(f"    - {item}")
                text_content.append("")
    
    # Central 端
    text_content.append("Central 端 (D:/Myprogramsystem/XT/Basic_service_processing)")
    text_content.append("=" * 50)
    for rel, items in result['Central'].items():
        text_content.append(f"{rel}:")
        for item in items:
            text_content.append(f"  - {item}")
        text_content.append("")

    # 保存到文本文件
    try:
        OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
            f.write("\n".join(text_content))
        print(f"目录扫描报告已保存至：{OUTPUT_FILE}")
    except Exception as e:
        print(f"[ERROR] 无法保存文本文件：{e}")

    return result


#&&% 5 检查指定文件夹下的无限循环脚本是否在运行

def check_scripts_running(scripts: List[str] = None) -> Dict[str, bool]:
    """
    检查指定目录下的 Python 脚本是否在运行，并将结果保存到带时间戳的文本文件。






    参数:
        scripts: List[str] - 要检查的脚本文件名列表。如果为 None，则使用默认列表。

    返回:
        Dict[str, bool] - 每个脚本的运行状态，键为脚本名，值为 True (运行中) 或 False (未运行)。
    """
    if scripts is None:
        scripts = [
            "allocation_script.py",
            "cad_dialog_killer.py",
            "central_dispatch.py",
            "sync_from_baidu.py",
            "external_processor.py",
        ]

    running_status = {script: False for script in scripts}
    target_dir = Path("D:/Myprogramsystem/CentralControlProgram")
    target_dir_str = str(target_dir.resolve()).replace('\\', '/').lower()

    for proc in psutil.process_iter(['name', 'cmdline']):
        try:
            if proc.info['name'].lower() in ('python', 'python.exe', 'python3', 'python3.exe', 'pythonw', 'pythonw.exe', 'py', 'py.exe'):
                cmdline = proc.info['cmdline']
                if cmdline:
                    cmd_str = ' '.join(cmdline)
                    cmd_str_normalized = cmd_str.replace('\\', '/').lower()
                    print(f"[DEBUG] Process: {proc.info['name']}, Cmdline: {cmd_str_normalized}")
                    for script in scripts:
                        script_normalized = unicodedata.normalize('NFC', script).lower()
                        script_path = str((target_dir / script).resolve()).replace('\\', '/').lower()
                        if script_normalized in cmd_str_normalized or script_path in cmd_str_normalized:
                            running_status[script] = True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass

    current_time = datetime.datetime.now()
    time_str = current_time.strftime("%Y-%m-%d %H:%M:%S JST")
    timestamp = current_time.strftime("%Y%m%d_%H%M%S")
    output_file = target_dir / f"script_status_{timestamp}.txt"

    text_content = [
        "脚本运行状态报告",
        f"检查时间: {time_str}",
        ""
    ]
    for script, is_running in running_status.items():
        status = "运行中" if is_running else "未运行"
        text_content.append(f"{script}: {status}")

    try:
        output_file.parent.mkdir(parents=True, exist_ok=True)
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("\n".join(text_content))
        print(f"脚本运行状态报告已保存至：{output_file}")
    except Exception as e:
        print(f"[ERROR] 无法保存文本文件：{e}")

    return running_status

# 6 修改系统文件夹名
def rename_auxiliary_folders(root_dir: str | Path) -> int:
    """
    遍历指定文件夹，递归查找所有01插图签/07插图签辅助子文件夹，重命名为01插图签/07插图签辅助输入。
    
    Args:
        root_dir (str | Path): 要遍历的根目录路径。
    
    Returns:
        int: 重命名的文件夹数量。


    rename_auxiliary_folders("D:/Myprogramsystem/XT/Basic_service_processing")

    """
    LD_FOLDER_NAME = "07插图签辅助"
    OLD_FOLDER_NAME = "07插图签辅助"
    NEW_FOLDER_NAME = "07插图签辅助输入"
    REL_PATH_PATTERN = f"01插图签/{OLD_FOLDER_NAME}"
    

    root_path = Path(root_dir)
    log(f"开始在 {root_path} 下递归重命名 01插图签/{OLD_FOLDER_NAME} 文件夹为 01插图签/{NEW_FOLDER_NAME}")
    renamed_count = 0
    
    try:
        # 递归查找所有匹配01插图签/07插图签辅助的文件夹
        for old_folder in root_path.rglob(REL_PATH_PATTERN):
            if not old_folder.is_dir():
                continue
            # 计算新文件夹路径
            new_folder = old_folder.parent / NEW_FOLDER_NAME
            try:
                # 确保目标文件夹不存在
                if new_folder.exists():
                    log(f"目标文件夹已存在，跳过：{new_folder}")
                    continue
                # 重命名
                old_folder.rename(new_folder)
                log(f"成功重命名：{old_folder} → {new_folder}")
                renamed_count += 1
            except Exception as e:
                log(f"[ERROR] 重命名失败：{old_folder} → {new_folder}，{traceback.format_exc()}")
        if renamed_count == 0:
            log(f"未在 {root_path} 下找到任何 01插图签/{OLD_FOLDER_NAME} 文件夹需要重命名")
        else:
            log(f"完成重命名，共处理 {renamed_count} 个文件夹")
    except Exception as e:
        log(f"[ERROR] 遍历目录 {root_path} 时发生异常：{traceback.format_exc()}")
    
    return renamed_count


# 强制中断无限循环程序

"""
cmd运行
查看所有py程序，可能包括打开的脚本
wmic process where "name like '%python%'" get processid,commandline
强制结束进程
taskkill /PID 15696 /

"""

def kill_script_processes(script_path: str, force_kill: bool = True, kill_children: bool = True):
    """
    通用函数：终止匹配脚本路径的进程，包括无限循环的顽固进程。
    - script_path: 脚本绝对路径。
    - force_kill: 如果True，直接使用kill()（SIGKILL），适合无限循环。
    - kill_children: 如果True，递归终止子进程。
    返回: 已终止的PID列表。
    """
    abs_path = os.path.abspath(script_path)
    terminated_pids = []
    current_pid = os.getpid()

    for proc in psutil.process_iter(['pid', 'cmdline']):
        try:
            cmd = proc.info['cmdline'] or []
            if not cmd:
                continue
            if (len(cmd) > 1 and 'python' in cmd[0].lower() and
                os.path.abspath(cmd[1]) == abs_path and
                proc.pid != current_pid):
                if kill_children:
                    for child in proc.children(recursive=True):
                        _terminate_process(child, force_kill)
                _terminate_process(proc, force_kill)
                terminated_pids.append(proc.pid)
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue

    if terminated_pids:
        print(f"终止了进程: {terminated_pids}")
    else:
        print("未找到匹配进程")
    return terminated_pids

def _terminate_process(proc, force_kill):
    try:
        if force_kill:
            proc.kill()
        else:
            proc.terminate()
            proc.wait(timeout=5)
    except psutil.TimeoutExpired:
        proc.kill()
    except Exception as e:
        print(f"终止失败 PID={proc.pid}: {e}")
    time.sleep(1)
    if proc.is_running():
        os.kill(proc.pid, signal.SIGKILL if force_kill else signal.SIGTERM)

def run_daemon_improved(script_path: str):
    kill_script_processes(script_path)
    subprocess.Popen(
        [sys.executable, script_path],
        creationflags=CREATE_NO_WINDOW | DETACHED_PROCESS if os.name == 'nt' else 0
    )
    print(f"启动新守护进程: {script_path}")


#守护启动无限循环脚本

"""

[](run_daemon)


"""



#&&% 每隔T分钟运行一个指定名称的脚本


def run_py1(pyname):
    try:
        result = subprocess.run(
            [sys.executable, pyname],
            check=True,
            text=True,
            capture_output=True,
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        print(f"✅ 程序 {pyname} 执行成功。输出:\n{result.stdout}")
    except subprocess.CalledProcessError as e:
        print(f"❌ 运行 {pyname} 时发生错误: {e}")
        print(f"错误信息: {e.stderr}")
    except FileNotFoundError:
        print(f"❌ 未找到程序 {pyname}。请检查文件名和路径。")

def run_script_periodically(pyname, T):
    """
    每隔 T 分钟运行指定名称的 Python 脚本，不会阻塞主线程。
    
    参数:
        pyname (str): 要运行的 Python 脚本文件名
        T (float): 时间间隔（分钟）
    """
    def periodic_task():
        while True:
            run_py1(pyname)  # 调用已定义的 run_py 函数
            time.sleep(T * 60)  # 将分钟转换为秒

    # 创建并启动一个后台线程来运行周期性任务
    thread = threading.Thread(target=periodic_task, daemon=True)
    thread.start()
    print(f"已启动周期性任务，每 {T} 分钟运行一次 {pyname}")






#&&&&%% 第十三部分  基础服务 

#_____________________________________________________________________________________________________________________________________________


#  模块使用说明

"""
该模块解决插图签、编目录、打印的基础问题

基于当前文件的操作，大大简化程序。例如编目录，也就是获取打印框线，图纸名称，目录底版几个操作步骤而已。整合到文件夹操作不是不可以，而是增加的这些
复杂性没有什么价值。




"""



def _select_by_type(etype: str, retry=5) -> List:
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    for k in range(1, retry+1):
        try:
            try: doc.SelectionSets.Item("SS_TMP2").Delete()
            except: pass
            ss = doc.SelectionSets.Add("SS_TMP2")
            ss.Select(5, 0, 0, vtInt([0]), vtVariant([etype]))
            return list(ss)
        except: time.sleep(0.2)
    return []

def entities_to_coord_info(entities, tol_z=1e-4):
    coord_info = []
    for ent in entities:
        try:
            p1, p2 = ent.GetBoundingBox()
            x1, y1, z1 = p1
            x2, y2, z2 = p2
            if abs(z1 - z2) > tol_z:   # 只留下最小的 Z
                z = min(z1, z2)
            pts2d = [(x1,y1),(x1,y2),(x2,y2),(x2,y1)]
            coord_info.append((pts2d,1))
        except: pass
    return coord_info




def print_com_info(info_dict):
    """
    接受 draw_pl_and_extract_info 返回的字典，
    将每个 COM 对象的信息分块分行打印。
    """
    for idx, data in info_dict.items():
        print(f"--- 第 {idx} 号块 ---")
        print(f"块 名       : {data['block_name']}")
        print(f"四 角 坐 标 :")
        for corner in data['corners']:
            print(f"    {corner}")
        print(f"图 幅       : {data['drawing_frame']}")
        print(f"比 例       : {data['ratio']}")
        print(f"规 格       : {data['spec']}")
        print()  # 空行分隔



#从多段线获取信息生成字典

def draw_pl_and_extract_info(resu, layer_name="测试辅助", width=0, color=256):
    """
    :param resu: List of (blk_entity, corners)  
                 corners = [(x_ll,y_ll,z), (x_lu,y_lu,z), (x_ru,y_ru,z), (x_rl,y_rl,z)]
    :param layer_name: 要绘制多段线的图层名
    :param width: 多段线常宽
    :param color: 多段线颜色索引
    :return: dict mapping index -> {
                'block_name': 原始块实体 Name,
                'corners': corners,
                'drawing_frame': 图幅,
                'ratio': 比例,
                'spec': 规格,
                'polyline': the new LWPOLYLINE COM object
             }

    """
    # 2) 确保图层存在
    try:
        lyr = doc.Layers.Item(layer_name)
    except Exception:
        lyr = doc.Layers.Add(layer_name)
    lyr.LayerOn = True

    results = {}

    for idx, (blk, corners) in enumerate(resu):
        # corners 是四个三元组，但 AddLightWeightPolyline 只要 XYZ 展平成 2D
        raw = []
        for x, y, z in corners:
            raw.extend((x, y))
        # 转 COM 数组
        arr = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, raw)

        # 绘制多段线
        lw = mp.AddLightWeightPolyline(arr)
        lw.Layer         = layer_name
        lw.ConstantWidth = width
        lw.Color         = color
        lw.Closed        = True

        # 调用你的分析函数
        try:
            frame, ratio, spec = generate_name_and_ratio_from_polyline(lw)
        except Exception as e:
            frame, ratio, spec = None, None, None
            print(f"[WARN] 第 {idx} 条多段线分析失败: {e}")

        # 记录
        results[idx] = {
            'block_name': blk.Name if hasattr(blk, 'Name') else None,
            'corners': corners,
            'drawing_frame': frame,
            'ratio': ratio,
            'spec': spec,
            'polyline': lw
        }

    print_com_info(results)

    #刷新

    doc.SendCommand("re\n")

    doc.SendCommand("Z\ne\n")

    time.sleep(1)

    return results



def draw_pl_and_extract_from_entities(
    entities,
    layer_name: str = "测试辅助",
    width: int = 0,
    color: int = 256,
    A3dy: int = 0,
    Fandy: Tuple[str, str, str, int] = ("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3", 0),
):
    """
    把 entities 统一绘制成闭合轻量多段线，并提取图幅/比例/规格信息。
    返回 results 字典，下游函数以之为 frame_info。
    """
    li()     

    coord_info, corners_list = [], []
    for ent in entities:
        if not hasattr(ent, "GetBoundingBox"):
            continue
        p1, p2 = ent.GetBoundingBox()
        minx, miny, _ = p1
        maxx, maxy, _ = p2
        corners = [
            (minx, miny, 0), (minx, maxy, 0),
            (maxx, maxy, 0), (maxx, miny, 0),
        ]
        corners_list.append(corners)
        coord_info.append(([(x, y) for x, y, _ in corners], 1))

    plines = plcoor_to_com(coord_info, layer_name=layer_name, width=width, color=color)

    results = {}
    for idx, (corners, pline) in enumerate(zip(corners_list, plines)):
        try:
            block_name = pline.Owner.Block.Name
        except Exception:
            block_name = None

        info = generate_name_and_ratio_from_com(pline, A3dy=A3dy, Fandy=Fandy)
        if not info or info == 0:
            drawing_frame = ratio = spec = orient = None
        elif len(info) == 3:
            drawing_frame, ratio, spec = info
            orient = None
        else:
            drawing_frame, ratio, spec, orient = info[:4]

        results[idx] = {
            "entity":        pline,
            "block_name":    block_name,
            "corners":       corners,
            "drawing_frame": drawing_frame,
            "ratio":         ratio,
            "spec":          spec,
            "orient":        orient,
        }

    print_com_info(results)
    doc.SendCommand("RE\n"); doc.SendCommand("Z\nE\n")
    time.sleep(0.3)
    return results


def insert_block_into_poly_area(block_name, poly_ent, k=1.0, max_retries=3):
    """
    在给定的多段线/多边形 COM 对象 poly_ent 所定义区域内插入已定义块，
    横向区域（宽 >= 高）在左下插入；竖向区域在左上插入并顺时针旋转 90°。

    :param block_name: AutoCAD 中已定义的块名称（字符串）
    :param poly_ent: COM 多段线/多边形，直线，块对象等，必须支持 GetBoundingBox()
    :param k: 块插入时的 X/Y/Z 比例因子
    :param max_retries: 最多重试次数
    :return: (orientation, insert_point, block_ref)
    """
    # 1) 验证 poly_ent
    if not hasattr(poly_ent, "GetBoundingBox"):
        raise TypeError("第一个参数 poly_ent 必须是具有 GetBoundingBox() 方法的 COM 对象")

    # 2) 计算包围盒和方向
    p1, p2 = poly_ent.GetBoundingBox()
    minx, miny, minz = p1
    maxx, maxy, maxz = p2
    width  = maxx - minx
    height = maxy - miny

    orientation = 0 if width >= height else 1
    ins_pt = (minx, miny, minz) if orientation == 0 else (minx, maxy, minz)

    # 3) 插入块前预备 COM 对象
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    mp   = doc.ModelSpace

    ins_var = win32com.client.VARIANT(
        pythoncom.VT_ARRAY | pythoncom.VT_R8,
        ins_pt
    )

    block_ref = None
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            block_ref = mp.InsertBlock(ins_var, block_name, k, k, k, 0.0)
            # 如果成功，这里立即跳出重试循环
            break
        except Exception as e:
            last_err = e
            print(f"⚠ 第 {attempt} 次插入块失败: {e}")
            time.sleep(0.5)
    else:
        # 如果所有尝试都失败
        raise RuntimeError(f"❌ 在 {max_retries} 次尝试后仍无法插入块 “{block_name}”：{last_err}")

    # 4) 竖向时顺时针 90°
    if orientation == 1 and block_ref is not None:
        try:
            block_ref.Rotation = -math.pi / 2
        except Exception as e:
            print("⚠ 设置 Rotation 失败：", e)

    return orientation, ins_pt, block_ref


def compute_insert_factors(entities, res, result):
    """
    根据 res 的 ratio/spec 与 result 的块定义，计算缩放系数 k。
    回传 [(entity, block_name, spec, k), ...]
    """
    result_by_spec = {
        info["spec"]: (info.get("block_name"), info.get("ratio"))
        for info in result.values()
        if info.get("spec")
    }

    outputs = []
    for ent in entities:
        # 1) 找到 res 中记录
        matched = next((info for info in res.values() if info["entity"] is ent), None)
        if not matched:
            outputs.append((ent, None, None, None))
            continue

        spec    = matched.get("spec")
        ratio1  = matched.get("ratio") or "1:1"

        # 2) 找 result 中同 spec
        block_name, ratio2 = result_by_spec.get(spec, (None, "1:1"))

        # 3) 计算 k
        def denom(s):  # 提取分母
            m = re.match(r".*:(\d+)$", s)
            return int(m.group(1)) if m else 1

        d1, d2 = denom(ratio1), denom(ratio2)
        k = d1 / d2 if d2 else None

        outputs.append((ent, block_name, spec, k))

    return outputs

def get_factor_for_entity(entity, factors):
    """
    从 factors 列表中，找到第一个第 0 项是 entity 的元组并返回它，
    如果找不到则返回 None。
    """
    for tup in factors:
        if tup[0] is entity:
            return tup
    return None




#查看一个图签块的标签和值

def inspect_block_attributes(block):
    """
    打印属性块 block 中所有属性标签及其当前值。

    :param block: 支持 GetAttributes() 的 COM 块引用对象
    """
    try:
        attrs = block.GetAttributes()
    except Exception as e:
        print("❌ 该对象不是属性块或无法获取属性：", e)
        return

    if not attrs:
        print("⚠ 该块没有属性。")
        return

    tags = []
    values = []
    print("—— 属性标签与当前值 ——")
    for attr in attrs:
        tag = attr.TagString
        val = attr.TextString
        tags.append(tag)
        values.append(val)
        print(f"• {tag!r}  = {val!r}")
    print("—— 共 %d 个属性 ——" % len(attrs))
    return tags, values

#查看当前打印文件默认打印框线图层上的图签块的标签和值

def 查看标签列表(layername="dy_quyu"):
    """
    打印属性块 block 中所有属性标签及其当前值。

    :param block: 支持 GetAttributes() 的 COM 块引用对象
    """
    LB,_ = fenchu_tuqian_dayinxian(layername=layername)    

    block = LB[0]

    try:
        attrs = block.GetAttributes()
    except Exception as e:
        print("❌ 该对象不是属性块或无法获取属性：", e)
        return

    if not attrs:
        print("⚠ 该块没有属性。")
        return

    tags = []
  
    for attr in attrs:
        tag = attr.TagString
        
        tags.append(tag)
      
        
    print("—— 共 %d 个属性 ——" % len(attrs))
    return tags



def highlight_entities_in_window_1(x1, y1, x2, y2, ty: float = 1.0, select_mode: str = "_C"):
    """Zoom + Select 高亮矩形窗口内对象。

    Parameters
    ----------
    x1, y1, x2, y2 : float
        对角坐标，可任意顺序。
    ty : float, default 1.0
        Zoom 后等待秒数；Select 后再等待 *ty / 2* 秒。
    select_mode : str, default "_C"
        '_C' Crossing ；'_W' Window。
    """
    # —— 归一化坐标 ——
    x_lo, x_hi = sorted((x1, x2))
    y_lo, y_hi = sorted((y1, y2))

    # —— 加 20% 缓冲做 Zoom ——
    buf = 0.20 * ((x_hi - x_lo) + (y_hi - y_lo)) / 2
    zoom_cmd = (
        "_.ZOOM\n_W\n"
        f"{x_lo - buf},{y_lo - buf}\n"
        f"{x_hi + buf},{y_hi + buf}\n"
    )
    doc.SendCommand(zoom_cmd)
    time.sleep(ty)

    sel_cmd = (
        f"_.SELECT\n{select_mode}\n"
        f"{x_lo},{y_lo}\n"
        f"{x_hi},{y_hi}\n\n"
    )
    doc.SendCommand(sel_cmd)
    time.sleep(ty / 2)
    print(f"✅ 高亮窗口 ({x_lo:.0f},{y_lo:.0f})–({x_hi:.0f},{y_hi:.0f}) 完成")




def sort_text_objs(objs, orient: str, cha=150):
    """
    将文字对象列表按阅读顺序排序并返回 **文本字符串列表**

    orient = 'h' 横向框（上→下，左→右）
    orient = 'v' 竖向框（右→左，上→下）
    """
    if not objs:
        return []

    # 左下角坐标
    def ll(ob):
        (x1, y1, _), _ = ob.GetBoundingBox()
        return x1, y1

    wrapped = [(ob, *ll(ob)) for ob in objs]   # (obj, x_ll, y_ll)

    if orient == 'h':                      # ── 横向 ──
        # 先按 y 降序（上→下）
        wrapped.sort(key=lambda t: -t[2])
        i = 0
        while i < len(wrapped) - 1:
            j = i + 1
            while j < len(wrapped) and abs(wrapped[i][2] - wrapped[j][2]) < cha:
                j += 1
            # 同一行内按 x 升序（左→右）
            wrapped[i:j] = sorted(wrapped[i:j], key=lambda t: t[1])
            i = j

    else:                                  # ── 竖向 ──
        # 先按 x 降序（右→左）
        wrapped.sort(key=lambda t: -t[1])
        i = 0
        while i < len(wrapped) - 1:
            j = i + 1
            while j < len(wrapped) and abs(wrapped[i][1] - wrapped[j][1]) < cha:
                j += 1
            # 同一列内按 y 降序（上→下）
            wrapped[i:j] = sorted(wrapped[i:j], key=lambda t: -t[2])
            i = j

    return [get_text_content(ob) for ob, _, _ in wrapped]


def sort_Handle(com_list, cha_Y=2000):
    """
    给定一组 COM 实体，按它们的 BoundingBox 左下角排序（同一行 Δy<cha_Y 时按 x 升序），
    返回排序后各实体的 Handle 列表。

    :param com_list: 可迭代的 COM 实体列表
    :param cha_Y:    同一“行”判定的 Y 差阈值，默认 2000
    :return:         按 sort_coms_by_llcorner 排序后，各实体的 Handle 列表
    """
    # 先按左下角排序拿回实体列表
    sorted_entities = sort_coms_by_llcorner(com_list, cha_Y=cha_Y)
    # 再提取 Handle
    return [getattr(ent, "Handle", None) for ent in sorted_entities]




#从指定标签名获取每个图签块对应的值返回


def get_tag_values_from_tqs(bqname: str, tqs):
    """
    按标签名 bqname 依次提取每个图签块的标签值。
    若某块缺少该标签则返回 None，占位保持顺序。

    :param bqname:   目标标签（TagString）
    :param tqs:      图签块实体列表
    :return:         长度 == len(tqs) 的值列表
    """
    import re

    _prefix_re = re.compile(r'^\\[A-Za-z]\d*\.?\d*;')

    if not tqs:
        return []

    values = []
    for blk in tqs:
        # 把当前块的属性做成 dict
        tags, zs = huoqukuai_shuxing_zhi(blk)   # 原函数，返回两个并行列表
        tagdict = dict(zip(tags, zs))

        raw = tagdict.get(bqname)
        if raw is None:
            values.append(None)
        else:
            cleaned = _prefix_re.sub('', raw)
            values.append(cleaned)

    # 如果目标标签在第一个块里都不存在，统一返回 None
    if all(v is None for v in values):
        print(f"⚠ 在任何图签块中都未找到标签 '{bqname}'")
        return None
    return values

#################



#测量已有文字长度

def celiang_wenzichangdu(TEXTCOM):

    text_copy = TEXTCOM.Copy()

    text_copy.Alignment = 2

    text_copy.TextAlignmentPoint =vtpnt(0,0,0)

    chang = abs(text_copy.InsertionPoint[0])

    text_copy.Delete()

    return chang

#测量新写文字长度

def celiang_wenzichangdu_write(ZF,style="图签",height=270,scalefactor=0.8):

    #根据字符串按样式字高宽度因子写入cad后的测量长度

    text_obj = acad.ActiveDocument.ModelSpace.AddText(ZF, vtpnt(0,0,0), height)

    text_obj.StyleName = style

    text_obj.ScaleFactor =scalefactor #宽度因子

    chang = celiang_wenzichangdu(text_obj)

    text_obj.Delete()

    return chang





#判定文字字体是否是shx字体
def is_shx_font(text_obj):
    """
    判断给定 CAD 文字对象是否使用 SHX 字体。

    参数：
        text_obj: COM 的 Text 或 MText 实体对象

    返回：
        True 表示使用 SHX 字体，False 表示使用 TrueType 字体或未知
    """
    try:
        style_name = text_obj.StyleName
        text_styles = doc.TextStyles
        style = text_styles.Item(style_name)

        font_file = style.FontFile
        if font_file.lower().endswith(".shx"):
            return True
        else:
            return False
    except Exception as e:
        print("⚠️ 无法判断字体类型：", e)
        return False



##判断是否广义标准打印线

def get_print_template_info(com_obj, tol=50):
    """
    判断给定 AutoCAD COM 对象的外包盒尺寸是否与下列 8 种标准之一在 tol 范围内匹配，
    并返回对应的打印模板信息。若没有匹配则返回 0。

    标准尺寸（单位与 CAD 外包盒单位一致）及对应返回值：
      65340.0 x 46200.0   → ("UserDefinedMetric (653.40 x 462.00毫米)", "1:100", "A2")
      71280.0 x 50400.0   → ("UserDefinedMetric (712.80 x 504.00毫米)", "1:100", "A2")
      83160.0 x 58800.0   → ("UserDefinedMetric (831.60 x 588.00毫米)", "1:150", "A2")
      77220.0 x 54600.0   → ("UserDefinedMetric (772.20 x 546.00毫米)", "1:150", "A2")
      92510.0 x 65340.0   → ("UserDefinedMetric (925.10 x 653.40毫米)", "1:100", "A1")
      100920.0 x 71280.0  → ("UserDefinedMetric (1009.20 x 712.80毫米)", "1:100", "A1")
      109330.0 x 77220.0  → ("UserDefinedMetric (1093.30 x 772.20毫米)", "1:150", "A1")
      117740.0 x 83160.0  → ("UserDefinedMetric (1177.40 x 831.60毫米)", "1:150", "A1")

    参数：
      com_obj -- 支持 GetBoundingBox() 的 AutoCAD COM 对象
      tol     -- 容差值（与标准尺寸比较时的绝对值阈值），默认为 20

    返回：
      若匹配成功，返回三元组：
        (
          模板名称字符串 (如 "UserDefinedMetric (653.40 x 462.00毫米)"),
          比例 (如 "1:100"),
          图纸尺寸 (如 "A2")
        )
      若任意一组都不匹配，返回 0。
    """
    # 1. 先获取外包盒的 minPt, maxPt
    min_pt, max_pt = com_obj.GetBoundingBox()
    dx = max_pt[0] - min_pt[0]
    dy = max_pt[1] - min_pt[1]

    # 2. 计算“长”和“宽”
    obj_length = max(dx, dy)
    obj_width  = min(dx, dy)

    # 3. 预定义所有标准尺寸及其对应的信息
    #    格式： (标准长, 标准宽, 比例, 图纸尺寸)
    standard_map = [
        (65340.0,   46200.0,  "1:100", "A2"),
        (71280.0,   50400.0,  "1:100", "A2"),
        (83160.0,   58800.0,  "1:150", "A2"),
        (77220.0,   54600.0,  "1:150", "A2"),
        (92510.0,   65340.0,  "1:100", "A1"),
        (100920.0,  71280.0,  "1:100", "A1"),
        (109330.0,  77220.0,  "1:150", "A1"),
        (117740.0,  83160.0,  "1:150", "A1"),
    ]

    # 4. 遍历每组标准，若在 tol 范围内，则格式化返回值
    for std_len, std_wid, scale, sheet in standard_map:
        if abs(obj_length - std_len) < tol and abs(obj_width - std_wid) < tol:
            # 将 CAD 单位转换为毫米，并保留两位小数
            mm_len = std_len / 100.0
            mm_wid = std_wid / 100.0
            template_name = f"UserDefinedMetric ({mm_len:.2f} x {mm_wid:.2f}毫米)"
            return template_name, scale, sheet

    # 5. 若都不匹配，返回 0
    return 0





#&&&% ▶ 基础服务整体框架
"""
任何进来的含标准打印区域的文件都会被操作
1 文件状态只允许下面的状态  1） 一个当前文件 2）两个要处理的文件 3）一个当前文件和00.dwg 4) 00.dwg  



ensure_active_cad_file和youxiao_jiance表明，我们能够在文件5种可能运行状态下，确保对文件的操作是有效的。这是全自动流程的基础之一。
在一些重要的环节，要确保文件被打开操作。

这并不意味着函数都要以filepath为输入变量。只有部分关键函数才这样。或者我们在某个函数操作前调用这样的带文件路径名输入变量的函数。
一般函数还是假设在当前激活文件上操作的。


先把目录模板已经插好、基本图签模板做好、清除原图签、从图层选取图名走通


#把调度程序研究清楚

CAD出错，重启弹出窗口需取消 图形修复处理器需关闭 考虑todesk实时传递消息远程操控 activate_and_click_aikeyun()应该的窗口被遮盖没起作用




20250908  从整个系统的成败来讲，应该先实现以文件夹为基础的个人版的基础服务。这样，每个人都可以在电脑上自行操作，减少复杂性。完成之后，再考虑
在服务器架构一个独立于每人电脑之外的基础服务，急就用自己的电脑处理，不急就用服务器处理，已经解决了实际的需要，原则上不需要考虑复杂的调度。



"""

#&&% ⮞ F1: 确保文件为当前唯一激活连接文件

# 🔧 主函数 
def ensure_active_cad_file(filepath,retry=3):


    result, attempts, error = chongfu_caozuo(
        ensure_active_cad_file_single,      # 要重复调用的函数
        args=(filepath,),           # 位置参数        
        max_retries=retry,               # 最多重试次数
        failure_value=None           # 全部失败时返回 None
    )
    if error:
        # 你可以在这里统一处理失败的情况
        print(f"在 {attempts} 次尝试后仍未成功：{error}")
    return result

# 🧪 测试函数:
@timeit
def ceshi(test_func):
    """
    测试入口：仅当传入的 test_func.__name__ == 'ensure_active_cad_file' 时才运行。
    这样就可以把这个测试函数留在主脚本里，而不会和其他函数互相污染。
    """
    # 只在函数名匹配时才执行
    if test_func.__name__ != "ensure_active_cad_file":
        print(f"跳过测试：函数名 `{test_func.__name__}` 与 `ensure_active_cad_file` 不匹配")
        return

    # 测试用 DWG 路径
    filepath = (
        r"D:/Myprogramsystem/BaiduSyncdisk/资料/测试备份/"
        "远程国际各专业最后CAD0905/建筑/"
        "远程国际建施2021.0903(LT4、LT5楼梯修改)_t7.dwg"
    )

    # 开启调试模式
    enable_debug()
    try:
        # 真正调用传入的函数
        result = test_func(filepath, retry=3)
        print(f"测试结束，返回值：{result}")
    finally:
        # 无论成功或失败，都把调试模式关掉
        disable_debug()


def _cad_proc_count():
    names = ("acad.exe", "t20", "tcad")  # 需根据实际进程名前缀补充
    return sum(1 for p in psutil.process_iter(['name'])
               if p.info['name'] and p.info['name'].lower().startswith(names))


def find_cad_process_for_filename(name):
    """
    返回打开指定 DWG 文件（不带路径）的 CAD 窗口及进程信息列表。
    """
    target = name.lower()
    matches = []
    def _enum(hwnd, _):
        if not win32gui.IsWindowVisible(hwnd):
            return True
        title = win32gui.GetWindowText(hwnd) or ""
        if target in title.lower():
            _, pid = win32process.GetWindowThreadProcessId(hwnd)
            matches.append(pid)
        return True
    win32gui.EnumWindows(_enum, None)
    return matches

def close_other_cad_processes(filename):
    """
    关闭所有 CAD 进程，除了正在打开 filename 的那些进程。

    参数：
        filename (str): 不带路径的 DWG 文件名，如 "sample.dwg"
    """
    # 找到所有打开目标文件的 CAD 进程 PID 列表
    keep_pids = set(find_cad_process_for_filename(filename))
    print(f"保留进程 PID: {keep_pids or '<无扫描到目标文件进程>'}")

    # 枚举所有 CAD 相关进程
    cad_names = ("acad.exe", "tcad.exe", "t20cad.exe", "t20.exe")
    for proc in psutil.process_iter(['pid', 'name']):
        name = proc.info['name']
        pid  = proc.info['pid']
        if not name:
            continue
        if name.lower() in cad_names:
            if pid in keep_pids:
                print(f"保留 CAD 进程: PID={pid}, Name={name}")
            else:
                try:
                    print(f"关闭 CAD 进程: PID={pid}, Name={name}")
                    proc.kill()
                except Exception as e:
                    print(f"⚠️ 无法终止 PID={pid}: {e}")

@debuggable            # ←← 只要给函数加这个装饰器即可
def ensure_active_cad_file_single(filepath):
    """
    确保 filepath 是唯一激活 DWG；调试信息由 node() 控制。
    返回最终打开的文件数。
    Open_By_Omission_wenjian函数是通过return True 判断文件正常打开，并不限定要知道打开文件的内容和时间 如果出现异常它会抛异常
    此时才会激起下一轮的尝试

    """
    filepath = os.path.abspath(filepath)
    filename = os.path.basename(filepath)

    if _cad_proc_count()==0:
        node("无CAD进程开始")
        start_applicationV9()
        li()
        name1=doc.Name
        node("当前文件名{}",name1)          
        time.sleep(1)
        Open_By_Omission_wenjian(filepath)
        li(); time.sleep(1)           
        name2=doc.Name
        node("当前文件名{}",name2)
        try:    
            close_dwg_by_name(name1)
            node("关闭之前文件")
        except:
            node("之前文件已经自动关闭")
            pass
        if _cad_proc_count() > 1:
            procs = [
                p for p in psutil.process_iter(['pid','name','create_time'])
                if p.info['name'] and p.info['name'].lower().startswith(("acad.exe","t20","tcad"))
            ]
            # 按创建时间升序：前面的最旧，最后一个最新
            procs_sorted = sorted(procs, key=lambda p: p.info['create_time'])
            # 杀掉所有旧进程
            for old in procs_sorted[:-1]:
                node("关闭旧 CAD 进程 PID={} NAME={}", old.pid, old.info['name'])
                try:
                    old.kill()
                except Exception as ex:
                    node("⚠️ 杀进程失败 {}: {}", old.pid, ex)
            node("只保留最新进程，结束")
        try:
            line = mp.AddLine(vtpnt(0,0,0),vtpnt(0,10000,0))  
            node("新绘直线的句柄值：",line.Handle)            
            line.Delete()
        except:
            time.sleep(1)
            node("没有绘制直线")               
        node("1:无CAD进程情况")
        return 1


    # ① 尝试获取 AutoCAD
    node("① Dispatch AutoCAD.Application()")
    acad = win32com.client.Dispatch("AutoCAD.Application")
    docs = acad.Documents
    node("①.1 当前 Documents.Count={}", docs.Count)
    # ② li() 连接
    if not li():
        node("文件数不能计算情况进程开始")
        close_all_cad_processes()
        start_applicationV9()           
        li()
        node("当前文件名{}",doc.Name)
        current_doc = win32com.client.Dispatch("AutoCAD.Application").ActiveDocument
        Open_By_Omission_wenjian(filepath)
        li(); time.sleep(1)
        try:    
            current_doc.Close(False)
        except:
            node("之前文件已经自动关闭")
            pass
        if _cad_proc_count() > 1:
            # 杀掉除最新之外的旧进程
            procs = [
                p for p in psutil.process_iter(['pid','name','create_time'])
                if p.info['name'] and p.info['name'].lower().startswith(("acad.exe","t20","tcad"))
            ]
            # 按创建时间升序：前面的最旧，最后一个最新
            procs_sorted = sorted(procs, key=lambda p: p.info['create_time'])
            # 杀掉所有旧进程
            for old in procs_sorted[:-1]:
                node("关闭旧 CAD 进程 PID={} NAME={}", old.pid, old.info['name'])
                try:
                    old.kill()
                except Exception as ex:
                    node("⚠️ 杀进程失败 {}: {}", old.pid, ex)
            node("只保留最新进程，结束")
        try:
            line = mp.AddLine(vtpnt(0,0,0),vtpnt(0,10000,0))   
            node("新绘直线的句柄值：",line.Handle)            
            line.Delete()
        except:
            time.sleep(1)
            node("没有绘制直线")               
        node("2:文件数不能计算情况")            
        return 1
    else:
        node("② li() 成功")       
    count  = docs.Count
    active = acad.ActiveDocument if count else None
    node("③ 文件数={} | Active={}", count, getattr(active, 'Name', None))
    current_doc = win32com.client.Dispatch("AutoCAD.Application").ActiveDocument
    
    # ④ 仅 1 个文件
    if count == 1:
        if active.FullName.lower() == filepath.lower():
            node("3:④✓ 唯一文件即目标 → 结束")
            return 1

        node("④✗ 唯一文件≠目标 → 打开目标并关闭旧")
        name1=doc.Name
        node("当前文件名{}",name1)           
        time.sleep(1)
        Open_By_Omission_wenjian(filepath)
        li(); time.sleep(1)           
        name2=doc.Name
        node("当前文件名{}",name2)
        try:    
            close_dwg_by_name(name1)
            node("关闭之前文件")
        except:
            node("之前文件已经自动关闭")
            pass
        if _cad_proc_count() > 1:
            procs = [
                p for p in psutil.process_iter(['pid','name','create_time'])
                if p.info['name'] and p.info['name'].lower().startswith(("acad.exe","t20","tcad"))
            ]
            # 按创建时间升序：前面的最旧，最后一个最新
            procs_sorted = sorted(procs, key=lambda p: p.info['create_time'])
            # 杀掉所有旧进程
            for old in procs_sorted[:-1]:
                node("关闭旧 CAD 进程 PID={} NAME={}", old.pid, old.info['name'])
                try:
                    old.kill()
                except Exception as ex:
                    node("⚠️ 杀进程失败 {}: {}", old.pid, ex)
            node("只保留最新进程，结束")
        try:
            line = mp.AddLine(vtpnt(0,0,0),vtpnt(0,10000,0))   
            node("新绘直线的句柄值：{}",line.Handle)            
            line.Delete()
        except:
            time.sleep(1)
            node("没有绘制直线")               
        node("4:一个非目标文件情况")
        return 1

    # ⑦ 文件数为 0 或 >1
    else:  

        node("⑦ 文件数={} (0 或 >1) → 全部重启", count)
        close_all_cad_processes()
        start_applicationV9()           
        li()
        node("当前文件名{}",doc.Name)
        current_doc = win32com.client.Dispatch("AutoCAD.Application").ActiveDocument
        Open_By_Omission_wenjian(filepath)
        li(); time.sleep(1)
        try:    
            current_doc.Close(False)
        except:
            node("之前文件已经自动关闭")
            pass
        if _cad_proc_count() > 1:
            # 杀掉除最新之外的旧进程
            procs = [
                p for p in psutil.process_iter(['pid','name','create_time'])
                if p.info['name'] and p.info['name'].lower().startswith(("acad.exe","t20","tcad"))
            ]
            # 按创建时间升序：前面的最旧，最后一个最新
            procs_sorted = sorted(procs, key=lambda p: p.info['create_time'])
            # 杀掉所有旧进程
            for old in procs_sorted[:-1]:
                node("关闭旧 CAD 进程 PID={} NAME={}", old.pid, old.info['name'])
                try:
                    old.kill()
                except Exception as ex:
                    node("⚠️ 杀进程失败 {}: {}", old.pid, ex)
            node("只保留最新进程，结束")
        try:
            line = mp.AddLine(vtpnt(0,0,0),vtpnt(0,10000,0))   
            node("新绘直线的句柄值：",line.Handle)            
            line.Delete()
        except:
            time.sleep(1)
            node("没有绘制直线")               
        node("5:2个以上文件情况")
        return 1

#&&% ⮞ F2: 确保对当前激活文件操作有效

def effect_on_file():

    
    li()
    lineobj = mp.AddLine(vtpnt(0,0,0),vtpnt(10000,10000,0))
    lineobj.Color=1

    print(lineobj.StartPoint,lineobj.EndPoint)
    res1 = lineobj.StartPoint
    res2 = lineobj.EndPoint
    chang = sqrt((res1[0]-res2[0])**2+(res1[1]-res2[1])**2)
    i=int(chang)
    print(lineobj.Handle)
    time.sleep(1)
    safe_delete(lineobj)

    return i == 14142     


def retry_effect_on_file():
    max_attempts = 7
    for attempt in range(max_attempts):
        try:
            if effect_on_file():
                return True
        except Exception as e:
            print(f"Attempt {attempt + 1} failed: {e}")
            if attempt < max_attempts - 1:
                time.sleep(5)
    return False  # 如果所有尝试失败，返回 "False"




#&&% ⮞ F3: Excle控制文件

# ➖➖ 固定坐标 ➖➖
COLS_D2N = [col2idx(c) for c in "DEFGHIJKLMN"]       # D~N 11 列
ROW_INFO1_KEY    = 18                                 # 图签基础信息  键行
ROW_INFO1_TAG    = 19                                 # 图签基础信息  标签行
ROW_INFO1_VALUE  = 20                                 # 图签基础信息  值行

ROW_INFO2_KEY    = 25                                 # 目录首行示例信息  键行
ROW_INFO2_TAG    = 26                                 # 目录首行示例信息  图层行
ROW_INFO2_VALUE  = 27                                 # 目录首行示例信息  值行

COL_P = col2idx("P")                                  # 图纸编号
COL_Q = col2idx("Q")                                  # 图纸名称
ROW_NAMES_KEY    = 31                                 # 图纸名称信息  键行
ROW_NAMES_START  = 32                                 # 图纸名称信息  数据起始行


# =============== 读取 ===============
def read_fixed_excel(path =r"D:/Myprogramsystem/BaiduSyncdisk/基础服务/宋岳/04Excel控制文件/不控制图纸名称/标准模板.xlsx" ):
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active

    # ── 1. 图签基础信息 ──
    keys_1   = [ws.cell(ROW_INFO1_KEY,   c).value or "" for c in COLS_D2N]
    tags_1   = [ws.cell(ROW_INFO1_TAG,   c).value or "" for c in COLS_D2N]
    vals_1   = [ws.cell(ROW_INFO1_VALUE, c).value or "" for c in COLS_D2N]

    info1: Dict[str, str] = dict(zip(keys_1, vals_1))
    info1["对应标签"]  = tags_1
    info1["对应取值"]  = vals_1

    # ── 2. 目录首行示例信息 ──
    keys_2   = [ws.cell(ROW_INFO2_KEY,   c).value or "" for c in COLS_D2N]
    layers_2 = [ws.cell(ROW_INFO2_TAG,   c).value or "" for c in COLS_D2N]
    vals_2   = [ws.cell(ROW_INFO2_VALUE, c).value or "" for c in COLS_D2N]

    info2: Dict[str, str] = dict(zip(keys_2, vals_2))
    info2["对应图层"]  = layers_2
    info2["对应取值"]  = vals_2

    # ── 3. 图纸名称信息（开放多行） ──
    nums, names = [], []
    r = ROW_NAMES_START
    while True:
        num  = ws.cell(r, COL_P).value
        name = ws.cell(r, COL_Q).value
        if (num is None or str(num).strip() == "") and (name is None or str(name).strip() == ""):
            break
        nums.append("" if num is None else str(num))
        names.append("" if name is None else str(name))
        r += 1
    info3 = {"图纸编号": nums, "图纸名称": names}

    return {
        "图签基础信息":      info1,
        "目录首行示例信息":  info2,
        "图纸名称信息":      info3,
    }


# =============== 写入 ===============
def write_fixed_excel(
    zd: Dict[str, dict],
    template_path = r"D:/Myprogramsystem/BaiduSyncdisk/基础服务/公用文件/Excel控制文件模板/Excel控制文件标准模板.xlsx",
    out_path =r"D:/Myprogramsystem/BaiduSyncdisk/基础服务/宋岳/04Excel控制文件/实际反馈/标准模板更新.xlsx" ,
):

    """
    重复相同名字的写入路径，则会自动覆盖
   
    """
    wb = load_workbook(str(template_path))
    ws = wb.active

    # 获取三块数据（若缺少则跳过）
    info1 = zd.get("图签基础信息", {})
    info2 = zd.get("目录首行示例信息", {})
    info3 = zd.get("图纸名称信息", {})

    # ── 1. 写图签基础信息 ──
    keys_1 = [ws.cell(ROW_INFO1_KEY, c).value for c in COLS_D2N]
    tags_1 = info1.get("对应标签", [])
    vals_1 = info1.get("对应取值", [info1.get(k, "") for k in keys_1])

    for c, tag in zip(COLS_D2N, tags_1):
        ws.cell(ROW_INFO1_TAG, c, tag)
    for c, val in zip(COLS_D2N, vals_1):
        ws.cell(ROW_INFO1_VALUE, c, val if val is not None else "")

    # ── 2. 写目录首行示例信息 ──
    keys_2 = [ws.cell(ROW_INFO2_KEY, c).value for c in COLS_D2N]
    layers_2 = info2.get("对应图层", [])
    vals_2   = info2.get("对应取值", [info2.get(k, "") for k in keys_2])

    for c, layer in zip(COLS_D2N, layers_2):
        ws.cell(ROW_INFO2_TAG, c, layer)
    for c, val in zip(COLS_D2N, vals_2):
        ws.cell(ROW_INFO2_VALUE, c, val if val is not None else "")

    # ── 3. 写图纸名称信息 ──
    nums  = info3.get("图纸编号", [])
    names = info3.get("图纸名称", [])
    max_len = max(len(nums), len(names))

    # 先清空旧行
    r = ROW_NAMES_START
    while not _row_is_empty(ws, r, COL_P, COL_Q):
        ws.cell(r, COL_P, "")
        ws.cell(r, COL_Q, "")
        r += 1

    # 再写新数据
    for i in range(max_len):
        ws.cell(ROW_NAMES_START + i, COL_P, nums[i]  if i < len(nums)  else "")
        ws.cell(ROW_NAMES_START + i, COL_Q, names[i] if i < len(names) else "")

    # 保存
    out_path = template_path if out_path is None else Path(out_path)
    wb.save(str(out_path))
    return Path(out_path)


# =============== 小工具 ===============
def _row_is_empty(ws, r: int, *cols) -> bool:
    return all((ws.cell(r, c).value in (None, "")) for c in cols)




#&&% ⮞ F4: 控制CAD弹出窗口、后台控制无限循环程序


def run_daemon(script_path: str):
    """
    后台守护脚本：先终止旧的同脚本进程，再以隐藏窗口方式启动新进程。

    run_daemon("D:/Myprogramsystem/CentralControlProgram/cad_dialog_killer.py")


    """
    # 1) 杀掉旧进程
    for proc in psutil.process_iter(['pid','cmdline']):
        try:
            cmd = proc.info['cmdline'] or []
            # 只要命令行里出现脚本路径，就视为旧守护进程
            if script_path in " ".join(cmd):
                proc.terminate()
                try:
                    proc.wait(timeout=3)
                except psutil.TimeoutExpired:
                    proc.kill()
                log(f"🔪 已终止旧守护: pid={proc.pid}")
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue

    # 2) 启动新进程
    try:
        subprocess.Popen(
            [sys.executable, script_path],
            creationflags=CREATE_NO_WINDOW | DETACHED_PROCESS
        )
        log(f"🛡️ 新守护启动: {script_path}")
    except Exception as e:
        log(f"❌ 守护启动失败: {script_path} 错误: {e}")



def log(msg):
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    try:
        os.makedirs(LOG_DIR, exist_ok=True)
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass




#&&% ⮞ F6: CAD文件循环处理函数的通用逻辑结构+



# 占位符函数定义（与之前相同）
def code_part1():
    # 核心功能：如果失败，抛出异常
    print("执行代码部分1")
    # 示例: raise Exception("模拟错误")  # 模拟失败
    # 实际核心代码...

def code_part2():
    # 成功记录
    print("代码部分2: 记录成功执行的消息")

def code_part4():
    # 失败记录
    print("代码部分4: 记录代码部分1产生错误的消息")

def code_part5():
    # 始终清理
    print("代码部分5: 无论成功失败都执行的清理操作")

# 主逻辑：使用 try-except-else-finally 建立可靠结构
max_attempts = 3
attempt = 0

try:
    while attempt < max_attempts:
        try:
            code_part1()  # 核心风险操作
        except Exception as e:
            print(f"捕获异常: {e}")  # 处理错误（可选日志）
            attempt += 1
            if attempt < max_attempts:
                # 失败时运行：延迟重试（代码部分3）
                try:
                    time.sleep(2)  # 失败专属操作：延迟
                except Exception as e2:
                    print(f"重试延迟失败: {e2}")  # 如果延迟出错，记录
            else:
                # 3次失败后：执行代码部分4
                code_part4()
        else:
            # 成功时：执行代码部分2
            code_part2()
            break  # 成功退出循环
finally:
    # 无论如何：执行代码部分5（清理）
    code_part5()























#*************************************************************************************************************************************************************

#&&&% 🧱 服务一 插图签完整流程



#&&&% （一） 放置目标dwg和文件和同名基本图签文件

"""


1 目标文件应使用标准图框或标准图框的1.2倍的单线框区域，左下角点水平排列整齐，按照从上到下、从左到右的顺序排列。只要不影响打印，打印区域可以是块或多段线。重复操作时，允许有插好的图签，部分

无图签等情况。

2 从D:/Myprogramsystem/BaiduSyncdisk/基础服务/公用文件/标准图签模板/标准图签模板.dwg 复制一个基本图签模板，改变名称为 '目标文件名+基本图签.dwg'，然后按视频指示修改

"""

#&&&% （二） 根据需要决定是否将目标文件放入  06插图签辅助

"""
1 把目标文件放在 D:/Myprogramsystem/BaiduSyncdisk/基础服务/宋岳/01插图签/07插图签辅助

这一步的目的是获得标准图签全部初步插好在各个图纸上，便于放置图纸名称有个参照,并能够在第一张属性标准图签上将图签基本信息即除了图纸名称和编号的信息写上。
也可以直接忽略这步，使用把图纸名称文字和图纸编号文字分别放入图层“图纸名称”和“图纸编号”的方法，或者直接使用Excel表格书写“图纸名称”和“图纸编号”的信息
也可以根据图纸的留白边距，将图纸名称或序号放入图签区域。1:100的边距是1000，因此1:150的是1500，1:50的是500,1:25的是250。设计人员是清楚某张图纸的比例的。

2 文件处理完后立即返回到D:/Myprogramsystem/BaiduSyncdisk/基础服务/宋岳/01插图签/04输出 ，标准的图签将放置到各个打印区域，但图纸名称和编号与图签内容都还没处理好。用户拿这个文件辅助放置好图纸名称和编号
再根据需要选择输入口


"""




#&&&% （三） 根据需要决定放入哪个输入文件夹

"""
插图签 → 编目录 → 打印 是一个复杂的过程，并非纯粹的单线流程

1把目标文件放在插图签的输入口D:/Myprogramsystem/BaiduSyncdisk/基础服务/宋岳/01插图签/01只插图签输入

根据需要，只插图签输入，则返回插好的图签，不仅仅是插好图签位置，也包括将图签上的文字都书写好了

或者既要插入图签，还要编目录，放对应的输入口  D:/Myprogramsystem/BaiduSyncdisk/基础服务/宋岳/01插图签/02插图签编目录输入

或者既要插入图签，还要编目录，乃至把打印搞好，就放对应的出入口 D:/Myprogramsystem/BaiduSyncdisk/基础服务/宋岳/01插图签/03插图签编目录打印输入

只是修改几个标签名对应的值，可以放入 D:/Myprogramsystem/XT/Basic_service_processing/01插图签/04修改标签值输入

"""



#&&&% （四） 确定打印区域及对应信息

"""
zd = determine_printing_area(tol = 50.0,cha_Y= 2000.0) 

考虑到实际情况可能面对重复操作，对全自动化处理插图签函数 这个仅仅只取打印区域的函数不够用 要清除属性标准图签 非属性标准图签不操作 因此应该使用
zd=process_print_areas_info(tol  = 50.0, tola = 100.0, aspect_tol = 4, area_min = 160000000,  area_max = 1000000000, tol2 = 3000.0,  cha_Y = 2000.0)
局部修改 普通图签文件应该自己补充完整  标准属性图签则不需要

从插图签流程的输入口这里获取到的属性标准图签，在全自动流程中将被清除。但是在“编目录流程”输入口就不会


测试问题是否解决





"""


#&&% ⮞ F1: 确定打印区域信息1

@timeit
@debuggable
def determine_printing_area(
    tol: float = 50.0,
    cha_Y: float = 2000.0
):
    """
    自动识别标准打印区域，并返回 LP_1 标准打印框线与其原始属性 zhi 的映射字典 prinfo
    属打印框线图层的对象都会被A3打印，在这里不用考虑
    标准尺寸放大110%,120%,130%等仍然按标准尺寸打印，这个后面再来处理20250704

    """
    li()
    enable_debug()


    # 1. 快速选择对象
    L1_zhun = []
    for fn in (select_polyline_chuantong, select_polyline, select_kuai):
        try:
            L1_zhun.extend(fn(5))
        except Exception as e:
            print(f"❌ {fn.__name__} 失败: {e}")

    L1 = [obj for obj in L1_zhun if getattr(obj, "Layer", "") != "print_tuqian_jiben"]
    node("[节点1] L1 总共选到的多段线和块对象数量: {}", len(L1))

    # 2. 识别 zhi 并筛选
    L3 = []
    zhi_dict = {}
    for comobj in L1:
        try:
            zhi = generate_name_and_ratio_from_com(
                comobj,
                A3dy=0,
                Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3", 0)
            )
            if zhi != 0:
                handle = getattr(comobj, "Handle", None)
                if handle:
                    zhi_dict[handle] = zhi
                L3.append(comobj)
        except Exception:
            continue
    node("[节点2] L3 : {}", len(L3))

    # 3. 去重
    L3_unique = dedupe_by_bbox_proximity(L3, tol=tol)

    # 4. 排序后生成标准打印框线
    now = datetime.datetime.now()
    timename = now.strftime("%Y-%m-%d-%H-%M")

    def _bbox_corners(obj):
        ll_point, ur_point = obj.GetBoundingBox()
        x1, y1, z1 = ll_point
        x2, y2, z2 = ur_point
        return [
            (x1, y1, z1),
            (x2, y1, z1),
            (x2, y2, z1),
            (x1, y2, z1)
        ]

    L3_sorted = sort_coms_by_llcorner(L3_unique, cha_Y=cha_Y)

    LP_1_raw = []
    for comobj in L3_sorted:
        try:
            coords3d = _bbox_corners(comobj)
            pl = draw_lwpolyline(
                coords3d,
                layer_name=f"标准打印框线{timename}",
                width=20.0,
                color=256,
                closed=True
            )
            LP_1_raw.append(pl)
        except Exception:
            continue

    LP_1_sorted = sort_coms_by_llcorner(LP_1_raw, cha_Y=cha_Y)
    handles_lp1 = [getattr(ent, "Handle", None) for ent in LP_1_sorted]
    node("[节点3] LP_1 标准打印区域数量 : {} handle{}", len(LP_1_sorted), handles_lp1)

    # 🔁 构造返回字典
    prinfo = {}
    for i in range(min(len(LP_1_sorted), len(L3_sorted))):
        h_new = getattr(LP_1_sorted[i], "Handle", None)
        h_old = getattr(L3_sorted[i], "Handle", None)
        if h_new and h_old in zhi_dict:
            prinfo[h_new] = zhi_dict[h_old]


    prinfo["dyx_coms"] = LP_1_sorted

    #  保存
    try: savefile()
    except: pass


    disable_debug()

    return prinfo



#&&% ⮞ F2: 确定打印区域信息2


"""
zd=process_print_areas_info(tol  = 50.0, tola = 100.0, aspect_tol = 4, area_min = 160000000,  area_max = 1000000000, tol2 = 3000.0,  cha_Y = 2000.0)

"""


#&&% ⮞ F3: 删除已有图签


"""
属性标准图签需要删除之前已有图签，普通图签模式也需要删除已有图签。如果属性标准图签仅仅是修改标签值，放入另外的专门输入口处理。普通标准图签的局部修改自己手工完成。

"""






#双线程插入公司基本图签

def insert_company_label_common_block(

    insertion_point=(0, 0, 0),

    filepath=r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/自动化(动态)/中亿图签3.dwg",

    scale=(1, 1, 1),
    rotation=0,
    wait=0.3
):

    """
    按照模板文件制作属性块图签，并导出为公共块
    中亿图签3.dwg就是一个公共块

    resu --- - 它是一个包含多对象的列表，每个元就是一个com打印图块（带图框和图签的块）的com实体及对应的4个角点坐标信息

    coms_tuqian——插入的图千块列表
    

   """
    #1插入公司图签

    resu = insert_and_explode_dwg(filepath,
                               insertion_point=(0, 0, 0),
                               scale=scale,
                               rotation=rotation,
                               wait=wait)

    resu =resu[0]   
    #显示插入的多个图签信息

    xinxi_tuqiankuai = draw_pl_and_extract_info(resu, layer_name="测试辅助", width=0, color=256)

    #2返回图签实体coms和对应信息字典

    coms_tuqian =[]
    names_tuqian =[]

    for ob in resu:

        coms_tuqian.append(ob[0])
        names_tuqian.append(ob[0].Name)



    time.sleep(1)

    savefile()

    return coms_tuqian ,names_tuqian,xinxi_tuqiankuai 

def f1_insert_company_getwindow(
    timeout_event, done_event,
    *,
    result_box: dict,                     # ★ 新增：线程安全的共享字典
    insertion_point=(0, 0, 0),
    filepath=r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/自动化(动态)/中亿图签3.dwg",
    scale=(1, 1, 1),
    rotation=0,
    wait=0.3,
):
    pythoncom.CoInitialize()
    try:
        node("线程1启动")
        li()

        # —— 真正的插入操作 ——
        result = insert_company_label_common_block(
            insertion_point=insertion_point,
            filepath=filepath,
            scale=scale,
            rotation=rotation,
            wait=wait,
        )
        node("线程1完成插入, 写入 result_box")
        result_box["data"] = result          # ★ 写回共享字典

        timeout_event.wait()                 # 等线程2“放行”
    except Exception as e:
        print("f1_insert_company_getwindow:", e, traceback.format_exc())
    finally:
        pythoncom.CoUninitialize()
        done_event.set()

def Insert_Company_Label_Common_Block(
    insertion_point=(0, 0, 0),
    filepath=r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/自动化(动态)/中亿图签3.dwg",
    scale=(1, 1, 1),
    rotation=0,
    wait=0.3,
    *,
    timeout_sec: int = 300,
):
    result_box = {}   # ★ 共享容器

    ok = run_dual_threads(
        f1=f1_insert_company_getwindow,
        f2=f2_delwindow,
        f1_kwargs={
            "result_box": result_box,              # ★ 传给线程1
            "insertion_point": insertion_point,
            "filepath": filepath,
            "scale": scale,
            "rotation": rotation,
            "wait": wait,
        },
        timeout_sec=timeout_sec,
    )

    # —— 线程结束后取回 ——
    if ok and "data" in result_box:
        node("高层拿到返回值")
        coms_tuqian, names_tuqian, xinxi_tuqiankuai = result_box["data"]
        print(f"🎉 成功插入 → {insertion_point}")
        return coms_tuqian, names_tuqian, xinxi_tuqiankuai
    else:
        print("🚨 插入图签模板失败")
        return None, None, None


#&&&% （五） 插入图签到打印区域

"""
zd_ctq=insert_and_scale_labels_area(
        zd['dyx_coms'],
        filepath=r"D:/Myprogramsystem/BaiduSyncdisk/基础服务/宋岳/01插图签/05基本图签/属性图签插图签样板文件1基本图签.dwg",
        layername = "dy_quyu",
        timestamp=None,
        delpan = 0

    )

"""
#&&% ⮞ F4: 插入图签

@timeit             
@debuggable         
def insert_and_scale_labels_area(
        coms_dayin,
        filepath=r"D:/Myprogramsystem/XT/标准图签模板.dwg",
        layername = "dy_quyu",

        timestamp=None,
        delpan = 0

    ):
    """
    D:/Myprogramsystem/XT/重要测试/图签测试1.dwg


    把公司公共图签块批量插入到 coms_dayin 定义的打印框区域，
    自动按比例缩放并建立绑定字典；全过程带 node() 调试输出。
    返回 bind_dict：{新块Handle: {"frame_info": ..., "title_block": COM}, ...}
    """
    def filter_coms_with_frame(coms_dayin):
       
        LBx = []
        LBy = []
        for ent in coms_dayin:
            handle = getattr(ent, "Handle", None)
            try:
                info = generate_name_and_ratio_from_com(ent)
            except Exception:
                # 调用失败则跳过
                continue
    
            if info != 0:
                LBx.append(ent)
            else:
                LBy.append(handle)
    
        return LBx, LBy

    #插图签之前，清空"dy_quyu"的图签块，用于第二次以上的重复操作  
    ensure_layer("dy_quyu")

    # ▶ 1 打印框排序
    node("▶ 1  打印框排序, 原数量={}", len(coms_dayin))
    coms_dayin = sort_coms_by_llcorner(coms_dayin, cha_Y=2000)

    coms_dayin,LBy = filter_coms_with_frame(coms_dayin)
    node("▶ 1.1  非标准打印区域handle{}", LBy)

    # ▶ 2 重绘打印框并提取信息
    node("▶ 2  绘制打印框并提取信息")
    ensure_layer("dy_quyu")
    res = draw_pl_and_extract_from_entities(
        coms_dayin,
        layer_name="dy_quyu",
        width=0,
        color=256,
        A3dy=0,
        Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3", 0),
    )
    node("    提取完毕, 共 {} 条", len(res))

    # ▶ 3 原对象移层
    for ent in coms_dayin:
        try: ent.Layer = "测试辅助"
        except: pass
    node("▶ 3  原打印框移至 '测试辅助'")

    # ▶ 4 调用公共块插入
    node("▶ 4  调用 Insert_Company_Label_Common_Block()")
    ret = Insert_Company_Label_Common_Block(filepath=filepath)
    if not ret or len(ret) < 3:
        raise RuntimeError("Insert_Company_Label_Common_Block 返回格式异常")
    ji_yuan, _, result_dict = ret
    node("    插入公共块成功, 新块数={}", len(ji_yuan))#文件不同，可能导致插入块文件是一个整体块或不是

    # ❶ ★ 关键：线程结束后重新获取 acad / doc / mp ---------------------
    def _refresh_cad_globals():
        global acad, doc, mp
        acad = win32com.client.Dispatch("AutoCAD.Application")
        doc  = acad.ActiveDocument
        mp   = doc.ModelSpace
    _refresh_cad_globals()

    # ❷ ★ 安全发送命令 ---------------------------------------------------
    def safe_send(cmd):
        try:
            doc.SendCommand(cmd)
        except pythoncom.com_error:
            # 重新刷新一次再重试
            _refresh_cad_globals()
            doc.SendCommand(cmd)

    # 刷新视图
    safe_send("RE\n"); safe_send("Z\nE\n")
    time.sleep(0.2)

    # ————— 下面逻辑保持不变 ……………—

    # ▶ 5 Handle → idx 映射
    ent2idx = {info["entity"].Handle: idx for idx, info in res.items()}
    node("▶ 5  生成映射 ent2idx, 共 {} 条", len(ent2idx))


    # ▶ 6 计算比例 k 并插块
    entities = [info["entity"] for info in res.values()]
    entities = sort_coms_by_llcorner(entities, cha_Y=2000)
    factors  = compute_insert_factors(entities, res, result_dict)

    bind_dict = {}
    for seq, ob in enumerate(entities, 1):
        idx = ent2idx.get(ob.Handle)
        if idx is None:
            node("⚠  区域 {} 未找到索引, 跳过", seq)
            continue

        _, blk_name, spec, k_val = get_factor_for_entity(ob, factors)
        node("▶  区域{}  spec={}  k={}", seq, spec, k_val)

        success = False
        for attempt in range(1, 6):
            try:
                _, _, blk_com = insert_block_into_poly_area(
                    blk_name, ob, k=k_val, max_retries=3
                )
                bind_dict[blk_com.Handle] = {
                    "frame_info": res[idx],
                    "title_block": blk_com,
                }
                node("    ✅ 插入成功, handle={}", blk_com.Handle)
                success = True
                break
            except Exception as e:
                node("    ⚠ 第{}次失败: {}", attempt, e)
                time.sleep(0.2)
        if not success:
            node("    ❌ 区域{} 放弃", seq)

    # ▶ 7 公共块移层
    for br in ji_yuan:
        try: br.Layer = "测试辅助"
        except: pass
    ensure_layer("测试辅助")
    node("▶ 7  清理 '测试辅助'")

    ##删除插入基本图签,它们应该放在"标准图签"图层

    time.sleep(2)

    ensure_layer("标准图签")

    time.sleep(2)

    node("▶ 8  delpan的值{}",delpan)  
    rename_and_delete_poly(layername = layername,timestamp=timestamp,delpan = delpan)        



    # ▶ 8 显示所有的图签属性，属性值就是标签名

    LK=stc("dy_quyu")

    LP = [ent for ent in LK if ent.ObjectName in ("AcDbPolyline","AcDb2dPolyline")]
    LB = [ent for ent in LK if ent.ObjectName in ("AcDbBlockReference")]
    
    LP = sort_coms_by_llcorner(LP, cha_Y=2000)

    LB = sort_coms_by_llcorner(LB, cha_Y=2000)

    fill_block_attributes_with_tag_name(LB)

    bind_dict["dyx_list"] =LP

    bind_dict["tq_list"] =LB


    # ▶ 9 清理未使用的标准图签块

    # 1) 获取所有块定义名
    lk = get_all_block_definitions()
    LBname = [blk.Name for blk in lk]
    
    # 2) 筛出“零实例”块
    LBname_wushili = [
        name for name in LBname
        if len(select_block_by_name(name)) == 0
    ]
    
    # 3) 只保留前两位是 A+数字 的
    pattern = re.compile(r'^A\d')  # ^A\d 匹配以 A 开头、第二位是数字
    LBname_to_purge = [
        name for name in LBname_wushili
        if pattern.match(name)
    ]
    
    # 4) 批量清理
    for name in LBname_to_purge:
        purge_block(name)

    # ▶ 10 保存
    try: savefile()
    except: pass
    name_zd = current_dwg_basename() + "_tuqian"
    save_print_dict_generic(name_zd, bind_dict)
    node("▶ 9  结果已保存至 '{}'", name_zd)

    return bind_dict



#清除标准图签块内'tuqian_neibu_pl上的多段线

def clean_internal_polylines(LB_bk):
    """
    删除列表 LB_bk 中每个块定义（Block）内部、图层为 'tuqian_neibu_pl' 的多段线。

    参数
    ----
    LB_bk : list[COMObject]
        块引用（InsertBlock）对象列表

    返回
    ----
    deleted_handles : list[str]
        被删除的多段线实体的 Handle 列表
    """
    deleted_handles = []
    for blk_ref in LB_bk:
        # 1) 获取块的定义名
        try:
            blk_name = blk_ref.Name
        except Exception:
            continue

        # 2) 在文档块表中取出块定义
        try:
            blk_def = doc.Blocks.Item(blk_name)
        except Exception:
            continue

        # 3) 快照所有实体，避免删除时集合变更
        ents = [blk_def.Item(i) for i in range(blk_def.Count)]

        # 4) 遍历并删除匹配的多段线
        for ent in ents:
            if ent.Layer == "tuqian_neibu_pl" and ent.ObjectName in ("AcDb2dPolyline", "AcDbPolyline"):
                deleted_handles.append(ent.Handle)
                safe_delete(ent)


    time.sleep(1)

    doc.SendCommand("re\n")

    return deleted_handles


####

def rename_and_delete_poly(layername="dy_quyu", timestamp=None, delpan=False):
    """
    从“打印框线”图层中获取所有块引用（BlockReference），
    将它们所引用的块定义重命名，原名后追加精确到分钟的时间戳；
    （可选）删除这些块内部图层为 'tuqian_neibu_pl' 的多段线。

    参数
    ----
    layername : str
        要扫描的图层名
    timestamp : str, optional
        时间戳，格式 'YYYYMMDDHHMM'。默认使用当前日期时间到分钟。
    delpan    : bool
        是否调用 clean_internal_polylines 删除内部多段线，默认 False。

    返回
    ----
    LB_bk : list[COMObject]
        图层上所有 BlockReference 引用对象列表
    """
    from datetime import datetime


    # 1. 准备时间戳，精确到分钟
    if timestamp is None:
        timestamp = datetime.now().strftime("%Y%m%d%H%M")

    # 2. 选取指定图层上的所有对象
    LBdy = stc(layername)

    # 3. 过滤出块引用
    LB_bk = [ent for ent in LBdy if ent.ObjectName == "AcDbBlockReference"]

    # 4. 收集所有不同的块定义名
    old_names = {blk.Name for blk in LB_bk}

    # 5. 遍历块定义表，按块定义重命名
    for old in old_names:
        new_name = f"{old}_{timestamp}"
        try:
            blk_def = doc.Blocks.Item(old)
            blk_def.Name = new_name
        except Exception as e:
            print(f"⚠ 无法重命名块定义 '{old}' → '{new_name}'：{e}")

    # 6. （可选）删除内部多段线
    if delpan:
        try:
            clean_internal_polylines(LB_bk)
        except Exception:
            try:
                clean_internal_polylines(LB_bk)
            except:
                pass

    return LB_bk



def delete_blocks_in_layer(layer_name="dy_quyu", max_attempts=3):
    """
    删除指定图层中的所有块实例（BlockReference）。
    最多尝试 max_attempts 次，失败对象跳过，避免引用已删除的对象。
    """
    try:
        ents = stc(layer_name)
        print(f"✅ 第 1 次尝试：成功选择图层 “{layer_name}” 上的 {len(ents)} 个对象")
    except Exception as e:
        print(f"❌ stc 选择失败: {e}")
        return

    # 复制列表，防止删除后原始引用失效
    blocks = [e for e in ents if e.ObjectName == "AcDbBlockReference"]
    print(f"📦 选中块实例数量: {len(blocks)}")

    for blk in blocks:
        safe_delete(blk)

    print("🧹 删除完成。")







def fill_block_attributes_with_tag_name(blocks):
    """
    将所有块中的属性值设置为其标签名本身（用于标注和检查），
    例如标签为 "图纸编号"，值也设为 "图纸编号"。
    
    参数：
        blocks (list): 包含块参照（BlockReference）的实体列表
    """
    count_total = 0
    count_success = 0

    for blk in blocks:
        if blk.ObjectName != "AcDbBlockReference":
            continue

        try:
            if not blk.HasAttributes:
                continue
            count_total += 1

            for att in blk.GetAttributes():
                tag = att.TagString
                att.TextString = tag  # 设置属性值为标签本身
            count_success += 1

        except Exception as e:
            print(f"⚠ 块 {getattr(blk, 'Handle', '?')} 属性设置失败: {e}")

    print(f"✅ 已处理块 {count_success}/{count_total} 个，属性值已替换为标签名。")



def move_entities_to_layer(layername1, layername2, max_attempts=3):
    """
    将 layername1 图层中的所有实体转移到 layername2 图层。
    若 layername2 不存在，则创建之。
    失败时等待 1 秒重试，最多尝试 max_attempts 次。
    """
    layers = doc.Layers

    # 1. 确保目标图层存在
    try:
        target_layer = layers.Item(layername2)
    except:
        try:
            target_layer = layers.Add(layername2)
            print(f"✅ 已创建目标图层：{layername2}")
        except Exception as e:
            print(f"❌ 创建图层失败：{e}")
            return

    # 2. 获取源图层实体
    try:
        ents = [ent for ent in doc.ModelSpace if ent.Layer == layername1]
        print(f"🎯 共找到 {len(ents)} 个实体在图层 '{layername1}' 上。")
    except Exception as e:
        print(f"❌ 获取图层实体失败：{e}")
        return

    # 3. 尝试逐个转移图层
    for ent in ents:
        for attempt in range(1, max_attempts + 1):
            try:
                ent.Layer = layername2
                break  # 成功则跳出 retry 循环
            except Exception as e:
                print(f"⚠ 实体 {getattr(ent, 'Handle', '?')} 第 {attempt} 次转移失败: {e}")
                if attempt < max_attempts:
                    time.sleep(1)
                else:
                    print(f"❌ 放弃实体 {getattr(ent, 'Handle', '?')} 的转移")

    print(f"✅ 所有实体尝试转移至图层 '{layername2}' 完成。")





#&&&% （六） 从边域获取图纸名称


#&&% ⮞ F5:从边域获取图纸名称和图纸编号信息

"""
这个函数直接将图纸名称和图纸编号信息放入了字典zd

如果图纸名称和编号文字放入了指定图层，就更快获取结果，将该函数优化，不需要另外写一个函数

zd = process_side_regions_info(zd, cha_Y=100, cha_X=100)

"""

#&&% ⮞ F6:从excel表格获取图签块标签信息


"""
还需要整前面流程的zd和zd_excel定义新函数 def 

zd_excel = read_fixed_excel(path =r"D:/Myprogramsystem/BaiduSyncdisk/基础服务/宋岳/04Excel控制文件/不控制图纸名称/属性图签插图签样板文件1.xlsx" )


"""

#&&&% （七）完成图签写入

"""
图签写入首先分普通图签和属性标准图签。在process_side_regions_info(zd, cha_Y=100, cha_X=100)中，普通图签的图纸编号已经更新，则意味着图签写入已经完成。可以返回输出。

对属性标准图签，它总是有值的。只是修正的问题。提供两种模式。Excel优先。只要Excel控制文件夹有文件，就按文件来。没有，则只能在dwg中处理。

对图签基本信息，我们约定写在第一个图签。从第一个图签读到的基本信息标签的标签值就是图签基本信息。图纸编号和图纸名称按前面流程逻辑处理。

图签写入完成之后，函数将文件送往输出接口或下一个流程。

我们还需要一个全自动处理函数。它不需要按照从开始一个文件输入到最后文件结束的模式。我们是在一个大系统中处理。但它应该把插图签从文件进入到写好图签完成。

还需要一个处理某个单独标签值的函数。可能是针对某个图纸名或基础信息标签。

考虑到实际需要修改某个标签，以及可能存在我们的系统未考虑到的标签，需要这样一个单独处理的 命令。它可以在Excel控制文件夹增加一个文件夹和文件。应该比改造原来的函数更方便。


针对整个文件夹的某个函数的批量处理，应该写成函数的函数

写图签函数改成以zd为输入变量




"""






#&&% ⮞ F7:写图签


def write_title_blocks(
    LB_bkcom,
    LC, LC_th, LC_ratio,
    *,
    constant_unit   = "麻阳苗族自治县工业园投资开发有限公司",
    constant_project= "怀化市麻阳苗族自治县省级工业集中区标准化厂房及配套基础设施建设项目",
    tags           = ('建设单位', '工程名称', '图纸名称', '图纸编号',
                      '65', '图纸时间', '设计号', '3'),
    cha_Y          = 2000,
):
    """
    按 LC 顺序对块列表 LB_bkcom 逐一写入属性：

        new_values = [
            constant_unit,                  # [0] 建设单位（固定）
            constant_project,               # [1] 工程名称（固定）
            LC[i],                          # [2] 图纸名称
            LC_th[i],                       # [3] 图纸编号
            len(LC),                        # [4] 65  (总张数)
            LC_ratio[i],                    # [5] 图纸时间 (比例)
            "",                             # [6] 设计号 (留空)
            str(i+1)                          # [7] 3  (当前张次)
        ]

    参数
    ----
    LB_bkcom   : list[AcadBlockReference]  # 待写入的块参照列表
    LC         : list[str]                 # 图纸名称
    LC_th      : list[str]                 # 图纸编号
    LC_ratio   : list[str]                 # 比例
    constant_* : str                       # 固定两项
    tags       : tuple[str]                # 8 个属性 Tag 顺序
    cha_Y      : float                     # 排序用容差
    """

    # ——— 1. 先按左下角排序，确保块顺序与 LC 对齐 ———
    LB_bkcom = sort_coms_by_llcorner(LB_bkcom, cha_Y=cha_Y)

    # ——— 2. 批量写属性 ———
    total = len(LC)
    for i, blk in enumerate(LB_bkcom):
        if i >= total:
            print(f"⚠ 块序号 {i} 超出 LC 长度，停止写入")
            break

        # 构造 new_values
        new_values = [
            constant_unit,
            constant_project,
            LC[i],
            LC_th[i],
            str(total),
            LC_ratio[i],
            "",
            str(i+1)
        ]

        # 写入块属性
        if getattr(blk, "ObjectName", "") != "AcDbBlockReference":
            print(f"⚠ 对象 {blk.Handle} 不是块参照，跳过")
            continue
        if not getattr(blk, "HasAttributes", False):
            print(f"⚠ 块 {blk.Handle} 无属性，跳过")
            continue

        try:
            for attr in blk.GetAttributes():
                tag = attr.TagString.strip()
                if tag in tags:
                    idx = tags.index(tag)
                    attr.TextString = new_values[idx]
        except Exception as e:
            print(f"⚠ 块 {blk.Handle} 写属性失败: {e}")

    print(f"✅ 已写入 {min(total, len(LB_bkcom))} 个图签属性")







@timeit             
@debuggable       
def 图签处理完整流程(filepath=r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/工业园整理/标准图签模板-JG.dwg", j=0, fujia="# 1", prefix="JZ-", zz="张", height=200.0, width=4500.0):
    li()
    enable_debug()

    # 节点1 获取打印线、图名信息
    try:
        zd = process_print_areas_info()
        zd1 = process_side_regions_info(zd, cha_Y=100)
        node("✅ 获取打印区域信息及图纸名称成功")
    except Exception as e:
        node("⚠ 出错: {}", e)

    # 检查 CAD 是否准备好
    check_cad_ready(retries=12, wait_time=3)

    # 节点2 插入图签模板
    try:
        insert_and_scale_labels_area(
            zd["dycoms"],
            filepath=filepath,
            layername="dy_quyu",
            timestamp=None,
            delpan=1
        )
        node("✅ 插入图签模板成功")
    except Exception as e:
        node("⚠ 出错: {}", e)

    # 检查 CAD 是否准备好
    check_cad_ready(retries=12, wait_time=3)

    # 节点3 检测模板插入是否成功
    node("🔍 开始检测模板插入情况")
    
    LB_BK = stc("dy_quyu")
    
    # 过滤出块实例
    LB_BK = [blk for blk in LB_BK if blk.ObjectName == "AcDbBlockReference"]
    
    retry_count = 0
    while len(LB_BK) == 0 and retry_count < 3:
        node(f"⚠ 无有效块实例，重新尝试插入图签模板，第 {retry_count + 1} 次重试")
        time.sleep(1)  # 等待 1 秒
        insert_and_scale_labels_area(
            zd["dycoms"],
            filepath=filepath,
            layername="dy_quyu",
            timestamp=None,
            delpan=1
        )
        
        LB_BK = stc("dy_quyu")
        LB_BK = [blk for blk in LB_BK if blk.ObjectName == "AcDbBlockReference"]
        
        retry_count += 1
    
    if len(LB_BK) == 0:
        node("❌ 尝试了 3 次仍然没有有效的块实例，图签模板插入失败，流程结束")
        return False  # 如果 3 次尝试后依然没有有效块实例，结束流程
    
    node(f"✅ 检测到 {len(LB_BK)} 个有效块实例，继续处理")

    # 节点4 生成图名等信息并写入图签    
    generate_all_names(zd1, j=j, fujia=fujia, prefix=prefix, zz=zz)

    node("🔍 LC,{} LC_th,{} LC_gg,{} LC_tf,{} LC_ratio,{} LC_zz{}", LC, LC_th, LC_gg, LC_tf, LC_ratio, LC_zz)

    # 获取去重后的块列表
    BK = get_dedup_blocks_on_layer("dy_quyu")
    
    # 执行图签数据写入
    write_title_blocks(
        BK,
        LC, LC_th, LC_ratio,
        constant_unit="麻阳苗族自治县工业园投资开发有限公司",
        constant_project="怀化市麻阳苗族自治县省级工业集中区标准化厂房及配套基础设施建设项目",
        tags=('建设单位', '工程名称', '图纸名称', '图纸编号', '65', '图纸时间', '设计号', '3'),
        cha_Y=2000
    )
    node("✅ 图签信息写入成功")

    # 节点5 修改标签字高
    node("🔍 开始修改标签字高")
    for obj in BK:
        resize_block_attribute(obj, "工程名称", height=height, width=width)
    node("✅ 标签字高修改成功")

    # 节点6 删除旧图签图名文字实体
    node("🔍 开始删除旧图签和图名")
    delete_old_title_and_frame(zd, zd1)
    node("✅ 旧图签图名删除成功")

    disable_debug()

    return zd,zd1



#&&% 改属性块某个标签值

def rename_block(bq=["工程名称"],new_zhi=["怀化市麻阳苗族自治县省级工业集中区标准化厂房及配套基础设施建设项目(三期)"]):

    BK=get_dedup_blocks_on_layer("dy_quyu")

    for obj in BK:

        set_attributes_values(obj, bq, new_zhi)


@timeit             
@debuggable        
def rename_block_all_files_in_folder(
    folder_path=r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/工业园整理/三期/CAD打印",
    bq=["工程名称"],
    new_zhi=["怀化市麻阳苗族自治县省级工业集中区标准化厂房及配套基础设施建设项目(三期)"],
    cha_Y=2000,
    tol=50,
    max_attempts: int = 5,
):


    """
    遍历文件夹 folder_path 中的每个 DWG 文件，操作rename_block，给属性标签改名

    """
    

    minimize_all_windows_d()

    # 恢复并定位 CAD 窗口
    restore_and_position(
        name="AutoCAD",
        width_ratio=1.0,
        height_ratio=1.0,
        x=0,
        y=0
    )
    time.sleep(1)

    enable_debug()
    li()
    
    # 获取文件夹及其子文件夹中的所有 dwg 文件
    dwg_files = []
    for root, _, files in os.walk(folder_path):
        # 在每个子文件夹中查找 .dwg 文件
        for file in files:
            if file.lower().endswith(".dwg"):
                dwg_files.append(os.path.join(root, file))


    # 确保当前文件（00.dwg）是已打开文件
    current_file = "D:/Myprogramsystem/XT/00.dwg"
   
    # 遍历文件夹内的每个 dwg 文件处理
    for dwg_file in dwg_files:
        dwg_filepath = os.path.join(folder_path, dwg_file)
               
        for attempt in range(1, max_attempts + 1):

            try:
                node("▷ [{}] 处理文件：{}", attempt, dwg_filepath)



                # 确保文件 00.dwg 是当前唯一激活的文件才打开目标操作文件，当try出错时ensure_active_cad_file(current_file)会被略过所以实测又打开了新文件

                if not keep_only_doc(r"D:/Myprogramsystem/XT/00.dwg"):

                    node("未能保持00.dwg是唯一激活文件状态，再次尝试") 

                    continue               

                doc = Open_By_Omission_wenjian(dwg_filepath)
        
                check_cad_ready(retries=12, wait_time=3)
                
                
                # 执行 rename_block
                rename_block(bq=bq,new_zhi=new_zhi)
        
                doc.SendCommand("Z\nE\n")
        
                time.sleep(1)
       
                # 处理完后关闭当前文件（确保关闭）

                # 尝试关闭当前文档，最多 3 次
                closed_ok   = False   # 记录是否真正调用过 doc.Close 并成功
                skip_closed = False   # 记录是否发现已经没有打开文档
                
                for _ in range(3):
                    try:
                        li()  # 刷新 COM 对象
                        if acad.Documents.Count==2:          # 还有2文档 → 尝试关闭dwg_filepath
                            doc.Close()
                            closed_ok = True
                        else:                             # 已无2个文档 → 视为成功，退出循环
                            skip_closed = True
                        break
                    except Exception:
                        time.sleep(0.5)                   # 等 500 ms 再重试
                else:
                    # 只有三次全部 raise 才会走到这里
                    node("[关键节点 2] ❌ 关闭文件 {} 失败 3 次，不再尝试", dwg_filepath)
                
                # 根据执行结果输出对应日志
                if closed_ok:
                    node("[关键节点 2] ✅ 已完成处理并关闭文件 {}", dwg_filepath)
                elif skip_closed:
                    node("[关键节点 2] ℹ️ AutoCAD 当前无打开文档，跳过关闭")




                # 确保文件 00.dwg 是当前唯一激活的文件
        
                li()#连接当前激活文件
                ensure_active_cad_file(current_file)
                      
                # 本次文件处理成功，跳出重试循环
                break    

            except Exception as e:
                node("⚠ [{}] 处理失败: {}", attempt, e)
                time.sleep(1)
                if attempt == max_attempts:
                    node("❌ [{}] 超过最大重试次数({})，跳过文件 {}", attempt, max_attempts, dwg_filepath)
        # end for attempts


    node("✅ 所有文件改名完毕 ")
    
    disable_debug()

    return

























#&&% ***🧠  CAD属性块的重要处理

"""
20250704 重新测试插图签 注意，标准图签模板必须放在“标准图签”图层上，用于集中销毁多余基本图签

我们给用户插入了万能图签，用户需自己改变图签中文字的位置，用户要记住改变的那个个体，编辑完必须在命令行输入 ATTSYNC 特别注意，是使用BE进行编辑，在块里面进行拷贝、移动、删除之后再在CAD命令窗口输入ATTSYNC 选择编辑改动的那个实例块 回车

提示选择块实例对象或块名 提示所有同步还是非所有 回车则所有位置改变 如果某个标签是多余的，删除它，通过ATTSYNC也能让其余块同步删除

每个标签值的改变还是word和边侧放置文字 今天就是重新测定了插入万能图签，明确了调整位置的方法 注意到该命令可能需要重新尝试多次 下一步就是将插图签流程完善


没有绝对的成功 如果尝试3次还不成功 那就要单独处理这个文件和程序了

#清空  
#矛盾在于第一次用户可能要用打印框线作标记 但第二次用户可能要删除之前的 打印框线的对象 所以最后返给用户的对象不能放在这个过程图层"打印框线"上

用来插图签的打印区域线，不要放入 打印框线 

"""







#*************************************************************************************************************************************************************






#&&&% 🧱  服务二 编目录完整流程
#&&%─────────────────────────────────────

"""
实际上我们需要一个后处理流程，针对重复的框线，隐藏的图层，保护的对象等

就编目录这个流程，我们不考虑别的附带处理 我们的核心功能就是 选取打印区域 文字边域 获取图名 获取编号 允许普通图签块 
"打印框线"图层是用于外界交流，决定非标准打印区域仍然按A3打印的设置。所以要修改之前的函数，不要使用这个名称

不允许混用有框图签和无框图签
有框图签只能是属性图签

整个体系的架构，应该按人来 ，因为每个人有自己负责的字体，打印样式等

应该将打印和基础信息分开  应该在获取边域文字这里重写序号或者写好图名序号 再另外针对属性有框图签完善图签内容即写完图签  编目录初始化 编目录



如果输入的文件是走完“插图签流程”的属性标准图签文件，那将是干净的文件，在dwg中不能获取任何信息，重复一遍process_side_regions_info就是空转

如果是采用插图签辅助模式获取图签块然后再放置图纸名称和编号信息的，正好在编目录流程的这个process_side_regions_info处理图纸名称和编号及写图签

如果是普通图签，已经处理好的图纸编号文字和信息，等于是多执行了一遍  所以编目录流程 需要设计一个不对外开放的  输入文件夹接口 但是这种改变代价太大意义不大 因此就让它重复损耗


对于Excel控制文件来说，在“插图签流程”的写图签这里重新修改即可 在“编目录流程”这里也是一样 不用增加函数，在原来函数的最后面加上一部分依据Excel文件重新修改内容即可



"""



#&&&% （一） 获取打印范围及基础信息

#&&%─────────────────────────────────────
"""
这个函数要处理无框图签和有框图签两种情况

默认的参数可能会失效，但在绝大多数情况下是有效的 。首先，我们针对无框图签来说，我们要求它是一个块。其次，我们要求它的长宽比是5以上。再者，我们要求它的右下角点和整个打印区域右下角点距离在3000之类。
在某些特殊情况下，仍然可能选择到不真实的打印框线。对这种极特殊的文件可以单独处理。以后再完善这个函数 

有框图签反而更容易识别

本函数的主要任务是：确定打印区域，确定图签和图签边域，确定目录模板信息
考虑到本函数的复杂和重要，前面加了li()连接

采用字典作为函数返回值，方便后续的极大扩展和体系的光滑连接

在本函数中清理整理多次使用造成的诸多问题


将enable_debug() disable_debug() 嵌入函数中直接控制函数运行的显示，并且对显示的节点重新定义


"""


@timeit            
@debuggable        
def process_print_areas_info(
    tol: float  = 20.0,
    tola: float = 100.0,
    aspect_tol: float = 4,
    area_min: float = 160000000,
    area_max: float = 1000000000,
    tol2: float = 3000.0,
    cha_Y: float = 2000.0

):
    """
    1. 使用 select_polyline_chuantong(5)、select_polyline(5)、select_kuai(5)
       快速选择所有多段线和块到列表 L1，并打印其长度。
    2. 对 L1 中每个对象调用 
       zhi = generate_name_and_ratio_from_polyline(
               comobj,
               A3dy=0,
               Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3",0)
           )
       如果 zhi == 0 跳过，否则将该对象放入 L3，并记录 zhi 到 zhi_dict：
         zhi_dict[handle] = zhi
    3. 对 L3 先按外包盒位置去重得到 L3_unique。
    4. 获取当前时间 timename="YYYY-MM-DD-HH-MM"。
       - 先对 L3_unique 使用 sort_coms_by_llcorner 排序，得到 L3_sorted。
       - 再从 L3_sorted 生成标准打印框线：对每个对象按顺时针提取外包盒四角，draw_lwpolyline(...)，
         并把返回的新多段线实体放入 LP_1_raw。
       - 对 LP_1_raw 使用 sort_coms_by_llcorner 排序得到 LP_1_sorted。
       - 由于 LP_1_sorted 与 L3_sorted 在排序逻辑上是一一对应的，因此可以
         按索引把 LP_1_sorted[i].Handle 映射到 L3_sorted[i] 的 zhi。
       - 打印 LP_1_sorted 的长度和对应的 Handle 列表。
    5. 从 L1 中筛选出块到 T1，再调用 
       filter_blocks_by_aspect_and_area(T1, aspect_tol, area_min, area_max) 得到 T1_filtered。
       再调用 dedupe_by_area(T1_filtered, tol) 得到 T2。对 T2 排序并打印 Handle 列表。T指无框图签块
    6. 从 L3_unique 中筛选出块到 F1，再调用 dedupe_by_area(F1, tol) 得到 F2_candidates，
       再调用 get_filtered_entities_from_blocks(F2_candidates, aspect_tol, area_min, area_max) 得到 F2。对 F2 排序并打印 Handle 列表。F是有框图签块
    7. 返回一个字典 result_dict，包含：

         - aspect_tol      图签区域的比例特征系数，默认为5 
         - area_min        图签区域的默认面积特征
         - "prinfo":       字典，键为 LP_1_sorted 中多段线的 Handle，值为对应的 L3_sorted 对象的 zhi（四元组）
         -"tol": = 20      两点距离容差
         -"tol2": = 3000   图框右下角角点距离特征 有的人并不按天正标准1000
         -"tola": = 100    面积容差
         -"cha_Y": =2000   打印线布置高度方向的间距容差，小于此值认为相等，按x坐标排序


         - 曾经的思想是用户决定是否按比例决定打印区域的选择，实际上应该改为正常找不到时自动启动按比例寻找
         - 为了更强的适应性、免得不可预估的复杂情况临阵混乱来不及分析，函数应提供默认为None的选择对象列表的方式 

    """

    # 0. 循环使用的处理，解冻之前打印时隐藏的文字分行分隔线
    li()
    #重复操作产生大量的多段线，需要及时处理
    Lb_qingchang = stc("dy_quyu")
    dycoms_qingchang = [ent for ent in Lb_qingchang if ent.ObjectName in ("AcDbPolyline", "AcDb2dPolyline")]
    for ob in dycoms_qingchang:
        safe_delete(ob)

    thaw_and_move_from_fengexian()#以计算目录行数
    enable_debug()


    delete_standard_print_frames()#不会删除“dy_quyu”图层

    # 1. 选择所有多段线和块到 L1,排除"print_tuqian_jiben"
    L1_zhun = []
    try:
        L1_zhun.extend(select_polyline_chuantong(5))
    except Exception as e:
        print(f"❌ select_polyline_chuantong 失败: {e}")
    try:
        L1_zhun.extend(select_polyline(5))
    except Exception as e:
        print(f"❌ select_polyline 失败: {e}")
    try:
        L1_zhun.extend(select_kuai(5))
    except Exception as e:
        print(f"❌ select_kuai 失败: {e}")



    L1 = []

    for obj in L1_zhun:

        try:
                if obj.Layer == "print_tuqian_jiben":
        
                    pass
                else:
        
                    L1.append(obj)
        
        except Exception as e:
            print(f"获取对象图层属性失败: {e}")
    
    print(f"L1 总共选到的多段线和块对象数量: {len(L1)}")


    #只要在“打印框线”图层上的图形对象都是合法打印区域，但这个修改后续再做20250704


    node(" L1 总共选到的多段线和块对象数量: {}", len(L1))
    # 2. 分组到 L3，并记录 zhi 到 zhi_dict
    L3 = []
    zhi_dict = {}  # 原始 L3 对象的 Handle -> zhi（四元组）
    for comobj in L1:
        try:
            zhi = generate_name_and_ratio_from_com(
                comobj,
                A3dy=0,
                Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3",0)
            )
        except Exception:
            continue

        if zhi != 0:
            handle = getattr(comobj, "Handle", None)
            if handle:
                zhi_dict[handle] = zhi
            L3.append(comobj)
    node("初步选择的标准打印区域L3 : {}", len(L3))
    # 3. 对 L3 按外包盒位置去重
    L3_unique = dedupe_by_bbox_proximity(L3, tol=tol)

    # 4. 获取当前时间戳
    now = datetime.datetime.now()
    timename = now.strftime("%Y-%m-%d-%H-%M")

    # 辅助：从外包盒生成按顺时针排序的 4 个角点
    def _bbox_corners(obj):
        ll_point, ur_point = obj.GetBoundingBox()
        x1, y1, z1 = ll_point
        x2, y2, z2 = ur_point
        return [
            (x1, y1, z1),
            (x2, y1, z1),
            (x2, y2, z1),
            (x1, y2, z1)
        ]

    # 4a. 对 L3_unique 排序，得到 L3_sorted
    L3_sorted = sort_coms_by_llcorner(L3_unique, cha_Y=cha_Y)

    # 4b. 从 L3_sorted 逐条生成标准打印框线，多段线实体放入 LP_1_raw
    LP_1_raw = []
    for comobj in L3_sorted:
        try:
            coords3d = _bbox_corners(comobj)
            pl = draw_lwpolyline(
                coords3d,
                layer_name=f"标准打印框线{timename}",
                width=20.0,
                color=256,
                closed=True
            )
            LP_1_raw.append(pl)
        except Exception:
            continue

    # 4c. 对 LP_1_raw 排序，得到 LP_1_sorted
    LP_1_sorted = sort_coms_by_llcorner(LP_1_raw, cha_Y=cha_Y)
    handles_lp1 = [getattr(ent, "Handle", None) for ent in LP_1_sorted]
    print(f"LP_1 标准打印区域数量: {len(LP_1_sorted)}，Handle 列表: {handles_lp1}")
    node("[节点1] LP_1 标准打印区域数量 : {} handle{}", len(LP_1_sorted),handles_lp1)

    # 构建 Bbox_lf：按 LP_sorted 顺序记录左右下角 + 右上角
    Bbox_lf = [pl.GetBoundingBox() for pl in LP_1_sorted]

    # 4d. 建立 prinfo：LP_1_sorted[i] 对应 L3_sorted[i]
    prinfo = {}
    for idx, pl_ent in enumerate(LP_1_sorted):
        pl_handle = getattr(pl_ent, "Handle", None)
        # 对应的 L3_sorted 对象
        src_obj = L3_sorted[idx]
        src_handle = getattr(src_obj, "Handle", None)
        if pl_handle and src_handle and src_handle in zhi_dict:
            prinfo[pl_handle] = zhi_dict[src_handle]


    ##_______________________________________________无框图签的处理_____________________________________________##


    # 5.1 从 L1 中筛选“块”对象到 T1，再过滤并按面积去重得到 T2
    def _is_block(obj):
        return obj.ObjectName == 'AcDbBlockReference'

    T1 = [obj for obj in L1 if _is_block(obj)]

    handle_T1 =[ent.Handle for ent in T1]

    node("初选的块T1 数量: {}handle {} ", len(T1),handle_T1)

    # 按面积和高宽比特征压缩选择范围
    T1_filtered = filter_blocks_by_aspect_and_area(
        T1, aspect_tol, area_min=area_min, area_max=area_max
    )

    # 按面积去重
    T2 = dedupe_by_area(T1_filtered, tol=tola)
    node("按面积去重块T2a 数量: {} ", len(T2))

    # 按位置特征确保所选对象只能是无框图签块
    T2 = filter_coms_by_proximity_to_frames(T2, LP_1_sorted, tol=tol2) 
    node("按位置去重块T2b 数量: {} ", len(T2))
    # 排序
    T2_sorted = sort_coms_by_llcorner(T2, cha_Y=cha_Y)

    handles_t2 = [getattr(ent, "Handle", None) for ent in T2_sorted]
    print(f"T2 不包含图框线的基本图签块数量: {len(T2_sorted)}，Handle 列表: {handles_t2}")

    node("[节点2] 不包含图框线的基本图签块T2 数量: {} handle{}", len(T2_sorted), handles_t2)


    # 5.2 获取所有位置去重的块

    T2_suoyou = dedupe_by_bbox_proximity(T1, tol=tol)#按位置去重，不能按面积

    node("按位置去重的块: {} ", len(T2_suoyou))

    T2suoyou_filtered = filter_blocks_by_aspect_and_area(
        T2_suoyou, aspect_tol, area_min=area_min, area_max=area_max
    )#按面积和长宽比特征


    # 按右下角特征确保所选对象只能是无框图签块
    T2_suoyou = filter_coms_by_proximity_to_frames(T2suoyou_filtered, LP_1_sorted, tol=tol2) 

    node("按右下角位置特征筛选无框图签: {} ", len(T2_suoyou))
    # 排序
    T2_suoyou = sort_coms_by_llcorner(T2_suoyou, cha_Y=cha_Y)

    handles_t2_s = [getattr(ent, "Handle", None) for ent in T2_suoyou]
    print(f"T2_suoyou 不包含图框线的图签块数量: {len(T2_suoyou)}，Handle 列表: {handles_t2_s}")

    node("[节点3] 不包含图框线的图签块T2_suoyou 数量: {} handle{}", len(T2_suoyou), handles_t2_s)



    ##_______________________________________________有框图签的处理_____________________________________________##


    # 6.1 从 L1 中筛选“块”对象到 F1，再按面积去重筛选，并进一步按   块内有“无图框图签特征” 筛选得到 F2
    F1 = [obj for obj in L1 if _is_block(obj)]

    handles_f1=[ent.Handle for ent in F1]

    node("所有块F1 数量: {} handle{}", len(F1),handles_f1) 

    F2_candidates = dedupe_by_area(F1, tol=tol)

    handles_f2cand=[ent.Handle for ent in F2_candidates]

    node("F1按面积去重 数量: {} handle{}", len(F2_candidates),handles_f2cand) 
    F2 = get_filtered_entities_from_blocks(
        F2_candidates, aspect_tol=aspect_tol,
        area_min=area_min, area_max=area_max
    )
    F2_sorted = sort_coms_by_llcorner(F2, cha_Y=cha_Y)
    handles_f2 = [getattr(ent, "Handle", None) for ent in F2_sorted]
    print(f"F2 包含图框线的基本图签块数量: {len(F2_sorted)}，Handle 列表: {handles_f2}")

    node("[节点4] 有框的基本图签块F2 数量: {} handle{}", len(F2_sorted),handles_f2) 

   # 6.2 获取所有有框图签

    F2_suoyou = dedupe_by_bbox_proximity(F1, tol=tol)#按位置去重
    F2_suoyou = get_filtered_entities_from_blocks(
        F2_suoyou, aspect_tol=aspect_tol,
        area_min=area_min, area_max=area_max
    )
    F2_suoyou = sort_coms_by_llcorner(F2_suoyou, cha_Y=cha_Y)
    handles_f2_s = [getattr(ent, "Handle", None) for ent in F2_suoyou]
    print(f"F2_suoyou 包含图框线的图签块数量: {len(F2_suoyou)}，Handle 列表: {handles_f2_s}")
    node("[节点5] 有框图签块F2_suoyou 数量: {} handle{}", len(F2_suoyou),handles_f2_s)


    ##_______________________________________________flag处理_____________________________________________##


    # 7.1 将文件是采用无框图签还是有框图签的信息明确下来 -1----无框图签  1----有框图签  0----无图签

    flag_wukuangtuqian = -1

    if  len(F2_suoyou) > 0  and  len(T2_suoyou) == 0 :

        flag_wukuangtuqian = 1

    elif len(F2_suoyou) == 0  and  len(T2_suoyou) == 0:

        flag_wukuangtuqian = 0 #无图签

    else:

        pass

    # 7.2 将文件是采用属性图签还是普通图签的信息明确下来 0----普通图签  1----属性图签

    flag_shuxingtuqian = 0

    try:
    
        bq=huoqukuai_shuxing_zhi(F2_suoyou[0])[0]
    
    
        if  len(bq) > 0  :
    
            flag_shuxingtuqian = 1
    
        else:
            pass
    
    except:
    
        pass

    # 7.3 将文件有无图纸目录模板的信息明确下来 None----无图纸目录模板  i----有图纸目录模板且在第i张

    flag_mulumoban =None

    muluhangshu = None

    muluzhizhangshu = 0     


    try:
    
        flag_mulumoban,muluhangshu,muluzhizhangshu = find_directory_frame(LP_1_sorted)

 
    except:

        pass

    node("[节点6]目录模板的序号数 {} ,目录行数 {}目录纸张数 {} ",flag_mulumoban,muluhangshu,muluzhizhangshu)

    node("flag[0]:-1即无框图签，1即有框图签，0为无图签；flag[1]:0即普通图签，1即属性图签；flag[2]:目录模板所在序号数，None即无目录模板")

    node("flag[3]:目录行数，无则为None；flag[4]:目录纸张数，None则未定")

   # 8 处理图签的边域

    T2_bianyu = get_bbox_corners_list(T2_suoyou)
    F2_bianyu = get_blocks_remaining_bbox(F2_suoyou)



    # 9. 返回结果字典
    result_dict = {
        "dyx_handles": handles_lp1,                                                         #所有打印线
        "wukuang_handles":    (handles_t2,handles_t2_s,T2_bianyu),                          #无框基本图签块、无框所有图签块、无框打印边域 
        "youkuang_handles":   (handles_f2,handles_f2_s,F2_bianyu),                          #有框基本图签块、有框所有图签块、有框打印边域

        "Bbox_lf": Bbox_lf,                                                                 #打印线外包盒对角点坐标
        "flag":[flag_wukuangtuqian,flag_shuxingtuqian,flag_mulumoban,muluhangshu,muluzhizhangshu],
        #无框图签还是有框图签还是没图签、是否属性图签、有无目录模板，有的话在第几张、目录行数，目录纸张数
        "dycoms":LP_1_sorted,#未关闭文件之前com对象是最方便的

        "wukuang_tq_coms":T2_suoyou,

        "youkuang_tq_coms":F2_suoyou,

        "prinfo":       prinfo
    }

    # 取消选择状态
    cancel_cad_selection()
    # 保存
    try: savefile()
    except: pass


    # 更新字典并保存到电脑本地

    
    try:
        dwg_name = current_dwg_basename()
        

    except Exception as e:

        time.sleep(1)
        dwg_name = current_dwg_basename()
        
        node("[ERROR] current_dwg_basename 失败，再尝试一次：{}",e)

    try:
        save_print_dict_generic(dwg_name, result_dict)
        node("[INFO] 已保存到字典 -> {}",dwg_name)
    except Exception as e:

        time.sleep(1)
        save_print_dict_generic(dwg_name, result_dict)
        node("[ERROR] save_print_dict_generic 失败：{}",e)

    disable_debug()

    return result_dict


def delete_standard_print_frames(timestamp: str = None,
                                 max_retries: int = 5,
                                 delay: float = 0.5,
                                 delete_retries: int = 5,
                                 delete_delay: float = 1.0):
    """
    删除图层名称以“标准打印框线”开头的对象及图层本身：
      - 如果 timestamp 不为 None，则只删除精确匹配 “标准打印框线{timestamp}” 的图层
      - 否则删除所有以 “标准打印框线” 开头的图层

    :param timestamp:      时间戳字符串，如 "2025-07-10-06-46"
    :param max_retries:    select_tuceng 最多重试次数
    :param delay:          select_tuceng 每次重试前等待秒数
    :param delete_retries: safe_delete 最多重试次数
    :param delete_delay:   safe_delete 每次重试前等待秒数
    """
    # 1. 收集目标图层名
    target_layers = []
    for lyr in doc.Layers:
        name = getattr(lyr, "Name", "")
        if name.startswith("标准打印框线"):
            if timestamp is None or name == f"标准打印框线{timestamp}":
                target_layers.append(name)

    if not target_layers:
        print(f"❌ 未找到任何‘标准打印框线{timestamp or ''}’图层")
        return

    # 2. 对每个图层依次选择、删除对象，然后删除图层
    for lyr_name in target_layers:
        print(f"🔍 处理图层 “{lyr_name}” …")
        objs = stc(lyr_name, max_retries=max_retries, delay=delay)
        if not objs:
            print(f"⚠ 图层 “{lyr_name}” 上没有可删除对象，直接尝试删除图层")
        else:
            deleted_count = 0
            for ent in objs:
                if safe_delete(ent, retries=delete_retries, delay=delete_delay):
                    deleted_count += 1
                else:
                    print(f"⚠ 删除对象 Handle={getattr(ent, 'Handle', '?')} 失败")
            print(f"✅ 已从图层 “{lyr_name}” 删除 {deleted_count}/{len(objs)} 个对象")

        # 等待一秒再删图层
        time.sleep(1)
        try:
            lyr_obj = doc.Layers.Item(lyr_name)
            lyr_obj.Delete()
            print(f"✅ 已删除图层 “{lyr_name}”")
        except Exception as e:
            print(f"⚠ 删除图层 “{lyr_name}” 失败：{e}")


#&&% 制作目录公共块流程 

"""

制作目录公共块的流程
1 内部分隔线""bianmulu_lp""和预设文字图层"ML-图纸序号"，"ML-图纸编号","ML-图纸规格","ML-图纸规格"设置要参考r"D:/Myprogramsystem/XT/mulugeshi_1.dwg"
2 先在文件1内部制作一个内部块将所有内容整合，移动块左下侧角点到0点
3 WB制作公共块，选择插入点仍然点击0点，对象选择刚才的块
4 保存文件命名好，制作完毕

这样做的目的是适应函数操作流程，确保按照预设点准确插入，并合乎块的炸开和不炸开规则
一个复杂的原因，不做内部块会造成插入后角度偏转，通过内部块配合控制修正这个旋转

"""

def insert_catalog_template(ins_pt=(0,0,0),
                         template_dwg =r"D:/Myprogramsystem/XT/mulugeshi_1.dwg",
                         
                         scale: tuple = (1, 1, 1),
                         rotation: float = 0,
                         wait: float = 0.3,
                         ):
    """
    从指定图层拆出图签块和打印框线，复制第一条打印框线，
    左移 offset 后取其最左下角作为插入点，插入并炸开外部 DWG。
    全程使用动态 Dispatch，避免 gencache 早绑定影响 TextString 等属性。

    :param ins_pt:       插入点
    :param template_dwg: 要插入的 DWG 路径
    :param offset:       复制框线后向左移动距离（X 负方向）
    :param scale:        插入时的缩放 (sx, sy, sz)
    :param rotation:     插入时的旋转角度（度）
    :param wait:         插入后等待秒数再炸开
    :param acad_instance: 已连接的 AutoCAD.Application（动态 Dispatch），
                          不传则内部动态 `Dispatch`
    """



    # ——  调用插入并炸开函数 ——  
    _,bk=insert_and_explode_dwg(template_dwg,
                           insertion_point=ins_pt,
                           scale=scale,
                           rotation=rotation,
                           wait=wait)
    time.sleep(1)

    LP=safe_explode_and_delete(bk)

    print(f"✅ 已在点 {ins_pt} 插入并炸开 '{template_dwg}'。")

    return LP


# 双线程插入目录模板

# ===============================================================
# 工具：检测并忽略“缺少 SHX 字体”对话框
# ===============================================================

def _find_shx_dialog(shx_titles=("缺少 SHX", "Missing SHX")) -> int:
    """返回对话框句柄；若不存在返回 0"""
    result = []
    def _enum(hwnd, res):
        title = win32gui.GetWindowText(hwnd)
        if any(k in title for k in shx_titles):
            res.append(hwnd)
    win32gui.EnumWindows(_enum, result)
    return result[0] if result else 0


def _ignore_shx_dialog(hwnd: int):
    """直接发送 Esc 关闭对话框，无需置前窗口"""
    win32gui.PostMessage(hwnd, win32con.WM_KEYDOWN, win32con.VK_ESCAPE, 0)
    win32gui.PostMessage(hwnd, win32con.WM_KEYUP,   win32con.VK_ESCAPE, 0xC0000000)

# ===============================================================
# 通用“双线程 GUI 调度器”
# ===============================================================

def run_dual_threads(
    f1, f2,
    f1_args=(), f1_kwargs=None,
    f2_args=(), f2_kwargs=None,
    *,
    timeout_sec: int = 180,
):
    """启动 f1 / f2 双线程，并在指定时限内协调退出"""
    f1_kwargs = f1_kwargs or {}
    f2_kwargs = f2_kwargs or {}

    timeout_event = threading.Event()
    done_event    = threading.Event()

    t1 = threading.Thread(target=f1, args=(timeout_event, done_event, *f1_args), kwargs=f1_kwargs, daemon=True)
    t2 = threading.Thread(target=f2, args=(timeout_event, done_event, *f2_args), kwargs=f2_kwargs, daemon=True)

    t1.start(); t2.start()

    t1.join(timeout=timeout_sec)
    t2.join(timeout=timeout_sec)

    if not done_event.is_set():
        timeout_event.set()
        print(f"⚠ 超时 {timeout_sec}s —— 强制放行线程1")
        return False
    print("✅ 双线程任务在时限内完成")
    return True

# ===============================================================
# 线程 1：执行插入操作（需根据自身项目调整 li()/insert_catalog_template）
# ===============================================================

def f1_insert_getwindow(timeout_event, done_event, *, ins_pt=(0, 0, 0),template_dwg =r"D:/Myprogramsystem/XT/mulugeshi_1.dwg",):
    pythoncom.CoInitialize()
    try:
        print("线程1启动")
        li()  # 你的 CAD 初始化函数
        insert_catalog_template(ins_pt=ins_pt,template_dwg =template_dwg)  # 这里触发缺字库对话框
        timeout_event.wait()                    # 等线程2放行
    except Exception as e:
        print("f1_insert_getwindow:", e, traceback.format_exc())
    finally:
        pythoncom.CoUninitialize()
        done_event.set()

# ===============================================================
# 线程 2：轮询对话框并 Esc 关闭
# ===============================================================

def f2_delwindow(
    timeout_event,
    done_event,
    *,
    poll_sec: float = 0.5,
    limit_sec: int  = 3,
    shx_titles=("缺少 SHX", "Missing SHX"),
):
    pythoncom.CoInitialize()
    try:
        print("线程2启动")
        t0 = time.time()
        while time.time() - t0 < limit_sec:
            if timeout_event.is_set():
                break
            hwnd = _find_shx_dialog(shx_titles)
            if hwnd:
                print("🖱 检测到‘缺少 SHX 字体’对话框，发送 Esc…")
                _ignore_shx_dialog(hwnd)
                timeout_event.set()
                break
            time.sleep(poll_sec)
        else:
            print(f"ℹ️ {limit_sec}s 内未出现 SHX 对话框，直接放行")
            timeout_event.set()
    except Exception as e:
        print("f2_delwindow 异常:", e, traceback.format_exc())
        timeout_event.set()
    finally:
        pythoncom.CoUninitialize()
        done_event.set()

# ===============================================================
# 高层封装：Insert_Catalog_Template
# ===============================================================

def Insert_Catalog_Template(
        ins_pt=(0, 0, 0),
        template_dwg =r"D:/Myprogramsystem/XT/mulugeshi_1.dwg",       
        scale = (1, 1, 1),
        rotation = 0,
        wait = 0.3,

        *, timeout_sec: int = 300

):



    """在 ins_pt 处插入目录模板，并自动忽略缺字库对话框"""
    ok = run_dual_threads(
        f1=f1_insert_getwindow,
        f2=f2_delwindow,
        f1_kwargs={"ins_pt": ins_pt,"template_dwg": template_dwg,"scale": scale,"rotation": rotation,"wait": wait},
        timeout_sec=timeout_sec,
    )
    if ok:
        print(f"🎉 成功插入 → {ins_pt}")
    else:
        print("🚨 插入目录模板失败")
    return ok

#&&&% （二） 插入标准目录模板
#&&%─────────────────────────────────────
        
@timeit
@debuggable
def supplementary_drawing_catalog_template(
        zd: dict,
        template_dwg : str = r"D:/Myprogramsystem/XT/mulugeshi_1.dwg",
        scale        : tuple = (1, 1, 1),
        rotation     : float = 0,
        wait         : float = 0.3,
        tol          : float = 10.0,

    ) -> dict:
    """
    · 插 1~2 份目录模板（总打印张数>57→2份）
    · 将新打印框多段线写回 prinfo / dyx_handles 且置于最前
    · 把旧打印框线图层统一改为新图层
    · 如采用无框图签，就不能使用默认目录模板，要自己带目录模板
    · template_dwg= r"D:/Myprogramsystem/XT/mulugeshi_2.dwg"时有些数据要改一下

    """

    # ── 内部工具 ───────────────────────────────────
    def _pt(x, y, z=0):
        return win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8,
            (float(x), float(y), float(z))
        )

    def _add_polyline_to_zd(pl):
        h = pl.Handle
        box_pl = pl.GetBoundingBox()        
        try:
            zhi = generate_name_and_ratio_from_com(
                    pl, A3dy=0,
                    Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3", 0))
            zd['prinfo'][h] = zhi
        except Exception:
            pass
        zd['dyx_handles'].insert(0, h)
        zd['dycoms'].insert(0, pl)
        zd['Bbox_lf'].insert(0, box_pl)    

    enable_debug()
    li()

    # ————————————————————————————
    # 0) 已有目录模板直接返回
    if zd.get("flag", [None, None, None])[2] is not None:
        node("[节点0] 已存在目录模板，跳过")
        disable_debug()
        return zd

    # 1) 基准点
    try:
        first_ent = zd['dycoms'][0]
    except Exception:
        first_ent = doc.HandleToObject(zd['dyx_handles'][0])
    (bl, _ur) = first_ent.GetBoundingBox()
    base_pt = (bl[0] - 100000, bl[1], 0)
    node("[节点1] base_pt = {}", base_pt)

    # 2) 份数（最多 2）
    copies = 2 if len(zd.get('dyx_handles', [])) > 57 else 1

    # 3) 插入目录块（只一次）
    Insert_Company_Label_Common_Block(
        insertion_point = (0, 0, 0),
        filepath        = template_dwg,
        scale           = scale,
        rotation        = rotation,
        wait            = wait,
        timeout_sec     = 300,
    )
    time.sleep(1.5); li(); doc.SendCommand("RE\n\n")

    # 4) 找块 / 就位
    blocks = [e for e in stc("mulugeshi")
              if e.ObjectName == 'AcDbBlockReference']
    if not blocks:
        node("❌ 未找到 mulugeshi 块")
        disable_debug()
        return zd

    bk0 = blocks[0]
    (bk0_bl, _ur) = bk0.GetBoundingBox()
    dx0, dy0 = base_pt[0] - bk0_bl[0], base_pt[1] - bk0_bl[1]
    if abs(dx0) > 1e-6 or abs(dy0) > 1e-6:
        bk0.Move(_pt(0,0,0), _pt(dx0,dy0,0))
    node("[节点2] 目录块#1 就位")

    # 如需第 2 张 → 复制右移 50 000
    bk_list = [bk0]
    if copies > 1:
        bk1 = bk0.Copy()
        bk1.Move(_pt(0,0,0), _pt(50000,0,0))
        bk_list.append(bk1)
        node("[节点2] 目录块#2 就位 (右移 50000)")

    # 5) 炸块并收集打印框多段线和图签块
    ensure_layer("增补目录模板")#后续要从该图层选取对象，先清场
    ensure_layer("标准图签")#后续要从该图层选取对象，先清场
    new_pls = []
    new_tqs = []

    for idx, bk in enumerate(bk_list, 1):
        try:
            lp = safe_explode_and_delete(bk)
            node("💥 块#{} 炸开 ({} 实体)", idx, len(lp))
            time.sleep(1)
            new_pls.extend(
                [pl for pl in lp if pl.ObjectName in ("AcDbPolyline", "AcDb2dPolyline") and pl.Layer == "增补目录模板" ]
            )

            new_tqs.extend(

                [tq_kuai for tq_kuai in lp if tq_kuai.ObjectName =='AcDbBlockReference' and  tq_kuai.Layer == "标准图签" ]
            )


        except RuntimeError as e:
            node("🚨 块#{} 炸开失败 {}", idx, e)

    # 6) 从增补目录模板图层再稳一次（避免炸块后图层改变）
    new_pls_layer = [pl for pl in stc("增补目录模板")
                     if pl.ObjectName in ("AcDbPolyline", "AcDb2dPolyline")]
    if new_pls_layer:
        new_pls = new_pls_layer

    new_tq_kuai_layer = [tq_kuai for tq_kuai in stc("标准图签")
                     if tq_kuai.ObjectName == 'AcDbBlockReference']
    if new_tq_kuai_layer:
        new_tqs = new_tq_kuai_layer

    # 7.1) 把新多段线写回 zd，并使用旧图层名
    old_layer = first_ent.Layer
    new_pls.sort(key=lambda p: p.GetBoundingBox()[0][0])
    for pl in reversed(new_pls):
        _add_polyline_to_zd(pl)
        try:
            pl.Layer = old_layer
        except Exception:
            pass
    node("[节点3.1] 新打印框写回并改层 → {}", old_layer)

    # 7.2) 把新图签块写回 zd，并使用旧图层名
    old_layer = "dy_quyu"
    new_tqs.sort(key=lambda p: p.GetBoundingBox()[0][0])
    # 确保键存在一次
    tgt = zd.setdefault('youkuang_tq_coms', [])


    # 倒序插入，新列表最终顺序同 new_tqs 的正序
    for tq_kuai in reversed(new_tqs):
        tgt.insert(0, tq_kuai)
        try:
            tq_kuai.Layer = old_layer
        except Exception:
            pass
    node("[节点3.2] 新图签块写回并改层 → {}", old_layer)
    
 
    # 7.3~7.5) —— 修正 zd['youkuang_handles'] ==================================
    youkuang = zd.setdefault('youkuang_handles', [[], [], []])
    basic_handles, all_handles, bbox_list = youkuang



    def size_equal(b1, b2, tol=tol):
        """只比较长边、短边尺寸是否相同（位置可不同）"""
        dx1, dy1 = abs(b1[1][0] - b1[0][0]), abs(b1[1][1] - b1[0][1])
        dx2, dy2 = abs(b2[1][0] - b2[0][0]), abs(b2[1][1] - b2[0][1])
        long1, short1 = sorted((dx1, dy1), reverse=True)
        long2, short2 = sorted((dx2, dy2), reverse=True)
        return abs(long1 - long2) < tol and abs(short1 - short2) < tol


    # —————————————————— 7.3 仅用 new_tqs[0] 更新基本图签列表 ——————————————————
    if new_tqs:
        tq0        = new_tqs[0]
        h0_new     = tq0.Handle
        bbox0_new  = tq0.GetBoundingBox()      # 直接用实体外包盒

        # 判重：把 basic_handles 转成 com 实体，再比对 GetBoundingBox
        dup_found = False
        try:
            for com in handles_to_coms(basic_handles):
                if size_equal(com.GetBoundingBox(), bbox0_new, tol):
                    dup_found = True
                    break
        except Exception:
            dup_found = h0_new in basic_handles  # 回退

        if not dup_found:
            basic_handles.insert(0, h0_new)
            
            node("[节点3.3] 基本图签 +1 Handle={}, bbox={}", h0_new, bbox0_new)
        else:
            node("[节点3.3] 跳过基本图签，发现重复 Handle={}", h0_new)
    else:
        node("[节点3.3] new_tqs 为空，基本图签不变")

    # —————————————————— 7.4 全部 Handle 插到最前 ——————————————————
    handles_new = [blk.Handle for blk in new_tqs]
    for h in reversed(handles_new):
        if h not in all_handles:
            all_handles.insert(0, h)
    node("[节点3.4] 汇总图签 Handle 数 = {}", len(all_handles))

    # —————————————————— 7.5 写入新“边域”数据 ——————————————————
    if new_tqs:
        try:
            bboxes_new = get_blocks_remaining_bbox(new_tqs)  # 边域数据
        except Exception:
            bboxes_new = [blk.GetBoundingBox() for blk in new_tqs]

        bbox_list[0:0] = bboxes_new
        node("[节点3.5] 外包盒列表长度 = {}", len(bbox_list))

    # 8) 目录识别

    try:

        dycom_list = zd.get('dycoms', [])
        flag_mulu0, mulu_rows, mulu_pages = find_directory_frame(dycom_list)
        handle_merge =[ent.Handle for ent in dycom_list]
        node("[节点4] 计算目录页数所依据的打印区域{}",handle_merge)
        # 更新 zd["flag"] ，确保长度足够
        flags = zd.setdefault("flag", [None]*5)
        flags[2] = flag_mulu0
        flags[3] = mulu_rows
        flags[4] = mulu_pages

        node("[节点5] 插入目录模板后：目录首序号 {} , 行数 {} , 目录页数 {}",
             flag_mulu0, mulu_rows, mulu_pages)
    except Exception as e:
        node("⚠ 目录识别失败：{}", e)

    # 9) 简单清理 (li+ql)
    try:
        ql()
    except Exception:
        try:
            li()
            ql()
        except Exception:
            pass


    # 取消选择状态
    cancel_cad_selection()
    # 保存
    try: savefile()
    except: pass


    # 更新字典并保存到电脑本地

    
    try:
        dwg_name = current_dwg_basename()
  

    except Exception as e:

        time.sleep(1)
        dwg_name = current_dwg_basename()
        
        node("[ERROR] current_dwg_basename 失败，再尝试一次：{}",e)

    try:
        save_print_dict_generic(dwg_name, zd)
        node("[INFO] 已保存到字典 -> {}",dwg_name)
    except Exception as e:

        time.sleep(1)
        save_print_dict_generic(dwg_name, zd)
        node("[ERROR] save_print_dict_generic 失败：{}",e)



    disable_debug()
    node("=== supplementary_drawing_catalog_template END ===")
    return zd



def filter_blocks_by_aspect_and_area(
    com_objects: list,
    aspect_tol: float,
    area_min: float = 160000000,
    area_max: float = 1000000000
) -> list:
    """
    从一组 COM 对象中筛选出满足以下两个条件的对象：
      1. 外包盒长边与短边的比值（长/短） >= ratio_threshold
      2. 面积（长边 * 短边）在 [area_min, area_max] 范围内

    :param com_objects:     要筛选的 COM 对象列表，每个对象需支持 GetBoundingBox() 方法。
    :param ratio_threshold: 长/短 比值阈值（例如 2.0 表示至少要是 2:1）。
    :param area_min:        面积下限，默认 160000000。
    :param area_max:        面积上限，默认 1000000000。
    :return:                符合条件的 COM 对象列表（保持原来顺序）。
    """
    filtered = []

    for obj in com_objects:
        try:
            # GetBoundingBox 返回 (ll_point, ur_point)，分别是最低点和最高点
            ll_point, ur_point = obj.GetBoundingBox()
            x1, y1, _ = ll_point
            x2, y2, _ = ur_point

            # 计算宽度和高度
            width = abs(x2 - x1)
            height = abs(y2 - y1)

            # 确定长边和短边
            long_side = max(width, height)
            short_side = min(width, height)

            # 避免除以零
            if short_side == 0:
                continue

            # 计算比值和面积
            aspect_ratio = long_side / short_side
            area = width * height

        except Exception:
            # 如果无法获取包围盒，就跳过此对象
            continue

        # 条件判断：比值和面积
        if aspect_ratio >= aspect_tol and area_min <= area <= area_max:
            filtered.append(obj)

    return filtered


#按面积相同去重，同名块可能进行了放缩

def dedupe_by_area(
    com_objects: list,
    tol: float = 100.0
) -> list:
    """
    从一组 COM 对象中，根据外包盒面积进行去重过滤：如果两个对象的面积差值小于 tol，
    则视为同一组，只保留第一次出现的对象，返回去重后的对象列表。

    :param com_objects: 要筛选的 COM 对象列表，每个对象需支持 GetBoundingBox() 方法。
    :param tol:         面积差值容差，小于该值时视为重复（默认 100）。
    :return:            去重后的 COM 对象列表（保持原始顺序）。
    """
    seen_areas = []
    result = []

    for obj in com_objects:
        try:
            ll_point, ur_point = obj.GetBoundingBox()
            x1, y1, _ = ll_point
            x2, y2, _ = ur_point

            width = abs(x2 - x1)
            height = abs(y2 - y1)
            area = width * height
        except Exception:
            # 如果无法获取包围盒，就跳过此对象
            continue

        # 检查当前面积是否与已记录面积差值小于 tol
        is_duplicate = False
        for existing_area in seen_areas:
            if abs(area - existing_area) < tol:
                is_duplicate = True
                break

        if not is_duplicate:
            seen_areas.append(area)
            result.append(obj)

    return result


#按位置去重

def dedupe_by_bbox_proximity(
    com_objects: list,
    tol: float = 20.0
) -> list:
    """
    对一组 COM 对象，根据其外包盒中心点的距离进行去重：如果两个对象的中心点距离小于 tol，
    则只保留第一个出现的对象，返回去重后的对象列表。

    :param com_objects: 要筛选的 COM 对象列表，每个对象需支持 GetBoundingBox() 方法。
    :param tol:         距离容差，小于该值时认为是“同一位置”，默认 20。
    :return:            去重后的 COM 对象列表（保持原始顺序）。
    """
    seen_centers = []
    result = []
    tol_sq = tol * tol

    for obj in com_objects:
        try:
            ll_point, ur_point = obj.GetBoundingBox()
            x1, y1, _ = ll_point
            x2, y2, _ = ur_point
            # 计算外包盒中心点
            cx = (x1 + x2) / 2.0
            cy = (y1 + y2) / 2.0
        except Exception:
            # 如果无法获取外包盒，则跳过此对象
            continue

        # 检查与已记录中心点的距离
        is_duplicate = False
        for (sx, sy) in seen_centers:
            dx = cx - sx
            dy = cy - sy
            if dx * dx + dy * dy < tol_sq:
                is_duplicate = True
                break

        if not is_duplicate:
            seen_centers.append((cx, cy))
            result.append(obj)

    return result


##筛选含图签和框线的图签块

@debuggable
def get_filtered_entities_from_blocks(
    block_refs: list,
    aspect_tol: float = 4,
    area_min: float = 160_000_000,
    area_max: float = 1_000_000_000
) -> list:
    """
    返回满足三重条件的块参照：
      ① 本身是 **属性块**（GetAttributes 返回非空）
      ② 内部实体中含 ≥1 个  (AcDbPolyline / AcDb2dPolyline / AcDbBlockReference)
      ③ 这些候选实体中存在 长/短 ≥ aspect_tol  且面积∈[area_min, area_max]
    """
    node("=== 进入 get_filtered_entities_from_blocks ===")
    qualified = []

    for blk in block_refs:
        blk_name   = getattr(blk, "Name", "<Unknown>")
        blk_handle = getattr(blk, "Handle", "<?>")
        node("→ 处理块 {} (Handle {})", blk_name, blk_handle)

        # ---------- 0) 属性块判定 ----------
        try:
            tags, vals = huoqukuai_shuxing_zhi(blk)
            if not tags and not vals:      # 既无 Tag 也无 Value
                node("  ✗ 非属性块，跳过")
                continue
            node("  ✓ 属性块 (属性数：{})", len(tags))
        except Exception as e:
            node("  ✗ 获取属性失败：{} — 跳过", e)
            continue

        # ---------- 1) 提取内部实体 ----------
        try:
            inner = get_entities_from_block_reference(blk)
            node("  内部实体数：{}", len(inner))
        except Exception as e:
            node("  ✗ 获取内部实体失败：{}", e)
            continue

        # ---------- 2) 类型初筛 ----------
        cand_types = ("AcDbPolyline", "AcDb2dPolyline", "AcDbBlockReference")
        cands = [e for e in inner if e.ObjectName in cand_types]
        node("  候选实体数：{}", len(cands))

        # ---------- 3) 长宽比 + 面积二次筛 ----------
        filtered = filter_blocks_by_aspect_and_area(
            cands,
            aspect_tol,
            area_min,
            area_max
        )
        node("  过滤后命中：{}", len(filtered))

        # ---------- 4) 最终是否收录 ----------
        if filtered:
            node("  ✓ 块 {} 被选中", blk_name)
            qualified.append(blk)

    node("=== 结束：共返回 {} 个块 ===", len(qualified))
    return qualified


#应该将多段线信息系统地以字典返回



"""
普通图签方式和属性图签方式两种，前者作为临时过渡 后者是标准化 成熟后只要选择好公司 就按自动化处理

纯 编目录打印服务 用户不提供图签

重写序号，适用于反复修改图纸，用户无需考虑图纸编号的变化

在用户提供的图纸目录模板上编写目录 ，用户需提供第一行文字的样板 也就是图纸目录这张图纸上的对应值 如果图纸数量很多需要两排或两张图纸目录 都应在最上方提供样板

对于已有目录的修改，会自动清零重排，适用于反复修改图纸的目录重编

锁定有框图签块 ，如果运算process_print_areas_info所得整体图签块列表非零，取第一个元，获取内部独立图签部分的外包盒边界右下角点 必定和某条打印线有确定的位置关系

锁定无框图签块，如果某个图块不是独立图签却在process_print_areas_info所得的独立图签块列表，则其外包盒的最右下角点 必定和某条打印线有确定的位置关系


于是我们就可以确定图签块

对于独立图签块情况，我们通过每张图纸的15米边域 块属性 连通图纸名称文字和序号文字的选择 一次性搞定 即通过一个函数选择到所有的图签块 图纸名字 序号 
并且可以判断无图签块的打印区域必定是项目名称 

包含  图纸目录  文字的打印区域就是目录这张图纸 还可能延伸到下一张 需要分析有大量水平等距线条并超过某个高度 

到达 图纸目录 的图纸后 分析均匀间隔垂直直线或多段线、水平直线或多段线 在均匀区域的第一排或上一排应该找到两排文字 它们在网格中

找到一个文字 ，其同样y坐标应该有4个文字以上 并且 这些文字的外包盒 与均匀网格的起点不会偏离多远
两排这样的文字，第二排就是

最左边的数字 是序号 有汉字并且数量大于3的是图纸名称 以A+数字构成的是图纸规格 不是



综合 

不管用户 采用方案A，还是方案B，还是如何，我们统一考虑各种情况，而不是繁琐分析如果对方采取A如何如何 

对一个功能的分析尽量到位，甚至每个小功能都是函数形式 这样其本身的结果和调试都很简明 组合功能也很简明 提倡尽量用函数写函数

将内部函数的作用结果的Handle打印出来 对于python数据 当然更简单了  这样就不用进行繁琐的反馈测试 直接可以查看中间的运行结果 

尽可能在一次操作中集成多方面的信息采集和分析 减少操作CAD 的次数 

函数的扁平化扩张并不改变函数的复杂度 虽然我们愿意把函数尽可能写得简单，但扁平化的扩展并不影响其复杂度

X = F(X) 的函数形式 便于修正调整函数


"""

# 1 确定图签块对象及有无图框线


#1)如果某个图块不是无框图签却在process_print_areas_info所得的无框图签块列表，则其外包盒的最右下角点 必定和某条打印线有确定的位置关系。以此清除可疑无框图签

#process_print_areas_info所得的有框图签块列表，不可能含有非图签块

#2)运算process_print_areas_info所得有框图签块列表非零，则取第一个元，获取内部独立图签部分的外包盒边界右下角点 必定和某条打印线有确定的位置关系，以此判定是有框还是无框
@debuggable
def filter_coms_by_proximity_to_frames(LBcoms, LB_dy, tol=3000.0):
    """
    对 com 对象列表 LBcoms 中的每个元素，获取其外包盒右下角点坐标 p = (maxX, minY, z)。
    然后令 p_db = (p[0] + 1000, p[1] - 1000, 0)。如果 p_db 与打印多段线列表 LB_dy
    中任意一根多段线的外包盒右下角坐标之间的距离 < tol，则将该 com 对象加入 LBcoms_new。

    参数：
      LBcoms : List of COM objects
        需要筛选的实体对象列表，每个对象都支持 GetBoundingBox()。
      LB_dy  : List of COM polyline objects
        打印多段线列表，每个对象都支持 GetBoundingBox()。
      tol    : float, default=20.0
        距离容差值，小于此值视为“匹配”。

    返回：
      LBcoms_new : List of COM objects
        满足与任意打印多段线的右下角距离 < tol 的 com 对象列表。
    """
    LBcoms_new = []

    # 1. 先计算 LB_dy 中每根多段线的“右下角”坐标字典
    dy_corners = []
    for dy in LB_dy:
        try:
            dy_min_pt, dy_max_pt = dy.GetBoundingBox()
        except Exception:
            # 如果多段线不支持外包盒，跳过
            continue

        # dy 的右下角：x = maxX, y = minY
        dy_x = dy_max_pt[0]
        dy_y = dy_min_pt[1]
        # z 坐标可以忽略，比较平面距离即可
        dy_corners.append((dy_x, dy_y))



    # 2. 遍历 LBcoms 中的每个对象
    for comobj in LBcoms:
        try:
            com_min_pt, com_max_pt = comobj.GetBoundingBox()
        except Exception:
            # 如果对象不支持外包盒，跳过
            continue

        # com_obj 的右下角：x = maxX, y = minY
        com_x = com_max_pt[0]
        com_y = com_min_pt[1]

        # 生成 p_db = (com_x + 1000, com_y - 1000)
        p_db_x = com_x + 1000
        p_db_y = com_y - 1000



        # 3. 与 LB_dy 中每根多段线的右下角坐标比较距离
        for dy_x, dy_y in dy_corners:
            dx = p_db_x - dy_x
            dy_dist = p_db_y - dy_y
            distance = math.hypot(dx, dy_dist)
            if distance < tol:
                # 只要有一根多段线满足条件，就把 comobj 加入结果并跳出内层循环
                LBcoms_new.append(comobj)
                break

    return LBcoms_new

def get_dedup_blocks_on_layer(layername):
    """
    从指定图层选取所有块参照，并按包围盒邻近度去重。

    参数
    ----
    layername : str
        图层名称

    返回
    ----
    list[COMObject]
        去重后的块参照列表
    """
    # 1. 选取图层上所有对象
    entities = stc(layername)
    # 2. 过滤出块参照
    LB1 = [ent for ent in entities if ent.ObjectName == "AcDbBlockReference"]
    # 3. 按包围盒邻近度去重
    LB_unique = dedupe_by_bbox_proximity(LB1)
    return LB_unique


# 2 确定边域

#1)有框图签
#从块内最大外包盒右下角点，往左或往上15000，由序号索引load的字典或process_print_areas_info所得的返回值字典断定横竖，在此区域选择对象，获取对象组的外包边界x方向宽度Lx

#2)无框图签
#从自建打印框线的Pj_out，往左或往上15000，由序号索引load的字典或process_print_areas_info所得的返回值字典断定横竖，在此区域选择合乎图签块特征的块Bk 

#3)按序返回列表


#处理无框
def get_bbox_corners_list(LBcoms):
    """
    对 com 对象列表 LBcoms 中的每个元素，获取其外包盒坐标并返回左下角和右上角。

    参数：
      LBcoms : list of COM objects
        每个对象都支持 GetBoundingBox()，返回 (minPt, maxPt)。

    返回：
      LB_bbox : list of tuple or None
        与 LBcoms 等长。LB_bbox[i] = ((minX, minY, minZ), (maxX, maxY, maxZ))，
        如果某个对象无法获取外包盒，则在对应位置放 None。
    """
    LB_bbox = []
    for comobj in LBcoms:
        try:
            min_pt, max_pt = comobj.GetBoundingBox()
            # min_pt 就是左下角 (minX, minY, minZ)
            # max_pt 就是右上角 (maxX, maxY, maxZ)
            LB_bbox.append((min_pt, max_pt))
        except Exception:
            LB_bbox.append(None)
    return LB_bbox

#处理有框
def get_blocks_remaining_bbox(
    LBcoms: list
) -> list[tuple[tuple[float, float, float], tuple[float, float, float]]]:
    """
    对列表 LBcoms 中的每个块参照，返回其内部剩余对象（排除图层“tuqian_neibu_pl”）
    的外包盒左下角和右上角的世界坐标。

    :param LBcoms: 块参照对象列表
    :return: 与 LBcoms 等长的列表，每项为 (world_bl, world_tr)，若获取失败则为 (None, None)
    """
    results = []
    for block_ref in LBcoms:
        # 1. 获取块内所有实体
        try:
            all_entities = get_entities_from_block_reference(block_ref)
        except Exception:
            results.append((None, None))
            continue

        # 2. 排除图层“tuqian_neibu_pl”上的对象
        rem_objs = [ent for ent in all_entities
                    if getattr(ent, 'Layer', '') != 'tuqian_neibu_pl']
        if not rem_objs:
            results.append((None, None))
            continue

        # 3. 计算本地坐标系下的外包盒
        bbox = group_bbox_corners(rem_objs)
        if not bbox:
            results.append((None, None))
            continue
        # 左下、右上
        local_bl = bbox[0]
        local_tr = bbox[1]

        # 4. 转换到世界坐标
        try:
            world_bl = transform_point_by_block(block_ref, local_bl)
            world_tr = transform_point_by_block(block_ref, local_tr)
        except Exception:
            results.append((None, None))
            continue

        results.append((world_bl, world_tr))

    return results



#&&&% （三） 获取图纸名称、图纸编号等信息
#&&%─────────────────────────────────────
        



"""

——以目标文件带路径名为输入参数，引入小工具--未打开则打开，已经打开则确保为当前激活。用于全自动流程

🧠 从边域选择得到图纸名称和图纸编号信息，其余信息用户自理

⮞ F1:process_side_regions_info(zd,cha_Y=100)



🧠 从边域选择得到图纸名称和图纸编号信息，从对应同名Word文件获取图纸名称和图纸编号之外的所有信息

⮞ F2:
🧠 从标准属性块获取全部相关信息，系统总是插入基本属性块，给用户处理。在得到允许时，清理图纸，插入修改好或未修改的标准属性图签基本块到所有图纸

⮞ F2:

🧠 选择到的对象列表为输入参数，直接获取。用于管理员临时处理这一环节

⮞ F3: 
zd=zd=process_print_areas_info()

process_side_regions_info(zd)



"""



def add_centered_text(x, y, content, height=250):
    """
    在 ModelSpace 中于 (x, y) 插入一行居中对齐的文本，使用当前默认字体样式和指定字高。
    返回该文本实体。
    """
    ms = doc.ModelSpace
    # 添加单行文本
    text_ent = ms.AddText(content, vtpnt(0.0, 0.0, 0.0), height)
    # 将插入点移到 (x, y)
    text_ent.TextAlignmentPoint = (x, y, 0.0)
    # 设置对齐方式为居中（Alignment=1 对于单行 Text 表示中间对齐）
    text_ent.Alignment = 1
    text_ent.Update()  # 更新显示
    return text_ent

# 确定图纸目录模板序号及行数、目录纸张数

@timeit
@debuggable
def find_directory_frame(LB):
    """
    检测打印框列表 LB 中的目录图纸。（2025-07-11 版，兼容横/竖向，高度计算修正）
    返回 (idx, row_cnt, muluzhizhangshu)。
    """
    enable_debug()
    max_width, max_height = 59400, 42000
    required_chars = set("图纸目录")
    search_limit = min(5, len(LB))
    tol, min_count, ci, cha_Y = 50, 10, 3, 100

    import re
    def _bbox_of(frame):
        corners = group_bbox_corners([frame])
        if not corners: return None
        bl, tr, *_ = corners
        return bl, tr

    def _extract_useful_texts(x1, y1, x2, y2):
        ents = select_entities_in_window(x1, y1, x2, y2, ty=1.0, select_mode="_W")
        raw = [e for e in ents if e.ObjectName in ("AcDbText","AcDbMText","TDbText","TDbMText")]
        out = []
        for e in raw:
            txt = (extract_text_content(e) or "").replace(" ", "")
            if re.fullmatch(r"[A-Za-z]+", txt): continue
            if "." in txt and re.search(r"[A-Za-z]", txt): continue
            out.append(e)
        return out

    for idx in range(search_limit):
        frame = LB[idx]
        bb = _bbox_of(frame)
        if not bb: continue
        (x1, y1, _), (x2, y2, _) = bb

        w, h = x2 - x1, y2 - y1
        if not ((w <= max_width and h <= max_height)
                or (w <= max_height and h <= max_width)):
            continue

        texts = _extract_useful_texts(x1, y1, x2, y2)
        page_txt = "".join(extract_text_content(e) or "" for e in texts)
        hit_text = required_chars.issubset(page_txt)
        if hit_text:
            node("图纸 {} 命中关键词『图 纸 目 录』", idx)

        groups = find_feature_horizontal_groups(
            x1, y1, x2, y2,
            tol=tol, min_count=min_count, ci=ci, cha_Y=cha_Y,
            layer_name="bianmulu_lp"
        )
        hit_group = bool(groups)
        if hit_group:
            node("图纸 {} 找到水平线簇", idx)

        if not (hit_text and hit_group):
            continue

        node("✅ 图纸 {} 同时命中关键词 + 水平线簇 → 目录页", idx)

        # —— 行数推算 —— #
        first_group = groups[0]
        spacing = first_group[2]
        lines = first_group[5]   # grp list: 所有水平线对象

        # 找到最上和最下水平线的左下角坐标
        bl_points = [e.GetBoundingBox()[0] for e in lines]
        top_bl = max(bl_points, key=lambda p: p[1])
        bot_bl = min(bl_points, key=lambda p: p[1])
        node("最上一根水平线左下角坐标: {}", top_bl)
        node("最下一根水平线左下角坐标: {}", bot_bl)

        # 用这两点 y 差值计算真实高度
        real_height = abs(top_bl[1] - bot_bl[1])
        row_cnt = int(round(real_height / spacing)) 

        node("行距: {}，真实外包盒高: {}，推算行数: {}", spacing, real_height, row_cnt)

        # —— 检测第二张目录页 —— #
        muluzhizhangshu = 1
        next_idx = idx + 1
        if next_idx < len(LB):
            bb2 = _bbox_of(LB[next_idx])
            if bb2:
                x1n, y1n, _ = bb2[0]; x2n, y2n, _ = bb2[1]
                g2 = find_feature_horizontal_groups(
                    x1n, y1n, x2n, y2n,
                    tol=tol, min_count=min_count, ci=ci, cha_Y=cha_Y,
                    layer_name="bianmulu_lp"
                )
                txt2 = _extract_useful_texts(x1n, y1n, x2n, y2n)
                page2 = "".join(extract_text_content(e) or "" for e in txt2)
                if g2 and required_chars.issubset(page2):
                    muluzhizhangshu = 2
                    node("检测到图纸 {} 也为目录页 → 目录共 2 张", next_idx)

        disable_debug()
        return idx, row_cnt, muluzhizhangshu

    disable_debug()
    return None, None, 0


#&&&%  ▶ F1 使用边域处理图纸名称和编号

@timeit
@debuggable
def process_side_regions_info(zd, *, cha_Y=100, cha_X=100):
    """
    1 图纸名称和序号的格式规定
    ----------------------------------------------------------
    • 序号：≤3 连续汉/英字 + 可选“-” + ≥1 数字（忽略空格）
    • 图纸名：至少 1 汉字且不符合序号规则


    2 竖向和横向区域
    ----------------------------------------------------------
    • 竖向区域：sort_coms_by_rbcorner  (右→左，上→下)
    • 横向区域：sort_coms_by_llcorner (左→右，上→下)

    返回字段均与侧边区域列表等长，对齐填充 None。

    序号只能写一个单行字符，图纸名称只能写多个单行字符
    竖向图纸的文字要整体旋转-90度

    3 写入图签和删除文字com
    ----------------------------------------------------------
    • 有框图签：将属性块的图纸名称和图纸序号写入，删除图纸名称com和序号com
    • 无框图签：更新原来的序号com实体的文字内容，保留图纸名称com实体

    """

    enable_debug(); node("[节点0] 开始侧边区域信息提取")

    # ——— 判定/工具 -------------------------------------------------------------------
    _serial_re = re.compile(r'^[A-Za-z\u4e00-\u9fa5]{1,3}-?\d+$')
    chinese_re = re.compile(r'[\u4e00-\u9fa5]')
    def is_serial(t): return _serial_re.match(t.replace(" ", "")) is not None

    def get_style(e):
        for _ in range(3):
            try:
                return e.TextStyle if e.ObjectName.startswith("TDb") else e.StyleName
            except Exception: time.sleep(1)
        return "Standard"

    # ——— 1. 确定边域类型 ----------------------------------------------------------------

    flag = zd.get("flag", [None])[0]
    zb   = (zd.get("wukuang_handles", [None,None,[]])[2] if flag==-1
            else zd.get("youkuang_handles", [None,None,[]])[2] if flag==1
            else None)
    if zb is None:
        print("打印线字典信息输入错误"); disable_debug(); return zd

    node("[节点1] 无框图签取值-1 有框图签取值1 {})", flag)
    N = len(zb)

    # ——— 初始化结果占位 ---------------------------------------------------------------
    xuhao_texts    = [None]*N
    xuhao_coms     = [None]*N
    xuhao_info     = [None]*N
    tuzhimingcheng = [None]*N
    tuming_coms    = [None]*N
    dyx_list       = zd.get("dyx_handles", [])
    prinfo         = zd.get("prinfo", {})

    # ——— 2. 遍历区域获取图纸名称和序号信息 ---------------------------------------------------------------------
    for i, region in enumerate(zb):
        (x1,y1,_),(x2,y2,_) = region
        x_lo,x_hi = sorted((x1,x2)); y_lo,y_hi = sorted((y1,y2))
        handle_i  = dyx_list[i] if i < len(dyx_list) else None
        orient    = prinfo.get(handle_i,(None,None,None,0))[3]   # 0=横 1=竖
        node("[节点2] 区域#{}, handle={}, orient={}", i, handle_i, orient)

        objs   = select_entities_in_window(x_lo,y_lo,x_hi,y_hi, ty=1.0, select_mode="_W")
        texts  = [e for e in objs if e.ObjectName in
                  ("AcDbText","AcDbMText","TDbText","TDbMText")]

        # —— 2-a 序号 --------------------------------------------------------------------
        cand_serial = [e for e in texts
                       if (txt:=extract_text_content(e)) and is_serial(txt)]
        if cand_serial:
            ent   = min(cand_serial, key=lambda e: e.GetBoundingBox()[0][1])
            txt   = extract_text_content(ent)
            p1,p2 = ent.GetBoundingBox()
            sty   = get_style(ent)

            try:  hgt = ent.Height
            except Exception: hgt = 250            # 默认高度

            xuhao_texts[i], xuhao_coms[i], xuhao_info[i] = txt, ent, (txt, ent, (p1,p2), sty,hgt)
            node("[节点2-a] 序号='{}' 样式='{}'", txt, sty)
        else:
            node("[节点2-a] 未检测到序号")

        # —— 2-b 图纸名称 -----------------------------------------------------------------
        cand_title = [e for e in texts
                      if (txt:=extract_text_content(e)) and chinese_re.search(txt)
                         and not is_serial(txt)]
        if cand_title:
            if orient == 0:   # 横向
                sorted_txt = sort_coms_by_llcorner(cand_title, cha_Y=cha_Y)
            else:            # 竖向
                sorted_txt = sort_coms_by_rbcorner(cand_title, cha_X=cha_X)

            joined = "、".join(extract_text_content(e) for e in sorted_txt)
            tuzhimingcheng[i], tuming_coms[i] = joined, sorted_txt
            node("[节点2-b] 图纸名称='{}'", joined)
        else:
            node("[节点2-b] 未检测到图纸名称")

    # ——— 3. 序号前缀 ------------------------------------------------------------------------
    def prefix_before_digit(s):
        s = s.replace(" ", "")
        m = re.search(r"\d", s)
        return s[:m.start()] if m else s
    xuhao_qz = prefix_before_digit(next((t for t in xuhao_texts if t), ""))

    # ——— 4. 信息写入字典 -------------------------------------------------------------------------
    zd.update({
        "xuhao_texts"     : xuhao_texts,
        "xuhao_coms"      : xuhao_coms,
        "xuhao_info"      : xuhao_info,
        "tuzhimingcheng"  : tuzhimingcheng,
        "tuming_coms"     : tuming_coms,
        "xuhao_qianzhui"  : xuhao_qz,
    })

    try:
        name = current_dwg_basename() + "_by"
        save_print_dict_generic(name, zd)
        node("[节点3] 已保存打印字典 -> {}", name)
    except Exception as e:
        node("[节点3] 保存打印字典失败: {}", e)


    # ——— 5. 编号信息重写 -------------------------------------------------------------------
    dir_first_idx = (zd.get("flag",[None,None,0,0,0])[2] or 0)
    dir_pages     = (zd.get("flag",[None,None,0,0,0])[4] or 0)
    prefix        = zd.get("xuhao_qianzhui", "")
    alpha         = "abcdefghijklmnopqrstuvwxyz"
    new_nums      = [None]*N
    run_no        = 1

    for i in range(N):
        # —— 属于目录页 —— 
        if dir_first_idx <= i < dir_first_idx + dir_pages:
            if dir_pages == 1:
                # 单张目录，写 SS-00
                new_nums[i] = f"{prefix}00"
            else:
                # 多张目录，从 a 开始
                suffix = alpha[i - dir_first_idx]
                new_nums[i] = f"{prefix}00{suffix}"
        else:
            # —— 普通图纸 —— 
            new_nums[i] = f"{prefix}{run_no:02d}"
            run_no += 1

    zd["xuhao_texts"] = new_nums
    node("[节点4] 序号重写完成：{}", new_nums)



    # ——— 6. 图纸名称信息重写 ---------------------------------------------------------------
    # flag[2] = 第一个目录页的索引，flag[4] = 目录页总数
    dir_first_idx = zd.get("flag", [None, None, 0, None, 0])[2] or 0
    dir_pages     = zd.get("flag", [None, None, 0, None, 0])[4] or 0
    # 将每个目录页对应的图纸名称都改为“图纸目录”
    for j in range(dir_first_idx, dir_first_idx + dir_pages):
        if 0 <= j < len(zd.get("tuzhimingcheng", [])):
            zd["tuzhimingcheng"][j] = "图纸目录"
    node("[节点5] 图纸名称重写完成：{}", zd["tuzhimingcheng"])



    # ——— 7. 有框图签写入处理 ---------------------------------------------------------------
    if zd.get("flag", [None])[0] == 1:
        # 取出所有“dy_quyu”图层上的目录块（去重后）
        blocks = zd["youkuang_tq_coms"]

        # 遍历
        for blk, title, num, title_entities, num_entity in zip(
            blocks,
            zd["tuzhimingcheng"],
            zd["xuhao_texts"],
            zd["tuming_coms"],
            zd["xuhao_coms"],
        ):
            # 1) 写入块属性
            set_attributes_values(blk, ["多行4图纸名称"], [title])
            set_attributes_values(blk, ["单行3图纸编号"], [num])#冒号要去掉，造成不稳定

            # 2) 删除原有图纸名称实体
            if title_entities:
                for ent in title_entities:
                    safe_delete(ent)

            # 3) 删除原有序号实体
            if num_entity:
                    safe_delete(num_entity)

        node("[节点6] 有框图签写入处理完成")


    # ——— 8. 无框图签写入处理 ---------------------------------------------------------------
    if zd.get("flag", [None])[0] == -1:
        # 取出所有“dy_quyu”图层上的目录块（去重后）
        blocks = zd["wukuang_tq_coms"]

        # 同时遍历：块、名称文本列表、编号文本列表、所有原始图纸名实体列表、所有原始序号实体 单行3：图纸编号
        for blk, title, num, title_entities, num_entity in zip(
            blocks,
            zd["tuzhimingcheng"],
            zd["xuhao_texts"],
            zd["tuming_coms"],
            zd["xuhao_coms"],
        ):

            # 1) 更新/写入新的序号文字实体
            if num_entity:
                try:
                    if num_entity.ObjectName == "TDbText":
                        num_entity.Text = num
                    else:
                        num_entity.TextString = num
                except Exception:
                    # 三次重试也可以根据需要补充
                    time.sleep(0.1)
                    try:
                        if num_entity.ObjectName == "TDbText":
                            num_entity.Text = num
                        else:
                            num_entity.TextString = num
                    except:
                        pass
        node("[节点6] 无框图签写入处理完成")

    # ——— 9. 结尾   --------------------------------------------------------------------------

    node("[节点7] 图纸序号 {} 条，图纸名称 {} 条",
         sum(t is not None for t in xuhao_texts),
         sum(t is not None for t in tuzhimingcheng))

    node("字典flag消息{}",zd["flag"])
    node("flag[0]:-1即无框图签，1即有框图签，0为无图签；flag[1]:0即普通图签，1即属性图签；flag[2]:目录模板所在序号数，None即无目录模板")
    node("flag[3]:目录行数，无则为None；flag[4]:目录纸张数，None则未定")




    # 取消选择状态
    cancel_cad_selection()
    # 保存
    try: savefile()
    except: pass


    # 更新字典并保存到电脑本地   
    try:
        dwg_name = current_dwg_basename()
    except Exception as e:
        time.sleep(1)
        dwg_name = current_dwg_basename()       
        node("[ERROR] current_dwg_basename 失败，再尝试一次：{}",e)
    try:
        save_print_dict_generic(dwg_name, zd)
        node("[INFO] 已保存到字典 -> {}",dwg_name)
    except Exception as e:
        time.sleep(1)
        save_print_dict_generic(dwg_name, zd)
        node("[ERROR] save_print_dict_generic 失败：{}",e)


    disable_debug()
    return zd




#&&&%  ▶ F2 使用图层处理图纸名称和编号

@timeit             
@debuggable         
def select_text_from_layer(zd, *, cha_Y=100, cha_X=100):

    """
    使用图层选择方法获取图纸名称实体和序号实体，其它与process_side_regions_info相同
    
    
    """

    node(">> 进入 select_text_from_layer, 打印框数量={}", len(zd["Bbox_lf"]))

    LB_Box     = zd["Bbox_lf"]
    all_name   = stc("图纸名称")
    all_number = stc("图纸编号")

    node("全图 '图纸名称' ={} 条, '图纸编号' ={} 条", len(all_name), len(all_number))

    tuming_by_layer, xuhao_by_layer = [], []

    for i, bbox in enumerate(LB_Box):
        node("[{}] 处理打印框 bbox={}", i, bbox)

        # — 图名 —
        wz_dy = [ent for ent in all_name if in_bbox(ent, bbox)]
        node("[{}]  ▶ 名称候选 {} 条", i, len(wz_dy))

        handle_i   = zd["dyx_handles"][i]
        orient     = zd["prinfo"].get(handle_i, [None, None, None, 0])[3]

        if orient == 0:
            wz_dy = sort_coms_by_llcorner(wz_dy, cha_Y=cha_Y)
            node("[{}]    横向排序完成", i)
        else:
            wz_dy = sort_coms_by_rbcorner(wz_dy, cha_X=cha_X)
            node("[{}]    竖向排序完成", i)

        tuming = "、".join(ent.TextString.strip() for ent in wz_dy)
        tuming_by_layer.append(tuming)
        node("[{}]    拼接图名='{}'", i, tuming)

        # — 编号 —
        wz_th = [ent for ent in all_number if in_bbox(ent, bbox)]
        node("[{}]  ▶ 编号候选 {} 条", i, len(wz_th))

        if wz_th:
            ent_num = min(wz_th, key=lambda e: e.GetBoundingBox()[0][1])  # y 最小
            center  = tuple(
                (a + b) / 2 for a, b in zip(*ent_num.GetBoundingBox()[:2])
            )
            info = (
                ent_num.TextString.strip(),
                center,
                getattr(ent_num, "Height", None),
                getattr(ent_num, "StyleName", None),
                getattr(ent_num, "ScaleFactor", None),
            )
            node("[{}]    选中编号='{}' @{}", i, info[0], center)
        else:
            info = ("", None, None, None, None)
            node("[{}]    ⚠ 未找到编号", i)

        xuhao_by_layer.append(info)

    zd["tuzhimingcheng"] = tuming_by_layer
    zd["xuhao_by_layer"]  = xuhao_by_layer

    node("<< 结束 select_text_from_layer")
    return zd

def in_bbox(ent, bbox):
    """
    支持两种 bbox 形态：
      1) ((x1,y1,_), (x2,y2,_))
      2) (x1, y1, x2, y2)
    """
    try:
        (ex1, ey1, _), (ex2, ey2, _) = ent.GetBoundingBox()
    except Exception:
        return False

    # ——— 统一提取 x1,y1,x2,y2 ———
    if len(bbox) == 2 and isinstance(bbox[0], (list, tuple)):
        (bx1, by1, *_), (bx2, by2, *_) = bbox
    else:
        bx1, by1, bx2, by2 = bbox

    # 确保左下 / 右上
    if bx1 > bx2:
        bx1, bx2 = bx2, bx1
    if by1 > by2:
        by1, by2 = by2, by1

    return ex1 >= bx1 and ex2 <= bx2 and ey1 >= by1 and ey2 <= by2




#&&&%  ▶ F3 完善图签

"""
使用通用Excel表格或与dwg同名Excel表格获取图签上除图纸名称和编号之外的信息
    
    
"""





#&&&% （四） 目录初始标准化
#&&%─────────────────────────────────────
        
"""
根本思想和步骤

确定目录所在图纸  确定水平线簇  确定正文首行分格中心点 确定目录标签行文字  确定目录正文首行文字  为文字分配好图层


"""

@timeit            
@debuggable 
def annotate_directory_template(
    zd,
    basic_layers =["ML-图纸序号","ML-图纸名称","ML-图纸规格","ML-图纸编号","ML-图纸比例","ML-其它","ML-其它1"],
    zhuanye_daihao="JZ-",
    tol=50,
    min_count=10,
    blxs=80,
    tol_x = 10000.0,
    ci = 3,
    cha_Y1= 100,
    cha_Y2= 50,
    height_bu = None,
    kuanduyinzi_bu = None


):
    """
    模式1 写好目录正文首行示例，并分别放入ML图层。最好使用CAD文字，或天正文字写好展开为CAD

    模式2 写好目录正文首行示例，系统自己识别并放入相应图层。最好使用CAD文字，或天正文字写好展开为CAD

    模式3 空白，系统自己生成

    模式4 空白，Excel文件写入  专业代号当属性图签未与图纸名称传入时，可以在Excel文件写入

    将针对图纸目录的首行示例文字内容，分别放入"ML-图纸序号","ML-图纸名称","ML-图纸规格","ML-图纸编号","ML-图纸比例"即可按用户自己的意愿排序和填写内容
    允许使用CAD单行文字和天正单行文字
    如果不想文字大小发生变化，就不要用天正单行文字写首行示例文字，而要直接使用CAD文字
    ---------------------------------------------------------------------------------------------------------------------------------------

    允许用户不设置图层而直接按常规名字书写首行示例文字。程序会自动分图层。第一个格子肯定是序号，最大格子肯定是图纸名称。含有A3开头的这些（还要细究）
    肯定是图纸规格。1:100这种比例也容易识别。编号也容易识别。不能识别的归于ML-其它，它的进一步书写采用重复。当规格附加了其它符号的时候，原则上也是规
    格加其它符号复制。
    采用字典存储的Handle，我们可以批量快速修改已经写好的目录内容和图纸对应文字。也可以处理这种复杂情况。

    ---------------------------------------------------------------------------------------------------------------------------------------


    针对图纸目录的首行示例文字内容缺失，则补写图纸目录首行文字
    如果图纸目录标签设置复杂，局部包含有复杂构成，则只能采用用户给出首行示例文字形式，即这种情况不能采用空的首行示例文字

    ---------------------------------------------------------------------------------------------------------------------------------------
    blxs : 天正文字高度矫正系数（默认 80）
    返回 : True(成功) / False(失败)

    tol_x :ML-其它 图层有多个对象在一行上时都需要保留的x方向距离限定
    ci =  3 尝试次数
    --------------------------------------------------------------------------------------------------------------------------------------

    用户自控补写文字的高度和宽度因子height_bu、kuanduyinzi_bu

    目录首行示例文字的文字内容和实体信息存入字典保存


    --------------------------------------------------------------------------------------------------------------------------------------

    """
    li()
    enable_debug()
    
    if zd["flag"][2]==None:
    
        return node("文件没有目录模板！")

    w = zd["flag"][2]

    dycom_mulu = zd['dycoms'][w]

    x1, y1, _ = dycom_mulu.GetBoundingBox()[0]

    x2, y2, _ = dycom_mulu.GetBoundingBox()[1]
    
    node("目录首张的左下角点{} 右上角点 {} ",dycom_mulu.GetBoundingBox()[0],dycom_mulu.GetBoundingBox()[1])




    LB_canzhaoji = []

    # 0.1. 确认ML-图层不空
        
    create_layers_from_list(["ML-图纸名称","ML-图纸序号","ML-图纸编号","ML-图纸规格","ML-图纸比例","ML-其它","ML-其它1","bianmulu_lp"])

    # 0.2. 循环操作的清理
    clean_directory_layers(
    basic_layers=["ML-图纸序号","ML-图纸名称","ML-图纸规格","ML-图纸编号","ML-图纸比例","ML-其它","ML-其它1",])

    time.sleep(2)
    # 1. 取水平线、bbox 及表头标签和正文首行示例文字
    res = find_feature_horizontalgroups_and_deltext(
        x1,y1,x2,y2,tol=tol,min_count=min_count,ci=ci,cha_Y1=cha_Y1,cha_Y2=cha_Y2
    )
    if not res or len(res) < 7:
        print("图纸目录模板内容错误"); return False

    first_h, _, hangju, bbox, intervals,handles_label, handles_body = res
    (x_lo,y_lo,_), (x_hi,y_hi,_) = bbox

 
   
    if not intervals:
        print("无法识别列间距"); return False


    node("[节点1]获取水平线簇、水平线簇外包盒bbox 及表头标签和正文首行示例文字 {} ",res)

    node("如果获取的表头标签文字为空，将影响正常运转")


    # 2.1. 若已有正文 → 先归层（新版）
    if handles_body:
        body = handles_to_coms(handles_body)
        # ① 计算 max_int, j_max, 累加值 leijia
        max_int = max(intervals)
        j_max   = intervals.index(max_int)
        leijia  = sum(intervals[:j_max + 1])     # 累加到 j_max （含本身）
    
        # ② first_line 左下 x0   →  X0 = (x0+leijia) - max_int/2
        first_line = doc.HandleToObject(first_h)
        x0, y0, _  = first_line.GetBoundingBox()[0]
        X0 = x0 + leijia - max_int * 0.5
    
        # ③ 取得正文各文本的 bbox 中心 x
        centers = [( (t.GetBoundingBox()[0][0] + t.GetBoundingBox()[1][0]) * 0.5 , idx )
                   for idx, t in enumerate(body)]
    
        # ③-a 图层“名称” → 距 X0 最近
        idx_name = min(centers, key=lambda c: abs(c[0] - X0))[1]
    
        # ③-b 图层“序号” → x 最小
        idx_num  = min(centers, key=lambda c: c[0])[1]
    
        # ④ regex 规则（去空格）
        patt_spec  = re.compile(r'^A\d+[^A-Za-z\u4e00-\u9fa5]*$')           # A 开头 + 数字
        patt_code  = re.compile(r'^[A-Za-z\u4e00-\u9fa5]+[^A-Za-z\u4e00-\u9fa5]?\d+$')
        patt_ratio = re.compile(r'^1:\d+$')
        patt_digit = re.compile(r'^\d+$')
    
        for idx, t in enumerate(body):
            raw = _norm_txt(extract_text_content(t).replace(" ", ""))  # 去空格
            if idx == idx_num:
                newL = "ML-图纸序号"
            elif idx == idx_name:
                newL = "ML-图纸名称"
            elif patt_spec.fullmatch(raw):
                newL = "ML-图纸规格"
            elif patt_code.fullmatch(raw):
                newL = "ML-图纸编号"
            elif patt_ratio.fullmatch(raw):
                newL = "ML-图纸比例"
            elif patt_digit.fullmatch(raw):
                newL = "ML-其它"
            else:
                newL = "ML-其它1"
    
            try:
                t.Layer = newL
            except Exception:
                print(f"⚠️ 无法写层 {t.Handle}")
    

        # 2.2. 若已有正文 → 若是天正文字，重写

        rewrite_tianzheng_texts(body, xs=80)
        time.sleep(2)


        # 2.3. 若已有正文 → 对可能的偏移进行中心重矫正
        first_line = doc.HandleToObject(first_h)
        x0,y0,_ = first_line.GetBoundingBox()[0]
      
        for attempt in range(1, 4):
            try:
                align_table_items_to_center(
                    x0,
                    y0,
                    hangju,
                    intervals,
                    layernames=basic_layers,
                    tol_move=10
                )
                print(f"✅ 第 {attempt} 次尝试：对齐成功")
                break
            except Exception as e:
                print(f"⚠️ 第 {attempt} 次尝试失败：{e}")
                if attempt < 3:
                    time.sleep(1)
                else:
                    print(f"❌ 对齐失败，共尝试 3 次，最后错误：{e}")

        # 永远先删除同名组（若存在）
        try:
    
            gr=nametogroup('编目录参照集')
            gr.Delete()
    
        except Exception:
            pass  # 组不存在 → 会抛异常，直接忽略

        LB_canzhaoji = stc("ML-图纸序号")+stc("ML-图纸名称")+stc("ML-图纸规格")+stc("ML-图纸编号")+stc("ML-图纸比例")+stc("ML-其它")+stc("ML-其它1")

        add_objects_to_group("编目录参照集", LB_canzhaoji)
  
        handle_shouhang = []
        text_shouhang   = []
        com_shouhang =    []
        try:
    
            handle_shouhang = [ent.Handle for ent in LB_canzhaoji]    
            text_shouhang =   [extract_text_content(ent) for ent in LB_canzhaoji]
            com_shouhang =   [ent for ent in LB_canzhaoji]

        except:
        
            pass     
        
        zd["shouhang"] = (handle_shouhang,text_shouhang,com_shouhang)

        node("[节点2]目录首行示例文字句柄{} 文字内容{}",handle_shouhang,text_shouhang)


        # 取消选择状态
        cancel_cad_selection()
        # 保存
        try: savefile()
        except: pass
    
    
        # 更新字典并保存到电脑本地   
        try:
            dwg_name = current_dwg_basename()
        except Exception as e:
            time.sleep(1)
            dwg_name = current_dwg_basename()       
            node("[ERROR] current_dwg_basename 失败，再尝试一次：{}",e)
        try:
            save_print_dict_generic(dwg_name, zd)
            node("[INFO] 已保存到字典 -> {}",dwg_name)
        except Exception as e:
            time.sleep(1)
            save_print_dict_generic(dwg_name, zd)
            node("[ERROR] save_print_dict_generic 失败：{}",e)
    
            return zd


    cand = handles_to_coms(handles_label)

    if not cand:
        print("缺少标签行"); return False
    # 3.  系数修正
    try:
        if cand[0].ObjectName  in ["AcDbText", "AcDbMText"]:
            blxs =1
        else:
            pass
        first_line = doc.HandleToObject(first_h)
        x0,y0,_ = first_line.GetBoundingBox()[0]
        cy = y0-hangju/2
        chouyang = cand[0]
        print("chouyang:",chouyang.GetBoundingBox()[0][1],"cy原值:",cy)
        if abs(chouyang.GetBoundingBox()[0][1]-cy) < hangju/3:
            cy =cy-hangju
            print("cy随表头标签变动的修改值:",cy)

        else:
            pass

    except:

        print("过程出现错误，未作系数修正")

    #  列中心坐标
    cum = 0; LB_center=[]
    for w in intervals:
        prev=cum; cum+=w
        cx = x0+prev+w/2
        
        LB_center.append((cx,cy,0))

    # 4. 需要补写首行 → 先找真实标签行


    # —— 将候选分多行，按 y 坐标分组 —— #
    rows = {}
    for t in cand:
        y = t.GetBoundingBox()[0][1]
        key = min(rows,key=lambda k:abs(k-y)) if any(abs(k-y)<50 for k in rows) else y
        rows.setdefault(key, []).append(t)

    # 选择包含关键词 “序号” 或 “图幅/规格/比例” 的行
    def row_score(txts):
        s = "".join(_norm_txt(extract_text_content(t)) for t in txts)
        keys = ["序号","图幅","规格","比例","图号","编号"]
        return any(k in s for k in keys)
    label_row = next((r for r in rows.values() if row_score(r)),
                     sorted(rows.values(), key=lambda v: -v[0].GetBoundingBox()[0][1])[0])

    label_row = sorted(label_row, key=lambda t:t.GetBoundingBox()[0][0])

    # 4.1 取得样式 / 高度
    styles = []; heights=[]

    for t in label_row:
        raw_h = getattr(t,"Height", getattr(t,"TextHeight",250))
        try: raw_h=float(raw_h)
        except: raw_h=250.0
        heights.append(raw_h*blxs)

        sty = getattr(t,"TextStyle", None) if t.ObjectName.startswith("TDb") \
              else getattr(t,"StyleName", None)
        styles.append(sty or "Standard")

    # 4.2 归一化文本列表
    labels_norm = [_norm_txt(extract_text_content(t)) for t in label_row]

    # 找列索引
    col_name = intervals.index(max(intervals))
    try: col_num  = labels_norm.index(next(x for x in labels_norm if "编号" in x or "图号" in x))
    except StopIteration: col_num = min(i for i in range(len(labels_norm)) if i not in (0,col_name))
    try: col_spec = labels_norm.index(next(x for x in labels_norm if "图幅" in x or "规格" in x))
    except StopIteration: col_spec = min(i for i in range(len(labels_norm)) if i not in (0,col_name,col_num))
    try: col_ratio= labels_norm.index(next(x for x in labels_norm if "比例" in x))
    except StopIteration: col_ratio=None

    # 5. 写入文字
    # 5.1 序号

    height = heights[0]
    
    if height_bu != None:
        height=height_bu
    
    kuanduyinzi = 0.8
    
    if kuanduyinzi_bu != None:
        kuanduyinzi = kuanduyinzi_bu

    cx,cy,cz = LB_center[0]
    t0 = mp.AddText("00", vtpnt(cx,cy,cz), height)
    t0.StyleName = styles[0]; _set_center_alignment(t0,(cx,cy,cz))
    t0.ScaleFactor=kuanduyinzi
    t0.Layer="ML-图纸序号"

    # 5.2 名称
    cx,cy,cz = LB_center[col_name]
    tn = mp.AddText("图纸目录", vtpnt(cx,cy,cz), height)
    tn.StyleName = styles[col_name]; _set_center_alignment(tn,(cx,cy,cz))
    tn.ScaleFactor=kuanduyinzi
    tn.Layer="ML-图纸名称"

    # 5.3 编号
    cx,cy,cz = LB_center[col_num]
    tb = mp.AddText(f"{zhuanye_daihao}00", vtpnt(cx,cy,cz), height)
    tb.StyleName = styles[col_num]; _set_center_alignment(tb,(cx,cy,cz))
    tb.ScaleFactor=kuanduyinzi
    tb.Layer="ML-图纸编号"

    # 5.4 规格
    cx,cy,cz = LB_center[col_spec]
    ts = mp.AddText("A2", vtpnt(cx,cy,cz), height)
    ts.StyleName = styles[col_spec]; _set_center_alignment(ts,(cx,cy,cz))
    ts.ScaleFactor=kuanduyinzi
    ts.Layer="ML-图纸规格"

    # 5.5 比例（可选）
    if col_ratio is not None:
        cx,cy,cz = LB_center[col_ratio]
        txt_ratio = next((s for s in labels_norm if "1:" in s), "1:100")
        tr = mp.AddText(txt_ratio, vtpnt(cx,cy,cz), height)
        tr.StyleName = styles[col_ratio]; _set_center_alignment(tr,(cx,cy,cz))
        tr.ScaleFactor=kuanduyinzi   
        tr.Layer="ML-图纸比例"



    # 6. —— 存放到 “编目录参照集” 组 —— #
    
   
    # 永远先删除同名组（若存在）
    try:

        gr=nametogroup('编目录参照集')
        gr.Delete()

    except Exception:
        pass  # 组不存在 → 会抛异常，直接忽略
    
    # 重新汇总对象
    LB_canzhaoji = (
        stc("ML-图纸序号")
        + stc("ML-图纸名称")
        + stc("ML-图纸规格")
        + stc("ML-图纸编号")
        + stc("ML-图纸比例")
        + stc("ML-其它")
        + stc("ML-其它1")
    )

    add_objects_to_group("编目录参照集", LB_canzhaoji)

    # 取消选择状态
    cancel_cad_selection()
    # 保存
    try: savefile()
    except: pass

    handle_shouhang = []
    text_shouhang   = []
    com_shouhang =    []
    try:

        handle_shouhang = [ent.Handle for ent in LB_canzhaoji]    
        text_shouhang =   [extract_text_content(ent) for ent in LB_canzhaoji]
        com_shouhang =   [ent for ent in LB_canzhaoji]

    except:
    
        pass     
    
    zd["shouhang"] = (handle_shouhang,text_shouhang,com_shouhang)

    node("[节点2]目录首行示例文字句柄{} 文字内容{}",handle_shouhang,text_shouhang)


    # 更新字典并保存到电脑本地   
    try:
        dwg_name = current_dwg_basename()
    except Exception as e:
        time.sleep(1)
        dwg_name = current_dwg_basename()       
        node("[ERROR] current_dwg_basename 失败，再尝试一次：{}",e)
    try:
        save_print_dict_generic(dwg_name, zd)
        node("[INFO] 已保存到字典 -> {}",dwg_name)
    except Exception as e:
        time.sleep(1)
        save_print_dict_generic(dwg_name, zd)
        node("[ERROR] save_print_dict_generic 失败：{}",e)

    return zd




#  找到区域内的均匀表格

def find_feature_horizontal_groups(
        x1, y1, x2, y2,
      
        tol        = 50,
        min_count  = 10,
        ci=3,
        cha_Y      = 100,#y坐标水平线容差
        layer_name = "bianmulu_lp"
    ):
    """
    综合：分组→尾条并入→补底线→统一图层→汇总 6 项信息
    返回 List[ info ], 见文档说明
    """
    msp = doc.ModelSpace
    poly_types = ("AcDb2dPolyline", "AcDbPolyline")
    ents_all = select_entities_in_window(x1, y1, x2, y2, select_mode="_W")

    # — 1. 过滤水平线 —
    horiz = []
    for e in ents_all:
        if e.ObjectName not in ("AcDb2dPolyline", "AcDbPolyline", "AcDbLine"):
            continue
        min_pt, max_pt = e.GetBoundingBox()
        if abs(max_pt[1] - min_pt[1]) <= tol:
            horiz.append(e)
    if not horiz:
        return []

    # — 2. 按 x 分列 —
    cols = []
    for e in horiz:
        x0 = e.GetBoundingBox()[0][0]
        for col in cols:
            if abs(x0 - col[0].GetBoundingBox()[0][0]) <= tol:
                col.append(e)
                break
        else:
            cols.append([e])

    # — 3. 列内处理 —
    info_list = []
    def ly(o): return o.GetBoundingBox()[0][1]

    for col in cols:
        col_sorted = sort_coms_by_llcorner(col, cha_Y=cha_Y)
        if not col_sorted:
            continue

        pending_small = []
        cur, t = [col_sorted[0]], None

        # 内部：完成一个组，做补线/图层/信息
        def finalize_group(grp, spacing):
            nonlocal pending_small, info_list
            if len(grp) >= min_count:
                # 把 pending_small 尾条接上？
                if pending_small and spacing:
                    br = abs(ly(pending_small[-1]) - ly(grp[0]))
                    if abs(br - spacing) <= tol:
                        grp = [pending_small.pop()] + grp
                pending_small.clear()

                # ====== 补底线 ======
                if spacing:
                    xs = [p[0] for e in grp for p in e.GetBoundingBox()]
                    x_lo, x_hi = min(xs), max(xs)
                    y_min = min(ly(e) for e in grp)
                    y_zeng = y_min - spacing

                    already = any(abs(ly(e) - y_zeng) <= tol for e in grp)
                    has_poly = any(
                        e.ObjectName in poly_types and
                        abs(e.GetBoundingBox()[0][1] - y_zeng) <= tol
                        for e in ents_all
                    )
                    if (not already) and has_poly:
                        p1, p2 = vtpnt(x_lo, y_zeng,0), vtpnt(x_hi, y_zeng,0)
                        line_zeng = msp.AddLine(p1, p2)
                        grp.append(line_zeng)

                # ====== 统一图层 ======
                for e in grp:
                    try:
                        e.Layer = layer_name
                    except Exception:
                        pass

                # ====== 汇总 6 项 ======
                grp_sorted = sorted(grp, key=ly, reverse=True)
                top_line, bot_line = grp_sorted[0], grp_sorted[-1]
                handle_top = top_line.Handle
                handle_bot = bot_line.Handle
                # 若 spacing=0（理论上不应发生）则再测一次
                if not spacing and len(grp_sorted) >= 2:
                    spacing = abs(ly(grp_sorted[0]) - ly(grp_sorted[1]))

                # bbox 点
                left_pt  = bot_line.GetBoundingBox()[0]   # 左下
                right_pt = top_line.GetBoundingBox()[1]   # 右上

                intervals = compute_vertical_intervals(
                    left_pt[0], left_pt[1], right_pt[0], right_pt[1],
                    tol=tol, ci=ci, cha_Y=cha_Y
                )

                info = [
                    handle_top,
                    handle_bot,
                    spacing,
                    (left_pt, right_pt),
                    intervals,
                    grp
                ]
                info_list.append(info)
            else:
                pending_small.extend(grp)

        # —— 切分 —
        for prev, nxt in zip(col_sorted, col_sorted[1:]):
            dy = abs(ly(prev) - ly(nxt))
            if t is None:
                t, cur = dy, cur + [nxt]
            elif abs(dy - t) <= tol:
                cur.append(nxt)
            else:
                finalize_group(cur, t)
                cur, t = [nxt], None
        finalize_group(cur, t if t else 0)
        pending_small.clear()   # 丢弃残余

    cancel_cad_selection()

    return info_list

#  确定目录水平线簇和标签示例文字
@timeit
@debuggable 
def find_feature_horizontalgroups_and_deltext(x1, y1, x2, y2,
                                             tol=50, min_count=10,
                                             ci=3, cha_Y1=100,
                                             cha_Y2=50):
    """
    从最左边的水平线簇确定标签示例文字
    有些图纸目录模板的中文下面含有英文，还包含一些点。如果编目录流程获取的标签文字为空，就要进一步检查标签异常
    需要将那些异常文字从基本的texts去掉 

    """
    # —— 节点1：归一化 & 拾取实体 ——
    x_lo, x_hi = sorted((x1, x2))
    y_lo, y_hi = sorted((y1, y2))
    ents = select_entities_in_window(x_lo, y_lo, x_hi, y_hi, select_mode="_W")
    node("[ffd] 节点1 拾取窗口实体: {} 个", len(ents))

    # 候选线、多段线
    cands = [e for e in ents if e.ObjectName in ("AcDb2dPolyline", "AcDbPolyline", "AcDbLine")]
    node("[ffd] 节点1 候选线实体: {} 条", len(cands))



    # 文本实体
    texts = [e for e in ents if e.ObjectName in ("AcDbText", "AcDbMText", "TDbText", "TDbMText")]
    texts = [
        e for e in texts
        if not re.fullmatch(r"[A-Za-z]+", (t:=extract_text_content(e).replace(" ", "")))
        and not ("." in t and re.search(r"[A-Za-z]", t))
    ]


    node("[ffd] 节点1 文本实体: {} 条", len(texts))



    # —— 节点2：查找均匀水平线簇 ——
    node("[ffd] 节点2 调用 find_feature_horizontal_groups")
    results = find_feature_horizontal_groups(x1, y1, x2, y2,
                                            tol=tol, min_count=min_count,
                                            ci=ci, cha_Y=cha_Y1,layer_name = "bianmulu_lp"
                                            )
    node("[ffd] 节点2 返回簇组数: {} 组", len(results))
    if not results:
        node("[ffd] 未找到符合条件的水平线簇，退出。")
        return None

    # 选择最左侧簇
    min_idx = min(range(len(results)), key=lambda i: results[i][3][0][0])
    node("[ffd] 选定最左簇索引: {}", min_idx)
    seq_first, seq_final, hangju, bbox, intervals, feature_seq = results[min_idx]
    node("[ffd] 簇 bbox: {}", bbox)




    # —— 节点3：筛选目录文本 ——
    tr_x = bbox[1][0]
    tr_x_left = bbox[0][0]
    node("[ffd] x 左侧阈值 tr_x_left: {}", tr_x_left)
    node("[ffd] x 右侧阈值 tr_x: {}", tr_x)

    # 初步筛选：限定在水平带内，并在左右边界之间
    mulu_texts = [t for t in texts
                   if (bbox[0][1] < t.GetBoundingBox()[0][1] < bbox[1][1])
                   and (tr_x_left <= t.GetBoundingBox()[0][0] <= tr_x)]
    mulu_texts = sort_coms_by_llcorner(mulu_texts, cha_Y=cha_Y2)
    node("[ffd] 节点3 初步目录文本: {} 条", len(mulu_texts))

    # 判断是否全中文
    regex_chinese = re.compile(r'^[\u4e00-\u9fa5\s]+$')
    all_chinese = bool(mulu_texts) and all(
        regex_chinese.fullmatch(extract_text_content(t)) for t in mulu_texts
    )
    node("[ffd] 文本全中文: {}", all_chinese)

    label_texts, nonlabel_texts = [], []
    if all_chinese:
        label_texts = mulu_texts
        node("[ffd] 全中文分支，标签文本内容: {}", [extract_text_content(t) for t in label_texts])
        nonlabel_texts = []
    else:
        # 非全中文分支，正文候选同初步筛选结果
        nonlabel_texts = list(mulu_texts)
        node("[ffd] 非全中文分支，正文候选初始: {} 条", len(nonlabel_texts))

        # 扩展上方区域寻找标签行
        upper_zone = [t for t in texts
                      if (bbox[1][1] < t.GetBoundingBox()[0][1] < bbox[1][1] + 2*hangju)
                      and (tr_x_left <= t.GetBoundingBox()[0][0] <= tr_x)]
        upper_zone = sort_coms_by_llcorner(upper_zone, cha_Y=cha_Y2)
        node("[ffd] upper_zone 文本: {} 条", len(upper_zone))

        # 按行分组并选最后一行为标签
        rows = []
        for t in upper_zone:
            y = t.GetBoundingBox()[0][1]
            if not rows or abs(rows[-1][0].GetBoundingBox()[0][1] - y) > cha_Y2:
                rows.append([t])
            else:
                rows[-1].append(t)
        for idx, row in enumerate(rows):
            node("[ffd] 第 {} 行文本: {} 条, 内容: {}", idx, len(row), [extract_text_content(tt) for tt in row])
        if rows and all(regex_chinese.fullmatch(extract_text_content(tt)) for tt in rows[-1]):
            label_texts = rows[-1]
            node("[ffd] 选定最后一行作为标签内容: {}", [extract_text_content(tt) for tt in label_texts])

    # 对正文首例再次限定在左右边界之间
    nonlabel_texts = [t for t in nonlabel_texts
                      if tr_x_left <= t.GetBoundingBox()[0][0] <= tr_x]
    node("[ffd] 限定后正文最终内容: {}", [extract_text_content(t) for t in nonlabel_texts])

    # 最终排序
    label_texts = sort_coms_by_llcorner(label_texts, cha_Y=cha_Y2)
    nonlabel_texts = sort_coms_by_llcorner(nonlabel_texts, cha_Y=cha_Y2)

    node("[ffd] 标签文本内容: {}", [extract_text_content(t) for t in label_texts])
    node("[ffd] 正文最终内容: {}", [extract_text_content(t) for t in nonlabel_texts])

    # 返回结果 handles
    handles_label = [t.Handle for t in label_texts]
    handles_nonlabel = [t.Handle for t in nonlabel_texts]


    cancel_cad_selection()

    return (seq_first, seq_final, hangju, bbox,
            intervals,handles_label, handles_nonlabel)


#分析表格横向间距
def compute_vertical_intervals(x1, y1, x2, y2, tol=50,ci = 3,cha_Y=100):
    """
    在给定矩形区域 (x1,y1)-(x2,y2) 内：
      1. 用 select_entities_in_window(..., select_mode="_C") 拾取所有实体；
      2. 筛选出 AcDb2dPolyline、AcDbPolyline、AcDbLine 类型且
         外包盒宽度 abs(maxX-minX) <= tol 的“垂直”线集合；
      3. 对它们按 sort_coms_by_llcorner(..., cha_Y=100) 排序；
      4. 取每条线的 minX 作为其横坐标，连同区域边界 x_lo,x_hi 
         合并后升序排序；
      5. 计算相邻横坐标差值列表并返回。

    参数：
      x1,y1,x2,y2 : float — 矩形对角坐标
      tol         : float — 判断垂直的宽度容差 & 去重容差

    返回：
      List[float] — 相邻 X 间距列表
    """
    # 坐标归一化
    x_lo, x_hi = sorted((x1, x2))
    y_lo, y_hi = sorted((y1, y2))

    # 1. 窗口交叉模式拾取
    ents = select_entities_in_window(x_lo, y_lo, x_hi, y_hi, select_mode="_C")

    # 2. 筛垂直线
    verts = []
    for e in ents:
        if e.ObjectName not in ("AcDb2dPolyline", "AcDbPolyline", "AcDbLine"):
            continue
        try:
            mn, mx = e.GetBoundingBox()
        except:
            continue
        if abs(mx[0] - mn[0]) <= tol:
            verts.append(e)

    # 3. 排序
    verts_sorted = sort_coms_by_llcorner(verts, cha_Y=cha_Y)

    set_layer_with_retry(verts_sorted, "bianmulu_lp", ci=ci)


    # 4. 收集 X 坐标 + 区域边界
    xs = [mn := e.GetBoundingBox()[0][0] for e in verts_sorted]
    xs = [x_lo] + xs + [x_hi]
    xs_unique = sorted(set(xs))

    # 5. 计算相邻差值
    intervals = [abs(xs_unique[i+1] - xs_unique[i])
                 for i in range(len(xs_unique)-1)]
    return intervals

# 编目录图层清理
"""
将一组目录文字设成不同的对齐方式则排序将会混乱，annotate_directory_template也会混乱

-873051.0818272297, 8613.201188054933, -813651.0818272297, 50613.20118805493

blxs=80=100*0,80是天正文字的经验系数 ，按100过大

清空图层有问题

"""
# ------------------------------------------------------------------ #
def _norm_txt(txt: str) -> str:            # 去掉所有空白
    return re.sub(r'\s+', '', txt)

def _set_center_alignment(txt_obj, pt):    # 中间
    if hasattr(txt_obj, "Alignment"):      # 单行
        try:
            txt_obj.Alignment = 4
            txt_obj.TextAlignmentPoint = vtpnt(*pt)
        except: pass
    elif hasattr(txt_obj, "AttachmentPoint"):  # 多行
        try:
            txt_obj.AttachmentPoint = 5
        except: pass

def clean_directory_layers(
    basic_layers=["ML-图纸序号","ML-图纸名称","ML-图纸规格","ML-图纸编号","ML-图纸比例","ML-其它"],
    ):
    """
    清理指定图层中的对象：
      对于 basic_layers 中的每个图层，
      删除该图层上所有不在 lb_canzhaoji_groups 列表中的对象。
      每个删除操作若出错，最多重试 3 次，每次失败后等待 1 秒。

    参数
    ----
    basic_layers : List[str]
        需清理的图层名列表
    lb_canzhaoji_groups : List
        保留对象的 文字列表，其它对象将被删除

    返回
    ----
    bool
        清理完成后返回 True
    """

    from win32com import client as dyn

    lb_canzhaoji_groups = get_com_from_groupname("编目录参照集")

    for layer in basic_layers:
        # 获取该图层上的所有对象
        objs = stc(layer)  
        for sel in objs:
            try:
                obj = dyn.Dispatch(sel._oleobj_)
            except Exception:
                continue
            
            # 如果不在保留列表，就删除
            if obj not in lb_canzhaoji_groups and len(lb_canzhaoji_groups)>0:
                for attempt in range(1, 4):
                    try:
                        obj.Delete()                        
                        break
                    except Exception as e:
                        print(f"⚠ 删除对象第 {attempt} 次失败: {e}")
                        if attempt < 3:
                            time.sleep(1)
                        else:
                            print(f"❌ 删除失败，跳过")
    return True


# 重写天正文字

def rewrite_tianzheng_texts(text_entities, xs=80):
    """
    重写天正文字为 CAD 单行文字：
      节点1：分离出天正文字 (TDbText/TDbMText) 与 CAD 文字
      节点2：获取天正文字样式 (TextStyle)、字高、内容、图层、包围盒中心和左下点
      节点3：插入 CAD 单行文字 (新高 = 原高 * xs)，居中对齐 (4)，并移动到原中心位置
      节点4：删除原天正文字
    """

    # —— 节点1：分离天正文字与 CAD 文字 ——
    node("[节点1] 输入实体总数 = {}", len(text_entities))
    tian_list, cad_list = [], []
    for ent in text_entities:
        obj = dyn.Dispatch(ent._oleobj_) if hasattr(ent, "_oleobj_") else ent
        cls = obj.ObjectName
        if cls in ("TDbText", "TDbMText"):
            tian_list.append(obj)
        else:
            cad_list.append(obj)
    node("[节点1] 天正文字数量 = {}，CAD 文字数量 = {}", len(tian_list), len(cad_list))

    # 节点2~4：处理每条天正文字
    for idx, obj in enumerate(tian_list, start=1):
        # 读取 TextStyle 属性（天正文字专用）
        style   = obj.TextStyle
        height  = float(obj.Height)
        new_h   = height * xs
        layer   = obj.Layer
        content = extract_text_content(obj)

        scale_x = obj.XScale
        min_pt, max_pt = obj.GetBoundingBox()
        old_min_x, old_min_y, old_min_z = min_pt
        old_cx = (min_pt[0] + max_pt[0]) / 2
        old_cy = (min_pt[1] + max_pt[1]) / 2

        node("[节点2-{}] TextStyle={} 原高={} 新高={} Layer={} 内容='{}'",
             idx, style, height, new_h, layer, content)

        # 插入 CAD 单行文字，初始放在原包围盒左下角
        ins_pt = vtpnt(old_min_x, old_min_y, old_min_z)
        new_txt = mp.AddText(content, ins_pt, new_h)
        # 使用 CAD Text 的 StyleName 设置样式
        new_txt.StyleName            = style
        new_txt.Layer                = layer
        new_txt.Alignment            = 4
        new_txt.TextAlignmentPoint   = ins_pt

        new_txt.ScaleFactor          = scale_x
        # —— 节点3：计算中心偏移并移动 ——
        new_min, new_max = new_txt.GetBoundingBox()
        new_cx = (new_min[0] + new_max[0]) / 2
        new_cy = (new_min[1] + new_max[1]) / 2
        dx = old_cx - new_cx
        dy = old_cy - new_cy
        new_txt.Move(vtpnt(0, 0, 0), vtpnt(dx, dy, 0))
        node("[节点3-{}] 新句柄={} 偏移(dx,dy)=({:.3f},{:.3f})", idx, new_txt.Handle, dx, dy)

        # —— 节点4：删除原天正文字 ——
        old_handle = obj.Handle
        obj.Delete()
        node("[节点4-{}] 删除原天正文字句柄={}", idx, old_handle)

    node("✅ rewrite_tianzheng_texts 完成，共处理 {} 条天正文字", len(tian_list))
    return True

# 校核首行示例文字到表格中心


def align_table_items_to_center(x0: float,
                               y0: float,
                               hangju: float,
                               intervals: list,
                               layernames: list=["ML-图纸序号","ML-图纸名称","ML-图纸规格","ML-图纸编号","ML-图纸比例","ML-其它"],
                               tol_move: float = 10):
    """将 `layernames` 图层首对象对齐到最近的表格中心。

    * `intervals[i]` 为第 *i* 列宽度。
    * 列中心 X = `x0 + prefix_sum - width/2`。
    * 表格有 **两行** 中心：`y0±hangju`。
    * 对每个对象：在 *所有* 中心点(两行 × 列数)中找到最近的一个并移动。
    """
    node("[ati] intervals={}  layernames={}", intervals, layernames)
    # 前缀和
    prefix_sums, acc = [], 0.0
    for w in intervals:
        acc += w
        prefix_sums.append(acc)

    # 列中心 X
    col_centers_x = [x0 + s - w/2 for s, w in zip(prefix_sums, intervals)]
    node("[ati] 列中心 X: {}", col_centers_x)

    # 生成两行中心坐标
    centers = [(x, y0 + hangju/2, 0.0) for x in col_centers_x] + \
              [(x, y0 - hangju/2, 0.0) for x in col_centers_x]
    node("[ati] 所有中心坐标: {}", centers)

    # 遍历图层
    for layer in layernames:
        ents = stc(layer)
        if not ents:
            node("[ati] ⛔ 图层 '{}' 无对象，跳过。", layer)
            continue
        ent = ents[0]
        try:
            mn, mx = ent.GetBoundingBox()
            x_c = (mn[0] + mx[0]) / 2
            y_c = (mn[1] + mx[1]) / 2
        except Exception as err:
            node("[ati] ❌ Handle={} bbox 失败: {}", getattr(ent,'Handle','?'), err)
            continue

        # 最近中心
        tgt = min(centers, key=lambda p: math.hypot(p[0]-x_c, p[1]-y_c))
        x_j, y_j, _ = tgt
        node("[ati] '{}' 当前({:.2f},{:.2f}) → 最近({:.2f},{:.2f})", layer, x_c, y_c, x_j, y_j)

        dx, dy = x_j - x_c, y_j - y_c
        if abs(dx) <= tol_move and abs(dy) <= tol_move:
            node("[ati] 位移({:.2f},{:.2f}) < tol {}, 跳过。", dx, dy, tol_move)
            continue
        try:
            ent.Move(vtpnt(0,0,0), vtpnt(dx,dy,0))
            node("[ati] ✅ Handle={} 已移动 ({:.2f},{:.2f})", ent.Handle, dx, dy)
        except Exception as err:
            node("[ati] ❌ Handle={} 移动失败: {}", ent.Handle, err)
    node("[ati] 处理完毕。")






#&&&% （五） 确定重新起行参数
#&&%─────────────────────────────────────
        

@timeit

@debuggable
def zhuanchang_mulu_canshu(zd,
                           tol: int = 50,
                           min_count: int = 10,
                           ci: int = 3,
                           cha_Y1: int = 100,
                           ):
    """


    提取所有目录图纸的目录列水平簇外包盒
    检查目录水平线误差，可能积少成多产生错误

    参数
    ------
    zd : dict            # 外部收集的中间数据
    tol, min_count, ci…  # 传递给 find_feature_horizontal_groups 的参数

    返回
    ------
    List[((x1,y1,z),(x2,y2,z))]  # 按 X 从左到右排好序的外包盒列表


    column_shifts —— 目录书写换列间距

    """


    # —— 节点1：读取标志位 ——
    biaoshi = zd.get("flag", (None, None, None))[2]
    node("[节点1] flag = {}", biaoshi)
    if not isinstance(biaoshi, int):
        node("  ✖ flag 无效，退出")
        return None

    # —— 节点1.1：预设图纸目录张数 ——

    zhizhangshu_tuzhimulu = 0

    if biaoshi is not None:
        zhizhangshu_tuzhimulu = 1


    # —— 节点2：获取打印线句柄与 prinfo ——
    handles = zd.get("dyx_handles", [])
    if biaoshi >= len(handles):
        node("  ✖ biaoshi 越界，退出")
        return None
    handle_line1 = handles[biaoshi]
    handle_line2 = handles[biaoshi + 1] if biaoshi + 1 < len(handles) else None
    info2 = zd.get("prinfo", {}).get(handle_line2)
    node("[节点2] line1={}, line2={}, prinfo2={} ", handle_line1, handle_line2, info2)
    if not info2 or info2[2] != "A2":
        node("  ✖ 第二张图不是 A2，退出")
        return None

    # —— 节点3：COM 对象安全转换 ——
    def safe_obj(h):
        try:
            return HandleToObject(h)
        except Exception as err:
            node("[节点3] HandleToObject({}) 失败: {}", h, err)
            return None

    line1 = safe_obj(handle_line1)
    frame2 = safe_obj(handle_line2)
    if not line1 or not frame2:
        node("  ✖ COM 对象获取失败，退出")
        return None

    # —— 节点4：获取两个打印框的 bbox ——
    try:
        mn1, mx1 = line1.GetBoundingBox()
        x1, y1 = mn1[0], mn1[1]
        x2, y2 = mx1[0], mx1[1]
    except Exception as err:
        node("[节点4] line1.GetBoundingBox 失败: {}", err)
        return None

    mn2, mx2 = frame2.GetBoundingBox()
    x1_f, y1_f = mn2[0], mn2[1]
    x2_f, y2_f = mx2[0], mx2[1]
    node("[节点4] bbox1=(({:.3f},{:.3f}),({:.3f},{:.3f}))  bbox2=(({:.3f},{:.3f}),({:.3f},{:.3f}))", x1,y1,x2,y2, x1_f,y1_f,x2_f,y2_f)

    zhizhangshu_tuzhimulu = 2

    node("目录纸张数量 {} 张",zhizhangshu_tuzhimulu)
    # —— 节点5：分别查找水平簇 ——
    groups1 = find_feature_horizontal_groups(x1, y1, x2, y2,
                                             tol=tol, min_count=min_count,
                                             ci=ci, cha_Y=cha_Y1)
    groups2 = find_feature_horizontal_groups(x1_f, y1_f, x2_f, y2_f,
                                             tol=tol, min_count=min_count,
                                             ci=ci, cha_Y=cha_Y1)
    node("[节点5] groups1={}  groups2={} (长度) ", len(groups1), len(groups2))

    # 提取外包盒 (元素[3])
    bboxes = [g[3] for g in groups1] + [g[3] for g in groups2]
    bboxes_sorted = sorted(bboxes, key=lambda b: b[0][0])
    node("[节点5] 排序后外包盒列表: {}", bboxes_sorted)


    # 将水平线簇图层设为"bianmulu_lp"

    ensure_layer_current("bianmulu_lp")

    for x in groups1:

        LB = x[-1]

        for obj in LB:

            try:
                obj.Layer = "bianmulu_lp"                               

            except:
                time.sleep(1)
                obj.Layer = "bianmulu_lp" 

    for x in groups2:

        LB = x[-1]

        for obj in LB:

            try:
                obj.Layer = "bianmulu_lp"                               

            except:
                time.sleep(1)
                obj.Layer = "bianmulu_lp" 


    # 1 计算换列参数
    xs_sorted,heights = get_sorted_lower_left_x_and_heights(bboxes_sorted)

    first_x = xs_sorted[0]
    column_shifts = [(x - first_x, 0) for x in xs_sorted]

    # 1）水平线簇表格的行距参数，从左到右，即按x的升序排列
    hangju_values = [sublist[2] for sublist in groups1 + groups2]

    # 2）每个水平线簇表格的高度参数，它就是heights列表
    

    # 2 确定目录图纸数量 zhizhangshu_tuzhimulu

    # 3 修改字典zd
                    
    zd["tuzhimulu"] = [column_shifts,hangju_values,heights,zhizhangshu_tuzhimulu]

    return  zd 





# 🧠 辅助函数： 
def get_sorted_lower_left_x_and_heights(bboxes):
    """
    从外包盒列表中提取每个盒子的左下角点 x 坐标，
    并计算对应的高度（y_max - y_min）的绝对值，
    最后按 x 坐标升序返回两个列表。

    :param bboxes: List of bounding boxes,
                   每个元素形如 ((x_min, y_min, z), (x_max, y_max, z))
    :return: (xs_sorted, heights)
             xs_sorted: 按升序排列的 x_min 列表
             heights:  对应每个 x_min 的高度列表
    """
    # 构造 (x_min, height) 的列表
    xs_heights = [
        (box[0][0], abs(box[1][1] - box[0][1]))
        for box in bboxes
    ]
    # 按 x_min 升序排序
    xs_heights.sort(key=lambda item: item[0])

    # 解压为两个列表
    xs_sorted   = [item[0] for item in xs_heights]
    heights     = [item[1] for item in xs_heights]

    return xs_sorted, heights



#&&&% （六） 写前文字内容整理
#&&%─────────────────────────────────────
        

@timeit
@debuggable
def word_processing_before_writing(zd, fujia: str = "", zz="", bz=""):
    """
    整理图纸相关列表并写回到 zd：
        • zz     : 不明参数
        • bz     : 对应备注
        • fujia  : 给规格 LC_gg 追加的后缀
    返回更新后的 zd

    没有恰当的目录标签名称 系统不会有对应的首行示例值 要强制按用户自己的意愿，就要自己写出首行示例文字
    系统只考虑[ML-图纸名称,ML-图纸序号,ML-图纸编号,ML-图纸规格,ML-图纸比例,ML-其它,ML-其它1"]7个图层对应的内容
    分别对应实际不同名字的图纸名称，序号，图纸编号，图纸规格，图纸比例，未明事项，备注。

    """
    li()

    enable_debug()
    # ——— 1) 图纸名称列表（去掉第一张） ———
    LC = zd.get("tuzhimingcheng", [])
    LC = LC[1:]

    # ——— 2) 图纸编号列表 LC_th ———
    LC_th = zd.get("xuhao_texts", [])
    LC_th = LC_th[1:]

    # ——— 3) 根据 dyx_handles / prinfo 生成图幅、比例、规格 ———
    dyx_h = zd.get("dyx_handles", [])[1:]
    prinfo = zd.get("prinfo", {})
    LC_tf, LC_ratio, LC_gg = [], [], []
    for i in range(len(LC)):
        handle = dyx_h[i] if i < len(dyx_h) else None
        info = prinfo.get(handle, (None, None, None, None))
        LC_tf.append(info[0])       # 图幅
        LC_ratio.append(info[1])    # 比例
        LC_gg.append(info[2])       # 规格

    # 给规格附加后缀（如果有）
    LC_gg = [ent + fujia for ent in LC_gg]
    # 跳过首行示例
    LC_tf = LC_tf[1:]
    LC_ratio = LC_ratio[1:]
    LC_gg = LC_gg[1:]

    # ——— 4) 未明列表 LC_zz（全部置 zz，预备） ———
    LC_zz = [zz for _ in LC]
    LC_zz = LC_zz[1:]

    # ——— 5) 备注列表 LC_bz（全部置 bz，预备） ———
    LC_bz = [bz for _ in LC]
    LC_bz = LC_bz[1:]

    # —— 合并到一级键 "bianxieneirong_mulu" 下 ——
    zd_bm = {
        "LC": LC,
        "LC_th": LC_th,
        "LC_tf": LC_tf,
        "LC_ratio": LC_ratio,
        "LC_gg": LC_gg,
        "LC_zz": LC_zz,
        "LC_bz": LC_bz,
    }
    zd["bianxieneirong_mulu"] = zd_bm

    node("LC {} LC_th {} LC_tf {} LC_ratio {} LC_gg {} LC_zz {} LC_bz {}",
         LC, LC_th, LC_tf, LC_ratio, LC_gg, LC_zz, LC_bz)

    # 更新字典并保存到电脑本地
    try:
        dwg_name = current_dwg_basename()
    except Exception as e:
        time.sleep(1)
        dwg_name = current_dwg_basename()
        node("[ERROR] current_dwg_basename 失败，再尝试一次：{}", e)

    try:
        save_print_dict_generic(dwg_name, zd)
        node("[INFO] 已保存到字典 -> {}", dwg_name)
    except Exception as e:
        time.sleep(1)
        save_print_dict_generic(dwg_name, zd)
        node("[ERROR] save_print_dict_generic 失败：{}", e)

    disable_debug()
    return zd




#&&&% （七） 模板上写目录
#&&%─────────────────────────────────────

#&&&%  ▶ F1 按字典信息编写

@timeit            
@debuggable 
def write_directory_template(
    zd,
    hangju=1000.0,
    n=28,                                   # 每列行数
    column_shifts=[(0, 0.0), (68198.12, 0.0), (92973.95, 0.0)],        # 每个数据对应一排，其值为相对第一排左侧起点的距离
    split_threshold=10000.0,                #文本分裂阈值，应该从水平线簇的最大横向分割间距得到，不能作为输入参数
    y_tol=300,
    P_start=(0, 0, 0)                       #LC 书写参照的相对原始基点的偏移向量 例如从第二行开始书写
):
    """
    输入字典，返回字典。所有书写的文字还可以单独控制修改

    在预期的ML图层上选择到非空对象才会启动列表内容的书写 这也是目录初始化的意义作用之一
    -----------------------------------------------------------------------------------
    如果CAD文件中首行示例文字中包含了ML-其它，ML-其它1的对象，则后续就是复制它们。否则
    按Excel控制文件。如果Excel控制文件也是空，那么也不会书写
    --------------------------------------------------

    批量插入目录文字，并删除分裂文本之间的水平分隔线
    --------------------------------------------------
    - column_shifts : 依列号提供 “参照起点” (dx, dy)。列号超出长度时按模循环
    - ML-比例 / ML-其它 图层可能无参考文字，函数自动跳过


    """

    enable_debug()

    # ——————————— 从字典获取准备好的书写内容列表 ———————————
    
    zd["bianxieneirong_mulu"] = zd_bm
    
    LC=zd_bm["LC"]
    
    LC_th=zd_bm["LC_th"]
    
    LC_tf=zd_bm["LC_tf"]
    
    LC_ratio=zd_bm["LC_ratio"]
    
    LC_gg=zd_bm["LC_gg"]
    
    LC_zz=zd_bm["LC_zz"]
    
    LC_bz=zd_bm["LC_bz"]
    

    from win32com import client as dyn

    # ——————————— 工具 ———————————
    def _get_shift(col_idx):
        """根据列号取得 (dx, dy)，超长时循环取模"""
        return column_shifts[col_idx % len(column_shifts)]

    # ————— 0. 预检查 —————
    for lay in ["ML-图纸序号","ML-图纸名称","ML-图纸规格",
                "ML-图纸编号","ML-图纸比例","ML-其它","ML-其它1","bianmulu_lp"]:
        ensure_layer_current(lay)

    clean_directory_layers(
        basic_layers=["ML-图纸序号","ML-图纸名称",
                      "ML-图纸规格","ML-图纸编号",
                      "ML-图纸比例","ML-其它","ML-其它1"]     
    )

    time.sleep(2)

    # ————— 1. 参考文字获取参数 —————
    refs = stc("ML-图纸名称")
    if not refs:
        node("⚠ 未找到 ML-图纸名称，退出"); return False
    ref_name = max([dyn.Dispatch(e._oleobj_) for e in refs],
                   key=lambda e: bbox_center_3(e)[1])
    cx0, cy0, cz0 = bbox_center_3(ref_name)
    cx0 += P_start[0]; cy0 += P_start[1]; cz0 += P_start[2]

    node("⭐ 名称参考 Handle={} Center=({:.1f},{:.1f})", ref_name.Handle, cx0, cy0)
    style, size = ref_name.StyleName, float(ref_name.Height)
    color, just = int(ref_name.Color), int(ref_name.Alignment)
    sf          = float(ref_name.ScaleFactor)
    # 第 0 列起点
    dx0, dy0 = _get_shift(0)
    base_x = cx0 + dx0
    base_y = cy0 + dy0 - hangju

    row_cnt = col_cnt = 0
    lb_fenlie = []    # 用于存放 (idx, [com1, com2]) 的分裂记录
    pos_map, split_pairs = {}, []

    # ——— 统一的行号函数 ———
    def next_pt():
        """
        每调用一次就占用一行：
          • 到 n 行自动换列
          • 返回 (x, y, z, col_idx)
        """
        nonlocal row_cnt, col_cnt
        if row_cnt >= n:
            row_cnt = 0
            col_cnt += 1
            node("🔀 换列 col_cnt={}",col_cnt)

        dx, dy = _get_shift(col_cnt)
        x = cx0 + dx
        y = (cy0 - hangju) + dy - row_cnt * hangju
        row_cnt += 1
        return x, y, cz0, col_cnt

    # =============== 2. 图纸名称 ===============
    node("▶[节点1] 插入图纸名称 共 {} 条",len(LC))

    LB_coms_tuzhimingcheng = []
    for idx, raw in enumerate(LC, 1):
        # 2-0) 宽度测量
        tmp = mp.AddText(raw, vtpnt(base_x, base_y, cz0), size)
        tmp.StyleName = style
        w = celiang_wenzichangdu(tmp)
        tmp.Delete()

        # 2-1) 单次拆行（超阈值且含“、”）
        lines = [raw]
        if w > split_threshold and "、" in raw:
            pos = min([m.start() for m in re.finditer("、", raw)],
                      key=lambda p: abs(p - len(raw)//2))
            lines = [raw[:pos+1], raw[pos+1:]]

        Ys, ents_this = [], []
        first_col = None
        for line in lines:
            x, y, z, col_idx = next_pt()
            pt = vtpnt(x, y, z)
            ent = mp.AddText(line, pt, size)
            ent.Layer, ent.StyleName = "ML-图纸名称", style
            ent.Color, ent.Alignment = color, just
            ent.ScaleFactor          = sf
            ent.TextAlignmentPoint   = pt
            cx, cy, _ = bbox_center_3(ent)

            Ys.append(cy)
            first_col = first_col or col_idx
            tag = "▶ LC[0]" if idx == 1 and line == lines[0] else "  •"

            node("{} {} Handle={} Center=({:.1f},{:.1f})", tag, line, ent.Handle, cx, cy)

            ents_this.append(ent)

        for ob in ents_this:
            LB_coms_tuzhimingcheng.append(ob) 

        pos_map[idx] = (Ys, first_col)
        if len(ents_this) == 2:
            node("分裂序号：{}", idx)
            split_pairs.append(tuple(ents_this))

            # 把当前 idx 及其拆分后的两个 COM 对象放入 lb_fenlie
            lb_fenlie.append((idx, ents_this))


    # =============== 3. 序号 ===============
    seq_objs = stc("ML-图纸序号")
    if seq_objs:
        ref_seq = max([dyn.Dispatch(e._oleobj_) for e in seq_objs],
                      key=lambda e: bbox_center_3(e)[1])
        seq_x0, _, _ = bbox_center_3(ref_seq)
        seq_x0 += P_start[0]

        node("⭐ [节点2]序号参考 Handle={} X={:.1f}", ref_seq.Handle, seq_x0)

        for idx in range(1, len(LC) + 1):
            Ys, col_idx = pos_map.get(idx, ([], 0))
            if not Ys: continue
            y_mid = sum(Ys)/len(Ys)
            dx, _ = _get_shift(col_idx)
            x  = seq_x0 + dx
            pt = vtpnt(x, y_mid, cz0)
            ent = mp.AddText(f"{idx:02d}", pt, size)
            ent.Layer, ent.StyleName = "ML-图纸序号", style
            ent.Color, ent.Alignment = color, just
            ent.ScaleFactor          = sf
            ent.TextAlignmentPoint   = pt
            cx, cy, _ = bbox_center_3(ent)

            node("✏ 序号[{}] Handle={} Center=({:.1f},{:.1f})", idx, ent.Handle, cx, cy)


    else:
        node("ℹ ML-图纸序号 为空，跳过")

    # =============== 4. 编号 / 规格 / 比例 / 其它/ 其它1 ===============
    def write_other(values, layer_name):
        if not values:
            node("ℹ {} 列表为空，跳过",layer_name); return
        objs = stc(layer_name)
        if not objs:
            node("⚠ {} 无参考文字，跳过",layer_name); return
        ref = max([dyn.Dispatch(e._oleobj_) for e in objs],
                  key=lambda e: bbox_center_3(e)[1])
        base_x, _, _ = bbox_center_3(ref)
        base_x += P_start[0]
        node("⭐ [节点3]{} 参考 Handle={} X={:.1f}", layer_name, ref.Handle, base_x)
        for idx, val in enumerate(values, 1):
            if str(val).strip() == "": continue
            Ys, col_idx = pos_map.get(idx, ([], 0))
            if not Ys: continue
            y_mid = sum(Ys)/len(Ys)
            dx, _ = _get_shift(col_idx)
            x  = base_x + dx
            pt = vtpnt(x, y_mid, cz0)
            ent = mp.AddText(str(val), pt, size)
            ent.Layer, ent.StyleName = layer_name, style
            ent.Color, ent.Alignment = color, just
            ent.ScaleFactor          = sf
            ent.TextAlignmentPoint   = pt
            try:
                cx, cy, _ = bbox_center_3(ent)

                node("✏ {}[{}] Handle={} Center=({:.1f},{:.1f})", layer_name, idx, ent.Handle, cx, cy)

            except Exception:
                node("✏ {}[{}] Handle={} (空文字，无 bbox)",layer_name,idx,ent.Handle)

    # —— 书写其余图层文字 ——
  
    qita_objs = stc("ML-其它")#对应未明参数

    if len(qita_objs)==0:
        LC_other = LC_zz
    else:
        fore_text = extract_text_content(qita_objs[0])
        LC_other = [fore_text for _ in LC]
        write_other(LC_other, "ML-其它")            
    
    qita_objs1 = stc("ML-其它1")#对应备注、也可是其它未明参数

    if len(qita_objs1)==0:
        LC_other = LC_bz
    else:
        fore_text = extract_text_content(qita_objs1[0])
        LC_other = [fore_text for _ in LC]
        write_other(LC_other, "ML-其它1")     

    th_objs = stc("ML-图纸编号")
    if len(th_objs)==0:
        pass
    else:
        write_other(LC_th, "ML-图纸编号")            

    gg_objs = stc("ML-图纸规格")
    if len(gg_objs)==0:
        pass
    else:
        write_other(LC_gg, "ML-图纸规格")            
    

    ra_objs = stc("ML-图纸比例")
    if len(ra_objs)==0:
        pass
    else:
        write_other(LC_ratio, "ML-图纸比例")            


    # =========================================
    # 5. 精准标记“分裂行”之间的水平线（按 lb_fenlie ）
    # =========================================
    if lb_fenlie:
        feng_raw = stc("bianmulu_lp")
        if not feng_raw:
            node("ℹ bianmulu_lp 图层无对象，跳过分裂线处理")
        else:
            # —— 预筛：真正的水平线 —— #
            hori = []
            for ent in feng_raw:
                try:
                    obj = dyn.Dispatch(ent._oleobj_)
                except:
                    continue
                if obj.ObjectName not in ("AcDbLine","AcDbPolyline","AcDb2dPolyline"):
                    continue
                (x1b, y1b, _), (x2b, y2b, _) = obj.GetBoundingBox()
                if abs(y2b - y1b) <= 20:
                    hori.append(obj)
    
            node("🗑 待处理 {} 处分裂行，候选水平线 {} 条",len(lb_fenlie),len(hori))
            used = set()
            for idx, (A, B) in lb_fenlie:
                # ① 两行中心
                x1, y1, _ = bbox_center_3(A)
                x2, y2, _ = bbox_center_3(B)
   
                node("分裂序号：{} → 上行({:.1f},{:.1f}) 下行({:.1f},{:.1f})", idx,  x1, y1, x2, y2)



                lo_y, hi_y = sorted((y1, y2))
                lo_x, hi_x = sorted((x1, x2))
    
                # ② 再筛：y 在区间内 & x 跨度覆盖
                cands = []
                for obj in hori:
                    if obj.Handle in used:
                        continue
                    ll, ur = obj.GetBoundingBox()
                    llx, lly = ll[0], ll[1]
                    urx = ur[0]
                    if lo_y < lly < hi_y and llx <= lo_x and urx >= hi_x:
                        cands.append(obj)
    
                if not cands:
                    node("  ⚠️ 序号 {} 找到覆盖行间的水平线",idx)
                    continue
    
                # 理论上只有一条或多条重叠，取第一
   
                # ③ 标红并移层
                try:
                    
                    for ob in cands:                   
                        ob.Color = 1
                        used.add(ob.Handle)
                    move_entity_to_frozen_layer(cands)

                except:
                    try:
  
                        for ob in cands:                   
                            ob.Color = 1
                            used.add(ob.Handle)
                        move_entity_to_frozen_layer(cands)

                    except:
                        node("    ⚠️ 上色失败 Handle={}",cands[0].Handle)
    
                y_line = cands[0].GetBoundingBox()[0][1]
                node("  🔴 已标记水平线 Handle={} y={:.1f}", cands[0].Handle, y_line)        
    
    time.sleep(2)
    ensure_layer("测试辅助")

    node("✅ 完美收官：所有列插入完成，分裂线已清除。")

    # —— 图纸名称的com实体对象信息存入字典 ——

    handles_LB_coms_tuzhimingcheng = [ent.Handle for ent in LB_coms_tuzhimingcheng]

    zd["mulu_tuming_coms"]=(LB_coms_tuzhimingcheng,handles_LB_coms_tuzhimingcheng)


    # 取消选择状态
    cancel_cad_selection()
    # 保存
    try: savefile()
    except: pass


    # 更新字典并保存到电脑本地   
    try:
        dwg_name = current_dwg_basename()
    except Exception as e:
        time.sleep(1)
        dwg_name = current_dwg_basename()       
        node("[ERROR] current_dwg_basename 失败，再尝试一次：{}",e)
    try:
        save_print_dict_generic(dwg_name, zd)
        node("[INFO] 已保存到字典 -> {}",dwg_name)
    except Exception as e:
        time.sleep(1)
        save_print_dict_generic(dwg_name, zd)
        node("[ERROR] save_print_dict_generic 失败：{}",e)

    disable_debug()

    return zd



# 将文字分隔线转入专有图层'bianmulu_fengexian'
@debuggable
def move_entity_to_frozen_layer(
        ents,
        layername: str = "bianmulu_fengexian",
        *,
        freeze_layer: bool = True,
):
    """
    将任意 COM 实体列表转移到指定图层，并（可选）冻结该图层。
    使用 node() 打印调试信息。
    - ents       : 可迭代的 COM 对象（支持 Layer 属性）
    - layername  : 目标图层名称
    - freeze_layer: True 则在所有视口冻结
    """
    import pythoncom, win32com.client
    global doc
    if doc is None:
        raise RuntimeError("doc 未初始化，请先运行 li() 或确保 doc = acad.ActiveDocument")

    layers = doc.Layers
    lname = layername.strip()

    # 确保图层存在
    node("-> 检查图层 '{}'", lname)
    try:
        layer = layers.Item(lname)
        node("· 图层已存在: Handle={} ", layer.Handle)
    except pythoncom.com_error:
        node("· 图层不存在，创建 '{}'", lname)
        layer = layers.Add(lname)
        node("· 创建成功: Handle={} ", layer.Handle)

    # 冻结或解冻
    if freeze_layer:
        node("-> 冻结图层 '{}'", lname)
        try:
            layer.Freeze = True
        except pythoncom.com_error:
            node("· Freeze 属性失败，尝试 LayerOn")
            try:
                layer.LayerOn = False
            except Exception:
                node("‼ 无法通过 LayerOn 冻结")
    else:
        node("-> 解冻图层 '{}'", lname)
        try:
            layer.Freeze = False
        except pythoncom.com_error:
            node("· Freeze 属性失败，尝试 LayerOn")
            try:
                layer.LayerOn = True
            except Exception:
                node("‼ 无法通过 LayerOn 解冻")

    # 批量移动
    total = len(ents)
    fail_list = []
    for ent in ents:
        try:
            ent.Layer = lname
            node("· 实体 {} 移动到图层 '{}'", ent.Handle, lname)
        except pythoncom.com_error as exc:
            node("‼ 实体 {} 移动失败: {}", ent.Handle, exc)
            fail_list.append((ent.Handle, exc))

    node("-> 完成移动: 成功={}，失败={}", total - len(fail_list), len(fail_list))
    return fail_list


# 将文字分隔线转出专有图层'bianmulu_fengexian'回"bianmulu_lp"

@debuggable
def thaw_and_move_from_fengexian(
        src_layer="bianmulu_fengexian",
        dst_layer="bianmulu_lp",
        *,
        refreeze_src=False,
):
    """
    ① 解冻 src_layer
    ② stc() 选实体
    ③ 移动到 dst_layer
    ④ （可选）再次冻结 src_layer
    """
    node("=== 函数开始 ===")
    global doc
    layers = doc.Layers

    # ---- 1. 图层准备 -------------------------------------------------
    src = get_or_create_layer(layers, src_layer.strip())
    dst = get_or_create_layer(layers, dst_layer.strip())

    # ---- 2. 解冻源图层 ------------------------------------------------
    try:
        node("· 解冻源图层 {}", src_layer)
        src.Freeze = False
    except Exception:
        node("· Freeze 属性失败，尝试 LayerOn")
        try: src.LayerOn = True
        except Exception as exc: node("‼ 无法解冻: {}", exc)

    # ---- 3. 选实体 ---------------------------------------------------
    ents = stc(src_layer.strip())
    node("· stc() 共捕获 {} 个实体", len(ents))

    # ---- 4. 批量移动 -------------------------------------------------
    fail = []
    for e in ents:
        try:
            e.Layer = dst_layer.strip()
        except Exception as exc:
            fail.append((e.Handle, exc))
    node("· 成功移动 {}/{} 个实体", len(ents) - len(fail), len(ents))

    # ---- 5. 再次冻结（可选）-------------------------------------------
    if refreeze_src:
        node("· 重新冻结源图层 {}", src_layer)
        try: src.Freeze = True
        except Exception:
            try: src.LayerOn = False
            except Exception: pass

    node("=== 函数结束：fail={} ===", len(fail))
    return fail

@debuggable
def get_or_create_layer(layers, lname):
    """保证 lname 图层存在，返回其 COM 对象。"""
    node(">>> 进入 get_or_create_layer('{}')", lname)
    try:
        lyr = layers.Item(lname)
        node("· 图层已存在 -> Handle={}", lyr.Handle)
    except Exception:
        node("· 图层不存在，准备新建")
        lyr = layers.Add(lname)
        node("· 新建成功 -> Handle={}", lyr.Handle)
    return lyr


#&&&%  ▶ F2 编目录后续处理

@timeit
@debuggable
def  cataloging_post_processing():

    # 1 将标准打印框线图层的打印线清除、图层清除

    # 2 根据word指令整理文字排版

    pass









#&&&%  ▶ F3 全自动编目录函数












#&&&% 🧱 服务三 打印完整流程


#&&&% 打印基础 

"""
模型空间上的对象在获取打印线后就打印了，虽然有空白情况，反馈后，简单处理一下就可以了

要理解透彻打印参数
参数顺序不能乱来，这是前置

oplot = acad.ActiveDocument.PlotConfigurations.Add("PDF", acad.ActiveDocument.ActiveLayout.ModelType)添加打印机再使用dir(oplot)

如果出错，直接用oplot = acad.ActiveDocument.PlotConfigurations.Add("PDF")
执行完lay.ConfigName         = "DWG To PDF.pc3" 执行 oplot = acad.ActiveDocument.PlotConfigurations.Add("PDF", acad.ActiveDocument.ActiveLayout.ModelType)
就不会错，就可以用oplot来dir测试了
但是它们还是不相同

oplot.CanonicalMediaName
'Sun_Hi-Res_(1600.00_x_1280.00_Pixels)'
lay.CanonicalMediaName
'ISO_A2_(594.00_x_420.00_MM)'

oplot.ConfigName
'C:\\Users\\Administrator\\AppData\\Roaming\\Autodesk\\AutoCAD 2021\\R24.0\\chs\\plotters\\AutoCAD 2021 - 简体中文 (Simplified Chinese) PC3 文件\\PublishToWeb JPG.pc3'
这是因为默认值

我们实际要的是lay的结果
oplot.GetCanonicalMediaNames() 所有尺寸


li()强化连接

lay = doc.ActiveLayout

添加"DWG To PDF.pc3"打印机

lay.ConfigName         = "DWG To PDF.pc3"

设置图幅

lay.CanonicalMediaName = "ISO_A2_(594.00_x_420.00_MM)"

直接设置lay.StandardScale=0不一定能采用布满图纸 还需要配合lay.UseStandardScale=True配合，它的真假可以取消其作用

参数顺序不能乱来，这是一般参数

居中
lay.CenterPlot         = True

图纸方向

lay.PlotRotation       = 0 1,2,3

打印样式 

lay.StyleSheet         = "monochrome.ctb"

按样式打印，可查看勾选情况

lay.PlotWithPlotStyles = True

lay.PlotHidden         = True并没有让后台打印点√ doc.SetVariable("BACKGROUNDPLOT", 1)会选择为后台打印

这是窗口

LowerLeft = [xs, ys]

UpperRight = [xe, ye]

# 这里假设你已经定义了 ConvertArrays2Variant 函数

LowerLeft = ConvertArrays2Variant(LowerLeft, "Double")

UpperRight = ConvertArrays2Variant(UpperRight, "Double")

currentLayout.SetWindowToPlot(LowerLeft, UpperRight)

currentLayout.PlotType = 4

它们都可以在交互界面看到效果

就使用pyautogui响应窗口，应对复杂的情况
LB=[(391260.99150432437, 1160938.196555899, 450660.99150432437, 1244938.196555899), (460660.99150432437, 1160938.196555899, 490360.99150432437, 1202938.196555899), (500360.99150432437, 1160938.196555899, 559760.9915043244, 1244938.196555899), (569760.9915043244, 1160938.196555899, 611760.9915043244, 1190638.196555899), (621760.9915043244, 1160938.196555899, 651460.9915043244, 1202938.196555899), (661460.9915043244, 1160938.196555899, 720860.9915043244, 1202938.196555899), (730860.9915043244, 1160938.196555899, 772860.9915043244, 1220338.196555899), (782860.9915043244, 1160938.196555899, 824860.9915043244, 1190638.196555899), (391260.99150432437, 760938.1965558991, 420960.99150432437, 802938.1965558991), (430960.99150432437, 760938.1965558991, 514960.99150432437, 820338.1965558991), (524960.9915043244, 760938.1965558991, 554660.9915043244, 802938.1965558991), (564660.9915043244, 760938.1965558991, 606660.9915043244, 790638.1965558991), (616660.9915043244, 760938.1965558991, 676060.9915043244, 844938.1965558991), (686060.9915043244, 760938.1965558991, 728060.9915043244, 820338.1965558991), (738060.9915043244, 760938.1965558991, 767760.9915043244, 802938.1965558991), (777760.9915043244, 760938.1965558991, 861760.9915043244, 820338.1965558991), (391260.99150432437, 360938.1965558991, 433260.99150432437, 390638.1965558991), (443260.99150432437, 360938.1965558991, 502660.99150432437, 402938.1965558991), (512660.99150432437, 360938.1965558991, 596660.9915043244, 420338.1965558991), (606660.9915043244, 360938.1965558991, 636360.9915043244, 402938.1965558991), (646360.9915043244, 360938.1965558991, 705760.9915043244, 402938.1965558991), (715760.9915043244, 360938.1965558991, 775160.9915043244, 444938.1965558991), (785160.9915043244, 360938.1965558991, 844560.9915043244, 402938.1965558991), (854560.9915043244, 360938.1965558991, 913960.9915043244, 402938.1965558991)]



一 plot_window_to_pdf函数专门处理打印成为白色的问题

二 按打印框对象绘制包含所有图形的框，在视口中全部显示，再将比例调为1：1能精准控制图纸空间的比例

三 线的宽度设为0，A0图按A3打印解决超大大图显示在A3问题

四 对pyautogui和双线程熟练了

五 竖向图形和横向图形必须分开使用export_window_to_pdf打印

六 export_window_to_pdf的各项设置可以在交互窗口看到，最关键的窗口设置与win32数据类型有关。

七 各项设置的顺序不能随意，有些必须先设置，否则一些参数会产生矛盾

八  .pc3右键可以设置打印质量

九 shx字体的内容获取可以用来直接在pdf检测关键字符内容或者关键名字直接使用仿宋或宋体以便从pdf就方便修改

十 缺字体解决方案

十一 

"""



def 打印测试():
    li()

    lay = doc.ActiveLayout                    # 或 acad.ActiveDocument.ActiveLayout
    lay.ConfigName         = "DWG To PDF.pc3" # 出图设备
    lay.CanonicalMediaName = "ISO_A2_(594.00_x_420.00_MM)"  # 纸张
    lay.StyleSheet         = "monochrome.ctb" # CTB
    lay.PlotRotation       = 0
    lay.CenterPlot         = True
    lay.UseStandardScale   = True
    lay.PlotWithPlotStyles = True
    lay.PlotHidden         = False

    # -- ① 设定窗口 --------------------------------------------------
    lower_left  = [391260.99150432437, 1160938.196555899]
    upper_right = [450660.99150432437, 1244938.196555899]

    # 你的工具函数：把 Python list → SafeArray (Variant)：
    lower_left  = ConvertArrays2Variant(lower_left,  "Double")
    upper_right = ConvertArrays2Variant(upper_right, "Double")

    lay.SetWindowToPlot(lower_left, upper_right)  # Window
    lay.PlotType = 4                              # 4 = Window

    # -- ② 打印到文件（**不会出现预览**）-----------------------------
               # 前台打印
    doc.Plot.QuietErrorMode = True                # 静默
    doc.Plot.PlotToFile(r"D:\CADXT\SHUC\TUZHI\00.pdf")

    print("✅ 直接生成 PDF 完成")


#双线程GUI方案
"""
测试


"""

#闪动窗口

def zd():

    minimize_all_windows_d()

    restore_and_position(
        "AutoCAD",
        width_ratio = 1,
        height_ratio = 1,
        x= 0,
        y= 0
    )


def auto_press_enter(filepath):
    """自动输入路径并按下回车键"""

    time.sleep(1)  # 等待2秒，确保窗口完全加载

    # Minimize all windows
    zd()

    list_open_window_titles()  # 打印当前窗口列表，调试用

    time.sleep(2)

    # 激活窗口
    x, y, _, _ = activate_window_by_title('创建打印文件', click_titlebar=True)

    # 设置命令行焦点
    focus_cmdline(x + 359, y + 415, delay=0.2)

    # 删除默认路径和文件名
    pyautogui.moveTo(x + 359, y + 415)  # 点击宽度窗口位置
    pyautogui.click(x + 359, y + 415)
    time.sleep(1)

    pyautogui.click()
    pyautogui.hotkey('ctrl', 'a')  # 全选
    time.sleep(1)

    pyautogui.press('delete')  # 删除选中的内容
    time.sleep(1)

    # 确保路径合乎语法
    filepath = f'{filepath}'
    pyautogui.write(filepath, interval=0.2)  # 写路径
    time.sleep(1)

    pyautogui.press('enter')  # 模拟按下回车键
    print("🖱 自动点击回车，关闭对话框")


def get_title(timeout_event, event, pt1=(391260.99150432437, 1160938.196555899, 0), pt2=(450660.99150432437, 1244938.196555899, 0)):
    """第一个线程：触发窗口并发送打印命令"""
    pythoncom.CoInitialize()
    try:
        # 调用plot_w()触发打印窗口弹出
        li()
        plot_w(pt1=pt1, pt2=pt2)
        print("CAD命令已发送，等待窗口操作完成...")

        # 增加延时，确保窗口完全加载
        time.sleep(3)

        # 调试：打印出所有当前窗口标题
        print("当前打开的窗口标题列表：", list_open_window_titles())

        # 等待timeout_event信号，如果收到信号，则退出
        timeout_event.wait()

    except Exception as e:
        print(f"get_title 出现错误: {e}")
    finally:
        pythoncom.CoUninitialize()
        event.set()  # 通知主线程完成


def deal_with_title(timeout_event, event, filepath):
    """第二个线程：处理弹出窗口并执行回车"""
    pythoncom.CoInitialize()
    try:
        print("第二个进程开始运行")
        li()

        # 等待窗口弹出并按下回车键
        auto_press_enter(filepath)
    except Exception as e:
        print(f"deal_with_title 出现错误: {e}")
    finally:
        pythoncom.CoUninitialize()
        timeout_event.set()
        event.set()  # 通知主线程完成


def plot_w(pt1=(391260.99150432437, 1160938.196555899, 0), pt2=(450660.99150432437, 1244938.196555899, 0)):
    """执行打印命令，生成打印窗口"""
    li()
    cmds = [
        "_-PLOT",  # 打印命令
        "",  # 是否需要详细打印配置？[是(Y)/否(N)] <否>: N
        "",  # 输入布局名或 [?] <模型>
        "",  # 输入页面设置名 <>:
        "",  # 输出设备名 <当前>
        "",  # 输入输出设备的名称或 [?] <DWG To PDF.pc3>:
        "",  # 是否保存对页面设置的修改 [是(Y)/否(N)]? <N>
        "",  # 是否继续打印？[是(Y)/否(N)] <Y>:
        f"{pt1[0]},{pt1[1]}",  # 打印区域左下角
        f"{pt2[0]},{pt2[1]}",  # 打印区域右上角
    ]

    # 拼接命令并发送给 AutoCAD
    doc.SendCommand("\n".join(cmds) + "\n")
    print("✅ 打印命令发送成功。")






def plot_window_to_pdf(pt1=(391260.99150432437, 1160938.196555899, 0), pt2=(450660.99150432437, 1244938.196555899, 0), filename_chunming="00.pdf"):
    """
    双线程模式打印窗口

    文件路径不能有汉字
    
    """
    #正名

    base_ming = r"D:/Myprogramsystem/cad/xitongjicuwenjian/dayin"

    pdf_fullpath = join_paths(base_ming, filename_chunming)

    t1 = time.time()

    # 创建一个Event，用于通知子线程终止
    timeout_event = threading.Event()
    # 创建一个Event，用于通知主线程完成
    event = threading.Event()

    # 创建并启动两个线程
    thread1 = threading.Thread(target=get_title, args=(timeout_event, event))
    thread2 = threading.Thread(target=deal_with_title, args=(timeout_event, event, pdf_fullpath))

    thread1.start()
    thread2.start()

    # 等待两个线程完成，最多等待 180 秒
    thread1.join(timeout=180)
    thread2.join(timeout=180)

    if not event.is_set():
        print("操作超时，正在中断...")
        timeout_event.set()

    print("文件打印结束")
    t2 = time.time()
    print("打印共用时", t2 - t1, "秒")


#&&% 打印线信息存储字典

"""
.Coordinates会因为轻量多段线和一般多段线而数据结构不同

"""

def generate_print_frame_dict(com_dys, cha_Y=2000):
    """
    1. 对 com_dys 先排序，再把它们传给 plcom_to_coor；
    2. 如果一开始调用报错，则剔除“坐标数组异常”的多段线；
    3. 其余逻辑与原来大致相同：排序 Handle，提取外包盒，写字典，最后 save。
    """
    # （1）排序
    try:
        sorted_coms = sort_coms_by_llcorner(com_dys, cha_Y=cha_Y)
    except Exception as e:
        print(f"[ERROR] sort_coms_by_llcorner 失败：{e}")
        return

    # （2）先尝试一次性提取所有坐标
    try:
        LB_coor = plcom_to_coor(sorted_coms)
    except Exception as e:
        # 理论上 plcom_to_coor 不会抛 IndexError 了，因为已做兼容
        print(f"[ERROR] plcom_to_coor 出现意外：{e}")
        return

    # （3）获取纯文件名
    try:
        dwg_name = current_dwg_basename()

        #针对文件打印线的字典信息带上特殊戳记以和其它如图签字典区别
        dwg_name =  dwg_name +"_dy" 

    except Exception as e:
        print(f"[ERROR] current_dwg_basename 失败：{e}")
        return

    # （4）排序 Handle
    try:
        LB_Handle = sort_Handle(sorted_coms, cha_Y=cha_Y)
    except Exception as e:
        print(f"[ERROR] sort_Handle 失败：{e}")
        return

    # （5）对齐长度并构建字典
    n1 = len(sorted_coms)
    n2 = len(LB_coor)
    n3 = len(LB_Handle)
    n = min(n1, n2, n3)
    if (n1, n2, n3) != (n, n, n):
        print(f"[WARN] 长度不一致：sorted_coms={n1}, LB_coor={n2}, LB_Handle={n3}；只取前 {n} 条。")
        sorted_coms = sorted_coms[:n]
        LB_coor     = LB_coor[:n]
        LB_Handle   = LB_Handle[:n]

    bind_dict = {"handles": list(LB_Handle[:n])}
    for idx in range(n):
        pl = sorted_coms[idx]
        handle = LB_Handle[idx]
        pts, closed_flag = LB_coor[idx]

        # （5.3）外包盒角点
        try:
            p1_box, p2_box = pl.GetBoundingBox()
            z0 = p1_box[2]
            bbox_corners = [
                (p1_box[0], p1_box[1], z0),
                (p1_box[0], p2_box[1], z0),
                (p2_box[0], p2_box[1], z0),
                (p2_box[0], p1_box[1], z0),
            ]
        except Exception:
            bbox_corners = []

        bind_dict[handle] = {
            "com_obj": pl,
            "coords":  pts,
            "closed":  closed_flag,
            "bbox":    bbox_corners
        }

    # （6）保存
    try:
        save_print_dict_generic(dwg_name, bind_dict)
        print(f"[INFO] 已保存打印框线字典 -> '{dwg_name}'")
    except Exception as e:
        print(f"[ERROR] save_print_dict_generic 失败：{e}")

    return bind_dict



#&&% 从字典库载入字典获取区域坐标列表（左下角、右上角）

def extract_bbox_extents(frame_dict: dict) -> list[tuple[float, float, float, float]]:
    """
    从已排序的打印框线字典中提取每个框线的外包盒坐标范围。

    输入字典 frame_dict 应满足以下结构：
      - frame_dict["handles"] 是一个按顺序排列的 Handle 字符串列表。
      - 对于 frame_dict 中的每个 Handle h，frame_dict[h] 是一个包含字段 "bbox" 的子字典，
        其中 "bbox" 是一个长度为 4 的列表，表示外包盒的 4 个角点坐标，
        例如 [(x0, y0, z0), (x0, y1, z0), (x1, y1, z0), (x1, y0, z0)]。

    本函数按 frame_dict["handles"] 中的顺序，对每个 Handle：
      1. 取子字典 info = frame_dict[handle]
      2. 从 info["bbox"] 中取第 1 个点 (索引 0) 作为左下角 (x1, y1)
      3. 从 info["bbox"] 中取第 3 个点 (索引 2) 作为右上角 (x2, y2)
      4. 将 (x1, y1, x2, y2) 组成一个元组，按顺序依次放入列表 LB。

    :param frame_dict: 已生成的打印框线信息字典，必须包含键 "handles"。
    :return: 一个列表 LB，其中每个元素为 (x1, y1, x2, y2)，
             与 frame_dict["handles"] 中的顺序一一对应。
    """
    # 确保 "handles" 存在
    if "handles" not in frame_dict:
        raise KeyError("输入字典中缺少 'handles' 键，无法提取排序信息。")

    handles = frame_dict["handles"]
    LB = []

    for handle in handles:
        # 确保该 Handle 在字典中存在
        if handle not in frame_dict:
            print(f"[WARN] Handle '{handle}' 在字典中不存在，已跳过。")
            continue

        info = frame_dict[handle]
        # 确保 "bbox" 字段合法
        if "bbox" not in info:
            print(f"[WARN] Handle '{handle}' 的 info 中缺少 'bbox' 字段，已跳过。")
            continue

        bbox = info["bbox"]
        # bbox 应该包含 4 个坐标点
        if not isinstance(bbox, list) or len(bbox) != 4:
            print(f"[WARN] Handle '{handle}' 的 'bbox' 格式不正确 (长度 != 4)，已跳过。")
            continue

        # 取第 1 个和第 3 个点
        try:
            x1, y1, _ = bbox[0]
            x2, y2, _ = bbox[2]
        except Exception as e:
            print(f"[WARN] 处理 Handle '{handle}' 的 'bbox' 时失败：{e}，已跳过。")
            continue

        LB.append((x1, y1, x2, y2))

    return LB






#&&&% 将目标文件按打印区域分解成多个纯粹文件

"""

系统的初始条件是00.dwg还是从天正启动的默认文件，将会对后续操作带来关键的影响
这当然可以监测判断


"""

@alias("分解文件")
@timeout_and_log2(600)  # 最多运行 300 秒，超时强制 kill CAD 并记录日志
def safe_batch_export_regions(region_list,
                              filename_prefix="区域导出",
                              ty=1,
                              timeout_sec=600):
    """
    带超时保护的“批量导出多个区域为 DWG”函数。

    如果单次执行超过 timeout_sec 秒，watchdog 会触发：
      - 调用 _kill_acad() 强制结束 CAD 进程
      - 记录“超时”日志
    并且抛出异常或 return，让上层脚本继续往下走，而不会卡死在这里。

    :param region_list:     [(x1,y1,x2,y2), ...]  要导出的区域列表
    :param filename_prefix: 导出文件名前缀
    :param ty:              传给 batch_export_regions_to_files 的 ty 参数
    :param timeout_sec:     本节点允许运行的最大秒数
    """
    start_time = time.time()
    dwg_name = current_dwg_basename()

    try:
        print(f"▶ [SafeExport] 开始批量导出，DWG = '{dwg_name}'，共 {len(region_list)} 个区域，超时阈值 = {timeout_sec}s")
        # 这里调用已有的、未经修改的批量导出函数
        batch_export_regions_to_files(region_list,
                                      filename_prefix=filename_prefix,
                                      ty=ty)

    except Exception as e:
        # 捕获 batch_export 内部抛出的异常（一旦出现，就记录日志并尝试恢复 CAD）
        elapsed = time.time() - start_time
        tb_str = traceback.format_exc()
        _log("safe_batch_export_regions",
             (region_list, filename_prefix, ty),
             {'timeout_sec': timeout_sec},
             f"内部异常，已运行 {elapsed:.1f}s，异常信息：\n{tb_str}")
        print(f"❌ [SafeExport] 出现异常，已中止。耗时 {elapsed:.1f}s，异常：{e}")

        # 尝试调用 st() 恢复 CAD 环境，供后续节点使用
        try:
            print("🔄 [SafeExport] 正在尝试恢复 CAD 环境 (st)…")

            chongfu_caozuo(
                huifu_xitong,
                dwg_instance=None,
                args=(),
                kwargs=None,
                max_retries=3,
                failure_value=None
            )
            print("🔄 [SafeExport] CAD 恢复完成。")
        except Exception as _:
            print("⚠ [SafeExport] 恢复 CAD 环境 (st) 失败，请手动检查 CAD 状态。")

        # 最终直接 return，不再往下执行
        return

    # 如果运行到这里，说明 batch_export 完全正常
    elapsed = time.time() - start_time
    print(f"✅ [SafeExport] 批量导出完成，耗时 {elapsed:.1f} 秒，生成的 DWG 前缀 = '{filename_prefix}'")




#&&&% 打印函数

#&&% 单张图纸打印


def export_window_to_pdf(
        lower_left_xy,      # (x, y，0)
        upper_right_xy,     # (x, y,0)
        folderpath = r"D:/Myprogramsystem/XT/dayinSHUCHU",
        file_chunming ="00.pdf",           # 不带路径的文件名"xxx.pdf"
        device   = "DWG To PDF.pc3",
        media    = "ISO_A3_(420.00_x_297.00_MM)",
        ctb      = "monochrome.ctb",
        fangxiang= 0        # 0,1,2,3

):
  
    """
    根据窗口坐标直接输出 PDF，不弹出对话框。
    打印之前清空文件夹

    """
    # ① 连接 / 准备
    li()

    pdf_path = os.path.join(folderpath,file_chunming)
        
    lower_left_xy  = (lower_left_xy[0],lower_left_xy[1])

    upper_right_xy = (upper_right_xy[0],upper_right_xy[1])

    lay = doc.ActiveLayout

    # ② 基本打印参数
    lay.ConfigName         = device
    lay.CanonicalMediaName = media
    lay.StyleSheet         = ctb
    lay.PlotRotation       = fangxiang
    lay.CenterPlot         = True
    lay.StandardScale      = 0     
    lay.UseStandardScale   = True
    lay.PlotWithPlotStyles = True
    lay.PlotHidden         = False

    # ③ 设置窗口
    ll = ConvertArrays2Variant(list(lower_left_xy),  "Double")
    ur = ConvertArrays2Variant(list(upper_right_xy), "Double")
    lay.SetWindowToPlot(ll, ur)
    lay.PlotType = 4          # 4 = Window

    # ⑤ 静默、前台出图
    doc.SetVariable("BACKGROUNDPLOT", 0)
    doc.Plot.QuietErrorMode = True   
    doc.Plot.PlotToFile(pdf_path)

    print(f"✅ 已输出 PDF → {pdf_path}")
    return True

#&&% 多张图纸打印

def printing_multiple_drawings(LBcom, cha_Y=2000, tol=10,
                               folderpath=r"D:/Myprogramsystem/XT/dayinSHUCHU",
                               ctb="monochrome.ctb"):
    """
    按节点调试的批量打印：
      节点1：排序打印区域，确定输出文件夹 (以当前 DWG 名称命名)
      节点2：遍历每个区域，调用 export_window_to_pdf；
              若 generate_name_and_ratio_from_com 返回 0，则跳过并报告；
              打印失败时自动重试，最多 3 次，每次间隔 1 秒。
    """
    # —— 节点1：排序 & 文件夹 ——
    node("[print] 节点1 开始排序打印区域，总数 = {}", len(LBcom))
    LBcom_sorted = sort_coms_by_llcorner(LBcom, cha_Y=cha_Y)
    # 生成子文件夹：folderpath/当前DWG名

    now = datetime.datetime.now()
    timename = now.strftime("%Y-%m-%d-%H-%M")

    name_dwg = current_dwg_basename()

    name_dwg = name_dwg + timename

    subfolder = os.path.join(folderpath, name_dwg)
    if not os.path.exists(subfolder):
        os.makedirs(subfolder, exist_ok=True)
        node("[print] 节点1 创建输出文件夹：{}", subfolder)
    else:
        node("[print] 节点1 输出文件夹已存在：{}", subfolder)

    # —— 节点2：遍历打印 ——
    for i, dy_com in enumerate(LBcom_sorted, start=1):
        p1, p2 = dy_com.GetBoundingBox()
        zhi = generate_name_and_ratio_from_com(
            dy_com,
            A3dy=0,
            Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3", 0),
            tol=tol
        )
        if zhi == 0:
            node("[print] 节点2-{} ⚠ 不合标准，跳过打印", i)
            continue

        media, ratio, size, fangxiang = zhi[0], zhi[1], zhi[2], zhi[3]
        node("[print] 节点2-{} media={} size={} fangxiang={} 开始打印", i, media, size, fangxiang)

        success = False
        for attempt in range(1, 4):  # 最多 3 次
            try:
                export_window_to_pdf(
                    p1, p2,
                    folderpath=subfolder,
                    file_chunming=str(i),
                    device="DWG To PDF.pc3",
                    media=media,
                    ctb=ctb,
                    fangxiang=fangxiang
                )
                node("[print] 节点2-{} 尝试 {} 成功", i, attempt)
                success = True
                break
            except Exception as e:
                node("[print] 节点2-{} 尝试 {} 失败：{}", i, attempt, e)
                time.sleep(1)
        if not success:
            node("[print] 节点2-{} ❌ 打印失败，已超过最大重试次数", i)

    node("✅ Printing_multiple_drawings 完成，输出目录：{}", subfolder)
    return True




#&&% 按单文件字典信息打印

def Print_by_file_dict_info(zd: dict,
                            cha_Y=2000,
                            tol=10,
                            folderpath=r"D:/Myprogramsystem/XT/dayinSHUCHU",
                            ctb="monochrome.ctb"):
    """根据 zd 中的 Bbox_lf + prinfo 批量导出 PDF"""
    bbox_list = zd.get("Bbox_lf", [])
    handles   = zd.get("dyx_handles", [])
    prinfo    = zd.get("prinfo", {})

    node("[pbfi] 节点1 Bbox_lf 条数 = {}", len(bbox_list))

    now = datetime.datetime.now()
    timename = now.strftime("%Y-%m-%d-%H-%M")

    name_dwg = current_dwg_basename()

    name_dwg = name_dwg + timename

    out_dir = os.path.join(folderpath, name_dwg)
    os.makedirs(out_dir, exist_ok=True)
    node("[pbfi] 节点1 输出目录 = {}", out_dir)

    for i, bbox in enumerate(bbox_list, 1):
        (min_x, min_y, _), (max_x, max_y, _) = bbox
        p1, p2 = (min_x, min_y, 0), (max_x, max_y, 0)
        h = handles[i-1] if i-1 < len(handles) else None
        info = prinfo.get(h)
        if not info:
            node("[pbfi] 节点2-{} ⚠ 无 prinfo, handle={} 跳过", i, h)
            continue
        media, ratio, size, fang向 = info
        node("[pbfi] 节点2-{} {} {} fang={} 开始", i, media, size, fang向)

        success = False
        for attempt in range(1, 4):
            try:
                export_window_to_pdf(
                    p1, p2,
                    folderpath=out_dir,
                    file_chunming=str(i),
                    device="DWG To PDF.pc3",
                    media=media,
                    ctb=ctb,
                    fangxiang=fang向
                )
                node("[pbfi] 节点2-{} 尝试{}成功", i, attempt)
                success = True
                break
            except Exception as e:
                node("[pbfi] 节点2-{} 尝试{}失败: {}", i, attempt, e)
                time.sleep(1)
        if not success:
            node("[pbfi] ❌ 节点2-{} 三次失败", i)

    node("✅ Print_by_file_dict_info 完成, 目录: {}", out_dir)
    return True














def update_prinfo(prinfo):
    """
    将字典 prinfo 中所有值的第一个元素修改为 "ISO_A3_(420.00_x_297.00_MM)"。
    
    参数
    ----
    prinfo : dict
        包含多个元素，其中每个元素的值为一个包含多个字段的元组
    """
    for key, value in prinfo.items():
        # 修改第一个元素
        prinfo[key] = ("ISO_A3_(420.00_x_297.00_MM)",) + value[1:]

    return prinfo



#&&% 图纸空间打印

#&&% 异常图纸打印

#&&&% 打印编目录插图签系统的整体流程

#  主函数
#  (1)
# 删除模型或图纸空间中的所有对象

#  该函数系列包括如下一些函数


def delete_all_modelspace_objects():##

    doc.SendCommand('_ai_selall' + chr(13) + chr(13))
        
    doc.SendCommand('E' + chr(13) + chr(13))

    print(f"✅ 清除完毕")


## 测试示例

##__________


#  主函数
#  (1)
# 获取当前文件所有含特定字符ZF的布局名称列表

#  该函数系列包括如下一些函数



def bujuming_jihe(zf):

    # 如果当前在模型空间，则切换到图纸空间
    if doc.ActiveSpace != 0:  # 0 = 图纸空间，1 = 模型空间
        doc.ActiveSpace = 0
    doc.MSpace = False

    layouts = doc.Layouts
    ALB = []

    for layout in layouts:
        layout_name = layout.Name
        if zf in layout_name:
            ALB.append(layout_name)

    print(f"✅ 含有 '{zf}' 的布局有 {len(ALB)} 个：{ALB}")
    return ALB


## 测试示例

##__________


             
#  主函数
#  (1)
# 清除指定布局上的默认视口

#  该函数系列包括如下一些函数

def is_close(a, b, tol=1):
    return abs(a - b) < tol

def is_near_point(pt1, pt2, tol=1.0):
    return math.hypot(pt1[0] - pt2[0], pt1[1] - pt2[1]) < tol

def delete_default_viewports(layout_name):
    """
    删除指定布局中的默认视口（包括两种典型尺寸）。
    """
    layoutx = doc.Layouts

    try:
        layout = layoutx.Item(layout_name)
    except Exception:
        print(f"布局 '{layout_name}' 不存在。")
        return

    deleted_count = 0
    block = layout.Block

    for item in block:
        try:
            if item.ObjectName == "AcDbViewport":
                center = item.Center
                width = float(item.Width)
                height = float(item.Height)

                # 默认视口常见尺寸 A：9×6，中心接近 (5.7, 3.9)
                is_default_a = (
                    is_near_point(center, (5.7, 3.9), 1.0) and
                    is_close(width, 9.0) and
                    is_close(height, 6.0)
                )

                # 默认视口常见尺寸 B：230.7675×158.2251
                is_default_b = (
                    is_close(width, 230.7675, tol=0.5) and
                    is_close(height, 158.2251, tol=0.5)
                )

                if is_default_a or is_default_b:
                    item.Delete()
                    deleted_count += 1
        except Exception as e:
            print(f"跳过对象，错误：{e}")

    print(f"删除默认视口 {deleted_count} 个。")


#&&% # 在图纸空间的布局创建对应模型空间图框区域的视口

def create_layout_viewports(*groups, layout_name="Layout1"):
    """
    在图纸空间创建布局"Layout1"，按指定区域建立多个视口。
    """
    li()
    layouts = doc.Layouts

    for i, group in enumerate(groups):
        if len(group) != 4:
            print(f"❌ 无效的坐标组: {group}")
            continue

        x1, y1, x2, y2 = group
        layout_suffix = f"{i+1:02d}"
        new_layout_name = layout_name + layout_suffix

        doc.SetVariable("TILEMODE", 1)
        delete_all_modelspace_objects()

        Center = vtpnt((x1 + x2) / 2, (y1 + y2) / 2, 0)
        Width = abs(x2 - x1)
        Height = abs(y1 - y2)

        Right_point = vtpnt(x2, (y1 + y2) / 2, 0)
        Left_point = vtpnt(x1, (y1 + y2) / 2, 0)
        Up_point = vtpnt((x1 + x2) / 2, y1, 0)
        Down_point = vtpnt((x1 + x2) / 2, y2, 0)

        rect_pts = [x1, y1, x2, y1, x2, y2, x1, y2, x1, y1]
        poly = doc.ModelSpace.AddLightWeightPolyline(vtFloat(rect_pts))
        poly.Closed = True

        line1 = doc.ModelSpace.AddLine(Center, Right_point)
        line2 = doc.ModelSpace.AddLine(Center, Left_point)
        line3 = doc.ModelSpace.AddLine(Center, Up_point)
        line4 = doc.ModelSpace.AddLine(Center, Down_point)

        doc.SetVariable("TILEMODE", 0)

        try:
            doc.Layouts.Add(new_layout_name)
            time.sleep(0.2)
        except Exception as e:
            print(f"布局名称 {new_layout_name} 已存在或添加失败: {e}")

        for attempt in range(3):
            try:
                li()
                layouts = doc.Layouts
                doc.ActiveLayout = layouts.Item(new_layout_name)
                time.sleep(0.5)
                break
            except Exception as e:
                print(f"⚠️ 第 {attempt+1} 次切换布局失败，重试中... 错误: {e}")
                time.sleep(1)
        else:
            print(f"❌ 第 {i+1} 个视口切换布局失败，跳过")
            continue

        delete_default_viewports(new_layout_name)

        try:
            pviewportObj1 = doc.PaperSpace.AddPViewport(Center, Width, Height)
            doc.SendCommand("Z" + chr(13) + "e" + chr(13))
            time.sleep(1)
            pviewportObj1.Display(True)
            time.sleep(1)
            pviewportObj1.StandardScale = 0
            time.sleep(1)
            pviewportObj1.StandardScale = 1
            time.sleep(1)
            pviewportObj1.CustomScale = 1
            print(f"✅ 创建了第 {i+1} 个视口")
        except Exception as e:
            print(f"❌ 创建视口失败: {e}")

        try:
            poly.Delete()
            line1.Delete()
            line2.Delete()
            line3.Delete()
            line4.Delete()
        except Exception as e:
            print(f"⚠️ 删除第{i+1}组辅助图形失败: {e}")

            
def create_layout_viewports_from_rects(rect_list, layout_name="Layout"):
    """
    直接从一组 (x1, y1, x2, y2) 矩形坐标创建多个布局视口。
    """
    create_layout_viewports(*rect_list, layout_name=layout_name)


#&&% 单视口创建
            
def create_layout_viewports_new(*groups, layout_name="Layout"):#按指定数据创建视口，不删除模型对象

    """
    这个函数不清除模型空间上的对象，适合一个文件只有单张图纸的打印
    这里的*groups实际是单个元组数据

    """    
    layouts = doc.Layouts

    def layout_exists(name):
        for i in range(layouts.Count):
            if layouts.Item(i).Name.lower() == name.lower():
                return True
        return False

    for i, group in enumerate(groups):
        if len(group) != 4:
            print(f"❌ 无效的坐标组: {group}")
            continue

        x1, y1, x2, y2 = group

        layout_suffix = f"{i+1:02d}"
        new_layout_name = layout_name + layout_suffix

        # 计算视口位置
        Center = vtpnt((x1 + x2) / 2, (y1 + y2) / 2, 0)
        Width = abs(x2 - x1)
        Height = abs(y1 - y2)

        doc.SetVariable("TILEMODE", 0)

        # 添加新布局
        if not layout_exists(new_layout_name):
            layouts.Add(new_layout_name)
            time.sleep(0.5)
        else:
            print(f"⚠️ 布局 {new_layout_name} 已存在，跳过添加")

        # 切换到新布局
        try:
            doc.ActiveLayout = layouts.Item(new_layout_name)
        except Exception as e:
            print(f"❌ 无法激活布局 {new_layout_name}：{e}")
            continue

        # 删除默认视口
        delete_default_viewports(new_layout_name)

        # 添加视口
        pviewportObj1 = doc.PaperSpace.AddPViewport(Center, Width, Height)
        doc.SendCommand("Z" + chr(13) + "e" + chr(13))
        time.sleep(1)

        pviewportObj1.Display(True)
        time.sleep(1)
        pviewportObj1.StandardScale = 0
        time.sleep(1)
        pviewportObj1.StandardScale = 1
        time.sleep(1)
        pviewportObj1.CustomScale = 1

        print(f"✅ 创建了第 {i+1} 个视口，布局名: {new_layout_name}")

        # 清理辅助图形（若存在）
        try:
            if "poly" in locals(): poly.Delete()
            if "line1" in locals(): line1.Delete()
            if "line2" in locals(): line2.Delete()
            if "line3" in locals(): line3.Delete()
            if "line4" in locals(): line4.Delete()
        except Exception as e:
            print(f"⚠️ 删除第{i+1}组辅助图形失败: {e}")



#  主函数
#  (1)
# 合并不同布局上的视口

#  该函数系列包括如下一些函数

def merge_layout_viewports(layout_names):


    # 0. 切换到图纸空间
    if doc.ActiveSpace != 0:
        doc.ActiveSpace = 0
    doc.MSpace = False

    layouts = doc.Layouts
    existing_names = [layout.Name for layout in layouts]

    # 1. 创建目标布局 My（如果不存在）
    if "My" not in existing_names:
        layouts.Add("My")

    # 2. 遍历输入的每个布局名
    for layout_name in layout_names:
        if layout_name not in existing_names:
            print(f"布局 {layout_name} 不存在，跳过。")
            continue

        print(f"处理布局 {layout_name}...")

        # 切换到源布局
        doc.ActiveLayout = layouts.Item(layout_name)
        doc.MSpace = False

        # 执行 COPYBASE 和全选

        doc.SendCommand('_ai_selall' + chr(13) + chr(13))
        
        doc.SendCommand('_copybase' + chr(13) + '0,0' + chr(13)+ chr(13))
        
        time.sleep(2.5)  # 等待复制完成

        # 切换到目标布局 My 并粘贴
        doc.ActiveLayout = layouts.Item("My")
        doc.MSpace = False
        doc.SendCommand('TPasteClip' + chr(13)+ '0,0'+ chr(13)+ chr(13))
        time.sleep(2.5)  # 等待粘贴完成

    print("所有布局内容已合并至 'My'。")

    delete_default_viewports("My")



## 测试示例

# merge_layout_viewports(["My1", "My2", "My3"])

##__________



##判断是否矩形

def is_axis_aligned_rectangle(Lb, tol=1.0):
    """
    判断4个3D点 (x, y, z) 是否构成一个标准轴对齐的矩形（图框意义上的）
    :param Lb: 4个点的列表 [(x1,y1,z1), (x2,y2,z2), ...]
    :param tol: 容差值，允许小范围误差
    :return: True/False
    """
    if len(Lb) != 4:
        return False

    # 提取所有 x 和 y 坐标
    xs = [pt[0] for pt in Lb]
    ys = [pt[1] for pt in Lb]

    # 判断 x-对（x 差值小于 tol 的点对数量）
    x_pairs = 0
    for i in range(4):
        for j in range(i + 1, 4):
            if abs(xs[i] - xs[j]) < tol:
                x_pairs += 1

    # 判断 y-对
    y_pairs = 0
    for i in range(4):
        for j in range(i + 1, 4):
            if abs(ys[i] - ys[j]) < tol:
                y_pairs += 1

    return x_pairs == 2 and y_pairs == 2







#  主函数
#  (1)
# 将任何文件的打印设置设为ZG_PDF指定样式

#  该函数系列包括如下一些函数

#&&%  进入当前激活文件的指定布局
def switch_to_layout(MyLayout):
    """
    将当前 CAD 文件切换到指定布局（图纸空间）
    参数：
        MyLayout：目标布局名称
    """
    try:
        li()
        layouts = doc.Layouts
        for i in range(layouts.Count):
            layout = layouts.Item(i)
            if layout.Name.lower() == MyLayout.lower():
                doc.ActiveLayout = layout
                doc.SetVariable("TILEMODE", 0)  # 确保在图纸空间
                print(f"✅ 成功切换到布局: {MyLayout}")
                print("🔍 当前布局是：", doc.ActiveLayout.Name)
                return
        print(f"❌ 布局 '{MyLayout}' 未找到。")
    except Exception as e:
        print(f"❌ 切换布局失败: {e}")





def copy_plot_configuration_to_another_doc(target_doc_name,  target_layout_name,plot_config_name="ZG_PDF"):# 将任何文件的打印设置设为ZG_PDF指定样式

    """
    直接设置打印参数很困难，但却可以从一个设置好的打印配置名称为ZG_PDF的文件复制过来

    打印样式等参数可以用一个单独的函数在打印之前来设置

    最好的模式是打开一个存在的文件作引子

    也可以使用天正刚开启时默认的文件为引子，可能会出现显示错误，显示文件名为ZG_PDF,新建一个文件即可查看到运行正确


    """
    # 🔹 确保桌面只有一个文件
    close_all_except_active_safe()
    li()

    #打开标准文件
    
    Open_By_Omission_wenjian(r"D:/Myprogramsystem/XT/ZG_PDF.dwg")
    li()
    # 获取当前活动的文档
    source_doc = acad.ActiveDocument

    # 使"布局1"为当前布局
    acad.ActiveDocument.SetVariable("TILEMODE",0)

    layouts = doc.Layouts

    existing_names = [layout.Name for layout in layouts]

    doc.ActiveLayout = layouts.Item("布局1")
    
    layout = source_doc.ActiveLayout

    # 获取 "ZG_PDF" 打印配置
    zg_pdf_config = source_doc.PlotConfigurations.Item("ZG_PDF")

    # 使用 CopyFrom 方法将 "ZG_PDF" 的设置复制到指定布局
    layout.CopyFrom(zg_pdf_config)

    

    # 打开目标文件

    Open_By_Omission_wenjian(target_doc_name)
    
    li()
    
    target_doc = acad.ActiveDocument

    # 转入图纸空间
    acad.ActiveDocument.SetVariable("TILEMODE",0)

    # 转入布局target_layout_name
    switch_to_layout(target_layout_name)

    # 使用 CopyFrom 方法将 "ZG_PDF" 的设置复制到 目标布局
    target_doc.ActiveLayout.CopyFrom(zg_pdf_config)

    op1 = acad.ActiveDocument.PlotConfigurations.Add("ZG_PDF")

    # 保存目标文件并关闭
    target_doc.Save()
    target_doc.Close()

    li()

    source_doc = acad.ActiveDocument

    wenjianshuliang = acad.Documents.Count

    print("文件数",wenjianshuliang)

    if wenjianshuliang == 1:

        Open_By_Omission_wenjian(r"D:/Myprogramsystem/XT/空白.dwg")

        source_doc.Close()

    else:

        source_doc.Close()        

    li()



def batch_copy_plot_config(target_files: list, target_layout_name: str, plot_config_name="ZG_PDF"):#批量处理,将一系列文件的打印设置设为ZG_PDF指定样式
    for f in target_files:
        try:
            copy_plot_configuration_to_another_doc(f, target_layout_name, plot_config_name)
        except Exception as e:
            print(f"⚠️ 处理文件 {f} 失败: {e}")


##测试示例
            
##batch_copy_plot_config([
##    r"D:/Myprogramsystem/打印/图纸01.dwg",
##    r"D:/Myprogramsystem/打印/图纸02.dwg"
##], "My")
##            


def copy_plot_config_to_current_doc_layout_1(layout_name, plot_config_name="ZG_PDF"):
    """
    将 ZG_PDF.dwg 文件中的打印配置复制到当前激活文件的指定布局。

    参数:
        layout_name (str): 当前文件中要设置打印样式的布局名称；
        plot_config_name (str): 源打印配置名，默认为 "ZG_PDF"。

    zg_pdf_config = acad.ActiveDocument.PlotConfigurations.Item("ZG_PDF")#从当前文档获取，只要它存在某个布局中
    acad.ActiveDocument.ActiveLayout.CopyFrom(zg_pdf_config)#使得当前布局采用这个配置    


    当前激活文件不能为天正刚开启的时候默认创建文件，因为打开新文件时它会自动关闭

    当前桌面只宜有一个文件
        
    """
    # 🔹 确保桌面只有一个文件
    close_all_except_active_safe()
    li()

    target_doc = acad.ActiveDocument

    #打开标准文件
    
    Open_By_Omission_wenjian(r"D:/Myprogramsystem/XT/ZG_PDF.dwg")
    li()
    # 获取当前活动的文档
    source_doc = acad.ActiveDocument

    # 使"布局1"为当前布局
    acad.ActiveDocument.SetVariable("TILEMODE",0)

    layouts = doc.Layouts

    existing_names = [layout.Name for layout in layouts]

    doc.ActiveLayout = layouts.Item("布局1")
    
    layout = source_doc.ActiveLayout

    # 获取 "ZG_PDF" 打印配置
    zg_pdf_config = source_doc.PlotConfigurations.Item("ZG_PDF")

    # 使用 CopyFrom 方法将 "ZG_PDF" 的设置复制到指定布局
    layout.CopyFrom(zg_pdf_config)

    

    # 激活目标文件

    set_active_doc(target_doc)
    
    li()
    
    target_doc = acad.ActiveDocument

    # 转入图纸空间
    acad.ActiveDocument.SetVariable("TILEMODE",0)

    # 转入布局target_layout_name
    switch_to_layout(layout_name)

    # 使用 CopyFrom 方法将 "ZG_PDF" 的设置复制到 目标布局
    target_doc.ActiveLayout.CopyFrom(zg_pdf_config)

    op1 = acad.ActiveDocument.PlotConfigurations.Add("ZG_PDF")

    # 保存目标文件
    target_doc.Save()

    set_active_doc(source_doc)

    li()

    source_doc.Close()        

    # 确保连回原文件
    try:
        set_active_doc(target_doc)
        li()
    except Exception as e:
        print("⚠️ 返回原始文档失败：", e)  

#&&% 将基本打印配置ZG_PDF赋给当前激活文件指定布局
def copy_plot_config_to_current_doc_layout(layout_name, plot_config_name="ZG_PDF"):#将 ZG_PDF.dwg 文件中的打印配置复制到当前激活文件的指定布局。
    """
    人工智能重写的    
    """
    try:
        # 确保桌面仅保留当前文件
        close_all_except_active_safe()
        li()

        target_doc = acad.ActiveDocument

        # 打开标准打印配置文件
        Open_By_Omission_wenjian(r"D:/Myprogramsystem/XT/ZG_PDF.dwg")
        li()
        source_doc = acad.ActiveDocument

        source_doc.SetVariable("TILEMODE", 0)
        source_doc.ActiveLayout = source_doc.Layouts.Item("布局1")

        # 获取打印配置对象
        zg_pdf_config = source_doc.PlotConfigurations.Item(plot_config_name)

        # 切换回目标文档并设置布局
        set_active_doc(target_doc)
        li()
        target_doc.SetVariable("TILEMODE", 0)
        switch_to_layout(layout_name)

        # 应用打印配置
        target_doc.ActiveLayout.CopyFrom(zg_pdf_config)
        target_doc.Save()

        # 关闭 ZG_PDF 文件
        set_active_doc(source_doc)
        source_doc.Close(False)

        # 返回目标文档
        set_active_doc(target_doc)
        li()

    except Exception as e:
        print(f"❌ 打印配置复制失败: {e}")



#  主函数
#  (1)
# 自动旋转布局中所有竖向视口

            

def select_paperspace_objects_in_window(x1, y1, x2, y2):
    """
    选择图纸空间中在左上 (x1, y1)、右下 (x2, y2) 所定义矩形区域内的对象。

    返回值：
        符合条件的对象组成的列表。
    """
    doc.SetVariable("TILEMODE", 0)  # 确保在图纸空间

    x_min, x_max = min(x1, x2), max(x1, x2)
    y_min, y_max = min(y1, y2), max(y1, y2)

    selected_objs = []

    for entity in doc.PaperSpace:
        try:
            min_pt, max_pt = entity.GetBoundingBox()
            ex1, ey1 = min_pt[0], min_pt[1]
            ex2, ey2 = max_pt[0], max_pt[1]

            # 判断实体的包围盒是否与选择框相交
            if (ex1 <= x_max and ex2 >= x_min and
                ey1 <= y_max and ey2 >= y_min):
                selected_objs.append(entity)
        except Exception as e:
            # 某些对象可能不支持 GetBoundingBox（如视口），直接跳过
            continue

    print(f"✅ 共选择图纸空间对象 {len(selected_objs)} 个。")
    return selected_objs




def rotate_entity_90(obj):
    """对一个对象绕其中心点旋转 90°"""
    try:
        min_pt, max_pt = obj.GetBoundingBox()
        center_x = (min_pt[0] + max_pt[0]) / 2
        center_y = (min_pt[1] + max_pt[1]) / 2
        center_z = (min_pt[2] + max_pt[2]) / 2
        base_point = vtpnt(center_x, center_y, center_z)

        obj.Rotate(base_point, math.radians(90))
        print(f"✅ 对象 {obj.Handle} 已旋转 90°")
    except Exception as e:
        print(f"❌ 无法旋转对象 {obj.Handle}：{e}")


#&&% 90度或负90度旋转布局中所有竖向视口

def rotate_portrait_viewports_in_layout(layout_name, angle_deg=90):
    """
    对指定布局中所有宽高比 < 1 的竖向视口进行旋转。
    
    参数:
        layout_name (str): 要处理的布局名称
        angle_deg (float): 旋转角度（默认 90 度，可设为 -90 逆时针）
    """
    try:
        doc.SetVariable("TILEMODE", 0)
        time.sleep(0.2)

        layout = doc.Layouts.Item(layout_name)
        doc.ActiveLayout = layout
        print(f"📐 正在处理布局：{layout_name}，旋转角度：{angle_deg}°")

        time.sleep(0.3)

        try:
            paperspace = doc.PaperSpace
        except Exception as e:
            print("❌ 无法访问 PaperSpace：", e)
            return

        count = 0

        for entity in paperspace:
            if entity.ObjectName != 'AcDbViewport':
                continue

            try:
                cx, cy = entity.Center[0], entity.Center[1]
                width, height = entity.Width, entity.Height

                if width / height >= 1:
                    continue

                pad = 5
                x1 = cx - width / 2 - pad
                y1 = cy + height / 2 + pad
                x2 = cx + width / 2 + pad
                y2 = cy - height / 2 - pad

                highlight_entities_in_window(x1, y1, x2, y2)

                cmd_rotate = (
                    '_ROTATE\n'
                    'P\n'
                    f"{cx},{cy}\n"
                    f"{angle_deg}\n"
                )
                doc.SendCommand(cmd_rotate)
                time.sleep(0.8)

                print(f"✅ 已旋转视口 Handle: {entity.Handle}")
                count += 1

            except Exception as e:
                print(f"⚠️ 旋转失败 Handle: {entity.Handle}，错误: {e}")
                continue

        print(f"🎯 布局 {layout_name} 中共旋转视口 {count} 个。")

    except Exception as e:
        print(f"❌ 函数运行失败: {e}")




#  主函数
#  (1)
# 图纸空间的窗口打印

    """
    设置了两种模式，一种是分步骤的，先完成ZG_PDF的基本配置，再设置打印样式，图纸尺寸，然后再打印，分三步。另一种模式就直接用一个函数


    """



#  该函数系列包括如下一些函数


def set_layout_plot_settings(layout_name, dyys="monochrome.ctb", tfcc="ISO_A2_(594.00_x_420.00_MM)", xz=0):#控制打印样式、图纸尺寸、图纸方向，需先搞好配置ZG_PDF 
    """
    设置指定布局的打印样式、图纸尺寸、方向等参数。

    参数:
        layout_name: 要设置的布局名称
        dyys: 打印样式表名称（如 'monochrome.ctb'）
        tfcc: 图纸尺寸（如 'ISO_A2_(594.00_x_420.00_MM)'）
        xz: 图纸方向（0=上，1=右，2=下，3=左）
    """
    try:
        switch_to_layout(layout_name)
        currentLayout = doc.ActiveLayout

        currentLayout.StyleSheet = dyys
        currentLayout.CanonicalMediaName = tfcc
        currentLayout.PlotRotation = xz

        print(f"✅ 布局 '{layout_name}' 打印设置更新成功")
        print(f"📎 打印样式: {dyys}")
        print(f"📐 图纸尺寸: {tfcc}")
        print(f"🔄 图纸方向: {xz}")

    except Exception as e:
        print(f"❌ 设置布局打印参数失败: {e}")




def print_layout_window_simple_1(layout_name, x1, y1, x2, y2, output_folder=r"D:/Myprogramsystem/XT/dayinSHUCHU"):#按当前设置作窗口打印
    """
    功能：
      1. 如果当前处于模型空间，则切换到图纸空间；
      2. 切换到指定名称 layout_name 的图纸空间布局（要求该布局已存在）；
      3. 以输入数据 (x1, y1) 为窗口左上角，(x2, y2) 为窗口右下角，
         实际打印时将转换为下左角 (x1, y2) 和上右角 (x2, y1)；
      4. 设置打印区域为“窗口”，居中打印，采用标准比例（布满图纸）；
      5. 调用 Plot.PlotToFile 输出 PDF 文件，文件名为 “layout_name01.pdf”。
      
    参数：
      layout_name - 目标布局名称（字符串）；
      x1, y1     - 窗口左上角纸空间坐标；
      x2, y2     - 窗口右下角纸空间坐标；
      output_folder - 输出 PDF 文件的文件夹路径（默认为 "D:/Myprogramsystem/XT/dayinSHUCHU"）。
    """
    # 如果当前为模型空间，则切换到图纸空间
    currentLayout = acad.ActiveDocument.ActiveLayout
    if currentLayout.Name.lower() == "model":
        acad.ActiveDocument.SetVariable("TILEMODE", 0)
        time.sleep(0.5)
        currentLayout = acad.ActiveDocument.ActiveLayout

    # 切换到指定布局（假定该布局已存在）
    acad.ActiveDocument.SendCommand('_-layout' + chr(13) + 'Set' + chr(13) + layout_name + chr(13))
    time.sleep(0.5)
    currentLayout = acad.ActiveDocument.ActiveLayout

    # 计算打印窗口：输入数据为左上 (x1,y1) 和右下 (x2,y2)，实际打印窗口要求下左角和上右角
    lower_left = [x1, y2]
    upper_right = [x2, y1]
    
    # 转换为 Variant 类型（请确保已定义 ConvertArrays2Variant 函数）
    LowerLeftVar = ConvertArrays2Variant(lower_left, "Double")
    UpperRightVar = ConvertArrays2Variant(upper_right, "Double")
    
    # 设置打印窗口及相关打印参数
    currentLayout.SetWindowToPlot(LowerLeftVar, UpperRightVar)
    currentLayout.PlotType = 4          # 4 表示窗口打印
    currentLayout.CenterPlot = True     # 居中打印
    currentLayout.UseStandardScale = True  # 使用标准比例（布满图纸）

    # 构造输出文件名，格式为 "布局名称01.pdf"
    output_file = output_folder + "\\" + layout_name + "01.pdf"

    # 调用 Plot.PlotToFile 进行打印
    result = acad.ActiveDocument.Plot.PlotToFile(output_file)
    print("打印完成，输出文件:", output_file, "结果:", result)



        




#&&% 单视口窗口打印

def print_layout_window_simple(
        lower_left_xy,                   # (x, y, z) 视口左下
        upper_right_xy,                  # (x, y, z) 视口右上
        *,                               # 之后全部用关键字
        layout_name   = "布局1",
        folderpath    = r"D:/Myprogramsystem/XT/dayinSHUCHU",
        output_folder = None,            # 兼容旧代码
        file_chunming = "00.pdf",
        device        = "DWG To PDF.pc3",
        media         = "ISO_A3_(420.00_x_297.00_MM)",
        ctb           = "monochrome.ctb",
        fangxiang     = 0,               # 0=0° 1=90° 2=180° 3=270°
        index         = None             # 序号，若有则拼到文件名
):
    """
    在图纸空间 `layout_name` 中，按给定窗口坐标输出 PDF。
    形参、返回值与 export_window_to_pdf 保持一致：成功 True / 失败 False
    --------------------------------------------------------------------------
    lower_left_xy / upper_right_xy 仅 x、y 有效；z 分量可给 0。
    """
    # -------- 0. 兼容旧参数名 -------------------------------------------------
    if output_folder is not None:               # 调用方若仍用旧名
        folderpath = output_folder              # 映射到新的 folderpath

    try:
        # -------- 1. 确保处于目标布局 --------------------------------------
        li()                                    # 刷新全局 acad/doc/mp/lay
        switch_to_layout(layout_name)           # ← 你已有的辅助函数
        lay = doc.ActiveLayout

        # 如果当前仍是模型空间，再保险地切一次 TILEMODE=0
        if doc.GetVariable("TILEMODE") != 0:
            doc.SetVariable("TILEMODE", 0)
            switch_to_layout(layout_name)       # 再切一次，绝对安全
            lay = doc.ActiveLayout

        # -------- 2. 打印参数 ----------------------------------------------
        lay.ConfigName         = device
        lay.CanonicalMediaName = media
        lay.StyleSheet         = ctb
        lay.PlotRotation       = fangxiang
        lay.CenterPlot         = True
        lay.UseStandardScale   = True
        lay.StandardScale      = 0              # Fit to paper
        lay.PlotWithPlotStyles = True
        lay.PlotHidden         = False
        lay.PlotType           = 4              # Window

        # -------- 3. 指定打印窗口 ------------------------------------------
        ll_var = ConvertArrays2Variant(list(lower_left_xy[:2]),  "Double")
        ur_var = ConvertArrays2Variant(list(upper_right_xy[:2]), "Double")
        lay.SetWindowToPlot(ll_var, ur_var)

        # -------- 4. 输出文件路径 ------------------------------------------
        if not os.path.isdir(folderpath):
            os.makedirs(folderpath, exist_ok=True)

        # 若给了 index，就把序号拼到文件名；否则用传入的 file_chunming
        if index is not None:
            _, ext = os.path.splitext(file_chunming)
            file_chunming = f"{index}{ext}"

        pdf_path = os.path.join(folderpath, file_chunming)

        # -------- 5. 静默前台出图 -----------------------------------------
        doc.SetVariable("BACKGROUNDPLOT", 0)    # 前台打印
        doc.Plot.QuietErrorMode = True
        ok = doc.Plot.PlotToFile(pdf_path)

        node("✅ 已输出 PDF → {}   result={}", pdf_path, ok)
        return True

    except Exception as e:
        node("❌ print_layout_window_simple 失败: {}", e)
        return False

#  主函数
#  (1)
# 获取布局上有效视口的对角点数据

"""
默认矩形对角点都是指左上角点、右下角点
之所以需要这么长的篇幅处理有效视口的对角点数据是因为图纸空间存在一个大隐性视口

"""
#  该函数系列包括如下一些函数


def convert_LB_to_points(LB):# 将表征左上角右下角的四维坐标点列表，按规则转为三维点列表
    """
    将四维坐标（x, y, z, w）转换为两个三维坐标点（x, y, 0）和（z, w, 0）
    
    参数:
        LB: 需要转换的四维坐标列表，例如：
            [(-670871.6477594106, 576917.4420590551, 1279318.874177304, -348728.54356674734),
             (-127998.52008571586, 85006.34675900426, -9098.520085715863, 906.3467590042565)]
    
    返回:
        转换后的坐标列表，格式为：
            [(-670871.6477594106, 576917.4420590551, 0), (1279318.874177304, -348728.54356674734, 0),
             (-127998.52008571586, 85006.34675900426, 0), (-9098.520085715863, 906.3467590042565, 0)]
    """
    points = []
    for item in LB:
        # 分离四个维度
        x1, y1, z1, w1 = item
        
        # 生成两个3D坐标点，z值固定为0
        point1 = (x1, y1, 0)
        point2 = (z1, w1, 0)
        
        # 将这两个点加入到结果列表
        points.append(point1)
        points.append(point2)
    
    return points


def get_non_pure_rectangles(rectangles):#分析一组矩形哪些是不纯净的，返回它们
    """
    检查哪些矩形是不纯净的，并返回它们本身的坐标值。消除隐含的伪视口和非法视口

    ✅ 不纯净矩形的完整判断逻辑：
    对于矩形 R1 和任意其他矩形 R2（R1 ≠ R2）：

    如果 R1 与 R2 有交集，并且：

    R1 不被 R2 完全包含，且

    R1 也不完全包含 R2，

    那么 R1 是 不纯净的。

    如果 R1 完全包含 R2，R1 也是不纯净的（即使交集是完整包含）

    参数:
        rectangles: [(x1, y1, x2, y2), ...]

    返回:
        不纯净矩形列表：[ (x1,y1,x2,y2), ... ]
    """

    def has_intersection(r1, r2):
        x1a, y1a, x2a, y2a = r1
        x1b, y1b, x2b, y2b = r2
        return not (x2a <= x1b or x2b <= x1a or y2a >= y1b or y2b >= y1a)

    def contains(ra, rb):
        x1a, y1a, x2a, y2a = ra
        x1b, y1b, x2b, y2b = rb
        return x1a <= x1b and y1a >= y1b and x2a >= x2b and y2a <= y2b

    non_pure_rects = []

    for i, r1 in enumerate(rectangles):
        for j, r2 in enumerate(rectangles):
            if i == j:
                continue

            if has_intersection(r1, r2):
                if contains(r1, r2) or not contains(r2, r1):
                    non_pure_rects.append(r1)
                    break

    print(f"✅ 检测到 {len(non_pure_rects)} 个不纯净矩形")
    return non_pure_rects



def get_viewports_coordinates(layout_name):
    """
    获取指定布局上的所有视口的左上角和右下角坐标。
    
    参数：
        layout_name - 布局名称
    
    返回：
        一个列表，包含每个视口的 (x1, y1, x2, y2) 坐标元组
    """
    try:
        # 切换到指定布局
        switch_to_layout(layout_name)
        time.sleep(0.5)

        # 删除默认生成的视口

        delete_default_viewports(layout_name)
        time.sleep(0.5)
        
        # 获取当前布局的视口对象
        viewports_coordinates = []
        
        # 遍历当前布局上的所有图形实体
        for entity in acad.ActiveDocument.PaperSpace:
            if entity.ObjectName == 'AcDbViewport':  # 如果是视口对象
                try:
                    # 获取视口的中心点、宽度和高度
                    cx, cy = entity.Center[0], entity.Center[1]  # 中心点
                    width, height = entity.Width, entity.Height  # 宽度和高度
                    
                    # 计算左上角坐标 (x1, y1)
                    x1 = cx - width / 2
                    y1 = cy + height / 2  # y坐标是反向的，所以加上高度的一半
                    
                    # 计算右下角坐标 (x2, y2)
                    x2 = cx + width / 2
                    y2 = cy - height / 2  # y坐标是反向的，所以减去高度的一半
                    
                    # 将坐标存入列表
                    viewports_coordinates.append((x1, y1, x2, y2))
                    print(entity.Handle)
                except Exception as e:
                    print(f"⚠️ 获取视口坐标失败: {e}")
        
        print(f"✅ 获取到 {len(viewports_coordinates)} 个视口的坐标。")
        return viewports_coordinates
    
    except Exception as e:
        print(f"❌ 获取视口坐标失败：{e}")
        return []



def get_pure_viewports_coordinates(layout_name):# #获取布局上所有   有效视口   的左上角、右下角的坐标点，以四维坐标组返回
    all_coords = get_viewports_coordinates(layout_name)
    non_pure = get_non_pure_rectangles(all_coords)
    pure_coords = [r for r in all_coords if r not in non_pure]
    print(f"✅ 获取到纯净视口 {len(pure_coords)} 个")
    return pure_coords




#  插入基本图签块

def insert_print_title_block(
    p,
    layer_name="print_tuqian_jiben",
    filepath=r"D:/Myprogramsystem/XT/tuqian_jiben.dwg"
):
    """
    在点 p 处插入基础图签块（tuqian_jiben.dwg）：
      1. 确保并清空指定图层
      2. 计算实际插入点 P = (p.x + dx, p.y + dy, 0)
      3. 调用 insert_and_explode_dwg 插入、分解，最多尝试 3 次
      4. 等待 1 秒
      5. 查找该图层上所有块参照，给每个块参照的 Name 加上时间戳
      6. 如果图层上块参照数量 > 1，返回 True，否则 False

    参数：
      p : tuple[float, float, float] — 目标点
      layer_name : str — 要操作的图层名
      filepath : str — 要插入的 DWG 文件路径

    返回：
      bool — 插入并命名成功且块数 > 1 则 True，否则 False
    """
    # 1. 确保并清空图层
    ensure_layer(layer_name)

    # 2. 计算插入点
    dx = 647340.35234613
    dy = -107806.62742633
    P = (p[0] + dx, p[1] + dy, 0.0)

    # 3. 插入并分解 DWG，最多尝试 3 次
    success = False
    for attempt in range(3):
        try:
            insert_and_explode_dwg(
                filepath,
                insertion_point=P,
                scale=(1.0, 1.0, 1.0),
                rotation=0.0,
                wait=0.3
            )
            success = True
            break
        except Exception:
            time.sleep(0.5)  # 间隔重试
    if not success:
        return False

    # 4. 等待 1 秒
    time.sleep(1.0)

    # 5. 给新插入的块加时间戳
    now = datetime.datetime.now()
    timename = now.strftime("%Y-%m-%d-%H-%M")

    Bks = stc(layer_name)  # 获取该图层所有实体
    count_blocks = 0
    for ent in Bks:
        if ent.ObjectName == "AcDbBlockReference":
            count_blocks += 1
            orig_name = ent.Name
            new_name = f"{orig_name}_{timename}"
            try:
                rename_block_entity(ent, new_name)
            except Exception:
                # 若无法直接修改，可按需实现复制/删除/再插入逻辑
                pass

    # 6. 检查数量
    return count_blocks > 1



#***************************************************************************************************************************************************************************








def 编目录():

    li()

    write_directory_template(
        LC,
        LC_th, LC_gg, LC_ratio, LC_other=None,
        hangju=1000.0,
        n=20,                                   # 每列行数
        column_shifts=[(0, 0.0), (68198.12, 0.0), (92973.95, 0.0)],        # 每个数据对应一排，其值为相对第一排左侧起点的距离
        split_threshold=10000.0,
        y_tol=300,
        P_start=(0, 0, 0)
    )


def A3打印():
    li()
    zd=process_print_areas_info()

    update_prinfo(zd["prinfo"])


    Print_by_file_dict_info(zd)




def 一般打印():
    li()
    zd=process_print_areas_info()

    Print_by_file_dict_info(zd)




#&&&% 文件夹打印

@timeit             
@debuggable        
def A3_print_all_files_in_folder(
    folder_path=r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/工业园整理/三期/测试",
    ctb="monochrome.ctb",
    cha_Y=2000,
    tol=50,
    max_attempts: int = 5,
):


    """
    遍历文件夹 folder_path 中的每个 DWG 文件，确保每次操作的文件唯一激活，并打印每个文件为 PDF。

    dwg_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".dwg")]# 获取文件夹内所有 dwg 文件
    先唯一打开00.dwg，基于它打开别的文件，操作完，关闭。再打开别的文件……这是比较稳定合理的多文件工作模式
    ensure_active_cad_file(current_file)当有两个文件时可以尝试关闭另一个文件实现目标，减少时间和算力损耗

    """
    

    minimize_all_windows_d()

    # 恢复并定位 CAD 窗口
    restore_and_position(
        name="AutoCAD",
        width_ratio=1.0,
        height_ratio=1.0,
        x=0,
        y=0
    )
    time.sleep(1)

    enable_debug()
    li()
    
    # 获取文件夹及其子文件夹中的所有 dwg 文件
    dwg_files = []
    for root, _, files in os.walk(folder_path):
        # 在每个子文件夹中查找 .dwg 文件
        for file in files:
            if file.lower().endswith(".dwg"):
                dwg_files.append(os.path.join(root, file))


    # 确保当前文件（00.dwg）是已打开文件
    current_file = "D:/Myprogramsystem/XT/00.dwg"
    zongshu =0   
    # 遍历文件夹内的每个 dwg 文件并打印
    for dwg_file in dwg_files:
        dwg_filepath = os.path.join(folder_path, dwg_file)
               
        for attempt in range(1, max_attempts + 1):

            try:
                node("▷ [{}] 处理文件：{}", attempt, dwg_filepath)


                # 确保文件 00.dwg 是当前唯一激活的文件才打开目标操作文件，当try出错时ensure_active_cad_file(current_file)会被略过所以实测又打开了新文件

                if not keep_only_doc(r"D:/Myprogramsystem/XT/00.dwg"):

                    node("未能保持00.dwg是唯一激活文件状态，再次尝试") 

                    continue               

                doc = Open_By_Omission_wenjian(dwg_filepath)
        
                check_cad_ready(retries=12, wait_time=3)
                
                
                # 执行 A3 打印
                dyn = A3_print_1(cha_Y=cha_Y, tol=tol,
                       folderpath=r"D:/Myprogramsystem/XT/dayinSHUCHU",
                       ctb="monochrome.ctb")
        
                doc.SendCommand("Z\nE\n")
        
                time.sleep(1)
        
                img = pyautogui.screenshot(region=(175,101,2350,1250))
        
                base_name, _ = os.path.splitext(dwg_file)
        
                # 保存到与 DWG 同级的文件夹
                dwg_dir = os.path.dirname(dwg_filepath)
                img_path = os.path.join(dwg_dir, f"{base_name}_screenshot.png")
            
                img.save(img_path)
        
                node("##____________##____________##____________##_____________")
        
                node("✅ 文件{}打印完，截图已保存：{}",dwg_filepath,img_path)
             
                node("[关键节点 1 ]打印了{}文件 {} 张图纸",dwg_filepath,dyn) 
                zongshu =zongshu+dyn   
        
                close_wps_window_by_click()
        
        
                # 打印完后关闭当前文件（确保关闭）

                # 尝试关闭当前文档，最多 3 次
                closed_ok   = False   # 记录是否真正调用过 doc.Close 并成功
                skip_closed = False   # 记录是否发现已经没有打开文档
                
                for _ in range(3):
                    try:
                        li()  # 刷新 COM 对象
                        if acad.Documents.Count==2:          # 还有2文档 → 尝试关闭dwg_filepath
                            doc.Close()
                            closed_ok = True
                        else:                             # 已无2个文档 → 视为成功，退出循环
                            skip_closed = True
                        break
                    except Exception:
                        time.sleep(0.5)                   # 等 500 ms 再重试
                else:
                    # 只有三次全部 raise 才会走到这里
                    node("[关键节点 2] ❌ 关闭文件 {} 失败 3 次，不再尝试", dwg_filepath)
                
                # 根据执行结果输出对应日志
                if closed_ok:
                    node("[关键节点 2] ✅ 已完成打印并关闭文件 {}", dwg_filepath)
                elif skip_closed:
                    node("[关键节点 2] ℹ️ AutoCAD 当前无打开文档，跳过关闭")


                # 确保文件 00.dwg 是当前唯一激活的文件
        
                li()#连接当前激活文件
                ensure_active_cad_file(current_file)
                      
                # 本次文件处理成功，跳出重试循环
                break    

            except Exception as e:
                node("⚠ [{}] 处理失败: {}", attempt, e)
                time.sleep(1)
                if attempt == max_attempts:
                    node("❌ [{}] 超过最大重试次数({})，跳过文件 {}", attempt, max_attempts, dwg_filepath)
        # end for attempts


    node("✅ 所有文件已打印完毕,共 {} 张 ",zongshu)
    
    disable_debug()

    return



def close_wps_window_by_click(
    title_keyword: str = "WPS Office",
    offset_x: int = 22,    # 距窗口右边缘的像素
    offset_y: int = 19,    # 距窗口上边缘的像素
    pause_before: float = 0.2
) -> bool:
    """
    在标题包含 title_keyword 的窗口右上角点击一次（×），尝试关闭该窗口。
    如果没找到窗口，则返回 False 并跳过。
    """
    # 枚举可见窗口，匹配标题
    handles = []
    def _enum(hwnd, lst):
        if win32gui.IsWindowVisible(hwnd):
            txt = win32gui.GetWindowText(hwnd) or ""
            if title_keyword.lower() in txt.lower():
                lst.append(hwnd)
    win32gui.EnumWindows(_enum, handles)

    if not handles:
        node("ℹ️ 未找到标题含 ‘{}’ 的窗口，跳过关闭", title_keyword)
        return False

    hwnd = handles[0]
    left, top, right, bottom = win32gui.GetWindowRect(hwnd)
    x = right - offset_x
    y = top   + offset_y

    time.sleep(pause_before)
    pyautogui.click(x, y)
    node("✅ 点击 '{}' 窗口(句柄 {}) 的关闭按钮坐标：({}, {})",
         title_keyword, hwnd, x, y)
    return True

#&&% 单文件A3打印

def A3_print_1(cha_Y=2000, tol=50,
               folderpath=r"D:/Myprogramsystem/XT/dayinSHUCHU",


               ctb="monochrome.ctb",

               media="ISO_A3_(420.00_x_297.00_MM)",

):
    """
    扫描当前文件“打印框线”图层，去重、排序后统一以 A3 纸幅窗口打印。
    """
    def bbox_area(ent):
     
        try:
            (x1, y1, _), (x2, y2, _) = ent.GetBoundingBox()
            return abs(x2 - x1) * abs(y2 - y1)
        except Exception:
            return None

    def _safe_bbox(ent, max_retry=5, delay=0.4):
        for _ in range(max_retry):
            try:
                return ent.GetBoundingBox()
            except Exception:
                time.sleep(delay)
        raise
    
    def _safe_set(obj, attr, value, max_retry=5, delay=0.4):
        for _ in range(max_retry):
            try:
                setattr(obj, attr, value)
                return
            except Exception:
                time.sleep(delay)
        raise
  

    li()
    # 1. 取并去重
    set_layer_color("打印框线", color = 6)

    LB1  = stc("打印框线")

    LB2 = [ent for ent in LB1 if ent.ObjectName in ("AcDbPolyline", "AcDb2dPolyline") ]

    LBco  = dedupe_by_bbox_proximity(LB2, tol=tol)
    LBco  = sort_coms_by_llcorner(LBco, cha_Y=cha_Y)

    #打印线设置
    for ent in LBco:

        _safe_set(ent, "Color", 6)
    
        if bbox_area(ent) < 1240000000:
            _safe_set(ent, "ConstantWidth", 0.2)
        else:
            _safe_set(ent, "ConstantWidth", 20)
    




    node("打印多段线区域有 {} 根",len(LBco))

    # 2. 生成输出目录
    subfolder = os.path.join(
        folderpath,
        f"{current_dwg_basename()}_{datetime.datetime.now():%Y-%m-%d-%H-%M}"
    )
    os.makedirs(subfolder, exist_ok=True)

    # 3. 打印循环

    for idx, ent in enumerate(LBco, 1):
        p1, p2 = _safe_bbox(ent)          # ← 用安全版本
    
        orient = 0 if (p2[0]-p1[0]) >= (p2[1]-p1[1]) else 1
    
        for attempt in range(1, 4):
            try:
                export_window_to_pdf(
                    p1, p2,
                    folderpath=subfolder,
                    file_chunming=str(idx),
                    device="DWG To PDF.pc3",
                    media=media,
                    ctb=ctb,
                    fangxiang=orient
                )
                node("[print] {} 成功(第 {} 次)", idx, attempt)
    
                # ★ 打完一张后让 AutoCAD 重生一口气
                try:
                    acad = win32com.client.Dispatch("AutoCAD.Application")
                    acad.ActiveDocument.Regen(1)     # acAllViewports = 1
                except Exception:
                    pass
                break
    
            except Exception as e:
                node("[print] {} 失败(第 {} 次): {}", idx, attempt, e)
                time.sleep(1)

        else:
            node("[print] {} ❌ 三次失败", idx)

    node("✅ A3_print_1 完成；输出：{}", subfolder)
    return len(LBco)


def set_layer_color(layername: str, color: int = 1):
    """
    将图层 layername 的颜色设为 color；如果该图层不存在就创建它。

    参数
    ----
    layername : str
        目标图层名称
    color     : int
        AutoCAD 颜色索引（默认为 1，即红色）

    返回
    ----
    layer : AcadLayer
        操作后的图层对象
    """
    # 获取 AutoCAD 应用和当前文档
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    layers = doc.Layers

    # 尝试获取已有图层
    try:
        layer = layers.Item(layername)
        # 图层存在，直接设置颜色
        layer.Color = color
    except Exception:
        # 图层不存在，创建并设置颜色
        layer = layers.Add(layername)
        layer.Color = color

    return layer

def disable_wps_auto_restore():
    """
    在注册表中设置 OpenLastDocsOnStartup = 0，
    禁用 WPS PDF 的自动恢复上次打开文档功能。
    """
    import winreg
    key_path = r"SOFTWARE\Kingsoft\Office\6.0\Common"
    try:
        # 打开（或不存在则创建）HKCU 下的该键
        key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, key_path)
        # 设置 DWORD 值 OpenLastDocsOnStartup = 0
        winreg.SetValueEx(
            key,
            "OpenLastDocsOnStartup",
            0,
            winreg.REG_DWORD,
            0
        )
        winreg.CloseKey(key)
        print("✅ 已禁用 WPS PDF 自动恢复上次文档")
    except Exception as e:
        print(f"❌ 设置失败: {e}")



#&&&% 文件夹图纸空间打印

@timeit             
@debuggable        
def A3_print_all_files_in_folder_sp(

    layout_name = "mylayout",
    folder_path=r"D:/Myprogramsystem/XT/CAD",
    ctb="monochrome.ctb",
    cha_Y=2000,
    tol=50,
    max_attempts: int = 5,
):


    """
    遍历文件夹 folder_path 中的每个 DWG 文件，确保每次操作的文件唯一激活，并打印每个文件为 PDF。

    dwg_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".dwg")]# 获取文件夹内所有 dwg 文件
    先唯一打开00.dwg，基于它打开别的文件，操作完，关闭。再打开别的文件……这是比较稳定合理的多文件工作模式
    ensure_active_cad_file(current_file)当有两个文件时可以尝试关闭另一个文件实现目标，减少时间和算力损耗

    理论上该函数已经包含了A3_print_all_files_in_folder，过段时间再统一
    """
    

    minimize_all_windows_d()

    # 恢复并定位 CAD 窗口
    restore_and_position(
        name="AutoCAD",
        width_ratio=1.0,
        height_ratio=1.0,
        x=0,
        y=0
    )
    time.sleep(1)

    enable_debug()
    li()
    
    # 获取文件夹及其子文件夹中的所有 dwg 文件
    dwg_files = []
    for root, _, files in os.walk(folder_path):
        # 在每个子文件夹中查找 .dwg 文件
        for file in files:
            if file.lower().endswith(".dwg"):
                dwg_files.append(os.path.join(root, file))


    # 确保当前文件（00.dwg）是已打开文件
    current_file = "D:/Myprogramsystem/XT/00.dwg"
    zongshu =0   
    # 遍历文件夹内的每个 dwg 文件并打印
    for dwg_file in dwg_files:
        dwg_filepath = os.path.join(folder_path, dwg_file)
               
        for attempt in range(1, max_attempts + 1):

            try:
                node("▷ [{}] 处理文件：{}", attempt, dwg_filepath)


                # 确保文件 00.dwg 是当前唯一激活的文件才打开目标操作文件，当try出错时ensure_active_cad_file(current_file)会被略过所以实测又打开了新文件

                if not keep_only_doc(r"D:/Myprogramsystem/XT/00.dwg"):

                    node("未能保持00.dwg是唯一激活文件状态，再次尝试") 

                    continue               

                doc = Open_By_Omission_wenjian(dwg_filepath)
        
                check_cad_ready(retries=12, wait_time=3)
                
                
                # 执行 A3 打印
                dyn = A3_print_2(layout_name =layout_name,cha_Y=cha_Y, tol=tol,
                       folderpath=r"D:/Myprogramsystem/XT/dayinSHUCHU",
                       ctb="monochrome.ctb")
        
                doc.SendCommand("Z\nE\n")
        
                time.sleep(1)
        
                img = pyautogui.screenshot(region=(175,101,2350,1250))
        
                base_name, _ = os.path.splitext(dwg_file)
        
                # 保存到与 DWG 同级的文件夹
                dwg_dir = os.path.dirname(dwg_filepath)
                img_path = os.path.join(dwg_dir, f"{base_name}_screenshot.png")
            
                img.save(img_path)
        
                node("##____________##____________##____________##_____________")
        
                node("✅ 文件{}打印完，截图已保存：{}",dwg_filepath,img_path)
             
                node("[关键节点 1 ]打印了{}文件 {} 张图纸",dwg_filepath,dyn) 
                zongshu =zongshu+dyn   
        
                close_wps_window_by_click()
        
        
                # 打印完后关闭当前文件（确保关闭）

                # 尝试关闭当前文档，最多 3 次
                closed_ok   = False   # 记录是否真正调用过 doc.Close 并成功
                skip_closed = False   # 记录是否发现已经没有打开文档
                
                for _ in range(3):
                    try:
                        li()  # 刷新 COM 对象
                        if acad.Documents.Count==2:          # 还有2文档 → 尝试关闭dwg_filepath
                            doc.Close()
                            closed_ok = True
                        else:                             # 已无2个文档 → 视为成功，退出循环
                            skip_closed = True
                        break
                    except Exception:
                        time.sleep(0.5)                   # 等 500 ms 再重试
                else:
                    # 只有三次全部 raise 才会走到这里
                    node("[关键节点 2] ❌ 关闭文件 {} 失败 3 次，不再尝试", dwg_filepath)
                
                # 根据执行结果输出对应日志
                if closed_ok:
                    node("[关键节点 2] ✅ 已完成打印并关闭文件 {}", dwg_filepath)
                elif skip_closed:
                    node("[关键节点 2] ℹ️ AutoCAD 当前无打开文档，跳过关闭")


                # 确保文件 00.dwg 是当前唯一激活的文件
        
                li()#连接当前激活文件
                ensure_active_cad_file(current_file)
                      
                # 本次文件处理成功，跳出重试循环
                break    

            except Exception as e:
                node("⚠ [{}] 处理失败: {}", attempt, e)
                time.sleep(1)
                if attempt == max_attempts:
                    node("❌ [{}] 超过最大重试次数({})，跳过文件 {}", attempt, max_attempts, dwg_filepath)
        # end for attempts


    node("✅ 所有文件已打印完毕,共 {} 张 ",zongshu)
    
    disable_debug()

    return











#&&% 单文件A3图纸空间打印
  
@debuggable
def A3_print_2(
    *,
    layout_name : str | None = None,
    cha_Y       : int  = 2000,                     # 行高阈值，传给 sort
    tol         : int  = 20,                       # A3_print_1 用
    folderpath  : str  = r"D:/Myprogramsystem/XT/dayinSHUCHU",
    ctb         : str  = "monochrome.ctb",
    frame_layer : str  = "打印框线",
    dedupe_tol  : float = 20.0 ,                    # 去重阈值
    media         = "ISO_A3_(420.00_x_297.00_MM)",

):
    """
    图纸空间：layout_name 布局 → 在 '打印框线' 图层找 Polyline
              stc + 去重 + 行列排序后依序打印
    模型空间：layout_name=None 或切换失败 → 调用 A3_print_1
    """

    # 0) 刷新 COM 句柄
    li()

    # 1) 切换布局（如指定）
    use_paperspace = False
    if layout_name:
        try:
            doc.SetVariable("TILEMODE", 0)
            switch_to_layout(layout_name)
            doc.Layouts.Item(layout_name)    # 验证存在
            use_paperspace = True
            node("▶ 切到布局 '{}'", layout_name)
        except Exception:
            node("⚠ 布局 '{}' 不存在/切换失败 → 回模型空间", layout_name)
            doc.SetVariable("TILEMODE", 1)

    # 2A) 图纸空间打印
    if use_paperspace:
        # 2A-1 取框线多段线
        raw_objs = [e for e in stc(frame_layer)
                    if e.ObjectName in ("AcDbPolyline", "AcDb2dPolyline")]
        if not raw_objs:
            node("ℹ️ 布局 '{}' 未找到图层 '{}' 多段线", layout_name, frame_layer)
            return 0

        # 2A-2 去重 & 排序
        deduped = dedupe_by_bbox_proximity(raw_objs, tol=dedupe_tol)
        frames  = sort_coms_by_llcorner(deduped, cha_Y=cha_Y)
        node("🛈 找到 {} 条，去重后 {} 条", len(raw_objs), len(frames))

        # 2A-3 准备输出目录
        out_dir = Path(folderpath) / f"{current_dwg_basename()}_{datetime.datetime.now():%Y-%m-%d-%H-%M}"
        out_dir.mkdir(parents=True, exist_ok=True)

        success = 0
        for idx, ent in enumerate(frames, 1):
            (x1,y1,_), (x2,y2,_) = ent.GetBoundingBox()
            ok = print_layout_window_simple(
                (x1,y1,0), (x2,y2,0),
                layout_name   = layout_name,
                folderpath    = str(out_dir),
                file_chunming = f"{idx}.pdf",
                device        = "DWG To PDF.pc3",
                media         = media,
                ctb           = ctb,
                fangxiang     = 0,
            )
            node("框 {:>2} → {}", idx, "✅" if ok else "❌")
            if ok:
                success += 1

        node("✅ 布局 '{}' 完成，共 {} 张 PDF，目录：{}", layout_name, success, out_dir)
        return success

    # 2B) 模型空间打印
    doc.SetVariable("TILEMODE", 1)
    return A3_print_1(
        cha_Y      = cha_Y,
        tol        = tol,
        folderpath = folderpath,
        ctb        = ctb,
        media      = media

    )

def A3_print_2_fast(
    *,
    layout_name : str | None = None,
    cha_Y       : int  = 2000,                     # 行高阈值，传给 sort
    tol         : int  = 20,                       # A3_print_1 用
    folderpath  : str  = r"D:/Myprogramsystem/XT/dayinSHUCHU",
    ctb         : str  = "monochrome.ctb",
    frame_layer : str  = "打印框线",
    dedupe_tol  : float = 20.0                     # 去重阈值
):
    """
    图纸空间：layout_name 布局 → 在 '打印框线' 图层找 Polyline
              stc + 去重 + 行列排序后依序打印
    模型空间：layout_name=None 或切换失败 → 调用 A3_print_1
    """

    # 0) 刷新 COM 句柄
    li()

    # 1) 切换布局（如指定）
    use_paperspace = False
    if layout_name:
        try:
            doc.SetVariable("TILEMODE", 0)
            switch_to_layout(layout_name)
            doc.Layouts.Item(layout_name)    # 验证存在
            use_paperspace = True
            node("▶ 切到布局 '{}'", layout_name)
        except Exception:
            node("⚠ 布局 '{}' 不存在/切换失败 → 回模型空间", layout_name)
            doc.SetVariable("TILEMODE", 1)

    # 2A) 图纸空间打印
    if use_paperspace:
        # 2A-1 取框线多段线
        raw_objs = [e for e in stc(frame_layer)
                    if e.ObjectName in ("AcDbPolyline", "AcDb2dPolyline")]
        if not raw_objs:
            node("ℹ️ 布局 '{}' 未找到图层 '{}' 多段线", layout_name, frame_layer)
            return 0

        # 2A-2 去重 & 排序
        deduped = dedupe_by_bbox_proximity(raw_objs, tol=dedupe_tol)
        frames  = sort_coms_by_llcorner(deduped, cha_Y=cha_Y)
        node("🛈 找到 {} 条，去重后 {} 条", len(raw_objs), len(frames))

        # 2A-3 准备输出目录
        out_dir = Path(folderpath) / f"{current_dwg_basename()}_{datetime.datetime.now():%Y-%m-%d-%H-%M}"
        out_dir.mkdir(parents=True, exist_ok=True)

        success = 0
        for idx, ent in enumerate(frames, 1):
            (x1,y1,_), (x2,y2,_) = ent.GetBoundingBox()
            ok = print_layout_window_simple(
                (x1,y1,0), (x2,y2,0),
                layout_name   = layout_name,
                folderpath    = str(out_dir),
                file_chunming = f"{idx}.pdf",
                device        = "DWG To PDF.pc3",
                media         = "ISO_A3_(420.00_x_297.00_MM)",
                ctb           = ctb,
                fangxiang     = 0,
            )
            node("框 {:>2} → {}", idx, "✅" if ok else "❌")
            if ok:
                success += 1

        node("✅ 布局 '{}' 完成，共 {} 张 PDF，目录：{}", layout_name, success, out_dir)
        return success

    # 2B) 模型空间打印
    doc.SetVariable("TILEMODE", 1)
    return A3_print_1(
        cha_Y      = cha_Y,
        tol        = tol,
        folderpath = folderpath,
        ctb        = ctb
    )











#&&% 获取布局中的视口
def get_pure_viewports(layout_name: str):
    """
    返回 [(vpEnt, (x1,y1,x2,y2)), …]  
    仅保留“纯矩形视口”，自动兼容不同 CAD 版本
    """
    switch_to_layout(layout_name)          # 你的 helper：切换并激活布局
    lay   = doc.ActiveLayout
    vps   = []

    for ent in doc.PaperSpace:             # 仅遍历当前布局
        if ent.ObjectName != "AcDbViewport":
            continue
        if not is_rectangular(ent):        # 过滤掉非矩形裁剪视口
            continue

        # 计算纸空间窗口包围盒
        cx, cy = ent.Center[0], ent.Center[1]
        w,  h  = ent.Width, ent.Height
        x1, y1 = cx - w/2, cy + h/2        # 左上
        x2, y2 = cx + w/2, cy - h/2        # 右下

        vps.append((ent, (x1, y1, x2, y2)))

    node("✅ 布局 '{}' 获取到纯矩形视口 {} 个", layout_name, len(vps))
    return vps

def is_rectangular(vp) -> bool:
    """兼容不同版本 AutoCAD，判断视口是否为“纯矩形”"""
    if hasattr(vp, "NonRectClipEntityId"):
        # 一些版本返回 long；0 = 没裁剪
        try:
            return int(vp.NonRectClipEntityId) == 0
        except Exception:
            return True
    elif hasattr(vp, "NonRectClipOn"):
        # 老版本布尔属性
        try:
            return not bool(vp.NonRectClipOn)
        except Exception:
            return True
    else:
        # 实在没有相关属性，默认当作矩形
        return True







#&&% ★⚡ 完善图纸名称等列表
#&&% ★⚡ 从属性图签获取图纸名称等全部信息
#&&% ★⚡ 从zd书写图纸编号



#&&% ★⚡ 从放好的目录模板和示例文字快速编目录


#&&% ★⚡ 从zd打印全部文件成PDF


#&&&% ▶ 函数的有效延迟衔接

"""
对于有打开，插入新对象的函数，合理衔接

re



"""


def check_cad_ready(retries=12, wait_time=3):
    """
    检测当前 CAD 环境是否已经准备好进行绘图处理。
    每隔 `wait_time` 秒检查一次，直到返回消息 14142，最多重试 `retries` 次。

    参数:
    ----
    retries   : int
        最大重试次数，默认 5 次
    wait_time : int
        每次重试的等待时间（秒），默认 3 秒

    返回:
    ----
    bool
        如果 CAD 准备好并返回 14142，返回 True，否则返回 False
    """
    attempt = 0

    while attempt < retries:
        try:
            # 执行绘图操作
            li()  # 初始化 CAD 环境（根据你的代码需求）

            # 创建一条直线
            lineobj = mp.AddLine(vtpnt(0, 0, 0), vtpnt(10000, 10000, 0))
            lineobj.Color = 1  # 设置颜色为红色

            # 获取线段的起始点和结束点
            print(lineobj.StartPoint, lineobj.EndPoint)
            res1 = lineobj.StartPoint
            res2 = lineobj.EndPoint

            # 计算直线的长度
            chang = math.sqrt((res1[0] - res2[0]) ** 2 + (res1[1] - res2[1]) ** 2)
            i = int(chang)

            # 如果长度符合预期，返回 True
            if i == 14142:
                safe_delete(lineobj)
                return True  # 返回 True，说明 CAD 环境准备就绪

            # 如果返回值不为 14142，则继续重试
            time.sleep(wait_time)

        except Exception as e:
            print(f"⚠ 尝试失败: {e}. 重试中...")
            time.sleep(wait_time)

        attempt += 1

    print("❌ CAD 环境准备失败，重试次数已达上限。")

    

    return False


#确保目标dwg是激活唯一的文件，关闭其余

def keep_only_doc(target_path: str,
                  save_before_close: bool = False,
                  max_checks: int = 10,
                  interval_sec: int = 3) -> bool:
    """
    让 target_path 成为唯一激活文档，并在最多 30 秒内
    持续轮询确认（每 interval_sec 秒检测一次）。

    返回值
    -------
    True  : 成功只剩 target_path 一个文档并已激活
    False : 超时仍未达到目标状态
    """
    pythoncom.CoInitialize()
    acad = win32com.client.Dispatch("AutoCAD.Application")
    target_path = os.path.abspath(target_path)

    # ---------- 找到或打开目标 ----------
    doc_keep = None
    for i in range(acad.Documents.Count):
        doc = acad.Documents.Item(i)
        try:
            if os.path.abspath(doc.FullName) == target_path:
                doc_keep = doc
                break
        except Exception:
            if doc.Name.lower() == os.path.basename(target_path).lower():
                doc_keep = doc
                break
    if doc_keep is None:                         # 未打开则打开
        doc_keep = acad.Documents.Open(target_path, False)
    doc_keep.Activate()

    # ---------- 重复尝试关闭其他文档并轮询 ----------
    for attempt in range(1, max_checks + 1):
        # 1) 关闭除 doc_keep 以外的所有文档（倒序）
        for idx in range(acad.Documents.Count - 1, -1, -1):
            doc = acad.Documents.Item(idx)
            if doc != doc_keep:
                try:
                    doc.Close(save_before_close)
                except Exception:
                    pass   # 可能正处于关闭流程

        # 2) 检查是否满足“只剩一个且为目标”
        if acad.Documents.Count == 1:
            active = acad.ActiveDocument
            try:
                if os.path.abspath(active.FullName) == target_path:
                    time.sleep(1)               # 再多等 1 秒，确保稳定
                    return True
            except Exception:
                pass

        # 3) 未满足则等待下一轮
        time.sleep(interval_sec)

    # —— 超时未成功 ——
    return False



#&&% PDF基本处理







#合并文件
@debuggable
def merge_and_combine_pdfs_1(
    src_root: str=r"D:/Myprogramsystem/XT/inputPDF",
    dst_root: str = r"D:/Myprogramsystem/XT/outPDF",
    
    combined_name: str = "全部图纸合并.pdf",
    overwrite: bool = True
):
    """
    把 src_root 下子文件夹内的 1.pdf,2.pdf… 复制到一个目录，并重命名为
    <字母数字>-<子编号>.pdf，再按同一顺序合并成单一 PDF (combined_name)。

    排序规则：字母 → 主编号(int) → 子编号(int；缺省 0)
    """

    #───────────────── 0. 准备 ─────────────────#
    src_root = Path(src_root).resolve()
    dst_root = Path(dst_root or (src_root.parent / "_MergedPDF")).resolve()
    dst_root.mkdir(parents=True, exist_ok=True)
    node("📂 [0] 输出目录 {}", dst_root)

    # 本地正则
    file_pat   = re.compile(r"^([A-Za-z]+)(\d+)(?:-(\d+))?\.pdf$", re.I)
    folder_pat = re.compile(r"^([A-Za-z]+)(\d+)")

    def parse_key(name: str):
        """返回 (letter, major, sub)；无法解析排最后"""
        m = file_pat.match(name)
        if not m:
            return ("~", float("inf"), float("inf"))
        return (m.group(1).upper(), int(m.group(2)),
                int(m.group(3)) if m.group(3) else 0)

    #───────────────── 1. 收集任务 ──────────────#
    tasks = []                                           # (folder_prefix, pdf_path)
    for sub in src_root.iterdir():
        if not sub.is_dir():
            continue
        fm = folder_pat.match(sub.name)
        if not fm:
            node("⚠ 跳过未知文件夹 {}", sub.name)
            continue
        prefix = fm.group(1).upper() + fm.group(2)       # 如 A12
        for pdf in sub.glob("*.pdf"):
            tasks.append((prefix, pdf))

    # 排序
    tasks.sort(key=lambda t: parse_key(f"{t[0]}-{t[1].name}"))
    node("📑 [1] 发现待复制 PDF {} 个", len(tasks))

    #───────────────── 2. 复制并重命名 ──────────#
    copied_paths: list[Path] = []                       # 按排序后的新文件路径
    for prefix, src in tasks:
        _, _, sub = parse_key(f"{prefix}-{src.name}")
        dst_name = f"{prefix}-{sub}.pdf"
        dst = dst_root / dst_name

        if dst.exists() and not overwrite:
            node("⚠ 已存在，跳过 {}", dst_name)
            continue

        shutil.copy2(src, dst)
        copied_paths.append(dst)
        node("✅ 已复制 {}", dst_name)

    node("🔧 [2] 复制完成，共 {} 个文件", len(copied_paths))

    if not copied_paths:
        node("❗ 无文件可合并，流程结束")
        return

    #───────────────── 3. 合并为单一 PDF ────────#
    writer = PdfWriter()
    for pdf_path in copied_paths:        # 保证顺序
        try:
            reader = PdfReader(pdf_path)
            for page in reader.pages:
                writer.add_page(page)
        except Exception as e:
            node("⚠ 读取 {} 失败: {}", pdf_path.name, e)

    combined_path = dst_root / combined_name
    if combined_path.exists() and not overwrite:
        timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        combined_path = combined_path.with_stem(combined_path.stem + "_" + timestamp)

    with combined_path.open("wb") as f:
        writer.write(f)
    node("📕 [3] 已生成合并文件 {}", combined_path.name)

    node("🎉 全流程完成！输出目录：{}", dst_root)

@debuggable
def merge_and_combine_pdfs(
    src_root,
    dst_root = r"D:/Myprogramsystem/XT/outPDF",
    combined_name = "全部图纸合并.pdf",
    overwrite = True
):
    """
    1. 扫描 src_root 下所有子文件夹里的 PDF；
    2. O0 文件夹按专用映射改名，其余按 <前缀>-<子编号>.pdf；
    3. 复制到 dst_root 并按规则排序；
    4. 合并成 combined_name。
    """

    # ───── 0. 目录准备 ─────
    src_root = Path(src_root).resolve()
    dst_root = Path(dst_root).resolve()
    dst_root.mkdir(parents=True, exist_ok=True)
    node("📂 输出目录 {}", dst_root)

    # ───── 正则 & 排序辅助 ─────
    file_pat   = re.compile(r"^([A-Za-z]+)(\d+)(?:-(\d+))?\.pdf$", re.I)
    folder_pat = re.compile(r"^([A-Za-z]+)(\d+)")

    def parse_key(fname: str) -> Tuple[str, int, int]:
        m = file_pat.match(fname)
        if not m:
            return ("~", 10**9, 10**9)
        return (m.group(1).upper(), int(m.group(2)), int(m.group(3) or 0))

    # ───── 1. 收集任务 ─────
    tasks: List[Tuple[int, str, Path]] = []  # (priority, dst_name, src_path)

    special_map = {
        "1.pdf":  "A0-0-1.pdf",
        "2.pdf":  "A0-0-2.pdf",
        "3.pdf":  "A0-0.pdf",
        "4.pdf":  "B0-0.pdf",
        "5.pdf":  "C0-0.pdf",
        "6.pdf":  "D0-0.pdf",
        "7.pdf":  "E0-0.pdf",
        "8.pdf":  "F0-0.pdf",
        "9.pdf":  "G0-0.pdf",
        "10.pdf": "H0-0.pdf",
        "11.pdf": "I0-0.pdf",
    }

    for sub in src_root.iterdir():
        if not sub.is_dir():
            continue
        m = folder_pat.match(sub.name)
        if not m:
            continue
        prefix = m.group(1).upper() + m.group(2)   # 例 A12

        if prefix == "O0":                         # —— A. O0 特殊 —— #
            for pdf in sub.glob("*.pdf"):
                low = pdf.name.lower()
                if low not in special_map:
                    continue
                dst_name = special_map[low]
                prio = 0 if low in ("1.pdf", "2.pdf") else (1 if low=="3.pdf" else 2)
                tasks.append((prio, dst_name, pdf))
            continue

        # —— B. 普通目录 —— #
        for pdf in sub.glob("*.pdf"):
            _, _, sub_no = parse_key(f"{prefix}-{pdf.name}")
            dst_name = f"{prefix}-{sub_no}.pdf"
            tasks.append((2, dst_name, pdf))

    node("[步骤1] 收集到 {} 个待复制 PDF", len(tasks))

    # ───── 2. 排序 & 复制 ─────
    tasks.sort(key=lambda t: (t[0], parse_key(t[1])))
    copied_paths: List[Path] = []

    for prio, dst_name, src in tasks:
        dst = dst_root / dst_name
        node("   → 准备复制  priority={}  {}  -> {}", prio, src.name, dst_name)
        if dst.exists() and not overwrite:
            node("     ⚠ 已存在，跳过 {}", dst_name)
            continue
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dst)
        copied_paths.append(dst)
        node("     ✅ 完成复制 {}", dst_name)

    if not copied_paths:
        node("❗ 无文件被复制，流程终止")
        return

    node("[步骤2] 共复制 {} 个 PDF", len(copied_paths))

    # ───── 3. 合并 PDF ─────
    writer = PdfWriter()
    for pdf_path in copied_paths:
        try:
            reader = PdfReader(pdf_path)
            for pg in reader.pages:
                writer.add_page(pg)
        except Exception as e:
            node("⚠ 读取 {} 失败: {}", pdf_path.name, e)

    out_path = dst_root / combined_name
    if out_path.exists() and not overwrite:
        ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        out_path = out_path.with_stem(out_path.stem + "_" + ts)

    node("[步骤3] 写出合并文件 {}", out_path)
    with out_path.open("wb") as f:
        writer.write(f)

    node("📕 合并完成 → {}", out_path.name)

def check_and_backup(
    src_dir = r"D:/Myprogramsystem/XT/dayinSHUCHU",
    backup_dir = r"D:/Myprogramsystem/XT/backup",
    max_retries = 3,              # ← 每项最多尝试次数
    retry_wait  = 1               # ← 失败后等待秒数
):
    """
    若 src_dir 非空，则把其内容移动到 backup_dir。
    单个文件/文件夹移动失败时，最多重试 max_retries 次，每次间隔 retry_wait 秒。
    全部处理完后等待 1 秒并返回 True；若 src_dir 原本为空则返回 False。
    """
    src = Path(src_dir)
    dst = Path(backup_dir)

    if not src.exists() or not src.is_dir():
        raise FileNotFoundError(f"源目录不存在: {src}")

    items = list(src.iterdir())
    if not items:
        return False

    dst.mkdir(parents=True, exist_ok=True)

    for item in items:
        target = dst / item.name

        # 如目标已存在，先删掉
        if target.exists():
            try:
                if target.is_dir():
                    shutil.rmtree(target, ignore_errors=True)
                else:
                    target.unlink(missing_ok=True)
            except Exception:
                pass

        moved = False
        for attempt in range(1, max_retries + 1):
            try:
                shutil.move(str(item), str(target))
                moved = True
                break
            except Exception as e:
                if attempt < max_retries:
                    time.sleep(retry_wait)
                else:
                    # 最终仍失败：记录一下错误，继续处理下一个
                    print(f"⚠ 无法移动 {item.name}: {e}")

        # （可选）如仍未移动成功，可在这里追加 copy+delete 的兜底逻辑

    time.sleep(1)
    return True


#&&&% 打印输出PDF


@timeit             
@debuggable        
@alias("打印输出PDF")
def orchestrate_cad_to_pdf(
    layout_name="mylayout",
    folder_path=r"D:/Myprogramsystem/XT/CAD",
    ctb="monochrome.ctb",
    cha_Y=2000,
    tol=50,
    max_attempts=5,
    check_retries=12,
    check_wait=3,
    dayin_folder=r"D:/Myprogramsystem/XT/dayinSHUCHU",
    input_pdf_folder=r"D:/Myprogramsystem/XT/inputPDF",
    dst_pdf_folder=r"D:/Myprogramsystem/XT/outPDF"   # ← 合并输出目录
):
    """
    0) 清空 dayinSHUCHU
    1) 打印 DWG → PDF
    2) 等待打印完成
    3) 复制 dayinSHUCHU → inputPDF（同名保留最新）
    4) 合并 inputPDF 下的 PDF 到 dst_pdf_folder
    """
    enable_debug()
    # 0) 清空打印输出目录
    node("0️⃣ 清理打印输出目录 {}", dayin_folder)
    dayin = Path(dayin_folder)
    if dayin.exists():
        for itm in dayin.iterdir():
            try:
                (shutil.rmtree(itm, ignore_errors=True) if itm.is_dir()
                 else itm.unlink(missing_ok=True))
            except Exception as e:
                node("⚠ 无法删除 {}: {}", itm.name, e)
    else:
        dayin.mkdir(parents=True, exist_ok=True)
    node("✅ dayinSHUCHU 已清空")

    # 同步清空 inputPDF / outPDF
    check_and_backup(input_pdf_folder)
    check_and_backup(dst_pdf_folder)

    # 1) 批量打印
    node("1️⃣ 开始批量打印 …")
    A3_print_all_files_in_folder_sp(
        layout_name, folder_path, ctb, cha_Y, tol, max_attempts
    )

    # 2) 等待打印完成
    node("2️⃣ 等待 CAD 打印完成 …")
    if not check_cad_ready(check_retries, check_wait):
        node("❌ CAD 打印未就绪，流程终止")
        return False
    node("✅ CAD 打印完成")

    # 3) 复制 dayin → inputPDF
    node("3️⃣ 2 秒后复制打印输出到 inputPDF 并去重")
    time.sleep(2)
    copy_dayin_to_input(dayin_folder, input_pdf_folder)

    # 4) 合并 PDF
    node("4️⃣ 2 秒后合并 inputPDF 下所有 PDF …")
    time.sleep(2)
    merge_and_combine_pdfs(
        src_root=input_pdf_folder,
        dst_root=dst_pdf_folder,         # ★ 明确目标目录
        combined_name="全部图纸合并.pdf",
        overwrite=True
    )

    node("🎉 全流程完成")
    disable_debug() 

    return True



def copy_dayin_to_input(dayin_folder, input_pdf_folder):
    """
    只复制 dayin_folder → input_pdf_folder。
    如果去掉末尾 '_YYYY-MM-DD-HH-MM' 时间戳后出现同名，只保留最新时间戳的那份。
    使用全局 node() 进行调试输出。
    """
    from datetime import datetime 
    ts_pat = re.compile(r"_(\d{4}-\d{2}-\d{2}-\d{2}-\d{2})$")

    dayin = Path(dayin_folder)
    dst   = Path(input_pdf_folder)
    dst.mkdir(parents=True, exist_ok=True)

    # ── 1) 找到每个“基名”最新的文件 / 文件夹 ──
    latest_map = {}           # {basename: (timestamp, Path)}

    for item in dayin.iterdir():
        if item.name.startswith("."):
            continue

        m = ts_pat.search(item.name)
        if m:
            ts_val = time.mktime(datetime.strptime(m.group(1),
                                     "%Y-%m-%d-%H-%M").timetuple())
            base = item.name[:m.start()]           # 去掉时间戳
        else:
            ts_val = item.stat().st_mtime
            base   = item.name                    # 本来就没时间戳

        if base not in latest_map or ts_val > latest_map[base][0]:
            latest_map[base] = (ts_val, item)

    # ── 2) 复制最新档 ──
    copied = 0
    for base, (_, src_path) in latest_map.items():
        target = dst / base
        try:
            if src_path.is_dir():
                if target.exists():
                    shutil.rmtree(target, ignore_errors=True)
                shutil.copytree(src_path, target, dirs_exist_ok=True)
            else:
                shutil.copy2(src_path, target)
            copied += 1
            node("   → 已复制最新 {}", base)
        except Exception as e:
            node("⚠ 复制 {} 失败: {}", src_path.name, e)

    node("📄 复制完成，共 {} 个（按基名去重）", copied)



#&&% 拆分 
def split_pdf(input_pdf, output_folder):
    doc = fitz.open(input_pdf)
    os.makedirs(output_folder, exist_ok=True)

    for i in range(len(doc)):
        output_path = os.path.join(output_folder, f"page_{i+1}.pdf")
        new_doc = fitz.open()
        new_doc.insert_pdf(doc, from_page=i, to_page=i)
        new_doc.save(output_path)
        new_doc.close()
        print(f"✅ 创建: {output_path}")

    doc.close()
    print("🎉 PDF 拆分完成！")



def measure_pdf_rect(pdf_path: str, zoom: float = 2.0, margin: int = 50):
    """
    在单页 PDF 上互动测量一个矩形区域的两个对角点坐标，
    并自动缩放显示以适应屏幕。

    :param pdf_path: PDF 文件路径（必须是单页）。
    :param zoom: 渲染分辨率倍数，越大越精细（亦会更大）。
    :param margin: 距离屏幕边缘的像素预留，默认 50px。
    :return: ((x1, y1), (x2, y2))，分别是两个角点在 PDF pt 单位下的坐标，
             原点在左下角。
    """
    # —— 1. 打开 PDF 并渲染成图像 —— 
    doc = fitz.open(pdf_path)
    if doc.page_count != 1:
        raise ValueError("PDF 必须是单页文件")
    page = doc[0]
    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat, alpha=False)
    arr = np.frombuffer(pix.samples, dtype=np.uint8)
    arr = arr.reshape(pix.height, pix.width, pix.n)
    if pix.n == 4:
        arr = cv2.cvtColor(arr, cv2.COLOR_BGRA2BGR)

    # —— 2. 读取屏幕分辨率，计算缩放比例 —— 
    root = tk.Tk()
    screen_w = root.winfo_screenwidth()
    screen_h = root.winfo_screenheight()
    root.destroy()

    max_w = screen_w - margin
    max_h = screen_h - margin
    # 如果图像本身就比屏幕小，则 scale_img=1
    scale_img = min(1.0, max_w / pix.width, max_h / pix.height)
    disp_w = int(pix.width * scale_img)
    disp_h = int(pix.height * scale_img)

    # —— 3. 显示可缩放窗口 —— 
    cv2.namedWindow("PDF Measure", cv2.WINDOW_NORMAL)
    cv2.resizeWindow("PDF Measure", disp_w, disp_h)
    # 在显示时直接让 OpenCV 再做一次缩放
    disp_img = cv2.resize(arr, (disp_w, disp_h), interpolation=cv2.INTER_AREA)

    points = []
    def _on_mouse(evt, x, y, flags, param):
        if evt == cv2.EVENT_LBUTTONDOWN:
            points.append((x, y))
            if len(points) >= 2:
                cv2.destroyAllWindows()

    cv2.setMouseCallback("PDF Measure", _on_mouse)
    cv2.imshow("PDF Measure", disp_img)
    cv2.waitKey(0)

    if len(points) < 2:
        raise RuntimeError("未获取到足够的点击点")

    # —— 4. 把窗口上的“显示坐标”映射回原始像素，再映射到 PDF pt —— 
    def disp2pt(dx, dy):
        # 先映射回原始像素位置
        px = dx / scale_img
        py = dy / scale_img
        # 再转成 PDF pt 坐标 (原点在左下)
        x_pt = px / zoom
        y_pt = (pix.height - py) / zoom
        return float(x_pt), float(y_pt)

    (dx1, dy1), (dx2, dy2) = points
    return disp2pt(dx1, dy1), disp2pt(dx2, dy2)

#&&% 测量pdf坐标

def measure_pdf_rect_zoomable(pdf_path: str,
                              zoom_pdf: float = 2.0,
                              margin: int = 50,
                              zoom_step: float = 1.25,
                              zoom_max: float = 20.0,
                              zoom_min: float = 1.0):
    """
    在单页 PDF 上交互测矩形：支持鼠标滚轮放大/缩小。

    返回值 ((x1, y1), (x2, y2)) —— 以 PDF pt 为单位，原点在左下角。
    """
    # ① 读 PDF → 渲染大图 (arr)
    doc = fitz.open(pdf_path)
    if doc.page_count != 1:
        raise ValueError("仅支持单页 PDF")
    page = doc[0]
    mat  = fitz.Matrix(zoom_pdf, zoom_pdf)
    pix  = page.get_pixmap(matrix=mat, alpha=False)
    arr  = np.frombuffer(pix.samples, np.uint8).reshape(pix.height, pix.width, pix.n)
    if pix.n == 4:                          # BGRA → BGR
        arr = cv2.cvtColor(arr, cv2.COLOR_BGRA2BGR)

    H_img, W_img = arr.shape[:2]

    # ② 计算窗口大小（让整个页面初始可见）
    root = tk.Tk();  screen_w, screen_h = root.winfo_screenwidth(), root.winfo_screenheight();  root.destroy()
    disp_w = min(W_img, screen_w - margin)
    disp_h = min(H_img, screen_h - margin)

    # ③ 视图状态
    view_scale = min(disp_w / W_img, disp_h / H_img)  # 初始整体缩放
    view_center = [W_img / 2, H_img / 2]              # 在原图中的中心
    clicks = []                                       # 存储原图像素坐标

    # ----- 工具函数 -----
    def clamp(v, lo, hi): return max(lo, min(v, hi))

    def refresh_window():
        """根据当前 view_center & view_scale 刷新显示"""
        # 原图中要截取的窗口大小
        roi_w = disp_w / view_scale
        roi_h = disp_h / view_scale
        x1 = clamp(int(view_center[0] - roi_w / 2), 0, W_img - int(roi_w))
        y1 = clamp(int(view_center[1] - roi_h / 2), 0, H_img - int(roi_h))
        x2, y2 = int(x1 + roi_w), int(y1 + roi_h)
        roi = arr[y1:y2, x1:x2]
        show = cv2.resize(roi, (disp_w, disp_h), interpolation=cv2.INTER_AREA)
        cv2.imshow("PDF Measure", show)
        return x1, y1, view_scale   # 返回当前左上原图坐标+缩放，用于坐标换算

    coords_cache = refresh_window()  # 初始化

    # ----- 鼠标回调 -----
    def on_mouse(event, x, y, flags, param):
        nonlocal view_scale, view_center, coords_cache
        # A. 左键点击 → 记录原图坐标
        if event == cv2.EVENT_LBUTTONDOWN:
            x1_img, y1_img, scale = coords_cache
            orig_x = x1_img + x / scale
            orig_y = y1_img + y / scale
            clicks.append((orig_x, orig_y))
            if len(clicks) >= 2:
                cv2.destroyAllWindows()
        # B. 滚轮缩放
        elif event == cv2.EVENT_MOUSEWHEEL:
            # flags>0 表示滚轮向前（放大）；<0 缩小
            delta = cv2.getMouseWheelDelta(flags) if hasattr(cv2, "getMouseWheelDelta") else (1 if flags > 0 else -1)
            zoom_dir = 1 if delta > 0 else -1
            new_scale = view_scale * (zoom_step ** zoom_dir)
            new_scale = clamp(new_scale, zoom_min, zoom_max)

            # 以光标为中心缩放：保证光标指向的原图点不变
            x1_img, y1_img, scale = coords_cache
            orig_x = x1_img + x / scale
            orig_y = y1_img + y / scale
            view_center = [orig_x, orig_y]
            view_scale = new_scale
            coords_cache = refresh_window()

    cv2.namedWindow("PDF Measure", cv2.WINDOW_NORMAL)
    cv2.resizeWindow("PDF Measure", disp_w, disp_h)
    cv2.setMouseCallback("PDF Measure", on_mouse)

    # 允许键盘 + / - 也能缩放
    while True:
        key = cv2.waitKey(50) & 0xFF
        if key in (27, ord('q')):   # ESC / q 退出
            cv2.destroyAllWindows();  raise RuntimeError("用户取消测量")
        elif key in (ord('+'), ord('=')):
            on_mouse(cv2.EVENT_MOUSEWHEEL, disp_w//2, disp_h//2, 1, None)
        elif key == ord('-'):
            on_mouse(cv2.EVENT_MOUSEWHEEL, disp_w//2, disp_h//2, -1, None)
        if len(clicks) >= 2:   # on_mouse 已关闭窗口
            break

    # ④ 原图像素 → PDF pt（左下）
    def img2pt(px, py):
        x_pt = px / zoom_pdf
        y_pt = (H_img - py) / zoom_pdf
        return float(x_pt), float(y_pt)

    (px1, py1), (px2, py2) = clicks
    return img2pt(px1, py1), img2pt(px2, py2)

def measure_pdf_rect_zoomable_pages(
    pdf_path: str | Path,
    page_no: int = 1,             # ← 新增：目标页码(1-based)
    zoom_pdf: float = 2.0,
    margin: int = 50,
    zoom_step: float = 1.25,
    zoom_max: float = 20.0,
    zoom_min: float = 1.0
) -> Tuple[Tuple[float, float], Tuple[float, float]]:
    """
    在多页 PDF 的指定 page_no 上交互测量矩形（支持鼠标滚轮缩放）。
    返回 ((x1, y1), (x2, y2)) —— 以 PDF pt 为单位，原点在左下角。
    """
    pdf_path = Path(pdf_path)
    doc      = fitz.open(pdf_path)
    if not (1 <= page_no <= doc.page_count):
        raise ValueError(f"page_no 越界 (1–{doc.page_count})")

    page = doc[page_no - 1]                       # 0-based 索引
    mat  = fitz.Matrix(zoom_pdf, zoom_pdf)
    pix  = page.get_pixmap(matrix=mat, alpha=False)
    arr  = np.frombuffer(pix.samples, np.uint8).reshape(pix.height, pix.width, pix.n)
    if pix.n == 4:
        arr = cv2.cvtColor(arr, cv2.COLOR_BGRA2BGR)

    H_img, W_img = arr.shape[:2]

    # ---------- 初始窗口尺寸 ----------
    root = tk.Tk()
    screen_w, screen_h = root.winfo_screenwidth(), root.winfo_screenheight()
    root.destroy()
    disp_w = min(W_img, screen_w - margin)
    disp_h = min(H_img, screen_h - margin)

    # ---------- 视图状态 ----------
    view_scale  = min(disp_w / W_img, disp_h / H_img)
    view_center = [W_img / 2, H_img / 2]
    clicks      = []

    # ---------- 辅助函数 ----------
    clamp = lambda v, lo, hi: max(lo, min(v, hi))

    def refresh_window():
        roi_w = disp_w / view_scale
        roi_h = disp_h / view_scale
        x1 = clamp(int(view_center[0] - roi_w / 2), 0, W_img - int(roi_w))
        y1 = clamp(int(view_center[1] - roi_h / 2), 0, H_img - int(roi_h))
        roi = arr[y1:y1+int(roi_h), x1:x1+int(roi_w)]
        show = cv2.resize(roi, (disp_w, disp_h), interpolation=cv2.INTER_AREA)
        cv2.imshow("PDF Measure", show)
        return x1, y1, view_scale

    coords_cache = refresh_window()

    # ---------- 鼠标回调 ----------
    def on_mouse(evt, x, y, flags, param):
        nonlocal view_scale, view_center, coords_cache
        if evt == cv2.EVENT_LBUTTONDOWN:
            x_left, y_top, scale = coords_cache
            orig_x = x_left + x / scale
            orig_y = y_top  + y / scale
            clicks.append((orig_x, orig_y))
            if len(clicks) >= 2:
                cv2.destroyAllWindows()
        elif evt == cv2.EVENT_MOUSEWHEEL:
            delta = cv2.getMouseWheelDelta(flags) if hasattr(cv2, "getMouseWheelDelta") else (1 if flags > 0 else -1)
            new_scale = clamp(view_scale * (zoom_step ** (1 if delta > 0 else -1)), zoom_min, zoom_max)
            x_left, y_top, scale = coords_cache
            orig_x = x_left + x / scale
            orig_y = y_top  + y / scale
            view_center[:] = [orig_x, orig_y]
            view_scale     = new_scale
            coords_cache   = refresh_window()

    cv2.namedWindow("PDF Measure", cv2.WINDOW_NORMAL)
    cv2.resizeWindow("PDF Measure", disp_w, disp_h)
    cv2.setMouseCallback("PDF Measure", on_mouse)

    # ---------- 键盘 + / - 缩放 ----------
    while True:
        key = cv2.waitKey(50) & 0xFF
        if key in (27, ord('q')):            # ESC / q 退出
            cv2.destroyAllWindows()
            doc.close()
            raise RuntimeError("用户取消测量")
        elif key in (ord('+'), ord('=')):
            on_mouse(cv2.EVENT_MOUSEWHEEL, disp_w//2, disp_h//2, 1, None)
        elif key == ord('-'):
            on_mouse(cv2.EVENT_MOUSEWHEEL, disp_w//2, disp_h//2, -1, None)
        if len(clicks) >= 2:
            break

    # ---------- 像素 → PDF pt（左下原点） ----------
    def img2pt(px, py):
        return float(px / zoom_pdf), float((H_img - py) / zoom_pdf)

    (px1, py1), (px2, py2) = clicks
    doc.close()
    return img2pt(px1, py1), img2pt(px2, py2)

#&&% 标准化pdf

def normalize_pdf_for_insert(src_pdf: str, dst_pdf: str | None = None):#标准化
    """
    消除页面初始 CTM + rotation + cropbox 偏移：
    1. page.wrap_contents()        → 让未来插图沿用页面坐标
    2. 反向旋转至 0°               → page.apply_transform()
    3. 平移到 (0,0)                → 同上
    写入 dst_pdf（None=原地增量）
    """
    src_pdf  = Path(src_pdf)
    out_pdf  = Path(dst_pdf) if dst_pdf else None

    doc = fitz.open(src_pdf)
    for p in doc:
        # 0) 把原始 CTM “包裹”进去
        p.wrap_contents()

        # 1) 若 /Rotate ≠0，反旋转并归零
        rot = p.rotation
        if rot:
            p.apply_transform(fitz.Matrix().prerotate(-rot))
            p.set_rotation(0)

        # 2) 若 CropBox 不从 (0,0) 起，整体左下归零
        cb = p.cropbox
        if cb.x0 or cb.y0:
            p.apply_transform(fitz.Matrix(1, 0, 0, 1, -cb.x0, -cb.y0))
            p.set_cropbox(fitz.Rect(0, 0, cb.width, cb.height))

    # 3) 保存
    if out_pdf:
        doc.save(out_pdf, deflate=True)
    else:
        doc.saveIncr()
    doc.close()

#某些情况要使用另一个标准化函数
def normalize_pdf_for_insert_1(src_pdf, out_pdf):
    """
    兼容旧版 PyMuPDF：把 CropBox=MediaBox、Rotate=0，
    尽量清掉内容流里的 CTM；若函数不存在则跳过。
    """
    src_pdf, out_pdf = map(Path, (src_pdf, out_pdf))
    doc = fitz.open(src_pdf)

    for page in doc:
        # —— 1. 把 CropBox 拉满 —— #
        try:
            mb = page.mediabox                # 老版属性
        except AttributeError:
            mb = page.rect                    # 再退一步
        try:
            # 若 rect 超界会报错，再退回用 mediabox
            page.set_cropbox(mb)
        except Exception:
            try:
                page.cropbox = mb
            except Exception:
                pass

        # —— 2. 清旋转 —— #
        try:
            page.set_rotation(0)
        except AttributeError:
            try:
                page.rotation = 0
            except Exception:
                pass

        # —— 3. 包装内容流，去 CTM —— #
        try:
            page.wrap_contents()
        except Exception:
            # 极老版没有 wrap_contents，也只能跳过
            pass

    # —— 4. 全量重写并压缩 —— #
    doc.save(out_pdf, incremental=False, deflate=True, garbage=4)
    doc.close()
    print(f"✅ 归一化完成 → {out_pdf}")
















#&&% 插入png到pdf文件中心

def insert_png_center(
    png_path, pdf_path, page_no: int,
    out_pdf=None, fit_ratio=0.65, dpi_default=300
):
    """
    前提：若目标 PDF 可能含初始 CTM，请先跑 normalize_pdf_for_insert().
    之后本函数即可保证 PNG 正立 & 居中。
    """
    png_path, pdf_path = map(Path, (png_path, pdf_path))
    # 1. PNG 尺寸
    with Image.open(png_path) as im:
        w_px, h_px = im.size
        dpi = im.info.get("dpi", (dpi_default, dpi_default))[0]
    w_pt, h_pt = w_px * 72 / dpi, h_px * 72 / dpi
    aspect = h_pt / w_pt

    # 2. PDF 页面
    doc = fitz.open(pdf_path)
    p    = doc[page_no]
    r    = p.rect                      # 此时坐标系已规整
    w_img = r.width * fit_ratio
    h_img = w_img * aspect
    if h_img > r.height * fit_ratio:
        h_img = r.height * fit_ratio
        w_img = h_img / aspect
    x0 = r.x0 + (r.width  - w_img)/2
    y0 = r.y0 + (r.height - h_img)/2
    box = fitz.Rect(x0, y0, x0+w_img, y0+h_img)

    p.insert_image(
        box,
        filename=str(png_path),
        keep_proportion=True,
        overlay=True
    )
    if out_pdf:
        doc.save(out_pdf, deflate=True)
    else:
        doc.saveIncr()
    doc.close()



#&&% 带b比例控制的png插入中心函数

def insert_png_center(
    png_path,
    pdf_path,
    page_no: int,
    out_pdf: str | None = None,
    fit_ratio: float = 0.65,
    k: float = 1.0,               # ← 新增缩放系数
    dpi_default: int = 300
):
    """
    将 png 插入 pdf 指定页面并居中。
    
    • fit_ratio: 相对页面尺寸的初次缩放比例；
    • k        : 在 fit_ratio 基础上再乘以的系数，用来微调大小。
      (k < 1 缩小，k > 1 放大，但不会突破 fit_ratio 的“边界”)
    
    若目标 PDF 可能有初始 CTM，请先跑 normalize_pdf_for_insert().
    """
    png_path, pdf_path = map(Path, (png_path, pdf_path))

    # 1. 读取 PNG 尺寸
    with Image.open(png_path) as im:
        w_px, h_px = im.size
        dpi = im.info.get("dpi", (dpi_default, dpi_default))[0]
    aspect = h_px / w_px          # 高宽比

    # 2. 打开 PDF 页面
    doc = fitz.open(pdf_path)
    page = doc[page_no]
    page_rect = page.rect         # 左下原点坐标

    # 3. 计算最终插入尺寸
    base_w = page_rect.width  * fit_ratio * k
    base_h = base_w * aspect
    if base_h > page_rect.height * fit_ratio * k:
        base_h = page_rect.height * fit_ratio * k
        base_w = base_h / aspect

    x0 = page_rect.x0 + (page_rect.width  - base_w) / 2
    y0 = page_rect.y0 + (page_rect.height - base_h) / 2
    box = fitz.Rect(x0, y0, x0 + base_w, y0 + base_h)

    # 4. 插图
    page.insert_image(
        box,
        filename=str(png_path),
        keep_proportion=True,
        overlay=True
    )

    # 5. 保存
    if out_pdf:
        doc.save(out_pdf, deflate=True)
    else:
        doc.saveIncr()
    doc.close()


def insert_png_center_pages_bili(
    png_path,
    pdf_path,
    out_pdf=None,
    n1: int = 1,
    n2: int | None = None,       # None=最后一页
    fit_ratio: float = 0.65,
    k: float = 1.0,              # 额外缩放系数
    dpi_default: int = 300
):
    """
    在 pdf 的 n1…n2 页（含）居中插入 png_path。
    - fit_ratio: 图片相对页面宽度的基准比例
    - k        : 在 fit_ratio 基础上的再缩放 (k<1 缩小, k>1 放大)
    """
    png_path, pdf_path = map(Path, (png_path, pdf_path))

    # ── 1. PNG 尺寸 & 高宽比 ──
    with Image.open(png_path) as im:
        w_px, h_px = im.size
        dpi = im.info.get("dpi", (dpi_default, dpi_default))[0]
    aspect = h_px / w_px                     # 高/宽

    # ── 2. 打开 PDF ──
    doc = fitz.open(pdf_path)
    if n2 is None:
        n2 = doc.page_count
    if not (1 <= n1 <= n2 <= doc.page_count):
        raise ValueError("页码范围非法")

    # ── 3. 批量插入 ──
    for idx in range(n1 - 1, n2):            # 0-based
        page = doc[idx]
        r = page.rect
        # 计算目标宽高
        base_w = r.width  * fit_ratio * k
        base_h = base_w * aspect
        if base_h > r.height * fit_ratio * k:
            base_h = r.height * fit_ratio * k
            base_w = base_h / aspect
        x0 = r.x0 + (r.width  - base_w)/2
        y0 = r.y0 + (r.height - base_h)/2
        box = fitz.Rect(x0, y0, x0+base_w, y0+base_h)
        page.insert_image(box, filename=str(png_path),
                          keep_proportion=True, overlay=True)
        node("Page {:>3}: 插入 {} ({}×{})".format(idx+1, png_path.name,
                                                  round(base_w,1), round(base_h,1)))

    # ── 4. 保存 ──
    if out_pdf:
        doc.save(out_pdf, deflate=True)
        node("✅ 已输出 {}", out_pdf)
    else:
        doc.saveIncr()
        node("✅ 已写回原 PDF")
    doc.close()

def insert_png_center_pages_ex(
    png_path,
    pdf_path,
    out_pdf=None,
    n1: int = 1,
    n2: int | None = None,
    fit_ratio: float = 0.65,
    k: float = 1.0,          # ← 额外缩放系数
    dx_pt: float = 0.0,      # ← 水平平移 (pt)
    dy_pt: float = 0.0,      # ← 垂直平移 (pt)
    dpi_default: int = 300
):
    """
    批量把 PNG 插入 PDF 指定页面 (n1..n2)，等比缩放并可微调：
      - fit_ratio: 相对页面宽度基准比例
      - k        : 在 fit_ratio 上再乘系数 (k<1 缩小; k>1 放大)
      - dx_pt    : 额外水平位移 (pt)
      - dy_pt    : 额外垂直位移 (pt)
    """
    png_path, pdf_path = map(Path, (png_path, pdf_path))

    # 1. PNG 高宽比
    with Image.open(png_path) as im:
        w_px, h_px = im.size
        dpi = im.info.get("dpi", (dpi_default, dpi_default))[0]
    aspect = h_px / w_px

    # 2. 打开 PDF
    doc = fitz.open(pdf_path)
    if n2 is None:
        n2 = doc.page_count
    if not (1 <= n1 <= n2 <= doc.page_count):
        raise ValueError("页码范围非法")

    # 3. 循环插图
    for page_idx in range(n1 - 1, n2):          # 0-based
        page = doc[page_idx]
        r = page.rect

        w_target = r.width  * fit_ratio * k
        h_target = w_target * aspect
        if h_target > r.height * fit_ratio * k:
            h_target = r.height * fit_ratio * k
            w_target = h_target / aspect

        # 居中位置 + 手动偏移
        x0 = r.x0 + (r.width  - w_target)/2 + dx_pt
        y0 = r.y0 + (r.height - h_target)/2 + dy_pt
        box = fitz.Rect(x0, y0, x0 + w_target, y0 + h_target)

        page.insert_image(
            box,
            filename=str(png_path),
            keep_proportion=True,
            overlay=True
        )
        node("Page {:>3}: 插入 {}  尺寸 {:.1f}×{:.1f} pt  偏移 ({:+.1f},{:+.1f})"
             .format(page_idx+1, png_path.name, w_target, h_target, dx_pt, dy_pt))

    # 4. 保存
    if out_pdf:
        doc.save(out_pdf, deflate=True)
        node("✅ 已输出 {}", out_pdf)
    else:
        doc.saveIncr()
        node("✅ 已写回原 PDF")
    doc.close()



def raster_insert_png(
    pdf_path,
    png_path,
    page_no:    int   = 0,
    out_pdf:    str | None = None,
    fit_ratio:  float = 0.65,
    k:          float = 1.0,
    dpi:        int   = 300,
) -> None:
    """
    ▸ 将 `png_path` 插入 `pdf_path` 第 `page_no` 页 (A3)，
      先把页面光栅化再合成，100% 兼容扫描 PDF / 受损 PDF。

    - fit_ratio : 目标 PNG 相对页面尺寸的初始缩放
    - k         : 在 fit_ratio 基础上的微调系数 (<1 缩小，>1 放大)
    - dpi       : 渲染 / 输出分辨率，建议 300 及以上
    """

    pdf_path, png_path = map(Path, (pdf_path, png_path))

    # ---------- 1. 读取要插入的 PNG ----------
    with Image.open(png_path).convert("RGBA") as im_overlay:
        ov_w, ov_h = im_overlay.size
        ov_aspect  = ov_h / ov_w

        # ---------- 2. 打开 PDF，渲染指定页 ----------
        doc   = fitz.open(pdf_path)
        page  = doc[page_no]
        rect  = page.rect                  # PDF 页面矩形 (pt)
        zoom  = dpi / 72.0                 # 1pt = 1/72 in
        mat   = fitz.Matrix(zoom, zoom)
        pix   = page.get_pixmap(matrix=mat, alpha=False)  # base RGB

        base = Image.frombytes(
            "RGB",
            [pix.width, pix.height],
            pix.samples
        ).convert("RGBA")                  # → RGBA 方便合成

    # ---------- 3. 计算插入图在光栅坐标的大小 ----------
    max_w = base.width  * fit_ratio * k
    max_h = base.height * fit_ratio * k

    tgt_w = min(max_w, max_h / ov_aspect)
    tgt_h = tgt_w * ov_aspect

    # Pillow 只能用整数像素
    tgt_w_i = int(round(tgt_w))
    tgt_h_i = int(round(tgt_h))

    im_overlay = im_overlay.resize((tgt_w_i, tgt_h_i), Image.LANCZOS)

    # ---------- 4. 计算位置 (居中) ----------
    x0 = (base.width  - tgt_w_i) // 2
    y0 = (base.height - tgt_h_i) // 2

    base.paste(im_overlay, (x0, y0), im_overlay)

    # ---------- 5. 把合成后的 PNG 封装为 PDF ----------
    # 存 PNG 到内存
    png_bytes = io.BytesIO()
    base.save(png_bytes, format="PNG")
    png_bytes.seek(0)

    # 新建 PDF，尺寸必须与原页一致
    patched = fitz.open()                   # 空文档
    new_page = patched.new_page(
        width = rect.width,
        height= rect.height
    )
    new_page.insert_image(
        rect, stream=png_bytes.read()
    )

    # ---------- 6. 输出 ----------
    if out_pdf:
        # 把原文档复制页到新文档→合并
        merged = fitz.open()
        for i in range(len(doc)):
            merged.insert_pdf(patched if i == page_no else doc, from_page=i, to_page=i)
        merged.save(out_pdf, garbage=4, deflate=True)
        merged.close()
    else:
        # 直接覆盖原页（先删除旧页再拼接）
        doc.delete_page(page_no)
        doc.insert_pdf(patched, from_page=0, to_page=0, start_at=page_no)
        doc.save(pdf_path.with_suffix(".incr.pdf"), garbage=4, deflate=True)  # 避免破坏原件
    doc.close()
    patched.close()



def _build_overlay(
    base_size: tuple[int, int],
    overlay_img: Image.Image,
    fit_ratio: float,
    k: float,
) -> Image.Image:
    """
    根据 base_size 计算缩放后的 overlay 图像并返回新对象（RGBA）。
    base_size: (width, height) —— 光栅基底尺寸
    """
    base_w, base_h = base_size
    ov_w, ov_h     = overlay_img.size
    aspect         = ov_h / ov_w

    max_w = base_w * fit_ratio * k
    max_h = base_h * fit_ratio * k

    # 保持比例缩放
    tgt_w = min(max_w, max_h / aspect)
    tgt_h = tgt_w * aspect

    tgt_w_i = int(round(tgt_w))
    tgt_h_i = int(round(tgt_h))

    return overlay_img.resize((tgt_w_i, tgt_h_i), Image.LANCZOS)

# 批量处理图片格式的pdf ，常用于盖章 ，个别能在WPS直接插入
def raster_insert_png_all(
    pdf_path ,
    png_path ,
    out_pdf:    str | None = None,
    fit_ratio:  float = 0.65,
    k:          float = 1.0,
    dpi:        int   = 300,
    pages = None,   # None=全部；否则给页号列表/范围[int]
    verbose:    bool = True,
) -> None:
    """
    把 `png_path` 代表的 PNG 叠加到 `pdf_path` 的所有（或指定）页面：
    1. 每页按 dpi 渲染为 RGBA 位图；
    2. 居中贴上 PNG（fit_ratio·k 缩放）；
    3. 再封装为同尺寸 PDF 页面；
    4. 整本保存为 `out_pdf`（默认在源文件名后加 _marked）。
    """
    pdf_path, png_path = map(Path, (pdf_path, png_path))

    # 预加载待插入 PNG（保持原透明度）
    with Image.open(png_path).convert("RGBA") as im_src:
        overlay_src = im_src.copy()

    src_doc = fitz.open(pdf_path)

    if pages is None:
        pages_to_do = range(len(src_doc))
    else:
        pages_to_do = pages

    out_doc = fitz.open()                       # 新文档

    zoom = dpi / 72.0
    mat  = fitz.Matrix(zoom, zoom)

    for i in range(len(src_doc)):
        page = src_doc[i]
        rect = page.rect

        # ---- 若本页需要处理，渲染并叠图 ----
        if i in pages_to_do:
            pix = page.get_pixmap(matrix=mat, alpha=False)
            base = Image.frombytes("RGB", [pix.width, pix.height], pix.samples).convert("RGBA")

            # 生成与当前页光栅尺寸匹配的 overlay
            overlay = _build_overlay(base.size, overlay_src, fit_ratio, k)

            x0 = (base.width  - overlay.width)  // 2
            y0 = (base.height - overlay.height) // 2
            base.paste(overlay, (x0, y0), overlay)

            # 把合成结果写入 BytesIO
            buf = io.BytesIO()
            base.save(buf, format="PNG")
            buf.seek(0)

            new_page = out_doc.new_page(width=rect.width, height=rect.height)
            new_page.insert_image(rect, stream=buf.read())

            if verbose:
                print(f"✅  Page {i + 1}/{len(src_doc)} done.")
        else:
            # 不处理的页直接复制
            out_doc.insert_pdf(src_doc, from_page=i, to_page=i)

    # ------- 保存 -------
    if out_pdf is None:
        out_pdf = str(pdf_path.with_name(pdf_path.stem + "_marked.pdf"))

    out_doc.save(out_pdf, garbage=4, deflate=True)
    out_doc.close()
    src_doc.close()

    if verbose:
        print(f"🎉 All finished. Output → {out_pdf}")




def _build_overlay_1(
    base_size: tuple[int, int],
    overlay_img: Image.Image,
    fit_ratio: float,
    k: float,
) -> Image.Image:
    """把 overlay_img 按 fit_ratio·k 缩放到适合 base_size 的大小"""
    base_w, base_h   = base_size
    ov_w, ov_h       = overlay_img.size
    aspect           = ov_h / ov_w

    max_w = base_w * fit_ratio * k
    max_h = base_h * fit_ratio * k

    tgt_w = min(max_w, max_h / aspect)
    tgt_h = tgt_w * aspect

    return overlay_img.resize(
        (int(round(tgt_w)), int(round(tgt_h))),
        Image.LANCZOS
    )

#可移动位置的批量处理图片格式的pdf
def raster_insert_png_all_1(
    pdf_path:   Union[str, Path],
    png_path:   Union[str, Path],
    out_pdf:    str | None = None,
    *,
    fit_ratio:  float = 0.65,
    k:          float = 1.0,
    dpi:        int   = 300,
    dx:         float = 0.0,             # ↔ 向右为正，单位 mm
    dy:         float = 0.0,             # ↑ 向上为正，单位 mm
    pages:      Sequence[int] | None = None,
    verbose:    bool = True,
) -> None:
    """
    把 png 叠加到整本 (或指定页) PDF，
    `dx`, `dy` 以 **毫米** 为单位：正值→向右/向上移动，负值→向左/向下移动。
    """
    pdf_path, png_path = map(Path, (pdf_path, png_path))

    # ---------- 预加载 PNG ----------
    with Image.open(png_path).convert("RGBA") as im_src:
        overlay_src = im_src.copy()

    src_doc = fitz.open(pdf_path)

    if pages is None:
        pages_to_do = range(len(src_doc))
    else:
        pages_to_do = pages

    out_doc = fitz.open()

    zoom = dpi / 72.0
    mat  = fitz.Matrix(zoom, zoom)
    mm_to_px = lambda mm: int(round(mm * 72 / 25.4 * zoom))  # mm → px

    dx_px = mm_to_px(dx)
    dy_px = mm_to_px(dy)

    for i in range(len(src_doc)):
        if i not in pages_to_do:
            # 原样复制未处理页
            out_doc.insert_pdf(src_doc, from_page=i, to_page=i)
            continue

        page = src_doc[i]
        rect = page.rect

        pix = page.get_pixmap(matrix=mat, alpha=False)
        base = Image.frombytes(
            "RGB", [pix.width, pix.height], pix.samples
        ).convert("RGBA")

        # —— 缩放 overlay 并计算粘贴坐标 ——
        overlay = _build_overlay_1(base.size, overlay_src, fit_ratio, k)

        x0 = (base.width  - overlay.width)  // 2 + dx_px
        y0 = (base.height - overlay.height) // 2 + dy_px

        base.paste(overlay, (x0, y0), overlay)

        # —— 写入新 PDF 页 ——
        buf = io.BytesIO()
        base.save(buf, format="PNG")
        buf.seek(0)

        new_page = out_doc.new_page(width=rect.width, height=rect.height)
        new_page.insert_image(rect, stream=buf.read())

        if verbose:
            print(f"✅  Page {i + 1}/{len(src_doc)} done.")

    # ---------- 保存 ----------
    if out_pdf is None:
        out_pdf = str(pdf_path.with_name(pdf_path.stem + "_marked.pdf"))

    out_doc.save(out_pdf, garbage=4, deflate=True)
    out_doc.close()
    src_doc.close()

    if verbose:
        print(f"🎉 Output saved → {out_pdf}")






























# ---------- 几何辅助 ----------
def rect_size(rect: fitz.Rect) -> Tuple[float, float]:
    return rect.width, rect.height

def rect_center(rect: fitz.Rect) -> Tuple[float, float]:
    return (rect.x0 + rect.x1) / 2, (rect.y0 + rect.y1) / 2


# ---------- 读取页面最新插入图片 ----------
def get_latest_image_rect(pdf_path: str, page_no: int = 0) -> fitz.Rect:
    """
    返回指定页 *xref 最大*（通常就是“最后插入”）那张图片的 Rect。
    如果页面没有位图图片 → 抛 ValueError。
    """
    pdf_path = Path(pdf_path)
    with fitz.open(pdf_path) as doc:
        page = doc[page_no]
        imgs = page.get_images(full=True)   # [(xref, smth, …), …]
        if not imgs:
            raise ValueError("页面上没有位图图片")

        # ❶ 按 xref 选取最大（最后插入）
        latest_xref = max(imgs, key=lambda t: t[0])[0]

        # ❷ 同一张图片可能被多次引用，这里取第一次出现的位置
        rects = page.get_image_rects(latest_xref)
        if not rects:
            raise RuntimeError("找不到图片的矩形")
        return rects[0]


#&&% 放置图片到目标区域

def place_png_in_pdf_region_bottom_origin(
    pdf_path, png_path,
    corner1_bl, corner2_bl,        # ← 你手里那两点：原点在左下
    output_pdf_path
):
    doc  = fitz.open(pdf_path)
    page = doc[0]

    # 页面高度（pt）
    H = page.rect.height

    # 把“左下原点坐标”翻成“左上原点坐标”
    x1, y1_bl = corner1_bl
    x2, y2_bl = corner2_bl
    xmin, xmax = sorted([x1, x2])
    ymin_bl, ymax_bl = sorted([y1_bl, y2_bl])      # bottom-origin
    ymin = H - ymax_bl    # 翻转
    ymax = H - ymin_bl

    rect = fitz.Rect(xmin, ymin, xmax, ymax)
    page.insert_image(rect, filename=png_path,
                      keep_proportion=False, overlay=True)

    doc.save(output_pdf_path)
    doc.close()



def batch_place_png_in_pdf_region_bottom_origin(
    pdf_path: str,
    png_path: str,
    corner1_bl: tuple[float, float],
    corner2_bl: tuple[float, float],
    output_pdf_path: str,
    start_page: int = 1,
    end_page: int | None = None,
    keep_proportion: bool = False,
    overlay: bool = True,
):
    """
    将 PNG 拉伸填满多页 PDF 中统一矩形区域 (start_page ≤ page ≤ end_page)，
    坐标以“左下角原点”给出，与 measure_pdf_rect 的输出一致。

    参数
    ----
    pdf_path          : 源 PDF（可多页）
    png_path          : 要插入的 PNG
    corner1_bl        : 第一个角点 (x, y) —— 左下原点坐标
    corner2_bl        : 第二个角点 (x, y) —— 左下原点坐标
    output_pdf_path   : 输出 PDF 路径
    start_page        : 起始页，1-based，默认为 1
    end_page          : 结束页，1-based，默认为最后一页
    keep_proportion   : 是否保持原 PNG 纵横比（默认 False：各向独立拉伸）
    overlay           : True 覆盖页面内容；False 在下层
    """
    doc = fitz.open(pdf_path)

    # 页码范围有效性
    if start_page < 1 or start_page > doc.page_count:
        raise ValueError(f"start_page 超出范围 (1-{doc.page_count})")
    if end_page is None:
        end_page = doc.page_count
    if end_page < start_page or end_page > doc.page_count:
        raise ValueError(f"end_page 超出范围 (start_page-{doc.page_count})")

    # 统一计算矩形（仍用 bottom-origin 坐标）
    x1, y1_bl = corner1_bl
    x2, y2_bl = corner2_bl
    xmin, xmax = sorted([x1, x2])
    ymin_bl, ymax_bl = sorted([y1_bl, y2_bl])

    # 逐页插图
    for page_no in range(start_page - 1, end_page):          # 转 0-based
        page = doc[page_no]
        H = page.rect.height

        # 翻成 top-origin
        ymin = H - ymax_bl
        ymax = H - ymin_bl
        rect = fitz.Rect(xmin, ymin, xmax, ymax)

        page.insert_image(
            rect,
            filename=png_path,
            keep_proportion=keep_proportion,
            overlay=overlay,
        )

    # 保存
    doc.save(output_pdf_path)
    doc.close()





#&&% 比较分析目标和当前区域

# ---------- 核心：对比函数 ----------
def compare_insert_vs_target(
    pdf_path: str,
    page_no: int,
    target_rect: fitz.Rect
) -> dict:
    """
    返回：
    {
        "insert_rect":  Rect(...),
        "target_rect":  Rect(...),
        "size_diff":    (dW, dH),      # 插入尺寸 - 目标尺寸
        "center_diff":  (dX, dY)       # 插入中心 - 目标中心
    }
    """
    ins_rect = get_latest_image_rect(pdf_path, page_no)

    # ① 尺寸差
    w_ins, h_ins = rect_size(ins_rect)
    w_tgt, h_tgt = rect_size(target_rect)
    size_diff = (w_ins - w_tgt, h_ins - h_tgt)

    # ② 位置差
    cx_ins, cy_ins = rect_center(ins_rect)
    cx_tgt, cy_tgt = rect_center(target_rect)
    center_diff = (cx_ins - cx_tgt, cy_ins - cy_tgt)

    return {
        "insert_rect": ins_rect,
        "target_rect": target_rect,
        "size_diff":   size_diff,
        "center_diff": center_diff,
    }

def rect_size(rect: fitz.Rect) -> Tuple[float, float]:
    return rect.width, rect.height

def rect_center(rect: fitz.Rect) -> Tuple[float, float]:
    return (rect.x0 + rect.x1) / 2, (rect.y0 + rect.y1) / 2


#&&% 获取最后插入的位图

def get_latest_image_rect(pdf_path: str, page_no: int = 0) -> fitz.Rect:
    """
    返回指定页 *xref 最大*（通常就是“最后插入”）那张图片的 Rect。
    如果页面没有位图图片 → 抛 ValueError。
    """
    pdf_path = Path(pdf_path)
    with fitz.open(pdf_path) as doc:
        page = doc[page_no]
        imgs = page.get_images(full=True)   # [(xref, smth, …), …]
        if not imgs:
            raise ValueError("页面上没有位图图片")

        # ❶ 按 xref 选取最大（最后插入）
        latest_xref = max(imgs, key=lambda t: t[0])[0]

        # ❷ 同一张图片可能被多次引用，这里取第一次出现的位置
        rects = page.get_image_rects(latest_xref)
        if not rects:
            raise RuntimeError("找不到图片的矩形")
        return rects[0]




def align_last_image_to_rect(
    pdf_path: str,
    page_no: int,
    target_rect: fitz.Rect,
    mode: Literal["contain", "cover"] = "contain",
    remove_old: bool = True
) -> None:
    """
    把“最后插入”的那张图片移动 + 缩放到 target_rect。

    参数
    ----
    pdf_path   : PDF 文件（已 wrap / normalize 过）
    page_no    : 页码 (0-based)
    target_rect: 拟对齐到的 Rect
    mode       : "contain" → 图片全部显示在矩形内
                 "cover"   → 填满矩形，必要时裁切
    remove_old : True 则删除旧图片；False 则保留旧图，新图会覆盖在上层
    """
    doc = fitz.open(pdf_path)
    page = doc[page_no]

    # --- 1. 找到最后一张位图（xref 最大） ---
    imgs = page.get_images(full=True)
    if not imgs:
        raise ValueError("该页没有位图图片")
    latest_xref = max(imgs, key=lambda t: t[0])[0]

    # 提取图片字节
    img_info = doc.extract_image(latest_xref)
    img_bytes = img_info["image"]
    img_ext   = img_info["ext"]  # png / jpeg / …

    # --- 2. 如需删除旧图片 ---
    if remove_old:
        try:
            page.delete_image(latest_xref)        # PyMuPDF ≥1.22
        except AttributeError:
            page._deleteObject(latest_xref)       # 兼容旧版

    # --- 3. 计算插入矩形（保持比例 contain/cover） ---
    box_w, box_h = target_rect.width, target_rect.height
    # 原图像素宽高
    w_px  = img_info["width"]
    h_px  = img_info["height"]
    aspect = h_px / w_px
    box_aspect = box_h / box_w

    if (mode == "contain" and aspect >= box_aspect) or \
       (mode == "cover"   and aspect <= box_aspect):
        h_img = box_h
        w_img = h_img / aspect
    else:
        w_img = box_w
        h_img = w_img * aspect

    cx, cy = (target_rect.x0 + target_rect.x1)/2, (target_rect.y0 + target_rect.y1)/2
    new_rect = fitz.Rect(
        cx - w_img/2, cy - h_img/2,
        cx + w_img/2, cy + h_img/2
    )

    # --- 4. 重新插图 ---
    page.insert_image(
        new_rect,
        stream = img_bytes,
        keep_proportion = False,     # 自己已算好 w_img/h_img
        overlay = True
    )

    doc.saveIncr()
    doc.close()

#比较两个矩形，计算缩放和移动


def _sorted_rect(p1, p2):
    x0, x1 = sorted((p1[0], p2[0]))
    y0, y1 = sorted((p1[1], p2[1]))
    return fitz.Rect(x0, y0, x1, y1)

def get_latest_image_rect(pdf_path: str, page_index: int) -> fitz.Rect | None:
    with fitz.open(pdf_path) as d:
        imgs = d[page_index].get_images(full=True)
        if not imgs:
            return None
        xref = max(imgs, key=lambda t: t[0])[0]
        rects = d[page_index].get_image_rects(xref)
        return rects[0] if rects else None


# --------- 主函数 ----------
def _rect_sorted(p1, p2) -> fitz.Rect:
    """确保 x0<x1, y0<y1"""
    x0, x1 = sorted((p1[0], p2[0]))
    y0, y1 = sorted((p1[1], p2[1]))
    return fitz.Rect(x0, y0, x1, y1)



#&&% 移动缩放png到目标位置

def move_and_scale_last_png(
    pdf_path: str,
    page_no: int,
    p1: Tuple[float, float],
    p2: Tuple[float, float],
    out_pdf: str | None = None,
    fit_mode: Literal["contain", "cover"] = "contain",
) -> None:
    """
    将“最后插入”的位图 (PNG/JPG) 移动并等比缩放到 p1‒p2 定义的矩形区域。

    参数
    ----
    pdf_path : PDF 文件路径
    page_no  : 页码 (1-based)
    p1, p2   : 目标矩形对角点 (x, y)，PDF 左下坐标系，单位 pt
    out_pdf  : None=原地增量保存；否则写入新文件
    fit_mode : "contain" (完整显示) | "cover" (填满后裁切)
    """
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(pdf_path)

    doc  = fitz.open(pdf_path)
    page = doc[page_no - 1]                      # 转 0-based

    # ——— 1. 找最后插入的位图 XObject ———
    imgs = page.get_images(full=True)
    if not imgs:
        doc.close()
        raise ValueError(f"第 {page_no} 页没有位图图片")
    latest_xref = max(imgs, key=lambda t: t[0])[0]

    # ——— 2. 提取图片字节与尺寸 ———
    info      = doc.extract_image(latest_xref)
    img_bytes = info["image"]
    iw, ih    = info["width"], info["height"]
    aspect    = ih / iw

    # ——— 3. 删除旧图片 ———
    try:
        page.delete_image(latest_xref)          # PyMuPDF ≥1.22
    except AttributeError:
        page._deleteObject(latest_xref)         # 旧版兼容

    # ——— 4. 若页面仍带 CropBox 偏移，回补到 p1,p2 ———
    cb = page.cropbox
    if cb.x0 or cb.y0:                          # 只要有偏移就补偿
        dx, dy = -cb.x0, -cb.y0                 # 把内容曾经移下来的量加回
        p1 = (p1[0] + dx, p1[1] + dy)
        p2 = (p2[0] + dx, p2[1] + dy)

    # ——— 5. 构造目标矩形（保证宽高正值） ———
    x0, x1 = sorted((p1[0], p2[0]))
    y0, y1 = sorted((p1[1], p2[1]))
    rect_tgt = fitz.Rect(x0, y0, x1, y1)
    if rect_tgt.is_empty:
        doc.close()
        raise ValueError(f"目标矩形无效：{rect_tgt}")

    # ——— 6. 计算插入矩形并插图 ———
    mode = fit_mode.strip().lower()
    if mode not in ("contain", "cover"):
        mode = "contain"

    if mode == "contain":
        # PyMuPDF 自带等比缩放并居中
        page.insert_image(
            rect_tgt,
            stream=img_bytes,
            keep_proportion=True,
            overlay=True,
        )
    else:  # cover
        bw, bh = rect_tgt.width, rect_tgt.height
        scale  = max(bw / iw, bh / ih)          # 填满所需放大倍数
        new_w, new_h = iw * scale, ih * scale
        cx, cy = rect_tgt.x0 + bw / 2, rect_tgt.y0 + bh / 2
        rect_ins = fitz.Rect(
            cx - new_w / 2, cy - new_h / 2,
            cx + new_w / 2, cy + new_h / 2,
        )
        page.insert_image(
            rect_ins,
            stream=img_bytes,
            keep_proportion=False,              # 我们已保证等比
            overlay=True,
        )

    # ——— 7. 保存 ———
    if out_pdf:
        doc.save(out_pdf, deflate=True, garbage=3)
    else:
        doc.saveIncr()
    doc.close()




# ------------- 你现成的工具：取“最后插入”位图矩形 -------------
def get_latest_image_rect(pdf_path: str, page_index: int) -> fitz.Rect | None:
    with fitz.open(pdf_path) as d:
        imgs = d[page_index].get_images(full=True)
        if not imgs:
            return None
        xref = max(imgs, key=lambda t: t[0])[0]
        rects = d[page_index].get_image_rects(xref)
        return rects[0] if rects else None
# --------------------------------------------------------------


#一般移动

def move_bitmap_xy(
    pdf_path: str,
    page_no: int,
    dst_center: Tuple[float, float],
    out_pdf: str | None = None,
) -> None:
    """
    将【最后插入】位图的中心移动到 dst_center，并打印前后坐标与位移量。

    参数
    ----
    pdf_path   : PDF 路径
    page_no    : 1-based 页码
    dst_center : (x, y) 目标中心 (pt，左下原点)
    out_pdf    : None=增量保存；否则另存为新文件
    """
    pdf_in = Path(pdf_path)
    if not pdf_in.exists():
        raise FileNotFoundError(pdf_in)

    # —— 1. 读取当前中心 —— #
    rect = get_latest_image_rect(str(pdf_in), page_no - 1)    # 0-based
    if rect is None:
        raise ValueError(f"第 {page_no} 页没有位图图片")

    cx_old = (rect.x0 + rect.x1) / 2
    cy_old = (rect.y0 + rect.y1) / 2
    cx_new, cy_new = dst_center
    dx = cx_new - cx_old
    dy = cy_new - cy_old

    print(f"📍 当前中心: ({cx_old:.2f}, {cy_old:.2f}) pt")
    print(f"🎯 目标中心: ({cx_new:.2f}, {cy_new:.2f}) pt")
    print(f"↔ 需平移 Δx={dx:.2f} pt, Δy={dy:.2f} pt")

    if dx == 0 and dy == 0:
        print("✅ 位图已在目标位置，无需移动。")
        return

    # —— 2. 打开 PDF 做移动 —— #
    doc  = fitz.open(pdf_in)
    page = doc[page_no - 1]

    # 获取对应 xref 再操作
    imgs = page.get_images(full=True)
    last_xref = max(imgs, key=lambda t: t[0])[0]

    if hasattr(page, "apply_transform"):
        # 新版：直接平移
        page.apply_transform(fitz.Matrix(1, 0, 0, 1, dx, dy), clip=rect)
    else:
        # 旧版：删旧 → 新矩形 → 重插
        img_bytes = doc.extract_image(last_xref)["image"]
        try:
            page.delete_image(last_xref)
        except AttributeError:
            page._deleteObject(last_xref)
        rect_new = fitz.Rect(rect.x0 + dx, rect.y0 + dy,
                             rect.x1 + dx, rect.y1 + dy)
        page.insert_image(rect_new, stream=img_bytes,
                          keep_proportion=False, overlay=True)

    # —— 3. 保存 —— #
    if out_pdf:
        doc.save(out_pdf, deflate=True, garbage=3)
    else:
        doc.saveIncr()

    # —— 4. 再次打印新中心 —— #
    new_rect = get_latest_image_rect(str(out_pdf or pdf_in), page_no - 1)
    if new_rect:
        cx_post = (new_rect.x0 + new_rect.x1) / 2
        cy_post = (new_rect.y0 + new_rect.y1) / 2
        print(f"✅ 移动后中心: ({cx_post:.2f}, {cy_post:.2f}) pt")

    doc.close()

#&&% 异常文件的批处理

def insert_png_center_pages(
    png_path: str | Path,
    pdf_path: str | Path,
    page_start: int = 1,
    page_end:   Optional[int] = None,
    out_pdf:    str | Path | None = None,
    fit_ratio:  float = 0.65,
    dpi_default: int = 300,
):
    """
    把 PNG 居中插到 page_start‒page_end，每页都等比缩放到 fit_ratio*页面宽度。
    """
    png_path, pdf_path = map(Path, (png_path, pdf_path))
    # —— 1. 读 PNG 尺寸 —— #
    with Image.open(png_path) as im:
        w_px, h_px = im.size
        dpi = im.info.get("dpi", (dpi_default, dpi_default))[0]
    w_pt, h_pt = w_px * 72 / dpi, h_px * 72 / dpi
    aspect     = h_pt / w_pt

    # —— 2. 逐页插图 —— #
    doc = fitz.open(pdf_path)
    if page_end is None:
        page_end = doc.page_count
    if not (1 <= page_start <= page_end <= doc.page_count):
        raise ValueError("page_start/page_end 越界")

    for idx in range(page_start - 1, page_end):
        p = doc[idx]
        r = p.rect
        w_img = r.width * fit_ratio
        h_img = w_img * aspect
        if h_img > r.height * fit_ratio:
            h_img = r.height * fit_ratio
            w_img = h_img / aspect
        x0 = r.x0 + (r.width  - w_img) / 2
        y0 = r.y0 + (r.height - h_img) / 2
        box = fitz.Rect(x0, y0, x0 + w_img, y0 + h_img)
        p.insert_image(box, filename=str(png_path),
                       keep_proportion=True, overlay=True)

    # —— 3. 保存 —— #
    out_pdf = out_pdf or pdf_path
    doc.save(out_pdf, incremental=False, deflate=True, garbage=4)
    doc.close()
    print(f"✅ PNG 已插入第 {page_start}–{page_end} 页，输出 → {out_pdf}")


def move_and_scale_last_png_pages(
    pdf_path: str | Path,
    page_start: int = 1,
    page_end:   Optional[int] = None,
    p1: Tuple[float, float] = (0, 0),
    p2: Tuple[float, float] = (100, 100),
    out_pdf: str | Path | None = None,
    fit_mode: str = "contain",
):
    """
    把 *每页* “最后插入的位图” 按 p1‒p2 统一缩放+居中。
    """
    pdf_path = Path(pdf_path)
    doc = fitz.open(pdf_path)
    if page_end is None:
        page_end = doc.page_count
    if not (1 <= page_start <= page_end <= doc.page_count):
        raise ValueError("page_start/page_end 越界")

    mode = fit_mode.strip().lower()
    if mode not in {"contain", "cover"}:
        mode = "contain"

    for idx in range(page_start - 1, page_end):
        page = doc[idx]
        imgs = page.get_images(full=True)
        if not imgs:
            continue
        latest_xref = max(imgs, key=lambda t: t[0])[0]
        info   = doc.extract_image(latest_xref)
        iw, ih = info["width"], info["height"]
        aspect = ih / iw
        # 新矩形
        x0, x1 = sorted((p1[0], p2[0]))
        y0, y1 = sorted((p1[1], p2[1]))
        rect_tgt = fitz.Rect(x0, y0, x1, y1)
        # 删除旧
        try:
            page.delete_image(latest_xref)
        except AttributeError:
            page._deleteObject(latest_xref)
        # 计算插入矩形
        if mode == "contain":
            page.insert_image(rect_tgt, stream=info["image"],
                              keep_proportion=True, overlay=True)
        else:
            bw, bh = rect_tgt.width, rect_tgt.height
            scale  = max(bw / iw, bh / ih)
            new_w, new_h = iw * scale, ih * scale
            cx, cy = rect_tgt.x0 + bw/2, rect_tgt.y0 + bh/2
            rect_ins = fitz.Rect(cx-new_w/2, cy-new_h/2,
                                 cx+new_w/2, cy+new_h/2)
            page.insert_image(rect_ins, stream=info["image"],
                              keep_proportion=False, overlay=True)

    out_pdf = out_pdf or pdf_path
    doc.save(out_pdf, incremental=False, deflate=True, garbage=4)
    doc.close()
    print(f"✅ 位图已缩放/移动第 {page_start}–{page_end} 页 → {out_pdf}")


def move_bitmap_xy_pages(
    pdf_path: str | Path,
    page_start: int = 1,
    page_end:   Optional[int] = None,
    dst_center: Tuple[float, float] = (0, 0),
    out_pdf: str | Path | None = None,
):
    """
    将每页最后插入位图的中心移动到同一个 dst_center。
    """
    pdf_path = Path(pdf_path)
    doc = fitz.open(pdf_path)
    if page_end is None:
        page_end = doc.page_count
    if not (1 <= page_start <= page_end <= doc.page_count):
        raise ValueError("page_start/page_end 越界")

    for idx in range(page_start - 1, page_end):
        page = doc[idx]
        imgs = page.get_images(full=True)
        if not imgs:
            continue

        last_xref = max(imgs, key=lambda t: t[0])[0]

        # —— ① 用 get_image_rects 取矩形 —— #
        rects = page.get_image_rects(last_xref)
        if not rects:
            continue                       # xref 找不到矩形就跳过
        rect = rects[0]

        # —— ② 计算平移量 —— #
        cx_old = (rect.x0 + rect.x1) / 2
        cy_old = (rect.y0 + rect.y1) / 2
        dx = dst_center[0] - cx_old
        dy = dst_center[1] - cy_old

        if dx == 0 and dy == 0:
            continue

        # —— ③ 平移 —— #
        if hasattr(page, "apply_transform"):
            page.apply_transform(fitz.Matrix(1, 0, 0, 1, dx, dy), clip=rect)
        else:
            info = doc.extract_image(last_xref)
            try:
                page.delete_image(last_xref)
            except AttributeError:
                page._deleteObject(last_xref)
            rect_new = fitz.Rect(rect.x0+dx, rect.y0+dy,
                                 rect.x1+dx, rect.y1+dy)
            page.insert_image(rect_new, stream=info["image"],
                              keep_proportion=False, overlay=True)


    out_pdf = out_pdf or pdf_path
    doc.save(out_pdf, incremental=False, deflate=True, garbage=4)
    doc.close()
    print(f"✅ 位图中心已移动到 {dst_center}，第 {page_start}–{page_end} 页 → {out_pdf}")


#旋转位图

##看不到，手工
def rotate_last_bitmap_pages_pillow(
    pdf_path: str | Path,
    out_pdf:  str | Path | None = None,
    page_start: int = 1,
    page_end:   Optional[int] = None,
    angle: int = 90
):
    pdf_path = Path(pdf_path)
    out_pdf  = Path(out_pdf) if out_pdf else pdf_path
    doc = fitz.open(pdf_path)

    if page_end is None:
        page_end = doc.page_count
    if not (1 <= page_start <= page_end <= doc.page_count):
        raise ValueError("页码范围越界")

    angle = angle % 360
    if angle not in (90, 180, 270):
        raise ValueError("angle 只能是 90 / 180 / 270")

    for idx in range(page_start - 1, page_end):
        page = doc[idx]
        imgs = page.get_images(full=True)
        if not imgs:
            continue

        # —— 找“最后一个有矩形”的位图 —— #
        rect0 = None
        img_bytes = None
        picked_xref = None
        for xref, *_ in sorted(imgs, key=lambda t: t[0], reverse=True):
            rects = page.get_image_rects(xref)
            if rects:
                rect0 = rects[0]
                picked_xref = xref
                try:
                    img_bytes = doc.extract_image(xref)["image"]
                except Exception:
                    # fallback 用 Pixmap 渲染
                    img_bytes = fitz.Pixmap(doc, xref).tobytes("png")
                break

        if rect0 is None:
            # 这一页所有 xref 都无矩形，跳过
            print(f"Page {idx+1}: ⚠ 没有可定位的位图，已跳过")
            continue

        # —— Pillow 实际旋转 —— #
        pil_img = Image.open(io.BytesIO(img_bytes))
        pil_rot = pil_img.rotate(-angle, expand=True)
        buf = io.BytesIO()
        pil_rot.save(buf, format="PNG")
        new_bytes = buf.getvalue()

        # —— 删除旧图像 —— #
        try:
            page.delete_image(picked_xref)
        except AttributeError:
            page._deleteObject(picked_xref)

        # —— 计算新框 (宽高互换或保持) —— #
        cx = (rect0.x0 + rect0.x1) / 2
        cy = (rect0.y0 + rect0.y1) / 2
        if angle in (90, 270):
            new_w, new_h = rect0.height, rect0.width
        else:
            new_w, new_h = rect0.width, rect0.height
        rect_new = fitz.Rect(cx - new_w/2, cy - new_h/2,
                             cx + new_w/2, cy + new_h/2)

        # —— 插回旋转后图片 —— #
        page.insert_image(rect_new, stream=new_bytes,
                          keep_proportion=False, overlay=True)

    doc.save(out_pdf, incremental=False, deflate=True, garbage=4)
    doc.close()
    print(f"✅ 已旋转 {angle}° 并保持面积，第 {page_start}–{page_end} 页 → {out_pdf}")


##复制位图

def copy_bitmap_to_pages(
    pdf_path: str | Path,
    src_page: int,
    dst_pages: Sequence[int] | Literal["all"] = "all",
    pick_mode: Literal["largest", "latest"] = "largest",
    out_pdf: str | Path | None = None,
    overlay: bool = True,
) -> None:
    """
    将 src_page 页选中的位图复制到 dst_pages 的同一矩形。

    参数
    ----
    pdf_path  : 源 PDF
    src_page  : 源页页码 (1-based)，从这里取位图
    dst_pages : 目标页序列；"all" = 除 src_page 外全部页
    pick_mode : "largest" = 面积最大那张位图；"latest" = xref 最大那张
    out_pdf   : 输出路径；None = 覆盖原文件
    overlay   : True 覆盖上层；False 置于底层
    """
    pdf_path = Path(pdf_path)
    out_pdf  = Path(out_pdf) if out_pdf else pdf_path
    doc = fitz.open(pdf_path)

    if not (1 <= src_page <= doc.page_count):
        raise ValueError("src_page 越界")
    src_pg = doc[src_page - 1]

    # —— 1. 在源页选图 —— #
    imgs = src_pg.get_images(full=True)
    if not imgs:
        raise ValueError(f"源页 {src_page} 没有位图")

    if pick_mode == "latest":
        # xref 最大
        imgs_sorted = sorted(imgs, key=lambda t: t[0], reverse=True)
    else:  # largest
        # 先按面积降序，再按 xref
        def img_area(t):
            rs = src_pg.get_image_rects(t[0])
            return max((r.width * r.height for r in rs), default=0)
        imgs_sorted = sorted(imgs, key=lambda t: (img_area(t), t[0]), reverse=True)

    # 找到第一张“有矩形”的
    for xref, *_ in imgs_sorted:
        rects = src_pg.get_image_rects(xref)
        if rects:
            rect_ref = rects[0]
            break
    else:
        raise RuntimeError("源页找不到可定位的位图")

    # —— 2. 取图片字节 —— #
    try:
        img_bytes = doc.extract_image(xref)["image"]
    except Exception:                         # 老版 fallback
        img_bytes = fitz.Pixmap(doc, xref).tobytes("png")

    # —— 3. 目标页集合 —— #
    if dst_pages == "all":
        dst_pages = [i for i in range(1, doc.page_count + 1) if i != src_page]
    else:
        dst_pages = [p for p in dst_pages if p != src_page]

    # —— 4. 批量插图 —— #
    for pno in dst_pages:
        if not (1 <= pno <= doc.page_count):
            print(f"⚠ 目标页 {pno} 越界，已跳过")
            continue
        pg = doc[pno - 1]
        if (abs(pg.rect.width - src_pg.rect.width) > 1e-2 or
            abs(pg.rect.height - src_pg.rect.height) > 1e-2):
            print(f"⚠ Page {pno}: 页面尺寸不同，已跳过")
            continue
        pg.insert_image(rect_ref, stream=img_bytes,
                        keep_proportion=False, overlay=overlay)

    # —— 5. 保存 —— #
    doc.save(out_pdf, incremental=False, deflate=True, garbage=4)
    doc.close()
    print(f"✅ 位图已从第 {src_page} 页复制到 {len(dst_pages)} 页 → {out_pdf}")


#&&% 提取pdf文件的字体字号

def extract_page_fonts(pdf_path: str, page_number: int) -> Dict[str, Set[float]]:
    """
    提取 pdf_path 指定文件第 page_number 页（1 基）上的所有文字 span 的
    字体资源名和字号（pt），包括 SHX 和非 SHX。

    返回值
    -----
    {
      "F1"       : {9.0, 10.0},
      "PDFSHX"   : {8.54},
      "Helvetica": {12.0},
      ...
    }

    参数
    ----
    pdf_path   : PDF 文件路径
    page_number: 要提取的页码（从 1 开始）
    """
    # 打开文档
    doc = fitz.open(pdf_path)
    try:
        # 1-based 转 0-based
        page = doc.load_page(page_number - 1)
    except IndexError:
        doc.close()
        raise ValueError(f"页码 {page_number} 超出范围，共有 {doc.page_count} 页")

    fonts: Dict[str, Set[float]] = {}
    # 用 dict 模式提取所有 text spans
    text_dict = page.get_text("dict")
    for block in text_dict["blocks"]:
        if block["type"] != 0:  # 0 = text block
            continue
        for line in block["lines"]:
            for span in line["spans"]:
                name = span["font"]
                size = round(span["size"], 2)  # 保留两位小数
                fonts.setdefault(name, set()).add(size)

    doc.close()
    return fonts



#检测是否有有效的shx字体

def debug_list_annots(pdf_path: str):
    reader = PdfReader(pdf_path)
    for pno, page in enumerate(reader.pages, 1):
        annots = page.get("/Annots", [])
        if not annots:
            continue
        print(f"--- Page {pno} ---")
        for ref in annots:
            a = ref.get_object()
            st = a.get("/Subtype")
            raw = a.get("/Contents")
            # 部分文件 Contents 是 ByteString，需要 decode
            try:
                txt = raw.decode('utf-16') if isinstance(raw, bytes) else raw
            except Exception:
                txt = repr(raw)[:60]
            print(st, ":", txt)

            
##debug_list_annots(r"D:/Myprogramsystem/TextImages/PDF测试/A0x.pdf")
##Multiple definitions in dictionary at byte 0x1c2e for key /PageMode
##--- Page 1 ---
##/Square : A0
##/Square : 测试
##/Square : PDF
##/Square : 合并编辑
##


from pathlib import Path
from typing  import Union, List
from pypdf   import PdfReader

def extract_shx_comments(pdf_path: Union[str, Path]) -> List[str]:
    """
    提取 AutoCAD 导出 PDF 中附加在批注里的 SHX 文字
    兼容 /FreeText, /Text, /Square, /Circle …
    """
    reader = PdfReader(str(pdf_path))
    texts  = []

    # 可根据需要再加其它批注类型
    good_types = {"/FreeText", "/Text", "/Square", "/Circle"}

    for p_no, page in enumerate(reader.pages, 1):
        for ref in page.get("/Annots", []):
            annot = ref.get_object()
            if annot.get("/Subtype") not in good_types:
                continue

            raw = annot.get("/Contents")
            if not raw:
                continue

            # 字符串或 ByteString → 解码
            if isinstance(raw, bytes):
                try:
                    txt = raw.decode("utf-16le")   # AutoCAD 常用 UTF‑16LE
                except UnicodeDecodeError:
                    txt = raw.decode(errors="replace")
            else:
                txt = str(raw)

            texts.append(f"[p{p_no}] {txt.strip()}")

    return texts

##if __name__ == "__main__":
##    shx_texts = extract_shx_comments(r"D:/Drawings/A0_SHX.pdf")
##    print("\n".join(shx_texts[:20]))  

#sddff
# HHHHHHHH   






##提取shx字体和位置
def extract_shx_texts_with_position(pdf_path):
    doc = fitz.open(pdf_path)
    results = []

    for page_number in range(len(doc)):
        page = doc[page_number]
        annot_list = page.annots()
        if annot_list:
            for annot in annot_list:
                content = annot.info.get("content", "")
                if "[p1]" in content:
                    results.append({
                        "page": page_number + 1,
                        "text": content.strip(),
                        "x": annot.rect.x0,
                        "y": annot.rect.y1
                    })
    return results


##提取shx文字内容和坐标

import fitz
from PyPDF2 import PdfReader
import re
from typing import Any, Dict, List, Set


def extract_pdf_fonts_and_shx_annotations(
    pdf_path: str,
    page_number: int
) -> Dict[str, Any]:
    """
    提取 PDF 第 page_number 页的信息，包括：

    1. text_fonts: 真实文字 span 的字体和字号
       Dict[font_name, Set[size_pt, ...]]
    2. resource_fonts: 本页所有 /Font 资源名集合
       Set[font_name, ...]
    3. shx_annots: SHX 批注中的文字内容、位置及字体信息
       List[{
         text: str,
         x: float,
         y: float,
         font_name: Optional[str],
         font_size: Optional[float],
         subtype: str
       }]

    参数
    ----
    pdf_path   : PDF 文件路径
    page_number: 要提取的页码，从 1 开始

    返回
    ----
    {
      "text_fonts":     Dict[str, Set[float]],
      "resource_fonts": Set[str],
      "shx_annots":     List[Dict[str,Any]]
    }
    """
    # —— 1. 文本 span 的字体 & 大小 —— 
    doc = fitz.open(pdf_path)
    try:
        page = doc.load_page(page_number - 1)
    except Exception:
        doc.close()
        raise ValueError(f"页码 {page_number} 超出范围，共 {doc.page_count} 页")

    text_fonts: Dict[str, Set[float]] = {}
    td = page.get_text("dict")
    for block in td["blocks"]:
        if block["type"] != 0:
            continue
        for line in block["lines"]:
            for span in line["spans"]:
                fn = span["font"]
                sz = round(span["size"], 2)
                text_fonts.setdefault(fn, set()).add(sz)

    # —— 2. 本页所有 /Font 资源 —— 
    resource_fonts: Set[str] = {font_rec[1] for font_rec in page.get_fonts()}

    doc.close()

    # —— 3. SHX 批注的提取 —— 
    reader = PdfReader(pdf_path)
    pdf_page = reader.pages[page_number - 1]
    shx_annots: List[Dict[str, Any]] = []
    good_types = {"/FreeText", "/Text", "/Square", "/Circle"}

    for ref in pdf_page.get("/Annots", []):
        annot = ref.get_object()
        subtype = annot.get("/Subtype")
        if subtype not in good_types:
            continue

        # 内容
        raw = annot.get("/Contents", "")
        if isinstance(raw, (bytes, bytearray)):
            try:
                text = raw.decode("utf-16le")
            except Exception:
                text = raw.decode(errors="ignore")
        else:
            text = str(raw)
        text = text.strip()

        # 位置（取左下角 x, 右上角 y）
        rect = annot.get("/Rect", [])
        x = rect[0] if len(rect) >= 1 else None
        y = rect[3] if len(rect) >= 4 else None

        # 默认外观 (/DA) 中的字体名和字号
        da = annot.get("/DA", "")
        m = re.search(r"/([^\s]+)\s+([\d.]+)\s+Tf", da)
        font_name = m.group(1) if m else None
        font_size = float(m.group(2)) if m else None

        shx_annots.append({
            "text":       text,
            "x":          x,
            "y":          y,
            "font_name":  font_name,
            "font_size":  font_size,
            "subtype":    subtype
        })

    return {
        "text_fonts":     text_fonts,
        "resource_fonts": resource_fonts,
        "shx_annots":     shx_annots
    }


def extract_pdf_fonts_and_shx_annotations_x(
    pdf_path: str,
    page_number: int
) -> Dict[str, Any]:
    """
    提取 PDF 第 page_number 页的信息，包括：

    1. text_fonts: 真实文字 span 的字体和字号
    2. resource_fonts: 本页所有 /Font 资源名集合
    3. shx_annots: SHX 注释（FreeText/Square/Circle）中的文字内容、
         位置、以及从注释 /AP 流中解析出的字体名和字号

    返回
    ----
    {
      "text_fonts":     Dict[str, Set[float]],
      "resource_fonts": Set[str],
      "shx_annots":     List[{
           text: str,
           x: float, y: float,
           font_name: Optional[str],
           font_size: Optional[float],
           subtype: str
         }]
    }
    """
    # ——— 1. 文字 span 级别字体/字号 ———
    doc = fitz.open(pdf_path)
    try:
        page = doc.load_page(page_number - 1)
    except Exception:
        doc.close()
        raise ValueError(f"页码 {page_number} 超出范围（共 {doc.page_count} 页）")

    text_fonts: Dict[str, Set[float]] = {}
    for block in page.get_text("dict")["blocks"]:
        if block["type"] != 0: continue
        for line in block["lines"]:
            for span in line["spans"]:
                fn = span["font"]
                sz = round(span["size"], 2)
                text_fonts.setdefault(fn, set()).add(sz)

    # ——— 2. 本页所有 /Font 资源 ———
    resource_fonts: Set[str] = { r[1] for r in page.get_fonts() }
    doc.close()

    # ——— 3. SHX 注释的解析 ———
    reader = PdfReader(pdf_path)
    pdf_page = reader.pages[page_number - 1]
    shx_annots: List[Dict[str, Any]] = []
    good = {"/FreeText", "/Text", "/Square", "/Circle"}

    for ref in pdf_page.get("/Annots", []):
        annot = ref.get_object()
        subtype = annot.get("/Subtype")
        if subtype not in good:
            continue

        # 3.1 内容
        raw = annot.get("/Contents", "")
        if isinstance(raw, (bytes, bytearray)):
            try:
                text = raw.decode("utf-16le")
            except:
                text = raw.decode(errors="ignore")
        else:
            text = str(raw)
        text = text.strip()

        # 3.2 位置
        rect = annot.get("/Rect", [])
        x = rect[0] if len(rect) >= 1 else None
        y = rect[3] if len(rect) >= 4 else None

        # 3.3 优先尝试注释字典的 /DA
        da = annot.get("/DA", "")
        m = re.search(r"/([^\s]+)\s+([\d.]+)\s+Tf", da)
        font_name = m.group(1) if m else None
        font_size = float(m.group(2)) if m else None

        # 3.4 如果 /DA 没给信息，就去注释外观流 /AP→/N 里找
        if font_name is None or font_size is None:
            ap = annot.get("/AP", {})
            n_ref = ap.get("/N")
            if n_ref:
                xobj = n_ref.get_object()               # Form XObject
                stream = xobj.get_data()               # bytes of the content stream
                m2 = re.search(rb"/([^\s]+)\s+([\d.]+)\s+Tf", stream)
                if m2:
                    font_name = m2.group(1).decode()
                    font_size = float(m2.group(2))

        shx_annots.append({
            "text":       text,
            "x":          x,
            "y":          y,
            "font_name":  font_name,
            "font_size":  font_size,
            "subtype":    subtype
        })

    return {
        "text_fonts":     text_fonts,
        "resource_fonts": resource_fonts,
        "shx_annots":     shx_annots
    }


#&&&% PDF处理



#&&&&%% 第十四部分  文本制作  




#&&&% 彩平生成流程


#&&% CAD和PNG的映射

import os
import cv2
import numpy as np
from PIL import Image


def draw_polygons_on_image(
    polygons,
    cad_bounds: tuple = ((0, 0), (118900, 84100)),
    image_path: str = r"D:/Myprogramsystem/BaiduSyncdisk/Text_production_services/Color_flat_generation/Basic_documents/Colorflat_production.png",
    output_path: str | None = None,
    line_color: tuple[int, int, int] = (0, 0, 255),
    thickness: int = 2
) -> np.ndarray:
    """
    在图片上按 CAD 坐标绘制封闭多边形。

    参数：
      polygons:     单个多边形或多组多边形。
                    - 单一多边形：[(x,y,z), ...]
                    - 多组：[[...], [...], ...]
      cad_bounds:   CAD 区域 ((min_x,min_y),(max_x,max_y))
      image_path:   输入位图路径；默认指向 Colorflat_production.png
      output_path:  输出图像路径；若为 None，则在原文件同目录生成 "<原名>_result<扩展名>"
      line_color:   BGR 画线颜色
      thickness:    线宽

    返回：
      返回绘制后的 OpenCV 图像数组，并将结果保存到 output_path
    """
    # 1. 单多边形自动包装
    if polygons and isinstance(polygons[0], tuple) and isinstance(polygons[0][0], (int, float)):
        polygons = [polygons]

    # 2. 验证输入路径
    image_path = os.path.normpath(image_path)
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"图像文件不存在：{image_path}")

    # 3. Unicode 路径读取
    data = np.fromfile(image_path, dtype=np.uint8)
    img = cv2.imdecode(data, cv2.IMREAD_UNCHANGED)
    if img is None:
        raise IOError(f"无法解码图像：{image_path}")

    h, w = img.shape[:2]
    (min_x, min_y), (max_x, max_y) = cad_bounds
    cad_w = max_x - min_x
    cad_h = max_y - min_y

    # 4. 计算比例
    scale_x = (w - 1) / cad_w
    scale_y = (h - 1) / cad_h

    # 5. 绘制多边形
    for poly in polygons:
        pts = []
        for pt in poly:
            # 直接取整 CAD 浮点
            x = int(pt[0])
            y = int(pt[1])
            # CAD→像素映射
            px = int(round((x - min_x) * scale_x))
            py = int(round((y - min_y) * scale_y))
            # 翻转 y 轴：CAD 原点左下 → OpenCV 原点左上
            py = h - 1 - py
            pts.append([px, py])
        pts = np.array(pts, dtype=np.int32)
        # 连接顶点
        for i in range(len(pts) - 1):
            cv2.line(img, tuple(pts[i]), tuple(pts[i + 1]), color=line_color, thickness=thickness)

    # 6. 保存结果
    if output_path is None:
        dir_, fname = os.path.split(image_path)
        name, ext = os.path.splitext(fname)
        output_path = os.path.join(dir_, f"{name}_result{ext}")
    cv2.imencode(ext, img)[1].tofile(output_path)

    return img




#&&% CAD上的多边形在PNG的填充


def draw_polygons_on_image_hatch(
    polygons,
    cad_bounds: tuple = ((0, 0), (118900, 84100)),
    image_path: str = r"D:/Myprogramsystem/BaiduSyncdisk/Text_production_services/Color_flat_generation/Basic_documents/Colorflat_production.png",
    fill_image_path: str | None = None,
    k: float = 1.0,
    output_path: str | None = None,
    line_color: tuple[int, int, int] = (0, 0, 255),
    thickness: int = 2
) -> np.ndarray:
    """
    在底图上按 CAD 坐标绘制封闭多边形，并可选地用纹理图填充。

    参数：
      polygons:         单个或多组多边形，
                        - 单一：[(x,y,z), ...]
                        - 多组：[[...], [...], ...]
      cad_bounds:       CAD 区域 ((min_x,min_y),(max_x,max_y))
      image_path:       底图路径，默认指向 Colorflat_production.png
      fill_image_path:  填充纹理图路径（支持 jpg/png/webp），若为 None 则不填充
      k:                纹理图缩放系数（>0），仅在 fill_image_path 不为 None 时生效
      output_path:      保存结果路径；若 None，则在原图目录生成 “<原名>_result<扩展名>”
      line_color:       多边形轮廓色（BGR）
      thickness:        轮廓线宽

    返回：
      返回绘制并填充后的 OpenCV 图像数组，并保存到输出路径。
    """
    # 1. 支持单一多边形自动包装
    if polygons and isinstance(polygons[0], tuple) and isinstance(polygons[0][0], (int, float)):
        polygons = [polygons]

    # 2. 读取底图，支持 Unicode 路径
    image_path = os.path.normpath(image_path)
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"底图不存在：{image_path}")
    data = np.fromfile(image_path, dtype=np.uint8)
    base_img = cv2.imdecode(data, cv2.IMREAD_UNCHANGED)
    if base_img is None:
        raise IOError(f"无法解码底图：{image_path}")

    # 获取尺寸与 CAD 比例
    h, w = base_img.shape[:2]
    (min_x, min_y), (max_x, max_y) = cad_bounds
    cad_w = max_x - min_x
    cad_h = max_y - min_y
    scale_x = (w - 1) / cad_w
    scale_y = (h - 1) / cad_h

    # 3. CAD → 像素 坐标映射
    pixel_polygons = []
    for poly in polygons:
        pts = []
        for x_f, y_f, *_ in poly:
            x = int(round(x_f))
            y = int(round(y_f))
            px = int(round((x - min_x) * scale_x))
            py = int(round((y - min_y) * scale_y))
            # 翻转 y 轴：CAD 原点左下 → 图像左上
            py = h - 1 - py
            pts.append([px, py])
        pixel_polygons.append(np.array(pts, dtype=np.int32))

    # 4. 若需纹理填充，读取、缩放、平铺、蒙版替换
    if fill_image_path:
        ft_path = os.path.normpath(fill_image_path)
        if not os.path.exists(ft_path):
            raise FileNotFoundError(f"纹理图不存在：{ft_path}")
        # 尝试 OpenCV 解码
        fd = np.fromfile(ft_path, dtype=np.uint8)
        tex = cv2.imdecode(fd, cv2.IMREAD_COLOR)
        if tex is None:
            # Pillow 回退
            pil_img = Image.open(ft_path).convert("RGB")
            tex = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
        # 缩放纹理
        if k != 1.0:
            new_w = max(1, int(tex.shape[1] * k))
            new_h = max(1, int(tex.shape[0] * k))
            tex = cv2.resize(tex, (new_w, new_h), interpolation=cv2.INTER_LINEAR)
        th, tw = tex.shape[:2]
        cols = int(np.ceil(w / tw))
        rows = int(np.ceil(h / th))
        tiled = np.tile(tex, (rows, cols, 1))[:h, :w]
        # 创建蒙版
        mask = np.zeros((h, w), dtype=np.uint8)
        for pts in pixel_polygons:
            cv2.fillPoly(mask, [pts], 255)
        # 使用 NumPy 逐通道替换
        for c in range(3):
            base_img[..., c] = np.where(mask == 255, tiled[..., c], base_img[..., c])

    # 5. 绘制多边形轮廓
    for pts in pixel_polygons:
        cv2.polylines(base_img, [pts], isClosed=False, color=line_color, thickness=thickness)

    # 6. 保存结果
    dir_, orig_name = os.path.split(image_path)
    name, ext = os.path.splitext(orig_name)
    if not output_path:
        output_path = os.path.join(dir_, f"{name}_result{ext}")
    cv2.imencode(ext, base_img)[1].tofile(output_path)
    return base_img


#&&% 将cad多段线转为有序坐标点列表


"""
get_unique_vertices_from_pl_com(pl_com)


"""

#&&% 将cad树木图块转变为图片插入

def insert_icons_on_image(
    centers,
    icon_path: str,
    cad_bounds: tuple = ((0, 0), (118900, 84100)),
    image_path: str = r"D:/Myprogramsystem/BaiduSyncdisk/Text_production_services/Color_flat_generation/Basic_documents/Colorflat_production.png",
    k: float = 1.0,
    output_path: str | None = None,
    transparent_background: bool = False
) -> np.ndarray:
    """
    根据 CAD 中的中心点列表，在底图上插入图块（图标），
    并可选生成透明背景，仅保留图标。

    参数：
      centers:                中心点列表 [(x,y,z), ...]
      icon_path:              图标路径（支持 PNG/JPG/WEBP，含 alpha 通道）
      cad_bounds:             CAD 区域 ((min_x,min_y),(max_x,max_y))
      image_path:             底图路径（仅用于获取尺寸）
      k:                      图标缩放系数 (>0)
      output_path:            输出文件路径；None 则保存到原图目录，文件名后缀“_icons”
      transparent_background: 若 True，输出图像仅含图标，其余透明（RGBA）；
                              若 False，则在原底图上合成（输出 BGRA）。

    返回：
      返回合成后的图像数组（BGRA），并保存到 output_path。
    """
    # 1. 读取底图并获取尺寸
    img_path = os.path.normpath(image_path)
    if not os.path.exists(img_path):
        raise FileNotFoundError(f"底图不存在：{img_path}")
    data = np.fromfile(img_path, dtype=np.uint8)
    img0 = cv2.imdecode(data, cv2.IMREAD_UNCHANGED)
    if img0 is None:
        raise IOError(f"无法解码底图：{img_path}")
    h, w = img0.shape[:2]

    # 2. 准备输出图像
    if transparent_background:
        out_img = np.zeros((h, w, 4), dtype=np.uint8)
    else:
        # 将底图转换为 BGRA
        if img0.ndim == 2:
            out_img = cv2.cvtColor(img0, cv2.COLOR_GRAY2BGRA)
        elif img0.shape[2] == 3:
            out_img = cv2.cvtColor(img0, cv2.COLOR_BGR2BGRA)
        else:
            out_img = img0.copy()

    # 3. 计算 CAD -> 像素 映射比例
    (min_x, min_y), (max_x, max_y) = cad_bounds
    cad_w = max_x - min_x
    cad_h = max_y - min_y
    sx = (w - 1) / cad_w
    sy = (h - 1) / cad_h

    # 4. 读取并缩放图标
    icon_norm = os.path.normpath(icon_path)
    if not os.path.exists(icon_norm):
        raise FileNotFoundError(f"图标文件不存在：{icon_norm}")
    idata = np.fromfile(icon_norm, dtype=np.uint8)
    icon = cv2.imdecode(idata, cv2.IMREAD_UNCHANGED)
    if icon is None:
        pil = Image.open(icon_norm).convert("RGBA")
        icon = cv2.cvtColor(np.array(pil), cv2.COLOR_RGBA2BGRA)
    # 确保图标有 4 通道
    if icon.ndim == 3 and icon.shape[2] == 3:
        b, g, r = cv2.split(icon)
        a = np.full_like(b, 255)
        icon = cv2.merge((b, g, r, a))
    ih, iw = icon.shape[:2]
    if k != 1.0:
        new_w = max(1, int(iw * k))
        new_h = max(1, int(ih * k))
        icon = cv2.resize(icon, (new_w, new_h), interpolation=cv2.INTER_AREA)
        ih, iw = new_h, new_w

    # 5. 插入图标到每个中心点
    for xf, yf, *_ in centers:
        # CAD 坐标 -> 像素坐标
        px = int(round((xf - min_x) * sx))
        py = int(round((yf - min_y) * sy))
        # 垂直翻转，使 CAD 左下角对齐图像左下角
        py = h - 1 - py
        # 计算图标放置区域
        x0 = px - iw // 2
        y0 = py - ih // 2
        x1 = max(x0, 0)
        y1 = max(y0, 0)
        x2 = min(x0 + iw, w)
        y2 = min(y0 + ih, h)
        ox1 = x1 - x0
        oy1 = y1 - y0
        ox2 = ox1 + (x2 - x1)
        oy2 = oy1 + (y2 - y1)

        roi = out_img[y1:y2, x1:x2]
        patch = icon[oy1:oy2, ox1:ox2]
        # alpha 合成
        alpha = patch[..., 3:] / 255.0
        # 合成颜色通道
        roi[..., :3] = (alpha * patch[..., :3] + (1 - alpha) * roi[..., :3]).astype(np.uint8)
        # 更新 alpha 通道
        roi[..., 3] = (alpha[..., 0] * 255 + (1 - alpha[..., 0]) * roi[..., 3]).astype(np.uint8)
        out_img[y1:y2, x1:x2] = roi

    # 6. 保存结果
    dir_, fname = os.path.split(img_path)
    base, ext = os.path.splitext(fname)
    if output_path is None:
        # 若透明背景，强制 PNG
        if transparent_background:
            output_path = os.path.join(dir_, f"{base}_icons.png")
        else:
            output_path = os.path.join(dir_, f"{base}_icons{ext}")
    save_ext = os.path.splitext(output_path)[1].lower()
    # 保存
    cv2.imencode(save_ext, out_img)[1].tofile(output_path)
    return out_img

##按中心点对应大小插入图片


def insert_icons_with_cad_size(
    centers_with_size,
    icon_path: str,
    cad_bounds: tuple = ((0, 0), (118900, 84100)),
    image_path: str = r"D:/Myprogramsystem/BaiduSyncdisk/Text_production_services/Color_flat_generation/Basic_documents/Colorflat_production.png",
    output_path: str | None = None
) -> np.ndarray:
    """
    在底图上，根据 CAD 中的中心点及指定的 CAD 单位尺寸插入图标，
    自动将图标缩放，使其最大边长对应目标 CAD 长度。

    参数：
      centers_with_size:  列表 [((x,y,z), target_size), ...]
      icon_path:          图标文件路径（支持 PNG/JPG/WEBP, 含 alpha）
      cad_bounds:         CAD 区域 ((min_x,min_y),(max_x,max_y))
      image_path:         底图路径（用于获取分辨率）
      output_path:        输出路径；None 则生成“_iconsize”后缀

    返回：
      合成后 BGRA 图像数组，并保存至 output_path。
    """
    # 1. 读取底图并准备 BGRA 图像
    img_norm = os.path.normpath(image_path)
    if not os.path.exists(img_norm):
        raise FileNotFoundError(f"底图不存在: {img_norm}")
    data = np.fromfile(img_norm, dtype=np.uint8)
    img0 = cv2.imdecode(data, cv2.IMREAD_UNCHANGED)
    if img0 is None:
        raise IOError(f"无法解码底图: {img_norm}")
    # 转为 BGRA
    if img0.ndim == 2:
        base = cv2.cvtColor(img0, cv2.COLOR_GRAY2BGRA)
    elif img0.ndim == 3 and img0.shape[2] == 3:
        base = cv2.cvtColor(img0, cv2.COLOR_BGR2BGRA)
    elif img0.ndim == 3 and img0.shape[2] == 4:
        base = img0.copy()
    else:
        raise ValueError("不支持的底图通道数: {}".format(img0.shape))
    h, w = base.shape[:2]

    # 2. 计算 CAD->像素 比例
    (min_x, min_y), (max_x, max_y) = cad_bounds
    cad_w, cad_h = max_x - min_x, max_y - min_y
    sx = (w - 1) / cad_w
    sy = (h - 1) / cad_h

    # 3. 读取并预备图标（BGRA）
    icon_norm = os.path.normpath(icon_path)
    if not os.path.exists(icon_norm):
        raise FileNotFoundError(f"图标不存在: {icon_norm}")
    idata = np.fromfile(icon_norm, dtype=np.uint8)
    icon0 = cv2.imdecode(idata, cv2.IMREAD_UNCHANGED)
    if icon0 is None:
        pil = Image.open(icon_norm).convert('RGBA')
        icon0 = cv2.cvtColor(np.array(pil), cv2.COLOR_RGBA2BGRA)
    # 确保有 4 通道
    if icon0.ndim == 3 and icon0.shape[2] == 3:
        b, g, r = cv2.split(icon0)
        a = np.full_like(b, 255)
        icon0 = cv2.merge((b, g, r, a))
    orig_h, orig_w = icon0.shape[:2]
    max_orig = max(orig_w, orig_h)

    # 4. 插入并缩放图标
    for (xf, yf, _), target_size in centers_with_size:
        # CAD->像素中心点
        px = int(round((xf - min_x) * sx))
        py = int(round((yf - min_y) * sy))
        py = h - 1 - py  # 垂直翻转
        # 目标像素长度
        target_px = target_size * max(sx, sy)
        k = target_px / max_orig
        new_w = max(1, int(round(orig_w * k)))
        new_h = max(1, int(round(orig_h * k)))
        icon = cv2.resize(icon0, (new_w, new_h), interpolation=cv2.INTER_AREA)
        ih, iw = new_h, new_w
        # 计算放置区域
        x0 = px - iw//2; y0 = py - ih//2
        x1 = max(x0, 0); y1 = max(y0, 0)
        x2 = min(x0+iw, w); y2 = min(y0+ih, h)
        ox1, oy1 = x1 - x0, y1 - y0
        ox2, oy2 = ox1 + (x2 - x1), oy1 + (y2 - y1)
        roi = base[y1:y2, x1:x2]
        patch = icon[oy1:oy2, ox1:ox2]
        # alpha 合成
        alpha = patch[..., 3:] / 255.0
        roi[..., :3] = (alpha * patch[..., :3] + (1 - alpha) * roi[..., :3]).astype(np.uint8)
        roi[..., 3] = (alpha[..., 0] * 255 + (1 - alpha[..., 0]) * roi[..., 3]).astype(np.uint8)
        base[y1:y2, x1:x2] = roi

    # 5. 保存结果
    dir_, fname = os.path.split(img_norm)
    name, ext = os.path.splitext(fname)
    if output_path is None:
        output_path = os.path.join(dir_, f"{name}_iconsize{ext}")
    cv2.imencode(ext, base)[1].tofile(output_path)
    return base

def insert_icons_with_cad_size_transparent(
    centers_with_size,
    icon_path: str,
    cad_bounds: tuple = ((0, 0), (118900, 84100)),
    image_path: str = r"D:/Myprogramsystem/BaiduSyncdisk/Text_production_services/Color_flat_generation/Basic_documents/Colorflat_production.png",
    output_path: str | None = None
) -> np.ndarray:
    """
    只在透明画布上，根据 CAD 中的中心点及指定 CAD 单位尺寸插入图标，
    其余区域全透明。

    参数：
      centers_with_size:  列表 [((x,y,z), target_size), ...]
      icon_path:          图标文件路径（支持 PNG/JPG/WEBP，含 alpha）
      cad_bounds:         CAD 区域 ((min_x,min_y),(max_x,max_y))
      image_path:         底图路径，仅用于拿分辨率
      output_path:        输出路径；None 则生成“_iconsize_transparent”后缀

    返回：
      仅图块的透明 BGRA 图像数组，并保存至 output_path。
    """
    # 1. 读取底图以获得分辨率 h,w
    img_norm = os.path.normpath(image_path)
    data = np.fromfile(img_norm, dtype=np.uint8)
    img0 = cv2.imdecode(data, cv2.IMREAD_UNCHANGED)
    if img0 is None:
        raise IOError(f"无法解码底图: {img_norm}")
    h, w = img0.shape[:2]

    # 2. 创建全透明的 BGRA 画布
    base = np.zeros((h, w, 4), dtype=np.uint8)

    # 3. 计算 CAD->像素 比例
    (min_x, min_y), (max_x, max_y) = cad_bounds
    cad_w, cad_h = max_x - min_x, max_y - min_y
    sx = (w - 1) / cad_w
    sy = (h - 1) / cad_h

    # 4. 读取并预备图标（BGRA）
    icon_norm = os.path.normpath(icon_path)
    idata = np.fromfile(icon_norm, dtype=np.uint8)
    icon0 = cv2.imdecode(idata, cv2.IMREAD_UNCHANGED)
    if icon0 is None:
        pil = Image.open(icon_norm).convert('RGBA')
        icon0 = cv2.cvtColor(np.array(pil), cv2.COLOR_RGBA2BGRA)
    # 确保 4 通道
    if icon0.ndim == 3 and icon0.shape[2] == 3:
        b, g, r = cv2.split(icon0)
        a = np.full_like(b, 255)
        icon0 = cv2.merge((b, g, r, a))
    orig_h, orig_w = icon0.shape[:2]
    max_orig = max(orig_w, orig_h)

    # 5. 插入并按 CAD 大小缩放图标
    for (xf, yf, _), target_size in centers_with_size:
        # CAD→像素中心点
        px = int(round((xf - min_x) * sx))
        py = int(round((yf - min_y) * sy))
        py = h - 1 - py  # 垂直翻转
        # 计算目标像素长度
        target_px = target_size * max(sx, sy)
        k = target_px / max_orig
        new_w = max(1, int(round(orig_w * k)))
        new_h = max(1, int(round(orig_h * k)))
        icon = cv2.resize(icon0, (new_w, new_h), interpolation=cv2.INTER_AREA)
        ih, iw = new_h, new_w

        # 计算放置区域并截取
        x0, y0 = px - iw//2, py - ih//2
        x1, y1 = max(x0, 0), max(y0, 0)
        x2, y2 = min(x0+iw, w), min(y0+ih, h)
        ox1, oy1 = x1 - x0, y1 - y0
        ox2, oy2 = ox1 + (x2 - x1), oy1 + (y2 - y1)
        roi = base[y1:y2, x1:x2]
        patch = icon[oy1:oy2, ox1:ox2]

        # 透明合成：patch.alpha + base
        alpha = patch[..., 3:] / 255.0
        roi[..., :3] = (alpha * patch[..., :3] + (1 - alpha) * roi[..., :3]).astype(np.uint8)
        roi[..., 3]   = (alpha[...,0] * 255 + (1 - alpha[...,0]) * roi[..., 3]).astype(np.uint8)
        base[y1:y2, x1:x2] = roi

    # 6. 保存结果
    dir_, fname = os.path.split(img_norm)
    name, ext = os.path.splitext(fname)
    suffix = "_iconsize_transparent"
    if output_path is None:
        output_path = os.path.join(dir_, f"{name}{suffix}.png")
    cv2.imencode(".png", base)[1].tofile(output_path)
    return base











#&&% 阴影效果
def shift_cad_image(
    image_path: str,
    p1: tuple[float, float, float],
    cad_bounds: tuple[tuple[float, float], tuple[float, float]] = ((0, 0), (118900, 84100)),
    output_path: str | None = None
) -> np.ndarray:
    """
    按 CAD 向量 p1 将一张 RGBA 图像整体偏移，留下透明背景。

    参数：
      image_path:   RGBA 图像路径（只有图块那部分不透明，背景透明）。
      p1:           CAD 坐标偏移向量终点 (x1, y1, 0)。
      cad_bounds:   CAD 区域 ((min_x,min_y),(max_x,max_y))。
      output_path:  输出路径；None 则在原文件同目录生成 “_shifted.png”。
    
    返回：
      偏移后的 RGBA 图像数组，并保存到 output_path。
    """
    # 1. 读取 RGBA 图像
    data = np.fromfile(image_path, dtype=np.uint8)
    img = cv2.imdecode(data, cv2.IMREAD_UNCHANGED)
    if img is None or img.ndim != 3 or img.shape[2] < 4:
        raise ValueError("请输入一张 RGBA 图像（背景透明，图块不透明）。")

    H, W = img.shape[:2]
    (min_x, min_y), (max_x, max_y) = cad_bounds
    cad_w, cad_h = max_x - min_x, max_y - min_y

    # 2. 计算 CAD->像素 偏移量
    x1, y1, _ = p1
    dx = int(round((x1 - min_x) * (W - 1) / cad_w))
    dy = int(round((y1 - min_y) * (H - 1) / cad_h))
    # CAD 向上为正，对应像素行向上 => 需要行负偏移
    row_off = -dy
    col_off =  dx

    # 3. 准备空白画布
    shifted = np.zeros_like(img)

    # 4. 计算可复制区域
    src_r0 = max(0, -row_off)
    src_r1 = min(H, H - row_off)
    dst_r0 = max(0, row_off)
    dst_r1 = dst_r0 + (src_r1 - src_r0)

    src_c0 = max(0, -col_off)
    src_c1 = min(W, W - col_off)
    dst_c0 = max(0, col_off)
    dst_c1 = dst_c0 + (src_c1 - src_c0)

    # 5. 执行整体偏移拷贝
    shifted[dst_r0:dst_r1, dst_c0:dst_c1, :] = img[src_r0:src_r1, src_c0:src_c1, :]

    # 6. 保存结果
    base, _ = os.path.splitext(image_path)
    save_path = output_path or f"{base}_shifted.png"
    cv2.imencode(".png", shifted)[1].tofile(save_path)

    return shifted

def make_cad_shadow(
    image_path: str,
    p1: tuple[float, float, float],
    cad_bounds: tuple[tuple[float, float], tuple[float, float]] = ((0, 0), (118900, 84100)),
    tou: float = 0.6,
    output_path: str | None = None
) -> np.ndarray:
    """
    在只含图块的 RGBA 图像上，生成向量阴影并与原图合成。

    1) 先按 CAD 向量偏移生成黑色半透明阴影层
    2) 再把原图完整地绘制到阴影之上

    参数：
      image_path:  RGBA 图像路径（背景透明，图块实心）
      p1:          CAD 坐标偏移向量终点 (x1,y1,0)
      cad_bounds:  CAD 区域 ((min_x,min_y),(max_x,max_y))
      tou:         阴影不透明度 0.0–1.0
      output_path: 保存路径；None 则产生“_shadow.png”

    返回：
      合成后 RGBA 图像数组，并保存到 output_path。
    """
    # ——— 1. 读取原图 ——————————————————————————————
    data = np.fromfile(image_path, dtype=np.uint8)
    img = cv2.imdecode(data, cv2.IMREAD_UNCHANGED)
    if img is None or img.ndim < 3 or img.shape[2] < 4:
        raise ValueError("请输入一张带 alpha 通道的 RGBA 图像。")
    H, W = img.shape[:2]

    # ——— 2. 计算偏移量 ————————————————————————————
    (min_x, min_y), (max_x, max_y) = cad_bounds
    cad_w, cad_h = max_x - min_x, max_y - min_y
    x1, y1, _ = p1
    dx = int(round((x1 - min_x) * (W - 1) / cad_w))
    dy = int(round((y1 - min_y) * (H - 1) / cad_h))
    row_off = -dy
    col_off =  dx

    # ——— 3. 构造阴影层 ————————————————————————————
    # 3.1 先整体偏移原图得到实心部分位置
    shifted = np.zeros_like(img)
    src_r0 = max(0, -row_off); src_r1 = min(H, H - row_off)
    src_c0 = max(0, -col_off); src_c1 = min(W, W - col_off)
    dst_r0 = max(0, row_off); dst_r1 = dst_r0 + (src_r1 - src_r0)
    dst_c0 = max(0, col_off); dst_c1 = dst_c0 + (src_c1 - src_c0)
    shifted[dst_r0:dst_r1, dst_c0:dst_c1] = img[src_r0:src_r1, src_c0:src_c1]

    # 3.2 把实心像素变黑并设定半透明
    alpha_val = int(round(tou * 255))
    mask = shifted[...,3] > 0
    shifted[..., :3][mask] = 0
    shifted[..., 3][mask]  = alpha_val

    # ——— 4. 合成原图在阴影之上 ——————————————————————
    # 将 shifted（阴影）和 img（原图）按 alpha 混合
    fg = img.astype(np.float32) / 255.0
    bg = shifted.astype(np.float32) / 255.0

    fg_a = fg[..., 3:4]
    comp_rgb = fg[..., :3] * fg_a + bg[..., :3] * (1 - fg_a)
    comp_a   = fg[..., 3]   + bg[..., 3]   * (1 - fg[..., 3])

    out = np.zeros_like(img, dtype=np.float32)
    out[..., :3] = comp_rgb
    out[..., 3]  = comp_a

    result = (out * 255).clip(0,255).astype(np.uint8)

    # ——— 5. 保存并返回 ——————————————————————————————
    base, _ = os.path.splitext(image_path)
    save_path = output_path or f"{base}_shadow.png"
    cv2.imencode(".png", result)[1].tofile(save_path)

    return result


#


#


#&&% 建筑物阴影

def add_architectural_shadows_from_entity(ent, p1, cad_bounds=((0,0),(118900,84100)), layer_name="Architectural_shadows"):
    """
    以多段线实体 ent 和 CAD 向量 p1 生成建筑阴影：
      1) 复制并 Move(ent, p1)
      2) 对每条边构造四边形 LWPOLY
      3) Hatch 为 SOLID
    坐标全部通过 vtpnt 转换为 VT_ARRAY R8。
    """
    # 连接 AutoCAD
    acad = win32com.client.GetActiveObject("AutoCAD.Application")
    doc  = acad.ActiveDocument
    ms   = doc.ModelSpace

    # 1. Copy + Move
    new_objs = doc.CopyObjects([ent], ms)
    moved = new_objs[0]
    # Move(基点, 偏移) 都用 vtpnt 生成
    moved.Move(vtpnt(0,0,0), vtpnt(*p1))

    # 2. 确保图层存在
    layers = doc.Layers
    try:
        layers.Item(layer_name)
    except:
        layers.Add(layer_name)

    # 设置原、多段线到阴影图层
    for obj in (ent, moved):
        obj.Layer = layer_name

    # 3. 取原多段线顶点
    coords = list(ent.Coordinates)  # [x0,y0, x1,y1, ...]
    pts = [(coords[i], coords[i+1], 0) for i in range(0, len(coords), 2)]
    # 复制后的顶点直接加 p1
    new_pts = [(x+p1[0], y+p1[1], 0) for x,y,_ in pts]

    # 4. 为每条边生成四边形 LWPOLYLINE
    quads = []
    n = len(pts)
    for i in range(n):
        A = pts[i];       B = pts[(i+1)%n]
        C = new_pts[(i+1)%n]; D = new_pts[i]
        # 扁平坐标数组
        arr = [A[0],A[1], B[0],B[1], C[0],C[1], D[0],D[1]]
        # 用 VT_ARRAY|VT_R8 包装
        vt_arr = win32com.client.VARIANT(pythoncom.VT_ARRAY|pythoncom.VT_R8, arr)
        quad = ms.AddLightWeightPolyline(vt_arr)
        quad.Closed = True
        quad.Layer  = layer_name
        quads.append(quad)

    # 5. Hatch SOLID
    hatch = ms.AddHatch(0, "SOLID", False)
    for q in quads:
        hatch.AppendLoop(0, win32com.client.VARIANT(pythoncom.VT_DISPATCH, q))
    hatch.Evaluate()
    hatch.Layer = layer_name

    return moved, quads, hatch


def draw_cad_polygons_shadow(
    polygons: list[list[tuple[float, float, float]]],
    image_path: str = r"D:/Myprogramsystem/BaiduSyncdisk/Text_production_services/Color_flat_generation/Basic_documents/Colorflat_production.png",
    cad_bounds: tuple[tuple[float, float], tuple[float, float]] = ((0, 0), (118900, 84100)),
    tou: float = 0.6,
    output_path: str | None = None
) -> np.ndarray:
    """
    在一个透明画布上根据 CAD 多边形列表绘制“黑色半透明”阴影。
    其余区域保持全透明。

    参数：
      polygons:    CAD 坐标多边形列表，每个多边形为 [(x,y,z), ...] (闭合)
      image_path:  用于获取输出画布大小的位图路径
      cad_bounds:  CAD 对应的左下、右上角 (min_x,min_y),(max_x,max_y)
      tou:         阴影的不透明度 (0.0–1.0)，默认 0.6
      output_path: 输出路径；None 则在原图同目录生成 “<原名>_shadow.png”

    返回：
      RGBA 图像数组（np.ndarray），并保存为 PNG。
    """
    # 1. 读取底图以获取尺寸
    img0 = cv2.imdecode(np.fromfile(image_path, np.uint8), cv2.IMREAD_UNCHANGED)
    if img0 is None:
        raise FileNotFoundError(f"无法打开图像：{image_path}")
    h, w = img0.shape[:2]

    # 2. 生成全透明 BGRA 画布
    out = np.zeros((h, w, 4), dtype=np.uint8)

    # 3. 计算 CAD→像素 缩放比例
    (min_x, min_y), (max_x, max_y) = cad_bounds
    cad_w, cad_h = max_x - min_x, max_y - min_y
    sx = (w - 1) / cad_w
    sy = (h - 1) / cad_h

    # 4. 构建 mask，将所有多边形填充为 255
    mask = np.zeros((h, w), dtype=np.uint8)
    for poly in polygons:
        pts = []
        for x, y, _ in poly:
            px = int(round((x - min_x) * sx))
            py = int(round((y - min_y) * sy))
            # CAD 原点左下 → 像素坐标左上，y 取反
            py = h - 1 - py
            pts.append([px, py])
        pts = np.array(pts, dtype=np.int32)
        cv2.fillPoly(mask, [pts], 255)

    # 5. 在 out 画布上填充黑色并设置 alpha
    alpha_val = int(round(tou * 255))
    out[..., 0] = 0   # B
    out[..., 1] = 0   # G
    out[..., 2] = 0   # R
    out[..., 3] = np.where(mask > 0, alpha_val, 0)

    # 6. 保存 PNG
    base, _ = os.path.splitext(image_path)
    save_path = output_path or f"{base}_shadow.png"
    # 强制保存为 PNG 以保留 alpha
    cv2.imencode(".png", out)[1].tofile(save_path)

    return out


##填充带透明度
def draw_cad_polygons_shadow_k(
    polygons: list[list[tuple[float, float, float]]],
    image_path: str = r"D:/Myprogramsystem/BaiduSyncdisk/Text_production_services/Color_flat_generation/Basic_documents/Colorflat_production.png",
    cad_bounds: tuple[tuple[float, float], tuple[float, float]] = ((0, 0), (118900, 84100)),
    color: tuple[int, int, int] = (0, 0, 0),  # 默认黑色填充
    tou: float = 0.6,
    output_path: str | None = None
) -> np.ndarray:
    """
    在一个透明画布上根据 CAD 多边形列表绘制指定颜色的半透明阴影。
    其余区域保持全透明。

    参数：
      polygons:    CAD 坐标多边形列表，每个多边形为 [(x,y,z), ...] (闭合)
      image_path:  用于获取输出画布大小的位图路径
      cad_bounds:  CAD 对应的左下、右上角 (min_x,min_y),(max_x,max_y)
      color:       填充颜色 BGR 三元组，默认 (0,0,0) 黑色
      tou:         阴影的不透明度 (0.0–1.0)，默认 0.6
      output_path: 输出路径；None 则在原图同目录生成 “<原名>_shadow.png”

    返回：
      RGBA 图像数组（np.ndarray），并保存为 PNG。
    """
    # 读取底图以获取尺寸
    src = cv2.imdecode(np.fromfile(image_path, np.uint8), cv2.IMREAD_UNCHANGED)
    if src is None:
        raise FileNotFoundError(f"无法打开图像：{image_path}")
    h, w = src.shape[:2]

    # 生成全透明 BGRA 画布
    out = np.zeros((h, w, 4), dtype=np.uint8)

    # 计算 CAD→像素 缩放比例
    (min_x, min_y), (max_x, max_y) = cad_bounds
    cad_w, cad_h = max_x - min_x, max_y - min_y
    sx = (w - 1) / cad_w
    sy = (h - 1) / cad_h

    # 构建 mask，将所有多边形填充为 255
    mask = np.zeros((h, w), dtype=np.uint8)
    for poly in polygons:
        pts = []
        for x, y, _ in poly:
            px = int(round((x - min_x) * sx))
            py = int(round((y - min_y) * sy))
            py = h - 1 - py
            pts.append([px, py])
        pts = np.array(pts, dtype=np.int32)
        cv2.fillPoly(mask, [pts], 255)

    # 在 out 画布上填充指定颜色并设置 alpha
    alpha_val = int(round(tou * 255))
    b, g, r = color
    out[..., 0] = np.where(mask > 0, b, 0)
    out[..., 1] = np.where(mask > 0, g, 0)
    out[..., 2] = np.where(mask > 0, r, 0)
    out[..., 3] = np.where(mask > 0, alpha_val, 0)

    # 保存 PNG
    base, _ = os.path.splitext(image_path)
    save_path = output_path or f"{base}_shadow.png"
    cv2.imencode(".png", out)[1].tofile(save_path)

    return out


#&&% PDF转png把 PNG 导出或缩放到 3508 × 4961 px（或横版时 4961 × 3508 px）

import os
import fitz    # pip install PyMuPDF
from PIL import Image  # pip install Pillow

def pdf_page_to_png_transparent(
    pdf_path: str,
    page_number: int,
    dpi: int = 300,
    threshold: int = 250,
    output_path: str | None = None
) -> str:
    """
    将 PDF 的第 page_number 页导出为 PNG，并把接近白色背景转为透明。

    参数：
      pdf_path:     源 PDF 文件路径
      page_number:  要导出的页码（从 1 开始）
      dpi:          渲染分辨率，默认 300
      threshold:    白色阈值，RGB 三通道同时 ≥ threshold 时视为白色
      output_path:  输出 PNG 路径；若为 None，则在 pdf 同目录生成后缀文件

    返回：
      最终保存的 PNG 路径
    """
    # 打开 PDF 并定位页
    doc = fitz.open(pdf_path)
    if not (1 <= page_number <= doc.page_count):
        raise ValueError(f"page_number 必须在 1 和 {doc.page_count} 之间")
    page = doc.load_page(page_number - 1)

    # 渲染为位图
    zoom = dpi / 72
    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat, alpha=False)

    # 转为 RGBA
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples).convert("RGBA")

    # 扫描并替换接近白色像素
    datas = img.getdata()
    new_data = []
    for r, g, b, a in datas:
        if r >= threshold and g >= threshold and b >= threshold:
            new_data.append((255, 255, 255, 0))
        else:
            new_data.append((r, g, b, a))
    img.putdata(new_data)

    # 构造输出路径
    base, _ = os.path.splitext(pdf_path)
    if output_path is None:
        output_path = f"{base}_page{page_number}.png"
    img.save(output_path, "PNG")
    return output_path

#&&% PNG 统一补白缩放到 3508 × 4961 px（或横版时 4961 × 3508 px）

from PIL import Image

def pad_to_ratio(img: Image.Image, ratio: float, fill_color=(255,255,255)) -> Image.Image:
    """
    按给定宽高比 ratio 在上下或左右补白，使画布达到该比例，原图居中。
    """
    w, h = img.size
    if w / h > ratio:
        # 过宽：以宽度为基准，增加高度
        new_w = w
        new_h = int(w / ratio)
    else:
        # 过高：以高度为基准，增加宽度
        new_h = h
        new_w = int(h * ratio)
    # 新画布
    canvas = Image.new(img.mode, (new_w, new_h), fill_color)
    # 计算粘贴位置，使原图居中
    left = (new_w - w) // 2
    top  = (new_h - h) // 2
    canvas.paste(img, (left, top))
    return canvas

def pad_and_resize_to_a3(
    png_path: str,
    output_path: str | None = None,
    target_size: tuple[int,int] = (4961, 3508),
    fill_color=(255,255,255)
) -> Image.Image:
    """
    将任何 PNG：
      1. 按 A3 纸横向比例补白（过宽上下补白，过高左右补白），原图居中；
      2. 缩放到 target_size（默认横向 A3：4961×3508 像素；纵向自动交换）。

    Args:
        png_path:       待处理的 PNG 文件路径
        output_path:    若不为 None，则保存到该路径
        target_size:    输出像素尺寸，默认 (4961, 3508)
        fill_color:     补白颜色，默认为白色

    Returns:
        Pillow Image 对象（已缩放后的图）。
    """
    img = Image.open(png_path)
    # A3 横向像素比例
    a3_ratio = target_size[0] / target_size[1]
    # 1. 补白到 A3 比例
    padded = pad_to_ratio(img, a3_ratio, fill_color)
    # 2. 根据补白后画布方向选择最终输出尺寸
    w, h = padded.size
    if w >= h:
        final_size = target_size
    else:
        final_size = (target_size[1], target_size[0])
    # 高质量缩放
    resized = padded.resize(final_size, Image.LANCZOS)
    if output_path:
        resized.save(output_path)
    return resized

#&&% 多张图片叠加

def overlay_images(
    image_paths: list[str],
    output_path: str | None = None
) -> Image.Image:
    """
    将 image_paths 列表中的图片依次叠加——
      1. 第一张做底板；
      2. 后续每张与当前复合图尺寸差异不得超过 3 像素，否则抛错；
      3. 若尺寸不完全相同（差异 ≤3），则使用 Lanczos 高质量缩放到底板尺寸；
      4. 用 alpha_composite 叠加（需 RGBA 模式）。

    Args:
        image_paths: 依次叠加的图片文件路径列表，至少包含 1 张图。
        output_path: 若不为 None，则保存最终结果到该路径。

    Returns:
        最终的复合 PIL.Image 对象（RGBA 模式）。
    """
    if not image_paths:
        raise ValueError("图片路径列表不能为空")

    # 打开第一张，作为底板
    base = Image.open(image_paths[0]).convert("RGBA")
    w0, h0 = base.size

    # 逐张叠加
    for idx, path in enumerate(image_paths[1:], start=2):
        im = Image.open(path).convert("RGBA")
        w, h = im.size

        # 尺寸检查
        if abs(w - w0) > 3 or abs(h - h0) > 3:
            raise ValueError(
                f"第 {idx} 张图尺寸 {w}×{h} 与底板 {w0}×{h0} 相差超过 3 像素"
            )
        # 若尺寸在容差内但不完全一致，则缩放到底板尺寸
        if (w, h) != (w0, h0):
            im = im.resize((w0, h0), Image.LANCZOS)

        # 叠加
        base = Image.alpha_composite(base, im)

    # 保存并返回
    if output_path:
        base.save(output_path)
    return base

#&&% 从点列表画虚线实线点划线
def draw_cad_polyline(
    pts_cad: list[tuple[float,float,float]],
    image_path: str = r"D:/Myprogramsystem/BaiduSyncdisk/Text_production_services/Color_flat_generation/Basic_documents/Colorflat_production.png",
    cad_bounds: tuple[tuple[float,float],tuple[float,float]] = ((0,0),(118900,84100)),
    line_type: str = 'solid',            # 'solid', 'dashed', or 'dotdash'
    line_color: tuple[int,int,int] = (0,0,255),  # BGR, 默认红色
    thickness: int = 1,                  # 线宽（像素）
    dash_length: int = 20                # 虚线及点划线中“划”或“点”的长度
) -> np.ndarray:
    """
    在指定底图上，用 CAD 坐标画一条多段线，支持实线、虚线和点划线。

    返回：绘制后的 BGRA 图像数组（保留原 alpha，若有）。
    """
    # 1. 读取图像（支持 unicode 路径）
    data = np.fromfile(image_path, dtype=np.uint8)
    img = cv2.imdecode(data, cv2.IMREAD_UNCHANGED)
    if img is None:
        raise FileNotFoundError(f"无法打开图像：{image_path}")
    h, w = img.shape[:2]

    # 2. CAD->像素 比例
    (min_x,min_y), (max_x,max_y) = cad_bounds
    cad_w, cad_h = max_x-min_x, max_y-min_y
    sx = (w-1)/cad_w
    sy = (h-1)/cad_h

    # 3. CAD 点映射到像素点
    pix_pts = []
    for x,y,_ in pts_cad:
        px = int(round((x-min_x)*sx))
        py = int(round((y-min_y)*sy))
        py = h-1-py  # 上下翻转
        pix_pts.append((px,py))

    # 4. 绘制函数：按线型切分
    def draw_solid(p1,p2):
        cv2.line(img, p1, p2, line_color, thickness, lineType=cv2.LINE_AA)

    def draw_dashed(p1,p2):
        # 在 p1->p2 上按 dash_length 画线段
        x0,y0 = p1; x1,y1 = p2
        dist = ((x1-x0)**2+(y1-y0)**2)**0.5
        dx = (x1-x0)/dist; dy=(y1-y0)/dist
        num = int(dist//dash_length)
        for i in range(num+1):
            start = (int(x0 + (i*dash_length)*dx),
                     int(y0 + (i*dash_length)*dy))
            end   = (int(x0 + ((i*dash_length)+dash_length//2)*dx),
                     int(y0 + ((i*dash_length)+dash_length//2)*dy))
            cv2.line(img, start, end, line_color, thickness, lineType=cv2.LINE_AA)

    def draw_dotdash(p1,p2):
        # pattern: [dash, gap, dot, gap] lengths in pixels
        x0,y0=p1; x1,y1=p2
        dist = ((x1-x0)**2+(y1-y0)**2)**0.5
        dx=(x1-x0)/dist; dy=(y1-y0)/dist
        dot_len = max(1, thickness*2)
        pattern = [dash_length, dash_length, dot_len, dash_length]
        pat_sum = sum(pattern)
        pos = 0
        i = 0
        while pos < dist:
            seg = pattern[i%4]
            if i%4==2:
                # dot: single point
                cx = int(x0 + pos*dx)
                cy = int(y0 + pos*dy)
                cv2.circle(img, (cx,cy), thickness, line_color, -1, lineType=cv2.LINE_AA)
            else:
                # dash or gap
                start = pos
                end   = min(pos+seg, dist)
                if i%4==0:
                    pstart = (int(x0 + start*dx), int(y0 + start*dy))
                    pend   = (int(x0 + end*dx),   int(y0 + end*dy))
                    cv2.line(img, pstart, pend, line_color, thickness, lineType=cv2.LINE_AA)
            pos += seg
            i += 1

    # 5. 对所有相邻点执行绘制
    for a,b in zip(pix_pts, pix_pts[1:]):
        if line_type=='solid':
            draw_solid(a,b)
        elif line_type=='dashed':
            draw_dashed(a,b)
        elif line_type=='dotdash':
            draw_dotdash(a,b)
        else:
            raise ValueError("不支持的 line_type: {}".format(line_type))

    # 6. 保存并返回
    base, ext = os.path.splitext(image_path)
    out_path = f"{base}_polyline{ext}"
    cv2.imencode(ext, img)[1].tofile(out_path)
    return img


#&&%% 第一步 导出红线在图片上

#冻结其余图层
def freeze_layers_except(layernames):
    """
    冻结当前文档中，除 layernames 列表外的所有图层。
    
    参数：

        layernames   : 要保留不冻结的图层名列表（字符串列表）
    """
    layers = doc.Layers
    # 将 layernames 转为集合以加速查找
    keep_set = set(layernames)
    
    # 遍历所有图层
    # AutoCAD COM 中，Layers 集合索引从 0 到 Count-1
    for i in range(layers.Count):
        layer = layers.Item(i)
        name = layer.Name
        if name not in keep_set:
            # 冻结图层：优先使用 Freeze 属性
            try:
                layer.Freeze = True
            except Exception:
                # 部分版本可能需用 Frozen 属性
                try:
                    layer.Frozen = True
                except Exception as e:
                    print(f"无法冻结图层 {name}: {e}")
        else:
            # 若想同时保证 keep_set 中的图层都解冻，可以在此处加上：
            # try: layer.Freeze = False
            # except: layer.Frozen = False
            pass

    # 最后重新生成（可选，但一般建议）
    try:
        # 0 = acAllViewports
        doc.Regen(0)
    except Exception:
        pass

##A3_print_2( ctb = "acad.ctb",media= "UserDefinedMetric (420.00 x 297.00毫米)")定义了周边留白0的A3图纸规格




#set_layer_properties("打印框线", color_index=7, linetype="Continuous", on=True, frozen=False)

#&&&% PDF处理


def _render_pdf_single_page_to_image(pdf_path: str, dpi: int) -> Tuple[Image.Image, Tuple[float, float]]:
    """
    渲染单页 PDF → PIL.Image（RGB），并返回 (图像, (页面宽pt, 页面高pt)).
    要求 pdf_path 仅 1 页；dpi 控制位图精度，但不改变最终 PDF 的页面物理尺寸（pt）。
    """
    doc = fitz.open(pdf_path)
    if len(doc) != 1:
        doc.close()
        raise ValueError(f"{pdf_path} 必须是单页 PDF（当前 {len(doc)} 页）")
    page = doc[0]
    w_pt, h_pt = page.rect.width, page.rect.height

    zoom = dpi / 72.0
    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat, alpha=False)  # RGB
    img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
    doc.close()
    return img, (w_pt, h_pt)

def _remove_white_to_alpha(img_rgb: Image.Image, white_threshold: int = 245) -> Image.Image:
    """
    将接近白色(三通道均>=white_threshold)的像素转为透明；其他像素设为不透明。
    返回 RGBA 图像。
    优先使用 numpy 加速；若未安装 numpy，退化为 PIL 像素循环。
    """
    img = img_rgb.convert("RGBA")
    try:
        import numpy as np  # 可选加速
        arr = np.array(img, dtype=np.uint8)
        r, g, b, a = arr[..., 0], arr[..., 1], arr[..., 2], arr[..., 3]
        mask_white = (r >= white_threshold) & (g >= white_threshold) & (b >= white_threshold)
        # 非白 = 不透明；白 = 透明
        a[~mask_white] = 255
        a[mask_white] = 0
        arr[..., 3] = a
        return Image.fromarray(arr, mode="RGBA")
    except Exception:
        # 纯 PIL 版本（慢）
        px = img.load()
        w, h = img.size
        thr = int(white_threshold)
        for y in range(h):
            for x in range(w):
                r, g, b, a = px[x, y]
                px[x, y] = (r, g, b, 0 if (r >= thr and g >= thr and b >= thr) else 255)
        return img

def _apply_uniform_alpha(img_rgb: Image.Image, alpha_ratio: float = 0.5) -> Image.Image:
    """
    将整张 RGB 图设为统一透明度 alpha_ratio (0~1)，返回 RGBA。
    """
    alpha_ratio = max(0.0, min(1.0, float(alpha_ratio)))
    img = img_rgb.convert("RGBA")
    a = Image.new("L", img.size, color=int(round(255 * alpha_ratio)))
    img.putalpha(a)
    return img

def compose_single_page_pdfs(
    top_pdf_remove_white: str,#第1张pdf
    bottom_pdf_half_transparent: str,#第2张pdf
    out_pdf: str,#合成pdf
    dpi: int = 300,
    white_threshold: int = 245,
    bottom_alpha: float = 0.5
) -> None:
    """
    功能：
      - 将 top_pdf_remove_white 的“白色区域”抠除为透明（依据 white_threshold）
      - 将 bottom_pdf_half_transparent 整体设为 bottom_alpha 透明度
      - 叠加（顶图在上，底图在下）并导出为 out_pdf
      - 最终 PDF 保持原始页面物理尺寸（pt 不变）。清晰度由渲染用 dpi 控制。

    参数：
      top_pdf_remove_white: 顶层单页 PDF 路径（去白为透明）
      bottom_pdf_half_transparent: 底层单页 PDF 路径（整体透明度）
      out_pdf: 输出 PDF 路径
      dpi: 渲染分辨率（300~600 常用；越高越清晰/文件越大）
      white_threshold: 去白阈值（0~255），建议 235~254 之间微调
      bottom_alpha: 底层整体透明度 (0~1)，例如 0.5 == 50%

    异常：
      - ValueError: 当任一 PDF 非单页、两者页面尺寸(pt)不同、或渲染后像素尺寸不同
    """
    # 渲染为位图
    top_rgb, (wpt_top, hpt_top) = _render_pdf_single_page_to_image(top_pdf_remove_white, dpi)
    bot_rgb, (wpt_bot, hpt_bot) = _render_pdf_single_page_to_image(bottom_pdf_half_transparent, dpi)

    # 校验页面物理尺寸（pt）
    if abs(wpt_top - wpt_bot) > 0.01 or abs(hpt_top - hpt_bot) > 0.01:
        raise ValueError(
            "两个 PDF 的页面尺寸（pt）不同："
            f" top={wpt_top:.2f}x{hpt_top:.2f} pt vs bottom={wpt_bot:.2f}x{hpt_bot:.2f} pt。"
            "请先统一页面大小。"
        )
    # 校验像素尺寸（通常会一致）
    if top_rgb.size != bot_rgb.size:
        raise ValueError(
            f"两个 PDF 渲染后的像素尺寸不同：top={top_rgb.size}, bottom={bot_rgb.size}。"
            "请确保原 PDF 的页面尺寸完全一致。"
        )

    # 去白/设透明
    top_rgba = _remove_white_to_alpha(top_rgb, white_threshold=white_threshold)
    bot_rgba = _apply_uniform_alpha(bot_rgb, alpha_ratio=bottom_alpha)

    # 覆盖合成（底→顶）
    composed = Image.alpha_composite(bot_rgba, top_rgba)

    # 写入 PNG 到内存，再按原始 pt 尺寸生成 PDF
    png_mem = io.BytesIO()
    composed.save(png_mem, format="PNG")
    png_mem.seek(0)

    out_doc = fitz.open()
    page = out_doc.new_page(width=wpt_top, height=hpt_top)
    rect = fitz.Rect(0, 0, wpt_top, hpt_top)
    page.insert_image(rect, stream=png_mem.getvalue())
    out_doc.save(out_pdf)
    out_doc.close()

"""
compose_single_page_pdfs(
    top_pdf_remove_white=r"D:/path/first.pdf",
    bottom_pdf_half_transparent=r"D:/path/second.pdf",
    out_pdf=r"D:/path/output.pdf",
    dpi=300,
    white_threshold=245,
    bottom_alpha=0.5
)

"""

#&&&&%% 第十五部分  报建图绘制 








#&&&&%% 第十六部分  三维建模           

##"""
##三维建模只要我们在skp中解决任意面插入一个切割组件的问题，其余的都交给CAD了
##
##问题是这个问题我起始已经解决了，这样就不需要过多研究skp的建模，我们在CAD中建好，或者提取信息到skp重构就可以了 一个cad三维对象，
##
##获取其图形信息，位置信息，传入skp重构就可以了 如果导入的模型有问题就重构
## 
##"""


#&&&&%% 第六部分 综合控制管理






#_____________________________________________________________________________________________________________________________________________________________________________
#_____________________________________________________________________________________________________________________________________________________________________________



print("__________________  CAD基本操作开始运行 _________________________")
































































































