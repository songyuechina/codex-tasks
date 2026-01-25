#!/usr/bin/env python3
# -*- coding: utf-8 -*-

#CAD_basic.py V1.0(系统整理版)-20260107

#&&&&%%   CAD基本操作 

#&&&&%%  第一部分 导入连接
#_____________________________________________________________________________________________________________________________________________


# 导入模块


# —— 脚本最顶部，就写这两行 —— 
import warnings
warnings.filterwarnings(
    "ignore",
    r".*Revert to STA COM threading mode.*",
    category=UserWarning,
    module=r"pywinauto\..*"
)
from pywinauto.application import Application


#A______________________________________________________________

import array

#B______________________________________________________________

import builtins


def _cad_safe_print(*args, **kwargs):
    """Wrap built-in print to drop unsupported characters (emoji) on GBK consoles."""
    try:
        return builtins._orig_print(*args, **kwargs)
    except UnicodeEncodeError:
        def _sanitize(text):
            if isinstance(text, str):
                try:
                    return text.encode("gbk", "ignore").decode("gbk", "ignore")
                except Exception:
                    return text.encode("ascii", "ignore").decode("ascii", "ignore")
            return text

        safe_args = [_sanitize(arg) for arg in args]
        try:
            return builtins._orig_print(*safe_args, **kwargs)
        except OSError:
            return builtins._orig_print(*safe_args)
    except OSError:
        try:
            return builtins._orig_print(*args)
        except Exception:
            return None


if not hasattr(builtins, "_orig_print"):
    builtins._orig_print = builtins.print
    builtins.print = _cad_safe_print



#C______________________________________________________________

from collections import defaultdict,deque

from concurrent.futures import ThreadPoolExecutor, TimeoutError

import cv2

import ctypes

from contextlib import redirect_stdout, redirect_stderr,suppress
import comtypes.client

#D______________________________________________________________


import datetime

#E______________________________________________________________

#F______________________________________________________________
import functools

from fractions import Fraction

from functools import cmp_to_key

from fitz import Rect




from functools import wraps

import functools

from functools import wraps

import fire

from func_timeout import func_timeout, func_set_timeout, FunctionTimedOut

import fitz  # PyMuPDF
#G______________________________________________________________

#H______________________________________________________________

#I______________________________________________________________

from itertools import chain

import importlib

import imageio

import inspect

import itertools

import io



#J______________________________________________________________


import  json


#K______________________________________________________________

#L______________________________________________________________


import logging





#M______________________________________________________________

import math

import multiprocessing

import mysql.connector

import matplotlib.pyplot as plt


from multiprocessing import Process



#N______________________________________________________________

import networkx as nx

import numpy as np 

import numbers

#O______________________________________________________________

import os


# ———— 1. 抑制 pygame 欢迎提示 ————
# 必须在任何 import pygame 之前设置
os.environ.setdefault("PYGAME_HIDE_SUPPORT_PROMPT", "1")


import openpyxl

from openpyxl import Workbook, load_workbook

from openpyxl.utils import column_index_from_string as col2idx

#P______________________________________________________________

import PyPDF2.errors

import pywintypes

from pywintypes import com_error


from  pyautocad  import Autocad,APoint,aDouble

import pyautocad.types

import pytesseract

from PIL import ImageGrab

from pythoncom import VT_ARRAY, VT_R8

import pyautogui

from pywinauto import findwindows

import psutil

import pythoncom

import pygame

import pygetwindow as gw

import pyttsx3

from PIL  import Image



import pprint

import pdb

from pathlib import Path



from pygetwindow import getWindowsWithTitle

from pypdf import PdfReader, PdfWriter













#Q______________________________________________________________

#R______________________________________________________________

import runpy

import regex as xre

import re

import random

#S______________________________________________________________

from shapely.geometry import Point, Polygon,LineString

import subprocess

import shutil

from sympy import *

import sys

from scipy.spatial import ConvexHull


from subprocess import DETACHED_PROCESS


import sys

import shutil
import signal
import subprocess
from subprocess import CREATE_NO_WINDOW, DETACHED_PROCESS
import subprocess

from subprocess import CREATE_NO_WINDOW, DETACHED_PROCESS


#T______________________________________________________________

import time

import tempfile

from typing import Tuple, Literal

from typing import Any, Dict, List ,Tuple, Optional

from typing import Sequence, Literal, Optional,Union

from typing import Dict, Set,List

import tkinter as tk

import traceback

import tkinter

import threading



#U______________________________________________________________
import uuid
import unicodedata

#V______________________________________________________________

#W______________________________________________________________



# ===== pywin32 / COM 基础 =====
import win32com                      # 确保顶层包存在

import win32com.client

import win32com.client as win32      # 常用入口，统一别名
from win32com.client import CastTo, makepy, constants, VARIANT,Dispatch

# 动态调度（少数场景用）
import win32com.client.dynamic as dyn

# Windows API / GUI / 进程
import win32gui
import win32api
import win32con
import win32process

import pythoncom                     # CoInitialize, COM 错误
import pywintypes                    # com_error


from win32com.client import gencache

#X______________________________________________________________

#Y______________________________________________________________

#Z______________________________________________________________


# --- COM 相关库 ---



#&&&% 可移植设置

# D:/claude-tasks/cad/scripts/CAD_basic.py
# V1.0

#&&%  系统自建模块
# ================= 2.基本模块导入 =================
import os
import sys
from pathlib import Path
current = Path(__file__).resolve()
while current.name != 'cad':
    if current.parent == current: raise Exception("找不到根目录")
    current = current.parent
sys.path.insert(0, str(current))
from system.project_setup import PathConfig

USERPATH=PathConfig.userpath#本脚本保留之前userpath就是一个自然的lujing，USERPATH

userpath= os.environ.get('USERPATH')
#是可以直接拼接的

# ================= 2.1 核心连接模块导入 =================
try:
    from system import licad
    from system.licad import C

    # 1. 桥接函数
    li = licad.li 

    # 2. 【关键修正】桥接全局变量！
    # 旧函数里满屏都是 doc.Something, acad.Something
    # 我们必须把 C 对象的属性“借”给全局变量，骗过旧函数
    # 注意：这里使用 property 动态获取可能不方便直接赋值，
    # 建议直接在 li() 之后或者脚本初始化时赋值，
    # 或者定义一个动态获取的 getattr 魔法（太复杂），
    # 最简单的办法是：让 acad 和 doc 指向 C 的属性
    
    # 定义获取函数，或者直接初始化
    # 为了防止 C.doc 在模块加载时还没连接导致报错，我们用一种更聪明的方法：
    # 在 li() 被调用时，顺便刷新一下模块级的 doc
    
    # 临时方案：先定义为 None，靠 li() 刷新
    acad = None
    doc = None
    
    # 覆盖 li 函数，让它每次连接成功后，顺便把全局变量 doc 给填上
    def li():
        global acad, doc, mp, sp
        if licad.li(): # 调用底层连接
            acad = C.acad
            doc = C.doc
            mp = C.mp
            sp = C.sp
            return True
        return False

    print("[成功] licad 核心连接模块已加载 (已建立全局变量桥接)")

except ImportError as e:
    sys_logger.info(f"[严重错误] 无法加载 licad.py: {e}")

    # 定义哑巴函数，防止后续代码报错
    def li(): print("[错误] licad 未加载"); return False
    
    # ⬇️⬇️⬇️ 【关键修正】定义一个“假装饰器”作为后备 ⬇️⬇️⬇️
    # 这样即使 licad 没加载，@retry_on_busy 也不会让程序崩溃，只是不起作用而已
    def retry_on_busy(func): 
        return func


# ================= 2.2 CAD协同机制导入 =================
try:
    from system.CAD_coordination import (
        wait_quiescent,
        wait_command_done,
        send_cmd_with_sync,        
        CADGuard        
    )
    print("[成功] CAD协同机制模块已加载")

except ImportError as e:
    sys_logger.info(f"[警告] CAD协同机制模块加载失败: {e}")

from   system.CAD_com_utils  import (

    SafeCOM,
    retry_on_busy ,
    retry_if_busy ,
    sys_logger,
    silent_mode,
    alias, 
    timeit,      # <--- 从 utils 导入
    debuggable,  # <--- 从 utils 导入
    node,     
)





# ================= 2.3 选择模块导入 =================
try:
    from system.CAD_selection import *
        
    print("[成功] CAD_selection选择与属性模块已加载")

except ImportError as e:
    sys_logger.info(f"[警告] CAD_selection选择与属性模块加载失败: {e}")



# ================= 2.4 调试模块导入 =================
from system.common_logger import (
             # 日志记录器 (info/error)
    CriticalSection,     # 【方式A】块状托管 (with ...)
    checkpoint,          # 【方式B】线性打点 (func(...))
    record_test_result,  # 底层 Excel 写入 (finally 中用)
    node,                # 简易日志格式化
    set_debug_mode       # ⚠️ 推荐补上这个：用于控制 AI/Human 模式
)



#&&&% 全局参数
# ================= 3.调试和静音控制 =================

# 👇 一键静音：之后的 info 和 debug 全变成空函数
#sys_logger.mute()



# 1: 开启静音 (极速模式，不打印Info)
# 0: 关闭静音 (调试模式，打印所有细节)
GLOBAL_SILENT_MODE =0



# 👇 一行代码应用配置
sys_logger.mute_mode = GLOBAL_SILENT_MODE








set_debug_mode(mode=1, who="AI", wait_time=30)





# ================= 4. 初始化检查 =================

sys_logger.info(f"[初始化] CAD_basic 环境加载完成，运行路径: {PathConfig.SCRIPTS_DIR}")







#&&% 控制函数运行时间
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --


# 日志配置
LOG_PATH = Path(r"D:/Myprogramsystem/XT/系统运行中断错误记录.xlsx")
LOG_SHEET = "错误记录"

# 初始化日志文件和工作表
if not LOG_PATH.exists():
    wb = Workbook()
    ws = wb.active
    ws.title = LOG_SHEET
    ws.append(["时间", "函数名", "参数repr", "备注"])
    wb.save(LOG_PATH)


def _log(func_name, args, kwargs, note):
    """
    将错误或超时信息写入 Excel 日志。如果工作表不存在则创建。
    """
    from time import strftime
    wb = load_workbook(LOG_PATH)
    # 确保日志表存在
    if LOG_SHEET in wb.sheetnames:
        ws = wb[LOG_SHEET]
    else:
        ws = wb.create_sheet(LOG_SHEET)
        ws.append(["时间", "函数名", "参数repr", "备注"])
    ws.append([strftime("%Y-%m-%d %H:%M:%S"),
               func_name,
               f"args={args!r}, kwargs={kwargs!r}",
               note])
    wb.save(LOG_PATH)


def _kill_acad():
    """
    强制结束所有以 acad 开头的 CAD 进程。
    """
    for proc in psutil.process_iter(("pid", "name")):
        name = proc.info.get("name")
        if name and name.lower().startswith("acad"):
            try:
                proc.kill()
            except Exception:
                pass


#&&% 超时控制


@func_set_timeout(3)
def connect_database_task():
    print("开始连接...")
    time.sleep(5)  # 模拟耗时操作 (故意超过 3 秒)
    return "连接成功"



def safe_save_cad(doc, filepath):
    try:
        # 意思：给 doc.SaveAs 最多 10 秒时间，参数是 filepath
        # func_timeout(超时时间, 函数名, args=(参数列表), kwargs={关键字参数})
        func_timeout(10, doc.SaveAs, args=(filepath,))
        print("保存成功！")
        return True
    except FunctionTimedOut:
        sys_logger.info(f"保存超时！CAD 可能卡住了。跳过此文件: {filepath}")
        return False
    except Exception as e:
        sys_logger.info(f"保存出错: {e}")
        return False





#&&% 双重超时保护装饰器
def timeout_and_log2(timeout_sec: float):
    """
    装饰器：对关键函数添加双重超时保护。
    - 守护线程监控主函数执行，超过 timeout_sec 秒将
      调用 _kill_acad() 并记录日志。
    - 函数内部可根据 timeout_sec+1 触发二次超时操作。

    用法：
        @timeout_and_log2(20)
        def Fx(..., timeout_sec=20):
            # 可选二次超时监控
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            finished = threading.Event()

            def watchdog():
                # 等待 timeout_sec 秒后，如函数尚未结束，杀CAD并记录
                if not finished.wait(timeout_sec):
                    _kill_acad()
                    _log(func.__name__, args, kwargs,
                         f"超时 {timeout_sec}s，强制终止 CAD")
            threading.Thread(target=watchdog, daemon=True).start()

            try:
                result = func(*args, **kwargs)
            except Exception:
                _log(func.__name__, args, kwargs,
                     f"函数内部异常:\n{traceback.format_exc()}")
                raise
            finally:
                finished.set()

            return result
        return wrapper
    return decorator

# ---------- 示例：双重超时测试函数 ----------
@timeout_and_log2(20)
def test_draw_circle_and_wait(center=(0,0,0), radius=10000, timeout_sec=20):
    """
    测试：在 CAD 中画圆并触发等待命令，
    后台线程在 timeout_sec+1 秒后发送无效命令以触发中断。
    """
    import threading
    from time import sleep, time

    # 获取当前 DWG 文件名
    try:
        current_file = current_dwg_basename()
    except Exception:
        current_file = '<unknown>'

    try:


        # 第1步：画圆
        circ = draw_circle(center, radius)
        if circ is None:
            print("[错误] 圆绘制失败，测试中断。")
            return


        # 第2步：触发 CAD 等待
        doc.SendCommand("h\n")

        # 第3步：二次超时线程：timeout_sec+1 秒后发无效命令中断CAD
        def timeout_action():
            start = time()
            while True:
                if time() - start > timeout_sec + 1:
                    try:
                        doc.SendCommand("_.POINT 0,0,0")
                    except Exception as e:
                        sys_logger.info(f"[警告] 第二重超时触发，中断 CAD: {e}")
                    break
                sleep(0.1)

        threading.Thread(target=timeout_action, daemon=True).start()
        return circ

    except Exception as e:
        sys_logger.info(f"[错误] test_draw_circle_and_wait 捕获到异常，退出: {e}")
        # 记录日志：函数名、输入参数、当前 DWG、异常信息
        try:
            current_file = getattr(doc, 'Name', '<unknown>')
            _log('test_draw_circle_and_wait', (center, radius), {'timeout_sec': timeout_sec},
                 f"内部异常，文件={current_file}, 异常={e}")
        except Exception:
            pass
        return


#&&&% 函数控制
#&&% 静音控制

"""

单句
with silent_mode():
    dy = smart_select_polylines(

            layout_name="平面分割图",

            operate_target="Layout",

            select_config=0,

            use_cache=False, 

            min_side=100.0,  

        )




"""
#&&% 测试轮询等待

def wait_quiescent_ceshi():
    """
    详细拆解 "._ZOOM _E "片段含义解析.原生别理会用户是否修改过这个命令，
    用原厂的。_国际化我说的是英文，请翻译成当前 CAD 语言执行。
    ZOOM主命令缩放视图命令。  (空格)确认相当于按了一下回车，
    执行 ZOOM，进入下一步。_国际化同样，把后面的参数 E 当作英文关键字处理。
    E参数Extents (范围)。即缩放到全图范围。  (空格)确认相当于按了一下回车，
    确认选项 E，命令结束。
    
    """
    
    from scripts.CAD_file_operations import open_file 

    open_file("D:/Mypro/基础服务/用户1/dwg文件/霞飞云果园【供水平面图-125_t7.dwg")
    wait_quiescent()
    C.doc.PostCommand('(command "._ZOOM" "_E") \n')#LISP
    

    wait_quiescent()

#&&% 测试PostCommand
def complex_operation_demo():
    """
    不要用旧窗口
    """
    import time
    from system.licad import C
    from system.CAD_coordination import wait_quiescent

    # === 数据准备 ===
    # 窗口坐标 (左下, 右上)
    x1, y1 = 0, 0
    x2, y2 = 500, 500
    
    # 移动向量
    move_from = "0,0"
    move_to   = "100,100"
    
    # 旋转中心与角度
    rot_center = "100,100"
    rot_angle  = "45"
    
    # 缩放中心与比例
    scale_center = "100,100"
    scale_factor = "0.5"

    # === 构建“长”命令串 ===
    # 技巧：用 list 并在最后 join，比用 + 号拼接清晰得多
    cmds = []

    # 1. 【选择阶段】使用 SELECT 命令预先选中对象
    # 语法: ._SELECT _W p1 p2 [空格结束选择]
    cmds.append(f"._SELECT _W {x1},{y1} {x2},{y2} ") 

    # 2. 【移动阶段】
    # 语法: ._MOVE _P [空格确认选择] p1 p2
    cmds.append(f"._MOVE _P  {move_from} {move_to}")

    # 3. 【旋转阶段】
    # 注意：移动后，物体位置变了，但 _P 依然指向那堆物体
    # 语法: ._ROTATE _P [空格确认选择] 中心点 角度
    cmds.append(f"._ROTATE _P  {rot_center} {rot_angle}")

    # 4. 【缩放阶段】
    # 语法: ._SCALE _P [空格确认选择] 中心点 比例
    cmds.append(f"._SCALE _P  {scale_center} {scale_factor}")

    # === 合并为一个原子命令 ===
    # 用空格连接每一步
    full_cmd = " ".join(cmds) + " " 
    
    # 打印出来检查一下 (调试时非常有必要)
    print(f"执行长指令: [{full_cmd}]")

    # === 发送与执行 ===
    if C.li(): # 刷新连接
        try:
            C.doc.Activate()
            
            # 发送这一长串
            C.doc.PostCommand(full_cmd)
            
            # 物理延时 (长命令解析需要时间)
            time.sleep(1.0) 
            
            # 智能拦截
            wait_quiescent()
            
        except Exception as e:
            print(f"执行失败: {e}")


#&&% 测试卡住状态


def draw_infinite_spiral():
    print("启动压力测试 V2 (抗干扰版)...")
    print("目的：即便 Monitor 在后台查询，此脚本也不会崩溃，用于长时间观察 Monitor 行为")
    
    try:
        acad = win32com.client.GetActiveObject("AutoCAD.Application")
        doc = acad.ActiveDocument
        msp = doc.ModelSpace
    except Exception as e:
        print(f"无法连接 CAD: {e}")
        return

    # 定义一个安全的画线函数
    @retry_if_busy(max_retries=20, delay=0.1)
    def safe_add_line(p1, p2, color):
        line = msp.AddLine(
            win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, p1),
            win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, p2)
        )
        set_attr(line,"Color", color)
        return line

    @retry_if_busy(max_retries=5, delay=0.2)
    def safe_update(obj):
        obj.Update()

    center_x, center_y = 0, 0
    angle = 0.0
    radius = 10.0
    i = 0

    print("开始绘图... (Monitor 应该显示 '空闲' 或偶尔显示瞬间占用)")
    
    try:
        while True:
            x1 = center_x + radius * math.cos(angle)
            y1 = center_y + radius * math.sin(angle)
            
            angle += 0.2
            radius += 0.5
            
            x2 = center_x + radius * math.cos(angle)
            y2 = center_y + radius * math.sin(angle)
            
            # 使用带重试的函数
            line = safe_add_line((x1, y1, 0), (x2, y2, 0), (i % 7) + 1)
            
            if i % 50 == 0:
                print(f"\r已绘制 {i} 条...", end="")
                if line: safe_update(line)
            
            i += 1
            # 这里的 sleep 不需要太长，因为 retry 机制会处理冲突
            time.sleep(0.01) 

    except KeyboardInterrupt:
        print("\n测试停止。")
    except Exception as e:
        print(f"\n未知错误: {e}")










#&&% 语音播报

# 尝试导入 pyttsx3，如果没有安装则静默失败或打印日志
try:
    import pyttsx3
    HAS_TTS = True
except ImportError:
    HAS_TTS = False
    print("⚠️ 提示: 未安装 pyttsx3 库，语音播报功能将不可用。(pip install pyttsx3)")

def speak_msg(text):
    """
    后台线程播放语音，避免阻塞 CAD 主进程
    """
    if not HAS_TTS:
        return

    def _run():
        try:
            engine = pyttsx3.init()
            # 设置语速 (可选, 默认通常约为 200)
            engine.setProperty('rate', 180) 
            engine.say(text)
            engine.runAndWait()
        except Exception:
            pass # 忽略语音引擎的偶发错误

    # 启动守护线程进行播报
    t = threading.Thread(target=_run, daemon=True)
    t.start()

#&&% JSON数据读取存放设置
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --

# ———————— 全局配置 ————————


SAVE_DIR = r"D:/Myprogramsystem/XT/文件字典信息"#系统文件夹及内容仍然保留，独立于用户文件夹
os.makedirs(SAVE_DIR, exist_ok=True)

# 全局内存缓存
PRINT_DICTS: dict[str, dict] = {}



# ———————— 加载指定 DWG 的打印字典 ————————
def com_to_handle(obj):
    """
    如果是 COM 对象，则返回它的 Handle（字符串）；否则原样返回 obj。
    """
    try:
        # 大多数 Dispatch 对象都有 .Handle 属性
        if hasattr(obj, "Handle"):
            return obj.Handle
    except pythoncom.com_error:
        pass
    return obj

def serialize(obj):
    """
    递归地将 dict/list/tuple/COMObject 转为只含基础类型(包括 handle 字符串)的结构。
    """
    if isinstance(obj, dict):
        return {k: serialize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [serialize(v) for v in obj]
    if isinstance(obj, tuple):
        return tuple(serialize(v) for v in obj)
    # 不是容器的，就尝试转换成 handle
    return com_to_handle(obj)

#&&% 保存打印字典
def save_print_dict_generic(dwg_name: str, bind_dict: dict):
    """
    将 bind_dict 序列化并写入 JSON 文件，键为 dwg_name。
    序列化时自动把所有 COM 对象转换为它们的 Handle。
    """
    ser = serialize(bind_dict)
    path = os.path.join(SAVE_DIR, f"{dwg_name}_打印字典.json")
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(ser, f, ensure_ascii=False, indent=2)
    PRINT_DICTS[dwg_name] = ser
    sys_logger.info(f"[OK] 已保存打印字典到 {path}")



#&&% 加载打印字典
def load(dwg_name: str) -> dict:
    """
    从 JSON 文件加载之前保存的打印字典（只含基本类型 & handle 字符串）。
    如果内存中已有，则优先返回内存中的。
    """
    if dwg_name in PRINT_DICTS:
        return PRINT_DICTS[dwg_name]

    path = os.path.join(SAVE_DIR, f"{dwg_name}_打印字典.json")
    if os.path.isfile(path):
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        PRINT_DICTS[dwg_name] = data
        sys_logger.info(f"[OK] 已加载打印字典自 {path}")
        return data

    sys_logger.info(f"[警告] 未找到 {path}，返回空字典")
    return {}

# ———————— 辅助：获取当前 DWG 的纯文件名 ————————
def current_dwg_basename() -> str:
    """
    acad.ActiveDocument.Name 可能带路径或带 .dwg，
    取不带扩展的文件名。
    """
    name = acad.ActiveDocument.Name
    base = os.path.splitext(os.path.basename(name))[0]
    return base

def current_dwg_folder():
    """
    【V2.0 安全版】获取当前激活文档所在的文件夹路径。
    如果未保存或无文档，返回 None，不再报错。
    """
    import os
    import win32com.client
    from system.licad import C
    
    doc = None
    
    # 渠道 1: 尝试从全局 C 对象获取
    if getattr(C, "doc", None):
        doc = C.doc
        
    # 渠道 2: 如果 C.doc 是 None，尝试直接从 COM 获取 (双重保险)
    if doc is None:
        try:
            app = win32com.client.GetActiveObject("AutoCAD.Application")
            doc = app.ActiveDocument
        except:
            pass

    # 开始获取路径
    try:
        if doc is None: 
            return None
            
        full_path = doc.FullName
        if not full_path: # 新建文件 FullName 可能为空
            return None
            
        return os.path.dirname(full_path)
        
    except Exception:
        # 默默失败，不要打印红色的 [错误]，交给调用者处理
        return None





#&&% 数据类型转换函数


def vtpnt(x, y, z):
    """坐标点转化为浮点数"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (x, y, z))

def vtobj(obj):
    """转化为对象数组"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj)

def vtFloat(lst):
    """列表转化为浮点数"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, lst)
    
def vtInt(lst):
    """列表转化为整数"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, lst)

def vtVariant(lst):
    """列表转化为变体"""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, lst)

def ConvertArrays2Variant(inputdata, vartype):#例如vartype="Variant"
    import pythoncom
    if vartype == "ArrayofObjects":  # 对象数组
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, inputdata)
    if vartype == "Double":  # 双精度
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, inputdata)
    if vartype == "ShortInteger":  # 短整型
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, inputdata)
    if vartype == "LongInteger":  # 长整型
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I4, inputdata)
    if vartype == "Variant":  # 变体
        outputdata = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, inputdata)
    return outputdata

def vtlist(obj_list):#将com对象列表转换为可用win32数据
    """

    将对象列表转为 VARIANT 类型以供 COM 接口使用

    依赖语句 from win32com.client import VARIANT

    """
    return VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj_list)


warnings.filterwarnings("ignore", category=UserWarning)




#&&% 启动天正CAD及守护进程
def start_applicationV9(
    PTH: str = r"C:\Tangent\TArchT20V9",
    max_retries: int = 3,
    retry_delay: float = 2.0
) -> subprocess.Popen | None:
    """
    【核心启动器】
    1. 启动天正 TGStart.exe (主程序)。
    2. 挂载 cad_dialog_killer.py (弹窗杀手)。
    3. 挂载 cad_command_monitor.py (命令监控 V6.0)。
    
    返回: Popen 对象（CAD主进程），失败返回 None。
    """
   
    exe = os.path.join(PTH, "TGStart.exe")
    
    # 确定脚本根目录 (假设当前文件在 scripts 目录，需向上两级找到 system)
    # 结构: D:/claude-tasks/cad/scripts/CAD_basic.py (当前)
    # 目标: D:/claude-tasks/cad/system/xxx.py
    current_dir = os.path.dirname(os.path.abspath(__file__))
    cad_root = os.path.dirname(current_dir) # D:/claude-tasks/cad
    system_dir = os.path.join(cad_root, "system")
    
    killer_script = os.path.join(system_dir, "cad_dialog_killer.py")
    monitor_script = os.path.join(system_dir, "cad_command_monitor.py")

    # 定义启动辅助脚本的内部函数
    def launch_helper_script(script_path, script_name):
        if not os.path.exists(script_path):
            sys_logger.info(f"[错误] {script_name} 不存在: {script_path}")
            return None
        
        try:
            # 检查该脚本是否已在运行 (利用脚本内部的单例锁机制)
            # 但为了保险，这里直接启动，脚本内部会自己处理"若运行则退出"的逻辑
            
            # 生产环境建议使用 CREATE_NO_WINDOW 隐藏黑窗口
            # 调试环境建议去掉 creationflags 以便看到心跳输出
            # 这里默认隐藏
            p = subprocess.Popen(
                [sys.executable, script_path],
                creationflags=subprocess.CREATE_NO_WINDOW
            )
            
            time.sleep(0.5) # 等待启动
            if p.poll() is None:
                sys_logger.info(f"[启动] {script_name} 成功 (PID: {p.pid})")
                print(f"[守护] {script_name} 已挂载。")
                return p
            else:
                # 可能是因为已经运行了，脚本自己退出了，这不算错误
                print(f"[提示] {script_name} 可能已在运行中 (Process Exited)。")
                return None
        except Exception as e:
            sys_logger.info(f"[警告] {script_name} 启动异常: {e}")
            return None

    # --- 主启动循环 ---
    for attempt in range(1, max_retries + 1):
        try:
            # 1. 启动 CAD 主程序
            if not os.path.exists(exe):
                print(f"[严重错误] 天正启动程序未找到: {exe}")
                return None

            print(f"[启动] 正在呼叫天正 CAD (尝试 {attempt}/{max_retries})...")
            # cwd参数很重要，确保天正能找到自己的配置文件
            proc = subprocess.Popen([exe], cwd=PTH)
            
            # 等待几秒，让 CAD 进程初始化，然后再启动守护脚本
            # 这样可以避免 Monitor 一上来就找不到 CAD 窗口
            time.sleep(3.0) 

            # 2. 启动 弹窗治理脚本
            launch_helper_script(killer_script, "cad_dialog_killer")

            # 3. 启动 命令监控脚本 (新增)
            launch_helper_script(monitor_script, "cad_command_monitor")
            
            print("[成功] 天正 CAD 及守护进程组启动完毕。")
            return proc

        except Exception as e:
            sys_logger.info(f"第 {attempt} 次启动失败: {e}")
            if attempt < max_retries:
                print(f"等待 {retry_delay} 秒后重试...")
                time.sleep(retry_delay)
                
    sys_logger.info(f"已达最大重试次数 ({max_retries})，启动失败。")
    return None

def force_show_cad_interface():
    """
    【辅助函数】强制显示并最大化现有的 AutoCAD 窗口
    """
    try:
        # 1. COM 连接设置可见性
        acad = win32com.client.GetActiveObject("AutoCAD.Application")
        acad.Visible = True
        
        # acMaximized = 3, acNormal = 1
        # 尝试最大化窗口
        try:
            acad.WindowState = 3 
        except:
            pass
            
        # 2. Windows API 强制置顶 (解决窗口被遮挡问题)
        # 获取句柄 (HWND) - 这里的 Handle 是 CAD 主窗口句柄
        hwnd = acad.HWND 
        
        if hwnd:
            # 还原窗口（如果被最小化）
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            # 置顶
            win32gui.SetForegroundWindow(hwnd)
            print("✅ 已强制显示并激活 CAD 窗口")
            return True
            
    except Exception as e:
        sys_logger.info(f"❌ 强制显示 CAD 界面失败 (可能进程存在但COM未响应): {e}")
        return False

#&&% 常规启动
def st():

    from CAD_file_operations import cad_zt_oneb
    
    cad_zt_oneb()

    jd()

    shu=jc()

    print("进程数：",shu)

    doc=C.doc

    Name=doc.Name

    if Name == 'Drawing1.dwg' and shu==1:

        return True

    else:
    
        return False

def get_acad_process_id(ming):#获取进程的ID
    for process in psutil.process_iter(attrs=['pid', 'name']):
        if str(ming) in process.info['name'].lower():
            return process.info['pid']
    return None
    
#&&% 获取CAD进程数

@alias("jc")
def jingchengshu_wenjian():#查看cad进程数

    CAD_PROCESS_NAME = "acad.exe"

    found_process_i = 0
    for process in psutil.process_iter(['pid', 'name']):
        if process.info['name'] == CAD_PROCESS_NAME:
            found_process_i = found_process_i+1            

    return  found_process_i 







    
#&&% 关闭所有CAD进程
def close_all_cad_processes():#关闭所有进程（使用taskkill强制终止）
    """
    强制关闭所有CAD进程
    使用系统taskkill命令确保即使有弹窗也能关闭
    返回: True表示成功，False表示失败
    """
    import subprocess

    max_retries = 3
    for attempt in range(max_retries):
        try:
            # 检查当前CAD进程数
            process_count = jingchengshu_wenjian()
            if process_count == 0:
                print("[信息] 没有CAD进程需要关闭")
                return True

            sys_logger.info(f"[清理] 检测到 {process_count} 个CAD进程，正在强制关闭...")

            # 使用taskkill强制终止所有acad.exe
            result = subprocess.run(
                ["taskkill", "/F", "/IM", "acad.exe"],
                capture_output=True,
                text=True,
                timeout=10
            )

            # 检查返回码
            if result.returncode == 0:
                print("[成功] CAD进程已关闭")
                time.sleep(2)

                # 验证是否真正关闭
                process_count_after = jingchengshu_wenjian()
                if process_count_after == 0:
                    print("[验证] 所有CAD进程已成功关闭")
                    return True
                else:
                    sys_logger.info(f"[警告] 仍有 {process_count_after} 个CAD进程未关闭")

            elif result.returncode == 128:
                # 没有找到进程（已经关闭）
                print("[信息] 没有CAD进程需要关闭")
                return True
            else:
                sys_logger.info(f"[警告] taskkill 返回码: {result.returncode}")
                sys_logger.info(f"[输出] {result.stdout}")

        except subprocess.TimeoutExpired:
            sys_logger.info(f"[错误] 第 {attempt + 1} 次关闭超时")
        except Exception as e:
            sys_logger.info(f"[错误] 第 {attempt + 1} 次关闭失败: {e}")

        if attempt < max_retries - 1:
            sys_logger.info(f"[重试] 等待 2 秒后重试...")
            time.sleep(2)

    print("[失败] 多次尝试关闭CAD进程失败，请手动检查")
    return False
    
#&&% 关闭最早CAD进程
def close_oldest_cad_process(process_name="acad.exe"):#关闭上一个进程
    cad_processes = [p for p in psutil.process_iter(['pid', 'name', 'create_time']) if p.info['name'] == process_name]
        
    # 检查是否有多个CAD进程
    if len(cad_processes) > 1:
        # 按启动时间排序，最早的进程在前
        oldest_process = sorted(cad_processes, key=lambda p: p.info['create_time'])[0]
            
        try:
            # 关闭最早的进程
            psutil.Process(oldest_process.pid).terminate()
            sys_logger.info(f"已关闭最早的CAD进程，PID: {oldest_process.pid}")
        except Exception as e:
            sys_logger.info(f"关闭进程时出错: {e}")
            pass
    else:
        print("没有多个CAD进程运行。")
        pass





def ensure_typelib_from_running():
    """从正在运行的 AutoCAD 直接生成 makepy（不依赖注册表）"""
    _coinit_once()
    app = win32.gencache.EnsureDispatch("AutoCAD.Application")
    tlb, _ = app._oleobj_.GetTypeInfo().GetContainingTypeLib()
    makepy.GenerateFromTypeLibSpec(tlb)

class _ComLiveProxy:
    """把 getter 包成“活体”对象：每次访问都拿最新/可用的 COM 实例"""
    def __init__(self, getter): self._getter = getter
    def __getattr__(self, name):
        obj = self._getter()
        return getattr(obj, name)
    # 可选：允许直接作为可迭代或 bool 使用时更自然
    def __dir__(self): return dir(self._getter())






#&&% 窗口缩放
def zoom_window(x1, y1, x2, y2, pad_ratio=0.1, doc=None):
    (x_lo,y_lo),(x_hi,y_hi) = normalize_rect(x1,y1,x2,y2)
    pad = pad_ratio * ((x_hi-x_lo + y_hi-y_lo)/2.0)
    send_cmd(f"_.ZOOM\n_W\n{x_lo-pad},{y_lo-pad}\n{x_hi+pad},{y_hi+pad}\n", wait_s=0.6)





#&&&% RGB色彩

def aci_to_rgb(ci: int):
    # 只给出常用 ACI 的近似色；其余返回 None
    table = {
        1:(255,0,0), 2:(255,255,0), 3:(0,255,0), 4:(0,255,255),
        5:(0,0,255), 6:(255,0,255), 7:(255,255,255)
    }
    return table.get(int(ci)) if ci is not None else None
#&&% 获取实体RGB
def get_entity_rgb(ent):
    """返回 (rgb_tuple_or_None, source_str)"""
    # 1) 先看 TrueColor 的方式
    tc = ent.TrueColor                 # IAcadAcCmColor
    method = getattr(tc, "ColorMethod", None)
    # acColorMethodByRGB 大多为 3（不同版本枚举值可能不同，但逻辑相同）
    if method == 3 or (hasattr(tc, "Red") and (tc.Red or tc.Green or tc.Blue)):
        return (tc.Red, tc.Green, tc.Blue), "ByRGB(TrueColor)"

    # 2) 看 ColorIndex / Color（ACI / ByLayer / ByBlock）
    ci = getattr(ent, "ColorIndex", None)
    if ci is None:
        ci = getattr(ent, "Color", None)

    # ByLayer
    if ci == 256:
        layer_name = ent.Layer
        layer = ent.Document.Layers.Item(layer_name)
        ltc = layer.TrueColor
        return (ltc.Red, ltc.Green, ltc.Blue), f"ByLayer({layer_name})"

    # ByBlock
    if ci == 0:
        # 真正的显示颜色取决于引用它的块参照/上级容器
        return None, "ByBlock(需根据上级块参照解析)"

    # 3) 纯 ACI（1..255）
    if isinstance(ci, int) and 1 <= ci <= 255:
        rgb = aci_to_rgb(ci)
        return (rgb if rgb else None), f"ACI({ci})"

    # 兜底
    return None, "Unknown"


"""
lj=pmxz()
pl=lj[0]
rgb, src = get_entity_rgb(pl)
print("RGB:", rgb, "| 来源:", src)
RGB: (255, 0, 0) | 来源: ByRGB(TrueColor)
"""

    

    



#&&&%  * 重复多次调用函数

#&&% 重复操作
def chongfu_caozuo(
    Fx,
    dwg_instance=None,
    args: tuple=(),
    kwargs: dict=None,
    max_retries: int=3,
    failure_value=None
):
    """
    对指定函数/方法进行重复调用，直到成功或耗尽重试次数。

    :param Fx:             待调用的函数对象（或方法），不要在这里写 Fx()
    :param dwg_instance:   如果要调用类的方法，就把实例传进来；否则为 None
    :param args:           位置参数，按顺序传给 Fx
    :param kwargs:         关键字参数 dict，传给 Fx
    :param max_retries:    最多尝试次数（>=1）
    :param failure_value:  全部尝试失败时的返回值

    :return: (result, attempts, error)
        - result:   Fx 的返回值，或 failure_value
        - attempts: 成功时为那次尝试编号（1-based），失败时等于 max_retries
        - error:    最后一次捕获到的 Exception 实例；成功时为 None
    """
    if kwargs is None:
        kwargs = {}

    last_exception = None

    for attempt in range(1, max_retries + 1):
        try:
            # 决定如何调用 Fx
            if dwg_instance is not None and not getattr(Fx, "__self__", None):
                # Fx 很可能是类中定义的函数，但还未绑定到实例
                result = Fx(dwg_instance, *args, **kwargs)
            else:
                # Fx 已经是普通函数或已绑定的方法
                result = Fx(*args, **kwargs)

            # 成功：返回 (结果, 尝试次数, None)
            return result, attempt, None

        except Exception as e:
            last_exception = e
            if attempt < max_retries:
                sys_logger.info(f"[警告] 操作出错，正在第 {attempt} 次重试…")
                # （可选）在重试前做些环境刷新、移动窗口等
                try:    li()  # 示例：重新排列 CAD/IDLE 窗口
                except: pass
                time.sleep(2)
            else:
                print("[错误] 多次尝试均失败，请检查环境或参数是否正确。")
                # 返回 (failure_value, 用尽的尝试次数, 最后一个异常)
                return failure_value, attempt, last_exception


#  自动计时装饰器

"""
用法
@simple_timer
def heavy_func():
    ...

heavy_func()

"""
#&&% 简单计时器
def simple_timer(func):
    def wrapper(*args, **kwargs):
        t0 = time.time()
        result = func(*args, **kwargs)
        t1 = time.time()
        sys_logger.info(f"⏱️ {func.__name__} 耗时：{t1 - t0:.4f} 秒")
        return result
    return wrapper            




#&&&&%% 第二部分 属性选择

#&&&% 选择

# RPC 错误码：忙碌 或 服务不可用
_RPC_BUSY = (-2147417846, -2147418111)
_RPC_DOWN = (-2147023174,)

def _coinit_once():
    """线程初始化防呆设计"""
    try: 
        pythoncom.CoInitialize()
    except pythoncom.error: 
        pass

def get_acad_doc(max_wait=7.0):
    """
    [底层原语] 获取/启动 AutoCAD 应用和文档 (安全修正版)暂时保留给引用20250106
    修正：仅当确实无文档时才新建；若有文档但无法激活(忙碌/弹窗)，则进入等待而非新建。
    """
    _coinit_once()
    t0 = time.time()
    app = None
    
    while True:
        try:
            # --- 阶段 A: 获取/启动 App ---
            if app is None:
                try:
                    app = win32com.client.GetActiveObject("AutoCAD.Application")
                    app = win32com.client.gencache.EnsureDispatch(app)
                except Exception:
                    app = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
            
            # --- 阶段 B: 获取文档 (带智能判断) ---
            try:
                # 1. 尝试直接获取
                doc = app.ActiveDocument
                _ = doc.Name 
                return app, doc
                
            except Exception as e_doc:
                # 🔥【关键逻辑修正】获取不到 ActiveDocument，必须区分是“没文档”还是“CAD忙/锁死”
                
                # 1. 先尝试探测当前文档数量
                doc_count = -1 # 初始化为 -1 表示探测失败
                try:
                    doc_count = app.Documents.Count
                except Exception:
                    # 如果连 Count 都读不到，说明 CAD 彻底僵死或有模态对话框阻挡 COM 通讯
                    # 此时千万不能新建，视为忙碌
                    print("[licad] 无法读取文档数量 (CAD可能正忙或有弹窗)，暂停操作...")
                    pass 

                # 2. 只有明确知道数量为 0 时，才执行新建
                if doc_count == 0:
                    try:
                        print("[licad] 检测到 CAD 运行中但无文档 (Count=0)，正在新建...")
                        doc = app.Documents.Add()
                        # 等待初始化
                        for _ in range(5):
                            time.sleep(0.2)
                            try:
                                if doc.Name: break
                            except: pass
                        return app, doc
                    except Exception as e_add:
                        sys_logger.info(f"[licad] 新建文档失败: {e_add}")
                        raise e_doc # 新建失败，抛出异常让外层重试

                else:
                    # 3. 如果 doc_count > 0，或者读取失败 (-1)
                    # 说明有文档但拿不到 ActiveDocument，这是 CAD 在忙（如正处于命令中、有弹窗等）
                    # 此时绝对不能新建，而是抛出异常进入外层的“等待重试”循环
                    sys_logger.info(f"[licad] 获取文档失败但存在 {doc_count} 个文档。CAD 可能正忙/有弹窗/在命令中。")
                    raise RuntimeError("AutoCAD 忙碌或被阻塞，暂时无法获取活动文档。")

        except pythoncom.com_error as e:
            # --- 阶段 C: 全局 COM 异常处理 (RPC 忙碌等) ---
            code = e.args[0] if e.args else None
            # RPC 忙碌 / 呼叫被拒绝 / 无法获取对象
            if (time.time() - t0 < max_wait):
                # 打印个简单的点，表示在等
                print(".", end="", flush=True) 
                time.sleep(0.5)
                continue
            else:
                raise RuntimeError("连接超时：AutoCAD 响应过慢或处于长期忙碌状态(请检查是否有未关闭的对话框)。")

        except Exception as e:
            # 其他未知错误
            if (time.time() - t0 < max_wait):
                time.sleep(0.5); continue
            raise e

def normalize_rect(x1, y1, x2, y2):
    x_lo, x_hi = sorted((x1, x2))
    y_lo, y_hi = sorted((y1, y2))
    return (x_lo, y_lo), (x_hi, y_hi)

# ==============================================================================
#                            API MANIFEST / 完整函数签名清单
# ==============================================================================
"""
1. 核心工具 (Core & Utils)
--------------------------------------------------------------------------------
[CORE-000]   com_retry(fn, retries=30, delay=0.05)
             : COM调用重试装饰器，自动处理 RPC_BUSY (-2147417846) 等错误。
             
[CORE-002-1] cast_object(obj)
             : [入口] 智能类型转换，将 IAcadEntity 转为具体接口 (如 IAcadLine)。
             
[CORE-002]   _maybe_cast(ent)
             : [核心] 转换核心，混合了查表(CastTo)、动态封装(Dynamic)和兜底策略。
             
[UTILS-001]  to_vt_int(seq)
             : 列表 -> COM 整数变体数组 (VT_I2)。
             
[UTILS-002]  to_vt_variant(seq)
             : 列表 -> COM 变体数组 (VT_VARIANT)。
             
[UTILS-003]  _to_vt_point(pt_tuple)
             : (x,y,z) -> COM 浮点变体数组 (VT_R8)。
             
[UTILS-004]  normalize_rect(x1, y1, x2, y2)
             : 坐标标准化，返回 ((min_x, min_y), (max_x, max_y))。
             
[UTILS-005]  expand_rectangle(p1, p2, offset)
             : 矩形向外扩充 offset 距离。

2. 基础选择引擎 (Selection Engine)
--------------------------------------------------------------------------------
[SEL-001]    ss_select(mode="all", p1=None, p2=None, filter_types=None, filter_data=None, autocast=True, prompt=None)
             : 通用选择集构造器。
             : mode: "all"|"window"|"crossing"|"onscreen"
             : filter_types/data: DXF 组码过滤 (如 [0], ["INSERT"])

3. 几何与空间选择 (Geometric Selection)
--------------------------------------------------------------------------------
[SEL-GEO-001] select_entities_through_point(p, tol=0.1)
              : 点选。通过构造 tol 大小的窗交区域模拟鼠标点击。
              
[SEL-GEO-002] select_objects_in_window_area(x1, y1, x2, y2, max_retry=5)
              : 隐性窗口选择。自动 Zoom 到区域防止选不中，支持重试。
              
[SEL-GEO-003] select_paperspace_objects_in_window(x1, y1, x2, y2)
              : 布局空间区域选择。结合 Window 选择和 BoundingBox 遍历。
              
[SEL-GEO-004] pmxz(prompt="\n请在屏幕拾取图元，以Enter键结束：", autocast=True)
              : 屏幕交互框选 (GetSelection)。
              
[SEL-GEO-005] get_last_n_objects(n=1, autocast=True)
              : 获取模型空间最后生成的 N 个对象。
              
[SEL-GEO-006] last_obj()
              : 获取最后一个对象 (快捷方式)。

4. 类型与图层选择 (Type & Layer Selection)
--------------------------------------------------------------------------------
[SEL-TYPE-001] select_tuceng(layer_names, max_retries=5, delay=0.5, autocast=True)
               : 按图层名选择 (支持 str 或 list[str])。
               
[SEL-ALIAS]    stc(layer_names, **kwargs)
               : select_tuceng 的简写别名。
               
[SEL-TYPE-002] select_kuai(max_retries=5, autocast=True)
               : 全选普通块引用 (INSERT)。
               
[SEL-TYPE-003] select_text(autocast=True) / select_mtext(autocast=True)
               : 全选单行文字 (TEXT) / 多行文字 (MTEXT)。
               
[SEL-TYPE-004] select_line(autocast=True) / select_circle(autocast=True) / select_ellipse(autocast=True)
               : 全选直线 / 圆 / 椭圆。
               
[SEL-TYPE-005] select_polyline(autocast=True)
               : 全选轻量多段线 (LWPOLYLINE)。
               
[SEL-TYPE-006] select_polyline_chuantong(autocast=True)
               : 全选二维多段线 (POLYLINE)。
               
[SEL-TYPE-007] select_spline(autocast=True)
               : 全选样条曲线 (SPLINE)。
               
[SEL-TYPE-011] select_all_texts_mixed(target_space="Model")
               : 混合全选。包含 CAD 文字(*TEXT) 和 天正文字(TDb*)。
               
[SEL-TYPE-012] select_pub_text_entities()
               : 专选 "PUB_TEXT" 图层上的 TDbText 和 TDbMText。
               
[SEL-TYPE-013] select_group_entities(group_obj)
               : 选择组 (Group) 内的所有对象。

5. 可视化与交互 (Visualization & Interaction)
--------------------------------------------------------------------------------
[VIS-001]    yin_to_xian_xuanze(LB, wait_s=0.6)
             : 隐转显 (Delete/Undo法)。强制将 LB 列表放入当前选择集 (兼容性好但慢)。
             
[VIS-002]    yin_to_xian_safe(LB, wait_s=0.1)
             : [推荐] 隐转显 (LISP sssetfirst法)。无副作用高亮选中对象。
             
[VIS-003]    xian_to_yin_pickfirst(clear_grips=True, autocast=True)
             : 获取当前编辑器选中的对象 (Pickfirst)。
             
[VIS-004]    select_entities_in_window(x1, y1, x2, y2, ty=1.0, select_mode="_W")
             : 区域高亮并返回对象。Zoom -> Select -> Pickfirst -> Return。
             
[VIS-005]    highlight_entities_in_window(x1, y1, x2, y2)
             : 仅高亮区域 (视觉提示)。
             
[VIS-006]    highlight_entity_by_bbox(entity)
             : 高亮单个对象 (基于包围盒，对 TDbWall 有额外偏移优化)。
             
[VIS-007]    set_entity_grip_state_precise(ent)
             : 独占式高亮。清除其他选中，仅显示该对象的夹点。
             
[VIS-008]    isolate_modelspace_area(x1, y1, x2, y2)
             : 隔离显示。仅显示区域内对象，隐藏其他 (IsolateObjects)。
             
[VIS-009]    unhide_all(space=None, filter_names=None, highlight=False)
             : 显示隐藏对象 (Visible = True)。

6. 属性与天正支持 (Properties & TArch)
--------------------------------------------------------------------------------
[PROP-000]   _resolve_attr_case_insensitive(obj, attr_name)
             : 属性名大小写智能解析 (PascalCase 优先，缓存加速)。
             
[PROP-001]   get_attr(obj, name)
             : [万能] 获取属性。自动判断天正(Invoke DISPID) 或 普通CAD对象。
             
[PROP-002]   set_attr(obj, name, value)
             : [万能] 设置属性。自动判断天正(Invoke DISPID) 或 普通CAD对象。
             
[PROP-OLD]   get_object_property(obj, property_name) / set_object_property(obj, property_name, value)
             : 兼容旧代码的属性读写接口。
             
[PROP-003]   brute_dump_tarch_props(ent, max_dispid=64)
             : [调试] 暴力扫描天正对象属性 ID (1~max_dispid)。
"""

#&&&% 属性


#&&% 获取选择集包围盒
def get_pmxz_group_bbox():
    """
    使用 pmxz() 选择一组对象，计算它们的整体外包盒。

    返回:
        (bbox_corners, rect_xy)

        bbox_corners: 四元组
            (
              (minX, minY, z),  # 左下角 bottom_left
              (maxX, maxY, z),  # 右上角 top_right
              (minX, maxY, z),  # 左上角 top_left
              (maxX, minY, z),  # 右下角 bottom_right
            )

        rect_xy: 四元组
            (x1, y1, x2, y2)
            其中 (x1, y1) = 左下角平面坐标，
                 (x2, y2) = 右上角平面坐标。

    若 pmxz() 未选到对象或全部对象不支持 GetBoundingBox()，返回 (None, None)。
    """
    print("========== [PMXZ_BBOX] get_pmxz_group_bbox BEGIN ==========")

    # 1. 让用户用 pmxz() 选择对象
    try:
        com_list = pmxz()
    except Exception as e:
        print("[PMXZ_BBOX] 调用 pmxz() 出错：", repr(e))
        print("========== [PMXZ_BBOX] END(ERROR_PMXZ) ==========")
        return None, None

    if not com_list:
        print("[PMXZ_BBOX] pmxz() 未返回任何对象。")
        print("========== [PMXZ_BBOX] END(NO_OBJECTS) ==========")
        return None, None

    sys_logger.info(f"[PMXZ_BBOX] pmxz() 返回对象数量 = {len(com_list)}")

    # 2. 计算整体外包盒四个角点
    bbox_corners = group_bbox_corners(com_list)
    if bbox_corners is None:
        print("[PMXZ_BBOX] group_bbox_corners 返回 None（可能所有对象都不支持 GetBoundingBox）。")
        print("========== [PMXZ_BBOX] END(NO_BBOX) ==========")
        return None, None

    bottom_left, top_right, top_left, bottom_right = bbox_corners
    x1, y1, _ = bottom_left
    x2, y2, _ = top_right

    rect_xy = (x1, y1, x2, y2)

    sys_logger.info(f"[PMXZ_BBOX] 外包盒四角点: {bbox_corners}")
    sys_logger.info(f"[PMXZ_BBOX] 矩形坐标 (x1,y1,x2,y2) = ({x1}, {y1}, {x2}, {y2})")
    print("========== [PMXZ_BBOX] get_pmxz_group_bbox END ==========")

    return bbox_corners, rect_xy



#&&% 获取选择集包围盒别名
def g():

    get_pmxz_group_bbox()


#&&% 最后对象对齐原点
def align_last_ms_obj_lb_to_origin():
    """
    选择当前激活图中“模型空间最后一个对象”，
    以其外包盒左下角对齐到世界坐标 (0, 0, 0)。


    """
    

    # 1. 连接当前 DWG
    doc=C.doc

    # 2. 获取“最后一个对象”
    try:
        ent = last_obj()
    except Exception as e:
        sys_logger.info(f"[错误] 调用 last_obj() 失败: {e}")
        return None

    # last_obj() 可能返回单个对象，也可能返回列表，这里做个保护
    try:
        _ = ent.ObjectName  # 若能访问说明是单个 COM 对象
    except Exception:
        try:
            ent = ent[-1]
        except Exception as e:
            sys_logger.info(f"[错误] last_obj() 返回值类型不支持: {e}")
            return None

    objname = getattr(ent, "ObjectName", "<无>")
    handle  = getattr(ent, "Handle", "<无>")
    sys_logger.info(f"[信息] 选中的最后一个对象: ObjectName={objname}, Handle={handle}")

    # 3. 使用“控制台同款”调用方式获取外包盒
    try:
        # ★ 关键：和你在控制台一样，不带参数调用 ★
        min_pt, max_pt = ent.GetBoundingBox()
        xmin, ymin, zmin = min_pt
        xmax, ymax, zmax = max_pt
        sys_logger.info(f"[外包盒-对齐前] min={min_pt}, max={max_pt}")
    except Exception as e:
        sys_logger.info(f"[错误] 获取外包盒失败: {e}")
        return None

    # 4. 以外包盒左下角为锚点，移动到 (0,0,0)
    try:
        anchor_pt = (float(xmin), float(ymin), float(zmin))  # 左下角
        target_pt = (0.0, 0.0, 0.0)                         # 原点

        from_pt = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, anchor_pt)
        to_pt   = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, target_pt)

        sys_logger.info(f"[对齐] 锚点(左下)={anchor_pt} → 目标点={target_pt}")
        ent.Move(from_pt, to_pt)
    except Exception as e:
        sys_logger.info(f"[错误] Move 对齐失败: {e}")
        return None

    # 5. 再次获取外包盒，验证对齐结果（同样用无参数版本）
    try:
        min_pt2, max_pt2 = ent.GetBoundingBox()
        sys_logger.info(f"[外包盒-对齐后] min={min_pt2}, max={max_pt2}")
    except Exception as e:
        sys_logger.info(f"[警告] 对齐后再次获取外包盒失败: {e}")

    print("[完成] 模型空间最后一个对象已按外包盒左下角对齐到 (0,0,0)。")
    return ent

def get_entity_full_info(entity):
    """
    基于优化后的 get_attr，获取实体的完整位置信息和图层属性。
    
    Args:
        entity: AutoCAD 图形对象 (支持早期/晚期绑定)
        
    Returns:
        dict: {
            'handle': 句柄,
            'space_type': 'Model' | 'Paper' | 'BlockDefinition',
            'layout_name': 布局名 (如 'Layout1'),
            'owner_block_name': 宿主块名 (如 '*Model_Space'),
            'layer_name': 图层名,
            'layer_locked': True/False,
            'layer_plottable': True/False,
            'layer_color': 颜色索引
        }
    """
    if entity is None:
        return None

    try:
        # 1. 获取 Document 对象
        # 用于后续将 OwnerID 转为对象，以及查找图层表
        doc = get_attr(entity, 'Document')
        if doc is None:
            # 某些旧版本或特殊状态下 entity.Document 可能获取失败
            # 回退方案：获取当前活动文档
            acad = win32com.client.Dispatch("AutoCAD.Application")
            doc = acad.ActiveDocument

        # 2. 获取宿主 (Owner) 信息
        # 必须使用 get_attr，因为 OwnerID 在早期绑定下必须严格匹配大小写
        owner_id = get_attr(entity, 'OwnerID')
        
        # 将 ID 转为 BlockTableRecord 对象
        owner_obj = doc.ObjectIdToObject(owner_id)
        
        # 安全获取块名 (Block Table Record Name)
        owner_name = get_attr(owner_obj, 'Name')
        
        # 初始化位置信息
        location_info = {
            'handle': get_attr(entity, 'Handle'),
            'owner_block_name': owner_name,
            'layout_name': None,
            'space_type': 'Unknown'
        }

        # 3. 判断空间逻辑
        if owner_name == '*Model_Space':
            # 模型空间
            location_info['layout_name'] = 'Model'
            location_info['space_type'] = 'Model'
            
        elif get_attr(owner_obj, 'IsLayout'): 
            # 布局空间 (通过 IsLayout 属性判断)
            # 注意：这里必须再次获取 Layout 对象才能拿到 tab 名
            layout_obj = get_attr(owner_obj, 'Layout')
            location_info['layout_name'] = get_attr(layout_obj, 'Name')
            location_info['space_type'] = 'Paper'
            
        else:
            # 块定义内部
            location_info['layout_name'] = 'Inside Block'
            location_info['space_type'] = 'BlockDefinition'

        # 4. 获取图层深度信息
        # 第一步：获取实体上的图层名
        layer_name = get_attr(entity, 'Layer')
        
        # 第二步：从文档图层表中获取图层对象 (IAcadLayer)
        # doc.Layers 是集合，Item 方法通常不区分大小写，但为了保险...
        # 这里直接用 .Item() 比较安全，因为它是方法调用不是属性访问
        layer_obj = doc.Layers.Item(layer_name)
        
        # 第三步：利用 get_attr 安全获取图层属性
        # 即使早期绑定生成的属性名是 LayerOn 还是 layerOn，get_attr 都能搞定
        layer_info = {
            'layer_name': layer_name,
            'layer_locked': get_attr(layer_obj, 'Lock'),         # 锁定
            'layer_plottable': get_attr(layer_obj, 'Plottable'), # 可打印
            'layer_frozen': get_attr(layer_obj, 'Freeze'),       # 冻结
            'layer_on': get_attr(layer_obj, 'LayerOn'),          # 开关
            'layer_color': get_attr(layer_obj, 'Color')          # 颜色
        }

        # 合并结果
        return {**location_info, **layer_info}

    except Exception as e:
        # 在调试阶段可以 print(e)，生产环境建议记录日志
        sys_logger.info(f"获取实体信息失败: {e}")
        return None







#&&&&%%  第三部分 线面分析 


#_____________________________________________________________________________________________________________________________________________

#  模块使用说明

"""
该模块研究dwg图纸中的线段、圆曲线、平面等基本几何问题 

"""
# 线段分析

#&&% 计算直线角度
def compute_line_angle(line):#按绘制顺序度量线段角度
    """
    计算直线的方向角（单位：度），基于 StartPoint / EndPoint。
    非直线对象将抛出异常。
    0-360，从起点处画横线，旋转到终点的角度
    """

    try:
        x1, y1, _ = line.StartPoint
        x2, y2, _ = line.EndPoint
        dx = x2 - x1
        dy = y2 - y1
        angle_rad = math.atan2(dy, dx)
        angle_deg = math.degrees(angle_rad)
        if angle_deg < 0:
            angle_deg += 360
        return angle_deg
    except AttributeError:
        print("[错误] 该对象不具备 StartPoint / EndPoint")
        return None


#&&% 绘制点
def draw_point(pt):
    """
    在模型空间绘制一个 AutoCAD 点实体。

    参数：
        pt: (x, y, z) 三维坐标元组

    返回：
        新创建的 Point 对象；失败时返回 None
    """
    try:
        # AutoCAD 的“点”由 AddPoint 创建，需传 VARIANT
        obj = mp.AddPoint(vtpnt(*pt))
        return obj
    except Exception as e:
        sys_logger.info(f"[错误] 无法绘制点: {e}")
        return None

#&&% 绘制直线
def draw_line(p1, p2):#从两点坐标返回直线段
    """
    在模型空间中绘制从 p1 到 p2 的直线段。

    参数：
        p1, p2: 三维坐标元组 (x, y, z)

    返回：
        新创建的直线对象（COM 对象）
    """
    try:
        line_obj = mp.AddLine(vtpnt(*p1), vtpnt(*p2))
        return line_obj
    except Exception as e:
        sys_logger.info(f"[错误] 无法绘制直线: {e}")
        return None


#&&% 绘制圆
def draw_circle(center, radius):
    """
    以 center 为圆心、radius 为半径绘制圆。

    参数：
        center: (x, y, z)
        radius: 浮点半径

    返回：
        新创建的 Circle 对象；失败时返回 None
    """
    try:
        obj = mp.AddCircle(vtpnt(*center), radius)
        return obj
    except Exception as e:
        sys_logger.info(f"[错误] 无法绘制圆: {e}")
        return None


#&&% 绘制正多边形
def draw_regular_polygon(center, radius, sides):
    """
    绘制正多边形（LWPolyline，已闭合）
    :param center: 圆心 (x,y,z)
    :param radius: 外接圆半径
    :param sides : 边数 ≥3
    """
    if sides < 3:
        print("[错误] 边数必须 ≥ 3"); return None
    cx, cy, cz = (center + (0.0,))[:3]

    pts_flat = []
    for k in range(sides):
        ang = 2 * math.pi * k / sides
        pts_flat.extend([cx + radius*math.cos(ang),
                         cy + radius*math.sin(ang)])

    try:
        v_pts = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8,
            pts_flat
        )
        poly = mp.AddLightWeightPolyline(v_pts)
        poly.Closed = True
        return poly
    except Exception as e:
        sys_logger.info(f"[错误] 无法绘制正多边形: {e}")
        return None



#&&% 优先水平线
def prioritize_horizontal(lines, tol=0.5):
    """
    将列表中所有“水平”直线段（起点和终点的 y 差小于 tol）放在最前面，
    其它直线保留原有相对顺序。

    :param lines: 直线对象列表，每个对象具有 .StartPoint 和 .EndPoint 属性，
                  这两个属性应返回 (x, y, z) 或类似可下标的三元组。
    :param tol:   判定为水平的 y 方向容差（默认 0.5）
    :return:      新列表，水平直线段在前，非水平在后
    """
    horizontals = []
    non_horizontals = []
    for ln in lines:
        y1 = ln.StartPoint[1]
        y2 = ln.EndPoint[1]
        if abs(y1 - y2) < tol:
            horizontals.append(ln)
        else:
            non_horizontals.append(ln)
    return horizontals,non_horizontals

    
#&&% 获取样条曲线长度
def get_spline_length_by_conversion(spline_entity):#返回样条曲线的长度（按默认10分断拟合）
    
    """
    将样条曲线对象复制、高亮并通过 _SPLINEDIT 转为多段线，
    然后读取长度，并删除该多段线。

    返回：
        样条曲线转换后的长度值（float）
    """
    try:
        # Step 1：复制 spline 对象
        new_spline = spline_entity.Copy()

        # Step 2：显性高亮
        highlight_entity_by_bbox(new_spline)

        # Step 3：模拟命令 SPLINEDIT → P → Enter → Enter
        doc.SendCommand("_SPLINEDIT\nP\n\n\n")
        time.sleep(1.2)  # 等待 CAD 完成处理（可根据机器速度调整）

        # Step 4：获取新生成的对象（最后一个）
        last_index = doc.ModelSpace.Count - 1
        poly = doc.ModelSpace.Item(last_index)

        # Step 5：检查 Length 属性
        if hasattr(poly, "Length"):
            length = poly.Length
            poly.Delete()  # 删除临时 polyline
            return length
        else:
            print("[错误] 转换后对象没有 Length 属性")
            return None

    except Exception as e:
        sys_logger.info(f"[错误] 获取样条曲线长度失败：{e}")
        return None

#&&% 估算椭圆周长
def estimate_ellipse_length(ellipse):#返回椭圆长度
    """
    估算椭圆对象的长度（周长），使用 Ramanujan 公式。
    """
    try:
        a = ellipse.MajorRadius
        b = ellipse.MinorRadius

        pi = math.pi
        h = 3 * (a + b) - math.sqrt((3 * a + b) * (a + 3 * b))
        length = pi * h
        return length
    except Exception as e:
        sys_logger.info(f"[错误] 无法估算椭圆长度: {e}")
        return None



#&&% 获取几何信息
def get_entity_geometry_info(obj):#返回图形关键几何信息
    """
    根据图元类型返回其关键几何信息：
    - 点：坐标
    - 直线：起点、终点、长度
    - 圆：圆心、半径、长度、面积
    - 椭圆：中心、主轴、次轴、长度、面积
    - 多段线：起点、终点、长度、面积（若闭合）
    - 样条曲线：起点、终点、长度（需转换），面积（若闭合）
    """
    try:
        name = obj.ObjectName.lower()

        # 点
        if "point" in name:
            return {"type": "Point", "position": obj.Coordinates}

        # 直线
        elif "line" in name and "xline" not in name:
            p1 = obj.StartPoint
            p2 = obj.EndPoint
            length = math.dist(p1, p2)
            return {"type": "Line", "start": p1, "end": p2, "length": length}

        # 圆
        elif "circle" in name:
            center = obj.Center
            radius = obj.Radius
            length = 2 * math.pi * radius
            area = math.pi * radius ** 2
            return {"type": "Circle", "center": center, "radius": radius, "length": length, "area": area}

        # 椭圆
        elif "ellipse" in name:
            center = obj.Center
            a = obj.MajorRadius
            b = obj.MinorRadius
            area = math.pi * a * b
            h = 3 * (a + b) - math.sqrt((3 * a + b) * (a + 3 * b))
            length = math.pi * h  # Ramanujan 公式
            return {
                "type": "Ellipse",
                "center": center,
                "major_radius": a,
                "minor_radius": b,
                "length": length,
                "area": area
            }

        # 多段线
        elif "polyline" in name:
            coords = obj.Coordinates
            start = (coords[0], coords[1], 0)
            end = (coords[-2], coords[-1], 0)
            length = getattr(obj, "Length", 0)
            area = obj.Area if obj.Closed else 0
            return {
                "type": "Polyline",
                "start": start,
                "end": end,
                "length": length,
                "area": area
            }

        # 样条曲线（需转换测量）
        elif "spline" in name:
            p1 = obj.GetFitPoint(0)
            p2 = obj.GetFitPoint(obj.NumberOfFitPoints - 1)
            length = get_spline_length_by_conversion(obj)
            area = obj.Area if obj.Closed else 0
            return {
                "type": "Spline",
                "start": p1,
                "end": p2,
                "length": length,
                "area": area
            }

        else:
            return {"type": "Unknown", "ObjectName": obj.ObjectName}

    except Exception as e:
        return {"type": "Error", "message": str(e)}



##在两点确定的方向上，返回与对象点指定距离的点

#&&% 直线定距点
def points_on_line_at_distance_3d(
    p1: Tuple[float, float, float],
    p2: Tuple[float, float, float],
    px: Tuple[float, float, float],
    distance: float
) -> List[Tuple[float, float, float]]:
    """
    已知 px 在由 p1->p2 确定的直线上，返回在该直线上与 px 距离为 distance 的两个点。

    :param p1: 起点 (x1, y1, z1)
    :param p2: 终点 (x2, y2, z2)
    :param px: 参考点 (x, y, z)，位于直线上
    :param distance: 与 px 的目标距离
    :return: [(ax, ay, az), (bx, by, bz)]，分别为正向和反向移动 distance 后的点
    """
    x1, y1, z1 = p1
    x2, y2, z2 = p2
    xx, yy, zz = px

    # 1) 计算方向向量并归一化
    dx, dy, dz = x2 - x1, y2 - y1, z2 - z1
    length = math.sqrt(dx*dx + dy*dy + dz*dz)
    if length == 0:
        raise ValueError("p1 和 p2 重合，无法确定方向向量")
    ux, uy, uz = dx / length, dy / length, dz / length

    # 2) 沿正向和反向各移动 distance
    ax = xx + ux * distance
    ay = yy + uy * distance
    az = zz + uz * distance

    bx = xx - ux * distance
    by = yy - uy * distance
    bz = zz - uz * distance

    return [(ax, ay, az), (bx, by, bz)]


# 找出一组直线段内的伪相交区域

#&&% 查找伪交点区域
def find_fake_intersection_regions(lines, tol=10, real_tol=0.01):
    """
    查找伪相交区域：对于任意线段 A 的端点 P，若：
    - 存在其他线段 B 满足 P 到 B 距离 < tol，且
    - 对所有 B，P 到 B 的距离 >= real_tol
    则判定为伪相交点。
    在模型空间中绘制圆（半径 1000）表示这些点。
    """
    ensure_layer("测试辅助")
    ms = doc.ModelSpace
    added = []

    def point_to_line_distance(p, a1, a2):
        x0, y0 = p[:2]
        x1, y1 = a1[:2]
        x2, y2 = a2[:2]
        dx, dy = x2 - x1, y2 - y1
        if dx == dy == 0:
            return math.hypot(x0 - x1, y0 - y1)
        t = max(0, min(1, ((x0 - x1) * dx + (y0 - y1) * dy) / (dx*dx + dy*dy)))
        proj_x = x1 + t * dx
        proj_y = y1 + t * dy
        return math.hypot(x0 - proj_x, y0 - proj_y)

    for A in lines:
        try:
            p1 = A.StartPoint
            p2 = A.EndPoint
        except Exception:
            continue

        for pt in [p1, p2]:
            pt_key = tuple(round(c, 3) for c in pt)
            if pt_key in added:
                continue

            min_dist = 1e10
            has_near = False
            has_real_near = False

            for B in lines:
                if B == A:
                    continue
                try:
                    b1, b2 = B.StartPoint, B.EndPoint
                    dist = point_to_line_distance(pt, b1, b2)
                    if dist < tol:
                        has_near = True
                    if dist < real_tol:
                        has_real_near = True
                        break
                except:
                    continue

            if has_near and not has_real_near:
                ms.AddCircle(vtpnt(*pt), 1000).Layer = "测试辅助"
                added.append(pt_key)
                sys_logger.info(f"[OK] 伪相交区域点: {pt}")

    print("[OK] 伪相交区域绘制完成")



# 把区域内的直线段交点打断

#&&% 直线打断
def lines_daduan(start_point,end_point):#全部脚本统一采用三维坐标点模式

    """
    这个命令对于避免天正墙体没有出现不相交的覆盖是非常重要的，直接应用天正的tlinebk

    还要先处理假相交点区域待优化20250409

    """

    # 使用 f-string 语法将三维坐标变量插入命令字符串中
    start_point_str = f"{start_point[0]},{start_point[1]},{start_point[2]}"

    end_point_str = f"{end_point[0]},{end_point[1]},{end_point[2]}"

    command = f"tlinebk{chr(13)}{start_point_str}{chr(13)}{end_point_str}{chr(13)}{chr(13)}{chr(13)}"

    acad.ActiveDocument.SendCommand(command)






#找出一组直线段中的所有直线段中所有重复的线段并删除

#&&% 删除重复直线
def delete_duplicate_lines(lines, tol=0.01):
    """
    删除重复的直线段，仅保留每组中一条。

    参数：
        lines: 模型空间中所有线段对象列表（ObjectName 为 'AcDbLine'）
        tol: 距离容差，小于此值认为两点重合
    """
    def is_same_point(p1, p2):
        return all(abs(a - b) < tol for a, b in zip(p1, p2))

    def is_duplicate(line1, line2):
        try:
            a1, a2 = line1.StartPoint, line1.EndPoint
            b1, b2 = line2.StartPoint, line2.EndPoint
            return (
                (is_same_point(a1, b1) and is_same_point(a2, b2)) or
                (is_same_point(a1, b2) and is_same_point(a2, b1))
            )
        except:
            return False

    keep = []
    to_delete = []

    for i, line in enumerate(lines):
        is_dup = False
        for j in range(i):
            if is_duplicate(line, lines[j]):
                is_dup = True
                break
        if is_dup:
            to_delete.append(line)
        else:
            keep.append(line)

    count = 0
    for dup in to_delete:
        try:
            dup.Delete()
            count += 1
        except:
            continue

    sys_logger.info(f"[OK] 删除了 {count} 条重复直线段，保留 {len(keep)} 条。")
    return keep



#删除完全或局部重复线段

#&&% 删除冗余直线
def delete_redundant_lines(lines, tol=0.01):
    """
    删除重复线段和局部重复线段，只保留每组中的一条。
    """
    def is_same_point(p1, p2):
        return abs(p1[0] - p2[0]) < tol and abs(p1[1] - p2[1]) < tol

    def point_on_segment(p, a, b):
        ax, ay = a[:2]
        bx, by = b[:2]
        px, py = p[:2]
        cross = abs((bx - ax) * (py - ay) - (by - ay) * (px - ax))
        if cross > tol:
            return False
        dot = (px - ax) * (bx - ax) + (py - ay) * (by - ay)
        if dot < 0:
            return False
        sq_len = (bx - ax)**2 + (by - ay)**2
        if dot > sq_len:
            return False
        return True

    def is_completely_duplicate(l1, l2):
        try:
            p1, p2 = l1.StartPoint, l1.EndPoint
            q1, q2 = l2.StartPoint, l2.EndPoint
            return (
                (is_same_point(p1, q1) and is_same_point(p2, q2)) or
                (is_same_point(p1, q2) and is_same_point(p2, q1))
            )
        except:
            return False

    def is_locally_duplicate(short_line, long_line):
        try:
            p1, p2 = short_line.StartPoint, short_line.EndPoint
            q1, q2 = long_line.StartPoint, long_line.EndPoint
            return point_on_segment(p1, q1, q2) and point_on_segment(p2, q1, q2)
        except:
            return False

    to_delete_handles = set()
    total = len(lines)

    for i in range(total):
        l1 = lines[i]
        h1 = l1.Handle
        if h1 in to_delete_handles:
            continue
        for j in range(i + 1, total):
            l2 = lines[j]
            h2 = l2.Handle
            if h2 in to_delete_handles:
                continue

            if is_completely_duplicate(l1, l2):
                to_delete_handles.add(h2)
            elif is_locally_duplicate(l2, l1):
                to_delete_handles.add(h2)
            elif is_locally_duplicate(l1, l2):
                to_delete_handles.add(h1)
                break

    deleted = 0
    for ent in lines:
        if ent.Handle in to_delete_handles:
            try:
                ent.Delete()
                deleted += 1
            except:
                continue

    sys_logger.info(f"[OK] 删除重复/局部重复线段 {deleted} 条，保留 {total - deleted} 条。")

#找出一组直线段中的孤立线段产生的交点


#&&% 查找孤立交点
def find_isolated_intersections(LB, tol=0.5):
    """
    找出线段列表 LB 中的孤立线段，并计算它们与其它线段的所有交点。

    用于人工标记的门窗位置

    参数：
      LB:   线段列表，每个元素是 [(x1,y1,z1), (x2,y2,z2)]
      tol:  端点重合判断容差

    返回：
      intersections: 交点列表，每个元素是 (x, y, z)
    """
    def segment_intersection(seg1, seg2, tol):
        """
        计算线段 seg1=(A,B) 与 seg2=(C,D) 的交点（二维），
        若相交且唯一，返回 (x, y, z)，否则返回 None。
        z 取 seg1 第一端点的 z。
        """
        (x1, y1, z1), (x2, y2, _) = seg1
        (x3, y3, _),  (x4, y4, _) = seg2

        # 方向向量
        r = (x2-x1, y2-y1)
        s = (x4-x3, y4-y3)
        # 叉积 r × s
        rxs = r[0]*s[1] - r[1]*s[0]
        if abs(rxs) < tol:
            return None  # 平行或共线，不处理
        # 解 t, u
        qp = (x3-x1, y3-y1)
        t = (qp[0]*s[1] - qp[1]*s[0]) / rxs
        u = (qp[0]*r[1] - qp[1]*r[0]) / rxs
        # 只考虑严格交于段内
        if -tol <= t <= 1+tol and -tol <= u <= 1+tol:
            xi = x1 + t*r[0]
            yi = y1 + t*r[1]
            zi = z1
            return (xi, yi, zi)
        return None

    # 1. 找出孤立线段
    isolated = []
    for i, seg in enumerate(LB):
        p1, p2 = seg
        shared = False
        for j, other in enumerate(LB):
            if i == j:
                continue
            q1, q2 = other
            if same_point(p1, q1, tol) or same_point(p1, q2, tol) \
            or same_point(p2, q1, tol) or same_point(p2, q2, tol):
                shared = True
                break
        if not shared:
            isolated.append(seg)

    # 2. 对每根孤立线段，与其余线段求交点
    intersections = []
    for seg in isolated:
        for other in LB:
            if other is seg:
                continue
            ip = segment_intersection(seg, other, tol)
            if ip is not None:
                intersections.append(ip)

    #删除孤立线段20250420
    for seg in isolated:

        seg.Delete()

    return intersections    


##doc.sendcommand("TSpOutline"+chr(13)+"41849.69465957, 12250.50102376, 0"+chr(13)+chr(13)+chr(13))

##doc.sendcommand("TRoflna"+chr(13)+"0"+chr(13))


#&&% 获取多边形内点
def get_inner_point_of_polygon(polygon: Polygon):
    """
    获取给定 polygon 的一个保证在其内部的点。

    参数：
        polygon (shapely.geometry.Polygon): 目标多边形

    返回：
        (x, y): 内部点坐标元组
    """
    if not isinstance(polygon, Polygon):
        raise ValueError("[错误] 输入必须是 shapely.geometry.Polygon")

    inner_point = polygon.representative_point()
    return inner_point.x, inner_point.y


#&&%____________________________________________________       获取一组直线段所有的封闭多边形和外轮廓线       ________________
#…………………………………………………………………………………………………………………………………………………………………
# 线面分析 - 获取一组直线段所有的封闭多边形和外轮廓线

#__________________________________________________________________________________________________________________________
#…………………………………………………………………………………………………………………………………………………………………
#  使用说明

"""
获取一组直线段所有的封闭多边形和外轮廓线

主要函数有：
(1)寻找直线段组中最右下角封闭多边性
(2)从com边或顶点坐标列表用PL复线绘制多边形
(3)获取一组直线段的外轮廓线
(4)获取最右下角的封闭多边形不影响其他封闭多边形的连续边的顶点列表
(5)获取全部封闭多边形，但不完全
(6)获取全部封闭多边形

"""
#&&% 获取房间轮廓
def get_room_outline_from_point(x, y, z=0):## 获取输入点所在房间的轮廓
    """
    自动发送 TSpOutline 命令，从指定点获取房间轮廓。

    参数:
        x, y, z: 点的坐标，z 默认为 0。

    也许会有用处，能从多边形内点迅速得到多边形
        
    """
    try:
        point_str = f"{x},{y},{z}"
        cmd = (
            "TSpOutline" + chr(13) +  # 启动命令
            point_str + chr(13) +     # 输入点
            chr(13) +                 # 确认默认设置
            chr(13)                   # 确认生成
        )
        doc.SendCommand(cmd)
        sys_logger.info(f"[OK] 已请求获取点 ({x},{y},{z}) 的房间轮廓。")

    except Exception as e:
        sys_logger.info(f"[错误] 获取房间轮廓失败：{e}")


#&&% 连接闭合多段线
def connect_lines_to_polyline_if_closed(lines, tol=0.5):##判断一组闭合直线段是否构成封闭多段线，是就返回PL复线
    """
    判断线段是否首尾连接成闭合多边形，如果是，则绘制PL多段线。
    
    参数:
        lines: AutoCAD中选中的 AcDbLine 对象列表。
        tol: 容许的端点闭合误差。
    
    返回:
        多段线对象，或 None。
    """

    try:
        # 提取二维端点集合（忽略z）
        segments = []
        for ln in lines:
            try:
                p1 = (ln.StartPoint[0], ln.StartPoint[1])
                p2 = (ln.EndPoint[0], ln.EndPoint[1])
                segments.append((p1, p2))
            except Exception:
                continue

        if not segments:
            print("[警告]️ 无有效线段")
            return None

        # 构建连接链条
        used = set()
        sequence = [segments[0][0]]
        current = segments[0][1]
        used.add(0)

        while True:
            found = False
            for idx, (a, b) in enumerate(segments):
                if idx in used:
                    continue
                if Point(current).distance(Point(a)) < tol:
                    sequence.append(current)
                    current = b
                    used.add(idx)
                    found = True
                    break
                elif Point(current).distance(Point(b)) < tol:
                    sequence.append(current)
                    current = a
                    used.add(idx)
                    found = True
                    break
            if not found:
                break

        # 检查是否闭合
        if Point(current).distance(Point(sequence[0])) > tol:
            print("[错误] 线段未构成闭合区域")
            return None
        sequence.append(sequence[0])  # 闭合环

        # 构造二维点数组
        pts = []
        for pt in sequence:
            pts.extend([pt[0], pt[1]])

        # 绘制PL
        poly = doc.ModelSpace.AddLightWeightPolyline(vtFloat(pts))
        poly.Closed = True
        print("[OK] 成功绘制封闭PL线")
        return poly

    except Exception as e:
        sys_logger.info(f"[错误] Polyline 创建失败: {e}")
        return None


#&&% 判断闭合多边形
def is_closed_polygon_from_lines(lines, tol=0.5):##判断一组闭合直线段是否构成封闭多段线，不返回PL复线
    """
    判断一组 AutoCAD 直线段是否首尾连接形成闭合多边形。
    
    参数:
        lines: AcDbLine 类型的 COM 对象列表
        tol: 闭合判断容差，单位与CAD一致（如mm）

    返回:
        True 表示首尾闭合形成多边形，False 否则
    """
    try:
        # 提取二维端点 (x, y)
        segments = []
        for ln in lines:
            try:
                p1 = (ln.StartPoint[0], ln.StartPoint[1])
                p2 = (ln.EndPoint[0], ln.EndPoint[1])
                segments.append((p1, p2))
            except Exception:
                continue

        if not segments:
            return False

        # 构造首尾连接链
        used = set()
        sequence = [segments[0][0]]
        current = segments[0][1]
        used.add(0)

        while True:
            found = False
            for idx, (a, b) in enumerate(segments):
                if idx in used:
                    continue
                if Point(current).distance(Point(a)) < tol:
                    sequence.append(current)
                    current = b
                    used.add(idx)
                    found = True
                    break
                elif Point(current).distance(Point(b)) < tol:
                    sequence.append(current)
                    current = a
                    used.add(idx)
                    found = True
                    break
            if not found:
                break

        # 判断是否回到起点（闭合）
        if Point(current).distance(Point(sequence[0])) < tol and len(used) == len(segments):
            return True
        else:
            return False

    except Exception as e:
        sys_logger.info(f"[错误] 判断失败: {e}")
        return False

#&&% 判断同点
def same_point(p1, p2, tol=0.5):
    """判断两个点是否在容差范围内相同（只比较 X、Y 坐标）"""
    return abs(p1[0] - p2[0]) <= tol and abs(p1[1] - p2[1]) <= tol


#&&% 判断同线
def same_line(ln1, ln2, tol=0.5):
    """
    判断两条线段 ln1 和 ln2 是否“相同”
    线段被认为相同的条件是：
      ln1 的 StartPoint 与 ln2 的 StartPoint 近似相同且 ln1 的 EndPoint 与 ln2 的 EndPoint 近似相同，
      或者 ln1 的 StartPoint 与 ln2 的 EndPoint 近似相同且 ln1 的 EndPoint 与 ln2 的 StartPoint 近似相同。
    """
    p1 = tuple(ln1.StartPoint)
    p2 = tuple(ln1.EndPoint)
    q1 = tuple(ln2.StartPoint)
    q2 = tuple(ln2.EndPoint)

    return (same_point(p1, q1, tol) and same_point(p2, q2, tol)) or \
           (same_point(p1, q2, tol) and same_point(p2, q1, tol))


#&&% 计算绝对角度
def calculate_absolute_angle(line, P, tol=0.5):
    """
    计算线段（line）从点 P 出发的绝对角度（0-360°）
    如果 line 的起点等于 P，则返回从 P 到终点的角度，否则返回从 P 到起点的角度。
    """
    sp = tuple(line.StartPoint)
    ep = tuple(line.EndPoint)
    if same_point(sp, P, tol):
        dx = ep[0] - P[0]
        dy = ep[1] - P[1]
    else:
        dx = sp[0] - P[0]
        dy = sp[1] - P[1]
    return math.degrees(math.atan2(dy, dx)) % 360

#&&% 计算相对角度
def calculate_relative_angle(line, P, current_line, tol=0.5):
    """
    计算当前参考线（current_line）与候选线段（line）之间的相对角度，
    角度是以共点 P 为中心，当前线从 P 到其非 P 的端点的绝对角度为基准，
    计算 candidate line 与该基准角度之间顺时针或逆时针的角度差（逆时针方向）。
    结果为 0 到 360 之间的数值。
    """
    sp_current = tuple(current_line.StartPoint)
    ep_current = tuple(current_line.EndPoint)
    sp = tuple(line.StartPoint)
    ep = tuple(line.EndPoint)

    # 选择当前线段中不等于 P 的端点作为参考
    if same_point(sp_current, P, tol):
        current_point = ep_current
    else:
        current_point = sp_current

    def angle(p1, p2):
        dx = p2[0] - p1[0]
        dy = p2[1] - p1[1]
        return math.degrees(math.atan2(dy, dx)) % 360

    angle_current = angle(P, current_point)
    # 对候选线段，选择不等于 P 的端点
    if same_point(sp, P, tol):
        target_point = ep
    else:
        target_point = sp
    angle_target = angle(P, target_point)
    angle_diff = (angle_target - angle_current) % 360
    return angle_diff

#####################
# 函数：查找给定点P经过的线段，按照绝对角度排序

#&&% 按角度查找线段
def find_lines_angle(lines, P, tol=0.5):
    """
    查找与指定点 P 共端点的所有线段，并按从 P 出发离开的绝对几何角度排序。

    参数:
        lines: 直线段对象列表，每个对象要求具备 StartPoint, EndPoint, Handle 属性。
        P: 三元组 (x, y, z)，目标共点。
        tol: 判断共点的容差（仅比较 x 和 y 坐标）。

    返回:
        按绝对角度从小到大排序的共点线段列表。
    """
    shared_lines = []
    P = tuple(P)
    sys_logger.info(f"调试：目标共点 P = {P}")
    
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
            dx_sp = abs(sp[0] - P[0])
            dy_sp = abs(sp[1] - P[1])
            dx_ep = abs(ep[0] - P[0])
            dy_ep = abs(ep[1] - P[1])
            sys_logger.info(f"线段 {ln.Handle}: sp={sp} 差值=({dx_sp:.4f},{dy_sp:.4f}), ep={ep} 差值=({dx_ep:.4f},{dy_ep:.4f})")
            if (dx_sp <= tol and dy_sp <= tol) or (dx_ep <= tol and dy_ep <= tol):
                shared_lines.append(ln)
        except Exception as e:
            sys_logger.info(f"[警告]️ 跳过无效线段 {getattr(ln, 'Handle', '未知')} : {e}")
            continue

    shared_lines.sort(key=lambda ln: calculate_absolute_angle(ln, P, tol))
    print("调试：按绝对角度排序后的共点线段：")
    for ln in shared_lines:
        sys_logger.info(f"  线段 {ln.Handle}：角度 = {calculate_absolute_angle(ln, P, tol):.2f}°")
    return shared_lines

#####################
# 函数：查找与P共点的线段，按照与当前线段的相对角度排序

#&&% 查找共点线段
def find_lines_sharing_point(lines, P, current_line, tol=0.5):
    """
    查找与指定点 P 共端点的所有线段，并按从 current_line 逆时针旋转到其他线段的相对角度排序。
    其中 current_line 的角度定义为 0°。
    
    参数:
        lines: 直线段对象列表，每个对象要求具备 StartPoint, EndPoint, Handle 属性。
        P: 共点，三元组 (x, y, z)。
        current_line: 当前参考线段，对应角度定义为 0°。
        tol: 判断共点的容差（只比较 x 和 y 坐标）。
        
    返回:
        按相对角度从小到大排序的经过 P 的线段列表（current_line 亦包括其中）。
    """
    shared_lines = []
    P = tuple(P)
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
            if (abs(sp[0]-P[0])<=tol and abs(sp[1]-P[1])<=tol) or (abs(ep[0]-P[0])<=tol and abs(ep[1]-P[1])<=tol):
                shared_lines.append(ln)
        except Exception as e:
            sys_logger.info(f"[警告]️ 跳过无效线段 {getattr(ln, 'Handle', '未知')} : {e}")
            continue
    # 根据从当前线段旋转（逆时针）的相对角度排序
    shared_lines.sort(key=lambda ln: calculate_relative_angle(ln, P, current_line, tol))
    return shared_lines

#####################
# 函数：根据当前线段和共点 P，选择下一条后继线段（选择相对角度最大的那条），返回 (后继线段, 新共点)

#&&% 查找最大转角后继线
def find_successor_line_max(current_line, lines, P, tol=0.5):
    """
    在给定共点 P 处，排除当前线段（current_line）后，
    选择在该点处与当前线段相对旋转角度最大的那条线段作为后继线段，
    返回 (后继线段, 新共点)。

    参数：
      current_line: 当前线段对象。
      lines: 所有直线段对象列表，每个对象必须具备 StartPoint, EndPoint, Handle 属性。
      P: 共点（三元组）。
      tol: 共点判断的容差（默认 0.5）。

    返回：
      (后继线段, 新共点)，若找不到合适的后继线段，则返回 (None, P)。
    """
    # 调用 find_lines_sharing_point 获取所有经过 P 的线段，并使用 current_line 的相对角度排序
    candidates = find_lines_sharing_point(lines, P, current_line, tol)
    # 排除当前线段
    candidates = [ln for ln in candidates if ln.Handle != current_line.Handle]

    if not candidates:
        sys_logger.info(f"[错误] 在点 {P} 处找不到除当前线外的候选后继线段")
        return None, P

    best_line = None
    max_angle = -1
    new_point = P
    for ln in candidates:
        relative_angle = calculate_relative_angle(ln, P, current_line, tol)
        if relative_angle > max_angle:
            max_angle = relative_angle
            if same_point(tuple(ln.StartPoint), P, tol):
                new_point = tuple(ln.EndPoint)
            else:
                new_point = tuple(ln.StartPoint)
            best_line = ln

    if best_line is None:
        sys_logger.info(f"[错误] 没有找到有效的后继线段")
        return None, P

    sys_logger.info(f"选中后继线段 {best_line.Handle}，新共点 {new_point}")
    return best_line, new_point

#&&%#####################
# 辅助函数：从所有线段中找出最右下角的点
#&&% 查找右下角点
def find_rightbottom_point(lines, tol=0.5):
    """
    从所有线段端点中，找出 y 值最小的点；若有多个，则选 x 最大的点作为最右下角点。
    """
    all_points = []
    for line in lines:
        if hasattr(line, "StartPoint") and hasattr(line, "EndPoint"):
            all_points.append(tuple(line.StartPoint))
            all_points.append(tuple(line.EndPoint))
    if not all_points:
        return None
    min_y = min(p[1] for p in all_points)
    candidates = [p for p in all_points if abs(p[1]-min_y) <= tol]
    rb = max(candidates, key=lambda p: p[0])
    sys_logger.info(f"[OK] 最右下角点为：{rb}")
    return rb



#  主函数
#  (1)
# 寻找直线段组中最右下角封闭多边性

#  该函数系列包括如下一些函数


"""
辅助函数

same_point(p1, p2, tol) 判断两点是否在容差范围内相同

calculate_absolute_angle(line, P, tol) 计算从点 P 出发到某条线段（取与 P 不相同的端点）的绝对角度

calculate_relative_angle(line, P, current_line, tol) 计算当前线（以 P 为起点，选取与 P 不相同的端点）

和候选线（以 P 为起点）的角度差（逆时针方向，结果在 0～360 之间）

函数 find_lines_angle
用于初始阶段：给定一个共点 P，找到所有经过该点的线段，并按它们的绝对角度排序（从小到大）。

函数 find_lines_sharing_point
给定共点 P 与一个“当前线段”（作为参考），返回所有经过 P 的线段，并按照从当前线段出发旋转（逆时针）的相对角度排序。

函数 find_successor_line_max
根据当前线段和共点 P，从通过 P 的候选线中选取“转角”最大的作为后继线段，并返回后继线段及该线段另一端的新共点。

函数 find_rightbottom_closed_polygon
利用上述函数构造封闭多边形。初始时先根据所有线段端点确定“最右下角”点（函数 find_rightbottom_point），然后在该点处调用

find_lines_angle 得到所有经过该点的线段，选取绝对角度最小的那一根作为第一条边（初始 current_line），再依次调用

find_successor_line_max 推进多边形直到闭合或达到最大步数。


"""

#&&% 查找右下角闭合多边形
def find_rightbottom_closed_polygon(lines, tol=0.5, max_steps=50):
    """
    利用所有线段构造封闭多边形：
      1. 先查找所有线段端点中的最右下角点（RB）。
      2. 在 RB 处获取所有经过该点的线段（按照绝对角度排序），选择绝对角度最小的那根作为初始边。
      3. 以该初始边为第一条边，之后依次利用 find_successor_line_max 选择后继边，
         直到新共点回到初始点（闭合）或超过最大步数。
         
    返回：
      构成封闭多边形的点列表（依次为每条边的终点），若无法构成则返回 None。
    """
    # 定位最右下角点 RB
    rb = find_rightbottom_point(lines, tol)
    if rb is None:
        print("[错误] 无右下角点")
        return None

    # 初始共点
    current_point = rb
    # 从 RB 处按绝对角度排序取候选线段
    candidates = find_lines_angle(lines, rb, tol)
    if not candidates:
        sys_logger.info(f"[错误] 在右下角点 {rb} 处没有找到经过的线段")
        return None
    # 选择绝对角度最小的线段作为初始边
    current_line = candidates[0]
    sys_logger.info(f"调试：选中初始线段 {current_line.Handle}，绝对角度 = {calculate_absolute_angle(current_line, rb, tol):.2f}°")

    # 得到初始边的另一端
    sp = tuple(current_line.StartPoint)
    ep = tuple(current_line.EndPoint)
    if same_point(sp, rb, tol):
        next_point = ep
    else:
        next_point = sp

    polygon_points = [rb, next_point]
    visited_handles = {current_line.Handle}
    current_point = next_point
    steps = 1

    while steps < max_steps:
        # 调用 find_successor_line_max 得到下一条线段及其另一端的点
        successor, new_point = find_successor_line_max(current_line, lines, current_point, tol)
        if successor is None:
            print("[错误] 无后继线段，构造失败")
            return None
        # 检查是否闭合（新共点与起始点接近）
        if same_point(new_point, rb, tol):
            polygon_points.append(rb)
            sys_logger.info(f"[OK] 成功构建封闭多边形，步数 = {steps}")
            return polygon_points

        if successor.Handle in visited_handles:
            print("🔁 检测到重复线段，构造失败")
            return None

        polygon_points.append(new_point)
        visited_handles.add(successor.Handle)
        current_line = successor
        current_point = new_point
        steps += 1

    print("[警告]️ 达到最大步数，未能构造出闭合多边形")
    return None


# 从com边或顶点坐标列表用PL复线绘制多边形


#&&% 绘制多边形
def draw_polygon_as_polyline(polygon, layer_name="测试辅助", tol=0.5):
    """
    将构造的多边形（polygon）转换为顶点序列，并在当前 AutoCAD 文档 doc 的 ModelSpace 中添加
    一个多段线（PLINE）。如果顶点序列的首尾两点重合，则绘制闭合多段线，否则绘制开放多段线。
    
    参数:
      polygon:
        1. 如果是由线段组成的列表，每个元素要求具备 StartPoint 和 EndPoint 属性，
           则按照这些线段的端点顺序构造顶点序列。
        2. 如果是顶点列表，例如 [(x, y, z), (x, y, z), ...]，则直接使用。
      layer_name: 绘制多段线所在的图层名称（默认为 "测试辅助"）。
      tol: 判断点是否重合的容差值（仅比较 x 和 y 坐标）。
      
    返回:
      成功时返回创建的多段线对象（PLINE），否则返回 None。
    """

    # 内部函数：判断两个点是否近似相等（仅比较 x 和 y 坐标）
##    def same_point(p1, p2, tol=tol):
##        return abs(p1[0] - p2[0]) <= tol and abs(p1[1] - p2[1]) <= tol

    if not polygon:
        print("[错误] 未提供有效的 polygon 数据")
        return None

    vertices = []
    is_closed = False  # 是否绘制闭合多段线

    # 判断 polygon 是顶点列表还是线段列表
    if isinstance(polygon[0], (tuple, list)):
        # 判断传入是否为顶点列表：检查第一个元素为 tuple/list 且长度>=3
        if len(polygon[0]) < 3:
            print("[错误] 顶点数据格式错误")
            return None
        # 直接使用顶点列表（转换为 tuple 形式）
        vertices = [tuple(pt) for pt in polygon]
        # 如果首尾已经重合，则视为闭合；否则保持开放（不自动补充首点）
        if same_point(vertices[0], vertices[-1], tol):
            is_closed = True
        else:
            is_closed = False
    else:
        # 假定 polygon 是线段列表，每个元素具有 StartPoint 和 EndPoint 属性
        first_line = polygon[0]
        start_pt = tuple(first_line.StartPoint)
        last_line = polygon[-1]
        # 判断哪一端与起始点相连，作为初始点
        if same_point(start_pt, tuple(last_line.StartPoint), tol) or same_point(start_pt, tuple(last_line.EndPoint), tol):
            # start_pt 可以作为起点
            pass
        else:
            # 否则选用第一条线段的 EndPoint作为起始点
            start_pt = tuple(first_line.EndPoint)
        vertices.append(start_pt)
        current_pt = start_pt
        # 遍历每条线段构造顶点序列
        for line in polygon:
            sp = tuple(line.StartPoint)
            ep = tuple(line.EndPoint)
            if same_point(current_pt, sp, tol):
                next_pt = ep
            elif same_point(current_pt, ep, tol):
                next_pt = sp
            else:
                sys_logger.info(f"[错误] 线段 {line.Handle} 与当前点 {current_pt} 未连接，构造多边形失败")
                return None
            vertices.append(next_pt)
            current_pt = next_pt
        # 检查是否闭合（首尾重合）
        if same_point(vertices[0], vertices[-1], tol):
            is_closed = True
        else:
            is_closed = False

    # 输出调试信息：打印顶点序列
    print("调试：多段线顶点序列：")
    for i, pt in enumerate(vertices):
        sys_logger.info(f"  顶点 {i}: {pt}")

    # 将顶点序列转换为一维坐标数组：[x1, y1, z1, x2, y2, z2, …]
    coords = []
    for pt in vertices:
        coords.extend([pt[0], pt[1], pt[2]])
    coords_tuple = tuple(coords)
    coords_variant = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, coords_tuple)

    # 确保图层存在，否则创建新图层
    try:
        _ = doc.Layers.Item(layer_name)
    except Exception:
        doc.Layers.Add(layer_name)

    # 在 ModelSpace 中添加多段线并设置相应属性
    try:
        ms = doc.ModelSpace
        polyline = ms.AddPolyline(coords_variant)
        polyline.Closed = is_closed
        polyline.Layer = layer_name
        # 可选设置颜色和宽度，按需要调整
        polyline.color = 1
        polyline.ConstantWidth = 20
        polyline.Update()
        doc.Regen(0)
        sys_logger.info(f"[OK] 成功在图层 '{layer_name}' 绘制多段线, Closed={is_closed}")
        return polyline
    except Exception as e:
        print("[错误] 绘制多段线失败：", e)
        return None



#&&%###外轮廓线

# -----------------------------------------------------
# 辅助函数：判断两个点是否近似相等（仅比较 x 和 y 坐标）
#&&% 近似相等
def is_nearly_equal(p1, p2, tol):
    return abs(p1[0] - p2[0]) <= tol and abs(p1[1] - p2[1]) <= tol



# -----------------------------------------------------
# 寻找后继线段：在共点 P 处，从当前边之外的候选边中选择相对角度最小的边
#&&% 查找最小转角后继线
def find_successor_line_min(current_line, lines, P, tol=0.5):
    # 先获取共点 P 的所有线段，使用 find_lines_sharing_point 逻辑，但这里直接采用 find_lines_angle 排序后筛选
    candidates = []
    P = tuple(P)
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
            if (abs(sp[0]-P[0]) <= tol and abs(sp[1]-P[1]) <= tol) or (abs(ep[0]-P[0]) <= tol and abs(ep[1]-P[1]) <= tol):
                # 排除当前线段自身
                if ln.Handle == current_line.Handle:
                    continue
                candidates.append(ln)
        except Exception as e:
            continue

    if not candidates:
        sys_logger.info(f"[错误] 在点 {P} 处找不到候选后继线段")
        return None, P

    best_line = None
    min_angle = 360
    new_point = P
    for candidate in candidates:
        angle_diff = calculate_relative_angle(candidate, P, current_line, tol)
        # 选择相对角度最小的候选
        if angle_diff < min_angle:
            min_angle = angle_diff
            best_line = candidate
            # 更新新共点：选择 candidate 中不等于 P 的端点
            sp_cand = tuple(candidate.StartPoint)
            ep_cand = tuple(candidate.EndPoint)
            if is_nearly_equal(sp_cand, P, tol):
                new_point = ep_cand
            else:
                new_point = sp_cand

    if best_line is None:
        sys_logger.info(f"[错误] 没有找到合适的后继线段")
        return None, P

    sys_logger.info(f"选中后继线段 {best_line.Handle}，新共点 {new_point}，相对转角 = {min_angle:.2f}°")
    return best_line, new_point



#  主函数
#  (3)
# 获取一组直线段的外轮廓线

#&&% 获取外轮廓
def get_outer_contour(lines, tol=0.5, max_steps=50):
    """
    获取一组直线段的外轮廓线
    规则：
      1. 计算所有线段端点的最右下角点 P（绝对方法）。
      2. 在 P 处，按绝对角度排序，选择绝对角度最小的边作为第一条边；
      3. 以当前边的另一端点作为新的共点，从当前边出发，选择相对于当前边逆时针转角最小的候选边，
         直到新共点回到初始 P 点（闭合）或达到最大步数为止。
    返回：
      外轮廓线构成的线段列表，如果无法构成封闭轮廓则返回空列表。
    """
    # 内部函数：获取所有线段的最右下角点
    rb = find_rightbottom_point(lines, tol)
    if rb is None:
        print("[错误] 无最右下角点")
        return []

    # 设置初始共点为 rb
    P = rb
    sys_logger.info(f"调试：最右下角点为 {P}")

    # 在 P 处，按照绝对角度排序获取所有共点的线段
    candidate_lines = find_lines_angle(lines, P, tol)
    if not candidate_lines:
        sys_logger.info(f"[错误] 在点 {P} 处未找到共点线段")
        return []

    # 根据题目要求，第一条边选择绝对角度最小的线段（即排序后第 1 根线段）
    initial_line = candidate_lines[0]
    # 确定初始线段的另一端点
    sp_init = tuple(initial_line.StartPoint)
    ep_init = tuple(initial_line.EndPoint)
    if is_nearly_equal(sp_init, P, tol):
        next_point = ep_init
    else:
        next_point = sp_init
    sys_logger.info(f"调试：选中初始线段 {initial_line.Handle}，起点 {P} -> 终点 {next_point}")
    
    contour_lines = [initial_line]
    visited_handles = {initial_line.Handle}
    current_line = initial_line
    current_point = next_point
    steps = 0

    # 迭代构造外轮廓
    while steps < max_steps:
        sys_logger.info(f"调试：目标共点 P = {current_point}")
        # 在当前共点处查找候选的后继线段，使用最小相对角度策略
        successor, new_point = find_successor_line_min(current_line, lines, current_point, tol)
        if successor is None:
            print("[错误] 无后继线段，构造失败")
            return []
        if successor.Handle in visited_handles:
            sys_logger.info(f"🔁 检测到重复线段 {successor.Handle}，构造失败")
            return []
        contour_lines.append(successor)
        visited_handles.add(successor.Handle)
        sys_logger.info(f"步 {steps+1}: 选中线段 {successor.Handle}，新共点 {new_point}")
        # 检查是否闭合：如果新共点与最右下角点近似相等，则认为闭合
        if is_nearly_equal(new_point, rb, tol):
            print("[OK] 成功构造封闭轮廓线")
            return contour_lines
        # 更新当前线段及共点
        current_line = successor
        current_point = new_point
        steps += 1

    print("[警告]️ 达到最大步数，未能构造封闭轮廓线")
    return []


#&&%##获取所有封闭多边形

#删除多边形的一些边

"""
有lines中多于两条直线段经过，就称该点为多分枝点

编制函数，PL是lines中的封闭多边形，用顶点列表表示，p1是lines的最右下角点，

且p1是PL的一个顶点，用坐标点表示。将PL的顶点排序，从p1按逆时针方向推进，

将遇到的非多分枝顶点放入列表LN，直到遇到一个多分枝点D1结束；再从p1按顺时

针方向推进，将遇到的顶点放入列表LN，直到遇到一个多分枝点D2结束。顺时针方

向推进时，把p1考虑在内，即p1如果为多分枝点则立即结束，结束时遇到的多分枝

点D2就是p1。逆时针方向推进时，不把p1考虑在内，即p1为多分枝点仍然推进，遇

到一个多分枝点D1结束。函数返回多分枝点D1，D2，以及它们之间的非多分枝顶点

列表LN。


"""
#&&% 顶点去重
def deduplicate_vertices(vertices, tol=0.5):
    """
    去掉顶点列表中相邻重复的顶点：
    如果两个相邻顶点（顺序出现）之间的二维距离小于 tol，则认为它们重复，只保留前一个，
    但如果这个重复点是列表中的最后一个且与第一个顶点相同（表示闭合多边形），则保留此重复点。
    
    参数:
      vertices: 顶点列表，每个顶点格式为 (x, y, z) 的元组。
      tol: 距离阈值。
      
    返回:
      处理后的顶点列表，只有中间连续重复的点被去除，保留闭合多边形的首尾重复。
    """
    if not vertices:
        return []
    
    deduped = [vertices[0]]
    n = len(vertices)
    
    for i in range(1, n):
        pt = vertices[i]
        prev = deduped[-1]
        dx = pt[0] - prev[0]
        dy = pt[1] - prev[1]
        dist = math.sqrt(dx*dx + dy*dy)
        
        # 如果距离大于等于 tol，则认为是不同的顶点，保留它
        if dist >= tol:
            deduped.append(pt)
        else:
            # 如果当前顶点是最后一个，并且与第一个顶点相同，则保留它（表示闭合）
            if i == n - 1 and same_point(pt, vertices[0], tol):
                deduped.append(pt)
            # 否则跳过该点
    return deduped


#  主函数
#  (4)
# 获取最右下角的封闭多边形不影响其他封闭多边形的连续边的顶点列表 

#&&% 分析多边形分支
def analyze_polygon_branches(PL, lines, p1, tol=0.5):
    """
    分析封闭多边形 PL 的分枝情况。PL 为封闭多边形的顶点列表（按逆时针顺序排列），
    p1 为最右下角点，必在 PL 中。
    
    规则：
      1. 从 p1 出发：
         - 沿逆时针方向（不将 p1 视为候选）推进，累计遇到的非多分枝顶点，直至遇到第一个多分枝点 D1；
         - 沿顺时针方向（将 p1 也考虑在内）推进，累计遇到的非多分枝顶点，直至遇到第一个多分枝点 D2。
      2. 将这两个方向累计得到的非多分枝顶点 LN（其中顺时针方向得到的顶点需要反转后与逆时针方向的顶点相接）打印出来。
      3. 由于 PL 是闭合多边形，从 D1 到 D2有两条连续顶点序列，从中选择包含 LN 的那条作为最终返回结果。

    返回：
      返回从 D1 到 D2（包含 D1 和 D2）的连续顶点序列，此序列包含了 LN 的顶点。
      如果任一方向未能找到多分枝点，则打印提示，并返回 None。
    """


    # 内部辅助：判断两个点是否近似相等（仅比较 x,y 坐标）
##    def same_point(a, b, tol=tol):
##        return abs(a[0] - b[0]) <= tol and abs(a[1] - b[1]) <= tol
##


    # 内部辅助：判断某顶点是否为多分枝点
    def is_multi_branch(vertex):
        cnt = 0
        for ln in lines:
            try:
                sp = tuple(ln.StartPoint)
                ep = tuple(ln.EndPoint)
            except Exception:
                continue
            if same_point(vertex, sp, tol) or same_point(vertex, ep, tol):
                cnt += 1
        return cnt > 2

    # 找出 p1 在 PL 中的索引
    try:
        idx = PL.index(p1)
    except Exception as e:
        print("错误：p1 不在 PL 中")
        return None

    n = len(PL)

    # 沿逆时针方向推进（不包括 p1），累计非多分枝顶点 LN_ccw，直到遇到第一个多分枝点 D1
    LN_ccw = []
    D1 = None
    i = (idx + 1) % n
    while i != idx:
        v = PL[i]
        if is_multi_branch(v):
            D1 = v
            break
        else:
            LN_ccw.append(v)
        i = (i + 1) % n
    if D1 is None:
        print("调试：逆时针方向未遇到多分枝点")
        return None

    # 沿顺时针方向推进（包含 p1），累计非多分枝顶点 LN_cw，直到遇到第一个多分枝点 D2
    LN_cw = []
    if is_multi_branch(PL[idx]):
        D2 = PL[idx]
    else:
        D2 = None
        i = (idx - 1) % n
        while i != idx:
            v = PL[i]
            if is_multi_branch(v):
                D2 = v
                break
            else:
                LN_cw.append(v)
            i = (i - 1) % n
        if D2 is None:
            print("调试：顺时针方向未遇到多分枝点")
            return None

    # 合并顺时针方向的顶点（需要反转）和逆时针方向的顶点
    LN = list(reversed(LN_cw)) + LN_ccw

    # 调试打印
    print("调试信息：")
    print("目标共点 p1 =", p1)
    print("逆时针方向收集的非多分枝顶点 (LN_ccw):")
    for pt in LN_ccw:
        print("  ", pt)
    print("顺时针方向收集的非多分枝顶点 (LN_cw, 原顺序):")
    for pt in LN_cw:
        print("  ", pt)
    print("合并后的非多分枝顶点 LN (从 D2 到 D1):")
    for pt in LN:
        print("  ", pt)
    print("逆时针方向遇到的多分枝点 D1:", D1)
    print("顺时针方向遇到的多分枝点 D2:", D2)

    # 计算在 PL 中 D1 和 D2 的索引
    try:
        idx_D1 = PL.index(D1)
        idx_D2 = PL.index(D2)
    except Exception as e:
        print("错误：无法在 PL 中查找 D1 或 D2:", e)
        return None

    # 由于 PL 是闭合的，我们有两条连接 D1 和 D2：
    if idx_D1 <= idx_D2:
        branch_A = PL[idx_D1: idx_D2 + 1]
        branch_B = PL[idx_D2:] + PL[:idx_D1 + 1]
    else:
        branch_A = PL[idx_D1:] + PL[:idx_D2 + 1]
        branch_B = PL[idx_D2: idx_D1 + 1]

    # 选择包含 LN 中顶点的那个分枝
    selected_branch = None
    if LN:
        found_in_A = any(same_point(v, LN[0], tol) for v in branch_A)
        found_in_B = any(same_point(v, LN[0], tol) for v in branch_B)
        if found_in_A and not found_in_B:
            selected_branch = branch_A
        elif found_in_B and not found_in_A:
            selected_branch = branch_B
        elif found_in_A and found_in_B:
            # 如果两条分枝都包含，则选择较短的那条，通常这才是“局部”的分枝
            selected_branch = branch_A if len(branch_A) <= len(branch_B) else branch_B
        else:
            print("调试：LN 中的顶点不在任一分枝上，默认选择 branch_A")
            selected_branch = branch_A
    else:
        print("调试：LN 为空，无法确定包含 LN 的分枝，默认选择 branch_A")
        selected_branch = branch_A

    print("最终返回的从 D1 到 D2（包含 LN 的分枝）顶点序列：")
    for pt in selected_branch:
        print("  ", pt)


    #去重
    selected_branch=deduplicate_vertices(selected_branch, tol=tol)

    return selected_branch

##根据输入的顶点列表判断，将lines的其顶点在该顶点列表的线段移出列表


#&&% 移除指定顶点线段
def remove_lines_in_LBv(lines, LB_v, tol=0.1):
    """
    从 COM 线段列表 lines 中移除那些其两个顶点都在 LB_v 中的线段。
    
    参数：
      lines: COM 线段对象列表，每个对象要求具有 StartPoint 与 EndPoint 属性，
             其值为形如 (x, y, z) 的可迭代对象。
      LB_v: 顶点列表，每个顶点为 (x, y, z) 的元组。
      tol: 判断两顶点近似相等的容差值（默认 0.1）。
    
    返回：
      返回移除满足条件（即两端点都在 LB_v 中）的线段后剩余的线段列表。
    """

    def same_point(pt1, pt2, tol=tol):
        # 只比较 x, y, z 坐标的差值，判断两点是否近似相等
        return abs(pt1[0] - pt2[0]) <= tol and abs(pt1[1] - pt2[1]) <= tol and abs(pt1[2] - pt2[2]) <= tol

    remaining_lines = []
    for ln in lines:
        try:
            sp = tuple(ln.StartPoint)
            ep = tuple(ln.EndPoint)
        except Exception as e:
            sys_logger.info(f"[警告]️ 跳过无效线段，原因: {e}")
            continue

        # 判断该线段的起点和终点是否均存在于 LB_v 中
        sp_in = any(same_point(sp, lb, tol) for lb in LB_v)
        ep_in = any(same_point(ep, lb, tol) for lb in LB_v)
        if not (sp_in and ep_in):
            remaining_lines.append(ln)
        else:
            sys_logger.info(f"删除线段 {ln.Handle}，其两个端点均在 LB_v 中：sp={sp}, ep={ep}")
    return remaining_lines

#  主函数
#  (5)
#&&% 获取全部封闭多边形，但不完全

#&&% 处理多边形
def process_polygons(lines, tol=0.5, max_steps=50, layer_name="测试辅助"):
    """
    递归提取并绘制直线段集 lines 中所有封闭多边形。
    
    流程：
      1. 从 lines 中提取最右下角封闭多边形（调用 find_rightbottom_closed_polygon）。
      2. 对该多边形：
            - 获取最右下角点 p1（调用 find_rightbottom_point）；
            - 将多边形（polygon）加入列表 LB；
            - 调用 analyze_polygon_branches 分析多边形分枝，得到用来判断移除线段的顶点列表 Lv；
            - 调用 remove_lines_in_LBv，将所有两端点均在 Lv 中的直线段从 lines 中移除；
            - 调用 draw_polygon_as_polyline 绘制该多边形，将生成的多段线 COM 对象加入 LBcom；
            - 然后调用 draw_polygon_as_polyline 绘制顶点列表 Lv（蓝色绘制），将生成的多段线 COM 对象加入 LB_yc；
      3. 重复上述过程直至无法提取出封闭多边形；
    
    返回：
      (LB, LBcom, LB_yc)，其中：
         LB: 封闭多边形顶点列表集合（每个为顶点序列）。
         LBcom: 绘制出的多段线 COM 对象列表。
         LB_yc: 绘制蓝色辅助多段线（用于检查移除的线段顶点）的 COM 对象列表。
    """
    LB = []
    LBcom = []
    LB_yc = []  # 存储蓝色辅助多段线 COM 对象，用于检查移除的顶点
    Ly = []  # 用来记录每次移除的直线段
    iteration = 0

    while True:
        iteration += 1
        sys_logger.info(f"\n【迭代 {iteration}】剩余直线段数量 = {len(lines)}")
        p1 = find_rightbottom_point(lines, tol)
        if p1 is None:
            print("未找到最右下角点，结束迭代")
            break
        sys_logger.info(f"当前最右下角点 p1 = {p1}")
        
        polygon = find_rightbottom_closed_polygon(lines, tol=tol, max_steps=max_steps)
        if polygon is None:
            print("无法提取封闭多边形，结束迭代")
            break
        LB.append(polygon)
        print("提取的封闭多边形顶点：")
        for pt in polygon:
            print("  ", pt)
        
        Lv = analyze_polygon_branches(polygon, lines, p1, tol=tol)
        if Lv is None:
            print("分析多分枝失败，结束本次迭代")
        else:
            print("用于移除线段的顶点列表 Lv：")
            for pt in Lv:
                print("  ", pt)
            # 移除 lines 中两端点均在 Lv 中的直线段（调用 remove_lines_in_LBv）
            new_lines = remove_lines_in_LBv(lines, Lv, tol=0.1)
            removed_count = len(lines) - len(new_lines)
            sys_logger.info(f"本次移除直线段数: {removed_count}")
            # 记录被移除的直线段
            for ln in lines:
                if ln not in new_lines:
                    Ly.append(ln)
            lines = new_lines
        
        # 绘制提取出的封闭多边形
        polyline = draw_polygon_as_polyline(polygon, layer_name=layer_name, tol=tol)
        if polyline:
            LBcom.append(polyline)
        
        # 调用 draw_polygon_as_polyline 绘制 Lv（辅助多段线，注意Lv是顶点列表）
        if Lv is not None and len(Lv) > 0:
            poly_blue = draw_polygon_as_polyline(Lv, layer_name=layer_name, tol=tol)
            if poly_blue:
                LB_yc.append(poly_blue)
        
        if len(lines) < 3:
            print("剩余直线段不足以构成封闭多边形，退出")
            break

    return LB, LBcom, LB_yc



#&&% 提取多边形
def extract_polygon_from_lines(lines, tol=0.5):
    """
    将表示封闭多边形边缘的线段（COM对象列表）转换为顶点列表（按顺序排列），
    消除中间相邻的重复顶点。若能成功构成闭合多边形，则返回顶点列表（闭环不重复），否则返回 None。
    
    参数：
      lines: 直线段对象列表，每个对象要求具有 StartPoint 和 EndPoint 属性
      tol: 判断顶点相等的容差（仅比较 x 和 y 坐标）。
    
    返回：
      顶点列表，如 [p1, p2, ..., pn]，其中 p1 表示多边形的起点（不重复列出闭合顶点）。
    """
    if not lines:
        return None
    
    # 定义同点判断函数
##    def same_point(a, b, tol=tol):
##        return abs(a[0] - b[0]) <= tol and abs(a[1] - b[1]) <= tol
    
    # 拷贝线段列表，方便修改（注意：此处不会复制COM对象，仅复制引用）
    remaining_lines = list(lines)
    
    try:
        start_pt = tuple(remaining_lines[0].StartPoint)
    except Exception as e:
        print("提取第一条线段起点失败：", e)
        return None
    vertices = [start_pt]
    current_pt = start_pt

    # 由于多边形是闭合的，最多尝试 2*len(remaining_lines) 次
    for _ in range(len(remaining_lines) * 2):
        found_edge = False
        for ln in remaining_lines:
            try:
                sp = tuple(ln.StartPoint)
                ep = tuple(ln.EndPoint)
            except Exception:
                continue

            nxt = None
            if same_point(current_pt, sp) and (not same_point(current_pt, ep)):
                nxt = ep
            elif same_point(current_pt, ep) and (not same_point(current_pt, sp)):
                nxt = sp

            if nxt is None:
                continue

            # 一旦找到与当前点相连的边，移除此边
            remaining_lines.remove(ln)
            # 如果得到的 nxt 与起点相同，则说明闭合结束
            if same_point(nxt, start_pt):
                vertices.append(start_pt)
                return deduplicate_vertices(vertices, tol)
            else:
                vertices.append(nxt)
                current_pt = nxt
                found_edge = True
                break
        if not found_edge:
            # 未能找到与当前点相连的边，构成闭合失败
            break
    # 如果循环结束后闭合未完成，则返回 None
    return None



#将多段线列表炸开为线段，返回线段列表



#&&% 炸开多段线
def explode_polylines(LB):
    """
    对多段线列表 LB 中的每一个多段线，调用 .Explode() 方法，
    并将炸开得到的所有线段合并为一个新的线段列表返回。

    参数:
      LB: 多段线 COM 对象列表，每个对象应支持 .Explode() 方法，
          如 pl = polyline_explode_object.Explode() 返回该多段线炸开后的线段集合。
          
    返回:
      一个包含所有炸开后直线段 COM 对象的列表。
    """
    exploded_lines = []
    for pl in LB:
        try:
            # 调用 Explode() 方法，返回一个集合（例如 COM Collection）
            exploded = pl.Explode()
            # 遍历返回的集合，将每根线段添加到列表中
            # 注意：遍历 COM Collection 的方法可能因环境不同而不同，此处假设可以直接迭代
            for ln in exploded:
                exploded_lines.append(ln)
        except Exception as e:
            handle = getattr(pl, 'Handle', '未知')
            sys_logger.info(f"[警告]️ 处理多段线 {handle} 时出错: {e}")
    return exploded_lines


#lines1 中那些不在 lines2 中的线段

#&&% 线段集相减
def subtract_line_sets(lines1, lines2, tol=0.5):
    """
    比较两个线段集合 lines1 和 lines2，返回 lines1 中那些不在 lines2 中的线段。

    参数：
      lines1: 第一组线段（候选集合），每个对象须具有 StartPoint 和 EndPoint 属性。
      lines2: 第二组线段，参照集合。
      tol: 判断点是否相同的容差，默认值为 0.5。
      
    返回：
      lines1 中所有不与 lines2 中任意线段相同的线段构成的列表。
    """
    result = []
    for ln in lines1:
        found = False
        for ln2 in lines2:
            if same_line(ln, ln2, tol):
                found = True
                break
        if not found:
            result.append(ln)
    return result

#  主函数
#  (6)
#&&% 获取全部封闭多边形


#&&% 最终处理
def process_final(lines, tol=0.5, max_steps=50, layer_name="测试辅助"):

    print("len(lines):",len(lines))

    L1, L2, L3 = process_polygons(lines, tol=tol, max_steps=max_steps, layer_name=layer_name)

    sys_logger.info(f"process_polygons 完成：\n  L1 数量 = {len(L1)}\n  L2 数量 = {len(L2)}\n  L3 数量 = {len(L3)}")
    
    print(">>> 对 L3 中的多段线执行 Explode() 操作...")

    exploded_lines=explode_polylines(L3)



    shengyu = subtract_line_sets(lines, exploded_lines, tol=tol)

    for x in shengyu:

        print(x.StartPoint,x.EndPoint)

    L_shengyu=extract_polygon_from_lines(shengyu, tol=tol)
   

    ld = draw_polygon_as_polyline(L_shengyu, layer_name=layer_name, tol=tol)

    L2.append(ld)

    L1.append(L_shengyu)
    for x in exploded_lines:
        x.Delete()

    return L1, L2, L3



#&&&% PL复线处理


"""
研究PL复线处理的问题

(1)多段线的基本操作 get_unique_vertices_from_pl_com

(2)PL打印线 generate_name_and_ratio_from_polyline(polyline,A3dy=0)

(3)将正交六边形多段线分成两个矩形区域 split_hexagon_combined(polygon, tol=0.1, simplify_tol=0.5)

(4)获取多段线的上下左右边界的直线段，返回线段端点列表 get_bbox_edge_segments(pl, tol=0.5)

(5)获取多段线的内部的文字 get_texts_in_polyline(com_pl, tol=0.5)

(6)多段线上的均分插入 distribute_points_on_entity(entity, n, block, scale_factor, ys)

(7)返回 pl1 中与 pl2 “共线且有重叠”的区段列表 common_segments_between_polylines(pl1, pl2, tol=0.5)

(8)找到全部“两根多段线耦合成一个矩形”的多段线 two_plines_making_rectangle(pl1, pl2, tol=0.5)






"""
#  主函数
#  (1)
# 多段线的基本操作

#  该函数系列包括如下一些函数
"""

标准顶点坐标列表,是像[(0, 0, 0), (0, 100, 0), (0, 900, 0)]这样的列表

它表示3个连续顶点2根连续线段，公共点坐标不重复

"""
## 0 轻量多段线和一般传统多段线各有用处，后者才能在三维中使用，更广泛




#&&% 绘制轻量多段线
@alias("画轻量多段线")
def draw_lwpolyline(
    coords3d: list[tuple[float, float, float]],
    layer_name: str = "0",
    width: float = 0.0,
    color: int = 256,
    closed: bool = False
):
    """
    根据一组 (x, y, z) 坐标点绘制轻量级多段线（LWPOLYLINE）。

    :param coords3d: 形如 [(x1, y1, z1), (x2, y2, z2), …] 的点列表，
                     仅使用 x,y 坐标，忽略 z。
    :param layer_name: 目标图层名称，不存在则自动创建。
    :param width:      多段线恒宽 (ConstantWidth)。
    :param color:      颜色索引 (AutoCAD Color Index)，256=BYLAYER。
    :param closed:     是否闭合多段线（首尾相连）。

    :return:           新建的轻量级多段线对象 (COM AddLightWeightPolyline)。

    pts = [
        (0.0, 0.0, 0.0),
        (100.0, 0.0, 0.0),
        (100.0, 50.0, 0.0),
        (0.0, 50.0, 0.0),
    ]
    poly = draw_lwpolyline(
        coords3d=pts,
        layer_name="dy_quyu",
        width=0.0,
        color=1,      # 红色
        closed=True
    )
    poly.Coordinates
    (0.0, 0.0, 100.0, 0.0, 100.0, 50.0, 0.0, 50.0)

    len((0.0, 0.0, 100.0, 0.0, 100.0, 50.0, 0.0, 50.0))
    8

    """
    # 1️⃣ 连接 AutoCAD

    # 2️⃣ 确保图层存在
    layers = doc.Layers
    try:
        lyr = layers.Item(layer_name)
    except Exception:
        lyr = layers.Add(layer_name)
    # Optional: 开启图层
    lyr.LayerOn = True

    # 3️⃣ 准备坐标数组：扁平化 x,y
    raw = []
    for x, y, _ in coords3d:
        raw.extend((x, y))
    # 转 COM VARIANT 数组
    arr = win32com.client.VARIANT(
        pythoncom.VT_ARRAY | pythoncom.VT_R8,
        raw
    )

    # 4️⃣ 绘制轻量级多段线

    try:

        pline = mp.AddLightWeightPolyline(arr)
        pline.Layer         = layer_name
        pline.ConstantWidth = width
        pline.color         = color
        pline.Closed        = bool(closed)

        sys_logger.info(f"[OK] 已在图层『{layer_name}』绘制多段线，Closed = {closed}")
        return pline
    except Exception as e:
        print("[错误] 绘制多段线失败:", e) 

   # 5️⃣ 返回新对象

    return pline

#&&% 绘制轻量多段线20260113
@alias("画轻量多段线")
def draw_lwpolyline(
    coords3d: list[tuple[float, float, float]],
    layer_name: str = "0",
    width: float = 0.0,
    color: int = 256,
    closed: bool = False
):
    """
    【通用版】支持在 模型空间 或 任意布局空间 绘制。
    它会自动检测当前激活的是哪个空间，就画在哪里。
    """
    # 1️⃣ 获取文档对象 (假设 C.doc 是全局的，如果不是，建议通过参数传入或者在函数内获取)
    doc = C.doc 
    
    # 2️⃣ 关键修改：动态获取绘制目标 (ModelSpace 或 PaperSpace)
    # doc.ActiveLayout.Block 代表当前激活界面的“容器”
    # 如果你在模型界面，它就是模型空间；如果你在布局界面，它就是当前那张纸
    target_space = doc.ActiveLayout.Block 

    # 3️⃣ 确保图层存在
    layers = doc.Layers
    try:
        lyr = layers.Item(layer_name)
    except Exception:
        lyr = layers.Add(layer_name)
    lyr.LayerOn = True

    # 4️⃣ 准备坐标数组
    raw = []
    for x, y, _ in coords3d:
        raw.extend((x, y))
    
    # 转 COM VARIANT 数组
    arr = win32com.client.VARIANT(
        pythoncom.VT_ARRAY | pythoncom.VT_R8,
        raw
    )

    # 5️⃣ 绘制
    try:
        # 使用 target_space 替代原来的 mp
        pline = target_space.AddLightWeightPolyline(arr)
        
        pline.Layer         = layer_name
        pline.ConstantWidth = width
        pline.color         = color
        pline.Closed        = bool(closed)

        # sys_logger.info(f"[OK] 已在图层『{layer_name}』绘制多段线")
        return pline
    except Exception as e:
        sys_logger.info(f"[错误] 绘制多段线失败: {e}") 
        return None



# 1 从com复线获取标准顶点坐标列表

#&&% 获取唯一顶点
def get_unique_vertices_from_pl_com(pl_com):
    """
    提取多段线的顶点列表，不重复连续线段的公共顶点，返回顶点列表。
    
    参数:
        pl_com: AutoCAD 中的 Polyline COM 对象（AcDbPolyline）
        
    返回:
        顶点列表，每两个点构成一条线段，格式：[ (x1, y1, z1), (x2, y2, z2), ... ]
        
    """
    # 获取二维坐标数据
    coords = pl_com.Coordinates
    vertices = []

    # 以每两个坐标为一组，构成线段
    for i in range(0, len(coords) - 1, 2):
        x1, y1 = coords[i], coords[i + 1]
        z1 = 0  # 假设z坐标为0，如果需要，可以通过某种方式获取真实z坐标
        if not vertices:
            vertices.append((x1, y1, z1))
        else:
            # 如果当前点与上一个点不重复，添加到顶点列表
            if (x1, y1, z1) != vertices[-1]:
                vertices.append((x1, y1, z1))
    
    # 添加最后一个点，避免遗漏
    last_x, last_y = coords[-2], coords[-1]
    z_last = 0  # 同样，假设最后的z坐标为0
    if (last_x, last_y, z_last) != vertices[-1]:
        vertices.append((last_x, last_y, z_last))
    
    return vertices

# 2 将com线段转成顶点坐标列表，一根线段一个列表

#&&% 线段转点集
def convert_lines_to_points(segments):
    """
    将com线段列表转换为顶点列表，每条线段的两个端点作为一个独立的列表。

    参数:
      segments: 线段对象列表，每个元素具有 StartPoint 和 EndPoint 属性。

    返回:
      包含多个线段顶点的列表，每个线段是一个包含两个端点坐标的列表。
    """
    points_list = []

    for segment in segments:
        # 提取线段的两个端点
        start_point = tuple(segment.StartPoint)
        end_point = tuple(segment.EndPoint)

        # 将线段的两个端点存入一个列表，添加到结果列表
        points_list.append([start_point, end_point])

    return points_list

# 3 合并顶点列表表示的连续线段，允许多根断开的连续线段

#&&% 合并线段
def merge_segments_new(LB, tol=0.5):
    """
    使用convert_lines_to_points 将线段实体转成顶点列表表达式后就可以使用此命令

    不断合并连接，能连接的都会连接   

    """
    def same(p, q):
        return abs(p[0]-q[0]) <= tol and abs(p[1]-q[1]) <= tol

    # 1) 为端点做哈希 — 用 round() 把 tol 纳入 key，避免浮点字典键难比较
    def key(pt):
        return (round(pt[0]/tol), round(pt[1]/tol))   # 只 hash XY

    buckets = defaultdict(list)   # key(pt)  ->  [(seg_index, dir), ...]
    for idx, seg in enumerate(LB):
        a, b = seg[0], seg[-1]
        buckets[key(a)].append((idx, +1))   #  +1 表示 seg[0] 方向
        buckets[key(b)].append((idx, -1))   #  -1 表示 seg[-1] 方向

    used = [False]*len(LB)
    sequences = []

    while True:
        # 找到尚未使用的第一条线段
        try:
            seed_idx = next(i for i,u in enumerate(used) if not u)
        except StopIteration:
            break                           # 全部用完
        used[seed_idx] = True
        seq = deque(LB[seed_idx])           # 双端队列便于首尾增长

        # 函数: 把可接的线段拼到 deque 的一头
        def grow(at_tail: bool):
            while True:
                end_pt = seq[-1] if at_tail else seq[0]
                bucket = buckets[key(end_pt)]
                # 移除已用完的骨牌
                bucket[:] = [pair for pair in bucket if not used[pair[0]]]
                if not bucket:              # 再也接不上
                    break
                idx, direction = bucket.pop()
                used[idx] = True
                seg = LB[idx]
                # 根据 direction 决定正向还是反向加入
                if direction == +1:         # bucket 点是 seg[0]
                    add = seg[1:]           # 去掉公共点再拼
                else:                       # bucket 点是 seg[-1]
                    add = seg[-2::-1]       # 反向、去掉公共点
                if at_tail:
                    seq.extend(add)
                else:
                    seq.extendleft(add[::-1])   # extendleft 要反转

        # 先往尾巴拼，再往头拼（顺序无所谓，都会做到极限）
        grow(True)
        grow(False)
        sequences.append(list(seq))

    return sequences

# 4 绘制连续PL多段线，断开的PL多段线要分开绘制

#&&% 绘制多段线

@retry_if_busy(max_retries=5, delay=0.2)
def draw_polyline(
    vertices,
    layer_name="tuqian_baobu",
    tol=0.5,
    width=50,
    color=256,
    target_space=None # ✨ 关键参数：指定绘制在哪个空间 (Block容器)
):
    """
    【修正版 V3】绘制轻量多段线 (LWPolyline)
    1. 移除了 li()，使用 C.doc。
    2. 支持 target_space 参数，实现 Model/Layout 自适应绘制。
    """
    
    # 1. 获取文档 (规范化使用 C.doc)
    doc = getattr(C, 'doc', None)
    if doc is None:
        sys_logger.error("[Draw] 失败: C.doc 未初始化")
        return None

    # 2. 确定绘制容器
    # 如果没传 target_space，默认为当前激活的布局/空间
    container = target_space
    if container is None:
        try:
            container = doc.ActiveLayout.Block
        except:
            container = doc.ModelSpace

    # ----------------- 内部工具 -----------------
    def same_point(p1, p2, _tol=tol):
        return abs(p1[0] - p2[0]) <= _tol and abs(p1[1] - p2[1]) <= _tol

    # ---------- 数据准备 ----------
    if not vertices or len(vertices) < 2:
        return None

    # 处理闭合
    is_closed = False
    if len(vertices) > 2 and same_point(vertices[0], vertices[-1]):
        is_closed = True

    # 展平坐标 (LWPolyline 需要 [x1, y1, x2, y2...])
    flat = []
    for pt in vertices:
        if len(pt) >= 2:
            flat.extend([pt[0], pt[1]])
        else:
            return None
            
    try:
        # 创建 COM 数组
        coords_var = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, tuple(flat))
    except Exception as e:
        sys_logger.error(f"[Draw] 坐标转换失败: {e}")
        return None

    # ---------- 绘制 ----------
    pl = None
    try:
        # 🟢 核心修正：在指定容器中创建，而不是 ModelSpace
        pl = container.AddLightWeightPolyline(coords_var)
        
        # --- 属性设置 ---
        # 分开设置以防单项失败导致整体崩溃
        try:
            if layer_name: pl.Layer = layer_name
        except: pass

        try:
            pl.Color = color
            pl.Closed = is_closed
            pl.ConstantWidth = width
        except: pass

        return pl

    except Exception as e:
        sys_logger.error(f"[Draw] AddLightWeightPolyline 失败: {e}")
        return None

# 5 获取多段线后的线段列表，原来的多段线仍然存在


"""
pl=  Pl_obj.Explode()

"""



# 6 将多条直线段（允许不连续）连接成PL复线

#&&% 线段转多段线

def lines_to_polylines(Lc=None,
                        tol=0.5,
                        layer_name="tuqian_baobu",
                        width=50,
                        color=256):
    """
    将直线段合并为多段线 (空间自适应版)
    """
    sys_logger.info(f"\n--- 🔍 开始 lines_to_polylines (Layer: {layer_name}) ---")

    # 0. 规范化连接 (替代 li())
    try:
        C.li() # 刷新连接
        doc = C.doc
    except Exception as e:
        sys_logger.error(f"CAD 连接失败: {e}")
        return []

    # 1. 获取源对象
    raw_objects = []
    if Lc is None:
        sys_logger.info(f"[Step 1] 搜索图层: {layer_name}")
        raw_objects = stc(layer_name)
    else:
        raw_objects = Lc

    if not raw_objects:
        sys_logger.info("[Step 1] 未找到对象")
        return []

    # 2. 筛选直线并确定宿主空间
    valid_lines = []
    target_container = None # 用于存放这些线所在的 Block (Model 或 Layout)

    for obj in raw_objects:
        try:
            if get_attr(obj, "ObjectName") == "AcDbLine":
                valid_lines.append(obj)
                
                # 🔥 核心逻辑：捕获第一条有效直线的“宿主”
                if target_container is None:
                    owner_id = obj.OwnerID
                    # 通过 OwnerID 反向获取容器对象 (Layout Block 或 ModelSpace)
                    try:
                        target_container = doc.ObjectIDToObject(owner_id)
                    except:
                        pass
        except: 
            continue

    if not valid_lines:
        return []

    # 兜底：如果没能通过 OwnerID 找到容器，就用当前激活的
    if target_container is None:
        target_container = doc.ActiveLayout.Block

    sys_logger.info(f"[Step 2] 有效直线: {len(valid_lines)} 条 | 目标容器: {get_attr(target_container, 'Name', 'Unknown')}")

    # 3. 转换点
    LB = convert_lines_to_points(valid_lines)
    if not LB: return []

    # 4. 合并段
    LK = merge_segments_new(LB, tol=tol)
    if not LK: return []

    # 5. 生成多段线
    PLs = []
    for verts in LK:
        if len(verts) < 2: continue
        
        # 🟢 传入探测到的 target_container
        pl = draw_polyline(verts,
                           layer_name=layer_name,
                           tol=tol,
                           width=width,
                           color=color,
                           target_space=target_container) # <--- 传入正确的空间
        if pl:
            PLs.append(pl)

    # 6. 删除原直线
    if PLs:
        deleted_count = 0
        for ln in valid_lines:
            try: 
                safe_delete(ln)
                deleted_count += 1
            except: pass
        sys_logger.info(f"✅ 生成 {len(PLs)} 条多段线，删除 {deleted_count} 条原直线")
            
    return PLs

# 7 找到多段线的最左下角的点

#&&% 查找最小点
def find_min_point(obj):
    """
    获取任意对象的左下角坐标（通过其外包盒）。

    :param obj: 支持 GetBoundingBox() 方法的 COM 对象（如多段线、块参照等）。
    :return:    (min_x, min_y) 元组，表示对象外包盒的左下角坐标
    """
    try:
        ll_point, _ = obj.GetBoundingBox()
        min_x, min_y, _ = ll_point
        return min_x, min_y
    except Exception as e:
        sys_logger.info(f"[错误] 获取外包盒失败: {e}")
        return None, None

# 8 找到多段线的最右上角的点

#&&% 查找最大点
def find_max_point(obj):
    """
    获取任意对象的右上角坐标（通过其外包盒）。

    :param obj: 支持 GetBoundingBox() 方法的 COM 对象（如多段线、块参照等）。
    :return:    (max_x, max_y) 元组，表示对象外包盒的右上角坐标
    """
    try:
        _, ur_point = obj.GetBoundingBox()
        max_x, max_y, _ = ur_point
        return max_x, max_y
    except Exception as e:
        sys_logger.info(f"[错误] 获取外包盒失败: {e}")
        return None, None


#&&% 计算距离
def distance(point1, point2):
    """计算两点之间的距离"""
    return ((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)**0.5



# 10 定义矩形

#&&% 定义矩形
def define_rectangle_by_diagonal(p1, p2):
    """
    使用两个对角顶点定义矩形，矩形的边与坐标轴平行或垂直。
    p1, p2: 分别为矩形的一对对角顶点，格式为 (x, y)。
    返回矩形的四个顶点、长和宽。
    """
    x1, y1 = p1
    x2, y2 = p2

    # 计算两个边长
    length_x = abs(x2 - x1)
    length_y = abs(y2 - y1)

    # 确定长和宽
    length = max(length_x, length_y)
    width = min(length_x, length_y)

    # 确定矩形的四个顶点
    rectangle_points = [(x1, y1), (x1, y2), (x2, y2), (x2, y1)]

    return rectangle_points, length, width

#&&% 定义矩形X
def define_rectangle_by_diagonal_x(p1, p2):
    """
    使用两个对角顶点定义矩形，矩形的边与坐标轴平行或垂直。
    p1, p2: 分别为矩形的一对对角顶点，格式为 (x, y)。
    返回矩形的四个顶点、长和宽。
    """
    x1, y1 = p1
    x2, y2 = p2

    # 计算两个边长
    length_x = abs(x2 - x1)
    length_y = abs(y2 - y1)

    # 确定长和宽
    length = max(length_x, length_y)
    width = min(length_x, length_y)

    # 确定矩形的四个顶点
    rectangle_points = [x1, y1, x2, y1, x2, y2, x1, y2]

    return rectangle_points






def expand_rectangle(p1, p2, offset=130):
    """
    给定矩形框的两个对角点（p1 和 p2），
    返回在四个方向扩展 offset 后的新对角点 P1 和 P2。
    """

    x1, y1, z1 = p1
    x2, y2, z2 = p2

    # 确保 p1 是左下角，p2 是右上角（即使输入反了）
    x_min, x_max = sorted([x1, x2])
    y_min, y_max = sorted([y1, y2])
    z = z1  # z 坐标保持一致

    # 四向扩展 offset，构造新的矩形框
    P1 = (x_min - offset, y_min - offset, z)
    P2 = (x_max + offset, y_max + offset, z)

    return P1, P2





def parse_rectangle_points(*args):
    """
    接收多种坐标格式输入，统一解析为矩形四角点：
    返回：
        (左下, 右上, 左上, 右下)，每个为三元组 (x, y, z)
    
    合法输入形式：
        - (x1, y1, z1), (x2, y2, z2)
        - [(x1, y1, z1), (x2, y2, z2)]
        - (x1, y1, x2, y2)
        - (x1, y1, 0, x2, y2, 0)
    """
    try:
        # 解包列表输入
        if len(args) == 1 and isinstance(args[0], list):
            args = args[0]

        # 标准化为两个三维点
        if len(args) == 2 and all(isinstance(pt, (tuple, list)) and len(pt) == 3 for pt in args):
            p1, p2 = args
        elif len(args) == 4:
            x1, y1, x2, y2 = args
            p1, p2 = (x1, y1, 0), (x2, y2, 0)
        elif len(args) == 6:
            x1, y1, z1, x2, y2, z2 = args
            p1, p2 = (x1, y1, z1), (x2, y2, z2)
        else:
            raise ValueError("输入格式不合法")

        # 解析 min/max 点坐标
        x_min = min(p1[0], p2[0])
        x_max = max(p1[0], p2[0])
        y_min = min(p1[1], p2[1])
        y_max = max(p1[1], p2[1])
        z = p1[2] if len(p1) > 2 else 0

        # 四角坐标
        left_bottom = (x_min, y_min, z)
        right_top = (x_max, y_max, z)
        left_top = (x_min, y_max, z)
        right_bottom = (x_max, y_min, z)

        return left_bottom, right_top, left_top, right_bottom

    except Exception as e:
        sys_logger.info(f"[错误] 解析矩形点失败: {e}")
        return None

#&&&% 模型空间选出矩形多段线

def get_rectangular_polylines(min_side=100.0, area_tolerance=0.02):
    """
    【智能筛选】获取所有“矩形”多段线 (兼容轻量线和老式线)。
    
    逻辑：
    1. 合并两种多段线列表 (物理互斥，直接相加)。
    2. 过滤掉尺寸太小的。
    3. 核心：通过 (自身面积 / 包围盒面积) 判断是否为矩形。
       完美排除 L形、三角形、凹多边形。

    Args:
        min_side: 最小边长 (默认100mm)
        area_tolerance: 面积容差 (默认 2%)

    Returns:
        List[COMObject]: 筛选出的矩形多段线列表
    """
    from  CAD_file_operations   import get_obj_loc    
    
    # 1. 获取并合并列表 (既然类型互斥，直接相加即可)
    candidates_1 = []
    try:
        l1 = select_polyline() or []
        if l1: candidates_1.extend(l1)
    except: pass
    
    try:
        l2 = select_polyline_chuantong() or []
        if l2: candidates_1.extend(l2)
    except: pass
    
    sys_logger.info(f"[筛选] 待检查多段线总数: {len(candidates_1)}")
    
    candidates=[ ent for ent in candidates_1 if  get_obj_loc(ent) == 1]



    results = []
    
    for pl in candidates:
        try:
            # --- A. 尺寸检查 (Bounding Box) ---
            # GetBoundingBox 对 LWPolyline 和 Polyline 都通用
            min_pt, max_pt = pl.GetBoundingBox()
            dx = abs(max_pt[0] - min_pt[0])
            dy = abs(max_pt[1] - min_pt[1])
            
            # 过滤掉“一条线”的情况 (宽度极小)
            if dx < 1e-3 or dy < 1e-3:
                continue
                
            # 过滤掉太小的 (如小于100mm)
            if max(dx, dy) < min_side or min(dx, dy) < min_side:
                continue
                
            # --- B. 形状检查 (核心逻辑: 面积法) ---
            # Area 属性对 LWPolyline 和 Polyline 也都通用
            real_area = abs(getattr(pl, "Area", 0))
            box_area = dx * dy
            
            if box_area <= 0: continue

            # 计算 (真实面积 - 包围盒面积) 的差异比例
            diff_ratio = abs(real_area - box_area) / box_area
            
            # 矩形的面积应该几乎填满包围盒
            # 容差 0.02 允许 2% 的绘图误差或圆角
            if diff_ratio <= area_tolerance:
                results.append(pl)

        except Exception:
            # 个别对象无法获取 Area 或 BoundingBox，直接跳过
            continue

    sys_logger.info(f"[筛选] 最终获得矩形数量: {len(results)}")
    return results

#&&%图纸空间矩形多段线


##def get_layout_rectangular_polylines_coords(layout_name, min_side=100.0):
    ##"""
    ##【早期绑定专用】基于 get_attr 的矩形分析 (已集成 sys_logger)
    ##"""
    ##doc = C.doc
    ##results = []
    ##
    ### [普通信息] -> info (可静音)
    ##sys_logger.info(f"🚀 [COM安全模式] 启动布局扫描: {layout_name}")
##
    ##try:
        ##target_layout = doc.Layouts.Item(layout_name)
        ##layout_block = target_layout.Block
    ##except Exception as e:
        ### [致命错误] -> error (永远显示)
        ##sys_logger.error(f"❌ 获取布局失败: {e}")
        ##return []
##
    ##sys_logger.info(f"📊 容器内对象总数: {layout_block.Count}")
    ##
    ### 遍历
    ##for i, obj in enumerate(layout_block):
        ##try:
            ### --- 1. 安全获取类型 ---
            ##obj_name = get_attr(obj, "ObjectName") 
            ##
            ##if not obj_name or "Polyline" not in str(obj_name):
                ##continue
##
            ##hdl = get_attr(obj, "Handle")
            ### [调试详情] -> info (可静音)
            ##sys_logger.info(f"\n🔍 [分析对象] Handle={hdl} | Type={obj_name}")
##
            ### --- 2. 安全获取坐标 ---
            ##coords = get_attr(obj, "Coordinates")
            ##
            ##if not coords:
                ##sys_logger.info("   ❌ 坐标为空 (None)")
                ##continue
                ##
            ##coords = list(coords)
            ##total_len = len(coords)
##
            ### --- 3. 智能判断维度 ---
            ##step = 2
            ##if "AcDb2dPolyline" in obj_name or "AcDb3dPolyline" in obj_name:
                ##step = 3
            ##
            ##num_points = total_len // step
            ##sys_logger.info(f"   📐 数据结构: 坐标总数={total_len}, 推断步长={step}, 顶点数={num_points}")
##
            ### 打印预览
            ##sys_logger.info(f"   💾 原始数据前预览: {coords[:6]}...")
##
            ##if num_points < 4:
                ##sys_logger.info("   ❌ 忽略: 顶点少于4个")
                ##continue
##
            ### --- 4. 提取 X 和 Y ---
            ##xs = coords[0::step]
            ##ys = coords[1::step]
##
            ### --- 5. 尺寸判定 ---
            ##min_x, max_x = min(xs), max(xs)
            ##min_y, max_y = min(ys), max(ys)
            ##
            ##w = max_x - min_x
            ##h = max_y - min_y
            ##
            ##sys_logger.info(f"   📏 计算尺寸: W={w:.2f}, H={h:.2f} (阈值 {min_side})")
            ##
            ##if w < min_side or h < min_side:
                ##sys_logger.info("   ❌ 忽略: 尺寸过小")
                ##continue
##
            ### --- 6. 正交矩形判定 ---
            ##coord_tol = 1.0 
            ##
            ##is_rect = True
            ##
            ### 检查 X 轴
            ##x_bad_count = 0
            ##for x in xs:
                ##if not (abs(x - min_x) <= coord_tol or abs(x - max_x) <= coord_tol):
                    ##x_bad_count += 1
            ##
            ### 检查 Y 轴
            ##y_bad_count = 0
            ##for y in ys:
                ##if not (abs(y - min_y) <= coord_tol or abs(y - max_y) <= coord_tol):
                    ##y_bad_count += 1
            ##
            ##if x_bad_count > 0 or y_bad_count > 0:
                ##sys_logger.info(f"   ❌ 形状不规则: X轴偏离点={x_bad_count}, Y轴偏离点={y_bad_count}")
                ##continue
##
            ### --- 7. 成功 ---
            ##sys_logger.info("   ✅ >> 匹配成功！")
            ##results.append(obj)
##
        ##except Exception as e:
            ### [非致命警告] -> warning (永远显示)
            ### 捕获单个对象处理时的异常，不中断整个循环
            ##sys_logger.warning(f"   ⚠️ [处理异常] Handle={get_attr(obj, 'Handle', '未知')}: {e}")
            ##continue
##
    ##sys_logger.info(f"\n✅ [扫描结束] 最终找到 {len(results)} 个矩形")
    ##return results
##
##

def get_layout_rectangular_polylines_coords(layout_name, min_side=100.0):
    """
    【早期绑定专用】基于 get_attr 的矩形分析 (已集成 sys_logger)
    【V2.0 动态容差版】修复大图框因微小形变被判定为不规则的问题。
    """
    doc = C.doc
    results = []
    
    # [普通信息] -> info
    sys_logger.info(f"🚀 [COM安全模式] 启动布局扫描: {layout_name}")

    try:
        target_layout = doc.Layouts.Item(layout_name)
        layout_block = target_layout.Block
    except Exception as e:
        sys_logger.error(f"❌ 获取布局失败: {e}")
        return []

    sys_logger.info(f"📊 容器内对象总数: {layout_block.Count}")
    
    # 遍历
    for i, obj in enumerate(layout_block):
        try:
            # --- 1. 安全获取类型 ---
            obj_name = get_attr(obj, "ObjectName") 
            
            if not obj_name or "Polyline" not in str(obj_name):
                continue

            hdl = get_attr(obj, "Handle")
            # [调试详情]
            # sys_logger.info(f"\n🔍 [分析对象] Handle={hdl} | Type={obj_name}")

            # --- 2. 安全获取坐标 ---
            coords = get_attr(obj, "Coordinates")
            
            if not coords:
                continue
                
            coords = list(coords)
            total_len = len(coords)

            # --- 3. 智能判断维度 ---
            step = 2
            if "AcDb2dPolyline" in obj_name or "AcDb3dPolyline" in obj_name:
                step = 3
            
            num_points = total_len // step

            if num_points < 4:
                continue

            # --- 4. 提取 X 和 Y ---
            xs = coords[0::step]
            ys = coords[1::step]

            # --- 5. 尺寸判定 ---
            min_x, max_x = min(xs), max(xs)
            min_y, max_y = min(ys), max(ys)
            
            w = max_x - min_x
            h = max_y - min_y
            
            # sys_logger.info(f"   📏 计算尺寸: W={w:.2f}, H={h:.2f} (阈值 {min_side})")
            
            if w < min_side or h < min_side:
                # sys_logger.info("   ❌ 忽略: 尺寸过小")
                continue

            # --- 6. 正交矩形判定 (核心修复) ---
            # 动态计算容差：短边 * 0.0005，限制在 [0.1, 10.0]
            current_min_side = min(w, h)
            coord_tol = current_min_side * 0.0005
            coord_tol = max(0.1, min(coord_tol, 10.0))
            
            # sys_logger.info(f"   📐 动态容差: {coord_tol:.4f} (基准短边: {current_min_side:.1f})")
            
            x_bad_count = 0
            for x in xs:
                if not (abs(x - min_x) <= coord_tol or abs(x - max_x) <= coord_tol):
                    x_bad_count += 1
            
            y_bad_count = 0
            for y in ys:
                if not (abs(y - min_y) <= coord_tol or abs(y - max_y) <= coord_tol):
                    y_bad_count += 1
            
            if x_bad_count > 0 or y_bad_count > 0:
                # sys_logger.info(f"   ❌ 形状不规则: X轴偏离={x_bad_count}, Y轴偏离={y_bad_count} (容差={coord_tol:.3f})")
                continue

            # --- 7. 成功 ---
            # sys_logger.info("   ✅ >> 匹配成功！")
            results.append(obj)

        except Exception as e:
            sys_logger.warning(f"   ⚠️ [处理异常] Handle={get_attr(obj, 'Handle', '未知')}: {e}")
            continue

    sys_logger.info(f"✅ [扫描结束] 最终找到 {len(results)} 个矩形")
    return results


#&&&%  分析打印线20260110


def generate_name_and_ratio_from_com(
    comobj,
    A3dy=0,
    Fandy=("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3", 0),
    tol=10, 
):
    """
    【V5.0 强制兜底版】
    逻辑：
    1. 严格筛选：排除非多段线、非矩形、尺寸过小(<100)的对象 -> return 0
    2. 优先匹配：在 tol 容差范围内寻找标准图框。
    3. 强制兜底：如果找不到标准图框，直接返回“最接近”的那个（将就模式），绝不返回 0。
    """
    import math

    # =========================================================================
    # 第一阶段：严格的合法性校验 (不合格直接踢出)
    # =========================================================================
    
    # 1. 类型检查
    try:
        obj_name = getattr(comobj, "ObjectName", "")
        if "Polyline" not in obj_name:
            return 0
    except: return 0

    # 2. 几何尺寸获取
    try:
        min_pt, max_pt = comobj.GetBoundingBox()
        dx = abs(max_pt[0] - min_pt[0])
        dy = abs(max_pt[1] - min_pt[1])
        
        # 3. 极小尺寸过滤 (短边小于100直接扔)
        if dx < 100 or dy < 100: 
            return 0

        # 4. 矩形形状校验 (防止三角形、L型)
        # 允许 2% 的面积误差
        try:
            real_area = abs(getattr(comobj, "Area", 0))
            box_area = dx * dy
            if box_area > 0 and abs(real_area - box_area) / box_area > 0.02:
                return 0 
        except: pass

        # 统一长宽方向
        length = max(dx, dy)
        width = min(dx, dy)
        orientation_flag = 1 if dy > dx else 0

    except Exception:
        return 0

    # 到这里，说明它肯定是一个合格的矩形多段线。
    # 接下来必须给它返回一个值，不能再 return 0 了。

    # —————————— 强制指定A3模式 ——————————
    if A3dy == 1:
        return (Fandy[0], Fandy[1], Fandy[2], orientation_flag)

    # —————————— 设定动态容差 ——————————
    if length > 1783.5:
        dynamic_tol = 10.0
    else:
        dynamic_tol = 1.0

    # —————————— 数据定义 ——————————
    LB_dayingkuang = [
        (118900, 84100, 100),  (178350, 126150, 150),   (59450, 42050, 50),    (29725, 21025, 25), 
        (133800, 84100, 100),  (200700, 126150, 150),   (66900, 42050, 50),    (33450, 21025, 25), 
        (148600, 84100, 100),  (222900, 126150, 150),   (74300, 42050, 50),    (37150, 21025, 25), 
        (84100,  59400, 100),  (126150, 89100,  150),   (42050, 29700, 50),    (21025, 14850, 25), 
        (105100, 59400, 100),  (157650, 89100,  150),   (52550, 29700, 50),    (26275, 14850, 25), 
        (126100, 59400, 100),  (189150, 89100,  150),   (63050, 29700, 50),    (31525, 14850, 25), 
        (147100, 59400, 100),  (220650, 89100,  150),   (73550, 29700, 50),    (36775, 14850, 25), 
        (59400,  42000, 100),  (89100,  63000,  150),   (29700, 21000, 50),    (14850, 10500, 25), 
        (74300,  42000, 100),  (111450, 63000,  150),   (37150, 21000, 50),    (18575, 10500, 25), 
        (89100,  42000, 100),  (133650, 63000,  150),   (44550, 21000, 50),    (22275, 10500, 25), 
        (104100, 42000, 100),  (156150, 63000,  150),   (52050, 21000, 50),    (26025, 10500, 25), 
        (42000,  29700, 100),  (63000,  44550,  150),   (21000, 14850, 50),    (10500, 7425,  25), 
    ]
    
    drawing_map_ml = [
        ("A0", "1:100"), ("A0", "1:150"), ("A0", "1:50"),  ("A0", "1:25"),
        ("A0+1/8", "1:100"), ("A0+1/8", "1:150"), ("A0+1/8", "1:50"),  ("A0+1/8", "1:25"),
        ("A0+1/4", "1:100"), ("A0+1/4", "1:150"), ("A0+1/4", "1:50"),  ("A0+1/4", "1:25"),
        ("A1", "1:100"), ("A1", "1:150"), ("A1", "1:50"), ("A1", "1:25"),
        ("A1+1/4", "1:100"), ("A1+1/4", "1:150"), ("A1+1/4", "1:50"),  ("A1+1/4", "1:25"),
        ("A1+1/2", "1:100"), ("A1+1/2", "1:150"), ("A1+1/2", "1:50"),  ("A1+1/2", "1:25"),
        ("A1+3/4", "1:100"), ("A1+3/4", "1:150"), ("A1+3/4", "1:50"),  ("A1+3/4", "1:25"),
        ("A2", "1:100"), ("A2", "1:150"), ("A2", "1:50"),  ("A2", "1:25"),
        ("A2+1/4", "1:100"), ("A2+1/4", "1:150"), ("A2+1/4", "1:50"),  ("A2+1/4", "1:25"),
        ("A2+1/2", "1:100"), ("A2+1/2", "1:150"), ("A2+1/2", "1:50"),  ("A2+1/2", "1:25"),
        ("A2+3/4", "1:100"), ("A2+3/4", "1:150"), ("A2+3/4", "1:50"),  ("A2+3/4", "1:25"),
        ("A3", "1:100"), ("A3", "1:150"), ("A3", "1:50"),  ("A3", "1:25")
    ]

    drawing_map = [
        "ISO_A0_(841.00_x_1189.00_MM)", "ISO_A0_(841.00_x_1189.00_MM)", "ISO_A0_(841.00_x_1189.00_MM)", "ISO_A0_(841.00_x_1189.00_MM)",
        "UserDefinedMetric (1338.00 x 841.00毫米)", "UserDefinedMetric (1338.00 x 841.00毫米)", "UserDefinedMetric (1338.00 x 841.00毫米)", "UserDefinedMetric (1338.00 x 841.00毫米)",
        "UserDefinedMetric (1486.00 x 841.00毫米)", "UserDefinedMetric (1486.00 x 841.00毫米)", "UserDefinedMetric (1486.00 x 841.00毫米)", "UserDefinedMetric (1486.00 x 841.00毫米)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "UserDefinedMetric (1051.00 x 594.00毫米)", "UserDefinedMetric (1051.00 x 594.00毫米)", "UserDefinedMetric (1051.00 x 594.00毫米)", "UserDefinedMetric (1051.00 x 594.00毫米)",
        "UserDefinedMetric (1261.00 x 594.00毫米)", "UserDefinedMetric (1261.00 x 594.00毫米)", "UserDefinedMetric (1261.00 x 594.00毫米)", "UserDefinedMetric (1261.00 x 594.00毫米)",
        "UserDefinedMetric (1471.00 x 594.00毫米)", "UserDefinedMetric (1471.00 x 594.00毫米)", "UserDefinedMetric (1471.00 x 594.00毫米)", "UserDefinedMetric (1471.00 x 594.00毫米)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "UserDefinedMetric (743.00 x 420.00毫米)", "UserDefinedMetric (743.00 x 420.00毫米)", "UserDefinedMetric (743.00 x 420.00毫米)", "UserDefinedMetric (743.00 x 420.00毫米)",
        "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)",
        "UserDefinedMetric (1041.00 x 420.00毫米)", "UserDefinedMetric (1041.00 x 420.00毫米)", "UserDefinedMetric (1041.00 x 420.00毫米)", "UserDefinedMetric (1041.00 x 420.00毫米)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)"
    ]

    # =========================================================================
    # 第二阶段：双轨匹配 (优先命中 -> 否则兜底)
    # =========================================================================
    draw_factors = [1.0, 0.5, 0.25, 1.5]
    multipliers = [1.0, 1.1, 1.2]
    
    # 追踪器 1: 完美匹配 (Difference <= dynamic_tol)
    strict_best_index = None
    strict_min_diff = float('inf')

    # 追踪器 2: 兜底匹配 (Global Minimum Difference)
    approx_best_index = None
    approx_min_diff = float('inf')
    
    for i, (std_len, std_wid, scale_val) in enumerate(LB_dayingkuang):
        for df in draw_factors:
            for mult in multipliers:
                
                # --- 计算两种模式的误差 ---
                # A. 模型空间模式
                tgt_len_m = std_len * df * mult
                tgt_wid_m = std_wid * df * mult
                diff_m = abs(length - tgt_len_m) + abs(width - tgt_wid_m)
                
                # B. 布局空间模式
                tgt_len_l = (std_len / scale_val) * df * mult
                tgt_wid_l = (std_wid / scale_val) * df * mult
                diff_l = abs(length - tgt_len_l) + abs(width - tgt_wid_l)
                
                # 取两者中较优的那个误差
                current_diff = min(diff_m, diff_l)
                
                # --- 逻辑 A: 记录全局最小误差 (用于兜底) ---
                if current_diff < approx_min_diff:
                    approx_min_diff = current_diff
                    approx_best_index = i
                
                # --- 逻辑 B: 记录严格匹配 (优先选择) ---
                # 判断标准：长宽误差都在容差内 (这里用总误差简化判断，或者沿用长宽独立判断均可)
                # 为了简化且有效，这里使用总误差 <= 2 * dynamic_tol 近似判断，
                # 或者严格判断: (dL <= tol and dW <= tol)。这里沿用之前的严格逻辑。
                
                # 重新计算单边误差用于严格判定
                if diff_m <= diff_l:
                     d_len = abs(length - tgt_len_m)
                     d_wid = abs(width - tgt_wid_m)
                else:
                     d_len = abs(length - tgt_len_l)
                     d_wid = abs(width - tgt_wid_l)
                
                if d_len <= dynamic_tol and d_wid <= dynamic_tol:
                    if current_diff < strict_min_diff:
                        strict_min_diff = current_diff
                        strict_best_index = i

    # =========================================================================
    # 第三阶段：决策输出
    # =========================================================================
    
    # 1. 优先返回严格匹配的结果
    if strict_best_index is not None:
        final_index = strict_best_index
    
    # 2. 如果没有严格匹配，返回兜底结果 (将就模式)
    elif approx_best_index is not None:
        final_index = approx_best_index
        # sys_logger.info(f"⚠️ [将就匹配] 未找到标准框，已匹配最近似标准 (误差 {approx_min_diff:.1f})")
    
    # 3. 理论上不应该发生 (除非列表为空)，防止 crash
    else:
        return 0

    res_name = drawing_map[final_index]
    res_ratio = drawing_map_ml[final_index][1]
    res_code = drawing_map_ml[final_index][0]
    
    return (res_name, res_ratio, res_code, orientation_flag)




#&&% 打印数据分析

def get_cad_app():
    """连接 CAD"""
    try:
        acad = win32com.client.GetActiveObject("AutoCAD.Application")
        return acad
    except Exception:
        print("请先打开 AutoCAD。")
        return None

def get_dimensions(ent):
    """获取实体长宽（长边在前）"""
    try:
        min_pt, max_pt = ent.GetBoundingBox()
        width = abs(max_pt[0] - min_pt[0])
        height = abs(max_pt[1] - min_pt[1])
        
        # 保持长边在前，短边在后
        length = max(width, height)
        wid = min(width, height)
        
        # 处理浮点数精度（可选：如果全是整数图框，可以直接取整）
        return length, wid
    except Exception:
        return 0, 0

def sort_coms_by_llcorner(com_list, cha_Y=2000):
    """
    按 BoundingBox 左下角坐标排序：
      · 先按 y 降序（越大越靠前，即自上而下）
      · 同一行(Δy<cha_Y)内按 x 升序（自左向右）
    """
    wrapped = []
    for ent in com_list:
        try:
            p1, _ = ent.GetBoundingBox()      # p1 已是左下
            x_ll, y_ll = p1[0], p1[1]
        except Exception:
            x_ll = y_ll = float('-inf')       # 取不到的一律放最后
        wrapped.append((ent, x_ll, y_ll))     # (实体, x, y)

    # 先按 y 降序
    wrapped.sort(key=lambda t: -t[2])

    i = 0
    while i < len(wrapped): # 注意：这里稍微改了一点点逻辑以确保处理最后一行
        j = i + 1
        # 寻找当前行的结束位置
        while j < len(wrapped) and abs(wrapped[i][2] - wrapped[j][2]) < cha_Y:
            j += 1
        # 行内再按 x 升序
        if j - i > 1:
            wrapped[i:j] = sorted(wrapped[i:j], key=lambda t: t[1])
        i = j

    return [ent for ent, _, _ in wrapped]

def main():
    acad = get_cad_app()
    if not acad: return
    doc = acad.ActiveDocument
    
    # 1. 手动选择
    print(">>> 请切换到 CAD 窗口，选择那 48 个多段线，然后按回车...")
    ss_name = "PythonFixScaleSelect"
    try:
        doc.SelectionSets.Item(ss_name).Delete()
    except Exception:
        pass
        
    ss = doc.SelectionSets.Add(ss_name)
    try:
        ss.SelectOnScreen()
    except Exception:
        print("选择已取消。")
        return
        
    if ss.Count == 0:
        print("未选择任何对象。")
        return

    # 提取对象到列表
    raw_objs = [ss.Item(i) for i in range(ss.Count)]
    
    # 2. 调用你的排序函数
    # 注意：cha_Y=1000 是你之前的容差，如果行间距较大，可以用默认2000
    sorted_objs = sort_coms_by_llcorner(raw_objs, cha_Y=1000)
    
    # 3. 构造数据并分配固定比例
    # 固定比例序列
    fixed_scales = [100, 150, 50, 25]
    
    output_lines = []
    current_line_str = "    "
    
    sys_logger.info(f"正在处理 {len(sorted_objs)} 个对象...")
    
    for idx, obj in enumerate(sorted_objs):
        # 获取尺寸
        l, w = get_dimensions(obj)
        
        # 格式化尺寸数值（去除无效的小数点）
        l_str = f"{l:.2f}".rstrip('0').rstrip('.')
        w_str = f"{w:.2f}".rstrip('0').rstrip('.')
        
        # 分配比例：使用模运算循环取值 (idx % 4)
        # 0->100, 1->150, 2->50, 3->25, 4->100...
        scale_val = fixed_scales[idx % 4]
        
        # 拼接到当前行字符串
        current_line_str += f"({l_str}, {w_str}, {scale_val}), "
        
        # 每4个换行，或者到了最后一个元素换行
        if (idx + 1) % 4 == 0 or (idx + 1) == len(sorted_objs):
            output_lines.append(current_line_str)
            current_line_str = "    " # 重置并缩进

    # 4. 打印结果
    print("\n" + "="*30 + " 结果 " + "="*30)
    print("LB_dayingkuang = [")
    for line in output_lines:
        print(line)
    print("]")
    print("="*66)
    
    ss.Delete()

def generate_relation_list(data_list):
    result_list = []
    
    for i, current in enumerate(data_list):
        best_match = None
        min_metric = float('inf') # 最小评价值
        
        for j, candidate in enumerate(data_list):
            if i == j: continue # 跳过自己
            
            # 计算 X 差值 和 Y 差值
            dx = abs(current[0] - candidate[0])
            dy = abs(current[1] - candidate[1])
            
            # 核心规则：x或y方向的最后选最小的那个
            current_metric = min(dx, dy)
            
            # 寻找最小差异
            if current_metric < min_metric:
                min_metric = current_metric
                best_match = candidate
        
        # 按照要求的形式生成字符串： "旧元素 : 最近的那个"
        formatted_str = f"{current} : {best_match}"
        result_list.append(formatted_str)
        
    return result_list


#&&&% 选择标准打印区域


def check_strict_standard_size(comobj, tol=10):
    """
    【函数编号】: MAP-CHECK-SIZE-004
    【功能】: 严格检查对象外包盒是否符合标准打印框尺寸。
             支持动态容差：虽然输入参数 tol 固定，但内部判定时会根据倍率缩放容差。
             即：当 scale=0.01 时，实际使用的容差是 tol * 0.01 = 0.1。
    
    【参数】:
        tol: 基准容差 (默认10)，对应 scale=1.0 时的允许误差。
    """
    
    # ————————————— 1. 基础数据定义 —————————————
    LB_dayingkuang = [
        (118900, 84100, 100),  (178350, 126150, 150),   (59450, 42050, 50),     (29725, 21025, 25), 
        (133800, 84100, 100),  (200700, 126150, 150),   (66900, 42050, 50),     (33450, 21025, 25), 
        (148600, 84100, 100),  (222900, 126150, 150),   (74300, 42050, 50),     (37150, 21025, 25), 
        (84100,  59400, 100),  (126150, 89100,  150),   (42050, 29700, 50),     (21025, 14850, 25), 
        (105100, 59400, 100),  (157650, 89100,  150),   (52550, 29700, 50),     (26275, 14850, 25), 
        (126100, 59400, 100),  (189150, 89100,  150),   (63050, 29700, 50),     (31525, 14850, 25), 
        (147100, 59400, 100),  (220650, 89100,  150),   (73550, 29700, 50),     (36775, 14850, 25), 
        (59400,  42000, 100),  (89100,  63000,  150),   (29700, 21000, 50),     (14850, 10500, 25), 
        (74300,  42000, 100),  (111450, 63000,  150),   (37150, 21000, 50),     (18575, 10500, 25), 
        (89100,  42000, 100),  (133650, 63000,  150),   (44550, 21000, 50),     (22275, 10500, 25), 
        (104100, 42000, 100),  (156150, 63000,  150),   (52050, 21000, 50),     (26025, 10500, 25), 
        (42000,  29700, 100),  (63000,  44550,  150),   (21000, 14850, 50),     (10500, 7425,  25), 
    ]

    drawing_map_ml = [
        ("A0", "1:100"), ("A0", "1:150"), ("A0", "1:50"),  ("A0", "1:25"),
        ("A0+1/8", "1:100"), ("A0+1/8", "1:150"), ("A0+1/8", "1:50"),  ("A0+1/8", "1:25"),
        ("A0+1/4", "1:100"), ("A0+1/4", "1:150"), ("A0+1/4", "1:50"),  ("A0+1/4", "1:25"),
        ("A1", "1:100"), ("A1", "1:150"), ("A1", "1:50"), ("A1", "1:25"),
        ("A1+1/4", "1:100"), ("A1+1/4", "1:150"), ("A1+1/4", "1:50"),  ("A1+1/4", "1:25"),
        ("A1+1/2", "1:100"), ("A1+1/2", "1:150"), ("A1+1/2", "1:50"),  ("A1+1/2", "1:25"),
        ("A1+3/4", "1:100"), ("A1+3/4", "1:150"), ("A1+3/4", "1:50"),  ("A1+3/4", "1:25"),
        ("A2", "1:100"), ("A2", "1:150"), ("A2", "1:50"),  ("A2", "1:25"),
        ("A2+1/4", "1:100"), ("A2+1/4", "1:150"), ("A2+1/4", "1:50"),  ("A2+1/4", "1:25"),
        ("A2+1/2", "1:100"), ("A2+1/2", "1:150"), ("A2+1/2", "1:50"),  ("A2+1/2", "1:25"),
        ("A2+3/4", "1:100"), ("A2+3/4", "1:150"), ("A2+3/4", "1:50"),  ("A2+3/4", "1:25"),
        ("A3", "1:100"), ("A3", "1:150"), ("A3", "1:50"),  ("A3", "1:25")
    ]

    drawing_map = [
        "ISO_A0_(841.00_x_1189.00_MM)", "ISO_A0_(841.00_x_1189.00_MM)", "ISO_A0_(841.00_x_1189.00_MM)", "ISO_A0_(841.00_x_1189.00_MM)",
        "UserDefinedMetric (1338.00 x 841.00毫米)", "UserDefinedMetric (1338.00 x 841.00毫米)",
        "UserDefinedMetric (1338.00 x 841.00毫米)", "UserDefinedMetric (1338.00 x 841.00毫米)",
        "UserDefinedMetric (1486.00 x 841.00毫米)", "UserDefinedMetric (1486.00 x 841.00毫米)",
        "UserDefinedMetric (1486.00 x 841.00毫米)", "UserDefinedMetric (1486.00 x 841.00毫米)",
        "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)", "ISO_A1_(841.00_x_594.00_MM)",
        "UserDefinedMetric (1051.00 x 594.00毫米)", "UserDefinedMetric (1051.00 x 594.00毫米)",
        "UserDefinedMetric (1051.00 x 594.00毫米)", "UserDefinedMetric (1051.00 x 594.00毫米)",
        "UserDefinedMetric (1261.00 x 594.00毫米)", "UserDefinedMetric (1261.00 x 594.00毫米)",
        "UserDefinedMetric (1261.00 x 594.00毫米)", "UserDefinedMetric (1261.00 x 594.00毫米)",
        "UserDefinedMetric (1471.00 x 594.00毫米)", "UserDefinedMetric (1471.00 x 594.00毫米)",
        "UserDefinedMetric (1471.00 x 594.00毫米)", "UserDefinedMetric (1471.00 x 594.00毫米)",
        "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)", "ISO_A2_(594.00_x_420.00_MM)",
        "UserDefinedMetric (743.00 x 420.00毫米)", "UserDefinedMetric (743.00 x 420.00毫米)",
        "UserDefinedMetric (743.00 x 420.00毫米)", "UserDefinedMetric (743.00 x 420.00毫米)",
        "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)",
        "UserDefinedMetric (891.00 x 420.00毫米)", "UserDefinedMetric (891.00 x 420.00毫米)",
        "UserDefinedMetric (1041.00 x 420.00毫米)", "UserDefinedMetric (1041.00 x 420.00毫米)",
        "UserDefinedMetric (1041.00 x 420.00毫米)", "UserDefinedMetric (1041.00 x 420.00毫米)",
        "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)", "ISO_A3_(420.00_x_297.00_MM)"
    ]

    # —————————— 2. 获取对象外包盒信息 ——————————
    try:
        PL_min = find_min_point(comobj)
        PL_max = find_max_point(comobj)
        # 依赖外部函数 define_rectangle_by_diagonal
        _, length, width = define_rectangle_by_diagonal(PL_min, PL_max)
        
        dx = abs(PL_max[0] - PL_min[0])
        dy = abs(PL_max[1] - PL_min[1])
        # 1表示竖向，0表示横向
        orientation_flag = 1 if dy > dx else 0
    except Exception as e:
        sys_logger.info(f"获取对象几何信息失败: {e}")
        return 0

    # —————————— 3. 严格匹配逻辑 ——————————
    allowed_scales = [1.0, 1.1, 1.2, 0.01]

    for i, (std_len, std_wid, _) in enumerate(LB_dayingkuang):
        for scale in allowed_scales:
            # 计算目标尺寸
            target_len = std_len * scale
            target_wid = std_wid * scale
            
            # 🔥 核心修正：使用传入的 tol 乘以 scale 得到当前容差
            # tol 是基准容差(针对1.0倍率)。
            # 当 scale=0.01 时，current_tol 变成 0.1 (假设 tol=10)
            current_tol = tol * scale
            
            # 判断是否命中
            if (abs(length - target_len) <= current_tol) and (abs(width - target_wid) <= current_tol):
                
                # --- 命中后的附加视觉处理 ---
                try:
                    if scale == 1.2:
                        comobj.Color = 5  # 蓝色 (×1.2)
                except Exception:
                    pass

                # --- 构造返回值 ---
                scale_str = drawing_map_ml[i][1]
                
                result = (
                    drawing_map[i],          # 纸张规范名
                    scale_str,               # 比例
                    drawing_map_ml[i][0],    # 图号
                    orientation_flag         # 竖向标志
                )
                return result

    # —————————— 4. 未命中 ——————————
    return 0


#&&%新版20260111
@debuggable
def check_strict_standard_size(comobj, tol=10):
    """
    【修正版 V5.0】
    逻辑：
    1. 根据长边判定环境：长边 < 2140.8 (1784*1.2) 则 tol_base=0.1，否则 tol_base=10。
    2. 支持比例：标准(1:100系)、缩小(1:1系) 及其 1.0, 1.1, 1.2 倍率。
    """
    # ————————————— 1. 基础数据定义 (1:100比例下的基准尺寸) —————————————
    # LB_dayingkuang 存储 (长, 宽)
    LB_dayingkuang = [
        (118900, 84100), (178350, 126150), (59450, 42050), (29725, 21025), # A0, A0+1/8, A0+1/4, A1
        (133800, 84100), (200700, 126150), (66900, 42050), (33450, 21025), 
        (148600, 84100), (222900, 126150), (74300, 42050), (37150, 21025), 
        (84100,  59400), (126150, 89100),  (42050, 29700), (21025, 14850), 
        (105100, 59400), (157650, 89100),  (52550, 29700), (26275, 14850), 
        (126100, 59400), (189150, 89100),  (63050, 29700), (31525, 14850), 
        (147100, 59400), (220650, 89100),  (73550, 29700), (36775, 14850), 
        (59400,  42000), (89100,  63000),  (29700, 21000), (14850, 10500), # A2, A2+1/4...
        (74300,  42000), (111450, 63000),  (37150, 21000), (18575, 10500), 
        (89100,  42000), (133650, 63000),  (44550, 21000), (22275, 10500), 
        (104100, 42000), (156150, 63000),  (52050, 21000), (26025, 10500), 
        (42000,  29700), (63000,  44550),  (21000, 14850), (10500, 7425),  # A3...
    ]

    # 对应的显示标签 (对应 1:100, 1:150, 1:50, 1:25)
    # 当检测到 1:1 环境时，标签会映射为 1:1, 1:1.5, 1:0.5, 1:0.25
    specs = [
        ("A0", "1:100"), ("A0", "1:150"), ("A0", "1:50"), ("A0", "1:25"),
        ("A0+1/8", "1:100"), ("A0+1/8", "1:150"), ("A0+1/8", "1:50"), ("A0+1/8", "1:25"),
        ("A0+1/4", "1:100"), ("A0+1/4", "1:150"), ("A0+1/4", "1:50"), ("A0+1/4", "1:25"),
        ("A1", "1:100"), ("A1", "1:150"), ("A1", "1:50"), ("A1", "1:25"),
        ("A1+1/4", "1:100"), ("A1+1/4", "1:150"), ("A1+1/4", "1:50"), ("A1+1/4", "1:25"),
        ("A1+1/2", "1:100"), ("A1+1/2", "1:150"), ("A1+1/2", "1:50"), ("A1+1/2", "1:25"),
        ("A1+3/4", "1:100"), ("A1+3/4", "1:150"), ("A1+3/4", "1:50"), ("A1+3/4", "1:25"),
        ("A2", "1:100"), ("A2", "1:150"), ("A2", "1:50"), ("A2", "1:25"),
        ("A2+1/4", "1:100"), ("A2+1/4", "1:150"), ("A2+1/4", "1:50"), ("A2+1/4", "1:25"),
        ("A2+1/2", "1:100"), ("A2+1/2", "1:150"), ("A2+1/2", "1:50"), ("A2+1/2", "1:25"),
        ("A2+3/4", "1:100"), ("A2+3/4", "1:150"), ("A2+3/4", "1:50"), ("A2+3/4", "1:25"),
        ("A3", "1:100"), ("A3", "1:150"), ("A3", "1:50"), ("A3", "1:25")
    ]

    # —————————— 2. 获取几何信息 ——————————
    try:
        min_p, max_p = comobj.GetBoundingBox()
        dx = abs(max_p[0] - min_p[0])
        dy = abs(max_p[1] - min_p[1])
        obj_L = max(dx, dy)
        obj_W = min(dx, dy)
        orientation = 1 if dy > dx else 0
    except: return 0

    # —————————— 3. 动态环境判定 (长边阈值) ——————————
    # 判定是否属于 1:1 系列环境 (长边 < 1784 * 1.2)
    is_mini_scale = True if obj_L < 2140.8 else False
    
    # 根据环境设定基准容差
    active_tol = 0.1 if is_mini_scale else 10.0
    
    # 设定检测的倍率 (1.0, 1.1, 1.2)
    growth_factors = [1.0, 1.1, 1.2]
    
    # 设定环境缩放系数 (1:100环境对应1.0, 1:1环境对应0.01)
    env_scale = 0.01 if is_mini_scale else 1.0

    # —————————— 4. 匹配循环 ——————————
    for i, (std_L, std_W) in enumerate(LB_dayingkuang):
        for factor in growth_factors:
            # 最终目标尺寸 = 基准尺寸 * 环境缩放 * 扩展倍率
            target_L = std_L * env_scale * factor
            target_W = std_W * env_scale * factor
            
            # 容差也随 factor 稍微波动 (确保 1.2 倍的框也有足够的容差空间)
            current_tol = active_tol * factor

            if abs(obj_L - target_L) <= current_tol and abs(obj_W - target_W) <= current_tol:
                # 命中！处理标签映射
                label_size = specs[i][0]
                raw_scale_label = specs[i][1]
                
                # 如果是 1:1 环境，修正比例标签 (例如 1:100 -> 1:1)
                if is_mini_scale:
                    scale_val = float(raw_scale_label.split(":")[1]) / 100.0
                    final_scale_label = f"1:{scale_val:g}"
                else:
                    final_scale_label = raw_scale_label
                
                # 如果倍率是 1.2，变蓝标记
                if abs(factor - 1.2) < 0.01:
                    try: comobj.Color = 5
                    except: pass
                
                return (
                    f"Paper_{label_size}", # 模拟 paper 规范名
                    final_scale_label,     # 比例 (1:1 或 1:100等)
                    label_size,            # 图号 (A0等)
                    orientation            # 旋转
                )

    return 0


#&&% 多段线排序
def polyline_sort(polyline_list):
    """对com多段线按照特定规则进行排序"""

    # 存储多段线及其最左下角点
    polylines_with_min_points = [(pl, find_min_point(pl)) for pl in polyline_list]

    # 先按照y值降序排序
    polylines_with_min_points.sort(key=lambda item: -item[1][1])

    # 再对y值差距在1000以内的多段线按照x值升序排序
    i = 0
    while i < len(polylines_with_min_points) - 1:
        j = i + 1
        # 查找所有y值差距在1000以内的多段线
        while j < len(polylines_with_min_points) and abs(polylines_with_min_points[i][1][1] - polylines_with_min_points[j][1][1]) < 1000:
            j += 1
        
        # 如果找到了y值相近的多段线，根据x值进行排序
        if j - i > 1:
            polylines_with_min_points[i:j] = sorted(polylines_with_min_points[i:j], key=lambda item: item[1][0])
        
        i = j

    # 只返回多段线对象
    return [item[0] for item in polylines_with_min_points]




#&&&%  *** 3 将PLcom线列表的坐标信息存储


#&&% 多段线转坐标
def plcom_to_coor(plines):
    """
    接受多根轻量级多段线或常规多段线的 COM 对象列表，返回它们的坐标列表及闭合状态。

    :param plines: 可迭代的一组 LWPOLYLINE 或 POLYLINE COM 对象
    :return: 列表，每个元素为 (pts, closed_flag)：
             - pts: 顶点列表 [(x0, y0), (x1, y1), …]
             - closed_flag: 1 表示闭合，0 表示未闭合

    兼容两种情况：
      1. LWPOLYLINE.Coordinates → 偶数长度，例如 [x0,y0, x1,y1, x2,y2, …]
      2. POLYLINE.Coordinates  → 3 的倍数长度，例如 [x0,y0,z0, x1,y1,z1, …]

    如果既不是偶数也不是 3 的倍数，将跳过该条多段线并打印 WARN 提示。
    """
    plines = ensure_list(plines)
    all_info = []

    for pl in plines:
        raw = list(pl.Coordinates)  # 可能是偶数长度 (LWPOLYLINE) 或 3 的倍数长度 (POLYLINE)

        pts = []
        # —— 情况 A：如果长度能被 3 整除，认为是常规 POLYLINE → 每 3 个数一组 (x,y,z)
        if len(raw) % 3 == 0 and len(raw) > 0:
            for i in range(0, len(raw), 3):
                x = raw[i]
                y = raw[i + 1]
                pts.append((x, y))

        # —— 情况 B：否则如果长度能被 2 整除，认为是轻量级 LWPOLYLINE → 每 2 个数一组 (x,y)
        elif len(raw) % 2 == 0 and len(raw) > 0:
            for i in range(0, len(raw), 2):
                x = raw[i]
                y = raw[i + 1]
                pts.append((x, y))

        else:
            # 既不是 2 的倍数也不是 3 的倍数：坐标数据异常，跳过这一条，打印 WARN
            handle = getattr(pl, "Handle", "<unknown>")
            sys_logger.info(f"[WARN] plcom_to_coor：跳过 Handle={handle} 的多段线，"
                  f"Coordinates 长度={len(raw)} 既非 2 的倍数也非 3 的倍数。")
            continue

        # 读取 Closed 属性，True 表示闭合
        closed_flag = 1 if getattr(pl, "Closed", False) else 0

        all_info.append((pts, closed_flag))

    return all_info


# 4 从坐标信息列表返回PLcom线列表

    
#&&% 坐标转多段线
def plcoor_to_com(coord_info, layer_name="测试辅助", width=0, color=256):
    """
    在当前 DWG 中根据坐标和封闭标志绘制多条轻量级多段线。

    :param coord_info: 列表，每个元素为 (pts, closed_flag)，
                       pts 为 [(x0,y0),…] 顶点列表，
                       closed_flag 为 1（闭合）或 0（不闭合）。
    :param layer_name: 目标图层名称（不存在则创建），默认 "测试辅助"
    :param width:      多段线宽度，默认 0
    :param color:      颜色索引，默认 256（BYLAYER）
    :return:           绘制的多段线对象列表
    """
    # 1) 连接 AutoCAD
    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    ms   = doc.ModelSpace

    # 2) 确保图层存在
    layers = doc.Layers
    try:
        lyr = layers.Item(layer_name)
    except Exception:
        lyr = layers.Add(layer_name)
    lyr.LayerOn = True

    created = []
    for pts, closed_flag in coord_info:
        # 将 pts 展平为 [x0,y0,x1,y1,…]
        raw = []
        for x, y in pts:
            raw.extend((x, y))
        # 转为 COM 数组
        arr = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8,
            raw
        )
        # 添加轻量级多段线
        lw = ms.AddLightWeightPolyline(arr)
        lw.Layer         = layer_name
        lw.ConstantWidth = width
        lw.color         = color
        lw.Closed        = bool(closed_flag)
        created.append(lw)

    # 可选：缩放到可见范围
    acad.ZoomExtents()
    sys_logger.info(f"[OK] 已绘制 {len(created)} 条轻量级多段线到图层 “{layer_name}”")
    return created






# 5 确定多段线打印框是否竖向

#&&% 判断竖向框
def panduan_shuxiangkuang(polyline):

    PL_min = find_min_point(polyline)
    
    PL_max = find_max_point(polyline)

    cha_x  = abs(PL_max[0] - PL_min[0])   

    cha_y  = abs(PL_max[1] - PL_min[1])

    if  cha_y > cha_x:

        return True

    else :

        return False


# 6 统一为A2的图幅打印

#&&% 统一图幅
def tongyi_tufu(LB,TFname):


    """
    将打印线列表的每根线对应的图纸尺寸统一为一个TFname
        
    """

    LB_xin = []

    for ob in LB:

        LB_xin.append(TFname)#"ISO_A2_(594.00_x_420.00_MM)","ISO_A3_(420.00_x_297.00_MM)"

    return LB_xin



#  主函数
#  (3)
#  将正交六边形多段线分成两个矩形区域

#  该函数系列包括如下一些函数

"""
该函数将正交六边形多段线分成两个矩形区域，只针对标准六点六边的PL多边形
在处理六点六边的正交PL形多边形之前，使用simplify_polygon(polygon, tol=0.5)将伪边点清除
            
"""

# 消除多边形的伪边点

#&&% 简化多边形
def simplify_polygon(poly, tol=0.5):
    """
    简化多边形顶点列表：如果某顶点 P 与其前后两点共线（在容差 tol 范围内），则将其移除。
    首尾相连处理：第一个点的“前点”是最后一个点，最后一个点的“次点”是第一个点。
    
    参数:
        poly: [(x, y, z), …]  原始顶点列表（可能首尾重复或有多余顶点）
        tol:  共线判断的容差（对应叉积绝对值）
    
    返回:
        简化后的顶点列表（同样保留首尾是否闭合的形式，不会自动去重首尾）
    """
    # 先做一次“去掉首尾重复”以免无限循环
    if len(poly) > 1 and poly[0] == poly[-1]:
        poly = poly[:-1]

    def is_colinear(p_prev, p, p_next):
        # 只比较 x,y，计算 (p - p_prev) × (p_next - p) 的“叉积”
        x1, y1 = p[0] - p_prev[0], p[1] - p_prev[1]
        x2, y2 = p_next[0] - p[0],   p_next[1] - p[1]
        cross = x1 * y2 - y1 * x2
        return abs(cross) <= tol

    # 重复扫一遍能删就删，直到一轮下来没有删除
    changed = True
    while changed and len(poly) >= 3:
        changed = False
        n = len(poly)
        for i in range(n):
            prev_idx = (i - 1) % n
            next_idx = (i + 1) % n
            if is_colinear(poly[prev_idx], poly[i], poly[next_idx]):
                # 删除第 i 个点，重置一轮扫描
                del poly[i]
                changed = True
                break
    return poly



# 1. 标准化多边形顶点列表，去掉相邻和首尾重复点
#&&% 标准化多边形
def normalize_polygon(polygon):
    """
    标准化多边形顶点列表：  
      - 去掉任意相邻重复的点  
      - 如果首尾相同，则去掉末尾那个  
      
    参数:
        polygon: 原始顶点列表，每个点为 (x, y, z)
    返回:
        去重后的顶点列表
    """
    if not polygon:
        return []
    normalized = [polygon[0]]
    for pt in polygon[1:]:
        if pt != normalized[-1]:
            normalized.append(pt)
    # 首尾相同则删尾
    if len(normalized) > 1 and normalized[0] == normalized[-1]:
        normalized.pop()
    return normalized


# 2. 找到某顶点的前驱/后继（按循环多边形）
#&&% 获取相邻点
def get_adjacent_points(polygon, p):
    """
    在多边形 polygon 中返回顶点 p 的前后相邻点（支持循环）。
    会先调用 normalize_polygon 清理重复点。
    """
    poly = normalize_polygon(polygon)
    if not poly:
        raise ValueError("多边形为空")
    try:
        idx = poly.index(p)
    except ValueError:
        raise ValueError(f"点 {p} 不在多边形顶点列表中")
    prev_pt = poly[idx - 1] if idx > 0 else poly[-1]
    next_pt = poly[idx + 1] if idx < len(poly) - 1 else poly[0]
    return prev_pt, next_pt


# 3. 点是否在多边形内部（射线法，仅在 XY 平面判断）
#&&% 点在多边形内
def point_in_polygon(pt, polygon):
    """
    判断三维点 pt=(x,y,z) 在多边形 polygon 的 XY 投影内否。
    polygon 中点格式为 (x,y,z)，首尾可重复或不重复均可。
    """
    x, y, _ = pt
    poly2d = [(q[0], q[1]) for q in normalize_polygon(polygon)]
    inside = False
    n = len(poly2d)
    for i in range(n):
        x1, y1 = poly2d[i]
        x2, y2 = poly2d[(i + 1) % n]
        if (y1 > y) != (y2 > y):
            xint = x1 + (y - y1) * (x2 - x1) / (y2 - y1)
            if xint > x:
                inside = not inside
    return inside


# 4. 无穷直线 vs 线段在 XY 平面相交
#&&% 线段相交
def line_segment_intersection_2d(p, d, a, b, tol=1e-8):
    """
    计算射线 L(t)=p + t·d 与线段 AB 在 XY 平面上的交点。
    p,d,a,b 皆为 (x,y,z)，但只取 x,y 分量参与计算。
    返回 (xi, yi, t) 或 None。
    """
    px, py, _ = p
    dx, dy, _ = d
    ax, ay, _ = a
    bx, by, _ = b
    ux, uy = bx - ax, by - ay

    det = dx * (-uy) - dy * (-ux)
    if abs(det) < tol:
        return None

    rhsx, rhsy = ax - px, ay - py
    t_param = (rhsx * (-uy) - rhsy * (-ux)) / det
    u_param = (dx * rhsy - dy * rhsx) / det

    if -tol <= u_param <= 1 + tol:
        xi = px + t_param * dx
        yi = py + t_param * dy
        return xi, yi, t_param
    return None


# 5. 计算 p 和其相邻点中点 c，如果 c 内部则返回 c，否则沿 p->c 的射线找到第一个交点
#&&% 获取辅助点
def get_auxiliary_point(p, p_prev, p_next, polygon, tol=1e-8):
    """
    对于多边形顶点 p 及其前后相邻点 p_prev, p_next，
    返回一个位于多边形内部的辅助点 q：
      1. 先取 c = (p_prev + p_next)/2；若 c 在内部，则返回 c
      2. 否则沿射线 p->c 与多边形其它边求最靠近 p 的交点
    返回点格式为 (x,y,z)。
    """
    # 1) 中点 c
    cx = (p_prev[0] + p_next[0]) / 2
    cy = (p_prev[1] + p_next[1]) / 2
    cz = (p_prev[2] + p_next[2]) / 2
    c = (cx, cy, cz)

    if point_in_polygon(c, polygon):
        return c

    # 2) 构造方向 d = c - p 并归一化
    dx, dy = cx - p[0], cy - p[1]
    mag = math.hypot(dx, dy)
    if mag < tol:
        raise RuntimeError("辅助点方向向量过小")
    d = (dx / mag, dy / mag, 0.0)

    # 在每条不含 p 的边上求交
    poly = normalize_polygon(polygon)
    intersects = []
    for i in range(len(poly)):
        a = poly[i]
        b = poly[(i + 1) % len(poly)]
        if a == p or b == p:
            continue
        res = line_segment_intersection_2d(p, d, a, b, tol)
        if res:
            xi, yi, t_param = res
            if abs(t_param) > tol:
                # 插值出对应的 z
                zi = p[2] + t_param * d[2]
                intersects.append((xi, yi, zi, t_param))

    if not intersects:
        raise RuntimeError("未找到有效交点")
    # 取最小 |t| 的那一个
    intersects.sort(key=lambda it: abs(it[3]))
    xi, yi, zi, _ = intersects[0]
    return (xi, yi, zi)


# 6. 计算 p 点的“凹凸度量角”
#&&% 凹凸度量
def concavity_measure(p, p_prev, p_next, q):
    """
    给定 p, p_prev, p_next, q（均为 (x,y,z)），
    计算度量角：
      angle = 360 - larger_angle + smaller_angle  
    其中 smaller/larger 是 pq→p_prev 与 pq→p_next 的逆时针夹角。
    凸点约 ~90°，凹点约 ~270°。
    """
    def angle_of(vx, vy):
        a = math.degrees(math.atan2(vy, vx))
        return a if a >= 0 else a + 360

    # 构造 2D 向量
    vq = (q[0] - p[0], q[1] - p[1])
    v1 = (p_prev[0] - p[0], p_prev[1] - p[1])
    v2 = (p_next[0] - p[0], p_next[1] - p[1])

    a_q = angle_of(*vq)
    a1  = angle_of(*v1)
    a2  = angle_of(*v2)

    d1 = (a1 - a_q) % 360
    d2 = (a2 - a_q) % 360

    small, large = (d1, d2) if d1 < d2 else (d2, d1)
    return 360 - large + small


# 7. 直接给出 p 在多边形上的度量角
#&&% 凹凸角
def concavity_angle(p, polygon):
    """
    直接计算多边形 polygon 上顶点 p 的凹凸度量角。
    """
    p_prev, p_next = get_adjacent_points(polygon, p)
    q = get_auxiliary_point(p, p_prev, p_next, polygon)
    return concavity_measure(p, p_prev, p_next, q)





# 8.合理分割PL正交六边形

#&&% 水平分割六边形
def split_orthogonal_hexagon(polygon, tol=0.1):#水平分割
    """
    将正交六边形 polygon 按凹顶点所在水平线切成两个矩形。
    polygon: 6 个 (x,y,z) 顶点的列表，允许首尾重合或相邻重复，会自动规范化。
    """
    # 1. 规范化，去掉相邻重复和首尾同点
    poly = normalize_polygon(polygon)
    if len(poly) != 6:
        raise ValueError("必须传入6点正交多边形")
    # 2. 找出唯一的凹点 p
    concaves = [pt for pt in poly
                if abs(concavity_angle(pt, poly) - 270) < tol]
    if len(concaves) != 1:
        raise RuntimeError(f"没能唯一定位凹点，找到 {len(concaves)} 个")
    p = concaves[0]
    y0 = p[1]

    # 3. 只对真正“跨越” y=y0 的边求交
    intersections = []
    n = len(poly)
    for i in range(n):
        a = poly[i]
        b = poly[(i+1) % n]
        y1, y2 = a[1], b[1]
        # 仅当严格跨越才算：一端在上 (y>y0)，一端在下 (y<y0)
        if (y1 - y0) * (y2 - y0) < -tol**2:
            t = (y0 - y1) / (y2 - y1)
            xi = a[0] + t * (b[0] - a[0])
            zi = a[2] + t * (b[2] - a[2])
            intersections.append((xi, y0, zi, i))

    # 4. 应该只剩一个真的 crossing 交点
    if len(intersections) != 1:
        raise RuntimeError(f"没能唯一定位 q，找到 {len(intersections)} 个候选点")
    xi, yi, zi, edge_idx = intersections[0]
    q = (xi, yi, zi)

    # 5. 把 q 插回那条边之后
    newpoly = []
    for i in range(n):
        newpoly.append(poly[i])
        if i == edge_idx:
            newpoly.append(q)
    # now len(newpoly)==7

    # 6. 找 p, q 的索引
    i_p = newpoly.index(p)
    i_q = newpoly.index(q)

    # 7. 分割成两段多边形
    if i_q < i_p:
        rect1 = newpoly[i_q:i_p+1]
        rect2 = newpoly[i_p:] + newpoly[:i_q+1]
    else:
        rect1 = newpoly[i_p:i_q+1]
        rect2 = newpoly[i_q:] + newpoly[:i_p+1]

    # 8. 计算2D面积
    def area2d(pts):
        s = 0
        m = len(pts)
        for j in range(m):
            x1,y1,_ = pts[j]
            x2,y2,_ = pts[(j+1)%m]
            s += x1*y2 - x2*y1
        return abs(s)/2

    A1, A2 = area2d(rect1), area2d(rect2)
    # 面积小的放前面
    return (rect1, rect2) if A1 <= A2 else (rect2, rect1)


#&&% 竖向分割六边形
def split_orthogonal_hexagon_vertical(polygon, tol=0.1):#竖向分割
    """
    将正交六边形 polygon 按凹顶点所在竖线切成两个矩形。
    polygon: 6 个 (x,y,z) 顶点的列表，允许首尾重合或相邻重复，会自动规范化。
    tol: 用于识别凹点和跨越判断的容差。
    返回: (rect1, rect2)，面积小的放前面。
    """
    # 规范化：去掉相邻重复和首尾同点
    poly = normalize_polygon(polygon)
    if len(poly) != 6:
        raise ValueError("必须传入6点正交多边形")

    # 找唯一凹点
    concaves = [pt for pt in poly
                if abs(concavity_angle(pt, poly) - 270) < tol]
    if len(concaves) != 1:
        raise RuntimeError(f"没能唯一定位凹点，找到 {len(concaves)} 个")
    p = concaves[0]
    x0 = p[0]

    # 求竖线 x=x0 与真正跨越边的交点
    intersections = []
    n = len(poly)
    for i in range(n):
        a = poly[i]
        b = poly[(i+1)%n]
        x1, x2 = a[0], b[0]
        # 仅对严格跨越的边求交
        if (x1 - x0)*(x2 - x0) < -tol**2:
            # 插值比例
            t = (x0 - x1)/(x2 - x1)
            yi = a[1] + t*(b[1] - a[1])
            zi = a[2] + t*(b[2] - a[2])
            intersections.append((x0, yi, zi, i))

    if len(intersections) != 1:
        raise RuntimeError(f"没能唯一定位 q，找到 {len(intersections)} 个候选点")
    xi, yi, zi, edge_idx = intersections[0]
    q = (xi, yi, zi)

    # 把 q 插回那条边
    newpoly = []
    for i in range(n):
        newpoly.append(poly[i])
        if i == edge_idx:
            newpoly.append(q)
    # newpoly 长度应为7

    # 定位 p、q 索引
    i_p = newpoly.index(p)
    i_q = newpoly.index(q)

    # 分割两段
    if i_q < i_p:
        rect1 = newpoly[i_q:i_p+1]
        rect2 = newpoly[i_p:] + newpoly[:i_q+1]
    else:
        rect1 = newpoly[i_p:i_q+1]
        rect2 = newpoly[i_q:] + newpoly[:i_p+1]

    # 面积计算（2D）
    def area2d(pts):
        s = 0
        m = len(pts)
        for j in range(m):
            x1,y1,_ = pts[j]
            x2,y2,_ = pts[(j+1)%m]
            s += x1*y2 - x2*y1
        return abs(s)*0.5

    A1, A2 = area2d(rect1), area2d(rect2)
    return (rect1, rect2) if A1 <= A2 else (rect2, rect1)


# 合理分割PL正交六边形

#&&% 计算面积
def area_of(verts):
    """多边形面积计算（顶点首尾闭合或不闭合均可）"""
    s = 0
    n = len(verts)
    for i in range(n):
        x1, y1, *_ = verts[i]
        x2, y2, *_ = verts[(i+1) % n]
        s += x1 * y2 - x2 * y1
    return abs(s) * 0.5



#&&% 综合分割六边形
def split_hexagon_combined(polygon, tol=0.1, simplify_tol=0.5):# 合理分割PL正交六边形
    """
    合理分割一个正交（近似）六边形 PL：
      1) 如果传入的是 COM PL 对象，则先提取唯一顶点列表；
      2) 对顶点列表做简化（去除伪顶点/伪边）；
      3) 先做横向分割，再做竖向分割，比较两种分割最小矩形面积，
         将最小的那对矩形排在前面，返回四个矩形顶点列表：
         [min1, partner1, min2, partner2]
    参数:
      polygon: 要分割的多段线，既可以是 [(x,y,z),...] 顶点列表，也可以是 COM PL 对象
      tol: 分割时判断凹角的容差
      simplify_tol: 简化多边形时去除“伪顶点/伪边”的容差
    返回:
      四个矩形的顶点列表：最小矩形、其同组矩形、次小矩形、其同组矩形
    """
    # —— 1. 如果是 COM PL 对象，先提取顶点列表 —— 
    # （假定已有 get_unique_vertices_from_pl_com(com_pl) -> [(x,y,z),...]）
    if not isinstance(polygon, list):
        polygon = get_unique_vertices_from_pl_com(polygon)

    # —— 2. 简化顶点列表 —— 
    # （假定已有 simplify_polygon(verts, tol) -> 清理后的顶点列表）
    polygon = simplify_polygon(polygon, simplify_tol)

    # —— 3. 横向与竖向分割 —— 
    rects_h = split_orthogonal_hexagon(polygon, tol)
    rects_v = split_orthogonal_hexagon_vertical(polygon, tol)

    A_h1, A_h2 = rects_h
    A_v1, A_v2 = rects_v

    # 计算每对的“更小”矩形和“配对”矩形
    if area_of(A_h1) <= area_of(A_h2):
        min_h, partner_h = A_h1, A_h2
    else:
        min_h, partner_h = A_h2, A_h1

    if area_of(A_v1) <= area_of(A_v2):
        min_v, partner_v = A_v1, A_v2
    else:
        min_v, partner_v = A_v2, A_v1

    # —— 4. 根据最小面积值决定输出顺序 —— 
    if area_of(min_h) <= area_of(min_v):
        return [min_h, partner_h, min_v, partner_v]
    else:
        return [min_v, partner_v, min_h, partner_h]



#  主函数
#  (4)
#  获取多段线的上下左右边界的直线段，返回线段端点列表

#  该函数系列包括如下一些函数

#&&% 获取包围盒边
def get_bbox_edge_segments(pl, tol=0.5):
    """
    获取对象 pl 的包围盒四条边，分别作为独立的列表返回：
      top    = [(xmin, ymax, z), (xmax, ymax, z)]
      bottom = [(xmin, ymin, z), (xmax, ymin, z)]
      left   = [(xmin, ymin, z), (xmin, ymax, z)]
      right  = [(xmax, ymin, z), (xmax, ymax, z)]
    并打印调试信息。
    """
    # ----- 1. 调用 GetBoundingBox -----
    try:
        min_pt, max_pt = pl.GetBoundingBox()
    except Exception:
        mins = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (0.0, 0.0, 0.0))
        maxs = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (0.0, 0.0, 0.0))
        pl.GetBoundingBox(mins, maxs)
        min_pt = tuple(mins.value)
        max_pt = tuple(maxs.value)

    xmin, ymin, zmin = min_pt
    xmax, ymax, _    = max_pt

    sys_logger.info(f"▶ BoundingBox Min 点: {min_pt}")
    sys_logger.info(f"▶ BoundingBox Max 点: {max_pt}")

    # ----- 2. 构造四个顶点（顺时针） -----
    p1 = (xmin, ymin, zmin)
    p2 = (xmax, ymin, zmin)
    p3 = (xmax, ymax, zmin)
    p4 = (xmin, ymax, zmin)

    print("▶ 矩形四个顶点 (顺时针):")
    for i, pt in enumerate((p1, p2, p3, p4), 1):
        sys_logger.info(f"   {i}: {pt}")

    # ----- 3. 四条边，各自用列表表达 -----
    top    = [p4, p3]    # y = ymax
    bottom = [p1, p2]    # y = ymin
    left   = [p1, p4]    # x = xmin
    right  = [p2, p3]    # x = xmax

    return top, bottom, left, right



#  主函数
#  (5)
#&&%  获取多段线的内部的文字

#  该函数系列包括如下一些函数



#&&% 获取多段线内文字
def get_texts_in_polyline(com_pl, tol=0.5):
    """
    在多段线 com_pl 内部筛选文字，并返回文字对象列表和对应的文字内容列表。

    参数:
      com_pl:  COM 多段线对象
      tol:     点-in-多边形时的容差（目前未用到，可留作将来扩展）

    返回:
      inside:   落在 com_pl 内部的文字 COM 对象列表
      contents: 对应 inside 中每个对象的文字内容列表
    """
    # 1) 多段线转顶点列表（标准化后的三维点）
    poly = get_unique_vertices_from_pl_com(com_pl)

    # 2) 收集所有文字实体（天正＋原生 CAD）
    tzh_text, tzh_mtext, cad_text, cad_mtext = collect_all_texts()
    all_texts = tzh_text + tzh_mtext + cad_text + cad_mtext

    inside = []
    contents = []

    for txt in all_texts:
        # 用左下角点判断是否在多段线内
        min_pt, _ = txt.GetBoundingBox()
        if point_in_polygon(min_pt, poly):
            inside.append(txt)

            # 根据对象类型取出内容
            name = getattr(txt, "ObjectName", "") or getattr(txt, "EntityName", "")
            if name in ("AcDbText", "AcDbMText"):
                # AutoCAD 原生单/多行
                contents.append(txt.TextString)
            elif name == "TDbText":
                # 天正单行
                contents.append(txt.Text)
            elif name == "TDbMText":
                # 天正多行，需要炸开取内容
                contents.append( TDbMText_content(txt))
            else:
                # 其它（万一有），留空或用 repr
                contents.append("")

    sys_logger.info(f"总共找到 {len(inside)} 条落在多段线内部的文字。")
    return inside, contents




# 获取单独一行的天正多行文字内容

#&&% 获取天正多行文字

def TDbMText_content(comobj, separator="\n"):
    """
    【函数】获取天正多行文字内容（副本炸开版，支持换行识别）
    
    参数:
      comobj: 天正文字对象
      separator: 换行符，默认为 "\n"，也可以设为 "\\" 或其他
    
    逻辑:
      1. 复制并炸开对象。
      2. 按 Y(容差0.3)+X 排序。
      3. 遍历碎片，检测 Y 坐标变化插入换行符。
      4. 清理碎片并返回。
    """
    # 确保环境已连接
    li() 
    
    fragments = []
    
    try:
        # 1. 复制副本
        try:
            copy_ent = comobj.Copy()
        except Exception as e:
            sys_logger.info(f"[错误] 复制天正对象失败: {e}")
            return ""

        # 2. 炸开副本
        fragments = explode_single_object_marker(copy_ent)
        
        if not fragments:
            return ""

        # 3. 排序 (Sort)
        # 容差设为 0.3 (可根据图纸精度调整)
        TOLERANCE = 0.3
        
        def get_sort_info(ent):
            """辅助函数：获取排序用的 y_bin 和 x"""
            ins = get_attr(ent, "InsertionPoint")
            if not ins:
                try:
                    min_p, _ = ent.GetBoundingBox()
                    x, y = min_p[0], min_p[1]
                except:
                    return 0, 0
            else:
                x, y, z = ins
            
            # y_bin 用于判定“行”
            y_bin = round(y / TOLERANCE)
            return y_bin, x

        # Python 的 sort key
        def sort_key(ent):
            y_bin, x = get_sort_info(ent)
            # Y 越大(越靠上)排越前 -> 取负
            return (-y_bin, x)

        fragments.sort(key=sort_key)

        # 4. 提取内容并处理换行 (Extract with Line Breaks)
        final_string = ""
        last_y_bin = None
        
        for i, frag in enumerate(fragments):
            # 获取内容
            txt = get_attr(frag, "TextString")
            if not txt:
                continue
                
            # 获取当前行的标识
            current_y_bin, _ = get_sort_info(frag)
            
            # 判断是否换行
            if last_y_bin is not None:
                # 如果当前行 Y (分桶后) 不等于上一行 Y，说明换行了
                if current_y_bin != last_y_bin:
                    final_string += separator
            
            final_string += txt
            
            # 更新 last_y_bin
            last_y_bin = current_y_bin
        
        return final_string

    except Exception as e:
        sys_logger.info(f"[错误] 提取天正文字内容失败: {e}")
        return ""

    finally:
        # 5. 清理碎片
        if fragments:
            for frag in fragments:
                try:
                    frag.Delete()
                except:
                    pass

#  主函数
#  (6)
#  多段线上的均分插入


"""
该函数用于在dwg文件沿着PL线快速均衡放置树木等图块

            
"""
#&&% 实体均分点
def distribute_points_on_entity(entity, n, block, scale_factor, ys):

    def distance(p1, p2):
        return math.sqrt((p1[0] - p2[0])**2 + (p1[1] - p2[1])**2)

    model_space = doc.ModelSpace
    block_name = block.Name  # 获取块的名称
    
    # 如果实体是直线
    if entity.ObjectName == "AcDbLine":
        start_point = entity.StartPoint
        end_point = entity.EndPoint
        for i in range(n):
            x = start_point[0] + i * (end_point[0] - start_point[0]) / (n - 1)
            y = start_point[1] + i * (end_point[1] - start_point[1]) / (n - 1)
            inserted_block = model_space.InsertBlock(vtpnt(x, y, 0), block_name, scale_factor, scale_factor, scale_factor, 0)
            inserted_block.color = ys  # 设置颜色为红色

    # 如果实体是圆弧
    elif entity.ObjectName == "AcDbArc":
        start_angle = entity.StartAngle
        end_angle = entity.EndAngle
        center = entity.Center
        radius = entity.Radius
        for i in range(n):
            angle = start_angle + i * (end_angle - start_angle) / (n - 1)
            x = center[0] + radius * math.cos(angle)
            y = center[1] + radius * math.sin(angle)
            inserted_block = model_space.InsertBlock(vtpnt(x, y, 0), block_name, scale_factor, scale_factor, scale_factor, 0)
            inserted_block.color = ys  # 设置颜色为红色

    # 如果实体是多段线
    elif entity.ObjectName == "AcDbPolyline":
        coords = entity.Coordinates
        points = [(coords[i], coords[i+1]) for i in range(0, len(coords), 2)]
        total_length = sum(distance(points[i], points[i+1]) for i in range(len(points)-1))
        segment_length = total_length / n
        current_length = 0

        for i in range(n):
            accumulated_length = 0
            for j in range(len(points) - 1):
                segment = distance(points[j], points[j+1])
                if accumulated_length + segment > current_length:
                    ratio = (current_length - accumulated_length) / segment
                    x = points[j][0] + ratio * (points[j+1][0] - points[j][0])
                    y = points[j][1] + ratio * (points[j+1][1] - points[j][1])
                    inserted_block = model_space.InsertBlock(vtpnt(x, y, 0), block_name, scale_factor, scale_factor, scale_factor, 0)
                    inserted_block.color = ys  # 设置颜色为红色
                    break
                accumulated_length += segment
            current_length += segment_length
            




#  主函数
#  (7)
# 返回 pl1 中与 pl2 “共线且有重叠”的区段列表

#  该函数系列包括如下一些函数

# 1 判断一条直线是否完全在另一条直线上

#&&% 判断线段包含
def is_segment_contained(seg_a, seg_b, tol=1e-4):
    """
    判断 seg_a 是否完全位于 seg_b 上（包含端点）。
    
    参数
    ----
    seg_a, seg_b :  ( (x1,y1,z1), (x2,y2,z2) )   或   AcDbLine
        两条待比较的线段。先判断 seg_a 是否被 seg_b 覆盖；
        若想双向判断，调用两次并对调顺序即可。
    tol : float
        距离与投影误差容差，默认 1e‑4 (CAD 单位)。

    返回
    ----
    bool
        True  —— seg_a 整段落在 seg_b 上  
        False —— 否则
    """
    # -------- 把输入统一转成端点元组 ----------
    def get_endpoints(entity):
        if hasattr(entity, "StartPoint"):            # COM 线段
            return (tuple(entity.StartPoint), tuple(entity.EndPoint))
        else:                                        # 纯坐标二元组
            return (tuple(entity[0]), tuple(entity[1]))

    a1, a2 = get_endpoints(seg_a)
    b1, b2 = get_endpoints(seg_b)

    # -------- 基本几何工具 ----------
    def dist(p, q):
        return math.hypot(p[0]-q[0], p[1]-q[1])

    def dot(u, v):
        return u[0]*v[0] + u[1]*v[1]

    def colinear(p, q, r, tol):
        """三点是否近似共线（面积≈0）"""
        return abs( (q[0]-p[0])*(r[1]-p[1]) - (q[1]-p[1])*(r[0]-p[0]) ) <= tol

    # -------- 先检测共线性 ----------
    if not (colinear(b1, b2, a1, tol) and colinear(b1, b2, a2, tol)):
        return False          # seg_a 端点不与 seg_b 共线 → 不可能包含

    # -------- 再检测投影是否在 b1‑b2 区间内 ----------
    # 令 b1 为原点，b_dir 为方向向量
    b_dir = (b2[0]-b1[0], b2[1]-b1[1])
    b_len2 = dot(b_dir, b_dir)
    if b_len2 == 0:           # seg_b 长度为 0 → 无法包含他段
        return False

    def proj_param(p):
        # 标准化投影参数 t，若 0<=t<=1 则投影落在 seg_b 上
        return dot( (p[0]-b1[0], p[1]-b1[1]), b_dir ) / b_len2

    t_a1 = proj_param(a1)
    t_a2 = proj_param(a2)

    # 允许介于 0±tol 到 1±tol 之间
    inside_1 = -tol <= t_a1 <= 1+tol
    inside_2 = -tol <= t_a2 <= 1+tol

    return inside_1 and inside_2

#&&% 2 返回 PL线pl1 中与 pl2 “共线且有重叠”的区段列表

#&&% 公共线段
def common_segments_between_polylines(pl1, pl2, tol=0.5):
    """
    返回 pl1 中与 pl2 “共线且有重叠”的区段列表，每个区段用
      [(x1,y1,0.0),(x2,y2,0.0)] 表示。

    参数
    ----
    pl1, pl2 : AutoCAD AcDbPolyline COM 对象 (或伪造对象，只需有 .Coordinates / .Closed)
    tol      : 共线 & 距离容差 (同 CAD 单位)

    返回
    ----
    overlaps : list[ list[(x,y,z),(x,y,z)] ]
    """

    # ────────── 内部小工具 ──────────
    def coords_to_xy_pairs(coords):
        """把 (x1,y1,x2,y2,…) 转成 [(x,y),…]；自动丢弃尾部残值"""
        pairs = []
        for i in range(0, len(coords) - 1, 2):
            pairs.append((coords[i], coords[i + 1]))
        return pairs

    def build_segments(verts, closed=False):
        """由顶点顺序生成线段 [P,Q] 列表"""
        segs = []
        for i in range(len(verts) - 1):
            segs.append([verts[i], verts[i + 1]])
        if closed and len(verts) > 2:
            segs.append([verts[-1], verts[0]])
        return segs

    def dist(p, q):
        return math.hypot(p[0] - q[0], p[1] - q[1])

    def colinear(p, q, r):
        """三点共线判定 |叉积| < tol · max(边长)"""
        return abs((q[0] - p[0]) * (r[1] - p[1]) -
                   (q[1] - p[1]) * (r[0] - p[0])) <= tol * max(dist(p, q), dist(p, r), dist(q, r), 1)

    def project(p, axis):
        """根据 axis==0 用 x，否则 y"""
        return p[axis]

    def segment_overlap(seg_a, seg_b):
        """
        若共线且区间有重叠，返回实际重叠区间端点 (pa, pb)（二维点）。
        否则返回 None
        """
        (p1, p2), (q1, q2) = seg_a, seg_b
        # 共线检测
        if not (colinear(p1, p2, q1) and colinear(p1, p2, q2)):
            return None

        # 选投影轴
        axis = 0 if abs(p2[0] - p1[0]) >= abs(p2[1] - p1[1]) else 1
        a1, a2 = project(p1, axis), project(p2, axis)
        b1, b2 = project(q1, axis), project(q2, axis)
        # 使 a1 <= a2, b1 <= b2
        if a1 > a2:
            p1, p2 = p2, p1
            a1, a2 = a2, a1
        if b1 > b2:
            q1, q2 = q2, q1
            b1, b2 = b2, b1

        # 计算 1‑D 重叠区间
        left, right = max(a1, b1), min(a2, b2)
        if right - left <= tol:          # “长度”≈0 视为无重叠
            return None

        # 把投影点还原到 2D 端点 —— 在线段 p1‑p2 上按比例取值
        def interp(t):
            """t 为 0~1"""
            return (p1[0] + t * (p2[0] - p1[0]),
                    p1[1] + t * (p2[1] - p1[1]))

        len_p = a2 - a1 if a2 != a1 else 1e-9
        pa = interp((left - a1) / len_p)
        pb = interp((right - a1) / len_p)
        return pa, pb

    # ────────── 主流程 ──────────
    v1 = coords_to_xy_pairs(pl1.Coordinates)
    v2 = coords_to_xy_pairs(pl2.Coordinates)

    # 若为闭合 polyline，补一个首尾顶点
    if getattr(pl1, "Closed", False) and v1 and v1[0] != v1[-1]:
        v1.append(v1[0])
    if getattr(pl2, "Closed", False) and v2 and v2[0] != v2[-1]:
        v2.append(v2[0])

    segs1 = build_segments(v1, closed=False)
    segs2 = build_segments(v2, closed=False)

    overlaps = []

    for s1 in segs1:
        for s2 in segs2:
            ov = segment_overlap(s1, s2)
            if ov:
                pa, pb = ov
                overlaps.append([(pa[0], pa[1], 0.0), (pb[0], pb[1], 0.0)])
                break        # 一条 s1 找到重叠就够了，可跳出

    # --- 打印摘要 ---
    print("★ 与 pl2 重叠 (或被包含) 的 pl1 线段数：", len(overlaps))
    for idx, seg in enumerate(overlaps, 1):
        sys_logger.info(f"  {idx}. {seg[0]}  →  {seg[1]}")

    return overlaps




#  主函数
#  (8)
# 找到全部“两根多段线耦合成一个矩形”的多段线

#  该函数系列包括如下一些函数

"""
函数是用来分析主房间带卫生间这种情况的，因此对输入变量是有较严格假定的，并非针对任意情况

"""
# 1 判断矩形是否包含另一个矩形

#&&% 矩形包含判断
def is_rect_inside_rect(rect_outer, rect_inner, tol=1e-6):
    """
    判定 axis‑aligned 的矩形 rect_inner 是否被（含边界）完全包在 rect_outer 内。

    参数
    ----
    rect_outer : ((xmin, ymin), (xmax, ymax))
    rect_inner : ((xmin, ymin), (xmax, ymax))
        两个元组分别给出矩形左下、右上坐标（假定 Z 全 0）。
    tol        : float
        容差（允许轻微数值误差；AutoCAD double 转 python float 时推荐 1e‑6～1e‑4）。

    返回
    ----
    bool   ——  rect_inner ⊆ rect_outer ？
    """
    (ox0, oy0), (ox1, oy1) = rect_outer
    (ix0, iy0), (ix1, iy1) = rect_inner

    return (
        ix0 >= ox0 - tol and
        iy0 >= oy0 - tol and
        ix1 <= ox1 + tol and
        iy1 <= oy1 + tol
    )




# 2 判断两条正交多段线拼在一起后是否正好是一个矩形


#&&% 两多段线组矩形
def two_plines_making_rectangle(pl1, pl2, tol=0.5):#
    """
    判断两条正交多段线拼在一起后是否正好是一个矩形。
    假设：两 PLine 没有面积重叠，只可能共用完整边或边的一部分。
    """

    import math

    def same_point(a, b, tol):
        return abs(a[0] - b[0]) <= tol and abs(a[1] - b[1]) <= tol

    def pline_vertices(pl):
        # 从 pl.Coordinates 提取 (x,y) 顶点列表，自动闭合
        c = pl.Coordinates
        verts = []
        for i in range(0, len(c), 2):
            verts.append((c[i], c[i+1]))
        if not same_point(verts[0], verts[-1], tol):
            verts.append(verts[0])
        return verts

    def poly_area(verts):
        # 计算首尾闭合的多边形面积
        s = 0
        for i in range(len(verts)-1):
            x1,y1 = verts[i]
            x2,y2 = verts[i+1]
            s += x1*y2 - x2*y1
        return abs(s)*0.5

    def collect_segments(verts):
        # 从顶点列表生成线段列表 [((x1,y1),(x2,y2)), ...]
        segs = []
        for i in range(len(verts)-1):
            segs.append((verts[i], verts[i+1]))
        return segs

    def covers_edge(edge, segs):
        """
        判断给定的 bbox 边 edge=((x0,y0),(x1,y1)) 能否被 segs 中若干共线线段
        连续覆盖（无缝隙）。
        """
        (x0,y0),(x1,y1) = edge
        # 算出方向向量和长度
        dx, dy = x1-x0, y1-y0
        L = math.hypot(dx, dy)
        if L < tol:
            return False
        ux, uy = dx/L, dy/L  # 单位方向向量

        # 投影每条 seg 到 [0,L] 参数区间
        intervals = []
        for (ax,ay),(bx,by) in segs:
            # 判断端点是否在同一直线上
            # cross == 0
            cross = (ax-x0)*dy - (ay-y0)*dx
            if abs(cross) > tol*L:  
                continue
            # 计算两端投影参数
            t1 = (  (ax-x0)*ux + (ay-y0)*uy )
            t2 = (  (bx-x0)*ux + (by-y0)*uy )
            a, b = min(t1,t2), max(t1,t2)
            # 只保留与 [0,L] 有交集的部分
            if b < -tol or a > L+tol:
                continue
            intervals.append((max(0.0, a), min(L, b)))

        if not intervals:
            return False

        # 合并所有区间
        intervals.sort(key=lambda iv: iv[0])
        cur_start, cur_end = intervals[0]
        for a,b in intervals[1:]:
            if a > cur_end + tol:
                # 出现间隙
                return False
            cur_end = max(cur_end, b)
        # 最后检测是否覆盖了整个 [0, L]
        return (cur_start <= tol) and (cur_end >= L - tol)

    # 1) 提取顶点及面积
    v1 = pline_vertices(pl1)
    v2 = pline_vertices(pl2)
    A1 = poly_area(v1)
    A2 = poly_area(v2)

    # 2) 计算公共外包矩形
    xs = [p[0] for p in v1[:-1]] + [p[0] for p in v2[:-1]]
    ys = [p[1] for p in v1[:-1]] + [p[1] for p in v2[:-1]]
    xmin, xmax = min(xs), max(xs)
    ymin, ymax = min(ys), max(ys)
    A_bbox = (xmax-xmin)*(ymax-ymin)

    # 3) 面积校验
    if abs((A1 + A2) - A_bbox) > tol:
        return False

    # 4) 检查四条边被覆盖
    bbox_edges = [
        ((xmin,ymin),(xmax,ymin)),
        ((xmax,ymin),(xmax,ymax)),
        ((xmax,ymax),(xmin,ymax)),
        ((xmin,ymax),(xmin,ymin))
    ]
    segs = collect_segments(v1) + collect_segments(v2)

    hits = 0
    for edge in bbox_edges:
        if covers_edge(edge, segs):
            hits += 1
    return hits == 4



#  主函数
#  (9)
#  判断多段线PL2是否在多段线PL1多边形中

#  该函数系列包括如下一些函数

#&&% 顶点全在内部
def are_all_vertices_inside(pl1, pl2):
    """
    判断多段线 pl2 的所有顶点是否都在多段线 pl1 构成的多边形内部。

    参数：
      pl1, pl2: COM 多段线对象（LWPOLYLINE 或 POLYLINE），视为封闭正多边形。
    返回：
      (all_inside, outside_pts)
      - all_inside: 如果 pl2 的每个顶点都在 pl1 内部，返回 True，否则 False。
      - outside_pts: 列表，包含所有落在 pl1 外部的 pl2 顶点 (x, y, z)。
    """
    # 先把 COM 多段线转成顶点列表 [(x,y,z), ...]
    verts1 = get_unique_vertices_from_pl_com(pl1)
    verts2 = get_unique_vertices_from_pl_com(pl2)

    outside_pts = []
    for pt in verts2:
        if not point_in_polygon(pt, verts1):
            outside_pts.append(pt)

    all_inside = len(outside_pts) == 0
    if all_inside:
        print("[OK] pl2 的所有顶点都在 pl1 的内部。")
    else:
        sys_logger.info(f"[错误] pl2 有 {len(outside_pts)} 个顶点不在 pl1 内部：")
        for p in outside_pts:
            print("   ", p)

    return all_inside, outside_pts



#&&&&%% 第四部分 一般对象


def ensure_list(input_data):
    """
    【通用工具】将输入参数统一转换为列表。
    
    增强功能：
    - 自动解包：如果输入是元组且第一个元素是列表（例如 (polylist, dict, ...)），
      则自动提取第一个元素返回，消除数据结构混乱。
    - 兼容 COM：支持 SelectionSet 等 COM 集合。
    """
    # 1. 处理 None
    if input_data is None:
        return []

    # 定义检测是否为“列表类”的内部函数 (List, Tuple, COM Collection)
    def is_list_like(obj):
        if isinstance(obj, (list, tuple)):
            return True
        # 针对 COM 集合 (有 Count 但没有 ObjectName)
        if hasattr(obj, "Count") and not hasattr(obj, "ObjectName"):
            return True
        return False

    # 2. 判断输入本身是否为列表/集合
    if is_list_like(input_data):
        # 先统一转为 Python 列表，方便操作
        try:
            current_list = list(input_data)
        except:
            # 极少数转换失败的情况（如损坏的 COM 对象），作为单对象处理
            return [input_data]

        # 3. 【核心优化】探测嵌套结构并解包
        if len(current_list) > 0:
            first_item = current_list[0]
            
            # 探测条件：第一个元素也是列表类
            if is_list_like(first_item):
                
                # 场景 A: 输入是元组 (Tuple) -> 强烈的“返回值容器”信号
                # 例子: dy = ([line1, line2], {mapping}) -> 解包返回 [line1, line2]
                if isinstance(input_data, tuple):
                    return ensure_list(first_item) # 递归调用以确保内部也被清理
                
                # 场景 B: 包含且仅包含一个列表的列表 -> 可能是多余的包装
                # 例子: [[line1, line2]] -> 解包返回 [line1, line2]
                if len(current_list) == 1:
                    return ensure_list(first_item)
                
                # 场景 C: 混合结构 (List + Dict) -> 可能是 [list, dict]
                # 例子: [[line1], {"info":1}] -> 判定为数据+元数据，解包
                if len(current_list) > 1 and isinstance(current_list[1], dict):
                    return ensure_list(first_item)

        # 4. 如果不符合解包条件（例如是坐标点列表 [[0,0], [10,10]]），则原样返回
        return current_list

    # 5. 默认：单个对象，包裹返回
    return [input_data]






# 按com实体对象中提取的坐标排序

#&&&%  排序
#&&% 元组排序
def sort_tuples(lst,cha_Y =2000):#对列表按插入点xy坐标排序
    
    """
    这是很有用的一个双值排序函数，对于COM对象，可以先将其转换为元组，即可使用这个函数

    它的价值在于，很容易拓展到n值排序
    """
    
    # 先按照m[1]降序排序
    lst.sort(key=lambda x: -x[2][1])

    i = 0
    while i < len(lst) - 1:
        j = i + 1
        # 查找所有m[1]差距在chaY以内的元素
        while j < len(lst) and abs(lst[i][2][1] - lst[j][2][1]) < cha_Y:
            j += 1
        
        # 如果找到了m[1]值相近的元素，根据m[0]值进行排序
        if j - i > 1:
            lst[i:j] = sorted(lst[i:j], key=lambda x: x[2][0])
        
        i = j

    return lst

#&&% 多维容差排序
def multi_dim_tolerance_sort(lst, key_index=2, tolerances=[10000, 1000, 0]):#高维排序
    """
    对 lst 列表中的元组按多维坐标字段排序，考虑每个维度的容差进行逐层排序。

    参数：
        lst: [(id, name, (x, y, z)), ...]
        key_index: 坐标在元组中的索引（默认是第3项，即元组[2]）
        tolerances: 每个维度允许的容差，例如 [Z差, Y差, X差]

    返回：
        排好序的新列表
    """
    # 最外层排序：按最高维降序排（Z从上往下）
    dim = len(tolerances)
    lst.sort(key=lambda x: -x[key_index][0])  # Z 倒序

    def recursive_sort(sublist, level):
        if level >= dim - 1:
            # 最后一级直接按该维升序
            return sorted(sublist, key=lambda x: x[key_index][level + 1])
        
        i = 0
        while i < len(sublist) - 1:
            j = i + 1
            while j < len(sublist) and abs(sublist[i][key_index][level] - sublist[j][key_index][level]) < tolerances[level]:
                j += 1
            if j - i > 1:
                sublist[i:j] = recursive_sort(sublist[i:j], level + 1)
            i = j
        return sublist

    return recursive_sort(lst, 0)


def get_ll_pt(ent):#提取函数，对象左下角点
    minpt, _ = ent.GetBoundingBox()
    return minpt[0], minpt[1],0

def get_center(ent):#提取函数，中心点
    minpt, maxpt = ent.GetBoundingBox()
    return ((minpt[0]+maxpt[0])/2, (minpt[1]+maxpt[1])/2)

#&&% 实体位置排序
def sort_entities_by_position( entity_list, extract_func, cha_Y=2000):#对com对象按提取坐标分别沿y,x方向排序
        """
        对实体列表根据其坐标（通过 extract_func 获取）进行排序：
        - 先按 Y 值降序（从上到下）
        - Y 值接近（差值 < cha_Y）者再按 X 值升序（从左到右）

        参数：
            entity_list: COM 实体对象列表
            extract_func: 提取坐标函数，返回 (x, y) 元组的函数即可
            cha_Y: 同一行判定的 Y 方向容差

        返回：按坐标顺序排列的新实体对象列表

        调用示例

        sorted_objs = sort_entities_by_position(LB, extract_func=get_ll_pt)
        
        """
        triples = [(ent, *extract_func(ent)) for ent in entity_list]

        # 按 Y 值降序排列
        triples.sort(key=lambda t: -t[2])

        i = 0
        while i < len(triples) - 1:
            j = i + 1
            while j < len(triples) and abs(triples[i][2] - triples[j][2]) < cha_Y:
                j += 1
            triples[i:j] = sorted(triples[i:j], key=lambda t: t[1])  # 按 X 升序
            i = j

        return [t[0] for t in triples]

def get_line_start(ent):
    """
    提取一条直线的起点 (x, y)
    ent: AcDbLine 或类似对象，具有 .StartPoint 属性
    """
    pt = ent.StartPoint   # 假设是一个 (x, y, z) 或 [x, y, z]
    return pt[0], pt[1]

#&&% * 对列表实体进行从上到下、从左到右的排序

#&&% 左下角排序
def sort_coms_by_llcorner(com_list, cha_Y=2000):
    """
    按 BoundingBox 左下角坐标排序：
      · 先按 y 降序（越大越靠前，即自上而下）
      · 同一行(Δy<cha_Y)内按 x 升序（自左向右）
    """
    wrapped = []
    for ent in com_list:
        try:
            p1, _ = ent.GetBoundingBox()      # p1 已是左下
            x_ll, y_ll = p1[0], p1[1]
        except Exception:
            x_ll = y_ll = float('-inf')       # 取不到的一律放最后
        wrapped.append((ent, x_ll, y_ll))     # (实体, x, y)

    # 先按 y 降序
    wrapped.sort(key=lambda t: -t[2])

    i = 0
    while i < len(wrapped) - 1:
        j = i + 1
        while j < len(wrapped) and abs(wrapped[i][2] - wrapped[j][2]) < cha_Y:
            j += 1
        # 行内再按 x 升序
        if j - i > 1:
            wrapped[i:j] = sorted(wrapped[i:j], key=lambda t: t[1])
        i = j

    return [ent for ent, _, _ in wrapped]




#&&% 右上角排序
def sort_coms_by_rbcorner(com_list, *, cha_X=100):
    """
    竖向图框（或已整体旋转 -90° 的图纸）使用 ——  
    · 先按列：x₁ 降序（右 → 左）  
    · 列内：  y₂ 降序（上 → 下）

    参数
    ----
    com_list : list[COM object]
        文字实体列表
    cha_X    : float
        判定“同一列”的容差，单位与图纸坐标一致
    """
    wrapped = []
    for ent in com_list:
        try:
            (x1, _, _), (_, y2, _) = ent.GetBoundingBox()  # x1 = 左, y2 = 上
        except Exception:
            x1 = y2 = float("-inf")                        # 失败的排最后
        wrapped.append((ent, x1, y2))

    # — ① 按 x₁ 降序：最右列在前 —
    wrapped.sort(key=lambda t: -t[1])

    # — ② 同列内再按 y₂ 降序：上 → 下 —
    i = 0
    while i < len(wrapped) - 1:
        j = i + 1
        while j < len(wrapped) and abs(wrapped[i][1] - wrapped[j][1]) < cha_X:
            j += 1
        wrapped[i:j] = sorted(wrapped[i:j], key=lambda t: -t[2])
        i = j

    return [t[0] for t in wrapped]


#&&% 自定义左下角排序
def sort_coms_by_llcorner_custom(objs, tol_x=100):
    """
    按左下角 x, y 坐标对 COM 对象列表 objs 排序：
      1. 先按 x 升序分组：相邻 x 差 ≤ tol_x 视为同一组
      2. 每组内按 y 降序排列（y 大的排前）
      3. 最终把各组按它们的 x 升序依次拼接

    参数
    ----
    objs : List[COMObject]
        要排序的 COM 对象列表，每个对象必须支持 GetBoundingBox()
    tol_x : float
        x 方向上的容差，小于等于此值的 x 差视为同组

    返回
    ----
    List[COMObject]
        排序后的对象列表
    """
    # 提取 (obj, llx, lly)
    items = []
    for e in objs:
        ll, ur = e.GetBoundingBox()
        llx, lly, _ = ll
        items.append((e, llx, lly))

    # 按 x 升序初排
    items.sort(key=lambda t: t[1])

    # 分组：相邻 x 差 ≤ tol_x 的归一组
    clusters = []
    rep_x, current = items[0][1], [items[0]]
    for item in items[1:]:
        _, x, _ = item
        if abs(x - rep_x) <= tol_x:
            current.append(item)
        else:
            clusters.append((rep_x, current))
            rep_x, current = x, [item]
    clusters.append((rep_x, current))

    # 按组的 x 升序排好后，组内按 y 降序，再平铺
    result = []
    for _, group in sorted(clusters, key=lambda c: c[0]):
        group_sorted = sorted(group, key=lambda t: t[2], reverse=True)
        result.extend([t[0] for t in group_sorted])

    return result


#&&% 中心点排序
def sort_coms_by_center(objs, tol_x=100):
    """
    按外包盒中心坐标对 COM 对象列表 objs 排序：
      1. 先计算每个对象的 bbox 中心 (cx, cy)
      2. 按 cx 升序初排序
      3. 将相邻 cx 差 ≤ tol_x 的对象归同一组
      4. 各组按 cy 降序排列（cy 大的排前）
      5. 最后按组的 cx 升序依次拼接各组内对象

    参数
    ----
    objs : List[COMObject]
        要排序的 COM 对象列表，每个对象必须支持 GetBoundingBox()
    tol_x : float
        x 方向上的容差，小于等于此值的 cx 差视为同组

    返回
    ----
    List[COMObject]
        排序后的对象列表
    """
    # 1. 提取 (obj, cx, cy)
    items = []
    for e in objs:
        ll, ur = e.GetBoundingBox()
        cx = (ll[0] + ur[0]) / 2.0
        cy = (ll[1] + ur[1]) / 2.0
        items.append((e, cx, cy))

    if not items:
        return []

    # 2. 按 cx 升序初排
    items.sort(key=lambda t: t[1])

    # 3. 分组：相邻 cx 差 ≤ tol_x 的归一组
    clusters = []
    rep_x, current = items[0][1], [items[0]]
    for item in items[1:]:
        _, x, _ = item
        if abs(x - rep_x) <= tol_x:
            current.append(item)
        else:
            clusters.append((rep_x, current))
            rep_x, current = x, [item]
    clusters.append((rep_x, current))

    # 4. 各组内部按 cy 降序排列，再平铺
    result = []
    for rep_x, group in sorted(clusters, key=lambda c: c[0]):
        group_sorted = sorted(group, key=lambda t: t[2], reverse=True)
        result.extend([t[0] for t in group_sorted])

    return result


##对列表实体进行正序或逆序编号

#&&% 实体编号
def number_entities_by_order(entity_list, prefix="", start=1, k=0):
    """
    对排序好的 COM 实体对象列表进行编号。

    参数：
        entity_list: 实体对象列表
        prefix: 编号前缀，默认为空字符串
        start: 编号起始数值，默认为 1
        k: 排序方向控制变量：
            - k = 0 正序编号
            - k = 1 逆序编号

    返回：
        编号字符串列表（如 ["1", "2", "3"] 或 ["图1", "图2", "图3"]）
    """
    n = len(entity_list)
    index_list = range(n) if k == 0 else reversed(range(n))
    
    result = []
    for i, idx in enumerate(index_list):
        label = f"{prefix}{start + i}"
        result.append(label)
    
    return result


# 重复操作列表对象

#&&% 列表遍历操作

def pr_list(P: list, f, *args, **kwargs):
    """
    args:  位置参数元组 (例如: 10, 20)
    kwargs: 关键字参数字典 (例如: layer="Wall", color=1)
    """
    results = []
    sys_logger.info(f"开始处理 {len(P)} 个对象...")
    
    for i, item in enumerate(P):
        try:
            # 关键点：将接收到的 item 和额外参数一起传给 f
            # item 是列表里的元素，args/kwargs 是你传入的固定参数
            res = f(item, *args, **kwargs)
            results.append(res)
        except Exception as e:
            sys_logger.info(f"❌ 第 {i} 个出错: {e}")
            results.append(None)
            
    return results





#&&% 列表提取操作
def apply_to_each2(obj_list, extract_func, action_func):#双层嵌套重复操作列表对象
    """
    对 obj_list 中的每个对象，先通过 extract_func 提取值，
    再将该值传入 action_func 中处理。

    参数：
        obj_list: 对象列表
        extract_func: 用于提取 (x, y) 或其他值的函数
        action_func: 用于处理提取结果的函数（如 srhd）
    """
    for obj in obj_list:
        value = extract_func(obj)
        action_func(value)

#&&&% 组
"""

创建组
group = doc.Groups.Add("mygroup")
LB=pmxz()
请在屏幕拾取图元，以Enter键结束

group.AppendItems(vtobj([LB[0], LB[1], LB[2]]))


从组名获取组合组中对象

group = doc.Groups.Item("G001")

entities = [group.Item(i) for i in range(group.Count)]  # 遍历组内对象

entities[0].Handle
'2C3'

entities[0].Move(vtpnt(0,0,0),vtpnt(0,10000,0))


解除组
group.Delete()


group1 = doc.Groups.Add("mygroup")
group1.AppendItems(vtobj([LB[0], LB[1], LB[2]]))
group2 = doc.Groups.Add("mygroupA")
group2.AppendItems(vtobj([LB[3], LB[4], LB[5],LB[6]]))
group3 = doc.Groups.Add("mygroupB")
group3.AppendItems(vtobj([group1,group2]))非法操作
group3.AppendItems(vtobj([LB[0],LB[1],LB[2],LB[3], LB[4], LB[5],LB[6]]))


# 提取 group1 和 group2 的所有成员
group1_entities = [group1.Item(i) for i in range(group1.Count)]
group2_entities = [group2.Item(i) for i in range(group2.Count)]

# 合并成新的列表
all_entities = group1_entities + group2_entities

# 创建 group3
group3 = doc.Groups.Add("mygroupB")
group3.AppendItems(vtobj(all_entities))

 get_boundingbox_from_objects(objs)
"""

# 建立全部列表com对象的最小边界框

#&&% 获取对象群包围盒
def get_boundingbox_from_objects(objs):#从列表com对象建立最小边界框
    """
    从一组图形对象（如 LB）中获取整体包围盒
    返回值：(min_x, min_y, min_z), (max_x, max_y, max_z)
    """
    min_point, max_point = None, None

    for obj in objs:
        try:
            min_pt, max_pt = obj.GetBoundingBox()
            if min_point is None:
                min_point, max_point = list(min_pt), list(max_pt)
            else:
                min_point = [min(min_point[i], min_pt[i]) for i in range(3)]
                max_point = [max(max_point[i], max_pt[i]) for i in range(3)]
        except Exception as e:
            sys_logger.info(f"跳过无法获取边界的对象: {obj.ObjectName}")
            continue

    return tuple(min_point), tuple(max_point)
        

# 建立组的最小边界框

#&&% 创建组
def chuangjian_zu(group_name):

    group = doc.Groups.Add(group_name)

    return group

#&&% 获取组对象
def nametogroup(group_name):#从组名获取实体com组对象
    group_obj = doc.Groups.Item(group_name)

    return group_obj

##获取所有组

#&&% 获取所有组名
def get_all_group_names():
    """
    获取当前 DWG 文档中所有组的名称列表。
    
    返回:
      List[str] — 包含所有组名称的列表
    """
    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    groups = doc.Groups
    return [groups.Item(i).Name for i in range(groups.Count)]

#&&% 获取所有组
def get_all_groups():
    """
    获取当前 DWG 文档中所有组的 COM 对象列表及其名称。
    
    返回:
      List[Tuple[str, COMObject]] — 每项为 (组名称, 组对象)
    """
    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument
    groups = doc.Groups
    result = []
    for i in range(groups.Count):
        grp = groups.Item(i)
        result.append((grp.Name, grp))
    return result



#将多个com对象对象加入名为group_name的组
#&&% 添加对象到组
def add_objects_to_group(group_name, obj_list):

    """
    将 obj_list 中的所有图形对象加入名为 group_name 的组中
    如果组已存在，使用原组；否则新建
    返回：Group 对象
    """
    groups = doc.Groups
    try:
        group = groups.Item(group_name)
    except:
        group = groups.Add(group_name)

    group.AppendItems(vtlist(obj_list))
    return group


#将单独com对象对象加入名为group_name的组中

#&&% 添加单对象到组
def add_object_to_group(group_name, obj):
    """
    将单个图形对象 obj 加入名为 group_name 的组中。
    如果组已存在，则使用该组；否则新建一个新组。
    
    参数：
      group_name (str)：组名称
      obj：要加入组的 COM 对象（如多段线、线段、块参照等）
    
    返回：
      COM Group 对象
    """
    groups = doc.Groups
    try:
        # 尝试获取已存在的组
        group = groups.Item(group_name)
    except Exception:
        # 不存在则新建
        group = groups.Add(group_name)
    
    # vtlist 工具将 Python list 转成 VBA 可接受的 SAFEARRAY
    group.AppendItems(vtlist([obj]))
    return group

#将单独com对象对象移出名为group_name的组
#&&% 移除组内对象
def remove_object_from_group(group_name, obj):
    """
    将单个 COM 对象 obj 从名为 group_name 的组中移出。
    如果组不存在或对象不在组中，则会打印错误信息但不抛异常。
    
    参数:
      group_name: 组名（字符串）
      obj:         要移出的 COM 对象
    
    返回:
      如果组存在，返回该 Group 对象；否则返回 None。
    """
    try:
        group = doc.Groups.Item(group_name)
    except Exception:
        sys_logger.info(f"[错误] 组 '{group_name}' 不存在")
        return None

    # 把单个对象包装成长度为1的 COM SafeArray
    variant = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, [obj])
    try:
        group.RemoveItems(variant)
        sys_logger.info(f"[OK] 已从组 '{group_name}' 中移除对象 {obj.Handle}")
    except Exception as e:
        sys_logger.info(f"[错误] 从组 '{group_name}' 移除对象失败：{e}")

    return group

#将多个com对象对象移出名为group_name的组
#&&% 批量移除组内对象
def remove_objects_from_group(group_name, obj_list):
    """
    将 obj_list 中的所有图形对象从名为 group_name 的组中移出。
    如果组不存在，会打印提示并返回 None；否则返回该组对象。
    
    :param group_name: 组名称
    :param obj_list: 要移除的 COM 对象列表
    :return: Group 对象 或 None
    """
    groups = doc.Groups
    try:
        group = groups.Item(group_name)
    except Exception:
        sys_logger.info(f"组 '{group_name}' 不存在，无法移除对象。")
        return None

    # 把 Python 列表包装成 VARIANT SafeArray，VT_DISPATCH 表示对象类型
    arr = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj_list)
    try:
        group.RemoveItems(arr)
        sys_logger.info(f"已从组 '{group_name}' 中移除 {len(obj_list)} 个对象。")
    except Exception as e:
        sys_logger.info(f"移除对象时发生错误：{e}")
    return group



#&&% 从名为group_name的组获取内部包含的实体对象


#&&% 获取组内实体
def get_com_from_groupname(group_name):
    """
    根据组名获取对应实体列表。
    - 若组不存在、或组中无实体，均返回空列表，不抛异常。
    """

    try:
        group = nametogroup(group_name)
    except Exception:
        # nametogroup 本身失败（组不存在等）
        return []

    if not group:
        # group 为 None 或空，也直接返回空列表
        return []

    entities = [group.Item(i) for i in range(group.Count)] #从com组中获取全部对象
    
    return entities

#从名为group_name的组返回按类型分类的字典
#&&% 获取组内实体分类
def get_com_from_groupname_by_type(group_name):
    """
    根据组名获取对应实体，并按类型名分类返回。

    :param group_name: 组名称
    :return: dict，键为实体类型名（ObjectName），值为该类型的实体列表
    """
    # nametogroup 是你已有的“组名→Group 对象”函数
    group = nametogroup(group_name)
    if group is None:
        sys_logger.info(f"组 '{group_name}' 不存在")
        return {}

    by_type = {}
    # Group.Count 是实体数量，Item(i) 取出第 i 个实体
    for i in range(group.Count):
        ent = group.Item(i)
        # AutoCAD COM 对象一般有 ObjectName 属性
        typ = getattr(ent, "ObjectName", None) or getattr(ent, "EntityName", "Unknown")
        by_type.setdefault(typ, []).append(ent)

    # 打印一下各类型数量，方便调试
    for typ, lst in by_type.items():
        sys_logger.info(f"  类型 {typ} ：{len(lst)} 个实体")

    return by_type

#从名为group_name的组返回按类型分类的字典，且类型按各自位置提取函数排好序
#&&% 获取组内实体排序
def get_group_entities_sorted(group_name, type_extractors, cha_Y=0.5):
    """
    从组中按类型获取实体，并对指定类型按坐标排序。

    :param group_name: str，组名
    :param type_extractors: dict，{ type_name: extract_func }，
           extract_func(ent) 返回 (x, y) 坐标，用于排序。
    :param cha_Y: float，同一行 Y 方向的容差
    :return: dict，{ type_name: [ent, ...] }，已排序或原序

    对天正单行多行文字可以使用GetBoundingBox获取的左下角点
    
    """
    def get_cadtext_pos(ent):
        # CAD单行、多行文字：Position 属性返回 (x, y, z)
        return float(ent.InsertionPoint[0]), float(ent.InsertionPoint[1])


    # 先按类型获取全部实体
    by_type = get_com_from_groupname_by_type(group_name)
    sorted_by_type = {}

    for typ, ents in by_type.items():
        if typ in type_extractors:
            extract_func = type_extractors[typ]
            # 排序
            sorted_list = sort_entities_by_position(ents, extract_func, cha_Y=cha_Y)
            sorted_by_type[typ] = sorted_list
            sys_logger.info(f"Type '{typ}' sorted with {len(sorted_list)} entities")
        else:
            # 保持原序
            sorted_by_type[typ] = list(ents)
            sys_logger.info(f"Type '{typ}' left unsorted ({len(ents)} entities)")

    return sorted_by_type


#从名为group_name的组返回按类型分类的字典，各类型统一按boundingbox中心排好序


#&&% 组内实体按中心排序
def get_group_entities_sorted_by_type_and_bbox(group_name, cha_Y=0.5):
    """
    将组 group_name 中的实体按类型分类，并对每种类型内部按包围盒中心排序：
      1) 先按 center_y 降序（从上到下）
      2) 同一“行”内（|ΔY|<cha_Y）再按 center_x 升序（从左到右）

    参数：
      group_name: 要操作的组名
      cha_Y: 同一“行”Y 方向容差

    返回：
      一个 dict，key=类型名(ObjectName)，value=排序后的实体列表
    """
    # 1. 取组
    group = nametogroup(group_name)
    ents = [group.Item(i) for i in range(group.Count)]

    # 2. 按类型分组
    by_type = {}
    for ent in ents:
        typ = getattr(ent, "ObjectName", None) or ent.EntityName
        by_type.setdefault(typ, []).append(ent)

    # 3. 辅助：计算包围盒中心点

    # 4. 对每个类型内部排序
    for typ, lst in by_type.items():
        triples = [(e, *bbox_center_2(e)) for e in lst]
        # Y 降序
        triples.sort(key=lambda t: -t[2])
        # 同行内按 X 升序
        i = 0
        while i < len(triples) - 1:
            j = i + 1
            while j < len(triples) and abs(triples[i][2] - triples[j][2]) < cha_Y:
                j += 1
            triples[i:j] = sorted(triples[i:j], key=lambda t: t[1])
            i = j
        # 覆盖原列表
        by_type[typ] = [t[0] for t in triples]

    return by_type


# 获取两个组中共有的实体，按类型分类并按包围盒中心排序
#&&% 共有组实体排序
def common_group_entities_sorted(group_name1, group_name2, cha_Y=0.5):
    """
    获取两个组中共有的实体，按类型分类并按包围盒中心排序。

    参数：
      group_name1, group_name2: 要比较的两个组名
      cha_Y: 同一行判定的 Y 方向容差

    返回：
      dict => key: ObjectName 类型名, value: 排序后的实体列表
    """
    # 1. 取组
    g1 = nametogroup(group_name1)
    g2 = nametogroup(group_name2)

    ents1 = [g1.Item(i) for i in range(g1.Count)]
    ents2 = [g2.Item(i) for i in range(g2.Count)]

    # 2. 建立 handle->entity 映射
    map1 = {e.Handle: e for e in ents1}
    map2 = {e.Handle: e for e in ents2}

    # 3. 找共有的 handles
    common_handles = set(map1.keys()) & set(map2.keys())

    # 4. 收集共有实体（这里取自 map1）
    common_ents = [map1[h] for h in common_handles]

    # 5. 按类型分组
    by_type = {}
    for ent in common_ents:
        typ = getattr(ent, "ObjectName", None) or ent.EntityName
        by_type.setdefault(typ, []).append(ent)

    # 6. 包围盒中心点计算
    def bbox_center(e):
        min_pt, max_pt = e.GetBoundingBox()
        x1, y1, _ = tuple(min_pt)
        x2, y2, _ = tuple(max_pt)
        return ((x1 + x2) / 2, (y1 + y2) / 2)

    # 7. 排序：Y 降序，同一行内(|ΔY|<cha_Y)按 X 升序
    for typ, lst in by_type.items():
        triples = [(e, *bbox_center(e)) for e in lst]
        triples.sort(key=lambda t: -t[2])
        i = 0
        while i < len(triples) - 1:
            j = i + 1
            while j < len(triples) and abs(triples[i][2] - triples[j][2]) < cha_Y:
                j += 1
            triples[i:j] = sorted(triples[i:j], key=lambda t: t[1])
            i = j
        by_type[typ] = [t[0] for t in triples]

    return by_type

#&&% 获取组包围盒
def get_boundingbox_from_group(group):#从com对象group建立最小边界框

    """
    并非组的实际BoundingBOX数据

    """

    entities = [group.Item(i) for i in range(group.Count)] #从组中获取全部对象

    p1,p2 = get_boundingbox_from_objects(entities)

    return p1,p2

#&&% 复制组S1
def copy_group_S1_from_doc1_to_doc2(doc1, doc2, group_name="S1"):#将名为“S1”的组复制到当前桌面上另一个文档
    """
    将 doc1 中名为 group_name 的组复制到 doc2 中，并重新组装组。
    粘贴点为 0,0,0。粘贴后通过 handle 差集识别新对象。

    mp要不断更新

    def refresh_modelspace(doc):
        return doc.ModelSpace
    逻辑更清晰

    直接发送命令复制粘贴或许更好
    """
    from CAD_file_operations import set_active_doc

    try:
        # 1. 激活源文档
        set_active_doc(doc1)
        li()

        group = doc.Groups.Item(group_name)
        handles = [ent.Handle for ent in group]
        objs = [ent for ent in mp if ent.Handle in handles]

        yin_to_xian_xuanze(objs)

        doc.SendCommand("_copybase\n0,0,0\n")
        time.sleep(0.5)
        doc.SendCommand("_copyclip\n\n")
        time.sleep(1)

        # 2. 激活目标文档
        set_active_doc(doc2)
        li()
        ms2 = doc.ModelSpace

        # 3. 粘贴前记录已有对象
        pre_map = get_handle_object_map(ms2)

        # 4. 粘贴
        doc.SendCommand("_pasteclip"+chr(13)+"0,0,0"+chr(13)+chr(13))#
        time.sleep(1.5)

        # 5. 粘贴后重新记录对象
        ms2 = doc.ModelSpace
        
        post_map = get_handle_object_map(ms2)

        # 6. 取出新对象（通过 handle 差集）
        new_handles = set(post_map) - set(pre_map)
        new_objs = [post_map[h] for h in new_handles]

        sys_logger.info(f"[OK] 粘贴完成，识别出 {len(new_objs)} 个新图元")

        # 7. 添加这些对象到组中（使用你提供的方法）
        add_objects_to_group(group_name, new_objs)

        sys_logger.info(f"[OK] 成功将粘贴对象加入组 '{group_name}'")

    except Exception as e:
        sys_logger.info(f"[错误] 复制组失败: {e}")

        



#&&&% Handle和Label


"""
Handle身份信息和天正标签Label

与列表和组不同，身份信息的标识使我们可以从字典存储的名称信息精准控制每个对象，而不是总依赖对全部对象的遍历


可以在对象创建时就打上标签，使对象归为某个类

可以在一批对象完成后调整标签

可以临时标记一组标签

可以通过Label给天正对象加标签
q1= doc.ModelSpace.Item(doc.ModelSpace.Count - 1)#自动获取图纸空间之前最后一个绘制的对象


从Handle回溯com对象

doc.HandleToObject('2BD')
q1= doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
q1.Label="A1"
q1= doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
q1.Label="B1"
LB=pmxz()
请在屏幕拾取图元，以Enter键结束
LB[0].Label
'B1'
LB[1].Label
'A1'


"""

# 将列表对象按分类将其handle身份标识存入字典

@alias("h")
#&&% 句柄转对象
def HandleToObject(ZF):#从Handle身份信息值回溯com对象

    """
    对连接在墙上的门窗测试无效

    """

    obj = doc.HandleToObject(ZF)

    return obj


def print_coms_handle(LB):

    LC=[]

    for x in LB:

        LC.append(x.Handle)

    sys_logger.info(f"com对象列表对应的Handle句柄列表：{LC} ")




@alias("H")
#&&% 批量句柄转对象
def handles_to_coms(LB_handles):

    """
    对连接在墙上的门窗测试无效

    """
    LC=[]

    for xx in LB_handles:

        obj = doc.HandleToObject(xx)
        LC.append(obj)

    return LC


#&&% 获取所有句柄
def get_all_handles():#获取所有Handle
    """
    获取当前图纸中所有对象（通常在 ModelSpace）的 Handle 值列表。

    返回：
        handle_list: 所有图元的 Handle 字符串列表
    """
    handle_list = []

    for obj in mp:  # 使用你已经定义好的全局 ModelSpace mp
        try:
            handle_list.append(obj.Handle)
        except:
            continue  # 跳过无 Handle 或异常对象

    sys_logger.info(f"[OK] 已获取 {len(handle_list)} 个对象的 Handle")
    return handle_list

#&&% 查找实体
def find_entity_by_handle(handle_str):#从Handle获取实体（适合包括天正的文件）
    """
    遍历当前图纸所有对象，手动比对 Handle 值，找到指定的实体对象。

    参数：
        handle_str: 目标 Handle（字符串）

    返回：
        对象（若找到），否则 None
    """
    for obj in mp:  # 可扩展：遍历 sp 也行
        try:
            if obj.Handle == handle_str:
                return obj
        except:
            continue

    return None



#&&% 按类型句柄分组
def group_objects_by_type_and_handle(LB):#将列表对象的Handle身份信息分类存入字典返回
    """
    将com对象列表 LB 中的对象按 ObjectName 分类，并存储其 Handle。
    每类按 LB 中出现顺序编号。

    参数：
        LB - AutoCAD 实体对象列表（如 select_objects_in_window_area() 返回）

    返回：
        ZD - dict 格式 {ObjectName: [Handle1, Handle2, ...]}
    """
    ZD = {}  # 初始化字典

    for obj in LB:
        try:
            obj_type = obj.ObjectName
            handle = obj.Handle

            if obj_type not in ZD:
                ZD[obj_type] = []

            ZD[obj_type].append(handle)

        except Exception as e:
            sys_logger.info(f"[警告]️ 跳过对象，原因: {e}")
            continue

    # 输出提示信息
    for obj_type, handles in ZD.items():
        sys_logger.info(f"[OK] {obj_type}: 共 {len(handles)} 个对象")

    return ZD

# 通过名称存储对象信息反回溯对象

#&&% 记录类型句柄
def record_handle_with_type(LB, typename, prefix="OBJ"):#将一批对象的 Handle 存储到结构化的字典中，并记录类型名和编号
    """
    替代 XData 方法：记录对象 Handle、类型名、编号，返回结构化字典。
    """
    ZD = {typename: {}}
    for i, obj in enumerate(LB, start=1):
        try:
            h = obj.Handle
            tag = f"{prefix}_{i:03d}"
            ZD[typename][h] = tag
        except:
            continue
    sys_logger.info(f"[OK] 已记录 {len(ZD[typename])} 个“{typename}”对象（Handle+编号）")
    return ZD

#&&% 转换命名字典
def convert_named_dict(ZD, typename):# 构建编号 → COM 对象 的映射字典
    """
    将 ZD["门"] 的结构由 Handle: 编号 转换为 编号: COM对象
    返回：新的字典 {编号: COM实体}
    """
    doc = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    named_dict = {}

    handle_map = ZD.get(typename, {})
    for handle, name in handle_map.items():
        try:
            obj = doc.HandleToObject(handle)
            named_dict[name] = obj
        except:
            sys_logger.info(f"[警告]️ 无法找到对象（handle={handle}）")
            continue

    return named_dict

#&&% 获取命名对象
def get_named_object(tag, ZD, typename="门"):#从标签获取对象
    named = convert_named_dict(ZD, typename)
    return named.get(tag)



#&&% 绘制固定标签
def draw_tags_on_objects_fixed(named_dict, height=250, offset=(1000, 1000, 0)):#直接将编号文字写在每个图元上，通常居中或偏移一点点
    """
    在每个对象的中心点附近绘制标注文字。
    
    参数:
        named_dict - 如 {"Men_001": <COMObject>, ...}
        height     - 文字高度
        offset     - 偏移量（用于防止文字盖住对象）
    """
    import win32com.client
    import pythoncom

    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc = acad.ActiveDocument
    ms = doc.ModelSpace

    for name, obj in named_dict.items():
        try:
            # 使用 GetBoundingBox 获取中心点
            min_pt, max_pt = obj.GetBoundingBox()
            center_pt = (
                (min_pt[0] + max_pt[0]) / 2,
                (min_pt[1] + max_pt[1]) / 2,
                (min_pt[2] + max_pt[2]) / 2
            )

            # 加上偏移量
            label_pt = (
                center_pt[0] + offset[0],
                center_pt[1] + offset[1],
                center_pt[2] + offset[2]
            )

            # 确保插入点是三维点
            pt = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, label_pt)
            
            # 添加文字
            ms.AddText(name, pt, height)
            sys_logger.info(f"[OK] 已标注对象: {name}")

        except Exception as e:
            sys_logger.info(f"[警告]️ 标注失败: {name}, 错误: {e}")

# 给天正对象打上标签存入字典，用于以名称反向回溯操作

#&&% 标记天正门
def label_tarch_doors(LB1, typename="门", prefix="men"):#给选中的天正门打上标签并存入字典返回，非天正图元没有Label属性
    """
    从对象列表 LB1 中筛选出天正门 (ObjectName == 'TDbOpening')，
    并为其按顺序打上 .Label 标签（如 men001, men002 ...）。

    返回：
        ZD = {typename: {编号: 对象}}
    """
    ZD = {typename: {}}
    LB2 = []

    for obj in LB1:
        try:
            if hasattr(obj, "ObjectName") and obj.ObjectName == "TDbOpening":
                LB2.append(obj)
        except:
            continue

    for i, obj in enumerate(LB2, start=1):
        try:
            tag = f"{prefix}{i:03d}"
            obj.Label = tag
            ZD[typename][tag] = obj
            sys_logger.info(f"[OK] 已标记: {tag}")
        except Exception as e:
            sys_logger.info(f"[警告]️ 设置标签失败：{e}")

    sys_logger.info(f"\n📦 共找到并标注 {len(LB2)} 个天正门")
    return ZD




# 获取模型空间上的Handle


"""
target_handles = ['5F', '60', '61']
map = get_handle_object_map(doc.ModelSpace)
objs = [map[h] for h in target_handles if h in map]
这比每次都遍历 ModelSpace 快得多，尤其是大图纸中上千个图元时。

"""

#&&% 获取句柄映射
def get_handle_object_map(ms):
    """返回 {handle: object} 映射"""
    return {ent.Handle: ent for ent in ms}



#&&% XData

"""

在 RegAppTable 中注册,在第一次给图元附加 XData 时，AutoCAD 内部会检查 RegAppTable（注册应用程序表）中是否已经存在 “TestApp” 这个名称。

如果不存在，AutoCAD 会自动往 RegAppTable 里插入一条记录，把 “TestApp” 注册进去。

如果你希望手动控制，也可以先调用 doc.Application.RegistryModes.Add("TestApp")（或使用 AutoLISP：(regapp "TestApp")）

app_name    = "TestApp"            # 自定义的应用程序名
data_types  = [1000, 1040, 1070]
data_values = ["示例文字", 3.14159, 12345]
set_xdata(lineObj, app_name, data_types, data_values)
types_out, data_out = get_xdata(lineObj, app_name)
types_out
[1001, 1000, 1040, 1070]
data_out
['TestApp', '示例文字', 3.14159, 12345]


"""
#&&% 设置扩展数据
def set_xdata(
    com_obj,
    app_name: str,
    data_types: list[int],
    data_values: list,
):
    """
    向任意 AutoCAD COM 对象（如 Line、Circle、BlockReference 等）附加 XData。

    参数：
      com_obj        -- 任意支持 SetXData() 方法的 COM 对象
      app_name       -- 注册过的应用程序名（字符串）；第一个 DataType 一定是 1001，对应的第一个 Data 存放此 app_name
      data_types     -- 后续的 DataType 列表（不含第一项 1001）；例如 [1000, 1040, 1070] 等
      data_values    -- 与 data_types 对应的数据值列表；长度与 data_types 一一对应。例如 ["文字串", 3.14, 42]

    说明：
      AutoCAD 规定 XData 的第一对元素必须是 (1001, 应用程序名)。后面才是按顺序出现的其他 (DataType, Data)。
      因此实际发送给 SetXData 的 DataType 数组第一个元素要放 1001，Data 数组第一个元素要放 app_name。
    """
    def vtint(val):
        """
        将 Python 列表转换为 VARIANT 类型的整数数组，
        以便传给 COM 对象作为 XData 的 DataType。
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, val)

    def vtvariant(var):
        """
        将 Python 列表转换为 VARIANT 类型的 VARIANT 数组，
        以便传给 COM 对象作为 XData 的 Data。
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, var)


    # 1. 拼接完整的 DataType 列表：第一个为 1001
    full_types = [1001] + data_types

    # 2. 拼接完整的 Data 列表：第一个为 app_name
    full_data = [app_name] + data_values

    # 3. 将 Python 列表转换为 VARIANT 数组
    vt_types = vtint(full_types)
    vt_data  = vtvariant(full_data)

    # 4. 调用 COM 的 SetXData 方法
    com_obj.SetXData(vt_types, vt_data)

#&&% 获取扩展数据
def get_xdata(
    com_obj,
    app_name: str,
):
    """
    从任意 AutoCAD COM 对象（如 Line、Circle、BlockReference 等）读取 XData。

    参数：
      com_obj   -- 任意支持 GetXData() 方法的 COM 对象
      app_name  -- 申请读取的应用程序名（必须与 set_xdata 时使用的相同）

    返回：
      (type_codes, data_values) 二元组，其中
        type_codes: Python 列表，对应每个 XData 项目的 DataType（包括第一项 1001）
        data_values: Python 列表，对应每个 XData 项目的 Data（包括第一项 app_name）
    
    如果该对象没有附加此 app_name 下的 XData，则 GetXData 会抛出错误；建议调用前先用 Error Handling 包裹或
    通过 com_obj.GetXData(app_name) 进行捕获并返回 None。
    """
    def vtint(val):
        """
        将 Python 列表转换为 VARIANT 类型的整数数组，
        以便传给 COM 对象作为 XData 的 DataType。
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, val)

    def vtvariant(var):
        """
        将 Python 列表转换为 VARIANT 类型的 VARIANT 数组，
        以便传给 COM 对象作为 XData 的 Data。
        """
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, var)


    try:
        type_codes, data_values = com_obj.GetXData(app_name)
        # 注意：返回的 type_codes 和 data_values 都是 tuple，转换为 list 更易处理
        return list(type_codes), list(data_values)
    except pythoncom.com_error:
        # 对象上不存在该 app_name 的 XData，或读取失败
        return None, None


#&&% Xdata标记

#&&% 设置打印标记
def set_xdata_tab(entitycom):

    app_name    = "PrintApp"
    data_types  = [1000]
    data_values = ["增补目录模板"]
    set_xdata(entitycom, app_name, data_types, data_values)

    return

#&&% 检查打印标记
def is_printApp_xdata_com(entitycom):

    try:

        get_xdata( entitycom, "PrintApp")

        return True

    except:

        return  False

#&&&% 文字

#&&% 写CAD单行文字

def write_cad_text(
    p=(0, 0, 0),
    text="单行文字",
    alignment="左下",
    height=350,
    width_factor=1.0,
    rotation=0.0,
    oblique=0.0,
    style="Standard",
    layer=None
):

    """
    【架构适配版】在指定位置写入 CAD 单行文字。

    """
   

    # =========================================================================
    # 1. 架构接入：获取环境
    # =========================================================================
    # 不再自己去 GetActiveObject，而是请求 CAD_basic 刷新环境
    if not C.li():
        print("[错误] 无法连接 CAD 或无激活文档，write_cad_text 中止。")
        return None


    ms = C.mp 

    # =========================================================================
    # 2. 内部辅助函数 (保持不变，逻辑很棒)
    # =========================================================================
    def local_set_attr(obj, name, value):
        try:
            setattr(obj, name, value)
        except Exception as e:
            # sys_logger.info(f"[警告] 属性 {name} 设置忽略") 
            pass

    def _align_text_dynamic(text_ent, target_pt, align_mode):
        try:
            # 强制刷新，确保 GetBoundingBox 能拿到真实尺寸
            text_ent.Update()
            
            # 获取包围盒 (MinPt, MaxPt)
            bbox = text_ent.GetBoundingBox()
            if not bbox: return
            
            min_pt, max_pt = bbox
            xmin, ymin, zmin = min_pt
            xmax, ymax, zmax = max_pt
            
            # 解析对齐模式，计算文字目前的“锚点”位置
            alg = str(align_mode).strip().lower()
            if alg in ("左下", "lb"):
                anchor = (xmin, ymin, zmin)
            elif alg in ("左上", "lt"):
                anchor = (xmin, ymax, zmin)
            elif alg in ("右下", "rb"):
                anchor = (xmax, ymin, zmin)
            elif alg in ("右上", "rt"):
                anchor = (xmax, ymax, zmin)
            elif alg in ("中心", "居中", "center", "c"):
                anchor = ((xmin + xmax)/2.0, (ymin + ymax)/2.0, (zmin + zmax)/2.0)
            else:
                return # 默认对齐（左下）不需要移动

            # 准备目标点 (Z轴补全)
            if len(target_pt) == 2:
                tp = (float(target_pt[0]), float(target_pt[1]), float(zmin))
            else:
                tp = (float(target_pt[0]), float(target_pt[1]), float(target_pt[2]))

            # 计算移动向量：从 anchor 移到 tp
            # 注意：Move 方法需要 VARIANT 类型的起点和终点
            vt_from = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, anchor)
            vt_to   = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, tp)
            
            text_ent.Move(vt_from, vt_to)
            
        except Exception as e:
            sys_logger.info(f"[警告] 文字对齐运算失败: {e}")

    # =========================================================================
    # 3. 核心执行逻辑
    # =========================================================================
    try:
        # --- 参数准备 ---
        p_list = [float(x) for x in p]
        if len(p_list) == 2: p_list.append(0.0)
        
        # 封装坐标点 (Early Binding 要求)
        insert_pt_variant = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, p_list)
        height_float = float(height)

        # --- 创建对象 ---
        # 使用 cb.mp (ms) 创建
        text_obj = ms.AddText(text, insert_pt_variant, height_float)

        # --- 属性设置 ---
        if style: local_set_attr(text_obj, "StyleName", style)
        if width_factor != 1.0: local_set_attr(text_obj, "ScaleFactor", float(width_factor))
        if rotation != 0.0: local_set_attr(text_obj, "Rotation", math.radians(rotation))
        if oblique != 0.0: local_set_attr(text_obj, "ObliqueAngle", math.radians(oblique))
        if layer: local_set_attr(text_obj, "Layer", layer)

        # --- 执行对齐 ---
        # 只有当需要特殊对齐时才调用移动逻辑
        if alignment and alignment not in ["左下", "lb", None]:
             _align_text_dynamic(text_obj, p_list, alignment)

        return text_obj

    except pythoncom.com_error as e:
        sys_logger.info(f"[COM 错误] WriteCadText 失败 (HR={e.hresult}): {e}")
        return None
    except Exception as e:
        sys_logger.info(f"[通用错误] WriteCadText 失败: {e}")
        return None



#&&% 写天正单行文字
def write_tianzheng_text(
    p=(0, 0, 0),
    text="天正单行文字tianzhengdanhangwenzi",
    alignment="左下",            # "左下"/"左对齐"/"左上"/"右下"/"右上"/"中心"/"center"
    height=3.5,
    width_factor=1.0,
    rotation=0.0,
    oblique=0.0,
    style="Standard",
    system_layer="xitong_tianzhengwenzi",
    system_file_name="tianzhengdanhangwenzi.dwg",
    delete_system_text=False,
):
    """
    在当前激活图中写入一段“天正单行文字”（通过系统模板 Copy 实现），
    然后通过 last_obj() 找到模型空间最后生成的对象，
    按其外包盒锚点对齐到指定点 p。

    alignment 支持：
        - "左下" / "左对齐" / "LB"
        - "左上" / "LT"
        - "右下" / "RB"
        - "右上" / "RT"
        - "中心" / "center" / "C"
      其它值默认按“左下”处理。
    """
    from CAD_file_operations import copy_file_content_pywin32

    # —— 小工具：根据外包盒对齐到目标点 —— #
    def _align_entity_by_bbox(ent, target_point, align="左下"):
        """
        使用无参数版 GetBoundingBox()：
          1. min_pt, max_pt = ent.GetBoundingBox()
          2. 根据 alignment 选锚点
          3. Move(锚点 → 目标点)
        """
        try:
            min_pt, max_pt = ent.GetBoundingBox()  # ★ 关键：和你在控制台里一样的调用方式
        except Exception as e:
            sys_logger.info(f"[错误] 获取外包盒失败: {e}")
            return

        xmin, ymin, zmin = min_pt
        xmax, ymax, zmax = max_pt
        sys_logger.info(f"[对齐-前]  BBox min={min_pt}, max={max_pt}")

        alg = str(align).strip().lower()
        if alg in ("左下", "左对齐", "lb"):
            anchor_pt = (xmin, ymin, zmin)
        elif alg in ("左上", "lt"):
            anchor_pt = (xmin, ymax, zmin)
        elif alg in ("右下", "rb"):
            anchor_pt = (xmax, ymin, zmin)
        elif alg in ("右上", "rt"):
            anchor_pt = (xmax, ymax, zmin)
        elif alg in ("中心", "center", "c"):
            anchor_pt = ((xmin + xmax) / 2.0,
                         (ymin + ymax) / 2.0,
                         (zmin + zmax) / 2.0)
        else:
            anchor_pt = (xmin, ymin, zmin)

        # 目标点：如果只传 (x, y)，z 用当前 zmin；否则用传入的 z
        if len(target_point) == 2:
            tx, ty = float(target_point[0]), float(target_point[1])
            tz = float(zmin)
        else:
            tx, ty, tz = float(target_point[0]), float(target_point[1]), float(target_point[2])

        target_pt = (tx, ty, tz)

        sys_logger.info(f"[对齐-计算] alignment='{align}', anchor_pt={anchor_pt} → target_pt={target_pt}")

        from_pt = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, anchor_pt)
        to_pt   = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, target_pt)
        ent.Move(from_pt, to_pt)

        # 再读一次外包盒看结果
        try:
            min_pt2, max_pt2 = ent.GetBoundingBox()
            sys_logger.info(f"[对齐-后]  BBox min={min_pt2}, max={max_pt2}")
        except Exception as e:
            sys_logger.info(f"[警告] 对齐后再次获取外包盒失败: {e}")

    try:
        # ===== 1. 确保连接到当前 DWG =====

        acad = C.acad
        doc  = C.doc

        # ===== 2. 查找/插入系统模板文字 =====
        lb = stc(system_layer)
        if len(lb) == 0:
            # 需要从系统文件插入一次模板
            system_file = Path(XITONG_DIR) / system_file_name
            if not system_file.exists():
                sys_logger.info(f"[错误] 系统天正文字文件不存在: {system_file}")
                return None

            current_file = doc.FullName
            sys_logger.info(f"[信息] 未找到系统文字图层 {system_layer}，从 {system_file} 插入...")

            try:
                
                ok = copy_file_content_pywin32(str(system_file), current_file)
                if not ok:
                    print("[警告] copy_file_content_pywin32 返回 False，可能部分失败，但继续尝试。")
            except Exception as e:
                sys_logger.info(f"[警告] 插入系统文字文件时抛出异常：{e}，继续尝试查找图层。")

            # 再 li 一次，重新查模板

            acad = C.acad
            doc  = C.doc
            lb   = stc(system_layer)

        if len(lb) == 0:
            sys_logger.info(f"[错误] 即使插入系统文件后，仍未在图层 {system_layer} 找到模板文字。")
            return None
        elif len(lb) > 1:
            sys_logger.info(f"[警告] 图层 {system_layer} 上找到 {len(lb)} 个对象，默认使用第一个作为模板。")

        template = lb[0]
        sys_logger.info(f"[信息] 使用模板文字: ObjectName={template.ObjectName}, Handle={template.Handle}")

        # ===== 3. 从模板 Copy 出新文字对象（不对齐） =====
        try:
            new_text = template.Copy()
        except Exception as e:
            sys_logger.info(f"[错误] 无法从模板文字 Copy 新对象：{e}")
            return None

        sys_logger.info(f"[DEBUG] 复制得到的新文字: ObjectName={new_text.ObjectName}, Handle={new_text.Handle}")

        # ===== 4. 设置天正文字属性（对新对象操作） =====
        try:
            set_object_property(new_text, "Text", text)
            set_object_property(new_text, "Height", height)
            set_object_property(new_text, "WidthFactor", width_factor)
            set_object_property(new_text, "Rotation", rotation)
            set_object_property(new_text, "Oblique", oblique)
            if style:
                set_object_property(new_text, "TextStyle", style)

            print(
                f"[步骤] 已设置天正文字属性: "
                f"text='{text}', 高度={height}, 宽度因子={width_factor}, "
                f"旋转={rotation}, 倾斜={oblique}, 样式={style or '[沿用模板]'}"
            )
        except Exception as e:
            sys_logger.info(f"[警告] 设置天正文字属性时出错：{e}")

        # ===== 5. 等待天正生成完对象，再通过 last_obj() 重新选取 =====
        try:
            time.sleep(1.5)     # 给天正一点时间
            try:
                doc.Regen(1)    # acAllViewports
            except Exception:
                pass
            try:
                pythoncom.PumpWaitingMessages()
            except Exception:
                pass
        except Exception:
            pass


        acad = C.acad
        doc  = C.doc

        try:
            ent = last_obj()
        except Exception as e:
            sys_logger.info(f"[错误] 调用 last_obj() 失败: {e}")
            return None

        # last_obj 可能返回单个对象或列表
        try:
            _ = ent.ObjectName
        except Exception:
            try:
                ent = ent[-1]
            except Exception as e:
                sys_logger.info(f"[错误] last_obj() 返回值类型无法识别: {e}")
                return None

        objname = getattr(ent, "ObjectName", "<无>")
        handle  = getattr(ent, "Handle", "<无>")
        sys_logger.info(f"[DEBUG] last_obj() 得到实体: ObjectName={objname}, Handle={handle}")

        # ===== 6. 用外包盒锚点对齐到 p（默认左下） =====
        _align_entity_by_bbox(ent, p, align=alignment or "左下")

        # ===== 7. 按需删除系统模板文字（保留这次新写的文字） =====
        if delete_system_text:
            try:
                print(
                    f"[信息] delete_system_text=True，正在删除图层 "
                    f"{system_layer} 上的系统模板文字（保留当前新文字）..."
                )
                cnt = 0
                ent_handle = str(getattr(ent, "Handle", "")).upper()
                objs = stc(system_layer)
                print(
                    f"[OK] 第 1 次尝试：选到图层 ['{system_layer}'] 上 "
                    f"{len(objs)} 个对象"
                )
                for obj in objs:
                    try:
                        h = str(getattr(obj, "Handle", "")).upper()
                        # 保护当前新文字：Handle 相同则跳过
                        if h == ent_handle:
                            continue
                        obj.Delete()
                        cnt += 1
                    except Exception:
                        pass
                sys_logger.info(f"[成功] 已删除 {cnt} 个系统文字模板对象")
            except Exception as e:
                sys_logger.info(f"[警告] 删除系统模板文字失败：{e}")

        sys_logger.info(f"[完成] 天正单行文字创建成功，位置约为 {p}\n")
        return ent  # 返回对齐后的实体

    except Exception as e:
        sys_logger.info(f"[错误] write_tianzheng_text 执行失败：{e}")
        import traceback
        traceback.print_exc()
        return None


# ====================  文字垂直对齐 ====================

#&&% 文字垂直对齐
def align_text_to_vertical_line(
    text_obj,
    x_position,
    align_side="左边界"
):
    """
    将文字按 BoundingBox 边界对齐到指定垂直线的 X 坐标。

    参数:
        text_obj:
            - 单个文字对象 (AcDbText / TDbText / 其它有 GetBoundingBox 的实体)
            - 或者多个文字对象组成的列表 / 元组 / 其它可迭代

        x_position:
            - 一个数字 x
            - 或一个点 (x, y)
            - 或一个点 (x, y, z)
            最终只使用 x 作为垂直线的 X 坐标。

        align_side:
            - "左边界": 使用外包盒左边界对齐到 x
            - "右边界": 使用外包盒右边界对齐到 x
    """

    C.li()
    # —— 1. 归一化 text_obj 为列表 —— #
    if text_obj is None:
        print("[错误] text_obj 为空，无法对齐。")
        return False

    # 单个对象：不是可迭代（或者是 COM 对象），就包装成列表
    objs = None
    if isinstance(text_obj, (list, tuple, set)):
        objs = list(text_obj)
    else:
        # 有些 COM 对象也会被当成可迭代，这里简单认为“有 GetBoundingBox 属性”的就是单个对象
        # 保守起见：直接包装成列表
        objs = [text_obj]

    if not objs:
        print("[错误] text_obj 列表为空，无法对齐。")
        return False

    # —— 2. 解析 x_position —— #
    if isinstance(x_position, numbers.Real):
        x_target = float(x_position)
    elif isinstance(x_position, (list, tuple)):
        if len(x_position) == 0:
            print("[错误] x_position 为空序列。")
            return False
        x_target = float(x_position[0])
    else:
        # 其它类型（例如 VARIANT），尝试转成 float
        try:
            x_target = float(x_position)
        except Exception:
            sys_logger.info(f"[错误] 无法从 x_position={x_position!r} 解析出 X 坐标。")
            return False

    sys_logger.info(f"[信息] 垂直对齐目标 X = {x_target}，处理对象数量 = {len(objs)}")

    # —— 3. 遍历对齐每一个对象 —— #
    success_count = 0

    for idx, obj in enumerate(objs, start=1):
        try:
            # 3.1 获取 BoundingBox（无参数版本）
            ll_pt, ur_pt = obj.GetBoundingBox()
            sys_logger.info(f"[对象#{idx}] 原外包盒: min={ll_pt}, max={ur_pt}")

            # 3.2 计算移动距离
            if align_side == "左边界":
                dx = x_target - float(ll_pt[0])
            elif align_side == "右边界":
                dx = x_target - float(ur_pt[0])
            else:
                sys_logger.info(f"[对象#{idx}] [错误] 不支持的对齐方式: {align_side}")
                continue

            # 3.3 执行移动（只沿 X 方向）
            base_pt = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0.0, 0.0, 0.0])
            move_vec = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [dx, 0.0, 0.0])
            obj.Move(base_pt, move_vec)

            # 3.4 再读一遍外包盒验证
            ll_pt2, ur_pt2 = obj.GetBoundingBox()
            sys_logger.info(f"[对象#{idx}] 新外包盒: min={ll_pt2}, max={ur_pt2}")
            sys_logger.info(f"[对象#{idx}] [成功] 文字{align_side}已对齐到 X={x_target}")
            success_count += 1

        except Exception as e:
            sys_logger.info(f"[对象#{idx}] [错误] 垂直对齐失败: {e}")

    if success_count == 0:
        print("[结果] 所有对象对齐都失败。")
        return False
    elif success_count < len(objs):
        sys_logger.info(f"[结果] 部分对齐成功: 成功 {success_count} / {len(objs)}")
        return False
    else:
        sys_logger.info(f"[结果] 全部 {success_count} 个对象已成功对齐到 X={x_target}")
        return True

# ====================  文字水平对齐 ====================
#&&% 文字水平对齐
def align_text_to_horizontal_line(
    text_obj,
    y_position,
    align_side="下边界"
):
    """
    将文字按 BoundingBox 边界对齐到指定水平线的 Y 坐标。

    参数:
        text_obj:
            - 单个文字对象 (AcDbText / TDbText / 其它有 GetBoundingBox 的实体)
            - 或者多个文字对象组成的列表 / 元组 / 其它可迭代

        y_position:
            - 一个数字 y
            - 或一个点 (x, y)
            - 或一个点 (x, y, z)
            最终只使用 y 作为水平线的 Y 坐标。

        align_side:
            - "下边界": 使用外包盒下边界对齐到 y
            - "上边界": 使用外包盒上边界对齐到 y
    """


    C.li()
    # —— 1. 归一化 text_obj 为列表 —— #
    if text_obj is None:
        print("[错误] text_obj 为空，无法对齐。")
        return False

    if isinstance(text_obj, (list, tuple, set)):
        objs = list(text_obj)
    else:
        # 直接包装成列表（和垂直对齐那边保持一致）
        objs = [text_obj]

    if not objs:
        print("[错误] text_obj 列表为空，无法对齐。")
        return False

    # —— 2. 解析 y_position —— #
    if isinstance(y_position, numbers.Real):
        y_target = float(y_position)
    elif isinstance(y_position, (list, tuple)):
        if len(y_position) < 2:
            print("[错误] y_position 序列长度不足 2，无法获取 Y 坐标。")
            return False
        y_target = float(y_position[1])
    else:
        # 尝试从其它类型（例如 VARIANT）解析为 float
        try:
            y_target = float(y_position)
        except Exception:
            sys_logger.info(f"[错误] 无法从 y_position={y_position!r} 解析出 Y 坐标。")
            return False

    sys_logger.info(f"[信息] 水平对齐目标 Y = {y_target}，处理对象数量 = {len(objs)}")

    # —— 3. 遍历对齐每一个对象 —— #
    success_count = 0

    for idx, obj in enumerate(objs, start=1):
        try:
            # 3.1 获取 BoundingBox（无参数版本）
            ll_pt, ur_pt = obj.GetBoundingBox()
            sys_logger.info(f"[对象#{idx}] 原外包盒: min={ll_pt}, max={ur_pt}")

            # 3.2 计算移动距离
            if align_side == "下边界":
                dy = y_target - float(ll_pt[1])
            elif align_side == "上边界":
                dy = y_target - float(ur_pt[1])
            else:
                sys_logger.info(f"[对象#{idx}] [错误] 不支持的对齐方式: {align_side}")
                continue

            # 3.3 执行移动（只沿 Y 方向）
            base_pt = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0.0, 0.0, 0.0])
            move_vec = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0.0, dy, 0.0])
            obj.Move(base_pt, move_vec)

            # 3.4 再读一遍外包盒验证
            ll_pt2, ur_pt2 = obj.GetBoundingBox()
            sys_logger.info(f"[对象#{idx}] 新外包盒: min={ll_pt2}, max={ur_pt2}")
            sys_logger.info(f"[对象#{idx}] [成功] 文字{align_side}已对齐到 Y={y_target}")
            success_count += 1

        except Exception as e:
            sys_logger.info(f"[对象#{idx}] [错误] 水平对齐失败: {e}")

    if success_count == 0:
        print("[结果] 所有对象水平对齐都失败。")
        return False
    elif success_count < len(objs):
        sys_logger.info(f"[结果] 部分水平对齐成功: 成功 {success_count} / {len(objs)}")
        return False
    else:
        sys_logger.info(f"[结果] 全部 {success_count} 个对象已成功对齐到 Y={y_target}")
        return True

# ====================  缩放天正文字高度 ====================

#&&% 缩放天正文字
def scale_tianzheng_text_to_cad(
    tianzheng_text_obj,
    cad_text_obj
):
    """
    使用 ScaleEntity 将天正文字的 BoundingBox 高度缩放到 CAD 文字的高度。

    参数:
        tianzheng_text_obj:
            - 单个天正文字对象（TDbText/TDbMText等，有 GetBoundingBox/ScaleEntity）
            - 或多个天正文字对象组成的列表/元组/集合

        cad_text_obj:
            - CAD 文字对象（AcDbText/AcDbMText 等），
              用它的 BoundingBox 高度作为“目标高度”
    """
    C.li()
    # —— 1. 归一化 tianzheng_text_obj 为列表 —— #
    if tianzheng_text_obj is None:
        print("[错误] tianzheng_text_obj 为空，无法缩放。")
        return False

    if isinstance(tianzheng_text_obj, (list, tuple, set)):
        tz_objs = list(tianzheng_text_obj)
    else:
        tz_objs = [tianzheng_text_obj]

    if not tz_objs:
        print("[错误] tianzheng_text_obj 列表为空，无法缩放。")
        return False

    # —— 2. 获取 CAD 文字的目标高度 —— #
    try:
        cad_ll_pt, cad_ur_pt = cad_text_obj.GetBoundingBox()
        cad_height = float(cad_ur_pt[1]) - float(cad_ll_pt[1])
    except Exception as e:
        sys_logger.info(f"[错误] 获取 CAD 文字 BoundingBox 失败: {e}")
        return False

    if not isinstance(cad_height, numbers.Real) or cad_height == 0:
        sys_logger.info(f"[错误] CAD 文字高度无效: {cad_height}")
        return False

    sys_logger.info(f"[信息] 目标 CAD 文字高度 = {cad_height:.4f}，待缩放对象数 = {len(tz_objs)}")

    # —— 3. 遍历缩放每一个天正文字 —— #
    success_count = 0

    for idx, tz_obj in enumerate(tz_objs, start=1):
        try:
            # 3.1 获取天正文字的 BoundingBox（无参数版）
            tz_ll_pt, tz_ur_pt = tz_obj.GetBoundingBox()
            tz_height = float(tz_ur_pt[1]) - float(tz_ll_pt[1])

            sys_logger.info(f"[对象#{idx}] 原外包盒: min={tz_ll_pt}, max={tz_ur_pt}")
            sys_logger.info(f"[对象#{idx}] 原高度: {tz_height:.4f}")

            if tz_height == 0:
                sys_logger.info(f"[对象#{idx}] [错误] 天正文字高度为 0，跳过。")
                continue

            # 3.2 计算缩放比例
            scale_factor = cad_height / tz_height

            # 3.3 以外包盒中心作为缩放基点
            center_x = (float(tz_ll_pt[0]) + float(tz_ur_pt[0])) / 2.0
            center_y = (float(tz_ll_pt[1]) + float(tz_ur_pt[1])) / 2.0
            if len(tz_ll_pt) > 2:
                center_z = (float(tz_ll_pt[2]) + float(tz_ur_pt[2])) / 2.0
            else:
                center_z = 0.0

            scale_pt = VARIANT(
                pythoncom.VT_ARRAY | pythoncom.VT_R8,
                [center_x, center_y, center_z]
            )

            # 3.4 执行缩放
            tz_obj.ScaleEntity(scale_pt, scale_factor)

            # 3.5 再获取一次外包盒，验证新的高度
            new_ll_pt, new_ur_pt = tz_obj.GetBoundingBox()
            new_height = float(new_ur_pt[1]) - float(new_ll_pt[1])

            sys_logger.info(f"[对象#{idx}] [成功] 已缩放天正文字")
            sys_logger.info(f"  原高度: {tz_height:.4f}")
            sys_logger.info(f"  目标高度: {cad_height:.4f}")
            sys_logger.info(f"  实际新高度: {new_height:.4f}")
            sys_logger.info(f"  缩放比例: {scale_factor:.6f}")
            success_count += 1

        except Exception as e:
            sys_logger.info(f"[对象#{idx}] [错误] 缩放失败: {e}")

    if success_count == 0:
        print("[结果] 所有天正文字缩放都失败。")
        return False
    elif success_count < len(tz_objs):
        sys_logger.info(f"[结果] 部分缩放成功: 成功 {success_count} / {len(tz_objs)}")
        return False
    else:
        sys_logger.info(f"[结果] 全部 {success_count} 个天正文字已成功缩放到 CAD 文字高度。")
        return True

#&&&% 非图形对象


#&&% *** 将屏幕所选对象赋予到指定图层


@alias("s1")
def sc_objs_to_layer(layer_name,cl=256):

    doc=C.doc

    def pmxz_new():
        """
        人工选择对象，返回所选实体对象列表。
        自动清理已有的 "SS1" 选择集。
        """
        try:
            # 如果 "SS1" 已存在，先删除
            try:
                ss = doc.SelectionSets.Item("SS1")
                ss.Delete()
            except:
                pass  # 如果不存在就忽略

            selection = doc.SelectionSets.Add("SS1")
            selection.SelectOnScreen()
            objs = [selection.Item(i) for i in range(selection.Count)]
            selection.Delete()
            return objs
        except Exception as e:
            print("[错误] 选择失败:", e)
            return []



    layers = doc.Layers

    try:
        layer = layers.Item(layer_name)
    except:
        layer = layers.Add(layer_name)
        sys_logger.info(f"🟢 已新建图层：{layer_name}")

    LB=pmxz_new()

    for x in LB:

        x.Layer = layer_name

        x.color = cl


    return LB

#&&% 删除图层
def delete_layer(layername: str):
    """
    删除当前 DWG 文件中名为 layername 的图层。
    - 如果图层不存在，直接返回。
    - 如果图层是当前层，则切换到 0 层后再删除。
    - 删除前会尝试解锁、去掉冻结/打印锁。
    """
    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc = acad.ActiveDocument
    layers = doc.Layers

    try:
        layer = layers.Item(layername)
    except Exception:
        sys_logger.info(f"[警告] 图层 {layername} 不存在。")
        return

    # 如果该图层是当前层，切换到 "0"
    if doc.ActiveLayer.Name == layername:
        doc.ActiveLayer = layers.Item("0")

    # 解锁、解冻、去掉打印锁定，避免删除失败
    try:
        layer.Lock = False
        layer.Freeze = False
        layer.Plottable = True
    except Exception:
        pass

    try:
        layer.Delete()
        sys_logger.info(f"[OK] 图层 {layername} 已删除。")
    except Exception as e:
        sys_logger.info(f"[错误] 删除失败：{e}")



#&&% 从列表创建图层
def create_layers_from_list(layer_names):
    """
    创建列表中指定的图层，如果图层已存在则跳过。

    参数：
        layer_names: 包含图层名称的字符串列表
    """
    try:
        _, doc = get_acad_doc()
        layers = doc.Layers
        created = 0
        skipped = 0

        for name in layer_names:
            try:
                _ = layers.Item(name)  # 检查是否已存在
                sys_logger.info(f"⏩ 图层已存在：{name}")
                skipped += 1
            except:
                layers.Add(name)
                sys_logger.info(f"[OK] 新建图层：{name}")
                created += 1

        sys_logger.info(f"\n📊 总计：新建 {created} 个图层，跳过 {skipped} 个已有图层")

    except Exception as e:
        print("[错误] 创建图层时出错：", e)


#&&% 从列表删除图层
def delete_layers_from_list(layer_names):
    """
    删除列表中指定的图层

    参数：
        layer_names: 包含图层名称的字符串列表

    返回：
        dict: {'deleted': 删除成功的图层列表, 'failed': 删除失败的图层列表}
    """
    try:
        _, doc = get_acad_doc()
        layers = doc.Layers
        deleted = []
        failed = []

        for name in layer_names:
            try:
                layer = layers.Item(name)
                # 检查是否为当前图层
                if doc.ActiveLayer.Name == name:
                    sys_logger.info(f"[警告] 图层 '{name}' 是当前图层，无法删除")
                    failed.append(name)
                    continue

                # 尝试删除图层
                layer.Delete()
                sys_logger.info(f"[成功] 已删除图层：{name}")
                deleted.append(name)
            except Exception as e:
                sys_logger.info(f"[失败] 无法删除图层 '{name}'：{e}")
                failed.append(name)

        sys_logger.info(f"\n[统计] 成功删除 {len(deleted)} 个图层，失败 {len(failed)} 个")
        return {'deleted': deleted, 'failed': failed}

    except Exception as e:
        sys_logger.info(f"[错误] 删除图层时出错：{e}")
        return {'deleted': [], 'failed': layer_names}


#&&% 逐点标注
def dim_by_points(*args):
    """
    使用天正逐点标注命令对倾斜对象进行标注

    参数：
        *args: 三个点坐标 (x1,y1,z1), (x2,y2,z2), (x3,y3,z3)
               或 p1, p2, p3 其中p1,p2是被标注对象的起始点和终点，p3是标注位置点

    返回：
        bool: 成功返回True
    """
    import pyautogui
    import time

    try:
        # 解析参数
        if len(args) == 3:
            p1, p2, p3 = args
        else:
            print("[错误] 需要3个点坐标")
            return False

        # 最小化所有窗口
        pyautogui.hotkey('win', 'd')
        time.sleep(0.5)

        # 激活AutoCAD窗口
        activate_window_by_title("AutoCAD", click_titlebar=True)
        time.sleep(0.5)

        # 发送天正逐点标注命令
        _, doc = get_acad_doc()
        cmd = f"zdbz\n{p1[0]},{p1[1]}\n{p2[0]},{p2[1]}\n{p3[0]},{p3[1]}\n\n"
        doc.SendCommand(cmd)

        print("[成功] 已执行逐点标注")
        return True

    except Exception as e:
        sys_logger.info(f"[错误] 标注失败：{e}")
        return False


#&&% 确保图层存在并清空
def ensure_layer(layer_name="jizhunwall"):
    """
    确保图层存在并设为当前图层，同时删除该图层上所有对象（最多重试 3 次）。
    """
    try:
        li()

        layers = doc.Layers
        # 1) 获取或新建图层
        try:
            layer = layers.Item(layer_name)
        except Exception:
            layer = layers.Add(layer_name)
            sys_logger.info(f"🟢 已新建图层：{layer_name}")
        # 2) 切换到图层
        doc.ActiveLayer = layer
        sys_logger.info(f"[OK] 当前图层已设置为：{layer_name}")

        # 3) 删除图层中全部对象，重试 up to 5
        for attempt in range(1, 6):
            ents = select_tuceng(layer_name)
            if not ents:
                # 已经没有对象，提前退出
                sys_logger.info(f"[CLEAN] 图层 '{layer_name}' 已清空（共尝试 {attempt - 1} 次）")
                break

            deleted = 0
            for ent in ents:
                try:
                    ent.Delete()
                    deleted += 1
                except:
                    continue
            sys_logger.info(f"  第 {attempt} 次删除：共删除 {deleted} 个对象")

            time.sleep(0.1)  # 短暂等待，确保对象被移除


            #刷新

            doc.SendCommand("RE\n")
            doc.SendCommand("Z\nE\n")

        else:
            # 五次都还有残留
            remaining = len(select_tuceng(layer_name))
            sys_logger.info(f"[警告] 重试 3 次后，图层 '{layer_name}' 仍有 {remaining} 个对象未能删除")

    except Exception as e:
        print("[错误] 创建/切换图层或清理失败：", e)
#&&% 只在模型空间上清理

def ensure_layer_model_only(layer_name="jizhunwall"):
    """
    确保图层存在并设为当前图层，同时【仅删除】该图层在模型空间（Model Space）中的对象。
    """
    from  CAD_file_operations   import get_obj_loc
    try:
        doc=C.doc  # 确保连接

        layers = doc.Layers
        # 1) 获取或新建图层
        try:
            layer = layers.Item(layer_name)
        except Exception:
            layer = layers.Add(layer_name)
            sys_logger.info(f"🟢 已新建图层：{layer_name}")
            
        # 2) 切换到图层
        doc.ActiveLayer = layer
        sys_logger.info(f"[OK] 当前图层已设置为：{layer_name}")

        # 3) 循环清理模型空间的对象
        for attempt in range(1, 6):
            # 获取该图层的所有对象（可能跨空间）
            ents = select_tuceng(layer_name)
            if not ents:
                sys_logger.info(f"[CLEAN] 图层 '{layer_name}' 已无对象（共尝试 {attempt - 1} 次）")
                break

            deleted_count = 0
            model_ents_found = False

            for ent in ents:
                # --- 空间判定核心逻辑 ---
                # 调用 get_obj_loc 判定，如果是 1 则代表模型空间
                if get_obj_loc(ent) == 1:
                    model_ents_found = True
                    try:
                        ent.Delete()
                        deleted_count += 1
                    except Exception as e:
                        continue
            
            if not model_ents_found:
                sys_logger.info(f"[CLEAN] 图层 '{layer_name}' 在模型空间中已清空")
                break

            sys_logger.info(f"  第 {attempt} 次尝试删除：模型空间已移除 {deleted_count} 个对象")
            
            # 短暂等待并刷新视图
            time.sleep(0.1)
            doc.SendCommand("RE\n")

        else:
            # 如果 5 次后仍有模型空间对象
            remaining_ents = select_tuceng(layer_name)
            still_in_model = sum(1 for e in remaining_ents if get_obj_loc(e) == 1)
            if still_in_model > 0:
                sys_logger.info(f"[警告] 经过 5 次尝试后，模型空间仍有 {still_in_model} 个对象无法删除")

    except Exception as e:
        print("[错误] 空间清理操作失败：", e)





#&&% 确保图层当前
@alias("s2")

def ensure_layer_current(layer_name="jizhunwall", max_retries=3):
    """
    确保图层存在并设为当前图层，失败时最多重试 max_retries 次。
    """
    layers = doc.Layers
    for attempt in range(1, max_retries + 1):
        try:
            # 获取或新建图层
            try:
                layer = layers.Item(layer_name)
            except Exception:
                layer = layers.Add(layer_name)
                sys_logger.info(f"🟢 已新建图层：{layer_name}")
            # 切换到图层
            doc.ActiveLayer = layer
            sys_logger.info(f"[OK] 当前图层已设置为：{layer_name} (尝试 {attempt})")
            return True
        except Exception as e:
            sys_logger.info(f"[错误] 尝试 {attempt} 创建/设置图层失败：{e}")
    sys_logger.info(f"[错误] 达到最大重试次数 ({max_retries})，无法创建或切换到图层：{layer_name}")
    return False


# 设置指定图层的颜色、线型、开关状态和冻结状态

#&&% 设置图层属性
@alias("s3")

def set_layer_properties(layer_name, color_index=9, linetype="Continuous", on=True, frozen=False):
    """
    设置指定图层的颜色、线型、开关状态和冻结状态。

    参数：
        layer_name (str): 图层名称
        color_index (int): 图层颜色索引（默认 9）
        linetype (str): 图层线型名称（默认 'Continuous'）
        on (bool): 图层是否打开（默认 True）
        frozen (bool): 图层是否冻结（默认 False）
    """
    li()
    try:
        layers = doc.Layers
        try:
            layer = layers.Item(layer_name)
        except:
            layer = layers.Add(layer_name)
            sys_logger.info(f"[OK] 已新建图层：{layer_name}")

        # 设置颜色
        layer.color = color_index

        # 设置线型
        try:
            ltype = doc.Linetypes.Item(linetype)
        except:
            doc.Linetypes.Load(linetype, linetype)
            ltype = doc.Linetypes.Item(linetype)
        layer.Linetype = linetype

        # 设置开关状态
        layer.LayerOn = on

        # 设置冻结状态
        layer.Freeze = frozen


        doc.SendCommand("re\n")


        sys_logger.info(f"🔧 图层属性已更新：{layer_name} | 颜色={color_index} | 线型={linetype} | 开关={'开' if on else '关'} | 冻结={'是' if frozen else '否'}")

    except Exception as e:
        sys_logger.info(f"[错误] 设置图层属性失败：{e}")




#&&% 将列表中的对象图层设为目标图层


def set_layer_with_retry(LB, layername, ci=3):
    """
    将给定 COM 对象列表 LB 中的每个对象的 Layer 属性设为 layername。
    【核心逻辑】:
    1. 自动创建不存在的图层。
    2. 使用 setattr 动态设置属性，避开部分第三方实体的早期绑定限制。
    3. 针对 CAD 繁忙状态提供重试机制。
    """

    li() 
    current_doc = globals().get('doc')
    if current_doc is None:
        print("❌ 无法获取 CAD 文档对象")
        return [], list(LB)
    
    success = []
    failed = []

    # --- 1. 检查并创建图层 ---
    try:
        layers = current_doc.Layers
        try:
            layers.Item(layername)
        except Exception:
            sys_logger.info(f"[信息] 图层“{layername}”不存在，正在创建...")
            layers.Add(layername)
    except Exception as e:
        sys_logger.info(f"[错误] 无法访问或创建图层“{layername}”: {e}")
        return [], list(LB)

    # --- 2. 遍历设置对象图层 ---
    for obj in LB:
        set_done = False
        for attempt in range(1, ci + 1):
            try:
                # 使用 setattr 动态设置属性，增强对第三方实体的兼容性
                setattr(obj, 'Layer', layername)
                success.append(obj)
                set_done = True
                break
            except Exception as e:
                if attempt == ci:
                    failed.append(obj)
                    handle = getattr(obj, 'Handle', '<未知Handle>')
                    sys_logger.info(f"[警告] 对象 {handle} 设置图层“{layername}”失败：{e}")
                else:
                    time.sleep(1) # 失败重试间隔
                    
    return success, failed






#&&% 强制改图层对象颜色

def force_layer_objects_color(layer_name, target_color=256, max_retries=3):
    """
    [最终修正版] 强制改色
    修复逻辑: 当 set_attr 成功但无法读取属性(None)时，视为成功。
    """
    sys_logger.info(f"\n🎨 [改色执行] 图层: {layer_name} -> 目标颜色: {target_color}")
    
    try:
        all_objs = stc(layer_name)
    except:
        sys_logger.info(f"   ❌ 无法选择图层 {layer_name}")
        return False

    if not all_objs:
        print("   ℹ️ 图层为空，跳过")
        return True

    pending_objs = list(all_objs)
    
    for attempt in range(1, max_retries + 1):
        if not pending_objs:
            print("   ✅ 所有对象已达标")
            break
            
        sys_logger.info(f"   🔄 第 {attempt} 轮 (剩 {len(pending_objs)} 个)...")
        next_round_objs = []
        
        for i, obj in enumerate(pending_objs):
            try:
                # 1. 执行修改
                # 使用小写 "color"，这是你实测有效的
                is_set_ok = set_attr(obj, "color", target_color)
                
                # 尝试刷新显示
                try: obj.Update()
                except: pass
                
                # 2. 读取验证
                new_color = get_attr(obj, "Color")
                
                # --- [关键逻辑修正] ---
                # 判定成功的两种情况:
                # A: 读回来的值等于目标值 (完美)
                # B: set_attr 说成功了(True)，但读回来是 None (无法读取，但信任写入)
                
                if new_color == target_color:
                    # sys_logger.info(f"      [Obj {i}] ✅ 验证成功")
                    pass 
                elif is_set_ok and new_color is None:
                    # sys_logger.info(f"      [Obj {i}] ⚠️ 验证受限 (读回None)，但写入返回True -> 视为成功")
                    pass 
                else:
                    # 只有当 写入失败 或者 读回来是明确的错误数值 时，才重试
                    # 获取 Handle 方便调试
                    h_val = "???"
                    try: h_val = get_attr(obj, "Handle")
                    except: pass
                    
                    sys_logger.info(f"      [Obj {i} | H={h_val}] ❌ 失败: Set={is_set_ok}, Get={new_color}")
                    next_round_objs.append(obj)
                    
            except Exception as e:
                sys_logger.info(f"      ❌ 异常: {e}")
                next_round_objs.append(obj)
        
        pending_objs = next_round_objs
        if pending_objs: time.sleep(0.1)

    if not pending_objs:
        sys_logger.info(f"   ✅ 改色完成: 图层 {layer_name} 全员 {target_color}")
        return True
    else:
        sys_logger.info(f"   ⚠️ 改色结束: 仍有 {len(pending_objs)} 个对象状态存疑")
        return False


#&&&&%% 第五部分 文件操作
#&&&% 原文件体系

#  该函数系列包括如下一些函数
"""
zhuancheng_t7()
zhuancheng_t3()
在li()连接激活文件后，直接执行该命令即可转换
一般理解所得文件和当前文件同文件夹，但这次测试结果却在文件夹
C:/Users/Administrator/Documents/WeChat Files/wxid_mhrzdlppqacc22/FileStorage/File/2025-04/
"""    
#&&% 对获得的两行大地坐标数据建立字典

"""
大量数据的整理读取，AI不能绝对保证正确

"""

#&&% 从文本构建J点
def build_J_points_from_selected_texts(LB, n_points=61, prefix_x="30", prefix_y="37"):
    num_re = re.compile(r"-?\d+(?:\.\d+)?")

    def first_number(s):
        m = num_re.search(str(s))
        if not m:
            raise ValueError(f"无法解析数字: {s!r}")
        return float(m.group(0))

    items = []
    for ent in LB:
        txt, ip = get_text_and_ip(ent)
        if not txt or not ip:
            continue
        txt = str(txt).strip().replace(" ", "")
        x0, y0, *_ = ip
        if txt.startswith(prefix_x):
            items.append(("x", first_number(txt), float(y0), float(x0)))
        elif txt.startswith(prefix_y):
            items.append(("y", first_number(txt), float(y0), float(x0)))

    # 从上到下（y降序），同行按x升序稳定
    items.sort(key=lambda t: (-t[2], t[3]))
    xs = [v for typ, v, _, _ in items if typ == "x"]
    ys = [v for typ, v, _, _ in items if typ == "y"]

    if len(xs) != n_points or len(ys) != n_points:
        raise RuntimeError(f"X={len(xs)} / Y={len(ys)} 与期望 {n_points} 不符。")

    pts_dict = {f"J{i+1}": (xs[i], ys[i]) for i in range(n_points)}
    pts_list = [(f"J{i+1}", xs[i], ys[i]) for i in range(n_points)]
    return pts_dict, pts_list


#从界点大地坐标计算经纬度
#&&% 坐标转经纬度
def convert_pts_dict_to_latlon(pts_dict, central_lon=111):
    """
    输入: pts_dict = {'J1': (N, E), 'J2': (N, E), ...}
          N=北坐标, E=东坐标 (CGCS2000 高斯-克吕格投影)
    输出: geo_dict = {'J1': (lon, lat), 'J2': (lon, lat), ...}  (经度, 纬度, 单位:度)
    """
    a = 6378137.0
    f = 1 / 298.257222101
    FE = 500000
    FN = 0
    e2 = 2*f - f*f
    e_prime_2 = e2 / (1.0 - e2)

    def one_point(N, E):
        m = N - FN
        mu = m / (a * (1 - e2/4 - 3*e2**2/64 - 5*e2**3/256))
        e1 = (1 - math.sqrt(1 - e2)) / (1 + math.sqrt(1 - e2))
        Bf = (mu
              + (3*e1/2 - 27*e1**3/32) * math.sin(2*mu)
              + (21*e1**2/16 - 55*e1**4/32) * math.sin(4*mu)
              + (151*e1**3/96) * math.sin(6*mu)
              + (1097*e1**4/512) * math.sin(8*mu))
        x_prime = E - FE
        cosBf = math.cos(Bf)
        tanBf = math.tan(Bf)
        eta2 = e_prime_2 * (cosBf**2)
        Nf = a / math.sqrt(1 - e2 * (math.sin(Bf)**2))
        Mf = a * (1 - e2) / (1 - e2 * (math.sin(Bf)**2))**1.5
        lat = (Bf
               - tanBf/(2* Mf* Nf) * (x_prime**2)
               + tanBf/(24* Mf* Nf**3) * (5 + 3*tanBf**2 + eta2 - 9*tanBf**2 * eta2) * (x_prime**4)
               - tanBf/(720* Mf* Nf**5) * (61 + 90*tanBf**2 + 45*tanBf**4) * (x_prime**6))
        lon = (math.radians(central_lon)
               + x_prime/(Nf* cosBf)
               - (1 + 2*tanBf**2 + eta2) * (x_prime**3)/(6* Nf**3* cosBf)
               + (5 + 28*tanBf**2 + 24*tanBf**4 + 6* eta2 + 8*tanBf**2 * eta2) * (x_prime**5)/(120* Nf**5* cosBf))
        return math.degrees(lon), math.degrees(lat)

    geo_dict = {}
    for k, (N, E) in pts_dict.items():
        geo_dict[k] = one_point(N, E)

    return geo_dict

# 文件基本操作



"""
使用共同的文件全局变量acad,mp,doc,sp是我们编制脚本控制不同文件的基础            


"""


    















    

#&&% 显示隐藏图形
def xianshi_yincangtuxing():#  显示文件中可能隐藏的对象
        
    acad.ActiveDocument.SendCommand("HFKJ"+chr(13)+chr(13))#在V4状态下可能要改成"HFKJ"+chr(13)+"Y"+chr(13)

    doc.Save()



# 创建一个事件对象

timeout_event = threading.Event()

event = threading.Event()

#&&% 运行CAD程序
def run_cad_program(timeout_event, event):
    pythoncom.CoInitialize()
    
    try:
        acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
        print(acad.ActiveDocument.Name)
        
        acad.ActiveDocument.SendCommand("TSaveAs"+chr(13))
        
        print("CAD命令已发送，等待窗口操作完成...")
        
        # 等待timeout_event信号，如果收到信号，则退出
        timeout_event.wait()

    except Exception as e:
        sys_logger.info(f"run_cad_program 出现错误: {e}")
        pass
    finally:
        pythoncom.CoUninitialize()
        event.set()  # 通知主线程完成

        
#&&% 自动化T7窗口
def automate_window_with_pywinauto_t7(timeout_event, event):
    pythoncom.CoInitialize()

    try:
        jincheng_acad = get_acad_process_id('acad.exe')
        handle = findwindows.find_windows(process=jincheng_acad)[0]

        app = Application().connect(handle=handle)
        time.sleep(2)

        try:
            window = app.window(title="图形导出", class_name="#32770")
        except Exception as e:
            print("尝试重新获取窗口:", e)
            time.sleep(2)
            window = app.window(title="图形导出", class_name="#32770")

        if window.exists():
            print("窗口存在")
            child_windows = window.children()

            for child_window in child_windows:
                try:
                    child_window.select("天正7文件 (*.dwg) ")
                    print("正在导出t7")
                    pass
                except Exception as e:
                    pass
        else:
            print("窗口不存在")
            pass

        save_button = window.child_window(title="保存(&S)", class_name="Button")

        try:
            save_button.set_focus()
            
            save_button.click()


        except Exception as e:
            print("pywinauto处理窗口出现问题:", e)
            time.sleep(2)
            save_button.set_focus()
            save_button.click()

        print("已使用pywinauto自动化窗口")
        pass

    except Exception as e:
        sys_logger.info(f"automate_window_with_pywinauto_t7 出现错误: {e}")
        pass
    finally:
        pythoncom.CoUninitialize()
        timeout_event.set()
        event.set()  # 通知主线程完成

#&&% # 天正转t7    
#&&% 转成T7
def zhuancheng_t7():
    
    """
    在li()连接激活文件后，直接执行该命令即可转换
    一般理解所得文件和当前文件同文件夹，但这次测试结果却在文件夹
    C:/Users/Administrator/Documents/WeChat Files/wxid_mhrzdlppqacc22/FileStorage/File/2025-04/
    """    
    t1 = time.time()

    chaoshibiaoji = 1

    # 创建一个Event，用于通知子线程终止
    timeout_event = threading.Event()
    
    # 创建一个Event，用于通知主线程完成
    event = threading.Event()

    # 创建两个线程，分别执行run_cad_program和automate_window_with_pywinauto_t7
    thread1 = threading.Thread(target=run_cad_program, args=(timeout_event, event))
    thread2 = threading.Thread(target=automate_window_with_pywinauto_t7, args=(timeout_event, event))

    thread1.start()
    thread2.start()

    thread1.join(timeout=180)
    thread2.join(timeout=180)

    if not event.is_set():
        print("操作超时，正在中断...")

        pass

        
        # 设置timeout_event，通知run_cad_program()终止
        timeout_event.set()

    print("文件转T7格式操作结束")

    t2 = time.time()
    print("文件转T7格式操作总共用时", t2 - t1, "秒")


# 创建一个事件对象

timeout_event = threading.Event()

event = threading.Event()


#&&% 自动化T3窗口
def automate_window_with_pywinauto_t3(timeout_event, event):
    pythoncom.CoInitialize()

    try:
        jincheng_acad = get_acad_process_id('acad.exe')
        handle = findwindows.find_windows(process=jincheng_acad)[0]

        app = Application().connect(handle=handle)
        time.sleep(2)

        try:
            window = app.window(title="图形导出", class_name="#32770")
        except Exception as e:
            print("尝试重新获取窗口:", e)
            time.sleep(2)
            window = app.window(title="图形导出", class_name="#32770")

        if window.exists():
            print("窗口存在")
            pass
            child_windows = window.children()

            for child_window in child_windows:
                try:
                    child_window.select("天正3文件 (*.dwg) ")
                    print("正在导出t3")
                    pass
                except Exception as e:
                    pass
        else:
            print("窗口不存在")
            pass

        save_button = window.child_window(title="保存(&S)", class_name="Button")

        try:
            save_button.set_focus()
            save_button.click()
        except Exception as e:
            print("pywinauto处理窗口出现问题:", e)
            pass
            time.sleep(2)
            save_button.set_focus()
            save_button.click()

        print("已使用pywinauto自动化窗口")
        pass

    except Exception as e:
        sys_logger.info(f"automate_window_with_pywinauto_t3 出现错误: {e}")
        pass
    finally:
        pythoncom.CoUninitialize()
        timeout_event.set()
        event.set()  # 通知主线程完成
#&&% # 天正转t3
        
#&&% 转成T3
def zhuancheng_t3():

    """
    在li()连接激活文件后，直接执行该命令即可转换
    一般理解所得文件和当前文件同文件夹，但这次测试结果却在文件夹
    C:/Users/Administrator/Documents/WeChat Files/wxid_mhrzdlppqacc22/FileStorage/File/2025-04/
    """    

    t1 = time.time()

    chaoshibiaoji = 1

    # 创建一个Event，用于通知子线程终止
    timeout_event = threading.Event()
    
    # 创建一个Event，用于通知主线程完成
    event = threading.Event()

    # 创建两个线程，分别执行run_cad_program和automate_window_with_pywinauto_t3
    thread1 = threading.Thread(target=run_cad_program, args=(timeout_event, event))
    thread2 = threading.Thread(target=automate_window_with_pywinauto_t3, args=(timeout_event, event))

    thread1.start()
    thread2.start()

    thread1.join(timeout=180)
    thread2.join(timeout=180)

    if not event.is_set():
        print("操作超时，正在中断...")

        pass

        
        # 设置timeout_event，通知run_cad_program()终止
        timeout_event.set()

    print("文件转T3格式操作结束")

    t2 = time.time()
    print("文件转T3格式操作总共用时", t2 - t1, "秒")














#&&&% 新加


#&&% 确保文件被删除
def ensure_file_absent(save_path: str, ty: float = 1.0) -> None:
    """
    确保指定路径的文件不存在。如果存在，则删除并等待 ty 秒。

    :param save_path: 带路径的文件名，例如 "D:/output/result.txt"
    :param ty: 删除后等待的秒数，默认为 1 秒
    """
    try:
        if os.path.isfile(save_path):
            os.remove(save_path)
            # 等待 ty 秒，以确保文件系统完成删除操作
            time.sleep(ty)
            sys_logger.info(f"[OK] 已删除文件：{save_path}，并等待了 {ty} 秒")
        else:
            sys_logger.info(f"ℹ 文件不存在，无需删除：{save_path}")
    except Exception as e:
        sys_logger.info(f"[错误] 删除文件时出错：{e}")


#&&% 遍历目录
def traverse_with_os_walk(root_dir: str):
    """
    遍历 root_dir 及其所有子目录，打印每个目录和文件的完整路径。
    """
    for dirpath, dirnames, filenames in os.walk(root_dir):
        # dirpath: 当前遍历到的目录路径
        sys_logger.info(f"Directory: {dirpath}")
        for dirname in dirnames:
            sys_logger.info(f"  Sub-dir: {os.path.join(dirpath, dirname)}")
        for filename in filenames:
            sys_logger.info(f"  File   : {os.path.join(dirpath, filename)}")



#&&% 按后缀查找文件
def find_files_with_extensions(directory, extensions):
    #找到以[".dwg"]结尾的文件directory为文件夹及其路径，extensions为后缀或中间位置字符列表
    matching_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if any(file.lower().endswith(ext) for ext in extensions):
                matching_files.append(os.path.join(root, file))
    return matching_files

#&&% 获取无后缀文件名
def get_filename_without_extension(FileandPath):
    # 从完整路径中获取文件名（包含扩展名）
    filename_with_extension = os.path.basename(FileandPath)
    
    # 分离文件名和扩展名
    filename_without_extension, _ = os.path.splitext(filename_with_extension)

    return filename_without_extension


#&&% 按模式删除文件
def delete_files_with_patterns(folder_path, patterns):
    """
    删除文件夹中符合指定模式的文件。

    Args:
        folder_path (str): 文件夹的路径。
        patterns (list): 包含要匹配的模式的列表，例如["_t7", "_t3", "_t8"]。

    Returns:
        None
    """
    # 遍历文件夹中的所有文件
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        
        # 检查文件名是否包含任何一个指定的模式
        if any(pattern in filename for pattern in patterns):
            # 如果符合条件，就删除该文件
            os.remove(file_path)
            sys_logger.info(f"Deleted file: {filename}")
            pass

    print("File deletion completed.")
    pass


#确保文件夹中名字含有特殊字符的文件被清空

#&&% 清除指定前缀文件
def clear_files_with_prefix(folder: str, filename_prefix: str = "区域导出", delay: float = 0.5):
    """
    清除指定文件夹中所有文件名包含给定前缀的文件。若文件正在被占用或删除失败，会尝试重试一次。

    :param folder:           要清理的文件夹路径
    :param filename_prefix:  需要匹配的文件名前缀（只要文件名中包含该字符串，就会被删除）
    :param delay:            删除失败时的等待时间（秒），默认 0.5 秒
    """
    if not os.path.isdir(folder):
        sys_logger.info(f"[错误] 目标路径不是有效文件夹：{folder}")
        return

    # 列出文件夹中所有条目
    entries = os.listdir(folder)
    # 过滤出文件名中包含指定前缀的文件
    to_delete = [fname for fname in entries if filename_prefix in fname and os.path.isfile(os.path.join(folder, fname))]

    if not to_delete:
        sys_logger.info(f"ℹ️ 文件夹中未发现文件名包含 “{filename_prefix}” 的文件。")
        return

    for fname in to_delete:
        full_path = os.path.join(folder, fname)
        try:
            os.remove(full_path)
            sys_logger.info(f"[OK] 已删除：{fname}")
        except Exception as e:
            sys_logger.info(f"[警告] 删除失败（第一次尝试）：{fname}，错误：{e}，稍后重试……")
            time.sleep(delay)
            # 再试一次
            try:
                os.remove(full_path)
                sys_logger.info(f"[OK] 重试成功删除：{fname}")
            except Exception as e2:
                sys_logger.info(f"[错误] 再次删除仍失败：{fname}，错误：{e2}")



#&&% 按字符串查找文件
def find_files_with_string(directory, search_string):
    #找到文件夹中含有指定字符串的文件
    matched_files = []
    for file in os.listdir(directory):
        if search_string in file:
            matched_files.append(file)
    return matched_files

#使路径名和文件名合并后合乎预期
#&&% 路径拼接
def join_paths(p1, p2):
    # 使用 os.path.join 合并路径
    result = os.path.join(p1, p2)
    # 替换反斜杠为正斜杠
    return result.replace("\\", "/")




#&&&&%%  第六部分 图块操作 


#&&&% 原块处理


#&&% 获取块实例块名

def get_block_name(obj):
    """获取块名，兼容动态块(EffectiveName)"""
    try:
        # 优先尝试获取动态块的真实名称
        return getattr(obj, "EffectiveName", obj.Name)
    except:
        return getattr(obj, "Name", "")


#&&% 获取块属性值
def huoqukuai_shuxing_zhi(XX):#XX为属性块实体


    attributes = XX.GetAttributes()

    tags=[]
    values=[]

    for attr in attributes:
               
        tag = attr.TagString

            
        value = attr.TextString

        tags.append(tag)

        values.append(value)

    return tags,values



#&&% 属性块标签编辑

def update_block_def_attributes_safe(
    block_ref_or_name, 
    target_tag, 
    *, 
    style=None,            
    height=None,           
    width_factor=None,     
    rotation_deg=None,     
    justify=None,         
    align_point=None,     
    boundary_width=None,  
    verbose=True
):
    """
    【函数编号】: BLK-007-Safe
    【功能描述】: 
        修改块定义属性（完全使用 get_attr/set_attr 封装，禁止直接属性访问）。
        保留了 V7 的“对正优先”逻辑。
    """
    doc=C.doc


    # ================= 1. 获取块定义名称 =================
    block_name = None
    if isinstance(block_ref_or_name, str):
        block_name = block_ref_or_name
    else:
        # 替换 .EffectiveName 和 .Name
        block_name = get_attr(block_ref_or_name, 'EffectiveName')
        if not block_name:
            block_name = get_attr(block_ref_or_name, 'Name')

    if not block_name:
        sys_logger.info(f"[错误] 无法解析块名")
        return False

    if verbose:
        sys_logger.info(f"\n=== [修改块定义] 目标: '{block_name}' / 标签: '{target_tag}' ===")

    try:
        # 获取 Blocks 集合 -> Item
        blocks_coll = get_attr(doc, 'Blocks')
        block_def = blocks_coll.Item(block_name)
    except Exception as e:
        sys_logger.info(f"[错误] 找不到块定义: {e}")
        return False

    target_def = None
    
    # ================= 2. 扫描并锁定属性定义对象 =================
    # 遍历块定义中的实体
    for entity in block_def:
        # 替换 entity.ObjectName
        obj_name = get_attr(entity, 'ObjectName')
        
        if obj_name == "AcDbAttributeDefinition":
            # 尝试转换接口
            attr_obj = cast_object(entity)
            
            # 替换 attr_obj.TagString
            current_tag = get_attr(attr_obj, 'TagString')
            
            if current_tag == target_tag:
                target_def = attr_obj
                break
    
    if target_def is None:
        if verbose: sys_logger.info(f"[失败] 未找到标签 '{target_tag}'。")
        return False

    # ================= 3. 执行修改 (使用 set_attr) =================
    any_change = False
    
    try:
        # —————————— [步骤 1] 优先设置对正 (Alignment) ——————————
        if justify is not None:
            # 替换 target_def.Alignment = int(justify)
            if set_attr(target_def, 'Alignment', int(justify)):
                any_change = True
                if verbose: sys_logger.info(f"  - [设置] 对正模式: {justify}")
            else:
                sys_logger.info(f"[警告] 设置对正失败")

        # —————————— [步骤 2] 基础外观属性 ——————————
        if style is not None:
            if set_attr(target_def, 'StyleName', str(style)):
                any_change = True

        if height is not None:
            if set_attr(target_def, 'Height', float(height)):
                any_change = True

        if width_factor is not None:
            if set_attr(target_def, 'ScaleFactor', float(width_factor)):
                any_change = True

        if rotation_deg is not None:
            rad = float(rotation_deg) * math.pi / 180.0
            if set_attr(target_def, 'Rotation', rad):
                any_change = True

        # —————————— [步骤 3] 多行属性与边界 (MText) ——————————
        if boundary_width is not None:
            # 1. 检测当前状态: get_attr(target_def, 'MTextAttribute')
            is_multiline = False
            current_mtext_attr = get_attr(target_def, 'MTextAttribute')
            if current_mtext_attr: 
                is_multiline = True
            
            # 2. 强制开启多行
            if not is_multiline:
                if set_attr(target_def, 'MTextAttribute', True):
                    is_multiline = True

            # 3. 设置宽度
            if is_multiline:
                if set_attr(target_def, 'MTextBoundaryWidth', float(boundary_width)):
                    any_change = True
                    if verbose: sys_logger.info(f"  - [设置] MText边界宽: {boundary_width}")
            else:
                if verbose: sys_logger.info(f"  - [忽略] 无法开启多行模式")

        # —————————— [步骤 4] 坐标移动 (.Move) ——————————
        # Move 是方法，不是属性，但我们需要先获取基点属性
        if align_point is not None:
            if len(align_point) == 2:
                tgt_x, tgt_y, tgt_z = float(align_point[0]), float(align_point[1]), 0.0
            else:
                tgt_x, tgt_y, tgt_z = float(align_point[0]), float(align_point[1]), float(align_point[2])
            
            # 获取当前对正
            current_align = get_attr(target_def, 'Alignment')
            
            start_raw = None
            pt_type = "插入点"
            
            if current_align == 0: # Left
                start_raw = get_attr(target_def, 'InsertionPoint')
            else:
                start_raw = get_attr(target_def, 'TextAlignmentPoint')
                pt_type = "对齐点"
            
            if start_raw:
                start_x, start_y, start_z = start_raw[0], start_raw[1], start_raw[2]

                if verbose:
                    sys_logger.info(f"  - [移动] 以{pt_type}为基准: {start_raw} -> ({tgt_x}, {tgt_y}, {tgt_z})")

                if "vtpnt" in globals():
                    p1_variant = vtpnt(start_x, start_y, start_z)
                    p2_variant = vtpnt(tgt_x, tgt_y, tgt_z)
                    # Move 是方法，直接调用 COM 对象的方法
                    target_def.Move(p1_variant, p2_variant)
                    any_change = True
                else:
                    print("[警告] 缺少 vtpnt 函数")


    except Exception as e:
        sys_logger.info(f"[错误] 修改属性时发生异常: {e}")
        return False

    if any_change:
        if verbose: print("[成功] 属性定义已更新。")
        return True
    return False


def update_block_def_attributes_v7(
    block_ref_or_name, 
    target_tag, 
    *, 
    style=None,           
    height=None,          
    width_factor=None,    
    rotation_deg=None,    
    justify=None,         # 对正 (0-14)
    align_point=None,     # 目标坐标
    boundary_width=None,  # 边界宽度
    verbose=True
):
    """
    【函数编号】: BLK-007
    【所属模块】: 块定义管理模块 (Block Definition)
    【功能描述】: 
        修改图块的“底层定义”(Block Definition) 中的属性特性。
        用于统一修正图签的文字高度、对正方式、MText 边界宽等格式。
        
        V7 版本核心修复：
        1. 顺序免疫：强制将 'Alignment (对正)' 的设置提前到 'MText (边界宽)' 之前。
           (原因：一旦先开启 MText 属性，AutoCAD 往往会锁定 Alignment 属性导致报错)
        2. 坐标修正：保留 .Move() + vtpnt 逻辑，确保基点正确移动。

    【参数详解】:
        - block_ref_or_name (str/Obj): 块名字符串或块引用对象。
        - target_tag (str): 目标属性标签 (如 "图纸名称")。
        - style (str): 文字样式名。
        - height (float): 文字高度。
        - width_factor (float): 宽度因子 (0.7, 0.8 等)。
        - rotation_deg (float): 旋转角度 (度)。
        - justify (int): 对正方式 (0=Left, 1=Center, ...)。
        - align_point (list/tuple): 对齐点/插入点坐标 [x, y]。
        - boundary_width (float): MText 换行边界宽度。若设置此值，会自动开启多行模式。
        - verbose (bool): 是否打印详细日志。

    【返回值】:
        - bool: 修改成功返回 True，无变化或失败返回 False。


    """
    C.li()
    
    # --- 1. 获取块名 ---
    if isinstance(block_ref_or_name, str):
        block_name = block_ref_or_name
    else:
        try: block_name = block_ref_or_name.EffectiveName
        except: block_name = block_ref_or_name.Name

    if verbose:
        sys_logger.info(f"\n=== [修改块定义] 目标: '{block_name}' / 标签: '{target_tag}' ===")

    try:
        block_def = doc.Blocks.Item(block_name)
    except Exception as e:
        sys_logger.info(f"[错误] 找不到块定义: {e}")
        return False

    target_def = None
    
    # --- 2. 扫描并锁定对象 ---
    for entity in block_def:
        if entity.ObjectName == "AcDbAttributeDefinition":
            try:
                attr_obj = win32com.client.CastTo(entity, "IAcadAttribute")
                if not hasattr(attr_obj, "TagString"):
                     attr_obj = win32com.client.dynamic.Dispatch(entity)
            except:
                continue

            if attr_obj.TagString == target_tag:
                target_def = attr_obj
                break
    
    if target_def is None:
        if verbose: sys_logger.info(f"[失败] 未找到标签 '{target_tag}'。")
        return False

    # --- 3. 执行修改 (注意顺序！) ---
    any_change = False
    
    try:
        # =======================================================
        # [步骤 1] 优先设置对正 (Alignment)
        # =======================================================
        # 在开启 MText 之前设置对正，兼容性最好
        if justify is not None:
            try:
                target_def.Alignment = int(justify)
                any_change = True
                if verbose: sys_logger.info(f"  - [设置] 对正: {justify}")
            except Exception as e:
                sys_logger.info(f"[警告] 设置对正失败: {e}")

        # =======================================================
        # [步骤 2] 基础属性
        # =======================================================
        if style is not None:
            target_def.StyleName = str(style)
            any_change = True

        if height is not None:
            target_def.Height = float(height)
            any_change = True

        if width_factor is not None:
            target_def.ScaleFactor = float(width_factor)
            any_change = True

        if rotation_deg is not None:
            target_def.Rotation = float(rotation_deg) * math.pi / 180.0
            any_change = True

        # =======================================================
        # [步骤 3] 多行属性与边界 (MText)
        # =======================================================
        if boundary_width is not None:
            # 1. 检测当前状态
            is_multiline = False
            try:
                if target_def.MTextAttribute: is_multiline = True
            except: pass
            
            # 2. 如果之前不是多行，现在强制开启
            if not is_multiline:
                try:
                    target_def.MTextAttribute = True
                    is_multiline = True
                except:
                    pass

            # 3. 设置宽度
            if is_multiline:
                target_def.MTextBoundaryWidth = float(boundary_width)
                any_change = True
                if verbose: sys_logger.info(f"  - [设置] 边界宽: {boundary_width}")
            else:
                if verbose: sys_logger.info(f"  - [忽略] 无法开启多行模式，忽略边界宽度")

        # =======================================================
        # [步骤 4] 坐标移动 (.Move)
        # =======================================================
        if align_point is not None:
            # 准备数据
            if len(align_point) == 2:
                tgt_x, tgt_y, tgt_z = float(align_point[0]), float(align_point[1]), 0.0
            else:
                tgt_x, tgt_y, tgt_z = float(align_point[0]), float(align_point[1]), float(align_point[2])
            
            # 智能判断基点 (基于最新的对正方式)
            current_align = target_def.Alignment
            if current_align == 0: # Left
                start_raw = target_def.InsertionPoint 
                pt_type = "插入点"
            else:
                start_raw = target_def.TextAlignmentPoint
                pt_type = "对齐点"
            
            start_x, start_y, start_z = start_raw[0], start_raw[1], start_raw[2]

            if verbose:
                sys_logger.info(f"  - [移动] 以{pt_type}为基准: {start_raw} -> ({tgt_x}, {tgt_y}, {tgt_z})")

            # 类型转换 (vtpnt)
            p1_variant = vtpnt(start_x, start_y, start_z)
            p2_variant = vtpnt(tgt_x, tgt_y, tgt_z)

            # 执行移动
            target_def.Move(p1_variant, p2_variant)
            any_change = True

    except Exception as e:
        sys_logger.info(f"[错误] 修改属性时发生异常: {e}")
        return False

    if any_change:
        if verbose: print("[成功] 属性定义已更新。请执行 ATTSYNC 刷新实例。")
        return True
    else:
        if verbose: print("[提示] 无变化。")
        return False

#&&% 属性块标签编辑生效

def attsync_block_instance(block_ref_obj):
    """
    【函数编号】: CMD-001
    【功能描述】: 
        强力同步属性块。
        连续执行 3 次 Base 操作，确保 AutoCAD 彻底刷新属性位置。
    """
    success_at_least_once = False
    
    # 获取块名用于日志（仅用于显示，失败不影响流程）
    try:
        b_name = get_attr(block_ref_obj,"Name")
    except:
        b_name = "未知块"

    sys_logger.info(f"🔄 [强力同步] 正在对 {b_name} 执行 3 轮 ATTSYNC...")

    for i in range(2):
        # 执行底层同步
        result = attsync_block_instance_base(block_ref_obj)
        
        if result:
            success_at_least_once = True
            # sys_logger.info(f"  -> 第 {i+1}/3 次指令已发送")
        
        # 【关键】稍微暂停，防止命令在 CAD 命令行中堆叠过快导致 "未知命令" 错误
        # 0.2秒通常足够 CAD 处理完上一条指令的提示符
        time.sleep(2)

    return success_at_least_once

def attsync_block_instance_base(block_ref_obj):
    """
    【函数编号】: CMD-001
    【功能描述】: 
        对指定的属性块执行 ATTSYNC (属性同步) 操作。
        用于在修改了块定义的属性位置/格式后，强制刷新实例显示。
    
    【参数】:
        - block_ref_obj: 块引用对象 (BlockReference)。
    
    【返回值】:
        - bool: 发送成功返回 True，失败返回 False。
    """
    # 0. 尝试刷新环境 (如果 li 是全局函数)
    try: C.li() 
    except: pass

    try:
        # 1. 获取块名 (用于 ATTSYNC 参数)
        # 优先取 EffectiveName (动态块)，失败则取 Name
        block_name = None
        if hasattr(block_ref_obj, 'EffectiveName'):
            block_name = block_ref_obj.EffectiveName
        if not block_name:
            block_name = getattr(block_ref_obj, 'Name', None)

        if not block_name:
            print("[错误] 无法获取块名，跳过同步。")
            return False

        sys_logger.info(f"🔄 [同步] 正在刷新图块实例: {block_name} ...")

        # 2. 构造命令字符串 (ATTSYNC -> Name -> 块名)
        cmd_str = f"_ATTSYNC\nN\n{block_name}\n"
        
        # 3. 发送命令
        # 【优化】优先使用对象自身的 Document 属性，比全局 doc 更安全
        target_doc = getattr(block_ref_obj, 'Document', None)

        if target_doc:
            target_doc.SendCommand(cmd_str)
            return True
        else:
            # 这里的 else 对应 "not target_doc"
            # 保底方案：使用全局单例 C.doc (它会自动重连)
            C.doc.SendCommand(cmd_str)
            return True

    except Exception as e:
        sys_logger.info(f"[警告] ATTSYNC 执行失败: {e}")
        return False



#&&% 设置属性块的标签值
def set_attribute_mtext(block, tags, new_texts, keep_prefix=True, verbose=True):
    """
    set_attribute_mtext(p[0],"图纸规格","A0")
    
    set_attribute_mtext(p[0],"项目名称",["某某未来城","工业园1#"])

    notes = ["1. 尺寸单位为毫米", "2. 未注公差为IT12"]
    dotes = ["县", "城乡住房和建设局"]
    set_attribute_mtext(p[0],["建设单位名称","图纸名称"],[notes,dotes]) 

    set_attribute_mtext(p[0],["专业名称","出图比例","设计编号"], ["结构","1：20","sjy-01"], keep_prefix=True, verbose=True)

    X混搭不仅仅无效还会破坏set_attribute_mtext(p[0],"图纸规格","项目名称",["A3",["远程国际","主楼"]])

    """
    C.li()
    # --- 1. Normalize Inputs ---
    # Ensure tags is a list
    if isinstance(tags, str):
        tag_list = [tags]
        single_tag_mode = True
    else:
        tag_list = list(tags)
        single_tag_mode = False

    # Ensure new_texts matches the structure of tag_list
    text_list = []
    if single_tag_mode:
        # If single tag, new_texts is the content for that tag
        text_list = [new_texts]
    else:
        # If multiple tags, new_texts must be a list corresponding to tags
        if not isinstance(new_texts, (list, tuple)):
            # Broadcast same text to all tags
            text_list = [new_texts] * len(tag_list)
        else:
            text_list = list(new_texts)
            # Pad or truncate to match tag_list length
            if len(text_list) < len(tag_list):
                text_list.extend([text_list[-1]] * (len(tag_list) - len(text_list)))
            elif len(text_list) > len(tag_list):
                text_list = text_list[:len(tag_list)]

    # --- 2. Get Attribute Objects ---
    block = cast_object(block)
    try:
        raw_attrs = block.GetAttributes()
    except Exception as e:
        if verbose:
            sys_logger.info(f"[Error] Failed to get attributes for block ({get_attr(block, 'Handle')}): {e}")
        return {tag: False for tag in tag_list}

    # Map TagString to Attribute Object (case-insensitive for robustness)
    attr_map = {}
    for ra in raw_attrs:
        attr = cast_object(ra)
        tag_str = get_attr(attr, "TagString")
        if tag_str:
            attr_map[tag_str] = attr
            attr_map[tag_str.upper()] = attr

    result = {}

    # --- 3. Process Each Attribute ---
    for tag, content in zip(tag_list, text_list):
        attr = attr_map.get(tag) or attr_map.get(tag.upper())
        
        if attr is None:
            if verbose: sys_logger.info(f"[Warning] Attribute tag '{tag}' not found.")
            result[tag] = False
            continue

        try:
            # --- A. Prepare Text Content ---
            is_multiline_content = isinstance(content, (list, tuple))
            
            if is_multiline_content:
                # Join list elements with AutoCAD's paragraph break code "\P"
                # Ensure all elements are strings
                clean_lines = [str(line) for line in content]
                body_text = "\\P".join(clean_lines)
            else:
                body_text = str(content)

            # --- B. Handle Prefix (Formatting Codes) ---
            old_text = get_attr(attr, "TextString") or ""
            prefix = ""
            if keep_prefix and old_text:
                # Look for MText formatting end marker ';' (e.g., \W0.8; or \C1;)
                # Simple heuristic: take everything up to the first semicolon if it looks like a format code
                # A more robust check might look for specific start characters like backslash
                if old_text.startswith("\\") and ";" in old_text:
                     pos = old_text.find(";")
                     prefix = old_text[:pos + 1]
            
            final_text = prefix + body_text

            # --- C. Configure Attribute Mode (Single vs Multi-line) ---
            # If the content implies multi-line (list input or contains \P), ensure MText mode
            # Note: Checking for "\\P" in string input allows users to manually pass multi-line strings
            requires_mtext = is_multiline_content or "\\P" in body_text
            
            # Check current mode
            try:
                is_currently_mtext = get_attr(attr, "MTextAttribute")
            except:
                is_currently_mtext = False

            if requires_mtext and not is_currently_mtext:
                # Try to enable MText mode if content requires it
                if not set_attr(attr, "MTextAttribute", True):
                    if verbose: sys_logger.info(f"[Info] Could not enable MText mode for '{tag}'. Text may display on one line.")
            
            # --- D. Set Text String ---
            if set_attr(attr, "TextString", final_text):
                if verbose: 
                    display_text = final_text if len(final_text) < 20 else final_text[:17] + "..."
                    sys_logger.info(f"[Success] Set '{tag}': {display_text}")
                result[tag] = True
            else:
                if verbose: sys_logger.info(f"[Error] Failed to set TextString for '{tag}'")
                result[tag] = False

            # Update the attribute entity
            attr.Update()

        except Exception as e:
            if verbose: sys_logger.info(f"[Error] Exception setting '{tag}': {e}")
            result[tag] = False

    # Update the block reference to reflect changes
    try:
        block.Update()
    except:
        pass

    return result





#&&% 获取属性块标签及标签值
def get_block_attributes_dict(
    block_ref,
    ignore_empty: bool = False,
    upper_tag: bool = True,
):
    """
    获取块参照 block_ref 的所有属性，返回 {标签: 纯文本值} 的字典。

    特别规则：
        若属性值形如 '\\W0.8000;1#楼'，则取 ';' 后面的部分 '1#楼' 作为真实值。
        即：总是优先使用第一个分号 ';' 后面的内容作为纯文本值，
        这用于剥离 MTEXT 格式控制前缀（如宽度控制 \\W0.8000; 等）。

    参数:
        block_ref   : IAcadBlockReference 实例
        ignore_empty: True 时忽略空字符串 / 全空白值
        upper_tag   : True 时将属性标签名统一转大写作为字典 key

    返回:
        dict，如:
        {
            "项目名称": "未来城",
            "图纸名称": "1#楼",
            ...
        }
    """

    attrs_dict: dict[str, str] = {}

    if block_ref is None:
        return attrs_dict

    # 不是块参照，直接返回空
    try:
        obj_name = getattr(block_ref, "ObjectName", "")
    except Exception:
        obj_name = ""
    if "BlockReference" not in str(obj_name):
        return attrs_dict

    # 没有属性，返回空
    try:
        has_attrs = getattr(block_ref, "HasAttributes", False)
    except Exception:
        has_attrs = False
    if not has_attrs:
        return attrs_dict

    try:
        att_refs = block_ref.GetAttributes()
    except Exception:
        return attrs_dict

    # 小工具：剥离 MTEXT 前缀，取分号后面的“真实值”
    def _clean_value(val):
        """把 '\\W0.8000;1#楼' → '1#楼'；其他情况尽量保持原值。"""
        if not isinstance(val, str):
            # 非字符串就直接转成 str
            return "" if val is None else str(val)

        # 统一去掉首尾空白，避免 '\W0.8; 1#楼' 之类
        s = val.strip()

        # 只处理以 '\' 开头且包含 ';' 的情况，避免误伤正常文本里本来就有分号的
        if s.startswith("\\") and ";" in s:
            # 只切第一个分号
            _, right = s.split(";", 1)
            s = right

        # 再 strip 一遍，保证结果干净
        return s.strip()

    # 主循环：收集所有属性
    for att in att_refs:
        try:
            tag = att.TagString
            raw_val = att.TextString
        except Exception:
            continue

        if not isinstance(tag, str):
            continue

        # 标签名大小写控制
        tag_key = tag.upper() if upper_tag else tag

        # 清洗值：去掉 MTEXT 控制前缀，得到纯文本
        val = _clean_value(raw_val)

        if ignore_empty and (val is None or str(val).strip() == ""):
            continue

        attrs_dict[tag_key] = val

    return attrs_dict
#&&% 筛选出指定块名外的对象

def separate_entities_by_block_names(entities, target_names):
    """
    将实体列表分为两组：
    1. 命中组：名字在 target_names 中的块实例。
    2. 保留组：名字不在 target_names 中的块，以及所有非块对象（线、圆等）。
    
    参数:
    entities: 待处理的对象列表
    target_names: 目标块名，可以是单个字符串 "A3-H"，也可以是列表 ["A3-H", "MyBlock"]
    
    依赖:
    外部函数 get_attr(obj, name)
    
    返回:
    (kept_entities, target_blocks) 元组
    """
    
    # 1. 参数标准化：将输入转为集合(set)，提高查找速度，并兼容单字符串输入
    if isinstance(target_names, str):
        target_set = {target_names}
    else:
        # 过滤掉 None 或空值，防止报错
        target_set = set(n for n in target_names if n)

    kept_entities = []   # 存放：非目标对象（保留下来的）
    target_blocks = []   # 存放：目标块实例（被选出来的）

    for obj in entities:
        is_hit = False
        
        # 1. 安全获取对象类型
        obj_type = get_attr(obj, 'ObjectName')
        
        # 2. 只有是块引用时，才去比对名字
        if obj_type == "AcDbBlockReference":
            e_name = get_attr(obj, 'EffectiveName')
            name = get_attr(obj, 'Name')
            
            # 3. 只要 EffectiveName 或 Name 任意一个命中目标集合，就算命中
            # (处理动态块兼容性)
            if (e_name in target_set) or (name in target_set):
                is_hit = True
        
        # 4. 根据命中结果分流
        if is_hit:
            target_blocks.append(obj)
        else:
            kept_entities.append(obj)
            
    return kept_entities, target_blocks




#&&% 获取块内多段线
def huoqu_kuai_pl(blocka):#输入实体块，得到实体块中多段线矩形的坐标，其坐标以插入点的定义点为原点
    # 连接到AutoCAD
    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    kuaiming=blocka.Name

    # 获取块定义
    block_def = doc.Blocks.Item(str(kuaiming))

    # 获取块定义中的所有对象
    block_objects = list(block_def)

    # 查找三角形并删除
    for obj in block_objects:

##        print(obj.ObjectName)
        
        if obj.ObjectName == "AcDbPolyline":

            print(obj.Coordinates)





#定义基点的块

#&&% 创建带基点块
def create_block_with_basepoint():
    # 连接到AutoCAD
    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # 定义块的基点位置
    base_point = vtpnt(10, 10, 0)

    # 创建一个新的块
    block = doc.Blocks.Add(base_point, "MyBlock")

    # 在块中添加一个圆形实体
    block.AddCircle(base_point, 5)

#块的添加


#&&% 创建三角形文字块
def create_block_with_triangle_and_text():
    # 连接到AutoCAD
    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # 创建新块
    grip = vtpnt(0, 0)
    blockObj = doc.Blocks.Add(grip, "MyBlock")

    # 在块中添加三角形
    pt1 = vtpnt(0, 0, 0)
    pt2 = vtpnt(10, 0, 0)
    pt3 = vtpnt(5, 10, 0)
    blockObj.AddLine(pt1, pt2)
    blockObj.AddLine(pt2, pt3)
    blockObj.AddLine(pt3, pt1)

    # 在块中添加文字对象
    text_point = vtpnt(2, 2, 0)
    blockObj.AddText("太美了", text_point, 2)

    print("块 'MyBlock' 创建成功")



def huoqu_kuai_pl(blocka):
    # 连接到AutoCAD
    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    kuaiming=blocka.Name

    # 获取块定义
    block_def = doc.Blocks.Item(str(kuaiming))

    # 获取块定义中的所有对象
    block_objects = list(block_def)

    # 查找三角形并删除
    for obj in block_objects:

        print(obj.ObjectName)
        if obj.ObjectName == "AcDbPolyline":

            print(obj.Coordinates)

    

# 块的边界

#&&% 获取块包围盒
def get_bounding_box_of_block(block_name):
    # 连接到AutoCAD
    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # 获取块定义
    block_def = doc.Blocks.Item(block_name)

    min_x, min_y, min_z = float('inf'), float('inf'), float('inf')
    max_x, max_y, max_z = float('-inf'), float('-inf'), float('-inf')

    # 遍历块定义中的所有对象
    for obj in block_def:
        try:
            # 尝试获取对象的边界框
            lower_left, upper_right = obj.GetBoundingBox()
            
            min_x = min(min_x, lower_left[0])
            min_y = min(min_y, lower_left[1])
            min_z = min(min_z, lower_left[2])
            
            max_x = max(max_x, upper_right[0])
            max_y = max(max_y, upper_right[1])
            max_z = max(max_z, upper_right[2])
        except:
            pass

    return ((min_x, min_y, min_z), (max_x, max_y, max_z))


#&&% 创建含插入和直线的块
def create_new_block_with_insert_and_line():
    # 连接到AutoCAD
    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc = acad.ActiveDocument

    # 检查块名称"块1"是否已经存在
    if "块3" in [blk.Name for blk in doc.Blocks]:
        print("块名称'块3'已经存在。请选择一个新的名称或删除现有的块。")
        return

    # 创建新块的插入点
    grip = vtpnt(0, 0, 0)
    blockObj1 = doc.Blocks.Add(grip, "块3")

    # 在块1中插入MyBlock块
    insertion_point_for_myblock = vtpnt(10, 10, 0)
    blockObj1.InsertBlock(insertion_point_for_myblock, "MyBlock", 1, 1, 1, 0)

    # 在块1中添加一根直线段
    start_point = vtpnt(0, 0, 0)
    end_point = vtpnt(50, 50, 0)
    blockObj1.AddLine(start_point, end_point)

    print("块1已创建并添加了MyBlock和直线段")


    
#&&% 复制并移动图层块
def copy_and_move_blocks_from_layer(layer_name, block_prefix):
    
        
    # 使用select_tuceng函数选择指定图层上的所有对象
    all_objects = select_tuceng(layer_name)
    
    # 过滤出块对象，且块名的前两个字母与指定的前缀匹配
    blocks = [obj for obj in all_objects if obj.ObjectName == "AcDbBlockReference" and obj.Name[:2] == block_prefix]
    
    # 定义移动的起始点和结束点
    vtpnt_from = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0, 0, 0])
    vtpnt_to = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0, 2000000, 0])
    
    # 对每个块进行复制和移动操作
    for block in blocks:
        # 复制块
        copied_block = block.Copy()
        
        # 移动复制的块
        copied_block.Move(vtpnt_from, vtpnt_to)

    sys_logger.info(f"Copied and moved {len(blocks)} blocks from layer {layer_name} with prefix {block_prefix}.")

#块名的清除

#&&&% 删除指定块名的实例及块

#&&% 旧版
def delete_block_instances_and_definition_retry(target_block_name, max_rounds=3):
    """
    删除指定名称的块实例和块定义，带重试机制。
    """
    _, doc = get_acad_doc()
    if not doc:
        print("❌ 无法获取 AutoCAD 文档对象")
        return

    sys_logger.info(f"\n🚀 开始清理块: [{target_block_name}] (最大尝试 {max_rounds} 轮)")

    for round_idx in range(1, max_rounds + 1):
        sys_logger.info(f"--- 第 {round_idx} 轮清理 ---")
        
        # 1. 获取所有块实例
        # 注意：select_kuai 返回的是所有 INSERT，我们需要自己筛选
        all_blocks = select_kuai(max_retries=3)
        
        # 2. 筛选出目标块
        target_instances = []
        for b in all_blocks:
            # 比较名称（忽略大小写）
            if get_block_name(b).upper() == target_block_name.upper():
                target_instances.append(b)
        
        count = len(target_instances)
        
        # 3. 验证与删除实例
        if count == 0:
            sys_logger.info(f"✅ 第 {round_idx} 轮检测：未发现块实例 [{target_block_name}]")
        else:
            sys_logger.info(f"🔨 发现 {count} 个实例，正在删除...")
            for obj in target_instances:
                safe_delete(obj)
            
            # 删除后简单验证一下，虽然下一轮循环会再次全面检查
            sys_logger.info(f"   已执行删除命令。")

        # 4. 尝试删除块定义 (Purge)
        # 只有当实例数为0（或刚删完）时，才尝试删定义
        definition_deleted = False
        try:
            # 从块表中获取块定义
            blk_def = doc.Blocks.Item(target_block_name)
            blk_def.Delete()
            sys_logger.info(f"🗑️ 块定义 [{target_block_name}] 已从块表中移除。")
            definition_deleted = True
        except Exception as e:
            # 常见错误：eKeyNotFound(块不存在), eBoInUse(还有实例或被嵌套)
            msg = str(e)
            if "eKeyNotFound" in msg:
                sys_logger.info(f"ℹ️ 块定义 [{target_block_name}] 已不存在。")
                definition_deleted = True # 视为成功
            elif "eBoInUse" in msg:
                sys_logger.info(f"⚠️ 无法删除定义：块仍被使用（可能被嵌套在其他块中）。")
            else:
                pass # 其他错误忽略

        # 5. 最终判定：如果没实例了，且定义也没了(或本就不存在)，则成功退出
        if count == 0 and definition_deleted:
            sys_logger.info(f"✨ 彻底清理完成：[{target_block_name}]")
            return True
        
        # 如果还有问题，等待一小会儿进入下一轮
        if round_idx < max_rounds:
            time.sleep(1)

    sys_logger.info(f"❌ 经过 {max_rounds} 轮尝试，未能彻底清除（可能是嵌套块导致）。")
    return False

#&&% 极速清理

def delete_block_instances_and_definition_optimized(target_name, max_retries=5):
    """
    【函数编号】: CLEAN-ROBUST-V34 (核验版)
    【功能】: 
        删除指定图块的实例和定义。
        核心改进：增加【块表反查】。不仅要实例删光，还要确认 Block Table 里彻底查无此人。
    """
    doc = C.doc
    
    # 0. 先预检：如果块表里压根没有，直接返回成功，省得费劲
    try:
        doc.Blocks.Item(target_name)
    except:
        # sys_logger.info(f"ℹ️ [预检] {target_name} 根本不存在，无需清理。")
        return True

    for attempt in range(1, max_retries + 1):
        try:
            # ====================================================
            # 阶段 A: 清理实例 (Model Space) - 过滤器极速版
            # ====================================================
            ss_name = "RobustDelete_SS"
            try: doc.SelectionSets.Item(ss_name).Delete()
            except: pass
            ss = doc.SelectionSets.Add(ss_name)
            
            p_filter_type = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, [0, 2])
            p_filter_data = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, ["INSERT", target_name])
            
            ss.Select(5, None, None, p_filter_type, p_filter_data) # 5=All
            
            if ss.Count > 0:
                ss.Erase() # 删除
                # 【关键】等待操作完成
                wait_command_done()
            
            ss.Delete()

            # ====================================================
            # 阶段 B: 清理定义 (Blocks Table)
            # ====================================================
            try:
                blk_def = doc.Blocks.Item(target_name)
                blk_def.Delete()
            except:
                # 这里报错很正常（比如被嵌套引用），先不管，下面统一验尸
                pass

            # ====================================================
            # 阶段 C: 【核心】块表反查验尸
            # ====================================================
            # 无论刚才有没有报错，我现在去查户口。
            # 如果还能查到，说明没死透；如果报错，说明彻底没了。
            try:
                # 尝试再次获取
                _ = doc.Blocks.Item(target_name)
                
                # 🔴 还能获取到？说明清理失败！
                # sys_logger.info(f"   ⚠️ 第 {attempt} 轮：[{target_name}] 定义仍存活 (可能被嵌套)。重试中...")
                
                # 强制刷新，试图断开引用锁
                if attempt % 2 == 0: # 偶数轮次尝试 Regen
                    doc.Regen(1)
                
                time.sleep(0.5)
                continue # 进入下一轮循环
                
            except Exception:
                # 🟢 报错了？说明 Block Table 里没有这个 key 了！
                sys_logger.info(f"🗑️ [清理成功] {target_name} 已彻底根除。")
                return True

        except Exception as e:
            sys_logger.info(f"   ⚠️ 异常: {e}")
            time.sleep(0.5)
            
    # 如果跑完循环还在
    sys_logger.info(f"❌ [清理失败] {target_name} 顽固残留 (检查是否被其他块嵌套引用)。")
    return False

#&&% 再次优化

def delete_block_instances_and_definition_optimized(target_name, max_retries=5):
    """
    【最终推荐版】
    利用选择集极速清理实例，利用验尸逻辑清理定义。
    """
    doc = C.doc
    
    # 0. 预检：如果块表里压根没有，直接返回成功
    try:
        doc.Blocks.Item(target_name)
    except Exception:
        # 查无此人，直接通过
        return True

    sys_logger.info(f"🚀 [极速清理] 开始移除: {target_name}")

    for attempt in range(1, max_retries + 1):
        try:
            # ====================================================
            # 阶段 A: 清理实例 (Model Space) - 过滤器极速版
            # ====================================================
            ss_name = "Clean_Temp_SS"
            
            # 【安全逻辑】先删再建，防止上次崩了残留
            try: doc.SelectionSets.Item(ss_name).Delete()
            except: pass
            
            ss = doc.SelectionSets.Add(ss_name)
            
            # 构建过滤器：只选 块参照(INSERT) 且 块名(2)=target_name
            # 注意：win32com 需要用 VARIANT 包装数组
            p_filter_type = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, [0, 2])
            p_filter_data = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, ["INSERT", target_name])
            
            # acSelectionSetAll = 5 (全图搜索，包含屏幕外的)
            ss.Select(5, None, None, p_filter_type, p_filter_data) 
            
            count = ss.Count
            if count > 0:
                ss.Erase() # 极速删除
                sys_logger.info(f"  🔨 第 {attempt} 轮：已销毁 {count} 个实例引用。")
            
            # 用完即删，保持卫生
            ss.Delete() 

            # ====================================================
            # 阶段 B: 清理定义 (Blocks Table)
            # ====================================================
            try:
                blk_def = doc.Blocks.Item(target_name)
                blk_def.Delete()
            except Exception:
                # 这里如果报错，通常是因为有嵌套引用，或者已经删掉了
                # 不要在这里做判断，去阶段 C 统一验证
                pass

            # ====================================================
            # 阶段 C: 【核心】块表反查验尸
            # ====================================================
            try:
                # 再次尝试获取，如果还能拿到，说明没死透
                _ = doc.Blocks.Item(target_name)
                
                # 🔴 还能活着走到这一步，说明清理失败
                # sys_logger.warning(f"  ⚠️ 定义仍存活，尝试强制刷新...")
                
                # 只有在删除失败时才重生成，节省时间
                doc.Regen(1) 
                time.sleep(0.5) # 给 CAD 喘息时间
                
            except Exception:
                # 🟢 触发异常 = 真的没了 = 成功
                sys_logger.info(f"✨ [清理成功] {target_name} 已根除。")
                return True

        except Exception as e:
            sys_logger.error(f"  ❌ 清理过程异常: {e}")
            time.sleep(1)
            
    # 如果跑完所有重试次数还在
    sys_logger.error(f"💀 [清理失败] {target_name} 可能是嵌套块或被锁定。")
    return False



#&&% 重命名块实体
def rename_block_entity(ent, new_name):
    """
    将给定块参照实体 ent 的块名改为 new_name。
    如果 new_name 在 Block 表中已存在，则该实体将指向已有定义；
    否则将重命名它当前所引用的块定义。
    
    参数：
      ent       -- 一个 COM 块参照对象（如 BlockReference）
      new_name  -- 目标块名（字符串）
    """
    # 获取当前文档和块表

    blocks = doc.Blocks
    old_name = ent.Name

    try:
        # 尝试查找 new_name 是否已存在
        blocks.Item(new_name)
        # 存在：直接让该实体引用此定义
        ent.Name = new_name
    except Exception:
        # 不存在：重命名它当前所引用的定义
        blk_def = blocks.Item(old_name)
        blk_def.Name = new_name

#&&% 由块名选择实例

def get_block_instances(block_name: str, max_retries: int = 5):
    """
    根据给定的块定义名，检索当前图形中所有对应的块参照实例（BlockReference），
    返回这些 COM 对象的列表。

    参数：
      block_name     – 块定义的名称（字符串），如 "MyBlock"
      max_retries    – 调用 select_kuai 时的最大重试次数

    返回：
      instances     – 包含所有匹配块参照的列表（如果未找到或者出错，返回空列表）
    """
    # ① 先调用 select_kuai() 拿到所有块实例 COM 对象（扁平列表）
    try:
        all_blocks = select_kuai(max_retries)
    except Exception as e:
        sys_logger.info(f"[错误] 调用 select_kuai 失败：{e}")
        return []

    instances = []
    # ② select_kuai 已经是所有块实例的列表，直接遍历
    for ent in all_blocks:
        try:
            # 对于块参照，EntityName 通常是 "AcDbBlockReference"
            # 且我们要筛选 Name 恰好等于 block_name 的那些
            if getattr(ent, "EntityName", "") == "AcDbBlockReference" and getattr(ent, "Name", "") == block_name:
                instances.append(ent)
        except Exception:
            # 如果某个 COM 对象没有 EntityName/Name 属性，就跳过
            continue

    sys_logger.info(f"选择到名为{block_name}的实例块{len(instances)}个")

    return instances



#&&% 从块实体对象获取其内部com对象

#&&% 获取块引用实体
def get_entities_from_block_reference(block_ref):
    """
    获取块引用对象中的所有子实体（COM对象形式）。

    参数：
        block_ref: 块引用对象（AcDbBlockReference）
        doc: 当前 AutoCAD 文档对象（Document）

    返回：
        entities: 子实体列表
    """
    try:
        block_name = block_ref.EffectiveName
        block_def = doc.Blocks.Item(block_name)
        entities = [ent for ent in block_def]
        sys_logger.info(f"[OK] 获取到 {len(entities)} 个子对象")
        return entities
    except Exception as e:
        sys_logger.info(f"[错误] 获取失败：{e}")
        return []

    




#以块的方式插入文件

#&&% 插入块到CAD
def insert_block_into_autocad(block_file_path, insertion_point=(0, 0, 0), scale=(1, 1, 1), rotation=0):
    """
    以块的方式插入 DWG 文件到 AutoCAD 中

    :param block_file_path: 块文件的路径，通常为 DWG 文件路径
    :param insertion_point: 插入位置的三元组 (x, y, z)
    :param scale: 块的缩放比例三元组 (sx, sy, sz)
    :param rotation: 块的旋转角度（弧度）
    """
    try:

        # 定义块插入点
        insertion_point = (insertion_point[0], insertion_point[1], insertion_point[2])

        # 插入块，使用 InsertBlock 方法
        block = ms.InsertBlock(insertion_point, block_file_path, scale[0], scale[1], scale[2], rotation)

        sys_logger.info(f"[OK] 块已插入，文件：{block_file_path}，插入点：{insertion_point}，缩放：{scale}，旋转角度：{rotation}")

    except Exception as e:
        sys_logger.info(f"[错误] 插入块时出错：{e}")



#不炸开

#&&% 插入标准块
def insert_standard_block(block_dwg,
                          insertion_point=(0, 0, 0),
                          scale=(1, 1, 1),
                          rotation=0,
                          wait=0.3):
    """
    不炸开，
    全程无交互对话框。

    block_dwg: 标准块 DWG 路径
    insertion_point: (x,y,z)
    scale: (sx,sy,sz)
    rotation: 旋转角度（度）
    """
    before = select_kuai()
    before_handles = {b.Handle for b in before}


    if not os.path.isfile(block_dwg):
        raise FileNotFoundError(block_dwg)

    # 准备参数
    x, y, z    = insertion_point
    sx, sy, sz = scale
    path       = os.path.abspath(block_dwg).replace("\\", "/")

    # 1) 插入块
    insert_cmd = (
        "-INSERT\n"
        f"\"{path}\"\n"    # 文件路径要加双引号
        f"{x},{y},{z}\n"
        f"{sx}\n"
        f"{sy}\n"
        f"{sz}\n"
        f"{rotation}\n"
    )
    doc.SendCommand(insert_cmd)
    time.sleep(wait)

    doc.SendCommand("RE\n")
    doc.SendCommand("Z\nE\n")
    time.sleep(wait)

    after = select_kuai()
    new_refs = [b for b in after if b.Handle not in before_handles]
    if not new_refs:
        print("[警告] 未检测到任何新插入的块引用")
        return []

    results = []
    for blk in new_refs:
        # 5. 先将它旋转归零
        try:
            blk.Rotation = 0
        except Exception:
            pass

        # 7. 取它的包围盒四角
        p1, p2 = blk.GetBoundingBox()
        minx, miny, minz = p1
        maxx, maxy, maxz = p2
        corners = [
            (minx, miny, minz),  # 左下
            (minx, maxy, minz),  # 左上
            (maxx, maxy, minz),  # 右上
            (maxx, miny, minz),  # 右下
        ]

        results.append((blk, corners))

    return results


#炸开


#&&% 插入并炸开DWG
def insert_and_explode_dwg(block_dwg,
                           insertion_point=(0, 0, 0),
                           scale=(1, 1, 1),
                           rotation=0,
                           wait=0.3):
    """
    将一个 WBLOCK 导出的标准块 DWG 插入到当前图，
    并立即 EXPLODE 成普通图元（不保留块引用）。

    参数:
        block_dwg: 标准块 DWG 路径
        insertion_point: 插入点 (x,y,z)
        scale: (sx,sy,sz)
        rotation: 旋转角度（度）
        wait: 每步命令后等待秒数
    """

    before = select_kuai()
    before_handles = {b.Handle for b in before}

    if not os.path.isfile(block_dwg):
        raise FileNotFoundError(block_dwg)

    # 准备参数
    x, y, z    = insertion_point
    sx, sy, sz = scale
    path       = os.path.abspath(block_dwg).replace("\\", "/")

    # 1) 插入块
    insert_cmd = (
        "-INSERT\n"
        f"\"{path}\"\n"    # 文件路径要加双引号
        f"{x},{y},{z}\n"
        f"{sx}\n"
        f"{sy}\n"
        f"{sz}\n"
        f"{rotation}\n"
    )
    doc.SendCommand(insert_cmd)
    time.sleep(wait)

    doc.SendCommand("RE\n")
    doc.SendCommand("Z\nE\n")
    time.sleep(wait)

    # 2) EXPLODE “Last”   （炸开最新插入的块引用）
    explode_cmd = (
        "EXPLODE\n"
        "L\n"    # Last
        "\n"     # 完成选择
    )
    doc.SendCommand(explode_cmd)
    time.sleep(wait)

    sys_logger.info(f"[OK] 已插入并炸开：{os.path.basename(path)} @ ({x},{y},{z})")

    after = select_kuai()
    new_refs = [b for b in after if b.Handle not in before_handles]
    if not new_refs:
        print("[警告] 未检测到任何新插入的块引用")
        return []

    results = []
    for blk in new_refs:
        # 5. 先将它旋转归零（容错，不成功就算了）
        try:
            blk.Rotation = 0
        except Exception:
            pass

        # 7. 取它的包围盒四角（加上 safe_get_bbox 防 CAD 忙）
        try:
            p1, p2 = safe_get_bbox(blk)
        except Exception as e:
            sys_logger.info(f"[警告] 获取块 {getattr(blk, 'Name', '?')} 外包盒失败：{e}")
            continue

        minx, miny, minz = p1
        maxx, maxy, maxz = p2
        corners = [
            (minx, miny, minz),  # 左下
            (minx, maxy, minz),  # 左上
            (maxx, maxy, minz),  # 右上
            (maxx, miny, minz),  # 右下
        ]

        results.append((blk, corners))

    # 兼容你当前调用方式：返回 (列表, 最后一个块)
    return results, blk if results else ([], None)


#&&% 新版本性能测试0109

@retry_on_busy
def insert_and_explode_dwg(
        block_dwg,
        insertion_point=(0, 0, 0),
        scale=(1, 1, 1),
        rotation=0,
        wait=0.3
    ):
    """
    【V3.0 重构版】插入并炸开 DWG
    
    改进点：
    1. 使用 CADGuard 保护 Insert 和 Explode 操作，防止死锁。
    2. 使用 SafeCOM 获取包围盒，增强稳定性。
    3. 保持原有返回格式，无缝兼容。
    """
    sys_logger.info(f"=====================确认insert_and_explode_dwg版本20260115=====================")    
    
    # 1. 参数与环境检查
    if not os.path.isfile(block_dwg):
        sys_logger.info(f"❌ 文件未找到: {block_dwg}")
        return [], None

    # 刷新一下，确保选择集最新
    # C.doc.SendCommand("RE\n") 
    
    # 2. 记录插入前的状态 (Snapshot)
    before = select_kuai() or []
    before_handles = {b.Handle for b in before}

    # 3. 准备参数
    x, y, z = insertion_point
    sx, sy, sz = scale
    path = os.path.abspath(block_dwg).replace("\\", "/")
    fname = os.path.basename(path)

    # 4. 执行插入 (关键操作，加卫士保护)
    # wait_after=True 确保插入动作完成后，CAD 恢复空闲才继续
    with CADGuard(f"插入-{fname}", wait_after=True):
        insert_cmd = (
            f"-INSERT\n"
            f"\"{path}\"\n"
            f"{x},{y},{z}\n"
            f"{sx}\n{sy}\n{sz}\n"
            f"{rotation}\n"
        )
        C.doc.SendCommand(insert_cmd)

    # 5. 视图调整 (防止炸开时对象在屏幕外报错)
    # with CADGuard("缩放范围"):
    #     C.doc.SendCommand("Z\nE\n") 

    # 6. 执行炸开 (Explode Last)
    with CADGuard(f"炸开-{fname}", wait_after=True):
        # 选中刚插入的块 (Last) 并炸开
        C.doc.SendCommand("EXPLODE\nL\n\n")

    sys_logger.info(f"✅ [OK] 已插入并炸开：{fname} @ ({x},{y},{z})")

    # 7. 捕获新对象 (Diff)
    # C.doc.SendCommand("RE\n") # 刷新数据库
    after = select_kuai() or []
    
    # 筛选出新增加的块引用 (即炸开后释放出来的那些)
    new_refs = [b for b in after if b.Handle not in before_handles]

    if not new_refs:
        print("⚠️ [警告] 炸开后未检测到新的块引用")
        return [], None

    # 8. 提取几何信息 (包围盒计算)
    results = []
    for blk in new_refs:
        # 8.1 归零旋转 (容错处理)
        try:
            if getattr(blk, "Rotation", 0) != 0:
                blk.Rotation = 0
        except: pass

        # 8.2 获取包围盒 (使用 SafeCOM 防止偶尔的 RPC 拒绝)
        try:
            # SafeCOM.call 会自动重试 GetBoundingBox
            p1, p2 = SafeCOM.call(blk.GetBoundingBox)
            minx, miny, minz = p1
            maxx, maxy, maxz = p2
            corners = [
                (minx, miny, minz), # 左下
                (minx, maxy, minz), # 左上
                (maxx, maxy, minz), # 右上
                (maxx, miny, minz), # 右下
            ]
            results.append((blk, corners))
        except Exception as e:
            sys_logger.info(f"⚠️ 无法获取块 {getattr(blk, 'Name', '?')} 包围盒: {e}")
            continue

    # 9. 返回结果
    # 兼容旧接口：返回 (列表, 最后一个块对象)
    last_blk = results[-1][0] if results else None
    return results, last_blk



#&&% 获取面积足够大的全部非同名块实例

#&&% 获取大块实例
def get_large_block_instances(
    area_threshold: float,
    tol: float = 100.0,
    max_retries: int = 5
) -> list:
    """
    获取模型空间中所有块实例，筛选出“包围盒面积大于 area_threshold” 的块，
    并按面积去重：如果两个块的面积差值小于 tol，则只保留其中一个。
    返回一个按出现顺序去重后的 COM 对象列表。

    :param area_threshold: 面积阈值（单位与 CAD 坐标相同，例如以平方图纸单位计）。
    :param tol:            面积去重的容差值（默认 100），若两个块面积差小于 tol，视为同一块。
    :param max_retries:    调用 select_kuai 时的最大重试次数，默认 5 次。
    :return:               包含所有“面积 > area_threshold”且去重后的块引用 COM 对象列表，
                          如果 select_kuai 调用失败，会返回空列表。
    """
    # ① 调用 select_kuai，每隔 max_retries 次重连一次
    try:
        all_blocks = select_kuai(max_retries)
    except Exception as e:
        sys_logger.info(f"[错误] 调用 select_kuai 失败：{e}")
        return []

    # 临时列表，用于收集已加入结果的块面积（用于去重）
    seen_areas = []
    large_blocks = []

    for blk in all_blocks:
        try:
            # GetBoundingBox 返回 (ll_point, ur_point)
            ll_point, ur_point = blk.GetBoundingBox()
            x1, y1, _ = ll_point
            x2, y2, _ = ur_point
            width = abs(x2 - x1)
            height = abs(y2 - y1)
            area = width * height
        except Exception:
            # 如果某些块无法获取包围盒，则跳过
            continue

        # 面积判断
        if area > area_threshold:
            # 检查当前 area 是否已近似出现在 seen_areas 中
            duplicate = False
            for existing_area in seen_areas:
                if abs(area - existing_area) < tol:
                    duplicate = True
                    break

            if not duplicate:
                seen_areas.append(area)
                large_blocks.append(blk)

    return large_blocks

#从com对象中，根据其外包盒的矩形的长边与短边的比值和面积在160000000到1000000000两个条件筛算

#&&% 确定合乎标准打印要求的自建多段线区域

def get_large_block_instances_with_tolerance(max_retries: int = 5, area_threshold: float = 70 ):
    """
    获取当前 DWG 中所有大尺寸块实例
    A3 1:100的正常值> 1240000000.000000
   
    """
    # ① 尝试获取所有块实例
    try:
        all_blocks = select_kuai(max_retries)
    except Exception as e:
        sys_logger.info(f"[错误] 调用 select_kuai 失败：{e}")
        return []

    # ② 先筛选面积 >= area_threshold 的块
    LB = []
    for blk in all_blocks:
        try:
            ll_point, ur_point = blk.GetBoundingBox()
            x1, y1, _ = ll_point
            x2, y2, _ = ur_point
            width = abs(x2 - x1)
            height = abs(y2 - y1)
            area = width * height
        except Exception:
            # 读取包围盒失败，跳过
            continue

        if area >= area_threshold:
            # 附带面积信息，便于后续比较
            LB.append((blk, area))

    return LC




#&&% 块内坐标转换成世界坐标（适合平面上的一般块）

def transform_point_by_block(block_ref, local_pt):
    """
    将块内部坐标 local_pt = (lx, ly, lz) 转换为世界坐标：
      1. 按 block_ref 的比例缩放 (XScaleFactor, YScaleFactor, ZScaleFactor)
      2. 按 block_ref.Rotation 旋转（绕 Z 轴，弧度制）
      3. 平移到 block_ref.InsertionPoint = (ix, iy, iz)

    返回 (wx, wy, wz)：
      wx = ix + (lx * sx * cosθ - ly * sy * sinθ)
      wy = iy + (lx * sx * sinθ + ly * sy * cosθ)
      wz = iz + (lz * sz)
    """
    ix, iy, iz = block_ref.InsertionPoint
    sx = block_ref.XScaleFactor
    sy = block_ref.YScaleFactor
    sz = block_ref.ZScaleFactor
    theta = block_ref.Rotation  # 单位：弧度

    lx, ly, lz = local_pt
    # 缩放后再旋转
    x_scaled = lx * sx
    y_scaled = ly * sy

    cos_t = math.cos(theta)
    sin_t = math.sin(theta)
    xr = x_scaled * cos_t - y_scaled * sin_t
    yr = x_scaled * sin_t + y_scaled * cos_t
    zr = lz * sz

    wx = ix + xr
    wy = iy + yr
    wz = iz + zr
    return (wx, wy, wz)


#&&% 按名称选择块
def select_block_by_name(block_name: str, max_retries: int = 5):
    """
    从 *模型空间* 快速选出指定块名的所有实例，返回实体列表。
    不遍历 ModelSpace，速度与 SelectionSet 相同量级。
    """
    t0, last_exc = time.time(), None

    for attempt in range(1, max_retries + 1):
        try:
            # 1) 删旧选择集
            with suppress(Exception):
                doc.SelectionSets.Item("SS_block_by_name").Delete()

            # 2) 新建选择集
            ss = doc.SelectionSets.Add("SS_block_by_name")

            # 3) 过滤器：0=INSERT, 2=块名, 410=布局名"Model"
            filterType  = vtInt([0, 2, 410])
            filterData  = vtVariant(["INSERT", block_name, "Model"])

            # acSelectionSetAll = 5
            ss.Select(5, 0, 0, filterType, filterData)

            lb = list(ss)
            sys_logger.info(f"[OK] 选到 {len(lb)} 个 “{block_name}”（{time.time() - t0:.3f}s，第{attempt}次）")
            return lb

        except Exception as e:
            last_exc = e
            sys_logger.info(f"[警告] 第 {attempt} 次失败：{e!r}")
            time.sleep(0.3)

    sys_logger.info(f"[错误] {max_retries} 次仍失败：{last_exc!r}")
    return []

#&&% 获取所有块定义
def get_all_block_definitions(max_retry: int = 3, quiet: bool = False):
    """
    返回当前 DWG 中所有块定义（BlockTableRecord）对象列表。

    实现策略：
    - 使用全局 li() / doc，不再重复 Dispatch
    - 使用 doc.Blocks.Item(i) 按索引获取，避免枚举器在 CAD 忙时抛 “应用程序正在使用中”
    - 对整个获取过程做最多 max_retry 次重试，若仍失败则返回当时已经获取到的部分列表

    参数:
        max_retry : 失败时最多重试次数（默认 3）
        quiet     : 是否静默，不打印警告信息

    返回:
        list[COM block object]
    """
    import time
    import pythoncom
    from contextlib import suppress

    global acad, doc, mp, sp

    def log(msg):
        if not quiet:
            print(msg)

    if not li():
        raise RuntimeError("get_all_block_definitions: li() 连接失败，无法获取当前 DWG。")

    blocks = []

    for attempt in range(1, max_retry + 1):
        blocks.clear()
        try:
            count = doc.Blocks.Count
        except pythoncom.com_error as e:
            # 可能是应用程序忙
            log(f"[警告] 获取 Blocks.Count 失败（第 {attempt} 次）：{e}")
            pythoncom.PumpWaitingMessages()
            time.sleep(0.2)
            continue

        try:
            for i in range(count):
                with suppress(Exception):
                    blk = doc.Blocks.Item(i)
                    blocks.append(blk)
            # 成功跑完一轮，直接返回
            return blocks
        except pythoncom.com_error as e:
            log(f"[警告] 遍历 Blocks 时出错（第 {attempt} 次）：{e}")
            pythoncom.PumpWaitingMessages()
            time.sleep(0.2)
            continue

    # 多次重试仍不完全成功，返回当前已经拿到的部分
    log(f"[警告] get_all_block_definitions 多次重试后仍存在问题，返回部分块定义，数量={len(blocks)}")
    return blocks


#&&% 获取所有块名
def get_all_block_names():
    """
    使用全局 li()/doc 获取当前 DWG 中所有块定义的名字列表。
    """
    import pythoncom

    blocks = get_all_block_definitions(quiet=True)
    names = []
    for blk in blocks:
        try:
            names.append(str(blk.Name))
        except pythoncom.com_error:
            continue
        except Exception:
            continue
    return names

#&&% 块清理

def purge_block(block_name: str, quiet: bool = False):
    """
    删除指定块的所有实例，并彻底清除该块定义。
    
    步骤：
      1. 在模型空间删除所有同名 INSERT 实例
      2. 在每个布局的 Block (PaperSpace) 删除所有同名 INSERT 实例
      3. 调用 PurgeAll() 清理未用定义
      4. 再次尝试删除块定义
    
    :param block_name: 要清理的块名称（区分大小写）
    :param quiet: True 则不打印过程信息
    """
    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument

    # --- 1) 模型空间 ---
    removed = 0
    for ent in list(doc.ModelSpace):
        if ent.ObjectName in ("AcDbBlockReference", "AcDbMInsertBlock") and getattr(ent, "Name", "") == block_name:
            with suppress(Exception):
                ent.Delete()
                removed += 1

    # --- 2) 各布局块空间（PaperSpace） ---
    for layout in doc.Layouts:
        with suppress(Exception):
            block_space = layout.Block
            for ent in list(block_space):
                if ent.ObjectName in ("AcDbBlockReference", "AcDbMInsertBlock") and getattr(ent, "Name", "") == block_name:
                    with suppress(Exception):
                        ent.Delete()
                        removed += 1

    time.sleep(0.2)
    if not quiet:
        sys_logger.info(f"ℹ 共删除 {removed} 个 “{block_name}” 实例")

    # --- 3) PurgeAll 清理所有未被引用的定义 ---
    with suppress(Exception):
        doc.PurgeAll()
        if not quiet:
            print("[OK] PurgeAll 清理未用定义")

    # --- 4) 删除块定义 ---
    try:
        blk = doc.Blocks.Item(block_name)
        blk.Delete()
        if not quiet:
            sys_logger.info(f"[OK] 已删除块定义：{block_name}")
    except Exception as e:
        if not quiet:
            sys_logger.info(f"[警告] 删除块定义失败（仍有隐性引用？）：{e}")

    if not quiet:
        sys_logger.info(f"ℹ 完成对 '{block_name}' 的清理。")

#&&% 清理未使用块
def purge_unused_blocks(quiet: bool = False):
    """
    一次性清除所有未被任何 INSERT 实例引用的块定义。
    速度快、可靠性高（不用逐个 SelectionSet 检测）。
    """
    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    doc  = acad.ActiveDocument

    # 1) 记录清理前的块名列表
    before = []
    for i in range(doc.Blocks.Count):
        with suppress(Exception):
            name = doc.Blocks.Item(i).Name
            before.append(name)

    # 2) 调用 PurgeAll 一次性清理
    t0 = time.time()
    with suppress(Exception):
        doc.PurgeAll()
    t1 = time.time()

    # 3) 记录清理后的块名列表
    after = []
    for i in range(doc.Blocks.Count):
        with suppress(Exception):
            name = doc.Blocks.Item(i).Name
            after.append(name)

    # 4) 计算差集
    removed = [name for name in before if name not in after]

    if not quiet:
        sys_logger.info(f"[OK] PurgeAll 清理完成，耗时 {t1 - t0:.3f}s")
        sys_logger.info(f"ℹ 共移除 {len(removed)} 个未使用块：")
        for nm in removed:
            print("   ·", nm)

    return removed



@debuggable
#&&% 清理块1
def purge_block_1(block_name: str, quiet: bool = False, max_delete_attempts: int = 2):
    """
    删除指定块的所有实例，并尽可能彻底清除该块定义。
    
    步骤：
      1. 在 *Model_Space / *Paper_Space 和其它块定义中删除所有同名 INSERT 实例
      2. 调用 PurgeAll() 清理未用定义
      3. 多次尝试删除块定义
      4. 如果仍然失败，将块名改为 lajikuai_时间戳_N，避免后续块名污染

    :param block_name: 要清理的块名称（区分大小写）
    :param quiet: True 则不打印过程信息
    :param max_delete_attempts: Delete 块定义的最大尝试次数
    """
    import time
    import pythoncom

    from datetime import datetime

    global acad, doc, mp, sp

    def log(msg):
        if not quiet:
            print(msg)

    # 小工具：判断是否块参照
    def is_block_ref(ent) -> bool:
        try:
            on = getattr(ent, "ObjectName", "")
        except Exception:
            return False
        return on in ("AcDbBlockReference", "AcDbMInsertBlock")

    # 小工具：生成垃圾块名
    def make_trash_name(base_prefix: str = "lajikuai", idx: int = 1) -> str:
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        return f"{base_prefix}_{ts}_{idx}"

    node("▶ PB0  purge_block：开始清理块 '{}'", block_name)

    # 0) 确保连接的是当前激活 DWG
    if not li():
        log(f"[错误] li() 连接失败，无法清理块 {block_name}")
        return

    # 1) 在所有块空间中删除该块的 INSERT 实例
    removed = 0

    # 1.1 模型空间
    with suppress(Exception):
        for ent in list(doc.ModelSpace):
            if is_block_ref(ent) and getattr(ent, "Name", "") == block_name:
                with suppress(Exception):
                    ent.Delete()
                    removed += 1

    # 1.2 所有布局的 Block（包括 PaperSpace）
    for layout in doc.Layouts:
        with suppress(Exception):
            block_space = layout.Block
            for ent in list(block_space):
                if is_block_ref(ent) and getattr(ent, "Name", "") == block_name:
                    with suppress(Exception):
                        ent.Delete()
                        removed += 1

    # 1.3 其它块定义内部的嵌套引用（避免“块中块”暗搓搓引用）
    for blk_def in doc.Blocks:
        with suppress(Exception):
            for ent in list(blk_def):
                if is_block_ref(ent) and getattr(ent, "Name", "") == block_name:
                    with suppress(Exception):
                        ent.Delete()
                        removed += 1

    if not quiet:
        log(f"ℹ 共删除 {removed} 个 “{block_name}” 实例（含模型/图纸/块内部）")

    time.sleep(0.1)

    # 2) PurgeAll 清理未用定义
    with suppress(Exception):
        doc.PurgeAll()
        if not quiet:
            log("[OK] PurgeAll 清理未用定义")

    # 3) 多次尝试删除块定义
    deleted = False
    for attempt in range(1, max_delete_attempts + 1):
        try:
            blk = doc.Blocks.Item(block_name)
        except pythoncom.com_error:
            # 已经不存在，视为删除成功
            deleted = True
            if not quiet:
                log(f"[OK] 块定义 '{block_name}' 已不存在，视为已清理。")
            break

        try:
            blk.Delete()
            deleted = True
            if not quiet:
                log(f"[OK] 已删除块定义：{block_name}（第 {attempt} 次尝试）")
            break
        except Exception as e:
            if not quiet:
                log(f"[警告] 第 {attempt} 次删除块定义 '{block_name}' 失败：{e}")
            time.sleep(0.1)

    # 4) 多次 Delete 仍失败 → 改名为 lajikuai_时间戳_N
    if not deleted:
        try:
            blk = doc.Blocks.Item(block_name)
        except pythoncom.com_error:
            # 刚刚已经被清掉了
            if not quiet:
                log(f"[OK] 块定义 '{block_name}' 在改名前就已不存在。")
            return

        idx = 1
        while True:
            new_name = make_trash_name("lajikuai", idx)
            try:
                _ = doc.Blocks.Item(new_name)
                idx += 1
            except pythoncom.com_error:
                break  # 找到未被占用的新名

        try:
            blk.Name = new_name
            if not quiet:
                log(f"[警告] 块定义 '{block_name}' 无法彻底删除，已改名为垃圾块 '{new_name}'")
        except Exception as e:
            if not quiet:
                log(f"[错误] 块 '{block_name}' Delete/改名均失败，可能仍有系统引用：{e}")

    if not quiet:
        log(f"ℹ 完成对 '{block_name}' 的清理。")

@debuggable
#&&% 清理未使用块1
def purge_unused_blocks_1(
    quiet: bool = False,
    protect_names=None,
    max_delete_attempts: int = 2,
    rename_prefix: str = "lajikuai",
):
    """
    一次性清除“当前文件中没有任何实例引用”的块定义。
    - 对每个无实例块，多次 Delete 不掉则改名为 lajikuai_时间戳_N。

    :param quiet: True 则不打印详细信息
    :param protect_names: 不参与清理的块名白名单（列表）
    :param max_delete_attempts: 每个块 Delete 最大尝试次数
    :param rename_prefix: 删除失败时垃圾块名前缀
    :return: List[str] - 所有被“处理”的块原名（包含已删除+已改名）
    """
    import time
    import pythoncom

    from datetime import datetime
    from collections import Counter

    global acad, doc, mp, sp

    if protect_names is None:
        protect_names = []

    def log(msg):
        if not quiet:
            print(msg)

    # 小工具：判断是否块参照
    def is_block_ref(ent) -> bool:
        try:
            on = getattr(ent, "ObjectName", "")
        except Exception:
            return False
        return on in ("AcDbBlockReference", "AcDbMInsertBlock")

    # 小工具：判断是否系统/匿名块（不要动）
    def is_system_block_name(name: str) -> bool:
        if name.startswith("*"):
            return True
        if "|" in name:
            return True
        return False

    # 小工具：生成垃圾块名
    def make_trash_name(base_prefix: str = "lajikuai", idx: int = 1) -> str:
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        return f"{base_prefix}_{ts}_{idx}"

    node("▶ PUB0 purge_unused_blocks：开始清理无实例块")

    if not li():
        log("[错误] li() 连接失败，无法清理无实例块。")
        return []

    # 1) 统计每个块名的引用次数（遍历所有块定义内部）
    usage_counter = Counter()
    block_names = []

    t0 = time.time()
    for blk_def in doc.Blocks:
        with suppress(Exception):
            name = str(blk_def.Name)
            block_names.append(name)

        with suppress(Exception):
            for ent in blk_def:
                if not is_block_ref(ent):
                    continue
                with suppress(Exception):
                    ref_name = str(ent.Name)
                    usage_counter[ref_name] += 1
    t1 = time.time()

    log(f"[INFO] 扫描 Blocks 完成，耗时 {t1 - t0:.3f}s，共 {len(block_names)} 个块定义。")

    # 2) 调用 PurgeAll 一次：先让 CAD 自己清一轮
    with suppress(Exception):
        doc.PurgeAll()
        log("[OK] PurgeAll 初步清理未用定义")

    # 3) 重建 block_names（因为部分已被 Purge）
    block_names = []
    for blk_def in doc.Blocks:
        with suppress(Exception):
            block_names.append(str(blk_def.Name))

    # 4) 过滤“无实例块候选”（usage == 0）
    unused_candidates = []
    for name in block_names:
        if is_system_block_name(name) or name in protect_names:
            continue
        if usage_counter.get(name, 0) == 0:
            unused_candidates.append(name)

    log(f"[INFO] 候选“无实例块”数量：{len(unused_candidates)}")

    if not unused_candidates:
        log("[OK] 未发现无实例块，清理结束。")
        return []

    processed = []  # 记录被删除或改名的“原始块名”

    # 5) 对每个无实例块尝试 Delete；失败则改名为 lajikuai_xxx
    for blk_name in unused_candidates:
        deleted = False

        for attempt in range(1, max_delete_attempts + 1):
            try:
                blk = doc.Blocks.Item(blk_name)
            except pythoncom.com_error:
                # 已不存在，当作成功
                deleted = True
                log(f"[OK] 块 '{blk_name}' 已不存在（可能刚刚被 PurgeAll 删除），视为已清理。")
                break

            try:
                blk.Delete()
                deleted = True
                processed.append(blk_name)
                log(f"[OK] 块 '{blk_name}' Delete 成功（第 {attempt} 次尝试）")
                break
            except Exception as e:
                log(f"[警告] 块 '{blk_name}' 第 {attempt} 次 Delete 失败：{e}")
                time.sleep(0.05)

        if deleted:
            continue

        # —— Delete 仍失败，改名为垃圾块 —— 
        try:
            blk = doc.Blocks.Item(blk_name)
        except pythoncom.com_error:
            # 又消失了
            log(f"[OK] 准备改名时，块 '{blk_name}' 已不存在。")
            continue

        idx = 1
        while True:
            new_name = make_trash_name(rename_prefix, idx)
            try:
                _ = doc.Blocks.Item(new_name)
                idx += 1
            except pythoncom.com_error:
                break

        try:
            blk.Name = new_name
            processed.append(blk_name)
            log(f"[警告] 块 '{blk_name}' 无实例但 Delete 失败，已改名为垃圾块 '{new_name}'")
        except Exception as e:
            log(f"[错误] 块 '{blk_name}' 无法 Delete 也无法改名：{e}")

    log(f"[OK] 无实例块清理完成，共处理 {len(processed)} 个块定义。")
    return processed

@debuggable
#&&% 预留新插入块名
def reserve_block_names_for_new_insert(
    block_names,
    rename_prefix: str = "oldblk",
    verbose: bool = True,
):
    """
    【核心目标】在插入新块之前，为一组块名“预留新定义空间”。

    逻辑：
    - 对于当前 DWG 中已经存在的同名块定义（包括已有实例）：
      → 不删除、不爆炸，只是把块定义统一改名为 rename_prefix_时间戳_N，
        让原有实例跟着一起使用这个“旧名”。
    - 这样，原来的块名（如 A3-H）在 Blocks 表中就“空出来”了，
      之后从标准模板插入同名块，就一定是全新的定义，不会被老文件干扰。

    参数:
        block_names  : 要预留的新块名列表，例如 ["A3-H", "A2-H", "A1-H", "A0-H"]
        rename_prefix: 旧定义改名使用的前缀（默认 "oldblk"）
        verbose      : 是否打印说明信息（node 日志不受此影响）

    返回:
        dict: { 原名 -> 新名 }，没有被占用的块名不会出现在返回字典中。
    """
    import pythoncom
    from datetime import datetime

    global acad, doc, mp, sp

    if isinstance(block_names, str):
        block_names = [block_names]

    def log(msg):
        if verbose:
            print(msg)

    # 生成唯一旧名
    def make_legacy_name(base_prefix: str, base_name: str, idx: int = 1) -> str:
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        # 带上原名，方便将来看出来是哪个演化而来
        return f"{base_prefix}_{base_name}_{ts}_{idx}"

    node("▶ RB0 reserve_block_names_for_new_insert：为块名 {} 预留新定义空间", block_names)

    if not li():
        log("[错误] li() 连接失败，无法预留块名。")
        return {}

    renamed_map = {}

    for name in block_names:
        try:
            blk_def = doc.Blocks.Item(name)
        except pythoncom.com_error:
            # 说明当前 DWG 里没有这个块定义，块名是“干净”的
            node("▶ RB1 块名 '{}' 在当前文件中不存在，可直接用于新定义。", name)
            continue

        # 已存在同名块定义 → 给它改个“旧块”名字，让位给后续新定义
        idx = 1
        while True:
            new_name = make_legacy_name(rename_prefix, name, idx)
            try:
                _ = doc.Blocks.Item(new_name)
                idx += 1
            except pythoncom.com_error:
                break  # 找到未占用的新名

        try:
            blk_def.Name = new_name
            renamed_map[name] = new_name
            node("▶ RB2 块定义 '{}' 已改名为旧块 '{}'", name, new_name)
            log(f"[INFO] 块定义 '{name}' 已让位给新定义，旧定义改名为 '{new_name}'。")
        except Exception as e:
            log(f"[警告] 块 '{name}' 改名为旧块失败：{e}")

    if verbose:
        if renamed_map:
            log("[OK] 已为以下块名预留新定义空间：")
            for old, new in renamed_map.items():
                log(f"   · {old} → {new}")
        else:
            log("[OK] 所有块名在当前 DWG 中本来就是干净的，无需让位。")

    return renamed_map





#&&% 获取选定块引用名
def get_selected_blockreference_names():
    """
    使用 pmxz() 选择实体，并返回所有块引用（AcDbBlockReference）的块名列表。

    :return: list of str，所选块引用的 Name 属性列表
    """
    try:
        # 调用已有函数获取当前选择集（返回 COM 对象列表）
        entities = pmxz()
    except Exception as e:
        sys_logger.info(f"[错误] 调用 pmxz() 失败：{e}")
        return []

    block_names = []
    for ent in entities:
        try:
            # 只处理块引用
            if getattr(ent, "ObjectName", "") == "AcDbBlockReference":
                # Name 属性就是块名
                name = getattr(ent, "Name", None)
                if name:
                    block_names.append(name)
        except Exception as e:
            # 某些 COM 对象可能不支持上述属性，忽略它
            continue

    return block_names


#&&&% 新加块处理

"""
20251205-07

1 对块处理函数进行了深入分析和完善，引入了系统模块D:/claude-tasks/cad/system/CAD_com_utils.py
处理RPC_E_CALL_REJECTED (-2147417846) 错误程序忙产生错误的问题。
有些函数发生了错误，没有完全预设。出问题再处理。

2 增加了ensure_list统一列表单元素和列表都可以作为参数的输入的函数。

3 在 CAD_basic 对属性块的格式操作，获取标签及其值，设置标签及其值，在 CAD_file_operations 对整体插图签需要的重构块内容作了深入到位的处理

4 geimini给出了一套LISP窗口静默传递变量的方法，其典型代码在函数redefine_block_with_entities

  ✅ 强烈推荐使用的场景（只要出现“选择对象:”提示）
  任何需要你指定特定一批对象进行操作的命令，用这个方法都是最稳的：
  移动/复制：_.-MOVE / _.-COPY （不想用 COM 接口的 Move 时，或者处理天正对象时）。
  旋转/缩放：_.-ROTATE / _.-SCALE。
  删除：_.-ERASE （比 COM 的 Delete() 更彻底，不容易残留）。
  写块：_.-WBLOCK （把特定对象存为新文件）。
  阵列/镜像：_.-ARRAY / _.-MIRROR。
  炸开：_.-EXPLODE （特别重要！ COM 接口没有 Explode 方法，只能用 SendCommand，而用 LISP 传参是最准的）。


5 CAD_basic新增3个重要函数

  TDbMText_content(comobj) 获取天正多行文字的内容

  set_entity_grip_state_precise 让处于混乱对象丛中的对象实体成为高亮，使用了LISP深度函数方法

  explode_single_object_marker(ent) 通过绘制辅助线段的handle，利用最后模型空间对象的-1，-2……回溯
  确定预期增加的对象，比图层法等方法大大改进


6 add_entities_to_block_direct提供了新的桌面模式引用，由块的桌面，换成直线也行

"""


#&&% 从区域创建CAD块
def create_block_from_region_cad(
    x1, y1, x2, y2,
    insert_point_option="左下",
    block_name_prefix="块",
    base_point=None,      # 不传 → 就地替换；传了则整体挪到 base_point
    ty: float = 1.0,
):
    """
    【纯 CAD 重绘版】从矩形区域创建块（重画为标准 CAD 实体），
    并用块实例替换原对象。

    特点：
      - 使用 select_entities_in_window 做窗口选取（实体进入夹点高亮状态）；
      - 使用 group_bbox_corners 计算整体外包盒，
        选择某个角点作为“块基点”；
      - 块定义基点固定为 (0,0,0)，所有几何以 base_corner 为原点重画；
      - 块实例插入点：
          base_point=None → = base_corner（就地替换）；
          否则插入到 base_point（整体平移到指定位置）；
      - 支持：
          AcDbLine / Circle / Arc / Polyline / LWPolyline / 2d/3dPolyline /
          Text / MText / BlockReference / Point
        TDbText 会转换为普通 TEXT。

    create_block_from_region_cad(
        68118.6404456771, 245584.27806091635, 351250.3972437666, 431600.55564468517,
        insert_point_option="左下",
        block_name_prefix="块",
        base_point=None,      # 不传 → 就地替换；传了则整体挪到 base_point
        ty= 1.0,
    )
    
    [成功] 已创建块 块06，插入块实例并替换原图形。
    <win32com.gen_py.AutoCAD 2021 Type Library.IAcadBlockReference instance at 0x3029450686608>
    
    D:/claude-tasks/cad/Functional_control/块处理/20251207.dwg

    """

    # ---------- 1. 获取应用与文档 ----------
    doc=C.doc

    # ---------- 2. 窗口选取实体 ----------
    entities = select_entities_in_window(
        x1, y1, x2, y2,
        ty=ty,
        select_mode="_W",
    )
    if not entities:
        print("[警告] 矩形区域内没有对象，无法创建块")
        return None

    sys_logger.info(f"[信息] 选中 {len(entities)} 个对象")

    # ---------- 3. 整体外包盒 + 基点 ----------
    bbox = group_bbox_corners(entities)
    if bbox is None:
        print("[错误] 无法计算整体外包盒")
        return None

    bottom_left, top_right, top_left, bottom_right = bbox
    corner_map = {
        "左下": bottom_left,
        "右上": top_right,
        "左上": top_left,
        "右下": bottom_right,
    }
    base_corner = corner_map.get(insert_point_option, bottom_left)
    bx, by, bz = base_corner
    sys_logger.info(f"[信息] 块基点(外包盒角点): {base_corner}")

    # 决定块实例插入点
    if base_point is None:
        insert_pt = base_corner      # 就地替换
    else:
        insert_pt = base_point       # 整体挪到指定点
    sys_logger.info(f"[信息] 块实例插入位置: {insert_pt}")

    # ---------- 4. 生成唯一块名 ----------
    block_name = block_name_prefix + "01"
    counter = 1
    while True:
        try:
            doc.Blocks.Item(block_name)
            counter += 1
            block_name = f"{block_name_prefix}{counter:02d}"
        except Exception:
            break

    sys_logger.info(f"[信息] 块名: {block_name}")

    # ---------- 5. 创建块定义（基点 = (0,0,0)） ----------
    block_base_pt = (0.0, 0.0, 0.0)
    block_base_variant = VARIANT(
        pythoncom.VT_ARRAY | pythoncom.VT_R8, list(block_base_pt)
    )
    block_def = doc.Blocks.Add(block_base_variant, block_name)

    # ---------- 小工具：GetBoundingBox ----------
    def bbox_two_points(e):
        try:
            bb_min, bb_max = e.GetBoundingBox()
            return bb_min, bb_max
        except Exception as ex:
            sys_logger.info(f"[警告] GetBoundingBox 失败: {ex}")
            return None, None

    # ---------- 6. 克隆函数：实体 → 块定义（相对 base_corner） ----------
    def clone_entity_to_block(ent):
        try:
            obj_name = com_retry(lambda: ent.ObjectName)
        except Exception as e:
            sys_logger.info(f"[警告] 无法获取实体类型: {e}")
            return False

        try:
            # ------ Line ------
            if obj_name == "AcDbLine":
                sp = get_object_property(ent, "StartPoint")
                ep = get_object_property(ent, "EndPoint")
                if sp is None or ep is None:
                    sp, ep = bbox_two_points(ent)
                    if sp is None or ep is None:
                        print("[警告] 无法获取线的两端点")
                        return False
                sp_rel = [sp[0] - bx, sp[1] - by, sp[2] - bz]
                ep_rel = [ep[0] - bx, ep[1] - by, ep[2] - bz]
                v_sp = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, sp_rel)
                v_ep = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, ep_rel)
                block_def.AddLine(v_sp, v_ep)
                return True

            # ------ Circle ------
            if obj_name == "AcDbCircle":
                center = get_object_property(ent, "Center")
                radius = get_object_property(ent, "Radius")
                if center is None or radius is None:
                    print("[警告] 无法获取圆的属性")
                    return False
                cen_rel = [center[0] - bx, center[1] - by, center[2] - bz]
                v_cen = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, cen_rel)
                block_def.AddCircle(v_cen, radius)
                return True

            # ------ Arc ------
            if obj_name == "AcDbArc":
                center = get_object_property(ent, "Center")
                radius = get_object_property(ent, "Radius")
                sa = get_object_property(ent, "StartAngle")
                ea = get_object_property(ent, "EndAngle")
                if None in (center, radius, sa, ea):
                    print("[警告] 无法获取圆弧的属性")
                    return False
                cen_rel = [center[0] - bx, center[1] - by, center[2] - bz]
                v_cen = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, cen_rel)
                block_def.AddArc(v_cen, radius, sa, ea)
                return True

            # ------ Polyline / LWPolyline / 2d/3dPolyline ------
            if obj_name in ("AcDbPolyline", "AcDb2dPolyline", "AcDb3dPolyline", "AcDbLWPolyline"):
                coords = get_object_property(ent, "Coordinates")
                if coords is None:
                    print("[警告] 无法获取多段线坐标")
                    return False
                coords = list(coords)
                for i in range(0, len(coords), 3):
                    coords[i]   -= bx
                    coords[i+1] -= by
                    coords[i+2] -= bz
                v_pts = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, coords)
                block_def.AddPolyline(v_pts)
                return True

            # ------ 单行文字 Text ------
            if obj_name == "AcDbText":
                ins = get_object_property(ent, "InsertionPoint")
                txt = get_object_property(ent, "TextString")
                h   = get_object_property(ent, "Height")
                if ins is None or txt is None or h is None:
                    print("[警告] 无法获取 Text 属性，保留原对象")
                    return False
                ins_rel = [ins[0] - bx, ins[1] - by, ins[2] - bz]
                v_ins = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, ins_rel)
                new_txt = block_def.AddText(txt, v_ins, h)
                rot = get_object_property(ent, "Rotation")
                sty = get_object_property(ent, "StyleName")
                try:
                    if rot is not None:
                        new_txt.Rotation = rot
                except Exception:
                    pass
                try:
                    if sty is not None:
                        new_txt.StyleName = sty
                except Exception:
                    pass
                return True

            # ------ MText ------
            if obj_name == "AcDbMText":
                ins = get_object_property(ent, "InsertionPoint")
                width = get_object_property(ent, "Width")
                contents = get_object_property(ent, "Contents")
                if ins is None or width is None or contents is None:
                    print("[警告] 无法获取 MText 属性，保留原对象")
                    return False
                ins_rel = [ins[0] - bx, ins[1] - by, ins[2] - bz]
                v_ins = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, ins_rel)
                new_mt = block_def.AddMText(v_ins, width, contents)
                ht  = get_object_property(ent, "TextHeight")
                rot = get_object_property(ent, "Rotation")
                try:
                    if ht is not None:
                        new_mt.Height = ht
                except Exception:
                    pass
                try:
                    if rot is not None:
                        new_mt.Rotation = rot
                except Exception:
                    pass
                return True

            # ------ BlockReference ------
            if obj_name == "AcDbBlockReference":
                ins = get_object_property(ent, "InsertionPoint")
                if ins is None:
                    bb_min, bb_max = bbox_two_points(ent)
                    if bb_min is None:
                        print("[警告] 无法获取块参照插入点")
                        return False
                    ins = bb_min

                name = get_object_property(ent, "Name")
                sx   = get_object_property(ent, "XScaleFactor") or 1.0
                sy   = get_object_property(ent, "YScaleFactor") or 1.0
                sz   = get_object_property(ent, "ZScaleFactor") or 1.0
                rot  = get_object_property(ent, "Rotation") or 0.0
                if name is None:
                    print("[警告] 无法获取块参照名称")
                    return False

                ins_rel = [ins[0] - bx, ins[1] - by, ins[2] - bz]
                v_ins = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, ins_rel)
                block_def.InsertBlock(v_ins, name, sx, sy, sz, rot)
                return True

            # ------ Point ------
            if obj_name == "AcDbPoint":
                bb_min, bb_max = bbox_two_points(ent)
                if bb_min is None:
                    print("[警告] 无法获取点坐标")
                    return False
                px = (bb_min[0] + bb_max[0]) / 2.0
                py = (bb_min[1] + bb_max[1]) / 2.0
                pz = (bb_min[2] + bb_max[2]) / 2.0
                pt_rel = [px - bx, py - by, pz - bz]
                v_pt = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, pt_rel)
                block_def.AddPoint(v_pt)
                return True

            # ------ TDbText：降级为普通 TEXT ------
            if obj_name == "TDbText":
                # 插入点：InsertionPoint / Position / 外包盒左下兜底
                ins = (get_object_property(ent, "InsertionPoint")
                       or get_object_property(ent, "Position"))
                if ins is None:
                    bb_min, bb_max = bbox_two_points(ent)
                    if bb_min is None:
                        print("[警告] TDbText 无法获取插入点，保留原对象")
                        return False
                    ins = bb_min

                txt = (get_object_property(ent, "TextString")
                       or get_object_property(ent, "Contents")
                       or "")
                h = (get_object_property(ent, "Height")
                     or get_object_property(ent, "TextHeight")
                     or 2500.0)
                rot = (get_object_property(ent, "Rotation")
                       or get_object_property(ent, "Angle")
                       or 0.0)

                ins_rel = [ins[0] - bx, ins[1] - by, ins[2] - bz]
                v_ins = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, ins_rel)
                new_txt = block_def.AddText(txt, v_ins, h)
                try:
                    new_txt.Rotation = rot
                except Exception:
                    pass
                return True

            # 其余对象暂不支持：保留在原位置
            sys_logger.info(f"[提示] 暂不支持的实体类型: {obj_name}，将保留在原位置。")
            return False

        except Exception as e:
            sys_logger.info(f"[警告] 克隆实体 {obj_name} 失败: {e}")
            return False

    # ---------- 7. 克隆 & 删除原对象 ----------
    cloned_count = 0
    skipped_count = 0

    for ent in list(entities):
        ok = False
        try:
            ok = clone_entity_to_block(ent)
        except Exception as e:
            sys_logger.info(f"[警告] 处理实体失败: {e}")
            ok = False

        if ok:
            try:
                ent.Delete()
                cloned_count += 1
            except Exception as e:
                sys_logger.info(f"[警告] 删除原实体失败: {e}")
                skipped_count += 1
        else:
            skipped_count += 1

    sys_logger.info(f"[信息] 成功克隆并删除 {cloned_count} 个实体，"
          f"保留 {skipped_count} 个实体(类型不支持或出错)。")

    if cloned_count == 0:
        print("[错误] 块内没有任何实体，取消插入块。")
        try:
            doc.Blocks.Item(block_name).Delete()
        except Exception:
            pass
        return None

    # ---------- 8. 插入块实例 ----------
    ins_variant = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, list(insert_pt))
    try:
        block_ref = mp.InsertBlock(ins_variant, block_name, 1.0, 1.0, 1.0, 0.0)
    except Exception as e:
        sys_logger.info(f"[错误] 插入块实例失败: {e}")
        return None

    sys_logger.info(f"[成功] 已创建块 {block_name}，插入块实例并替换原图形。")
    return block_ref


#&&% 从区域创建CMD块
@retry_on_busy 
def create_block_from_region_cmd(
    x1, y1, x2, y2,
    insert_point_option="左下",
    block_name_prefix="块",
    base_point=None,      # 不传则就地替换，传了则整体挪到 base_point
    ty: float = 1.0,
):
    """
    【命令行版】通过 -BLOCK 从矩形区域创建块（保留天正对象），
    已集成 SafeCOM 防崩溃机制。

    create_block_from_region_cmd(
        68118.6404456771, 245584.27806091635, 351250.3972437666, 431600.55564468517,
        insert_point_option="左下",
        block_name_prefix="块",
        base_point=None,      # 不传则就地替换，传了则整体挪到 base_point
        ty= 1.0,
    )
    
    [成功] 已创建块 块07，插入块实例。
    <win32com.gen_py.AutoCAD 2021 Type Library.IAcadBlockReference instance at 0x3029450694608>
    
    D:/claude-tasks/cad/Functional_control/块处理/20251207.dwg

    使用这个命令将属性文字和其它对象可以做成属性块
    create_block_from_region_cmd(
        432574.7599864453, 847782.1082262965, 464037.73142620176, 869475.5330589646,
        insert_point_option="左下",
        block_name_prefix="块",
        base_point=None,      # 不传则就地替换，传了则整体挪到 base_point
        ty = 1.0,
    )
    当前桌面文件： 20251207.dwg
    [成功] 已创建块 块09，插入块实例。
    <win32com.gen_py.AutoCAD 2021 Type Library.IAcadBlockReference instance at 0x2308131932240>
    
    """
    # 确保连接到 CAD
    doc = C.doc
     

    # --- 0. 状态检查：防止在命令交互中强行执行 ---
    try:
        cmd_active = doc.GetVariable("CMDACTIVE")
        if cmd_active != 0:
            sys_logger.info(f"[警告] AutoCAD 处于交互状态 (CMDACTIVE={cmd_active})，尝试取消...")
            doc.SendCommand("\x1b\x1b") # 发送 ESC
            time.sleep(0.5)
    except:
        pass

    # ---------- 小工具：归一化矩形 ----------
    def _normalize_rect(a1, b1, a2, b2):
        x_lo = min(a1, a2)
        x_hi = max(a1, a2)
        y_lo = min(b1, b2)
        y_hi = max(b1, b2)
        return (x_lo, y_lo), (x_hi, y_hi)

    (x_lo, y_lo), (x_hi, y_hi) = _normalize_rect(x1, y1, x2, y2)

    # ---------- 1. 用 SelectionSet.Window 做“后台选取” ----------
    ss_name = "ZB_TMP_BLOCK_SEL_SAFE"
    try:
        doc.SelectionSets.Item(ss_name).Delete()
    except Exception:
        pass

    ss = doc.SelectionSets.Add(ss_name)

    p1 = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [x_lo, y_lo, 0.0])
    p2 = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [x_hi, y_hi, 0.0])

    try:
        ss.Select(constants.acSelectionSetWindow, p1, p2)
    except Exception as e:
        sys_logger.info(f"[错误] SelectionSet.Window 选择失败: {e}")
        return None

    # === 【关键修改点 1】使用 SafeCOM 安全获取列表 ===
    # 替换掉了原来的 entities = list(ss)
    print("[信息] 正在获取选中对象列表...")
    entities = SafeCOM.list_selection(ss) 

    # 用完立即删除选择集，释放资源
    try:
        ss.Delete()
    except Exception:
        pass

    if not entities:
        print("[警告] 矩形区域内没有对象，无法创建块")
        return None

    sys_logger.info(f"[信息] 后台选中 {len(entities)} 个对象（含天正对象）。")

    # ---------- 2. 用 group_bbox_corners 计算整体外包盒 + 基点 ----------
    bbox = group_bbox_corners(entities)
    if bbox is None:
        print("[错误] 无法计算整体外包盒")
        return None

    bottom_left, top_right, top_left, bottom_right = bbox
    corner_map = {
        "左下": bottom_left,
        "右上": top_right,
        "左上": top_left,
        "右下": bottom_right,
    }
    base_corner = corner_map.get(insert_point_option, bottom_left)
    bx, by, bz = base_corner
    sys_logger.info(f"[信息] 块基点(外包盒角点): {base_corner}")

    # 决定块实例插入点
    if base_point is None:
        insert_pt = base_corner
    else:
        insert_pt = base_point
    sys_logger.info(f"[信息] 块实例插入位置: {insert_pt}")

    # ---------- 3. 生成唯一块名 ----------
    block_name = block_name_prefix + "01"
    counter = 1
    while True:
        try:
            doc.Blocks.Item(block_name)
            counter += 1
            block_name = f"{block_name_prefix}{counter:02d}"
        except Exception:
            break

    sys_logger.info(f"[信息] 计划创建块名: {block_name}")

    # ---------- 4. 用 -BLOCK + W 窗口，只建块定义 ----------
    base_pt_str = f"{bx},{by}"
    win_p1_str = f"{x_lo},{y_lo}"
    win_p2_str = f"{x_hi},{y_hi}"

    cmd_lines = [
        "_.-BLOCK",
        block_name,
        base_pt_str,
        "W",
        win_p1_str,
        win_p2_str,
        "",
    ]
    cmd = "\n".join(cmd_lines) + "\n"

    print("[信息] 发送命令流 -BLOCK（窗口选择，只建块定义）...")
    doc.SendCommand(cmd)
    
    # 必须等待命令消化
    time.sleep(ty)
    
    # 额外的忙碌等待
    wait_count = 0
    while doc.GetVariable("CMDACTIVE") != 0 and wait_count < 10:
        time.sleep(0.5)
        wait_count += 1

    # ---------- 5. 用 COM 插入块实例到 insert_pt ----------
    ins_variant = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, list(insert_pt))
    
    block_ref = None
    # 简单的本地重试，或者也可以用 SafeCOM.call(mp.InsertBlock, ...)
    for i in range(5):
        try:
            block_ref = mp.InsertBlock(ins_variant, block_name, 1.0, 1.0, 1.0, 0.0)
            break
        except Exception:
            time.sleep(0.5)
            
    if block_ref is None:
        sys_logger.info(f"[错误] 插入块实例失败")
        return None

    # ---------- 6. 删除原对象 ----------
    # === 【关键修改点 2】使用 SafeCOM 安全删除对象 ===
    # 很多对象可能已经被 -BLOCK 吸纳，所以这里是为了清理残留
    deleted = 0
    for ent in entities:
        try:
            # 安全调用 Delete，如果对象已失效会自动忽略
            SafeCOM.call(ent.Delete)
            deleted += 1
        except Exception:
            pass

    sys_logger.info(f"[成功] 已创建块 {block_name}，插入块实例。")
    return block_ref

#&&% 从列表对象创建块
@retry_on_busy
def create_block_from_list_cmd(
    entities,
    insert_point_option="左下",
    block_name_prefix="块",
    base_point=None, 
    ty: float = 0.5,
):
    """
    【命令行版】将指定的 Python 对象列表封装成块。
    
    原理：
      虽然 SendCommand 不接受 vtlist (VARIANT)，但我们可以提取对象的 Handle，
      通过 LISP 将其转换为命令行可识别的选择集。
      
    优势：
      1. 完美保留天正/自定义实体特性（非 COM 复制）。
      2. 即使对象不连续、不在矩形内也能精准建块。

    p=cb.pmxz()
                                         
    create_block_from_list_cmd(
        p,
        insert_point_option="左下",
        block_name_prefix="块",
        base_point=None, 
        ty = 0.5,
    )
                                         
    当前桌面文件： 20251207.dwg
    win32 已连接（复用现有 acad，第 1 次成功，Handle=5481）
    [成功] 块 块08 创建完成。
    <win32com.gen_py.AutoCAD 2021 Type Library.IAcadBlockReference instance at 0x3029422644624>
    
    使用这个命令可以创建处属性块，如果使用向普通块追加属性文字不能成属性块
    cb.get_block_attributes_dict(
        kuai,
        ignore_empty= False,
        upper_tag = True,
    )
    {'MINGCHENG': '', 'RIQI': ''}
    
    D:/claude-tasks/cad/Functional_control/块处理/20251207.dwg







    """

    # 1. 列表预处理
    if not entities:
        print("[警告] 列表为空")
        return None
    entities = ensure_list(entities)

    doc = C.doc


    # 2. 计算包围盒 & 基点 (自动处理)
    # 调用现有的 group_bbox_corners 计算几何中心
    bbox = group_bbox_corners(entities)
    if not bbox:
        print("[错误] 无法计算包围盒")
        return None

    bottom_left, top_right, top_left, bottom_right = bbox
    corner_map = {
        "左下": bottom_left, "右上": top_right,
        "左上": top_left,    "右下": bottom_right,
    }
    
    # 确定最终基点
    final_base = base_point if base_point else corner_map.get(insert_point_option, bottom_left)
    bx, by, bz = final_base
    sys_logger.info(f"[信息] 块基点: {bx:.2f}, {by:.2f}, {bz:.2f}")

    # 3. 生成块名
    block_name = block_name_prefix + "01"
    i = 1
    while True:
        try:
            doc.Blocks.Item(block_name)
            i += 1
            block_name = f"{block_name_prefix}{i:02d}"
        except:
            break # 报错说明名字不存在，可用

    # =========================================================
    # 【关键步骤】 将 Python 列表转换为命令行选择集
    # =========================================================
    # 原理：我们不能传 VARIANT 给 SendCommand，但可以传 Handle 字符串
    sys_logger.info(f"[处理] 正在将 {len(entities)} 个对象加入选择集...")
    
    # 定义一个临时的 LISP 变量名
    ss_var = "myss" 
    
    # 1. 清空/新建选择集: (setq myss (ssadd))
    doc.SendCommand(f"(setq {ss_var} (ssadd))\n")
    
    # 2. 遍历列表，将每个对象的 Handle 加入选择集
    # 构造 LISP 语句: (ssadd (handent "句柄值") myss)
    # 为了速度，我们分批拼接字符串发送
    
    cmd_buffer = []
    for ent in entities:
        try:
            h = ent.Handle
            # 将句柄加入 LISP 变量
            cmd_buffer.append(f'(ssadd (handent "{h}") {ss_var})')
        except:
            pass

    # 将巨大的命令拆分成小块发送（避免命令行缓冲区溢出）
    chunk_size = 20 
    for i in range(0, len(cmd_buffer), chunk_size):
        chunk = cmd_buffer[i:i+chunk_size]
        # 用 progn 包裹可以一次执行多条
        full_lisp = "(progn " + " ".join(chunk) + ")\n"
        doc.SendCommand(full_lisp)

    # =========================================================
    # 4. 执行 -BLOCK 命令
    # =========================================================
    sys_logger.info(f"[执行] -BLOCK 创建块: {block_name}")
    
    # 格式: -BLOCK <名> <基点> !<LISP变量名> <回车结束>
    # 注意 !myss 是告诉 CAD "去读 LISP 变量 myss 的内容"
    cmd_str = f"_.-BLOCK\n{block_name}\n{bx},{by},{bz}\n!{ss_var}\n\n"
    
    doc.SendCommand(cmd_str)

    # 等待命令完成
    time.sleep(ty) 
    
    # 清理 LISP 变量
    doc.SendCommand(f"(setq {ss_var} nil)\n")

    # =========================================================
    # 5. 原地插入块参照
    # =========================================================
    # -BLOCK 会把原对象吸走，我们需要原地插回来
    print("[收尾] 原地插入块参照...")
    try:
        ins_pt = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [bx, by, bz])
        # 简单的重试机制
        for _ in range(3):
            try:
                blk_ref = mp.InsertBlock(ins_pt, block_name, 1.0, 1.0, 1.0, 0.0)
                sys_logger.info(f"[成功] 块 {block_name} 创建完成。")
                return blk_ref
            except:
                time.sleep(0.2)
    except Exception as e:
        sys_logger.info(f"[错误] 插入块失败: {e}")

    return None




#&&% 获取块内实体
def get_block_contents_at_same_location(block_ref):
    """
    【函数2】获取块内图形并在原位置复制,它的作用是获取到实体内部对象

    思路：
        1. 复制一份块参照（与 block_ref 完全重合）；
        2. 对复制出来的块参照执行 Explode()；
        3. 删除临时块参照；
        4. 返回 Explode 产生的新实体列表。

    优点：
        - 位置、比例、旋转全部由 AutoCAD 自己处理；
        - 包括嵌套块在内的所有内容都会展开到当前空间；
        - 原 block_ref 保留不动。

    20251207测试完毕



    """
    C.li() 

    copied_entities = []

    try:
        # 防止传进来的不是 IAcadBlockReference，可以适当 CastTo（可选）
        # from win32com.client import CastTo
        # block_ref = CastTo(block_ref, "IAcadBlockReference")

        # 1. 复制一份块参照（复制品和原块重合）
        temp_ref = block_ref.Copy()
        print("[信息] 已创建临时块参照作为爆炸对象。")

        # 2. Explode：返回一组新实体，这些实体已经应用了块的变换
        exploded = temp_ref.Explode()  # 一般返回 tuple/列表
        if exploded is None:
            exploded = []
        copied_entities = list(exploded)
        sys_logger.info(f"[信息] Explode 产生 {len(copied_entities)} 个实体。")

        # 3. 删除临时块参照，保留原块参照不动
        try:
            temp_ref.Delete()
        except Exception:
            pass

        sys_logger.info(f"[成功] 已在原位置复制块内容（保留原块参照），复制实体数: {len(copied_entities)}")
        return copied_entities

    except Exception as e:
        sys_logger.info(f"[错误] 操作失败: {e}")
        return []



#&&% 添加实体到块

@retry_on_busy
def add_entities_to_block_direct(block_ref, entities, delete_original=True):
    """
    【函数】进入块定义内部添加对象（强类型修正版）
    1. 修复 CopyObjects 的 Variant 数组传参。
    2. 修复 CopyObjects 返回值嵌套元组解包。
    3. 【本次修复】修复 Move/Rotate/Scale 的点坐标参数无效问题 (强制转为 COM Variant)。

    保留原块的所有属性机制

    p=cb.pmxz()
    当前桌面文件： 20251207.dwg
    win32 已连接（复用现有 acad，第 1 次成功，Handle=566B）
    q=cb.pmxz()
    当前桌面文件： 20251207.dwg
    win32 已连接（复用现有 acad，第 1 次成功，Handle=566D）
    add_entities_to_block_direct(p[0], q, delete_original=True)
    [进入] 正在操作块定义: 块07
    [处理] 正在校正 9 个对象的位置...
    [成功] 已将 9 个对象加入块 块07 并校正位置。
    True
    测试完毕
    D:/claude-tasks/cad/Functional_control/块处理/20251207.dwg

    使用这个命令向普通块追加属性文字就是属性块，但还需要cb.attsync_block_instance(p[0])后才能生效
    add_entities_to_block_definition_1(p[0], q, ty = 0.5)
    
    [完成] 块 cfff 已包含新对象并原位重定义。
    True
    
    cb.get_block_attributes_dict(
    p[0],
    ignore_empty= False,
    upper_tag = True,
    )
    {}


    """
    
    # --- 内部辅助函数：将点转换为 COM 变体 ---
    def com_point(pt_tuple):
        """将 Python 元组 (x,y,z) 转换为 ActiveX 需要的 Variant (Array of Doubles)"""
        # 确保转为 float，防止 int 导致类型错误
        pt_list = [float(x) for x in pt_tuple]
        # 如果是二维点，补全为三维
        if len(pt_list) == 2:
            pt_list.append(0.0)
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, pt_list)

    # 1. 检查输入
    if not entities:
        return False
    entities = ensure_list(entities)

    doc = block_ref.Document ##这是一种特殊语法，换成直线等对象也可以，没有出错，也许多文档有用
    


    # 2. 获取块定义 (修改后)
    try:
        # A. 获取块名：优先取 EffectiveName (处理动态块)，失败则取 Name
        #    利用 get_attr 安全读取，避免报错
        block_name = cb.get_attr(block_ref, 'EffectiveName')
        if not block_name:
            block_name = cb.get_attr(block_ref, 'Name')

        # B. 判空检查：这是最关键的一步！
        #    如果 block_name 是 None，传给 doc.Blocks.Item() 会直接崩溃
        if not block_name:
            sys_logger.info(f"[错误] 无法读取块对象的名称 (Handle={getattr(block_ref, 'Handle', '未知')})")
            return False

        # C. 获取块定义
        #    此时 block_name 肯定是字符串，安全通过
        block_def = doc.Blocks.Item(block_name)

    except Exception as e:
        # 使用 locals() 检查变量是否存在，防止打印错误日志时自身又报错
        safe_name = block_name if 'block_name' in locals() and block_name else "未知"
        sys_logger.info(f"[错误] 获取块定义失败: '{safe_name}'. 错误信息: {e}")
        return False







    sys_logger.info(f"[进入] 正在操作块定义: {block_name}")

    # =========================================================
    # 3. 复制对象 (构建 Dispatch 数组)
    # =========================================================
    try:
        dispatch_array = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, 
            entities
        )
        ret_obj = doc.CopyObjects(dispatch_array, block_def)
        
        # 解包逻辑
        new_inner_objects = []
        if not isinstance(ret_obj, (list, tuple)):
            new_inner_objects = [ret_obj]
        else:
            if len(ret_obj) > 0 and isinstance(ret_obj[0], tuple):
                new_inner_objects = list(ret_obj[0])
            else:
                new_inner_objects = list(ret_obj)
        
        # 二次防御
        if new_inner_objects and isinstance(new_inner_objects[0], tuple):
             new_inner_objects = list(new_inner_objects[0])

    except Exception as e:
        sys_logger.info(f"[错误] CopyObjects 失败: {e}")
        return False

    # =========================================================
    # 4. 计算并执行位置逆变换 (使用 com_point 包装参数)
    # =========================================================
    # 获取参数


    ins_pt_raw = get_attr(block_ref, 'InsertionPoint')


    rot_angle = block_ref.Rotation    
    scale_factor = block_ref.XScaleFactor 
    origin_raw = (0.0, 0.0, 0.0)

    # 【关键步骤】预先将点转换为 COM Variant
    # Move/Rotate/ScaleEntity 方法对参数类型极其敏感
    try:
        com_ins_pt = com_point(ins_pt_raw)
        com_origin = com_point(origin_raw)
    except Exception as e:
        sys_logger.info(f"[致命错误] 坐标点转换失败: {e}")
        return False

    success_count = 0
    sys_logger.info(f"[处理] 正在校正 {len(new_inner_objects)} 个对象的位置...")

    for obj in new_inner_objects:
        try:
            # A: 逆向移动
            # Move(Point1, Point2) - 必须传入 Variant
            obj.Move(com_ins_pt, com_origin)
            
            # B: 逆向旋转
            # Rotate(BasePoint, RotationAngle)
            if abs(rot_angle) > 1e-6:
                obj.Rotate(com_origin, -rot_angle)
            
            # C: 逆向缩放
            # ScaleEntity(BasePoint, ScaleFactor)
            if abs(scale_factor - 1.0) > 1e-6:
                 obj.ScaleEntity(com_origin, 1.0 / scale_factor)
            
            success_count += 1
                 
        except Exception as e:
            # 打印更详细的 Handle 以便追踪
            h = "Unknown"
            try: h = obj.Handle
            except: pass
            sys_logger.info(f"[警告] 对象(H:{h}) 变换失败: {e}")

    # =========================================================
    # 5. 删除外部原对象
    # =========================================================
    if delete_original and success_count > 0:
        for ent in entities:
            try:
                ent.Delete()
            except:
                pass

    block_ref.Update()
    sys_logger.info(f"[成功] 已将 {success_count} 个对象加入块 {block_name} 并校正位置。")
    return True


@retry_on_busy
def add_entities_to_block_definition_explode(block_ref, new_entities, ty: float = 0.5):
    """
    【函数】向块定义中追加对象（独立版）
    不清除原块对象，而是将新对象与原块内容合并，原地重定义块。
    
    逻辑：
    1. 利用 ActiveX Explode 方法获取块的原有几何体（生成新图元，原块引用保持不变）。
    2. 将“新图元”与“原有几何体”合并。
    3. 使用 -BLOCK 命令覆盖原定义（由于选中的是炸开后的图元，命令执行后它们会消失进入块定义）。
    4. 原有的 block_ref 会自动更新显示新内容。

    :param block_ref: 目标块引用对象 (IAcadBlockReference)
    :param new_entities: 待加入的新图元列表
    :param ty: 等待时间
    """

    # 1. 检查新对象
    if not new_entities:
        return False
    new_entities = ensure_list(new_entities)

    # 调试信息（假设 li() 是你的日志函数）
    try:
        doc = C.doc
    except:
        pass


        
    # 2. 获取锁定目标信息
    # 优先取 EffectiveName (处理动态块)，如果你有 get_attr 函数，这是最稳健的写法
    block_name = cb.get_attr(block_ref, 'EffectiveName')
    
    # 如果 EffectiveName 获取失败（返回 None），降级尝试获取 Name
    if not block_name:
        block_name = cb.get_attr(block_ref, 'Name')

    # 【关键】必须进行判空检查
    # 因为如果 block_ref 既没有 EffectiveName 也没有 Name，block_name 会是 None
    if not block_name:
        sys_logger.info(f"[错误] 无法获取块名称 (Handle={cb.get_attr(block_ref, 'Handle')})")
        return False  # 或者 continue，视上下文而定






    # 获取基点 (必须是原插入点，保证视觉不动)


    ins_pt = get_attr(block_ref, 'InsertionPoint')


    bx, by, bz = ins_pt

    sys_logger.info(f"[追加对象] 锁定目标块: {block_name}")

    # =========================================================
    # 3. 关键步骤：获取原块内的几何体
    # =========================================================
    try:
        # Explode() 方法会在原地生成块内的图元对象，并返回一个元组/列表
        # 注意：这不会删除 block_ref 本身，仅仅是生成了它的组成部分
        # 这些部分位于当前模型空间，且坐标已经是变换后的世界坐标
        exploded_objects = block_ref.Explode()
        
        # 将 tuple 转为 list 以便合并
        exploded_list = list(exploded_objects)
        sys_logger.info(f"[处理] 原块分解出 {len(exploded_list)} 个图元用于重组")
    except Exception as e:
        sys_logger.info(f"[错误] 无法炸开块引用获取原几何体: {e}")
        return False

    # 合并：新对象 + 原块炸开的对象
    total_entities = new_entities + exploded_list

    # =========================================================
    # 4. 构建 LISP 选择集
    # =========================================================
    ss_var = "zb_append_ss"
    
    # 4.1 初始化
    doc.SendCommand(f"(setq {ss_var} (ssadd))\n")
    
    # 4.2 注入 Handle (包含新对象和原块分解出的对象)
    cmd_buffer = []
    for ent in total_entities:
        try:
            h = ent.Handle
            cmd_buffer.append(f'(ssadd (handent "{h}") {ss_var})')
        except:
            pass

    # 4.3 发送 LISP 填充命令 (分块发送防止命令行溢出)
    chunk_size = 20
    for i in range(0, len(cmd_buffer), chunk_size):
        chunk = cmd_buffer[i:i+chunk_size]
        doc.SendCommand("(progn " + " ".join(chunk) + ")\n")

    # =========================================================
    # 5. 发送覆盖命令
    # =========================================================
    print("[执行] 发送重定义指令(追加模式)...")
    
    # _.-BLOCK -> 名字 -> Y(覆盖) -> 基点 -> !变量 -> 结束
    # 原理：BLOCK 命令会将选择集中的对象（exploded_list 和 new_entities）
    # 从模型空间移除，并存入块定义中。
    cmd_str = f"_.-BLOCK\n{block_name}\nY\n{bx},{by},{bz}\n!{ss_var}\n\n"
    
    doc.SendCommand(cmd_str)
    
    # 等待命令执行完成
    time.sleep(ty)
    wait_count = 0
    while doc.GetVariable("CMDACTIVE") != 0 and wait_count < 20:
        time.sleep(0.2)
        wait_count += 1

    # 清理 LISP 变量
    doc.SendCommand(f"(setq {ss_var} nil)\n")

    # =========================================================
    # 6. 刷新与善后
    # =========================================================
    try:
        # 强制刷新显示
        doc.Regen(1)
        # 尝试更新块引用（虽然重定义后通常会自动更新）
        block_ref.Update()
    except:
        pass

    sys_logger.info(f"[完成] 块 {block_name} 已包含新对象并原位重定义。")
    return True



#&&% 重定义块内容

@retry_on_busy
def redefine_block_with_entities(block_ref, entities, ty: float = 0.5, debug_log_path="C:\\cad_debug_log.txt"):
    """
    【调试专用版 V2】redefine_block_with_entities
    修复了 'IAcadEntity object has no attribute Name' 的接口类型问题。
    """
    sys_logger.info(f"\n--- [DEBUG模式 V2] 正在执行严格版重定义 (ty={ty}) ---")
    
    # ================= 0. 环境准备 & 接口修复 =================
    if not entities:
        print("[错误] 输入实体列表为空。")
        return False
        
    try:
        acad = win32com.client.GetActiveObject("AutoCAD.Application")
        doc = acad.ActiveDocument
    except:
        acad = win32com.client.Dispatch("AutoCAD.Application")
        doc = acad.ActiveDocument

    # 【关键修复】尝试强制类型转换，确保它是块引用
    try:
        if block_ref.ObjectName == "AcDbBlockReference":
            block_ref = win32com.client.CastTo(block_ref, "IAcadBlockReference")
    except Exception as cast_e:
        # 如果转换失败，不阻断，后续用 getattr 尝试
        sys_logger.info(f"[类型警告] 接口转换失败，将尝试直接读取: {cast_e}")

    # 获取块名和 Handle
    try:
        # 1. 尝试获取属性（兼容 generic entity）
        # getattr(obj, name, default) 可以避免报错
        block_name = getattr(block_ref, 'EffectiveName', None)
        if not block_name:
            block_name = getattr(block_ref, 'Name', None)
            
        if not block_name:
            # 如果还是拿不到，可能是对象失效了
            sys_logger.info(f"[致命错误] 无法获取块名 (ObjectName={getattr(block_ref, 'ObjectName', 'Unknown')})")
            return False

        target_handle = getattr(block_ref, 'Handle', None)
        ins_pt = getattr(block_ref, 'InsertionPoint', None)
        
        if not target_handle or not ins_pt:
            print("[致命错误] 无法获取 Handle 或 插入点")
            return False
            
        bx, by, bz = ins_pt
    except Exception as e:
        sys_logger.info(f"[致命错误] 读取目标块信息时发生异常: {e}")
        return False

    # ================= 1. 现状快照 (基准线) =================
    try:
        block_def = doc.Blocks.Item(block_name)
        old_count = block_def.Count
        sys_logger.info(f"[现状] 块定义 '{block_name}' (Handle:{target_handle}) 当前对象数: {old_count}")
    except:
        sys_logger.info(f"[警告] 无法获取块定义 '{block_name}'，可能是匿名块。")
        old_count = -1

    # ================= 2. 输入审计 (Input Audit) =================
    valid_handles = []
    
    try:
        with open(debug_log_path, "a", encoding="utf-8") as f:
            f.write(f"\n====== {time.strftime('%H:%M:%S')} 处理块: {block_name} ======\n")
            f.write(f"目标 Handle: {target_handle}\n")
            
            for i, ent in enumerate(entities):
                try:
                    h = ent.Handle
                    if h == target_handle:
                        f.write(f"索引 {i}: Handle={h} -> [剔除] 目标块自身\n")
                        sys_logger.info(f"[警报] 发现目标块自身混入列表 (Handle={h})，已剔除。")
                        continue
                    
                    valid_handles.append(h)
                    f.write(f"索引 {i}: Handle={h} -> [有效] 类型={getattr(ent, 'ObjectName', 'Unknown')}\n")
                except:
                    pass
    except Exception as log_e:
        sys_logger.info(f"[日志警告] 写日志失败: {log_e}")
        valid_handles = []
        for ent in entities:
            if hasattr(ent, 'Handle') and ent.Handle != target_handle:
                valid_handles.append(ent.Handle)

    input_count = len(valid_handles)
    sys_logger.info(f"[审计] 准备注入 {input_count} 个对象。")

    if input_count == 0:
        print("[终止] 有效对象为 0，取消操作。")
        return False

    # ================= 3. 构建 LISP 选择集 & 握手 =================
    ss_var = "zb_debug_ss"
    doc.SetVariable("USERI1", 0) 
    doc.SendCommand(f"(setq {ss_var} (ssadd))\n")

    cmd_buffer = []
    for h in valid_handles:
        cmd_buffer.append(f'(ssadd (handent "{h}") {ss_var})')
    
    chunk_size = 10
    for i in range(0, len(cmd_buffer), chunk_size):
        chunk = cmd_buffer[i:i+chunk_size]
        doc.SendCommand("(progn " + " ".join(chunk) + ")\n")

    # === 握手验证 ===
    doc.SendCommand(f'(setvar "USERI1" (sslength {ss_var}))\n')
    
    # 等待 LISP 回写
    lisp_count = 0
    retries = 0
    while retries < 20: 
        time.sleep(0.1)
        lisp_count = doc.GetVariable("USERI1")
        if lisp_count > 0:
            break
        retries += 1

    sys_logger.info(f"[握手] Python发送: {input_count} | LISP接收: {lisp_count}")

    if lisp_count == 0:
        print("[失败] LISP 选择集为空，操作终止。")
        return False

    # ================= 4. 执行核心指令 =================
    print("[执行] 发送 -BLOCK 覆盖指令...")
    
    cmd = f"_.-BLOCK\n{block_name}\nY\n{bx},{by},{bz}\n!{ss_var}\n\n"
    doc.SendCommand(cmd)
    
    wait_cycles = 0
    while doc.GetVariable("CMDACTIVE") != 0 and wait_cycles < 50:
        time.sleep(0.1)
        wait_cycles += 1
    
    doc.SendCommand(f"(setq {ss_var} nil)\n")
    
    # ================= 5. 结果验尸 =================
    try:
        doc.Regen(1)
        block_ref.Update()
    except:
        pass
    
    try:
        new_count = doc.Blocks.Item(block_name).Count
        delta = new_count - old_count
        sys_logger.info(f"[结果] 块定义 '{block_name}' 旧数:{old_count} -> 新数:{new_count} (变化:{delta})")

        if new_count == old_count:
            print("❌ [失败确认] 块内部实体数量未变化！")
            return False
        else:
            print("✅ [成功确认] 块定义已更新。")
            return True
    except:
        return True


















#&&% 提取非块实体
@retry_on_busy
def extract_specific_entities_from_block(block_ref, mode: str = 'all', keep_in_block: bool = True):
    """
    【函数】从块中提取指定类型的对象（筛选版）
    
    :param block_ref: 块引用对象
    :param mode: 提取模式 
                 'text'    - 提取文字 (AcDbText, MText, AttributeDef, TDb...)
                 'block'   - 提取嵌套块 (BlockReference)
                 'polyline'- 提取多段线 (LWPolyline, 2d/3dPolyline)
                 'all'     - 提取所有对象
    :param keep_in_block: True=仅复制出来(保留原块内容); False=剪切出来(删除块内对应内容)
    :return: 提取出的新对象列表 (位于模型空间)


    lb=extract_specific_entities_from_block(p[0], mode = 'text', keep_in_block = True)
    
    [开始] 提取模式: text | 保留原块: True
    [提取] 成功筛选出 6 个目标对象。
    lb[0].ObjectName
    'TDbText'
    lb[1].ObjectName
    'TDbMText'
    lb[2].ObjectName
    'AcDbMText'
    lb[3].ObjectName
    'AcDbText'
    lb[4].ObjectName
    'AcDbAttributeDefinition'
    lb[5].ObjectName
    'AcDbAttributeDefinition'

    D:/claude-tasks/cad/Functional_control/块处理/20251207.dwg    

    """
    
    # ================= 配置区域 =================
    # 1. 预定义目标类型集合 (ObjectName)
    target_map = {
        'text': {
            'AcDbText', 
            'AcDbMText', 
            'AcDbAttributeDefinition', 
            'TDbMText',  # 天正多行文字
            'TDbText'    # 天正单行文字
        },
        'block': {
            'AcDbBlockReference', 
            'AcDbMInsertBlock'
        },
        'polyline': {
            'AcDbPolyline', 
            'AcDbLWPolyline', 
            'AcDb2dPolyline', 
            'AcDb3dPolyline'
        }
    }
    
    mode = mode.lower()
    extracted_entities = []
    
    sys_logger.info(f"\n[开始] 提取模式: {mode} | 保留原块: {keep_in_block}")

    try:
        doc = block_ref.Document
        
        # 0. 检查是否为 MInsertBlock
        if hasattr(block_ref, "MInsertCount"):
            if block_ref.MInsertCount > 1 or block_ref.MInsertRow > 1:
                print("[警告] 无法处理多重插入块 (MInsertBlock)。")
                return []

        # =========================================================
        # 第一步：炸开获取副本 (ModelSpace)
        # =========================================================
        try:
            # Explode 生成新对象在原位
            exploded_variants = block_ref.Explode()
        except Exception as e:
            sys_logger.info(f"[错误] 炸开失败: {e}")
            return []
            
        if not exploded_variants:
            return []

        # =========================================================
        # 第二步：筛选与清理副本
        # =========================================================
        # 逻辑：遍历炸开后的碎片，符合要求的保留，不符合的由程序立即删除
        
        for ent in exploded_variants:
            try:
                obj_name = ent.ObjectName
                is_match = False
                
                # --- 匹配逻辑 ---
                if mode == 'all':
                    is_match = True
                elif mode in target_map:
                    if obj_name in target_map[mode]:
                        is_match = True
                else:
                    sys_logger.info(f"[错误] 未知模式: {mode}")
                    return []

                # --- 动作 ---
                if is_match:
                    extracted_entities.append(ent)
                else:
                    # 关键：这不是我想要的东西，这是 Explode 产生的“废料”，立即删除
                    ent.Delete()
                    
            except Exception:
                # 异常对象防御性删除
                try: ent.Delete() 
                except: pass

        sys_logger.info(f"[提取] 成功筛选出 {len(extracted_entities)} 个目标对象。")

        # =========================================================
        # 第三步：处理原块定义 (如果需要“剪切”效果)
        # =========================================================
        if not keep_in_block and len(extracted_entities) > 0:
            try:

                
                # 获取块名 (优先取 EffectiveName 以支持动态块，失败则退回 Name)
                b_name = cb.get_attr(block_ref, 'EffectiveName')
                if not b_name:
                    b_name = cb.get_attr(block_ref, 'Name')
                
                # 2. 安全获取块定义
                # 必须判空！如果 b_name 为 None，直接传给 Item() 会导致 COM 崩溃
                if b_name:
                    try:
                        block_def = doc.Blocks.Item(b_name)
                    except Exception:
                        # 只有在名字存在但块定义表中找不到时才会进这里（极少见）
                        block_def = None
                else:
                    block_def = None



                sys_logger.info(f"[内部] 正在清理块定义 {b_name} 中的对应对象...")
                count_deleted = 0
                
                # 遍历块定义内部的所有实体
                # 注意：不能直接在循环中删除集合元素，建议先收集再删除，或者小心处理
                items_to_delete = []
                
                for item in block_def:
                    obj_name = item.ObjectName
                    should_delete = False
                    
                    if mode == 'all':
                        # 注意：全选模式下，不建议删除里面的 EndBlk 等特殊对象，通常只需删除几何体
                        # 但简单起见，且 Explode 无法还原属性，慎用 all + keep_in_block=False
                        if obj_name != "AcDbBlockEnd" and obj_name != "AcDbBlockBegin":
                             should_delete = True
                    elif mode in target_map:
                        if obj_name in target_map[mode]:
                            should_delete = True
                            
                    if should_delete:
                        items_to_delete.append(item)

                # 执行删除
                for item in items_to_delete:
                    try:
                        item.Delete()
                        count_deleted += 1
                    except:
                        pass
                
                sys_logger.info(f"[完成] 从块定义中移除了 {count_deleted} 个对象。")
                
                # 强制刷新
                block_ref.Update()
                doc.Regen(1)
                
            except Exception as e:
                sys_logger.info(f"[警告] 修改块定义失败: {e}")

        return extracted_entities

    except Exception as e:
        sys_logger.info(f"[异常] extract_specific_entities 发生错误: {e}")
        return []

#&&&% 确定炸开块

#&&% 确保简单炸开块

@retry_on_busy(max_retries=5, base_delay=0.2)
def safe_explode(block_entity):
    """
    【原子操作】
    只负责炸开和删除。如果 CAD 忙，@retry_on_busy 会自动原地重试。
    如果 5 次后依然失败（如对象锁定），抛出异常给上层处理。
    """
    # 1. 炸开 (Explode 保留原对象)
    result=block_entity.Explode()

    return  result



@retry_on_busy(max_retries=5, base_delay=0.2)
def _atomic_explode_and_delete(block_entity):
    """
    【原子操作】
    只负责炸开和删除。如果 CAD 忙，@retry_on_busy 会自动原地重试。
    如果 5 次后依然失败（如对象锁定），抛出异常给上层处理。
    """
    # 1. 炸开 (Explode 保留原对象)
    result=block_entity.Explode()

    try:
        # 这里的 wait 不需要太久，旨在让指令队列“落地”
        wait_quiescent(min_quiet=0.1, timeout=1.0) 
    except:
        time.sleep(0.1) # 兜底



    # 2. 删除 (Delete 移除原对象)
    block_entity.Delete()

    return  result

@timeit
def safe_explode_retry(entity, max_retries=5, rescue_retries=5, interval=1.0, verbose=True):
    """
    【通用原子函数 - 深度搜救修正版 (V5.0)】
    
    功能：
        尝试炸开实体并返回新对象。
        如果 Explode 返回空但原对象已消失，启用“深度搜救”在数据库中查找新生成的对象。
    
    返回契约 (Strict Mode):
        list: [新对象...] 或 [] -> 成功 (原对象已彻底消失)
        None:                  -> 失败 (原对象依然健在)
    
    修复记录:
        - 修正搜救索引算法：从 count_before - 1 开始扫描，解决多抓取一个旧对象的问题。
    """
    
    # 1. 保存身份信息与容器
    try:
        ent_handle = entity.Handle
        ent_name = entity.Name if hasattr(entity, 'Name') else "<Block>"
        # 获取 Document 和 容器 (ModelSpace/PaperSpace) 用于人口普查
        doc = entity.Document
        owner_space = doc.ObjectIdToObject(entity.OwnerID)
    except:
        if verbose: print("⚠️ [前置检查] 对象无效或无法读取容器，视为已处理。")
        return [] # 死对象视为成功处理

    # 闭包：存活检查
    def is_entity_alive(ent):
        try:
            _ = ent.Layer
            return True
        except:
            return False

    # 0. 开局检查
    if not is_entity_alive(entity):
        if verbose: sys_logger.info(f"ℹ️ [前置] 对象 {ent_name} 已不存在，无需炸开。")
        return []

    for i in range(1, max_retries + 1):
        # —————— 📸 拍摄快照 (人口普查) ——————
        count_before = 0
        try: count_before = owner_space.Count
        except: pass
        # ————————————————————————————————————

        try:
            # 1. 执行原子操作 (Explode + Delete)
            # 注意：_atomic_explode_and_delete 自带 @retry_on_busy 抗忙碌
            result = _atomic_explode_and_delete(entity)

            # 2. 【完美路径】直接拿到了返回对象
            if result and len(result) > 0:
                return list(result)

            # 3. 【搜救路径】返回空，但原对象没了 -> 启动深度搜救
            if not is_entity_alive(entity):
                if verbose: sys_logger.info(f"✅ [尝试 {i}] 炸开成功(无返回)，启动深度搜救 (Max {rescue_retries}次)...")
                
                # —————— ⛑️ 深度搜救循环 (索引修正版) ——————
                # 逻辑推导：
                # 旧数量 N -> 删1剩 N-1 (索引 0~N-2) -> 新对象从 N-1 开始
                # 修正：start_index = count_before - 1
                start_index = max(0, count_before - 1) 
                
                for r_attempt in range(1, rescue_retries + 1):
                    try:
                        # a. 等待数据库同步 (时间递增)
                        wait_time = 0.5 * r_attempt
                        wait_quiescent(min_quiet=wait_time, timeout=3.0)
                        
                        # b. 获取当前数量
                        count_after = owner_space.Count
                        
                        # c. 如果数量确实增加了 (排除仅仅是删除了的情况)
                        if count_after > count_before: 
                            recovered_objs = []
                            
                            # 遍历新增区间
                            # range(start, end) 含头不含尾
                            for idx in range(start_index, count_after):
                                try:
                                    item = owner_space.Item(idx)
                                    # 双重保险：排除掉自己(虽然应该已经死了)
                                    if hasattr(item, 'Handle') and item.Handle != ent_handle:
                                        recovered_objs.append(item)
                                except: pass
                            
                            # 如果捞到了东西
                            if recovered_objs:
                                if verbose: sys_logger.info(f"   🎉 [搜救 {r_attempt}] 成功找回 {len(recovered_objs)} 个新对象。")
                                return recovered_objs
                        
                        # d. 如果还没刷新出来，尝试 Regen 刺激一下 CAD
                        if verbose: sys_logger.info(f"   ⏳ [搜救 {r_attempt}] 数据库未刷新 (Cnt: {count_before}->{count_after})，尝试 Regen...")
                        try: doc.Regen(0) 
                        except: pass
                        
                    except Exception as rescue_err:
                        if verbose: sys_logger.info(f"   ⚠️ 搜救出错: {rescue_err}")

                # —————— 搜救结束 ——————
                if verbose: sys_logger.info(f"⚠️ [搜救结束] 尽力了，数据库未返回新对象。返回空列表。")
                return [] # 实在捞不到，只能返回空，但因为对象没了，算成功

            # 4. 逻辑失败: 对象还在
            if verbose: sys_logger.info(f"⚠️ [尝试 {i}/{max_retries}] 对象依然健在，重试...")

        except pythoncom.com_error as e:
            # 5. 捕获 "对象已删除" -> 视为成功
            if e.hresult == -2147352567 or "对象已被删除" in str(e):
                if verbose: sys_logger.info(f"✅ [尝试 {i}] 捕获删除信号。")
                return []
            
            if verbose: sys_logger.info(f"⚠️ [尝试 {i}] COM 报错: {e}")
            time.sleep(interval)
        
        except Exception as e:
            if verbose: sys_logger.info(f"⚠️ [尝试 {i}] 未知错误: {e}")
            time.sleep(interval)

    # 循环结束，最后确认一眼
    if not is_entity_alive(entity):
        return [] # 虽然坎坷，但结果是对象没了，算成功
    
    if verbose:
        sys_logger.info(f"❌ [最终失败] {ent_name} (Handle: {ent_handle}) 依然健在。")
    return None # ❌ 唯一返回 None 的情况：失败






#&&% 炸开对象并回溯

def explode_single_object_marker(ent):
    """
    【主函数】炸开单个对象（辅助线回溯版）
    
    逻辑：
      1. 创建一根辅助线（Marker）。
      2. 选中目标对象 (set_entity_grip_state_precise)。
      3. 发送 EXPLODE。
      4. 倒序遍历模型空间，收集对象，直到遇到辅助线为止。
      5. 删除辅助线，返回收集到的碎片。
    """


    li()

    if not ent: return []

    marker = None
    exploded_objs = []

    try:
        # 1. 创建辅助线 (Marker)
        # 随便画在原点附近即可，位置不重要，重要的是 Handle
        # 使用 VARIANT 创建坐标点
        pt1 = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [0.0, 0.0, 0.0])
        pt2 = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, [1.0, 1.0, 0.0])
        
        try:
            marker = mp.AddLine(pt1, pt2)
            marker_handle = marker.Handle
        except Exception as e:
            sys_logger.info(f"[错误] 无法创建辅助标记线: {e}")
            return []

        # 2. 选中目标对象 (进入夹点状态)
        # 调用之前写好的精确选中函数
        if not set_entity_grip_state_precise(ent):
            print("[中断] 无法选中对象，取消炸开。")
            return []

        # 3. 发送炸开命令
        # print("[操作] 发送 EXPLODE...")
        doc.SendCommand("._EXPLODE\n")

        # 4. 等待炸开完成
        try:
            # 调用已有的等待函数
            wait_command_done()
        except:
            # 简单回退策略
            time.sleep(0.5)

        # 5. 回溯模型空间 (Backtracking)
        # 原理：AutoCAD 新生成的对象（碎片）一定在 Database 的最后面
        # 我们从最后往前找，直到撞到我们的 Marker 辅助线
        
        current_count = mp.Count
        
        # 倒序遍历
        for i in range(current_count - 1, -1, -1):
            try:
                obj = mp.Item(i)
                
                # 如果遇到了我们的辅助线，说明后面的都是新炸出来的
                if obj.Handle == marker_handle:
                    break
                
                # 收集碎片
                exploded_objs.append(obj)
                
            except Exception:
                continue

    except Exception as e:
        sys_logger.info(f"[错误] 炸开过程异常: {e}")

    finally:
        # 6. 清理现场 (删除辅助线)
        if marker:
            try:
                marker.Delete()
            except:
                pass

    sys_logger.info(f"[成功] 炸开回收 {len(exploded_objs)} 个碎片。")
    return exploded_objs

#&&% 安全炸开并删除
def safe_explode_and_delete(bk, ci=3, delay=1.0):
    """
    对块对象 bk 执行安全的 Explode 与 Delete 操作：
      1. 最多尝试 ci 次调用 bk.Explode()，
         每次调用后等待 delay 秒，再检查返回的实体列表长度是否 > 1。
      2. 如果 ci 次都没有成功（len(LP) <= 1），抛出 RuntimeError。
      3. 如果 Explode 成功，最多尝试 ci 次调用 bk.Delete()，遇错则继续重试。
      4. 返回第一次成功 Explode 时得到的实体列表 LP。

    :param bk:    要炸开的块引用 COM 对象
    :param ci:    最大尝试次数，默认 3
    :param delay: 每次 Explode 后的等待时间（秒），默认 1.0
    :return:      成功 Explode 后返回的新实体列表 LP
    :raises:      RuntimeError 如果 Explode 在 ci 次内都失败
    """
    LP = []
    # 1️⃣ 重试 Explode
    for attempt in range(1, ci + 1):
        try:
            LP = bk.Explode()
        except Exception as e:
            LP = []  # 如果调用失败，视为没有炸开
        time.sleep(delay)

        # 尝试获取长度
        try:
            count = len(LP)
        except Exception:
            # 如果 LP 不是标准序列，就把它转换一下
            LP = list(LP)
            count = len(LP)

        if count > 1:
            # 炸开成功
            break
        else:
            # 第 attempt 次 Explode 未成功，继续重试
            continue
    else:
        raise RuntimeError(f"Explode 在 {ci} 次尝试后仍未成功 (len(LP)={len(LP)})")

    # 2️⃣ Explode 成功后，重试 Delete
    for attempt in range(1, ci + 1):
        try:
            bk.Delete()
            break
        except Exception:
            # Delete 出错，继续下次重试
            continue
    # 如果 ci 次都失败，也不抛错误，只是放弃删除
    return LP

#&&&&%% 第七部分 综合控制

#&&&% 系统操作

#&&% 清理缓存
"""
在你的 li() 初始化函数或错误处理逻辑中，建议加入一个**“万能重试机制”**：

如果捕获到任何 AttributeError 且涉及 win32com，自动执行一次缓存清理，然后提示用户重启程序。

这样可以极大地降低维护成本，不用每次报错都去手动删文件夹。

"""
def fix_com_cache():
    """
    【急救】清理 win32com 的 gen_py 缓存
    解决 'CLSIDToPackageMap' 或 'AttributeError' 等 COM 接口损坏问题。
    """
    import os
    import shutil
    import sys
    
    print("🧹 正在寻找损坏的 COM 缓存...")
    
    # 1. 尝试动态获取缓存路径
    try:
        import win32com
        #通常在 %TEMP%\gen_py 或 Python安装目录\Lib\site-packages\win32com\gen_py
        if hasattr(win32com, '__gen_path__'):
            gen_path = win32com.__gen_path__
        else:
            import win32com.client
            gen_path = win32com.client.gencache.GetGeneratePath()
    except Exception as e:
        sys_logger.info(f"无法自动定位路径，尝试标准位置... ({e})")
        gen_path = os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py')

    sys_logger.info(f"📍 定位到缓存路径: {gen_path}")

    # 2. 执行删除
    if os.path.exists(gen_path):
        try:
            # 必须删除整个文件夹，不仅仅是里面的文件
            shutil.rmtree(gen_path)
            print("✅ 成功删除损坏的缓存文件夹！")
            print("🚀 下次连接 CAD 时，Python 会自动重新生成它（可能会慢 1-2 秒，正常现象）。")
        except Exception as e:
            sys_logger.info(f"❌ 删除失败 (可能文件被占用): {e}")
            print("请手动删除该文件夹！")
    else:
        print("❓ 未找到缓存文件夹，可能已经被清理过了。")

    # 3. 还有一种情况：在 site-packages 里的只读缓存
    try:
        import site
        for site_package in site.getsitepackages():
            win32_path = os.path.join(site_package, "win32com", "gen_py")
            if os.path.exists(win32_path):
                sys_logger.info(f"🔎 发现系统级缓存: {win32_path}")
                try:
                    shutil.rmtree(win32_path)
                    print("✅ 系统级缓存已清理。")
                except:
                    print("⚠️ 系统级缓存清理失败 (可能需要管理员权限)，通常忽略即可。")
    except: pass




#&&% 清除nul

def delete_all_nul_under_folder(folder_path):
    """
    输入文件夹路径（如 D:\claude-tasks），
    自动扫描该文件夹及所有子文件夹，强制删除所有名为 'nul' 的文件。
    """
    # 1. 获取绝对路径，确保格式正确
    target_dir = os.path.abspath(folder_path)
    
    if not os.path.exists(target_dir):
        sys_logger.info(f"❌ 错误：找不到文件夹 {target_dir}")
        return

    sys_logger.info(f"🔎 正在扫描文件夹: {target_dir} ...")
    
    count = 0
    
    # 2. 递归遍历目录树 (root: 当前路径, dirs: 子文件夹, files: 文件列表)
    for root, dirs, files in os.walk(target_dir):
        # 检查文件列表中是否有 'nul' (Windows下通常不区分大小写，但也做个.lower()保险)
        for filename in files:
            if filename.lower() == "nul":
                # 构造普通完整路径
                full_path = os.path.join(root, filename)
                
                # 3. 构造强制删除路径 (UNC Path)
                # 核心技巧：加上 \\.\ 前缀绕过 Windows 设备检查
                unc_path = r"\\.\%s" % full_path
                
                try:
                    os.remove(unc_path)
                    sys_logger.info(f"✅ [删除成功] {full_path}")
                    count += 1
                except Exception as e:
                    sys_logger.info(f"❌ [删除失败] {full_path} -> 原因: {e}")

    if count == 0:
        print("✨ 扫描完成，未发现 'nul' 文件。")
    else:
        sys_logger.info(f"🎉 扫描完成，共删除了 {count} 个 'nul' 文件。")

#&&% 终止弹窗程序

def kill_dialog_killer():
    """
    查找并终止名为 'cad_dialog_killer.py' 的 Python 进程
    """
    print("正在检查并清理干扰进程...")
    # 使用 wmic 获取所有 python 进程的命令行参数
    try:
        # 获取进程列表 (PID 和 命令行)
        cmd = 'wmic process where "name=\'python.exe\'" get commandline,processid'
        proc = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        stdout, _ = proc.communicate()
        
        # 解码输出 (Windows通常是mbcs或gbk)
        output_str = stdout.decode('mbcs', errors='ignore')
        
        lines = output_str.strip().split('\n')
        killed = False
        
        for line in lines:
            if "cad_dialog_killer.py" in line:
                # 提取 PID (一般在行末)
                parts = line.strip().split()
                if parts:
                    pid = parts[-1]
                    if pid.isdigit():
                        sys_logger.info(f"发现干扰进程 (PID={pid})，正在终止...")
                        os.system(f"taskkill /F /PID {pid}")
                        killed = True
        
        if not killed:
            print("未发现 cad_dialog_killer.py 在运行。")
            
    except Exception as e:
        sys_logger.info(f"尝试清理进程时出错: {e}")





#&&% 终止指定py脚本 (psutil版)
def kill_python_script_by_name(target_script_name):
    """
    【推荐】使用 psutil 精准终止指定名称的 Python 脚本。
    能够准确处理命令行参数中包含空格路径的情况。
    """
    sys_logger.info(f"--- 正在检查进程: {target_script_name} ---")
    current_pid = os.getpid()
    target_lower = target_script_name.lower()
    killed_count = 0

    # 遍历所有进程
    for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
        try:
            # 1. 过滤掉非 Python 进程 (可选，为了提高效率)
            proc_name = proc.info['name'].lower() if proc.info['name'] else ""
            if "python" not in proc_name and "py.exe" not in proc_name:
                continue

            # 2. 获取命令行参数列表
            cmdline = proc.info['cmdline']
            if not cmdline:
                continue
            
            # 将命令行列表转换为字符串以便查找
            # cmdline 通常是 ['python.exe', 'D:/path/to/script.py', ...]
            cmd_str = " ".join(cmdline).lower()

            # 3. 核心判断：脚本名是否在命令行中
            if target_lower in cmd_str:
                pid = proc.info['pid']
                
                # 保护机制：不杀自己
                if pid == current_pid:
                    sys_logger.info(f"跳过自身进程 (PID={pid})")
                    continue

                sys_logger.info(f"发现目标进程 '{target_script_name}' (PID={pid})，正在终止...")
                
                # 4. 执行终止
                proc.kill() # 或者 proc.terminate()
                killed_count += 1
                
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            # 进程可能在遍历过程中已经消失或无法访问
            pass

    if killed_count == 0:
        sys_logger.info(f"未发现 {target_script_name} 在运行。")
    else:
        sys_logger.info(f"成功清理 {killed_count} 个目标进程。")
    
    return killed_count





#&&&% 重要基础知识（python、算法、数据结构）

#&&% ***🧠  抛出异常的处理逻辑

"""
1 try+except+finaly

成功后执行finaly后语句，失败后也会执行finaly后语句，必定预期整个代码段要做的就是finaly提供的功能


2 try+except+else

成功后执行else后语句，失败后不会会执行else后语句，必定预期成功情况下 要做的就是else提供的功能

3
try:
    risky_operation()
except SomeError as e:
    handle_error(e)
    # 这段只在 try 失败时运行
    try:
        do_special_on_failure()
    except Exception as e2:
        log("failure-only block failed:", e2)
else:
    # 成功时走这里
    do_if_success()
finally:
    cleanup()

do_special_on_failure() 就是专门为代码段失败时预期必定要执行的语句。它不同于直接放在except中，因为except可能的失败操作不能保证它被预期必然执行。

我们需要在某个操作失败时，必须执行某个操作，就使用这个逻辑



"""




#&&&% 脚本注释标准模块
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --

#📌 节点 1：形成正式“打印框线”并重排图形
#─────────────────────────────────────

"""🔍 输入:"""

"""🔧 关键函数："""
"""   > F1: `find_standard_printframes()` – 返回标准图框列表"""

"""🧠 处理逻辑:"""

"""📤 输出:"""


 

"""
🄐  🄑  🄒  🄓  🄔  🄕  🄖  🄗  🄘  🄙  🄚  🄛  🄜  🄝  🄞  🄟  🄠  🄡  🄢  🄣  🄤  🅅  🅆  🅇   🅈   🅉







⓿ ① ② ③ ④ ⑤ ⑥ ⑦ ⑧ ⑨ ⑩ ⑪ ⑫ ️⃣  

# ❖❖❖—— 重要函数：foo() ——❖❖❖

# ✺✺✺—— 关键逻辑：process_data()——✺✺✺

# ⚡⚡⚡—— 高优先：init_config() ——⚡⚡⚡

# ★★★—— 核心入口：main() ——★★★
常用符号推荐：

❖ （四角菱形）

✺ （放射状花瓣）

⚡ ★ （闪电）

★ （实心星）

✦ （中实菱形）

❗ （粗感叹号）

 
+--------------------------+
| 方法1：从打印图签块或打印线获取DX |
+--------------------------+
      ↓ 成功          ↓ 失败
+------------------+   +--------------------------+
| 从图签块重建框线 |   | 方法2：从"公司图签"图块提取DX |
+------------------+   +--------------------------+
           ↓                     ↓ 成功    ↓ 失败


+--------------------------------+
| 初始状态：未开始打印           |
+--------------------------------+
               ↓


        +---------+
        | F3函数  | ← 识别拟合框线对应图形
        +---------+
             ↓
 
📌 节点 1：确定“准打印框线”
─────────────────────────────────────
🔁 初始化阶段：
    > 清空 "准打印框线" 图层
    > 若已存在旧对象，则全部删除

🔍 提取打印框候选集合 DX 的三种方法：

① 方法一（推荐方式）
    > 查找 “图签块” + “打印框线” 块
    > 若存在图签块：
        > 提取包围盒 → 重绘为矩形 → 存入“准打印框线”
        > 清除原“打印框线”对象
        > [OK] 完成本节点目标


📦 输出：DX 对象集合 + “准打印框线”图层绘制

------------------------------------------------------

📌 节点 2：形成正式“打印框线”并重排图形
─────────────────────────────────────
🔍 输入：DX 集合（来自上节点）

🧠 图形分析与处理：
    > 识别标准打印框（F1）
    > 识别拟合打印框（F2）

🧱 图框替换与排版：
    > 拟合框 → 替换为最近 A1 标准比例图框
    > 重新排列打印图框，避免重叠

📦 输出：
    > 正式“打印框线”写入目标图层
    > 准打印框图层清空
    > Handle → 信息存储用于后续匹配

🔧 函数：
    > F1: 标准框识别
    > F2: 拟合框识别
    > F3: 图形识别归属（局部）
    > F4: 图形重排整体逻辑

┌────────────┐> F1: 标准框识别 
│ 节点 1：   │
│ 准打印框线 │
└────┬───────┘
     │ 清空图层
     ▼
┌────────────┐
│ 方法一：图签块或框线 │◄─────┐
└────┬──────────────┘      │
     │ 找不到             │
     ▼                    │

┌────────────┐
│ 排序 + 移动图形 F3/F4 │
└────┬────────┘
     ▼
┌────────────────────┐
│ 写入“打印框线”图层 │
└────────────────────┘

📌 节点 X：<简要标题>
─────────────────────────────────────
🔍 输入：
    > 输入数据或图元说明（如 DX 集合）
    > 来源说明（如来自上游节点）
    
🧠 处理逻辑：
    > 步骤说明 1（如：识别标准打印框）
    > 步骤说明 2（如：判断拟合框）

🔧 关键函数：
    > F1: 函数名 – 功能说明
    > F2: 函数名 – 功能说明

📤 输出：
    > 输出数据目标（如写入正式图层）
    > 副作用（如清除中间层、更新 handle 记录）

[警告]️ 异常处理：
    > 未匹配图框 → 输出警告
    > 多图重叠 → 排序并修正重排

🗂 示例调用：
    process_printframes(DX)


📌	节点起始	📌 节点 2：形成正式“打印框线”并重排图形
────────────	分割线	──────── 节点分隔 ────────
🧭	逻辑导向	🧭 下一步识别拟合图框
🔀	路径分支	🔀 分流处理：标准框 vs 拟合框
⤵ / ⤴	子流程跳转	⤵ 进入子流程 F3
🔂	循环处理	🔂 针对所有 F2 框执行替换逻辑

🔧 功能或模块说明
符号	用途	示例
🔧	函数定义	🔧 F1: 标准框识别函数
🧠	处理逻辑	🧠 图形分析与处理
🧪	测试/调试块	🧪 临时打印 bbox 信息
📦	输出结果	📦 输出图层对象
📥 / 📤	输入/输出	📤 输出写入 LAYER_PRINT 正式图框
🧱	数据结构/实体对象	🧱 实体对象替换操作

🔍 分析与识别相关
符号	用途	示例
🔍	搜索/识别	🔍 查找符合 A1 比例的图框
🧬	特征提取	🧬 提取图形外包盒坐标与宽高比
🔎	精细识别	🔎 判断是否为完整封闭矩形

[OK] 状态反馈
符号	用途	示例
[OK]	成功	[OK] 图框替换完成
[警告]️	警告	[警告]️ 找不到合适图框匹配项
[错误]	错误	[错误] 无法识别多段线边界
⏳	等待	⏳ 等待 CAD 对象响应

🔁 控制结构（可嵌入注释中）
符号	用途	示例
⬅️ ➡️ ⬆️ ⬇️	方位、移动方向	⬆️ 上移图框 1000 单位避免重叠
🔄	重复执行	🔄 对所有图框执行一次旋转检测
↪️	返回	↪️ 返回原图层处理状态

📊 数据操作与计算
符号	用途	示例
📐	几何计算	📐 计算图框宽高比
📏	测量	📏 测量距离与偏移
🧮	数值处理	🧮 总计图框数量 = len(F1) + len(F2)



"""

#&&% 结束WPS进程
def kill_wps(verbose: bool = False):
    """
    结束所有 WPS/金山办公相关进程，特别是 wpspdf.exe。
    verbose=True 时打印被终止的映像名。
    """
    # 需要关掉的映像清单（全部小写）
    targets = {
        "wps.exe",          # WPS 主进程
        "wpspdf.exe",       # ★ PDF 预览器
        "wpp.exe",          # 演示
        "et.exe",           # 表格
        "ksolaunch.exe",    # 启动器
        "wpscloudsvr.exe",  # 云同步
        "wpsupdate.exe",    # 更新器
        "wpscenter.exe",    # 消息中心
    }

    killed = set()
    for name in targets:
        # /T 连子进程一起；/F 强制；/IM 按映像名
        rc = subprocess.call(
            f'taskkill /T /F /IM {name}',
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        if rc == 0:      # 返回 0 表示确有此进程并已结束
            killed.add(name)

    if verbose:
        if killed:
            print("[OK] 已结束进程:", ", ".join(sorted(killed)))
        else:
            print("ℹ️ 未检测到 WPS 相关进程")


#&&%关闭excel进程

def close_all_excel_processes():
    """
    【函数编号】: SYS-KILL-XLS
    【功能描述】: 
        强制关闭所有 EXCEL.EXE 进程。
        使用 taskkill /F 强制终止，包含重试机制。
        
    【返回】: 
        bool: True (成功/无进程), False (失败)
    """
    import subprocess
    import time

    # --- 内部辅助：检测Excel进程数 ---
    def _get_excel_count():
        try:
            # 使用 tasklist 查找映像名称为 EXCEL.EXE 的进程
            # /NH 不显示列标题, /FO CSV 输出格式便于解析
            cmd = 'tasklist /FI "IMAGENAME eq EXCEL.EXE" /NH /FO CSV'
            result = subprocess.run(
                cmd, 
                capture_output=True, 
                text=True, 
                shell=True
            )
            # 统计输出中 "EXCEL.EXE" 出现的次数 (忽略大小写)
            if result.stdout:
                return result.stdout.lower().count("excel.exe")
            return 0
        except Exception:
            return 0

    max_retries = 3
    
    for attempt in range(max_retries):
        try:
            # 1. 检查当前 Excel 进程数
            process_count = _get_excel_count()
            
            if process_count == 0:
                # 如果是第一次循环就没发现，说明本来就没开
                if attempt == 0:
                    print("[信息] 没有 Excel 进程需要关闭")
                return True

            sys_logger.info(f"[清理] 检测到 {process_count} 个 Excel 进程，正在强制关闭 (第 {attempt + 1} 次)...")

            # 2. 使用 taskkill 强制终止
            # /F: 强制终止
            # /IM: 指定映像名称
            # /T: 终止指定的进程及其启动的任何子进程 (Excel常有子挂起进程)
            result = subprocess.run(
                ["taskkill", "/F", "/IM", "EXCEL.EXE", "/T"],
                capture_output=True,
                text=True,
                timeout=10,
                shell=True # Windows下建议开启shell以兼容路径问题
            )

            # 3. 检查结果
            # returncode 0: 成功终止
            # returncode 128: 进程未找到 (可能在检查和执行之间已自行关闭)
            if result.returncode == 0 or result.returncode == 128:
                print("[执行] 终止指令已发送，等待进程释放...")
                time.sleep(1.5) # 给系统一点反应时间

                # 4. 再次验证
                process_count_after = _get_excel_count()
                if process_count_after == 0:
                    print("[验证] 所有 Excel 进程已成功关闭")
                    return True
                else:
                    sys_logger.info(f"[警告] 仍有 {process_count_after} 个 Excel 进程顽固残留")
            else:
                sys_logger.info(f"[警告] taskkill 返回码: {result.returncode}")
                # sys_logger.info(f"[调试] {result.stderr}") 

        except subprocess.TimeoutExpired:
            sys_logger.info(f"[错误] 第 {attempt + 1} 次关闭操作超时")
        except Exception as e:
            sys_logger.info(f"[错误] 关闭 Excel 时发生异常: {e}")

        # 如果还没成功，等待后重试
        if attempt < max_retries - 1:
            sys_logger.info(f"[重试] 等待 1 秒后重试...")
            time.sleep(1)

    print("[失败] 达到最大重试次数，无法彻底关闭所有 Excel 进程")
    return False














#&&&% 确保安全删除
def safe_delete(ob, retries: int = 3, delay: float = 1) -> bool:
    """
    安全删除 CAD 对象。
    
    改进点：
    1. 删除前先检查对象有效性，避免对已删除对象报错重试。
    2. 缩短默认等待时间 (1.0s -> 0.1s)。
    3. 成功删除或对象本就不存在，都返回 True。
    """
    
    # --- 1. 空对象检查 ---
    if ob is None:
        return True

    # --- 2. 有效性预检查 ---
    # 尝试访问 Handle 属性。如果访问失败，说明对象已经失效/被删除。
    try:
        _ = ob.Handle
    except Exception:
        # 访问属性失败，说明对象极可能已经不存在了，视为“删除成功”
        return True

    # --- 3. 尝试删除循环 ---
    for attempt in range(1, retries + 1):
        try:
            ob.Delete()
            return True
        
        except pywintypes.com_error as e:
            # 获取错误码 (HRESULT)
            # -2147418111 (0x80010001)通常是 Call rejected by callee (CAD正忙)
            # 还有可能是对象被锁定
            
            error_code = e.hresult
            
            # 如果错误提示对象已失效 (例如 e.excepinfo 可能包含 'eWasErased')
            # 不同的 CAD 版本/接口错误信息可能不同，但通常如果 Delete 失败，
            # 我们再检查一次 Handle，如果 Handle 也没了，就说明其实删掉了。
            try:
                _ = ob.Handle
            except Exception:
                # 再次确认 Handle 无法访问，说明刚才的 Delete 可能其实生效了，或者并发被删了
                return True

            # 如果对象还在，且报错，说明是“繁忙/锁定”，需要等待重试
            if attempt < retries:
                time.sleep(delay)
            else:
                # 此时 debug 模式下可以打印错误信息
                # sys_logger.info(f"删除失败 Handle={ob.Handle}: {e}")
                pass

    return False



#&&&%  * 高亮选择转隐性移动区域内全部对象


"""
高亮选择窗口操作更可靠从而更快，但要考虑窗口不能挡

"""

#&&% 区域实体移动
def move_entities_in_region(coms, target=(0,0,0), ty=1, max_iter=3):
    """
    将 `coms` 对象的包围盒内所有实体，沿向量 (target - 左下角) 移动，
    每轮等待 `ty` 秒，最多循环 `max_iter` 次。

    参数：
      coms      -- 支持 GetBoundingBox() 的 COM 对象（如多段线、块参照等）
      target    -- 目标基点坐标，默认为 (0,0,0)
      ty        -- 每轮移动后等待秒数
      max_iter  -- 最多尝试的轮数
    """
    # 1. 读取包围盒，计算左下角 (x1,y1) 和右上角 (x2,y2)
    p1, p2 = coms.GetBoundingBox()
    x1, y1 = min(p1[0], p2[0]), min(p1[1], p2[1])
    x2, y2 = max(p1[0], p2[0]), max(p1[1], p2[1])

    # 2. 计算位移向量 Δ = target - (x1,y1)
    dx = target[0] - x1
    dy = target[1] - y1
    dz = target[2] - 0.0

    for i in range(1, max_iter + 1):
        # 清空显性选择集
        try:
            doc.PickfirstSelectionSet.Clear()
        except Exception:
            pass

        # 3. 在 CAD 里用窗口选择高亮区域内对象
        highlight_entities_in_window(x1, y1, x2, y2)

        # 4. 拿到显性选择集里的实体
        pick = doc.PickfirstSelectionSet
        count = pick.Count if hasattr(pick, 'Count') else len(list(pick))
        if count == 0:
            sys_logger.info(f"[OK] 第 {i} 轮：区域已清空，停止。")
            break

        sys_logger.info(f"♻️ 第 {i} 轮：检测到 {count} 个对象，正在移动…")
        # 5. 对每个实体按计算好的向量执行 Move
        for ent in pick:
            try:
                # 从 (x1,y1,0) 移动到 (x1+dx, y1+dy, dz)
                ent.Move(vtpnt(x1, y1, 0.0),
                         vtpnt(x1 + dx, y1 + dy, dz))
            except Exception as e:
                sys_logger.info(f"  [警告] 对象 {ent.Handle} 移动失败：{e}")

        # 6. 等待 CAD 完成命令再进入下一轮
        time.sleep(ty)
    else:
        print("[警告] 达到最大迭代次数，可能仍有残留对象。")



#&&% 设置点样式
def 圆点(tz=1):


    """
    控制点的显示

    """

    zhi=0

    if tz ==1:

        zhi=35

    acad.ActiveDocument.SetVariable("PDMODE", zhi)#就是点最好的圆加十字形显示



#&&% 设置图纸背景色
def 图纸背景(zhi = 16777215 ):

    acad.ActiveDocument.Application.preferences.Display.GraphicsWinModelBackgrndColor = zhi#0即变成黑色





#&&&%  *** 按区域调整视图

#&&% 视图区域缩放
def  shitu_region(x1,y1,x2,y2):

    """
    按按对象外包盒调整视图
    使用时可shitu_region(*p),p=(x1,y1,x2,y2)

    """

    h=0.3*(abs(x1-x2)+abs(y1-y2))/2

    # 1️⃣ 缩放视图到合适窗口
    zoom_cmd = (
        f"_.ZOOM\n"      # 调用 Zoom
        f"_W\n"          # 窗口选项
        f"{x1-h},{y1-h}\n"   # 第一个角点
        f"{x2+h},{y2+h}\n"   # 第二个角点
    )
    doc.SendCommand(zoom_cmd)

#按对象外包盒调整视图

#&&% 视图实体缩放
def  shitu_entity(obj):

    """
    按按对象外包盒调整视图

    """
    p1,p2=obj.GetBoundingBox()

    x1,y1=p1[0],p1[1]

    x2,y2=p2[0],p2[1]

    h=0.3*(abs(x1-x2)+abs(y1-y2))/2

    # 1️⃣ 缩放视图到合适窗口
    zoom_cmd = (
        f"_.ZOOM\n"      # 调用 Zoom
        f"_W\n"          # 窗口选项
        f"{x1-h},{y1-h}\n"   # 第一个角点
        f"{x2+h},{y2+h}\n"   # 第二个角点
    )
    doc.SendCommand(zoom_cmd)




#&&% 截屏
"""
# 1. 截取全屏，返回一个 PIL.Image 对象
img = pyautogui.screenshot()

# 2. 如果想直接保存到文件：
pyautogui.screenshot('full_screen.png')

# 3. 只截取屏幕的一部分（x, y, width, height）
#    例如：从左上角 (100,100) 开始，截取 300×200 的区域
region_img = pyautogui.screenshot(region=(100, 100, 300, 200))
region_img.save('partial.png')


"""

#&&% 录制屏幕GIF
def record_screen_gif(
    output_path: str,
    duration: float = 5.0,
    fps: int = 10,
    region: tuple[int,int,int,int] | None = None
):
    """
    录制屏幕并保存为 GIF。

    :param output_path: 输出 GIF 文件路径，比如 "demo.gif"
    :param duration:    录制时长（秒）
    :param fps:         帧率（每秒截多少帧）
    :param region:      可选，(left, top, width, height) 只录制这一区域
    """
    import imageio 

    frames = []
    interval = 1.0 / fps
    end_time = time.time() + duration

    sys_logger.info(f"▶ 开始录制：时长={duration}s，帧率={fps}fps，区域={region or '全屏'}")
    while time.time() < end_time:
        img = pyautogui.screenshot(region=region)  # PIL.Image 对象
        frames.append(img)
        time.sleep(interval)

    sys_logger.info(f"🛑 录制结束，共捕获 {len(frames)} 帧，正在合成 GIF……")
    # imageio 能直接接受 PIL.Image
    imageio.mimsave(output_path, frames, fps=fps)
    sys_logger.info(f"[OK] 已保存为 {output_path}")



#&&% 最小化所有窗口
def minimize_all_windows():
    """
    模拟按下 Win+M，将所有窗口最小化。

    特性	Win + D (显示桌面)	Win + M (最小化所有)
    核心逻辑	切换开关 (Toggle)	单向命令 (Command)
    动作行为	
    第1次：隐藏所有窗口
    
    
    第2次：恢复原窗口布局
    
    第1次：最小化所有窗口
    
    
    第2次：无变化 (继续保持最小化)
    
    覆盖范围	极强。能隐藏几乎所有类型的窗口（包括无法最小化的对话框）。	普通。只能最小化支持“最小化”属性的窗口（某些固定对话框可能关不掉）。
    自动化风险	高。如果脚本判断失误连续运行了两次，窗口会弹回来，导致后续截图或操作失败。	低。它是幂等的（Idempotent），无论运行多少次，窗口都是最小化状态，适合“强制清场”。
    恢复方式	再按一次 Win+D	需要按 Win+Shift+M
    

    """

    import ctypes

    user32 = ctypes.windll.user32
    VK_LWIN = 0x5B   # 左 Win 键
    VK_M    = 0x4D   # M 键
    KEYEVENTF_KEYUP = 0x2

    # 按下 Win
    user32.keybd_event(VK_LWIN, 0, 0, 0)
    # 按下 M
    user32.keybd_event(VK_M,    0, 0, 0)
    # 松开 M
    user32.keybd_event(VK_M,    0, KEYEVENTF_KEYUP, 0)
    # 松开 Win
    user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)

    # 给系统一点反应时间
    time.sleep(0.1)



#&&%#控制CAD屏幕窗口在左上角


#&&% CAD窗口置左上
def set_autocad_window_to_top_left(resize_half: bool = True):
    """
    将 AutoCAD 窗口还原并移动到屏幕左上角，可选将其调整为半屏大小。
    """
    # 1️⃣ 找到可见或最小化的 AutoCAD 窗口
    windows = [w for w in gw.getWindowsWithTitle('AutoCAD')]
    if not windows:
        print("[错误] 未找到 AutoCAD 窗口！")
        return
    win = windows[0]

    # 2️⃣ 如果窗口最小化，先还原
    if win.isMinimized:
        win.restore()
        time.sleep(0.3)

    # 3️⃣ 激活窗口，确保我们操作的是真正的前台窗口
    try:
        win.activate()
    except Exception:
        # 有时 activate 也会失败，短暂等待再试
        time.sleep(0.2)
        win.activate()
    time.sleep(0.2)

    # 4️⃣ 移动到左上角 (0,0)
    win.moveTo(0, 0)
    time.sleep(0.2)

    # 5️⃣ 可选：将窗口调整为屏幕一半大小
    if resize_half:
        screen_w, screen_h = pyautogui.size()
        win.resizeTo(screen_w // 2, screen_h // 2)
        time.sleep(0.2)
        sys_logger.info(f"[OK] AutoCAD 窗口已恢复并移动到左上角，尺寸设为 {screen_w//2} x {screen_h//2}")
    else:
        print("[OK] AutoCAD 窗口已恢复并移动到左上角")



#&&% CAD窗口置左上别名
def l():

    set_autocad_window_to_top_left()



#&&% 更合理控制窗口函数
# — — — — -- -- -- -- --  — — — — -- -- -- -- -- — — — — -- -- -- -- --  — — — — -- --

#&&% 最小化窗口Win+D
def minimize_all_windows_d():
    """
    模拟 Win + D，将所有窗口最小化（切换）。
    """
    # VK_LWIN = 0x5B, VK_D = 0x44, KEYEVENTF_KEYUP = 0x2
    ctypes.windll.user32.keybd_event(0x5B, 0, 0, 0)       # 按下 Win
    ctypes.windll.user32.keybd_event(0x44, 0, 0, 0)       # 按下 D
    ctypes.windll.user32.keybd_event(0x44, 0, 0x2, 0)     # 松开 D
    ctypes.windll.user32.keybd_event(0x5B, 0, 0x2, 0)     # 松开 Win
    time.sleep(0.3)


#&&% 最小化窗口Win+M
def minimize_all_windows_m():
    """
    模拟按下 Win+M，将所有窗口最小化。
    """
    user32 = ctypes.windll.user32
    VK_LWIN = 0x5B   # 左 Win 键
    VK_M    = 0x4D   # M 键
    KEYEVENTF_KEYUP = 0x2

    # 按下 Win
    user32.keybd_event(VK_LWIN, 0, 0, 0)
    # 按下 M
    user32.keybd_event(VK_M,    0, 0, 0)
    # 松开 M
    user32.keybd_event(VK_M,    0, KEYEVENTF_KEYUP, 0)
    # 松开 Win
    user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)

    # 给系统一点反应时间
    time.sleep(0.1)


#&&% 恢复并定位窗口
def restore_and_position(
    name: str = "AutoCAD",
    width_ratio: float = 0.5,
    height_ratio: float = 0.5,
    x: int = 0,
    y: int = 0
) -> bool:
    """
    将第一个标题中包含 name 的窗口恢复、激活，并调整到指定位置和大小。

    :param name:          窗口标题关键字，默认 "AutoCAD"
    :param width_ratio:   窗口宽度占屏幕宽度的比例 (0 < ratio ≤ 1)
    :param height_ratio:  窗口高度占屏幕高度的比例 (0 < ratio ≤ 1)
    :param x:             窗口左上角 X 坐标，默认 0
    :param y:             窗口左上角 Y 坐标，默认 0
    :return:              找到并操作成功返回 True，否则 False
    """
    # ① 查找窗口
    candidates = [w for w in gw.getWindowsWithTitle(name) if w.title]
    if not candidates:
        sys_logger.info(f"[错误] 未找到标题包含 “{name}” 的窗口。")
        return False
    win = candidates[0]

    # ② 如果最小化，则还原
    if win.isMinimized:
        try:
            win.restore()
            time.sleep(0.2)
        except Exception as e:
            sys_logger.info(f"[警告] 无法还原窗口：{e}")

    # ③ 激活窗口（置于最前）
    try:
        win.activate()
    except Exception:
        time.sleep(0.1)
        try:
            win.activate()
        except Exception as e:
            sys_logger.info(f"[警告] 激活窗口失败：{e}")

    time.sleep(0.2)

    # ④ 移动到指定位置
    try:
        win.moveTo(x, y)
    except Exception as e:
        sys_logger.info(f"[警告] 移动窗口失败：{e}")

    time.sleep(0.2)

    # ⑤ 调整窗口大小
    sw, sh = pyautogui.size()
    # 限制比例范围
    wr = max(0.01, min(1.0, width_ratio))
    hr = max(0.01, min(1.0, height_ratio))
    new_w = int(sw * wr)
    new_h = int(sh * hr)
    try:
        win.resizeTo(new_w, new_h)
        time.sleep(0.2)
    except Exception as e:
        sys_logger.info(f"[警告] 调整窗口大小失败：{e}")

    sys_logger.info(f"[OK] 已将窗口“{win.title}”移动到 ({x},{y})，并调整为 {new_w}×{new_h}（占屏幕 {wr*100:.0f}% × {hr*100:.0f}%）")

    return True



#&&% 列出打开窗口标题
def list_open_window_titles() -> list[str]:
    """
    获取当前所有可见窗口的标题列表包括子窗口。

    :return: 一个字符串列表，每一项都是一个非空窗口标题。
    """
    titles = []
    for w in gw.getAllWindows():
        title = w.title.strip()
        if title:
            titles.append(title)
    return titles

#&&% 测鼠标位置
def ceshubiao_weizhi():
    """
    提示用户 5 秒内将鼠标移动到 AutoCAD 命令栏输入位置，
    然后采集当前鼠标坐标并返回 (x, y)。
    """
    print("请在 5 秒钟内，将鼠标精确地放在 AutoCAD 命令栏的输入位置…")
    time.sleep(5)
    x, y = pyautogui.position()
    sys_logger.info(f"已获取坐标：({x}, {y})")
    return x, y

#&&% 后台运行IDLE
def run_idle_background(script_path: str):
    """
    用后台模式启动 IDLE 去运行某个脚本，返回 Popen 实例。
    """
    # Windows 上隐藏窗口的 flag
    CREATE_NO_WINDOW = 0x08000000
    proc = subprocess.Popen([
            sys.executable, "-m", "idlelib", "-r", script_path
        ],
        creationflags=CREATE_NO_WINDOW,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL
    )
    return proc



"""
idle_proc = run_idle_background(r"D:/Myprogramsystem/cad/CAD基本操作.py")
# 结束刚才启动的 IDLE
idle_proc.terminate()
idle_proc.wait(timeout=5)
"""


#&&% 点击并拖动
def click_and_drag(x: int, y: int, juli: int):
    """
    在屏幕坐标 (x, y) 按下左键，然后向纵向拖动距离 juli。
    若 juli 为正值，则向上拖动；若为负值，则向下拖动。
    """
    # 1. 移动到 (x, y)
    pyautogui.moveTo(x, y)

    time.sleep(2) 

    # 2. 按住左键
    pyautogui.mouseDown(button='left')

    # 3. 拖动：拖动到 (x, y + juli)
    #    如果仅需相对拖动，也可用 dragRel(0, juli)。
    dest_x = x
    dest_y = y - juli  # 注意：屏幕坐标里，向上是 y 减小，向下是 y 增大
    pyautogui.moveTo(dest_x, dest_y, duration=0.2)

    # 4. 松开左键
    pyautogui.mouseUp(button='left')


#&&% 点击并找图
def click_and_find_image_shape(x: int, y: int, tupian_path: str, timeout: float = 10.0):
    """
    在 (x, y) 单击一次，然后不断在整个屏幕上查找与 tupian_path 对应的图片形状，
    一旦找到，就把鼠标移到它的中心并返回中心坐标 (Python int)；超时仍未找到则返回 None。

    :param x, y:        单击的初始坐标（可触发界面更新）
    :param tupian_path: 要识别的目标图片路径（如微信笑脸）
    :param timeout:     最长等待时间（秒）
    :return:            找到时返回 (x, y)；否则返回 None
    """
    if not os.path.isfile(tupian_path):
        raise FileNotFoundError(f"找不到图片文件：{tupian_path}")

    # 1. 右键点击 (x, y)
    pyautogui.moveTo(x, y)
    pyautogui.click()
    time.sleep(0.3)  # 等待界面刷新

    start = time.time()
    while True:
        # 2. 搜索图片并获取中心坐标
        loc = pyautogui.locateCenterOnScreen(tupian_path, confidence=0.9)
        if loc:
            cx, cy = loc
            # 转为 Python int 再移动
            cx, cy = int(cx), int(cy)
            pyautogui.moveTo(cx, cy)
            return cx, cy

        # 3. 检查超时
        if time.time() - start > timeout:
            print("[错误] 超时，未能识别到指定图片形状。")
            return None

        time.sleep(0.2)


#&&% 右键点击并移动
def right_click_and_move(x: int, y: int, x_xiangdui: int, y_xiangdui: int):
    """
    在屏幕坐标 (x, y) 处执行右键点击，然后将鼠标相对于 (x, y) 
    移动到 (x + x_xiangdui, y + y_xiangdui) 并停留。

    :param x:            初始水平坐标
    :param y:            初始垂直坐标
    :param x_xiangdui:   相对水平偏移（正值向右，负值向左）
    :param y_xiangdui:   相对垂直偏移（正值向下，负值向上）
    """
    # 移动到目标位置并右键点击
    pyautogui.moveTo(x, y)
    pyautogui.click(button='right')
    time.sleep(2)
    # 计算相对目标位置
    dest_x = x + x_xiangdui
    dest_y = y + y_xiangdui
    # 平滑移动到新位置
    pyautogui.moveTo(dest_x, dest_y, duration=0.2)

#&&% 结束所有IDLE进程
def kill_all_idle():
    """
    终止所有名为 'idle' 或 'idle.exe' 的进程（不再需要任务管理器）。
    """
    for p in psutil.process_iter(("pid", "name")):
        name = (p.info["name"] or "").lower()
        if name.startswith("idle"):
            try:
                p.terminate()
            except Exception:
                pass


#&&% IDLE窗口置右上
def set_idle_window_to_top_right():

    # 获取IDLE窗口句柄
    windows = [w for w in gw.getWindowsWithTitle('IDLE') if w.visible]
    
    if not windows:
        print("[错误] 未找到IDLE窗口！")
        return
    
    # 获取IDLE主窗口
    win = windows[0]
    
    # 获取屏幕分辨率
    screen_width, screen_height = pyautogui.size()

    # 计算窗口大小：横竖各占1/2
    half_width = screen_width // 2
    half_height = screen_height // 2

    # 将窗口移动到屏幕的右上角 (屏幕宽度的一半，顶部为0)
    win.moveTo(half_width, 0)

    # 调整窗口大小，设置为屏幕的1/2宽度和1/2高度
    win.resizeTo(half_width, half_height)
    
    sys_logger.info(f"[OK] IDLE窗口已调整到屏幕右上角，尺寸为 {half_width} x {half_height}")


def r():

    set_idle_window_to_top_right()

##控制OBS窗口在右下角
#&&% OBS窗口置右下
def place_obs_bottom_right():
    """
    将 OBS Studio 主窗口移动到屏幕右下角，并缩放为屏幕宽高的一半。
    - 若找不到 OBS 窗口，会打印错误信息。
    - 若有多个 OBS 窗口，仅操作第一个可见窗口。
    """
    # 1️⃣ 获取 OBS 主窗口
    obs_windows = [w for w in gw.getWindowsWithTitle('OBS') if w.visible]
    if not obs_windows:
        print('[错误] 未找到可见的 OBS Studio 窗口')
        return

    obs = obs_windows[0]            # 取第一个
    print(f'🔍 找到窗口: {obs.title}')

    # 2️⃣ 计算目标尺寸与位置 —— 右下角 1/4 区域
    screen_w, screen_h = pyautogui.size()
    half_w, half_h = screen_w // 2, screen_h // 2
    target_left = screen_w - half_w      # 右半屏起点
    target_top  = screen_h - half_h      # 下半屏起点

    # 3️⃣ 移动并缩放
    obs.moveTo(target_left, target_top)
    time.sleep(0.1)                      # 给系统一点缓冲
    obs.resizeTo(half_w, half_h)

    print(f'[OK] 已将 OBS 窗口调整到右下角 {half_w}×{half_h}')


def r2():

    place_obs_bottom_right()


#&&% 最小化指定窗口
def minimize_window(window_keyword: str = 'OBS') -> bool:
    """
    通用：最小化第一个标题包含 window_keyword 的可见窗口。

    :param window_keyword: 要匹配的窗口标题关键字（子串匹配），默认 'OBS'
    :return: 如果成功最小化返回 True，否则返回 False
    """
    # 1) 找到所有匹配的可见窗口
    windows = [w for w in gw.getWindowsWithTitle(window_keyword) if w.visible]
    if not windows:
        print(f'[错误] 未找到标题包含 "{window_keyword}" 的可见窗口')
        return False

    # 2) 取第一个并最小化
    win = windows[0]
    print(f'🔍 找到窗口: "{win.title}"，执行最小化')
    win.minimize()
    return True

#&&% 最大化CAD窗口
def maximize_autocad_window(window_keyword: str = 'AutoCAD') -> bool:
    """
    强制最大化第一个标题包含 window_keyword 的可见窗口。
    优先尝试使用 win32gui，如不可用则退回 ctypes 调用 user32。
    """
    # 1) 找到目标窗口
    wins = [w for w in gw.getWindowsWithTitle(window_keyword) if w.visible]
    if not wins:
        sys_logger.info(f"[错误] 未找到标题包含 “{window_keyword}” 的可见窗口")
        return False

    win = wins[0]
    hWnd = win._hWnd

    # 2) 先恢复（避免最小化状态），再最大化
    #    尝试使用 win32gui
    try:
        import win32gui, win32con
        win32gui.ShowWindow(hWnd, win32con.SW_RESTORE)
        time.sleep(0.1)
        win32gui.ShowWindow(hWnd, win32con.SW_MAXIMIZE)
    except ImportError:
        # 如果没有 pywin32，就退回 ctypes
        SW_RESTORE  = 9
        SW_MAXIMIZE = 3
        ctypes.windll.user32.ShowWindow(hWnd, SW_RESTORE)
        time.sleep(0.1)
        ctypes.windll.user32.ShowWindow(hWnd, SW_MAXIMIZE)

    time.sleep(0.2)  # 确保窗口完成最大化
    sys_logger.info(f"[OK] 已将窗口 “{win.title}” 最大化")
    return True


#&&% 点击开始OBS录制
def start_obs_recording_by_click(x: int = 1768, y: int = 872,
                                 button: str = 'left',
                                 clicks: int = 1,
                                 move_duration: float = 0.2):
    """
    通过鼠标点击屏幕上 (x,y) 坐标来控制 OBS 开始/停止录制。
    
    :param x: 目标点击点 X 坐标
    :param y: 目标点击点 Y 坐标
    :param button: 鼠标按钮，'left'、'right' 或 'middle'
    :param clicks: 点击次数，默认为单击
    :param move_duration: 从当前位置移动到目标的耗时（秒）
    """
    # 安全设置（可选）
    pyautogui.FAILSAFE = True
    pyautogui.PAUSE    = 0.1

    # 1) 平滑移动到目标
    pyautogui.moveTo(x, y, duration=move_duration)
    # 2) 点击
    pyautogui.click(x=x, y=y, clicks=clicks, button=button)
    sys_logger.info(f"[OK] 已点击 ({x}, {y})，请检查 OBS 是否已开始/停止录制。")


#&&&% *** 录屏
#&&% 发送微信
def fs(x1,y1):
    """
    微信调到0.5窗口

    """
    
    pyautogui.moveTo(x1+595,y1+190)
    pyautogui.click(x1+595,y1+190)
    pyautogui.press("enter")

#&&% 选微信群
def xuanqun(x1,y1,neirong):
    
    copy_to_clipboard(neirong)

    pyautogui.moveTo(x1+128,y1+33) 
    pyautogui.click(x1+128,y1+33)   
    time.sleep(2)
    activate_window_by_title("微信")

    time.sleep(2)

    pyautogui.moveTo(x1+128,y1+33)

    activate_window_by_title("微信")

    time.sleep(2)

    pyautogui.click(x1+128,y1+33)
  

    pyautogui.hotkey('ctrl', 'v') 


    pyautogui.moveTo(x1+158,y1+49) 
    pyautogui.click(x1+158,y1+49)   
 

    pyautogui.press("enter")


#&&% 复制到剪贴板
def copy_to_clipboard(text: str):
    """
    将传入的 text 文本写入系统剪贴板，供后续右键→粘贴使用。
    
    """
    
    # 创建一个隐藏的 tk 根窗口
    r = tkinter.Tk()
    r.withdraw()  # 隐藏主窗口

    # 清空剪贴板并写入新的文本
    r.clipboard_clear()
    r.clipboard_append(text)
    # 必须 update() 一下，确保数据真正存到剪贴板
    r.update()

    # 销毁隐藏窗口
    r.destroy()

#&&% 写微信
def xieweixin(x1,y1,neirong):
        
        copy_to_clipboard(neirong)

        time.sleep(2)
   
        pyautogui.moveTo(x1+532,y1+151)

        time.sleep(2)
        activate_window_by_title("微信")

        
        pyautogui.moveTo(x1+532,y1+151)

        time.sleep(2)

        pyautogui.click(x1+532,y1+151)
      

        pyautogui.hotkey('ctrl', 'v')        

        pyautogui.click(x1+532,y1+151)
    
#&&% 主操作函数
def 主操作函数():
    restore_and_position(
        name = "微信",
        width_ratio = 0.5,
        height_ratio = 0.5,
        x = 0,
        y = 0
    )  
 
    time.sleep(1)


    x1,y1,_,_ = activate_window_by_title("微信")

    xuanqun(x1,y1,"华新工作群")

    time.sleep(3)

    x2,y2=click_and_find_image_shape(358, 594, r"D:/Myprogramsystem/XT/weixin_xiaolian.png", timeout = 10.0)

    time.sleep(5)

    pyautogui.click(x2,y2) 

    time.sleep(5)

    x3,y3=click_and_find_image_shape(358, 594, r"D:/Myprogramsystem/XT/weixin_daixao_biaoqingbao.png", timeout = 10.0)

    time.sleep(2)
    pyautogui.click(x3,y3) 
    time.sleep(2)


 
    xieweixin(x1,y1,"Hello!我是公司管理员小化身，从今天起将参与公司管理，与各位一起奋进！")

    fs(x1,y1)

    
    time.sleep(1)


#&&% 主函数入口
def  main_func(folder_path=r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/工业园整理/三期/测试"):

    打印输出PDF() 



#&&% 录屏
def luping(main_func, *args, **kwargs):
    """
    1) 开始 OBS 录制  
    2) 最小化所有窗口；恢复并定位 OBS  
    3) 切换到 CAD 窗口，调用 main_func(*args, **kwargs）  
    4) 切换回 OBS，停止录制  

    main_func : 任意可调用对象  
    *args, **kwargs : 会原样传给 main_func  
    """
    # —— 最小化所有窗口 ——  
    minimize_all_windows_d()

    # —— 定位并激活 OBS ——  
    restore_and_position(name="OBS", width_ratio=0.5, height_ratio=0.5, x=0, y=0)
    time.sleep(1)

    # 点击 OBS “开始录制”  
    x0, y0, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x0 + 824, y0 + 328)
    time.sleep(0.2)
    pyautogui.click(x0 + 824, y0 + 328)

    # —— 切换到 CAD ——  
    minimize_all_windows_d()
    time.sleep(0.5)
    restore_and_position(name="AutoCAD", width_ratio=1.0, height_ratio=1.0, x=0, y=0)
    time.sleep(1)

    # —— 3️⃣ 执行主操作 ——  
    main_func(*args, **kwargs)
    print("主操作函数执行完毕")

    # —— 停止录制 ——  
    time.sleep(0.5)
    minimize_all_windows_d()
    restore_and_position(name="OBS", width_ratio=0.5, height_ratio=0.5, x=0, y=0)
    time.sleep(0.2)

    x4, y4, _, _ = activate_window_by_title("OBS")
    pyautogui.moveTo(x4 + 824, y4 + 328)
    time.sleep(0.2)
    pyautogui.click(x4 + 824, y4 + 328)

    print("录制已停止")


#&&% 魔方
def 魔方():

    import 魔方分析

    魔方分析.魔方控制台(r"D:/Myprogramsystem/BaiduSyncdisk/宋岳/自动化(动态)/魔方/分析")



#&&% 运行Python脚本
def run_py(pyname):
    try:
        # 运行指定的 Python 程序，隐藏命令行窗口
        result = subprocess.run(
            [sys.executable, pyname],
            check=True,
            text=True,
            capture_output=True,
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        sys_logger.info(f"[OK] 程序 {pyname} 执行成功。输出:\n{result.stdout}")
    except subprocess.CalledProcessError as e:
        sys_logger.info(f"[错误] 运行 {pyname} 时发生错误: {e}")
        sys_logger.info(f"错误信息: {e.stderr}")
    except FileNotFoundError:
        sys_logger.info(f"[错误] 未找到程序 {pyname}。请检查文件名和路径。")


#&&% 聚焦命令行
def focus_cmdline(cmd_x, cmd_y, delay=0.2):
    """
    把鼠标移到命令行并单击，确保焦点回到 AutoCAD 命令栏。

    
    """
    pyautogui.moveTo(cmd_x, cmd_y, duration=delay)
    pyautogui.click()



#&&% 激活窗口和子窗口


def activate_window_by_title(title: str, click_titlebar: bool = True) -> bool:
    """
    激活一个指定标题的窗口。
    
    1. 用 pygetwindow 找到窗口对象；
    2. 如果窗口被最小化，先还原；
    3. 将窗口置为前台（优先使用 win.activate()，失败时用 ctypes 强制）；
    4. （可选）在标题栏中点击一次，确保焦点。
    
    :param title: 要激活的窗口标题（部分匹配）。
    :param click_titlebar: 是否模拟一次点击标题栏，以确保窗口获得焦点。
    :return: True=激活成功，False=未找到窗口。
    """
    USER32 = ctypes.windll.user32
    SW_RESTORE = 9


    # ① 查找标题包含关键字的窗口
    wins = [w for w in gw.getWindowsWithTitle(title) if title in w.title]
    if not wins:
        sys_logger.info(f"[错误] 未找到标题包含 “{title}” 的窗口")
        return False
    win = wins[0]

    # ② 如果最小化，先还原
    if win.isMinimized:
        win.restore()
        time.sleep(0.2)

    
    # ③ 置为前台：先尝试 pygetwindow.activate()
    try:
        win.activate()
        time.sleep(0.2)
    except Exception:
        # ctypes 强制还原并置前
        hwnd = win._hWnd
        USER32.ShowWindow(hwnd, SW_RESTORE)          # 还原
        USER32.SetForegroundWindow(hwnd)             # 置前
        time.sleep(0.2)

    # ④ 可选：在标题栏中点一下
    if click_titlebar:
        x = win.left + win.width // 2
        y = win.top + max(1, min(30, win.height // 10))
        pyautogui.click(x, y)
        time.sleep(0.1)

    sys_logger.info(f"[成功] 已激活窗口: {win.title} 位置({win.left},{win.top}) 大小{win.width}x{win.height}")

    return  win.left,win.top,win.width,win.height


#&&% 窗口内点击
def click_in_window(title: str, offset_x: float, offset_y: float, click_titlebar: bool = False) -> bool:
    """
    在指定窗口的某个相对像素位置点击一次（相对于窗口左上角的偏移量）。

    :param title:         窗口标题关键字（部分匹配）
    :param offset_x:      从窗口左上角算起的水平像素偏移量（0 表示左边缘，width 表示右边缘）
    :param offset_y:      从窗口左上角算起的垂直像素偏移量（0 表示顶边缘，height 表示底边缘）
    :param click_titlebar: 是否先在标题栏点击一次以确保窗口获取焦点
    :return:              True=点击成功，False=未找到窗口
    """
    # 1) 查找窗口
    wins = [w for w in gw.getWindowsWithTitle(title) if title in w.title]
    if not wins:
        sys_logger.info(f"[错误] 未找到标题包含 “{title}” 的窗口")
        return False
    win = wins[0]

    # 2) 如果最小化，就先还原
    if win.isMinimized:
        win.restore()
        time.sleep(0.2)

    # 3) 尝试激活
    try:
        win.activate()
        time.sleep(0.2)
    except Exception:
        # ctypes 强制还原并置前
        hwnd = win._hWnd
        SW_RESTORE = 9
        ctypes.windll.user32.ShowWindow(hwnd, SW_RESTORE)
        ctypes.windll.user32.SetForegroundWindow(hwnd)
        time.sleep(0.2)

    # 4) 可选：点击标题栏让焦点真正到窗口
    if click_titlebar:
        tx = win.left + win.width // 2
        ty = win.top + 10  # 标题栏一般在窗口顶部 10px 左右
        pyautogui.click(tx, ty)
        time.sleep(0.1)

    # 5) 计算目标点的绝对屏幕坐标
    abs_x = int(win.left + offset_x)
    abs_y = int(win.top  + offset_y)

    # 6) 点击
    pyautogui.moveTo(abs_x, abs_y, duration=0.2)
    pyautogui.click(abs_x, abs_y)
    time.sleep(0.1)
    sys_logger.info(f"🔘 已在窗口 “{win.title}” 内部点击 ({offset_x}, {offset_y}) → 绝对 ({abs_x}, {abs_y})")
    return True

"""
click_in_window("图形导出", offset_x=600-10, offset_y=600-10, click_titlebar=True)

"""

#&&% 激活并点击艾可云
def activate_and_click_aikeyun():


    try:
    
        left, top, width, height = activate_window_by_title("艾可云", click_titlebar=True)
        
        # Calculate relative offsets from provided data
        rel_x1 = 887 - 682  # 205
        rel_y1 = 542 - 93   # 449
        
        # Move to first click position and left click
        pyautogui.moveTo(left + rel_x1, top + rel_y1)
        pyautogui.click(button='left')
        
        # Wait for 2 seconds
        time.sleep(2)
        
        # Calculate close position: using width - 22 for x, as it aligns with top-right close button
        rel_x2 = width - 22  # Equivalent to 400 - 22 = 378 in data
        rel_y2 = 18          # 111 - 93
        
        # Move to close position and left click
        pyautogui.moveTo(left + rel_x2, top + rel_y2)
        pyautogui.click(button='left')
    
    
    
    except:
    
        pass
    



#&&% 窗口内简单拖拽
def drag_in_window_simple(
    title: str,
    start: tuple[float,float],
    offset: tuple[float,float],
    duration: float = 0.3,
    button: str = 'left',
    absolute_start: bool = False
):
    """
    拖拽函数，支持相对或绝对起点：

    :param title: 窗口标题关键字
    :param start: 起点坐标，(x, y)；
                  如果 absolute_start==False，则当成“窗口内部”坐标
                  如果 absolute_start==True，则当成“屏幕”坐标
    :param offset: 拖拽向量 (dx, dy)
    :param duration: 拖动时长
    :param button: 'left' 或 'right'
    :param absolute_start: True 则 start 视为屏幕绝对坐标
    """
    left, top, w, h = activate_window_by_title(title)

    if absolute_start:
        x1, y1 = start
    else:
        x1 = left + start[0]
        y1 = top  + start[1]

    x2 = x1 + offset[0]
    y2 = y1 + offset[1]

    pyautogui.moveTo(x1, y1)
    pyautogui.mouseDown(button=button)
    pyautogui.moveTo(x2, y2, duration=duration)
    pyautogui.mouseUp(button=button)
    time.sleep(0.1)
    sys_logger.info(f"已在“{title}”窗口拖拽：起点{(x1,y1)} → 终点{(x2,y2)}")


"""
drag_in_window_simple(
    "图形导出",
    start=(10,10),
    offset=(100,50),
    absolute_start=False
)

"""


#&&% 纯窗口操作炸开区域内对象
def run_auto_explode_area(x1, y1, x2, y2, cmd_x, cmd_y, delay=2):

    """
    这是一个未使用pywin32API控制天正CAD的典型函数
    它炸开区域x1, y1, x2, y2的所有对象，不适合反复操作

    合理的处理方式还是配合视窗调整（发送z\na或e\n）先slect显性选择对象，再发送命令

    """
    
    script = Path(__file__).with_name('auto_EXPLODE.py')
    cmd = [
        sys.executable, str(script),
        f'--x1={x1}', f'--y1={y1}',
        f'--x2={x2}', f'--y2={y2}',
        f'--cmd_x={cmd_x}', f'--cmd_y={cmd_y}',
        f'--delay={delay}'
    ]
    # 隐藏子进程窗口
    flags = subprocess.CREATE_NO_WINDOW
    try:
        subprocess.run(cmd, check=True, creationflags=flags)
        print('▶ 子程序完成')
    except subprocess.CalledProcessError as e:
        print(f'[错误] 子程序退出码 {e.returncode}')


#&&% 列出所有窗口
def list_all_windows():
    # 获取所有可见的窗口
    windows = gw.getWindowsWithTitle('')
    
    if not windows:
        print("[错误] 没有找到任何可见窗口。")
    else:
        print("当前桌面上的所有窗口：")
        for win in windows:
            sys_logger.info(f"窗口标题: {win.title}, 窗口大小: {win.width}x{win.height}, 位置: ({win.left}, {win.top})")

##最小、最大化窗口
def minimize_window(window_keyword: str = 'OBS') -> bool:
    """
    通用：最小化第一个标题包含 window_keyword 的可见窗口。

    :param window_keyword: 要匹配的窗口标题关键字（子串匹配），默认 'OBS'
    :return: 如果成功最小化返回 True，否则返回 False
    """
    # 1) 找到所有匹配的可见窗口
    windows = [w for w in gw.getWindowsWithTitle(window_keyword) if w.visible]
    if not windows:
        print(f'❌ 未找到标题包含 "{window_keyword}" 的可见窗口')
        return False

    # 2) 取第一个并最小化
    win = windows[0]
    print(f'🔍 找到窗口: "{win.title}"，执行最小化')
    win.minimize()
    return True

def maximize_autocad_window(window_keyword: str = 'AutoCAD') -> bool:
    """
    强制最大化第一个标题包含 window_keyword 的可见窗口。
    优先尝试使用 win32gui，如不可用则退回 ctypes 调用 user32。
    """
    # 1) 找到目标窗口
    wins = [w for w in gw.getWindowsWithTitle(window_keyword) if w.visible]
    if not wins:
        sys_logger.info(f"❌ 未找到标题包含 “{window_keyword}” 的可见窗口")
        return False

    win = wins[0]
    hWnd = win._hWnd

    # 2) 先恢复（避免最小化状态），再最大化
    #    尝试使用 win32gui
    try:
        import win32gui, win32con
        win32gui.ShowWindow(hWnd, win32con.SW_RESTORE)
        time.sleep(0.1)
        win32gui.ShowWindow(hWnd, win32con.SW_MAXIMIZE)
    except ImportError:
        # 如果没有 pywin32，就退回 ctypes
        SW_RESTORE  = 9
        SW_MAXIMIZE = 3
        ctypes.windll.user32.ShowWindow(hWnd, SW_RESTORE)
        time.sleep(0.1)
        ctypes.windll.user32.ShowWindow(hWnd, SW_MAXIMIZE)

    time.sleep(0.2)  # 确保窗口完成最大化
    sys_logger.info(f"✅ 已将窗口 “{win.title}” 最大化")
    return True




# 天正墙中轴线显示与隐藏

#  该函数系列包括如下一些函数

"""
墙是否加粗边线，墙中线显示，墙中线隐藏是文件属性，因此我们用D:/Myprogramsystem/cad/xitongjicuwenjian/墙基线打开.dwg，

D:/Myprogramsystem/cad/xitongjicuwenjian/墙基线关闭.dwg，D:/Myprogramsystem/cad/xitongjicuwenjian/墙线加粗.dwg三个空文件来

控制文件和程序的运行。因为获取天正墙信息的函数依赖墙中线的显示，所以对不熟悉的文件进行处理时，就可以通过这几个基本文件实现确定

的控制，即我们预期要文件的墙中线显示出来，或者墙边线要加粗，不加粗，依赖实际的需要。




"""

#&&% 设置单位精度
def set_dwg_units_precision():
    """
    设置当前 DWG 文件的单位及精度：
    - 长度单位：单位类型不变，仅设置精度为 0.00000000
    - 角度单位：单位类型不变，精度为 0.00000000
    """
    doc=C.doc

    try:
        vars = doc.GetVariable

        # 设置长度精度（LUPREC = 8 表示 8 位小数）
        doc.SetVariable("LUPREC", 8)

        # 设置角度精度（AUPREC = 8 表示 8 位小数）
        doc.SetVariable("AUPREC", 8)

        print("[OK] 已将长度和角度单位精度设置为 8 位小数 (0.00000000)")
    except Exception as e:
        sys_logger.info(f"[错误] 设置失败: {e}")


def jd():
    set_dwg_units_precision()



#&&% 列出标注样式
def list_dim_styles():
    """
    列出当前 DWG 文件中所有标注样式名称。
    """
    try:
        styles = doc.DimStyles
        names = [styles.Item(i).Name for i in range(styles.Count)]
        print("📐 当前标注样式列表：")
        for name in names:
            print(" -", name)
        return names
    except Exception as e:
        sys_logger.info(f"[错误] 获取标注样式失败：{e}")
        return []



#&&% 设置当前标注样式
def set_current_dimstyle_via_command(style_name="_TCH_ARCH"):
    """
    使用命令行方式设置当前标注样式，兼容天正。

    参数:
        style_name (str): 要设为当前标注样式的名称（如 "_TCH_ARCH"）
    """
    try:
        doc.SendCommand(f"-DIMSTYLE\nR\n{style_name}\n")
        sys_logger.info(f"[OK] 已尝试通过命令行设置标注样式为：{style_name}")
    except Exception as e:
        sys_logger.info(f"[错误] 命令行设置标注样式失败：{e}")


        
#&&% 设置当前文字样式
def set_current_text_style(style_name="Standard"):
    """
    设置当前文字样式（通过 COM 接口方式）。

    参数:
        style_name (str): 要设置为当前的文字样式名称
    """
    try:
        text_styles = doc.TextStyles
        style = text_styles.Item(style_name)  # 获取指定样式
        doc.ActiveTextStyle = style           # 设置为当前样式
        sys_logger.info(f"[OK] 当前文字样式已设置为：{style_name}")
    except Exception as e:
        sys_logger.info(f"[错误] 设置文字样式失败：{e}")


#&&% 获取字体样式
def huoqu_ziti_style():

    styles = {ts.Name for ts in doc.TextStyles}
    
    return styles




#&&% 设置字体样式

"""
1，可多次对某种样式设置再更新

2，对汉字字体设置，不存在要设大字体。可用字体在C:/Windows/Fonts查找，直接用汉字名，使用create_text_style命令

3， 可以单独设置shx字体，使用set_text_style_onlyshx(style_name="style01", font_file="gbenor.shx", big_font_file=None)

所造字体缺失大字体，对汉字会产生问号

4，如不是单独设置汉字字体，完整的字体设置应包括英文部分的设置和大字体设置，用于中文显示而且是shx字体

 

"""
#&&% 创建文字样式
def create_text_style(sty_name="style01", ziti="宋体"):
    """
    在当前 DWG 中创建（或更新）一个中文文字样式。
    

    :param sty_name: 样式名称，默认 "style01"
    :param ziti:     字体名称，AutoCAD 会在系统字体中查找，例如 "宋体"
    # acad.ActiveDocument.ActiveTextStyle.SetFont(Typeface, Bold, Italic, charSet, PitchandFamily)
    # Typeface 字体名称；
    # Bold 加粗，布尔值，False为不加粗字体；
    # Italic 倾斜，布尔值，False为倾斜字体；
    # CharSet 字体字符集，1为默认字符集；
    # PitchAndFamily 字节及笔画形式。

    ts = doc.TextStyles.Item("HIT_TxtStyle")对shx

    ts.fontFile = "bigfont.shx"



    """

    # 1️⃣ 确保样式存在
    styles = doc.TextStyles
    try:
        ts = styles.Item(sty_name)
        sys_logger.info(f"[警告] 样式 '{sty_name}' 已存在，正在更新其属性。")
    except Exception:
        ts = styles.Add(sty_name)
        sys_logger.info(f"[OK] 已创建文字样式 '{sty_name}'。")

    # 2️⃣ 设置字体
    try:
        acad.ActiveDocument.TextStyles.Item(sty_name).SetFont(ziti, False, False, 1, 0 or 0)
    except Exception:
        try:
            acad.ActiveDocument.TextStyles.Item(sty_name).SetFont(ziti, False, False, 1, 0 or 0)
        except Exception as e:
            sys_logger.info(f"[警告] 无法设置字体为 '{ziti}'：{e}")

    # 3️⃣ 置为当前

    acad.ActiveDocument.ActiveTextStyle = acad.ActiveDocument.TextStyles.Item(sty_name)
    
    # 5️⃣ 通知用户
    sys_logger.info(f"[OK] 样式 '{sty_name}' 属性已更新")




#&&% 设置SHX文字样式
def set_text_style_onlyshx(style_name="style01", font_file="gbenor.shx", big_font_file=None):
    """
    C:/Program Files/Autodesk/AutoCAD 2021/Fonts查找可用shx字体    
    设置 AutoCAD 的文字样式：
    - font_file：英文字体（.shx 或 .ttf）
    - big_font_file：可选的大字体（如 gbcbig.txt），默认为 None 表示不设置，仅设置英文字体
    set_text_style_onlyshx(style_name="TEST_STYLE", font_file="gbxxx.shx", big_font_file=None)
    不成功消息可用于判定shx字体在不在左侧英文字体


    """
    try:
        import win32com.client

        acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
        doc = acad.ActiveDocument
        styles = doc.TextStyles

        # 查找或创建字体样式
        if style_name in [s.Name for s in styles]:
            text_style = styles.Item(style_name)
        else:
            text_style = styles.Add(style_name)

        text_style.FontFile = font_file

        # 只有传入合法的大字体路径才设置 BigFontFile，否则跳过设置
        if big_font_file and isinstance(big_font_file, str):
            text_style.BigFontFile = big_font_file

        sys_logger.info(f"字体样式 '{style_name}' 设置成功：英文字体 = {font_file}，大字体 = {big_font_file or '未设置'}")
        return True

    except Exception as e:
        sys_logger.info(f"设置字体样式失败：{e}")
        return False



#&&% 设置文字样式
def set_text_style(style_name="style01", font_file="gbenor.shx", big_font_file="gbcbig.shx"):
    """
    设置CAD中文字样式：英文shx文件 + 中文大字体文件
    """
    try:

        styles = doc.TextStyles

        # 判断是否已有该样式
        if style_name in [s.Name for s in styles]:
            text_style = styles.Item(style_name)
        else:
            text_style = styles.Add(style_name)

        # 设置字体和大字体
        text_style.FontFile = font_file
        text_style.BigFontFile = big_font_file

        sys_logger.info(f"字体样式 '{style_name}' 设置成功，英文字体: {font_file}, 中文大字体: {big_font_file}")
        return True

    except Exception as e:
        sys_logger.info(f"设置字体样式失败：{e}")
        return False

#&&% 列出可用shx非大字表 shx大字表 

"""

到CAD字体设置下拉菜单找

"""

##&&% 问号字体替换
"""
网友的ftst方案已经彻底解决此问题，我们需要的是备份好文件和进一步编制点击窗口的函数
我建议使用Standard样式的gbenor.shx和大字gbcbig.shx替换更安全
"""



#&&% 重命名冲突文字样式
def rename_conflicting_text_styles(file1_path: str,
                                   file2_path: str,
                                   suffix: str = "_1",
                                   retry_delay: float = 0.2,
                                   max_retries: int = 10):
    """
    在两个 DWG 中找出同名（用户）文字样式，
    并在第一个文件中将它们重命名（原名 + suffix）：
      1) 通过 -RENAME 命令重命名样式
      2) 确认样式表里旧名已消失、新名已出现
      3) REGEN 强制刷新
      4) 遍历 ModelSpace，将引用旧样式的实体指向新样式
      5) 保存并关闭
    系统默认样式会被自动跳过。
    """
    SYSTEM_STYLES = {"Standard", "ASHADE", "Annotative", "BigFont"}

    acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application")
    acad.Visible = True

    doc1 = acad.Documents.Open(os.path.abspath(file1_path))
    doc2 = acad.Documents.Open(os.path.abspath(file2_path))

    try:
        # 1) 收集两文件的样式
        styles1 = {ts.Name for ts in doc1.TextStyles}
        styles2 = {ts.Name for ts in doc2.TextStyles}
        conflicts = (styles1 & styles2) - SYSTEM_STYLES
        if not conflicts:
            print("[OK] 未发现需要重命名的用户样式。")
            return

        sys_logger.info(f"[警告] 发现同名用户样式：{conflicts}，将在 “{os.path.basename(file1_path)}” 中重命名：")
        ms = doc1.ModelSpace

        for old_name in conflicts:
            new_name = old_name + suffix
            # 如果新名也冲突，就多加后缀，直到独一
            while new_name in styles1:
                new_name += suffix

            # 2) 发送 RENAME 命令
            cmd = f"-RENAME\nStyle\n{old_name}\n{new_name}\n\n"
            doc1.SendCommand(cmd)
            # 等待命令被处理、样式表更新
            for attempt in range(max_retries):
                time.sleep(retry_delay)
                # 强制刷新图形状态
                doc1.SendCommand("REGEN\n")
                try:
                    # 如果旧名已不存在并且新名存在，就跳出重试
                    doc1.TextStyles.Item(new_name)
                    try:
                        doc1.TextStyles.Item(old_name)
                        # 旧名仍然存在，继续等
                        continue
                    except Exception:
                        # 旧名被正确移除
                        break
                except Exception:
                    # 新名还没出现，继续等
                    continue
            else:
                sys_logger.info(f"  ❗ 重命名 “{old_name}” → “{new_name}” 可能未生效（超时）。")

            # 3) 再把所有实体中引用旧样式的改成新样式
            for ent in ms:
                try:
                    ename = getattr(ent, "EntityName", "").upper()
                    oname = getattr(ent, "ObjectName", "")
                    # 原生 TEXT/MTEXT
                    if ename in ("TEXT", "MTEXT"):
                        if ent.TextStyle == old_name:
                            ent.TextStyle = new_name
                    # 天正文字
                    elif oname in ("TDbText", "TDbMText"):
                        if ent.TextStyle == old_name:
                            ent.TextStyle = new_name
                except Exception:
                    # 有些实体可能不允许改样式，忽略它们
                    pass

            # 更新本地样式集合
            styles1.discard(old_name)
            styles1.add(new_name)
            sys_logger.info(f"  · 样式 “{old_name}” → “{new_name}”")

        # 4) 最后保存并反馈
        doc1.Save()
        sys_logger.info(f"[OK] 已保存改动到 “{os.path.basename(file1_path)}”。")

    finally:
        # 关闭文档，不保存对第二个文件的任何改动
        doc1.Close(False)
        doc2.Close(False)

#&&&%CAD连接问题

"""

在 pywin32 (win32com) 操作 AutoCAD 的上下文中，Dispatch 和 EnsureDispatch 的核心区别在于 “后期绑定” (Late Binding) 与 “早期绑定” (Early Binding)，这直接影响到代码提示、运行速度以及常量 (Constants) 的使用。

以下是详细对比和 pywin32 中其他常见的 Dispatch 方式。

1. Dispatch vs. EnsureDispatch
win32com.client.Dispatch("AutoCAD.Application")
机制 (后期绑定):

这是最基础的连接方式。Python 在运行时才去询问 AutoCAD：“你有没有这个属性？你有没有这个方法？”

它不会生成 Python 包装文件（wrapper files）。

缺点:

无常量: 你无法使用 AutoCAD 的内置常量（如 acRed, acLine, acSelectionSetWindow）。你必须手动查找这些常量对应的整数值（例如用 1 代表红色）。

无代码提示: 在 IDE（如 PyCharm, VS Code）中，你基本上看不到任何自动补全，因为编辑器不知道返回的对象是什么。

速度稍慢: 每次调用都需要动态查询接口。

优点:

启动快（不需要生成中间文件）。

对版本不敏感（不需要重新生成缓存）。

win32com.client.EnsureDispatch("AutoCAD.Application")
机制 (早期绑定):

它会强制运行 makepy 过程。它会读取 AutoCAD 的类型库 (Type Library)，并在你的 Python site-packages/win32com/gen_py 目录下生成一套对应的 Python 源代码文件。

如果缓存文件已存在，它就直接加载；如果不存在，它会自动生成。

优点:

可以使用常量: 这是最大的优势。你可以直接使用 win32com.client.constants.acRed，而不需要查表找数字。

更好的类型支持: 某些特定的 COM 方法（涉及 ByRef 参数或复杂类型）在 Dispatch 下可能会报错，但在 EnsureDispatch 下能正常工作。

可能略快: 对于大量重复调用，因为它已经知道了方法的内存地址（Dispid），调用速度会略快。

缺点:

第一次运行会慢（因为要生成文件）。

如果 AutoCAD 升级了，可能需要手动删除 gen_py 缓存文件夹来强制重新生成。

直观对比示例
Python

import win32com.client

# --- 使用 Dispatch (后期绑定) ---
acad_late = win32com.client.Dispatch("AutoCAD.Application")
# 你必须知道 1 代表红色，0 代表 ByLayer
line = acad_late.ActiveDocument.ModelSpace.AddLine((0,0,0), (10,10,0))
line.Color = 1  # 必须写数字

# --- 使用 EnsureDispatch (早期绑定) ---
acad_early = win32com.client.EnsureDispatch("AutoCAD.Application")
# 现在可以使用常量了
from win32com.client import constants
line = acad_early.ActiveDocument.ModelSpace.AddLine((0,0,0), (10,10,0))
line.Color = constants.acRed  # 代码可读性大大增强
2. 总共有哪些 Dispatch 及其变体？
在 win32com.client 模块中，主要有以下几种获取 COM 对象的方式：

A. Dispatch
用途: 创建一个新的对象，或者连接到一个现有的对象（取决于该软件的 COM 实现，AutoCAD 通常是如果已打开则连接，未打开则新建）。

语法: win32com.client.Dispatch("ProgID")

B. EnsureDispatch (属于 gencache 模块)
用途: 强制生成类型库缓存（早期绑定），然后返回对象。

语法: win32com.client.EnsureDispatch("ProgID")

注意: win32com.client.gencache.EnsureDispatch 是其完整路径，client.EnsureDispatch 只是一个快捷方式。

C. DispatchEx
用途: 强制创建一个新的实例。

场景: 如果你已经打开了一个 AutoCAD 窗口，但你想在后台启动另一个完全独立的 AutoCAD 进程（不干扰前台操作），使用这个。

语法: win32com.client.DispatchEx("AutoCAD.Application")

区别: 普通的 Dispatch 会查看 Windows 的运行对象表 (ROT)，如果有现成的就拿来用；DispatchEx 忽略 ROT，直接让系统启动新进程。

D. GetActiveObject (及其重要)
用途: 仅连接到正在运行的实例。

场景: 如果 AutoCAD 没有打开，这个命令会直接报错 (com_error)。这通常用于脚本的开头，判断用户是否已经打开了软件。

语法: win32com.client.GetActiveObject("AutoCAD.Application")

E. CastTo (虽不是 Dispatch，但在 CAD 中必用)
用途: 类型转换（QueryInterface）。

场景: 当你通过选择集或遍历获得一个对象，它可能被泛型包装为 IAcadEntity。此时如果你想访问该对象特有的属性（例如动态块的 EffectiveName），直接访问可能报错。

语法: win32com.client.CastTo(obj, "IAcadBlockReference")

注意: 必须配合 EnsureDispatch 生成的库使用，因为它需要知道目标类型的定义。

3. AutoCAD 开发的最佳实践代码
为了兼顾稳定性和常量的使用，通常推荐的连接代码如下：

Python

import win32com.client
from win32com.client import constants
import sys

def connect_to_acad():
    try:
        # 1. 尝试连接已打开的 AutoCAD
        # 使用 gencache.EnsureDispatch 包装 GetActiveObject 以获得常量支持
        acad = win32com.client.GetActiveObject("AutoCAD.Application")
        
        # 这一步是为了确保 gen_py 缓存被加载，否则 GetActiveObject 拿到的可能还是后期绑定对象
        # 另一种方法是显式调用 gencache.EnsureModule(...)
        acad = win32com.client.gencache.EnsureDispatch("AutoCAD.Application") 
        
        print("已连接到正在运行的 AutoCAD 实例。")
        
    except Exception:
        try:
            # 2. 如果没打开，则启动新的实例
            acad = win32com.client.EnsureDispatch("AutoCAD.Application")
            acad.Visible = True # 默认可能不可见
            print("已启动新的 AutoCAD 实例。")
        except Exception as e:
            sys_logger.info(f"无法连接或启动 AutoCAD: {e}")
            return None

    return acad

acad_app = connect_to_acad()

# 测试常量是否可用
if acad_app:
    try:
        sys_logger.info(f"AutoCAD 红色常量值: {constants.acRed}") 
    except AttributeError:
        print("警告：常量未加载，可能是缓存问题。尝试删除 gen_py 文件夹。")
总结
简单脚本/不想生成缓存文件 → 用 Dispatch。

需要代码提示/需要使用 constants/复杂项目 → 用 EnsureDispatch。

需要强制多开后台进程 → 用 DispatchEx。

你是否遇到过 AttributeError: <unknown>.Name 这种错误？这通常就是因为使用了 Dispatch (后期绑定) 导致 Python 无法正确解析某些特定接口，切换到 EnsureDispatch 配合 CastTo 通常能解决。

"""

##将一个对象属性传给多个对象

#&&% 格式刷属性传递
def transfer_props_by_matchprop(entity, Ob, max_try=3, delay=0.4):

    CR = chr(13)

    def wait_idle(acad, dt=0.2, n=50):
        """轮询 IsQuiescent，最多 n 次，每次 dt 秒"""
        for _ in range(n):
            if acad.GetAcadState().IsQuiescent:
                return
            time.sleep(dt)

    def expand_rectangle(p1, p2, offset):
        return (p1[0]-offset, p1[1]-offset), (p2[0]+offset, p2[1]+offset)


    """
    把 entity 的属性批量复制到 Ob。若 Layer 未变化则重试，最多 3 次。
    """
    li()

    src_layer   = entity.Layer
    orig_layer  = Ob.Layer        # 复制前目标的图层

    # 目标包围盒窗口
    p1, p2 = Ob.GetBoundingBox()
    x1, y1, x2, y2 = p1[0], p1[1], p2[0], p2[1]
    h = 0.1 * (abs(x1 - x2) + abs(y1 - y2)) / 2
    (wx1, wy1), (wx2, wy2) = expand_rectangle(p1, p2, h)

    match_cmd = (
        "_MATCHPROP" + CR +
        "P"          + CR +            # Previous 作为源
        "_W"         + CR +
        f"{wx1},{wy1}" + CR +
        f"{wx2},{wy2}" + CR + CR
    )

    for attempt in range(1, max_try + 1):
        try:
            # ——— 1. 设定源对象为 Previous ———
            try:
                highlight_entity_by_bbox(entity)
            except Exception as e:
                # 退而求其次：使用 LISP 通过 Handle 选择对象
                sys_logger.info(f"[FALLBACK] highlight_entity_by_bbox 失败，使用 Handle 选择: {e}")
                handle = com_retry(lambda: entity.Handle)
                # 使用 LISP 的 ssget 创建选择集并显示
                doc.SendCommand(f"(sssetfirst nil (ssget \"_X\" '((5 . \"{handle}\"))))\n")
                time.sleep(0.5)
            time.sleep(delay)

            # ——— 2. 发送 MATCHPROP ———
            doc.SendCommand(match_cmd)
            wait_idle(acad)

            # ——— 3. 判断是否成功 ———
            if Ob.Layer == src_layer:
                sys_logger.info(f"[OK] 第 {attempt} 次匹配成功，Layer 改为 {src_layer}")
                return True

            sys_logger.info(f"[WARN] 第 {attempt} 次后 Layer 未变，重试…")
            time.sleep(delay)

        except Exception as e:
            sys_logger.info(f"[ERR] 第 {attempt} 次匹配异常：{e}")

    sys_logger.info(f"[FAIL] 连续 {max_try} 次仍未把属性复制给目标")
    return False


#视图合理化控制
"""
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_swiso"+chr(13))#西南轴测
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_seiso"+chr(13))#东南轴测
acad.ActiveDocument.SendCommand(
    "_-view" + chr(13) +
    "_nwiso" + chr(13)
)#西北轴测
acad.ActiveDocument.SendCommand(
    "_-view" + chr(13) +
    "_neiso" + chr(13)
)#东北轴测
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_top"+chr(13))#俯视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_bottom"+chr(13))#仰视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_front"+chr(13))#前视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_Right"+chr(13))#右视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_Back"+chr(13))#后视图
acad.ActiveDocument.SendCommand("_-view"+chr(13)+"_Left"+chr(13))#左视图
acad.ActiveDocument.SendCommand("_vscurrent"+chr(13)+"_R"+chr(13))#真实视觉样式

##合理显示对象

acad.ActiveDocument.SendCommand(
    "_z" + chr(13) +
    "_e" + chr(13)
)

acad.ActiveDocument.SendCommand(
    "_zoom" + chr(13) +
    "s"     + chr(13) +
    "0.8xp" + chr(13)
)




"""
#&&% 双线程生成器

#&&% 双线程运行1
def run_dual_threads_1(f1,                     # 线程1 函数

                     f2,                     # 线程2 函数

                     f1_args=(), f1_kwargs=None,

                     f2_args=(), f2_kwargs=None,

                     timeout_sec=180):



    """
    通用“双线程-GUI”调度器

    - f1 必须负责触发一个阻塞性窗口（如 SendCommand、ShowModalDlg 等）。
    - f2 负责侦测并自动化处理该窗口，当处理完毕后通知 f1 继续或退出。
    - f1、f2 的函数签名都应以 (timeout_event, done_event, …) 开头：
    
      def f1(timeout_event, done_event, …):
          pythoncom.CoInitialize()
          try:
              # ……主体代码：例如触发打印对话框
              timeout_event.wait()    # 等待线程2通知或超时
          except Exception:
              # 打印日志或其他处理
          finally:
              pythoncom.CoUninitialize()
              done_event.set()        # 通知“双线程”总控：我结束了

      def f2(timeout_event, done_event, …):
          pythoncom.CoInitialize()
          try:
              # ……主体代码：例如等待“创建打印文件”对话框出现并点击“回车”键
          except Exception:
              # 打印日志或其他处理
          finally:
              pythoncom.CoUninitialize()
              timeout_event.set()    # 通知 f1 线程：窗口已处理完，可以结束等待
              done_event.set()       # 通知“双线程”总控：我结束了

    - 参数说明：
        f1, f2：传入上面那种签名的函数
        f1_args、f1_kwargs：给 f1 传递的位置参数和关键字参数
        f2_args、f2_kwargs：给 f2 传递的位置参数和关键字参数
        timeout_sec：最多等待多少秒，如果超过则报告超时并返回 False

    - 返回值：
        True  ：两个子线程在 time_limit 内都完成了
        False ：超时，至少一个线程没及时调用 done_event.set()
    """
    if f1_kwargs is None:
        f1_kwargs = {}
    if f2_kwargs is None:
        f2_kwargs = {}

    # ① 为线程 1、2 准备同一个事件对
    timeout_event = threading.Event()
    done_event    = threading.Event()

    # ② 启动“线程1”：负责弹出、阻塞窗口
    t1 = threading.Thread(
        target=f1,
        args=(timeout_event, done_event, *f1_args),
        kwargs=f1_kwargs,
        daemon=True
    )
    # ③ 启动“线程2”：负责侦测窗口并点击、关闭
    t2 = threading.Thread(
        target=f2,
        args=(timeout_event, done_event, *f2_args),
        kwargs=f2_kwargs,
        daemon=True
    )

    start_time = time.time()
    t1.start()
    t2.start()

    # ④ 等待线程1 在 timeout_sec 秒内结束
    t1.join(timeout=timeout_sec)
    
    t2.join(timeout=timeout_sec)
    
        
        

    # ⑥ 检查 done_event 是否已被 set()
    if not done_event.is_set():
        # 超时：由调度器负责给线程1 发送退出信号
        timeout_event.set()
        sys_logger.info(f"[警告] 双线程执行超过 {timeout_sec}s —— 已触发 timeout_event")
        return False
    else:
        print("[OK] 双线程任务在时限内完成")
        return True



#&&% 取消CAD选择
def cancel_cad_selection(attempts: int = 3, delay: float = 0.5) -> bool:

    for i in range(1, attempts + 1):
        try:
            highlight_entities_in_window(0, 0, 0, 0)
            sys_logger.info(f"[OK] 第{i}次尝试：cancel_cad_selection 成功")
            return True
        except Exception as e:
            sys_logger.info(f"[警告] 第{i}次尝试失败：{e}")
            if i < attempts:
                time.sleep(delay)
    print("[错误] 已重试多次，仍未能执行 cancel_cad_selection")
    return False



#&&&% 打印辅助

def close_wps_window_by_click(
    title_keyword: str = "WPS Office",
    offset_x: int = 22,    # 距窗口右边缘的像素
    offset_y: int = 19,    # 距窗口上边缘的像素
    pause_before: float = 0.2
) -> bool:
    """
    在标题包含 title_keyword 的窗口右上角点击一次（×），尝试关闭该窗口。
    如果没找到窗口，则返回 False 并跳过。
    """
    # 枚举可见窗口，匹配标题
    handles = []
    def _enum(hwnd, lst):
        if win32gui.IsWindowVisible(hwnd):
            txt = win32gui.GetWindowText(hwnd) or ""
            if title_keyword.lower() in txt.lower():
                lst.append(hwnd)
    win32gui.EnumWindows(_enum, handles)

    if not handles:
        node("ℹ️ 未找到标题含 ‘{}’ 的窗口，跳过关闭", title_keyword)
        return False

    hwnd = handles[0]
    left, top, right, bottom = win32gui.GetWindowRect(hwnd)
    x = right - offset_x
    y = top   + offset_y

    time.sleep(pause_before)
    pyautogui.click(x, y)
    node("✅ 点击 '{}' 窗口(句柄 {}) 的关闭按钮坐标：({}, {})",
         title_keyword, hwnd, x, y)
    return True






#&&&% 测试辅助

#&&% 最小化窗口
def min_w():
    import ctypes
    
    VK_MENU = 0x12  # Alt键
    VK_TAB = 0x09   # Tab键
    VK_LWIN = 0x5B  # 左Win键
    VK_M = 0x4D     # M键
    KEYEVENTF_KEYUP = 0x2

    # 模拟Alt + Tab
    ctypes.windll.user32.keybd_event(VK_MENU, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_TAB, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_TAB, 0, KEYEVENTF_KEYUP, 0)
    ctypes.windll.user32.keybd_event(VK_MENU, 0, KEYEVENTF_KEYUP, 0)

    # 模拟Win + M
    ctypes.windll.user32.keybd_event(VK_LWIN, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_M, 0, 0, 0)
    ctypes.windll.user32.keybd_event(VK_M, 0, KEYEVENTF_KEYUP, 0)
    ctypes.windll.user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)




#&&% 清除测试图层
def ql():#清除测试辅助图层上的对象

    ensure_layer("测试辅助")

    




    
#&&% 模型空间画点
def srhd(*args):#根据输入坐标在模型空间画点
    """
    在模型空间绘制点并标注序号，支持以下调用形式：
    - srhd((x1,y1,z1), (x2,y2,z2))   # 多个点元组
    - srhd([(x1,y1,z1), (x2,y2,z2)]) # 列表形式
    - srhd((x1,y1,z1))               # 单点
    """
    doc = acad.ActiveDocument
    ms = doc.ModelSpace
    layer_name = "测试辅助"

    # 创建图层（如果不存在）
    try:
        doc.Layers.Item(layer_name)
    except:
        doc.Layers.Add(layer_name)
        sys_logger.info(f"[OK] 已创建图层：{layer_name}")

    # 统一格式处理：允许单点、多个点、列表传入
    if len(args) == 1:
        if isinstance(args[0], (list, tuple)):
            if len(args[0]) == 3 and all(isinstance(i, (int, float)) for i in args[0]):
                points = [args[0]]
            else:
                points = args[0]
        else:
            print("[错误] 输入格式不正确")
            return
    else:
        points = args

    # 绘制点与编号
    for idx, P in enumerate(points, 1):
        try:
            pt = vtpnt(*P)
            point = ms.AddPoint(pt)
            point.Layer = layer_name

            text = ms.AddText(str(idx), pt, 10)
            text.Layer = layer_name
        except Exception as e:
            sys_logger.info(f"[错误] 添加点失败: {e}")

    return "[OK] 点与编号已绘制"



#&&% 图纸空间画点
def srhd_p(*args):#根据输入坐标在图纸空间画点
    """
    在图纸空间绘制点和编号，支持：
    - 多个坐标元组：srhd_p((x1,y1,z1), (x2,y2,z2))
    - 单个列表：srhd_p([(x1,y1,z1), (x2,y2,z2)])
    - 单个点：srhd_p((x1,y1,z1))
    """
    ps = doc.PaperSpace
    layer_name = "测试辅助"

    # 创建图层（如果不存在）
    try:
        doc.Layers.Item(layer_name)
    except:
        doc.Layers.Add(layer_name)
        sys_logger.info(f"[OK] 已创建图层：{layer_name}")

    # 判断参数结构
    if len(args) == 1:
        if isinstance(args[0], (list, tuple)):
            if len(args[0]) == 3 and all(isinstance(i, (int, float)) for i in args[0]):
                # 单个点元组
                points = [args[0]]
            else:
                # 列表形式
                points = args[0]
        else:
            print("[错误] 输入格式不正确")
            return
    else:
        # 多个点元组
        points = args

    for idx, P in enumerate(points, 1):
        try:
            pt = vtpnt(*P)
            point = ps.AddPoint(pt)
            point.Layer = layer_name

            text = ps.AddText(str(idx), pt, 10)
            text.Layer = layer_name
        except Exception as e:
            sys_logger.info(f"[错误] 添加点失败: {e}")

    return "[OK] 图纸空间中的点与编号已绘制"


#&&% COM点转数学点
def comtomath(LBcom):#将com点列表转为数学点列表

    LB_point=[]

    for i in range(0,len(LBcom)):

        point = LBcom[i].Coordinates

        LB_point.append(point)

    return LB_point        

def p():

    li()

    oblist= pmxz()
    LBpoint=comtomath(oblist)
    return print(LBpoint)




        

#&&% 隔远查看

#&&% 复制查看
def fuzhi_chakan(LBcom,K=1):#K为放大倍数

    LK=[]

    for xx in LBcom:

        try:

            copy_obj = xx.Copy()
    
            point1 = vtpnt(0,0,0)
    
            point2 = vtpnt(0,K*1000000,0)
    
            copy_obj.Move(point1,point2)


            LK.append(copy_obj)

        except Exception as e:

            sys_logger.info(f"[错误] 移动对象{xx.Handle}失败: {e}")               

    return LK



#&&% 测量文字长度
def celiang_wenzichangdu(TEXTCOM):

    text_copy = TEXTCOM.Copy()

    text_copy.Alignment = 2

    text_copy.TextAlignmentPoint =vtpnt(0,0,0)

    chang = abs(text_copy.InsertionPoint[0])

    text_copy.Delete()

    return chang

#测量新写文字长度

#&&% 写入并测量文字长度
def celiang_wenzichangdu_write(ZF,style="图签",height=270,scalefactor=0.8):

    #根据字符串按样式字高宽度因子写入cad后的测量长度

    text_obj = acad.ActiveDocument.ModelSpace.AddText(ZF, vtpnt(0,0,0), height)

    text_obj.StyleName = style

    text_obj.ScaleFactor =scalefactor #宽度因子

    chang = celiang_wenzichangdu(text_obj)

    text_obj.Delete()

    return chang




##清空文件夹
#&&% 清空文件夹
def qingkong_wenjianjia(FolderPath):

     #清空文件夹B
    folder_path_1 = FolderPath 

    # 遍历文件夹中的所有文件
    for filename in os.listdir(folder_path_1):
        
        file_pathx = os.path.join(folder_path_1, filename)
        
        # 确保它是一个文件而不是文件夹
        if os.path.isfile(file_pathx):
            
            os.remove(file_pathx)  # 删除文件

        sys_logger.info(f"{FolderPath}文件夹已清空")


#&&% 返回对象外包盒的长，宽，横竖向，角点信息

#&&% 获取包围盒信息
def get_bbox_info(com_obj):
    """
    获取传入 AutoCAD COM 对象（如 Line、Circle、BlockReference 等）的外包盒信息，
    并计算其长（length）、宽（width）以及横向或竖向（orientation）状态。

    参数：
      com_obj -- 任意支持 GetBoundingBox() 方法的 AutoCAD COM 对象

    返回：
      (length, width, orientation)
        length      -- 外包盒较长一边的长度（在 X/Y 平面上）
        width       -- 外包盒较短一边的长度（在 X/Y 平面上）
        orientation -- 字符串：
                         "horizontal" 表示 X 方向跨度 ≥ Y 方向跨度，
                         "vertical"   表示 Y 方向跨度 >  X 方向跨度
    如果对象不支持 GetBoundingBox，会抛出异常；也可根据需要自行捕获处理。
    """
 


   # 调用 GetBoundingBox 方法，返回两个点：minPt、maxPt
    # minPt、maxPt 都是 3 元素的 tuple 或 list，形如 (x, y, z)
    




    try:

        minPt, maxPt = com_obj.GetBoundingBox()

    except Exception as e:

        sys_logger.info(f"获取外包盒失败: {e}")               

        return None


    # 计算 X、Y 方向跨度
    dx = maxPt[0] - minPt[0]
    dy = maxPt[1] - minPt[1]

    # 将较大值定义为 length，较小值定义为 width
    length = max(dx, dy)
    width  = min(dx, dy)

    # 判断横向（X 跨度 ≥ Y 跨度）还是竖向（Y 跨度 > X 跨度）
    if dx >= dy:
        orientation = "horizontal"
    else:
        orientation = "vertical"

    return minPt, maxPt,length, width, orientation


#&&% 包围盒方向标志
def bbox_orientation_flag(com_obj):
    """
    判断任意 COM 对象的外包盒是竖向、横向还是正方形：
      - 如果 Y 方向跨度 > X 方向跨度，返回 1（竖向）
      - 否则（包括 X 方向跨度 >= Y 方向跨度），返回 0
        —— 即当外包盒为正方形（X 跨度 == Y 跨度）时，也返回 0

    参数：
      com_obj -- 支持 GetBoundingBox() 的 AutoCAD COM 对象

    返回：
      int -- 竖向返回 1，横向或正方形返回 0
    """
    # 获取外包盒的最小点和最大点
    min_pt, max_pt = com_obj.GetBoundingBox()
    # 计算 X、Y 方向跨度
    dx = abs(max_pt[0] - min_pt[0])
    dy = abs(max_pt[1] - min_pt[1])
    # 如果是竖向（dy > dx），返回 1；否则（横向或正方形）返回 0
    return 1 if dy > dx else 0

#&&% 获取多个对象的外包盒数据

def group_bbox_corners(com_objs, max_retry=3, delay=0.05):
    """
    【通用加强版】计算一组 COM 对象的整体外包盒，并按顺序返回四个角点。
    
    集成特性：
    1. 包含 safe_get_bbox 的重试与刷新逻辑，防止因单个对象 "Call Rejected" 导致失败。
    2. 自动过滤无法获取几何信息的对象（如空文字、未生成的实体）。
    3. 自动解包 COM 数据类型。

    参数：
        com_objs: 可迭代的 COM 对象列表
        max_retry: 单个对象获取失败时的重试次数
        delay: 重试间隔

    返回：
        四元组：(左下, 右上, 左上, 右下)
        每个点为 (x, y, z)。如果不含有效对象，返回 None。
    """
    import math
    import time
    from win32com.client import VARIANT

    # 初始化为极端值
    global_min_x = float('inf')
    global_min_y = float('inf')
    global_max_x = float('-inf')
    global_max_y = float('-inf')
    global_min_z = float('inf')
    
    valid_count = 0

    for obj in com_objs:
        if obj is None: continue

        # --- 单个对象的安全获取逻辑 (内嵌 safe_get_bbox 核心) ---
        p1, p2 = None, None
        
        for k in range(max_retry):
            try:
                # [关键] 强制刷新，防止新对象无几何数据
                obj.Update()
                
                # 获取包围盒
                min_var, max_var = obj.GetBoundingBox()
                
                # [清洗] 确保转为 list/tuple，防止 SafeArray 报错
                p1 = list(min_var) if hasattr(min_var, '__iter__') else min_var
                p2 = list(max_var) if hasattr(max_var, '__iter__') else max_var
                
                # 成功获取，跳出重试循环
                break 
            except Exception:
                # 遇到错误（如 CAD 忙），稍作休眠后重试
                if k < max_retry - 1:
                    time.sleep(delay)
                else:
                    # 彻底失败，p1 p2 保持为 None
                    pass
        
        # 如果该对象获取失败，跳过它，处理下一个
        if p1 is None or p2 is None:
            continue
        
        # --- 数据更新 ---
        valid_count += 1
        x1, y1, z1 = p1[0], p1[1], p1[2]
        x2, y2, z2 = p2[0], p2[1], p2[2]

        # 更新 X/Y/Z 极值
        if x1 < global_min_x: global_min_x = x1
        if y1 < global_min_y: global_min_y = y1
        if x2 > global_max_x: global_max_x = x2
        if y2 > global_max_y: global_max_y = y2
        
        # 统一使用所有对象中最低的 Z 值
        if z1 < global_min_z: global_min_z = z1

    # 结果校验：如果没有一个有效对象，返回 None
    if valid_count == 0:
        return None

    # 构造结果
    z = global_min_z
    
    # 1. 左下 (Min X, Min Y)
    bottom_left  = (global_min_x, global_min_y, z)
    # 2. 右上 (Max X, Max Y)
    top_right    = (global_max_x, global_max_y, z)
    # 3. 左上 (Min X, Max Y)
    top_left     = (global_min_x, global_max_y, z)
    # 4. 右下 (Max X, Min Y)
    bottom_right = (global_max_x, global_min_y, z)

    return bottom_left, top_right, top_left, bottom_right



#&&% zip的用法
"""
seq_x0  = (x0, y0, z0)
P_start = (dx, dy, dz)
要得到新坐标 (x0+dx, y0+dy, z0+dz)，就可以写：
seq_x0 = tuple(a + b for a, b in zip(seq_x0, P_start))
zip(seq_x0, P_start) 会产出 (x0, dx), (y0, dy), (z0, dz)
a + b for a, b in ... 就对每一对做相加
最后用 tuple(...) 把结果收成三元组
对两个列表求和
xs = [10, 20, 30]
ys = [1, 2, 3]
sums = [x + y for x, y in zip(xs, ys)]
# sums == [11, 22, 33]
并行遍历三组数据
names = ["A", "B", "C"]
ages  = [30, 25, 40]
scores= [85, 92, 78]
for n, a, s in zip(names, ages, scores):
    sys_logger.info(f"{n} 年龄{a} 分数{s}")
解压
如果有一个列表 pairs = [(1,4),(2,5),(3,6)]，要拆成两个列表：
a, b = zip(*pairs)
# a == (1,2,3), b == (4,5,6)

"""

#&&% 从两点绘制矩形


#&&% 包围盒中心2D
def bbox_center_2(e):
    # GetBoundingBox 返回两个 Point (minPt, maxPt)
    min_pt, max_pt = e.GetBoundingBox()
    x1, y1, _ = tuple(min_pt)
    x2, y2, _ = tuple(max_pt)
    return ((x1 + x2) / 2, (y1 + y2) / 2)

#&&% 包围盒中心3D
def bbox_center_3(ent):
    """返回实体外包盒中心 (cx, cy, cz)"""
    mn, mx = ent.GetBoundingBox()
    return ((mn[0] + mx[0]) / 2.0,
            (mn[1] + mx[1]) / 2.0,
            (mn[2] + mx[2]) / 2.0)

#&&% 安全获取包围盒
def safe_get_bbox(ent, max_retry=5, delay=0.05):
    """
    【通用加强版】安全获取 AutoCAD 图元外包盒 (GetBoundingBox)
    
    解决痛点：
    1. 新创建的对象（如 Text/Block）未计算几何数据导致的 "Empty Extents" (-2145386308)。
    2. AutoCAD 忙碌导致的 "Call Rejected" (-2147417846)。
    3. 偶发的通用 COM 错误。
    
    参数:
        ent: AutoCAD 图元对象 (IAcadEntity)
        max_retry: 最大重试次数 (默认5次)
        delay: 失败后的休眠秒数 (默认0.05秒)
        
    返回:
        (min_point, max_point) 元组，坐标为 (x, y, z)。
        如果彻底失败，返回 None。
    """


    # 如果对象本身就是空的，直接返回
    if ent is None:
        return None

    for i in range(max_retry):
        try:
            # [关键步骤] 强制刷新对象数据库
            # 这是解决 "刚创建的文字取不到包围盒" 的核心
            ent.Update()

            # 获取包围盒
            min_var, max_var = ent.GetBoundingBox()

            # [数据清洗] 确保拿到的是 Python 可用的数据
            # 有时 win32com 返回的是 tuple，有时是 SafeArray，这里做个转换保平安
            p1 = list(min_var) if hasattr(min_var, '__iter__') else min_var
            p2 = list(max_var) if hasattr(max_var, '__iter__') else max_var

            return p1, p2

        except com_error as e:
            hresult = e.args[0] if e.args else 0
            
            # 常见错误代码分析：
            # -2147417846: RPC_E_CALL_REJECTED (CAD 忙，例如正在自动保存或执行命令)
            # -2145386308: eInvalidExtents (对象无几何范围，常见于空文字或新对象)
            # -2147352567: 通用错误
            
            # 如果是最后一次尝试，或者错误严重，则记录日志（可选）
            if i == max_retry - 1:
                # sys_logger.info(f"[BBox最终失败] Handle={getattr(ent, 'Handle', 'Unknown')}, Err={hresult}")
                pass
            else:
                # 还有机会，休眠后重试
                time.sleep(delay)
                continue
                
        except Exception as e:
            # 捕获其他非 COM 错误（如 Python 逻辑错误）
            if i < max_retry - 1:
                time.sleep(delay)
                continue

    return None




#&&&&%% 第八部分 工程图纸基础服务系统


"""
标准图签模板在D:/Myprogramsystem/XT/标准图签模板.dwg
属性图签的同名图块，在进行内部的属性文字移动位置后，命令行窗口执行ATTSYNC，按提示操作完，所有同名属性块文字都会移动到新的位置
这就解决了图签内容随不同公司图签变换位置的问题。
每个公司的自定义内容属于固定块，我们不需要管。
我们给出的是通用语义的值，例如项目名称的值，你叫项目大名称也无所谓。
通过统一的属性文字操作，能够快速将某个标签值设为空，这就适应了不同公司的需要，实现了万能图签的目标。
重复操作插入之前会先清空默认的"dy_quyu"，以不影响函数的运行
属性块制作
1 ATT 输入命令行
2 选择多行文字，字体，字体大小，文字边界，正中
3 ATT 输入命令行
4 选择单行文字，字体，字体大小，文字边界，正中
5 检查好字体，图层，每个公司内容有点差别，文字图层统一为“图签目录”，注意每个图签的多行文字边界
6 ATTMODE 2 ATTDISP ON，确保设置好这两个值再做块

根据20251126的最新函数，文字边界宽度可以设到12000,不需要通过文字边界换行，而是可以自由换行

LB=select_print_areas_smart()
ok,bind,info =stable_insert_and_scale_labels_area(
    LB,
    filepath=r"D:/Myprogramsystem/XT/标准图签模板.dwg",
    layername="dy_quyu",
    timestamp=None,
    
    core_layer="dy_quyu_H",
    max_try=3,
    verbose=True,
    aggressive_purge=False,  # 新增：是否启用强力无实例块清理
)

特别是使用混合样式，一定要清理干净文件，不仅仅是块，不要有乱七八糟的选区。




"""


#&&&% ===（零）调试和数据处理 ===


#&&% 调试风格规范


#
"""
核心设计思想：verbose_level (详细等级)
用一个整数 verbose_level 代替 mute_logs：

0 (Silent): 全局静音，生产模式。

1 (Summary): 仅显示主函数的关键步骤（如“开始扫描”、“找到5个”），不显示循环内的细节。

2 (Trace): 显示主函数的循环细节（如“对象A跳过”，“对象B入选”），但强行静音被引用的子函数。

3 (Deep Debug): 火力全开，主函数和子函数的所有细节都打印。

示例
select_print_areas_paperspace

check_valid_rect_pro



"""
#XVX 调试
# ==========================================
# 1. 全局总控 (建议放在脚本头部)
# ==========================================
# 0=静音, 1=摘要, 2=追踪, 3=调试
GLOBAL_DEFAULT_VERBOSE = 1 

def resolve_log_level(local_setting):
    """解析日志等级: 参数 > 全局"""
    if local_setting is not None:
        return int(local_setting)
    return GLOBAL_DEFAULT_VERBOSE


#&&% 数据处理中心
# ========================================================
# 1. 基础配置与连接模块
# ========================================================

def get_data_root():
    """获取数据存储根目录"""
    env_path = os.environ.get('USERPATH')
    if env_path and os.path.exists(env_path):
        return env_path
    return r"D:\Myprogramsystem\XT\文件字典信息"

def _resolve_json_path(file_input, folder_name):
    """统一路径解析工具"""
    root = get_data_root()
    save_dir = os.path.join(root, "配置", folder_name)
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    if not file_input:
        try:
            doc = C.doc # 尝试从全局获取
        except: 
            # 临时连接兜底
            try: doc = win32com.client.Dispatch("AutoCAD.Application").ActiveDocument
            except: doc = None
            
        if doc:
            base = os.path.splitext(doc.Name)[0]
            try: layout = doc.ActiveLayout.Name
            except: layout = "Model"
            fname = f"{base}_{layout}.json"
        else:
            fname = "Unknown_Data.json"
    else:
        fname = str(file_input)
        if not fname.lower().endswith(".json"):
            fname += ".json"
    
    return os.path.join(save_dir, fname)

# ========================================================
# 2. 核心数据处理：提取与自适应恢复
# ========================================================

def extract_poly_data(poly_obj):
    """
    【存储】提取多段线数据 (改为 getattr 安全读取)
    """
    try:
        # 安全获取对象类型
        obj_name = getattr(poly_obj, "ObjectName", "")
        if not obj_name.endswith("Polyline"):
            return None

        # 使用 getattr 安全获取属性，提供默认值
        data = {
            "handle": getattr(poly_obj, "Handle", None),
            "layer":  getattr(poly_obj, "Layer", "0"),
            "closed": getattr(poly_obj, "Closed", False),
            "width":  getattr(poly_obj, "ConstantWidth", 0.0),
            "coords": list(getattr(poly_obj, "Coordinates", [])) # 转list存储
        }
        
        # 如果没有 Handle，说明对象无效
        if not data["handle"]: return None
        return data
        
    except Exception as e:
        # sys_logger.info(f"提取失败: {e}")
        return None


def restore_poly_adaptive(data_dict):
    """
    【加载】自适应恢复 (改为 set_attr 安全赋值)
    """
    try:
        doc = C.doc
    except:
        try: doc = win32com.client.Dispatch("AutoCAD.Application").ActiveDocument
        except: return None

    # --- 方案 A: 尝试通过 Handle 找回 ---
    handle = data_dict.get("handle")
    if handle:
        try:
            obj = doc.HandleToObject(handle)
            # 验证存活：尝试访问 Layer
            _ = getattr(obj, "Layer", None)
            return obj 
        except Exception:
            pass # Handle无效，进入重绘逻辑

    # --- 方案 B: 重绘 (复活) ---
    try:
        raw_coords = data_dict.get("coords")
        if not raw_coords or len(raw_coords) < 4: 
            return None
        
        # 坐标清洗
        safe_coords = [float(x) for x in raw_coords]
        if len(safe_coords) % 2 != 0: safe_coords.pop()

        # 强制转换为双精度浮点数组
        points_variant = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, safe_coords)

        # 属性准备
        layer = data_dict.get("layer", "0")
        width = data_dict.get("width", 0.0)
        is_closed = data_dict.get("closed", False)

        # 确定绘制空间
        try: target_space = doc.ActiveLayout.Block
        except: target_space = doc.ModelSpace

        # 绘制
        new_pl = target_space.AddLightWeightPolyline(points_variant)
        
        # --- 属性恢复 (使用 set_attr 风格) ---
        # 如果全局有 set_attr 则使用，否则使用本地安全赋值
        if 'set_attr' in globals():
            set_attr(new_pl, "Layer", layer)
            set_attr(new_pl, "ConstantWidth", float(width))
            set_attr(new_pl, "Closed", is_closed)
        else:
            # 本地简易版 set_attr
            try: new_pl.Layer = layer
            except: pass
            try: new_pl.ConstantWidth = float(width)
            except: pass
            try: new_pl.Closed = is_closed
            except: pass

        return new_pl
        
    except Exception as e:
        sys_logger.info(f"❌ 重绘失败: {e}")
        return None


#&&% 多段线列表信息存取

# ========================================================
# 3. 业务功能：多段线列表存取
# ========================================================

def save_poly_list(poly_list, file_name=None):
    """保存多段线列表到 JSON"""


    poly_list = ensure_list(poly_list)
    
    if not poly_list: return False
    
    path = _resolve_json_path(file_name, "poly_lists")
    data_list = []
    
    for pl in poly_list:
        d = extract_poly_data(pl)
        if d: data_list.append(d)
        
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data_list, f, indent=4)
        sys_logger.info(f"✅ 保存列表成功: {os.path.basename(path)}")
        return True
    except Exception as e:
        sys_logger.info(f"❌ 保存列表失败: {e}")
        return False

def load_poly_list(file_name=None):
    """加载 JSON 并恢复多段线列表"""
    path = _resolve_json_path(file_name, "poly_lists")
    if not os.path.exists(path):
        print("❌ 文件不存在")
        return []
    
    try:
        with open(path, 'r', encoding='utf-8') as f:
            raw_data = json.load(f)
            
        restored = []
        for d in raw_data:
            obj = restore_poly_adaptive(d)
            if obj: restored.append(obj)
            
        sys_logger.info(f"✅ 加载完成，有效对象: {len(restored)} 个")
        return restored
    except Exception as e:
        sys_logger.info(f"❌ 加载列表失败: {e}")
        return []

#&&% ctq信息存取


# ========================================================
# 4. 业务功能：CTQ (图签系统) 存取
# ========================================================

def save_ctq(ctq_data, file_name=None):
    """
    【功能】: 保存 (polys, blocks, mapping) 到 JSON。
    """
    if not ctq_data or len(ctq_data) < 3 or not ctq_data[2]: return False
    
    polys, blocks, mapping = ctq_data
    export_list = []
    
    # 建立 Handle -> Object 的查找表 (Handle 是唯一的，做 Key 没问题)
    p_map = {}
    for p in polys:
        h = getattr(p, "Handle", None)
        if h: p_map[h] = p
    
    sorted_keys = sorted(mapping.keys())
    for seq in sorted_keys:
        p_handle, b_handle = mapping[seq]
        
        # 通过 Handle 找对象
        p_obj = p_map.get(p_handle)
        p_data = extract_poly_data(p_obj) if p_obj else None
        
        if p_data:
            export_list.append({
                "seq": seq,
                "poly_data": p_data,      
                "block_handle": b_handle  
            })
            
    f_path = _resolve_json_path(file_name, "ctq_data")
    try:
        with open(f_path, 'w', encoding='utf-8') as f:
            json.dump(export_list, f, indent=4)
        return True
    except: return False


def load_ctq(file_name=None, verbose=1):
    """
    【功能】: 加载 JSON -> 恢复对象 -> 生成 CTQ
    【修复】: 解决 unhashable type 报错
    """
    path = _resolve_json_path(file_name, "ctq_data")
    if not os.path.exists(path): return [], [], {}
    
    if 'li' in globals(): li()
    try:
        doc = C.doc
    except:
        doc = win32com.client.Dispatch("AutoCAD.Application").ActiveDocument

    try:
        with open(path, 'r', encoding='utf-8') as f:
            raw_list = json.load(f)
    except: return [], [], {}
    
    valid_polys = []
    
    # 🔥🔥🔥 修复点：Key 必须是 Handle (字符串)，不能是 COM 对象 🔥🔥🔥
    poly_handle_to_blk_handle = {} 
    
    # --- 1. 恢复阶段 (Recovery) ---
    for item in raw_list:
        p_data = item.get("poly_data")
        b_handle = item.get("block_handle")
        
        p_obj = restore_poly_adaptive(p_data) 
        
        if p_obj:
            valid_polys.append(p_obj)
            # 使用 Handle 作为 Key
            h = getattr(p_obj, "Handle", None)
            if h:
                poly_handle_to_blk_handle[h] = b_handle

    if not valid_polys: return [], [], {}

    # --- 2. 排序阶段 (Resort) ---
    if 'sort_coms_by_llcorner' in globals():
        sorted_polys = sort_coms_by_llcorner(valid_polys)
    else:
        sorted_polys = valid_polys

    # --- 3. 重组阶段 (Rebuild) ---
    final_polys = []
    final_blocks = []
    final_mapping = {}
    
    for i, p_obj in enumerate(sorted_polys):
        seq = i + 1
        final_polys.append(p_obj)
        
        # 🔥🔥🔥 修复点：通过 Handle 查找 Block Handle 🔥🔥🔥
        p_h = getattr(p_obj, "Handle", None)
        b_handle = poly_handle_to_blk_handle.get(p_h)
        
        b_obj = None
        if b_handle:
            try:
                temp_blk = doc.HandleToObject(b_handle)
                # 验证 Handle 是否存在
                if getattr(temp_blk, "Handle", None): 
                    b_obj = temp_blk
            except: 
                b_obj = None 
        
        if b_obj: final_blocks.append(b_obj)
        
        # 建立映射
        final_mapping[seq] = (p_h, getattr(b_obj, "Handle", None) if b_obj else None)
        
    return final_polys, final_blocks, final_mapping






#&&&% ===（一）配置图签 ===


#&&% 重定义标准图签文件中的核心块
 
@timeit
def Redefine_standard_blocks(source_file=None, target_file=None):
    """
    【函数编号】: BLOCK-UPDATE-001
    【功能描述】: 
      自动化更新标准图框库的核心流程。
      
      流程分为两个阶段：
      阶段一：重定义 (Redefine)
        1. 复制核心源文件内容。
        2. 抓取实体并重定义标准块。
        3. 保存并关闭文件（确保数据库写入磁盘）。
        
      阶段二：同步 (Sync)
        4. 重新打开文件（强制刷新内存）。
        5. 执行 batch_attsync_loop 循环同步。
        6. 再次保存关闭。
    """
    
    # 引入必要的依赖

    from CAD_file_operations import close_file, open_file,insert_file_exploded


    # 1. 获取环境变量
    base_path_user = os.environ.get('USERPATH')
    
    # 安全检查
    if not base_path_user:
        raise ValueError("错误：系统环境变量 'USERPATH' 未设置！请先设置该变量。")

    # 2. 动态构建路径
    if source_file is None:
        source_file = os.path.join(base_path_user, "dwg文件", "核心图签.dwg")
    
    if target_file is None:
        target_file = os.path.join(base_path_user, "dwg文件", "标准图签.dwg")

    sys_logger.info(f"当前工作基准路径: {base_path_user}")
    sys_logger.info(f"正在处理文件: {target_file}")

    # ==========================================
    # 阶段一：重定义块定义 (Redefine Phase)
    # ==========================================
    try:
        print("\n=== [阶段一] 开始重定义图块定义 ===")
        
        # 1. 显式打开目标文件 (因为 insert 需要操作 active document)
        open_file(target_file)
        wait_command_done()
        
        # 2. 获取文档对象
        doc = C.doc
        
        # 3. 执行插入炸开 (替代原来的复制粘贴)
        # 使用新函数，传入 doc 对象，位置 (0,0,0)
        success = insert_file_exploded(source_file,  x=0, y=0, z=0)
        
        if not success:
            raise RuntimeError("核心图签插入失败")
            
        # 2. 清理删除辅助对象
        wait_command_done()
        
        to_delete_layers = ["tuqian_neibu_pl", "TWT_TITLE"]
        for layer_name in to_delete_layers:
            # 假设 stc 是 select_by_layer 的缩写
            try:
                obs = stc(layer_name) 
                for ent in obs:
                    safe_delete(ent)
            except: pass

        # 3. 定义各图幅的处理配置
        layout_configs = [
            (43626.52, 91007.41, 85626.52, 120707.41, "A3-H"),
            (325411.14, 80912.14, 384811.14, 122912.14, "A2-H"),
            (706618.00, 80912.14, 790718.00, 140312.14, "A1-H"),
            (1210920.21, 64916.99, 1329820.21, 149016.99, "A0-H")
        ]

        # 4. 循环处理所有图幅
        for x_min, y_min, x_max, y_max, blk_name in layout_configs:
            wait_command_done()
            
            # 选框选择
            ents = select_entities_in_window(
                x_min, y_min, x_max, y_max,
                ty=1.0, select_mode="_W"
            )
            
            # 分离块和其他实体
            ents_components, target_blocks = separate_entities_by_block_names(ents, [blk_name])

            time.sleep(1)
            cancel_cad_selection()
            time.sleep(1)

            # 安全性检查与重定义
            if target_blocks:
                wait_command_done()
                redefine_block_with_entities(target_blocks[0], ents_components, ty=0.5)
                sys_logger.info(f"成功重定义: {blk_name}")
            else:
                sys_logger.info(f"警告: 在指定范围内未找到块 {blk_name}，跳过重定义。")

    except Exception as e:
        sys_logger.info(f"Redefine_standard_blocks [阶段一] 出错: {e}")
        raise e 

    finally:
        # 阶段一结束：强制保存并关闭
        # 这确保了新的块定义被物理写入磁盘
        wait_command_done()
        close_file(save_option="auto_save")
        print("阶段一完成：文件已保存并关闭。")

    # ==========================================
    # 阶段二：重新打开并强制同步 (Reload & Sync Phase)
    # ==========================================
    print("\n=== [阶段二] 重新打开文件进行属性同步 ===")
    
    try:
        # 1. 重新打开刚才处理过的文件
        # 这会强制 AutoCAD 重新加载数据库，清除缓存
        open_file(target_file)
        

        
        # 2. 确保环境就绪
        wait_command_done()

        # 3. 执行多轮强制同步
        # 这里调用上一条回复中定义的 batch_attsync_loop
        # 假设该函数已在脚本中定义或导入
        batch_attsync_loop(rounds=3, delay_per_cmd=0.5)



        # 4. 确保环境就绪
        wait_command_done()

        # 定义要删除的目标块名
        target_del_names = ["A0-H", "A1-H", "A2-H", "A3-H"]

        # 1. 使用 select_kuai() 找出所有块
        all_blocks = select_kuai()

        if all_blocks:
            sys_logger.info(f"正在扫描并删除以下块实例: {target_del_names} ...")
            deleted_count = 0
            
            for blk in all_blocks:
                try:
                    # 2. 使用 get_attr 获取块名
                    b_name = get_attr(blk, "Name")
                    
                    # 3. 如果名字匹配，则删除该实例
                    if b_name in target_del_names:
                        # 直接调用 COM 对象的 Delete 方法，或者使用 safe_delete(blk)
                         
                        safe_delete(blk)                               
                        deleted_count += 1
                except:
                    # 防止对象已删除或无法访问导致的报错
                    continue
            
            sys_logger.info(f"✅ 清理完成，共删除 {deleted_count} 个标准图框实例。")
        else:
            print("未找到任何块实例，无需清理。")


    except Exception as e:
        sys_logger.info(f"Redefine_standard_blocks [阶段二] 同步出错: {e}")
        # 这里不 raise，保证文件能关闭，因为定义已经改好了，同步失败可以后续手动补
        
    finally:
        # 4. 最终关闭
        wait_command_done()
        close_file(save_option="auto_save")
        print("✅ 标准化维护全部完成 (已重定义 + 已同步)。")







#&&% 图名排序
@timeit
def get_sorted_titles_by_areas_final(layer_name="DIM_SYMB"):
    """
    【最终重命名版】
    按 sdy 顺序遍历 -> 必须使用 Zoom 配合 SendCommand 选择 -> 强制顺序重命名
    """
    # 1. 初始化
    C.li()
    
    # 2. 获取已排序的打印区域
    sdy = select_print_areas_maxrect_from_polylines(
        lm=70, tol_single=0.01, layer_name="dy_zhuanyong",
        width=0.0, color=1, z=0.0, duanbian=70,
        debug=False, print_rejection_reason=False, cha_Y=5, 
    )
    polylines = sdy[0]

    sys_logger.info(f"开始处理 {len(polylines)} 个打印区域，并将执行【顺序重命名】...")

    # === 全局计数器 (核心) ===
    global_counter = 1
    
    # 总列表
    final_sorted_titles = []

    # 3. 遍历处理
    for i, pl in enumerate(polylines):
        
        # A. 获取包围盒
        try:
            min_pt, max_pt = pl.GetBoundingBox()
            x1, y1 = min_pt[0], min_pt[1]
            x2, y2 = max_pt[0], max_pt[1]
        except Exception:
            continue

        # B. 执行选择 (使用 SendCommand 模式，ty=0.2 足够)
        # 这一步保证了我们选到的绝对是当前框内的，不多也不少
        lk = select_entities_in_window(x1, y1, x2, y2, ty=0.2, select_mode="_W")

        # C. 过滤图层
        found_objs = []
        if lk:
            for xx in lk:
                try:
                    if xx.Layer == layer_name:
                        found_objs.append(xx)
                except:
                    continue
        
        sys_logger.info(f"第 {i} 打印区：", end="")

        xx = generate_name_and_ratio_from_com(pl)
        print(xx)
    
        
        # D. 组内排序 + 重命名
        if found_objs:
            # 1. 先按几何位置排序 (上->下，左->右)
            # 确保同一张图里的两个图名，上面的叫 N，下面的叫 N+1
            sorted_objs = sort_coms_by_llcorner(found_objs, cha_Y=50)
            
            sys_logger.info(f" (命中 {len(sorted_objs)} 个)")
            
            for obj in sorted_objs:
                # 2. 生成新名字
                new_name = f"未命名{global_counter}"
                
                # 3. 【关键步骤】写入新属性
                # 这会覆盖掉之前的 "未命名13"
                set_attr(obj, "图名文字", new_name)
                
                sys_logger.info(f"  - 修改为: {new_name}")
                
                # 4. 计数器递增
                global_counter += 1
                final_sorted_titles.append(obj)
        else:
            print(" (空)")

    sys_logger.info(f"\n处理完成：共重命名 {len(final_sorted_titles)} 个图名。")
    return final_sorted_titles


#&&% 选择测试
@timeit
def get_sorted_titles_ce(layer_name="DIM_SYMB"):
    """
    【最终重命名版】
    按 sdy 顺序遍历 -> 必须使用 Zoom 配合 SendCommand 选择 -> 强制顺序重命名
    """
    # 1. 初始化
    C.li()
    
    # 2. 获取已排序的打印区域
    sdy = select_print_areas_maxrect_from_polylines(
        lm=70, tol_single=0.01, layer_name="dy_zhuanyong",
        width=0.0, color=1, z=0.0, duanbian=70,
        debug=False, print_rejection_reason=False, cha_Y=5, 
    )
    polylines = sdy[0]

    sys_logger.info(f"开始处理 {len(polylines)} 个打印区域，并将执行【顺序重命名】...")

    # === 全局计数器 (核心) ===
    global_counter = 1
    
    # 总列表
    final_sorted_titles = []

    # 3. 遍历处理
    for i, pl in enumerate(polylines):
        
        # A. 获取包围盒
        try:
            min_pt, max_pt = pl.GetBoundingBox()
            x1, y1 = min_pt[0], min_pt[1]
            x2, y2 = max_pt[0], max_pt[1]
        except Exception:
            continue

        # B. 执行选择 (使用 SendCommand 模式，ty=0.2 足够)
        # 这一步保证了我们选到的绝对是当前框内的，不多也不少
        lk = select_objects_in_window_area(x1, y1, x2, y2)

        # C. 过滤图层
        found_objs = []
        if lk:
            for xx in lk:
                try:
                    if xx.Layer == layer_name:
                        found_objs.append(xx)
                except:
                    continue
        
        sys_logger.info(f"第 {i} 打印区：", end="")

        xx = generate_name_and_ratio_from_com(pl)
        print(xx)
    
        
        # D. 组内排序 + 重命名
        if found_objs:
            # 1. 先按几何位置排序 (上->下，左->右)
            # 确保同一张图里的两个图名，上面的叫 N，下面的叫 N+1
            sorted_objs = sort_coms_by_llcorner(found_objs, cha_Y=50)
            
            sys_logger.info(f" (命中 {len(sorted_objs)} 个)")
            
            for obj in sorted_objs:
                # 2. 生成新名字
                new_name = f"未命名{global_counter}"
                
                # 3. 【关键步骤】写入新属性
                # 这会覆盖掉之前的 "未命名13"
                set_attr(obj, "图名文字", new_name)
                
                sys_logger.info(f"  - 修改为: {new_name}")
                
                # 4. 计数器递增
                global_counter += 1
                final_sorted_titles.append(obj)
        else:
            print(" (空)")

    sys_logger.info(f"\n处理完成：共重命名 {len(final_sorted_titles)} 个图名。")
    return final_sorted_titles




#&&% 编辑块生效的强化

def batch_attsync_loop(rounds=3, delay_per_cmd=0.5):
    """
    【函数编号】: CMD-002 (Select版)
    【功能描述】: 
        利用 select_kuai() 获取所有图块，
        通过 get_attr(obj, "Name") 筛选出标准图框 (A3-H ~ A0-H)，
        执行多轮强制属性同步。
    """
    doc=C.doc   
    # 1. 定义目标块名集合
    target_names = ["A3-H", "A2-H", "A1-H", "A0-H"]
    
    sys_logger.info(f"🚀 开始执行标准图框同步循环 (共 {rounds} 轮)...")

    # ================= 循环开始 =================
    for r in range(1, rounds + 1):
        sys_logger.info(f"\n--- [第 {r}/{rounds} 轮] 同步开始 ---")
        
        # 1. 获取所有块实例 (调用现有函数)
        try:
            # 假设 select_kuai() 返回的是一个 COM 对象列表
            all_blocks = select_kuai() 
        except Exception as e:
            sys_logger.info(f"❌ 调用 select_kuai() 失败: {e}")
            break

        if not all_blocks:
            print("⚠️ 未选择到任何图块。")
            if r == 1: break # 如果第一轮就没抓到，直接退出
            continue

        # 2. 筛选目标图块 (去重)
        # ATTSYNC 是对“块定义”操作的，所以 A3-H 只要找到 1 个实例去触发命令就够了
        # 不需要对 100 个 A3-H 执行 100 次同步
        found_targets_map = {} # 格式: {'A3-H': obj, 'A1-H': obj}

        for blk in all_blocks:
            try:
                # 使用 get_attr 获取名字
                b_name = get_attr(blk, "Name")
                
                # 兼容动态块: 如果 Name 是匿名块(*U...), 尝试获取 EffectiveName
                if b_name and "*" in b_name:
                    eff_name = get_attr(blk, "EffectiveName")
                    if eff_name: b_name = eff_name
                
                # 检查是否在目标列表中
                if b_name in target_names:
                    # 只要记录一个代表即可
                    if b_name not in found_targets_map:
                        found_targets_map[b_name] = blk
            except:
                continue
        
        # 3. 执行同步
        if not found_targets_map:
            print("⚠️ 本轮未找到 A3-H ~ A0-H 目标图块。")
        else:
            for name, blk_obj in found_targets_map.items():
                # 调用同步函数
                result = attsync_block_instance(blk_obj)
                
                if result:
                    sys_logger.info(f"   ✅ {name}: 同步命令已发送")
                else:
                    sys_logger.info(f"   ⚠️ {name}: 同步失败")
                
                # 强制等待，防止命令拥堵
                time.sleep(delay_per_cmd)

        # 轮次间缓冲
        time.sleep(0.5)

    # 4. 最终刷新
    try:
        doc.Regen(1)
        print("🔄 全局重生成 (Regen) 完成。")
    except: pass

    sys_logger.info(f"🏁 同步任务结束。")
    return True

#&&&% ===（二）打印区域 ===


#&&% 统一选择

@timeit
@retry_on_busy
def smart_select_polylines(
    layout_name=None,
    operate_target="Model",
    select_config=0,
    use_cache=False, #复用已有数据必须是在清醒的情况下才可以，否则会造成巨大隐患混乱
    min_side=100.0,  
):
    """
    【函数编号】: CACHE-LAYER-001 (智能缓存选择器 - V12.1 参数优化版)
    【功能】: 
        作为 universal_select_polylines 的上层封装。
        负责处理“文件缓存”逻辑。缓存文件名现在会包含 min_side 的指纹，防止精度不同导致混淆。
    
    【参数】:
        layout_name (str/None): 布局名称
        operate_target (str): "Model" 或 "Layout"
        select_config (int): 兼容旧逻辑，1=强制高精模式(min_side置0)，0=常规
        use_cache (bool): 是否优先读取 JSON 缓存
        min_side (float): 最小边长限制 (默认 100.0)
    """
    
    # --- 1. 简单的日志打印 ---
    # 既然移除了 verbose 参数，这里使用默认的 print 或者全局日志等级
    def log(msg):
        sys_logger.info(f"🧠 [SmartCache] {msg}")

    # --- 2. 生成缓存指纹 (Context Awareness) ---
    doc = C.doc
    try:
        doc_name = os.path.splitext(doc.Name)[0]
    except:
        doc_name = "UnknownDoc"

    # 命名规则: [文档名]_[空间]_[布局]_[精度参数].json
    # 为了防止 min_side=100 和 min_side=10 的缓存混用，建议把 min_side 加入文件名
    side_tag = f"L{int(min_side)}" 
    
    if operate_target == "Model":
        if layout_name:
            suffix = f"Model_Hybrid_{layout_name}_{side_tag}"
        else:
            suffix = f"Model_Pure_{side_tag}"
    else: # Layout
        target_layout = layout_name if layout_name else "ActiveLayout"
        suffix = f"Layout_{target_layout}_{side_tag}"
        
    cache_filename = f"{doc_name}_{suffix}.json"

    # ================= 分支 A: 尝试使用缓存 =================
    if use_cache:
        log(f"📂 尝试加载缓存: {cache_filename} ...")
        cached_data = load_poly_list(file_name=cache_filename)
        
        if cached_data:
            log(f"✅ 缓存命中! 获取 {len(cached_data)} 个对象。")
            return cached_data
        else:
            log("⚠️ 缓存未找到或失效，转为【实时计算】...")

    # ================= 分支 B: 重新计算 =================
    if not use_cache:
        log("⚡ 模式: 强制重算 (Skip Cache)")
    
    # 调用下层通用选择器，传入 min_side
    fresh_data = universal_select_polylines(
        layout_name=layout_name, 
        operate_target=operate_target, 
        select_config=select_config,
        min_side=min_side
    )

    # ================= 分支 C: 更新缓存 =================
    if fresh_data:
        log(f"💾 计算完成，正在写入缓存: {cache_filename}")
        save_poly_list(fresh_data, file_name=cache_filename)
    else:
        log("⚠️ 计算结果为空，跳过缓存更新。")

    return fresh_data


def universal_select_polylines(
    layout_name=None,
    operate_target="Model",
    select_config=0, 
    min_side=100.0 # ✅ 接收上层传递的 min_side
):
    """
    【函数编号】: SELECT-UNIFIED-002 (V12.1 - 综合分发器)
    【功能】: 
        根据 min_side 和 select_config 计算最终精度参数，
        分发给 select_print_areas_paperspace 或 select_maxrect_polylines_1。
    """
    
    # --- 1. 参数融合逻辑 ---
    # select_config 是旧代码的遗产，min_side 是新算法的核心
    # 逻辑：如果 select_config=1 (高精模式)，强制覆盖 min_side 为 0
    # 否则使用传入的 min_side
    
    final_min_side = min_side
    precision_mode = False
    
    if select_config == 1:
        final_min_side = 0.0
        precision_mode = True
        sys_logger.info(f"🚀 [统一选择] 触发高精模式 (select_config=1) -> 覆盖 MinSide=0")
    else:
        # 智能推断：如果用户手动传了很小的 min_side，也自动开启 precision_mode
        if final_min_side < 10.0:
            precision_mode = True
            
    sys_logger.info(f"🚀 [统一选择] 目标: {operate_target} | 布局: {layout_name} | MinSide: {final_min_side}")

    # ================= 分支 A: 布局空间模式 =================
    if operate_target == "Layout":
        if not layout_name:
            print("❌ [错误] 布局模式必须提供 layout_name")
            return []
            
        # 调用新函数 select_print_areas_paperspace
        res_list, _ = select_print_areas_paperspace(
            layout_name=layout_name,
            precision_mode=precision_mode,
            min_side=final_min_side, # ✅ 传递参数
            width=0.0,
            color=1
        )
        return res_list

    # ================= 分支 B: 模型空间模式 =================
    elif operate_target == "Model":
        
        # 强制锁定到模型空间
        target_layout_for_model = "Model" 
        
        # 调用新函数 select_maxrect_polylines_1
        res_list, _ = select_maxrect_polylines_1(
            layout_name=target_layout_for_model, 
            precision_mode=precision_mode,
            min_side=final_min_side, # ✅ 传递参数
            width=0.0,
            color=1
        )
        return res_list

    # ================= 异常兜底 =================
    else:
        sys_logger.info(f"❌ [错误] 未知的操作目标: {operate_target}")
        return []




#&&% 模型空间极大矩形选择打印区域

@timeit
@debuggable
def select_print_areas_maxrect_from_polylines(**kwargs):
    """
    【兼容层】旧参数 lm, duanbian, tol_single 等已被 kwargs 自动吃掉并忽略。
    """
    return select_maxrect_polylines_1(**kwargs)

@timeit
@debuggable
def select_maxrect_polylines_1(
    layer_name: str = "dy_zhuanyong",#绘制打印多段线图层
    precision_mode: bool = False,    #任意小矩形多段线
    width: float = 0.0,              #绘制打印多段线宽度
    color: int = 1,                  #绘制打印多段线颜色

    min_side: float = 100.0,         #矩形多段线的最小边长
    **kwargs                         #距离容差和排序容差是自适应的所以没有参数了
):                                   #其它硬编码参数是对一般图纸情况的归纳，没必要暴露给用户
    """
    【V12.0 步进式审计版】
    严格遵循用户定义的 5 步日志输出，用于排查矩形识别与过滤流程。
    """
    from  CAD_file_operations   import set_space_mode

    doc = C.doc
    _set_attr = globals().get('set_attr', setattr)
    
    set_space_mode(1)

    # —————————— 初始化 ——————————
    ##try:
        ##lyr_obj = ensure_layer(layer_name)
        ##if lyr_obj: _set_attr(lyr_obj, 'Color', color)
        ##sys_logger.info(f"🧹 [初始化] 已清空图层 {layer_name}")
    ##except Exception as e:
        ##sys_logger.warning(f"初始化图层警告: {e}")

    # —————————— 参数设定 ——————————
    # 根据 precision_mode 设定初始扫描参数
    if precision_mode:
        scan_min_side = 0.0
        active_tol = 0.1
        sys_logger.info("⚙️ [模式] 高精模式 (MinSide=0, Tol=0.1)")
    else:
        scan_min_side = min_side
        # 初始容差先给个默认值，后面根据图形动态调整
        active_tol = 10.0 
        sys_logger.info("⚙️ [模式] 自适应模式 (MinSide=100)")

    final_output_ents = []
    loop_count = 0
    
    # —————————— 循环递归分析 ——————————
    while loop_count < 5:
        loop_count += 1
        sys_logger.info(f"\n🔄 --- 开始第 {loop_count} 次分析循环 ---")

        # =======================================================
        # [步骤 1] 输出选择到的矩形多段线数量
        # =======================================================
        raw_polys = get_rectangular_polylines(min_side=scan_min_side, area_tolerance=0.05)
        sys_logger.info(f"📍 [步骤1] 原始选择矩形数量: {len(raw_polys)} 个")
        
        if not raw_polys:
            sys_logger.error("❌ 未找到任何矩形，流程终止。")
            return [], {}

        # =======================================================
        # [步骤 4] (提前执行) 排除非模型空间上的矩形
        # 解释：必须先过滤，否则布局空间的框会干扰模型空间的包含关系判断
        # =======================================================
        ms_polys = []
        for p in raw_polys:
            try:
                # 检查 OwnerID 是否指向模型空间
                if doc.ObjectIdToObject(p.OwnerID).Name.upper() == "*MODEL_SPACE":
                    ms_polys.append(p)
            except: pass
        
        sys_logger.info(f"📍 [步骤4] 经模型空间过滤后数量: {len(ms_polys)} 个 (剔除 {len(raw_polys) - len(ms_polys)} 个非模型对象)")
        
        if not ms_polys:
            sys_logger.warning("❌ 模型空间无有效矩形，流程终止。")
            break

        # --- 动态计算几何参数 (不计入步骤，属于中间计算) ---
        rect_data = []
        all_bounds = []
        for p in ms_polys:
            try:
                ll, ur = p.GetBoundingBox()
                w, h = abs(ur[0]-ll[0]), abs(ur[1]-ll[1])
                rect_data.append({
                    "ent": p, "h": p.Handle, "ll": ll, "ur": ur,
                    "diag": (w**2 + h**2)**0.5
                })
                all_bounds.append(min(w, h))
            except: pass
        
        # =======================================================
        # [步骤 1.5] 前置高精度去重 (新增逻辑)
        # =======================================================
        if rect_data:
            # 1. 计算去重容差 (基于当前收集到的最小边长)
            current_min_side = min(all_bounds) if all_bounds else 100.0
            
            dedup_tol = current_min_side * 0.0005
            dedup_tol = max(0.1, min(dedup_tol, 10.0)) # 限制范围 [0.1, 10]

            sys_logger.info(f"🔍 [前置去重] 动态容差: {dedup_tol:.5f} (基准短边: {current_min_side:.2f})")

            # 2. 执行物理去重
            ents_to_check = [r["ent"] for r in rect_data]
            unique_ents_pre = remove_duplicate_polylines(
                polylines=ents_to_check, 
                tol=dedup_tol, 
                priority_layer=layer_name
            )

            # 3. 重新过滤 rect_data (仅保留存活的 Handle)
            surviving_handles = set()
            for e in unique_ents_pre:
                try: surviving_handles.add(e.Handle)
                except: pass
            
            old_count = len(rect_data)
            rect_data = [r for r in rect_data if r["h"] in surviving_handles]
            
            # 4. 同步更新 all_bounds (这点很重要，防止已删除的小图框影响后续容差)
            if len(rect_data) < old_count:
                sys_logger.info(f"📉 [前置去重] 剔除 {old_count - len(rect_data)} 个重叠对象")
                all_bounds = []
                for r in rect_data:
                    w = abs(r["ur"][0] - r["ll"][0])
                    h = abs(r["ur"][1] - r["ll"][1])
                    all_bounds.append(min(w, h))
        else:
            rect_data = []


        # 动态更新容差
        if not precision_mode and all_bounds:
            min_L = min(all_bounds)
            active_tol = max(0.1, min_L * 0.0005)
            active_cha_Y = max(10.0, min_L * 0.03)
        else:
            active_cha_Y = 5.0

        # --- 分析极大矩形候选 (中间计算) ---
        maxima_candidates = []
        for i, ri in enumerate(rect_data):
            is_contained = False
            for j, rj in enumerate(rect_data):
                if i == j: continue
                # 包含判定
                if (rj["ll"][0] <= ri["ll"][0] + active_tol and rj["ll"][1] <= ri["ll"][1] + active_tol and
                    rj["ur"][0] >= ri["ur"][0] - active_tol and rj["ur"][1] >= ri["ur"][1] - active_tol):
                    if rj["diag"] > ri["diag"] + active_tol:
                        is_contained = True; break
            if not is_contained:
                maxima_candidates.append(ri)

        # =======================================================
        # [步骤 2] 输出伪极大矩形多段线数量 (并炸开)
        # =======================================================
        pseudo_rects = []
        true_candidates = []
        
        for mx in maxima_candidates:
            is_pseudo = False
            # 检查 mx 内部是否包含标准图框
            for item in rect_data:
                if item["h"] == mx["h"]: continue
                # 几何包含检查
                if (mx["ll"][0] <= item["ll"][0] + active_tol and mx["ll"][1] <= item["ll"][1] + active_tol and
                    mx["ur"][0] >= item["ur"][0] - active_tol and mx["ur"][1] >= item["ur"][1] - active_tol):
                    
                    # 核心校验：内部是否有标准尺寸
                    if check_strict_standard_size(item["ent"], tol=10) != 0:
                        is_pseudo = True; break
            
            if is_pseudo:
                pseudo_rects.append(mx)
            else:
                true_candidates.append(mx["ent"])

        sys_logger.info(f"📍 [步骤2] 发现伪极大矩形数量: {len(pseudo_rects)} 个")


        # =======================================================
        # [步骤 2] 优化版：安全移除伪极大矩形 (仅删除，不炸开)
        # =======================================================
        if pseudo_rects:
            sys_logger.warning(f"🧹 发现 {len(pseudo_rects)} 个伪极大矩形 (外框)，执行安全移除...")
            
            # 1. 执行删除 (只操作 ID，不产生新碎片)
            count_del = 0
            for p_item in pseudo_rects:
                try:
                    # 再次确认对象是否存在，防止重复删除报错
                    if not p_item["ent"].EntityName: continue 
                    p_item["ent"].Delete()
                    count_del += 1
                except: pass
            
            sys_logger.info(f"✅ 已移除 {count_del} 个外框。")

            # 2. 【关键】清理现场
            # 这一步至关重要，告诉 CAD 数据库更新状态
            doc.Regen(1) 
            
            # 3. 【关键】重启循环
            # 绝对不要复用当前的 raw_polys 或 ms_polys，因为里面可能包含刚才删掉的对象
            # 直接 continue，跳回步骤 1，重新 get_rectangular_polylines
            sys_logger.info("🔄 数据库已变更，正在重新扫描...")
            continue

        else:
            # =======================================================
            # [步骤 3] 输出真实的极大多段线打印区域数量
            # =======================================================
            final_output_ents = true_candidates
            sys_logger.info(f"📍 [步骤3] 真实的极大多段线打印区域数量: {len(final_output_ents)} 个")
            break # 分析完成，跳出循环

    # =======================================================
    # [步骤 5] 输出去重后的多段线数量 直接改图层，不重绘
    # =======================================================
    unique_ents = remove_duplicate_polylines(final_output_ents, tol=active_tol, priority_layer=layer_name)
    sys_logger.info(f"📍 [步骤5] 去重后的最终多段线数量: {len(unique_ents)} 个")

    # —————————— 排序与重绘 ——————————
    sorted_ents = sort_coms_by_llcorner(unique_ents, cha_Y=active_cha_Y)
    
    final_dict = {}
    draw_func = globals().get('draw_lwpolyline')
    info_func = globals().get('generate_name_and_ratio_from_com')

    sys_logger.info(f"🎨 开始重绘 {len(sorted_ents)} 个区域...")

    sorted_new_entities = []
    for ent in sorted_ents:
        try:
            ll, ur = ent.GetBoundingBox()
            # 补齐 3D 坐标，防止 draw_lwpolyline 解包报错
            pts = [
                (ll[0], ll[1], 0.0), 
                (ll[0], ur[1], 0.0), 
                (ur[0], ur[1], 0.0), 
                (ur[0], ll[1], 0.0)
            ]
            
            new_pl = draw_func(pts, layer_name=layer_name, width=width, color=color, closed=True)
            
            if new_pl:
                # 使用安全属性设置
                _set_attr(new_pl, 'Color', color)
                
                info = info_func(new_pl, tol=active_tol) if info_func else "Ready"
                final_dict[new_pl.Handle] = info

                sorted_new_entities.append(new_pl)
        except Exception as e:
            sys_logger.error(f"重绘失败: {e}")

    try:
        for ent in sorted_ents:
            safe_delete(ent)
    
        doc.Regen(1)
    
    except:

        pass

    return sorted_new_entities, final_dict



#&&% 图纸空间打印区域的选择

#20260116

def select_print_areas_paperspace(
    layout_name: str,                   
    layer_name: str = "dy_zhuanyong", 
    precision_mode: bool = False,       
    width: float = 0.0,                 
    color: int = 3,  # 默认绿色                    
    min_side: float = 100.0,             
    **kwargs
):
    """
    【函数编号】: SEL-PAPER-002 (V16.4 - 抗干扰增强版)
    【修复】: 
        1. 增加布局切换的死循环重试机制 (针对 'Call rejected')。
        2. 确保必须在正确布局下才开始扫描。
    """
    import time
    import pythoncom

    from  CAD_file_operations   import set_space_mode    
    doc = C.doc
    _get = globals().get('get_attr', getattr) 
    
    sys_logger.info(f"🚀 [启动] 图纸空间选择 - 目标布局: {layout_name}")
    
    # 尝试切换空间模式 (防止还在视口里)
    try: set_space_mode(0)
    except: pass

    # ================= 1. 布局切换 (增强重试) =================
    switched = False
    for attempt in range(5):
        try:
            # 如果已经在该布局，直接跳过
            if doc.ActiveLayout.Name == layout_name:
                switched = True
                break
            
            # 尝试切换
            doc.ActiveLayout = doc.Layouts.Item(layout_name)
            time.sleep(0.5) # 给一点反应时间
            
            # 二次确认
            if doc.ActiveLayout.Name == layout_name:
                switched = True
                break
        except Exception as e:
            err_msg = str(e)
            if "rejected" in err_msg or "busy" in err_msg:
                sys_logger.info(f"   ⏳ CAD忙碌，重试切换布局 ({attempt+1}/5)...")
                time.sleep(1.0)
            else:
                sys_logger.error(f"❌ 布局切换致命错误: {e}")
                break
    
    if not switched:
        sys_logger.error(f"❌ 最终未能切换到布局 '{layout_name}'，无法扫描。")
        return [], {}

    # ================= 2. 解锁图层 =================
    try:
        lay = doc.Layers.Item(layer_name)
        if lay.Lock: lay.Lock = False
    except: pass

    # ================= 3. 扫描多段线 =================
    scan_min_side = 0.0 if precision_mode else min_side
    
    valid_polylines = get_layout_rectangular_polylines_coords(layout_name, min_side=scan_min_side)
    
    if not valid_polylines:
        sys_logger.warning(f"❌ 在布局 '{layout_name}' 层 '{layer_name}' 未找到符合条件的矩形。")
        return [], {}

    # ================= 4. 数据适配与筛选 (保持原逻辑) =================
    rect_data = []
    for obj in valid_polylines:
        try:
            min_pt, max_pt = obj.GetBoundingBox()
            w, h = abs(max_pt[0] - min_pt[0]), abs(max_pt[1] - min_pt[1])
            hdl = _get(obj, "Handle")
            rect_data.append({
                "ent": obj, "h": hdl, 
                "ll": min_pt, "ur": max_pt,
                "w": w, "h_dim": h, "diag": (w**2 + h**2)**0.5, "area": w * h
            })
        except: continue

    # 极大矩形保留策略
    all_bounds = [min(r["w"], r["h_dim"]) for r in rect_data]
    if not precision_mode and all_bounds:
        active_tol = max(1.0, min(all_bounds) * 0.001)
        active_cha_Y = max(2.0, min(all_bounds) * 0.03)
    else:
        active_tol, active_cha_Y = 1.0, 2.0

    maxima_candidates = []
    for i, ri in enumerate(rect_data):
        is_contained = False
        for j, rj in enumerate(rect_data):
            if i == j: continue
            if (rj["ll"][0] <= ri["ll"][0] + active_tol and rj["ll"][1] <= ri["ll"][1] + active_tol and
                rj["ur"][0] >= ri["ur"][0] - active_tol and rj["ur"][1] >= ri["ur"][1] - active_tol):
                if rj["diag"] > ri["diag"] + active_tol:
                    is_contained = True; break
        if not is_contained:
            maxima_candidates.append(ri)

    final_ents = [m["ent"] for m in maxima_candidates]
    
    # 去重与排序
    unique_ents = remove_duplicate_polylines(final_ents, tol=active_tol, priority_layer=layer_name)
    sorted_ents = sort_coms_by_llcorner(unique_ents, cha_Y=active_cha_Y)

    # ================= 5. 重绘与返回 =================
    sys_logger.info(f"🎨 重绘 {len(sorted_ents)} 个有效区域")
    
    final_dict = {}
    draw_func = globals().get('draw_lwpolyline')
    info_func = globals().get('generate_name_and_ratio_from_com')
    new_objs = []

    for ent in sorted_ents:
        try:
            ll, ur = ent.GetBoundingBox()
            pts = [(ll[0], ll[1], 0.0), (ll[0], ur[1], 0.0), (ur[0], ur[1], 0.0), (ur[0], ll[1], 0.0)]
            
            # 绘制新线
            new_pl = draw_func(pts, layer_name=layer_name, width=width, color=color, closed=True)
            
            if new_pl:
                info = info_func(new_pl, tol=active_tol) if info_func else "Ready"
                final_dict[_get(new_pl, "Handle")] = info
                new_objs.append(new_pl)
                
                # 删除旧线 (防残留)
                try: ent.Delete()
                except: pass
                
        except Exception as e:
            sys_logger.error(f"重绘失败: {e}")

    try: doc.Regen(1)
    except: pass
    
    return new_objs, final_dict


#&&% 标准框选择打印区域

@timeit
@debuggable
def select_standard_print_areas(
    lm: float = 70,
    layer_name: str = "dy_zhuanyong",
    tol_single: float = 0.01,
    cha_Y: float = 20,
    mute_logs: bool = False,
):
    """
    【函数编号】: RECOG-002
    【所属模块】: 核心识别算法 (Core Recognition)
    【功能描述】: 
        标准图框专用筛选与重构函数（保护模式）。
        
        核心机制：
        1. 【只读扫描】: 遍历图纸中的多段线，进行严格的标准尺寸匹配。
        2. 【原图保护】: 绝不修改、移动或删除被选中的原始图元。
        3. 【重绘输出】: 提取合规对象的包围盒坐标，在专用图层 (layer_name) 生成全新的矩形框。
        4. 【数据映射】: 自动将原图的尺寸识别结果映射给新生成的打印框。

    【参数详解】:
        - lm (float): 
            最小尺寸阈值。短边小于此值的多段线将被直接忽略。
        - layer_name (str): 
            输出图层。程序会清空该图层并重新绘制识别到的图框。
        - tol_single (float): 
            判断原图多段线是否闭合的端点距离容差。
        - cha_Y (float): 
            结果排序时的行高容差。用于最终生成列表的空间排序。
        - mute_logs (bool): 
            静默模式。True 则不输出日志（适合被其他高频函数调用时使用）。

    【返回值】:
        - (list, dict): 
            1. sorted_pls: 已排序的【新生成】多段线对象列表。
            2. final_dict: {NewHandle: SizeCode} 字典，记录了每个新框对应的标准尺寸代码。
    dy_2=select_standard_print_areas()
    dy_2[1]
    {'713E': ('UserDefinedMetric (1051.25 x 594.00毫米)', '1:100', 'A1+1/4', 0), '713F': ('UserDefinedMetric (1261.50 x 594.00毫米)')}
    {'713E': ('UserDefinedMetric (1051.25 x 594.00毫米)', '1:100', 'A1+1/4', 0), '713F': 'UserDefinedMetric (1261.50 x 594.00毫米)'}
    


    """

    # —— 日志小工具 —— #
    def log(msg, *args):
        try: 
            text = msg.format(*args) if args else msg
        except: 
            text = msg
        
        try: 
            print(text)
        except: 
            pass
        
        n = globals().get("node", None)
        if n: 
            try: 
                n(text)
            except: 
                pass

    # —— 0. 环境准备 —— #
    li() 
    
    try:
        ensure_layer(layer_name)
        # 1. 清空目标图层（只会删除上次生成的框，不会删原图）
        objs = stc(layer_name)
        if objs:
            cnt = 0
            iter_objs = []
            if isinstance(objs, list): 
                iter_objs = objs
            elif hasattr(objs, "Item"): 
                iter_objs = [objs.Item(i) for i in range(objs.Count)]
            else: 
                iter_objs = [objs]

            for o in iter_objs:
                try: 
                    o.Delete()
                    cnt += 1
                except: 
                    pass
            log("✅ 已清空图层 '{}' (删除 {} 个旧打印框)", layer_name, cnt)
    except Exception as e:
        log("⚠ 环境准备阶段报错: {}", e)

    # ========= 内部工具 =========
    def get_attr_safe(ent, name):
        try: return getattr(ent, name)
        except: return None

    def get_coords(ent):
        try:
            raw = get_attr_safe(ent, "Coordinates")
            if not raw: return []
            coords = list(raw)
            n = len(coords)
            step = 3 if (n % 3 == 0 and n % 2 != 0) else 2
            pts = []
            for i in range(0, n, step):
                pts.append((float(coords[i]), float(coords[i+1])))
            return pts
        except: return []

    # ==================== Step 2: 聚合选择 (收集原图多段线) ====================
    all_polys = []
    funcs = ['select_polyline', 'select_polyline_chuantong']
    
    for fname in funcs:
        func = globals().get(fname)
        if not func: continue
        
        try:
            if 'autocast' in func.__code__.co_varnames:
                res = func(autocast=True)
            else:
                res = func()
            
            if not res: continue
            
            batch = []
            if isinstance(res, list): 
                batch = res
            elif hasattr(res, "Item"): 
                batch = [res.Item(i) for i in range(res.Count)]
            else: 
                batch = [res]
            
            # 收集原图对象
            valid_batch = [e for e in batch if hasattr(e, "ObjectName")] 
            all_polys.extend(valid_batch)
            log("📥 [{}] 贡献了 {} 个原图对象", fname, len(valid_batch))
            
        except Exception as e:
            log("⚠ {} 执行异常: {}", fname, e)

    log("📊 聚合完成，共收集到 {} 个原图候选", len(all_polys))

    if not all_polys:
        return [], {}

    # ==================== Step 3: 筛选流程 (分析原图对象) ====================
    
    candidates = [] # 存储通过筛选的原图实体
    handle_size_map_temp = {} 
    check_func = globals().get('check_strict_standard_size')

    for idx, ent in enumerate(all_polys):
        h = get_attr_safe(ent, "Handle") or f"Idx{idx}"
        
        try:
            obj_name = str(get_attr_safe(ent, "ObjectName"))
            if "Polyline" not in obj_name: continue

            closed = get_attr_safe(ent, "Closed") or get_attr_safe(ent, "Closed2")
            pts = get_coords(ent)
            
            if not closed and len(pts) >= 2:
                p1, p2 = pts[0], pts[-1]
                dist_sq = (p1[0]-p2[0])**2 + (p1[1]-p2[1])**2
                if dist_sq > (tol_single ** 2): continue 

            if len(pts) < 4: continue
            
            try:
                ll, ur = ent.GetBoundingBox()
                w = abs(ur[0] - ll[0])
                h = abs(ur[1] - ll[1])
                if min(w, h) < lm: continue
            except: continue

            res = 0
            if check_func: 
                res = check_func(ent, tol=10)
            
            is_valid = False
            if isinstance(res, (tuple, list)): is_valid = True
            elif res != 0 and res is not None: is_valid = True
            
            if is_valid:
                candidates.append(ent)
                # 记录【原图Handle】对应的尺寸
                handle_size_map_temp[h] = res
            
        except Exception as e:
            pass

    log("✅ 筛选结束，找到 {} 个有效标准图框源对象", len(candidates))

    # ==================== Step 4: 几何去重 (基于原图位置) ====================
    
    unique_candidates = candidates
    if 'remove_duplicate_polylines' in globals() and candidates:
        try:
            # 这里的 remove_duplicate_polylines 只是为了算出哪些是重复的
            # 我们传入原图列表，返回去重后的原图列表
            unique_candidates = remove_duplicate_polylines(candidates, tol=0.01)
            log("✂️ 基于原图位置去重后剩余: {}", len(unique_candidates))
        except Exception as e:
            log("⚠ 几何去重失败，使用全部候选: {}", e)

    # ==================== Step 5: 重绘 (核心：不改原图，画新图) ====================
    
    new_drawn_polys = [] # 存放新画的线
    final_dict = {}      # 存放 {新Handle: 尺寸代码}
    
    draw_func = globals().get('draw_lwpolyline')
    if not draw_func:
        log("‼ 致命错误：未找到 draw_lwpolyline 函数")
        return [], {}

    count_drawn = 0
    
    # 遍历筛选出的【原图对象】
    for origin_ent in unique_candidates:
        try:
            # 1. 提取原图几何信息
            ll, ur = origin_ent.GetBoundingBox()
            minx, miny = float(ll[0]), float(ll[1])
            maxx, maxy = float(ur[0]), float(ur[1])
            z_val = 0.0 
            
            # 2. 构造新坐标 (5点闭合)
            coords3d = [
                (minx, miny, z_val),
                (minx, maxy, z_val),
                (maxx, maxy, z_val),
                (maxx, miny, z_val),
                (minx, miny, z_val)
            ]
            
            # 3. 在专用图层画【新实体】
            new_pl = draw_func(
                coords3d=coords3d,
                layer_name=layer_name, 
                width=0.0,
                color=1, 
                closed=True
            )
            
            if new_pl:
                new_drawn_polys.append(new_pl)
                count_drawn += 1
                
                # 4. 传递数据：将【原图尺寸】赋给【新Handle】
                origin_h = get_attr_safe(origin_ent, "Handle")
                new_h = get_attr_safe(new_pl, "Handle")
                
                if origin_h in handle_size_map_temp:
                    final_dict[new_h] = handle_size_map_temp[origin_h]
                else:
                    # 备用：重新计算
                    if check_func:
                        final_dict[new_h] = check_func(origin_ent, tol=10)
                    else:
                        final_dict[new_h] = "Unknown"

        except Exception as e:
            log("⚠ 重绘失败: {}", e)

    log("🎨 已在 '{}' 图层重绘 {} 个打印框 (原图未变动)", layer_name, count_drawn)

    # ==================== Step 6: 排序 (对新画的实体排序) ====================
    
    sorted_pls = new_drawn_polys
    if 'sort_coms_by_llcorner' in globals() and new_drawn_polys:
        try:
            sorted_pls = sort_coms_by_llcorner(new_drawn_polys, cha_Y=cha_Y)
            log("🔢 新实体排序完成")
        except: 
            pass

    return sorted_pls, final_dict

#&&% 从指定图层列表的图块建立打印区域

@timeit
@debuggable
def select_print_areas_from_blocks(
    block_layers: tuple = ("dy_quyu", "tuqian_neibu_pl"), # 需搜索的块所在图层列表单图层要加,
    rect_layer: str = "dy_zhuanyong",                     # 结果绘制到的专用图层
    width: float = 0.0,
    color: int = 1,
    z: float = 0.0,
    cha_Y: float = 2000,                                  # 排序行容差
    debug: bool = False,
):

    """
    【函数编号】: RECOG-003
    【所属模块】: 核心识别算法 (Core Recognition)
    【功能描述】: 
        基于图块 (Block) 的打印区域识别函数。
        
        核心流程：
        1. 【广撒网】: 选取模型空间中的所有图块。
        2. 【精过滤】: 根据 block_layers 过滤出特定图层的图块。
        3. 【侦探模式】: 若未找到目标，会自动分析并报告实际选到的图层，辅助排查。
        4. 【转换输出】: 获取图块的包围盒，在 rect_layer 绘制对应的多段线框。
        5. 【信息生成】: 生成标准比例或图名信息，并按空间位置排序。

    【参数详解】:
        - block_layers (tuple): 
            需搜索的图层名称列表。注意：单元素元组需加逗号，如 ("layer1",)。
        - rect_layer (str): 
            结果输出图层。程序会在该图层绘制与图块外框一致的矩形。
        - cha_Y (float): 
            空间排序时的行高容差。
        - width, color, z: 
            新绘制矩形的线宽、颜色索引和 Z 轴标高。

    【返回值】:
        - (list, dict): 
            1. sorted_pls: 已排序的【新生成】多段线对象列表。
            2. final_dict: {NewHandle: Info} 字典。
    """


    # —— 日志小工具 (已严格修复缩进) —— #
    def log(msg, *args):
        try: 
            text = msg.format(*args) if args else msg
        except: 
            text = msg
        
        try: 
            print(text)
        except: 
            pass
            
        n = globals().get("node", None)
        if n: 
            try: 
                n(text)
            except: 
                pass



    # —— 0. 环境准备 —— #
    li() 
    
    try:
        ensure_layer(rect_layer)
        # 清空目标图层
        objs = stc(rect_layer)
        if objs:
            cnt = 0
            iter_objs = []
            if isinstance(objs, list): 
                iter_objs = objs
            elif hasattr(objs, "Item"): 
                iter_objs = [objs.Item(i) for i in range(objs.Count)]
            else: 
                iter_objs = [objs]

            for o in iter_objs:
                try: 
                    o.Delete()
                    cnt += 1
                except: 
                    pass
            log("✅ 已清空结果图层 '{}' (删除 {} 个旧对象)", rect_layer, cnt)
    except Exception as e:
        log("⚠ 环境准备阶段报错: {}", e)

    # ==================== Step 1: 收集块并过滤 ====================
    
    # 调用外部函数选择所有块
    all_blocks = []
    select_kuai_func = globals().get('select_kuai')
    
    if select_kuai_func:
        try:
            res = select_kuai_func()
            if res:
                if isinstance(res, list): 
                    all_blocks = res
                elif hasattr(res, "Item"): 
                    all_blocks = [res.Item(i) for i in range(res.Count)]
                else: 
                    all_blocks = [res]
        except Exception as e:
            log("⚠ select_kuai 调用失败: {}", e)
    else:
        log("‼ 错误：未找到 select_kuai 函数")
        return [], {}
        
    log("📥 select_kuai 返回了 {} 个对象", len(all_blocks))

    # 过滤图层
    target_layers = set(block_layers) 
    valid_blocks = []
    
    # 用于侦探调试
    found_info = set()

    for blk in all_blocks:
        try:
            # 确保是块引用
            if "AcDbBlockReference" not in str(getattr(blk, "ObjectName", "")):
                continue
            
            # 记录侦探信息
            try:
                b_name = blk.EffectiveName if hasattr(blk, "EffectiveName") else blk.Name
                found_info.add(f"图层='{blk.Layer}', 块名='{b_name}'")
            except: 
                found_info.add(f"图层='{blk.Layer}'")

            # 检查图层
            if blk.Layer in target_layers:
                valid_blocks.append(blk)
        except: pass

    # 如果没找到任何块，启动侦探模式报告
    if not valid_blocks:
        log("❌ 经图层筛选后，剩余 0 个目标块！")
        log("--------------------------------------------------")
        log("🕵️‍♂️【侦探报告】select_kuai 实际选到了以下内容的块：")
        for info in list(found_info)[:10]: # 只打印前10种
             log("   👉 {}", info)
        log("--------------------------------------------------")
        log("💡 提示：请检查 block_layers 参数是否与上述 '图层=' 完全一致。")
        return [], {}

    log("✅ 经图层筛选后，剩余 {} 个目标块", len(valid_blocks))

    # ==================== Step 2: 绘制矩形并生成信息 ====================

    draw_func = globals().get('draw_lwpolyline')
    info_func = globals().get('generate_name_and_ratio_from_com')
    
    if not draw_func:
        log("‼ 致命错误：缺少 draw_lwpolyline 函数")
        return [], {}
    
    new_polys_unsorted = []
    final_dict = {} # {新Handle: Info}
    count_drawn = 0
    
    for blk in valid_blocks:
        try:
            # 1. 获取块的外包盒
            ll, ur = blk.GetBoundingBox()
            minx, miny = float(ll[0]), float(ll[1])
            maxx, maxy = float(ur[0]), float(ur[1])
            
            if abs(maxx - minx) < 1 or abs(maxy - miny) < 1:
                continue

            # 2. 构造矩形坐标
            coords3d = [
                (minx, miny, z),
                (minx, maxy, z),
                (maxx, maxy, z),
                (maxx, miny, z),
                (minx, miny, z)
            ]
            
            # 3. 在专用图层绘制
            new_pl = draw_func(
                coords3d=coords3d,
                layer_name=rect_layer,
                width=width,
                color=color,
                closed=True
            )
            
            if new_pl:
                new_polys_unsorted.append(new_pl)
                count_drawn += 1
                
                # 4. 生成打印信息
                h = new_pl.Handle
                if info_func:
                    try:
                        info = info_func(new_pl, tol=10)
                        final_dict[h] = info
                    except Exception as e:
                        final_dict[h] = f"InfoError: {e}"
                else:
                    final_dict[h] = "InfoFuncMissing"
                    
        except Exception as e:
            pass

    log("🎨 已在 '{}' 绘制 {} 个打印框", rect_layer, count_drawn)

    # ==================== Step 3: 排序 ====================
    
    sorted_pls = sort_coms_by_llcorner(new_polys_unsorted, cha_Y=cha_Y)
    
    log("🔢 排序完成，返回结果。")
    return sorted_pls, final_dict


#&&% 从指定图层获得打印区域

@timeit
@debuggable
def select_print_areas_from_layer(
    source_layer: str = "标准图签",       # 数据源图层
    rect_layer: str = "dy_zhuanyong",   # 结果绘制到的专用图层
    width: float = 0.0,
    color: int = 1,
    z: float = 0.0,
    cha_Y: float = 2000,                # 排序行容差
    debug: bool = False,
):
    """
    【函数编号】: RECOG-004
    【所属模块】: 核心识别算法 (Core Recognition)
    【功能描述】: 
        基于指定图层的全量对象打印区域生成。
        
        核心机制：
        1. 【全量扫描】: 选中 source_layer 上的所有图元。
        2. 【重绘输出】: 计算包围盒并在 rect_layer 上绘制新框。
        3. 【智能去重】: 调用 remove_duplicate_polylines 清除位置重叠的冗余框。
        4. 【数据映射】: 生成打印比例或图名信息。
        
        ★ 优化特性：
        - 支持源图层与输出图层相同 (source_layer == rect_layer)，自动跳过清理并执行去重，
          实现对现有打印框的清洗和重排序。

    【参数详解】:
        - source_layer (str): 数据源图层名称。
        - rect_layer (str): 结果输出图层。
        - cha_Y (float): 空间排序时的行高容差。

    【返回值】:
        - (list, dict): 
            1. sorted_pls: 已去重并排序的多段线列表。
            2. final_dict: {NewHandle: Info} 字典。
    """

    # —— 日志小工具 —— #
    def log(msg, *args):
        try: text = msg.format(*args) if args else msg
        except: text = msg
        try: print(text)
        except: pass
        n = globals().get("node", None)
        if n: 
            try: n(text)
            except: pass

    # —— 0. 环境准备 —— #
    li() 
    
    # 清理结果图层 (★ 优化：仅当输出层与源层不同时才清理)
    # 如果相同，说明是“自清洗”模式，保留原对象，后面通过去重函数处理
    if rect_layer != source_layer:
        try:
            ensure_layer(rect_layer)
            objs = stc(rect_layer)
            if objs:
                cnt = 0
                iter_objs = []
                if isinstance(objs, list): iter_objs = objs
                elif hasattr(objs, "Item"): iter_objs = [objs.Item(i) for i in range(objs.Count)]
                else: iter_objs = [objs]

                for o in iter_objs:
                    try: o.Delete(); cnt += 1
                    except: pass
                log("✅ 已清空结果图层 '{}' (删除 {} 个旧对象)", rect_layer, cnt)
        except Exception as e:
            log("⚠ 环境准备阶段报错: {}", e)
    else:
        log("🛡️ 源图层与输出图层相同，跳过清理，将在 Step 3 执行合并去重。")

    # ==================== Step 1: 从源图层选择对象 ====================
    raw_objs = []
    try:
        res = stc(source_layer)
        if res:
            if isinstance(res, list): raw_objs = res
            elif hasattr(res, "Item"): raw_objs = [res.Item(i) for i in range(res.Count)]
            else: raw_objs = [res]
    except Exception as e:
        log("⚠ stc('{}') 调用失败: {}", source_layer, e)
        return [], {}

    if not raw_objs:
        log("❌ 图层 '{}' 上未选到任何对象。", source_layer)
        return [], {}
        
    log("📥 从图层 '{}' 选到了 {} 个源对象", source_layer, len(raw_objs))

    # ==================== Step 2: 绘制矩形并生成信息 ====================
    
    draw_func = globals().get('draw_lwpolyline')
    info_func = globals().get('generate_name_and_ratio_from_com')
    
    if not draw_func:
        log("‼ 致命错误：缺少 draw_lwpolyline 函数")
        return [], {}
    
    new_polys_unsorted = [] # 此时可能包含重叠对象
    final_dict = {} 
    count_drawn = 0
    count_skipped = 0
    
    for obj in raw_objs:
        try:
            # 1. 获取包围盒
            try: ll, ur = obj.GetBoundingBox()
            except: count_skipped += 1; continue 

            minx, miny = float(ll[0]), float(ll[1])
            maxx, maxy = float(ur[0]), float(ur[1])
            
            if abs(maxx - minx) < 1 or abs(maxy - miny) < 1: 
                count_skipped += 1; continue

            # 2. 构造坐标
            coords3d = [
                (minx, miny, z), (minx, maxy, z), (maxx, maxy, z), (maxx, miny, z), (minx, miny, z)
            ]
            
            # 3. 绘制
            new_pl = draw_func(
                coords3d=coords3d,
                layer_name=rect_layer,
                width=width,
                color=color,
                closed=True
            )
            
            if new_pl:
                new_polys_unsorted.append(new_pl)
                count_drawn += 1
                
                # 4. 生成信息
                h = new_pl.Handle
                if info_func:
                    try: final_dict[h] = info_func(new_pl, tol=10)
                    except Exception as e: final_dict[h] = f"InfoError: {e}"
                else:
                    final_dict[h] = "InfoFuncMissing"
                    
        except Exception:
            count_skipped += 1
            pass

    log("🎨 初步绘制 {} 个打印框 (跳过 {} 个)", count_drawn, count_skipped)

    # ==================== Step 3: 几何去重 (新增步骤) ====================
    
    unique_polys = new_polys_unsorted
    dedup_func = globals().get('remove_duplicate_polylines')
    
    if dedup_func and new_polys_unsorted:
        try:
            # 调用去重函数，tol=1.0 表示 1mm 内的框视为同一个
            unique_polys = dedup_func(
                new_polys_unsorted, 
                tol=1.0, 
                priority_layer=rect_layer
            )
            
            diff = len(new_polys_unsorted) - len(unique_polys)
            if diff > 0:
                log("✂️ 执行智能去重: 删除 {} 个重叠框，剩余 {} 个。", diff, len(unique_polys))
            else:
                log("✅ 无重复对象，无需清理。")
                
        except Exception as e:
            log("⚠ 去重步骤执行异常: {}", e)
    else:
        if not dedup_func: log("ℹ️ 未找到去重函数，跳过清理。")

    # ==================== Step 4: 空间排序 ====================
    
    sorted_pls = sort_coms_by_llcorner(unique_polys, cha_Y=cha_Y)
    
    log("🔢 排序完成，返回 {} 个最终结果。", len(sorted_pls))

    poly_z = stc("dy_zhuanyong")
    remove_duplicate_polylines(poly_z)
    return sorted_pls, final_dict

#&&% 从屏幕选择获得打印区域

@timeit
@debuggable
def select_print_areas_from_screen(
    rect_layer: str = "dy_zhuanyong",      # 结果绘制到的专用图层
    width: float = 0.0,
    color: int = 1,
    z: float = 0.0,
    cha_Y: float = 2000,                   # 排序行容差
    debug: bool = False,
):
    """
    【函数编号】: RECOG-005
    【所属模块】: 交互式识别 (Interactive Recognition)
    【功能描述】: 
        屏幕手动选择对象生成打印区域。
        
        核心流程：
        1. 【交互选择】: 暂停脚本，等待用户在 CAD 窗口中框选任意对象。
        2. 【环境清理】: 清空 rect_layer (注意：若用户选中了该层对象，可能会导致引用丢失，需谨慎)。
        3. 【重绘输出】: 计算选中对象的包围盒，在 rect_layer 上绘制新框。
        4. 【数据映射】: 生成打印比例或图名信息。
        5. 【空间排序】: 对生成的结果进行从上到下、从左到右的排序。

    【参数详解】:
        - rect_layer (str): 
            结果输出图层。生成的打印框将绘制在此层。
        - cha_Y (float): 
            空间排序时的行高容差。

    【返回值】:
        - (list, dict): 
            1. sorted_pls: 已排序的【新生成】多段线对象列表。
            2. final_dict: {NewHandle: Info} 字典。


    """

    # —— 日志小工具 —— #
    def log(msg, *args):
        try: 
            text = msg.format(*args) if args else msg
        except: 
            text = msg
        try: 
            print(text)
        except: 
            pass
        n = globals().get("node", None)
        if n: 
            try: 
                n(text)
            except: 
                pass

 

    # —— 0. 环境准备 —— #
    li() 
    
    # ==================== Step 1: 屏幕交互选择 ====================
    log("⏳ 请在 CAD 窗口中框选对象 (按回车结束)...")
    
    raw_objs = []
    try:
        # 调用 pmxz 屏幕选择
        res = pmxz()
        
        if res:
            if isinstance(res, list): 
                raw_objs = res
            elif hasattr(res, "Item"): 
                raw_objs = [res.Item(i) for i in range(res.Count)]
            else: 
                raw_objs = [res]
    except Exception as e:
        log("⚠ pmxz() 调用失败: {}", e)
        return [], {}

    if not raw_objs:
        log("❌ 未选择任何对象。")
        return [], {}
        
    log("📥 用户选择了 {} 个对象", len(raw_objs))

    # ==================== Step 2: 清理结果图层 ====================
    # 注意：要在用户选择完成后再清理，避免清理掉用户可能想参考的旧框
    try:
        ensure_layer(rect_layer)
        objs = stc(rect_layer)
        if objs:
            cnt = 0
            iter_objs = []
            if isinstance(objs, list): 
                iter_objs = objs
            elif hasattr(objs, "Item"): 
                iter_objs = [objs.Item(i) for i in range(objs.Count)]
            else: 
                iter_objs = [objs]

            for o in iter_objs:
                try: 
                    o.Delete()
                    cnt += 1
                except: 
                    pass
            log("✅ 已清空结果图层 '{}' (删除 {} 个旧对象)", rect_layer, cnt)
    except Exception as e:
        log("⚠ 环境准备阶段报错: {}", e)

    # ==================== Step 3: 绘制矩形并生成信息 ====================

    draw_func = globals().get('draw_lwpolyline')
    info_func = globals().get('generate_name_and_ratio_from_com')
    
    if not draw_func:
        log("‼ 致命错误：缺少 draw_lwpolyline 函数")
        return [], {}
    
    new_polys_unsorted = []
    final_dict = {} # {新Handle: Info}
    count_drawn = 0
    count_skipped = 0
    
    for obj in raw_objs:
        try:
            # 1. 尝试获取外包盒
            try:
                ll, ur = obj.GetBoundingBox()
            except:
                count_skipped += 1
                continue 

            minx, miny = float(ll[0]), float(ll[1])
            maxx, maxy = float(ur[0]), float(ur[1])
            
            # 过滤无效尺寸 (点或极小物体)
            if abs(maxx - minx) < 1 or abs(maxy - miny) < 1: 
                count_skipped += 1
                continue

            # 2. 构造矩形坐标
            coords3d = [
                (minx, miny, z), (minx, maxy, z), (maxx, maxy, z), (maxx, miny, z), (minx, miny, z)
            ]
            
            # 3. 在专用图层绘制
            new_pl = draw_func(
                coords3d=coords3d,
                layer_name=rect_layer,
                width=width,
                color=color,
                closed=True
            )
            
            if new_pl:
                new_polys_unsorted.append(new_pl)
                count_drawn += 1
                
                # 4. 生成打印信息
                h = new_pl.Handle
                if info_func:
                    try:
                        info = info_func(new_pl, tol=10)
                        final_dict[h] = info
                    except Exception as e:
                        final_dict[h] = f"InfoError: {e}"
                else:
                    final_dict[h] = "InfoFuncMissing"
                    
        except Exception as e:
            count_skipped += 1
            pass

    log("🎨 已在 '{}' 绘制 {} 个打印框 (跳过 {} 个无效对象)", rect_layer, count_drawn, count_skipped)

    # ==================== Step 4: 排序 ====================
    
    sorted_pls = sort_coms_by_llcorner(new_polys_unsorted, cha_Y=cha_Y)
    
    log("🔢 排序完成，返回 {} 个结果。", len(sorted_pls))
    return sorted_pls, final_dict







def check_valid_rect_pro(raw_ent, verbose=None):
    """
    【辅助函数】矩形有效性深度检查
    参数 verbose: 只有 >= 3 时才会打印内部详细日志
    """
    # 简单的等级解析
    try:
        level = int(verbose) if verbose is not None else 1
    except: level = 1

    def log(msg):
        if level >= 3: sys_logger.info(f"      [Check] {msg}")

    try:
        # 兼容性转换
        ent = _maybe_cast(raw_ent) if '_maybe_cast' in globals() else raw_ent
        obj_name = getattr(ent, "ObjectName", "Unknown")
        
        # 1. 类型检查
        if "Viewport" in obj_name:
            log("失败: 是视口(Viewport)")
            return False, "是视口"
        if "AcDbLine" in obj_name:
            log("失败: 是直线(Line)")
            return False, "是直线"
        if "Polyline" not in obj_name:
            log(f"失败: 类型不符({obj_name})")
            return False, f"类型不符({obj_name})"
        
        # 2. 闭合检查
        is_closed = getattr(ent, "Closed", False)
        if not is_closed:
            coords = getattr(ent, "Coordinates", [])
            if not coords or len(coords) < 4: 
                log("失败: 非闭合且无坐标")
                return False, "非闭合且无坐标"
            
            # 几何闭合计算
            try:
                stride = 2
                if "2dPolyline" in obj_name: stride = 3
                if len(coords) >= 2*stride:
                    sx, sy = coords[0], coords[1]
                    ex, ey = coords[-stride], coords[-stride+1]
                    dist = ((sx-ex)**2 + (sy-ey)**2)**0.5
                    if dist < 1.0: 
                        # log(f"几何闭合通过 (间距{dist:.2f})") # 过于啰嗦，可不打
                        pass
                    else: 
                        log(f"失败: 首尾未闭合 (间距{dist:.2f})")
                        return False, f"未闭合(间距{dist:.1f})"
                else:
                    log("失败: 坐标点不足")
                    return False, "坐标点不足"
            except: pass
        
        log(f"检查通过 ({obj_name})")
        return True, "OK"

    except Exception as e:
        log(f"异常: {e}")
        return False, f"异常:{e}"






#&&% 对多段线去重处理

def remove_duplicate_polylines(
    polylines: list,
    tol: float = 1,
    priority_layer: str = "dy_zhuanyong"
):
    """
    【函数编号】: UTIL-CLEAN-001
    【所属模块】: 数据清洗与优化 (Data Cleaning)
    【功能描述】: 
        高效多段线空间去重算法 (Sweep-Line Algorithm variant)。
        
        核心逻辑：
        1. 【缓存】: 提取所有多段线的包围盒 (BoundingBox) 和图层信息，避免重复调用 COM 接口。
        2. 【排序】: 将对象按 Min-X 坐标排序，这是算法高效的关键。
        3. 【扫掠】: 只对比空间位置相邻的对象。一旦 X 轴距离超过 tol，立即停止后续比对。
        4. 【决策】: 发现重叠时，根据 `priority_layer` 决定保留哪一个（通常保留新建的专用层对象）。
        5. 【清理】: 物理删除 (Delete) 被判定为冗余的对象。

    【参数详解】:
        - polylines (list): 
            待处理的 CAD COM 多段线对象列表。
        - tol (float): 
            重叠判定容差。四个边界坐标差值均在此范围内视为重复。
        - priority_layer (str): 
            “权重”图层。当发生冲突时，位于此图层的对象拥有免死金牌（保留它，删另一个）。

    【返回值】:
        - (list): 去重后幸存的对象列表。
    """
    
    # --- 1. 数据缓存与预处理 ---
    cached_data = []
    priority_layer_lower = priority_layer.lower()
    
    # 用于统计
    initial_count = len(polylines)
    sys_logger.info(f"正在分析 {initial_count} 个对象进行去重...")

    for pl in polylines:
        try:
            # 获取外包盒 (需要处理 COM 异常)
            ll, ur = pl.GetBoundingBox()
            minx, miny = float(ll[0]), float(ll[1])
            maxx, maxy = float(ur[0]), float(ur[1])
            
            # 规范化坐标
            x1, x2 = (minx, maxx) if minx <= maxx else (maxx, minx)
            y1, y2 = (miny, maxy) if miny <= maxy else (maxy, miny)
            
            # 获取图层
            lay = pl.Layer.lower()
            
            cached_data.append({
                'obj': pl,
                'bbox': (x1, y1, x2, y2), # minx, miny, maxx, maxy
                'layer': lay,
                'removed': False,        # 标记删除状态
                'id': id(pl)
            })
        except Exception:
            # 忽略无效对象
            continue

    # --- 2. 排序 (扫掠算法的核心) ---
    # 按 minx 从小到大排序
    cached_data.sort(key=lambda item: item['bbox'][0])

    count_removed = 0
    n = len(cached_data)

    # --- 3. 核心去重循环 ---
    for i in range(n):
        current = cached_data[i]
        
        # 如果当前对象已经被之前的操作标记删除了，直接跳过
        if current['removed']:
            continue

        # 向后搜索潜在的重复项
        for j in range(i + 1, n):
            candidate = cached_data[j]
            
            if candidate['removed']:
                continue

            # 【重要优化】如果 candidate.minx 已经超出了容差范围，
            # 说明后面的对象都在更远的 X 轴位置，无需再比，直接中断内层循环。
            if candidate['bbox'][0] - current['bbox'][0] > tol:
                break

            # 详细对比 4 个坐标 (Y轴和X轴远端)
            # bbox: (minx, miny, maxx, maxy)
            c_box = current['bbox']
            t_box = candidate['bbox']

            is_duplicate = (
                abs(c_box[1] - t_box[1]) <= tol and # miny
                abs(c_box[2] - t_box[2]) <= tol and # maxx
                abs(c_box[3] - t_box[3]) <= tol and # maxy
                abs(c_box[0] - t_box[0]) <= tol     # minx (再次确认，虽然已排序)
            )

            if is_duplicate:
                # 发现重复！根据图层优先级决定去留
                
                w_curr = 1 if current['layer'] == priority_layer_lower else 0
                w_cand = 1 if candidate['layer'] == priority_layer_lower else 0

                if w_curr >= w_cand:
                    # 当前对象优先级更高(或相等) -> 删除候选对象
                    candidate['removed'] = True
                    safe_delete(candidate['obj']) # 使用安全删除
                    count_removed += 1
                else:
                    # 候选对象优先级更高 -> 删除当前对象
                    current['removed'] = True
                    safe_delete(current['obj'])   # 使用安全删除
                    count_removed += 1
                    
                    # 当前主对象 current 既然已经牺牲了，就没资格再去 PK 别人了
                    # 跳出内层循环，让外层循环进入 i+1
                    break 

    # --- 4. 整理幸存者 ---
    survivors = [item['obj'] for item in cached_data if not item['removed']]

    sys_logger.info(f"✅ 去重完成：处理 {len(cached_data)} 个，删除 {count_removed} 个，保留 {len(survivors)} 个。")
    return survivors

#&&&% ===（三）插图签 ===

#&&% 统一插图签
@timeit
@retry_on_busy
def universal_insert_labels_dispatch(
        layout_name=None, 
        operate_target="Model", 
        Select_Config=0,       # ✨ [核心] 0=常规选择, 1=精细选择
        manual_dy_list=None,   
        filepath=None,
        layername="dy_quyu",
        
        debug=False,
        ref_width=None, 
        use_cache=False
    ):
    from  CAD_file_operations   import set_space_mode

    """
    【函数编号】: DISPATCH-001 (V3.5 - 修复整合版)
    【功能】: 
        1. 负责调用 smart_select_polylines 获取对象列表。
        2. 负责计算基准 (Ref_width)。
        3. 负责分发给 Power版(模型) 或 PaperSpace版(布局)。
        4. ✨ 负责调用对应的后续 repair 修复逻辑。
    """
    
    # ============================================================
    # 🚩 【Part 0】 基础连接
    # ============================================================
    try:
        if not C.acad.Visible: 
            C.acad.Visible = True
            time.sleep(0.5)
    except Exception as e:
        sys_logger.error(f"❌ [致命] CAD 连接失败: {e}")
        return False

    # ============================================================
    # 🚩 【Part 0.5】 环境强制对齐
    # ============================================================
    try:
        current_layout = C.doc.ActiveLayout.Name
        
        # 🟢 情况 A: 目标是模型空间
        if operate_target == "Model":
            if current_layout != "Model":
                sys_logger.info("🕹️ [调度器] 正在强制切回模型空间 (Model)...")
                set_space_mode(1) 
                
        # 🔵 情况 B: 目标是布局空间
        elif operate_target == "Layout":
            if layout_name and current_layout != layout_name:
                sys_logger.info(f"🕹️ [调度器] 正在切换至目标布局: {layout_name}")
                try:
                    set_space_mode(0)
                except Exception as e:
                    sys_logger.error(f"❌ 无法切换到布局 '{layout_name}': {e}")
                    return False

    except Exception as e:
        sys_logger.error(f"❌ [调度器] 环境切换失败: {e}")
        return False

    # ============================================================
    # 🚩 【Part 1】 空间架构判定
    # ============================================================
    is_layout_mode = (operate_target == "Layout")
    is_pure_model  = (operate_target == "Model" and (not layout_name or layout_name == "Model"))
    
    mode_desc = "常规模式" if Select_Config == 0 else "精细模式"
    sys_logger.info(f"🏗️ [Dispatch] 模式: {operate_target} | 布局: {layout_name or 'Model'} | 策略: {mode_desc}")

    # ============================================================
    # 🚩 【Part 2】 获取打印区域 (数据源)
    # ============================================================
    final_dy_list = []

    if manual_dy_list is not None:
        sys_logger.info(f"🎮 使用手动输入区域: {len(manual_dy_list)} 个")
        final_dy_list = manual_dy_list
    else:
        # ✨ [核心] 调用智能选择
        scan_result = smart_select_polylines(
            layout_name=layout_name,
            operate_target=operate_target,
            select_config=Select_Config, 
            use_cache=use_cache,
            min_side=100.0, 
        )
        
        if isinstance(scan_result, tuple):
            final_dy_list = scan_result[0]
        elif isinstance(scan_result, list):
            final_dy_list = scan_result
        else:

            final_dy_list = []


    if not final_dy_list:
        sys_logger.warning("⚠️ [中止] 未获取到有效的打印区域框线。")
        return False

    count_areas = len(final_dy_list)
    speak_msg(f" {count_areas} 个打印区域")
    # ============================================================
    # 🚩 【Part 3】 修复基准 (Ref_width)
    # ============================================================
    final_ref_width = 1.0 

    if ref_width is not None:
        final_ref_width = float(ref_width)
    elif is_pure_model:
        final_ref_width = 100.0 # 纯模型
    else:
        final_ref_width = 1.0   # 布局/混合

    # ============================================================
    # 🚩 【Part 4】 任务分发与后处理修复
    # ============================================================
    
    dispatch_success = False

    # 🟢 分支 A: 布局空间
    if is_layout_mode:
        target_layout = layout_name if layout_name else "布局1"
        
        # 1. 执行插入
        insert_ok = insert_and_scale_labels_paper_space(
            dy=final_dy_list,  
            filepath=filepath,
            layername=layername,
            layout_name=target_layout,
            debug=debug,
            Ref_width=final_ref_width,
            operate_target="Layout"
        )
        
        # 2. 如果插入成功，执行布局专用修复
        if insert_ok:
            sys_logger.info("✨ [Layout] 插入成功，开始执行布局修复流程...")
            time.sleep(2)
            wait_quiescent()

            dispatch_success = repair_sp_insert(target_layout_name=target_layout)
        else:
            dispatch_success = False

    # 🔵 分支 B: 模型/混合空间
    else:
        # 1. 执行插入
        insert_ok = insert_and_scale_labels_area_power(
            coms_dayin=final_dy_list, 
            filepath=filepath,
            layername=layername,
            debug=debug,
            layout_name=layout_name,
            Ref_width=final_ref_width,
            operate_target="Model"
        )
        
        # 2. 如果插入成功，执行模型专用修复
        if insert_ok:
            sys_logger.info("✨ [Model] 插入成功，开始执行模型修复流程...")
            # 注意: operate_target 传入原始参数以处理可能的借道逻辑
            time.sleep(2)
            wait_quiescent()

            dispatch_success = repair_mp_insert(target_layout_name=layout_name, operate_target=operate_target)
        else:
            dispatch_success = False

    # ============================================================
    # 🚩 【Part 5】 最终状态稳定
    # ============================================================
    if dispatch_success:
        sys_logger.info("⏳ 任务完成，正在等待系统静止...")

        speak_msg(" 成功插入图签并剥出修复")
        return True
    else:
        sys_logger.error("❌ 任务分发或修复过程中出现错误。")

        speak_msg(" 插入图签失败")
        return False






#&&% 幂等增强插入加速版

@timeit
def insert_and_scale_labels_area_power(
        coms_dayin=None, 
        filepath=None,
        layername="dy_quyu",
        timestamp=None,
        
        debug=False,
        layout_name=None,
        Ref_width=None,     
        operate_target="Model",
        **kwargs          
    ):
    """
    【函数编号】: BLOCK-INSERT-001 (V34 - 同文件直调版)
    【逻辑】: 直接调用本文件内的 select_maxrect_polylines_1
    """
    
    # 1. 路径补全
    final_filepath = filepath
    if not final_filepath:
        import os
        from system.project_setup import PathConfig
        try:
            u_path = os.environ.get('USERPATH') or str(PathConfig.userpath)
            final_filepath = os.path.join(u_path, "dwg文件", "标准图签.dwg")
        except:
            sys_logger.warning("无法自动构建路径，尝试使用默认值")

    # 2. 智能对象抓取
    target_coms = coms_dayin
    
    if target_coms is None:
        sys_logger.info(f"🔍 [兼容层] 启动 Power 专用扫描 (Target={operate_target})...")
        
        try:
            # ✅ [关键修正]
            # 因为 select_maxrect_polylines_1 就在本脚本(CAD_basic.py)里，
            # 所以不需要 "from ... import ..."，直接用名字调用！
            
            # 注意：select_maxrect_polylines_1 原版似乎没有 layout_name 参数，
            # 如果它不支持切布局，我们需要在这里手动切一下，或者透传 kwargs
            if operate_target == "Layout" and layout_name:
                try:
                    from system.licad import C
                    if C.doc.ActiveLayout.Name != layout_name:
                        C.doc.ActiveLayout = C.doc.Layouts.Item(layout_name)
                except: pass

            result = select_maxrect_polylines_1(
                layer_name="dy_zhuanyong", 
                precision_mode=False,
                width=0.0,
                color=1,  
                min_side=100.0,
                # 如果你的 select_maxrect_polylines_1 定义支持 **kwargs，这些参数会传进去
                # 如果不支持，它可能会报错，请根据该函数的实际定义调整
            )
            
            # select_maxrect_polylines_1 返回 (list, dict)
            if isinstance(result, tuple):
                target_coms = result[0]
            else:
                target_coms = result if result else []
            
            if target_coms:
                sys_logger.info(f"✅ [兼容层] 自动捕获 {len(target_coms)} 个对象")
            else:
                sys_logger.warning("⚠️ 未找到符合 dy_zhuanyong/Color=1 的矩形。")

        except Exception as e:
            sys_logger.error(f"❌ 自动抓取过程出错: {e}")
            import traceback
            traceback.print_exc()
            return False

    # 3. 空值阻断
    if not target_coms:
        return False

    # 4. 核心转发
    sys_logger.info(f"🔄 [兼容层] Power版 -> 请求流水线...")

    try:
        from Insert_chart.insert_labels import run_title_block_assembly_pipeline
    except ImportError as e:
        sys_logger.error(f"❌ 无法加载业务模块: {e}")
        return False

    return run_title_block_assembly_pipeline(
        external_coms=target_coms,       
        external_filepath=final_filepath
    )



#&&% 图纸空间的图签插入



# -------------------------------------------------------------
# 主流程函数
# -------------------------------------------------------------

#新

@timeit
def insert_and_scale_labels_paper_space(
        dy=None,                
        filepath=None,
        layername="dy_quyu",
        layout_name="布局1", 
        timestamp=None,
        
        debug=False,
        Ref_width=1.0,
        operate_target="Layout",
        **kwargs
    ):
    """
    【函数编号】: BLOCK-INSERT-002 (V32 - 返回值解包修复版)
    """
    from CAD_file_operations import save_file ,switch_to_layout   
    
    # 1. 环境准备
    try:
        doc = C.doc
        if not C.acad.Visible: C.acad.Visible = True
    except Exception as e:
        sys_logger.info(f"❌ CAD 连接失败: {e}")
        return False

    sys_logger.info(f"\n🚀 [流程启动] 布局 '{layout_name}' 全自动处理模式...")

    # 2. 智能切换布局
    if not switch_to_layout(layout_name, retry=5):
        sys_logger.info(f"❌ 无法切换到目标布局: {layout_name}，流程终止。")
        return False

    # 3. 数据源准备 (智能抓取)
    target_coms = dy
    if target_coms is None:
        sys_logger.info(f"🔍 [自动抓取] 未传入打印框，正在扫描布局 '{layout_name}'...")
        
        # ✅ [关键修正] 获取结果并解包
        # select_print_areas_paperspace 返回 (entities_list, info_dict)
        selection_result = select_print_areas_paperspace(
            layout_name=layout_name,
            min_side=100.0,
            # contain_tol=0.01, # 如果你的函数定义里没这个参数，就删掉，或者通过 kwargs 传
            **kwargs 
        )
        
        # 解包逻辑：我们要的是列表
        if isinstance(selection_result, tuple):
            target_coms = selection_result[0]
        elif isinstance(selection_result, list):
            target_coms = selection_result
        else:
            target_coms = []
        
    # 二次检查
    if not target_coms:
        sys_logger.info(f"⚠️ [流程中止] 在布局 '{layout_name}' 未找到有效的打印区域多段线。")
        return False

    # 打印真实的图元数量
    sys_logger.info(f"✅ 锁定待处理对象: {len(target_coms)} 个")

    # 4. 执行核心流水线 (Direct Call)
    # 直接在当前布局空间操作
    try:
        # 这里传入的 target_coms 现在是纯粹的 COM 对象列表了
        result = insert_and_scale_labels_area_power(
            coms_dayin=target_coms,   # 传入对象列表
            filepath=filepath,
            layername=layername,
            timestamp=timestamp,
            
            debug=debug,
            layout_name=layout_name,
            Ref_width=Ref_width,
            operate_target=operate_target,
            **kwargs
        )
        
        if result is False: 
            sys_logger.warning("❌ 核心流水线执行返回失败。") # 改用 sys_logger 记录
            return False

    except Exception as e:
        sys_logger.info(f"❌ 核心流水线执行报错: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # 5. 结尾收官
    wait_command_done()
    
    sys_logger.info(f"✅ [流程结束] 布局 '{layout_name}' 处理完毕。")
    
    try:
        doc.Regen(1) 
        save_file()
    except: pass
        
    return True



def clean_blocks_until_vanished(target_names, max_retry_loops=3):
    """
    【函数编号】: BLK-CLEAN-001
    【功能描述】: 
        循环清理指定的图块列表，直到检测不到任何残留，或达到最大循环次数。
        
    【执行流程】:
        1. 遍历名单执行删除。
        2. 扫描当前模型空间的所有块实例 (select_kuai)。
        3. 扫描当前图纸的所有块定义 (doc.Blocks)。
        4. 如果仍有残留，进入下一轮循环；否则提前结束。
        
    【参数】:
        - target_names (list): 要删除的块名列表。
        - max_retry_loops (int): 最大重复轮数 (默认3次)。
    """
    
    sys_logger.info(f"🚀 开始深度清理 {len(target_names)} 个目标图块...")

    for round_idx in range(1, max_retry_loops + 1):
        sys_logger.info(f"\n🔄 --- 第 {round_idx}/{max_retry_loops} 轮清理 ---")
        
        # ================= 1. 执行删除操作 =================
        for name in target_names:
            try:
                # 调用你原有的删除函数 (假设它已定义)
                # max_rounds=2 是单次删除内部的重试，这里是外部大循环
                if "delete_block_instances_and_definition_retry" in globals():
                    delete_block_instances_and_definition_retry(name, max_rounds=2)
                else:
                    sys_logger.info(f"❌ 缺少核心删除函数，跳过: {name}")
            except Exception as e:
                sys_logger.info(f"⚠️ 删除出错 [{name}]: {e}")

        # ================= 2. 校验残留 (Verification) =================
        sys_logger.info(f"🔍 第 {round_idx} 轮校验中...")
        
        remaining_targets = set()
        
        # --- A. 检查场景中的实例 (Instances) ---
        try:
            # 获取所有块实例 (假设 select_kuai 返回 COM 对象列表)
            all_instances = select_kuai() 
            if all_instances:
                for blk in all_instances:
                    # 获取块名 (优先取动态块真实名)
                    b_name = get_attr(blk, "EffectiveName")
                    if not b_name: b_name = get_attr(blk, "Name")
                    
                    if b_name and b_name in target_names:
                        remaining_targets.add(b_name)
        except Exception as e:
            sys_logger.info(f"⚠️ 校验实例时出错: {e}")

        # --- B. 检查块定义 (Definitions) ---
        # 即使实例删光了，定义可能还在 (未 Purge)
        try:
            # 遍历目标名单，直接去 Blocks 集合里查，比遍历整个 Blocks 集合更快
            for t_name in target_names:
                if t_name in remaining_targets: continue # 已经发现残留了，不用查定义了
                
                try:
                    # 尝试获取块定义，不报错说明还存在
                    doc.Blocks.Item(t_name)
                    remaining_targets.add(t_name)
                except:
                    # 报错说明找不到了，说明删干净了
                    pass
        except Exception as e:
            sys_logger.info(f"⚠️ 校验定义时出错: {e}")

        # ================= 3. 判定结果 =================
        if not remaining_targets:
            sys_logger.info(f"✅ 完美！所有目标图块已彻底清除 (耗时 {round_idx} 轮)。")
            return True
        else:
            sys_logger.info(f"⚠️ 本轮结束后仍有 {len(remaining_targets)} 个顽固分子: {list(remaining_targets)}")
            if round_idx < max_retry_loops:
                print("⏳ 准备执行下一轮强力清除...")
            else:
                print("❌ 达到最大尝试次数，清理结束 (仍有残留)。")

    return False


#&&% 坐标转多段线COM
def plcoor_to_com(coord_info, layer_name="测试辅助", width=0, color=256):
    """
    根据坐标数据批量创建轻量多段线并返回 COM 对象列表。

    coord_info 结构：[(coords, flag), ...]
        coords: [(x, y), (x, y), ...]
        flag:   预留标志位（当前没用也无所谓）
    """

    doc = C.doc

    # 图层准备
    try:
        layer = doc.Layers.Item(layer_name)
    except Exception:
        layer = doc.Layers.Add(layer_name)
    set_attr(layer, "LayerOn", True)

    polys = []
    for coords, flag in coord_info:
        data = []
        for x, y in coords:
            data.extend((x, y))
        arr = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8, data
        )
        poly = mp.AddLightWeightPolyline(arr)
        set_attr(poly, "Layer", layer_name)
        set_attr(poly, "ConstantWidth", width)
        set_attr(poly, "Color", color)
        set_attr(poly, "Closed", True)
        polys.append(poly)

    node("[plcoor_to_com] 已在图层 '{}' 上生成 {} 条多段线", layer_name, len(polys))
    return polys


#&&% 绘制PL并提取信息
def draw_pl_and_extract_info(resu, layer_name="测试辅助", width=0, color=256):
    """
    resu: List[(blk_entity, corners)]
        corners = [(x_ll, y_ll, z), (x_lu, y_lu, z), (x_ru, y_ru, z), (x_rl, y_rl, z)]
    在指定图层绘制外包盒多段线，并调用 generate_name_and_ratio_from_polyline
    提取图幅 / 比例 / 规格。
    """
    doc=C.doc
    mp= C.mp

    # 确保图层存在
    try:
        layer = doc.Layers.Item(layer_name)
    except Exception:
        layer = doc.Layers.Add(layer_name)
    try:
        set_attr(layer, "LayerOn", True)
    except Exception:
        pass

    results = {}

    for idx, (blk, corners) in enumerate(resu):
        # corners 是四个三元组，但 LWPOLYLINE 只要 2D XY
        raw = []
        for x, y, _ in corners:
            raw.extend((x, y))
        arr = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_R8,
            raw
        )

        lw = mp.AddLightWeightPolyline(arr)
        set_attr(lw, "Layer", layer_name)
        set_attr(lw, "ConstantWidth", width)
        set_attr(lw, "Color", color)
        set_attr(lw, "Closed", True)

        # 图幅/比例/规格分析
        try:
            info = generate_name_and_ratio_from_polyline(lw)
        except Exception:
            info = None

        if info and len(info) >= 3:
            frame, ratio, spec = info[:3]
        else:
            frame = ratio = spec = None

        blk_name = None
        try:
            blk_name = get_attr(blk, "Name")
        except Exception:
            pass

        results[idx] = {
            "block_name":    blk_name,
            "corners":       corners,
            "drawing_frame": frame,
            "ratio":         ratio,
            "spec":          spec,
            "polyline":      lw,
        }

    try:
        print_com_info(results)
    except Exception:
        pass

    try:
        doc.SendCommand("RE\n")
        doc.SendCommand("Z\nE\n")
    except Exception:
        pass
    time.sleep(0.3)
    return results


#&&% 从实体绘制PL并提取
def draw_pl_and_extract_from_entities(
    entities,
    layer_name: str = "测试辅助",
    width: int = 0,
    color: int = 256,
    A3dy: int = 0,
    Fandy: Tuple[str, str, str, int] = ("ISO_A3_(420.00_x_297.00_MM)", "0:0", "A3", 0),
):
    """
    把 entities 统一转换为闭合多段线，并提取图幅 / 比例 / 规格信息。
    返回 results 字典，下游以 info["entity"] 作为区域多段线引用。
    """
    doc = C.doc
    mp=C.mp
    coord_info, corners_list = [], []
    for ent in entities:
        if not hasattr(ent, "GetBoundingBox"):
            continue
        try:
            p1, p2 = ent.GetBoundingBox()
        except Exception:
            continue

        minx, miny, _ = p1
        maxx, maxy, _ = p2
        corners = [
            (minx, miny, 0),
            (minx, maxy, 0),
            (maxx, maxy, 0),
            (maxx, miny, 0),
        ]
        corners_list.append(corners)
        coord_info.append(([(x, y) for x, y, _ in corners], 1))

    
    plines = plcoor_to_com(
        coord_info, layer_name=layer_name, width=width, color=color
    )

    results = {}
    for idx, (corners, pline) in enumerate(zip(corners_list, plines)):
        # 所属块名（如果在块内）
        block_name = None
        try:
            owner = get_attr(pline, "Owner")
            owner_block = get_attr(owner, "Block") if owner is not None else None
            target = owner_block or owner
            if target is not None:
                block_name = get_attr(target, "Name")
        except Exception:
            block_name = None

        # 使用 COM 版本的图幅分析函数
        try:
            info = generate_name_and_ratio_from_com(
                pline, A3dy=A3dy, Fandy=Fandy
            )
        except Exception:
            info = 0

        if info and len(info) >= 3:
            drawing_frame, ratio, spec = info[:3]
            orient = info[3] if len(info) > 3 else None
        else:
            drawing_frame = ratio = spec = orient = None

        results[idx] = {
            "entity":        pline,
            "block_name":    block_name,
            "corners":       corners,
            "drawing_frame": drawing_frame,
            "ratio":         ratio,
            "spec":          spec,
            "orient":        orient,
        }

    try:
        print_com_info(results)
    except Exception:
        pass

    try:
        doc.SendCommand("RE\n")
        doc.SendCommand("Z\nE\n")
    except Exception:
        pass
    time.sleep(0.3)
    return results


#&&% 区域内插入块
def insert_block_into_poly_area(block_name, poly_ent, k=1.0, max_retries=3):
    """
    在多段线/多边形 poly_ent 所定义区域内插入已定义块，
    横向区域（宽 >= 高）在左下插入；竖向区域在左上插入并顺时针旋转 90°。
    """
    import time
    import math
    import win32com.client
    import pythoncom

    if not hasattr(poly_ent, "GetBoundingBox"):
        raise TypeError("poly_ent 必须具备 GetBoundingBox 方法")

    li()
    global doc, mp
    if doc is None or mp is None:
        raise RuntimeError("[insert_block_into_poly_area] doc/mp 为 None")

    p1, p2 = poly_ent.GetBoundingBox()
    minx, miny, minz = p1
    maxx, maxy, _ = p2
    width = maxx - minx
    height = maxy - miny

    orientation = 0 if width >= height else 1
    ins_pt = (minx, miny, minz) if orientation == 0 else (minx, maxy, minz)

    ins_var = win32com.client.VARIANT(
        pythoncom.VT_ARRAY | pythoncom.VT_R8,
        ins_pt,
    )

    block_ref = None
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            block_ref = mp.InsertBlock(ins_var, block_name, k, k, k, 0.0)
            break
        except Exception as exc:
            last_err = exc
            sys_logger.info(f"⚠ 第 {attempt} 次插入块失败: {exc}")
            time.sleep(0.5)
    else:
        raise RuntimeError(f"❌ 多次尝试仍无法插入块 {block_name}: {last_err}")

    # 竖向时顺时针 90°
    if orientation == 1 and block_ref is not None:
        try:
            set_attr(block_ref, "Rotation", -math.pi / 2)
        except Exception as exc:
            print("⚠ 设置 Rotation 失败：", exc)

    return orientation, ins_pt, block_ref

#&&% 区域内插入块20260109
def insert_block_into_poly_area(block_name, poly_ent, k=1.0, max_retries=3):
    """
    【修正版 V2】自动识别 poly_ent 所在空间（模型或布局），直接使用 C.doc。
    横向区域在左下插入；竖向区域在左上插入并顺时针旋转 90°。
    """

    if not hasattr(poly_ent, "GetBoundingBox"):
        raise TypeError("poly_ent 必须具备 GetBoundingBox 方法")

    # ———— 修改点：直接使用 C.doc，不再依赖 li() 或 global mp ————
    doc = C.doc
    if doc is None:
        raise RuntimeError("[insert_block_into_poly_area] C.doc 为 None，请检查 CAD 连接")

    # ———— 核心修复：通过 OwnerID 动态获取目标空间 ————
    try:
        # poly_ent.OwnerID 指向它所在的 BlockTableRecord（即模型空间或某个图纸空间布局）
        target_space = doc.ObjectIdToObject(poly_ent.OwnerID)
    except Exception as e:
        sys_logger.info(f"⚠ 无法自动获取空间，降级使用 ModelSpace: {e}")
        target_space = doc.ModelSpace
    # ——————————————————————————————————————————————

    p1, p2 = poly_ent.GetBoundingBox()
    minx, miny, minz = p1
    maxx, maxy, _ = p2
    width = maxx - minx
    height = maxy - miny

    orientation = 0 if width >= height else 1
    # 坐标系已通过 target_space 对齐，直接使用原始坐标即可
    ins_pt = (minx, miny, minz) if orientation == 0 else (minx, maxy, minz)

    ins_var = win32com.client.VARIANT(
        pythoncom.VT_ARRAY | pythoncom.VT_R8,
        ins_pt,
    )

    block_ref = None
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            # 使用动态获取的 target_space (而不是 mp) 进行插入
            block_ref = target_space.InsertBlock(ins_var, block_name, k, k, k, 0.0)
            break
        except Exception as exc:
            last_err = exc
            sys_logger.info(f"⚠ 第 {attempt} 次插入块失败: {exc}")
            time.sleep(0.5)
    else:
        raise RuntimeError(f"❌ 多次尝试仍无法插入块 {block_name}: {last_err}")

    # 竖向时顺时针 90°
    if orientation == 1 and block_ref is not None:
        try:
            # 确保 set_attr 在作用域内可用
            set_attr(block_ref, "Rotation", -math.pi / 2)
        except Exception as exc:
            print("⚠ 设置 Rotation 失败：", exc)

    return orientation, ins_pt, block_ref



#&&% 计算插入因子
def compute_insert_factors(entities, res, result_dict):
    """
    根据 res 的 ratio/spec 与 result_dict（图签模板信息）的块定义，计算缩放系数 k。
    回传 [(entity, block_name, spec, k), ...]
    """
    import re

    def _denom(s: str) -> int:
        m = re.match(r".*:(\d+)$", s or "")
        return int(m.group(1)) if m else 1

    mapping = {
        info.get("spec"): (info.get("block_name"), info.get("ratio"))
        for info in result_dict.values()
        if info.get("spec")
    }

    outputs = []
    for ent in entities:
        matched = next(
            (info for info in res.values() if info.get("entity") is ent),
            None
        )
        if not matched:
            outputs.append((ent, None, None, None))
            continue

        spec = matched.get("spec")
        ratio1 = matched.get("ratio") or "1:1"
        block_name, ratio2 = mapping.get(spec, (None, "1:1"))

        d1, d2 = _denom(ratio1), _denom(ratio2)
        k = d1 / d2 if d2 else None
        outputs.append((ent, block_name, spec, k))

    return outputs


#&&% 获取实体因子
def get_factor_for_entity(entity, factors):
    """在 factors 中找到第一项为 entity 的元组。"""
    for tup in factors:
        if tup[0] is entity:
            return tup
    return None


# =============================
#  插入公司通用图签块（单线程）
# =============================

#&&% 插入公司通用图签单线程
def insert_company_label_common_block(
    insertion_point=(0, 0, 0),
    filepath=r"D:/Myprogramsystem/XT/标准图签模板.dwg",
    scale=(1, 1, 1),
    rotation=0,
    wait=0.3,
):
    """
    直接插入“公司通用图签块”DWG，并炸开，生成外包盒多段线信息。

    返回:
        (coms_tuqian, names_tuqian, info_dict)

    r"D:/Myprogramsystem/XT/标准图签模板.dwg"并不影响最后的实际使用

    但是修改为相对路径就会出错，不要修改它了20251216

    """
    import time
    from CAD_file_operations import save_file
    C.li()
    # 1) 插入并炸开 DWG
    resu_all = insert_and_explode_dwg(
        filepath,
        insertion_point=insertion_point,
        scale=scale,
        rotation=rotation,
        wait=wait,
    )
    # insert_and_explode_dwg 通常返回 [resu]
    resu = resu_all[0]

    # 2) 外包盒 → 多段线 → 图幅信息
    info_dict = draw_pl_and_extract_info(
        resu, layer_name="测试辅助", width=0, color=256
    )

    # 3) 整理块引用与块名
    coms_tuqian = []
    names_tuqian = []
    for ob in resu:
        blk = ob[0]
        coms_tuqian.append(blk)
        names_tuqian.append(get_attr(blk, "Name"))

    try:
        time.sleep(1.0)
        save_file()
    except Exception:
        pass

    return coms_tuqian, names_tuqian, info_dict


# =============================
#  双线程：忽略 SHX 字体对话框
# =============================

#&&% 线程1:插入公司图签
def f1_insert_company_getwindow(
    timeout_event,
    done_event,
    *,
    result_box: dict,
    insertion_point=(0, 0, 0),
    filepath=r"D:/Myprogramsystem/XT/标准图签模板.dwg",
    scale=(1, 1, 1),
    rotation=0,
    wait=0.3,
):
    """

    线程1：插入公司图签块，并将结果写入 result_box['data']。

    r"D:/Myprogramsystem/XT/标准图签模板.dwg"并不影响最后的实际使用

    但是修改为相对路径就会出错，不要修改它了20251216

    """






    import pythoncom
    import traceback

    pythoncom.CoInitialize()
    try:
        node("线程1启动（Insert_Company_Label_Common_Block）")
        li()
        result = insert_company_label_common_block(
            insertion_point=insertion_point,
            filepath=filepath,
            scale=scale,
            rotation=rotation,
            wait=wait,
        )
        node("线程1完成插入, 写入 result_box")
        result_box["data"] = result
        timeout_event.wait()
    except Exception as exc:
        print("f1_insert_company_getwindow:", exc, traceback.format_exc())
    finally:
        pythoncom.CoUninitialize()
        done_event.set()



def _find_shx_dialog(shx_titles=("缺少 SHX", "Missing SHX")) -> int:
    """查找“缺少 SHX 字体”对话框窗口句柄。"""
    import win32gui

    result = []

    def _enum(hwnd, _):
        title = win32gui.GetWindowText(hwnd)
        if any(k in title for k in shx_titles):
            result.append(hwnd)

    win32gui.EnumWindows(_enum, None)
    return result[0] if result else 0



def _ignore_shx_dialog(hwnd: int):
    """向“缺少 SHX 字体”对话框发送 ESC 键关闭。"""
    import win32gui
    import win32con

    win32gui.PostMessage(hwnd, win32con.WM_KEYDOWN, win32con.VK_ESCAPE, 0)
    win32gui.PostMessage(hwnd, win32con.WM_KEYUP,   win32con.VK_ESCAPE, 0xC0000000)


#&&% 线程2:删除窗口
def f2_delwindow(
    timeout_event,
    done_event,
    *,
    poll_sec: float = 0.5,
    limit_sec: int = 3,
    shx_titles=("缺少 SHX", "Missing SHX"),
):
    """线程2：轮询“缺少 SHX”对话框并自动 ESC 关闭。"""
    import pythoncom
    import time
    import traceback

    pythoncom.CoInitialize()
    try:
        node("线程2启动（监控 SHX 对话框）")
        t0 = time.time()
        while time.time() - t0 < limit_sec:
            if timeout_event.is_set():
                break
            hwnd = _find_shx_dialog(shx_titles)
            if hwnd:
                node("🖱 检测到 SHX 对话框，发送 ESC")
                _ignore_shx_dialog(hwnd)
                timeout_event.set()
                break
            time.sleep(poll_sec)
        else:
            # 在 limit_sec 内未见对话框，也放行
            timeout_event.set()
    except Exception as exc:
        print("f2_delwindow 异常:", exc, traceback.format_exc())
        timeout_event.set()
    finally:
        pythoncom.CoUninitialize()
        done_event.set()



def run_dual_threads(
    f1,
    f2,
    f1_args=(),
    f1_kwargs=None,
    f2_args=(),
    f2_kwargs=None,
    *,
    timeout_sec: int = 300,
    spawn_f1: bool = True,
):
    """
    启动 f1 / f2 协作：
      - f1(timeout_event, done_event, ...)
      - f2(timeout_event, done_event, ...)

    spawn_f1 = False 表示在当前线程执行 f1（适合 COM 主线程场景）。
    """
    import threading
    import time

    if f1_kwargs is None:
        f1_kwargs = {}
    if f2_kwargs is None:
        f2_kwargs = {}

    timeout_event = threading.Event()
    done_event = threading.Event()

    def _run_f1():
        try:
            f1(timeout_event, done_event, *f1_args, **f1_kwargs)
        finally:
            # 确保无论如何 timeout_event 会被 set
            timeout_event.set()

    def _run_f2():
        f2(timeout_event, done_event, *f2_args, **f2_kwargs)

    # 启动线程2（永远子线程）
    t2 = threading.Thread(target=_run_f2, daemon=True)
    t2.start()

    # f1 可以选择在当前线程，或子线程
    if spawn_f1:
        t1 = threading.Thread(target=_run_f1, daemon=True)
        t1.start()
        t1.join(timeout=timeout_sec)
    else:
        _run_f1()
        t1 = None

    t2.join(timeout=timeout_sec)

    if not done_event.is_set():
        timeout_event.set()
        node("[WARN] 双线程执行超时 {}s", timeout_sec)
        return False

    node("[OK] 双线程任务完成")
    return True


#&&% 插入公司通用图签(双线程)
def Insert_Company_Label_Common_Block(
    insertion_point=(0, 0, 0),
    filepath=r"D:/Myprogramsystem/XT/标准图签模板.dwg",
    scale=(1, 1, 1),
    rotation=0,
    wait=0.3,
    *,
    timeout_sec: int = 300,
):
    """
    高层：在忽略 SHX 对话框的前提下插入公司图签块。
    返回: (coms_tuqian, names_tuqian, info_dict) 或 (None, None, None)
    """
    result_box = {}
    ok = run_dual_threads(
        f1=f1_insert_company_getwindow,
        f2=f2_delwindow,
        f1_kwargs={
            "result_box":     result_box,
            "insertion_point": insertion_point,
            "filepath":        filepath,
            "scale":           scale,
            "rotation":        rotation,
            "wait":            wait,
        },
        timeout_sec=timeout_sec,
        # 对 COM 更安全：f1 在当前主线程执行
        spawn_f1=False,
    )
    if ok and "data" in result_box:
        return result_box["data"]
    return None, None, None


# =============================
#  图签块的清理/重命名/调试填属性
# =============================

#&&% 清理内部多段线
def clean_internal_polylines(block_refs):
    """删除块定义内图层为 'tuqian_neibu_pl' 的多段线。"""
    import time

    li()
    global doc

    deleted = []
    for blk_ref in block_refs:
        try:
            blk_name = get_attr(blk_ref, "Name")
            if not blk_name:
                continue
            blk_def = doc.Blocks.Item(blk_name)
        except Exception:
            continue

        count = getattr(blk_def, "Count", 0)
        entities = [blk_def.Item(i) for i in range(count)]
        for ent in entities:
            layer_name = get_object_property(ent, "Layer")
            obj_name = get_object_property(ent, "ObjectName")
            if layer_name == "tuqian_neibu_pl" and obj_name in (
                "AcDb2dPolyline",
                "AcDbPolyline",
            ):
                handle = get_object_property(ent, "Handle")
                if handle:
                    deleted.append(handle)
                safe_delete(ent)

    try:
        time.sleep(1)
        doc.SendCommand("RE\n")
    except Exception:
        pass
    return deleted





def fill_block_attributes_with_tag_name(blocks):
    """
    将块属性文字设置为对应的标签名（用于调试，直接在图签上显示 Tag）。
    """
    for blk in blocks:
        if get_object_property(blk, "ObjectName") != "AcDbBlockReference":
            continue

        try:
            if not get_attr(blk, "HasAttributes"):
                continue
            attrs = blk.GetAttributes()
        except Exception:
            continue

        for att in attrs:
            tag = get_attr(att, "TagString")
            if tag is None:
                continue
            set_attr(att, "TextString", tag)



def _make_bind_dict_serializable(bind_dict):
    """
    将 insert_and_scale_labels_area 生成的 bind_dict 压缩成
    只包含 JSON 可序列化的简单结构（数字、字符串、dict、list）。

    原始结构示意：
        {
            <blk_handle>: {
                "frame_info": {..., "entity": <COM Polyline>, ...},
                "title_block": <COM BlockReference>,
            },
            "dyx_list": [<COM Polyline>, ...],
            "tq_list": [<COM BlockReference>, ...],
        }

    转换后示意：
        {
            <blk_handle>: {
                "frame_info": {
                    ... 原 frame_info 去掉 COM ...
                    "entity_handle": "...",
                },
                "title_block_handle": "...",
            },
            "dyx_handles": ["...", "..."],
            "tq_handles": ["...", "..."],
        }
    """
    ser = {}

    # 1) 先处理每个图签块条目（键是 Handle 的那部分）
    for key, val in bind_dict.items():
        # 跳过后面会单独处理的几个特殊键
        if key in ("dyx_list", "tq_list"):
            continue

        # 大部分 key 是字符串 Handle
        if not isinstance(key, str):
            try:
                k_str = str(key)
            except Exception:
                k_str = repr(key)
        else:
            k_str = key

        if not isinstance(val, dict):
            ser[k_str] = val
            continue

        frame_info = val.get("frame_info", {}) or {}
        title_block = val.get("title_block")

        # ---- 压缩 frame_info ----
        fi_new = {}
        for k2, v2 in frame_info.items():
            # entity/polyline 是 COM，需要改成 Handle
            if k2 in ("entity", "polyline"):
                try:
                    h = get_attr(v2, "Handle")
                except Exception:
                    h = None
                fi_new[k2 + "_handle"] = h
            else:
                # 其他字段一般是字符串/数字，可以直接放
                fi_new[k2] = v2

        # ---- 压缩 title_block COM ----
        try:
            h_title = get_attr(title_block, "Handle")
        except Exception:
            h_title = None

        ser[k_str] = {
            "frame_info": fi_new,
            "title_block_handle": h_title,
        }

    # 2) 打印区域多段线列表 / 图签块列表：只保留 Handle 数组
    dyx_list = bind_dict.get("dyx_list", [])
    tq_list = bind_dict.get("tq_list", [])

    def _handles_from_list(objs):
        hs = []
        for o in objs:
            try:
                h = get_attr(o, "Handle")
            except Exception:
                h = None
            hs.append(h)
        return hs

    ser["dyx_handles"] = _handles_from_list(dyx_list)
    ser["tq_handles"] = _handles_from_list(tq_list)

    return ser


def normalize_core_title_blocks_by_layer(
    core_layer: str = "dy_quyu_H",
    core_base_names=None,
    verbose: bool = True,
):
    """
    【桥接函数】
    功能：调用业务层 insert_labels.py 中的规范化函数。
    注意：使用了函数内延迟导入，防止循环引用。
    """
    # ⬇️⬇️⬇️ 关键点：在函数内部导入，而不是文件顶部 ⬇️⬇️⬇️
    try:
        # 这里假设 Insert_chart 包在 sys.path 路径下
        from Insert_chart import insert_labels as ilb
    except ImportError as e:
        sys_logger.info(f"❌ 无法加载业务模块: {e}")
        return False

    # 转发调用，并将参数原样传递
    return ilb.normalize_core_title_blocks_by_layer_new1(
        core_layer=core_layer,
        core_base_names=core_base_names,
        verbose=verbose
    )

#&&% 炸开图签壳块
@debuggable
def explode_title_wrappers_to_core_layer(
    wrapper_layer: str = "dy_quyu",
    core_layer: str = "dy_quyu_H",
    core_base_names=None,
    verbose: bool = True,
):
    """
    【步骤 1】炸开 "dy_quyu" 上的图签壳块，把内部的核心块 A?-H 释放到 "dy_quyu_H" 层。

    约定：
    - 外壳块实例所在图层：wrapper_layer（默认 dy_quyu）
    - 核心块基础名：A3-H, A2-H, A1-H, A0-H（可以通过 core_base_names 自定义）
    - 核心块实例在壳块块定义中必须放在 core_layer（默认 dy_quyu_H）上

    返回：
        int: 实际炸开的壳块数量
    """
    import pythoncom

    global acad, doc, mp, sp

    if core_base_names is None:
        core_base_names = ["A3-H", "A2-H", "A1-H", "A0-H"]

    def log_err(msg):
        if verbose:
            print("[错误]", msg)

    def log_warn(msg):
        if verbose:
            print("[警告]", msg)

    # 小工具：判断是否块参照
    def is_block_ref(ent) -> bool:
        try:
            on = getattr(ent, "ObjectName", "")
        except Exception:
            return False
        return "BlockReference" in str(on)

    # 小工具：取基础名（去掉后缀）
    def get_base_name(name: str) -> str:
        return str(name).split("_")[0]

    if not li():
        log_err("li() 连接失败，无法爆炸图签壳块。")
        return 0

    node("▶ 0  explode_title_wrappers_to_core_layer：开始炸开壳块")
    node("▶ 0.1 wrapper_layer='{}', core_layer='{}'", wrapper_layer, core_layer)

    # 1. 从壳层选对象
    node("▶ 1  从图层 '{}' 选择壳块候选（stc）…", wrapper_layer)
    objs = stc(wrapper_layer)
    wrappers = []
    for e in objs:
        if not is_block_ref(e):
            continue
        try:
            name = str(e.Name)
        except Exception:
            continue
        base = get_base_name(name)
        # 壳块：基础名不在核心名列表
        if base not in core_base_names:
            wrappers.append(e)

    node("▶ 1.1 图层 '{}' 共 {} 个对象，其中壳块 {} 个",
         wrapper_layer, len(objs), len(wrappers))

    if not wrappers:
        log_warn(f"图层 '{wrapper_layer}' 上未发现壳块（基础名不在 {core_base_names}）。")
        return 0

    exploded_ok = 0
    exploded_fail = 0

    # 2. 逐个 Explode + Erase 壳块
    node("▶ 2  开始逐个 Explode + Erase 壳块，共 {} 个…", len(wrappers))
    for e in wrappers:
        try:
            w_name = str(e.Name)
        except Exception:
            w_name = "<?>"

        try:
            _ = safe_explode_and_delete(e)  # 内部对象（含核心块参照）弹出到模型空间
            e.Erase()         # 删除壳块
            exploded_ok += 1
            node("▶ 2.1 已炸开壳块：'{}'", w_name)
        except pythoncom.com_error as exc:
            exploded_fail += 1
            log_warn(f"炸开壳块失败：{w_name}，错误：{exc}")

    node(
        "▶ 2.2 壳块处理完成：成功 {} 个，失败 {} 个",
        exploded_ok, exploded_fail
    )

    # 3. 简单 RE 刷新视图（可选）
    try:
        doc.SendCommand("RE\n")
    except Exception:
        pass

    return exploded_ok






#&&% 模型空间修复插入

# ============================================================
# 1. 主修复入口 (已移除 width_factor)
# ============================================================
def repair_mp_insert(target_layout_name=None, operate_target="Model"):
    """
    【函数功能】: 修复模型空间插入对象的线宽与图层 (V2.1 - 纯净版)
    【变更】: 彻底移除了 width_factor，全权委托给 smart_repair_frame_polyline_widths_m 进行几何分析。
    """
    from CAD_file_operations import save_file, set_space_mode 

    # --- 1. 环境准备 ---
    current_doc = C.doc
    def log(msg):
        sys_logger.info(f"[Repair] {msg}")

    # --- 2. 判定模式 ---
    is_borrow_mode = (operate_target == "Layout" and target_layout_name is not None)

    try:
        # --- 3. 切换环境 ---
        set_space_mode(1)
        time.sleep(0.5)

        # --- 4. 转换边框 (仅负责转多段线，线宽设为0等待智能修复) ---
        log("处理图层: tuqian_baobu (转多段线)...")
        lines_to_polylines(Lc=None, tol=0.5, layer_name="tuqian_baobu", width=0.0, color=256)
        wait_command_done()
        
        # --- 5. 转换内部线条 ---
        log("处理图层: tuqian_neibu_pl (细线化)...")
        lines_to_polylines(Lc=None, tol=0.5, layer_name="tuqian_neibu_pl", width=0.0, color=256)
        wait_command_done()

        # --- 6. 图层迁移 ---
        if not is_borrow_mode:
            layers_to_migrate = ["tuqian_neibu_pl", "dy_quyu_H", "tuqian_baobu"]
            for old_layer in layers_to_migrate:
                new_layer = f"{old_layer}_m"
                objs = stc(old_layer) 
                if objs:
                    log(f"迁移 {old_layer} -> {new_layer}")
                    set_layer_with_retry(objs, new_layer, ci=3)
        else:
            log("🛑 借道模式: 跳过图层后缀添加 (_m)，保持原始图层名。")

        time.sleep(1)

        # --- 7. 智能线宽修复 (无参调用，全自动) ---
        smart_repair_frame_polyline_widths_m()

        time.sleep(1)

        # 清理工作
        ensure_layer_model_only(layer_name="dy_quyu")

        # —————— 清理 ——————

        try:
            set_space_mode(1)                     

            wait_quiescent()
            ensure_layer("测试辅助")

            ensure_layer("dy_zhuanyong")

        except:
        
            pass   

        return True

    except Exception as e:
        sys_logger.info(f"❌ repair_mp_insert 执行出错: {e}")
        return False




#&&% 图纸空间样式修复插入 

def repair_sp_insert(target_layout_name="布局1"):
    """
    【函数功能】: 直接在布局空间修复图签 (无需借道模型空间)
    【逻辑】: 
        1. 切换到目标布局。
        2. 原地将线条转换为多段线 (tuqian_baobu, tuqian_neibu_pl)。
        3. 调用 smart_repair_frame_polyline_widths_p 进行智能线宽归整。
    """

    from  CAD_file_operations   import  switch_to_layout
    import time
    current_doc = C.doc
    
    def log(msg):
        sys_logger.info(f"[Repair-Layout] {msg}")

 

    try:
        
        switch_to_layout(target_layout_name, retry=10, delay=0.5)


        time.sleep(0.5)

        # --- 2. 第一轮转换: 外部边框 (tuqian_baobu) ---
        log("处理图层: tuqian_baobu (转多段线)...")
        # 线宽设为 0.0，全权委托给后续的 smart_repair 处理
        lines_to_polylines(Lc=None, tol=0.5, layer_name="tuqian_baobu", width=0.0, color=256)
        wait_command_done()

        # --- 3. 第二轮转换: 内部线条 (tuqian_neibu_pl) ---
        log("处理图层: tuqian_neibu_pl (细线化)...")
        lines_to_polylines(Lc=None, tol=0.5, layer_name="tuqian_neibu_pl", width=0.0, color=256)
        wait_command_done()

        # --- 4. 智能线宽修复 (核心) ---
        # 此时对象就在布局空间，函数会自动应用 "Small(1:1)" 模式或 "Layout" 规则
        time.sleep(0.5)
        log("启动智能线宽修复...")
        smart_repair_frame_polyline_widths_p()
        
        # --- 5. 收尾 ---
        log(f"✅ 布局 '{target_layout_name}' 图签修复完成。")



        # —————— 清理 ——————

        try:
            set_space_mode(1)                     

            wait_quiescent()
            ensure_layer("测试辅助")

        except:
        
            pass   

        try:
            set_space_mode(0)                     
            ensure_layer("dy_zhuanyong")
            wait_quiescent()


        except:
        
            pass   

        return True


    except Exception as e:
        sys_logger.info(f"❌ repair_layout_direct 执行出错: {e}")
        return False




#&&% 线宽修复


def smart_repair_frame_polyline_widths_m(verbose=True):
    """
    【函数编号】: FIX-002-Smart (V9.2 - 尺寸优先分流版)
    【修复逻辑】: 
        1. 不再盲目根据图层/空间判断匹配模式。
        2. 基于几何尺寸(短边阈值)自动判断是 "1:1系列" 还是 "1:100系列"。
        3. 解决 287x390 这种 1:1 尺寸被误判为大线宽的问题。
    """
    C.li()

    # --- 核心数据源 (原始数据均为 1:100 基准) ---
    # 格式: (Width, Height, LineWidth)
    # 例如: (42000, 29700, 100) -> 代表 A3 1:100, 线宽 100
    LB_dayingkuang = [
        (118900, 84100, 100),  (178350, 126150, 150),   (59450, 42050, 50),     (29725, 21025, 25), 
        (133800, 84100, 100),  (200700, 126150, 150),   (66900, 42050, 50),     (33450, 21025, 25), 
        (148600, 84100, 100),  (222900, 126150, 150),   (74300, 42050, 50),     (37150, 21025, 25), 
        (84100,  59400, 100),  (126150, 89100,  150),   (42050, 29700, 50),     (21025, 14850, 25), 
        (105100, 59400, 100),  (157650, 89100,  150),   (52550, 29700, 50),     (26275, 14850, 25), 
        (126100, 59400, 100),  (189150, 89100,  150),   (63050, 29700, 50),     (31525, 14850, 25), 
        (147100, 59400, 100),  (220650, 89100,  150),   (73550, 29700, 50),     (36775, 14850, 25), 
        (59400,  42000, 100),  (89100,  63000,  150),   (29700, 21000, 50),     (14850, 10500, 25), 
        (74300,  42000, 100),  (111450, 63000,  150),   (37150, 21000, 50),     (18575, 10500, 25), 
        (89100,  42000, 100),  (133650, 63000,  150),   (44550, 21000, 50),     (22275, 10500, 25), 
        (104100, 42000, 100),  (156150, 63000,  150),   (52050, 21000, 50),     (26025, 10500, 25), 
        (42000,  29700, 100),  (63000,  44550,  150),   (21000, 14850, 50),     (10500, 7425,  25), 
    ]

    # 定义处理图层 (不再区分 model/paper 参数，内部自动判断)
    layer_tasks = ["tuqian_baobu_m"]
    
    total_fixed = 0
    sys_logger.info(f"🛠️ [智能修复] 启动尺寸自适应分析...")

    def calculate_width_for_obj(obj):
        try:
            # 1. 获取物理尺寸
            min_pt, max_pt = obj.GetBoundingBox()
            w = abs(max_pt[0] - min_pt[0])
            h = abs(max_pt[1] - min_pt[1])
            short_side = min(w, h)
            current_dim = (max(w, h), min(w, h))

            # =========================================================
            # ✨ 核心分流：根据物理尺寸决定是 "1:1系列" 还是 "1:100系列"
            # =========================================================
            # 阈值设为 2000mm (A0长边是1189，放大100倍是118900)
            # 凡是短边小于 2000 的，一律视为 1:1 系列 (包括布局和微型模型)
            is_small_series = (short_side < 2000.0)

            # 准备匹配参数
            if is_small_series:
                # [1:1 系列]
                # 数据库缩放因子: 1/100 (将库里的 42000 变成 420)
                # 线宽缩放因子: 1/100 (将库里的 100 变成 1.0)
                # 容差: 0.5 (允许 287 vs 287.5)
                # 比例值映射: 分析出 100 -> 期望 1.0
                db_divisor = 100.0
                match_tol = 0.5 
                mode_name = "Small(1:1)"
            else:
                # [1:100 系列]
                # 数据库缩放因子: 1.0 (保持原样)
                # 线宽缩放因子: 1.0 (保持原样)
                # 容差: 10.0 (允许 28700 vs 28710)
                # 比例值映射: 分析出 100 -> 期望 100
                db_divisor = 1.0
                match_tol = 10.0
                mode_name = "Large(1:100)"

            # 2. 获取比例信息 (辅助仲裁)
            # generate_name... 只能识别 1:100, 1:50 等，无法识别 1:1
            # 所以这里传入 min_lm 帮助其过滤干扰
            ratio_info = generate_name_and_ratio_from_com(
                obj, tol=match_tol, min_lm=(1000 if not is_small_series else 10)
            )
            # 默认比例 100 (即 1:100)
            detected_ratio = ratio_info[1] if (ratio_info and len(ratio_info) >= 2) else 100.0

            # 3. 准备匹配候选
            scale_series = [1.0, 1.1, 1.2]
            candidates = [] # 存 (线宽, 距离)

            for ref_w, ref_h, ref_lw in LB_dayingkuang:
                # 动态变换标准库数据
                t_w = ref_w / db_divisor
                t_h = ref_h / db_divisor
                t_lw = float(ref_lw) / db_divisor # 关键：线宽也随之缩小

                # 检查系列 (1.0, 1.1, 1.2)
                for factor in scale_series:
                    scaled_w = t_w * factor
                    scaled_h = t_h * factor
                    
                    # 检查是否命中尺寸
                    if abs(current_dim[0] - scaled_w) <= match_tol and \
                       abs(current_dim[1] - scaled_h) <= match_tol:
                        candidates.append(t_lw)
                        break 
            
            # 4. 决策逻辑
            final_w = None

            if len(candidates) == 1:
                # A. 唯一精确匹配
                final_w = candidates[0]
                
            elif len(candidates) > 1:
                # B. 多个匹配 (例如 1:100 和 1:50 尺寸重叠时)，使用 detected_ratio 仲裁
                # 关键映射：如果 detected_ratio 是 100
                # Large模式 -> 期望 100
                # Small模式 -> 期望 1.0 (100 / 100)
                expected_lw = detected_ratio / db_divisor
                
                # 优先找完全一致的
                for cand in candidates:
                    if abs(cand - expected_lw) < 0.001:
                        final_w = cand
                        break
                # 找不到就取第一个
                if final_w is None:
                    final_w = candidates[0] 
            
            else:
                # C. 无精确匹配 -> 找最近的 (Fallback)
                # 比如 287x390 这种非标内框，它不等于库里的 297x420
                # 但我们需要在 Small模式下，找到最接近的 Small库线宽
                best_dist = float('inf')
                # 默认兜底：Small模式 1.0, Large模式 100.0
                best_lw = 1.0 if is_small_series else 100.0

                for ref_w, ref_h, ref_lw in LB_dayingkuang:
                    # 同样要缩小库数据来对比
                    t_w = ref_w / db_divisor
                    t_h = ref_h / db_divisor
                    t_lw = float(ref_lw) / db_divisor

                    # 欧氏距离
                    dist = ((current_dim[0] - t_w)**2 + (current_dim[1] - t_h)**2) ** 0.5
                    if dist < best_dist:
                        best_dist = dist
                        best_lw = t_lw
                
                final_w = best_lw

            # 调试日志 (可选)
            # if verbose:
            #    sys_logger.info(f"OBJ {int(current_dim[0])}x{int(current_dim[1])} [{mode_name}] -> 匹配线宽 {final_w}")

            return final_w

        except Exception as e:
            # 终极兜底：根据尺寸给一个安全值
            try:
                min_pt, max_pt = obj.GetBoundingBox()
                if abs(max_pt[0] - min_pt[0]) < 2000:
                    return 1.0 # 1:1 安全值
                else:
                    return 100.0 # 1:100 安全值
            except:
                return 1.0

    # --- 执行循环 ---
    for layer_name in layer_tasks:
        try:
            objs = stc(layer_name)
            if not objs: continue
            
            count = 0
            for obj in objs:
                if "AcDbPolyline" not in get_attr(obj, "ObjectName"): continue
                
                # 计算
                target_width = calculate_width_for_obj(obj)
                
                # 应用 (减少不必要的写入)
                if abs(obj.ConstantWidth - target_width) > 0.0001:
                    obj.ConstantWidth = target_width
                    count += 1
            
            if count > 0:
                sys_logger.info(f"  - [修复] '{layer_name}': 更新 {count} 个")
                total_fixed += count

        except Exception as e:
            sys_logger.error(f"处理图层 {layer_name} 失败: {e}")

    if total_fixed > 0:
        try: C.doc.Regen(1)
        except: pass
        sys_logger.info(f"✅ 智能线宽修复完成，共更新 {total_fixed} 条。")
        return True
    return False

def smart_repair_frame_polyline_widths_p(verbose=True):
    """
    【函数编号】: FIX-002-Smart (V9.2 - 尺寸优先分流版)
    【修复逻辑】: 
        1. 不再盲目根据图层/空间判断匹配模式。
        2. 基于几何尺寸(短边阈值)自动判断是 "1:1系列" 还是 "1:100系列"。
        3. 解决 287x390 这种 1:1 尺寸被误判为大线宽的问题。
    """
    C.li()

    # --- 核心数据源 (原始数据均为 1:100 基准) ---
    # 格式: (Width, Height, LineWidth)
    # 例如: (42000, 29700, 100) -> 代表 A3 1:100, 线宽 100
    LB_dayingkuang = [
        (118900, 84100, 100),  (178350, 126150, 150),   (59450, 42050, 50),     (29725, 21025, 25), 
        (133800, 84100, 100),  (200700, 126150, 150),   (66900, 42050, 50),     (33450, 21025, 25), 
        (148600, 84100, 100),  (222900, 126150, 150),   (74300, 42050, 50),     (37150, 21025, 25), 
        (84100,  59400, 100),  (126150, 89100,  150),   (42050, 29700, 50),     (21025, 14850, 25), 
        (105100, 59400, 100),  (157650, 89100,  150),   (52550, 29700, 50),     (26275, 14850, 25), 
        (126100, 59400, 100),  (189150, 89100,  150),   (63050, 29700, 50),     (31525, 14850, 25), 
        (147100, 59400, 100),  (220650, 89100,  150),   (73550, 29700, 50),     (36775, 14850, 25), 
        (59400,  42000, 100),  (89100,  63000,  150),   (29700, 21000, 50),     (14850, 10500, 25), 
        (74300,  42000, 100),  (111450, 63000,  150),   (37150, 21000, 50),     (18575, 10500, 25), 
        (89100,  42000, 100),  (133650, 63000,  150),   (44550, 21000, 50),     (22275, 10500, 25), 
        (104100, 42000, 100),  (156150, 63000,  150),   (52050, 21000, 50),     (26025, 10500, 25), 
        (42000,  29700, 100),  (63000,  44550,  150),   (21000, 14850, 50),     (10500, 7425,  25), 
    ]

    # 定义处理图层 (不再区分 model/paper 参数，内部自动判断)
    layer_tasks = ["tuqian_baobu"]
    
    total_fixed = 0
    sys_logger.info(f"🛠️ [智能修复] 启动尺寸自适应分析...")

    def calculate_width_for_obj(obj):
        try:
            # 1. 获取物理尺寸
            min_pt, max_pt = obj.GetBoundingBox()
            w = abs(max_pt[0] - min_pt[0])
            h = abs(max_pt[1] - min_pt[1])
            short_side = min(w, h)
            current_dim = (max(w, h), min(w, h))

            # =========================================================
            # ✨ 核心分流：根据物理尺寸决定是 "1:1系列" 还是 "1:100系列"
            # =========================================================
            # 阈值设为 2000mm (A0长边是1189，放大100倍是118900)
            # 凡是短边小于 2000 的，一律视为 1:1 系列 (包括布局和微型模型)
            is_small_series = (short_side < 2000.0)

            # 准备匹配参数
            if is_small_series:
                # [1:1 系列]
                # 数据库缩放因子: 1/100 (将库里的 42000 变成 420)
                # 线宽缩放因子: 1/100 (将库里的 100 变成 1.0)
                # 容差: 0.5 (允许 287 vs 287.5)
                # 比例值映射: 分析出 100 -> 期望 1.0
                db_divisor = 100.0
                match_tol = 0.5 
                mode_name = "Small(1:1)"
            else:
                # [1:100 系列]
                # 数据库缩放因子: 1.0 (保持原样)
                # 线宽缩放因子: 1.0 (保持原样)
                # 容差: 10.0 (允许 28700 vs 28710)
                # 比例值映射: 分析出 100 -> 期望 100
                db_divisor = 1.0
                match_tol = 10.0
                mode_name = "Large(1:100)"

            # 2. 获取比例信息 (辅助仲裁)
            # generate_name... 只能识别 1:100, 1:50 等，无法识别 1:1
            # 所以这里传入 min_lm 帮助其过滤干扰
            ratio_info = generate_name_and_ratio_from_com(
                obj, tol=match_tol, min_lm=(1000 if not is_small_series else 10)
            )
            # 默认比例 100 (即 1:100)
            detected_ratio = ratio_info[1] if (ratio_info and len(ratio_info) >= 2) else 100.0

            # 3. 准备匹配候选
            scale_series = [1.0, 1.1, 1.2]
            candidates = [] # 存 (线宽, 距离)

            for ref_w, ref_h, ref_lw in LB_dayingkuang:
                # 动态变换标准库数据
                t_w = ref_w / db_divisor
                t_h = ref_h / db_divisor
                t_lw = float(ref_lw) / db_divisor # 关键：线宽也随之缩小

                # 检查系列 (1.0, 1.1, 1.2)
                for factor in scale_series:
                    scaled_w = t_w * factor
                    scaled_h = t_h * factor
                    
                    # 检查是否命中尺寸
                    if abs(current_dim[0] - scaled_w) <= match_tol and \
                       abs(current_dim[1] - scaled_h) <= match_tol:
                        candidates.append(t_lw)
                        break 
            
            # 4. 决策逻辑
            final_w = None

            if len(candidates) == 1:
                # A. 唯一精确匹配
                final_w = candidates[0]
                
            elif len(candidates) > 1:
                # B. 多个匹配 (例如 1:100 和 1:50 尺寸重叠时)，使用 detected_ratio 仲裁
                # 关键映射：如果 detected_ratio 是 100
                # Large模式 -> 期望 100
                # Small模式 -> 期望 1.0 (100 / 100)
                expected_lw = detected_ratio / db_divisor
                
                # 优先找完全一致的
                for cand in candidates:
                    if abs(cand - expected_lw) < 0.001:
                        final_w = cand
                        break
                # 找不到就取第一个
                if final_w is None:
                    final_w = candidates[0] 
            
            else:
                # C. 无精确匹配 -> 找最近的 (Fallback)
                # 比如 287x390 这种非标内框，它不等于库里的 297x420
                # 但我们需要在 Small模式下，找到最接近的 Small库线宽
                best_dist = float('inf')
                # 默认兜底：Small模式 1.0, Large模式 100.0
                best_lw = 1.0 if is_small_series else 100.0

                for ref_w, ref_h, ref_lw in LB_dayingkuang:
                    # 同样要缩小库数据来对比
                    t_w = ref_w / db_divisor
                    t_h = ref_h / db_divisor
                    t_lw = float(ref_lw) / db_divisor

                    # 欧氏距离
                    dist = ((current_dim[0] - t_w)**2 + (current_dim[1] - t_h)**2) ** 0.5
                    if dist < best_dist:
                        best_dist = dist
                        best_lw = t_lw
                
                final_w = best_lw

            # 调试日志 (可选)
            # if verbose:
            #    sys_logger.info(f"OBJ {int(current_dim[0])}x{int(current_dim[1])} [{mode_name}] -> 匹配线宽 {final_w}")

            return final_w

        except Exception as e:
            # 终极兜底：根据尺寸给一个安全值
            try:
                min_pt, max_pt = obj.GetBoundingBox()
                if abs(max_pt[0] - min_pt[0]) < 2000:
                    return 1.0 # 1:1 安全值
                else:
                    return 100.0 # 1:100 安全值
            except:
                return 1.0

    # --- 执行循环 ---
    for layer_name in layer_tasks:
        try:
            objs = stc(layer_name)
            if not objs: continue
            
            count = 0
            for obj in objs:
                if "AcDbPolyline" not in get_attr(obj, "ObjectName"): continue
                
                # 计算
                target_width = calculate_width_for_obj(obj)
                
                # 应用 (减少不必要的写入)
                if abs(obj.ConstantWidth - target_width) > 0.0001:
                    obj.ConstantWidth = target_width
                    count += 1
            
            if count > 0:
                sys_logger.info(f"  - [修复] '{layer_name}': 更新 {count} 个")
                total_fixed += count

        except Exception as e:
            sys_logger.error(f"处理图层 {layer_name} 失败: {e}")

    if total_fixed > 0:
        try: C.doc.Regen(1)
        except: pass
        sys_logger.info(f"✅ 智能线宽修复完成，共更新 {total_fixed} 条。")
        return True
    return False


#&&% 从模型空间剪切到图纸空间


def cut_model_to_paper_and_switch(target_layout_name="平面分割图"):
    """
    【函数编号】: MOVE-SPACE-V3.2 (自信版)
    【改进】: 
        修正了 CopyObjects 返回值计数逻辑。
        不再因为 CAD 内部对象合并导致的数量差异而报假警。
    """
    # ... (头部引用保持不变) ...
    try:
        from licad import C 
    except ImportError:
        import sys, os
        sys.path.append(os.path.dirname(os.path.dirname(__file__)))
        from system.licad import C

    from win32com.client import VARIANT
    import pythoncom

    doc = C.doc
    
    def log(msg): sys_logger.info(f"🚚 [MoveToPaper] {msg}")
    
    log(f"准备将对象从 模型空间 -> 布局 [{target_layout_name}] ...")

    # 1. 获取目标布局
    try:
        layout = doc.Layouts.Item(target_layout_name)
        target_block = layout.Block 
    except Exception as e:
        log(f"❌ 找不到目标布局 '{target_layout_name}': {e}")
        return False

    # 2. 筛选目标对象
    target_layers = ["tuqian_baobu", "tuqian_neibu_pl", "dy_quyu_H", "dy_quyu", ]
    target_layers_str = ",".join(target_layers)
    
    objs_to_move = []
    
    try:
        ss_name = "Move_SS"
        try: doc.SelectionSets.Item(ss_name).Delete()
        except: pass
        ss = doc.SelectionSets.Add(ss_name)
        
        p_filter_type = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, [8, 410])
        p_filter_data = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, [target_layers_str, "Model"])
        
        ss.Select(5, None, None, p_filter_type, p_filter_data)
        
        if ss.Count == 0:
            log("⚠️ 模型空间为空，无需搬运。")
            ss.Delete()
            return True
            
        for i in range(ss.Count):
            objs_to_move.append(ss.Item(i))
            
        log(f"🔍 锁定 {len(objs_to_move)} 个源对象...")
        ss.Delete() 
        
    except Exception as e:
        log(f"❌ 筛选对象失败: {e}")
        return False

    # 3. 执行 CopyObjects
    try:
        variant_objs = VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, objs_to_move)
        
        # 执行内存克隆
        cloned_objs = doc.CopyObjects(variant_objs, target_block)
        
        # 🟢【修正逻辑】智能判断
        created_count = len(cloned_objs) if cloned_objs else 0
        original_count = len(objs_to_move)
        
        if created_count > 0:
            if created_count == original_count:
                log(f"✅ 完美搬运: {created_count} 个对象。")
            else:
                # 即使数量不等，只要有生成，且没有抛出异常，通常也是成功的
                # 这属于“对象合并”现象
                log(f"✅ 搬运成功 (对象合并): 源 {original_count} -> 新 {created_count}")
        else:
            log(f"❌ 搬运失败: 未生成任何新对象。")
            return False
        
    except Exception as e:
        log(f"❌ CopyObjects 内存操作失败: {e}")
        return _fallback_copy_method(doc, target_layout_name, target_layers_str)

    # 4. 删除源对象
    try:
        log("🗑️ 清理模型空间源对象...")
        for obj in objs_to_move:
            try: obj.Delete()
            except: pass
        try: doc.Regen(1)
        except: pass
    except Exception as e:
        log(f"⚠️ 源对象清理轻微异常: {e}")

    # 5. 切换视角
    try:
        doc.ActiveLayout = layout
        doc.MSpace = False
    except: pass
    
    return True

# ------------------------------------------------------------------
# 保底方案：如果 COM 还是不行，使用 COPYBASE (比 CUTCLIP 稳)
# ------------------------------------------------------------------
def _fallback_copy_method(doc, layout_name, layer_filter):
    print("🔄 [MoveToPaper] 启动保底方案: SendCommand (COPYBASE)...")
    try:
        doc.ActiveLayout = doc.Layouts.Item("Model")
        # 使用 COPYBASE (带基点复制)，基点设为 0,0,0，防止粘贴错位
        cmd = f'(setq ss (ssget "_X" \'((8 . "{layer_filter}") (410 . "Model")))) '
        cmd += '(if ss (command "_.COPYBASE" "0,0,0" ss "")) '
        cmd += '(setq ss nil) '
        doc.SendCommand(cmd + "\n")
        
        # 切换布局
        doc.ActiveLayout = doc.Layouts.Item(layout_name)
        doc.MSpace = False
        
        # 原位粘贴
        doc.SendCommand("_.PASTECLIP 0,0,0 ")
        
        # 回去删除 (分开做，防止粘贴失败却删了)
        doc.ActiveLayout = doc.Layouts.Item("Model")
        cmd_del = f'(setq ss (ssget "_X" \'((8 . "{layer_filter}") (410 . "Model")))) '
        cmd_del += '(if ss (command "_.ERASE" ss "")) '
        doc.SendCommand(cmd_del + "\n")
        
        # 最后切回布局
        doc.ActiveLayout = doc.Layouts.Item(layout_name)
        print("✅ 保底方案执行完毕。")
        return True
    except Exception as e:
        sys_logger.info(f"❌ 保底方案也失败了: {e}")
        return False




#&&% 从模型空间屏幕选择对象复制到图纸空间


def cut_screen_selection_to_paper(target_layout_name="布局1"):
    """
    【功能】: 屏幕选择模型空间对象 -> 原坐标复制到指定布局空间
    【依赖】: 需要全局存在 li(), doc, vtobj(), pmxz(), switch_to_layout()
    实际是剪切

    """
    from  CAD_file_operations   import  switch_to_layout

    # 1. 连接 AutoCAD
    try: li()
    except: pass

    # 2. 获取目标布局的 Block 容器 (图纸空间容器)
    target_block = None
    try:
        # 遍历查找，忽略大小写
        for i in range(doc.Layouts.Count):
            l = doc.Layouts.Item(i)
            if l.Name.lower() == target_layout_name.lower():
                target_block = l.Block
                break
        
        if not target_block:
            sys_logger.info(f"❌ 找不到布局 '{target_layout_name}'")
            return
    except Exception as e:
        sys_logger.info(f"❌ 获取目标布局失败: {e}")
        return

    # 3. 调用屏幕选择 (pmxz)
    sys_logger.info(f"\n🚀 请在屏幕上选择要复制到 '{target_layout_name}' 的对象 (按回车结束)...")
    
    # --- 这里的 pmxz() 应该是你封装好的 SelectOnScreen 函数 ---
    # 假设它返回的是一个包含 COM 对象的列表 (list) 或元组 (tuple)
    try:
        pm = pmxz() 
    except NameError:
        print("❌ 错误: 未找到 pmxz() 函数。")
        return
    except Exception as e:
        sys_logger.info(f"❌ 选择过程出错: {e}")
        return

    # 4. 验证选择结果
    if not pm or len(pm) == 0:
        print("⚠️ 未选择任何对象，操作取消。")
        return

    sys_logger.info(f"选中 {len(pm)} 个对象，准备复制...")

    # 5. 执行复制
    try:
        # ---------------------------------------------------------
        # 【数据转换】
        # 确保 pm 是列表格式，然后转为 COM 变体数组
        # ---------------------------------------------------------
        py_obj_list = list(pm) # 强转列表，防止 pmxz 返回的是元组
        com_obj_array = vtobj(py_obj_list)
        
        # ---------------------------------------------------------
        # 【核心复制】
        # CopyObjects 会保持对象的几何属性（包括坐标 X,Y,Z）不变
        # 只是将其“宿主”从 模型空间 Block 变成了 布局空间 Block
        # ---------------------------------------------------------
        doc.CopyObjects(com_obj_array, target_block)
        print("✅ 复制成功！(坐标已保持)")
        
        # 6. 切换视图到目标布局查看结果
        try:
            if 'switch_to_layout' in globals():
                switch_to_layout(target_layout_name)
            else:
                doc.ActiveLayout = doc.Layouts.Item(target_layout_name)
                doc.SetVariable("TILEMODE", 0) # 确保进入图纸模式
            
            # 稍微刷新一下显示
            doc.Regen(1)
            # doc.Application.ZoomExtents() # 可选：是否缩放全图
            
        except Exception as e:
            sys_logger.info(f"⚠️ 切换布局时遇到小问题: {e}")

    except Exception as e:
        sys_logger.info(f"❌ 复制对象失败: {e}")
        import traceback
        traceback.print_exc()





#&&% 从图纸空间复制打印区域到模型空间

def copy_layout_polylines_to_model(layout_name, polylines_list):
    """
    【函数编号】: MAP-COPY-LAYOUT-004 (布局切换+列表直传版)
    【功能】: 
        1. 切换到指定布局 (确保上下文正确)。
        2. 将输入的 polylines_list (来自布局的对象) 复制到模型空间。
        3. 将新生成的对象归档到 'dy_paper_to_model' 图层。
    【参数】:
        layout_name: str - 来源布局名称 (用于 switch_to_layout)
        polylines_list: list - 待复制的 COM 对象列表
    """

    from  CAD_file_operations   import  switch_to_layout
    def log(msg):
        sys_logger.info(f"[CopyLayout->Model] {msg}")

    # 1. 基础校验
    if not polylines_list:
        log("⚠️ 输入列表为空，跳过复制。")
        return []

    # 2. 切换到源布局 (确保 COM 上下文稳定)
    # 假设 switch_to_layout 已在全局定义
    if not switch_to_layout(layout_name):
        log(f"❌ 无法切换到布局: {layout_name}，停止复制。")
        return []
    
    # 确保不在视口内 (MSpace=False)
    try:
        li()
        doc = globals().get('doc')
        doc.MSpace = False
        model_space = doc.ModelSpace
    except Exception as e:
        log(f"❌ 获取环境对象失败: {e}")
        return []

    # 3. 准备目标图层
    target_layer = "dy_paper_to_model"
    try: ensure_layer(target_layer)
    except: pass

    # ================= Step 1: 执行复制 (快照法) =================
    new_model_objects = []
    
    try:
        log(f"正在将 {len(polylines_list)} 个对象从 '{layout_name}' 复制到模型空间...")
        
        

        # A. 记录快照
        count_before = model_space.Count

        # B. 构造参数
        vt_objects = win32com.client.VARIANT(
            pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, 
            polylines_list
        )
        
        # C. 执行复制
        doc.CopyObjects(vt_objects, model_space)
        
        # D. 计算增量
        count_after = model_space.Count
        diff = count_after - count_before
        
        if diff > 0:
            # E. 捕获新对象
            for i in range(count_before, count_after):
                try:
                    obj = model_space.Item(i)
                    new_model_objects.append(obj)
                except: pass
            
            log(f"✅ 成功捕获 {len(new_model_objects)} 个新对象。")
            
            # ================= Step 2: 后处理 (归档图层) =================
            if new_model_objects:
                count_moved = 0
                for obj in new_model_objects:
                    try:
                        obj.Layer = target_layer
                        obj.Color = 256 # ByLayer
                        count_moved += 1
                    except: pass
                log(f"已归档至图层 '{target_layer}': {count_moved} 个")

        else:
            log("⚠️ 警告: 复制指令执行完成，但模型空间对象数量未增加。")

        return new_model_objects

    except Exception as e:
        log(f"❌ 复制流程异常: {e}")
        return []



#&&% 清空图纸空间

def clear_layout_objects(layout_name="布局1", max_retries=5):
    """
    【函数编号】: MAP-CLEAR-LAYOUT-003 (多轮重试版)
    【功能】: 彻底清空指定布局空间内的所有对象（保留布局本身和背景视口）。
    【机制】: 采用多轮清理机制，解决 'Call Rejected' 和残留对象问题。
    """
    def log(msg):
        sys_logger.info(f"[ClearLayout] {msg}")

    # 1. 环境准备
    try: li() # 你的初始化函数
    except: pass
    
    # 确保 doc 存在 (兼容你的上下文)
    try: 
        doc = win32com.client.GetActiveObject("AutoCAD.Application").ActiveDocument
    except:
        log("❌ 无法获取 ActiveDocument")
        return False

    # 2. 切换到目标布局
    try:
        if doc.ActiveLayout.Name != layout_name:
            layouts = doc.Layouts
            try:
                target_layout = layouts.Item(layout_name)
                doc.ActiveLayout = target_layout
                # 切换后稍微等待，防止 AutoCAD 还没渲染完
                time.sleep(0.2) 
            except Exception:
                log(f"❌ 找不到布局: {layout_name}")
                return False
    except Exception as e:
        log(f"❌ 切换布局失败: {e}")
        return False

    # 3. 多轮清理循环
    paper_space = doc.PaperSpace
    
    for round_idx in range(1, max_retries + 1):
        # 获取当前对象总数
        current_count = paper_space.Count
        
        # === 终止条件 1: 布局已空 (或只剩1个删不掉的基础视口) ===
        if current_count == 0:
            log(f"✅ (第{round_idx}轮) 布局已完全清空。")
            return True
        
        log(f"🔄 第 {round_idx}/{max_retries} 轮清理... (当前对象数: {current_count})")
        
        deleted_in_this_round = 0
        error_rpc_count = 0 # 记录“呼叫被拒绝”的次数

        # 倒序遍历
        for i in range(current_count - 1, -1, -1):
            try:
                item = paper_space.Item(i)
                item.Delete()
                deleted_in_this_round += 1
            except Exception as e:
                # 检查是否为 RPC 忙碌错误 (Call Rejected)
                # 错误码 -2147418111 对应 0x80010001 (RPC_E_CALL_REJECTED)
                if "-2147418111" in str(e):
                    error_rpc_count += 1
                    time.sleep(0.05) # 稍微让渡一下CPU
                else:
                    # 其他错误通常是“无法删除的对象”（如基础视口），这是正常的
                    pass
        
        # === 策略调整 ===
        if error_rpc_count > 0:
            log(f"   ⚠️ 本轮发生 {error_rpc_count} 次忙碌拒绝，将在稍后重试。")
            time.sleep(0.5) # 既然忙，就多歇会儿

        # === 终止条件 2: 没有任何进展 ===
        # 如果本轮尝试了，但没有删除任何东西，说明剩下的都是删不掉的（如基础视口）
        if deleted_in_this_round == 0:
            # 通常布局至少剩下一个对象(Layout Viewport)是删不掉的，这是成功的标志
            log(f"✅ 清理结束。剩余 {current_count} 个系统保留对象(通常是视口)。")
            try: doc.Regen(1) 
            except: pass
            return True

    # 4. 最终刷新
    try: doc.Regen(1) 
    except: pass
    
    final_count = paper_space.Count
    log(f"⚠️ 达到最大重试次数。仍有 {final_count} 个对象残留。")
    return True


#&&% 清除无实例块

def clean_unused_blocks_global_scan(verbose=True):
    """
    【函数编号】: CLEAN-005 (精准概念版)
    """
    try:
        doc = C.doc
    except:
        return

    # ==========================================
    # 1. 从“实体”中提取“名字” (白名单)
    # ==========================================
    active_names_set = set()
    
    # lb 是 [实体对象, 实体对象, ...]
    lb = select_kuai() 
    
    if lb:
        for blk_ref in lb: # blk_ref 是 引用(Reference)
            try:
                # 提取名字字符串
                n = get_attr(blk_ref, "Name")
                if n: active_names_set.add(n)
                
                # 动态块还要提取 EffectiveName
                eff = get_attr(blk_ref, "EffectiveName")
                if eff: active_names_set.add(eff)
            except:
                pass
    
    if verbose:
        sys_logger.info(f"📋 [保护名单] 发现 {len(active_names_set)} 个正在使用的块名。")

    # ==========================================
    # 2. 遍历“定义”，对比“名字”
    # ==========================================
    deleted_count = 0
    blocks_collection = doc.Blocks
    count = blocks_collection.Count
    
    # 先收集要删的名字字符串，不要一边遍历一边删，容易乱索引
    names_to_delete = []

    for i in range(count):
        try:
            # blk_def 是 定义对象 (Definition)
            blk_def = blocks_collection.Item(i)
            
            # 获取定义的字符串名字
            def_name_str = blk_def.Name
            
            # 1. 跳过系统块 (*Model_Space 等)
            if def_name_str.startswith("*"):
                continue
            
            # 2. 核心比对：名字是否在白名单里？
            if def_name_str not in active_names_set:
                names_to_delete.append(def_name_str)
                
        except Exception:
            pass

    # ==========================================
    # 3. 执行删除
    # ==========================================
    if verbose and names_to_delete:
        sys_logger.info(f"🔍 [发现] {len(names_to_delete)} 个无实例的定义 (含 '{names_to_delete[0]}' 等)。")

    for target_name in names_to_delete:
        try:
            # 重新获取定义对象来删除
            target_def = blocks_collection.Item(target_name)
            target_def.Delete()
            deleted_count += 1
            if verbose:
                sys_logger.info(f"🗑️ [已清理] 块定义: {target_name}")
        except:
            # 删不掉说明可能有嵌套引用
            pass

    if verbose:
        sys_logger.info(f"✅ 清理结束，销毁 {deleted_count} 个定义。")

#&&&% ===（四）重建字典 ===

#&&% 统一重建字典

@timeit
def smart_rebuild_print_info(
    layout_name=None,
    operate_target="Model",
    select_config=None,
    use_cache=False,
    verbose=None,
    **kwargs
):
    """
    【函数编号】: MAP-UNIFIED-001 (V2.1 - 终极简化版)
    【功能】: 
        集成缓存与标准化重建流程。
        简化分支逻辑：
          - 所有 "Model" 操作 -> rebuild_print_area_title_mapping
          - 所有 "Layout" 操作 -> rebuild_print_area_title_mapping_paper
    """
    
    # 1. 日志与环境
    lvl = int(verbose) if verbose is not None else globals().get('GLOBAL_DEFAULT_VERBOSE', 1)
    def log(msg):
        if lvl >= 1: sys_logger.info(f"🧠 [SmartCTQ] {msg}")

    try:
        doc = C.doc
        doc_base = os.path.splitext(doc.Name)[0]
    except: 
        doc_base = "Unknown"

    # --- 1.5 智能参数分配逻辑 ---
    # 如果用户没有指定配置 (为 None)，则根据环境自动设定默认值
    if select_config is None:
        if operate_target == "Model":
            # 模型空间：默认使用 0 (常规/自适应模式)
            # 无论是纯模型还是混合借道，通常 0 都能较好处理
            select_config = 0
        else:
            # 布局空间：通常是 1:1，使用 1 (高精模式) 或 0 均可
            # 根据经验，布局空间 0 (宽容模式) 其实更稳健
            select_config = 0 


    # 2. 缓存 Key 计算
    mode_tag = "Unknown"
    if operate_target == "Layout":
        target_name = layout_name if layout_name else "UnknownLayout"
        mode_tag = f"Layout_{target_name}"
    elif operate_target == "Model":
        if layout_name: mode_tag = f"Model_Hybrid_{layout_name}"
        else: mode_tag = "Model_Pure"
    
    cache_key = f"{doc_base}_{mode_tag}_CTQ"

    # ================= 分支 A: 尝试加载缓存 =================
    if use_cache:
        log(f"尝试加载缓存: {cache_key}.json ...")
        try:
            c_polys, c_blocks, c_mapping = load_ctq(file_name=cache_key, verbose=lvl)
            if c_mapping and len(c_mapping) > 0:
                log(f"✅ 缓存命中! 成功恢复 {len(c_mapping)} 组关系。")
                return c_polys, c_blocks, c_mapping
            else:
                log("⚠️ 缓存无效或不存在，转为重新计算...")
        except Exception as e:
            log(f"⚠️ 读取缓存出错: {e}，转为重新计算。")

    # ================= 分支 B: 重新计算 =================
    if not use_cache:
        log("⚡ 模式: 强制重新计算")

    result = ([], [], {})
    
    try:
        # ✅ 核心逻辑分支简化
        
        if operate_target == "Layout":
            # 🟡 布局空间
            # 必须传入目标布局名
            target_layout = layout_name if layout_name else "布局1"
            
            result = rebuild_print_area_title_mapping_paper(
                layout_target=target_layout,
                select_config=select_config, 
                verbose=lvl,
                **kwargs
            )
            
        elif operate_target == "Model":
            # 🟢 模型空间 (统一入口)
            # 无论是纯模型，还是混合模式，都走这一个函数
            # 函数内部会自动 set_space_mode(1) 切换到模型空间
            
            result = rebuild_print_area_title_mapping(
                select_config=select_config, 
                verbose=lvl,
                **kwargs
            )
            
        else:
            log(f"❌ 未知操作目标: {operate_target}")
            return [], [], {}
                
    except Exception as e:
        log(f"❌ 计算过程致命异常: {e}")
        import traceback
        traceback.print_exc()
        return [], [], {}

    # 3. 更新缓存
    polys, blocks, mapping = result
    if polys:
        log(f"💾 更新缓存: {cache_key}.json")
        try:
            save_ctq(result, file_name=cache_key)
        except Exception as e:
            log(f"❌ 缓存保存失败: {e}")
    else:
        log("⚠️ 扫描结果为空，跳过保存缓存。")

    return result




#&&% 重建字典_模型空间


@timeit
def rebuild_print_area_title_mapping(
    core_layer: str = "dy_quyu_H_m",             
    final_poly_layer: str = "tuqian_neibu_pl_m", 
    select_config = 0,  
 
    verbose=None,
    **kwargs
):
    """
    【函数编号】: MAP-MODEL-FINAL (V40 - 简化判定版)
    【功能】: 
        1. 强制切换至模型空间。
        2. 调用 select_maxrect_polylines_1 获取标准化、排序后的打印区域。
        3. 搜集图签块。
        4. 使用“中心点包含”逻辑建立 区域->图签 的映射关系。
    
    【参数 select_config】:
        - 0: 常规模式 (Precision=False)
        - 1: 高精模式 (Precision=True)
    """
    from  CAD_file_operations   import set_space_mode

    doc = C.doc
    lvl = int(verbose) if verbose is not None else 1
    def log(msg): 
        if lvl >= 1: 
            sys_logger.info(f"🔨 [Model-Map] {msg}")

    # ============================================================
    # 🚩 1. 环境准备与区域获取
    # ============================================================
    # 强制进入模型空间
    try:
        set_space_mode(1)
    except Exception as e:
        log(f"⚠️ 切换空间失败: {e}")
        return [], [], {}

    log(f"🚀 Step1: 获取打印区域 (Config={select_config})...")
    
    # 映射配置参数
    is_precision = True if select_config == 1 else False

    # 调用核心选择函数
    # 注意：直接传入 final_poly_layer，确保重绘后的线在目标图层
    polys_sorted, _ = select_maxrect_polylines_1(
        layer_name = final_poly_layer,  # 目标图层 "tuqian_neibu_pl_m"
        precision_mode = is_precision,  # 0->False, 1->True
        width = 0.0,
        color = 256,
        min_side = 100.0
    )

    if not polys_sorted: 
        log("❌ 未获取到有效的打印区域多段线。")
        return [], [], {}

    log(f"✅ 获得排序后的区域: {len(polys_sorted)} 个")

    # ============================================================
    # 🚩 2. 搜集图签块 (候选池)
    # ============================================================
    log("🚀 Step2: 搜集图签块...")
    block_pool = []
    
    try:
        # 搜索策略：同时查找 原始层(dy_quyu_H_m) 和 处理层(dy_quyu_H_m_m) 以及 兜底层
        target_layers = list(set([core_layer, f"{core_layer}_m", "dy_quyu"]))
        log(f"   -> 搜索图层: {target_layers}")
        
        for lay in target_layers:
            res = stc(lay)
            if res:
                batch = []
                if isinstance(res, list): batch = res
                elif hasattr(res, "Item"): batch = [res.Item(k) for k in range(res.Count)]
                else: batch = [res]
                
                for e in batch:
                    # 1. 必须在模型空间 (OwnerID 检查)
                    try:
                        if e.OwnerID != doc.ModelSpace.ObjectID: continue
                    except: continue # 某些对象可能无法获取 OwnerID

                    # 2. 必须是块引用 (BlockReference / MInsertBlock)
                    obj_name = getattr(e, "ObjectName", "")
                    if "Block" in obj_name or "Insert" in obj_name: 
                        block_pool.append(e)
                        
        log(f"   -> 候选池: 找到 {len(block_pool)} 个块对象")
        
    except Exception as e:
        log(f"⚠️ 搜集图块时出错: {e}")
        return [], [], {}

    # ============================================================
    # 🚩 3. 匹配算法 (中心点包含判定)
    # ============================================================
    
    def is_center_in_poly(blk, poly):
        """
        简化版判定：只要块的中心点在多段线的包围盒范围内即可。
        不使用额外的 buffer 容差。
        """
        try:
            # 获取块的包围盒 -> 计算中心
            min_b, max_b = blk.GetBoundingBox()
            cx = (min_b[0] + max_b[0]) / 2.0
            cy = (min_b[1] + max_b[1]) / 2.0
            
            # 获取多段线的包围盒
            min_p, max_p = poly.GetBoundingBox()
            
            # 判定：中心点是否在矩形范围内
            # min_x <= cx <= max_x  AND  min_y <= cy <= max_y
            if (min_p[0] <= cx <= max_p[0]) and (min_p[1] <= cy <= max_p[1]):
                return True
            return False
        except Exception:
            return False

    mapping_dict = {}         # {序号: (PolyHandle, BlockHandle)}
    blocks_sorted_result = [] # 按 Poly 顺序排列的 Block 对象
    matched_blk_handles = set()
    
    log(f"🚀 Step3: 执行几何匹配 ({len(polys_sorted)} vs {len(block_pool)})...")

    for i, poly in enumerate(polys_sorted):
        seq = i + 1
        found_blk = None
        
        # 遍历块池寻找匹配
        for blk in block_pool:
            # 优化：跳过已匹配的块 (假设是一对一关系)
            if blk.Handle in matched_blk_handles: continue
            
            if is_center_in_poly(blk, poly): 
                found_blk = blk
                break
        
        if found_blk:
            # 记录匹配
            mapping_dict[seq] = (poly.Handle, found_blk.Handle)
            blocks_sorted_result.append(found_blk)
            matched_blk_handles.add(found_blk.Handle)
        else: 
            # 未找到匹配的块
            try:
                min_p, max_p = poly.GetBoundingBox()
                center = ((min_p[0]+max_p[0])/2, (min_p[1]+max_p[1])/2)
                log(f"⚠️ 警告: 序号 {seq} 未匹配到图签 (Poly中心: {center[0]:.0f}, {center[1]:.0f})")
            except:
                log(f"⚠️ 警告: 序号 {seq} 未匹配到图签")

    log(f"✅ 配对完成: 成功匹配 {len(mapping_dict)} / {len(polys_sorted)} 组")
    
    return polys_sorted, blocks_sorted_result, mapping_dict





#&&% 重建字典_图纸空间


@timeit
def rebuild_print_area_title_mapping_paper(
    layout_target: str = "布局1",   
    core_layer: str = "dy_quyu_H",              
    final_poly_layer: str = "tuqian_neibu_pl", 
    select_config = 0, 
    verbose=None,
    **kwargs
):
    """
    【函数编号】: MAP-PAPER-FINAL (V40 - 统一逻辑版)
    【功能】: 
        1. 强制切换至目标布局。
        2. 调用 select_print_areas_paperspace 获取标准化、排序后的打印区域。
        3. 搜集本布局内的图签块。
        4. 使用“中心点包含”逻辑建立 区域->图签 的映射关系。
    
    【参数 select_config】:
        - 0: 常规模式 (Precision=False)
        - 1: 高精模式 (Precision=True)
    """
    from  CAD_file_operations   import  switch_to_layout
    doc = C.doc
    lvl = int(verbose) if verbose is not None else 1
    def log(msg): 
        if lvl >= 1: 
            sys_logger.info(f"🔨 [Paper-Map] {msg}")

    # ============================================================
    # 🚩 1. 环境准备与区域获取
    # ============================================================
    
    # 1.1 切换布局
    if not switch_to_layout(layout_target):
        log(f"❌ 无法进入布局 '{layout_target}'，任务终止。")
        return [], [], {}

    # 获取当前布局的 OwnerID (用于后续过滤对象)
    try:
        target_owner_id = doc.PaperSpace.ObjectID
    except:
        log("❌ 无法获取布局空间ID")
        return [], [], {}

    # 1.2 获取打印区域
    log(f"🚀 Step1: 获取布局打印区域 (Config={select_config})...")
    
    # 映射配置参数
    is_precision = True if select_config == 1 else False

    # 调用核心选择函数 (图纸空间版)
    polys_sorted, _ = select_print_areas_paperspace(
        layout_name = layout_target,      # 必须传入当前布局名
        layer_name = final_poly_layer,    # "tuqian_neibu_pl"
        precision_mode = is_precision,
        width = 0.0,
        color = 256,
        min_side = 100.0
    )

    if not polys_sorted: 
        log("❌ 未获取到有效的打印区域多段线。")
        return [], [], {}

    log(f"✅ 获得排序后的区域: {len(polys_sorted)} 个")

    # ============================================================
    # 🚩 2. 搜集图签块 (候选池)
    # ============================================================
    log("🚀 Step2: 搜集本布局图签块...")
    block_pool = []
    
    try:
        # 搜索策略：查找 core_layer 和 兜底层 dy_quyu
        target_layers = list(set([core_layer, "dy_quyu"]))
        log(f"   -> 搜索图层: {target_layers}")
        
        for lay in target_layers:
            res = stc(lay)
            if res:
                batch = []
                if isinstance(res, list): batch = res
                elif hasattr(res, "Item"): batch = [res.Item(k) for k in range(res.Count)]
                else: batch = [res]
                
                for e in batch:
                    # 1. 必须在当前布局空间 (OwnerID 检查)
                    try:
                        if e.OwnerID != target_owner_id: continue
                    except: continue 

                    # 2. 必须是块引用
                    obj_name = getattr(e, "ObjectName", "")
                    if "Block" in obj_name or "Insert" in obj_name: 
                        block_pool.append(e)
                        
        log(f"   -> 候选池: 找到 {len(block_pool)} 个块对象")
        
    except Exception as e:
        log(f"⚠️ 搜集图块时出错: {e}")
        return [], [], {}

    # ============================================================
    # 🚩 3. 匹配算法 (中心点包含判定)
    # ============================================================
    
    def is_center_in_poly(blk, poly):
        """
        简化版判定：只要块的中心点在多段线的包围盒范围内即可。
        """
        try:
            # 获取块的包围盒 -> 计算中心
            min_b, max_b = blk.GetBoundingBox()
            cx = (min_b[0] + max_b[0]) / 2.0
            cy = (min_b[1] + max_b[1]) / 2.0
            
            # 获取多段线的包围盒
            min_p, max_p = poly.GetBoundingBox()
            
            # 判定：中心点是否在矩形范围内
            if (min_p[0] <= cx <= max_p[0]) and (min_p[1] <= cy <= max_p[1]):
                return True
            return False
        except Exception:
            return False

    mapping_dict = {}         
    blocks_sorted_result = [] 
    matched_blk_handles = set()
    
    log(f"🚀 Step3: 执行几何匹配 ({len(polys_sorted)} vs {len(block_pool)})...")

    for i, poly in enumerate(polys_sorted):
        seq = i + 1
        found_blk = None
        
        # 遍历块池寻找匹配
        for blk in block_pool:
            if blk.Handle in matched_blk_handles: continue
            
            if is_center_in_poly(blk, poly): 
                found_blk = blk
                break
        
        if found_blk:
            mapping_dict[seq] = (poly.Handle, found_blk.Handle)
            blocks_sorted_result.append(found_blk)
            matched_blk_handles.add(found_blk.Handle)
        else: 
            try:
                min_p, max_p = poly.GetBoundingBox()
                center = ((min_p[0]+max_p[0])/2, (min_p[1]+max_p[1])/2)
                log(f"⚠️ 警告: 序号 {seq} 未匹配到图签 (Poly中心: {center[0]:.0f}, {center[1]:.0f})")
            except:
                log(f"⚠️ 警告: 序号 {seq} 未匹配到图签")

    log(f"✅ 配对完成: 成功匹配 {len(mapping_dict)} / {len(polys_sorted)} 组")
    
    return polys_sorted, blocks_sorted_result, mapping_dict




#&&% excel字典信息转换

#常量

# 第一部分：每张图纸的字段（列头）
DRAWING_KEYS = [
    "序号",
    "图纸编号",
    "图纸名称",
    "图纸规格",
    "出图比例",
    "纸张数",

]

# 第二部分：整体信息字段（列头）
PROJECT_KEYS = [
    "专业名称",
    "专业代号",
    "项目名称",
    "子项目名称",
    "建设单位名称",
    "设计阶段",
    "版本号",
    "出图日期",
    "设计编号",
    "设计院名称",
]


#&&% 构建表头映射
def build_header_map(ws) -> Dict[str, int]:
    """
    根据第 1 行表头构建 {表头文字: 列号} 的映射。
    列号为 openpyxl 使用的 1-based 整数。
    """
    header_map: Dict[str, int] = {}
    for cell in ws[1]:
        v = cell.value
        if v is None:
            continue
        key = str(v).strip()
        if key:
            header_map[key] = cell.column
    return header_map




#&&% 读取Excel到字典

def read_xlsx_to_dict(xlsx_path: str) -> Dict[str, Any]:
    """
    读取 Excel，转成统一的数据字典结构。
    
    【修改说明】：
    1. 移除了对 PROJECT_KEYS 的依赖，改为扫描所有非图纸列的表头，实现自动扩展。
    2. 只要 Excel 第一行有标题，第二行有值，就会被自动读入 project 字典。
    """
    
    # 1. 定义图纸列表专用的字段 (这些字段属于 drawings 数组，不属于 project)
    #    根据需要，这里列出所有“随每一行变化”的列名
    DRAWING_KEYS = {
        "序号", "图纸编号", "图纸名称", "图纸规格", "出图比例", "纸张数",
        "_layout_name", "_block_handle" # 如果Excel里有这些元数据列
    }

    # 加载 Excel
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
    except Exception as e:
        sys_logger.info(f"❌ 打开 Excel 失败: {e}")
        return {}

    # --- 0. 构建表头映射 (Header Map) ---
    # 格式: { "专业名称": 1, "图纸编号": 2, ... }
    header_map = {}
    if ws.max_row >= 1:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                # 去除可能的空格，并转为字符串
                header_name = str(val).strip()
                header_map[header_name] = col

    # ---------- 1) 整体信息（Project）：自动扫描所有扩展列 ----------
    # 逻辑：只要表头存在，且不在 DRAWING_KEYS 里，就认为是项目全局信息
    project_row_index = 2
    project_info: Dict[str, Any] = {}
    
    for header_name, col_idx in header_map.items():
        # 如果这个标题是图纸列表专用的（如“图纸编号”），则跳过，不放进 project 字典
        if header_name in DRAWING_KEYS:
            continue
            
        # 读取第 2 行对应列的值
        # 注意：这里假设项目信息只填在第 2 行。如果后续行也有不同值，通常那是图纸信息。
        cell_val = ws.cell(row=project_row_index, column=col_idx).value
        
        # 存入字典 (去除 None 值，或者保留 None 看你需求，这里保留 None 以便知道列存在)
        project_info[header_name] = cell_val

    # ---------- 2) 图纸信息（Drawings）：从第 2 行往下读取 ----------
    drawings: List[Dict[str, Any]] = []
    row = 2
    empty_row_streak = 0 

    while True:
        # 只读取属于 DRAWING_KEYS 的列
        row_data: Dict[str, Any] = {}
        has_effective_data = False # 用于判断这一行是不是空行
        
        # 遍历我们关心的图纸字段
        for key in DRAWING_KEYS:
            col = header_map.get(key)
            if col is None:
                continue # Excel里没这一列，跳过
            
            value = ws.cell(row=row, column=col).value
            row_data[key] = value
            
            # 判断行是否有数据 (排除序号列，因为有时候序号是公式自动生成的，但内容是空的)
            if value not in (None, ""):
                has_effective_data = True

        # --- 空行检测逻辑 ---
        if not has_effective_data:
            empty_row_streak += 1
            if empty_row_streak >= 3: # 连续3行空行则结束
                break
        else:
            empty_row_streak = 0
            # 只有当这一行有实际图纸数据时，才加入列表
            # 也可以加一个判断：必须有“图纸编号”或“图纸名称”才算有效
            if row_data.get("图纸编号") or row_data.get("图纸名称"):
                drawings.append(row_data)

        row += 1
        if row > ws.max_row + 10: # 安全熔断
            break

    return {
        "project": project_info,
        "drawings": drawings,
    }

#&&% 写入字典到Excel
def write_dict_to_xlsx(
    data: Dict[str, Any],
    output_xlsx_path: str,    # 必填在前
    template_xlsx_path=None,  # 选填在后
) -> None:
    """
    【最终修正版】将字典写回 Excel。
    
    核心逻辑：
    1. 【读取默认值】：先读取模板第2行的数据存入内存（保留设计院名称等固定信息）。
    2. 【彻底清空】：将第 2 行及之后的所有单元格内容清空（保留样式），仅保留第 1 行标题。
    3. 【写入数据】：
       - 图纸列：每一行都写。
       - 项目列：只写第 2 行。
       - 扩展性：自动识别所有表头。
    """
    
    # 0. 环境与路径处理
    if template_xlsx_path is None:
        userpath = os.environ.get('USERPATH')
        if not userpath:
            print("❌ 错误：未找到系统环境变量 USERPATH")
            return
        template_xlsx_path = os.path.join(userpath, "标准模板", "项目图纸信息模板.xlsx")
    
    # 定义图纸行数据列
    DRAWING_KEYS = {
        "序号", "图纸编号", "图纸名称", "图纸规格", "出图比例", "纸张数", 
        "备注", "_layout_name" 
    }

    try:
        wb = load_workbook(template_xlsx_path)
        ws = wb.active
    except Exception as e:
        sys_logger.info(f"❌ 加载模板失败: {e}")
        return

    # 获取表头映射
    header_map = build_header_map(ws)
    if not header_map:
        return

    # =================================================
    # Step 1: 读取默认值 (Read Defaults)
    # =================================================
    # 先把模板第2行里的“设计院”、“审核人”等固定信息存下来
    default_values: Dict[str, Any] = {}
    for header, col_idx in header_map.items():
        val = ws.cell(row=2, column=col_idx).value
        default_values[header] = val

    # =================================================
    # Step 2: 彻底清空模板值 (Clear Values)
    # =================================================
    # 从第 2 行开始，直到最后一行，把值设为 None
    # 注意：这里不用 delete_rows，因为那样会删除行高和边框样式
    # 我们只清空 value，保留 style
    if ws.max_row >= 2:
        # 使用 iter_rows 效率更高
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.value = None

    # =================================================
    # Step 3: 准备写入数据
    # =================================================
    project_info = data.get("project", {}) or {}
    drawings = data.get("drawings", []) or []
    
    # 即使没有图纸，也至少要处理第2行（为了填项目信息）
    rows_to_process = len(drawings) if len(drawings) > 0 else 1
    start_row = 2
    
    # =================================================
    # Step 4: 循环写入
    # =================================================
    for i in range(rows_to_process):
        current_row = start_row + i
        
        # 获取当前行的图纸数据
        drawing_data = drawings[i] if i < len(drawings) else {}

        # 遍历 Excel 所有列
        for header, col_idx in header_map.items():
            
            val_to_write = None

            # --- 分支 A: 项目信息列 (只写第2行) ---
            if header not in DRAWING_KEYS:
                if current_row == 2: # 只有第2行才写入
                    val_to_write = project_info.get(header)
                    # 兜底：如果数据源没值，回填模板默认值
                    if val_to_write is None or str(val_to_write).strip() == "":
                        val_to_write = default_values.get(header)

            # --- 分支 B: 图纸信息列 (每行都写) ---
            else:
                if header == "序号":
                    val_to_write = drawing_data.get("序号")
                    if not val_to_write and i < len(drawings):
                        val_to_write = f"{i + 1:02d}"
                else:
                    val_to_write = drawing_data.get(header)

                # 图纸列通常不使用模板默认值(除非是空的备注之类)，可视情况决定
                # 这里逻辑：如果为空，尝试用默认值(例如模板里备注写了"无")
                if val_to_write is None or str(val_to_write).strip() == "":
                    val_to_write = default_values.get(header)

            # --- 写入单元格 ---
            if val_to_write is not None:
                ws.cell(row=current_row, column=col_idx, value=val_to_write)

    # 保存文件
    try:
        wb.save(output_xlsx_path)
        sys_logger.info(f"✅ Excel 生成成功: {output_xlsx_path}")
    except Exception as e:
        sys_logger.info(f"❌ 保存失败 (文件可能被占用): {e}")




#&&&% ===（五）从Cad写入Excel ===

#&&% 自动导出
def auto_export_excel_with_fallback(
    layout_name=None, 
    operate_target="Model", 
    select_config=None, 
    template_path=None,
    output_path=None,  
    start_index=1,
    use_cache=False,
    verbose=1
):
    """
    【函数编号】: XLS-AUTO-FORCE-001
    【功能】: 增强型导出调度器。
             具备 auto_export_excel_by_style 的全部功能。
             【新增特性】: 即使提取不到任何图签块 (Blocks)，只要有多段线 (Polys)，
             也会强制导出 Excel，并自动填充“无图纸名称”、“无编号”。
    """
    
    # 1. 智能推断 select_config
    if select_config is None:
        select_config = 1 if (operate_target != "Model" or layout_name) else 0

    if verbose:
        sys_logger.info(f"🚀 [强制导出模式] 目标: {operate_target} | 布局: {layout_name} | 策略: {select_config}")

    # 2. 获取数据 (返回三元组)
    polys, blocks, mapping = smart_rebuild_print_info(
        layout_name=layout_name,
        operate_target=operate_target,
        select_config=select_config,
        use_cache=use_cache,
        verbose=verbose
    )

    # 3. 数据完整性检查与修复 (关键逻辑)
    # -----------------------------------------------------------
    # 如果连框线都没有，那就真的没法导出了
    if not polys:
        print("⚠️ [自动导出] 未检测到任何打印框线 (Polys)，停止导出。")
        return False

    # 检查图签块是否存在
    # 如果 blocks 为空，或者数量与 polys 严重不符(例如为0)，则进行填充
    if not blocks:
        if verbose: sys_logger.info(f"ℹ️ [信息] 未检测到图签块，检测到 {len(polys)} 个框线。正在构造虚拟数据以强制导出...")
        # 生成与 polys 等长的 None 列表，作为虚拟块
        blocks = [None] * len(polys)
        # mapping 保持为空即可，底层函数会忽略
        mapping = {}
    elif len(blocks) < len(polys):
        # 可选：如果块的数量少于框线，也可以选择补全 None，防止数据丢失
        # 这里演示简单的补全逻辑
        missing_count = len(polys) - len(blocks)
        blocks.extend([None] * missing_count)

    # 4. 构造修正后的 CTQ 数据包
    ctq_data = (polys, blocks, mapping)

    # 5. 决定 Excel 文件名后缀
    excel_layout_arg = layout_name if operate_target == "Layout" else None
    
    # 6. 执行导出
    return build_full_print_dict_and_export_excel(
        ctq=ctq_data,
        layout_target=excel_layout_arg, 
        template_path=template_path,
        output_path=output_path,        
        start_index=start_index
    )






#&&% 基础导出

def build_full_print_dict_and_export_excel(
    ctq,
    layout_target=None, 
    template_path=None,
    output_path=None,
    start_index=1,  
):
    """
    【函数编号】: XLS-001 (V3.9 - 极简回退版)
    【修改说明】: 
        基于 V3.1 原版。
        新增逻辑：当 ctq[1] (图签块) 为空但 ctq[0] (多段线) 有值时，
        自动进入“纯框线模式”，生成一份仅包含图幅和比例的骨架 Excel。
    """
    # 打印当前模式，方便调试
    mode_str = '图纸空间 - '+layout_target if layout_target else '模型空间'
    sys_logger.info(f"版本:v3.9 (极简回退版) | 模式: {mode_str}")

    import re
    import time
    from pathlib import Path # 确保引入 Path

    # ======================== 0. 字段常量 ========================
    PROJECT_KEYS = [
        "专业名称", "专业代号", "项目名称", "子项目名称", "建设单位名称", 
        "设计阶段", "版本号", "出图日期", "设计编号", "设计院名称",
        "法人代表","注册建筑师","注册结构师",
    ]

    INHERIT_KEYS = [
        "设计院名称", 
        "法人代表", 
        "注册建筑师", 
        "注册结构师"
    ]

    INVALID_PLACEHOLDERS = ["某设计院", "某某", "某工程师", "DEFAULT", "XXXX", "xxxx"]

    t0 = time.time()

    # ======================== 1. CAD 数据准备 ========================

    current_doc = C.doc

    try: doc_name = Path(current_doc.Name).stem
    except: doc_name = "未命名图纸"

    userpath = os.environ.get('USERPATH')

    # --- 1.0 切换空间 ---
    if layout_target:
        try:
            sys_logger.info(f"📂 切换至布局: {layout_target} (以确保环境一致)")
            current_doc.ActiveLayout = current_doc.Layouts.Item(layout_target)
            current_doc.MSpace = False
        except Exception as e:
            sys_logger.info(f"❌ 无法切换到布局 '{layout_target}': {e}")
            return False

    # --- 1.1 确定模板路径 ---
    if template_path is None:
        if not userpath:
            print("❌ 错误：未找到系统环境变量 USERPATH (无法确定默认模板路径)")
            return False
        template_path = os.path.join(userpath, "标准模板", "项目图纸信息模板.xlsx")

    # --- 1.2 确定输出路径 ---
    if output_path is None:
        try:
            base_path = Path(current_doc.FullName)
            current_stem = base_path.stem  
            
            if layout_target:
                excel_name = f"{current_stem}_{layout_target}.xlsx"
            else:
                excel_name = f"{current_stem}.xlsx"
            
            output_path = base_path.parent / excel_name
            sys_logger.info(f"📍 Excel 输出路径锁定: {output_path.name}")
            
        except Exception as e:
            sys_logger.info(f"❌ 路径生成失败 (请先保存DWG文件): {e}")
            return False
    else:
        output_path = Path(output_path)


    # ======================== 1.3 数据分支逻辑 (修改核心) ========================
    if not isinstance(ctq, (list, tuple)) or len(ctq) < 1:
        print("⚠ 输入数据格式错误"); return False

    dyx_list = ctq[0] # 多段线
    # 兼容处理：如果 ctq 长度不够，视为空列表
    tq_list  = ctq[1] if len(ctq) >= 2 else []

    paired = []
    count = 0

    # —————————— 分支判定 ——————————
    # 情况 A: 正常模式 (有框线 且 有图签)
    if len(dyx_list) > 0 and len(tq_list) > 0:
        count = min(len(dyx_list), len(tq_list))
        paired = list(zip(dyx_list[:count], tq_list[:count]))
        sys_logger.info(f"📌 [标准模式] 匹配到 {count} 组图纸数据 (按列表顺序配对)")

    # 情况 B: 兜底模式 (有框线 但 无图签) -> 您的新需求
    elif len(dyx_list) > 0 and len(tq_list) == 0:
        count = len(dyx_list)
        # 构造 (Poly, None) 的配对
        paired = [(dyx, None) for dyx in dyx_list]
        sys_logger.info(f"ℹ️ [兜底模式] 未检测到图签，将基于 {count} 个打印框线生成骨架信息")
    
    # 情况 C: 无数据
    else:
        print("⚠ 无有效图纸数据 (无框线)"); return False


    # ======================== 1.5 预扫描：确定全局专业代号 ========================
    global_discipline_code = "JZ"
    found_code_source = "默认"
    
    # 只有在标准模式下才尝试扫描，兜底模式直接用 JZ
    if len(tq_list) > 0:
        print("🔍 正在预扫描专业代号...")
        for _, br in paired:
            if br is None: continue
            try:
                temp_attrs = get_block_attributes_dict(br, ignore_empty=True, upper_tag=True)
                raw_no = str(temp_attrs.get("图纸编号", "")).strip()
                temp_code = None
                clean_no = raw_no.replace("—", "-").replace("_", "-").replace(" ", "-")
                if "-" in clean_no:
                    parts = clean_no.split("-")
                    if parts[0].strip(): temp_code = parts[0].strip()
                else:
                    match = re.match(r"^([^\d]+)", clean_no)
                    if match: temp_code = match.group(1).strip()
                
                if temp_code and temp_code.upper() != "JZ":
                    global_discipline_code = temp_code.upper()
                    found_code_source = f"提取自 '{raw_no}'"
                    break 
            except: pass
    
    sys_logger.info(f"✅ 全局专业代号: {global_discipline_code} ({found_code_source})")

    # ======================== 2. 构建数据 ========================
    print_infos = []
    
    for idx, (dyx, br) in enumerate(paired, start=1):
        attrs = {}

        # —————————— 分支提取 ——————————
        if br is not None:
            # A. 正常提取属性
            try:
                attrs = get_block_attributes_dict(br, ignore_empty=True, upper_tag=True)
            except: attrs = {}
            
            # 清洗换行符
            for k, v in attrs.items():
                if isinstance(v, str) and "\\P" in v:
                    attrs[k] = v.replace("\\P", "//")
        else:
            # B. 兜底填充 (无图签块)
            attrs = {
                "图纸名称": "无图纸名称",
                "图纸编号": "无编号",
                # 其他字段留空，等待 Excel 继承或用户手填
            }

        # B. 几何信息 (通用逻辑：只要有 dyx 就能算)
        geom_info = None
        try:
            if "generate_name_and_ratio_from_com" in globals():
                res = generate_name_and_ratio_from_com(dyx)
                if isinstance(res, (tuple, list)) and len(res) >= 3:
                    geom_info = res
        except: pass

        if geom_info:
            if not attrs.get("图纸规格"): attrs["图纸规格"] = str(geom_info[2])
            if not attrs.get("出图比例"): attrs["出图比例"] = str(geom_info[1])

        # C. 强制应用全局代号
        attrs["专业代号"] = global_discipline_code 
        
        print_infos.append({
            "index": idx - 1,
            "attrs": attrs,
            "discipline_code": global_discipline_code
        })

    # D. 提取项目整体信息
    overall_info = {k: None for k in PROJECT_KEYS}
    overall_info["专业代号"] = global_discipline_code 
    
    for info in print_infos:
        attrs = info["attrs"]
        for k in PROJECT_KEYS:
            if k == "专业代号": continue 
            if overall_info[k] is None:
                val = attrs.get(k)
                # 修改点：如果是默认的"无图纸名称"等占位符，不要作为项目信息提取
                if val and str(val).strip() and "无图纸" not in str(val) and "无编号" not in str(val):
                    overall_info[k] = str(val)

    # ======================== 3. 操作 Excel ========================
    excel = None; wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False; excel.DisplayAlerts = False
    except Exception as e:
        sys_logger.info(f"[错误] Excel 启动失败: {e}"); return False

    # —————————— 【历史数据继承】 ——————————
    C_PROJ_START = 7 
    
    if output_path.exists():
        sys_logger.info(f"👀 发现同名文件 ({output_path.name})，正在检查历史数据...")
        try:
            # 临时打开旧文件
            wb_old = excel.Workbooks.Open(str(output_path))
            try: ws_old = wb_old.Worksheets(1)
            except: ws_old = wb_old.ActiveSheet
            
            # 遍历需要继承的 Key
            for key in INHERIT_KEYS:
                if key in PROJECT_KEYS:
                    offset = PROJECT_KEYS.index(key)
                    col_idx = C_PROJ_START + offset
                    
                    # 读取旧值 (第2行)
                    old_val = ws_old.Cells(2, col_idx).Value
                    if old_val:
                        old_val_str = str(old_val).strip()
                        if old_val_str and old_val_str not in INVALID_PLACEHOLDERS:
                            overall_info[key] = old_val_str
                            sys_logger.info(f"   ↺ 继承历史值 [{key}]: {old_val_str}")
            
            wb_old.Close(SaveChanges=0) 
        except Exception as e:
            sys_logger.info(f"⚠ 读取旧文件失败 (将使用CAD默认值): {e}")
            try: wb_old.Close(SaveChanges=0)
            except: pass
    # ————————————————————————————————————————————————————

    template_path = Path(template_path)
    if not template_path.is_file():
        sys_logger.info(f"[错误] 模板缺失: {template_path}"); excel.Quit(); return False

    try:
        wb = excel.Workbooks.Open(str(template_path))
        try: ws = wb.Worksheets(1)
        except: ws = wb.ActiveSheet

        # ---- 增删行 ----
        ROW_START = 2
        template_rows_count = 0
        r = ROW_START
        while True:
            if ws.Cells(r, 1).Value in [None, ""]: break
            template_rows_count += 1; r += 1; 
            if template_rows_count > 500: break
        
        needed_rows = count
        if needed_rows > template_rows_count:
            rows_to_add = needed_rows - template_rows_count
            last_row = ROW_START + template_rows_count - 1
            ws.Rows(last_row).Copy()
            ws.Range(f"{last_row+1}:{last_row+rows_to_add}").Insert()
            excel.CutCopyMode = False
        elif needed_rows < template_rows_count:
            del_start = ROW_START + needed_rows
            del_end = ROW_START + template_rows_count - 1
            if del_start <= del_end: ws.Range(f"{del_start}:{del_end}").Delete()

        try:
            ws.Range(f"A{ROW_START}:A{ROW_START+count}").NumberFormatLocal = "@"
            ws.Range(f"F{ROW_START}:F{ROW_START+count}").NumberFormatLocal = "@"
        except: pass


        # ======================== 4. 填入数据 ========================
        print("✍ 正在填入 Excel...")
        C_SER=1; C_CODE=2; C_NAME=3; C_SPEC=4; C_SCALE=5; C_SHEETS=6
        # C_PROJ_START = 7  (前面已经定义)

        for i in range(count):
            row_idx = ROW_START + i
            info = print_infos[i]
            attrs = info["attrs"]
            
            # 使用 start_index 计算当前的显示编号
            current_num = i + start_index 

            ws.Cells(row_idx, C_SER).Value = f"{current_num:02d}"
            ws.Cells(row_idx, C_CODE).Value = f"{info['discipline_code']}-{current_num:02d}"

            if v := attrs.get("图纸名称"): ws.Cells(row_idx, C_NAME).Value = str(v)
            if v := attrs.get("图纸规格"): ws.Cells(row_idx, C_SPEC).Value = str(v)
            if v := attrs.get("出图比例"): ws.Cells(row_idx, C_SCALE).Value = str(v)
            ws.Cells(row_idx, C_SHEETS).Value = f"{count}/1"

            if row_idx > ROW_START:
                try: ws.Range(f"G{row_idx}:P{row_idx}").ClearContents()
                except: pass

        # ======================== 5. 填入项目信息 ========================
        for offset, key in enumerate(PROJECT_KEYS):
            val = overall_info.get(key)
            if val is not None:
                val_str = str(val).strip()
                
                # 防覆盖保护
                if key in INHERIT_KEYS and val_str in INVALID_PLACEHOLDERS:
                    continue

                ws.Cells(ROW_START, C_PROJ_START + offset).Value = val_str

        # ======================== 6. 保存 ========================
        save_path = str(output_path.resolve())
        sys_logger.info(f"💾 保存路径: {save_path}")

        if output_path.exists():
            try: output_path.unlink()
            except PermissionError:
                sys_logger.info(f"❌ 文件被占用: {save_path}"); wb.Close(SaveChanges=0); return False
            except: pass

        try: wb.SaveAs(Filename=save_path, FileFormat=51)
        except: wb.SaveCopyAs(save_path)

        sys_logger.info(f"✅ 导出成功!")
        return True

    except Exception as e:
        import traceback; traceback.print_exc()
        sys_logger.info(f"❌ 错误: {e}")
        return False
    finally:
        if wb: wb.Close(SaveChanges=0)
        if excel: excel.Quit()
        del wb, excel



#&&&% ===（六）从图纸获取信息写入图签 ===

#&&% 自动_图名标注写入图签

def auto_process_drawing_names_by_style(
    layout_name=None, 
    operate_target="Model", 
    select_config=None, 
    # --- 下面是算法微调参数 ---
    layername="DIM_SYMB",  # 图名所在的图层
    num1=14,               # 智能换行阈值
    num2=3,                # 孤儿行合并阈值
    start_index=1,         # 起始编号
    prefix=None,           # 强制指定前缀(如"JG")，None则自动探测
    use_cache=False,
    verbose=1
):
    """
    【函数编号】: XLS-AUTO-003 (V1.0 - 全自动图名抓取版)
    【所属模块】: 自动化处理模块 (Auto Processing)
    【功能】: 
        智能图名处理调度器。
        根据空间样式，自动提取 CTQ 打印框，并在相应的空间内利用【几何包围盒】算法，
        抓取图框内部的图名文字，清洗后回填到图签属性中。
    
    【样式匹配逻辑】:
        1. 纯模型 (Model, None):
           -> 提取 Model 打印框 -> 在 Model 空间扫描图名 -> 几何匹配
        2. 混合模型 (Model, "Layout1"):
           -> 提取 Model (视口映射) 打印框 -> 在 Model 空间扫描图名 -> 几何匹配
           *注意*: 混合模式下，通常图名标注在模型空间，与图框坐标一致。
        3. 纯布局 (Layout, "Layout1"):
           -> 提取 Layout 打印框 -> 在 Layout 空间扫描图名 -> 几何匹配

    【参数】:
        - layout_name, operate_target, select_config: 核心定位参数 (同上)。
        - layername: 告诉程序图名文字在哪个图层 (支持 TDbDrawingName 和普通文字)。
    """

    # =========================================================================
    # 1. 智能推断 select_config (CTQ 提取策略)
    # =========================================================================
    if select_config is None:
        if operate_target == "Model":
            select_config = 1 if layout_name else 0
        else:
            select_config = 1

    if verbose >= 1:
        sys_logger.info(f"🚀 [自动图名] 目标: {operate_target} | 布局: {layout_name} | 策略: {select_config}")

    # =========================================================================
    # 2. 获取目标数据 (调用 smart_rebuild_print_info)
    # =========================================================================
    # 我们需要打印框 (Poly) 作为搜捕范围，需要图签块 (Block) 作为写入目标
    polys, blocks, mapping = smart_rebuild_print_info(
        layout_name=layout_name,
        operate_target=operate_target,
        select_config=select_config,
        use_cache=use_cache,
        verbose=verbose
    )

    if not polys:
        if verbose >= 1: print("⚠️ [自动图名] 未提取到任何图框，无法执行抓取。")
        return False

    # =========================================================================
    # 3. 构造 CTQ 数据包
    # =========================================================================
    ctq_data = (polys, blocks, mapping)

    # =========================================================================
    # 4. 决定“图名搜索空间” (Search Space Logic)
    # =========================================================================
    # process_drawing_names_and_fill_titleblocks 的 layout_target 参数决定了去哪里找文字
    
    search_target_arg = None
    
    if operate_target == "Layout":
        # 【纯布局模式】
        # 打印框在布局空间，图名也在布局空间
        # 传入布局名，底层函数会切换到该布局并扫描 PaperSpace
        search_target_arg = layout_name
        
    elif operate_target == "Model":
        # 【纯模型 OR 混合模型】
        # 打印框坐标是模型空间坐标 (即使是混合模式，smart_rebuild 返回的也是映射后的 Model 坐标)
        # 因此，我们必须在 Model 空间搜索图名，才能进行几何碰撞检测 (Box in Box)。
        # 如果在混合模式下图名写在布局空间，坐标系是不重叠的，无法匹配。
        # 结论：只要 operate_target 是 Model，就去 Model 找图名。
        search_target_arg = None

    # =========================================================================
    # 5. 执行自动抓取与回填
    # =========================================================================
    return process_drawing_names_and_fill_titleblocks(
        ctq=ctq_data,
        layout_target=search_target_arg, # 决定去哪个空间搜寻图名对象
        num1=num1,
        num2=num2,
        layername=layername,
        start_index=start_index,
        prefix=prefix
    )





#&&% 基础_图名标注写入图签

def process_drawing_names_and_fill_titleblocks(
    ctq, 
    layout_target=None, # 【新增参数】布局名称 (模型空间传None, 布局传具体名称)
    num1=14,    # 单行字数阈值
    num2=3,     # 孤儿行阈值
    layername="DIM_SYMB",
    start_index=1, 
    prefix=None # 默认自动探测
):
    """
    【函数编号】: XLS-003 (V3.1 - 双空间适配版)
    【功能描述】: 
        全自动提取 CAD 内部图名，支持模型空间和图纸空间两种模式。
        根据 layout_target 参数自动切换抓取策略。

    【参数详解】:
        - ctq (tuple): (polys, blocks) 配对数据。
        - layout_target (str): 
            * None: 模型空间模式 (默认)。
            * "布局名称": 图纸空间模式，将在指定布局内搜索图名。
        - num1, num2, layername, start_index, prefix: 同原版。
    """

    import re 
    from CAD_file_operations import save_file

    # ================= 1. 基础校验与环境准备 =================
    # 确保连接

    current_doc = C.doc

    if not ctq or len(ctq) < 2: 
        print("❌ 错误: CTQ 数据无效")
        return False
    
    polys = ctq[0]  # 打印框 (多段线)
    blocks = ctq[1] # 图签块 (BlockReference)
    count = min(len(polys), len(blocks))
    if count == 0: 
        print("❌ 错误: 匹配到的图框/图签数量为 0")
        return False

    mode_str = f"图纸空间 [{layout_target}]" if layout_target else "模型空间"
    sys_logger.info(f"🚀 开始处理 {count} 张图纸 | 模式: {mode_str}")
    sys_logger.info(f"   (参数: 换行={num1}, 孤儿行={num2}, 图层={layername})")

    # ================= 2. 切换空间 & 获取图名对象 =================
    # 核心改造点：根据模式决定去哪里找 "TDbDrawingName"
    
    all_drawing_names = []

    try:
        if layout_target:
            # ——————————【分支 B: 图纸空间模式】——————————
            
            try:
                # 1. 激活指定布局 (确保 PaperSpace 指向正确)
                current_doc.ActiveLayout = current_doc.Layouts.Item(layout_target)
                current_doc.MSpace = False # 确保不是视口激活状态
                sys_logger.info(f"📂 已切换至布局: {layout_target}")
                
                # 2. 仅遍历该布局空间的对象
                paper_space = current_doc.PaperSpace
                sys_logger.info(f"🔍 正在布局中扫描图层 '{layername}' ...")
                
                for ent in paper_space:
                    # 过滤图层 (不区分大小写)
                    if ent.Layer.lower() == layername.lower():
                        # 过滤对象类型 (兼容天正图名 TDbDrawingName 和普通文字)
                        if ent.ObjectName == "TDbDrawingName":
                            all_drawing_names.append(ent)
            
            except Exception as e:
                sys_logger.info(f"❌ 切换布局或扫描失败: {e}")
                return False

        else:
            # ——————————【分支 A: 模型空间模式】——————————
            # 保持原有逻辑，使用 stc 全局/模型选择
            sys_logger.info(f"🔍 正在模型空间扫描图层 '{layername}' ...")
            selection = stc(layername) # 假设 stc 是您环境中已定义的函数
            
            if selection:
                for ent in selection:
                    if ent.ObjectName == "TDbDrawingName":
                        all_drawing_names.append(ent)
    
    except Exception as e:
        sys_logger.info(f"⚠️ 扫描图名对象时发生异常: {e}")

    sys_logger.info(f"✅ 共找到 {len(all_drawing_names)} 个潜在图名对象")

    # ================= 3. 前缀确定逻辑 (保持不变) =================
    target_prefix = "JZ" 
    found_source = "默认"

    if prefix is not None:
        target_prefix = prefix
        found_source = "用户指定"
    else:
        scan_limit = min(5, len(blocks))
        for i in range(scan_limit):
            try:
                blk = blocks[i]
                if not blk.HasAttributes: continue
                atts = blk.GetAttributes()
                for att in atts:
                    if att.TagString == "图纸编号":
                        val = att.TextString.strip()
                        match = re.match(r"^([A-Za-z0-9]+)[\-_]", val)
                        if match:
                            target_prefix = match.group(1)
                            found_source = f"提取自图签{i+1}"
                            break
                if found_source != "默认": break
            except: pass
            
    sys_logger.info(f"🔧 使用编号前缀: {target_prefix} ({found_source})")

    # ================= 4. 循环匹配与写入 (逻辑复用) =================
    # 由于 poly, block, drawing_name 对象在 COM 接口中自带坐标属性
    # 只要它们在同一个空间内 (Model或Paper)，GetBoundingBox 的逻辑是通用的。

    success_count = 0
    
    for i in range(count):
        poly = polys[i]
        block = blocks[i]
        
        # --- A. 几何计算 ---
        size_val = ""; scale_val = ""; orientation_flag = 0
        try:
            # 兼容性调用：如果 geometry 函数需要 doc，它通常使用全局 doc
            res = generate_name_and_ratio_from_com(poly, A3dy=0)
            if isinstance(res, (tuple, list)):
                scale_val = str(res[1])
                size_val = str(res[2])
                if len(res) >= 4: orientation_flag = int(res[3])
        except: pass

        # --- B. 空间搜索 (Box in Box) ---
        # 判断哪个图名文字落在了当前这个打印框(poly)里
        valid_names_objs = []
        try:
            min_p, max_p = poly.GetBoundingBox()
            x_min, x_max = min_p[0], max_p[0]
            y_min, y_max = min_p[1], max_p[1]
            
            for name_obj in all_drawing_names:
                try:
                    n_min, n_max = name_obj.GetBoundingBox()
                    cx = (n_min[0] + n_max[0]) / 2
                    cy = (n_min[1] + n_max[1]) / 2
                    # 判定中心点是否在框内
                    if x_min <= cx <= x_max and y_min <= cy <= y_max:
                        valid_names_objs.append(name_obj)
                except: pass
        except: pass

        # --- C. 排序 (Sort) ---
        sorted_objs = []
        if valid_names_objs:
            try:
                if orientation_flag == 1: # 竖图
                    sorted_objs = sort_coms_by_rbcorner(valid_names_objs, cha_X=100)
                else: # 横图
                    sorted_objs = sort_coms_by_llcorner(valid_names_objs, cha_Y=100)
            except: sorted_objs = valid_names_objs
        
        # --- D. 提取文字 & 智能分行 ---
        
        raw_texts = []
        for obj in sorted_objs:
            # 尝试获取属性，兼容不同类型的对象
            txt = str(get_attr(obj, "图名文字") or "").strip()
            # 如果是普通 Text/MText，可能没有 "图名文字" 属性，尝试 .TextString
            if not txt and hasattr(obj, 'TextString'):
                txt = obj.TextString.strip()
            if txt: raw_texts.append(txt)

        final_name = "未命名图纸"
        
        if raw_texts:
            # --- 智能分行算法 ---
            lines = []
            current_line_items = []
            current_len = 0
            
            for text in raw_texts:
                item_len = len(text)
                sep_len = 1 if current_line_items else 0
                
                if (current_len + sep_len + item_len <= num1) or (not current_line_items):
                    current_line_items.append(text)
                    current_len += (sep_len + item_len)
                else:
                    lines.append("、".join(current_line_items))
                    current_line_items = [text]
                    current_len = item_len
            
            if current_line_items:
                lines.append("、".join(current_line_items))
            
            # 孤儿行合并
            if len(lines) > 1:
                last_line = lines[-1]
                if len(last_line) < num2:
                    final_name = "、".join(raw_texts) # 强制单行
                else:
                    final_name = lines # 列表形式
            else:
                final_name = lines[0]

        # --- E. 写入图签 ---
        current_num = i + start_index
        dwg_no = f"{target_prefix}-{current_num:02d}"
        
        keys = ["图纸名称", "图纸编号", "图纸规格", "出图比例"]
        vals = [final_name, dwg_no, size_val, scale_val]
        
        print_name = str(final_name) if isinstance(final_name, str) else str(final_name)
        sys_logger.info(f"   ✍ [{i:02d}] {dwg_no} | {print_name}")
        
        if "set_attribute_mtext" in globals():
            # 假设 set_attribute_mtext 已经处理了 block 写入逻辑
            # 注意：block 对象必须是有效的 COM 对象，且要在当前上下文中可写
            set_attribute_mtext(block, keys, vals, keep_prefix=True, verbose=False)
            success_count += 1

    sys_logger.info(f"✅ 处理完成，共更新 {success_count} 个图签")

    # 仅在有操作时保存，防止空跑
    if success_count > 0:
        if 'wait_command_done' in globals(): wait_command_done()
        save_file()

    return True

#&&&% ===（七）从Excel写入Cad ===

#&&% 自动写入

def auto_import_excel_to_cad(
    layout_name=None, 
    operate_target="Model", 
    select_config=None, 
    excel_path=None,  # 允许外部强制指定 Excel 路径，否则自动推断
    start_index=1,
    use_cache=False,
    verbose=1
):
    """
    【函数编号】: XLS-AUTO-002 (V2.0 - 完整标准版)
    【所属模块】: 数据导入模块 (Data Import)
    【功能】: 
        智能导入调度器。
        根据空间样式（纯模型/混合/布局）自动从 Excel 读取数据并逆向更新到 CAD 图签。
    
    【逻辑流程】:
        1. 策略推断: 根据 operate_target 和 layout_name 自动判定 CTQ 提取策略。
        2. 目标锁定: 调用 smart_rebuild_print_info 获取内存中的图块对象 (这是我们要更新的目标)。
        3. 路径匹配: 根据样式决定去读取哪个 Excel 文件。
           - 纯模型 (Model, None)     -> 读取 "文件名.xlsx"
           - 混合模型 (Model, "A3")   -> 读取 "文件名.xlsx" (核心：混合模式Excel不带后缀)
           - 布局空间 (Layout, "A3")  -> 读取 "文件名_A3.xlsx"
        4. 执行更新: 调用底层 read_excel_and_update_cad_titleblocks 完成写入。

    【参数】:
        - layout_name: 布局名称 (混合模型或布局空间需提供)。
        - operate_target: "Model" 或 "Layout"。
        - select_config: 0或1。None时自动推断。
        - excel_path: 如果不为None，强制读取此文件，忽略自动命名规则。
    """
    
    # =========================================================================
    # 1. 智能推断 select_config (CTQ 提取策略)
    # =========================================================================
    # 逻辑必须与 auto_export_excel_by_style 保持 100% 一致
    if select_config is None:
        if operate_target == "Model":
            # 纯模型(无layout_name) -> 0 (大坐标/框选)
            # 混合模型(有layout_name) -> 1 (1:1/视口)
            select_config = 1 if layout_name else 0
        else:
            # 布局空间 -> 1 (1:1)
            select_config = 1

    if verbose >= 1:
        sys_logger.info(f"🚀 [自动导入] 目标: {operate_target} | 布局: {layout_name} | 策略: {select_config}")

    # =========================================================================
    # 2. 获取目标数据 (调用 smart_rebuild_print_info)
    # =========================================================================
    # 我们必须先扫描图纸，拿到内存中的 BlockReference 对象列表，才能往里写数据
    
    # 接收完整的三元组返回值
    polys, blocks, mapping = smart_rebuild_print_info(
        layout_name=layout_name,
        operate_target=operate_target,
        select_config=select_config,
        use_cache=use_cache,
        verbose=verbose
    )

    # 简单校验
    if not blocks:
        if verbose >= 1:
            print("⚠️ [自动导入] 未提取到任何图框块，无需更新。")
        return False

    # =========================================================================
    # 3. 构造 CTQ 数据包 (严格三元组结构)
    # =========================================================================
    # 包含: (打印线列表, 图块列表, 映射字典)
    ctq_data = (polys, blocks, mapping)

    # =========================================================================
    # 4. 决定 Excel 文件名查找规则 (Mapping Logic)
    # =========================================================================
    # 这里的 import_layout_arg 是传给底层 read_excel_... 函数用来拼接文件名的
    
    import_layout_arg = None 
    
    if operate_target == "Layout":
        # 【纯布局模式】
        # 告诉底层函数去寻找 "文件名_布局名.xlsx"
        # 并且底层函数会尝试切换 active layout 以确保环境一致
        import_layout_arg = layout_name
        
    elif operate_target == "Model":
        # 【纯模型 OR 混合模型】
        # 告诉底层函数去寻找 "文件名.xlsx" (即 import_layout_arg=None)
        # 即使是混合模式 (operate_target="Model", layout_name="A3")，
        # 我们依然是在 Model 空间更新，且 Excel 文件通常是主文件。
        import_layout_arg = None

    # =========================================================================
    # 5. 执行底层导入
    # =========================================================================
    return read_excel_and_update_cad_titleblocks(
        ctq=ctq_data,
        layout_target=import_layout_arg, # 决定 Excel 文件名后缀逻辑
        excel_path=excel_path,           # 允许外部强制指定
        start_index=start_index
    )







#&&% 基础写入

def read_excel_and_update_cad_titleblocks(
    ctq, 
    layout_target=None, # 模型空间传None, 布局传具体名称)
    excel_path=None, 
    start_index=1 
):
    """
    【函数编号】: XLS-002 (V3.1 - 双空间适配版)
    【功能描述】: 
        从 Excel 读取项目信息逆向更新到 CAD。
        支持模型空间和图纸空间两种模式。
        
    【路径逻辑】:
        若 excel_path 为 None，则自动推断：
        - Model模式:  {当前图纸路径}/{文件名}.xlsx
        - Layout模式: {当前图纸路径}/{文件名}_{布局名}.xlsx
    """
    import re

    from CAD_file_operations import save_file

    # ================= 0. 环境与依赖检查 =================

    current_doc = C.doc

    try: doc_name = Path(current_doc.Name).stem
    except: doc_name = "未命名图纸"

    # --- 0.5 切换空间 (新增逻辑) ---
    if layout_target:
        try:
            sys_logger.info(f"📂 切换至布局: {layout_target} (以确保环境一致)")
            current_doc.ActiveLayout = current_doc.Layouts.Item(layout_target)
            current_doc.MSpace = False
        except Exception as e:
            sys_logger.info(f"❌ 无法切换到布局 '{layout_target}': {e}")
            return False

    # ================= 1. 路径智能构建 (关键修改) =================
    if excel_path is None:
        try:
            base_path = Path(current_doc.FullName)
            if layout_target:
                # Layout模式: 寻找 {文件名}_{布局名}.xlsx
                new_name = f"{base_path.stem}_{layout_target}.xlsx"
                excel_path = base_path.parent / new_name
            else:
                # Model模式: 寻找 {文件名}.xlsx
                excel_path = base_path.with_suffix(".xlsx")
                
            sys_logger.info(f"🔍 推断 Excel 路径: {excel_path}")
        except Exception as e:
            sys_logger.info(f"❌ 无法确定路径 (图纸可能未保存): {e}")
            return False
    else:
        excel_path = Path(excel_path)

    # ================= 2. 文件存在性校验 =================
    if not excel_path.exists():
        sys_logger.info(f"❌ 找不到 Excel 文件: {excel_path}")
        print("   (请确认文件已导出，且文件名符合规则)")
        return False

    # ================= 3. 数据校验 =================
    if not isinstance(ctq, (list, tuple)) or len(ctq) < 2:
        print("⚠ 图块数据 (ctq) 格式错误")
        return False
    
    tq_list = ctq[1]
    if not tq_list:
        print("⚠ 未找到有效的图签块")
        return False

    t0 = time.time()
    sys_logger.info(f"📖 正在读取: {excel_path.name} ...")

    # ================= 4. 打开 Excel 处理逻辑 =================
    excel = None; wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False; excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(excel_path))
        
        try: ws = wb.Worksheets(1)
        except: ws = wb.ActiveSheet
        
        # --- A. 读取项目通用信息 (第2行, G列起) ---
        PROJECT_KEYS = [
            "专业名称", "专业代号", "项目名称", "子项目名称", "建设单位名称", 
            "设计阶段", "版本号", "出图日期", "设计编号", "设计院名称",
            "法人代表", "注册建筑师", "注册结构师"
        ]
        
        project_data = {}
        # G列是第7列
        for idx, key in enumerate(PROJECT_KEYS):
            val = ws.Cells(2, 7 + idx).Value 
            project_data[key] = str(val).strip() if val is not None else ""

        global_discipline_code = project_data.get("专业代号", "JZ").strip() or "JZ"

        # --- B. 遍历行读取 ---
        start_row = 2
        success_count = 0
        MULTI_LINE_KEYS = ["图纸名称", "项目名称", "子项目名称", "建设单位名称"]
        
        while True:
            idx_val = ws.Cells(start_row, 1).Value
            if idx_val is None or str(idx_val).strip() == "": break
            
            # 计算对应的 CAD 列表索引
            # Excel序号 = index + start_index
            # index = Excel序号 - start_index
            try: 
                tq_index = int(float(idx_val)) - start_index
            except ValueError: 
                start_row += 1; continue

            if tq_index < 0 or tq_index >= len(tq_list):
                start_row += 1; continue

            # 读取独立属性
            sheet_data = {
                "图纸名称": str(ws.Cells(start_row, 3).Value or "").strip(),
                "图纸规格": str(ws.Cells(start_row, 4).Value or "").strip(),
                "出图比例": str(ws.Cells(start_row, 5).Value or "").strip()
            }

            # 强制重构编号
            current_calc_num = tq_index + start_index
            new_dwg_no = f"{global_discipline_code}-{current_calc_num:02d}"
            sheet_data["图纸编号"] = new_dwg_no

            # 合并数据
            full_attributes = project_data.copy()
            full_attributes.update(sheet_data)

            # 准备写入
            final_keys = []
            final_values = []
            for k, v in full_attributes.items():
                final_keys.append(k)
                if k in MULTI_LINE_KEYS and v:
                    # 分割换行符 // 或 \\ 或 \n
                    parts = re.split(r'//|\\\\|\n', v)
                    clean_parts = [p.strip() for p in parts if p.strip()]
                    final_values.append(clean_parts if len(clean_parts) > 1 else v)
                else:
                    final_values.append(v)

            # 写入 CAD

            block_target = tq_list[tq_index]
            
            # 兼容性检查：如果是布局模式，确保 block 属于当前布局
            # (ctq生成时已经过滤好了，这里只要不报错即可)
            try:
                set_attribute_mtext(block_target, final_keys, final_values, keep_prefix=True, verbose=False)
                sys_logger.info(f"   ✅ 更新 [{tq_index}]: {new_dwg_no}")
                success_count += 1
            except Exception as e:
                sys_logger.info(f"   ⚠️ 更新失败 [{tq_index}]: {e}")
        
            start_row += 1

    except Exception as e:
        import traceback; traceback.print_exc()
        sys_logger.info(f"❌ 错误: {e}")
        return False
    finally:
        if wb: wb.Close(SaveChanges=0)
        if excel: excel.Quit()
        del wb, excel

    sys_logger.info(f"✅ 完成! 更新 {success_count} 个，耗时 {time.time() - t0:.2f}s")

    wait_command_done()

    save_file()

    return True





#&&% 自动_修改图块标签属性

def auto_update_titleblock_format_by_style(
    att_config,       # 【核心参数】格式配置字典
    layout_name=None, 
    operate_target="Model", 
    select_config=None, 
    use_cache=False,
    verbose=1
):
    """
    【函数编号】: XLS-AUTO-004 (V1.0 - 格式刷)
    【所属模块】: 自动化处理模块 (Auto Processing)
    【功能】: 
        智能格式调整调度器。
        根据空间样式自动提取图签块，并批量应用格式配置（高度、样式、对齐等）。
    
    【核心逻辑】:
        1. 定位: 利用 smart_rebuild_print_info 找到当前样式下的所有有效图签块。
           (这一步至关重要，因为它能过滤掉非目标图签，比如模型空间里的垃圾块)
        2. 执行: 将提取到的块列表传递给 batch_update_block_attributes_config。
        3. 同步: 底层函数会自动修改块定义(Block Definition)并同步实例(ATTSYNC)。

    【参数】:
        - att_config: 格式配置字典 (见 batch_update_block_attributes_config 说明)。
        - layout_name, operate_target: 空间样式定位参数。
    """

    # =========================================================================
    # 1. 智能推断 select_config (CTQ 提取策略)
    # =========================================================================
    # 保持 AUTO 系列的一致性
    if select_config is None:
        if operate_target == "Model":
            select_config = 1 if layout_name else 0
        else:
            select_config = 1

    if verbose >= 1:
        sys_logger.info(f"🚀 [自动格式刷] 目标: {operate_target} | 布局: {layout_name} | 策略: {select_config}")

    # =========================================================================
    # 2. 获取目标数据 (调用 smart_rebuild_print_info)
    # =========================================================================
    # 我们需要获取 blocks 列表，以便让底层函数知道要去修改哪些块定义。
    # 虽然修改的是定义(Definition)，但我们需要通过实例(Reference)来反查定义名。
    
    polys, blocks, mapping = smart_rebuild_print_info(
        layout_name=layout_name,
        operate_target=operate_target,
        select_config=select_config,
        use_cache=use_cache,
        verbose=verbose
    )

    if not blocks:
        if verbose >= 1: print("⚠️ [自动格式刷] 未找到任何图签块，无法调整格式。")
        return False

    # =========================================================================
    # 3. 构造 CTQ 数据包
    # =========================================================================
    ctq_data = (polys, blocks, mapping)

    # =========================================================================
    # 4. 执行批量更新
    # =========================================================================
    # 直接调用您提供的底层函数
    return batch_update_block_attributes_config(
        ctq=ctq_data,
        att_config=att_config
    )



#&&% 基础_修改图块标签属性

def batch_update_block_attributes_config(
    ctq, 
    att_config
):
    """
    【函数编号】: BLK-008-Smart
    【功能描述】: 
        批量调整图签格式，支持“字典模式”精准指定修改项。
        
    【参数 att_config 写法】:
        {
            "标签名": { "height": 3.5, "justify": 4 },  <-- 推荐：只写要改的
            "标签名2": (3.5, 0.7, 0, 4, 120)            <-- 兼容：旧式全写
        }
        
    【支持的字典 Key】:
        - style (样式)
        - height (高度)
        - width_factor (宽度因子)
        - rotation_deg (旋转角度)
        - justify (对正 0-14)
        - boundary_width (MText边界宽)


    config = {
            "图纸名称": { 
                "height": 600, 
                "style": "Standard"
            },
            "设计阶段": { 
                "height": 600, 
                "style": "Standard"
            }
        }
    
    batch_update_block_attributes_config(ctq, config)

    20251217

    """
    from CAD_file_operations import save_file

    # ================= 1. 基础校验 =================

    C.li()
    if not ctq or len(ctq) < 2: return False
    blocks = ctq[1]
    if not blocks: return False
    if not att_config: return False

    sys_logger.info(f"🚀 [智能模式] 开始批量调整格式...")

    # ================= 2. 提取唯一的块名 =================
    unique_block_names = set()
    name_to_obj_map = {} 

    for blk in blocks:
        try:
            # 1. 优先获取 EffectiveName (动态块)
            name = get_attr(blk, 'EffectiveName')
            if not name: 
                # 2. 回退到 Name (普通块)
                name = get_attr(blk, 'Name')
            
            # 3. 记录映射关系
            if name and (name not in unique_block_names):
                unique_block_names.add(name)
                # 【关键】保存这个活的实例对象，供最后一步 ATTSYNC 使用
                name_to_obj_map[name] = blk 
        except Exception:
            # 如果某个块对象损坏或 RPC 失败，跳过它，不要中断整个流程
            continue

    sys_logger.info(f"🔍 涉及 {len(unique_block_names)} 种块定义")

    # ================= 3. 循环处理 =================
    global_success = True
    
    for blk_name in unique_block_names:
        sys_logger.info(f"\n------ 处理块定义: [{blk_name}] ------")
        
        def_modified = False
        
        for tag, params in att_config.items():
            
            # --- 核心修改：参数解析逻辑 ---
            kwargs = {}
            
            # 情况A：用户传入的是字典（推荐） { "height":3.5 }
            if isinstance(params, dict):
                kwargs = params # 直接由底层函数去匹配参数名
                
            # 情况B：用户传入的是元组（旧版） (3.5, 0.7, ...)
            elif isinstance(params, (list, tuple)) and len(params) >= 5:
                kwargs = {
                    'height': params[0],
                    'width_factor': params[1],
                    'rotation_deg': params[2],
                    'justify': params[3],
                    'boundary_width': params[4]
                }
            else:
                continue # 格式不对跳过

            # --- 调用底层修改函数 ---
            # 使用 **kwargs 将字典拆解为参数
            # 例如：update_block_def_attributes_safe(..., height=3.5)
            res = update_block_def_attributes_safe(
                blk_name, 
                tag, 
                verbose=True,
                **kwargs 
            )
            
            if res: def_modified = True

        # --- 执行同步 ---
        if def_modified:
            print("⏳ 缓冲中 (Wait)...")
            wait_command_done() 

            representative_blk = name_to_obj_map.get(blk_name)
            if representative_blk:
                if attsync_block_instance(representative_blk):
                    sys_logger.info(f"✅ [{blk_name}] 同步指令已发送")
                else:
                    sys_logger.info(f"❌ [{blk_name}] 同步发送失败")
            
            wait_command_done()
        else:
            sys_logger.info(f"⚪ [{blk_name}] 无变化，跳过同步。")

    print("\n" + "="*50)
    print("✅ 批量格式调整完成")


    wait_command_done()
    save_file()
    

    return global_success


#&&&% ===（八）目录文件初步生成 ===


#&&% 统一编目录1

timeit
def bianmulu_func1_h(
    layout_name=None, 
    operate_target="Model", 
    mulu_xuhao=1, 
    select_config=None, 
    use_cache=False, 
    verbose=1
):
    """
    【函数编号】: CAT-UNIFIED-001 (V9.0 - 严格命名版)
    【功能描述】: 
        编目录 Step 1: 扫描当前图纸，生成独立的目录 DWG 文件。
    
    【修改说明】: 
        1. 实施 "唯一互逆性" 命名规则: 只要 layout_name 存在，必须生成带后缀的目录文件。
        2. 确保混合模型空间 (Model + layout_name) 生成独立文件，不覆盖主模型目录。
    """
    from CAD_file_operations import new_file, close_file, open_file, save_file, copy_file_content_pywin32,switch_to_layout
    from pathlib import Path
    import os
    import time

    # ================= 1. 智能策略推断 =================
    if select_config is None:
        if operate_target == "Model":
            select_config = 1 if layout_name else 0
        else:
            select_config = 1

    if verbose >= 1:
        sys_logger.info(f"\n🚀 [编目录 Step 1] 生成文件 | 目标: {operate_target} | 布局: {layout_name}")

    # ================= 2. 获取核心数据 (扫描图框) =================
    # 这一步决定了我们需要生成多少行目录
    polys, blocks, mapping = smart_rebuild_print_info(
        layout_name=layout_name,
        operate_target=operate_target,
        select_config=select_config,
        use_cache=use_cache,
        verbose=verbose
    )

    if not polys:
        print("❌ [中断] 未提取到任何有效图框，无法生成目录。")
        return False

    src_drawing_count = len(polys)
    sys_logger.info(f"📊 扫描完成: 检测到 {src_drawing_count} 张图纸")

    # ================= 3. 路径推导 (🟢 严格唯一模式) =================
    doc = C.doc
    try:
        src_full_path = doc.FullName
        if not src_full_path:
            print("❌ 当前文档未保存，无法生成目录。")
            return False
            
        src_path_obj = Path(src_full_path)
        current_stem = src_path_obj.stem
        parent_dir = src_path_obj.parent

        # 🟢 [修改点] 严格后缀逻辑
        # 只要 layout_name 有值，文件名就必须带后缀，与 Excel 导出逻辑保持 100% 一致
        name_suffix = f"_{layout_name}" if layout_name else ""
            
        mulu_dwg_name = f"{current_stem}{name_suffix}_目录.dwg"
        mulu_dwg_path = str(parent_dir / mulu_dwg_name)
        
    except Exception as e:
        sys_logger.info(f"❌ 路径计算失败: {e}")
        return False

    # ================= 4. 创建目录文件 =================
    sys_logger.info(f"📂 [新建] 准备创建: {mulu_dwg_name}")
    
    # 如果文件已在 CAD 中打开，先强制关闭（不保存）
    try: C.acad.Documents.Item(mulu_dwg_name).Close(False)
    except: pass
    
    # 创建新文件
    new_file(mulu_dwg_path, close_after=False) 
    wait_command_done()
    
    # 暂时关闭新文件以进行模板注入 (CopyFile 需要文件未被占用)
    print("🔒 暂时关闭文件以进行模板注入...")
    try: 
        doc_mulu = C.acad.Documents.Item(mulu_dwg_name)
        doc_mulu.Close(True)
    except: pass
    time.sleep(1.0) 

    # ================= 5. 插入模板 (后台注入) =================
    userpath = os.environ.get('USERPATH')
    if not userpath: 
        print("❌ 环境变量 USERPATH 未设置。")
        return False
        
    config_mulu_path = os.path.join(userpath, "标准模板", f"目录模板{mulu_xuhao}.dwg")
    
    sys_logger.info(f"🔨 [注入] 正在后台插入模板...")
    copy_success = False
    
    # 增加重试机制，防止文件系统占用延迟
    for attempt in range(1, 4):
        try:
            if os.path.exists(config_mulu_path):
                copy_success = copy_file_content_pywin32(config_mulu_path, mulu_dwg_path)
                if copy_success:
                    print("✅ 模板插入成功。")
                    break 
            else:
                sys_logger.info(f"❌ 模板文件缺失: {config_mulu_path}")
                break
                
            sys_logger.info(f"⏳ 插入重试 {attempt}/3...")
            time.sleep(1.5)
        except Exception as e:
            sys_logger.info(f"   出错: {e}")
            time.sleep(1.5)
    
    if not copy_success:
        print("❌ 模板插入失败，流程终止。")
        # 恢复打开源文件
        open_file(src_full_path)
        return False

    # ================= 6. 剪裁页面 (Page Trimming) =================
    sys_logger.info(f"✂️ [剪裁] 重新打开并整理页面...")
    if not open_file(mulu_dwg_path): return False
    wait_command_done()
    
    try: C.doc.SetVariable("TILEMODE", 1)
    except: pass
    
    # --- 剪裁逻辑 (保留原有的临界值判断) ---
    paper_n = 2 
    # 根据模板序号和图纸数量判断是否需要保留第2页
    if mulu_xuhao == 1 and src_drawing_count <= 58: paper_n = 1
    elif mulu_xuhao == 2 and src_drawing_count <= 58: paper_n = 1
    elif mulu_xuhao == 3 and src_drawing_count <= 28: paper_n = 1            
    elif mulu_xuhao == 4 and src_drawing_count <= 28: paper_n = 1 
    
    sys_logger.info(f"🧮 页面计算: 图纸{src_drawing_count}张 -> 保留{paper_n}页")

    # 如果只需要1页，删除第2页的框线及内容
    if paper_n == 1:
        try:
            # 扫描 mulu_zhuanyong 图层获取页面边框
            dy_res = select_maxrect_polylines_1(
                layer_name = "mulu_zhuanyong",
                precision_mode = False,   
                width = 0.0,             
                color = 256,                 
                min_side = 100.0,         
            )
            
            mulu_polys = dy_res[0] if dy_res else []
            
            # 如果检测到超过1个框，说明有多页
            if len(mulu_polys) > 1:
                # 假设第二个框就是第二页 (通常按创建顺序或坐标排序)
                # 这里简单取第二个，建议结合坐标排序更稳健，但保持原逻辑不变
                target_frame = mulu_polys[1] 
                min_p, max_p = target_frame.GetBoundingBox()
                
                if 'select_entities_in_window' in globals():
                    # 窗选删除第2页范围内的所有物体
                    lb = select_entities_in_window(min_p[0], min_p[1], max_p[0], max_p[1], ty=0.5, select_mode="_W")
                    if lb:
                        for obj in lb:
                            try: obj.Delete()
                            except: pass
                        sys_logger.info(f"   ✅ 已移除第2页及其内容。")
        except Exception as e:
            sys_logger.info(f"⚠️ 剪裁过程出错: {e}")

    # ================= 7. 收尾与保存 =================
    save_file()
    sys_logger.info(f"✅ [完成] 目录文件已生成: {mulu_dwg_name}")
    time.sleep(0.5)

    try: C.acad.ActiveDocument.Close(True)
    except: pass
    
    sys_logger.info(f"📂 [恢复] 切回源文件...")
    open_file(src_full_path)
    
    # 智能恢复视图
    if layout_name and operate_target == "Layout":
        if 'switch_to_layout' in globals():
            switch_to_layout(layout_name)
    elif operate_target == "Model":
        try: C.doc.SetVariable("TILEMODE", 1)
        except: pass
        
    return True

#&&&% ===（九）目录文件深化 ===

#&&% 编目录2_统一

@timeit
def bianmulu_func2_h(
    layout_name=None, 
    operate_target="Model", 
    select_config=None, 
    use_cache=False,
    verbose=1
):
    """
    【函数编号】: CAT-UNIFIED-002 (V9.0 - 严格寻址版)
    【功能描述】: 
        编目录 Step 2: 填写图纸列表（目录图签）。
    
    【修改说明】: 
        1. 废弃所有模糊匹配逻辑，采用 "唯一互逆性" 命名规则。
        2. 严格锁定 Excel 数据源，防止读取错误的项目信息。
    """
    from CAD_file_operations import open_file, close_file, save_file,switch_to_layout
    import time
    from pathlib import Path
    import os

    if verbose >= 1:
        sys_logger.info(f"\n🚀 [编目录 Step 2] 填写列表 | 目标: {operate_target} | 缓存: {use_cache}")

    # ================= 1. 获取主文件信息 =================
    doc = C.doc
    try:
        main_dwg_path = doc.FullName
        if not main_dwg_path:
            print("❌ 主文件未保存，无法定位目录文件。")
            return False
        
        src_path_obj = Path(main_dwg_path)
        current_stem = src_path_obj.stem
        parent_dir = src_path_obj.parent
        
    except Exception as e:
        sys_logger.info(f"❌ 获取主文件信息失败: {e}")
        return False

    # ================= 2. 推导路径 (严格唯一模式) =================
    try:
        # 🟢 [修改点 1] 严格后缀逻辑 (必须与 Export/Step1 保持一致)
        name_suffix = f"_{layout_name}" if layout_name else ""
            
        # A. 锁定的目录文件名
        mulu_dwg_name = f"{current_stem}{name_suffix}_目录.dwg"
        mulu_dwg_path = str(parent_dir / mulu_dwg_name)
        
        # B. 锁定的 Excel 文件名
        target_excel_name = f"{current_stem}{name_suffix}.xlsx"
        target_excel_path_strict = str(parent_dir / target_excel_name)
        
        # C. 存在性检查
        if not os.path.exists(mulu_dwg_path):
            sys_logger.info(f"❌ [错误] 目录文件不存在: {mulu_dwg_name}")
            sys_logger.info(f"   (请先运行 Step 1 生成目录文件)")
            return False
            
        # D. Excel 锁定 (找不到就置空，绝不找备胎)
        real_excel_path = None
        if os.path.exists(target_excel_path_strict):
            real_excel_path = target_excel_path_strict
            sys_logger.info(f"🎯 锁定 Excel 数据源: {target_excel_name}")
        else:
            sys_logger.info(f"⚠️ [警告] 未找到专属 Excel: {target_excel_name}")
            sys_logger.info(f"   (将跳过数据填写，仅处理图签结构)")
            # 注意：此处不返回 False，因为 Step 2 还肩负着生成图签的任务，
            # 但后续填写步骤会被跳过。
            
    except Exception as e:
        sys_logger.info(f"❌ 路径推导错误: {e}")
        return False

    # ================= 3. 净室启动 =================
    sys_logger.info(f"🧹 [隔离] 切换至目录文件: {mulu_dwg_name}")
    
    if not open_file(mulu_dwg_path):
        print("❌ 无法打开目录文件。")
        return False
    wait_command_done()
    
    # 强制关闭主文件及其他无关文件
    target_name_lower = Path(mulu_dwg_path).name.lower()
    try:
        docs_to_close = [d for d in C.acad.Documents if d.Name.lower() != target_name_lower]
        if docs_to_close:
            sys_logger.info(f"   👋 关闭 {len(docs_to_close)} 个背景文件...")
            for d in docs_to_close:
                try: d.Close(True)
                except: pass
    except: pass

    # 验证环境
    if C.acad.ActiveDocument.Name.lower() != target_name_lower:
        sys_logger.info(f"❌ 环境隔离失败，终止操作。")
        open_file(main_dwg_path)
        return False

    current_doc = C.acad.ActiveDocument

    # ================= 4. 核心业务: 插入与映射 =================
    
    # 4.1 检查图签是否存在 (原版逻辑：查 dy_quyu_H_m)
    lb_tu = stc("dy_quyu_H_m") 
    lb_tu_kuai = [e for e in lb_tu if e.ObjectName=="AcDbBlockReference"] if lb_tu else []

    force_rebuild = False 

    if len(lb_tu_kuai) > 0:
        sys_logger.info(f"✅ 检测到已有图签 ({len(lb_tu_kuai)} 个)，准备填写。")
    else:
        print("🔄 未检测到图签，执行【线条->图签】转化...")
        
        # A. 复制线条 (从 mulu_zhuanyong -> dy_zhuanyong)
        try:
            if not stc("dy_zhuanyong"):
                mulu_objs = stc("mulu_zhuanyong")
                if mulu_objs:
                    try: current_doc.Layers.Add("dy_zhuanyong")
                    except: pass
                    
                    count_copy = 0
                    for o in mulu_objs: 
                        try: 
                            new_o = o.Copy(); new_o.Layer="dy_zhuanyong"
                            count_copy += 1
                        except: pass
                    sys_logger.info(f"   已从模板层复制 {count_copy} 个框线。")
        except: pass

        # B. 识别与插入
        dy_list = []
        try:
            # 获取绘制打印多段线
            raw_ret = select_maxrect_polylines_1(
                layer_name = "mulu_zhuanyong_1",
                precision_mode = False,  
                width = 0.0,            
                color = 3,                 
            )
            dy_list = raw_ret[0] if raw_ret else []
        except: pass

        if dy_list:
            if 'insert_and_scale_labels_area_power' in globals():
                # 插入图签 (这步操作会将块炸开处理并放入 dy_quyu)
                insert_and_scale_labels_area_power(
                    dy_list, layername="dy_quyu", 
                    debug=False
                )
                print("✅ 插入完成，等待数据库刷新...")
                time.sleep(3)

                # ================= 修复与图层迁移逻辑 =================
                
                # 1. 确保辅助层存在
                ensure_layer("测试辅助")
                time.sleep(1)
                C.doc.Regen(1)
                
                # 2. 修复线宽
                smart_repair_frame_polyline_widths_m(verbose=True)

                # 3. 将 Line 合并为 Polyline (tuqian_neibu_pl)
                lines_to_polylines(Lc=None, tol=0.5, layer_name="mulu_zhuanyong_1", width=0.0, color=256)
                
                # 4. 【图层迁移】 tuqian_neibu_pl -> tuqian_neibu_pl_m
                objs_created = select_tuceng("tuqian_neibu_pl")
                if objs_created:
                    sys_logger.info(f"🚚 [图层迁移] 正在迁移 {len(objs_created)} 个对象...")
                    for o in objs_created:
                        try:
                            o.Layer = "tuqian_neibu_pl_m"
                            o.Color = 256 # 随层
                        except: pass
                else:
                    sys_logger.info("⚠️ 警告: 未检测到生成的边界线对象！")

                time.sleep(1) 
                try: C.doc.Regen(1)
                except: pass
                
                force_rebuild = True
            else:
                print("❌ 缺少插入函数 insert_and_scale_labels_area_power，无法自动修复。")
        else:
            print("❌ 错误：目录文件为空或无法识别框线 (检查 mulu_zhuanyong_1 图层)。")
            open_file(main_dwg_path)
            return False
        
        wait_command_done()

    # 4.2 重建映射
    # -------------------------------------------------
    final_ctq = None
    
    if force_rebuild or not final_ctq:
        sys_logger.info(f"🔍 [Rebuild] 扫描目录文件结构...")
        
        protect_layer = "mulu_zhuanyong"
        try: current_doc.Layers.Item(protect_layer).Lock = True
        except: pass

        final_ctq = rebuild_print_area_title_mapping(
            core_layer="dy_quyu_H",             
            final_poly_layer="tuqian_neibu_pl_m", 
        )
        
        try: current_doc.Layers.Item(protect_layer).Lock = False
        except: pass

    if not (final_ctq and len(final_ctq) >= 3 and len(final_ctq[2]) > 0):
        print("❌ 目录文件映射无效，无法填写。")
        open_file(main_dwg_path)
        return False

    # ================= 5. 数据填写 (严格匹配) =================
    if real_excel_path:
        sys_logger.info(f"📖 读取数据源: {Path(real_excel_path).name}")
        try:
            if 'update_catalog_titleblocks_from_excel' in globals():
                # 传入完整三元组
                update_catalog_titleblocks_from_excel(
                    ctq=final_ctq, 
                    excel_path=real_excel_path,
                    catalog_name="图纸目录",
                    custom_suffixes=None
                )
                print("✅ 数据填写完成。")
                current_doc.Regen(1)
            else:
                print("❌ 缺少填写函数 update_catalog_titleblocks_from_excel")
        except Exception as e:
            sys_logger.info(f"❌ 填写过程异常: {e}")
    else:
        # 🟢 [修改点] 严格模式下，如果没找到专属 Excel，直接跳过，不报 Fatal Error 但也不填乱数据
        sys_logger.info(f"⏭️ 跳过填写: 未找到匹配的 Excel 文件。")

    # ================= 6. 保存与恢复 =================
    sys_logger.info(f"💾 保存目录文件...")
    save_file()
    
    try: current_doc.Close(False)
    except: pass
    
    sys_logger.info(f"🔙 恢复主文件: {Path(main_dwg_path).name}")
    open_file(main_dwg_path)
    
    # 智能恢复视图
    if layout_name and operate_target == "Layout" and 'switch_to_layout' in globals():
        switch_to_layout(layout_name)
    elif operate_target == "Model":
        try: C.doc.SetVariable("TILEMODE", 1)
        except: pass
        
    return True

#&&&% ===（十）目录编写 ===
#&&% 编好目录

@timeit
def bianmulu_func3_h(
    layout_name=None, 
    operate_target="Model", 
    mubanxuhao=1,
    select_config=None, 
    use_cache=False,
    verbose=1,
    start_index=0, 
    # --- 样式与功能参数 ---
    text_height=250, 
    text_style="宋体", 
    text_alignment="中心",            
    text_rotation=0.0, 
    text_oblique=0.0, 
    text_width_factor=1.0, 
    width_factors_map=None
):
    """
    【函数编号】: CAT-UNIFIED-003 (V9.0 - 严格寻址版)
    【功能描述】: 
        编目录 Step 3: 读取 Excel 数据并填写到目录 DWG 文件中。
    
    【修改说明】: 
        1. 实施严格的 "唯一互逆性" 命名规则。
        2. 废弃 Excel 备选搜索逻辑，找不到指定后缀的 Excel 即视为错误。
    """
    from CAD_file_operations import open_file, close_file, save_file,switch_to_layout
    from pathlib import Path
    import os
    import time
   
    if verbose >= 1:
        sys_logger.info(f"\n🚀 [编目录 Step 3] 填写表格内容 | 目标: {operate_target}")

    # ================= 1. 获取主文件信息 =================
    doc = C.doc
    try:
        main_dwg_path = doc.FullName
        if not main_dwg_path:
            print("❌ 主文件未保存，无法定位。")
            return False
        
        src_path_obj = Path(main_dwg_path)
        current_stem = src_path_obj.stem
        parent_dir = src_path_obj.parent
        
    except Exception as e:
        sys_logger.info(f"❌ 获取主文件信息失败: {e}")
        return False

    # ================= 2. 推导路径 (严格唯一模式) =================
    try:
        # 🟢 [修改点] 严格后缀逻辑
        name_suffix = f"_{layout_name}" if layout_name else ""
            
        # 1. 锁定的目录文件名
        mulu_dwg_name = f"{current_stem}{name_suffix}_目录.dwg"
        mulu_dwg_path = str(parent_dir / mulu_dwg_name)
        
        # 2. 锁定的 Excel 文件名
        target_excel_name = f"{current_stem}{name_suffix}.xlsx"
        target_excel_path = str(parent_dir / target_excel_name)
        
        # 3. 严格校验 (不许猜测)
        if not os.path.exists(mulu_dwg_path):
            sys_logger.info(f"❌ [错误] 目录文件不存在: {mulu_dwg_name}")
            sys_logger.info(f"   (请确认 Step 1/2 已成功执行)")
            return False
            
        if not os.path.exists(target_excel_path):
            sys_logger.info(f"❌ [错误] 找不到匹配的 Excel: {target_excel_name}")
            sys_logger.info(f"   (规则: layout_name='{layout_name}' -> 必须存在带后缀的Excel)")
            return False
            
    except Exception as e:
        sys_logger.info(f"❌ 路径推导错误: {e}")
        return False

    # ================= 3. 净室启动 (隔离环境) =================
    try: 
        sys_logger.info(f"🧹 [隔离] 切换至目录文件: {mulu_dwg_name}")
        open_file(mulu_dwg_path)
        wait_command_done()
        
        # 关闭其他文档 (防止干扰)
        target_name_lower = Path(mulu_dwg_path).name.lower()
        try:
            docs_to_close = [d for d in C.acad.Documents if d.Name.lower() != target_name_lower]
            if docs_to_close:
                sys_logger.info(f"   👋 关闭 {len(docs_to_close)} 个背景文件...")
                for d in docs_to_close:
                    try: d.Close(True) # 保存并关闭
                    except: pass
        except: pass
        
        # 二次确认当前文档
        doc = C.doc
        if not doc or doc.Name.lower() != target_name_lower:
            sys_logger.info(f"❌ 净室建立失败，当前文档: {doc.Name if doc else 'None'}")
            return False

        # ================= 4. 扫描目录结构 =================
        print("🔍 正在扫描目录文件布局...")
        
        # 重建映射关系 (扫描 dy_quyu_H -> tuqian_neibu_pl)
        ctq_mulu_local = rebuild_print_area_title_mapping(
            core_layer="dy_quyu_H",
            final_poly_layer="tuqian_neibu_pl", 
            protected_layers=["mulu_zhuanyong", "dy_quyu"]
        )

        if not ctq_mulu_local or len(ctq_mulu_local) < 3 or not ctq_mulu_local[2]:
            print("❌ 错误：未能识别到有效的目录图签结构。")
            return False

        # ================= 5. 执行填表 =================
        sys_logger.info(f"✍️ 正在写入数据 (源: {Path(target_excel_path).name})...")
        
        if 'write_catalog_from_excel_to_cad' in globals():
            write_catalog_from_excel_to_cad(
                ctq_mulu_local,            # 位置参数1: CTQ数据
                data_excel_path=target_excel_path, 
                mubanxuhao=mubanxuhao,
                start_index=start_index,   # 传递序号偏移
                text_height=text_height,
                text_style=text_style,
                text_alignment=text_alignment,
                text_rotation=text_rotation,
                text_oblique=text_oblique,
                text_width_factor=text_width_factor,
                width_factors_map=width_factors_map
            )
            
            print("✅ 目录表格填写完成！")
            wait_command_done()
            
            # 保存
            save_file()
            sys_logger.info(f"💾 目录文件已保存。")
            
            # 自动关闭 (save_option="auto_save" 通常意味着不再次弹窗)
            close_file(save_option="auto_save") 
            return True
        else:
            print("❌ 致命错误: 缺少 write_catalog_from_excel_to_cad 函数")
            return False

    except Exception as e:
        sys_logger.info(f"⚠️ [CAT-GEN-H] 运行时发生异常: {e}")
        import traceback
        traceback.print_exc()
        return False

    finally:
        # ================= 6. 现场恢复 =================
        sys_logger.info(f"🔙 正在恢复主文件环境: {Path(main_dwg_path).name}")
        try:
            if os.path.exists(main_dwg_path):
                open_file(main_dwg_path)
                # 尝试恢复之前的视图状态 (可选)
                if operate_target == "Layout" and layout_name and 'switch_to_layout' in globals():
                    switch_to_layout(layout_name)
            else:
                sys_logger.info(f"❌ 无法恢复主文件，路径未找到: {main_dwg_path}")
        except Exception as e:
            sys_logger.info(f"⚠️ 恢复主文件时出错: {e}")

#&&&% ===（十一）合并目录文件到主文件 ===
#&&% 合并

@timeit
def bianmulu_func4_h(
    layout_name=None, 
    operate_target="Model", 
    select_config=None, 
    verbose=1,
    tol=0.01
):
    """
    【函数编号】: CAT-UNIFIED-004 (V9.8 - 规范展开版)
    【修复说明】: 
        1. 逻辑核心: 继承 V9.7 (布局左移3倍宽 + 模型直插)。
        2. 代码格式: 彻底移除所有 "一行多语句" 写法，改为标准缩进，防止语法错误和调试困难。
    """
    from CAD_file_operations import open_file, save_file
    import time
    from pathlib import Path
    import os
    import pythoncom
    import win32com.client

    lvl = int(verbose) if verbose is not None else 1
    def log(msg, min_lvl=1):
        if lvl >= min_lvl: sys_logger.info(f"{'  ' * (min_lvl - 1)}[Catalog-Merge] {msg}")

    log(f"🚀 启动合并 | 目标: {operate_target} | 布局: {layout_name}", min_lvl=1)

    # ================= 1. 路径锁定 =================
    try:
        main_dwg_path = C.doc.FullName
        if not main_dwg_path:
            log("❌ 主文件未保存。", 0)
            return False
        
        p = Path(main_dwg_path) 
        name_suffix = f"_{layout_name}" if layout_name else ""
        mulu_name = f"{p.stem}{name_suffix}_目录.dwg"
        mulu_path = str(p.parent / mulu_name)
        
        if not os.path.exists(mulu_path):
            log(f"❌ 找不到目录文件: {mulu_name}", 0)
            return False
            
    except Exception as e:
        log(f"路径推导错误: {e}", 0)
        return False

    # =================================================================
    #  分支 A: Layout 模式 (剪贴板方案 + 智能定位)
    # =================================================================
    if operate_target == "Layout":
        log("🔧 [模式A] 执行 Layout 剪贴板方案...", 1)
        
        # 1. 准备源数据
        open_file(mulu_path)
        time.sleep(1.0)
        
        res_src = select_maxrect_polylines_1(
            layer_name="dy_zhuanyong", 
            width=0.0, 
            color=1, 
            min_side=100.0
        )
        
        if not res_src or not res_src[0]:
            log("❌ 目录文件内容为空", 0)
            open_file(main_dwg_path)
            return False
            
        min_x, min_y = 1e20, 1e20
        max_x, max_y = -1e20, -1e20
        
        for poly in res_src[0]: 
            mn, mx = poly.GetBoundingBox()
            min_x = min(min_x, mn[0])
            min_y = min(min_y, mn[1])
            max_x = max(max_x, mx[0])
            max_y = max(max_y, mx[1])
        
        src_width = max_x - min_x
        base_pt = f"{min_x},{min_y}"
        
        # 智能缩放
        scale_f = 0.01 if src_width > 5000 else 1.0
        
        try:
            doc_src = C.doc
            if scale_f != 1.0:
                doc_src.SendCommand(f'(command "_.SCALE" "_All" "" "{base_pt}" "{scale_f}")\n')
                time.sleep(1.0)
            
            doc_src.SendCommand(f'(command "_.COPYBASE" "{base_pt}" "_All" "")\n')
            time.sleep(1.5)
            
            doc_src.Close(False) # 不保存关闭
        except:
            open_file(main_dwg_path)
            return False

        # 2. 回到主文件并定位
        log(f"📂 回到主文件...", 1)
        open_file(main_dwg_path)
        time.sleep(1.0)
        
        try:
            C.doc.ActiveLayout = C.doc.Layouts.Item(layout_name)
            C.doc.MSpace = False
            
            insert_x, insert_y = 0.0, 0.0
            
            try:
                # 尝试扫描布局现有图框
                try:
                    res_layout = select_print_areas_paperspace(
                        layout_name, 
                        layer_name="dy_zhuanyong", 
                        precision_mode=False, 
                        width=0.0, 
                        color=256, 
                        min_side=100.0
                    )
                except:
                    res_layout = select_maxrect_polylines_1(
                        layer_name="dy_zhuanyong", 
                        precision_mode=False, 
                        width=0.0, 
                        color=1, 
                        min_side=100.0
                    )

                if res_layout and len(res_layout) > 0:
                    objs = res_layout[0] if isinstance(res_layout, tuple) else res_layout
                    
                    if objs:
                        min_ex = 1e20
                        for poly in objs:
                            mn, _ = poly.GetBoundingBox()
                            if mn[0] < min_ex:
                                min_ex = mn[0]
                        
                        real_width = src_width * scale_f
                        
                        # 🟢 定位逻辑: 插入点 = 最左X - (3.0 * 目录宽) - 50
                        insert_x = min_ex - (real_width * 3.0) - 50.0
                        
                        leftmost_poly = min(objs, key=lambda p: p.GetBoundingBox()[0][0])
                        insert_y = leftmost_poly.GetBoundingBox()[0][1]
                        
                        log(f"📍 智能定位(修正): 左侧偏移3倍宽 ({insert_x:.1f}, {insert_y:.1f})", 1)
                    else:
                        log("⚠️ 布局图框未识别，使用原点。", 1)
                else:
                    log("ℹ️ 布局是空的，使用原点。", 1)
                    
            except Exception as loc_err:
                log(f"⚠️ 定位出错: {loc_err} -> 使用(0,0)", 1)

            # 3. 粘贴
            C.doc.SendCommand(f'(command "_.PASTECLIP" "{insert_x},{insert_y}")\n')
            send_cmd_with_sync("_ZOOM _E")
            return True
            
        except Exception as e:
            log(f"❌ 布局粘贴失败: {e}", 0)
            return False

    # =================================================================
    # 分支 B: Model 模式 (标准缩进版 WBlock 方案)
    # =================================================================
    else:
        log("🔧 [模式B] 执行 Model 直插法 (WBlock Logic)...", 1)
        
        # 1. 打开源文件计算范围
        open_file(mulu_path)
        
        res_src = select_maxrect_polylines_1(
            layer_name="dy_zhuanyong", 
            width=0.0, 
            color=1, 
            min_side=100.0
        )
        
        if not res_src or not res_src[0]:
            log("❌ 源文件为空", 0)
            return False
        
        min_p, max_p = res_src[0][0].GetBoundingBox()
        src_x, src_y = min_p[0], min_p[1]
        src_w = max_p[0] - min_p[0]
        
        # 关闭源文件
        C.doc.Close(False)
        time.sleep(0.5)
        
        # 2. 打开目标文件计算位置
        open_file(main_dwg_path)
        C.doc.SetVariable("TILEMODE", 1)
        
        res_tar = select_maxrect_polylines_1(
            layer_name="dy_zhuanyong", 
            width=0.0, 
            color=1, 
            min_side=100.0
        )
        
        tgt_x, tgt_y = 0, 0
        if res_tar and res_tar[0]:
            leftmost = min(res_tar[0], key=lambda poly: poly.GetBoundingBox()[0][0])
            rmin, rmax = leftmost.GetBoundingBox()
            
            # 模型空间定位逻辑：左侧紧贴插入
            tgt_x = rmin[0] - src_w - (rmax[0]-rmin[0])/4.0
            tgt_y = rmin[1]
            
        insert_x = tgt_x - src_x
        insert_y = tgt_y - src_y
        
        try:
            block_name = Path(mulu_path).stem
            
            # 清理旧块定义
            try:
                C.doc.Blocks.Item(block_name).Delete()
            except:
                pass
            
            # 插入
            pt = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (insert_x, insert_y, 0.0))
            block_ref = C.doc.ModelSpace.InsertBlock(pt, mulu_path, 1.0, 1.0, 1.0, 0.0)
            time.sleep(1.0)
            
            # 炸开
            exploded = False
            for attempt in range(5):
                try:
                    block_ref.Explode()
                    exploded = True
                    break
                except:
                    time.sleep(1.0)
            
            if exploded:
                try:
                    block_ref.Delete()
                except:
                    pass
                    
                try:
                    C.doc.Blocks.Item(block_name).Delete()
                except:
                    pass
                    
                send_cmd_with_sync("_ZOOM _E")
                return True
            else:
                save_file()
                return True
        except Exception as e:
            log(f"❌ 插入过程崩溃: {e}", 0)
            return False








#&&% 写入目录图签 

def update_catalog_titleblocks_from_excel(
    ctq, 
    excel_path=r"D:/Mypro/基础服务/用户1/dwg文件/插图签测试文件1.xlsx",
    catalog_name="图纸目录",   
    custom_suffixes=None       
):
    """
    【目录专用 - 修正版】读取 Excel 项目信息 -> 生成目录图签
    
    修正内容：
    1. 自动调用 generate_name_and_ratio_from_com 计算 ctq[0] 的打印框比例。
    2. 将计算出的比例填入图签的“出图比例”字段。
    """
    import re    
    print("\n🔍 [Debug] 开始执行 update_catalog_titleblocks_from_excel ...")
    
    # 尝试刷新连接
    try: li()
    except: pass

    # ================= 1. 基础校验 =================
    if not ctq:
        print("❌ [错误] 输入的 ctq 为空")
        return False
        
    if not isinstance(ctq, (list, tuple)) or len(ctq) < 2:
        sys_logger.info(f"❌ [错误] ctq 结构错误，期望 (polys, blocks)")
        return False

    # 提取多段线列表和图签块列表
    polys = ctq[0] 
    blocks = ctq[1] 
    
    if not blocks:
        print("❌ [错误] 图签块列表为空")
        return False
        
    count = min(len(polys), len(blocks)) # 确保配对安全
    sys_logger.info(f"✅ [校验通过] 准备处理 {count} 组目录配对。")

    # ================= 1.1 Excel 路径校验 =================
    excel_path = Path(excel_path)
    if not excel_path.exists():
        sys_logger.info(f"⚠️ [提示] 指定路径不存在: {excel_path}")
        try:
            if 'doc' in globals():
                doc_name = Path(doc.Name).stem
                inferred_path = Path(r"D:/Mypro/基础服务/用户1/dwg文件") / f"{doc_name}.xlsx"
                sys_logger.info(f"   --> 尝试推断路径: {inferred_path}")
                if inferred_path.exists():
                    sys_logger.info(f"✅ [自动修正] 找到同名文件，自动切换。")
                    excel_path = inferred_path
                else:
                    return False
            else:
                return False
        except: return False

    sys_logger.info(f"📖 [目录模式] 正在读取 Excel: {excel_path.name} ...")

    # ================= 2. 读取 Excel 项目头信息 =================
    excel = None; wb = None; project_data = {}
    
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False; excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(excel_path))
        try: ws = wb.Worksheets(1)
        except: ws = wb.ActiveSheet
        
        # 定义项目通用字段
        project_keys = [
            "专业名称", "专业代号", "项目名称", "子项目名称", "建设单位名称", 
            "设计阶段", "版本号", "出图日期", "设计编号", "设计院名称",
            "法人代表", "注册建筑师", "注册结构师"
        ]
        
        # 读取第2行
        for idx, key in enumerate(project_keys):
            val = ws.Cells(2, 7 + idx).Value # G列开始
            project_data[key] = str(val).strip() if val is not None else ""
            
    except Exception as e:
        sys_logger.info(f"❌ Excel 读取失败: {e}")
        return False
    finally:
        if wb: wb.Close(SaveChanges=0)
        if excel: excel.Quit()
        del wb, excel

    prefix = project_data.get("专业代号", "JZ") or "JZ"

    # ================= 3. 生成目录数据并写入 =================
    sys_logger.info(f"🚀 开始写入 {count} 个目录图签 (前缀: {prefix})...")
    
    MULTI_LINE_KEYS = ["图纸名称", "项目名称", "子项目名称", "建设单位名称"]
    success_count = 0

    for i in range(count):
        try:
            target_poly = polys[i]  # 获取多段线 (打印框)
            target_block = blocks[i] # 获取图签块
            
            # --- A. 计算比例与图幅 (新增逻辑) ---
            calculated_scale = ""
            calculated_size = ""
            
            if "generate_name_and_ratio_from_com" in globals():
                try:
                    # 调用几何分析函数
                    # 返回结构: (图幅名, 比例字符串, 图号, 竖向标志)
                    res = generate_name_and_ratio_from_com(target_poly, A3dy=0)
                    
                    if res and isinstance(res, (tuple, list)) and len(res) >= 3:
                        calculated_size = str(res[2]) # 例如 "A3"
                        calculated_scale = str(res[1]) # 例如 "1:100"
                    else:
                        sys_logger.info(f"   ⚠️ 第{i+1}个框无法识别标准尺寸，比例将置空。")
                except Exception as calc_err:
                    sys_logger.info(f"   ⚠️ 计算比例出错: {calc_err}")
            else:
                print("   ⚠️ 缺少 generate_name_and_ratio_from_com 函数，跳过比例计算。")

            # --- B. 生成图号 ---
            dwg_no_suffix = ""
            if custom_suffixes and isinstance(custom_suffixes, list) and i < len(custom_suffixes):
                dwg_no_suffix = str(custom_suffixes[i])
            else:
                if count == 1: dwg_no_suffix = "00"
                else: dwg_no_suffix = f"00-{i+1}"
            
            final_dwg_no = f"{prefix}-{dwg_no_suffix}"
            
            # --- C. 组合数据 ---
            sheet_data = {
                "图纸编号": final_dwg_no,
                "图纸名称": catalog_name, 
                "图纸规格": calculated_size,  # 填入计算出的图幅
                "出图比例": calculated_scale  # 填入计算出的比例
            }
            
            full_attributes = project_data.copy()
            full_attributes.update(sheet_data)
            
            # --- D. 格式清洗 ---
            final_keys = []
            final_values = []
            for k, v in full_attributes.items():
                final_keys.append(k)
                if k in MULTI_LINE_KEYS and v:
                    parts = re.split(r'//|\\\\|\n', v)
                    clean_parts = [p.strip() for p in parts if p.strip()]
                    final_values.append(clean_parts if len(clean_parts) > 1 else v)
                else:
                    final_values.append(v)

            sys_logger.info(f"   ✍ 目录 [{i+1}/{count}]: {final_dwg_no} | {catalog_name} | 比例:{calculated_scale}")
            
            # --- E. 写入 ---
            if "set_attribute_mtext" in globals():
                set_attribute_mtext(target_block, final_keys, final_values, keep_prefix=True, verbose=False)
                success_count += 1
            else:
                print("❌ [错误] 缺少 set_attribute_mtext 函数")
                break
                
        except Exception as e:
            sys_logger.info(f"❌ [错误] 处理第 {i+1} 个块时出错: {e}")

    sys_logger.info(f"✅ 目录更新完成! 共 {success_count} 张")
    return True





















def update_catalog_titleblocks_from_excel_y(
    ctq, 
    excel_path=r"D:/Mypro/基础服务/用户1/dwg文件/插图签测试文件1.xlsx",
    catalog_name="图纸目录",   # <--- 可自定义目录名称
    custom_suffixes=None       # <--- 可自定义编号后缀列表
):
    """
    【目录专用】读取 Excel 项目信息 -> 生成目录图签
    
    逻辑说明：
    1. 仍然读取 Excel 获取 [项目名称/建设单位/专业代号] 等通用信息。
    2. 忽略 Excel 中的具体图号和图名。
    3. 图名统一为 catalog_name (默认"图纸目录")。
    4. 图号生成规则 (基于专业代号 PREFIX):
       - 情况A (默认): 
         - 如果只有1张图: PREFIX-00
         - 如果有多张图: PREFIX-00-1, PREFIX-00-2 ...
       - 情况B (用户传入 custom_suffixes=["00a", "00b"]):
         - 按顺序生成: PREFIX-00a, PREFIX-00b ...
    """

    import re 

    # ================= 1. 基础校验 =================
    if not ctq or len(ctq) < 2: return False
    # 注意：这里我们遍历的是 CAD 里选中的块，而不是 Excel 的行
    blocks = ctq[1] 
    count = len(blocks)
    if count == 0: return False

    excel_path = Path(excel_path)
    if not excel_path.exists():
        # 尝试自动寻找同名 excel
        try:
            doc_name = Path(doc.Name).stem
            inferred_path = Path(r"D:/Mypro/基础服务/用户1/dwg文件") / f"{doc_name}.xlsx"
            if inferred_path.exists():
                sys_logger.info(f"⚠️ 自动切换到: {inferred_path}")
                excel_path = inferred_path
            else:
                return False
        except: return False

    t0 = time.time()
    sys_logger.info(f"📖 [目录模式] 正在读取 Excel 项目头信息: {excel_path.name} ...")

    # ================= 2. 仅读取 Excel 项目头信息 =================
    excel = None; wb = None
    project_data = {}
    
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False; excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(excel_path))
        try: ws = wb.Worksheets(1)
        except: ws = wb.ActiveSheet
        
        # 定义需要读取的项目通用字段 (必须与 Excel 表头顺序一致)
        project_keys = [
            "专业名称", "专业代号", "项目名称", "子项目名称", "建设单位名称", 
            "设计阶段", "版本号", "出图日期", "设计编号", "设计院名称",
            "法人代表", "注册建筑师", "注册结构师"
        ]
        
        # 只读取第2行 (Header Info)
        for idx, key in enumerate(project_keys):
            val = ws.Cells(2, 7 + idx).Value # G列开始
            project_data[key] = str(val).strip() if val is not None else ""
            
    except Exception as e:
        sys_logger.info(f"❌ Excel 读取失败: {e}")
        return False
    finally:
        if wb: wb.Close(SaveChanges=0)
        if excel: excel.Quit()
        del wb, excel

    # 获取全局专业代号
    prefix = project_data.get("专业代号", "JZ")
    if not prefix: prefix = "JZ"

    # ================= 3. 生成目录数据并写入 =================
    sys_logger.info(f"🚀 开始写入 {count} 个目录图签 (前缀: {prefix})...")
    
    # 定义多行处理字段 (虽然目录名一般很短，但项目名可能很长)
    MULTI_LINE_KEYS = ["图纸名称", "项目名称", "子项目名称", "建设单位名称"]
    
    success_count = 0

    for i in range(count):
        target_block = blocks[i]
        
        # --- A. 核心逻辑：生成图号 ---
        dwg_no_suffix = ""
        
        if custom_suffixes and isinstance(custom_suffixes, list):
            # 1. 用户自定义模式
            if i < len(custom_suffixes):
                dwg_no_suffix = str(custom_suffixes[i])
            else:
                # 列表不够长，回退到默认逻辑
                dwg_no_suffix = f"00-{i+1}" 
        else:
            # 2. 默认智能模式
            if count == 1:
                dwg_no_suffix = "00"
            else:
                # 多页目录：00-1, 00-2...
                dwg_no_suffix = f"00-{i+1}"
        
        # 组合最终图号
        final_dwg_no = f"{prefix}-{dwg_no_suffix}"
        
        # --- B. 组合数据 ---
        sheet_data = {
            "图纸编号": final_dwg_no,
            "图纸名称": catalog_name, 
            "图纸规格": "", # 目录通常不需要填规格比例，或者可设为固定值
            "出图比例": ""
        }
        
        # 合并 (项目信息 + 目录独立信息)
        full_attributes = project_data.copy()
        full_attributes.update(sheet_data)
        
        # --- C. 格式清洗 (换行处理) ---
        final_keys = []
        final_values = []

        for k, v in full_attributes.items():
            final_keys.append(k)
            if k in MULTI_LINE_KEYS and v:
                parts = re.split(r'//|\\\\|\n', v)
                clean_parts = [p.strip() for p in parts if p.strip()]
                if len(clean_parts) > 1: final_values.append(clean_parts)
                else: final_values.append(v)
            else:
                final_values.append(v)

        sys_logger.info(f"   ✍ 目录 [{i+1}/{count}]: {final_dwg_no} | {catalog_name}")
        
        # --- D. 写入 ---
        set_attribute_mtext(target_block, final_keys, final_values, keep_prefix=True, verbose=False)
        success_count += 1

    sys_logger.info(f"✅ 目录更新完成! 共 {success_count} 张")
    return True



#&&% 从excel写入目录dwg文件

def write_catalog_from_excel_to_cad(
    ctq, 
    data_excel_path=None,
    mubanxuhao = 1,
    start_index=0,
    text_height=250, text_style="Standard", text_alignment="中心",       
    text_rotation=0.0, text_oblique=0.0, text_width_factor=1.0, width_factors_map=None        
):

    """
    【函数编号】: CAT-001
    【所属模块】: 目录自动生成模块 (Catalog Automation)
    【功能描述】: 
        全自动 CAD 图纸目录生成器 (最终稳定版)。
        执行 "环境复原 -> 模板解析 -> 空间扫描 -> 数据融合 -> 智能回写 -> 归档收尾" 的完整闭环流程。

        核心特性：
        1. 头部解冻与复原 (Rescue): 
           在开始前强制解冻相关图层，将上次生成的（被隐藏的）横线归还至源图层，确保旧线能被重新利用，防止线段丢失或重复绘制。
        2. 智能路径推断:
           若未提供数据源路径，自动根据当前 DWG 文件名推断 Excel 路径（支持去除 "_目录" 后缀逻辑）。
        3. 空间投影匹配:
           扫描 CAD 中的横线几何位置，结合配置文件定义的列宽范围，通过投影算法精准确定文字落点。
        4. 双源数据融合:
           优先读取 CAD 图中选中的图签块信息 (is_self=True)，再合并 Excel 中的外部图纸信息。

    【参数详解】:
        - ctq (tuple): 
            CAD 选择集结果 (polys, blocks)。
            用于提取当前图纸中已存在的图签信息（如封面、目录本身的图号）。
        - data_excel_path (str, optional): 
            数据源 Excel 路径。若为 None，则自动根据当前 CAD 文件名推断。
        - mubanxuhao (int): 
            模板配置文件序号。决定读取 "配置/目录结构{n}.xlsx"。
        - start_index (int): 
            生成的目录序号起始值（默认为 0）。
        - text_height (float): 
            生成的文字高度（默认 250）。
        - text_style (str): 
            使用的 CAD 文字样式名（默认 "Standard"）。
        - text_alignment (str): 
            文字对齐方式描述（如 "中心"）。
        - width_factors_map (dict): 
            宽度因子映射表。可针对不同列（如名称列）指定特殊的字宽压缩比。

    【返回值】:
        - bool: 执行成功返回 True，失败（如路径错误、模板读取失败）返回 False。    
    
    """

    
    import re


    print("\n" + "="*50)
    print("🚀 [最终稳定版] 目录生成程序")
    print("   修正: 头部解冻逻辑，确保旧线能被复原")
    print("="*50)






    
    # 如果未传入 data_excel_path，则自动从当前图纸推断
    if data_excel_path is None:
        try:
            # 1. 获取当前激活图纸的完整路径
            # 例如: "D:\项目\住宅楼_目录.dwg"
            current_path = doc.FullName
        except:
            current_path = ""
    
        # 校验：防止针对未保存的新建文件操作
        if not current_path:
            print("❌ 错误：当前图纸尚未保存，无法推断路径。")
            return False
    
        # 2. 拆分路径与后缀
        # root = "D:\项目\住宅楼_目录", ext = ".dwg"
        root, ext = os.path.splitext(current_path)
    
        # 3. 核心清洗逻辑：如果文件名以 "_目录" 结尾，则去除它
        if root.endswith("_目录"):
            clean_root = root[:-3] # 去掉最后3个字符
            sys_logger.info(f"🔍 检测到目录文件，自动定位到主文件: {clean_root}")
        else:
            clean_root = root
    
        # 4. 生成 Excel 路径 (强制改为 .xlsx)
        data_excel_path = clean_root + ".xlsx"
        
        sys_logger.info(f"✅ 推断数据 Excel: {data_excel_path}")


    # =========================================================
    # 0. 内部辅助函数
    # =========================================================
    def wait_cad():
        try: doc.Application.Update()
        except: pass
        time.sleep(0.1)

    # 简化的图层属性设置，增加异常捕获
    def set_layer_status(name, freeze=None, lock=None, show=None):
        try:
            l = doc.Layers.Item(name)
            # 如果要操作冻结，且该层是当前层，必须先切走
            if freeze is True and doc.ActiveLayer.Name == name:
                doc.ActiveLayer = doc.Layers.Item("0")
            
            if freeze is not None: set_attr(l, "Freeze", freeze)
            if lock is not None: set_attr(l, "Lock", lock)
            if show is not None: set_attr(l, "LayerOn", show)
        except: pass

    # =========================================================
    # 1. 头部: 初始化与复原 (The "Rescue" Phase)
    # =========================================================
    li() 
    LAYER_TEXT   = "mulu_wenzi"
    LAYER_SOURCE = "mulu_hengxian"
    LAYER_HIDDEN = "mulu_hengxian_d"

    print("♻️ [Step 1] 环境复原与清理...")

    # 1.1 【关键修正】先确保图层存在，并强制解冻、解锁、打开
    # 如果不解冻，stc 选不到对象，复原就会失败！
    for lname in [LAYER_HIDDEN, LAYER_SOURCE, LAYER_TEXT]:
        try: doc.Layers.Add(lname)
        except: pass
        set_layer_status(lname, freeze=False, lock=False, show=True)
    
    wait_cad() # 等待 CAD 反应过来

    # 1.2 归还旧线 (从 隐藏层 -> 源图层)
    # 现在图层已解冻，stc 可以正常工作了
    try:
        objs_hidden = stc(LAYER_HIDDEN)
        if objs_hidden:
            sys_logger.info(f"   🔙 正在归还 {len(objs_hidden)} 条横线到 {LAYER_SOURCE}...")
            for obj in objs_hidden:
                try:
                    set_attr(obj, "Layer", LAYER_SOURCE)
                    set_attr(obj, "Color", 256) # 恢复随层
                except: pass
    except Exception as e:
        sys_logger.info(f"   ⚠️ 复原旧线时出错: {e}")

    # 1.3 清空文字层
    try:
        objs_text = stc(LAYER_TEXT)
        if objs_text:
            sys_logger.info(f"   🧹 清空文字层: 删除 {len(objs_text)} 个对象...")
            for obj in objs_text:
                try: obj.Delete()
                except: pass
    except: pass
    
    wait_cad()


    # =========================================================
    # 2. 解析配置 (Config) - 【路径修正版】
    # =========================================================
    print("\n📏 [Step 2] 解析排版配置...")
    
    import os
    userpath = os.environ.get('USERPATH')
    
    if not userpath:
        print("❌ 错误：环境变量 USERPATH 未设置")
        return False

    # 1. 定义配置文件路径
    # 逻辑: D:/Mypro/基础服务/用户1/配置/目录结构{n}.xlsx
    config_filename = f"目录结构{mubanxuhao}.xlsx"
    config_excel_path = os.path.join(userpath, "配置", config_filename)

    # 2. 检查文件是否存在
    if not os.path.exists(config_excel_path):
        sys_logger.info(f"❌ 错误：找不到排版配置文件 -> {config_excel_path}")
        return False

    sys_logger.info(f"   👉 读取配置: {config_filename}")

    # 3. 读取参数
    try:
        template_config = read_catalog_template_config(config_excel_path)
        if not template_config: return False
    except Exception as e:
        sys_logger.info(f"❌ 读取配置函数报错: {e}")
        return False

    # 4. 提取常用变量
    hangju = template_config['hangju']
    y_start_1 = template_config['Y_START_1']
    y_start_2 = template_config['Y_START_2']
    
    # 5. 定义列范围 (保持原逻辑)
    col_x_ranges = {}
    col_keys = ['col1', 'col2', 'col3', 'col4']
    
    # 增加容错：确保 template_config['cols'] 存在且是列表
    cols_data = template_config.get('cols', [])
    if not cols_data:
        print("❌ 配置错误: 'cols' 数据为空")
        return False

    for i, col_data in enumerate(cols_data):
        if i >= 4: break # 防止越界
        center_x = col_data['name_x']
        half_width = 5500.0 
        col_x_ranges[col_keys[i]] = (center_x - half_width, center_x + half_width)


    # =========================================================
    # 3. 扫描源横线 (Scanning)
    # =========================================================
    print("\n🔍 [Step 3] 扫描源横线...")
    lines_pool = []
    try: raw_objs = stc(LAYER_SOURCE)
    except: raw_objs = []
    
    if raw_objs:
        for obj in raw_objs:
            try:
                safe_ent = cast_object(obj)
                min_pt, max_pt = safe_ent.GetBoundingBox()
                if abs(min_pt[1] - max_pt[1]) < 5.0:
                    mid_y = (min_pt[1] + max_pt[1]) / 2.0
                    x_start = min(min_pt[0], max_pt[0])
                    x_end = max(min_pt[0], max_pt[0])
                    lines_pool.append({
                        'y': mid_y, 
                        'x_min': x_start, 'x_max': x_end,
                        'mid_x': (x_start + x_end)/2,
                        'safe_obj': safe_ent
                    })
            except: pass
    sys_logger.info(f"   📦 有效横线库: {len(lines_pool)} 条")





    # =========================================================
    # 4. 准备数据 & 坐标生成 (Data Preparation)
    # =========================================================
    print("\n📊 [Step 3.5] 动态生成坐标点 (位置修正: 上移一行)...")

    all_idx = []; all_no = []; all_name = []; all_spec = []
    
    # 定义一个生成器函数
    def generate_col_points(col_index, start_y, count, x_config):
        """生成一整列的坐标点 (idx, no, name, spec)"""
        p_idx, p_no, p_name, p_spec = [], [], [], []
        
        for r in range(count):
            # -----------------------------------------------------------
            # 【最终修正】：
            # 现象：之前低了一整行。
            # 措施：在原基础上 + hangju (向上平移一行)。
            # 公式：Start - r*H - H/2 + H  ==>  Start - r*H + H/2
            # -----------------------------------------------------------
            current_y = start_y - (r * hangju) + (hangju / 2.0)
            
            # 生成坐标点 (X, Y, Z)
            p_idx.append( (x_config['idx_x'],  current_y, 0.0) )
            p_name.append((x_config['name_x'], current_y, 0.0) )
            p_no.append(  (x_config['no_x'],   current_y, 0.0) )
            p_spec.append((x_config['spec_x'], current_y, 0.0) )
            
        return p_idx, p_name, p_no, p_spec

    # --- 循环生成 4 栏的数据 ---
    cols_data = template_config['cols'] # 列表，包含4个字典
    
    # 第 1 栏
    i, n, no, s = generate_col_points(0, y_start_1, template_config['row_count_1'], cols_data[0])
    all_idx.extend(i); all_name.extend(n); all_no.extend(no); all_spec.extend(s)
    
    # 第 2, 3, 4 栏 (它们共享 Y_START_2 和 row_count_others)
    for col_i in range(1, 4): # 1, 2, 3
        i, n, no, s = generate_col_points(col_i, y_start_2, template_config['row_count_others'], cols_data[col_i])
        all_idx.extend(i); all_name.extend(n); all_no.extend(no); all_spec.extend(s)

    # 计算总容量
    max_cap = len(all_name)
    sys_logger.info(f"   ✅ 已生成坐标网格，最大容量: {max_cap} 条目")














    full_list = []
    
    # --- A. 读取自身图签信息 (目录页) ---
    # 【修改点】：同时遍历 poly 和 block，计算图幅
    if ctq and len(ctq) >= 2:
        polys = ctq[0]
        blocks = ctq[1]
        cnt = min(len(polys), len(blocks))
        
        for i in range(cnt):
            poly = polys[i]
            blk = blocks[i]
            
            # 1. 尝试从属性获取基础信息
            try:
                attrs = get_block_attributes_dict(blk, False, False)
                dwg_no = attrs.get("图纸编号","") or attrs.get("图号","")
                dwg_name = attrs.get("图纸名称","图纸目录")
            except: 
                dwg_no = ""; dwg_name = "图纸目录"

            # 2. 【新增】调用几何分析函数获取规格
            # 默认给个空值，如果算出来了就覆盖
            calculated_spec = "" 
            if "generate_name_and_ratio_from_com" in globals():
                try:
                    # 返回结构: (图幅全名, 比例, 图幅简称, 竖向标志)
                    # 例如: ('ISO_A3...', '1:100', 'A3', 0)
                    res = generate_name_and_ratio_from_com(poly, A3dy=0)
                    
                    if res and isinstance(res, (tuple, list)) and len(res) >= 3:
                        calculated_spec = str(res[2]) # 取第3个元素，即 "A3"
                except: pass
            
            # 3. 存入列表
            full_list.append({
                "no": dwg_no, 
                "name": dwg_name, 
                "spec": calculated_spec, # <--- 这里填入计算出的 A1/A2/A3
                "is_self": True
            })

    # --- B. 读取 Excel 外部图纸信息 (保持不变) ---

    sys_logger.info(f"📖 [Data] 正在读取数据源: {Path(data_excel_path).name} ...")
    
    excel_data = None
    wb_data = None
    
    try:
        # 1. 强制启动一个新的 Excel 进程 (避免干扰)
        excel_data = win32com.client.DispatchEx("Excel.Application")
        excel_data.Visible = False
        excel_data.DisplayAlerts = False
        
        # 2. 打开数据文件
        wb_data = excel_data.Workbooks.Open(str(data_excel_path), ReadOnly=True)
        ws = wb_data.Worksheets(1)
        
        # 3. 循环读取
        r = 2
        added_count = 0
        while True:
            # 安全获取值 (转字符串并去空格)
            v_no = str(ws.Cells(r, 2).Value or "").strip()   # B列: 图号
            v_name = str(ws.Cells(r, 3).Value or "").strip() # C列: 图名
            
            # 停止条件: 图号和图名都为空
            if not v_no and not v_name: 
                break
            
            # 读取规格 (D列)
            v_spec = str(ws.Cells(r, 4).Value or "").strip()
            
            # 加入总列表
            full_list.append({
                "no": v_no, 
                "name": v_name, 
                "spec": v_spec, 
                "is_self": False
            })
            
            r += 1
            added_count += 1
            
        sys_logger.info(f"   ✅ 成功从 Excel 读取 {added_count} 条图纸数据")

    except Exception as e:
        sys_logger.info(f"❌ [数据读取失败] 无法读取 Excel 数据: {e}")
        import traceback
        traceback.print_exc() # 打印详细报错信息
        
    finally:
        # 4. 清理资源
        if wb_data: 
            try: wb_data.Close(False)
            except: pass
        if excel_data: 
            try: excel_data.Quit()
            except: pass
        del wb_data, excel_data



    # =========================================================
    # 5. 书写与标记 (Writing) - 【新版智能排版逻辑】
    # =========================================================
    
    # --- 1. 定义样式与辅助变量 ---
    wf_name = width_factors_map.get("name", 0.7) if width_factors_map else 0.7
    # 基础参数包
    base_kw = {
        "layer": LAYER_TEXT, 
        "rotation": text_rotation, 
        "style": text_style, 
        "alignment": text_alignment
    }

    # --- 2. 定义内部辅助函数 (测量与分行) ---
    
    def get_text_visual_width(txt, height, width_factor):
        """物理测量：在CAD原点生成临时文字，测完即删"""
        if not txt: return 0.0
        try:
            # 创建临时文字对象
            temp_obj = doc.ModelSpace.AddText(txt, win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (0,0,0)), height)
            temp_obj.ScaleFactor = width_factor
            temp_obj.StyleName = text_style
            # 获取包围盒计算宽度
            min_p, max_p = temp_obj.GetBoundingBox()
            w = abs(max_p[0] - min_p[0])
            temp_obj.Delete() # 立即删除
            return w
        except:
            return len(txt) * height * width_factor * 0.8 # 兜底估算

    def smart_split_by_width(full_text, max_w, height, wf):
        """智能分行：按顿号切分，累加测试，防止孤儿行"""
        # 将字符串按顿号炸开，保留顿号在词尾
        # "A图、B图" -> ["A图、", "B图"]
        raw_atoms = full_text.split("、")
        atoms = []
        for idx, a in enumerate(raw_atoms):
            if idx < len(raw_atoms) - 1: atoms.append(a + "、")
            else: atoms.append(a) # 最后一个词不加顿号
        
        lines = []
        current_line = ""
        
        for atom in atoms:
            test_str = current_line + atom
            # 预测宽度
            w = get_text_visual_width(test_str, height, wf)
            
            if w <= max_w:
                current_line = test_str
            else:
                # 超宽换行
                if current_line: 
                    # 换行前去掉行尾多余顿号
                    clean_prev = current_line[:-1] if current_line.endswith("、") else current_line
                    lines.append(clean_prev)
                current_line = atom
        
        # 处理最后剩下的
        if current_line:
            clean_last = current_line[:-1] if current_line.endswith("、") else current_line
            lines.append(clean_last)
            
        return lines

    # --- 3. 开始循环写入 ---
    slot = 0
    count_marked = 0

    for i, item in enumerate(full_list):
        if slot >= max_cap: break
        
        raw_n = item['name']
        cur_idx_str = f"{start_index + i:02d}"
        
        # [A] 数据清洗：统一换行符为顿号，并去重
        clean_name = re.sub(r'//|\\\\|\n', '、', raw_n)
        clean_name = re.sub(r'、+', '、', clean_name) # 防止出现 "、、"
        
        # [B] 获取当前列的物理限宽
        # all_name[slot] 是文字插入点坐标
        current_x = all_name[slot][0]
        limit_width = 11000.0 # 默认值
        
        # 遍历列范围配置，找到当前文字属于哪一列 (col1? col2?)
        for _, (x_min, x_max) in col_x_ranges.items():
            if x_min - 100 <= current_x <= x_max + 100:
                limit_width = (x_max - x_min) * 0.95 # 使用 95% 宽度
                break
        
        # [C] 排版决策三部曲
        final_lines = []
        final_height = text_height
        
        # 1. 尝试单行 (标准高度)
        w_single = get_text_visual_width(clean_name, text_height, wf_name)
        
        if w_single <= limit_width:
            final_lines = [clean_name] # 完美放下
        else:
            # 2. 尝试分行 (标准高度)
            lines_b = smart_split_by_width(clean_name, limit_width, text_height, wf_name)
            
            if len(lines_b) <= 2:
                final_lines = lines_b # 两行能放下，接受
            else:
                # 3. 实在放不下 -> 缩小字高至 80% 再分行
                reduced_h = text_height * 0.8
                lines_c = smart_split_by_width(clean_name, limit_width, reduced_h, wf_name)
                
                final_lines = lines_c
                final_height = reduced_h # 标记字高已缩小

        # [D] 执行写入
        is_multiline = (len(final_lines) > 1)
        
        if is_multiline:
            # ---------------------------------------------------------
            # 双行模式
            # ---------------------------------------------------------
            if slot + 1 >= max_cap: 
                sys_logger.info(f"⚠️ [警告] 空间不足，无法写入多行: {clean_name}")
                break
            
            # 1. 判定是否跨列
            x1 = all_name[slot][0]
            x2 = all_name[slot+1][0]
            is_cross_column = abs(x1 - x2) > 500

            # 2. 写入文字 (两行都写)
            kw_multi = base_kw.copy()
            kw_multi["height"] = final_height
            
            write_cad_text(p=all_name[slot], text=final_lines[0], width_factor=wf_name, **kw_multi)
            txt_2 = final_lines[1] if len(final_lines) > 1 else ""
            write_cad_text(p=all_name[slot+1], text=txt_2, width_factor=wf_name, **kw_multi)
            
            # 3. 计算其他信息位置 & 处理横线
            target_y = 0.0
            
            if is_cross_column:
                # === A: 跨列 (禁止隐藏线，序号跟随第一行) ===
                sys_logger.info(f"   ℹ️ [排版] 跨列分行: {cur_idx_str}")
                target_y = all_name[slot][1] 
                
                # 坐标跟随 slot (第一列)
                p_idx  = (all_idx[slot][0], target_y, 0.0)
                p_no   = (all_no[slot][0], target_y, 0.0)
                p_spec = (all_spec[slot][0], target_y, 0.0)
                
            else:
                # === B: 同列 (必须隐藏横线，序号居中) ===
                y1 = all_name[slot][1]
                y2 = all_name[slot+1][1]
                target_y = (y1 + y2) / 2.0
                
                # 坐标 X 跟随 slot，Y 居中
                p_idx  = (all_idx[slot][0], target_y, 0.0)
                p_no   = (all_no[slot][0], target_y, 0.0)
                p_spec = (all_spec[slot][0], target_y, 0.0)
                
                # —————— 核心修正：寻找并隐藏横线 ——————
                # 逻辑：找到 X 范围在当前列内，且 Y 坐标接近 target_y 的横线
                current_col_range = None
                current_text_x = all_name[slot][0]
                
                # a. 确定当前在哪一列
                for _, (x_min, x_max) in col_x_ranges.items():
                    if x_min - 100 <= current_text_x <= x_max + 100:
                        current_col_range = (x_min, x_max)
                        break
                
                # b. 扫描线库
                if current_col_range:
                    c_min, c_max = current_col_range
                    cand_line = None
                    min_dist = 99999.0
                    
                    for line_info in lines_pool:
                        # 条件1: 线的中心 X 必须在列宽范围内
                        if not (c_min <= line_info['mid_x'] <= c_max): 
                            continue
                        
                        # 条件2: 线的 Y 必须接近两行文字的中间 (target_y)
                        dist = abs(line_info['y'] - target_y)
                        
                        # 容差设为 50 (只要比行距小很多即可)
                        if dist < 50.0 and dist < min_dist:
                            min_dist = dist
                            cand_line = line_info['safe_obj']
                    
                    # c. 执行隐藏
                    if cand_line:
                        try:
                            set_attr(cand_line, "Layer", LAYER_HIDDEN)
                            set_attr(cand_line, "Color", 1) # 临时变红，便于调试观察
                            count_marked += 1
                        except: pass
                # ——————————————————————————————————————————

            # 4. 写入其他信息 (序号/图号/规格)
            kw_other = base_kw.copy()
            kw_other["height"] = text_height 
            
            write_cad_text(p=p_idx, text=cur_idx_str, width_factor=1.0, **kw_other)
            write_cad_text(p=p_no, text=item['no'], width_factor=0.8, **kw_other)
            write_cad_text(p=p_spec, text=item['spec'], width_factor=1.0, **kw_other)
            
            slot += 2
                    
        else:
            # --- 单行模式 (保持不变) ---
            kw_single = base_kw.copy()
            kw_single["height"] = final_height 
            
            write_cad_text(p=all_idx[slot], text=cur_idx_str, width_factor=1.0, **kw_single)
            write_cad_text(p=all_no[slot], text=item['no'], width_factor=0.8, **kw_single)
            write_cad_text(p=all_name[slot], text=final_lines[0], width_factor=wf_name, **kw_single)
            write_cad_text(p=all_spec[slot], text=item['spec'], width_factor=1.0, **kw_single)
            
            slot += 1





    print("\n🏁 [Step 5] 最终收尾...")

    wait_cad()
    
    # 6.1 强制改色: 隐藏层全员变回 256 (ByLayer)
    # 这一步是为了下一次运行时的“复原”做准备
    # 如果不改回 256，下次复原到源图层后，它们还是红色的
    
    # 定义改色函数 (闭环版)
    def force_color_256(layer_name):
        sys_logger.info(f"   🎨 正在归档图层 {layer_name} (颜色 -> 256)...")
        try: objs = stc(layer_name)
        except: return
        if not objs: return
        
        pending = list(objs)
        for _ in range(3): # 尝试 3 次
            if not pending: break
            next_p = []
            for o in pending:
                try:
                    set_attr(o, "color", 256)
                    if get_attr(o, "Color") != 256: next_p.append(o)
                except: next_p.append(o)
            pending = next_p
            if pending: time.sleep(0.1)

    force_color_256(LAYER_HIDDEN)
    
    # 6.2 冻结隐藏层
    set_layer_status(LAYER_HIDDEN, freeze=True)
    
    wait_cad()
    try: doc.Regen(1)
    except: pass
    
    sys_logger.info(f"\n✅ 全部完成: 已处理 {slot} 行数据")
    return True


#&&% 读取目录结构的excel文件

def read_catalog_template_config(excel_path):
    """
    【配置读取器 - 兼容增强版】
    读取目录结构模板的数学定义参数。
    增强特性：能够处理单元格里的 "x1, x2" 格式，自动计算中心值。
    """
    import win32com.client
    
    # --- 内部辅助：智能数值转换 ---
    def _smart_float(val):
        """
        将 Excel 单元格内容转为浮点数。
        1. 如果是数字，直接返回。
        2. 如果是 "num1, num2" 字符串，返回 (num1+num2)/2。
        3. 如果是空或乱码，返回 0.0。
        """
        if val is None: return 0.0
        if isinstance(val, (int, float)): return float(val)
        
        s = str(val).replace('，', ',').strip() # 兼容中文逗号
        if not s: return 0.0
        
        try:
            # 尝试直接转换 "200.0"
            return float(s)
        except ValueError:
            # 转换失败，尝试处理逗号分隔 "200, 300"
            if ',' in s:
                try:
                    parts = [float(x) for x in s.split(',') if x.strip()]
                    if len(parts) >= 2:
                        # 如果是两个数，取平均值（中心点）
                        # 如果是旧格式 x1,y1,x2,y2 (4个数)，取 (x1+x3)/2
                        if len(parts) >= 4:
                            return (parts[0] + parts[2]) / 2.0
                        return sum(parts) / len(parts)
                    elif len(parts) == 1:
                        return parts[0]
                except:
                    pass
            sys_logger.info(f"⚠️ [警告] 无法解析配置数值: '{val}'，已重置为 0")
            return 0.0

    config = {}
    excel = None
    wb = None
    
    try:
        # 使用 DispatchEx 强制启动新进程
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(excel_path), ReadOnly=True)
        ws = wb.Worksheets(1)
        
        # --- A. 读取全局参数 (第2行) ---
        config['Y_START_1']       = _smart_float(ws.Cells(2, 1).Value)
        config['Y_START_2']       = _smart_float(ws.Cells(2, 2).Value)
        config['row_count_1']     = int(_smart_float(ws.Cells(2, 3).Value))
        config['row_count_others']= int(_smart_float(ws.Cells(2, 4).Value))
        config['hangju']          = _smart_float(ws.Cells(2, 9).Value)
        
        if config['hangju'] == 0: config['hangju'] = 800.0 # 兜底默认值

        # --- B. 读取列坐标参数 (第2-5行) ---
        cols_config = []
        for r in range(2, 6): # 2, 3, 4, 5 行
            col_data = {
                "idx_x":  _smart_float(ws.Cells(r, 5).Value),
                "name_x": _smart_float(ws.Cells(r, 6).Value), # 这里会触发智能计算
                "no_x":   _smart_float(ws.Cells(r, 7).Value),
                "spec_x": _smart_float(ws.Cells(r, 8).Value),
            }
            cols_config.append(col_data)
        
        config['cols'] = cols_config
        
        sys_logger.info(f"📏 [配置加载成功] 行距={config['hangju']}, Y1={config['Y_START_1']}")
        # 打印一下第一列的图名X，确认是否解析正确
        sys_logger.info(f"   (调试) 第1栏图名中心 X: {config['cols'][0]['name_x']}")
        
        return config

    except Exception as e:
        sys_logger.info(f"❌ 读取模板配置失败: {e}")
        return None
    finally:
        if wb: 
            try: wb.Close(False)
            except: pass
        if excel: 
            try: excel.Quit()
            except: pass











#&&% 处理目录全局信息块

#&&% 将目录dwg文件合并到主文件



class CatalogConfigBuilder:
    """
    目录模板配置生成器：将简单的行列参数转换为复杂坐标字典
    """
    def __init__(self, row_height=800):
        self.row_height = row_height
        self.column_layouts = [] 
        self.field_x_definitions = {}

    def add_column_layout(self, top_y, row_count):
        """定义第 N 排的纵向布局"""
        self.column_layouts.append({
            "top": top_y,
            "count": int(row_count)
        })

    def set_field_x_ranges(self, field_key, x_ranges_list):
        """定义某个字段在每一排的 X 轴范围 [(min, max), (min, max)...]"""
        self.field_x_definitions[field_key] = x_ranges_list

    def generate(self):
        config = {}
        for field, x_ranges in self.field_x_definitions.items():
            config[field] = {}
            for i, (x_min, x_max) in enumerate(x_ranges):
                if i >= len(self.column_layouts): continue
                layout = self.column_layouts[i]
                y_start = layout["top"]
                count = layout["count"]
                
                coords_lines = []
                for r in range(count):
                    curr_top = y_start - (r * self.row_height)
                    curr_bottom = curr_top - self.row_height
                    line = f"{x_min}, {curr_bottom}, {x_max}, {curr_top}"
                    coords_lines.append(line)
                config[field][f"col{i+1}"] = "\n".join(coords_lines)
        return config


def get_my_template_config_from_excel(config_path):
    """
    读取配置目录结构Excel，解析参数，返回 template_config 字典
    """
    config_path = Path(config_path)
    if not config_path.exists():
        sys_logger.info(f"❌ [配置错误] 找不到配置文件: {config_path}")
        return None

    excel = None
    wb = None
    try:


        excel = win32com.client.DispatchEx("Excel.Application")

        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(config_path), ReadOnly=True)
        try: ws = wb.Worksheets(1)
        except: ws = wb.ActiveSheet

        # --- 1. 读取单值参数 (第2行, 第1-4列) ---
        # Cell(行, 列)
        y_start_1 = float(ws.Cells(2, 1).Value)      # Col 1: Y_START_1
        y_start_others = float(ws.Cells(2, 2).Value) # Col 2: Y_START_2
        row_count_1 = int(ws.Cells(2, 3).Value)      # Col 3: row_count_1
        row_count_others = int(ws.Cells(2, 4).Value) # Col 4: row_count_2_3_4
        
        # --- 2. 初始化生成器 ---
        # 假设行高固定800，如果Excel里也有行高设置，可以加一列读取
        builder = CatalogConfigBuilder(row_height=800)

        # 添加排布局 (4排)
        builder.add_column_layout(y_start_1, row_count_1)        # 第1排
        builder.add_column_layout(y_start_others, row_count_others) # 第2排
        builder.add_column_layout(y_start_others, row_count_others) # 第3排
        builder.add_column_layout(y_start_others, row_count_others) # 第4排

        # --- 3. 读取列表参数 (第2-5行, 第5-8列) ---
        # 辅助函数：解析单元格 "min, max" -> (float, float)
        def parse_x_range(cell_val):
            try:
                parts = str(cell_val).replace('，', ',').split(',')
                return (float(parts[0]), float(parts[1]))
            except:
                sys_logger.info(f"⚠️ [配置警告] 无法解析X范围: {cell_val}")
                return (0.0, 0.0)

        # 我们需要读取 4 行数据 (对应 CAD 的 4 排)
        # Excel 行号: 2, 3, 4, 5
        rows_to_read = [2, 3, 4, 5]

        # 5列: 序号 (idx)
        idx_ranges = [parse_x_range(ws.Cells(r, 5).Value) for r in rows_to_read]
        # 6列: 图名 (name)
        name_ranges = [parse_x_range(ws.Cells(r, 6).Value) for r in rows_to_read]
        # 7列: 编号 (no)
        no_ranges = [parse_x_range(ws.Cells(r, 7).Value) for r in rows_to_read]
        # 8列: 规格 (spec)
        spec_ranges = [parse_x_range(ws.Cells(r, 8).Value) for r in rows_to_read]

        # --- 4. 填入生成器 ---
        builder.set_field_x_ranges("idx", idx_ranges)
        builder.set_field_x_ranges("name", name_ranges)
        builder.set_field_x_ranges("no", no_ranges)
        builder.set_field_x_ranges("spec", spec_ranges)

        print("✅ 目录结构配置读取成功！")
        return builder.generate()

    except Exception as e:
        sys_logger.info(f"❌ [配置读取失败] Excel 解析出错: {e}")
        import traceback; traceback.print_exc()
        return None
    finally:
        if wb: wb.Close(SaveChanges=0)
        if excel: excel.Quit()




#&&% 文件名修改1


def rename_drawings(
    folderpath, 
    split_prefix="-", 
    gu1=None, 
    gu2=None, 
    tumings_path=os.path.join(userpath,"配置","图纸名称修复.xlsx"), 
    naming_order=None
):
    """
    功能：根据Excel内容和原有编号重命名图纸，支持自定义排序列表。
    
    参数：
    - naming_order: 列表，决定命名顺序。
      支持的关键字: 'gu1', 'gu2', 'tuming' (Excel图名), 'idx' (原文件名切分后的右半部分)
      默认值: ['gu1', 'gu2', 'idx', 'tuming'] -> 结果如: 1#楼-建施-01-首层平面图.pdf
    """
    import pandas as pd
    import re    
    # 默认排序逻辑
    if naming_order is None:
        naming_order = ['gu1', 'gu2', 'idx', 'tuming']

    # --- 1. 环境准备 ---
    if not os.path.exists(folderpath):
        sys_logger.info(f"❌ 文件夹不存在: {folderpath}")
        return
    if not os.path.exists(tumings_path):
        sys_logger.info(f"❌ Excel文件不存在: {tumings_path}")
        return

    # --- 2. 读取 Excel 图名 ---
    try:
        df = pd.read_excel(tumings_path, header=None, usecols="A")
        # 转换为列表，去除NaN
        tuming_source = df.iloc[:, 0].fillna("").astype(str).tolist()
    except Exception as e:
        sys_logger.info(f"❌ 读取Excel失败: {e}")
        return

    # --- 3. 获取并排序文件 (按文件名中包含的数字) ---
    files = [f for f in os.listdir(folderpath) if f.lower().endswith('.pdf')]
    
    def get_sort_key(filename):
        name_no_ext = os.path.splitext(filename)[0]
        # 取分隔符右边的部分找数字
        if split_prefix in name_no_ext:
            suffix = name_no_ext.rpartition(split_prefix)[2]
            nums = re.findall(r'\d+', suffix)
            if nums: return int(nums[0]) # 按提取到的第一个数字排序
            return suffix
        return name_no_ext

    files.sort(key=get_sort_key)
    sys_logger.info(f"📂 准备处理 {len(files)} 个文件 (已按数字排序)")

    # --- 4. 循环处理 ---
    count = 0
    for i, filename in enumerate(files):
        if i >= len(tuming_source):
            print("⚠️ Excel 图名数量不足，停止处理剩余文件。")
            break

        name_no_ext, ext = os.path.splitext(filename)
        
        # 提取 'idx' (右边留下的字符)
        if split_prefix in name_no_ext:
            idx_val = name_no_ext.rpartition(split_prefix)[2]
        else:
            idx_val = name_no_ext

        # 准备数据字典
        data_map = {
            'gu1': str(gu1) if gu1 else "",
            'gu2': str(gu2) if gu2 else "",
            'tuming': tuming_source[i],
            'idx': idx_val,
            'idex': idx_val # 兼容你手误写的 idex
        }

        # --- 核心逻辑：按列表顺序拼接 ---
        parts = []
        for key in naming_order:
            val = data_map.get(key, "")
            # 只有当值不为空时才加入，避免出现 "--"
            if val: 
                parts.append(val)
        
        # 用 split_prefix (例如 "-") 连接所有部分
        new_base_name = split_prefix.join(parts)
        
        # 组装完整路径
        new_filename = f"{new_base_name}{ext}"
        old_path = os.path.join(folderpath, filename)
        new_path = os.path.join(folderpath, new_filename)

        # 重命名操作
        try:
            os.rename(old_path, new_path)
            sys_logger.info(f"✅ {filename} -> {new_filename}")
            count += 1
        except Exception as e:
            sys_logger.info(f"❌ 重命名失败 {filename}: {e}")

    sys_logger.info(f"🎉 全部完成，共处理 {count} 个文件。")

#&&% 文件名修改2







#&&% 模拟键盘调整打印图幅


# =============================================================================
# 1. 核心算法：带“防粘连”的双阶段捕获
# =============================================================================
def get_mouse_target_v3(step_idx, total_steps, prompt_text, prev_pos=None, dwell_time=2.0):
    """
    参数:
    - step_idx: 当前步骤序号
    - total_steps: 总步骤数
    - prev_pos: 上一步录入的坐标 (用于判断是否位置重叠)
    """
    step_str = f"[第 {step_idx}/{total_steps} 步]"
    sys_logger.info(f"\n🔹 {step_str} 目标: {prompt_text}")
    
    # --- 阶段 0: 防粘连检测 (Anti-Stick) ---
    # 如果上一步就在这儿，用户可能根本不动鼠标。
    # 强制要求用户移开，确保程序能检测到"为了当前步骤而做的新动作"。
    if prev_pos is not None:
        curr_x, curr_y = pyautogui.position()
        dist_from_prev = math.sqrt((curr_x - prev_pos[0])**2 + (curr_y - prev_pos[1])**2)
        
        # 如果距离上一步的位置小于 50 像素，提示移开
        if dist_from_prev < 50:
            sys_logger.info(f"   ⚠️  检测到鼠标仍在原位！请【移开鼠标再移回来】以激活此步骤...", end="\r")
            while True:
                curr_x, curr_y = pyautogui.position()
                dist_now = math.sqrt((curr_x - prev_pos[0])**2 + (curr_y - prev_pos[1])**2)
                if dist_now > 50: # 用户移开了
                    break
                time.sleep(0.1)

    # --- 阶段 1: 唤醒检测 (等待移动) ---
    sys_logger.info(f"   👁️  请移动鼠标去操作/寻找目标...                        ", end="\r")
    start_x, start_y = pyautogui.position()
    
    while True:
        curr_x, curr_y = pyautogui.position()
        dist = math.sqrt((curr_x - start_x)**2 + (curr_y - start_y)**2)
        
        if dist > 10: # 检测到移动
            break
        time.sleep(0.1)

    # --- 阶段 2: 驻留锁定 ---
    sys_logger.info(f"   🚀 已激活！请在目标位置【松手静止 {dwell_time} 秒】...         ", end="\r")
    
    last_x, last_y = pyautogui.position()
    stable_start_time = None
    
    while True:
        curr_x, curr_y = pyautogui.position()
        dist = math.sqrt((curr_x - last_x)**2 + (curr_y - last_y)**2)
        
        # 5像素容差
        if dist < 5:
            if stable_start_time is None:
                stable_start_time = time.time()
            
            elapsed = time.time() - stable_start_time
            
            # 进度条
            progress = int((elapsed / dwell_time) * 10)
            bar = "█" * progress + "░" * (10 - progress)
            sys_logger.info(f"   ⏳ 正在锁定 {step_str}: {bar} {elapsed:.1f}s", end="\r")
            
            if elapsed >= dwell_time:
                # --- 核心改进：即时反馈 ---
                sys_logger.info(f"   ✅ {step_str} 【{prompt_text.split(':')[0]}】 录入成功！坐标:({curr_x}, {curr_y})      ")
                print('\a') # 提示音
                return (curr_x, curr_y)
        else:
            stable_start_time = None
            last_x, last_y = curr_x, curr_y
            sys_logger.info(f"   🖱️  寻找中... 请在 {prompt_text[:10]}... 处停住      ", end="\r")
            
        time.sleep(0.1)

# =============================================================================
# 2. 安全输入逻辑
# =============================================================================
def safe_input_text(coords, text, desc="输入"):
    pyautogui.click(coords)
    time.sleep(0.5) 
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.5)
    pyautogui.press('delete')
    time.sleep(0.3)
    pyautogui.write(str(text), interval=0.1)
    # sys_logger.info(f"      ...{desc}: {text}") # 减少刷屏，保持清爽
    time.sleep(1.0) 

# =============================================================================
# 3. 主流程
# =============================================================================
def auto_setup_custom_paper_sizes(data_list):

    """
    auto_setup_custom_paper_sizes(dy_yonghu)

    """

    pyautogui.FAILSAFE = True 
    coords = {}

    print("\n" + "="*70)
    print("🛠️  CAD 批量助手 (终极防呆版)")
    print("="*70)
    
    # --- 关键前置提醒 ---
    print("❗❗❗ 【操作前必读 - 避免数据错误】 ❗❗❗")
    print("1. 打印机必须选为: DWG TO PDF.pc3")
    print("2. 基础图幅必须选择: ISO A1 (594.00 x 841.00 MM) 或其他【MM单位】图幅！")
    print("   ❌ 切勿选择 Inches (英寸) 单位的图幅，否则输入的数据会全部错误。")
    print("3. 【同位置操作提醒】: 如果连续两步按钮位置相同（如连续点“下一步”）：")
    print("   👉 必须先将鼠标【移开】一段距离，然后再【移回来】悬停，程序才能识别。")
    print("-" * 70)
    
    input("👉 确认已理解上述 3 点？按【回车】开始录入流程...")

    # --- 录入步骤定义 ---
    steps_map = [
        ("btn_props",      "特性按钮: 点击【特性(R)】"),
        ("tab_custom",     "选项卡: 点击【自定义图纸尺寸】"), 
        ("btn_add",        "添加按钮: 点击【添加(A)】"),
        ("btn_next_1",     "向导1: 点击【下一步(N)】 (开始)"),
        ("input_width",    "宽度框: 点击【宽度(W)】输入区域"),
        ("input_height",   "高度框: 点击【高度(E)】输入区域"),
        ("btn_next_2",     "向导2: 点击【下一步(N)】 (宽高页)"),
        ("input_top",      "上边距: 点击【上(T)】输入区域"),
        ("input_bottom",   "下边距: 点击【下(O)】输入区域"),
        ("input_left",     "左边距: 点击【左(L)】输入区域"),
        ("input_right",    "右边距: 点击【右(R)】输入区域"),
        ("btn_next_3",     "向导3: 点击【下一步(N)】 (边距页)"),
        ("btn_next_4",     "向导4: 点击【下一步(N)】 (跳过改名)"),
        ("btn_finish",     "完成按钮: 点击【完成】"),
        ("btn_ok",         "确定按钮: 点击底部的【确定】")
    ]

    # --- 循环录入 ---
    total = len(steps_map)
    prev_coordinate = None # 用于记录上一步的坐标

    for idx, (key, prompt) in enumerate(steps_map, 1):
        # 调用 V3 版捕获函数
        current_coord = get_mouse_target_v3(
            step_idx=idx, 
            total_steps=total, 
            prompt_text=prompt, 
            prev_pos=prev_coordinate, # 传入上一步坐标用于防粘连
            dwell_time=2.0
        )
        
        coords[key] = current_coord
        prev_coordinate = current_coord # 更新上一步坐标
        time.sleep(0.5) 

    print("\n🎉 坐标校准全部完成！请将鼠标移至安全区域，3秒后自动执行...")
    time.sleep(3)

    # --- 批量执行 ---
    CLICK_WAIT = 1.0
    WINDOW_WAIT = 2.0
    
    print("\n🚀 开始批量写入数据...")

    for i, item in enumerate(data_list):
        w, h, t, b, l, r = item
        sys_logger.info(f"[{i+1}/{len(data_list)}] 正在写入: {w} x {h}")

        try:
            # 1. 特性
            pyautogui.click(coords["btn_props"])
            time.sleep(WINDOW_WAIT) 
            # 2. 选项卡
            pyautogui.click(coords["tab_custom"])
            time.sleep(CLICK_WAIT)
            # 3. 添加
            pyautogui.click(coords["btn_add"])
            time.sleep(CLICK_WAIT)
            # 4. 下一步
            pyautogui.click(coords["btn_next_1"])
            time.sleep(CLICK_WAIT)
            # 5. 宽高
            safe_input_text(coords["input_width"], w, "宽度")
            safe_input_text(coords["input_height"], h, "高度")
            # 下一步
            pyautogui.click(coords["btn_next_2"])
            time.sleep(CLICK_WAIT)
            # 6. 边距
            safe_input_text(coords["input_top"], t, "上边距")
            safe_input_text(coords["input_bottom"], b, "下边距")
            safe_input_text(coords["input_left"], l, "左边距")
            safe_input_text(coords["input_right"], r, "右边距")
            # 下一步
            pyautogui.click(coords["btn_next_3"])
            time.sleep(CLICK_WAIT)
            # 下一步
            pyautogui.click(coords["btn_next_4"])
            time.sleep(CLICK_WAIT)
            # 完成
            pyautogui.click(coords["btn_finish"])
            time.sleep(CLICK_WAIT)
            # 确定
            pyautogui.click(coords["btn_ok"])
            
            # PMP 保存处理
            time.sleep(WINDOW_WAIT) 
            pyautogui.press('enter') 
            time.sleep(WINDOW_WAIT) 

        except Exception as e:
            sys_logger.info(f"❌ 出错: {e}")
            break

    print("\n✅ 所有图纸尺寸添加完成！")

# 数据准备
dy_yonghu = [
    ("1337.63", "841.00", "13", "13", "13", "13"), 
    ("1486.25", "841.00", "13", "13", "13", "13"), 
    ("1051.25", "594.00", "13", "13", "13", "13"), 
    ("1261.50", "594.00", "13", "13", "13", "13"), 
    ("1471.75", "594.00", "13", "13", "13", "13"), 
    ("742.50", "420.00", "13", "13", "13", "13"),  
    ("891.00", "420.00", "13", "13", "13", "13"),  
    ("1039.50", "420.00", "13", "13", "13", "13")  
]

#天正数据

dy_yonghu = [
    ("1338.00", "841.00", "0", "0", "0", "0"), 
    ("1486.00", "841.00", "0", "0", "0", "0"), 
    ("1051.00", "594.00", "0", "0", "0", "0"), 
    ("1261.00", "594.00", "0", "0", "0", "0"), 
    ("1471.00", "594.00", "0", "0", "0", "0"), 
    ("743.00", "420.00", "0", "0", "0", "0"),  
    ("891.00", "420.00", "0", "0", "0", "0"),  
    ("1041.00", "420.00", "0", "0", "0", "0")  
]



#&&% 获取当前图幅尺寸


def list_current_printer_papers(printer_name="DWG To PDF.pc3"):
    """
    列出指定打印机的所有纸张名称 (内部名 vs 显示名)
    """
    sys_logger.info(f"🔌 正在连接 AutoCAD...")
    try:
        acad = win32com.client.GetActiveObject("AutoCAD.Application")
        doc = acad.ActiveDocument
        layout = doc.ActiveLayout
    except Exception as e:
        sys_logger.info(f"❌ 无法连接 AutoCAD: {e}")
        return

    sys_logger.info(f"⚙️ 正在切换打印机配置为: {printer_name} ...")
    try:
        # 必须先设置 ConfigName，否则获取的是默认打印机的纸张
        layout.RefreshPlotDeviceInfo()
        layout.ConfigName = printer_name
        # 再次刷新以加载该打印机的图幅列表
        layout.RefreshPlotDeviceInfo() 
    except Exception as e:
        sys_logger.info(f"❌ 设置打印机失败 (可能拼写错误或驱动不存在): {e}")
        return

    # 获取所有纸张的内部名
    canonical_names = layout.GetCanonicalMediaNames()
    
    print("\n" + "="*90)
    sys_logger.info(f"🖨️  打印机 [{printer_name}] 纸张列表侦探")
    sys_logger.info(f"📋 共发现 {len(canonical_names)} 种纸张")
    print("="*90)
    sys_logger.info(f"| {'内部名 (写代码用这个!)':<50} | {'显示名 (你在CAD里看到的)':<35} |")
    print("-" * 90)

    count = 0
    user_defined_count = 0

    for c_name in canonical_names:
        # 获取对应的本地显示名
        l_name = layout.GetLocaleMediaName(c_name)
        
        # 过滤逻辑：我们重点关注 "User" (自定义) 或 "Full bleed" (扩展)
        # 如果你想看所有纸张，把下面的 if 注释掉即可
        is_custom = "User" in c_name or "Custom" in c_name
        is_full_bleed = "full_bleed" in c_name
        
        # 这里我设置只显示 自定义纸张 和 A0/A1/A2 的扩展纸张，防止列表太长刷屏
        if is_custom or (is_full_bleed and "A0" in c_name): 
            sys_logger.info(f"| {c_name:<50} | {l_name:<35} |")
            count += 1
            if is_custom:
                user_defined_count += 1
    
    print("-" * 90)
    sys_logger.info(f"✅ 扫描结束。显示了 {count} 个相关结果。")
    sys_logger.info(f"🔍 其中包含 {user_defined_count} 个用户自定义图幅 (UserDefined)。")
    print("💡 请将左侧的【内部名】复制到你的 Python 代码中，替换掉报错的纸张名称。")


#&&% 配置字体


def replace_cad_fonts_incremental():
    # === 路径配置 ===
    cad_root_path = r"C:\Program Files\Autodesk\AutoCAD 2021"
    current_fonts_path = os.path.join(cad_root_path, "Fonts")
    user_source_path = r"D:\Mypro\基础服务\用户1\配置\Fonts"

    # === 1. 权限检查 ===
    if not is_admin():
        print("错误：此操作需要修改 C 盘 Program Files 文件夹，请以【管理员身份】运行此脚本。")
        return

    # === 2. 预检查 ===
    if not os.path.exists(user_source_path):
        sys_logger.info(f"错误：用户源文件夹不存在 -> {user_source_path}")
        return

    # === 3. 关闭 CAD 进程 ===
    # 在进行任何文件操作前，先杀进程并等待
    close_all_cad_processes()

    print("--- 开始处理 CAD 字体 (增量备份模式) ---")

    try:
        # === 4. 处理备份 (Fonts -> FontsbatN) ===
        if os.path.exists(current_fonts_path):
            backup_base_name = "Fontsbat"
            counter = 0
            
            # 循环查找下一个可用的备份文件夹名 (Fontsbat0, Fontsbat1...)
            while True:
                candidate_name = f"{backup_base_name}{counter}"
                candidate_path = os.path.join(cad_root_path, candidate_name)
                
                if not os.path.exists(candidate_path):
                    final_backup_path = candidate_path
                    break
                counter += 1
            
            # 执行重命名
            sys_logger.info(f"正在备份原厂字体...")
            os.rename(current_fonts_path, final_backup_path)
            sys_logger.info(f"✅ 成功备份为：{os.path.basename(final_backup_path)}")
            
        else:
            print("注意：原始 Fonts 文件夹不存在，跳过备份步骤。")

        # === 5. 复制用户字体 ===
        sys_logger.info(f"正在复制用户字体...")
        sys_logger.info(f"源：{user_source_path}")
        sys_logger.info(f"目标：{current_fonts_path}")
        
        shutil.copytree(user_source_path, current_fonts_path)
        print("--- ✅ 字体替换全部完成 ---")

    except PermissionError:
        print("\n[关键错误]：权限不足！无法访问 C:\Program Files。")
        print("请右键点击脚本或 IDE，选择“以管理员身份运行”。")
    except OSError as e:
        sys_logger.info(f"\n[错误]：文件操作失败。")
        sys_logger.info(f"尽管尝试关闭了 CAD，文件可能仍被占用，或者有其他程序正在使用该目录。")
        sys_logger.info(f"系统报错信息：{e}")
    except Exception as e:
        sys_logger.info(f"\n[未知错误]：{e}")
def is_admin():
    """检查当前是否获得管理员权限"""
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False





def sanitize_filename(name):
    """【COMMON-001-AUX】文件名清洗"""
    if not name: return ""
    return re.sub(r'[\\/*?:"<>|]', '_', str(name))



#&&% 模型空间窗口打印

def export_model_window_pure(
        point_a,
        point_b,
        pdf_fullpath,
        *,
        device="DWG To PDF.pc3",
        media="ISO_A3_(420.00_x_297.00_MM)",
        ctb="monochrome.ctb",
        rotation=0,
        xiubukuan = 25

):
    """
    【函数编号】: PRINT-001 (A0修正版)
    【所属模块】: 打印核心模块 (Print Automation Core)
    【功能描述】: 
        模型空间窗口打印原子函数（增强版：坐标自适应 + 视觉反馈 + A0特殊旋转修正）。

    【A0 特殊修正】:
        针对 ISO_A0_(841.00_x_1189.00_MM) 图纸，自动将旋转角取反 (0<->1)。
    """

    # 0. 连接环境
    doc=C.doc

    # ———————————————— 1. 坐标标准化 ————————————————
    try:
        x1, y1 = float(point_a[0]), float(point_a[1])
        x2, y2 = float(point_b[0]), float(point_b[1])
        
        min_x, max_x = min(x1, x2), max(x1, x2)
        min_y, max_y = min(y1, y2), max(y1, y2)
        
        # 内部使用
        lower_left_standard = (min_x, min_y)
        upper_right_standard = (max_x, max_y)
    except Exception as e:
        sys_logger.info(f"❌ [PRINT-001] 坐标解析失败: {e}")
        return False

    # ———————————————— 2. 内部辅助：视觉反馈 ————————————————
    def _draw_boundary_markers():
        """在打印区域边缘绘制临时标记线"""
        try:
            marker_width = xiubukuan 
            
            # 绘制左侧边缘 (垂直线)
            draw_lwpolyline(
                coords3d=[(min_x, min_y, 0), (min_x, max_y, 0)],
                layer_name="dy_zhuanyong",
                width=marker_width,
                color=256,
                closed=False
            )
            # 绘制底侧边缘 (水平线)
            draw_lwpolyline(
                coords3d=[(min_x, min_y, 0), (max_x, min_y, 0)],
                layer_name="dy_zhuanyong",
                width=marker_width,
                color=256,
                closed=False
            )
        except Exception as e:
            sys_logger.info(f"⚠️ [PRINT-001] 绘制标记线警告: {e}")

    try:
        # ———————————————— 3. 执行打印逻辑 ————————————————
        
        # A. 绘制边界标记
        _draw_boundary_markers()

        # B. 获取并配置布局对象
        lay = doc.ActiveLayout 
        lay.RefreshPlotDeviceInfo()
        
        # ——————————————————————————————————————————————————————————————
        # ⚙️【A0 旋转修正逻辑】 (Special Logic for A0)
        # 描述：A0 图纸定义的宽长方向通常与 A3 等相反，需要对调旋转角。
        # ——————————————————————————————————————————————————————————————
        final_rotation = rotation # 默认保持原值
        
        if media == "ISO_A0_(841.00_x_1189.00_MM)":
            if rotation == 0:
                final_rotation = 1
                sys_logger.info(f"ℹ️ [A0修正] 检测到 A0 图纸，强制旋转: 0 -> 1")
            elif rotation == 1:
                final_rotation = 0
                sys_logger.info(f"ℹ️ [A0修正] 检测到 A0 图纸，强制旋转: 1 -> 0")
        # ——————————————————————————————————————————————————————————————

        # 基础配置
        lay.ConfigName         = device
        lay.PaperUnits         = 1 # acMillimeters
        lay.CanonicalMediaName = media
        lay.StyleSheet         = ctb
        lay.PlotRotation       = final_rotation  # <--- 使用修正后的角度
        lay.CenterPlot         = True
        lay.StandardScale      = 0        
        lay.UseStandardScale   = True 
        lay.PlotWithPlotStyles = True
        lay.PlotHidden         = False

        # C. 窗口设置
        ll = VARIANT(VT_ARRAY | VT_R8, [min_x, min_y])
        ur = VARIANT(VT_ARRAY | VT_R8, [max_x, max_y])
        lay.SetWindowToPlot(ll, ur)
        lay.PlotType = 4  # acWindow

        # D. 执行物理输出
        doc.SetVariable("BACKGROUNDPLOT", 0)
        doc.Plot.QuietErrorMode = True
        
        if os.path.exists(pdf_fullpath):
            try: os.remove(pdf_fullpath)
            except: pass

        doc.Plot.PlotToFile(pdf_fullpath)
        
        if os.path.exists(pdf_fullpath):
            sys_logger.info(f"✅ [Model] 输出成功: {os.path.basename(pdf_fullpath)}")
            return True
        else:
            sys_logger.info(f"❌ [Model] 文件未生成: {os.path.basename(pdf_fullpath)}")
            return False

    except Exception as e:
        sys_logger.info(f"❌ [Model] 打印流程异常: {e}")
        return False

#&&% 模型空间LISP窗口打印export_model_window_lisp_fit

def export_model_window_lisp_fit(
        point_a,
        point_b,
        pdf_fullpath,  # <--- 参数名是这个
        *,
        device="DWG To PDF.pc3",
        media="ISO_A3_(420.00_x_297.00_MM)",
        ctb="monochrome.ctb",
        rotation=0,
        xiubukuan=25
):
    """
    【函数编号】: PRINT-002 (LISP 强力修正版 - V2修复Bug)
    """
    import os
    import time

    # 0. 连接环境
    doc = C.doc

    # ———————————————— 1. 坐标标准化 ————————————————
    try:
        x1, y1 = float(point_a[0]), float(point_a[1])
        x2, y2 = float(point_b[0]), float(point_b[1])
        
        min_x, max_x = min(x1, x2), max(x1, x2)
        min_y, max_y = min(y1, y2), max(y1, y2)
    except Exception as e:
        sys_logger.info(f"❌ [PRINT-LISP] 坐标解析失败: {e}")
        return False

    # ———————————————— 2. 内部辅助：视觉反馈 ————————————————
    def _draw_boundary_markers():
        try:
            if 'draw_lwpolyline' in globals():
                draw_lwpolyline(
                    coords3d=[(min_x, min_y, 0), (min_x, max_y, 0)],
                    layer_name="dy_zhuanyong",
                    width=xiubukuan, color=256, closed=False
                )
                draw_lwpolyline(
                    coords3d=[(min_x, min_y, 0), (max_x, min_y, 0)],
                    layer_name="dy_zhuanyong",
                    width=xiubukuan, color=256, closed=False
                )
        except Exception:
            pass 

    # ———————————————— 3. 执行打印逻辑 ————————————————
    try:
        # A. 绘制边界标记
        #_draw_boundary_markers()

        # B. A0 旋转修正逻辑
        final_rotation = rotation
        if media == "ISO_A0_(841.00_x_1189.00_MM)":
            if rotation == 0:
                final_rotation = 1
                sys_logger.info(f"ℹ️ [A0修正] 检测到 A0 图纸，强制旋转: 0 -> 1")
            elif rotation == 1:
                final_rotation = 0
                sys_logger.info(f"ℹ️ [A0修正] 检测到 A0 图纸，强制旋转: 1 -> 0")

        # C. 构造 LISP 参数
        p1_str = f"{min_x},{min_y}"
        p2_str = f"{max_x},{max_y}"

        # 🔥🔥🔥 修正点：使用正确的变量名 pdf_fullpath 🔥🔥🔥
        pdf_path_lisp = pdf_fullpath.replace("\\", "/") 

        orientation_str = "Portrait" if final_rotation == 1 else "Landscape"

        # 预清理文件
        if os.path.exists(pdf_fullpath):
            try: os.remove(pdf_fullpath)
            except: pass

        # D. 构造 LISP 命令流
        lisp_cmd = f"""(command "-plot" 
            "Yes" "Model" "{device}" "{media}" "Millimeters" 
            "{orientation_str}" "No" "Window" "{p1_str}" "{p2_str}" 
            "Fit" "Center" "Yes" "{ctb}" "Yes" "As displayed" 
            "{pdf_path_lisp}" "No" "Yes"
        ) """

        # E. 发送命令
        clean_cmd = " ".join([line.strip() for line in lisp_cmd.split('\n') if line.strip()])
        doc.SendCommand(clean_cmd + "\n")

        # F. 等待文件生成
        max_wait = 60 # 稍微延长等待时间
        start_time = time.time()
        while time.time() - start_time < max_wait:
            if os.path.exists(pdf_fullpath):
                time.sleep(0.5) 
                if os.path.getsize(pdf_fullpath) > 0:
                    sys_logger.info(f"✅ [LISP-Model] 输出成功: {os.path.basename(pdf_fullpath)}")
                    return True
            time.sleep(0.5)
        
        sys_logger.info(f"❌ [LISP-Model] 超时未生成文件: {os.path.basename(pdf_fullpath)}")
        return False

    except Exception as e:
        sys_logger.info(f"❌ [LISP-Model] 打印流程异常: {e}")
        return False




#&&% 图纸空间窗口打印


def export_layout_window_pure(
        point_a,
        point_b,
        pdf_fullpath,           # <--- 直接传入完整的输出路径
        layout_name,            # <--- 布局必须指定
        *,
        device="DWG To PDF.pc3",
        media="ISO_A3_(420.00_x_297.00_MM)",
        ctb="monochrome.ctb",
        rotation=0
):
    """
    【函数编号】: PRINT-002
    【所属模块】: 打印核心模块 (Print Automation Core)
    【功能描述】: 
        图纸空间(布局)窗口打印原子函数（增强版：坐标自适应）。

        本函数专用于处理 Layout (布局) 空间的打印任务。与模型空间打印不同，
        布局打印需要严格管理 ActiveLayout 的切换和刷新机制。
        
        核心特性：
        1. 【坐标自适应】: 自动标准化输入的两个对角点坐标，调用者无需区分左下/右上。
        2. 【时序修正】: 针对 COM 接口特性，强制执行 "先注入窗口坐标 -> 后切换打印模式" 
           的顺序，修复了布局空间下 SetWindowToPlot 容易失效的 Bug。
        3. 【布局守护】: 打印前强制校验并切换至目标布局，防止在错误的空间执行打印。
        4. 【静默输出】: 包含异常捕获、旧文件清理及后台弹窗抑制。

    【参数详解】:
        - point_a (tuple/list): 
            打印窗口的第一个角点坐标 (x, y)。
        - point_b (tuple/list): 
            打印窗口的对角点坐标 (x, y)。
        - pdf_fullpath (str): 
            输出 PDF 文件的完整绝对路径。
        - layout_name (str): 
            目标布局的名称（如 "Layout1" 或 "总图"）。
        - device, media, ctb, rotation: 
            参考 PRINT-001 的标准打印参数。

    【返回值】:
        - bool: 打印成功返回 True，失败返回 False。
    """
    from  CAD_file_operations   import  switch_to_layout
    # 0. 连接环境
    doc=C.doc

    # ———————————————— 1. 坐标标准化 ————————————————
    try:
        x1, y1 = float(point_a[0]), float(point_a[1])
        x2, y2 = float(point_b[0]), float(point_b[1])
        
        min_x, max_x = min(x1, x2), max(x1, x2)
        min_y, max_y = min(y1, y2), max(y1, y2)
    except Exception as e:
        sys_logger.info(f"❌ [PRINT-002] 坐标解析失败: {e}")
        return False

    try:
        # ———————————————— 2. 环境准备 ————————————————
        
        # A. 切换到目标布局
        if not switch_to_layout(layout_name): 
            sys_logger.info(f"❌ [Layout] 无法切换到布局: {layout_name}")
            return False
            
        lay = doc.ActiveLayout

        # B. 挂载打印机
        lay.ConfigName = device
        
        # C. 刷新设备信息
        lay.RefreshPlotDeviceInfo()

        # ———————————————— 3. A0 旋转修正逻辑 ————————————————
        # 必须在赋值 PlotRotation 之前计算好
        final_rotation = rotation
        
        if media == "ISO_A0_(841.00_x_1189.00_MM)":
            if rotation == 0:
                final_rotation = 1
                sys_logger.info(f"ℹ️ [Layout-A0] 检测到 A0 图纸，强制旋转: 0 -> 1")
            elif rotation == 1:
                final_rotation = 0
                sys_logger.info(f"ℹ️ [Layout-A0] 检测到 A0 图纸，强制旋转: 1 -> 0")

        # ———————————————— 4. 注入窗口坐标 ————————————————
        # 先注入坐标
        p1_args = [min_x, min_y]
        p2_args = [max_x, max_y]
        
        lay.SetWindowToPlot(
            VARIANT(VT_ARRAY | VT_R8, p1_args), 
            VARIANT(VT_ARRAY | VT_R8, p2_args)
        )

        # ———————————————— 5. 配置打印参数 ————————————————
        # 后切换模式
        lay.PlotType = 4  # acWindow

        lay.CanonicalMediaName = media
        lay.StyleSheet         = ctb
        lay.PlotRotation       = final_rotation  # <--- 使用修正后的角度
        lay.CenterPlot         = True
        
        lay.UseStandardScale   = True # Fit
        lay.StandardScale      = 0      
        lay.PlotWithPlotStyles = True
        
        # ———————————————— 6. 执行物理输出 ————————————————
        doc.SetVariable("BACKGROUNDPLOT", 0)
        doc.Plot.QuietErrorMode = True

        if os.path.exists(pdf_fullpath):
            try: os.remove(pdf_fullpath)
            except: pass
            
        doc.Plot.PlotToFile(pdf_fullpath)
        
        if os.path.exists(pdf_fullpath):
            sys_logger.info(f"✅ [Layout] 输出成功: {os.path.basename(pdf_fullpath)} ({layout_name})")
            return True
        else:
            sys_logger.info(f"❌ [Layout] 文件未生成: {os.path.basename(pdf_fullpath)}")
            return False

    except Exception as e:
        sys_logger.info(f"❌ [Layout] 打印流程异常 ({layout_name}): {e}")
        return False


#&&% 图纸空间打印边距修正版备用

def export_layout_window_pure_bianju(
        point_a,
        point_b,
        pdf_fullpath,
        layout_name,
        *,
        device="DWG To PDF.pc3",
        media="ISO_A3_(420.00_x_297.00_MM)",
        ctb="monochrome.ctb",
        rotation=0
):
    """
    【函数编号】: PRINT-002-V3 (A0修正版)
    【所属模块】: 打印核心模块 (Print Automation Core)
    【功能描述】: 
        布局空间窗口打印原子函数（增强版：坐标自适应 + 边距微调 + A0旋转修正）。

        本函数主要用于解决 CAD 打印时“外框线被切除”或“图纸内容稍微偏移”的问题。
        同时内置了针对 ISO A0 图纸的旋转角自动翻转逻辑。

    【参数详解】:
        - point_a, point_b: 打印区域的任意对角点坐标。
        - 其他参数同 PRINT-002。
    """
    from  CAD_file_operations   import  switch_to_layout

    # ——————————————————————————————————————————————————————
    # ⚙️ 内部边距微调区 (单位: 图纸单位/mm)
    # ——————————————————————————————————————————————————————
    # 场景：如果打印出来的 PDF 经常看不到最左边的图框线，尝试将 pad_L 设为 1.0 或 2.0
    pad_L = 0  # 左边距外扩 (Left Padding)
    pad_R = 0  # 右边距外扩 (Right Padding)
    pad_T = 0  # 上边距外扩 (Top Padding)
    pad_B = 0  # 下边距外扩 (Bottom Padding)
    # ——————————————————————————————————————————————————————

    doc=C.doc

    # ———————————————— 1. 坐标计算与修正 ————————————————
    try:
        # 1.1 解析输入并计算原始包围盒 (Min/Max)
        x1, y1 = float(point_a[0]), float(point_a[1])
        x2, y2 = float(point_b[0]), float(point_b[1])
        
        raw_min_x, raw_max_x = min(x1, x2), max(x1, x2)
        raw_min_y, raw_max_y = min(y1, y2), max(y1, y2)

        # 1.2 应用边距修正 (Padding Application)
        # 向左扩=X减小，向右扩=X增大，向下扩=Y减小，向上扩=Y增大
        final_min_x = raw_min_x - pad_L
        final_min_y = raw_min_y - pad_B
        final_max_x = raw_max_x + pad_R
        final_max_y = raw_max_y + pad_T

        # 准备 COM 参数
        p1_args = [final_min_x, final_min_y]
        p2_args = [final_max_x, final_max_y]

    except Exception as e:
        sys_logger.info(f"❌ [PRINT-002-V3] 坐标计算失败: {e}")
        return False

    try:
        # ———————————————— 2. 环境配置 ————————————————
        # 切换布局
        if not switch_to_layout(layout_name): 
            return False
        lay = doc.ActiveLayout

        # 必须先配置设备并刷新
        lay.ConfigName = device
        lay.RefreshPlotDeviceInfo()

        # ———————————————— 3. A0 旋转修正逻辑 ————————————————
        # 必须在赋值 PlotRotation 之前计算好
        final_rotation = rotation
        
        if media == "ISO_A0_(841.00_x_1189.00_MM)":
            if rotation == 0:
                final_rotation = 1
                sys_logger.info(f"ℹ️ [Layout-V3] 检测到 A0 图纸，强制旋转: 0 -> 1")
            elif rotation == 1:
                final_rotation = 0
                sys_logger.info(f"ℹ️ [Layout-V3] 检测到 A0 图纸，强制旋转: 1 -> 0")

        # ———————————————— 4. 注入窗口 (核心步骤) ————————————————
        # 先设置窗口范围
        lay.SetWindowToPlot(
            VARIANT(VT_ARRAY | VT_R8, p1_args), 
            VARIANT(VT_ARRAY | VT_R8, p2_args)
        )

        # ———————————————— 5. 打印参数 ————————————————
        # 后设置模式 (关键顺序)
        lay.PlotType = 4  # acWindow
        
        lay.CanonicalMediaName = media
        lay.StyleSheet         = ctb
        lay.PlotRotation       = final_rotation # <--- 使用修正后的角度
        lay.CenterPlot         = True
        lay.UseStandardScale   = True 
        lay.StandardScale      = 0 # Fit
        lay.PlotWithPlotStyles = True
        
        # ———————————————— 6. 执行输出 ————————————————
        doc.SetVariable("BACKGROUNDPLOT", 0)
        doc.Plot.QuietErrorMode = True

        if os.path.exists(pdf_fullpath):
            try: os.remove(pdf_fullpath)
            except: pass
        
        doc.Plot.PlotToFile(pdf_fullpath)
        
        # 日志 (带 Padding 信息)
        pad_info = ""
        if any([pad_L, pad_R, pad_T, pad_B]):
            pad_info = f" (Pad L:{pad_L}/R:{pad_R}/T:{pad_T}/B:{pad_B})"
            
        sys_logger.info(f"✅ [Layout] 输出成功: {os.path.basename(pdf_fullpath)}{pad_info}")
        return True

    except Exception as e:
        sys_logger.info(f"❌ [Layout] 打印异常: {e}")
        return False

#&&% 布局空间LISP窗口打印

def export_layout_window_lisp_fit_v1(
        point_a,
        point_b,
        pdf_fullpath,
        layout_name,  # <--- 布局名称必传，LISP需要用它来切换或指定目标
        *,
        device="DWG To PDF.pc3",
        media="ISO_A3_(420.00_x_297.00_MM)",
        ctb="monochrome.ctb",
        rotation=0
):
    """
    【函数编号】: PRINT-004 (LISP-Layout 终极版 - 旧版本，已重命名)
    【所属模块】: 打印核心模块
    【功能描述】: 
        图纸空间(Layout)窗口打印原子函数（LISP 命令版）。
        完全绕过 COM 接口的 Window/Layout 切换 BUG，直接使用命令行精准控制。
    
    【核心优势】:
        1. 稳健切换: 显式激活目标布局，确保打印发生在正确的空间。
        2. 强制充满: 使用 "Fit" 比例，确保 1:1 的图纸空间内容撑满图纸。
        3. 零偏移: 利用 LISP 的 "Center" 命令强制居中。
    """
    import os
    import time

    # 0. 环境准备
    doc = C.doc
    
    # ———————————————— 1. 布局激活保护 ————————————————
    # LISP 的 -plot 命令可以直接指定打印哪个布局，但为了保险起见，
    # 我们先尝试切换到那个布局，确保坐标是在该布局坐标系下获取的。
    try:
        # 如果当前布局不是目标布局，切过去
        if doc.ActiveLayout.Name != layout_name:
            # 尝试通过 COM 切换 (比较温和)
            doc.ActiveLayout = doc.Layouts.Item(layout_name)
    except:
        # 如果切换失败，可能不影响，因为 LISP 命令里还会指定一次 LayoutName
        sys_logger.info(f"⚠️ 布局预切换警告: 可能已经在目标布局或切换受阻")

    # ———————————————— 2. 坐标标准化 ————————————————
    try:
        x1, y1 = float(point_a[0]), float(point_a[1])
        x2, y2 = float(point_b[0]), float(point_b[1])
        
        min_x, max_x = min(x1, x2), max(x1, x2)
        min_y, max_y = min(y1, y2), max(y1, y2)
    except Exception as e:
        sys_logger.info(f"❌ [LISP-Layout] 坐标解析失败: {e}")
        return False

    try:
        # ———————————————— 3. 构造 LISP 参数 ————————————————
        
        # A. A0 修正逻辑
        final_rotation = rotation
        if media == "ISO_A0_(841.00_x_1189.00_MM)":
            final_rotation = 1 if rotation == 0 else 0
            
        # B. 坐标与路径
        p1_str = f"{min_x},{min_y}"
        p2_str = f"{max_x},{max_y}"
        pdf_path_lisp = pdf_fullpath.replace("\\", "/")
        
        # C. 纸张方向
        orientation_str = "Portrait" if final_rotation == 1 else "Landscape"

        # D. 清理旧文件
        if os.path.exists(pdf_fullpath):
            try: os.remove(pdf_fullpath)
            except: pass

        # ———————————————— 4. 构造 LISP 命令流 ————————————————
        # 注意: 布局空间的 -PLOT 参数顺序与模型空间略有不同
        # 关键参数:
        # "Yes"     -> 详细配置
        # layout_name -> 显式输入布局名称 (代替 "Model")
        # "Window"  -> 打印区域类型
        # "1:1"     -> 布局空间通常是 1:1 打印，但为了容错，这里也用 "Fit"? 
        #              建议：布局空间如果是画在 1:1 的图纸上，用 "1:1" 最准。
        #              但是为了统一“充满”的效果，"Fit" 更安全。这里采用 "Fit"。
        
        lisp_cmd = f"""(command "-plot" 
            "Yes"                       ;详细配置
            "{layout_name}"             ;输入布局名称 🔥
            "{device}"                  ;打印机
            "{media}"                   ;图纸尺寸
            "Millimeters"               ;单位
            "{orientation_str}"         ;方向
            "No"                        ;不反向
            "Window"                    ;窗口打印
            "{p1_str}"                  ;角点1
            "{p2_str}"                  ;角点2
            "Fit"                       ;比例 (Fit布满 / 1:1)
            "Center"                    ;偏移 (居中)
            "Yes"                       ;按样式打印
            "{ctb}"                     ;样式表
            "Yes"                       ;打印线宽
            "No"                        ;布局空间通常不需要“按显示着色”，选 No 即可 (Scale lineweights?) -> 这里是 Plot with lineweights
                                        ;注意：布局空间接下来的提示可能略有不同
                                        ;通常是: Plot with plotting styles? Yes
                                        ;       Plot with lineweights? Yes
                                        ;       Scale lineweights? No (通常不缩放线宽)
                                        ;       Plot paper space first? No
                                        ;       Hide paperspace objects? No
            "No"                        ;Scale lineweights (缩放线宽)
            "No"                        ;Plot paper space first (先打印图纸空间? - 通常布局就是图纸空间，选No)
            "No"                        ;Hide paperspace objects (隐藏图纸空间对象?)
            "{pdf_path_lisp}"           ;文件名
            "No"                        ;保存修改
            "Yes"                       ;继续打印
        ) """
        
        # 【特别注意】：布局空间的 -PLOT 交互流程比模型空间多几个选项
        # 上面的流程是标准流程，如果卡住，说明某些 CAD 版本提示不同。
        # 稳健做法：使用简化版参数，或者测试一下。
        # 针对 AutoCAD 2021+ 的布局打印流程优化：
        
        lisp_cmd_optimized = f"""(command "-plot" 
            "Yes" "{layout_name}" "{device}" "{media}" "Millimeters" 
            "{orientation_str}" "No" "Window" "{p1_str}" "{p2_str}" 
            "Fit" "Center" "Yes" "{ctb}" "Yes" 
            "No" "No" "No" 
            "{pdf_path_lisp}" "No" "Yes"
        ) """

        # ———————————————— 5. 执行命令 ————————————————
        clean_cmd = " ".join([line.strip() for line in lisp_cmd_optimized.split('\n') if line.strip()])
        
        # sys_logger.info(f"🚀 发送 Layout 打印指令...") # 调试用
        doc.SendCommand(clean_cmd + "\n")

        # ———————————————— 6. 等待结果 ————————————————
        # 🔥 关键改进：大幅增加超时时间，因为打印可能需要较长时间
        # 用户反馈：实际上PDF已经打印出来了，只是超过了15秒超时
        max_wait = 60  # 从15秒增加到120秒
        start_time = time.time()
        last_log_time = start_time

        while time.time() - start_time < max_wait:
            if os.path.exists(pdf_fullpath) and os.path.getsize(pdf_fullpath) > 0:
                elapsed = time.time() - start_time
                sys_logger.info(f"✅ [LISP-Layout] 输出成功: {os.path.basename(pdf_fullpath)} (耗时: {elapsed:.1f}s)")
                return True

            # 每10秒输出一次等待进度
            current_time = time.time()
            if current_time - last_log_time >= 10:
                elapsed = current_time - start_time
                sys_logger.info(f"⏳ [LISP-Layout] 等待PDF生成中... (已等待: {elapsed:.0f}s / {max_wait}s)")
                last_log_time = current_time

            time.sleep(0.5)

        sys_logger.info(f"❌ [LISP-Layout] 超时: {os.path.basename(pdf_fullpath)} (等待了{max_wait}秒)")
        return False

    except Exception as e:
        sys_logger.info(f"❌ [LISP-Layout] 异常: {e}")
        return False



def export_layout_window_lisp_fit(
        point_a,
        point_b,
        pdf_fullpath,
        layout_name,
        *,
        device="DWG To PDF.pc3",
        media="ISO_A3_(420.00_x_297.00_MM)",
        ctb="monochrome.ctb",
        rotation=0
):
    """
    【函数编号】: PRINT-LAYOUT-ANTI-TARCH (天正穿透版)
    【修复原理】: 使用 ._-plot 强制绕过天正软件的命令劫持。
    """
    import time
    from system.CAD_coordination import wait_quiescent

    doc = C.doc

    # 路径与坐标处理
    pdf_path_final = pdf_fullpath.replace("\\", "/")
    p1_str = f"{point_a[0]},{point_a[1]}"
    p2_str = f"{point_b[0]},{point_b[1]}"
    orientation_str = "Portrait" if rotation == 1 else "Landscape"

    # ———————————————— 1. 构造指令队列 ————————————————
    # 🔥 核心修改：使用 ._-plot 绕过天正劫持
    commands_sequence = [
        "._-plot",         # <--- 关键修改！加点(.)强制使用原生命令，加下划线(_)适配中文版
        "Yes",             # 详细配置
        layout_name,       # 布局名称
        device,            # 打印机
        media,             # 纸张
        "Millimeters",     # 单位
        orientation_str,   # 方向
        "No",              # 反向?
        "Window",          # 窗口模式
        p1_str,            # 角点1
        p2_str,            # 角点2
        "Fit",             # 比例:布满
        "Center",          # 居中
        "Yes",             # 打印样式?
        ctb,               # 样式表
        "Yes",             # 线宽?
        "No",              # 缩放线宽? (Fit模式下)
        "No",              # 先打图纸空间?
        "No",              # 隐藏对象?
        pdf_path_final,    # 文件名
        "No",              # 保存设置?
        "Yes"              # 继续?
    ]

    sys_logger.info(f"🛡️ [天正穿透] 正在向 {layout_name} 发送打印指令...")

    # ———————————————— 2. 激活窗口 ————————————————
    try:
        if doc.WindowState == 2: # 如果最小化了
            doc.WindowState = 1  # 恢复正常
        doc.Activate()
    except: pass

    # ———————————————— 2.5. 等待CAD就绪 ————————————————
    if not wait_quiescent(min_quiet=0.5, timeout=10.0):
        sys_logger.warning("⚠️ CAD未完全就绪，尝试继续...")

    # ———————————————— 2.6. 切换到目标布局 ————————————————
    from scripts.CAD_file_operations import switch_to_layout
    if not switch_to_layout(layout_name):
        sys_logger.error(f"❌ 无法切换到布局: {layout_name}")
        return False

    # 布局切换后等待稳定
    time.sleep(1.0)

    # ———————————————— 3. 步进发送循环 ————————————————
    try:
        # 清理命令行（如果失败则跳过）
        try:
            doc.SendCommand("\x1b\x1b")
            time.sleep(0.2)
        except Exception as esc_err:
            sys_logger.debug(f"⚠️ ESC命令失败（已忽略）: {esc_err}")
            time.sleep(0.2)
        
        for i, cmd in enumerate(commands_sequence):
            # 简单的防错处理：给包含空格的名称加引号
            send_str = cmd
            if " " in str(cmd) and "," not in str(cmd) and not str(cmd).startswith("."):
                 send_str = f'"{cmd}"'
            
            try:
                doc.SendCommand(send_str + "\n")
            except Exception as e:
                # 🔥 天正特异性处理：
                # 天正有时会抛出 "输入无效" 但实际上命令已经进入缓冲区。
                # 如果是第一条命令报错，可能是真报错；如果是中间报错，可能是假警报。
                err_msg = str(e)
                if "天正" in err_msg or "输入无效" in err_msg:
                    sys_logger.info(f"   ⚠️ 无视天正干扰报错 (Step {i}): {cmd}")
                    time.sleep(0.2)
                    continue # 尝试继续发下一条
                else:
                    raise e # 其他错误正常抛出
            
            # 保持 0.15s 步进延迟
            time.sleep(0.15) 
            
    except Exception as e:
        sys_logger.info(f"❌ 发送指令致命中断: {e}")
        return False

    # ———————————————— 4. 等待结果 ————————————————
    max_wait = 20
    start_time = time.time()
    while time.time() - start_time < max_wait:
        if os.path.exists(pdf_fullpath):
            try:
                os.rename(pdf_fullpath, pdf_fullpath)
                sys_logger.info(f"✅ [输出成功] {os.path.basename(pdf_fullpath)}")
                return True
            except OSError:
                time.sleep(0.5) 
                continue
        time.sleep(0.5)
    
    sys_logger.info(f"❌ 超时未生成: {os.path.basename(pdf_fullpath)}")
    return False

#&&% 文件夹打印

def print_batch_custom_list(
        files_list,             # [必填] 文件任务列表 (List[dict])
        global_config=None,     # [可选] 全局默认配置 (dict)
        file_interval=20        # [可选] 文件处理间隔秒数
):
    """
    【函数编号】: PRINT-006 (Router Version)
    【所属模块】: 打印流程控制 (Workflow / Custom Batch)
    【功能描述】: 
        基于字典配置的高级批量打印路由函数。
        它充当一个“调度员”，根据配置将任务分发给 print_dwg_file_model 或 print_dwg_file_layout。

    【分发逻辑】:
        1. 读取合并后的配置中的 "mode" 字段。
        2. 若 mode="Model" (默认) -> 调用 print_dwg_file_model
        3. 若 mode="Layout"      -> 调用 print_dwg_file_layout (需提供 layout_name)

    【输入参数】:
        - files_list (list[dict]): 包含 "file_path" 及其他差异化配置的字典列表。
        - global_config (dict): 全局通用配置。
        - file_interval (int): 文件间隔时间。


    【使用示例】:
        common = {
            "mode": "Model", 
            "device": "DWG To PDF.pc3",
            "media": "ISO_A3_(420.00_x_297.00_MM)",
            "output_folder_root": r"D:/Myprogramsystem/XT/dayinSHUCHU",
            "force_fixed_media": False
        }

        tasks = [
            { "file_path": r"C:/1.dwg" },  # 使用 common 所有配置
            { 
              "file_path": r"C:/2.dwg", 
              "mode": "Layout",            # 覆盖 common 中的 mode
              "force_fixed_media": True    # 覆盖 common 中的设置
            }
        ]
        
        print_batch_custom_list(tasks, common, file_interval=15)

    """
    import os
    import time
    import shutil
    import datetime
    import openpyxl

    # ———————————————— 内部辅助：写 Excel 日志 ————————————————
    def _write_log_to_excel(file_path_str, status_msg, detail_msg):
        try:
            # 1. 获取用户目录
            userpath = os.environ.get('USERPATH') 
            log_file = os.path.join(userpath, "处理日志.xlsx") if userpath else "处理日志.xlsx"
            
            # 2. 加载 or 新建
            if os.path.exists(log_file):
                wb = openpyxl.load_workbook(log_file)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["日期", "时间", "文件名", "状态", "详细信息"])
            
            # 3. 寻找写入行
            last_row = ws.max_row
            while last_row > 1:
                if ws.cell(row=last_row, column=1).value is not None:
                    break
                last_row -= 1
            
            target_row = last_row + 2 if last_row > 0 else 2
            
            # 4. 写入
            now = datetime.datetime.now()
            fname = os.path.basename(file_path_str) if file_path_str else "Unknown"
            
            ws.cell(row=target_row, column=1, value=now.strftime("%Y-%m-%d"))
            ws.cell(row=target_row, column=2, value=now.strftime("%H:%M:%S"))
            ws.cell(row=target_row, column=3, value=fname)
            ws.cell(row=target_row, column=4, value=status_msg)
            ws.cell(row=target_row, column=5, value=detail_msg)
            
            wb.save(log_file)
            
        except Exception as e:
            sys_logger.info(f"❌ Excel日志写入失败: {e}")
    # ——————————————————————————————————————————————————————————

    # 1. 初始化检查
    li() # 连接 CAD
    
    if not files_list or not isinstance(files_list, list):
        print("❌ 错误: files_list 必须是非空列表")
        return False

    # 2. 定义基准默认值
    base_defaults = {
        "output_folder_root": r"D:/Myprogramsystem/XT/dayinSHUCHU",
        "mode": "Model",            # 默认为模型空间
        "layout_name": "布局1",     # 布局模式下的默认名
        "start_index": 0,
        "digit_width": 2,
        "force_fixed_media": False,
        "safety_delay": 10,
        "wps_close_threshold": 6,
        "device": "DWG To PDF.pc3",
        "media": "ISO_A3_(420.00_x_297.00_MM)",
        "ctb": "monochrome.ctb"
    }
    
    # 合并全局配置
    if global_config and isinstance(global_config, dict):
        base_defaults.update(global_config)

    total_files = len(files_list)
    sys_logger.info(f"📦 [批量路由] 任务数: {total_files} | 间隔: {file_interval}s | 错误重试: 3次")

    results = []

    # 3. 任务循环
    for i, file_task in enumerate(files_list):
        
        # A. 路径检查
        current_fpath = file_task.get("file_path")
        if not current_fpath or not os.path.exists(current_fpath):
            err = f"文件不存在: {current_fpath}"
            sys_logger.info(f"\n❌ 跳过任务 {i+1}: {err}")
            _write_log_to_excel("未知", "失败", err)
            continue

        file_basename = os.path.basename(current_fpath)
        file_name_pure = os.path.splitext(file_basename)[0]
        sys_logger.info(f"\nProcessing Task [{i+1}/{total_files}]: {file_basename}")
        
        # B. 参数合并 (Task > Global > Base)
        current_params = base_defaults.copy() 
        current_params.update(file_task)
        
        # 提取关键路由参数
        mode = current_params.pop("mode", "Model") # 提取并从kwargs中删除，避免传给具体函数报错
        layout_name_target = current_params.pop("layout_name", "布局1")
        
        # 清理掉不需要传给子函数的参数
        if "file_path" in current_params: del current_params["file_path"]

        # C. 预计算清理路径 (用于重试)
        # 根据子函数的逻辑推导生成文件夹路径
        folder_to_clean = ""
        if mode == "Model":
            folder_to_clean = os.path.join(current_params["output_folder_root"], file_name_pure)
        else:
            folder_to_clean = os.path.join(current_params["output_folder_root"], f"{file_name_pure}_{layout_name_target}")

        # D. 重试循环
        max_retries = 3
        success_flag = False
        final_msg = ""

        for attempt in range(1, max_retries + 1):
            try:
                if attempt > 1: sys_logger.info(f"    🔄 第 {attempt} 次尝试...")
                
                # ———————————————— 核心分发逻辑 ————————————————
                res = ""
                if mode == "Model":
                    # 调用模型空间打印函数
                    res = print_dwg_file_model(current_fpath, **current_params)
                
                elif mode == "Layout":
                    # 调用布局空间打印函数 (显式传入 layout_name)
                    res = print_dwg_file_layout(
                        current_fpath, 
                        layout_name=layout_name_target, 
                        **current_params
                    )
                else:
                    raise ValueError(f"未知的模式: {mode}")
                # ——————————————————————————————————————————————

                if "✅" in res:
                    success_flag = True
                    final_msg = res
                    break 
                else:
                    # 如果返回的是错误字符串
                    raise RuntimeError(res)

            except Exception as e:
                err_info = str(e)
                sys_logger.info(f"    ⚠️ 尝试 {attempt} 失败: {err_info}")
                
                if attempt < max_retries:
                    # 失败后清空目录
                    if os.path.exists(folder_to_clean):
                        try:
                            sys_logger.info(f"      正在清空输出目录: {os.path.basename(folder_to_clean)}")
                            shutil.rmtree(folder_to_clean)
                            time.sleep(1)
                        except Exception: pass
                    
                    # 重连 CAD
                    try: li() 
                    except: pass
                    time.sleep(2)
                else:
                    final_msg = f"3次重试均失败. 模式: {mode}, 错误: {err_info}"

        # E. 结果记录
        results.append(f"{file_basename}: {final_msg}")
        status_short = "成功" if success_flag else "失败"
        _write_log_to_excel(current_fpath, status_short, final_msg)

        if success_flag:
            sys_logger.info(f"✅ 文件完成: {file_basename}")
        else:
            sys_logger.info(f"❌ 文件彻底失败: {file_basename}")

        print("-" * 60)
        
        # F. 冷却时间
        if i < total_files - 1:
            sys_logger.info(f"⏳ 冷却 {file_interval} 秒...")
            time.sleep(file_interval)

    # 4. 总结
    success_count = len([r for r in results if "✅" in r])
    summary = f"🏁 任务全流程结束。成功: {success_count} / 总计: {total_files}"
    print(summary)
    return summary







#&&% 典型文件的绘制


def mark_print_areas_final(polylines_list, layer_name="PRINT_MARKER_FINAL"):
    """
    【函数编号】: DRAW-007 (最终版 - 瞄准镜样式)
    【功能】: 在多段线中心绘制：圆 + 矩形框 + 局部十字线 + 序号。
    """

    
    # 1. 初始化
    try:
        if 'doc' not in globals(): li()
    except:
        pass

    if not polylines_list:
        return

    new_objs = []
    sys_logger.info(f"🎯 [DRAW-007] 开始绘制标准标记 (共 {len(polylines_list)} 个)...")

    # 2. 准备图层 (防止 draw_circle 没地方放)
    try:
        layers = doc.Layers
        try:
            lyr = layers.Item(layer_name)
        except:
            lyr = layers.Add(layer_name)
        lyr.LayerOn = True
        doc.ActiveLayer = lyr
    except:
        pass

    # 3. 循环绘制
    for i, pl in enumerate(polylines_list):
        try:
            # --- A. 几何计算 (沿用 debug 函数逻辑) ---
            pl = cast_object(pl)
            bbox = safe_get_bbox(pl)
            if not bbox: continue

            min_pt, max_pt = bbox
            # 强制 float 转换
            x1, y1 = float(min_pt[0]), float(min_pt[1])
            x2, y2 = float(max_pt[0]), float(max_pt[1])
            
            cx = (x1 + x2) / 2.0
            cy = (y1 + y2) / 2.0
            z  = float(min_pt[2]) # 保持高度
            
            width = abs(x2 - x1)
            height = abs(y2 - y1)
            min_side = min(width, height)

            # --- B. 尺寸设定 ---
            # 文字高度设为短边的 1/5
            text_h = min_side / 5.0
            # 圆半径设为文字高度的 1.2 倍
            radius = text_h * 1.2
            # 颜色循环
            color_idx = (i % 6) + 1 

            # 打印调试信息：如果这里 radius 很小(例如 0.5)，屏幕上就看不见
            sys_logger.info(f"  > 对象 {i}: 中心=({cx:.0f},{cy:.0f}) 半径 R={radius:.1f}")

            # --- C. 绘制图形 ---
            center_3d = (cx, cy, z)

            # 1. 画圆 (内部圈)
            c_obj = draw_circle(center_3d, radius)
            if c_obj:
                c_obj.Layer = layer_name
                c_obj.color = color_idx
                new_objs.append(c_obj)

            # 2. 画矩形框 (外部框 - 稍微大一点)
            box_r = radius * 1.6 # 矩形半宽
            box_pts = [
                (cx - box_r, cy - box_r, z),
                (cx + box_r, cy - box_r, z),
                (cx + box_r, cy + box_r, z),
                (cx - box_r, cy + box_r, z)
            ]
            # 使用 draw_lwpolyline 绘制闭合矩形
            box_obj = draw_lwpolyline(box_pts, layer_name=layer_name, width=0, color=color_idx, closed=True)
            if box_obj: new_objs.append(box_obj)

            # 3. 画十字线 (不穿过圆心，制造“瞄准镜”缺口感)
            gap = radius * 1.1 # 缺口起始点
            len_line = radius * 2.5 # 线条长度
            
            # 四段短线：上、下、左、右
            lines_data = [
                [(cx, cy + gap, z), (cx, cy + len_line, z)], # 上
                [(cx, cy - gap, z), (cx, cy - len_line, z)], # 下
                [(cx - gap, cy, z), (cx - len_line, cy, z)], # 左
                [(cx + gap, cy, z), (cx + len_line, cy, z)], # 右
            ]
            
            for pts in lines_data:
                l_obj = draw_lwpolyline(pts, layer_name=layer_name, width=0, color=color_idx, closed=False)
                if l_obj: new_objs.append(l_obj)

            # 4. 写文字
            txt_str = f"{i+1:03d}"
            txt_obj = write_cad_text(
                p=center_3d,
                text=txt_str,
                alignment="中心",
                height=text_h,
                width_factor=0.8,
                style="Standard",
                layer=layer_name
            )
            if txt_obj:
                txt_obj.color = color_idx
                new_objs.append(txt_obj)

        except Exception as e:
            sys_logger.info(f"⚠️ 第 {i} 个处理异常: {e}")

    # 4. 刷新
    try:
        doc.Regen(0)
    except:
        pass
    sys_logger.info(f"✅ 完成。")
    return new_objs

#&&% 生成图名标注

def generate_tarch_drawing_names_v5(source_obj, print_areas, layer_name="DIM_SYMB"):
    """
    【函数功能】: DRAW-008 (v5.0 - 网格自适应终极版)
    【解决痛点】:
        1. 彻底解决竖向图纸重叠问题（自动切换为 3行2列）。
        2. 放弃固定间距，使用图框尺寸自适应计算。
    """

    name_choices = [
        "一层平面图", "二层平面图", "屋顶平面图", "地下一层平面图",
        "东立面图", "西立面图", "南立面图", "北立面图",
        "1-1剖面图", "楼梯详图", "节点大样", "门窗表"
    ]

    new_objs = []
    
    # 辅助函数
    def vtpnt(x, y, z=0):
        return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (x, y, z))

    sys_logger.info(f"🚀 开始生成 (模式：图框尺寸自适应)...")

    # --- 循环处理每个打印区域 ---
    for i, area in enumerate(print_areas):
        try:
            # A. 获取图框边界
            bbox = safe_get_bbox(area)
            if not bbox: continue

            p_min, p_max = bbox
            x_min, y_min = float(p_min[0]), float(p_min[1])
            x_max, y_max = float(p_max[0]), float(p_max[1])
            
            # 计算宽高
            width = abs(x_max - x_min)
            height = abs(y_max - y_min)
            
            # B. 判断横竖向，决定网格策略
            is_portrait = height > width
            
            if is_portrait:
                # === 竖向图纸 (Portrait) ===
                # 策略：3行 x 2列 (利用高度优势)
                rows = 3
                cols = 2
                tarch_angle = -90.0  # 图名旋转 -90度
                
                # 定义生成区域：使用图框的 下半截 (Bottom 50%)
                # 留一点边距 (margin)
                work_y_min = y_min + height * 0.05
                work_y_max = y_min + height * 0.45 
                work_x_min = x_min + width * 0.1
                work_x_max = x_max - width * 0.1
                
            else:
                # === 横向图纸 (Landscape) ===
                # 策略：2行 x 3列 (利用宽度优势)
                rows = 2
                cols = 3
                tarch_angle = 0.0
                
                # 定义生成区域：图框 下方 1/3 区域
                work_y_min = y_min + height * 0.05
                work_y_max = y_min + height * 0.35
                work_x_min = x_min + width * 0.1
                work_x_max = x_max - width * 0.1

            # C. 网格计算 (Grid Calculation)
            # 计算工作区宽高
            work_w = work_x_max - work_x_min
            work_h = work_y_max - work_y_min
            
            # 计算单元格步长 (Step)
            # 为了居中，我们把工作区切分为 rows * cols 个格子，并将对象放在格子中心
            step_x = work_w / cols
            step_y = work_h / rows
            
            origin_pt = vtpnt(0, 0, 0) # 源对象位置

            # --- D. 生成矩阵 ---
            # 外层循环：行 (从上往下排，或者从下往上，这里选择从上往下)
            # 注意：work_y_max 是上方，work_y_min 是下方
            
            for r in range(rows):
                # 计算当前行的中心 Y
                # 如果 r=0 (第一行)，位于最上方
                # Y = Top - (RowIndex + 0.5) * StepHeight
                # 或者简单的插值：
                center_y = work_y_max - (r * step_y) - (step_y / 2)
                
                for c in range(cols):
                    # 计算当前列的中心 X
                    # X = Left + (ColIndex + 0.5) * StepWidth
                    center_x = work_x_min + (c * step_x) + (step_x / 2)

                    try:
                        # 1. 复制 & 移动
                        new_obj = source_obj.Copy()
                        new_obj.Move(origin_pt, vtpnt(center_x, center_y, 0))
                        
                        # 2. 设置旋转
                        set_attr(new_obj, "布局转角", tarch_angle)

                        # 3. 修改内容
                        rand_name = random.choice(name_choices)
                        set_attr(new_obj, "Layer", layer_name)
                        set_attr(new_obj, "图名文字", rand_name)
                        
                        new_obj.Update()
                        new_objs.append(new_obj)

                    except Exception as e:
                        sys_logger.info(f"⚠️ 生成失败: {e}")

        except Exception as e:
            sys_logger.info(f"⚠️ 区域 {i} 异常: {e}")

    # 刷新
    try: doc.Regen(0) 
    except: pass

    sys_logger.info(f"✅ 生成完成，共 {len(new_objs)} 个 (自适应布局)。")
    return new_objs






#&&&% ===（十二）模型空间文件打印 ===


#&&% 模型空间文件打印

def print_dwg_file_model(
        file_path=None,               # 目标文件路径 (None表示当前文件)
        *,
        # --- 命名与序号控制 ---
        start_index=0,                # 起始序号
        digit_width=2,                # 序号位数
        
        # --- 智能与安全控制 ---
        force_fixed_media=False,      # 是否强制使用固定图幅
        safety_delay=10,              # 横竖切换安全等待时间(秒)
        wps_close_threshold=6,        # 触发WPS关闭的打印张数阈值
        
        # --- 路径与设备配置 ---
        output_folder_root=None,      # 根输出目录
        device="DWG To PDF.pc3",
        media="ISO_A3_(420.00_x_297.00_MM)", # 根输出目录force_fixed_media为真的打印图幅
        ctb="monochrome.ctb",

        # --- 空间定向 ---

        layout_name = None,
        operate_target = "Model",
        select_config = 0,                   #0为常规模式1为精细模式
        use_cache = False,

        #边线补偿宽度
        xiubukuan = 25,

):
    """
    【函数编号】: PRINT-MODEL-MANAGER (V1 - 模型空间专用)
    【功能】: 
        1. 专用于处理模型空间 (Model Space) 的批量打印任务。
        2. 自动打开文件、提取多段线和图签。
        3. 调用 print_polylines_list 引擎执行。
    """
    # 0. 路径初始化
    if output_folder_root is None:
        if userpath: output_folder_root = os.path.join(userpath, "输出pdf")
        else: output_folder_root = r"D:/Myprogramsystem/XT/dayinSHUCHU"

    if select_config == 1:
        xiubukuan = 0.25 

    # 1. 打开文件 & 连接
    doc=C.doc
    if file_path:
        try:
            from CAD_file_operations import open_file
            sys_logger.info(f"📂 打开: {file_path}")
            open_file(file_path)
            
        except Exception as e:
            return f"❌ 打开失败: {e}"

    # 2. 准备输出目录
    try:
        doc_name = doc.Name
        file_name_pure = os.path.splitext(doc_name)[0]
        target_folder = os.path.join(output_folder_root, file_name_pure)
        
        if os.path.exists(target_folder):
            shutil.rmtree(target_folder)
        os.makedirs(target_folder)
    except: pass

    # 3. 强制进入模型空间 & 提取数据
    print("🔍 [Model] 正在提取图框与图签...")
    try:
        doc.SetVariable("TILEMODE", 1) # 🔥 强制模型空间
        
        # 调用你的核心提取函数
        # 该函数返回 [多段线列表, 图签块列表, ...]

        ctq = smart_rebuild_print_info(
        layout_name=layout_name,
        operate_target=operate_target,
        select_config=select_config, 
        use_cache=use_cache,
        )
   
    except Exception as e:
        return f"❌ 数据提取失败: {e}"

    if not ctq or not ctq[0]:
        return "❌ 未找到打印框"
        
    polys_list = ctq[0]
    # 如果有图签列表则获取，没有则为 None
    titles_list = ctq[1] if len(ctq) > 1 else None

    # 4. 委托给打印引擎

    print_polylines_list(
        polylines_list=polys_list,
        title_blocks_list=titles_list, # 传入图签，让引擎自动命名
        
        start_index=start_index,
        digit_width=digit_width,
        folderpath=target_folder,
        
        device=device,
        media=media,
        ctb=ctb,
        
        mode="Model", # 🔥 显式指定模型模式
        force_fixed_media=force_fixed_media,
        safety_delay=safety_delay,
        wps_close_threshold=wps_close_threshold,
        xiubukuan = xiubukuan

    )
    
    # 5. 简单校验
    try:
        if os.path.exists(target_folder):
            generated_files = [f for f in os.listdir(target_folder) if f.lower().endswith(".pdf")]
            actual_count = len(generated_files)
        else:
            actual_count = 0
        
        return f"✅ [Model] 打印完成: {actual_count}/{len(polys_list)} (目录: {file_name_pure})"
    except Exception as e:
        return f"❌ 校验异常: {e}"



#&&% 模型空间批量打印

def print_polylines_list(
        polylines_list,
        title_blocks_list=None,       # [可选] 对应的图签块列表，用于生成文件名
        subproject=None,      
        dwg_num=None,         
        dwg_name=None,        
        *,
        # --- 基础配置 ---
        folderpath=os.path.join(userpath,"输出pdf"),
        start_index=0,
        digit_width=2,
        
        # --- 打印控制 ---
        force_fixed_media=False,
        safety_delay=10,              # 横竖切换的休息时间
        wps_close_threshold=6,
        
        # --- 设备参数 ---
        device="DWG To PDF.pc3",
        media="ISO_A3_(420.00_x_297.00_MM)",
        ctb="monochrome.ctb",
        mode="Model",
        layout_name="布局1",
        xiubukuan = 25,   
):
    """
    【函数编号】: PRINT-FINAL (V14 - 旋转逻辑修复版)
    【修复内容】: 
        1. 废弃正则表达式判定旋转的旧逻辑。
        2. 直接使用 generate_name_and_ratio_from_com 返回的 orientation_flag (0或1) 控制旋转。
        3. 完美支持 UserDefinedMetric 等非标自定义图纸的自动旋转。
    """

    import re
    
    # 0. 环境准备
    doc=C.doc
    
    # 兼容 rebuild_print_area_title_mapping 返回的元组结构
    if isinstance(polylines_list, tuple): polylines_list = polylines_list[0]
    
    if not polylines_list: return False
    if not os.path.exists(folderpath): os.makedirs(folderpath, exist_ok=True)

    try:
        current_doc_name = os.path.splitext(doc.Name)[0]
    except:
        current_doc_name = "Unnamed"
    
    # 1. 切换空间
    try:
        if mode.lower() == "model":
            doc.SetVariable("TILEMODE", 1)
        else:
            doc.SetVariable("TILEMODE", 0)
            doc.ActiveLayout = doc.Layouts.Item(layout_name)
    except: pass

    # ==================== A. 预处理：生成任务单 (含文件名和分组) ====================
    sys_logger.info(f"🚀 [引擎启动] 处理 {len(polylines_list)} 个对象 (图签: {'有' if title_blocks_list else '无'})...")
    
    tasks_landscape = [] # 横向组
    tasks_portrait = []  # 竖向组
    
    for i, pl in enumerate(polylines_list):
        # --- 1. 几何与纸张分析 ---
        bbox = safe_get_bbox(pl) 
        if not bbox: continue
        
        # 获取几何原始宽高，仅用于最后的任务分组 (减少打印机切换卡顿)
        width_raw = float(abs(bbox[1][0] - bbox[0][0]))
        height_raw = float(abs(bbox[1][1] - bbox[0][1]))
        dwg_is_landscape = width_raw > height_raw 
        
        # 调用分析函数
        target_fandy = (media, "Fit", "Fixed", 0)
        print_info = generate_name_and_ratio_from_com(
            pl, A3dy=1 if force_fixed_media else 0, Fandy=target_fandy
        )
        
        if print_info == 0:
            sys_logger.info(f"   ⚠️ 跳过: 第 {i} 个对象无法识别图幅。")
            continue
            
        # 🔥【核心修改】正确解包所有4个参数
        # print_info 结构: (图纸名, 比例, 图号名, 旋转标志0或1)
        calc_media, calc_scale_str, _, calc_orientation = print_info
        
        # --- 2. 旋转计算 ---
        # 🔥【核心修改】直接使用计算好的方向标志，不再猜测
        target_rotation = calc_orientation

        # --- 3. 文件名生成 ---
        idx_str = f"{i + start_index:0{digit_width}d}"
        fname = f"{current_doc_name}-{idx_str}.pdf" 
        
        if title_blocks_list and i < len(title_blocks_list):
            blk = title_blocks_list[i]
            try:
                attrs = get_block_attributes_dict(blk, ignore_empty=True, upper_tag=True)
                
                s_p = attrs.get("子项目名称") or attrs.get("子项名称") or subproject or ""
                d_n = attrs.get("图纸名称") or attrs.get("图名") or dwg_name or ""
                d_u = attrs.get("图纸编号") or attrs.get("图号") or dwg_num or ""
                
                parts = [p for p in [s_p, d_u, d_n] if p]
                if parts: fname = f"{'-'.join(parts)}.pdf"
            except: pass
        elif any([subproject, dwg_num, dwg_name]):
             parts = [p for p in [subproject, dwg_num, dwg_name] if p]
             fname = f"{'-'.join(parts)}-{idx_str}.pdf"



        # 文件名清洗
        fname = "".join([c if c not in '<>:"/\\|?*\n\r\t' else '_' for c in fname])
        full_path = os.path.join(folderpath, fname)

        # --- 4. 存入任务单 ---
        task = {
            "id": i,
            "p_min": bbox[0], "p_max": bbox[1],
            "full_path": full_path,
            "media": calc_media,
            "rotation": target_rotation, # 使用修正后的 0 或 1
            "desc": f"{int(width_raw)}x{int(height_raw)}"
        }
        
        # 按图纸几何形状分组
        if dwg_is_landscape:
            tasks_landscape.append(task)
        else:
            tasks_portrait.append(task)

    sys_logger.info(f"✅ 分组完毕: 横向 {len(tasks_landscape)} 张 | 竖向 {len(tasks_portrait)} 张")
    
    # ==================== B. 执行逻辑 ====================
    total_success = 0
    
    def execute_batch(batch_list, group_name):
        nonlocal total_success
        if not batch_list: return
        
        print("-" * 100)
        sys_logger.info(f"▶️  开始打印【{group_name}】组...")
        for t in batch_list:
            sys_logger.info(f"   [ID={t['id']}] {t['desc']:<12} -> {os.path.basename(t['full_path'])[:40]}...")
            
            try:
                res = False
                # 根据模式调用底层函数
                if mode.lower() == "model":
                    res = export_model_window_lisp_fit(
                        point_a = t["p_min"],     
                        point_b = t["p_max"],     
                        pdf_fullpath = t["full_path"],
                        device = device,
                        media = t["media"], 
                        ctb = ctb,
                        rotation = t["rotation"],
                        xiubukuan = xiubukuan,

                    )
                else:
                    # 布局打印保持原有参数命名
                    try:
                        res = export_layout_window_lisp_fit(
                            point_a = t["p_min"],
                            point_b = t["p_max"],
                            pdf_fullpath = t["full_path"],
                            layout_name = layout_name,
                            device = device,
                            media = t["media"],
                            ctb = ctb,
                            rotation = t["rotation"]
                        )
                    except:
                        # 兼容性尝试 (旧版可能是 lower_left_xy)
                        pass
                
                if res: total_success += 1

                # ---------------------------------------------------------
                # WPS 窗口清理逻辑 (增强版：重试3次 + 先最大化唤醒)
                # ---------------------------------------------------------
                if wps_close_threshold > 0 and total_success % wps_close_threshold == 0:
                    # 尝试 3 次
                    for attempt in range(3):
                        try:
                            import win32gui, win32con
                            
                            # 使用 nonlocal 标记是否在本次扫描中找到了窗口
                            wps_found = False

                            def cb_maximize_and_close(hwnd, ex):
                                nonlocal wps_found
                                # 获取窗口标题
                                title = win32gui.GetWindowText(hwnd)
                                
                                # 判断条件：标题包含 WPS 且 窗口可见
                                if "WPS Office" in title and win32gui.IsWindowVisible(hwnd):
                                    wps_found = True
                                    try:
                                        # 窗口管理
                                        minimize_all_windows()
                                        time.sleep(0.5)
                                        activate_window_by_title("WPS Office", click_titlebar=True)
                                        time.sleep(0.5)

                                        # 1. 先最大化 (唤醒窗口，确保它能接收消息)
                                        win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
                                        
                                        # 2. 发送关闭指令
                                        win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
                                    except:
                                        pass

                            # 执行遍历
                            win32gui.EnumWindows(cb_maximize_and_close, None)
                            
                            # 如果这一轮扫描没发现任何 WPS 窗口，说明已经清理干净了，提前退出循环
                            if not wps_found:
                                break
                            
                            # 如果发现了并处理了，休息 0.5 秒等待窗口响应关闭，再进行下一轮确认
                            time.sleep(0.5)

                        except Exception as e:
                            # sys_logger.info(f"清理 WPS 出错: {e}") # 调试时可打开
                            pass

                    
            except Exception as e:
                sys_logger.info(f"   ❌ 打印异常: {e}")

    # 1. 执行横向
    execute_batch(tasks_landscape, "横向")
    
    # 2. 中场休息 (给打印机缓冲时间)
    if len(tasks_landscape) > 0 and len(tasks_portrait) > 0:
        print("\n" + "="*50)
        sys_logger.info(f"⏸️  横向结束。等待 {safety_delay} 秒调整打印机方向...")
        print("="*50 + "\n")
        time.sleep(safety_delay)
        
    # 3. 执行竖向
    execute_batch(tasks_portrait, "竖向")

    sys_logger.info(f"\n🏁 全部完成: {total_success}/{len(polylines_list)}")
    return True


#&&&% ===（十三）图纸空间文件打印 ===

#&&% 图纸空间文件打印

def print_dwg_file_layout(
        file_path=None,               # 目标文件路径 (None 则处理当前打开的文件)
        layout_name="布局1",          # 必须指定布局名称
        *,
        # --- 命名与序号控制 ---

        start_index=0,                # 起始序号
        digit_width=2,                # 序号位数 (如 01, 02)

        # --- 路径与设备配置 ---
        output_folder_root=None,      # 输出根目录
        device="DWG To PDF.pc3",      # 打印驱动
        media="ISO_A3_(420.00_x_297.00_MM)", # 默认纸张 (备用)
        ctb="monochrome.ctb",         # 打印样式

        # --- 智能与安全控制 ---
        force_fixed_media=False,      # 是否强制使用固定纸张
        safety_delay=10,              # 横竖分组切换延时
        wps_close_threshold=6,         # 每打几张清理一次 PDF 预览进程

        # --- 空间定向 ---
        operate_target = "Layout",
        select_config = 1,                   #1为小比例模式0为大比例模式元组为自定义
        use_cache = False,
        

):
    """
    【功能】: 布局空间批量打印主管理器 (V7 - Polyline适配版)
    """
    # 0. 路径初始化
    userpath = os.environ.get('USERPATH') 
    if output_folder_root is None:
        output_folder_root = os.path.join(userpath, "输出pdf") if userpath else r"D:/Myprogramsystem/XT/dayinSHUCHU"

    # 1. 环境准备
    doc=C.doc
    


    if file_path:
        from CAD_file_operations import open_file
        open_file(file_path)
        doc=C.doc        

    # 2. 准备输出子目录
    try:
        file_pure_name = os.path.splitext(doc.Name)[0]
    except:
        file_pure_name = "Unnamed"
        
    target_folder = os.path.join(output_folder_root, f"{file_pure_name}_{layout_name}")
    
    # 清理旧文件夹 (可选，根据需求决定是否保留)
    if os.path.exists(target_folder): 
        try: shutil.rmtree(target_folder)
        except: pass 
    os.makedirs(target_folder, exist_ok=True)

    # 3. 提取打印区域与图签映射 (核心侦察 - 修改点1)
    sys_logger.info(f"🔍 [Layout] 正在分析布局 '{layout_name}' 的多段线区域...")
    
    polylines, titles, _ = smart_rebuild_print_info(
    layout_name=layout_name, 
    operate_target=operate_target,
    select_config=select_config,
    use_cache=use_cache,   
    )

    if not polylines:
        return f"❌ 错误: 在布局 '{layout_name}' 中未找到有效的【矩形多段线】打印区域"

    # 4. 执行批量打印任务 (修改点2)
    # 🔥🔥🔥 改为调用新的 Polyline 打印列表函数 🔥🔥🔥
    res = print_layout_polylines_list(
        polylines_list=polylines,     # 传入多段线列表
        title_blocks_list=titles,     # 传入对应的图签块
        layout_name=layout_name,
        folderpath=target_folder,
        start_index=start_index,
        digit_width=digit_width,
        device=device,
        media=media,
        ctb=ctb,
        force_fixed_media=force_fixed_media,
        safety_delay=safety_delay,
        wps_close_threshold=wps_close_threshold,
        # 传递文档元数据用于命名
        dwg_name=file_pure_name,
        dwg_num="" # 如果有需要，可以从这里传递额外的编号
    )
    
    return f"✅ [Layout] 打印完成: {len(polylines)} 张图纸已输出至 {target_folder}"


#&&% 图纸空间列表打印

def print_layout_polylines_list(
        polylines_list,       
        title_blocks_list=None,
        subproject=None,       
        dwg_num=None,          
        dwg_name=None,         
        *,
        layout_name="布局1",
        folderpath=os.path.join(userpath,"输出pdf"),
        start_index=0,
        digit_width=2,
        force_fixed_media=False,
        safety_delay=10,
        wps_close_threshold=6,
        device="DWG To PDF.pc3",
        media="ISO_A3_(420.00_x_297.00_MM)", 
        ctb="monochrome.ctb"
):
    """
    【函数编号】: PRINT-LAYOUT-ENGINE (V7.1 - 侦察版)
    【功能】: 接收多段线列表进行打印，并输出详细的匹配调试信息。
    """
    from  CAD_file_operations   import  switch_to_layout

    doc=C.doc
    
    if not polylines_list: return False
    if not os.path.exists(folderpath): os.makedirs(folderpath, exist_ok=True)

    try: current_doc_name = os.path.splitext(doc.Name)[0]
    except: current_doc_name = "Unnamed"
    
    # 1. 切换布局
    if not switch_to_layout(layout_name):
        sys_logger.info(f"❌ 无法切换到布局: {layout_name}")
        return False
    time.sleep(1.0)
    # ==================== A. 数据预读取 ====================
    sys_logger.info(f"🚀 [Layout引擎] 正在分析 {layout_name} 中的 {len(polylines_list)} 个打印区域...")
    
    # 读取图签数据的逻辑
    titles_data = []
    if title_blocks_list:
        sys_logger.info(f"📋 正在读取图签属性...")
        for idx, blk in enumerate(title_blocks_list):
            try:
                if blk is None:
                    titles_data.append(("", "", ""))
                    continue
                attrs = get_block_attributes_dict(blk, ignore_empty=False, upper_tag=True)
                s_p = attrs.get("子项目名称") or attrs.get("子项名称") or attrs.get("SUB_TITLE") or ""
                d_n = attrs.get("图纸名称") or attrs.get("图名") or attrs.get("TITLE") or attrs.get("DWG_NAME") or ""
                d_u = attrs.get("图纸编号") or attrs.get("图号") or attrs.get("DWG_NO") or attrs.get("DRAWING_NO") or ""
                titles_data.append((s_p, d_n, d_u))
                sys_logger.info(f"   [图块{idx}] {d_u} {d_n}")
            except:
                titles_data.append(("", "", ""))
    
    tasks_landscape = [] 
    tasks_portrait = []  
    
    # 遍历多段线列表
    for i, pl_obj in enumerate(polylines_list):
        
        # 1. Geometry
        bbox = safe_get_bbox(pl_obj) 
        if not bbox: 
            sys_logger.info(f"   ⚠️ 第 {i} 个区域获取包围盒失败，跳过。")
            continue
            
        p_min, p_max = bbox
        width_raw = float(abs(p_max[0] - p_min[0]))
        height_raw = float(abs(p_max[1] - p_min[1]))
        
        dwg_is_landscape = width_raw > height_raw 
        
        # 2. 图幅匹配逻辑
        try:
            print_info = generate_name_and_ratio_from_com(pl_obj, A3dy=0)
        except:
            print_info = 0

        # 🔥🔥🔥🔥🔥 【侦察探针】 START 🔥🔥🔥🔥🔥
        sys_logger.info(f"   🔎 [侦察 ID={i}] 尺寸: {int(width_raw)}x{int(height_raw)}")
        if print_info != 0:
            # 这里的 print_info[0] 就是导致后面报错的“嫌疑犯”
            sys_logger.info(f"      -> 匹配到的纸张名: '{print_info[0]}'") 
            sys_logger.info(f"      -> 匹配到的比例: '{print_info[1]}'")
        else:
            sys_logger.info(f"      -> 匹配失败 (将使用默认 A3)")
        # 🔥🔥🔥🔥🔥 【侦察探针】 END 🔥🔥🔥🔥🔥

        if print_info == 0:
            calc_orientation = 0 if dwg_is_landscape else 1
            calc_media = media 
        else:
            calc_media, _, _, calc_orientation = print_info

        target_rotation = calc_orientation 

        # 3. Filename Generation
        idx_str = f"{i + start_index:0{digit_width}d}"
        final_s_p = subproject
        final_d_n = dwg_name
        final_d_u = dwg_num

        # 如果有图签数据，优先使用图签数据覆盖
        if i < len(titles_data):
            t_sp, t_dn, t_du = titles_data[i]
            if t_sp: final_s_p = t_sp
            if t_dn: final_d_n = t_dn
            if t_du: final_d_u = t_du

        # --- 🔥 核心修改逻辑开始 ---
        # 收集非空的字段
        valid_parts = [p for p in [final_s_p, final_d_u, final_d_n] if p]

        if valid_parts:
            # 【情况1】：获取到了有效信息 -> 拼接名称，不带序号
            # 例如：子项-图号-图名.pdf
            fname = f"{'-'.join(valid_parts)}.pdf"
        else:
            # 【情况2】：完全没有信息 -> 使用默认兜底格式，必须带序号防止重名
            # 例如：文件名-布局名-01.pdf
            fname = f"{current_doc_name}-{layout_name}-{idx_str}.pdf"
        # --- 🔥 核心修改逻辑结束 ---

        # 文件名字符清洗
        fname = "".join([c if c not in '<>:"/\\|?*\n\r\t' else '_' for c in fname])
        full_path = os.path.join(folderpath, fname)


        # 4. Packaging
        task = {
            "id": i,
            "p_min": p_min, "p_max": p_max,
            "full_path": full_path,
            "media": calc_media,
            "rotation": target_rotation,
            "desc": f"{int(width_raw)}x{int(height_raw)}"
        }
        
        if dwg_is_landscape: tasks_landscape.append(task)
        else: tasks_portrait.append(task)

    sys_logger.info(f"✅ 分组完毕: 横向 {len(tasks_landscape)} | 竖向 {len(tasks_portrait)}")

    # ==================== B. 执行逻辑 ====================
    total_success = 0
    

    def execute_batch(batch_list, group_name):
            nonlocal total_success
            if not batch_list: return
            
            sys_logger.info(f"▶️  开始打印【{group_name}】组...")
            
            # 引入 time 模块 (如果外部没引入，这里保险起见引入一下)
            import time 
    
            for t in batch_list:
                sys_logger.info(f"   [ID={t['id']}] {t['desc']:<12} -> {os.path.basename(t['full_path'])[:40]}")
                
                try:
                    # 1. 执行打印指令（使用v1版本 - LISP命令方式）
                    res = export_layout_window_lisp_fit_v1(
                        point_a = t["p_min"],
                        point_b = t["p_max"],
                        pdf_fullpath = t["full_path"],
                        layout_name = layout_name,
                        device = device,
                        media = t["media"],
                        ctb = ctb,
                        rotation = t["rotation"]
                    )
                    
                    if res: total_success += 1
    
                    # 🔥🔥🔥【位置 1：打印冷却期】(关键！) 🔥🔥🔥
                    # 原因：AutoCAD 刚收到指令正在生成 PDF，此时若立即发下一条指令或操作窗口，容易造成“消化不良”卡死。
                    # 建议：最少 1.0 秒，推荐 1.5 秒
                    time.sleep(1.5)
    
                    # ---------------------------------------------------------
                    # WPS 窗口清理逻辑
                    # ---------------------------------------------------------
                    if wps_close_threshold > 0 and total_success > 0 and total_success % wps_close_threshold == 0:
                        
                        sys_logger.info("🧹 触发 WPS 内存清理...")
                        
                        # 🔥🔥🔥【位置 2：操作前缓冲】🔥🔥🔥
                        # 原因：即将进行窗口最小化/最大化的大动作，先让 CPU 喘口气
                        time.sleep(0.5)
    
                        # 尝试 3 次清理
                        for attempt in range(3):
                            try:
                                import win32gui, win32con
                                wps_found = False
    
                                def cb_maximize_and_close(hwnd, ex):
                                    nonlocal wps_found
                                    title = win32gui.GetWindowText(hwnd)
                                    if "WPS Office" in title and win32gui.IsWindowVisible(hwnd):
                                        wps_found = True
                                        try:
                                            # 窗口操作
                                            minimize_all_windows() # 慎用：这会把 CAD 也最小化
                                            time.sleep(0.5) 
                                            activate_window_by_title("WPS Office", click_titlebar=True)
                                            time.sleep(0.5)
                                            win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
                                            
                                            # 发送关闭
                                            win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
                                        except: pass
    
                                win32gui.EnumWindows(cb_maximize_and_close, None)
                                
                                if not wps_found:
                                    break
                                
                                time.sleep(0.5) # 等待窗口响应关闭
    
                            except Exception as e:
                                pass
                        
                        # 🔥🔥🔥【位置 3：回归 CAD 焦点】(非常重要！) 🔥🔥🔥
                        # 原因：WPS 清理完后，焦点可能丢失。
                        # 如果不加延迟直接进入下一次循环发送 SendCommand，命令可能发到虚空里。
                        sys_logger.info("🔙 正在恢复 CAD 焦点...")
                        try:
                            # 尝试把 CAD 激活回前台 (防止刚才 minimize_all_windows 把 CAD 也收起来了)
                            activate_window_by_title("AutoCAD", click_titlebar=False) 
                        except:
                            pass
                        
                        # 给 CAD 重新获取焦点的时间
                        time.sleep(1.0) 
    
                except Exception as e:
                    sys_logger.info(f"   ❌ 打印异常: {e}")
    
    
    
    
    
    






    execute_batch(tasks_landscape, "横向")
    if len(tasks_landscape) > 0 and len(tasks_portrait) > 0:
        print("\n" + "="*50)
        sys_logger.info(f"⏸️  横向结束。等待 {safety_delay} 秒调整打印机方向...")
        print("="*50 + "\n")
        time.sleep(safety_delay)
    execute_batch(tasks_portrait, "竖向")

    sys_logger.info(f"\n🏁 [Layout] 打印结束: {total_success}/{len(polylines_list)}")
    return True



def print_layout_polylines_list_y(
        polylines_list,        
        title_blocks_list=None,
        subproject=None,        
        dwg_num=None,           
        dwg_name=None,          
        *,
        layout_name="布局1",
        folderpath=os.path.join(os.environ.get('USERPATH', 'D:/'), "输出pdf"), # 防止 USERPATH 报错
        start_index=0,
        digit_width=2,
        force_fixed_media=False,
        safety_delay=10,
        wps_close_threshold=6,
        device="DWG To PDF.pc3",
        media="ISO_A3_(420.00_x_297.00_MM)", 
        ctb="monochrome.ctb"
):
    """
    【函数编号】: PRINT-LAYOUT-ENGINE (V7.2 - 稳健量产版)
    """
    from CAD_file_operations import switch_to_layout
    import time
    
    # 确保 C.doc 可用
    try: doc = C.doc
    except: return False
    
    if not polylines_list: return False
    if not os.path.exists(folderpath): os.makedirs(folderpath, exist_ok=True)

    try: current_doc_name = os.path.splitext(doc.Name)[0]
    except: current_doc_name = "Unnamed"
    
    # 1. 切换布局
    if not switch_to_layout(layout_name):
        sys_logger.info(f"❌ 无法切换到布局: {layout_name}")
        return False

    # 2. 预处理数据 (同前，省略部分非核心代码...)
    titles_data = []
    # ... (此处保留原本的图签读取逻辑) ...
    
    tasks = []
    for i, pl_obj in enumerate(polylines_list):
        bbox = safe_get_bbox(pl_obj) 
        if not bbox: continue
            
        # 🔥 数据清洗：确保坐标是纯净的 float tuple，防止 COM 对象类型干扰
        try:
            p_min = (float(bbox[0][0]), float(bbox[0][1]), 0.0)
            p_max = (float(bbox[1][0]), float(bbox[1][1]), 0.0)
        except:
            sys_logger.info(f"⚠️ 坐标转换失败 ID={i}")
            continue

        width_raw = abs(p_max[0] - p_min[0])
        height_raw = abs(p_max[1] - p_min[1])
        dwg_is_landscape = width_raw > height_raw 
        
        # ... (此处保留原本的图幅匹配逻辑) ...
        # 假设匹配到了 print_info
        calc_media = media # 默认
        target_rotation = 0 if dwg_is_landscape else 1
        
        # 简化的文件名生成 (确保不报错)
        idx_str = f"{i + start_index:0{digit_width}d}"
        fname = f"{current_doc_name}-{layout_name}-{idx_str}.pdf"
        full_path = os.path.join(folderpath, fname)

        tasks.append({
            "id": i,
            "p_min": p_min, 
            "p_max": p_max,
            "full_path": full_path,
            "media": calc_media,
            "rotation": target_rotation
        })

    sys_logger.info(f"✅ 准备打印 {len(tasks)} 张图纸...")

    # ==================== 3. 稳健执行循环 ====================
    success_count = 0
    
    for t in tasks:
        sys_logger.info(f"🖨️ [ID={t['id']}] 正在打印 -> {os.path.basename(t['full_path'])}")
        
        try:
            # 🔥 调用那个测试成功的函数
            res = export_layout_window_lisp_fit(
                point_a = t["p_min"],
                point_b = t["p_max"],
                pdf_fullpath = t["full_path"],
                layout_name = layout_name,
                device = device,
                media = t["media"], # 使用任务中匹配的纸张
                ctb = ctb,
                rotation = t["rotation"]
            )
            
            if res:
                success_count += 1
                
            # 🔥🔥🔥 关键修改：强制休息，防止 CAD 缓冲区溢出 🔥🔥🔥
            time.sleep(1.5) 

            # ---------------------------------------------------------
            # WPS 清理逻辑 (⚠️ 暂时建议注释掉，排除干扰)
            # ---------------------------------------------------------
            # if wps_close_threshold > 0 and success_count % wps_close_threshold == 0:
            #     sys_logger.info("🧹 (跳过 WPS 清理以保持稳定性)")
            #     pass 

        except Exception as e:
            sys_logger.info(f"❌ 循环内异常: {e}")

    sys_logger.info(f"\n🏁 全部完成: {success_count}/{len(tasks)}")
    return True


#&&% 统一打印


def smart_print_dispatch(
        file_path=None,               # 目标文件
        operate_target="Model",       # "Model" 或 "Layout"
        layout_name=None,             # 布局名称
        select_config=None,           # 🔥 关键：默认为 None，用于触发智能判断逻辑
        *,
        # --- 透传参数 (Pass-through arguments) ---
        use_cache=False,
        start_index=0,
        digit_width=2,
        output_folder_root=None,
        device="DWG To PDF.pc3",
        media="ISO_A3_(420.00_x_297.00_MM)",
        ctb="monochrome.ctb",
        force_fixed_media=False,
        safety_delay=10,
        wps_close_threshold=6,
        xiubukuan=25 # 已经取消暂时保留
):
    """
    【函数编号】: PRINT-DISPATCHER (智能分发中心)
    【功能】: 根据空间类型和布局名称，智能决定 select_config 的默认值并调用对应引擎。
    
    【逻辑规则】:
    1. 纯模型 (Target=Model, Layout=None) -> 默认 Config=0 (大比例), 调 Model引擎
    2. 混合模型 (Target=Model, Layout=有) -> 默认 Config=1 (小比例), 调 Model引擎
    3. 纯布局/混合布局 (Target=Layout)     -> 默认 Config=1 (小比例), 调 Layout引擎
    
    * 任何情况下，如果用户传入了 select_config (不为None)，则覆盖默认值。
    """
    
    # --- 1. 智能参数决策逻辑 ---
    final_config = select_config # 初始值
    
    # 场景 A: 模型空间操作
    if operate_target == "Model":
        # 如果用户没有强制指定 config
        if final_config is None:
            if layout_name is None:
                # Case 1: 纯模型空间 -> 默认为 0 (大比例)
                final_config = 0
                sys_logger.info(f"🤖 [智能判断] 纯模型模式 -> 自动设定 select_config = 0")
            else:
                # Case 2: 混合空间模型 (有布局名但操作模型) -> 默认为 1 (小比例)
                final_config = 1
                sys_logger.info(f"🤖 [智能判断] 混合模型模式 (布局: {layout_name}) -> 自动设定 select_config = 1")
        
        # 针对 config=1 自动调整修补宽度的便利性逻辑 (可选)
        current_xiubukuan = xiubukuan
        if final_config == 1 and xiubukuan == 25: 
             current_xiubukuan = 0.25
        


        # 调用 Model 引擎
        return print_dwg_file_model(
            file_path=file_path,
            layout_name=layout_name,
            operate_target="Model",
            select_config=final_config,  # 传入最终决定的配置
            xiubukuan=current_xiubukuan,
            # 透传其他参数
            use_cache=use_cache,
            start_index=start_index,
            digit_width=digit_width,
            output_folder_root=output_folder_root,
            device=device,
            media=media,
            ctb=ctb,
            force_fixed_media=force_fixed_media,
            safety_delay=safety_delay,
            wps_close_threshold=wps_close_threshold
        )

    # 场景 B: 布局空间操作
    elif operate_target == "Layout":
        # 容错：如果布局模式没传名字，默认给布局1
        final_layout_name = layout_name if layout_name else "布局1"

        # Case 3: 纯布局/混合布局 -> 默认为 1 (小比例)
        if final_config is None:
            final_config = 1
            sys_logger.info(f"🤖 [智能判断] 布局模式 ({final_layout_name}) -> 自动设定 select_config = 1")

        # 调用 Layout 引擎
        return print_dwg_file_layout(
            file_path=file_path,
            layout_name=final_layout_name,
            operate_target="Layout",
            select_config=final_config, # 传入最终决定的配置
            # 透传其他参数 (Layout不需要xiubukuan)
            use_cache=use_cache,
            start_index=start_index,
            digit_width=digit_width,
            output_folder_root=output_folder_root,
            device=device,
            media=media,
            ctb=ctb,
            force_fixed_media=force_fixed_media,
            safety_delay=safety_delay,
            wps_close_threshold=wps_close_threshold
        )
    
    else:
        return f"❌ 参数错误: 未知的 operate_target='{operate_target}'"




#_____________________________________________________________________________________________________________________________________________________________________________
#_____________________________________________________________________________________________________________________________________________________________________________



print("__________________  CAD基本操作开始运行 _________________________")

























































































