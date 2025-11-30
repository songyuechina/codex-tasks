"""Generate sample Excel files and JSON data for Excel dict conversion demos."""
from __future__ import annotations

from pathlib import Path
import json

from openpyxl import Workbook

ROOT = Path(__file__).parent
SAMPLES_DIR = ROOT / "samples"
SAMPLES_DIR.mkdir(parents=True, exist_ok=True)

TEMPLATE_PATH = SAMPLES_DIR / "project_template.xlsx"
SAMPLE_XLSX_PATH = SAMPLES_DIR / "sample_project.xlsx"
SAMPLE_JSON_PATH = SAMPLES_DIR / "sample_project_data.json"

HEADERS = [
    "序号",
    "图纸编号",
    "图纸名称",
    "图纸规格",
    "出图比例",
    "纸张数",
    "专业名称",
    "项目名称",
    "子项目名称",
    "建设单位名称",
    "设计阶段",
    "版本号",
    "出图日期",
    "设计编号",
    "设计院名称",
]

PROJECT_INFO = {
    "项目名称": "未来城住宅一期",
    "子项目名称": "1#住宅楼",
    "建设单位名称": "未来城置业",
    "专业名称": "建筑",
    "设计阶段": "施工图",
    "版本号": "V1.0",
    "出图日期": "2025-11-15",
    "设计编号": "FC-2025-001",
    "设计院名称": "未来设计院",
}

DRAWINGS = [
    {
        "序号": 1,
        "图纸编号": "A-101",
        "图纸名称": "一层平面图",
        "图纸规格": "A1",
        "出图比例": "1:100",
        "纸张数": 1,
        "专业名称": "建筑",
    },
    {
        "序号": 2,
        "图纸编号": "A-201",
        "图纸名称": "立面图",
        "图纸规格": "A1",
        "出图比例": "1:100",
        "纸张数": 1,
        "专业名称": "建筑",
    },
]


def build_workbook(with_data: bool) -> Workbook:
    wb = Workbook()
    ws = wb.active
    for idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=idx, value=header)
    if with_data:
        # Row offset for drawings (start at row 2)
        for row_offset, drawing in enumerate(DRAWINGS, start=0):
            row = 2 + row_offset
            for key, value in drawing.items():
                col = HEADERS.index(key) + 1
                ws.cell(row=row, column=col, value=value)
            if row_offset == 0:
                # populate project info on the first drawing row
                for key, value in PROJECT_INFO.items():
                    col = HEADERS.index(key) + 1
                    ws.cell(row=row, column=col, value=value)
    return wb


def main():
    template_wb = build_workbook(with_data=False)
    template_wb.save(TEMPLATE_PATH)

    sample_wb = build_workbook(with_data=True)
    sample_wb.save(SAMPLE_XLSX_PATH)

    sample_json = {
        "project": PROJECT_INFO,
        "drawings": DRAWINGS,
    }
    SAMPLE_JSON_PATH.write_text(json.dumps(sample_json, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"[OK] Template: {TEMPLATE_PATH}")
    print(f"[OK] Sample Excel: {SAMPLE_XLSX_PATH}")
    print(f"[OK] Sample JSON: {SAMPLE_JSON_PATH}")


if __name__ == "__main__":
    main()

